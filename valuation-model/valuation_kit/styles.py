"""Excel 样式和基础写入函数。"""

import re

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BLUE = Font(color="0000FF")
BLACK = Font(color="000000")
GREEN = Font(color="008000")
GREY = Font(color="808080")
HF = Font(color="FFFFFF", bold=True)
BF = Font(bold=True)

HFILL = PatternFill("solid", fgColor="1F4E79")
BAND = PatternFill("solid", fgColor="305496")
CH = PatternFill("solid", fgColor="D9E1F2")
OUT = PatternFill("solid", fgColor="BDD7EE")
CUR = PatternFill("solid", fgColor="FCE4D6")
GREYF = PatternFill("solid", fgColor="F2F2F2")

PCT = "0.0%"
N1 = "#,##0.0;(#,##0.0);-"
N0 = "#,##0;(#,##0);-"
N2 = "0.00"
MX = '0.0"x"'
PX = "#,##0.00"

BORD = Border(bottom=Side(style="thin", color="BFBFBF"))
_YEAR = re.compile(r"^20\d\d[AE]?$")


def _esc(value):
    if isinstance(value, str) and value.startswith("="):
        return "＝" + value[1:]
    return value


def R(sheet, cell):
    return f"'{sheet}'!{cell}"


def hdr(ws, row, text, span=12):
    ws.cell(row, 1, text).font = HF
    for column in range(1, span + 1):
        ws.cell(row, column).fill = HFILL
    ws.row_dimensions[row].height = 18


def band(ws, row, text, span=12):
    ws.cell(row, 1, text).font = Font(color="FFFFFF", bold=True)
    for column in range(1, span + 1):
        ws.cell(row, column).fill = BAND
    ws.row_dimensions[row].height = 15


def inp(ws, cell, value, src=None, fmt=N1):
    target = ws[cell]
    target.value = value
    target.font = BLUE
    target.number_format = fmt or N1
    return target


def introw(ws, row, cols, vals, src=None, fmt=N1):
    for col, value in zip(cols, vals):
        if value is not None:
            inp(ws, f"{col}{row}", value, fmt=fmt)


def fml(ws, cell, formula, fmt=N1, link=False):
    target = ws[cell]
    target.value = formula
    target.font = GREEN if link else BLACK
    target.number_format = fmt
    return target


def lab(ws, cell, text, b=False, note=False):
    target = ws[cell]
    target.value = _esc(text)
    target.font = BF if b else (GREY if note else BLACK)
    return target


def logic(ws, cell, text):
    if text:
        text = re.sub("。(?=\\S)", "。\n", _esc(text).replace("**", ""))
    ws[cell].value = text
    ws[cell].font = BLACK
    ws[cell].alignment = Alignment(wrap_text=True, vertical="top")


def txt(ws, row, text, span_cols, h=30, bold=False):
    ws.merge_cells(f"A{row}:{span_cols[-1]}{row}")
    cell = ws[f"A{row}"]
    cell.value = _esc(text)
    cell.font = BF if bold else BLACK
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if h:
        ws.row_dimensions[row].height = h


def mtext(ws, row, text, last_col, nrows=2):
    logic(ws, f"A{row}", text)
    ws.merge_cells(f"A{row}:{last_col}{row + nrows - 1}")
    return row + nrows


def set_widths(ws, a_width, year_cols, year_width, logic_col=None, logic_width=56,
               cur_col=None, cur_width=14):
    ws.column_dimensions["A"].width = a_width
    for col in year_cols:
        ws.column_dimensions[col].width = year_width
    if logic_col:
        ws.column_dimensions[logic_col].width = logic_width
    if cur_col:
        ws.column_dimensions[cur_col].width = cur_width


def finalize(wb, freeze=None, default_freeze="B3"):
    freeze = freeze or {}
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=6):
            for cell in row:
                if isinstance(cell.value, str) and _YEAR.match(cell.value.strip()):
                    cell.alignment = Alignment(horizontal="right")
        ws.freeze_panes = freeze.get(ws.title, default_freeze)
