"""统一输入的工作簿生成器。"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .input import validate_input
from .styles import N0, N1, MX, PCT, PX, band, finalize, fml, hdr, inp, lab, logic, set_widths


CASES = ("Bear", "Base", "Bull")


def build_workbook(data, output_path):
    validate_input(data)
    company = data["company"]
    historical = data["years"]["historical"]
    forecast = data["years"]["forecast"]
    sources = data["sources"]
    currency = data["currency"]["price"]
    fx_now = data["currency"]["local_per_usd"]["value"]
    shares_now = company["shares_millions"]["value"]
    net_debt = company["net_debt_usd_billions"]["value"]

    def price_formula(method, case, profit_rows, multiple_ref=None):
        target_col = get_column_letter(forecast.index(method["target_year"]) + 2)
        name = method["name"]
        if name == "P/E":
            return f"'利润预测'!{target_col}{profit_rows['net_income']}*1000/{shares_now}*{multiple_ref}*{fx_now}"
        if name == "P/B":
            return f"'利润预测'!{target_col}{profit_rows['equity']}*1000/{shares_now}*{multiple_ref}*{fx_now}"
        if name == "EV/Sales":
            return f"('利润预测'!{target_col}{profit_rows['revenue']}*{multiple_ref}-{net_debt})*1000/{shares_now}*{fx_now}"
        if name == "DCF":
            assumptions = method["assumptions"][case]
            rate = assumptions["discount_rate"]
            growth = assumptions["terminal_growth"]
            terms = []
            for index, margin in enumerate(assumptions["fcf_margins"], 1):
                col = get_column_letter(index + 1)
                terms.append(f"'利润预测'!{col}{profit_rows['revenue']}*{margin}/(1+{rate})^{index}")
            last_col = get_column_letter(len(forecast) + 1)
            terminal = (f"'利润预测'!{last_col}{profit_rows['revenue']}*{assumptions['fcf_margins'][-1]}"
                        f"*(1+{growth})/({rate}-{growth})/(1+{rate})^{len(forecast)}")
            return f"(({'+'.join(terms)}+{terminal})-{net_debt})*1000/{shares_now}*{fx_now}"
        parts = []
        for part in method["assumptions"][case]:
            col = get_column_letter(forecast.index(part["target_year"]) + 2)
            parts.append(f"'利润预测'!{col}{profit_rows[part['metric']]}*{part['multiple']}")
        return f"(({'+'.join(parts)})-{net_debt})*1000/{shares_now}*{fx_now}"

    def src(source_id):
        item = sources[source_id]
        return f"{item['publisher']}，{item['title']}，{item['as_of']}，{item['currency_basis']}"

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("封面")
    ws["A1"] = f"{company['name']}（{company['ticker']}）估值模型"
    ws["A1"].font = Font(bold=True, size=18, color="305496")
    ws.merge_cells("A1:F1")
    for row, pair in enumerate((
        ("报告日期", company["report_date"]), ("数据截止", company["data_date"]),
        ("现价", f"{company['price']['value']:,.2f} {currency}"),
        ("财务口径", "收入、利润、权益和净债务统一使用十亿美元。"),
        ("估值方法", "至少使用两种估值方法交叉核对。"),
        ("Key Takeaways", "预测从业务驱动量开始，三种情景共用同一套公式。"),
    ), 2):
        ws[f"A{row}"] = pair[0]
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = pair[1]
        ws.merge_cells(f"B{row}:F{row}")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25

    ws = wb.create_sheet("历史数据")
    hdr(ws, 1, "历史数据（财务金额：$B）", len(historical) + 2)
    ws["A2"] = "项目"
    for index, year in enumerate(historical, 2):
        ws.cell(2, index, f"{year}A")
    note_col = get_column_letter(len(historical) + 2)
    ws[f"{note_col}2"] = "来源"
    hist_rows = {}
    labels = {
        "revenue": "总收入 ($B)", "net_income": "净利润 ($B)", "equity": "股东权益 ($B)",
        "shares_millions": "股本 (mn)", "price": f"年末股价 ({currency})",
        "local_per_usd": f"汇率 ({currency}/USD)",
    }
    for row, (key, title) in enumerate(labels.items(), 3):
        hist_rows[key] = row
        lab(ws, f"A{row}", title, b=key in ("revenue", "net_income", "equity"))
        series = data["history"][key]
        for index, value in enumerate(series["values"], 2):
            if value is not None:
                inp(ws, f"{get_column_letter(index)}{row}", value, fmt=PX if key == "price" else N1)
        logic(ws, f"{note_col}{row}", src(series["source_id"]))
    set_widths(ws, 24, [get_column_letter(i) for i in range(2, len(historical) + 2)], 13,
               logic_col=note_col, logic_width=58)

    ws = wb.create_sheet("业务驱动量")
    all_years = historical + forecast
    hdr(ws, 1, "[ANCHOR] 业务驱动量", len(all_years) + 2)
    ws["A2"] = "指标"
    for index, year in enumerate(all_years, 2):
        ws.cell(2, index, f"{year}{'A' if year in historical else 'E'}")
    note_col = get_column_letter(len(all_years) + 2)
    ws[f"{note_col}2"] = "来源"
    driver_rows = []
    for row, driver in enumerate(data["drivers"], 3):
        driver_rows.append(row)
        lab(ws, f"A{row}", f"{driver['name']} ({driver['unit']})", b=True)
        values = driver["historical"]["values"] + driver["forecast_base"]["values"]
        for index, value in enumerate(values, 2):
            if value is not None:
                inp(ws, f"{get_column_letter(index)}{row}", value)
        logic(ws, f"{note_col}{row}", src(driver["historical"]["source_id"]) + "；预测：" + src(driver["forecast_base"]["source_id"]))
    set_widths(ws, 28, [get_column_letter(i) for i in range(2, len(all_years) + 2)], 12,
               logic_col=note_col, logic_width=58)

    ws = wb.create_sheet("情景切换")
    hdr(ws, 1, "情景切换", len(forecast) + 2)
    lab(ws, "A2", "当前情景", b=True)
    ws["B2"] = "Base"
    validation = DataValidation(type="list", formula1='"Bear,Base,Bull"')
    ws.add_data_validation(validation)
    validation.add(ws["B2"])
    ws["A3"] = "项目"
    for index, year in enumerate(forecast, 2):
        ws.cell(3, index, f"{year}E")
    note_col = get_column_letter(len(forecast) + 2)
    ws[f"{note_col}3"] = "来源"
    scenario_rows = {case: {"drivers": []} for case in CASES}
    row = 4
    anchor_start = len(historical) + 2
    for case in CASES:
        band(ws, row, f"{case} 参数", len(forecast) + 2)
        row += 1
        case_data = data["scenarios"][case]
        for driver_index, driver in enumerate(data["drivers"]):
            scenario_rows[case]["drivers"].append(row)
            lab(ws, f"A{row}", f"{case} {driver['name']}")
            adjustment = case_data["driver_adjustments"][driver_index]
            for year_index in range(len(forecast)):
                col = get_column_letter(year_index + 2)
                anchor_col = get_column_letter(anchor_start + year_index)
                fml(ws, f"{col}{row}", f"='业务驱动量'!{anchor_col}{driver_rows[driver_index]}*(1+{adjustment})", N1, link=True)
            logic(ws, f"{note_col}{row}", src(case_data["source_id"]))
            row += 1
        scenario_rows[case]["margin"] = row
        lab(ws, f"A{row}", f"{case} 净利率")
        for index, value in enumerate(case_data["net_margin"], 2):
            inp(ws, f"{get_column_letter(index)}{row}", value, fmt=PCT)
        logic(ws, f"{note_col}{row}", src(case_data["source_id"]))
        row += 1
    set_widths(ws, 30, [get_column_letter(i) for i in range(2, len(forecast) + 2)], 13,
               logic_col=note_col, logic_width=58)

    ws = wb.create_sheet("利润预测")
    hdr(ws, 1, "利润预测（财务金额：$B）", len(forecast) + 1)
    ws["A2"] = "项目"
    for index, year in enumerate(forecast, 2):
        ws.cell(2, index, f"{year}E")
    profit_rows = {}
    row = 3
    for case in CASES:
        band(ws, row, f"{case} 案", len(forecast) + 1)
        revenue_row, margin_row, ni_row, equity_row = row + 1, row + 2, row + 3, row + 4
        profit_rows[case] = {"revenue": revenue_row, "margin": margin_row, "net_income": ni_row, "equity": equity_row}
        for target_row, title in ((revenue_row, f"{case} 总收入 ($B)"), (margin_row, f"{case} 净利率"),
                                  (ni_row, f"{case} 净利润 ($B)"), (equity_row, f"{case} 股东权益 ($B)")):
            lab(ws, f"A{target_row}", title, b=target_row != margin_row)
        for year_index in range(len(forecast)):
            col = get_column_letter(year_index + 2)
            refs = [f"'情景切换'!{col}{r}" for r in scenario_rows[case]["drivers"]]
            fml(ws, f"{col}{revenue_row}", "=" + "+".join(refs), N1, link=True)
            fml(ws, f"{col}{margin_row}", f"='情景切换'!{col}{scenario_rows[case]['margin']}", PCT, link=True)
            fml(ws, f"{col}{ni_row}", f"={col}{revenue_row}*{col}{margin_row}")
            if year_index == 0:
                last_hist_col = get_column_letter(len(historical) + 1)
                formula = f"='历史数据'!{last_hist_col}{hist_rows['equity']}+{col}{ni_row}"
            else:
                formula = f"={get_column_letter(year_index + 1)}{equity_row}+{col}{ni_row}"
            fml(ws, f"{col}{equity_row}", formula, N1, link=year_index == 0)
        row += 6
    set_widths(ws, 28, [get_column_letter(i) for i in range(2, len(forecast) + 2)], 13)

    ws = wb.create_sheet("估值方法")
    hdr(ws, 1, "估值方法及参数", 6)
    for col, value in zip("ABCDEF", ("方法", "目标年", "Bear", "Base", "Bull", "来源")):
        ws[f"{col}2"] = value
    method_rows = {}
    for row, method in enumerate(data["valuation_methods"], 3):
        method_rows[method["name"]] = row
        lab(ws, f"A{row}", method["name"], b=True)
        inp(ws, f"B{row}", method["target_year"], fmt=N0)
        for col, case in zip("CDE", CASES):
            if method["name"] in {"P/E", "P/B", "EV/Sales"}:
                value, fmt = method["multiples"][case], MX
            elif method["name"] == "DCF":
                value, fmt = method["assumptions"][case]["discount_rate"], PCT
            else:
                value, fmt = len(method["assumptions"][case]), N0
            inp(ws, f"{col}{row}", value, fmt=fmt)
        note = src(method["source_id"])
        if method["name"] == "DCF":
            note += "；C-E 列显示折现率，公式同时使用各情景的自由现金流率和永续增长率。"
        elif method["name"] == "SOTP":
            note += "；C-E 列显示各情景纳入估值的分部数量。"
        logic(ws, f"F{row}", note)
    set_widths(ws, 18, list("BCDE"), 13, logic_col="F", logic_width=68)

    ws = wb.create_sheet("情景估值")
    hdr(ws, 1, "当前情景估值", 4)
    ws["A2"] = "在“情景切换”B2 选择 Bear、Base 或 Bull；本页显示所选情景的估值。"
    for row, method in enumerate(data["valuation_methods"], 3):
        refs = []
        for case in CASES:
            pr = profit_rows[case]
            multiple = None
            if method["name"] in {"P/E", "P/B", "EV/Sales"}:
                multiple = f"'估值方法'!{dict(Bear='C', Base='D', Bull='E')[case]}{method_rows[method['name']]}"
            refs.append(price_formula(method, case, pr, multiple))
        lab(ws, f"A{row}", f"{method['name']} 隐含价 ({currency})", b=True)
        fml(ws, f"B{row}", f'=IF(\'情景切换\'!$B$2="Bear",{refs[0]},IF(\'情景切换\'!$B$2="Base",{refs[1]},{refs[2]}))', PX, link=True)

    ws = wb.create_sheet("估值对比")
    hdr(ws, 1, "估值方法对照", max(7, len(historical) + 1))
    ws["A2"] = "历史回测"
    for index, year in enumerate(historical, 2):
        col = get_column_letter(index)
        ws[f"{col}2"] = f"{year}A"
        if data["history"]["price"]["values"][index - 2] is None:
            continue
        hcol = get_column_letter(index)
        p = f"'历史数据'!{hcol}{hist_rows['price']}"
        e = f"'历史数据'!{hcol}{hist_rows['equity']}"
        s = f"'历史数据'!{hcol}{hist_rows['shares_millions']}"
        fx = f"'历史数据'!{hcol}{hist_rows['local_per_usd']}"
        fml(ws, f"{col}3", f"=((({e}*1000/{s})*({p}/({e}*1000/{s}*{fx}))*{fx})/{p})-1", PCT, link=True)
    ws["A3"] = "P/B 历史回测"
    row = 5
    multiple_cols = {"Bear": "C", "Base": "D", "Bull": "E"}
    for case in CASES:
        band(ws, row, f"{case} 案 — 估值结果", 7)
        row += 1
        price_rows = []
        for method in data["valuation_methods"]:
            price_rows.append(row)
            pr = profit_rows[case]
            multiple = None
            if method["name"] in {"P/E", "P/B", "EV/Sales"}:
                multiple = f"'估值方法'!{multiple_cols[case]}{method_rows[method['name']]}"
            lab(ws, f"A{row}", f"{method['name']} 隐含价 ({currency})", b=True)
            formula = "=" + price_formula(method, case, pr, multiple)
            fml(ws, f"B{row}", formula, PX, link=True)
            row += 1
        average = "AVERAGE(" + ",".join(f"B{value}" for value in price_rows) + ")"
        first_method = data["valuation_methods"][0]
        target_col = get_column_letter(forecast.index(first_method["target_year"]) + 2)
        lab(ws, f"A{row}", "隐含 forward P/E", b=True)
        fml(ws, f"B{row}", f"={average}/('利润预测'!{target_col}{profit_rows[case]['net_income']}*1000/{shares_now}*{fx_now})", MX, link=True)
        row += 1
        lab(ws, f"A{row}", "平均隐含价 vs 现价", b=True)
        fml(ws, f"B{row}", f"={average}/{company['price']['value']}-1", PCT)
        row += 2
    set_widths(ws, 28, list("BCDE"), 14)

    ws = wb.create_sheet("综合判断仪表盘")
    hdr(ws, 1, "综合判断仪表盘", 5)
    for row, values in enumerate((
        ("业务驱动量", "查看实际数据是否偏离 Base 路径。"),
        ("估值方法", "比较至少两种方法；结果差异较大时检查参数口径。"),
        ("情景调整", "只有新数据足以改变假设时才切换情景。"),
    ), 3):
        ws[f"A{row}"], ws[f"B{row}"] = values
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 68

    ws = wb.create_sheet("来源")
    hdr(ws, 1, "来源", 6)
    for col, value in zip("ABCDEF", ("ID", "发布方", "资料", "数据日期", "币种口径", "网址")):
        ws[f"{col}2"] = value
    for row, (source_id, item) in enumerate(sources.items(), 3):
        for col, value in zip("ABCDEF", (source_id, item["publisher"], item["title"], item["as_of"], item["currency_basis"], item["url"])):
            ws[f"{col}{row}"] = value
    for col, width in {"A": 20, "B": 22, "C": 45, "D": 14, "E": 30, "F": 68}.items():
        ws.column_dimensions[col].width = width

    finalize(wb, freeze={"封面": "A2", "来源": "A3"})
    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    return target
