# -*- coding: utf-8 -*-
"""
build_mu.py — 美光科技 (MU) 估值模型, 基于 build_kit v2。
物理锚: AIDC CapEx × 存储收入强度 (compute-semiconductor 锚, 引用 shared-base)。
USD 公司: FX 全部 = 1。
数据 SOT: `VALUATION_INPUT_DIR/MU_input.json` 对应的历史快照(2026-06-12)。
"""
import os
from openpyxl import Workbook
import build_kit as K

# ════════════ 0. 全局轴 ════════════
ALLC = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
ALLY = ["2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
HC, HY = ["B", "C", "D", "E", "F"], ["2021", "2022", "2023", "2024", "2025"]
FC = ["F", "G", "H", "I", "J", "K"]
FCf = FC[1:]
FX_FC = 1
FX_HIST = [1, 1, 1, 1, 1]
CASES = ["Bear", "Base", "Bull"]

S_COVER, S_HIST, S_PX, S_CONS = "封面", "历史财务与估值", "股价走势", "卖方研报共识"
S_HMULT, S_MULT, S_SW = "历史估值倍数", "估值倍数假设", "情景切换"
S_ANCHOR, S_SEG, S_FUND = "AIDC Capex预测", "分部测算", "利润与收入假设"
S_VAL, S_CMP, S_DASH = "情景估值", "估值对比", "综合判断仪表盘"

MONTHLY = [("2021-12", 91), ("2022-06", 55), ("2022-12", 49),
           ("2023-06", 64), ("2023-12", 85), ("2024-06", 132), ("2024-12", 84),
           ("2025-06", 105), ("2025-09", 155), ("2025-12", 285),
           ("2026-03", 470), ("2026-05", 920), ("2026-06", 1211)]
PX_NOW = 1211

wb = Workbook()
wb.remove(wb["Sheet"])

# ════════════ 1. 封面 ════════════
K.write_cover(wb.create_sheet(S_COVER), {
    "title": "美光科技 (MU) 估值模型",
    "meta": [
        ("报告日期", "2026-06-23"),
        ("数据截止", "Q2 FY2026 财报(2026-03-18, 实际营收$23.86B/GM 74.4%) + Q3 指引($33.5B/81%) + 卖方共识 + 现价 $1211"),
        ("现价", f"${PX_NOW} | 市值 ~$1.37T"),
        ("时效声明", "Q3 FY2026 财报 2026-06-24 盘后发布, 财报后必须更新本模型。"),
        ("方法一句话", "物理锚(AIDC capex)→ AI段收入=capex×强度 + 商品段周期 → 段利润 → P/E主线×FY2027前瞻EPS → 隐含价; P/B 交叉验证。"),
    ],
    "takeaways": [
        ("① 当下估值位置", "TTM P/E 55x / P/B 19.2x, 均为历史极值(周期顶低P/E规律下, 高P/E反映市场按成长股给MU定价)。"),
        ("② 核心引擎", "AI段收入强度: 2025A实际3.5% → 2026E 9.0%(价格翻倍+HBM4放量) → 2028E回落6.0%(价格normalize)。"),
        ("③ 周期判断", "2026全年紧缺(HBM售罄锁价+库存3.4周); 见顶共识2027(4Q27 DRAM); 本模型Base假设2027高位维持、2028价格normalize回落。"),
        ("④ 三情景目标价", "Bear $461 / Base $938 / Bull $1,460 (FY2027 目标年), 概率加权(25/50/25) $949 vs 现价 $1211 = -22%。"),
        ("⑤ 评级与风险", "HOLD(持有): 好公司、好周期、贵价格——现价已price in接近Bull的乐观情景。踏空险=结构性重估叙事被接受(LTAs锁价); 追高险=2027周期顶。"),
    ],
})

# ════════════ 2. 历史财务与估值 ════════════
ha = K.write_history(wb.create_sheet(S_HIST), {
    "title": "美光 历史财务与估值 ($B) — FY2021-2025A + 当下",
    "hist_cols": HC, "hist_years": HY,
    "fx_hist": FX_HIST, "fx_now": FX_FC,
    "vals_in_usd": True,
    "segments": [
        ("AI直驱 收入", [2.5, 3.5, 2.0, 5.0, 17.0], True),
        ("商品存储 收入", [21.7, 23.3, 10.5, 16.1, 15.9], True),
        ("其他(车规/工业) 收入", [3.5, 4.0, 3.0, 4.0, 4.5], False),
    ],
    "total_now": 58.0,
    "gm_pct": [0.379, 0.452, -0.091, 0.224, 0.398], "gm_now": 0.60,
    "ni": [5.86, 8.69, -5.83, 0.78, 8.54], "ni_now": 25.0,
    "eq": [43.9, 49.9, 44.1, 45.1, 54.2], "eq_now": 72.0,
    "shares": [1141, 1122, 1093, 1118, 1125], "shares_now": 1142,
    "px_end": [74, 57, 69, 96, 115],
    "px_now": PX_NOW,
    "px_avg": [82, 65, 62, 105, 110],
    "band_note": "P/B 历史常态 1-3x(2018峰值3.2x) → 当下 17.3x 远超历史带 = AI 重估 regime",
    "notes": [
        ("AI直驱 收入", "HBM+服务器DRAM(高密度DIMM/LP server)+企业级SSD; FY2025 公司口径 HBM+高密度+LP=$10B, 加eSSD约$17B。"),
        ("商品存储 收入", "Mobile/PC DRAM + 消费NAND, 强周期段; FY2023 腰斩为底。"),
        ("其他(车规/工业) 收入", "车规+工业+嵌入式, 结构性增长段。"),
        ("HREV", "总营收=10-K实际(FY8月末); 分部拆分按公司季度commentary估算。"),
        ("HGMP", "GAAP毛利率: FY2023 -9.1%为存储寒冬; 当下=TTM(Q2 FY26单季已74.9%)。"),
        ("HNI", "净利: 10-K实际; 当下=TTM ~$24.3B。"),
        ("HEQ", "股东权益: 10-K实际; 当下估$65B(高盈利快速累积)。"),
        ("HSH", "稀释股本(M股); 当下~1,130M。"),
        ("HPX", "FY年末(8月底)收盘价; 当下=2026-06-11现价$996。"),
        ("HPXA", "FY年均股价, 月度收盘均值。"),
    ],
})

# ════════════ 3. 股价走势 ════════════
def phase_fn(ym):
    if ym <= "2023-09":
        return "① 存储寒冬"
    if ym <= "2025-06":
        return "② AI叙事确立"
    if ym <= "2026-03":
        return "③ 盈利爆发"
    return "④ 万亿市值"

px = K.write_price_chart(wb.create_sheet(S_PX), MONTHLY, {
    "fn": phase_fn,
    "rows": [("① 存储寒冬", "FY2023营收腰斩至$15.5B, 毛利-9%, 股价$50-70区间磨底"),
             ("② AI叙事确立", "HBM3E量产+NVIDIA认证, 数据中心收入翻倍, 股价$85→$115"),
             ("③ 盈利爆发", "Q1-Q2 FY26营收/毛利大超预期(Q2 GM 74.4%), 股价$285→$920, 半年约+220%"),
             ("④ 万亿市值", "2026-05-26单日+19%破$1T; 现价$1211已高于卖方平均目标$945, YTD+325%")],
}, title="美光 月度股价 (USD)")

# ════════════ 4. 卖方研报共识 ════════════
K.write_consensus(wb.create_sheet(S_CONS), {
    "title": "卖方研报共识 — 36家; 评级与目标价自相矛盾是本案最大看点",
    "overview": "40家覆盖, 约94%买入(9 strongBuy/30 Buy/4 Hold/1 Sell), 平均目标$945(低于现价$1211!), 区间$249-$1,750跨度7倍 = 史诗级分歧。",
    "assumptions": [
        ("FY2026 营收/EPS", "共识 $108.7B / EPS $57.71。", "H1实际$37.5B+Q3指引$33.5B, 共识基本是指引外推, 分歧小。", "base 直接对齐共识 EPS $57。"),
        ("FY2027 EPS", "共识 ~$44.55(WSJ口径), 区间$25-$70。", "最大分歧: 价格何时normalize——H2'27(主流) vs 2028后(激进多头)。", "base $45 对齐共识, 假设 H2'27 开始回落。"),
        ("DRAM价格路径", "Gartner: 2026全年或再涨125%; Morgan Stanley: 2027-28剧烈修正。", "supercycle论 vs 周期律回归论, 完全对立。", "base: 2026涨价兑现, 2027H2转跌, 2028 -30%。"),
        ("目标倍数", "隐含P/E区间 5x(熊)~20x(牛)。", "周期股trough倍数 vs '基础设施长约'重估倍数。", "base 15x = mid-cycle上沿+结构溢价, 见『估值倍数假设』。"),
    ],
    "divergences": [
        "① 周期见顶时点: 2027H2(本模型base) vs 2028后(牛) vs 2027H1(熊) —— 决定FY2027 EPS落在$25-70哪里。",
        "② 评级vs目标价矛盾: 34/36 BUY但平均目标低于现价 —— 卖方在'追赶上调'而非'前瞻定价', 本模型不跟随。",
    ],
    "stances": [
        "UBS(买入, TP $1,625, 2026-05-26): 15x NTM P/E × C2029E EPS $117; LTAs长约把through-cycle EPS锚到>$100, 周期股变'半结构性盈利平台'。",
        "Morgan Stanley(增持, TP $1,050, 2026-06-22): 29.5x × through-cycle EPS $35, HBM mix带来结构性ASP溢价。",
        "Bernstein(跑赢, TP $510, 2026-06): 2x 2-year forward BVPS(P/B法); 担心2HCY27价格见顶、CY28 normalize。",
        "Goldman Sachs(中性, TP $400, 2026-03): 18x normalized EPS $22; 明确点名2027供给放量+CXMT, risk/reward均衡。",
    ],
})

# ════════════ 5. 历史估值倍数 ════════════
hm = K.write_hist_multiples(wb.create_sheet(S_HMULT), {
    "title": "历史估值倍数 — 自身历史带 + 当下 + 同业对照",
    "intro": "①MU自己历史上值多少(P/B常态1-3x, P/E中位16.78x, 2018盈利顶P/E仅3x) ②现在市场给多少(P/E 42x / P/B 17.3x, 历史极值) ③同业值多少。看完再去下一页拍三案倍数。",
    "s_hist": S_HIST, "ha": ha, "hist_cols": HC, "hist_years": HY,
    "yhigh": px["yhigh"], "ylow": px["ylow"],
    "fwd_note": "forward P/E ≈17.5x(现价÷FY2026E EPS $57) / ≈22x(÷FY2027E EPS $45)",
    "self_name": "美光 (MU)",
    "self_fwd_pe_label": "≈17.5x (FY26E)",
    "self_note": "本模型标的; forward 推导见『情景估值』。",
    "peers": [
        {"name": "SK海力士(HBM龙头)", "yearly": [1.8, 1.0, 1.5, 1.6, 6.0], "cur_pb": 7.5, "cur_pe": 12.0, "fwd_pe": 9.0,
         "note": "最直接可比; HBM份额62%; 同样在重估但倍数低于MU——MU溢价部分来自美国本土稀缺性。"},
        {"name": "三星电子(综合体)", "yearly": [1.8, 1.2, 1.4, 1.1, 1.5], "cur_pb": 2.8, "cur_pe": 18.0, "fwd_pe": 11.0,
         "note": "存储+代工+手机综合体, 纯度低; HBM落后但产能最大, 是供给纪律最大变量。"},
        {"name": "NVIDIA(AI龙头, 参照上沿)", "yearly": None, "cur_pb": None, "cur_pe": 48.0, "fwd_pe": 30.0,
         "note": "AI硬件估值光谱上沿; MU的P/E 42x已接近NVDA——但MU是周期股, NVDA有定价权。"},
        {"name": "标普500(参照下沿)", "yearly": None, "cur_pb": None, "cur_pe": 24.0, "fwd_pe": 21.0,
         "note": "大盘光谱; MU forward 17.5x低于大盘, 这是多头'便宜'论据的来源。"},
    ],
    "ratio": {"peer": "SK海力士(HBM龙头)",
              "note": "MU/Hynix P/B比值从历史~1x扩张到当下~2.3x: MU相对核心同业也在超额重估, 溢价依据=美国制造+1γ节点领先, 但2.3x偏高。"},
    "reading": "① 自己: 当下P/E 42x/P/B 17.3x均为历史极值, TTM倍数已不可作锚——周期股低TTM P/E=盈利顶(2018年3x), 高P/E=盈利底。② 同业: 全行业re-rating, 但MU倍数高于Hynix(HBM份额3倍于MU)。③ forward口径: 17.5x(FY26)看似合理, 但FY2026是周期峰值盈利——peak EPS × normal倍数 = 经典周期陷阱。→ 下一页: 用FY2027(含回落)做目标年。",
})

# ════════════ 6. 估值倍数假设 ════════════
ma = K.write_multiple_assumptions(wb.create_sheet(S_MULT), {
    "title": "估值倍数假设 — P/E 主线 + 三案目标倍数",
    "intro": "镜头判断+三层分解。『情景切换』引用并切换, 『情景估值』套用当前案, 『估值对比』三案并排。",
    "why_text": ("镜头选择: MU是商品型重资产周期股, 传统主线应为P/B(资产是穿越周期的存量)。但本轮HBM带来长约(5年)+定制化(逐客户认证), AI直驱段已占收入~60%, 盈利的'结构成分'显著上升 → 主线用P/E×FY2027前瞻EPS(目标年含价格normalize, 避开peak EPS陷阱), P/B做交叉验证防低P/E陷阱。"
                 "若2027-28验证'盈利底大幅抬升'(下行年EPS>$15), 市场将永久换P/E镜头; 若回到EPS<$5的深亏, 镜头退回P/B。"),
    "why_rows": 4,
    "method_text": "三层分解: ①历史mid-cycle P/E锚 ~12x(中位16.78x偏高因含亏损年高P/E; 盈利正常年10-15x) × ②结构溢价 1.25x(HBM 5年长约+寡头纪律+美国制造稀缺) × ③情绪值(三案)。一致性检验: 12×1.25×1.0=15x ≈ 当下forward P/E 17.5x量级。",
    "peak": 12.0, "peak_note": "历史mid-cycle P/E锚: 盈利正常年(FY2017/2021/2025)的P/E 10-15x取中。不用TTM 42x(极值)也不用2018年3x(盈利顶陷阱)。",
    "premium": 1.25, "premium_note": "结构溢价: HBM 5年长约(收入可见性↑) + 三寡头供给纪律 + CHIPS Act美国制造稀缺性。对账线: MU/Hynix P/B比值~2.3x偏高, 故溢价只给1.25x不给1.5x。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "hpb_row": ha["HPB"],
    "cases": [
        ("Bear", [1.20, 1.07, 1.07, 1.00, 0.95], "AI capex减速+三星破坏纪律, 价格2027H1崩塌; EPS腰斩后市场反而给高一点的'周期底倍数'(P/E 16x)。"),
        ("Base", [1.10, 1.00, 0.90, 0.85, 0.85], "2026情绪偏热(1.1)→2027回归中性(1.0, P/E 15x)→2028下行年市场前瞻复苏(0.9)。"),
        ("Bull", [1.30, 1.20, 1.10, 1.00, 0.95], "'结构性重估'叙事被接受, P/E 18x(=12×1.25×1.2); 类比2024-25年NVDA从30x→50x的re-rating。"),
    ],
    "reconcile_text": "卖方隐含目标P/E区间5-20x, 本模型三案15-18x落在区间上半段——因为我们认可结构改善(长约+纪律), 但拒绝用peak EPS×peak倍数双重乐观。",
    "source_text": "第一层=Macrotrends MU历史P/E; 第二层=MU/Hynix相对倍数对账; 第三层依据『综合判断仪表盘』D块。",
})

# ════════════ 7. 情景切换 ════════════
sw = K.write_scenario_switch(wb.create_sheet(S_SW), {
    "title": "情景切换 — 全模型唯一的情景参数库 + 切换开关 (默认 Base)",
    "usage": ("怎么用: B2 是唯一入口——下拉选案 → 各杠杆『当前案』行跟着切 → 整条链(锚→测算→利润→倍数→估值)变档。"
              "三案对比不用切: 『估值对比』恒常三列并排。情景参数只在本页改(蓝字)。"),
    "cases": CASES, "default": "Base",
    "triggers": [
        ("Bear", "Hyperscaler capex指引YoY<20% + DRAM合约价连续2季跌>10% + 三星capex执行>80% → 价格2027H1提前崩塌。"),
        ("Base", "2026紧缺兑现, 2027H2新产能(Boise Fab1+三星)释放、价格开始normalize, 2028下行年但幅度温和(-30% vs 历史-60%)。"),
        ("Bull", "AI capex 2027-28持续加速(Barclays: 共识低估$225B), HBM售罄延续至2028, 价格回调推迟且温和。"),
    ],
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "levers": [
        {"key": "capex", "name": "AIDC capex 增速", "fmt": K.PCT, "cols": FC[2:],
         "vals": {"Bear": [0.08, 0.02, 0.05, 0.07], "Base": [0.27, 0.18, 0.11, 0.09], "Bull": [0.39, 0.28, 0.15, 0.10]},
         "desc": "物理锚。2026三案共用$830B(指引已锁); 分歧在2027+: Bear=折旧悬崖踩刹车(+8%), Base=GS基线(+27%), Bull=Barclays上修(+39%)。",
         "stories": {"Bear": "折旧费用增速超过收入增速, hyperscaler集体踩刹车。", "Base": "GS 'Tracking Trillions'基线。", "Bull": "推理需求爆发, 企业级+主权AI接力。"},
         "hist": [None, 1.00, 1.33, 1.86, 1.44]},
        {"key": "aint", "name": "AI 收入强度", "fmt": K.PCT,
         "vals": {"Bear": [0.082, 0.060, 0.040, 0.036, 0.034], "Base": [0.090, 0.085, 0.060, 0.052, 0.048], "Bull": [0.098, 0.098, 0.075, 0.064, 0.058]},
         "desc": "AI段收入÷AIDC capex。历史反推: 2024A=2.5%, 2025A=3.5%(17.0/488)。2026E跳至7.6%: 价格翻倍(~2x)+HBM4放量, 有build-up支撑。2028回落=价格normalize。",
         "stories": {"Bear": "HBM份额被Hynix/三星挤压+合约价被压, 强度快速回落。", "Base": "份额守住20%, 价格normalize拖低强度。", "Bull": "HBM4份额升至25%+价格高位维持。"},
         "hist": [0.167, 0.117, 0.029, 0.025, 0.035]},
        {"key": "cprc", "name": "商品 价格涨幅", "fmt": K.PCT,
         "vals": {"Bear": [0.70, -0.20, -0.38, -0.05, 0.05], "Base": [0.90, 0.10, -0.30, 0.00, 0.08], "Bull": [1.00, 0.25, -0.12, -0.05, 0.05]},
         "desc": "商品段(Mobile/PC/消费NAND)混合ASP路径。2026: 涨价红利(DRAM合约价Q1'26环比+90%, 全年混合+45%)。分歧在2027后: 何时转跌、跌多深。",
         "stories": {"Bear": "高价破坏需求+新产能集中释放, 2027即转跌-25%。", "Base": "2027H2转跌(全年-8%), 2028 -33%。", "Bull": "缺货拖到2027底, 2028温和-15%。"},
         "hist": [0.18, -0.10, -0.45, 0.45, 0.05]},
        {"key": "hopm", "name": "AI 段营业利润率", "fmt": K.PCT,
         "vals": {"Bear": [0.62, 0.52, 0.38, 0.36, 0.38], "Base": [0.68, 0.66, 0.50, 0.48, 0.50], "Bull": [0.70, 0.68, 0.58, 0.56, 0.56]},
         "desc": "HBM/服务器DRAM段。当前Q2 FY26全司GM 74.9%/OM 67.6%为历史极值; 锚最新实际, 前瞻随价格normalize回落。",
         "stories": {"Bear": "三供成型+客户压价, OPM腰斩。", "Base": "定价权随竞争稀释, 但长约托底。", "Bull": "HBM4代际溢价维持。"},
         "hist": [None, None, None, 0.25, 0.50]},
        {"key": "copm", "name": "商品 段营业利润率", "fmt": K.PCT,
         "vals": {"Bear": [0.55, 0.25, -0.05, 0.08, 0.15], "Base": [0.62, 0.52, 0.15, 0.20, 0.26], "Bull": [0.66, 0.58, 0.30, 0.28, 0.32]},
         "desc": "纯周期段, 利润率几乎完全由价格路径决定; FY2023走过-30%深亏。",
         "stories": {"Bear": "价格崩塌, 回到盈亏线下。", "Base": "2028回落到10%(历史mid-cycle)。", "Bull": "高价多吃一年。"},
         "hist": [0.25, 0.30, -0.30, 0.10, 0.30]},
    ],
    "linked": [
        {"key": "sent", "name": "情绪值(倍数第三层)", "fmt": K.N2,
         "src_sheet": S_MULT, "src_row0": ma["sent_row0"],
         "note": "三案取值与依据见『估值倍数假设』; 本页只做切换。"},
    ],
})
_pk = f"'{S_MULT}'!{ma['pk_cell']}"
_pr = f"'{S_MULT}'!{ma['pr_cell']}"
_sent_act = sw["SWACT"]["sent"]
_r = sw["next_row"]
K.lab(wb[S_SW], f"A{_r}", "目标倍数(当前案)", b=True)
for _c in ALLC:
    K.fml(wb[S_SW], f"{_c}{_r}", f"={_pk}*{_pr}*{_c}{_sent_act}", K.MX, link=True)
K.logic(wb[S_SW], f"L{_r}", "= mid-cycle锚12x × 结构溢价1.25x × 当前案情绪值 → 喂『情景估值』的前瞻P/E。")
SWPE = _r

# ════════════ 8. 物理锚 ════════════
anchor = K.write_anchor(wb.create_sheet(S_ANCHOR), {
    "title": "全球 AI 数据中心 CapEx ($B) — 需求物理盘子",
    "all_cols": ALLC, "all_years": ALLY,
    "series": [("AI 数据中心 capex ($B)",
                [15, 30, 70, 200, 488, 830, None, None, None, None],
                "2021-25=实际(CreditSights/公司财报); 2026E=$830B(TrendForce+四大云厂指引>$700B, 三案共用); 2027+=锚×当前案增速", K.N0)],
    "yoy_row": "AI 数据中心 capex ($B)",
    "source_note": "口径=全球AI数据中心专项capex(非hyperscaler总额)。引自 shared-base/compute-aidc-base.json v1.0.0 (2026-06-03), 与 Hynix/NVDA/TSMC 模型共用同一基座。",
    "role_note": "作用: AI段收入=capex×收入强度挂在它上面。改capex → 收入 → 估值全链动。",
})
CAPEX_ROW = anchor["row_of"]["AI 数据中心 capex ($B)"]
for _i, _c in enumerate(FC[2:]):
    K.fml(wb[S_ANCHOR], f"{_c}{CAPEX_ROW}", f"={FC[1:][_i]}{CAPEX_ROW}*(1+{K.R(S_SW, _c + str(sw['SWACT']['capex']))})", K.N0, link=True)

# ════════════ 9. 分部测算 ════════════
seg = K.write_segment_model(wb.create_sheet(S_SEG), {
    "title": "分部测算 — AI直驱段(capex×强度) + 商品段(周期) ($B)",
    "all_cols": ALLC, "all_years": ALLY, "logic_col": "N",
    "groups": [
        ("AIDC capex 物理锚", [
            ("AIDC capex ($B)", None, K.N0, "= 引自『AIDC Capex预测』。改capex, AI收入跟着动。"),
        ]),
        ("AI直驱段 = capex × 收入强度", [
            ("AI 收入强度 (%)", None, K.PCT,
             "历史=实际AI段收入÷当年capex(公式反推): 2024A 2.5%, 2025A 3.5%。2021-23标n.m.(AI capex未成规模, 大基底÷小capex失真)。前瞻=『情景切换』当前案。"),
            ("AI 收入 ($B)", None, K.N1, "历史取实数; 前瞻=capex×强度。FY2026E base $63B(=830×7.6%), 对齐'H1实际+指引'外推。"),
        ]),
        ("商品段 = 周期(上年×(1+bit)×(1+价))", [
            ("商品 bit增速", [None, 0.05, -0.10, 0.08, 0.04, -0.08, 0.02, 0.05, 0.05, 0.04], K.PCT,
             "2026 bit负增长: 产能优先HBM(1颗HBM耗4x晶圆)+涨价需求破坏; normalize后恢复。"),
            ("商品 价格变化", None, K.PCT, "历史=实际混合ASP; 前瞻=『情景切换』当前案。"),
            ("商品 收入 ($B)", None, K.N1, "历史取实数; 前瞻=上年×(1+bit)×(1+价)。不挂capex, 走周期。"),
        ]),
    ],
})
m = seg["m"]
AI_HROW = ha["seg_rows"]["AI直驱 收入"]
CM_HROW = ha["seg_rows"]["商品存储 收入"]
for col in ALLC:
    K.fml(wb[S_SEG], f"{col}{m['AIDC capex ($B)']}", f"={K.R(S_ANCHOR, col + str(CAPEX_ROW))}", K.N0, link=True)
for col in HC:
    K.fml(wb[S_SEG], f"{col}{m['AI 收入强度 (%)']}", f"={K.R(S_HIST, col + str(AI_HROW))}/{col}{m['AIDC capex ($B)']}", K.PCT, link=True)
    K.fml(wb[S_SEG], f"{col}{m['AI 收入 ($B)']}", f"={K.R(S_HIST, col + str(AI_HROW))}", K.N1, link=True)
    K.fml(wb[S_SEG], f"{col}{m['商品 收入 ($B)']}", f"={K.R(S_HIST, col + str(CM_HROW))}", K.N1, link=True)
K.introw(wb[S_SEG], m["商品 价格变化"], HC, [0.18, -0.10, -0.45, 0.45, 0.05], None, K.PCT)
for col in FCf:
    K.fml(wb[S_SEG], f"{col}{m['AI 收入强度 (%)']}", f"={K.R(S_SW, col + str(sw['SWACT']['aint']))}", K.PCT, link=True)
    K.fml(wb[S_SEG], f"{col}{m['AI 收入 ($B)']}", f"={col}{m['AIDC capex ($B)']}*{col}{m['AI 收入强度 (%)']}", K.N1)
    K.fml(wb[S_SEG], f"{col}{m['商品 价格变化']}", f"={K.R(S_SW, col + str(sw['SWACT']['cprc']))}", K.PCT, link=True)
_prevs = [HC[-1]] + list(FCf[:-1])
for _p, _c in zip(_prevs, FCf):
    K.fml(wb[S_SEG], f"{_c}{m['商品 收入 ($B)']}",
          f"={_p}{m['商品 收入 ($B)']}*(1+{_c}{m['商品 bit增速']})*(1+{_c}{m['商品 价格变化']})", K.N1)
for col in FCf:
    wb[S_SEG][f"{col}{m['AI 收入 ($B)']}"].fill = K.OUT

# ════════════ 10. 利润与收入假设 ════════════
fr = K.write_fundamentals(wb.create_sheet(S_FUND), {
    "title": "利润与收入假设 — 其他增速 + 段OPM + 净利转换 + 留存 + 分部营收→利润→EPS/BPS",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
    "assum_groups": [
        ("营收驱动(AI/商品已在分部测算; 此处仅其他段)", [
            {"name": "其他 增速", "vals": [None, 0.14, -0.25, 0.33, 0.13, 1.00, 0.13, -0.11, 0.13, 0.11],
             "fmt": K.PCT, "logic": "车规/工业段; 2026涨价红利+ADAS渗透(到$8B), 2028随周期回落。"},
        ]),
        ("利润率假设(历史实际锚 + 前瞻; 粗颗粒)", [
            {"name": "AI 营业利润率", "vals": [None, None, None, 0.25, 0.50, None, None, None, None, None],
             "fmt": K.PCT, "logic": "历史=估算(早年n.m.); 前瞻=『情景切换』当前案。", "nm_cols": ["B", "C", "D"],
             "link": {"sheet": S_SW, "row": sw["SWACT"]["hopm"]}},
            {"name": "商品 营业利润率", "vals": [0.25, 0.30, -0.30, 0.10, 0.30, None, None, None, None, None],
             "fmt": K.PCT, "logic": "强周期段, FY2023深亏-30%; 前瞻=『情景切换』当前案。",
             "link": {"sheet": S_SW, "row": sw["SWACT"]["copm"]}},
            {"name": "其他 营业利润率", "vals": [0.15, 0.18, 0.05, 0.15, 0.20, 0.30, 0.28, 0.18, 0.20, 0.22],
             "fmt": K.PCT, "logic": "车规毛利稳定高于消费, 随大周期小幅波动。"},
        ]),
        ("净利转换与留存", [
            {"name": "净利转换率(净利/营业利润)", "vals": [None, 0.92, None, 0.55, 0.85, 0.92, 0.90, 0.85, 0.87, 0.88],
             "fmt": K.PCT, "logic": "营业利润扣税/利息到净利; 税率~15%; 亏损年/微利年n.m.。", "nm_cols": ["B", "D", "E"]},
            {"name": "留存率", "vals": [0.97, 0.95, None, 0.40, 0.94, 0.96, 0.96, 0.93, 0.93, 0.93],
             "fmt": K.PCT, "logic": "股息$0.60/年很小+回购受CHIPS条款限制 → 高留存扩产; 亏损年n.m.。", "nm_cols": ["D"]},
        ]),
    ],
    "segments": [
        {"name": "AI直驱 收入", "hist_row": "AI直驱 收入", "fwd": {"sheet": S_SEG, "row": m["AI 收入 ($B)"]}},
        {"name": "商品存储 收入", "hist_row": "商品存储 收入", "fwd": {"sheet": S_SEG, "row": m["商品 收入 ($B)"]}},
        {"name": "其他(车规/工业) 收入", "hist_row": "其他(车规/工业) 收入", "fwd": {"growth": "其他 增速"}},
    ],
    "profit_terms": [
        (["AI直驱 收入"], "AI 营业利润率", True),
        (["商品存储 收入"], "商品 营业利润率", False),
        (["其他(车规/工业) 收入"], "其他 营业利润率", False),
    ],
    "conv_assum": "净利转换率(净利/营业利润)", "retention_assum": "留存率",
    "note_text": "分部营收(AI段=capex×强度+商品段周期)→段驱动营业利润→净利→权益(留存递推)→EPS/BPS/ROE。下游『情景估值』直接引本表每股。",
})

# ════════════ 11. 情景估值 ════════════
sv = K.write_scenario_valuation(wb.create_sheet(S_VAL), {
    "title": "情景估值 — 当前案的逐年隐含价 (P/E 主线; P/B 交叉验证)",
    "intro": "本表输出=『情景切换』当前案(默认Base)。隐含价=目标P/E(当前案)×前瞻EPS。历史列用实际年末价反推倍数(事实); 前瞻是预测、不拟合现价。三案并排见『估值对比』。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf, "hist_years": HY,
    "fx_fwd": FX_FC, "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
    "s_fund": S_FUND, "fr": fr,
    "s_switch": S_SW, "target_row": SWPE, "sw_cell": "B2",
    "yend": px["yend"], "yavg": px["yavg"],
    "lens": "PE",
    "reading": "P/E主线读法: 隐含价=目标P/E×当年EPS。注意FY2028 EPS回落年隐含价骤降——这不是模型错误, 是周期股的本质: 估值锚定目标年FY2027(Base隐含价$938), 2028下行已在倍数的情绪折价里体现。P/B交叉验证: Base隐含价$938÷FY27 BPS $164=5.7x, 高于历史带(峰值3.2x), 反映市场对'盈利平台抬升'的重估预期——这正是P/B镜头对周期股的警示, 故主线用P/E、P/B作约束。",
    "method": "方法: 整体公司、P/E主线逐年估。基本面在『利润与收入假设』; 目标倍数在『估值倍数假设』(三层); 本表只做最后一步: 目标P/E × 前瞻EPS → 隐含价 + 市值。",
    "concl": "结论(方向性): Base目标价$938(FY2027目标年), vs现价$1211 = -22.5%。三情景概率加权$949 = -22%。现价已price in接近Bull的乐观情景, 评级HOLD(持有)——本质是'好公司、好周期、贵价格', 回调至Base区间($900附近)再评估加仓。",
})

# ════════════ 12. 估值对比 ════════════
SWB = sw["SWB"]
SH_F = K.R(S_HIST, f"$F${ha['HSH']}")
PX_NOW_REF = K.R(S_HIST, f"G{ha['HPX']}")
_ogrow = fr["am"]["其他 增速"]
_oopm = fr["am"]["其他 营业利润率"]
_conv = fr["am"]["净利转换率(净利/营业利润)"]
_ret = fr["am"]["留存率"]
_bit = m["商品 bit增速"]


def _fwdprev(j, A, key):
    return (HC[-1] if j == 0 else FCf[j - 1]) + str(A[key])


cmp_rows = [
    {"key": "cap", "label": "AIDC capex ($B)", "fmt": K.N0,
     "hist": lambda c, ci, A: f"={K.R(S_ANCHOR, c + str(CAPEX_ROW))}",
     "fwd": lambda c, j, ci, A: (f"={K.R(S_ANCHOR, 'G' + str(CAPEX_ROW))}" if j == 0
                                 else f"={FCf[j-1]}{A['cap']}*(1+{K.R(S_SW, c + str(SWB['capex'] + ci))})")},
    {"key": "ai", "label": "AI直驱 收入 ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(AI_HROW))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['cap']}*{K.R(S_SW, c + str(SWB['aint'] + ci))}"},
    {"key": "cm", "label": "商品 收入 ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(CM_HROW))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'cm')}*(1+{K.R(S_SEG, c + str(_bit))})*(1+{K.R(S_SW, c + str(SWB['cprc'] + ci))})"},
    {"key": "oth", "label": "其他 ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['seg_rows']['其他(车规/工业) 收入']))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'oth')}*(1+{K.R(S_FUND, c + str(_ogrow))})"},
    {"key": "rev", "label": "总收入 ($B)", "fmt": K.N1, "bold": True,
     "hist": lambda c, ci, A: f"={c}{A['ai']}+{c}{A['cm']}+{c}{A['oth']}",
     "fwd": lambda c, j, ci, A: f"={c}{A['ai']}+{c}{A['cm']}+{c}{A['oth']}"},
    {"key": "ni", "label": "净利 ($B)", "fmt": K.N0, "bold": True,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HNI']))}",
     "fwd": lambda c, j, ci, A: (f"=({c}{A['ai']}*{K.R(S_SW, c + str(SWB['hopm'] + ci))}"
                                 f"+{c}{A['cm']}*{K.R(S_SW, c + str(SWB['copm'] + ci))}"
                                 f"+{c}{A['oth']}*{K.R(S_FUND, c + str(_oopm))})*{K.R(S_FUND, c + str(_conv))}")},
    {"key": "eps", "label": "EPS ($)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={c}{A['ni']}*1000/{K.R(S_HIST, c + str(ha['HSH']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['ni']}*1000/{SH_F}"},
    {"key": "sent", "label": "情绪值(该案)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_MULT, c + str(ma['sent_row0'] + ci))}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_MULT, c + str(ma['sent_row0'] + ci))}"},
    {"key": "pe", "label": "目标P/E(该案)", "fmt": K.MX,
     "hist": lambda c, ci, A: f"={_pk}*{_pr}*{c}{A['sent']}",
     "fwd": lambda c, j, ci, A: f"={_pk}*{_pr}*{c}{A['sent']}"},
    {"key": "px", "label": "隐含价 ($)", "fmt": K.PX, "bold": True, "out": True,
     "hist": lambda c, ci, A: f"={c}{A['pe']}*{c}{A['eps']}",
     "fwd": lambda c, j, ci, A: f"={c}{A['pe']}*{c}{A['eps']}"},
    {"key": "up", "label": "历史: vs 实际年末价 / 前瞻: vs 现价", "fmt": K.PCT,
     "hist": lambda c, ci, A: f"={c}{A['px']}/{K.R(S_HIST, c + str(ha['HPX']))}-1",
     "fwd": lambda c, j, ci, A: f"={c}{A['px']}/{PX_NOW_REF}-1"},
]
cm_sheet = K.write_comparison(wb.create_sheet(S_CMP), {
    "title": "估值对比 — Bear / Base / Bull 三情景目标价并排",
    "intro": ("三情景各自完整推演: 物理锚 → 分部收入 → 净利 → EPS → 目标P/E → 逐年隐含价。"
              "本表三案恒常并排, 不随『情景切换』变化。未列入情景矩阵的假设三案共用Base取值。"
              "历史列=同一条链填实际值, 隐含价历史列应接近实际年末价(内置回测; 注意MU亏损年P/E失效, 历史回测看正常盈利年)。"),
    "case_names": CASES,
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "block_start": 16,
    "rows": cmp_rows,
    "summary": {
        "band": "三案汇总 (目标年 FY2027E; 各案触发条件见『情景切换』)",
        "target_col": "H",
        "rows": [
            ("净利($B)", "ni", K.N0, "由该案capex/价格/利润率假设逐年推导", False),
            ("总收入($B)", "rev", K.N0, "= AI段(capex×强度) + 商品段(该案价格路径) + 其他", False),
            ("EPS($)", "eps", K.N2, "= 净利 ÷ 稀释股本1,130M", False),
            ("目标P/E", "pe", K.MX, "= mid-cycle锚12x × 结构溢价1.25x × 该案情绪值", False),
            ("隐含价($)", "px", K.PX, "= 目标P/E × FY2027E EPS", True),
            ("vs 现价", "up", K.PCT, "对照现价$996的上行/下行空间", True),
        ],
        "mcap": {"label": "隐含市值($B)", "key": "px", "expr": f"*{SH_F}/1000",
                 "note": "= 隐含价 × 股本1,130M"},
        "concl": "风险收益比(FY2027 vs 现价$1211): Bear -62% / Base -22% / Bull +21%。概率加权(25/50/25) = $949, -22%。三案中只有Bull(超级周期延续)能打平现价之上 = 现价已price in偏乐观情景, 评级HOLD(持有)。证伪触发器见『综合判断仪表盘』。",
    },
})

# ════════════ 13. 综合判断仪表盘 ════════════
EPS27 = K.R(S_FUND, "H" + str(fr["EPS"]))
PXD = K.R(S_HIST, "G" + str(ha["HPX"]))
dash = K.write_dashboard(wb.create_sheet(S_DASH), {
    "title": "综合判断仪表盘 — A 基本面拐点 · B 估值错位 · C 催化剂 · D 情绪确认",
    "usage": ("预测引擎是B(错位)+C(催化剂); 情绪D只做timing确认+过热刹车。"
              "验收=回测: 2023年存储底部时, A块(HBM拐点)+B块(P/B 1.2x vs 该给2x)都指向买入——这套表当时能看到那波。"),
    "blocks": [
        {"title": "A. 基本面拐点 — 业务在结构性变好吗?", "rows": [
            ("产品组合迁移", "AI直驱段占收入: FY2024 20% → FY2025 45% → FY2026E ~60%", "真实的结构迁移, 不只是周期反弹——这是本案多头最强论据。"),
            ("可持续 ROE", "FY2025 ROE 16% → 当下TTM ~37% → 周期顶或达50%+", "盈利质量在升级(长约+定制), 但当前水平含大量周期成分。"),
            ("A 判断", "【强】", "结构改善是真的; 问题从来不在基本面, 在价格。", True),
        ]},
        {"title": "B. 估值错位(预测引擎 ★)— 市场现在给的 vs 基本面该给的", "rows": [
            ("市场现在给(forward P/E vs FY27)", {"fml": f"={PXD}/{EPS27}", "fmt": K.MX, "fill": True},
             "= 现价 ÷ FY2027E EPS(公式算, 随模型走)。"),
            ("基本面该给(justified P/E)", {"inp": 15.0, "fmt": K.MX},
             "= mid-cycle 12x × 结构溢价1.25x(三层分解的前两层, 情绪中性)。"),
            ("错位 GAP = 该给÷市场给 − 1",
             {"fml": lambda ro: f"=B{ro['基本面该给(justified P/E)']}/B{ro['市场现在给(forward P/E vs FY27)']}-1", "fmt": K.PCT},
             "GAP为负 = 市场给的已超过基本面该给 = 进入情绪定价区, 重估空间已被吃完。"),
            ("回测: 2023年存储底的读数", "市场给P/B 1.2x vs 该给2x+, GAP +65%", "当时错位为正且大 → 该买; 现在反过来了。"),
        ]},
        {"title": "C. 催化剂 — 什么会逼市场闭合 GAP", "rows": [
            ("Q3 FY26财报(2026-06-24)", "待", "指引EPS $19.15已是共识, beat才有增量; miss即触发回调——双向不对称(高位miss杀伤大)。"),
            ("HBM4量产爬坡+NVIDIA Rubin放量", "进行中", "Bull情景核心支撑; 看月度出货与份额数据。"),
            ("2027新产能投放(Boise Fab1+三星)", "待(2027H2)", "Base/Bear情景的周期转向扳机, 物理时点基本锁定。"),
            ("C 判断", "利好已大部分兑现", "剩余催化剂(财报beat)赔率不对称, 利空催化剂(产能投放)时点确定。", True),
        ]},
        {"title": "D. 情绪确认 — 只做 timing + 刹车", "rows": [
            ("量价温度计", "YTD+325%(一年约+10倍), 数十个交易日$0.5T→$1T, 单日+19%", "典型加速赶顶形态; 仅作温度计, 不进倍数。"),
            ("现价倍数 vs 基本面该给", "forward P/E 19x(vs FY27 EPS$63) vs 该给15x", "市场已付出超过基本面该给约28% = 情绪定价区。"),
            ("当前档位", "【过热】", "对照2018年1月/2021年11月的存储股顶部形态: 盈利完美+股价加速+卖方追赶上调, 三特征齐全。", True),
            ("衰减扳机", "6 条", "DRAM合约价连跌2季>10% / HBM份额<15% / 三星capex执行>80% / hyperscaler capex增速<20% / 毛利率环比-5pp / 现货合约价比<0.8。任一翻 → 降档+下调情绪值。"),
        ]},
    ],
    "final": {"band": "★ 综合判断(A+B+C+D 收成一句可执行的话)",
              "text": "A强(结构改善真实)但B负(错位已反向)+C不对称(利好兑现/利空定时)+D过热(赶顶形态) → 现价$1211已price in接近Bull情景, 概率加权目标$949(-22%)。评级HOLD(持有)/已持仓者不追加。这不是看空公司, 是看空价格——若回调至$850-940(Base区间), 重新评估买入。"},
    "tracking": {
        "intro": "哪个指标恶化 → 哪个假设先崩 → 触发什么动作。",
        "rows": [
            ("__band__", "一、核心驱动链"),
            ("AI 收入强度(HBM份额×价格)", "2025A 3.5% → 2026E 7.6%", "关键敏感项: AI收入=capex×强度", "季报分部数据+Counterpoint份额月报", "强度低于6.5% → 下调Base, 重算"),
            ("__band__", "二、商品周期"),
            ("DRAM合约价+库存周数", "Q1'26 +90% QoQ; 库存2-4周", "关键敏感项: 价格normalize时点", "TrendForce月度合约价", "连跌2季>10% → 见顶确认, 切Bear"),
            ("__band__", "三、需求总盘子"),
            ("AIDC capex指引", "2026E $830B(+70%)", "关键敏感项: 物理锚盘子", "四大云厂季报capex指引", "合计增速<20% → 2027E下调至<$950B全链重算"),
            ("__band__", "四、供给纪律"),
            ("三星capex执行率", "FY2026计划$73B(+130%)", "关键敏感项: 寡头纪律是否破坏", "三星季报+TrendForce产能跟踪", "执行>80% → normalize前移到2027H1"),
        ],
    },
})

# ════════════ 全局格式 + 落盘 ════════════
K.finalize(wb, freeze={
    S_HIST: "B3", S_PX: "B4", S_CONS: "A2", S_HMULT: "B5", S_MULT: "B4", S_SW: "B3",
    S_ANCHOR: "B3", S_SEG: "B3", S_FUND: "B3", S_VAL: "B4", S_CMP: "B6", S_DASH: "B6",
    S_COVER: "A2",
})
out = os.path.join(os.path.dirname(__file__), "..", "out", "MU_valuation_model.xlsx")
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("saved:", os.path.abspath(out))
print("sheets:", wb.sheetnames)
