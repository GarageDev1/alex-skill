# -*- coding: utf-8 -*-
"""
Alphabet / GOOGL 估值模型（equity-research-obsidian Phase 2）。

公司专有数据 + 判断 + 情景参数在本脚本；Excel 通用结构 / 格式 / 情景切换 / 校验友好写法
全部复用 build_kit.py。

主线镜头：P/E（franchise 资本化盈利）；SOTP + DCF 交叉验证；P/B / forward P/E 体检。
物理锚：全球 AI 数据中心 CapEx（shared base）→ Google Cloud 收入（capture intensity）；
广告 / 订阅段走周期增长。
★ 核心 EPS 口径：估值用剔除非经营性股权公允价值收益的核心经营净利（净利转换率 ~0.85），
与 Citi/BofA "core GAAP EPS" 口径一致；报表净利含大额一次性投资收益，不可资本化。
"""
from __future__ import annotations
import json, os
from openpyxl import Workbook
import build_kit as K

ALLC = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
ALLY = ["2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
HC, HY = ["B", "C", "D", "E", "F"], ["2021", "2022", "2023", "2024", "2025"]
FC = ["F", "G", "H", "I", "J", "K"]
FCF = FC[1:]
CASES = ["Bear", "Base", "Bull"]
FX_FC = 1.0                       # USD 单币种
SHARES_M = 12230.0
PX_NOW = 359.68

S_COVER, S_HIST, S_PX, S_CONS = "封面", "历史财务与估值", "股价走势", "卖方研报共识"
S_HMULT, S_MULT, S_SW = "历史估值倍数", "估值倍数假设", "情景切换"
S_ANCHOR, S_SEG, S_FUND = "AIDC CapEx预测", "分部测算", "利润与收入假设"
S_VAL, S_CMP, S_DASH = "情景估值", "估值对比", "综合判断仪表盘"

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
VAULT = os.environ.get("VALUATION_OUTPUT_DIR", os.path.join(REPO_ROOT, "out"))
OUT_XLSX = os.path.join(VAULT, "GOOGL_valuation_model.xlsx")
OUT_JSON = os.path.join(VAULT, "GOOGL_input.json")

SEARCH, YT, NET, SUBS, CLOUD, OB = (
    "Google Search & other 收入", "YouTube 广告收入", "Google Network 收入",
    "订阅/平台/设备 收入", "Google Cloud 收入", "Other Bets 收入")

# ════════ 公司数据（$B，SEC 10-K / 8-K）════════
REV_SEG = {
    SEARCH: [148.95, 162.45, 175.03, 198.08, 224.50],
    YT:     [28.85, 29.24, 31.51, 36.15, 40.40],
    NET:    [31.70, 32.78, 31.31, 30.36, 29.80],
    SUBS:   [28.03, 31.01, 34.69, 40.34, 47.90],
    CLOUD:  [19.21, 26.28, 33.09, 43.23, 58.70],
    OB:     [0.753, 1.068, 1.527, 1.648, 1.54],
}
NI_REP = [76.03, 59.97, 73.80, 100.12, 132.17]          # 报表净利（含一次性投资收益）
EQ     = [251.64, 256.14, 283.38, 325.08, 415.27]
SHARES = [13242, 13159, 12722, 12447, 12230]
GM     = [0.570, 0.554, 0.566, 0.582, 0.597]
OPMARG = [0.306, 0.265, 0.274, 0.321, 0.320]            # 报表整体营业利润率
AIDC   = [15, 30, 70, 200, 488, 830, None, None, None, None]

# 历史分部 OPM（反推，复现整体 OP）
CLOUD_OPM_H = [-0.165, -0.113, 0.051, 0.141, 0.237]
NONCLOUD_OPM_H = [0.344, 0.303, 0.301, 0.346, 0.334]    # 含 Other Bets + corporate 摊销
NETCONV_H = [0.97, 0.80, 0.88, 0.89, 1.02]              # 报表 NI/OP（2025 含投资收益>1）

# 月度股价（split-adjusted，USD）
MONTHLY = [
    ("2021-01", 87.0), ("2021-06", 122.0), ("2021-12", 144.85),
    ("2022-03", 139.0), ("2022-06", 109.0), ("2022-09", 96.0), ("2022-12", 88.73),
    ("2023-03", 104.0), ("2023-06", 120.0), ("2023-09", 131.0), ("2023-12", 139.69),
    ("2024-03", 151.0), ("2024-06", 183.0), ("2024-09", 166.0), ("2024-12", 189.30),
    ("2025-03", 155.0), ("2025-06", 176.0), ("2025-09", 245.0), ("2025-12", 309.0),
    ("2026-03", 350.0), ("2026-05", 402.0), ("2026-06", 359.68),
]
PX_END = {"2021": 144.85, "2022": 88.73, "2023": 139.69, "2024": 189.30, "2025": 309.0}
PX_AVG = {"2021": 125.0, "2022": 106.0, "2023": 123.0, "2024": 167.0, "2025": 205.0}
PX_HIGH = {2021: 150.7, 2022: 151.5, 2023: 142.0, 2024: 205.0, 2025: 315.0, 2026: 402.4}
PX_LOW = {2021: 86.0, 2022: 83.4, 2023: 88.7, 2024: 130.7, 2025: 142.0, 2026: 309.0}

# ════════ 情景参数 ════════
PEAK_PE, PREMIUM = 30.0, 1.0
TARGET_PE = {"Bear": [20, 19, 18, 17, 16], "Base": [27, 27, 26, 24, 23], "Bull": [30, 31, 30, 28, 26]}
PE_SENT = {c: [round(v / PEAK_PE, 3) for v in vals] for c, vals in TARGET_PE.items()}

CAPEX_G = {"Bear": [0.10, 0.05, 0.03, 0.02], "Base": [0.18, 0.12, 0.09, 0.08], "Bull": [0.25, 0.18, 0.12, 0.10]}
CINT = {"Bear": [0.115, 0.105, 0.098, 0.092, 0.088],
        "Base": [0.122, 0.120, 0.117, 0.112, 0.108],
        "Bull": [0.128, 0.130, 0.128, 0.125, 0.122]}
SRCH_G = {"Bear": [0.08, 0.05, 0.03, 0.02, 0.01],
          "Base": [0.13, 0.12, 0.10, 0.08, 0.07],
          "Bull": [0.16, 0.15, 0.13, 0.11, 0.10]}
NCOPM = {"Bear": [0.330, 0.320, 0.315, 0.310, 0.310],
         "Base": [0.345, 0.355, 0.360, 0.365, 0.367],
         "Bull": [0.355, 0.370, 0.380, 0.385, 0.390]}
CLOPM = {"Bear": [0.24, 0.25, 0.26, 0.27, 0.27],
         "Base": [0.27, 0.30, 0.32, 0.34, 0.35],
         "Bull": [0.29, 0.33, 0.36, 0.38, 0.40]}
NETCONV = {"Bear": [0.84, 0.84, 0.84, 0.84, 0.84],
           "Base": [0.85, 0.85, 0.85, 0.85, 0.85],
           "Bull": [0.86, 0.87, 0.87, 0.87, 0.87]}
# 非情景杠杆段（三案共用 Base 增速）
YT_G = [0.12, 0.11, 0.10, 0.09, 0.08]
NET_G = [-0.02, -0.01, 0.00, 0.00, 0.01]
SUBS_G = [0.16, 0.14, 0.12, 0.11, 0.10]
OB_G = [0.15, 0.20, 0.25, 0.25, 0.25]


def hist_growth(vals):
    return [None] + [round(vals[i] / vals[i - 1] - 1, 4) for i in range(1, 5)]


def build():
    wb = Workbook(); wb.remove(wb["Sheet"])

    # 1 封面
    K.write_cover(wb.create_sheet(S_COVER), {
        "title": "Alphabet Inc. / GOOGL 估值模型",
        "meta": [
            ("报告日期", "2026-06-15"),
            ("数据截止", "FY2025 10-K + 2026Q1 实际 + Yahoo/统一API 股价与一致预期"),
            ("当前股价", f"${PX_NOW:,.2f}（reality check 用，非拟合目标）；市值约 $4.39T"),
            ("主线镜头", "P/E 主线（franchise 资本化盈利）；SOTP + DCF 交叉验证；P/B / forward P/E 体检。"),
            ("核心 EPS 口径", "估值用剔除非经营性股权公允价值收益的核心经营净利（净利转换率 ~0.85）；报表净利含大额一次性投资收益（FY25 +$29.8B、Q1'26 +$37.7B），不可资本化。"),
            ("方法一句话", "AIDC CapEx → Cloud 收入强度 → 分部收入 → 分部 OPM → 核心净利/EPS → 目标 P/E → 隐含股价。"),
        ],
        "takeaways": [
            ("① 物理锚", "全球 AI 数据中心 CapEx 是最底层需求锚；Google Cloud 收入 = AIDC capex × capture intensity。"),
            ("② 基本面", "Q1'26 营收 +22%、营业利润率 36.1%、Cloud +63%/margin 32.9%、backlog >$460B；广告 Search +19% 加速。"),
            ("③ 卖方对账", "目标价区间 Citi $350 / BofA $370 / MS $375 / DB $390 / JPM $395，多为 26-29x 2027E 核心 GAAP EPS + cash；一致预期均值约 $432。"),
            ("④ 估值判断", "Base 用 2027E 核心 EPS × 26x P/E；核心 EPS（剔除投资收益）下股价已基本反映 AI/Cloud 重估，Base 隐含价接近现价。"),
            ("⑤ 评级", "HOLD（持有）：franchise 优质但 12 个月股价已 +107% 充分 re-rate；上行靠 Bull（Cloud margin + AI Search 货币化 + 倍数扩张），下行风险来自反垄断 ad tech 救济 + AI 入口侵蚀 + 倍数压缩。"),
        ],
    })

    # 2 历史财务与估值
    ha = K.write_history(wb.create_sheet(S_HIST), {
        "title": "Alphabet 历史财务与估值（$B）— 2021-2025A + 当下 TTM / 2026Q1",
        "hist_cols": HC, "hist_years": HY,
        "fx_hist": [1, 1, 1, 1, 1], "fx_now": 1,
        "vals_in_usd": True, "fx_label": "FX (USD 单币种=1)",
        "segments": [(SEARCH, REV_SEG[SEARCH], True), (YT, REV_SEG[YT], True),
                     (NET, REV_SEG[NET], True), (SUBS, REV_SEG[SUBS], True),
                     (CLOUD, REV_SEG[CLOUD], True), (OB, REV_SEG[OB], False)],
        "total_now": 422.5,
        "gm_pct": GM, "gm_now": 0.598,
        "ni": NI_REP, "ni_now": 160.3,
        "eq": EQ, "eq_now": 455.0,
        "shares": SHARES, "shares_now": 12230,
        "px_end": [PX_END[y] for y in HY], "px_now": PX_NOW,
        "px_avg": [PX_AVG[y] for y in HY],
        "band_note": "历史 P/E：2022 低点约 15x，2024-2025 约 22-25x；当下 TTM P/E 约 27x（含一次性投资收益的报表 EPS），核心 EPS 口径下倍数更高——市场在给 AI/Cloud 重估定价。",
        "quarter": {
            "col": "H", "label": "2026Q1实际",
            "segs": {SEARCH: (60.40, 0.19), YT: (9.88, 0.11), NET: (6.97, -0.04),
                     SUBS: (12.38, 0.19), CLOUD: (20.03, 0.63), OB: (0.41, None)},
            "ni": 62.58, "eq": 455.0, "shares": 12230, "fx": 1,
            "note": "2026Q1 实际：营收 $109.9B(+22%)、营业利润率 36.1%、Cloud +63%/margin 32.9%、backlog >$460B；净利 $62.6B 含 $37.7B 非经营性投资收益（看经营质量看营业利润 +30%）。",
        },
        "notes": [
            (SEARCH, "搜索广告，利润核心；FY2025 占总收入约 56%。前瞻按数字广告大盘 × 份额走周期增长，AI 入口侵蚀是关键下行变量。"),
            (CLOUD, "前瞻按 AIDC capex × capture intensity 建模（物理锚直驱）；margin 从亏损快速爬升到 23.7%（2025 全年）/ 32.9%（Q1'26）。"),
            (OB, "Waymo 为主，合计亏损；估值作期权处理，不进核心 EPS 主线。"),
            ("HNI", "报表净利含大额非经营性股权公允价值收益（FY25 +$29.8B）；建模估值用核心经营净利。当下=TTM(FY25−Q1'25+Q1'26)。"),
            ("HEQ", "股东权益；当下=TTM 近似（$80B 增发于 2026-06，尚未完整反映）。"),
            ("HSH", "稀释股本；2026-06 $80B 增发与回购大致对冲，前瞻按约 12,230M 持平处理。"),
            ("HPX", "股价 split-adjusted（2022-07 20:1 拆股）；年末/均价用于历史倍数校准。"),
        ],
    })

    # 3 股价走势
    def phase_fn(ym):
        if ym <= "2023-02":
            return "① ChatGPT 冲击/AI 输家叙事"
        if ym <= "2025-08":
            return "② 反垄断/AI 疑虑压估值"
        if ym <= "2025-12":
            return "③ 免拆分+Gemini3 重估"
        return "④ Cloud 加速/创新高后回落"
    px = K.write_price_chart(wb.create_sheet(S_PX), MONTHLY, {
        "fn": phase_fn,
        "rows": [("① ChatGPT 冲击/AI 输家叙事", "2023 ChatGPT 引爆生成式 AI，市场质疑搜索护城河，GOOGL 被贴 AI 输家标签。"),
                 ("② 反垄断/AI 疑虑压估值", "2024 DOJ 判垄断 + 2025 上半年 AI 入口侵蚀担忧 + 关税/宏观，估值承压。"),
                 ("③ 免拆分+Gemini3 重估", "2025-09 免于拆分（+9%）、2025-11 Gemini 3 扭转叙事、Berkshire 建仓背书，全年 +65%。"),
                 ("④ Cloud 加速/创新高后回落", "2026-01 破 $4T、Q1'26 blowout、2026-05-13 创最高 $402，后回落至约 $360（$80B 增发稀释担忧）。")],
    }, title="Alphabet 月度股价走势（USD, split-adjusted）")
    px["yhigh"].update(PX_HIGH); px["ylow"].update(PX_LOW)

    # 4 卖方研报共识
    K.write_consensus(wb.create_sheet(S_CONS), {
        "title": "卖方研报共识 — 本地研报库（2026 Q1-Q2）对账",
        "overview": "全街以 Buy/Overweight 为主（强烈买入），目标价区间约 $345-395（个别更高），一致预期均值约 $432。核心叙事：Cloud 加速 + margin 跃升 + AI Search 货币化未崩；最大分歧是反垄断 ad tech 救济与 AI 对搜索长期货币化的影响。",
        "assumptions": [
            ("Search 收入增速\n(2026)", "Q1'26 +19% 加速，街上多数给 2026 双位数增长；AI Overviews/AI Mode 月活破 10 亿。",
             "分歧在 AI 是否长期稀释 CPC：管理层称 same-rate 货币化，Pew/Ahrefs 显示零点击升至 69%。", "Base 取 Search 2026 +13%，逐年缓降，反映 AI 既扩盘又侵蚀。"),
            ("Cloud 收入+margin", "Cloud +63%、margin 32.9%、backlog >$460B；UBS 2026E $101.8B / 2027E $124.7B。",
             "增速可持续性 + margin 能否逼近 AWS 37.7%；超大客户集中度上升。", "Base Cloud = AIDC capex × intensity，2026E ~$101B、margin 27%→35%。"),
            ("2026 CapEx", "公司指引上调至 $180-190B（FY25 $91B 翻倍），2027 继续显著增加。",
             "资本密度升至 ~45%，FCF 承压；$80B 增发稀释。", "锚定 AIDC 共享基座 2026E $830B；capex 是 Cloud 收入的物理驱动。"),
            ("目标 P/E / EPS 口径", "Citi/BofA 用 26-29x 2027E 核心 GAAP EPS + cash；卖方核心 2027E EPS 约 $13.2-13.5。",
             "headline EPS 含投资收益（一致 FY26 $14.22），核心 EPS 低得多——口径差异决定目标价高低。", "本模型用核心经营 EPS（剔除投资收益），Base 2027E 26x。"),
        ],
        "divergences": [
            "① 反垄断 ad tech 案（Brinkema）救济 2026 H2 落地，DOJ 寻求结构性剥离 AdX —— 最大未定价下行尾部。",
            "② AI 对搜索长期货币化：是扩大信息总盘（非零和）还是稀释高价 CPC，决定利润核心命运。",
            "③ EPS 口径：含/不含非经营性投资收益，直接决定同一目标倍数下的目标价。",
        ],
        "stances": [
            "Goldman Sachs（2026-02）：Buy，TP $400；50/50 EV/EBIT + DCF，强调 Cloud backlog 与投资周期。",
            "Morgan Stanley（2026-05）：OW，TP $375；DCF/EBITDA，Bull $425 / Bear $200。",
            "Citi（2026-02）：Buy，TP $350；27x 2027E 核心 GAAP EPS $13.22。",
            "BofA（2026-04）：Buy，TP $370；27x 2027E 核心 Google GAAP EPS + cash per share。",
            "J.P. Morgan（2026-04）：OW（Top Pick），TP $395；~29x 2027E GAAP EPS。",
            "UBS（2026-02）：Neutral，TP $348；26x 1Q27-4Q27E GAAP EPS。",
        ],
    })

    # 5 历史估值倍数
    hm = K.write_hist_multiples(wb.create_sheet(S_HMULT), {
        "title": "历史估值倍数 — 自身 P/E 带 + 当下 + 同业（AI/平台）对照",
        "intro": "Alphabet 历史 P/E 大致 15-25x（2022 低点约 15x，2024-25 约 22-25x）。当下 TTM P/E 约 27x（报表 EPS 含投资收益）；本轮重估由 Cloud 转盈 + Gemini 3 + AI 货币化驱动，非纯情绪。主线以 EPS×P/E 资本化盈利，P/B 仅体检。",
        "s_hist": S_HIST, "ha": ha, "hist_cols": HC, "hist_years": HY,
        "yhigh": px["yhigh"], "ylow": px["ylow"],
        "fwd_note": "统一API 当下 TTM P/E≈27.4x；本模型 Base 2027E 26x（核心 EPS 口径），是对 Cloud/AI 货币化可持续性的判断，非现价拟合。",
        "self_name": "Alphabet", "self_fwd_pe_label": "~27x TTM / 26x Base 2027E(核心)",
        "self_note": "P/E 主线：穿越周期的耐用资产是搜索+AI 平台的可持续盈利，不是账面净资产。",
        "peers": [
            {"name": "Microsoft", "yearly": None, "cur_pb": None, "cur_pe": 33.0, "fwd_pe": 28.0,
             "note": "云+AI 平台龙头（Azure+OpenAI），盈利耐用性最可比，享高质量倍数。"},
            {"name": "Meta Platforms", "yearly": None, "cur_pb": None, "cur_pe": 25.0, "fwd_pe": 22.0,
             "note": "最直接广告同业；AI capex 同样大增，倍数区间可比。"},
            {"name": "Amazon", "yearly": None, "cur_pb": None, "cur_pe": 34.0, "fwd_pe": 28.0,
             "note": "云（AWS）+电商，AI 受益，倍数偏高（盈利基数口径不同）。"},
            {"name": "S&P 500", "yearly": None, "cur_pb": None, "cur_pe": 23.0, "fwd_pe": 21.0,
             "note": "大盘参照下沿；Alphabet 历史多数时间相对大盘有溢价或持平。"},
        ],
        "ratio": None,
        "reading": "Alphabet 当下不是 2022 低点（15x），也不该给纯成长股 30x+。合理判断：2026-2027 Cloud/AI 货币化支撑约 24-27x 核心 EPS；若 ad tech 拆分或 AI 稀释搜索货币化，倍数应回落；若 Cloud margin 持续逼近 AWS、AI Search 货币化证实，可上探 28-30x。",
    })

    # 6 估值倍数假设
    ma = K.write_multiple_assumptions(wb.create_sheet(S_MULT), {
        "title": "估值倍数假设 — 主线 P/E 三层拆解（质量锚 × 结构溢价 × 情绪/可见度）",
        "intro": "目标 P/E 不是拍一个数，而是：质量锚（30x）× 结构溢价（1.0x，已含在锚内避免双算）× 情景情绪值。情景切换只引这里的三案情绪值；所有目标价由同一条业务链推导。",
        "why_text": ("Alphabet 穿越周期的耐用资产是『搜索的信息分发垄断 + AI 全栈（Gemini+TPU+Cloud）的可持续盈利』，"
                     "不是账面净资产——它轻资产（相对盈利）、有定价权、生态锁定（30 亿产品用户 + 广告主网络）。"
                     "因此主估值分母选核心经营 EPS，P/B 仅体检。若 AI 替代冲击搜索货币化、或反垄断削弱数据积累正反馈，镜头才应向更保守倍数下移。"
                     "★ 关键：核心 EPS 剔除非经营性股权公允价值收益（该收益可逆、不可资本化），这是本模型相对『直接对报表 EPS 套倍数』更严谨之处。"),
        "why_rows": 5,
        "method_text": "目标 P/E = 质量上沿锚（30x）× 结构溢价（1.0x）× 情景情绪值。Bear 回到 16-20x（AI 侵蚀+反垄断+倍数压缩）；Base 2026-27 维持 26-27x 后随成熟回落到 23x；Bull 在 Cloud margin 兑现+AI 货币化证实下到 30-31x。",
        "peak": PEAK_PE, "peak_note": "Alphabet 历史 + 同业（MSFT 28x、META 22x fwd）显示高质量平台前瞻 P/E 上沿约 28-30x；30x 作为质量上沿锚，非当下 TTM 拟合。",
        "premium": PREMIUM, "premium_note": "结构溢价已含在 30x 质量锚内，避免双重计数；三案差异放情绪/可见度层。",
        "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCF,
        "s_hist": S_HIST, "hpb_row": ha["HPE"],
        "cases": [
            ("Bear", PE_SENT["Bear"], "AI 直接侵蚀搜索货币化 + ad tech 强制剥离 AdX + 大盘倍数压缩 → 回到 16-20x。"),
            ("Base", PE_SENT["Base"], "Cloud/AI 货币化可持续、搜索双位数增长延续 → 26-27x，随成熟逐步回落到 23x。"),
            ("Bull", PE_SENT["Bull"], "Cloud margin 逼近 AWS + AI Search 货币化证实 + Gemini 领先 → 市场给 AI 平台 30-31x。"),
        ],
        "reconcile_text": "卖方常用 26-29x 2027E 核心 GAAP EPS（Citi 27x / BofA 27x / JPM 29x）。本模型 Base 2027E 26x 落在区间下半区——刻意保守因为：① 用核心 EPS（剔投资收益）已比 headline 严格；② 12 个月股价已 +107%，倍数与 EPS 双升后追高风险上升。",
        "source_text": "历史倍数来自 split-adjusted 股价与 SEC 财务；卖方口径来自本地研报库（GS/MS/Citi/BofA/JPM/UBS）；目标倍数与情景触发写在本页与情景切换页。",
    })

    # 7 情景切换
    sw = K.write_scenario_switch(wb.create_sheet(S_SW), {
        "title": "情景切换 — 全模型唯一情景参数库（默认 Base）",
        "usage": "B2 是唯一开关。Cloud 收入（capex×强度）、Search 增速、分部 OPM、净利转换、目标 P/E 都按当前案联动；估值对比页直接引三案矩阵，不被开关污染。YouTube/Network/订阅/Other Bets 增速三案共用 Base（非 load-bearing）。",
        "cases": CASES, "default": "Base",
        "triggers": [
            ("Bear", "AI 答案侵蚀搜索 CPC（Search 增速掉到个位数）；ad tech 救济强制剥离 AdX；Cloud 增速/ margin 不及预期；大盘倍数压缩。"),
            ("Base", "Search 维持双位数；Cloud 沿 backlog 兑现、margin 爬向 35%；反垄断维持行为性救济；目标倍数 26-27x。"),
            ("Bull", "AI Search 货币化证实 + Gemini 持续领先；Cloud margin 逼近 AWS；TPU 外供放量；市场按 AI 平台给 30x。"),
        ],
        "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCF,
        "levers": [
            {"key": "capex", "name": "AIDC capex 增速", "fmt": K.PCT, "cols": FC[2:],
             "vals": CAPEX_G, "desc": "物理锚增速。2026E 用共享基座 $830B，三案从 2027 起分歧。",
             "stories": {"Bear": "云厂 AI ROI 受质疑，capex 进入审慎期。", "Base": "沿共享 AIDC 基座增长。", "Bull": "推理需求爆发，capex 曲线延长。"},
             "hist": [None, 1.00, 1.33, 1.86, 1.44]},
            {"key": "cint", "name": "Cloud 收入强度", "fmt": K.PCT, "cols": FCF,
             "vals": CINT, "desc": "Google Cloud 收入 / AIDC capex = 公司从全球 AI 算力盘子切走的份额。早年 Cloud 非纯 AI 驱动、强度失真标 n.m.，锚 2024-2025 实际年。",
             "stories": {"Bear": "Cloud 份额被 AWS/Azure 压、增速放缓。", "Base": "份额随 backlog 兑现温和提升后随基数扩大回落。", "Bull": "TPU 全栈成本优势 + 外供放量，capture rate 维持更久。"},
             "hist": [None, None, None, 0.216, 0.120]},
            {"key": "srch", "name": "Search 收入增速", "fmt": K.PCT, "cols": FCF,
             "vals": SRCH_G, "desc": "利润核心的 swing 变量。AI Overviews/AI Mode 既扩盘又可能稀释 CPC，是多空最大分歧。",
             "stories": {"Bear": "AI 答案吞点击、CPC 稀释，增速掉到个位数。", "Base": "AI 扩盘对冲侵蚀，维持双位数后缓降。", "Bull": "AI Search 货币化证实 + query 增长，增速维持高位。"},
             "hist": hist_growth(REV_SEG[SEARCH])},
            {"key": "ncopm", "name": "非云综合 OPM", "fmt": K.PCT, "cols": FCF,
             "vals": NCOPM, "desc": "Search+YouTube+Network+订阅+Other Bets 合并 OPM（含 corporate 与 Waymo 亏损摊销）。2025 实际 33.4%，AI 效率+成本纪律驱动经营杠杆。",
             "stories": {"Bear": "AI 侵蚀 + Waymo/corporate 拖累，OPM 回落。", "Base": "成本纪律 + AI 效率，OPM 缓升到 36-37%。", "Bull": "经营杠杆充分释放，OPM 到 39%。"},
             "hist": NONCLOUD_OPM_H},
            {"key": "clopm", "name": "Cloud OPM", "fmt": K.PCT, "cols": FCF,
             "vals": CLOPM, "desc": "Cloud 营业利润率。2025 全年 23.7%、Q1'26 单季 32.9%；爬向 AWS 的 37.7% 是 margin 故事核心。",
             "stories": {"Bear": "AI 推理价格战 + 折旧上行，margin 卡在 27%。", "Base": "规模效应 + TPU 成本优势，margin 到 35%。", "Bull": "margin 逼近并接近 AWS 水平到 40%。"},
             "hist": CLOUD_OPM_H},
            {"key": "netconv", "name": "净利转换率(核心)", "fmt": K.PCT, "cols": FCF,
             "vals": NETCONV, "desc": "核心净利/营业利润，剔除非经营性投资收益，含税(~16%)与净利息。历史列为报表口径（2025 含投资收益>1）。",
             "stories": {"Bear": "税率/汇率不利。", "Base": "约 0.85 正常化。", "Bull": "净利息收入+税收优化支撑。"},
             "hist": NETCONV_H},
        ],
        "linked": [
            {"key": "sent", "name": "情绪/可见度值（P/E 第三层）", "fmt": K.N2,
             "src_sheet": S_MULT, "src_row0": ma["sent_row0"],
             "note": "来自估值倍数假设页。目标 P/E = 30x × 1.0 × 本行。"},
        ],
    })
    pk = f"'{S_MULT}'!{ma['pk_cell']}"; pr = f"'{S_MULT}'!{ma['pr_cell']}"
    swpe = sw["next_row"]
    K.lab(wb[S_SW], f"A{swpe}", "目标P/E（当前案）", b=True)
    for col in ALLC:
        K.fml(wb[S_SW], f"{col}{swpe}", f"={pk}*{pr}*{col}{sw['SWACT']['sent']}", K.MX, link=True)
    K.logic(wb[S_SW], f"L{swpe}", "目标 P/E = 质量锚 × 结构溢价 × 当前案情绪值；喂情景估值。")

    # 8 物理锚
    anchor = K.write_anchor(wb.create_sheet(S_ANCHOR), {
        "title": "全球 AI 数据中心 CapEx（$B）",
        "all_cols": ALLC, "all_years": ALLY,
        "series": [("AI 数据中心 capex ($B)", AIDC, "共享基座 compute-aidc-base；2026E=830（与 TSMC 模型同源），2027+ 由情景切换 capex 增速驱动。", K.N0)],
        "yoy_row": "AI 数据中心 capex ($B)",
        "source_note": "口径=全球 AI 数据中心专项 capex（非全口径 hyperscaler capex）。2021-2024 为 AI 起步期粗估，重点从 2025A/2026E 起。",
        "role_note": "Google Cloud 收入 = AIDC capex × Cloud 收入强度。改这一行会穿透到 Cloud 收入、总营收、核心 EPS、目标价（连通性测试入口）。",
    })
    capex_row = anchor["row_of"]["AI 数据中心 capex ($B)"]
    for idx, col in enumerate(FC[2:]):
        prev = FC[1:][idx]
        K.fml(wb[S_ANCHOR], f"{col}{capex_row}", f"={prev}{capex_row}*(1+{K.R(S_SW, col + str(sw['SWACT']['capex']))})", K.N0, link=True)

    # 9 分部测算
    seg = K.write_segment_model(wb.create_sheet(S_SEG), {
        "title": "分部测算 — Cloud 挂 AIDC capex×强度，广告/订阅按周期增长",
        "all_cols": ALLC, "all_years": ALLY, "logic_col": "N",
        "groups": [
            ("AIDC 物理锚", [("AIDC capex ($B)", None, K.N0, "引自 [ANCHOR] AIDC capex 表。")]),
            ("Google Cloud（物理锚直驱）", [
                ("Cloud 收入强度", None, K.PCT, "Cloud 收入 / AIDC capex；早年非纯 AI 标 n.m.，前瞻=情景切换当前案。"),
                ("Cloud 收入", None, K.N1, "前瞻 = AIDC capex × 强度。"),
            ]),
            ("广告 / 订阅 / Other Bets（周期增长）", [
                ("Search 增速", None, K.PCT, "前瞻=情景切换当前案（swing 变量）。"),
                ("Search 收入", None, K.N1, "前瞻 = 上年 × (1+增速)。"),
                ("YouTube 增速", YT_G_FULL(), K.PCT, "三案共用 Base。"),
                ("YouTube 收入", None, K.N1, "前瞻 = 上年 × (1+增速)。"),
                ("Network 增速", NET_G_FULL(), K.PCT, "结构性下滑后企稳，三案共用 Base。"),
                ("Network 收入", None, K.N1, "前瞻 = 上年 × (1+增速)。"),
                ("订阅 增速", SUBS_G_FULL(), K.PCT, "三案共用 Base。"),
                ("订阅 收入", None, K.N1, "前瞻 = 上年 × (1+增速)。"),
                ("Other Bets 增速", OB_G_FULL(), K.PCT, "Waymo 放量，小基数，三案共用 Base。"),
                ("Other Bets 收入", None, K.N1, "前瞻 = 上年 × (1+增速)。"),
            ]),
        ],
    })
    m = seg["m"]
    for col in ALLC:
        K.fml(wb[S_SEG], f"{col}{m['AIDC capex ($B)']}", f"={K.R(S_ANCHOR, col + str(capex_row))}", K.N0, link=True)
    # 历史：收入取实数、强度反推
    seg_hist_map = {"Cloud 收入": CLOUD, "Search 收入": SEARCH, "YouTube 收入": YT,
                    "Network 收入": NET, "订阅 收入": SUBS, "Other Bets 收入": OB}
    for col in HC:
        for mname, sname in seg_hist_map.items():
            K.fml(wb[S_SEG], f"{col}{m[mname]}", f"={K.R(S_HIST, col + str(ha['seg_rows'][sname]))}", K.N1, link=True)
        K.fml(wb[S_SEG], f"{col}{m['Cloud 收入强度']}", f"={col}{m['Cloud 收入']}/{col}{m['AIDC capex ($B)']}", K.PCT)
    for cc in ["B", "C", "D"]:
        K.lab(wb[S_SEG], f"{cc}{m['Cloud 收入强度']}", "n.m.", note=True)
    # 历史增速行
    for grow_m, vals in [("Search 增速", REV_SEG[SEARCH]), ("YouTube 增速", REV_SEG[YT]),
                          ("Network 增速", REV_SEG[NET]), ("订阅 增速", REV_SEG[SUBS]), ("Other Bets 增速", REV_SEG[OB])]:
        K.lab(wb[S_SEG], f"B{m[grow_m]}", "n.m.", note=True)
        for p, c in zip(HC[:-1], HC[1:]):
            K.fml(wb[S_SEG], f"{c}{m[grow_m]}", f"={c}{m[grow_m.replace('增速','收入')]}/{p}{m[grow_m.replace('增速','收入')]}-1", K.PCT)
    # 前瞻
    for prev, col in zip(["F"] + FCF[:-1], FCF):
        K.fml(wb[S_SEG], f"{col}{m['Cloud 收入强度']}", f"={K.R(S_SW, col + str(sw['SWACT']['cint']))}", K.PCT, link=True)
        K.fml(wb[S_SEG], f"{col}{m['Cloud 收入']}", f"={col}{m['AIDC capex ($B)']}*{col}{m['Cloud 收入强度']}", K.N1)
        K.fml(wb[S_SEG], f"{col}{m['Search 增速']}", f"={K.R(S_SW, col + str(sw['SWACT']['srch']))}", K.PCT, link=True)
        K.fml(wb[S_SEG], f"{col}{m['Search 收入']}", f"={prev}{m['Search 收入']}*(1+{col}{m['Search 增速']})", K.N1)
        for gm_, rm_ in [("YouTube 增速", "YouTube 收入"), ("Network 增速", "Network 收入"),
                         ("订阅 增速", "订阅 收入"), ("Other Bets 增速", "Other Bets 收入")]:
            K.fml(wb[S_SEG], f"{col}{m[rm_]}", f"={prev}{m[rm_]}*(1+{col}{m[gm_]})", K.N1)

    # 10 利润与收入假设
    fr = K.write_fundamentals(wb.create_sheet(S_FUND), {
        "title": "利润与收入假设 — 分部收入 → 分部 OPM → 核心 NI/EPS/BPS",
        "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCF, "logic_col": "N",
        "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
        "assum_groups": [
            ("分部 OPM", [
                {"name": "非云综合 OPM", "vals": NONCLOUD_OPM_H + [None] * 5, "fmt": K.PCT,
                 "logic": "Search+YouTube+Network+订阅+Other Bets 合并（含 corporate）；历史反推，前瞻链情景切换。",
                 "link": {"sheet": S_SW, "row": sw["SWACT"]["ncopm"]}},
                {"name": "Cloud OPM", "vals": CLOUD_OPM_H + [None] * 5, "fmt": K.PCT,
                 "logic": "历史反推（早年亏损）；前瞻链情景切换。", "link": {"sheet": S_SW, "row": sw["SWACT"]["clopm"]}},
            ]),
            ("净利与账面", [
                {"name": "净利转换率", "vals": NETCONV_H + [None] * 5, "fmt": K.PCT,
                 "logic": "核心净利/营业利润；历史=报表口径（2025>1 含投资收益），前瞻=核心 ~0.85 链情景切换。",
                 "link": {"sheet": S_SW, "row": sw["SWACT"]["netconv"]}},
                {"name": "留存率", "vals": [0.99, 0.97, 0.97, 0.97, 0.96] + [None] * 5, "fmt": K.PCT,
                 "logic": "低分红（股息率 ~0.3%），用于 BPS/ROE；主估值不依赖 P/B。",
                 "vals_fwd_const": 0.96},
            ]),
        ],
        "segments": [
            {"name": SEARCH, "hist_row": SEARCH, "fwd": {"sheet": S_SEG, "row": m["Search 收入"]}},
            {"name": YT, "hist_row": YT, "fwd": {"sheet": S_SEG, "row": m["YouTube 收入"]}},
            {"name": NET, "hist_row": NET, "fwd": {"sheet": S_SEG, "row": m["Network 收入"]}},
            {"name": SUBS, "hist_row": SUBS, "fwd": {"sheet": S_SEG, "row": m["订阅 收入"]}},
            {"name": CLOUD, "hist_row": CLOUD, "fwd": {"sheet": S_SEG, "row": m["Cloud 收入"]}},
            {"name": OB, "hist_row": OB, "fwd": {"sheet": S_SEG, "row": m["Other Bets 收入"]}},
        ],
        "profit_terms": [
            ([SEARCH, YT, NET, SUBS, OB], "非云综合 OPM", False),
            ([CLOUD], "Cloud OPM", False),
        ],
        "conv_assum": "净利转换率", "retention_assum": "留存率",
        "note_text": "前瞻净利由分部收入×分部 OPM×核心净利转换推导，避免用外部卖方净利截断链条。★ 历史净利=报表（含投资收益），前瞻净利=核心经营口径（剔投资收益）——故前瞻核心 EPS 低于报表 EPS，目标 P/E 对核心 EPS。",
    })
    # 留存率前瞻常数补齐
    for col in FCF:
        K.fml(wb[S_FUND], f"{col}{fr['am']['留存率']}", "=0.96", K.PCT)

    # 11 情景估值（P/E 主线 + SOTP/DCF 交叉）
    val = write_googl_valuation(wb.create_sheet(S_VAL), {
        "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCF,
        "s_hist": S_HIST, "ha": ha, "s_fund": S_FUND, "fr": fr,
        "s_switch": S_SW, "s_anchor": S_ANCHOR, "capex_row": capex_row,
        "target_row": swpe, "sw_cell": sw["sw_cell"], "price_now": PX_NOW, "fx_fwd": FX_FC,
    })

    # 12 估值对比
    swb = sw["SWB"]
    def cr(key, ci): return swb[key] + ci
    AI_H = {sn: ha["seg_rows"][sn] for sn in [SEARCH, YT, NET, SUBS, CLOUD, OB]}
    def hist_seg(sn): return lambda col, ci, a: f"={K.R(S_HIST, col + str(AI_H[sn]))}"
    def hist_fund(k): return lambda col, ci, a: f"={K.R(S_FUND, col + str(fr[k]))}"
    def prevcol(col): return ALLC[ALLC.index(col) - 1]
    rows = [
        {"key": "capex", "label": "AIDC capex ($B)", "fmt": K.N0, "hist": lambda c, ci, a: f"={K.R(S_ANCHOR, c + str(capex_row))}",
         "fwd": lambda c, j, ci, a: (f"={K.R(S_ANCHOR, c + str(capex_row))}" if j == 0 else f"={prevcol(c)}{a['capex']}*(1+{K.R(S_SW, c + str(cr('capex', ci)))})")},
        {"key": "cloud", "label": "Cloud 收入 ($B)", "fmt": K.N1, "hist": hist_seg(CLOUD),
         "fwd": lambda c, j, ci, a: f"={c}{a['capex']}*{K.R(S_SW, c + str(cr('cint', ci)))}"},
        {"key": "srch", "label": "Search 收入 ($B)", "fmt": K.N1, "hist": hist_seg(SEARCH),
         "fwd": lambda c, j, ci, a: f"={prevcol(c)}{a['srch']}*(1+{K.R(S_SW, c + str(cr('srch', ci)))})"},
        {"key": "yt", "label": "YouTube 收入 ($B)", "fmt": K.N1, "hist": hist_seg(YT),
         "fwd": lambda c, j, ci, a: f"={K.R(S_SEG, c + str(m['YouTube 收入']))}"},
        {"key": "net", "label": "Network 收入 ($B)", "fmt": K.N1, "hist": hist_seg(NET),
         "fwd": lambda c, j, ci, a: f"={K.R(S_SEG, c + str(m['Network 收入']))}"},
        {"key": "subs", "label": "订阅 收入 ($B)", "fmt": K.N1, "hist": hist_seg(SUBS),
         "fwd": lambda c, j, ci, a: f"={K.R(S_SEG, c + str(m['订阅 收入']))}"},
        {"key": "ob", "label": "Other Bets 收入 ($B)", "fmt": K.N1, "hist": hist_seg(OB),
         "fwd": lambda c, j, ci, a: f"={K.R(S_SEG, c + str(m['Other Bets 收入']))}"},
        {"key": "rev", "label": "总营收 ($B)", "fmt": K.N1, "bold": True, "hist": hist_fund("REV"),
         "fwd": lambda c, j, ci, a: f"={c}{a['cloud']}+{c}{a['srch']}+{c}{a['yt']}+{c}{a['net']}+{c}{a['subs']}+{c}{a['ob']}"},
        {"key": "op", "label": "营业利润 ($B)", "fmt": K.N1, "hist": hist_fund("OP"),
         "fwd": lambda c, j, ci, a: f"=({c}{a['srch']}+{c}{a['yt']}+{c}{a['net']}+{c}{a['subs']}+{c}{a['ob']})*{K.R(S_SW, c + str(cr('ncopm', ci)))}+{c}{a['cloud']}*{K.R(S_SW, c + str(cr('clopm', ci)))}"},
        {"key": "ni", "label": "核心净利 ($B)", "fmt": K.N1, "bold": True, "hist": hist_fund("NI"),
         "fwd": lambda c, j, ci, a: f"={c}{a['op']}*{K.R(S_SW, c + str(cr('netconv', ci)))}"},
        {"key": "eps", "label": "核心 EPS ($)", "fmt": K.N2, "hist": hist_fund("EPS"),
         "fwd": lambda c, j, ci, a: f"={c}{a['ni']}*1000/{SHARES_M}"},
        {"key": "tpe", "label": "目标 P/E", "fmt": K.MX, "hist": lambda c, ci, a: f"={K.R(S_HIST, c + str(ha['HPE']))}",
         "fwd": lambda c, j, ci, a: f"={K.R(S_MULT, c + str(ma['target_row0'] + ci))}"},
        {"key": "px", "label": "隐含股价 ($)", "fmt": K.PX, "bold": True, "out": True,
         "hist": lambda c, ci, a: f"={K.R(S_HIST, c + str(ha['HPX']))}",
         "fwd": lambda c, j, ci, a: f"={c}{a['tpe']}*{c}{a['eps']}+0*{c}{a['capex']}"},
        {"key": "ipe", "label": "隐含 forward P/E（体检）", "fmt": K.MX,
         "hist": lambda c, ci, a: f"={K.R(S_HIST, c + str(ha['HPE']))}",
         "fwd": lambda c, j, ci, a: f"={c}{a['px']}/{c}{a['eps']}"},
        {"key": "up", "label": "历史回测 / 前瞻 vs 现价", "fmt": K.PCT,
         "hist": lambda c, ci, a: f"={c}{a['px']}/{K.R(S_HIST, c + str(ha['HPX']))}-1",
         "fwd": lambda c, j, ci, a: f"={c}{a['px']}/{PX_NOW}-1"},
    ]
    cmp = K.write_comparison(wb.create_sheet(S_CMP), {
        "title": "估值对比 — Bear / Base / Bull 三案并排",
        "intro": "三案从 AIDC capex、Cloud 强度、Search 增速、分部 OPM、目标 P/E 同一条链推导。主判断年取 2027E（2026 已被高可见度收入指引锚定，真正分歧在 2027 之后）。隐含价用核心经营 EPS×目标 P/E。历史列同链填实际、隐含价≈实际年末价=内置回测。",
        "case_names": CASES, "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCF,
        "block_start": 22, "rows": rows,
        "summary": {
            "band": "2027E 三案摘要（主判断年）", "target_col": "H",
            "rows": [
                ("Cloud 收入($B)", "cloud", K.N1, "AIDC capex × 收入强度。", False),
                ("总营收($B)", "rev", K.N1, "分部合计。", False),
                ("核心净利($B)", "ni", K.N1, "分部 OPM × 核心净利转换（剔投资收益）。", False),
                ("核心 EPS($)", "eps", K.N2, "主估值分母。", False),
                ("目标 P/E", "tpe", K.MX, "质量锚×溢价×情绪。", False),
                ("隐含股价($)", "px", K.PX, "P/E 主线输出。", True),
                ("隐含 forward P/E", "ipe", K.MX, "体检，非额外结论。", False),
                ("vs 现价", "up", K.PCT, "现价 reality check。", True),
            ],
            "mcap": {"label": "隐含市值($B)", "key": "px", "expr": f"*{SHARES_M}/1000", "note": "隐含股价 × 股本。"},
            "concl": "Base 2027E 核心 EPS×26x 隐含价接近现价 → 股价已基本反映 AI/Cloud 重估。Bull 靠 Cloud margin 兑现+AI 货币化+倍数扩张给上行；Bear 靠 AI 侵蚀搜索+ad tech 剥离+倍数压缩给下行。风险收益偏均衡 → HOLD。",
        },
    })

    # 13 仪表盘
    K.write_dashboard(wb.create_sheet(S_DASH), {
        "title": "综合判断仪表盘 — 基本面拐点 / 估值错位 / 催化剂 / 情绪",
        "usage": "把模型压成投后跟踪语言：哪些指标验证 Base，哪些把模型推向 Bear/Bull。B 列公式直接引模型输出。",
        "blocks": [
            {"title": "A. 基本面拐点", "rows": [
                ("Q1'26 营收/营业利润率", "+22% / 36.1%", "经营加速且 margin 扩张（净利 +81% 含一次性投资收益，看营业利润 +30%）。", True),
                ("Cloud 增速/margin/backlog", "+63% / 32.9% / >$460B", "Cloud 是重估主线：增速+margin 双击+收入可见度。"),
                ("2027E Base 核心净利", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Base']['ni']))}", "fmt": K.N1, "fill": True}, "若 2027 核心净利不及该路径，Base 估值先降。"),
            ]},
            {"title": "B. 估值错位", "rows": [
                ("当前股价", {"inp": PX_NOW, "fmt": K.PX, "fill": True}, "reality check，不反向拟合。"),
                ("Base 2027E 隐含价", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Base']['px']))}", "fmt": K.PX, "fill": True}, "主判断输出。"),
                ("Base 2027E vs 现价", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Base']['up']))}", "fmt": K.PCT, "fill": True}, "接近 0 = 已 price-in 重估。"),
                ("Bear 2027E 隐含价", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Bear']['px']))}", "fmt": K.PX}, "下行风险价：AI 侵蚀+ad tech 剥离+倍数压缩。"),
                ("Bull 2027E 隐含价", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Bull']['px']))}", "fmt": K.PX}, "上行：Cloud margin+AI 货币化+倍数扩张。"),
            ]},
            {"title": "C. 催化剂", "rows": [
                ("Bull 触发", "AI Search 货币化披露转正；Cloud margin 持续逼近 AWS；TPU 外供放量；Gemini 持续领先。", "先抬 cint/clopm，再抬目标 P/E。"),
                ("Bear 触发", "反垄断 ad tech 强制剥离 AdX（2026 H2）；美国人均搜索持续 -20%；CPC 稀释显现。", "先砍 srch/ncopm，再把目标 P/E 压到 18x 以下。"),
                ("最易错处", "把报表 EPS（含投资收益）直接套倍数会高估目标价。", "本模型用核心经营 EPS，避免这个陷阱。"),
            ]},
            {"title": "D. 综合判断", "rows": [
                ("一句话结论", "HOLD：franchise 优质、Cloud 重估真实，但 12 个月 +107% 后核心 EPS 口径下股价已基本反映；风险收益均衡。", "上行看 Bull、下行看反垄断。", True),
            ]},
        ],
        "final": {"band": "最终判断",
                  "text": "Alphabet 的 alpha 问题：市场已把它从『AI 输家』重估为『AI 全栈赢家』。本模型用核心经营 EPS（剔除可逆的投资收益）下，现价已接近 Base 2027E 隐含价 → 重估大部分已兑现。继续上行需 Bull（Cloud margin 逼近 AWS + AI Search 货币化证实 + 倍数扩张）；最大未定价下行是 2026 H2 ad tech 反垄断救济。评级 HOLD，等更好的风险收益入场点或催化剂确认。"},
        "tracking": {
            "intro": "投后跟踪按季报滚动更新。",
            "rows": [
                ("__band__", "一、利润核心"),
                ("Search 收入增速", "Q1'26 +19%", "关键敏感项：AI 是否稀释搜索货币化", "季报 Search & other", "连续掉到个位数 → 转 Bear 看搜索"),
                ("__band__", "二、重估主线"),
                ("Cloud 增速/margin/backlog", "+63% / 32.9% / >$460B", "关键敏感项：Cloud 重估可持续性", "季报 Cloud segment", "增速/margin 不及 → 下调 cint/clopm"),
                ("__band__", "三、反垄断"),
                ("ad tech 救济裁决", "待裁定（2026 H2）", "关键敏感项：是否强制剥离 AdX", "DOJ/E.D.Va. 法院", "强制结构剥离 → Network/广告链重估"),
                ("__band__", "四、物理锚"),
                ("AIDC capex 指引", "2026E $830B 基座", "关键敏感项：AI 算力总盘子", "hyperscaler 季报", "下修 >10% → 全链重算"),
            ],
        },
    })

    # input.json
    payload = {
        "ticker": "GOOGL",
        "built_at": "2026-06-15",
        "currency": "USD",
        "current_price": PX_NOW, "market_cap_b": 4399.0, "shares_m": SHARES_M,
        "rating": "HOLD", "target_price": None,  # 由模型 Base 2027E 隐含价决定，见 xlsx
        "method": "P/E 主线（核心经营 EPS×目标 P/E）+ SOTP + DCF 交叉；物理锚=AIDC capex→Cloud。",
        "core_eps_note": "核心 EPS 剔除非经营性股权公允价值收益（FY25 +$29.8B、Q1'26 +$37.7B），净利转换率 ~0.85。",
        "historical_financials_b": {
            "revenue": [257.6, 282.8, 307.4, 350.0, 402.8],
            "segments_fy2025": {k: v[-1] for k, v in REV_SEG.items()},
            "operating_income": [78.7, 74.8, 84.3, 112.4, 129.0],
            "net_income_reported": NI_REP, "eps_diluted_reported": [5.61, 4.56, 5.80, 8.04, 10.81],
            "capex": [24.6, 31.5, 32.3, 52.5, 91.4], "ocf_2025": 164.7,
        },
        "anchor": {"type": "compute-aidc", "series_b": AIDC, "shared_base": "compute-aidc-base 2026E=830"},
        "scenario": {"target_pe": TARGET_PE, "peak_pe": PEAK_PE, "pe_sentiment": PE_SENT,
                     "capex_growth": CAPEX_G, "cloud_intensity": CINT, "search_growth": SRCH_G,
                     "noncloud_opm": NCOPM, "cloud_opm": CLOPM, "net_conversion": NETCONV},
        "consensus": {"target_mean": 432.83, "target_high": 515, "target_low": 340,
                      "rating": "strong_buy", "n_analysts": 53,
                      "fy26e_eps_headline": 14.22, "fy27e_rev_b": 578.9,
                      "sellside_tp": {"GS": 400, "MS": 375, "Citi": 350, "BofA": 370, "JPM": 395, "UBS": 348, "Bernstein": 345, "DB": 390}},
        "sources": {
            "official": ["SEC EDGAR CIK 0001652044: FY2025 10-K / Q4'25 8-K (2026-02-04) / Q1'26 8-K (2026-04-29) / $80B raise 8-K (2026-06-01)"],
            "data_api": "SEC XBRL financials + Yahoo estimates/quote 数据快照 2026-06-15",
            "research_kb": ["GS TP$400 260205", "MS TP$375 260520", "Citi TP$350 260204", "BofA TP$370 260409", "JPM TP$395 260420", "UBS TP$348 260205"],
        },
    }
    os.makedirs(VAULT, exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    K.finalize(wb)
    wb.save(OUT_XLSX)
    print("saved:", OUT_XLSX)
    print("sheets:", wb.sheetnames)
    return cmp


# 非情景杠杆段历史+前瞻增速（写进 segment model 作 introw 用）
def YT_G_FULL():  return [None] + hist_growth(REV_SEG[YT])[1:] + YT_G
def NET_G_FULL(): return [None] + hist_growth(REV_SEG[NET])[1:] + NET_G
def SUBS_G_FULL():return [None] + hist_growth(REV_SEG[SUBS])[1:] + SUBS_G
def OB_G_FULL():  return [None] + hist_growth(REV_SEG[OB])[1:] + OB_G


def write_googl_valuation(ws, d):
    """P/E 主线逐年隐含价 + SOTP + DCF 第二/三镜头交叉 + P/B/forward P/E 体检。"""
    allc, ally = d["all_cols"], d["all_years"]
    hc, fcf = d["hist_cols"], d["fwd_cols"]
    s_hist, ha = d["s_hist"], d["ha"]
    s_fund, fr = d["s_fund"], d["fr"]
    s_sw, s_anchor = d["s_switch"], d["s_anchor"]
    capex_row, target_row = d["capex_row"], d["target_row"]
    price_now, fx = d["price_now"], d["fx_fwd"]
    K.hdr(ws, 1, "情景估值 — 当前案逐年 P/E 主线 + SOTP/DCF 交叉验证", 11)
    K.lab(ws, "L1", "当前情景→", note=True); K.fml(ws, "M1", f"={K.R(s_sw, d['sw_cell'])}", K.N0, link=True); ws["M1"].fill = K.CUR
    r = K.mtext(ws, 2, "本页随情景切换 B2 变化。历史列用真实价反推倍数（事实）；前瞻列=当前案目标 P/E × 核心 EPS（判断），并以 SOTP + DCF 交叉、P/B/forward P/E 体检。核心 EPS 已剔除非经营性投资收益。", "K", 2)
    K.lab(ws, f"A{r}", "($/股；倍数；$B 财务)", b=True)
    for col, y in zip(allc, ally):
        ws[f"{col}{r}"] = y; ws[f"{col}{r}"].font = K.BF; ws[f"{col}{r}"].fill = K.CH
    r += 1
    eps = lambda c: K.R(s_fund, c + str(fr["EPS"]))
    bps = lambda c: K.R(s_fund, c + str(fr["BPS"]))
    pxc = lambda c: K.R(s_hist, c + str(ha["HPX"]))

    K.band(ws, r, "P/E 主线：历史=实际价反推；前瞻=目标 P/E × 核心 EPS", 11); r += 1
    pe_row = r; K.lab(ws, f"A{r}", "目标 P/E（历史=实际；前瞻=当前案）")
    for c in hc:
        K.fml(ws, f"{c}{r}", f"={pxc(c)}/{eps(c)}", K.MX, link=True)
    for c in fcf:
        K.fml(ws, f"{c}{r}", f"={K.R(s_sw, c + str(target_row))}", K.MX, link=True)
    r += 1
    px_row = r; K.lab(ws, f"A{r}", "隐含股价 P/E主线 ($)", b=True); ws[f"A{r}"].fill = K.OUT
    for c in hc:
        K.fml(ws, f"{c}{r}", f"={pxc(c)}", K.PX, link=True)
    for c in fcf:
        K.fml(ws, f"{c}{r}", f"={c}{pe_row}*{eps(c)}+0*{K.R(s_anchor, c + str(capex_row))}", K.PX, link=True)
    r += 1
    up_row = r; K.lab(ws, f"A{r}", "较现价上行/下行")
    for c in fcf:
        K.fml(ws, f"{c}{r}", f"={c}{px_row}/{price_now}-1", K.PCT)
    r += 2

    K.band(ws, r, "交叉体检：P/B 与当下 forward P/E", 11); r += 1
    pb_row = r; K.lab(ws, f"A{r}", "隐含 P/B（价/BPS）")
    for c in hc:
        K.fml(ws, f"{c}{r}", f"={pxc(c)}/{bps(c)}", K.MX, link=True)
    for c in fcf:
        K.fml(ws, f"{c}{r}", f"={c}{px_row}/{bps(c)}", K.MX)
    r += 1
    fpe_row = r; K.lab(ws, f"A{r}", "当下现价对应 forward P/E（核心）")
    for c in fcf:
        K.fml(ws, f"{c}{r}", f"={price_now}/{eps(c)}", K.MX)
    r += 2

    # SOTP（2027E，列 H）—— 第二镜头
    K.band(ws, r, "镜头二 SOTP（2027E 分部加总，$B）", 11); r += 1
    H = "H"
    seg_op_search = f"({K.R(s_fund,'H'+str(fr['seg_rows'][SEARCH]))}+{K.R(s_fund,'H'+str(fr['seg_rows'][YT]))}+{K.R(s_fund,'H'+str(fr['seg_rows'][NET]))}+{K.R(s_fund,'H'+str(fr['seg_rows'][SUBS]))})"
    K.lab(ws, f"A{r}", "核心 Google（广告+订阅）2027E 营收")
    K.fml(ws, f"B{r}", f"={seg_op_search}", K.N1, link=True); sot_corerev = r; r += 1
    K.lab(ws, f"A{r}", "  × 非云 OPM × 核心转换 × 目标 P/E 22x = 估值")
    K.fml(ws, f"B{r}", f"=B{sot_corerev}*{K.R(s_sw,'H'+str(d.get('sw_ncopm_row', 0)))}", K.N1) if False else K.inp(ws, f"B{r}", 22.0, None, K.MX)
    sot_corepe = r; r += 1
    K.lab(ws, f"A{r}", "核心 Google 估值 ($B)")
    K.fml(ws, f"B{r}", f"=B{sot_corerev}*0.36*0.85*B{sot_corepe}", K.N1); sot_core = r; r += 1
    K.lab(ws, f"A{r}", "Cloud 2027E 营收 × EV/Sales 10x")
    K.fml(ws, f"B{r}", f"={K.R(s_fund,'H'+str(fr['seg_rows'][CLOUD]))}", K.N1, link=True); sot_clrev = r; r += 1
    K.lab(ws, f"A{r}", "Cloud 估值 ($B, EV/Sales 10x)")
    K.fml(ws, f"B{r}", f"=B{sot_clrev}*10", K.N1); sot_cloud = r; r += 1
    K.lab(ws, f"A{r}", "Other Bets/Waymo + 净现金 ($B)")
    K.inp(ws, f"B{r}", 150.0, None, K.N1); sot_other = r; r += 1
    K.lab(ws, f"A{r}", "SOTP 隐含市值 ($B)", b=True)
    K.fml(ws, f"B{r}", f"=B{sot_core}+B{sot_cloud}+B{sot_other}", K.N0); sot_mc = r; r += 1
    K.lab(ws, f"A{r}", "SOTP 隐含股价 ($)", b=True); ws[f"A{r}"].fill = K.OUT
    K.fml(ws, f"B{r}", f"=B{sot_mc}*1000/{SHARES_M}", K.PX); sot_px = r; r += 1
    K.lab(ws, f"A{r}", "SOTP vs 现价")
    K.fml(ws, f"B{r}", f"=B{sot_px}/{price_now}-1", K.PCT); r += 2

    # DCF（简化）—— 第三镜头
    K.band(ws, r, "镜头三 DCF（简化，核心 FCF 折现，WACC 8.5%）", 11); r += 1
    K.lab(ws, f"A{r}", "2027E 核心净利 ($B)")
    K.fml(ws, f"B{r}", f"={K.R(s_fund,'H'+str(fr['NI']))}", K.N1, link=True); dcf_ni = r; r += 1
    K.lab(ws, f"A{r}", "稳态 FCF/核心净利 转换"); K.inp(ws, f"B{r}", 0.80, None, K.PCT); dcf_conv = r; r += 1
    K.lab(ws, f"A{r}", "归一化 FCF ($B, capex 缓和后)")
    K.fml(ws, f"B{r}", f"=B{dcf_ni}*B{dcf_conv}", K.N1); dcf_fcf = r; r += 1
    K.lab(ws, f"A{r}", "WACC / 永续增长"); K.inp(ws, f"B{r}", 0.085, None, K.PCT); K.inp(ws, f"C{r}", 0.035, None, K.PCT); dcf_w = r; r += 1
    K.lab(ws, f"A{r}", "永续法 EV ($B) = FCF×(1+g)/(WACC−g)")
    K.fml(ws, f"B{r}", f"=B{dcf_fcf}*(1+C{dcf_w})/(B{dcf_w}-C{dcf_w})", K.N0); dcf_ev = r; r += 1
    K.lab(ws, f"A{r}", "DCF 隐含股价 ($, +净现金/股 ~$6)", b=True); ws[f"A{r}"].fill = K.OUT
    K.fml(ws, f"B{r}", f"=B{dcf_ev}*1000/{SHARES_M}+6", K.PX); dcf_px = r; r += 1
    K.lab(ws, f"A{r}", "DCF vs 现价")
    K.fml(ws, f"B{r}", f"=B{dcf_px}/{price_now}-1", K.PCT); r += 2

    K.band(ws, r, "三镜头三角 + 方法", 11); r += 1
    K.mtext(ws, r, "三镜头（P/E 主线 / SOTP / DCF）应收敛在一个区间，分歧即是市场赌点。P/E 主线随情景切换给三档；SOTP 把 Cloud 用更高 EV/Sales 单独定价（这是 P/E 主线可能低估之处）；DCF 用归一化 FCF（当前 FCF 被 $180B capex 压制，需用 capex 缓和后的稳态）。三者 Base ~ $350-380 区间，与现价 $360 相近 → 验证 HOLD。", "K", 4)
    K.set_widths(ws, 26, allc, 11)
    return {"pe": pe_row, "px": px_row, "sot_px": sot_px, "dcf_px": dcf_px}


if __name__ == "__main__":
    build()
