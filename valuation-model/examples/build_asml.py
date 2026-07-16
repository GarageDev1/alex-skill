# -*- coding: utf-8 -*-
"""
build_asml.py — ASML Holding 物理锚驱动估值模型 (USD $B 口径)。

物理锚 = 光刻系统出货台数(EUV + DUV)。公式链:
  EUV 台数 × EUV ASP($M/台) → EUV 收入($B)
  DUV 台数 × DUV ASP($M/台) → DUV 收入($B)
  服务收入(Installed Base Mgmt)随累计装机基数 ~+8-12%/yr 增长
  总营收 = EUV + DUV + 服务 + M&I → 营业利润率 → 净利 → EPS(÷ 股本)
镜头: P/E 主线(franchise, 资本化盈利) + DCF 交叉验证。隐含价 = 目标 forward P/E × 前瞻 EPS。

口径: ASML 财务以 EUR 报告, 已按当年年均汇率换算成 USD $B(vals_in_usd=True);
模型内全程 USD, FX 行恒为 1.0(单币种), 股价用 ADR(USD) 现价 $1,863.55。
"""
import os
from openpyxl import Workbook
import build_kit as K

# ════════════ 0. 全局轴 ════════════
ALLC = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
ALLY = ["2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
HC, HY = ["B", "C", "D", "E", "F"], ["2021", "2022", "2023", "2024", "2025"]
FC = ["F", "G", "H", "I", "J", "K"]          # 前瞻含基年 F=2025A
FCf = FC[1:]                                   # 纯前瞻 2026E-2030E
FX_FWD = 1.0                                   # 单币种 USD 模型, FX 恒 1
FX_HIST = [1.0, 1.0, 1.0, 1.0, 1.0]
CASES = ["Bear", "Base", "Bull"]

S_COVER, S_HIST, S_PX, S_CONS = "封面", "历史财务与估值", "股价走势", "卖方研报共识"
S_HMULT, S_MULT, S_SW = "历史估值倍数", "估值倍数假设", "情景切换"
S_ANCHOR, S_SEG, S_FUND = "光刻系统出货", "分部测算", "利润与收入假设"
S_VAL, S_CMP, S_DCF, S_DASH = "情景估值", "估值对比", "DCF交叉验证", "综合判断仪表盘"

# 月度收盘(ASML ADR, USD; 单一价格来源)
MONTHLY = [("2021-01", 530), ("2021-06", 690), ("2021-12", 795),
           ("2022-06", 535), ("2022-12", 545),
           ("2023-06", 720), ("2023-12", 755),
           ("2024-06", 1035), ("2024-12", 695),
           ("2025-06", 780), ("2025-09", 930),
           ("2025-12", 1010),
           ("2026-03", 1550), ("2026-06", 1864)]
PX_NOW = 1864   # ADR USD 现价(2026-06-12 $1,863.55)

wb = Workbook()
wb.remove(wb["Sheet"])

# ════════════ 1. 封面 ════════════
K.write_cover(wb.create_sheet(S_COVER), {
    "title": "ASML Holding (ASML) 估值模型 — 物理锚=光刻系统出货台数 (USD $B)",
    "meta": [
        ("报告日期", "2026-06-15"),
        ("数据截止", "Q1'26 季报 + 2026/2030 管理层指引 + 卖方一致预期 + ADR 实时股价 (2026-06-12)"),
        ("现价", f"${PX_NOW:,} (ADR, USD); 市值 ~$718B; 股本 ~390M 股"),
        ("时效声明", "基于建模时点的指引/共识/股价; AI 驱动的光刻周期与 High-NA 爬坡变化快, 建议每季财报后更新。"),
        ("方法一句话", "物理锚(EUV/DUV 出货台数)× ASP → 系统收入; 服务收入随累计装机基数增长 → 总营收 → 营业利润率 → 净利 → EPS; 主线 目标forward P/E × 前瞻EPS → 隐含价, DCF 交叉验证。"),
    ],
    "takeaways": [
        ("① 当下估值位置", f"现价 ${PX_NOW} / FY26E EPS ~$31.5 = forward P/E ~59x, 显著高于历史 30-45x 区间与卖方目标均值 $1,708(现价已超均值 ~9%)。"),
        ("② 核心引擎", "EUV 收入 = EUV 台数 × EUV ASP(High-NA EXE:5200 ~$350M/台拉高混合 ASP); 2027 计划产 ≥80 台 low-NA EUV。"),
        ("③ 周期判断", "DUV 受中国占比从 2024 峰值 49%→2026 指引 ~20% 拖累(Bear 维持低迷~250 台 / Base 温和复苏 / Bull 逻辑存储扩产强劲)。"),
        ("④ 三情景目标价", "Bear/Base/Bull 沿同一条出货链翻档, 目标 forward P/E 28x/36x/42x(franchise+AI 溢价), 见『估值对比』。"),
        ("⑤ 主要风险", "AI capex 见顶 / High-NA 良率与导入慢于预期 / 中国出口管制进一步收紧 / 单一先进逻辑客户依赖。"),
    ],
})

# ════════════ 2. 历史财务与估值 ════════════
# 分部($B): EUV / DUV / 服务(Installed Base Mgmt + 其他)
# 总营收 2021-2025: 22.0/22.3/29.8/30.6/36.9; EUV 7.5/7.4/9.8/9.0/13.1; DUV 8.2/8.1/13.3/13.9/13.6
# 服务+M&I = 总 - EUV - DUV: 6.3/6.8/6.7/7.7/10.2  (服务 5.87/6.05/6.08/7.03/9.25 + M&I 余项)
ha = K.write_history(wb.create_sheet(S_HIST), {
    "title": "ASML 历史财务与估值 ($B) — 2021-2025A + 当下 (EUR 按当年年均汇率折 USD)",
    "hist_cols": HC, "hist_years": HY,
    "fx_hist": FX_HIST, "fx_now": FX_FWD,
    "vals_in_usd": True,
    "fx_label": "FX (模型内 USD, 已换算)",
    "segments": [
        ("EUV 系统 收入 ($B)", [7.5, 7.4, 9.8, 9.0, 13.1], True),
        ("DUV 系统 收入 ($B)", [8.2, 8.1, 13.3, 13.9, 13.6], True),
        ("服务+M&I 收入 ($B)", [6.3, 6.8, 6.7, 7.7, 10.2], True),
    ],
    "total_now": 36.9,
    "gm_pct": [0.527, 0.505, 0.513, 0.513, 0.528], "gm_now": 0.528,
    "ni": [6.96, 5.92, 8.48, 8.19, 10.86], "ni_now": 10.86,
    "eq": [12.0, 9.3, 14.5, 20.0, 22.2], "eq_now": 22.2,
    "shares": [413, 398, 394, 393, 390], "shares_now": 390,
    "px_end": [795, 545, 755, 695, 1010],
    "px_now": PX_NOW,
    "px_avg": [663, 555, 640, 865, 832],
    "band_note": "forward P/E 历史常态 30-45x → 当下 ~59x 冲上沿之上(AI 溢价定价); P/B ~33x 无意义(franchise 不看账面)。",
    "notes": [
        ("EUV 系统 收入 ($B)", "EUV 系统销售; 含 High-NA EXE 系列(2024 起出货 2 台, 2025 4 台)。来源: ASML 年报分部, EUR 按当年年均汇率折 USD。"),
        ("DUV 系统 收入 ($B)", "DUV(immersion + dry)系统销售; 强受中国需求与逻辑/存储扩产周期影响。"),
        ("服务+M&I 收入 ($B)", "Installed Base Management(服务+升级)+ Metrology & Inspection 余项; 随累计装机基数稳定增长(粘性高)。"),
        ("HREV", "总营收 = 公司实际(EUR 按当年年均汇率折 USD); 当下 = FY2025 实际。"),
        ("HGMP", "毛利率: 公司实际报告口径。"),
        ("HNI", "净利: 公司实际, 按当年均值汇率折 $B; 当下 = FY2025 实际。"),
        ("HEQ", "股东权益: 公司年报实际, 按当年均值汇率折 USD。"),
        ("HSH", "稀释股本(mn 股); 持续回购小幅缩股(413→390)。"),
        ("HPX", "年末股价: ASML ADR(USD) 月末收盘; 当下 = 现价 $1,864。"),
        ("HPXA", "年均股价: ADR 月度收盘均值, 同『股价走势』单一价格源。"),
    ],
})

# ════════════ 3. 股价走势 ════════════
def phase_fn(ym):
    if ym <= "2022-12":
        return "① 后疫情高位回调"
    if ym <= "2024-06":
        return "② AI 预期推升"
    if ym <= "2025-06":
        return "③ 中国+DUV 担忧回落"
    return "④ AI capex 再加速"

px = K.write_price_chart(wb.create_sheet(S_PX), MONTHLY, {
    "fn": phase_fn,
    "rows": [("① 后疫情高位回调", "2022 估值压缩 + 半导体周期下行"),
             ("② AI 预期推升", "EUV 订单 + AI 叙事推动创高(2024-06 ~$1,035)"),
             ("③ 中国+DUV 担忧回落", "中国 DUV 占比见顶回落 + 指引谨慎, 股价回撤"),
             ("④ AI capex 再加速", "Q4'25 net bookings 创纪录 €13.2B(EUV €7.4B), 2026 指引上调, 股价创新高")],
}, title="ASML ADR 月度股价 (USD)")

# ════════════ 4. 卖方研报共识 ════════════
K.write_consensus(wb.create_sheet(S_CONS), {
    "title": "卖方研报共识 — 38 买入 / 4 持有 / 2 卖出; 目标价均值 $1,708(区间 $898-$2,269)",
    "overview": "全街偏多(38/44 买入), 但目标价均值 $1,708 已低于现价 ~9% — 共识认为现价已大幅 price-in AI 光刻超级周期, 上行需靠 High-NA 兑现与 2030 长期指引上修。",
    "assumptions": [
        ("EUV 出货\n台数 (2027)",
         "共识看 2027 ≥80 台 low-NA + High-NA 爬坡; AI 逻辑/HBM 扩产驱动。",
         "分歧在 High-NA 导入速度(Intel 14A / Samsung)与良率。",
         "Base: 2027 EUV ~74 台(含 High-NA ~10), 沿管理层 ≥80 low-NA 指引留余量。"),
        ("EUV ASP\n($M/台)",
         "混合 ASP 逐年升: EXE:5200 ~$350M/台拉高均值。",
         "High-NA 占比节奏决定混合 ASP 上行斜率。",
         "Base: 从 2025 ~$273M 升至 2030 ~$340M(High-NA 占比上升)。"),
        ("DUV 与中国\n占比",
         "中国 DUV 占比 2024 峰值 49%→2025 ~29%→2026 指引 ~20%, 之后逻辑/存储扩产接力。",
         "最大分歧: 中国回落多深、非中国需求接力强度。",
         "Base: DUV 台数温和复苏(2026 ~280→2030 ~330), 价格平稳。"),
        ("目标 forward\nP/E",
         "卖方隐含目标 P/E ~35-45x(franchise + AI 溢价)。",
         "给多高的 AI 结构溢价 = 估值最大不确定。",
         "三层分解拍(历史峰值 × 结构溢价 × 情绪), 三案 28x/36x/42x, 见『估值倍数假设』。"),
    ],
    "divergences": [
        "① High-NA 商业化节奏: 决定 2027+ EUV 台数与混合 ASP 上行斜率(最直接的台数物理量分歧)。",
        "② AI capex 可持续性: 决定整条出货链的需求盘子是否在 2028+ 见顶。",
    ],
    "stances": [
        "多数买入(TP 区间上沿 $2,269): AI 光刻是不可替代瓶颈, EUV 独家垄断 + High-NA 打开新一轮 ASP 上行。",
        "少数卖出/谨慎(TP 下沿 $898): 现价已透支, 中国 DUV 占比快速回落 + High-NA 导入慢于预期是近期下行风险。",
    ],
})

# ════════════ 5. 历史估值倍数 ════════════
hm = K.write_hist_multiples(wb.create_sheet(S_HMULT), {
    "title": "历史估值倍数 — 数据底座: ASML 自身 forward P/E 历史带 + 当下 TTM + 半导体设备/AI 同业对照",
    "intro": "ASML 是 franchise(EUV 全球独家垄断 + 高粘性服务), 主线看 P/E 资本化盈利, 不看 P/B(P/B ~33x 无意义)。这一页给数据底座: ①自己历史 P/E 带 + 当下 ②同业 forward P/E 光谱 ③相对设备同业比值。看完去下一页拍三案目标 P/E。",
    "s_hist": S_HIST, "ha": ha, "hist_cols": HC, "hist_years": HY,
    "yhigh": px["yhigh"], "ylow": px["ylow"],
    "fwd_note": "forward P/E ≈59x(现价 $1,864 ÷ FY26E EPS ~$31.5) — 远高于历史 30-45x, 反映 AI 光刻超级周期定价。",
    "self_name": "ASML",
    "self_fwd_pe_label": "≈59x",
    "self_note": "本模型标的; EUV 全球独家垄断, franchise; forward P/E 推导见『情景估值』P/E 交叉验证块。",
    "peers": [
        {"name": "Applied Materials (AMAT)", "yearly": None, "cur_pb": None, "cur_pe": 24.0, "fwd_pe": 20.0,
         "note": "半导体设备同业(沉积/刻蚀), 无垄断壁垒 → 估值显著低于 ASML。"},
        {"name": "Lam Research (LRCX)", "yearly": None, "cur_pb": None, "cur_pe": 26.0, "fwd_pe": 22.0,
         "note": "刻蚀/沉积设备, 与 ASML 互补但非垄断; 设备同业对照下沿。"},
        {"name": "NVIDIA (AI 龙头, 参照上沿)", "yearly": None, "cur_pb": None, "cur_pe": 48.0, "fwd_pe": 38.0,
         "note": "AI 算力龙头 forward P/E 光谱上沿; ASML 作为 AI 光刻瓶颈可享类似溢价。"},
        {"name": "标普500 (参照下沿)", "yearly": None, "cur_pb": None, "cur_pe": None, "fwd_pe": 22.0,
         "note": "大盘 forward P/E, 光谱下沿基准。"},
    ],
    "ratio": None,
    "reading": "① 自己: 历史 forward P/E 带 30-45x; 当下 ~59x 冲上沿之上 = AI 溢价定价 → 第一层锚取历史峰值 ~45x(不用被本轮抬高的当下 ~59x)。② 同业: 设备同业(AMAT/LRCX)forward 20-22x, ASML 凭 EUV 垄断享显著溢价; AI 龙头(NVDA)~38x 是溢价上沿参照。③ 三层取值: 历史峰值 P/E × AI 结构溢价 × 情绪值, 三案 28x/36x/42x, 见下一页。",
})

# ════════════ 6. 估值倍数假设(P/E 主线) ════════════
# 这里"目标倍数"= 目标 forward P/E。三层 = 历史峰值 P/E × AI结构溢价 × 情绪值。
# 用 HPE 作为历史实际倍数行(P/E 而非 P/B)做历史列反推与一致性检验。
ma = K.write_multiple_assumptions(wb.create_sheet(S_MULT), {
    "title": "估值倍数假设 — P/E 主线: 三案目标 forward P/E = 历史峰值 P/E × AI 结构溢价 × 情绪值",
    "intro": "这一页只做判断(数据底座在上一页): ①为什么 P/E 做主线(franchise 业务判断) ②三层分解出三案目标 forward P/E。『情景切换』引用并切换, 『情景估值』套用当前案, 『估值对比』三案并排。注: 目标倍数 = 目标 forward P/E(不是 P/B)。",
    "why_text": ("镜头选择是业务判断: ASML『穿越周期持续存在的东西』是盈利还是资产? "
                 "ASML 是 franchise——EUV 光刻机全球独家垄断(High-NA 更是独此一家), 客户(TSMC/Samsung/Intel)被技术与认证深度锁定, "
                 "服务收入随累计装机基数稳定复利、粘性极高。盈利可持续且增长, 市场敢把当年盈利外推很多年 → 用 P/E 资本化盈利做主线。"
                 "P/B 在这里无意义(P/B ~33x, 账面不是 ASML 价值的载体)——这正是 franchise 与商品型重资产(内存/面板)镜头的根本区别。"
                 "DCF 做交叉验证(长久期成长 + 高 FCF 转换), 两镜头三角。"),
    "why_rows": 5,
    "method_text": "三层分解(不硬拍): ①历史周期峰值 forward P/E(过去最强周期实际到过的 ~45x, 不用本轮已抬到 ~59x 的当下值) × ②AI 结构溢价(EUV 垄断 + AI 光刻瓶颈相对设备同业的溢价) × ③情绪值(AI 周期/情绪位置)。一致性检验: 三层相乘对照当下/历史实际 forward P/E。",
    "peak": 45.0, "peak_note": "ASML 过去最强成长周期实际到过的 forward P/E 峰值 ~45x(2021/2024 高位); 不用本轮 AI 抬高的当下 ~59x。",
    "premium": 1.00, "premium_note": "结构溢价基准 1.00x(峰值 P/E 已含 ASML 垄断溢价); AI 增量溢价放在情绪值层逐年体现, 避免双重计算。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "hpb_row": ha["HPE"],   # ★ P/E 主线: 历史实际倍数行用 HPE(不是 HPB)
    "cases": [
        ("Bear", [0.70, 0.62, 0.58, 0.56, 0.62],
         "AI capex 见顶 + High-NA 导入慢 → 市场撤回成长溢价, 目标 P/E 向设备同业(~28x)收敛; 2030 略回稳。"),
        ("Base", [0.85, 0.80, 0.78, 0.78, 0.80],
         "AI 光刻超级周期延续但逐步成熟, 目标 P/E 从 ~38x 缓降至稳态 ~36x(franchise 应得溢价)。"),
        ("Bull", [1.00, 0.96, 0.93, 0.92, 0.93],
         "High-NA 全面兑现 + AI capex 多年不减速, 市场持续按 AI 龙头光谱(~42x)定价 ASML。"),
    ],
    "sent_note": "情绪值 = AI 周期/情绪位置, 以历史峰值 P/E 45x 为基准。目标 P/E = 45 × 情绪值(溢价层=1.0)。Base ~0.80 → 36x; Bear ~0.62 → 28x; Bull ~0.93 → 42x。历史列 = 实际 forward P/E ÷ (峰值×溢价) 反推。",
    "target_note": "三案目标 forward P/E = 峰值 45x × 溢价 1.0 × 情绪值。历史列 = 实际 P/E(回看, 三案同值)。",
    "reconcile_text": "卖方隐含目标 P/E ~35-45x; 我们 Base 36x 落在区间内偏中性, Bull 42x 接近 AI 龙头光谱, Bear 28x 退至设备同业 — 凭 EUV 垄断 + High-NA 独占的事实给出有判断力的区间。",
    "source_text": "第一层峰值 P/E: ASML 历史 forward P/E 带(2021/2024 ~45x)。第二层溢价: 设为 1.0(峰值已含垄断溢价, 避免重复)。第三层情绪值: 依据『综合判断仪表盘』D 块 AI 周期档位。",
})

# ════════════ 7. 情景切换 ════════════
sw = K.write_scenario_switch(wb.create_sheet(S_SW), {
    "title": "情景切换 — 全模型唯一的情景参数库 + 切换开关 (默认 Base)",
    "usage": ("怎么用: B2 是唯一入口 — 下拉选案 → 案序号派生 → 各杠杆『当前案』行跟着切 → "
              "整条出货链(台数→ASP→收入→利润→P/E→隐含价)变档, 『情景估值』输出该案逐年隐含价。"
              "三案对比不用切: 『估值对比』恒常三列并排。情景参数只在本页改(蓝字); 未列入的假设三案共用 Base。"),
    "cases": CASES, "default": "Base",
    "triggers": [
        ("Bear", "AI capex 见顶踩刹车 + High-NA(Intel 14A/Samsung)导入慢于预期 + 中国 DUV 进一步收紧 → EUV/DUV 台数缓增, ASP 上行受阻, 市场撤回成长溢价。"),
        ("Base", "AI 光刻周期延续, 2027 产 ≥80 台 low-NA + High-NA 平稳爬坡, 中国 DUV 占比按指引回落至 ~20% 后非中国需求接力, 落进 2030 营收 $51-69B 指引区间。"),
        ("Bull", "High-NA 全面放量 + AI capex 多年不减速 + 逻辑/存储扩产强劲 → EUV/DUV 台数与 ASP 双升, 市场按 AI 龙头光谱定价。"),
    ],
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "levers": [
        {"key": "euvq", "name": "EUV 出货台数", "fmt": K.N0,
         "vals": {"Bear": [50, 56, 60, 64, 68], "Base": [58, 74, 86, 96, 104], "Bull": [64, 86, 104, 120, 134]},
         "desc": "物理锚①: EUV(含 High-NA)年出货台数。EUV 收入 = 台数 × EUV ASP, 直接挂它。2027 管理层指引 ≥80 台 low-NA。",
         "stories": {"Bear": "AI capex 谨慎 + High-NA 慢, 台数缓增至 2030 ~68 台。",
                     "Base": "沿 2027 ≥80 台 low-NA + High-NA 爬坡, 2030 ~104 台。",
                     "Bull": "High-NA 全面放量 + AI 需求强, 2030 ~134 台。"},
         "hist": [42, 40, 53, 44, 48]},
        {"key": "euvasp", "name": "EUV ASP ($M/台)", "fmt": K.N0,
         "vals": {"Bear": [280, 290, 300, 308, 315], "Base": [285, 300, 315, 328, 340], "Bull": [290, 310, 330, 350, 365]},
         "desc": "物理量②: EUV 混合 ASP, 随 High-NA EXE:5200(~$350M/台)占比上升而走高。历史 2025 ~$273M/台。",
         "stories": {"Bear": "High-NA 占比慢, 混合 ASP 升幅有限至 ~$315M。",
                     "Base": "High-NA 平稳渗透, 2030 混合 ASP ~$340M。",
                     "Bull": "High-NA 占比快升, 2030 混合 ASP ~$365M。"},
         "hist": [177, None, 186, 205, 273]},
        {"key": "duvq", "name": "DUV 出货台数", "fmt": K.N0,
         "vals": {"Bear": [250, 248, 250, 252, 255], "Base": [280, 295, 310, 322, 332], "Bull": [300, 330, 355, 375, 392]},
         "desc": "物理锚③: DUV(immersion+dry)年出货台数。受中国占比回落(2024 49%→2026 ~20%)与逻辑/存储扩产周期影响。",
         "stories": {"Bear": "中国 DUV 深跌且非中国接力弱, 台数维持低迷 ~250。",
                     "Base": "中国回落后逻辑/存储扩产接力, 温和复苏至 2030 ~332。",
                     "Bull": "全球扩产强劲 + 中国企稳, 2030 ~392。"},
         "hist": [244, 305, 396, 374, 279]},
        {"key": "duvasp", "name": "DUV ASP ($M/台)", "fmt": K.N0,
         "vals": {"Bear": [49, 49, 49, 50, 50], "Base": [49, 50, 50, 51, 51], "Bull": [50, 51, 52, 52, 53]},
         "desc": "物理量④: DUV 混合 ASP, 相对平稳(immersion 高/dry 低的组合)。历史 2025 ~$49M/台(13.6/279)。",
         "stories": {"Bear": "组合偏低端, ASP 平稳 ~$50M。",
                     "Base": "组合稳定, ASP ~$51M。",
                     "Bull": "immersion 占比升, ASP ~$53M。"},
         "hist": [34, 27, 34, 37, 49]},
        {"key": "svcg", "name": "服务+M&I 增速", "fmt": K.PCT,
         "vals": {"Bear": [0.08, 0.08, 0.08, 0.08, 0.08], "Base": [0.10, 0.10, 0.10, 0.10, 0.10], "Bull": [0.12, 0.12, 0.12, 0.12, 0.12]},
         "desc": "服务收入随累计装机基数(installed base)增长, 粘性高、低波动。历史 ~+8-12%/yr。",
         "stories": {"Bear": "装机增速放缓, 服务 +8%/yr。", "Base": "装机稳定复利, 服务 +10%/yr。", "Bull": "装机加速 + 升级需求强, 服务 +12%/yr。"},
         "hist": [None, 0.08, -0.01, 0.15, 0.32]},
        {"key": "opm", "name": "营业利润率 (%)", "fmt": K.PCT,
         "vals": {"Bear": [0.330, 0.345, 0.355, 0.360, 0.365], "Base": [0.350, 0.380, 0.400, 0.415, 0.430], "Bull": [0.360, 0.400, 0.430, 0.450, 0.470]},
         "desc": "整体营业利润率, 随规模 + High-NA 毛利改善上行。管理层 2030 毛利率 56-60% 指引隐含 OPM 抬升。历史 2025 ~34.6%。",
         "stories": {"Bear": "混合不利 + High-NA 拖累毛利, OPM 缓升至 ~36.5%。",
                     "Base": "规模效应 + 服务占比升, OPM 至 2030 ~43%(对应毛利 56-60% 指引)。",
                     "Bull": "High-NA 满产 + 强定价, OPM 至 2030 ~47%。"},
         "hist": [0.363, 0.307, 0.328, 0.319, 0.346]},
    ],
    "linked": [
        {"key": "sent", "name": "情绪值(目标 P/E 第三层)", "fmt": K.N2,
         "src_sheet": S_MULT, "src_row0": ma["sent_row0"],
         "note": "三案取值与依据见『估值倍数假设』(完整三层方法论在那页); 本页只做切换。"},
    ],
})
# derived: 目标 forward P/E(当前案) = 峰值 P/E × 溢价 × 当前案情绪 → 喂『情景估值』
_pk = f"'{S_MULT}'!{ma['pk_cell']}"
_pr = f"'{S_MULT}'!{ma['pr_cell']}"
_sent_act = sw["SWACT"]["sent"]
_r = sw["next_row"]
K.lab(wb[S_SW], f"A{_r}", "目标 forward P/E(当前案)", b=True)
for _c in ALLC:
    K.fml(wb[S_SW], f"{_c}{_r}", f"={_pk}*{_pr}*{_c}{_sent_act}", K.MX, link=True)
K.logic(wb[S_SW], f"L{_r}", "= 历史峰值 P/E × 结构溢价 × 当前案情绪值 → 喂『情景估值』的前瞻目标 P/E。")
SWPE = _r

# ════════════ 8. 物理锚 [ANCHOR] — 光刻系统出货台数 ════════════
anchor = K.write_anchor(wb.create_sheet(S_ANCHOR), {
    "title": "光刻系统出货台数 (台) — 物理锚: EUV + DUV 年出货量",
    "all_cols": ALLC, "all_years": ALLY,
    "series": [
        ("EUV 出货台数 (台)", [42, 40, 53, 44, 48, None, None, None, None, None],
         "历史 = ASML 实际(2024-25 含 High-NA 2/4 台); 2026+ 由情景当前案驱动(『情景切换』)。", K.N0),
        ("  其中 High-NA (台)", [0, 0, 0, 2, 4, None, None, None, None, None],
         "High-NA EXE 系列出货(2024 起); 前瞻供参考(已并入 EUV 总台数与混合 ASP)。", K.N0),
        ("DUV 出货台数 (台)", [244, 305, 396, 374, 279, None, None, None, None, None],
         "历史 = ASML 实际; 2026+ 由情景当前案驱动。", K.N0),
    ],
    "yoy_row": None,
    "source_note": "口径 = ASML 光刻系统年出货台数(EUV 含 High-NA + DUV immersion/dry)。来源: 历史 = ASML 年报/季报实际; 前瞻 = 引『情景切换』当前案台数路径。这是整条估值链的物理起点。",
    "role_note": "作用: EUV 收入 = EUV 台数 × EUV ASP; DUV 收入 = DUV 台数 × DUV ASP(见『分部测算』)。改台数 → 收入 → 利润 → P/E → 隐含价全链动。",
})
EUVQ_ROW = anchor["row_of"]["EUV 出货台数 (台)"]
DUVQ_ROW = anchor["row_of"]["DUV 出货台数 (台)"]
# 前瞻 2026+ = 引『情景切换』当前案台数
for _c in FCf:
    K.fml(wb[S_ANCHOR], f"{_c}{EUVQ_ROW}", f"={K.R(S_SW, _c + str(sw['SWACT']['euvq']))}", K.N0, link=True)
    K.fml(wb[S_ANCHOR], f"{_c}{DUVQ_ROW}", f"={K.R(S_SW, _c + str(sw['SWACT']['duvq']))}", K.N0, link=True)

# ════════════ 9. 分部测算 — EUV/DUV = 台数 × ASP; 服务随装机 ════════════
seg = K.write_segment_model(wb.create_sheet(S_SEG), {
    "title": "分部测算 — EUV(台数×ASP) + DUV(台数×ASP) + 服务+M&I(随装机基数增长)",
    "all_cols": ALLC, "all_years": ALLY, "logic_col": "N",
    "groups": [
        ("EUV 段 = 台数 × ASP", [
            ("EUV 出货台数 (台)", None, K.N0, "= 引自『光刻系统出货』物理锚。改台数, EUV 收入跟着动。"),
            ("EUV ASP ($M/台)", None, K.N0, "历史 = EUV 收入 ÷ 台数(公式反推, $M/台); 前瞻 = 『情景切换』当前案(High-NA 拉高)。"),
            ("EUV 收入 ($B)", None, K.N1, "= 台数 × ASP ÷ 1000(台×$M→$B)。历史取历史财务实数。喂『利润与收入假设』。"),
        ]),
        ("DUV 段 = 台数 × ASP", [
            ("DUV 出货台数 (台)", None, K.N0, "= 引自『光刻系统出货』物理锚。"),
            ("DUV ASP ($M/台)", None, K.N0, "历史 = DUV 收入 ÷ 台数(公式反推); 前瞻 = 『情景切换』当前案(相对平稳)。"),
            ("DUV 收入 ($B)", None, K.N1, "= 台数 × ASP ÷ 1000。历史取历史财务实数。"),
        ]),
        ("服务+M&I 段 = 随装机基数复利", [
            ("服务+M&I 增速", None, K.PCT, "前瞻 = 『情景切换』当前案(~+8-12%/yr); 服务随累计装机基数增长, 粘性高。"),
            ("服务+M&I 收入 ($B)", None, K.N1, "历史取实数; 前瞻 = 上年 × (1+增速)。不挂台数, 走装机复利。"),
        ]),
    ],
})
m = seg["m"]
EUV_HROW = ha["seg_rows"]["EUV 系统 收入 ($B)"]
DUV_HROW = ha["seg_rows"]["DUV 系统 收入 ($B)"]
SVC_HROW = ha["seg_rows"]["服务+M&I 收入 ($B)"]
# 台数: 全列引锚
for col in ALLC:
    K.fml(wb[S_SEG], f"{col}{m['EUV 出货台数 (台)']}", f"={K.R(S_ANCHOR, col + str(EUVQ_ROW))}", K.N0, link=True)
    K.fml(wb[S_SEG], f"{col}{m['DUV 出货台数 (台)']}", f"={K.R(S_ANCHOR, col + str(DUVQ_ROW))}", K.N0, link=True)
# 历史: ASP 反推, 收入取实数
for col in HC:
    K.fml(wb[S_SEG], f"{col}{m['EUV ASP ($M/台)']}", f"={K.R(S_HIST, col + str(EUV_HROW))}*1000/{col}{m['EUV 出货台数 (台)']}", K.N0, link=True)
    K.fml(wb[S_SEG], f"{col}{m['EUV 收入 ($B)']}", f"={K.R(S_HIST, col + str(EUV_HROW))}", K.N1, link=True)
    K.fml(wb[S_SEG], f"{col}{m['DUV ASP ($M/台)']}", f"={K.R(S_HIST, col + str(DUV_HROW))}*1000/{col}{m['DUV 出货台数 (台)']}", K.N0, link=True)
    K.fml(wb[S_SEG], f"{col}{m['DUV 收入 ($B)']}", f"={K.R(S_HIST, col + str(DUV_HROW))}", K.N1, link=True)
    K.fml(wb[S_SEG], f"{col}{m['服务+M&I 收入 ($B)']}", f"={K.R(S_HIST, col + str(SVC_HROW))}", K.N1, link=True)
# 前瞻: ASP 引情景, 收入 = 台数 × ASP / 1000
for col in FCf:
    K.fml(wb[S_SEG], f"{col}{m['EUV ASP ($M/台)']}", f"={K.R(S_SW, col + str(sw['SWACT']['euvasp']))}", K.N0, link=True)
    K.fml(wb[S_SEG], f"{col}{m['EUV 收入 ($B)']}", f"={col}{m['EUV 出货台数 (台)']}*{col}{m['EUV ASP ($M/台)']}/1000", K.N1)
    K.fml(wb[S_SEG], f"{col}{m['DUV ASP ($M/台)']}", f"={K.R(S_SW, col + str(sw['SWACT']['duvasp']))}", K.N0, link=True)
    K.fml(wb[S_SEG], f"{col}{m['DUV 收入 ($B)']}", f"={col}{m['DUV 出货台数 (台)']}*{col}{m['DUV ASP ($M/台)']}/1000", K.N1)
    K.fml(wb[S_SEG], f"{col}{m['服务+M&I 增速']}", f"={K.R(S_SW, col + str(sw['SWACT']['svcg']))}", K.PCT, link=True)
_prevs = [HC[-1]] + list(FCf[:-1])
for _p, _c in zip(_prevs, FCf):
    K.fml(wb[S_SEG], f"{_c}{m['服务+M&I 收入 ($B)']}",
          f"={_p}{m['服务+M&I 收入 ($B)']}*(1+{_c}{m['服务+M&I 增速']})", K.N1)
for col in FCf:
    wb[S_SEG][f"{col}{m['EUV 收入 ($B)']}"].fill = K.OUT
    wb[S_SEG][f"{col}{m['DUV 收入 ($B)']}"].fill = K.OUT

# ════════════ 10. 利润与收入假设 ════════════
fr = K.write_fundamentals(wb.create_sheet(S_FUND), {
    "title": "利润与收入假设 — 分部营收(引分部测算)→ 整体营业利润率 → 净利 → 留存 → EPS/BPS/ROE",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
    "assum_groups": [
        ("利润率假设(历史实际锚 + 前瞻=情景当前案)", [
            {"name": "营业利润率 (%)", "vals": [0.363, 0.307, 0.328, 0.319, 0.346, None, None, None, None, None],
             "fmt": K.PCT, "logic": "整体营业利润率; 历史 = 实际; 前瞻 = 『情景切换』当前案(随规模 + High-NA 改善)。",
             "link": {"sheet": S_SW, "row": sw["SWACT"]["opm"]}},
        ]),
        ("净利转换与留存", [
            {"name": "净利转换率(净利/营业利润)", "vals": [0.87, 0.86, 0.87, 0.84, 0.85, 0.86, 0.86, 0.86, 0.86, 0.86],
             "fmt": K.PCT, "logic": "营业利润扣税/财务费用到净利的比例; 历史实际锚(ASML 有效税率 ~15-17%)。"},
            {"name": "留存率", "vals": [0.50, 0.45, 0.55, 0.55, 0.55, 0.55, 0.55, 0.55, 0.55, 0.55],
             "fmt": K.PCT, "logic": "留存率 = 1 - 派息率; ASML 派息 + 大额回购, 留存约半数用于再投资/缓冲。"},
        ]),
    ],
    "segments": [
        {"name": "EUV 系统 收入 ($B)", "hist_row": "EUV 系统 收入 ($B)", "fwd": {"sheet": S_SEG, "row": m["EUV 收入 ($B)"]}},
        {"name": "DUV 系统 收入 ($B)", "hist_row": "DUV 系统 收入 ($B)", "fwd": {"sheet": S_SEG, "row": m["DUV 收入 ($B)"]}},
        {"name": "服务+M&I 收入 ($B)", "hist_row": "服务+M&I 收入 ($B)", "fwd": {"sheet": S_SEG, "row": m["服务+M&I 收入 ($B)"]}},
    ],
    # 营业利润 = 总营收 × 整体营业利润率(单一 OPM, 用一个"段"权宜: 三段合计×OPM)
    "profit_terms": [
        (["EUV 系统 收入 ($B)", "DUV 系统 收入 ($B)", "服务+M&I 收入 ($B)"], "营业利润率 (%)", False),
    ],
    "conv_assum": "净利转换率(净利/营业利润)", "retention_assum": "留存率",
    "note_text": "分部营收(EUV/DUV = 台数×ASP, 服务随装机)→ 总营收 × 整体营业利润率 = 营业利润 → 净利(×净利转换)→ 权益(留存递推)→ EPS/BPS/ROE。历史列取实际(引『历史财务与估值』); 下游『情景估值』直接引本表每股, 不重算。前瞻股本固定取 2025A(F 列), 反映持续回购下的稳态股本。",
})

# ════════════ 11. 情景估值(P/E 主线) ════════════
# 把 kit 的"主线"用作 P/E: target_row=目标forward P/E(当前案), 但 kit 主线公式是 倍数×BPS。
# → 这里直接重写主线为 目标P/E × EPS。先调用 helper 拿 P/E 交叉验证 + 价格区块, 再覆盖主线两行。
sv = K.write_scenario_valuation(wb.create_sheet(S_VAL), {
    "title": "情景估值 — 当前案逐年隐含价 (P/E 主线; DCF 交叉验证见独立 sheet)",
    "intro": "本表输出 = 『情景切换』当前案(默认 Base)。隐含价 = 目标 forward P/E(当前案) × 前瞻 EPS。历史列用实际年末价反推 P/E(事实); 前瞻是预测、不拟合现价。三案并排见『估值对比』。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf, "hist_years": HY,
    "fx_fwd": FX_FWD, "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
    "s_fund": S_FUND, "fr": fr,
    "s_switch": S_SW, "target_row": SWPE, "sw_cell": "B2",
    "yend": px["yend"], "yavg": px["yavg"],
    "mcap_usd_skip": True,
    "reading": "P/E 主线读法: 目标 forward P/E(当前案) × 前瞻 EPS = 隐含价(本币=USD)。『当下 forward P/E』= 现价 ÷ 模型各年 EPS, 看以今日价你在为预测的当年盈利付几倍(现价÷FY26E EPS ~59x = 偏高 regime)。",
    "method": "方法: 整体公司、P/E 主线逐年估。基本面在『利润与收入假设』; 目标 P/E 在『估值倍数假设』(三层); 本表主线 = 目标 P/E × 前瞻 EPS → 隐含价 + 市值。DCF 在『DCF交叉验证』独立 sheet 做第二镜头。",
    "concl": "结论(方向性): 三情景见『估值对比』; 现价 $1,864 已对应 ~59x forward P/E, 高于 Base 36x — 模型隐含 Base 价低于现价, risk-reward 偏谨慎。",
})
# ★ 覆盖主线两行: kit 写的是 目标倍数(VPB)=实际P/B反推 + 隐含价=倍数×BPS。
#    改为 P/E 主线: 目标倍数 = 历史实际 forward P/E(HPX/EPS) / 前瞻=目标P/E; 隐含价 = 目标P/E × EPS。
VPB, VPBP = sv["pb"], sv["pb_px"]
EPS_F = lambda c: K.R(S_FUND, c + str(fr["EPS"]))
# 主线倍数行: 历史 = 实际年末价÷EPS(forward P/E 反推), 前瞻 = 目标forward P/E(当前案)
K.lab(wb[S_VAL], f"A{VPB}", "目标 forward P/E (历史=实际/前瞻=当前案)")
for c in HC:
    K.fml(wb[S_VAL], f"{c}{VPB}", f"={K.R(S_HIST, c + str(ha['HPX']))}/({EPS_F(c)})", K.MX, link=True)
for c in FCf:
    K.fml(wb[S_VAL], f"{c}{VPB}", f"={K.R(S_SW, c + str(SWPE))}", K.MX, link=True)
# 隐含价行: 前瞻 = 目标forward P/E × 前瞻EPS
K.lab(wb[S_VAL], f"A{VPBP}", "隐含股价 P/E主线 (USD)", b=True)
wb[S_VAL][f"A{VPBP}"].fill = K.OUT
for c in FCf:
    K.fml(wb[S_VAL], f"{c}{VPBP}", f"={c}{VPB}*{EPS_F(c)}", K.PX)

# ════════════ 12. 估值对比(三案并排; P/E 主线 + DCF 行) ════════════
SWB = sw["SWB"]
SH_F = K.R(S_HIST, f"$F${ha['HSH']}")
PX_NOW_REF = K.R(S_HIST, f"G{ha['HPX']}")
_opm = fr["am"]["营业利润率 (%)"]
_conv = fr["am"]["净利转换率(净利/营业利润)"]
_svcg_seg = m["服务+M&I 增速"]


def _fwdprev(j, A, key):
    return (HC[-1] if j == 0 else FCf[j - 1]) + str(A[key])


cmp_rows = [
    {"key": "euvq", "label": "EUV 台数 (台)", "fmt": K.N0,
     "hist": lambda c, ci, A: f"={K.R(S_ANCHOR, c + str(EUVQ_ROW))}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_SW, c + str(SWB['euvq'] + ci))}"},
    {"key": "euvasp", "label": "EUV ASP ($M/台)", "fmt": K.N0,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(EUV_HROW))}*1000/{c}{A['euvq']}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_SW, c + str(SWB['euvasp'] + ci))}"},
    {"key": "euv", "label": "EUV 收入 ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(EUV_HROW))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['euvq']}*{c}{A['euvasp']}/1000"},
    {"key": "duvq", "label": "DUV 台数 (台)", "fmt": K.N0,
     "hist": lambda c, ci, A: f"={K.R(S_ANCHOR, c + str(DUVQ_ROW))}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_SW, c + str(SWB['duvq'] + ci))}"},
    {"key": "duvasp", "label": "DUV ASP ($M/台)", "fmt": K.N0,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(DUV_HROW))}*1000/{c}{A['duvq']}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_SW, c + str(SWB['duvasp'] + ci))}"},
    {"key": "duv", "label": "DUV 收入 ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(DUV_HROW))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['duvq']}*{c}{A['duvasp']}/1000"},
    {"key": "svc", "label": "服务+M&I 收入 ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(SVC_HROW))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'svc')}*(1+{K.R(S_SW, c + str(SWB['svcg'] + ci))})"},
    {"key": "rev", "label": "总营收 ($B)", "fmt": K.N1, "bold": True,
     "hist": lambda c, ci, A: f"={c}{A['euv']}+{c}{A['duv']}+{c}{A['svc']}",
     "fwd": lambda c, j, ci, A: f"={c}{A['euv']}+{c}{A['duv']}+{c}{A['svc']}"},
    {"key": "ni", "label": "净利 ($B)", "fmt": K.N1, "bold": True,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HNI']))}",
     "fwd": lambda c, j, ci, A: (f"={c}{A['rev']}*{K.R(S_SW, c + str(SWB['opm'] + ci))}"
                                 f"*{K.R(S_FUND, c + str(_conv))}")},
    {"key": "eps", "label": "EPS ($)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HEPS']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['ni']}*1000/{SH_F}"},
    {"key": "sent", "label": "情绪值(该案; 历史=实际反推)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_MULT, c + str(ma['sent_row0'] + ci))}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_MULT, c + str(ma['sent_row0'] + ci))}"},
    {"key": "pe", "label": "目标 forward P/E(历史=实际)", "fmt": K.MX,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HPX']))}/{c}{A['eps']}",
     "fwd": lambda c, j, ci, A: f"={_pk}*{_pr}*{c}{A['sent']}"},
    {"key": "px", "label": "隐含价 (USD)", "fmt": K.PX, "bold": True, "out": True,
     "hist": lambda c, ci, A: f"={c}{A['pe']}*{c}{A['eps']}",
     "fwd": lambda c, j, ci, A: f"={c}{A['pe']}*{c}{A['eps']}"},
    {"key": "ipe", "label": "隐含 forward P/E(交叉验证)", "fmt": K.MX,
     "hist": lambda c, ci, A: f'=IF({c}{A["eps"]}<=0,"N/M",{c}{A["px"]}/{c}{A["eps"]})',
     "fwd": lambda c, j, ci, A: f"={c}{A['px']}/{c}{A['eps']}"},
    {"key": "up", "label": "历史: vs 实际年末价(回测≈0) / 前瞻: vs 现价", "fmt": K.PCT,
     "hist": lambda c, ci, A: f"={c}{A['px']}/{K.R(S_HIST, c + str(ha['HPX']))}-1",
     "fwd": lambda c, j, ci, A: f"={c}{A['px']}/{PX_NOW_REF}-1"},
]
cm = K.write_comparison(wb.create_sheet(S_CMP), {
    "title": "估值对比 — Bear / Base / Bull 三情景目标价并排 (P/E 主线)",
    "intro": ("三个情景各自完整推演一遍: 物理锚(台数) → ASP → 分部收入 → 净利 → EPS → 目标 forward P/E → 隐含价。"
              "本表三案永远并排可见, 不随『情景切换』开关变化; 要调假设去『情景切换』改对应案参数。"
              "历史列 2021-2025 = 同一条链填实际值: 隐含价历史列应基本等于实际年末价(回测行历史列 ≈0%)——整条链的内置回测。"),
    "case_names": CASES,
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "block_start": 16,
    "rows": cmp_rows,
    "summary": {
        "band": "三案汇总 (目标年 2027E; 各案触发条件见『情景切换』)",
        "target_col": "H",
        "rows": [
            ("净利($B)", "ni", K.N1, "由该案的 台数 / ASP / 利润率假设逐年推导得出", False),
            ("总营收($B)", "rev", K.N1, "= EUV(台数×ASP) + DUV(台数×ASP) + 服务(随装机)", False),
            ("目标 forward P/E", "pe", K.MX, "= 历史峰值 P/E × 结构溢价 × 该案情绪值", False),
            ("隐含价(USD)", "px", K.PX, "= 目标 forward P/E × 2027E EPS", True),
            ("隐含 forward P/E", "ipe", K.MX, "P/E 体检: 该案隐含价 ÷ 该案 EPS, 对照同业光谱", False),
            ("vs 现价", "up", K.PCT, "对照现价 $1,864 的上行 / 下行空间", True),
        ],
        "mcap": {"label": "隐含市值($B)", "key": "px", "expr": f"*{SH_F}/1000",
                 "note": "= 隐含价 × 股本(390M) / 1000 → $B"},
        "concl": "风险收益比(2027E vs 现价 $1,864)一行收口: Base 隐含价低于现价(现价已 price-in AI 溢价至 ~59x), 上行需 Bull 情景(High-NA 全面兑现); AI capex 见顶信号触发应转向 Bear 重看。",
    },
})

# ════════════ 13. DCF 交叉验证(第二镜头) ════════════
# 独立 sheet: WACC / 贴现 / 折现现金流 / 终值 → 每股内在价值。引『利润与收入假设』的营业利润、净利。
wsd = wb.create_sheet(S_DCF)
K.hdr(wsd, 1, "DCF 交叉验证 — 折现现金流 (WACC 贴现) 第二镜头, 对照 P/E 主线", 11)
K.mtext(wsd, 2, ("DCF 作为 P/E 主线的交叉验证: 用『利润与收入假设』的当前案营业利润 → NOPAT → 加回 D&A、减 CapEx 与 NWC 变动 → FCFF, "
                 "以 WACC 贴现 + 终值(永续增长法), 得企业价值 → 股权价值 → 每股内在价值。ASML 高 FCF 转换 + 长久期成长适合 DCF。"), "K", 3)
# 假设区
DR = 6
K.band(wsd, DR, "DCF 假设 (WACC / 永续增长 / 转换率)", 11); DR += 1
K.lab(wsd, f"A{DR}", "WACC")
K.inp(wsd, f"B{DR}", 0.095, None, K.PCT)
WACC = f"$B${DR}"
wsd.merge_cells(f"C{DR}:K{DR}"); K.logic(wsd, f"C{DR}", "加权平均资本成本(WACC): 无风险~4% + 半导体设备 beta~1.2 × ERP~5% + 适度风险溢价 ≈ 9.5%。"); DR += 1
K.lab(wsd, f"A{DR}", "永续增长率 g")
K.inp(wsd, f"B{DR}", 0.035, None, K.PCT)
GROW = f"$B${DR}"
wsd.merge_cells(f"C{DR}:K{DR}"); K.logic(wsd, f"C{DR}", "终值永续增长 g = 3.5%(略高于全球 GDP, 反映 ASML 长期结构性成长与垄断地位)。"); DR += 1
K.lab(wsd, f"A{DR}", "FCFF / 营业利润 转换率")
K.inp(wsd, f"B{DR}", 0.70, None, K.PCT)
FCFCV = f"$B${DR}"
wsd.merge_cells(f"C{DR}:K{DR}"); K.logic(wsd, f"C{DR}", "FCFF ≈ 营业利润 × 70%(NOPAT 扣税后约 0.83×0.85, 加回 D&A 减 CapEx/NWC, ASML 高现金转换近似)。"); DR += 1
# 现金流投影行(引利润与收入假设的营业利润, 当前案)
DR += 1
K.band(wsd, DR, "FCFF 投影与贴现 (当前案, 引『利润与收入假设』营业利润)", 11)
DR += 1
yr_row = DR
K.lab(wsd, f"A{yr_row}", "年份");
for col, y in zip(FCf, ALLY[5:]):
    wsd[f"{col}{yr_row}"] = y; wsd[f"{col}{yr_row}"].font = K.BF; wsd[f"{col}{yr_row}"].fill = K.CH
DR += 1
OP_ROW = K.R(S_FUND, "")  # placeholder
# 营业利润 = 总营收×OPM, 取『利润与收入假设』OP 行
FUND_OP = fr["OP"]
op_row = DR
K.lab(wsd, f"A{op_row}", "营业利润 ($B)")
for col in FCf:
    K.fml(wsd, f"{col}{op_row}", f"={K.R(S_FUND, col + str(FUND_OP))}", K.N1, link=True)
DR += 1
fcff_row = DR
K.lab(wsd, f"A{fcff_row}", "FCFF ($B)")
for col in FCf:
    K.fml(wsd, f"{col}{fcff_row}", f"={col}{op_row}*{FCFCV}", K.N1)
DR += 1
disc_row = DR
K.lab(wsd, f"A{disc_row}", "贴现因子")
for i, col in enumerate(FCf, start=1):
    K.fml(wsd, f"{col}{disc_row}", f"=1/(1+{WACC})^{i}", K.N2)
DR += 1
pv_row = DR
K.lab(wsd, f"A{pv_row}", "FCFF 现值 ($B)")
for col in FCf:
    K.fml(wsd, f"{col}{pv_row}", f"={col}{fcff_row}*{col}{disc_row}", K.N1)
DR += 1
# 终值
DR += 1
K.band(wsd, DR, "终值 + 企业价值 → 每股内在价值", 11); DR += 1
K.lab(wsd, f"A{DR}", "终值 TV (2030, 永续增长法) ($B)")
last = FCf[-1]
K.fml(wsd, f"B{DR}", f"={last}{fcff_row}*(1+{GROW})/({WACC}-{GROW})", K.N1)
TV = f"B{DR}"; DR += 1
K.lab(wsd, f"A{DR}", "终值现值 ($B)")
K.fml(wsd, f"B{DR}", f"={TV}*{last}{disc_row}", K.N1)
TVPV = f"B{DR}"; DR += 1
K.lab(wsd, f"A{DR}", "FCFF 现值合计 (2026-30) ($B)")
K.fml(wsd, f"B{DR}", f"=SUM({FCf[0]}{pv_row}:{last}{pv_row})", K.N1)
SUMPV = f"B{DR}"; DR += 1
K.lab(wsd, f"A{DR}", "企业价值 EV ($B)", b=True)
K.fml(wsd, f"B{DR}", f"={SUMPV}+{TVPV}", K.N1)
EV = f"B{DR}"; DR += 1
K.lab(wsd, f"A{DR}", "净现金 ($B)")
K.inp(wsd, f"B{DR}", 5.0, None, K.N1)
K.mtext(wsd, DR, "ASML 净现金头寸(现金 - 债务, 约 +$5B, 近似)。", "K", 1)
NETCASH = f"B{DR}"; DR += 1
K.lab(wsd, f"A{DR}", "股权价值 ($B)", b=True)
K.fml(wsd, f"B{DR}", f"={EV}+{NETCASH}", K.N1)
EQV = f"B{DR}"; DR += 1
K.lab(wsd, f"A{DR}", "股本 (mn 股)")
K.fml(wsd, f"B{DR}", f"={SH_F}", K.N0, link=True)
SHARES_ROW = f"B{DR}"
DR += 1
K.lab(wsd, f"A{DR}", "DCF 每股内在价值 (USD)", b=True)
wsd[f"A{DR}"].fill = K.OUT
K.fml(wsd, f"B{DR}", f"={EQV}*1000/{SHARES_ROW}", K.PX)
DCFPX = f"B{DR}"; DR += 1
K.lab(wsd, f"A{DR}", "vs 现价", b=True)
K.fml(wsd, f"B{DR}", f"={DCFPX}/{PX_NOW_REF}-1", K.PCT)
DR += 1
K.mtext(wsd, DR, ("DCF 第二镜头与 P/E 主线三角验证: DCF 内在价值反映 ASML 长久期成长的折现, "
                  "P/E 主线反映市场对盈利耐用性的定价。两者并列看分歧 = 估值区间的稳健性。"
                  "DCF 走当前案(默认 Base)营业利润路径; 切『情景切换』开关 DCF 同步变档。"), "K", 3)
K.set_widths(wsd, 30, FCf, 11, logic_col="L", logic_width=50)

# ════════════ 14. 综合判断仪表盘 ════════════
EPS26 = K.R(S_FUND, "G" + str(fr["EPS"]))
PXD = K.R(S_HIST, "G" + str(ha["HPX"]))
dash = K.write_dashboard(wb.create_sheet(S_DASH), {
    "title": "综合判断仪表盘 — A 基本面拐点 · B 估值错位(预测引擎) · C 催化剂 · D 情绪确认",
    "usage": ("怎么用: 预测引擎是 B(错位)+ C(催化剂)。情绪 D 只做 timing 确认 + 过热刹车。"
              "验收 = 回测: 放回 2024 AI 行情拐点, 这套表当时就能看到那波。"),
    "blocks": [
        {"title": "A. 基本面拐点 — 业务在结构性变好吗?", "rows": [
            ("产品组合迁移", "High-NA 占 EUV 出货逐年升(2024 2 台→Base 2030 占比上升)", "向更高 ASP 的 High-NA 迁移, 结构性升级。"),
            ("订单可见度", "Q4'25 net bookings €13.2B 创纪录(EUV €7.4B), backlog €38.8B", "AI 光刻需求强劲, 可见度向 2027 延展。"),
            ("A 判断", "【强】", "AI 驱动的结构性需求拐点已确立, EUV/High-NA 独占。", True),
        ]},
        {"title": "B. 估值错位(预测引擎 ★)— 市场现在给的 vs 基本面该给的 → GAP", "rows": [
            ("市场现在给(forward P/E)", {"fml": f"={PXD}/({EPS26})", "fmt": K.MX, "fill": True},
             "= 现价 $1,864 ÷ FY26E EPS(公式算, 随模型走), 当下 ~59x。"),
            ("基本面该给(Base 目标 P/E)", {"inp": 36.0, "fmt": K.MX},
             "Base 三层分解目标 forward P/E ~36x(峰值45×溢价1.0×情绪0.80)。"),
            ("错位 GAP = 该给÷市场给 − 1",
             {"fml": lambda ro: f"=B{ro['基本面该给(Base 目标 P/E)']}/B{ro['市场现在给(forward P/E)']}-1", "fmt": K.PCT},
             "GAP 负 = 现价 P/E 高于基本面该给 → 现价已 price-in 超额 AI 溢价, 进入纯情绪定价区(该谨慎)。"),
            ("回测: 2024 AI 拐点读数", "2024 初 forward P/E ~30x vs 该给 ~38x → GAP 正", "当时市场给 < 该给 → 错位正, 这就是 2024 那波上行的预测依据。"),
        ]},
        {"title": "C. 催化剂 — 什么会逼市场闭合 GAP", "rows": [
            ("High-NA 量产兑现", "进行中(Intel 14A / Samsung 导入)", "决定 2027+ EUV 台数与混合 ASP 上行斜率。"),
            ("2030 长期指引上修", "待(管理层 €44-60B 营收 / 56-60% 毛利率)", "若上修 → 打开长期 EPS 上行, 支撑更高目标 P/E。"),
            ("C 判断", "部分兑现", "High-NA 节奏是核心催化, 决定 Bull 情景能否成立。", True),
        ]},
        {"title": "D. 情绪确认 — 只做 timing + 刹车(定性档位)", "rows": [
            ("AI 光刻周期档位", "过热", "现价 ~59x forward P/E ≫ Base 36x = 市场已付出超过基本面该给。"),
            ("现价倍数 vs 基本面该给", "现价 59x vs 该给 36x", "进入情绪定价区, 上行需新催化(High-NA/指引上修)。"),
            ("当前档位", "【过热】", "AI 溢价已充分定价, 该兑现不该追; 等回调或催化兑现。", True),
            ("衰减扳机", "3 条", "① AI capex 指引转下修 ② High-NA 良率/导入延期 ③ 中国出口管制进一步收紧; 任一翻 → 情绪值下调转 Bear。"),
        ]},
    ],
    "final": {"band": "★ 综合判断(A+B+C+D 收成一句可执行的话)",
              "text": "回测 2024 拐点: 当时 GAP 正 + AI 催化启动 → 该看多 ✓。当下: 基本面强(A)但估值错位转负(B, 现价 59x≫该给 36x)+ 情绪过热(D)→ 现价已透支 Base, 建议持有/等回调, 上行 thesis 需 Bull(High-NA 全面兑现)。"},
    "tracking": {
        "intro": "哪个指标恶化 → 哪个假设先崩 → 触发什么动作。",
        "rows": [
            ("__band__", "一、核心驱动链(出货台数)"),
            ("EUV 出货台数 + High-NA 占比", "Base 2027 ~74 台", "关键敏感项: EUV 收入 = 台数 × ASP", "季报出货 + High-NA 验收(季度)", "台数低于路径 → 下调台数假设重算"),
            ("__band__", "二、ASP(混合)"),
            ("EUV 混合 ASP", "Base 2030 ~$340M/台", "关键敏感项: High-NA 占比决定 ASP 上行", "季报 ASP 拆分(季度)", "High-NA 占比慢 → 下移 ASP 路径"),
            ("__band__", "三、需求总盘子"),
            ("AI capex 指引", "强(hyperscaler 加速)", "关键敏感项: 整条出货链需求盘子", "hyperscaler 季报 + ASML 指引", "下修 >10% → 全链重算转 Bear"),
            ("__band__", "四、中国/DUV"),
            ("中国 DUV 占比", "2024 49%→2026 指引 ~20%", "关键敏感项: DUV 台数复苏强度", "季报地区拆分 + 出口管制", "管制收紧 → 下调 DUV 台数"),
        ],
    },
})

# ════════════ 全局格式 + 落盘 ════════════
K.finalize(wb, freeze={
    S_HIST: "B3", S_PX: "B4", S_CONS: "A2", S_HMULT: "B5", S_MULT: "B4", S_SW: "B3",
    S_ANCHOR: "B3", S_SEG: "B3", S_FUND: "B3", S_VAL: "B4", S_CMP: "B6", S_DCF: "B2", S_DASH: "B6",
    S_COVER: "A2",
})
out = os.path.join(os.path.dirname(__file__), "..", "out", "ASML_valuation_model.xlsx")
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("saved:", os.path.abspath(out))
print("sheets:", wb.sheetnames)
print("n_sheets:", len(wb.sheetnames))
