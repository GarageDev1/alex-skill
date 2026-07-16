# -*- coding: utf-8 -*-
"""
build_now.py — ServiceNow (NOW) 估值模型，基于 build_kit(v2)。

SaaS 锚：客户数 × ACV × NRR（存量复利 + 新增）。
分部：核心工作流(ITSM/CSM/HRSD/GRC) + 安全&Creator + NowAssist AI，三组不同增长动力学。
估值镜头：P/B 主线(kit 默认，资产视角 sanity) + P/E 平行镜头(主结论，盈利视角)。
口径：2025-12-18 完成 5:1 拆股，全程拆股后（股本~1.04B，股价 split-adjusted）。
"""
import os, sys
from openpyxl import Workbook
import build_kit as K

# ════════════ 0. 全局轴 ════════════
ALLC = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
ALLY = ["2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
HC, HY = ["B", "C", "D", "E", "F"], ["2021", "2022", "2023", "2024", "2025"]
FC = ["F", "G", "H", "I", "J", "K"]      # 含基年 F=2025A
FCf = FC[1:]                                # 纯前瞻 2026E-2030E
FX = 1.0                                    # USD 本币
CASES = ["Bear", "Base", "Bull"]

S_COVER, S_HIST, S_PX, S_CONS = "封面", "历史财务与估值", "股价走势", "卖方研报共识"
S_HMULT, S_MULT, S_SW = "历史估值倍数", "估值倍数假设", "情景切换"
S_ANCHOR, S_SEG, S_FUND = "SaaS物理锚", "分部测算", "利润与收入假设"
S_VAL, S_CMP, S_DASH = "情景估值", "估值对比", "综合判断仪表盘"

# 月度股价（split-adjusted，来自 now_price_6y.json，60 个月末）
import json
_repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_input_dir = os.environ.get("VALUATION_INPUT_DIR", os.path.join(_repo_root, "out"))
_pxfile = os.path.join(_input_dir, "now_price_6y.json")
with open(_pxfile, "r", encoding="utf-8") as _f:
    _pxdata = json.load(_f)
_monthly = {}
for _b in _pxdata["bars"]:
    _monthly[_b["date"][:7]] = _b["close"]
MONTHLY = [(ym, round(_monthly[ym], 2)) for ym in sorted(_monthly.keys())]
PX_NOW = MONTHLY[-1][1]

wb = Workbook(); wb.remove(wb["Sheet"])

# ════════════ 1. 封面 ════════════
K.write_cover(wb.create_sheet(S_COVER), {
    "title": "ServiceNow (NOW) 估值模型",
    "meta": [
        ("报告日期", "2026-07-01"),
        ("数据截止", "2025 FY 10-K + Q4 2025 财报 + 2026-06-30 收盘价"),
        ("现价", f"${PX_NOW:.2f}（拆股后；2025-12-18 完成 5:1 拆股）"),
        ("市值", "$103.9B  |  股本 ~1.04B（拆股后）"),
        ("时效声明", "基于 2026-07-01 时点的财报/共识/股价；SaaS 与 AI 转型假设变化快，建议每季财报后更新。"),
        ("方法一句话", "物理锚(客户数×ACV×NRR)→ 分部收入(核心工作流/安全Creator/AI 三组)→ 段驱动利润 → EPS → P/E 平行镜头×前瞻每股 → 隐含价。"),
    ],
    "takeaways": [
        ("① 当下估值位置", "现价 $99.28 对应 TTM P/E 59x、forward P/E ~48x，低于自身历史 5 年最低（2023 年 84x）——2026 SaaS 大跌创造的低位。"),
        ("② 核心引擎", "收入 = 存量客户×NRR(125%) + 新增 logo；NRR>120% 时存量自然扩张，是估值最敏感变量。"),
        ("③ AI 转型判断", "NowAssist(GenAI) 占比 5% 但增速 +50%，能否成为第二增长曲线决定增速换挡节奏。"),
        ("④ 三情景目标价", "Bear/Base/Bull 从 NRR + 利润率扩张 + 估值倍数 三杠杆翻档，沿同一条链算出，见『估值对比』。"),
        ("⑤ 主要风险", "Salesforce 入侵 ITSM + AI 投入拉低毛利 + 大客户渗透见顶 + 利率压制高估值。"),
    ],
})

# ════════════ 2. 历史财务与估值 ════════════
ha = K.write_history(wb.create_sheet(S_HIST), {
    "title": "ServiceNow 历史财务与估值 ($B) — 2021-2025A（拆股后口径）",
    "hist_cols": HC, "hist_years": HY,
    "fx_hist": [1, 1, 1, 1, 1], "fx_now": 1.0,
    "vals_in_usd": True, "fx_label": "FX (USD 单币种=1)",
    "segments": [
        # 分部历史（估算，合计≈实际总收入；公司不披露精确模块，按平台分组估）
        ("核心工作流 收入", [5.20, 6.35, 7.80, 9.20, 10.63], True),
        ("安全&Creator 收入", [0.50, 0.70, 1.10, 1.50, 1.99], True),
        ("NowAssist AI 收入", [0.00, 0.00, 0.07, 0.28, 0.66], False),  # 2021-22=0，关 YoY 防 0/0
    ],
    "total_now": 13.278,
    "gm_pct": [0.770, 0.783, 0.786, 0.792, 0.775], "gm_now": 0.775,
    "ni": [0.230, 0.325, 1.731, 1.425, 1.748], "ni_now": 1.748,
    "eq": [3.695, 5.032, 7.628, 9.609, 12.964], "eq_now": 12.964,
    "shares": [1015.8, 1017.7, 1027.5, 1038.1, 1046.7], "shares_now": 1046.7,  # 拆股后百万股
    "px_end": [129.82, 77.65, 141.30, 212.02, 153.19],
    "px_now": PX_NOW,
    "px_avg": [128.28, 93.80, 110.38, 167.37, 183.62],
    "band_note": "P/E 历史高位（2021-22 利润低致 PE 失真 574/243x）；正常年 2023-25 区间 84-154x。当前 59x 处历史低位。",
    "notes": [
        ("核心工作流 收入", "ITSM+CSM+HRSD+GRC+其他，基本盘，2025 占 80%（估）。"),
        ("安全&Creator 收入", "SecOps+Creator 低代码，高增长段，2025 占 15%（估）。"),
        ("NowAssist AI 收入", "GenAI 助手，2023 起量，2025 占 5% 增速 +50%（估）。"),
        ("HREV", "总营收=10-K 实际；分部按公司季度 commentary + 卖方拆分估。"),
        ("HGMP", "毛利率：订阅模式 77-79% 稳定，2025 回落 77.5%（AI 基础设施成本）。"),
        ("HNI", "净利：10-K 实际；2023 含一次性税项/投资收益偏高。"),
        ("HEQ", "股东权益：10-K 实际。"),
        ("HSH", "稀释股本（拆股后，百万股）；2025-12 5:1 拆股已追溯调整。"),
        ("HPX", "年末股价：split-adjusted（拆股后等价）；2025-12-18 拆股。"),
    ],
})

# ════════════ 3. 股价走势 ════════════
def phase_fn(ym):
    if ym <= "2022-10": return "① 利率估值倍数下修"
    if ym <= "2024-06": return "② 复苏扩张"
    if ym <= "2025-06": return "③ AI 叙事顶峰"
    return "④ SaaSpocalypse 回撤"
px = K.write_price_chart(wb.create_sheet(S_PX), MONTHLY, {
    "fn": phase_fn,
    "rows": [("① 利率估值倍数下修", "2022 加息周期，高估值 SaaS 普跌"),
             ("② 复苏扩张", "2023-24 利润率释放，股价修复"),
             ("③ AI 叙事顶峰", "2024末-2025初 GenAI 主题推至 $239 周期高"),
             ("④ SaaSpocalypse 回撤", "2025中-2026 AI 泡沫担忧+SaaS 集体回撤，拆股后跌至 $99")],
}, title="ServiceNow 月度股价 ($，split-adjusted)")

# ════════════ 4. 卖方研报共识 ════════════
K.write_consensus(wb.create_sheet(S_CONS), {
    "title": "卖方研报共识 — 45 位分析师；目标价均值 $141.48（vs 现价 $99.28）",
    "overview": "Strong Buy 共识：9 强买+34 买+4 持+1 卖。目标价区间 $85-$236，均值 $141.48（+42% 上行）。核心叙事：GenAI 驱动平台扩张 + 运营杠杆释放利润。",
    "assumptions": [
        ("收入增速\n(2026E)", "街上共识 +22%（rev $16.2B）。", "分歧在 AI 兑现节奏。", "base 取 +20%（略低于共识，基数效应谨慎）。"),
        ("NRR", "街上跟踪 125%，共识预期维持 120%+。", "NRR 是否见顶回落是最大分歧。", "base 取 122%（温和回落），bull 125%。"),
        ("营业利润率", "共识 2026E non-GAAP 30%+；GAAP ~15%。", "AI 投入对利润率的拖累程度。", "base GAAP 14.5%→18% 扩张路径。"),
        ("目标 P/E", "卖方隐含 forward P/E 30-40x（目标价÷2026E EPS）。", "给高增长 SaaS 多少溢价。", "base forward P/E 45-50x（三案见估值假设）。"),
    ],
    "divergences": [
        "① AI 货币化节奏：NowAssist 能否从功能升级变成独立增长曲线。",
        "② 竞争格局：Salesforce 入侵 ITSM 的实际威胁程度。",
        "③ 利率环境：高估值 SaaS 对 10Y 利率敏感。",
    ],
    "stances": [
        "GS（买入，TP ~$200）：AI 平台化龙头，NRR 维持 120%+。",
        "MS（买入，TP ~$180）：运营杠杆持续释放，利润率扩张路径清晰。",
        "Bernstein（持有，TP ~$120）：估值已反映大部分利好，AI 兑现存疑。",
    ],
})

# ════════════ 5. 历史估值倍数 ════════════
hm = K.write_hist_multiples(wb.create_sheet(S_HMULT), {
    "title": "历史估值倍数 — 自身 P/E·P/B 带 + 当下 + 同业对照",
    "intro": "数据底座：①自己历史上值多少（2023-25 正常利润年 P/E 84-154x，2021-22 低利润失真）②现在市场给多少（TTM 59x）③同行（CRM/WDAY/ADBE）+ 相对比值。看完再拍三案倍数。",
    "s_hist": S_HIST, "ha": ha, "hist_cols": HC, "hist_years": HY,
    "yhigh": px["yhigh"], "ylow": px["ylow"],
    "fwd_note": "forward P/E ~48x（现价÷base 2026E EPS $2.06）；forward P/B ~7.5x",
    "self_name": "ServiceNow",
    "self_fwd_pe_label": "≈48x",
    "self_note": "当前 TTM P/E 59x，低于自身 2023-25 正常区间下沿 84x——SaaSpocalypse 回撤所致。",
    "peers": [
        {"name": "Salesforce (CRM)", "yearly": [None, None, 30.0, 28.0, 22.0], "cur_pb": 4.2, "cur_pe": 45.0, "fwd_pe": 22.0,
         "note": "CRM 龙头，增速放缓(8-10%)，估值低于 NOW；正入侵 ITSM。"},
        {"name": "Workday (WDAY)", "yearly": [None, None, 40.0, 45.0, 38.0], "cur_pb": 7.5, "cur_pe": 55.0, "fwd_pe": 38.0,
         "note": "HCM/ERP，增速 17%，PE 与 NOW 接近。"},
        {"name": "Adobe (ADBE)", "yearly": [None, None, 35.0, 38.0, 30.0], "cur_pb": 9.0, "cur_pe": 35.0, "fwd_pe": 28.0,
         "note": "创意软件龙头，利润率高(40%+)，PE 低于 NOW。"},
        {"name": "大盘 SaaS 中位", "yearly": None, "cur_pb": None, "cur_pe": None, "fwd_pe": 30.0,
         "note": "高增长 SaaS forward P/E 光谱中位 ~30x（参照）。"},
    ],
    "ratio": {"peer": "Workday (WDAY)",
              "note": "NOW/WDAY forward PE 比 ~1.3x——NOW 享受增速+AI 叙事溢价。溢价能否维持取决于 NowAssist 兑现。"},
    "reading": "① 自己：当前 59x 显著低于 2023-25 正常下沿 84x → 估值已回落，非泡沫区。② 同行：NOW forward 48x 高于 CRM(22)/ADBE(28)/WDAY(38)，溢价来自更高增速+AI。③ 相对 WDAY 比值 1.3x，AI 兑现则撑住、不兑现则收敛。→ 下一页 P/E 三案 36-55x。",
})

# ════════════ 6. 估值倍数假设（P/B 三层 = kit 主线；P/E 在情景切换页另设）════════════
ma = K.write_multiple_assumptions(wb.create_sheet(S_MULT), {
    "title": "估值倍数假设 — P/B 主线(kit) + P/E 平行镜头(ServiceNow 主结论)",
    "intro": "ServiceNow 是轻资产 SaaS，P/B 意义有限（资产少、ROE 由无形资产驱动）。kit 主线用 P/B 作资产视角 sanity check；真正的估值主结论用 P/E 平行镜头（盈利视角，市场对 SaaS 的实际定价方式）。P/E 三案目标在『情景切换』页。",
    "why_text": ("镜头选择：ServiceNow 是已盈利的稳态高增长 SaaS（净利率 13%+、FCF margin 34%），市场用 P/E 定价。"
                 "kit 的 P/B 主线对轻资产 SaaS 只作资产视角参考（BPS 含少量现金+递延收入，P/B 历史波动大）。"
                 "主结论走 P/E 平行镜头：目标 P/E × 前瞻 EPS → 隐含价。P/B 镜头同步给出做三角验证（两个镜头差距大 = 对利润可持续性的分歧）。"),
    "why_rows": 5,
    "method_text": "P/B 三层（kit 主线，参考用）：①历史峰值 P/B ~35x（2021）× ②结构溢价 0.4x（轻资产，当前 ROE 低于峰值）× ③情绪值。当前 P/B 8x 已大幅低于历史，反映拆股后+回撤。P/E 三案在『情景切换』页直接设置目标（不经三层，因 SaaS P/E 由增速+利润率扩张+利率综合决定）。",
    "peak": 35.0, "peak_note": "2021 P/B 峰值 35.7x（低 BPS 高估值）。",
    "premium": 0.40, "premium_note": "轻资产 SaaS，BPS 含义弱，结构溢价折扣。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "hpb_row": ha["HPB"],
    "cases": [
        ("Bear", [0.25, 0.22, 0.20, 0.18, 0.17], "情绪退潮，P/B 回到下沿。"),
        ("Base", [0.30, 0.28, 0.25, 0.22, 0.20], "温和，P/B 随 BPS 增长逐步回升。"),
        ("Bull", [0.35, 0.33, 0.30, 0.27, 0.24], "AI 兑现，P/B 修复但难回 2021 峰值。"),
    ],
    "reconcile_text": "卖方目标价 $141 隐含 2026E P/E ~30x（用卖方 EPS）或 forward ~68x（用真实拆股后 EPS）。我们 base forward P/E 45-50x，比卖方隐含略保守——因我们对 AI 货币化节奏更谨慎。",
    "source_text": "P/B 历史峰值=『历史财务与估值』实际 P/B；同业=Yahoo/卖方；P/E 三案=本研究判断（增速+利润率+利率综合）。",
})

# ════════════ 7. 情景切换（SaaS 杠杆：分部增速 + NRR + OPM + 目标P/E）════════════
sw = K.write_scenario_switch(wb.create_sheet(S_SW), {
    "title": "情景切换 — SaaS 参数库（分部增速/NRR/利润率/目标P/E）+ 开关（默认 Base）",
    "usage": ("B2 下拉切案 → 各杠杆『当前案』行跟切 → 分部收入→利润→EPS→隐含价 全链变档。"
              "三案并排见『估值对比』。SaaS 核心杠杆：①三组分部增速 ②综合营业利润率 ③目标 P/E（主镜头）。"),
    "cases": CASES, "default": "Base",
    "triggers": [
        ("Bear", "Salesforce 入侵 ITSM 见效 + NowAssist 货币化不及预期 + NRR 跌破 118% + 利率上行 → 增速降至 10%、估值去溢价至 24-30x P/E。"),
        ("Base", "GenAI 稳步兑现、NRR 维持 120%+、利润率扩张至 18%、估值温和去溢价（36-45x P/E）。"),
        ("Bull", "NowAssist 成第二曲线（占比→15%）+ NRR 升至 125%+ + 利润率破 20% + 估值维持高溢价（46-55x P/E）。"),
    ],
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "levers": [
        {"key": "g_core", "name": "核心工作流 增速", "fmt": K.PCT, "cols": FCf,
         "vals": {"Bear": [0.15, 0.12, 0.09, 0.07, 0.06], "Base": [0.18, 0.16, 0.14, 0.12, 0.10], "Bull": [0.21, 0.19, 0.17, 0.15, 0.13]},
         "desc": "核心工作流(ITSM/CSM/HRSD/GRC) 增速，占收入 80%。SaaS 基本盘，随基数扩大温和回落。",
         "stories": {"Bear": "大客户渗透见顶+竞争加剧。", "Base": "稳健扩张，NRR 120%+ 支撑。", "Bull": "AI 提升客单价，增速维持。"},
         "hist": [None, 0.221, 0.229, 0.180, 0.155]},
        {"key": "g_sc", "name": "安全&Creator 增速", "fmt": K.PCT, "cols": FCf,
         "vals": {"Bear": [0.22, 0.18, 0.14, 0.11, 0.09], "Base": [0.28, 0.24, 0.20, 0.17, 0.15], "Bull": [0.34, 0.30, 0.26, 0.23, 0.20]},
         "desc": "安全+Creator 低代码，高增长段，占 15%。",
         "stories": {"Bear": "低代码竞争(MS Power Apps)加剧。", "Base": "企业自动化需求支撑。", "Bull": "AI 代码生成放量。"},
         "hist": [None, 0.400, 0.571, 0.364, 0.327]},
        {"key": "g_ai", "name": "NowAssist AI 增速", "fmt": K.PCT, "cols": FCf,
         "vals": {"Bear": [0.30, 0.25, 0.20, 0.15, 0.12], "Base": [0.50, 0.40, 0.32, 0.25, 0.20], "Bull": [0.80, 0.60, 0.45, 0.35, 0.28]},
         "desc": "GenAI 助手，2025 占 5%，增速最快但基数小。能否成第二曲线的关键变量。",
         "stories": {"Bear": "AI 沦为功能升级，无独立定价。", "Base": "稳步渗透，占比 2030→12%。", "Bull": "成第二曲线，占比 2030→18%。"},
         "hist": [None, None, None, 3.0, 1.357]},
        {"key": "opm", "name": "综合营业利润率", "fmt": K.PCT, "cols": FCf,
         "vals": {"Bear": [0.125, 0.115, 0.108, 0.102, 0.10], "Base": [0.145, 0.158, 0.168, 0.175, 0.18], "Bull": [0.155, 0.172, 0.185, 0.195, 0.20]},
         "desc": "GAAP 营业利润率，运营杠杆+AI 投入平衡。2025 实际 13.7%。",
         "stories": {"Bear": "AI 成本侵蚀+竞争压价。", "Base": "运营杠杆释放，扩张至 18%。", "Bull": "AI 提效+规模效应，破 20%。"},
         "hist": [0.044, 0.049, 0.085, 0.124, 0.137]},
        {"key": "tpe", "name": "目标 P/E（主镜头）", "fmt": K.MX, "cols": FCf,
         "vals": {"Bear": [42, 36, 30, 26, 24], "Base": [55, 50, 45, 40, 36], "Bull": [65, 60, 55, 50, 46]},
         "desc": "★ 主结论镜头：目标 P/E × 前瞻 EPS → 隐含价。随增速换挡去溢价。当前 forward ~48x。",
         "stories": {"Bear": "去溢价至 24-30x（低增长 SaaS 区间）。", "Base": "温和去溢价至 36-45x。", "Bull": "维持高溢价 46-55x。"},
         "hist": [574.4, 243.4, 83.9, 154.4, 91.7]},
    ],
    "linked": [
        {"key": "sent", "name": "情绪值(P/B 第三层)", "fmt": K.N2,
         "src_sheet": S_MULT, "src_row0": ma["sent_row0"],
         "note": "P/B 镜头情绪值（kit 主线用）；P/E 主镜头不经过此，P/E 三案直接在上方杠杆行。"},
    ],
})
# derived: 目标 P/B（当前案）= 峰值 × 溢价 × 当前案情绪 → 喂情景估值主线
_pk = f"'{S_MULT}'!{ma['pk_cell']}"; _pr = f"'{S_MULT}'!{ma['pr_cell']}"
_sent_act = sw["SWACT"]["sent"]; _r = sw["next_row"]
K.lab(wb[S_SW], f"A{_r}", "目标 P/B(当前案, kit 主线)", b=True)
for _c in ALLC:
    K.fml(wb[S_SW], f"{_c}{_r}", f"={_pk}*{_pr}*{_c}{_sent_act}", K.MX, link=True)
K.logic(wb[S_SW], f"L{_r}", "= P/B 峰值 × 结构溢价 × 当前案情绪 → 喂情景估值 P/B 主线（参考镜头）。")
SWPB = _r

# ════════════ 8. 物理锚 [ANCHOR] — SaaS 核心量 ════════════
anchor = K.write_anchor(wb.create_sheet(S_ANCHOR), {
    "title": "SaaS 物理锚 — 客户数 × ACV × NRR（存量复利 + 新增）",
    "all_cols": ALLC, "all_years": ALLY,
    "series": [
        ("订阅收入 ($B)", [5.51, 6.92, 8.70, 10.76, 13.02, None, None, None, None, None],
         "订阅收入(≈总收入 98%)，历史取 10-K；前瞻由分部测算驱动", K.N1),
        ("cRPO ($B)", [4.20, 5.60, 7.60, 9.80, 12.85, None, None, None, None, None],
         "current RPO(未来12月确认)，+25% YoY，订单动能先行指标", K.N1),
        ("NRR (%)", [1.19, 1.22, 1.24, 1.25, 1.25, None, None, None, None, None],
         "美元净留存率，>120%=强；存量扩张乘数，估值最敏感变量", K.PCT),
        ("大企业客户数(>$1M ARR)", [1.50, 1.62, 1.75, 1.90, 2.15, None, None, None, None, None],
         "ARR>$1M 客户数(k)，大客户驱动高 ACV", K.N0),
    ],
    "source_note": "口径：订阅收入/cRPO/NRR 来自 ServiceNow Q4 季报披露；客户数为估算（公司披露 ARR>$1M 客户增长但不给精确总数）。",
    "role_note": "作用：SaaS 物理盘子。核心工作流 ARR 增速 = 留存(NRR≈125%)扩张 + 新增 logo 的净 ARR 增速——这是驱动收入主力(80% 占比)的物理量，历史 NRR/cRPO/客户数是它的证据行。改本表『核心 ARR 增速』或情景 → 核心收入 → 净利 → EPS → 隐含价 全链动（连通性铁律的落点）。",
})
SUB_ROW = anchor["row_of"]["订阅收入 ($B)"]
# ★ 连通性锚点：核心工作流 ARR 增速(当前案) 落在物理锚 sheet，前瞻由情景切换驱动，
#   下游分部测算的核心收入从这里读——把物理锚接进"锚→收入→隐含价"主链。
_anc_ws = wb[S_ANCHOR]
_anc_g_row = anchor["row_of"]["大企业客户数(>$1M ARR)"] + 2
K.lab(_anc_ws, f"A{_anc_g_row}", "核心 ARR 增速(当前案)", b=True)
for _c in FCf:
    K.fml(_anc_ws, f"{_c}{_anc_g_row}", f"={K.R(S_SW, _c + str(sw['SWACT']['g_core']))}", K.PCT, link=True)
# 历史列填实际核心增速（供对照）
for _i in range(1, len(HC)):
    _c, _p = HC[_i], HC[_i-1]
    _core_hrow = ha["seg_rows"]["核心工作流 收入"]
    K.fml(_anc_ws, f"{_c}{_anc_g_row}", f'=IFERROR({K.R(S_HIST, _c + str(_core_hrow))}/{K.R(S_HIST, _p + str(_core_hrow))}-1,"n.m.")', K.PCT, link=True)
K.logic(_anc_ws, f"L{_anc_g_row}", "核心 ARR 净增速 = NRR 存量扩张 + 新增 logo（前瞻按情景，历史=实际反推）。分部测算核心收入引此行 → 物理锚进主链。")
ANC_GCORE = _anc_g_row

# ════════════ 9. 分部测算（3 组 SaaS 模块，增速驱动）════════════
seg = K.write_segment_model(wb.create_sheet(S_SEG), {
    "title": "分部测算 — 核心工作流 / 安全&Creator / NowAssist AI 三组（增速驱动）",
    "all_cols": ALLC, "all_years": ALLY, "logic_col": "N",
    "groups": [
        ("核心工作流 (ITSM/CSM/HRSD/GRC) — 基本盘", [
            ("核心 增速", None, K.PCT, "历史=实际 YoY；前瞻=『情景切换』当前案 g_core。占收入 80%，稳健。"),
            ("核心 收入 ($B)", None, K.N1, "历史取实数；前瞻=上年×(1+增速)。喂利润表。"),
        ]),
        ("安全 & Creator — 高增长段", [
            ("安全C 增速", None, K.PCT, "前瞻=『情景切换』g_sc。占 15%，企业自动化+安全需求。"),
            ("安全C 收入 ($B)", None, K.N1, "历史实数；前瞻=上年×(1+增速)。"),
        ]),
        ("NowAssist AI — 第二曲线候选", [
            ("AI 增速", None, K.PCT, "前瞻=『情景切换』g_ai。2025 占 5%，能否成第二曲线的关键。"),
            ("AI 收入 ($B)", None, K.N1, "历史实数(2023 起量)；前瞻=上年×(1+增速)。"),
        ]),
    ],
})
m = seg["m"]
CORE_HROW = ha["seg_rows"]["核心工作流 收入"]
SC_HROW = ha["seg_rows"]["安全&Creator 收入"]
AI_HROW = ha["seg_rows"]["NowAssist AI 收入"]
for col in HC:
    K.fml(wb[S_SEG], f"{col}{m['核心 收入 ($B)']}", f"={K.R(S_HIST, col + str(CORE_HROW))}", K.N1, link=True)
    K.fml(wb[S_SEG], f"{col}{m['安全C 收入 ($B)']}", f"={K.R(S_HIST, col + str(SC_HROW))}", K.N1, link=True)
    K.fml(wb[S_SEG], f"{col}{m['AI 收入 ($B)']}", f"={K.R(S_HIST, col + str(AI_HROW))}", K.N1, link=True)
# 历史增速（YoY，从收入行反推；B 列无前值标 n.m.）
for _rev_key, _grow_key in [("核心 收入 ($B)", "核心 增速"), ("安全C 收入 ($B)", "安全C 增速"), ("AI 收入 ($B)", "AI 增速")]:
    K.lab(wb[S_SEG], f"B{m[_grow_key]}", "n.m.", note=True)
    for _i in range(1, len(HC)):
        _c, _p = HC[_i], HC[_i-1]
        K.fml(wb[S_SEG], f"{_c}{m[_grow_key]}", f'=IFERROR({_c}{m[_rev_key]}/{_p}{m[_rev_key]}-1,"n.m.")', K.PCT)
# 前瞻增速 + 收入
# ★ 核心增速引『物理锚』的「核心 ARR 增速」行（物理锚接进主链：锚→收入→隐含价）
for col in FCf:
    K.fml(wb[S_SEG], f"{col}{m['核心 增速']}", f"={K.R(S_ANCHOR, col + str(ANC_GCORE))}", K.PCT, link=True)
    K.fml(wb[S_SEG], f"{col}{m['安全C 增速']}", f"={K.R(S_SW, col + str(sw['SWACT']['g_sc']))}", K.PCT, link=True)
    K.fml(wb[S_SEG], f"{col}{m['AI 增速']}", f"={K.R(S_SW, col + str(sw['SWACT']['g_ai']))}", K.PCT, link=True)
# 前瞻收入 = 上年 × (1+增速)
for key_growth, key_rev in [("核心 增速", "核心 收入 ($B)"), ("安全C 增速", "安全C 收入 ($B)"), ("AI 增速", "AI 收入 ($B)")]:
    _prevs = [HC[-1]] + list(FCf[:-1])
    for _p, _c in zip(_prevs, FCf):
        K.fml(wb[S_SEG], f"{_c}{m[key_rev]}", f"={_p}{m[key_rev]}*(1+{_c}{m[key_growth]})", K.N1)
for col in FCf:
    wb[S_SEG][f"{col}{m['核心 收入 ($B)']}"].fill = K.OUT
    wb[S_SEG][f"{col}{m['安全C 收入 ($B)']}"].fill = K.OUT
    wb[S_SEG][f"{col}{m['AI 收入 ($B)']}"].fill = K.OUT

# ════════════ 10. 利润与收入假设 ════════════
fr = K.write_fundamentals(wb.create_sheet(S_FUND), {
    "title": "利润与收入假设 — 分部营收(3组) → 综合OPM → 净利 → EPS/BPS",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
    "assum_groups": [
        ("利润率假设", [
            {"name": "综合营业利润率", "vals": [0.044, 0.049, 0.085, 0.124, 0.137, None, None, None, None, None],
             "fmt": K.PCT, "logic": "历史=10-K 实际 GAAP OP；前瞻=『情景切换』当前案 opm。",
             "link": {"sheet": S_SW, "row": sw["SWACT"]["opm"]}},
            {"name": "净利转换率(净利/营业利润)", "vals": [0.90, 0.92, 2.27, 1.05, 0.96, 0.93, 0.91, 0.90, 0.90, 0.90],
             "fmt": K.PCT, "logic": "营业利润→净利（扣税/利息）；2023 异常高(一次性)，前瞻 normalize 0.90-0.93。", "nm_cols": []},
            {"name": "留存率", "vals": [0.95, 0.95, 0.95, 0.95, 0.95, 0.95, 0.95, 0.95, 0.95, 0.95],
             "fmt": K.PCT, "logic": "无派息，留存率~0.95（回购小幅抵消）。"},
        ]),
    ],
    "segments": [
        {"name": "核心工作流 收入", "hist_row": "核心工作流 收入", "fwd": {"sheet": S_SEG, "row": m["核心 收入 ($B)"]}},
        {"name": "安全&Creator 收入", "hist_row": "安全&Creator 收入", "fwd": {"sheet": S_SEG, "row": m["安全C 收入 ($B)"]}},
        {"name": "NowAssist AI 收入", "hist_row": "NowAssist AI 收入", "fwd": {"sheet": S_SEG, "row": m["AI 收入 ($B)"]}},
    ],
    "profit_terms": [
        (["核心工作流 收入", "安全&Creator 收入", "NowAssist AI 收入"], "综合营业利润率", False),
    ],
    "conv_assum": "净利转换率(净利/营业利润)", "retention_assum": "留存率",
    "note_text": "分部营收(3组增速驱动)→ 综合营业利润(收入×OPM)→ 净利(×转换率)→ 权益(留存递推)→ EPS/BPS/ROE。历史列取 10-K 实际；前瞻 EPS 喂 P/E 主镜头隐含价。",
})

# ════════════ 11. 情景估值（P/B 主线 + P/E 平行镜头=主结论）════════════
sv = K.write_scenario_valuation(wb.create_sheet(S_VAL), {
    "title": "情景估值 — 当前案逐年隐含价（P/E 平行镜头=主结论；P/B 主线=参考）",
    "intro": "本表=『情景切换』当前案(默认 Base)。P/B 主线(kit)=目标P/B×BPS→隐含价(资产视角参考)；P/E 平行镜头=目标P/E×EPS→隐含价(主结论，SaaS 市场定价方式)。历史列=实际价反推倍数；前瞻不拟合现价。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf, "hist_years": HY,
    "fx_fwd": FX, "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
    "s_fund": S_FUND, "fr": fr,
    "s_switch": S_SW, "target_row": SWPB, "sw_cell": "B2",
    "yend": px["yend"], "yavg": px["yavg"],
    "pe_lens": {"target_row": sw["SWACT"]["tpe"], "mcap_label": "市值 前瞻·P/E镜头($B)"},
    "reading": "P/E 镜头读法：『当下 forward P/E』= 现价÷模型各年 EPS（光谱里标的的出处）；『P/E 前瞻隐含』= 目标P/E×EPS 的隐含价÷EPS。P/B 与 P/E 两镜头差距大 = 对利润可持续性/BPS 含义的分歧（轻资产 SaaS 正常）。",
    "method": "方法：P/E 平行镜头逐年估（主结论）。目标 P/E(情景切换当前案) × 前瞻 EPS → 隐含价。基本面在『利润与收入假设』；目标 P/E 在『情景切换』。",
    "concl": "结论(方向性)：三情景见『估值对比』；当前 $99 处历史低位，base 隐含价高于现价=有上行空间，但需 AI 兑现+利率配合。",
})

# ════════════ 12. 估值对比（三案 P/E 链为主）════════════
SWB = sw["SWB"]
SH_F = K.R(S_HIST, f"$F${ha['HSH']}")
PX_NOW_REF = K.R(S_HIST, f"G{ha['HPX']}")
_opm = fr["am"]["综合营业利润率"]
_conv = fr["am"]["净利转换率(净利/营业利润)"]
_ret = fr["am"]["留存率"]
_gc = m["核心 增速"]; _gsc = m["安全C 增速"]; _gai = m["AI 增速"]

def _fwdprev(j, A, key):
    return (HC[-1] if j == 0 else FCf[j - 1]) + str(A[key])

# 三案 block：收入(3组)→总营收→OP→净利→EPS→目标P/E→隐含价(P/E)→隐含价(P/B 参考)→vs现价
cmp_rows = [
    {"key": "core", "label": "核心工作流 ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(CORE_HROW))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'core')}*(1+{K.R(S_SW, c + str(SWB['g_core'] + ci))})"},
    {"key": "sc", "label": "安全&Creator ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(SC_HROW))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'sc')}*(1+{K.R(S_SW, c + str(SWB['g_sc'] + ci))})"},
    {"key": "ai", "label": "NowAssist AI ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(AI_HROW))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'ai')}*(1+{K.R(S_SW, c + str(SWB['g_ai'] + ci))})"},
    {"key": "rev", "label": "总收入 ($B)", "fmt": K.N1, "bold": True,
     "hist": lambda c, ci, A: f"={c}{A['core']}+{c}{A['sc']}+{c}{A['ai']}",
     "fwd": lambda c, j, ci, A: f"={c}{A['core']}+{c}{A['sc']}+{c}{A['ai']}"},
    {"key": "op", "label": "营业利润 ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HNI']))}*0+{c}{A['rev']}*{K.R(S_FUND, c + str(_opm))}" if False else f"={c}{A['rev']}*{K.R(S_FUND, c + str(_opm))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['rev']}*{K.R(S_SW, c + str(SWB['opm'] + ci))}"},
    {"key": "ni", "label": "净利 ($B)", "fmt": K.N1, "bold": True,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HNI']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['op']}*{K.R(S_FUND, c + str(_conv))}"},
    {"key": "eq", "label": "期末权益 ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HEQ']))}",
     "fwd": lambda c, j, ci, A: (f"={K.R(S_HIST, 'F' + str(ha['HEQ']))}+{c}{A['ni']}*{K.R(S_FUND, c + str(_ret))}" if j == 0
                                 else f"={FCf[j-1]}{A['eq']}+{c}{A['ni']}*{K.R(S_FUND, c + str(_ret))}")},
    {"key": "eps", "label": "EPS ($)", "fmt": K.N2, "bold": True,
     "hist": lambda c, ci, A: f"={c}{A['ni']}*1000/{K.R(S_HIST, c + str(ha['HSH']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['ni']}*1000/{SH_F}"},
    {"key": "tpe", "label": "目标 P/E（该案）", "fmt": K.MX,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HPE']))}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_SW, c + str(SWB['tpe'] + ci))}"},
    {"key": "px", "label": "隐含价 P/E镜头 ($)", "fmt": K.PX, "bold": True, "out": True,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HPX']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['tpe']}*{c}{A['eps']}"},
    {"key": "ipe", "label": "隐含 forward P/E(交叉)", "fmt": K.MX,
     "hist": lambda c, ci, A: f'=IF({c}{A["ni"]}<=0,"N/M",{K.R(S_HIST, c + str(ha["HPX"]))}/({c}{A["eps"]}))',
     "fwd": lambda c, j, ci, A: f"={c}{A['px']}/{c}{A['eps']}"},
    {"key": "up", "label": "历史:vs年末价(回测≈0)/前瞻:vs现价", "fmt": K.PCT,
     "hist": lambda c, ci, A: f"={c}{A['px']}/{K.R(S_HIST, c + str(ha['HPX']))}-1",
     "fwd": lambda c, j, ci, A: f"={c}{A['px']}/{PX_NOW_REF}-1"},
]
cm = K.write_comparison(wb.create_sheet(S_CMP), {
    "title": "估值对比 — Bear/Base/Bull 三案目标价并排（P/E 主镜头）",
    "intro": ("三案各自完整推演：分部增速(3组)→总收入→净利→EPS→目标P/E→隐含价。三案恒常并排。"
              "历史列 2021-2025 = 实际值，隐含价历史列=实际年末价（回测行历史列≈0%，链的内置回测）。"
              "主结论=P/E 镜头隐含价；P/B 镜头见『情景估值』作参考。"),
    "case_names": CASES,
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "block_start": 18,
    "rows": cmp_rows,
    "summary": {
        "band": "三案汇总 (目标年 2027E；触发条件见『情景切换』)",
        "target_col": "H",
        "rows": [
            ("总收入($B)", "rev", K.N1, "= 核心+安全Creator+AI 三组增速驱动", False),
            ("净利($B)", "ni", K.N1, "= 总收入×OPM×净利转换率", False),
            ("EPS ($)", "eps", K.N2, "= 净利÷股本(~1.04B)", False),
            ("目标 P/E", "tpe", K.MX, "= 该案 P/E（增速换挡去溢价）", False),
            ("隐含价 P/E镜头 ($)", "px", K.PX, "= 目标P/E × EPS（主结论）", True),
            ("隐含 forward P/E", "ipe", K.MX, "= 隐含价÷EPS，交叉验证", False),
            ("vs 现价", "up", K.PCT, "= 隐含价÷$99.28−1", True),
        ],
        "mcap": {"label": "隐含市值($B)", "key": "px", "expr": f"*{SH_F}/1000",
                 "note": "= 隐含价 × 股本(~1.04B)"},
        "concl": "风险收益比(2027E vs $99.28)：base 隐含价上行空间取决于 AI 兑现+利润率扩张；Bear 下行风险=竞争+估值去溢价。详见决策 memo。",
    },
})

# ════════════ 13. 综合判断仪表盘 ════════════
EPS27 = K.R(S_FUND, "H" + str(fr["EPS"]))
PXD = K.R(S_HIST, "G" + str(ha["HPX"]))
dash = K.write_dashboard(wb.create_sheet(S_DASH), {
    "title": "综合判断仪表盘 — A 基本面拐点 · B 估值错位(预测引擎) · C 催化剂 · D 情绪确认",
    "usage": "预测引擎=B(错位)+C(催化剂)，当下可观测；D 情绪只做 timing 确认。验收=回测：放回 2022 低点或 2025 高点，这套表当时能看到。",
    "blocks": [
        {"title": "A. 基本面拐点 — 业务在结构性变好吗?", "rows": [
            ("收入增速持续性", "2025 +21%，2026E 共识 +22%；核心 NRR 125% 支撑存量扩张", "高增长 SaaS 顶部 10-15% 仍有路径。"),
            ("利润率扩张路径", "GAAP OP 4.4%(2021)→13.7%(2025)，运营杠杆释放", "AI 投入是上行风险也是下行风险。"),
            ("AI 货币化", "NowAssist 占 5% 增速 +50%，能否成第二曲线", "决定增速换挡节奏。"),
            ("A 判断", "【中强】", "基本面稳健，AI 兑现是增量。", True),
        ]},
        {"title": "B. 估值错位(预测引擎 ★)— 市场给 vs 基本面该给 → GAP", "rows": [
            ("市场现在给(forward P/E)", {"fml": f"={PXD}/{EPS27}", "fmt": K.MX, "fill": True},
             "= 现价 ÷ 2027E EPS（公式算）。"),
            ("基本面该给(justified forward P/E)", {"inp": 52.0, "fmt": K.MX},
             "= base 三案中枢（增速 17%+利润率扩张+利率正常化该给的稳态高增长 SaaS P/E）。"),
            ("错位 GAP = 该给÷市场给 − 1",
             {"fml": lambda ro: f"=B{ro['基本面该给(justified forward P/E)']}/B{ro['市场现在给(forward P/E)']}-1", "fmt": K.PCT},
             "GAP 正=重估空间(该买)；转负=已透支。当前 forward PE ~48x < 该给 52x = 小幅上行空间。"),
            ("回测: 2025 初高点读数", "当时 forward PE ~60x vs 该给 ~50x → GAP 转负 → 那波回撤的预测依据", "现价已闭合大部分错位。"),
        ]},
        {"title": "C. 催化剂 — 什么逼市场闭合 GAP", "rows": [
            ("Q2/Q3 2026 财报", "✓ 进行中", "NowAssist ARR 披露+ NRR 走势是核心。"),
            ("利率下行", "待", "10Y 回落直接抬升高估值 SaaS。"),
            ("AI 货币化里程碑", "待", "NowAssist 占比突破 10%。"),
            ("C 判断", "部分兑现", "财报是最近 catalyst；利率看宏观。", True),
        ]},
        {"title": "D. 情绪确认 — timing + 刹车（定性档位）", "rows": [
            ("SaaSpocalypse 位置", "从 52 周高 $211 跌 53% 至 $99", "恐慌情绪释放中，但未见明确反转信号。"),
            ("现价 vs 基本面该给", "forward 48x 略低于该给 52x", "接近合理偏低，非极端低估。"),
            ("当前档位", "【退潮末期/萌芽】", "下行恐慌接近尾声，反转需 catalyst。", True),
            ("衰减扳机", "NRR 跌破 118% / NowAssist 增速跌破 30% / OPM 转跌", "任一翻→降档+P/E 下调。"),
        ]},
    ],
    "final": {"band": "★ 综合判断(A+B+C+D)",
              "text": "基本面中强(A)+估值小幅错位上行(B)+催化剂部分兑现(C)+情绪退潮末期(D)→ 偏多但非重仓信号；base 隐含价高于现价，建议逢 catalyst(Q2 财报/利率下行)分批建仓，Bear 档作止损参照。"},
    "tracking": {
        "intro": "哪个指标恶化 → 哪个假设先崩 → 触发什么动作（盯的优先级）。",
        "rows": [
            ("__band__", "一、核心驱动链（NRR + AI 兑现）"),
            ("NRR（美元净留存率）", "125%", "关键敏感项：存量扩张乘数，>120% 强", "季报披露", "跌破 118% → 全链下调、转 Bear"),
            ("NowAssist ARR/占比", "5% 占比 +50% 增速", "关键敏感项：第二曲线兑现", "季报 segment 披露", "增速跌破 30% → 下调 g_ai"),
            ("__band__", "二、利润率"),
            ("GAAP 营业利润率", "13.7%", "关键敏感项：AI 投入 vs 运营杠杆", "季报", "环比转跌 → 下调 opm 路径"),
            ("__band__", "三、估值与利率"),
            ("forward P/E", "~48x", "关键敏感项：高估值对利率敏感", "股价/一致预期", "10Y 突破 4.8% → P/E 目标下调"),
        ],
    },
})

# ════════════ 全局格式 + 落盘 ════════════
K.finalize(wb, freeze={
    S_HIST: "B3", S_PX: "B4", S_CONS: "A2", S_HMULT: "B5", S_MULT: "B4", S_SW: "B3",
    S_ANCHOR: "B3", S_SEG: "B3", S_FUND: "B3", S_VAL: "B4", S_CMP: "B6", S_DASH: "B6",
    S_COVER: "A2",
})
out = os.path.join(os.path.dirname(__file__), "..", "out", "NOW_valuation_model.xlsx")
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("saved:", os.path.abspath(out))
print("sheets:", wb.sheetnames)
print("PX_NOW:", PX_NOW)
