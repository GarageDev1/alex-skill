# -*- coding: utf-8 -*-
"""
build_template.py — 用 build_kit(v2, 对齐 v47 架构)给新公司建模的瘦实例模板。

演示"新公司只填 data dict + 调 build_kit 的 sheet 骨架函数"就能出 13-sheet 机构级
Excel。下面用占位/示意数据(虚构内存/AI 半导体公司 DEMOCO),数据本身不重要——
重要的是调用骨架。新公司建模时:
  1. 复制本文件;
  2. 改全局轴(年份/列/汇率)+ 把每个 write_* 的 data dict 换成真实数据;
  3. 公司专有的传导公式(分部测算回填 / 估值对比 lambdas)按自家链改;
  4. PYTHONUTF8=1 python examples/build_template.py
  5. python scripts/validate_valuation.py out/_template_demo.xlsx

调用顺序 = sheet 物理顺序 = 只向前引用的 DAG(05 §1):
  封面 → 历史财务与估值 → 股价走势 → 卖方研报共识 → 历史估值倍数 → 估值倍数假设
  → 情景切换 → 物理锚[ANCHOR] → 分部测算 → 利润与收入假设 → 情景估值 → 估值对比
  → 综合判断仪表盘 → K.finalize()
"""
import os
from openpyxl import Workbook
import build_kit as K

# ════════════ 0. 全局轴(年份 / 列 / 汇率 / 价格序列)════════════
ALLC = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
ALLY = ["2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
HC, HY = ["B", "C", "D", "E", "F"], ["2021", "2022", "2023", "2024", "2025"]
FC = ["F", "G", "H", "I", "J", "K"]          # 前瞻含基年 F=2025A
FCf = FC[1:]                                   # 纯前瞻 2026E-2030E
FX_FC = 1410
FX_HIST = [1144, 1292, 1306, 1364, 1397]
CASES = ["Bear", "Base", "Bull"]

S_COVER, S_HIST, S_PX, S_CONS = "封面", "历史财务与估值", "股价走势", "卖方研报共识"
S_HMULT, S_MULT, S_SW = "历史估值倍数", "估值倍数假设", "情景切换"
S_ANCHOR, S_SEG, S_FUND = "AIDC Capex预测", "分部测算", "利润与收入假设"
S_VAL, S_CMP, S_DASH = "情景估值", "估值对比", "综合判断仪表盘"

# 月度收盘(单一价格来源, 示意; 实际填完整真实序列, tvremix get_ohlcv resample 月末)
MONTHLY = [("2021-01", 118000), ("2021-06", 122500), ("2021-12", 120500),
           ("2022-06", 95000), ("2022-12", 88500),
           ("2023-06", 110000), ("2023-12", 134700),
           ("2024-06", 180000), ("2024-12", 199200),
           ("2025-06", 270000), ("2025-12", 909000),
           ("2026-03", 1500000), ("2026-06", 2360000)]
PX_NOW = 2360000

wb = Workbook()
wb.remove(wb["Sheet"])

# ════════════ 1. 封面 ════════════
K.write_cover(wb.create_sheet(S_COVER), {
    "title": "DEMOCO 估值模型(示意)",
    "meta": [
        ("报告日期", "YYYY-MM-DD(示意)"),
        ("数据截止", "最新季报 + 卖方共识 + 实时股价的时点(示意)"),
        ("现价", f"{PX_NOW:,} 本币(示意)"),
        ("时效声明", "基于建模时点的共识/指引/股价;关键假设变化快, 建议每季财报后更新(示意)。"),
        ("方法一句话", "物理锚(AIDC capex)→ 分部收入(AI 段=capex×强度, 商品段=周期)→ 段驱动利润 → 三层目标倍数 × 前瞻每股 → 隐含价(示意)。"),
    ],
    "takeaways": [
        ("① 当下估值位置", "现价对应的 TTM/forward 倍数处历史带什么位置(示意)。"),
        ("② 核心引擎", "AI 收入 = AIDC capex × 收入强度, 强度锚最近实际年反推(示意)。"),
        ("③ 周期判断", "商品价格何时见顶/normalize 的物理依据与时点(示意)。"),
        ("④ 三情景目标价", "Bear/Base/Bull 从业务杠杆翻档, 沿同一条链算出, 见『估值对比』(示意)。"),
        ("⑤ 主要风险", "价格周期见顶 / 份额被切 / 单一客户 / normalize 时点(示意)。"),
    ],
})

# ════════════ 2. 历史财务与估值 ════════════
ha = K.write_history(wb.create_sheet(S_HIST), {
    "title": "DEMOCO 历史财务与估值 ($B) — 2021-2025A + 当下(示意数据)",
    "hist_cols": HC, "hist_years": HY,
    "fx_hist": FX_HIST, "fx_now": FX_FC,
    "segments": [
        ("AI内存 收入", [0.5, 1.0, 2.0, 11.0, 28.0], True),
        ("商品内存 收入", [40.0, 41.0, 28.0, 52.0, 64.4], True),
        ("其他 收入", [2.5, 2.6, 1.8, 4.2, 5.1], False),
    ],
    "total_now": 132.0,
    "gm_pct": [0.45, 0.36, -0.02, 0.48, 0.60], "gm_now": 0.68,
    "ni": [9.6, 2.4, -9.1, 19.8, 42.9], "ni_now": 75.0,
    "eq": [64.6, 66.8, 53.5, 73.1, 113.8], "eq_now": 120.0,
    "shares": [710, 710, 710, 705, 700], "shares_now": 700,
    "px_end": [120500, 88500, 134700, 199200, 909000],
    "px_now": PX_NOW,
    "px_avg": [120300, 91750, 122350, 189600, 589500],
    "band_note": "P/B 历史常态 1-2x → 2025末冲上沿 → 当下远超(示意)",
    "notes": [
        ("AI内存 收入", "分部拆分按公司季度 commentary 估; AI 段爆发期(示意)。"),
        ("商品内存 收入", "商品段=总营收−AI−其他, 强周期(示意)。"),
        ("其他 收入", "小体量段, 随消费电子周期(示意)。"),
        ("HREV", "总营收=交易所实际; 分部拆分按公司季度 commentary 估(示意)。"),
        ("HGMP", "毛利率: 卖方 blended 口径, 早年粗估(示意)。"),
        ("HNI", "净利: 公司实际, 按当年均值汇率折$B; 当下=TTM(示意)。"),
        ("HEQ", "股东权益: 公司年报实际, 按当年均值汇率(示意)。"),
        ("HSH", "基本流通股(扣库存); 无增发, 小幅回购(示意)。"),
        ("HPX", "年末股价: 交易所月K真实收盘; 当下=现价(示意)。"),
        ("HPXA", "年均股价: 月度收盘均值, 同『股价走势』单一价格源(示意)。"),
    ],
})

# ════════════ 3. 股价走势 ════════════
def phase_fn(ym):
    if ym <= "2023-03":
        return "① 周期低迷"
    if ym <= "2024-12":
        return "② 复苏"
    if ym <= "2025-12":
        return "③ 爆发"
    return "④ 超级周期"

px = K.write_price_chart(wb.create_sheet(S_PX), MONTHLY, {
    "fn": phase_fn,
    "rows": [("① 周期低迷", "下行周期, 区间震荡(示意)"),
             ("② 复苏", "AI 浪潮启动(示意)"),
             ("③ 爆发", "AI 段放量(示意)"),
             ("④ 超级周期", "创历史新高(示意)")],
}, title="DEMOCO 月度股价 (本币)")

# ════════════ 4. 卖方研报共识(按"模型假设"组织的卖方对账单)════════════
K.write_consensus(wb.create_sheet(S_CONS), {
    "title": "卖方研报共识 — N 家;这张表是后面测算的'卖方对账单'(示意)",
    "overview": "一句话总览: 全街评级 / 目标价区间 / 核心叙事(示意)。",
    "assumptions": [
        ("AI 收入增速\n(2026)", "街上共识区间, 说人话(示意)。", "分歧在哪、为什么(示意)。", "base 取了谁的口径 + 为什么(示意)。"),
        ("商品价格\n(2026)", "共识暴涨, 但口径不同(示意)。", "混合口径 vs 纯商品口径(示意)。", "取中段(示意)。"),
        ("缺口何时转松", "最大分歧: 偏早 vs 偏晚(示意)。", "周期何时见顶 = 估值最大不确定(示意)。", "base 挂物理事件(新厂出 bit 时点)(示意)。"),
        ("目标倍数", "卖方区间(示意)。", "给多高的结构溢价(示意)。", "三层分解拍, 见『估值倍数假设』(示意)。"),
    ],
    "divergences": [
        "① 周期何时见顶: 决定 normalize 时点拍在哪年(示意)。",
        "② 竞品份额: 决定 AI 段强度路径(示意)。",
    ],
    "stances": [
        "BrokerA(买入, TP xxx): 一句话核心观点(示意)。",
        "BrokerB(买入, TP xxx): 一句话核心观点(示意)。",
    ],
})

# ════════════ 5. 历史估值倍数(数据底座)════════════
hm = K.write_hist_multiples(wb.create_sheet(S_HMULT), {
    "title": "历史估值倍数 — 先看数据再做假设: 自身历史带 + 当下倍数 + 同业对照(示意)",
    "intro": "这一页是『估值倍数假设』的数据底座: ①自己历史上值多少(逐年+年内高低带) ②现在市场给多少 ③同行值多少 + 相对核心同行的比值(结构溢价对账线)。看完这页再去下一页拍三案倍数。",
    "s_hist": S_HIST, "ha": ha, "hist_cols": HC, "hist_years": HY,
    "yhigh": px["yhigh"], "ylow": px["ylow"],
    "fwd_note": "forward P/E ≈10x · forward P/B ≈4.9x(现价÷模型前瞻每股, 示意)",
    "self_name": "DEMOCO",
    "self_fwd_pe_label": "≈10x",
    "self_note": "本模型标的; forward 推导见『情景估值』(示意)。",
    "peers": [
        {"name": "PeerA(综合体)", "yearly": [1.8, 1.1, 1.5, 0.9, None], "cur_pb": 5.0, "cur_pe": 25.8, "fwd_pe": 13.0,
         "note": "核心可比; 2025 年末缺公开口径标 n.a. 不硬补(示意)。"},
        {"name": "PeerB(纯同业)", "yearly": [2.27, 1.11, 2.20, 2.11, 6.04], "cur_pb": 19.7, "cur_pe": 44.8, "fwd_pe": 11.0,
         "note": "最直接可比; 同业也在重估(示意)。"},
        {"name": "AI 龙头(参照上沿)", "yearly": None, "cur_pb": None, "cur_pe": 32.0, "fwd_pe": 28.0,
         "note": "forward PE 光谱上沿(示意)。"},
        {"name": "大盘(参照下沿)", "yearly": None, "cur_pb": None, "cur_pe": None, "fwd_pe": 10.0,
         "note": "光谱下沿(示意)。"},
    ],
    "ratio": {"peer": "PeerA(综合体)",
              "note": "结构溢价的对账线: 比值从 <1 一路扩张 = 结构溢价持续扩张 → 下一页第二层取值有据(示意)。"},
    "reading": "① 自己: 当下 TTM 倍数 vs 历史带位置 → 第一层锚取历史峰值, 不用被本轮抬高的当下倍数。② 同行: 行业性 re-rating 还是独贵; forward 口径错位即重估空间。③ 相对核心同行比值走到哪 → 第二层结构溢价的依据。→ 下一页: 峰值 × 溢价 × 情绪值(三案)。(示意)",
})

# ════════════ 6. 估值倍数假设(三案倍数在此拍)════════════
ma = K.write_multiple_assumptions(wb.create_sheet(S_MULT), {
    "title": "估值倍数假设 — 主线方法论 + 三案目标倍数(= 历史周期峰值 × 结构溢价 × 情绪值)",
    "intro": "这一页只做判断(数据底座在上一页): ①为什么这个镜头做主线 ②三层分解出三案目标倍数。『情景切换』引用并切换, 『情景估值』套用当前案, 『估值对比』三案并排。",
    "why_text": ("镜头选择是业务判断: 这家公司『穿越周期持续存在的东西』是盈利还是资产? 用那个做分母。"
                 "商品型重资产无定价权 → 盈利是周期状态量, 资产/产能才是结构存量 → P/B 主线, P/E 永远做支线体检(每个情景都做)。"
                 "若 franchise 段长出定价权, 镜头会随业务迁移(re-rating 的本质 = 市场换镜头)——写成判断+触发条件, 不是永久标签。(示意, 新标的按 04 §镜头选择现写)"),
    "why_rows": 4,
    "method_text": "三层分解(不硬拍): ①历史周期峰值(过去最强周期实际到过的倍数, 不用本轮已重估的当下值) × ②结构溢价(锚同业相对法) × ③情绪值(周期/情绪位置, 依据『综合判断仪表盘』D块)。一致性检验: 三层相乘应复现最近实际倍数。(示意)",
    "peak": 2.0, "peak_note": "过去最强周期实际到过的峰值(示意; 实拉研报库口径)。",
    "premium": 1.30, "premium_note": "相对核心同行的结构溢价(示意; 对账线见上一页比值行)。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "hpb_row": ha["HPB"],
    "cases": [
        ("Bear", [0.90, 0.70, 0.50, 0.40, 0.35], "任一衰减扳机触发 → 情绪快速退潮(示意故事: 事件→时点→数字→出处)。"),
        ("Base", [1.20, 1.10, 0.75, 0.50, 0.45], "可见度撑 18-24 个月, 之后退潮(示意故事)。"),
        ("Bull", [1.35, 1.40, 1.00, 0.70, 0.55], "市场接受重估叙事, 超涨延长(示意故事)。"),
    ],
    "reconcile_text": "卖方目标倍数区间 vs 我们: 凭什么敢给非主流数——事实+逻辑链(示意)。",
    "source_text": "第一层口径来源 / 第二层口径来源 / 情绪值依据=『综合判断仪表盘』D块(示意)。",
})

# ════════════ 7. 情景切换 ════════════
sw = K.write_scenario_switch(wb.create_sheet(S_SW), {
    "title": "情景切换 — 全模型唯一的情景参数库 + 切换开关 (默认 Base)",
    "usage": ("怎么用: B2 是唯一入口——下拉选案 → 案序号派生 → 各杠杆『当前案』行跟着切 → "
              "整条明细链(锚→测算→利润→倍数→估值)变档, 『情景估值』输出该案逐年隐含价。"
              "三案对比不用切: 『估值对比』恒常三列并排(引本页矩阵行)。情景参数只在本页改(蓝字); 未列入的假设三案共用(跟 Base)。"),
    "cases": CASES, "default": "Base",
    "triggers": [
        ("Bear", "什么发生 = 落进 Bear(示意: 竞品过认证 + 价格提前见顶 + 衰减扳机触发)。"),
        ("Base", "什么发生 = 落进 Base(示意: 缺口按物理时点转松, 可见度撑住)。"),
        ("Bull", "什么发生 = 落进 Bull(示意: 代际溢价兑现 + 缺货拖更久 + 重估叙事被接受)。"),
    ],
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "levers": [
        {"key": "capex", "name": "AIDC capex 增速", "fmt": K.PCT, "cols": FC[2:],
         "vals": {"Bear": [0.10, 0.05, 0.03, 0.02], "Base": [0.25, 0.18, 0.11, 0.09], "Bull": [0.35, 0.25, 0.15, 0.10]},
         "desc": "物理锚: AI 段收入都挂它。2026 三案共用同一个锚(指引已给, 改锚页 G3 三案全动); 分歧在 2027 之后(示意)。",
         "stories": {"Bear": "投资回报证伪, 集体踩刹车(示意)。", "Base": "基线路径(示意)。", "Bull": "推理需求爆发, 几年不减速(示意)。"},
         "hist": [None, 1.00, 1.33, 1.86, 1.44]},
        {"key": "aint", "name": "AI 收入强度", "fmt": K.PCT,
         "vals": {"Bear": [0.035, 0.038, 0.035, 0.032, 0.030], "Base": [0.037, 0.044, 0.042, 0.040, 0.038], "Bull": [0.039, 0.048, 0.047, 0.045, 0.043]},
         "desc": "AI 收入 ÷ AIDC capex = 『AI 段占 capex 份额』×『公司占 AI 段份额』, 锚最近实际年反推(示意)。",
         "stories": {"Bear": "份额被切 + 合约价被压(示意)。", "Base": "份额缓降、单价对冲(示意)。", "Bull": "代际溢价 + 份额守住(示意)。"},
         "hist": [0.029, 0.026, 0.022, 0.040, 0.041]},
        {"key": "cprc", "name": "商品 价格涨幅", "fmt": K.PCT,
         "vals": {"Bear": [1.10, -0.10, -0.25, -0.15, -0.05], "Base": [1.30, 0.10, -0.15, -0.12, -0.05], "Bull": [1.40, 0.30, 0.05, -0.10, -0.10]},
         "desc": "弹性最大的变量; 三案分歧全在『何时见顶』(示意)。",
         "stories": {"Bear": "高价破坏需求 + 新产能放量, 提前转跌(示意)。", "Base": "缺口按物理时点转松(示意)。", "Bull": "缺货拖更久(示意)。"},
         "hist": [0.10, -0.18, -0.37, 0.77, 0.28]},
        {"key": "hopm", "name": "AI 段营业利润率", "fmt": K.PCT,
         "vals": {"Bear": [0.65, 0.55, 0.48, 0.45, 0.42], "Base": [0.72, 0.70, 0.66, 0.62, 0.60], "Bull": [0.75, 0.74, 0.70, 0.66, 0.64]},
         "desc": "定制段利润率结构性高于商品段; 锚最新季报实际(示意)。",
         "stories": {"Bear": "三供成型, 客户反过来压价(示意)。", "Base": "定价权随竞争缓慢稀释(示意)。", "Bull": "绑定更深, 定价权维持更久(示意)。"},
         "hist": [None, None, None, 0.55, 0.62]},
        {"key": "copm", "name": "商品 段营业利润率", "fmt": K.PCT,
         "vals": {"Bear": [0.60, 0.42, 0.30, 0.25, 0.25], "Base": [0.72, 0.66, 0.48, 0.41, 0.40], "Bull": [0.74, 0.70, 0.55, 0.48, 0.45]},
         "desc": "纯周期段, 利润率几乎完全由上面『商品价格』路径决定(示意)。",
         "stories": {"Bear": "价格转跌, 利润率坍回中周期之下(示意)。", "Base": "峰值后回中周期(示意)。", "Bull": "高价多吃 1-2 年(示意)。"},
         "hist": [0.28, 0.10, -0.24, 0.35, 0.42]},
    ],
    "linked": [
        {"key": "sent", "name": "情绪值(倍数第三层)", "fmt": K.N2,
         "src_sheet": S_MULT, "src_row0": ma["sent_row0"],
         "note": "三案取值与依据见『估值倍数假设』(完整三层方法论在那页); 本页只做切换——要改情绪值, 去那页改蓝字。"},
    ],
})
# derived: 目标倍数(当前案) = 峰值 × 溢价 × 当前案情绪 → 喂『情景估值』
_pk = f"'{S_MULT}'!{ma['pk_cell']}"
_pr = f"'{S_MULT}'!{ma['pr_cell']}"
_sent_act = sw["SWACT"]["sent"]
_r = sw["next_row"]
K.lab(wb[S_SW], f"A{_r}", "目标倍数(当前案)", b=True)
for _c in ALLC:
    K.fml(wb[S_SW], f"{_c}{_r}", f"={_pk}*{_pr}*{_c}{_sent_act}", K.MX, link=True)
K.logic(wb[S_SW], f"L{_r}", "= 历史周期峰值 × 结构溢价 × 当前案情绪值 → 喂『情景估值』的前瞻倍数。")
SWPB = _r

# ════════════ 8. 物理锚 [ANCHOR] ════════════
anchor = K.write_anchor(wb.create_sheet(S_ANCHOR), {
    "title": "全球 AI 数据中心 CapEx ($B) — 需求物理盘子(示意)",
    "all_cols": ALLC, "all_years": ALLY,
    "series": [("AI 数据中心 capex ($B)",
                [15, 30, 70, 200, 488, 830, None, None, None, None],
                "历史粗估 + 当年指引锚(三案共用); 2027+ 由情景增速驱动(示意)", K.N0)],
    "yoy_row": "AI 数据中心 capex ($B)",
    "source_note": "口径=全球 AI 数据中心专项 capex。来源: 当年=公司/行业指引(三案共用锚); 前瞻=锚×当前案增速(『情景切换』)(示意)。",
    "role_note": "作用: AI 段收入按『收入 = capex × 收入强度』挂在它上面(见分部测算)。改 capex → 收入 → 估值全链动。",
})
CAPEX_ROW = anchor["row_of"]["AI 数据中心 capex ($B)"]
# 前瞻 2027+ = 上年 × (1 + 当前案增速)
for _i, _c in enumerate(FC[2:]):
    K.fml(wb[S_ANCHOR], f"{_c}{CAPEX_ROW}", f"={FC[1:][_i]}{CAPEX_ROW}*(1+{K.R(S_SW, _c + str(sw['SWACT']['capex']))})", K.N0, link=True)

# ════════════ 9. 分部测算(AI 段=capex×强度 + 商品段=周期)════════════
seg = K.write_segment_model(wb.create_sheet(S_SEG), {
    "title": "分部测算 — AI 段(capex×收入强度, 强度锚最近实际年) + 商品段(周期)(示意)",
    "all_cols": ALLC, "all_years": ALLY, "logic_col": "N",
    "groups": [
        ("AIDC capex 物理锚", [
            ("AIDC capex ($B)", None, K.N0, "= 引自『AIDC Capex预测』物理锚。改 capex, AI 收入跟着动。"),
        ]),
        ("AI 段 = capex × 收入强度", [
            ("AI 收入强度 (%)", None, K.PCT,
             "历史 = 实际 AI 收入 ÷ 当年 capex(公式反推, 锚最近实际年); 前瞻 = 『情景切换』当前案。早年若失真标 n.m.(示意)。"),
            ("AI 收入 ($B)", None, K.N1, "历史取历史财务实数; 前瞻 = capex × 强度。喂『利润与收入假设』。"),
        ]),
        ("商品段 = 周期(上年 × (1+bit) × (1+价))", [
            ("商品 bit增速", [None, 0.10, 0.05, 0.12, 0.15, -0.05, -0.03, 0.06, 0.06, 0.06], K.PCT,
             "缺货年 bit 被压(产能优先 AI + 涨价需求破坏), normalize 后恢复(示意)。"),
            ("商品 价格变化", None, K.PCT, "历史=实际; 前瞻=『情景切换』当前案价格路径(示意)。"),
            ("商品 收入 ($B)", None, K.N1, "历史取实数; 前瞻 = 上年×(1+bit)×(1+价)。不挂 capex, 走周期。"),
        ]),
    ],
})
m = seg["m"]
AI_HROW = ha["seg_rows"]["AI内存 收入"]
CM_HROW = ha["seg_rows"]["商品内存 收入"]
for col in ALLC:
    K.fml(wb[S_SEG], f"{col}{m['AIDC capex ($B)']}", f"={K.R(S_ANCHOR, col + str(CAPEX_ROW))}", K.N0, link=True)
for col in HC:   # 历史强度 = 实数反推; 历史收入 = 实数
    K.fml(wb[S_SEG], f"{col}{m['AI 收入强度 (%)']}", f"={K.R(S_HIST, col + str(AI_HROW))}/{col}{m['AIDC capex ($B)']}", K.PCT, link=True)
    K.fml(wb[S_SEG], f"{col}{m['AI 收入 ($B)']}", f"={K.R(S_HIST, col + str(AI_HROW))}", K.N1, link=True)
    K.fml(wb[S_SEG], f"{col}{m['商品 收入 ($B)']}", f"={K.R(S_HIST, col + str(CM_HROW))}", K.N1, link=True)
K.introw(wb[S_SEG], m["商品 价格变化"], HC, [0.10, -0.18, -0.37, 0.77, 0.28], None, K.PCT)
for col in FCf:  # 前瞻
    K.fml(wb[S_SEG], f"{col}{m['AI 收入强度 (%)']}", f"={K.R(S_SW, col + str(sw['SWACT']['aint']))}", K.PCT, link=True)
    K.fml(wb[S_SEG], f"{col}{m['AI 收入 ($B)']}", f"={col}{m['AIDC capex ($B)']}*{col}{m['AI 收入强度 (%)']}", K.N1)
    K.fml(wb[S_SEG], f"{col}{m['商品 价格变化']}", f"={K.R(S_SW, col + str(sw['SWACT']['cprc']))}", K.PCT, link=True)
_prevs = [HC[-1]] + list(FCf[:-1])
for _p, _c in zip(_prevs, FCf):
    K.fml(wb[S_SEG], f"{_c}{m['商品 收入 ($B)']}",
          f"={_p}{m['商品 收入 ($B)']}*(1+{_c}{m['商品 bit增速']})*(1+{_c}{m['商品 价格变化']})", K.N1)
for col in FCf:
    wb[S_SEG][f"{col}{m['AI 收入 ($B)']}"].fill = K.OUT

# ════════════ 10. 利润与收入假设 ════════════
fr = K.write_fundamentals(wb.create_sheet(S_FUND), {
    "title": "利润与收入假设 — 其他增速 + 段OPM + 净利转换 + 留存 + 分部营收→利润→EPS/BPS(估值倍数在『估值倍数假设』)",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
    "assum_groups": [
        ("营收驱动(AI/商品已在分部测算; 此处仅其他段)", [
            {"name": "其他 增速", "vals": [None, 0.04, -0.15, -0.04, -0.15, 0.08, 0.08, 0.08, 0.08, 0.08],
             "fmt": K.PCT, "logic": "2022-25 = 实际 YoY; 前瞻小体量外推(示意)。"},
        ]),
        ("利润率假设(历史实际锚 + 前瞻; 粗颗粒)", [
            {"name": "AI 营业利润率", "vals": [None, None, None, 0.55, 0.62, None, None, None, None, None],
             "fmt": K.PCT, "logic": "历史=实际/卖方估(早年 n.m.); 前瞻=『情景切换』当前案。", "nm_cols": ["B", "C", "D"],
             "link": {"sheet": S_SW, "row": sw["SWACT"]["hopm"]}},
            {"name": "商品 营业利润率", "vals": [0.28, 0.10, -0.24, 0.35, 0.42, None, None, None, None, None],
             "fmt": K.PCT, "logic": "强周期段, 历史走过深亏; 前瞻=『情景切换』当前案。",
             "link": {"sheet": S_SW, "row": sw["SWACT"]["copm"]}},
            {"name": "其他 营业利润率", "vals": [0.10, 0.08, 0.05, 0.10, 0.10, 0.10, 0.10, 0.10, 0.10, 0.10],
             "fmt": K.PCT, "logic": "小体量, 简单外推(示意)。"},
        ]),
        ("净利转换与留存", [
            {"name": "净利转换率(净利/营业利润)", "vals": [None, None, None, 0.86, 0.90, 0.93, 0.91, 0.89, 0.89, 0.89],
             "fmt": K.PCT, "logic": "营业利润扣税/利息到净利的比例; 历史实际锚(亏损年 n.m.)(示意)。", "nm_cols": ["B", "C", "D"]},
            {"name": "留存率", "vals": [0.91, 0.64, None, 0.94, 0.95, 0.90, 0.90, 0.90, 0.90, 0.90],
             "fmt": K.PCT, "logic": "留存率=1−派息率; 低派息扩产优先(亏损年 n.m.)(示意)。", "nm_cols": ["D"]},
        ]),
    ],
    "segments": [
        {"name": "AI内存 收入", "hist_row": "AI内存 收入", "fwd": {"sheet": S_SEG, "row": m["AI 收入 ($B)"]}},
        {"name": "商品内存 收入", "hist_row": "商品内存 收入", "fwd": {"sheet": S_SEG, "row": m["商品 收入 ($B)"]}},
        {"name": "其他 收入", "hist_row": "其他 收入", "fwd": {"growth": "其他 增速"}},
    ],
    "profit_terms": [
        (["AI内存 收入"], "AI 营业利润率", True),
        (["商品内存 收入"], "商品 营业利润率", False),
        (["其他 收入"], "其他 营业利润率", False),
    ],
    "conv_assum": "净利转换率(净利/营业利润)", "retention_assum": "留存率",
    "note_text": "分部营收(AI 段=capex×强度 + 商品段周期)→ 段驱动营业利润 → 净利(×净利转换)→ 权益(留存递推)→ EPS/BPS/ROE。历史列取实际(引『历史财务与估值』); 下游『情景估值』直接引本表每股, 不重算。",
})

# ════════════ 11. 情景估值 ════════════
sv = K.write_scenario_valuation(wb.create_sheet(S_VAL), {
    "title": "情景估值 — 当前案的逐年隐含价 (P/B 主线; P/E 交叉验证)",
    "intro": "本表输出=『情景切换』当前案(默认 Base)。隐含价 = 目标倍数(当前案) × 前瞻BPS; P/E 仅交叉验证。历史列用实际年末价反推倍数(事实); 前瞻是预测、不拟合现价。三案并排见『估值对比』。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf, "hist_years": HY,
    "fx_fwd": FX_FC, "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
    "s_fund": S_FUND, "fr": fr,
    "s_switch": S_SW, "target_row": SWPB, "sw_cell": "B2",
    "yend": px["yend"], "yavg": px["yavg"],
    "reading": "P/E 交叉验证读法: 『当下 forward P/E』= 现价÷模型各年每股收益(光谱里标的的出处); 『P/E 前瞻』= 隐含价÷前瞻每股收益, 对照同业光谱看是否荒谬; 盈利回落年同价 P/E 跳升 = 低 P/E 陷阱, 这正是 P/E 只作支线的原因。(示意)",
    "method": "方法: 整体公司、P/B 主线逐年估。基本面在『利润与收入假设』; 目标倍数在『估值倍数假设』(三层); 本表只做最后一步: 目标倍数 × 前瞻BPS → 隐含价 + 市值。(示意)",
    "concl": "结论(方向性, 示意): 三情景见『估值对比』; risk-reward 一行收口; 风险列举。",
})

# ════════════ 12. 估值对比(三案恒常并排; 防污染: 只引矩阵行/未翻档行/静态锚)════════════
SWB = sw["SWB"]
SH_F = K.R(S_HIST, f"$F${ha['HSH']}")
PX_NOW_REF = K.R(S_HIST, f"G{ha['HPX']}")
_ogrow = fr["am"]["其他 增速"]
_oopm = fr["am"]["其他 营业利润率"]
_conv = fr["am"]["净利转换率(净利/营业利润)"]
_ret = fr["am"]["留存率"]
_bit = m["商品 bit增速"]


def _fwdprev(j, A, key):
    """对比 block 前瞻列的上一列坐标(j=0 → 基年 F)。"""
    return (HC[-1] if j == 0 else FCf[j - 1]) + str(A[key])


cmp_rows = [
    {"key": "cap", "label": "AIDC capex ($B)", "fmt": K.N0,
     "hist": lambda c, ci, A: f"={K.R(S_ANCHOR, c + str(CAPEX_ROW))}",
     "fwd": lambda c, j, ci, A: (f"={K.R(S_ANCHOR, 'G' + str(CAPEX_ROW))}" if j == 0
                                 else f"={FCf[j-1]}{A['cap']}*(1+{K.R(S_SW, c + str(SWB['capex'] + ci))})")},
    {"key": "ai", "label": "AI 收入 ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(AI_HROW))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['cap']}*{K.R(S_SW, c + str(SWB['aint'] + ci))}"},
    {"key": "cm", "label": "商品 收入 ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(CM_HROW))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'cm')}*(1+{K.R(S_SEG, c + str(_bit))})*(1+{K.R(S_SW, c + str(SWB['cprc'] + ci))})"},
    {"key": "oth", "label": "其他 ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['seg_rows']['其他 收入']))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'oth')}*(1+{K.R(S_FUND, c + str(_ogrow))})"},
    {"key": "rev", "label": "总收入 ($B)", "fmt": K.N1, "bold": True,
     "hist": lambda c, ci, A: f"={c}{A['ai']}+{c}{A['cm']}+{c}{A['oth']}",
     "fwd": lambda c, j, ci, A: f"={c}{A['ai']}+{c}{A['cm']}+{c}{A['oth']}"},
    {"key": "ni", "label": "净利 ($B)", "fmt": K.N0, "bold": True,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HNI']))}",
     "fwd": lambda c, j, ci, A: (f"=({c}{A['ai']}*{K.R(S_SW, c + str(SWB['hopm'] + ci))}"
                                 f"+{c}{A['cm']}*{K.R(S_SW, c + str(SWB['copm'] + ci))}"
                                 f"+{c}{A['oth']}*{K.R(S_FUND, c + str(_oopm))})*{K.R(S_FUND, c + str(_conv))}")},
    {"key": "eq", "label": "期末权益 ($B)", "fmt": K.N0,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HEQ']))}",
     "fwd": lambda c, j, ci, A: (f"={K.R(S_HIST, 'F' + str(ha['HEQ']))}+{c}{A['ni']}*{K.R(S_FUND, c + str(_ret))}" if j == 0
                                 else f"={FCf[j-1]}{A['eq']}+{c}{A['ni']}*{K.R(S_FUND, c + str(_ret))}")},
    {"key": "bps", "label": "BPS ($)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={c}{A['eq']}*1000/{K.R(S_HIST, c + str(ha['HSH']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['eq']}*1000/{SH_F}"},
    {"key": "sent", "label": "情绪值(该案; 历史=实际反推)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_MULT, c + str(ma['sent_row0'] + ci))}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_MULT, c + str(ma['sent_row0'] + ci))}"},
    {"key": "pb", "label": "目标倍数(历史=实际)", "fmt": K.MX,
     "hist": lambda c, ci, A: f"={_pk}*{_pr}*{c}{A['sent']}",
     "fwd": lambda c, j, ci, A: f"={_pk}*{_pr}*{c}{A['sent']}"},
    {"key": "px", "label": "隐含价 (本币)", "fmt": K.PX, "bold": True, "out": True,
     "hist": lambda c, ci, A: f"={c}{A['pb']}*{c}{A['bps']}*{K.R(S_HIST, c + str(ha['HFX']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['pb']}*{c}{A['bps']}*{FX_FC}"},
    {"key": "ipe", "label": "隐含 forward P/E(交叉验证)", "fmt": K.MX,
     "hist": lambda c, ci, A: f'=IF({c}{A["ni"]}<=0,"N/M",{c}{A["px"]}/({c}{A["ni"]}*1000/{K.R(S_HIST, c + str(ha["HSH"]))}*{K.R(S_HIST, c + str(ha["HFX"]))}))',
     "fwd": lambda c, j, ci, A: f"={c}{A['px']}/({c}{A['ni']}*1000/{SH_F}*{FX_FC})"},
    {"key": "up", "label": "历史: vs 实际年末价(回测≈0) / 前瞻: vs 现价", "fmt": K.PCT,
     "hist": lambda c, ci, A: f"={c}{A['px']}/{K.R(S_HIST, c + str(ha['HPX']))}-1",
     "fwd": lambda c, j, ci, A: f"={c}{A['px']}/{PX_NOW_REF}-1"},
]
cm = K.write_comparison(wb.create_sheet(S_CMP), {
    "title": "估值对比 — Bear / Base / Bull 三个情景的目标价并排对比",
    "intro": ("三个情景各自完整推演一遍: 物理锚 → 分部收入 → 净利 → 每股净资产 → 目标倍数 → 逐年隐含价。"
              "本表三案永远并排可见, 不随『情景切换』开关变化; 要调假设去『情景切换』改对应案参数。"
              "未列入情景矩阵的假设三案共用 Base 取值。历史列 2021-2025 = 同一条链填实际值: 隐含价历史列应基本等于实际年末价"
              "(回测行历史列 ≈0%)——这是整条估值链的内置回测。"),
    "case_names": CASES,
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "block_start": 16,
    "rows": cmp_rows,
    "summary": {
        "band": "三案汇总 (目标年 2027E; 各案触发条件见『情景切换』)",
        "target_col": "H",
        "rows": [
            ("净利($B)", "ni", K.N0, "由该案的 capex / 价格 / 利润率假设逐年推导得出", False),
            ("总收入($B)", "rev", K.N0, "= AI 段(capex × 收入强度) + 商品段(该案价格路径) + 其他", False),
            ("目标倍数", "pb", K.MX, "= 历史周期峰值 × 结构溢价 × 该案情绪值", False),
            ("隐含价(本币)", "px", K.PX, "= 目标倍数 × 2027E 每股净资产 × 汇率", True),
            ("隐含 forward P/E", "ipe", K.MX, "P/E 体检(任何情景都做): 该案隐含价 ÷ 该案每股收益, 对照『历史估值倍数』同业光谱", False),
            ("vs 现价", "up", K.PCT, "对照现价的上行 / 下行空间", True),
        ],
        "mcap": {"label": "隐含市值($B)", "key": "px", "expr": f"*{SH_F}/{FX_FC}/1000",
                 "note": "= 隐含价 × 股本, 折回美元"},
        "concl": "风险收益比(目标年 vs 现价)一行收口: 下行 X% / 上行 Y%, 偏向哪边、成立前提是什么; 衰减信号触发应转向 Bear 重看。(示意)",
    },
})

# ════════════ 13. 综合判断仪表盘 ════════════
BPS26 = K.R(S_FUND, "G" + str(fr["BPS"]))
PXD = K.R(S_HIST, "G" + str(ha["HPX"]))
dash = K.write_dashboard(wb.create_sheet(S_DASH), {
    "title": "综合判断仪表盘 — A 基本面拐点 · B 估值错位(预测引擎) · C 催化剂 · D 情绪确认",
    "usage": ("怎么用: 预测引擎是 B(错位)+ C(催化剂)——两样都当下可观测; 情绪 D 只做 timing 确认 + 过热刹车(定性档位)。"
              "验收=回测: 放回上一轮行情拐点, 这套表当时就能看到那波。(示意)"),
    "blocks": [
        {"title": "A. 基本面拐点 — 业务在结构性变好吗?", "rows": [
            ("产品组合迁移", "AI 段占营收逐年升(示意)", "结构迁移、不是周期反弹(示意)。"),
            ("可持续 ROE", "周期常态 → franchise 拉高(示意)", "盈利质量升级还是周期顶虚高(示意)。"),
            ("A 判断", "【强/中/弱】", "一句话定性(示意)。", True),
        ]},
        {"title": "B. 估值错位(预测引擎 ★)— 市场现在给的 vs 基本面该给的 → GAP", "rows": [
            ("市场现在给(forward 倍数)", {"fml": f"={PXD}/({BPS26}*{FX_FC})", "fmt": K.MX, "fill": True},
             "= 现价 ÷ 前瞻每股净资产 × 汇率(公式算, 随模型走)。"),
            ("基本面该给(justified)", {"inp": 3.2, "fmt": K.MX},
             "=(ROE−g)/(COE−g)(示意取值)。注: 目标倍数 = 该给 × 情绪溢价; 本块 GAP 衡量纯基本面错位。"),
            ("错位 GAP = 该给÷市场给 − 1",
             {"fml": lambda ro: f"=B{ro['基本面该给(justified)']}/B{ro['市场现在给(forward 倍数)']}-1", "fmt": K.PCT},
             "GAP 正且大 = 重估空间(该买); 转负 = 基本面空间已被价格吃完, 进入纯情绪定价区。"),
            ("回测: 上一轮拐点的读数", "(示意)", "当时市场给 vs 该给 → 错位多少, 这就是那波的预测依据。"),
        ]},
        {"title": "C. 催化剂 — 什么会逼市场闭合 GAP", "rows": [
            ("催化剂 1", "✓/进行中/待(示意)", "为什么重要(示意)。"),
            ("催化剂 2", "✓/进行中/待(示意)", "为什么重要(示意)。"),
            ("C 判断", "已兑现/待兑现", "剩余催化剂决定超涨还能不能撑(示意)。", True),
        ]},
        {"title": "D. 情绪确认 — 只做 timing + 刹车(定性档位; 不精确量化)", "rows": [
            ("量价温度计(只当温度计, 不进倍数)", "(示意读数)", "量价拟合进目标倍数=循环论证, 只用来看与基本面的背离。"),
            ("现价倍数 vs 基本面该给", "(示意)", "市场已付出超过基本面该给 = 进入情绪定价区。"),
            ("当前档位", "【萌芽/启动/过热/退潮】", "档位读法 + 对照上一轮拐点(示意)。", True),
            ("衰减扳机", "N 条", "逐条列最强先行信号; 任一翻 → 档位降档 + 情绪值下调(示意)。"),
        ]},
    ],
    "final": {"band": "★ 综合判断(A+B+C+D 收成一句可执行的话)",
              "text": "回测上一轮拐点: 当时这套表的读数 → 该看多 ✓。当下: 各块读数 → 一句可执行结论(买/持/收)。(示意)"},
    "tracking": {
        "intro": "哪个指标恶化 → 哪个假设先崩 → 触发什么动作(盯的优先级)(示意)。",
        "rows": [
            ("__band__", "一、核心驱动链"),
            ("AI 收入强度", "(示意)", "关键敏感项: AI 收入 = capex × 强度", "季报/行业 tracker(示意)", "强度回落 → 下调假设重算(示意)"),
            ("__band__", "二、商品周期"),
            ("商品合约价 + 库存", "(示意)", "关键敏感项: 价格路径与 normalize 时点", "月度合约价 tracker(示意)", "连续转跌 → 前移见顶年(示意)"),
            ("__band__", "三、需求总盘子"),
            ("AI capex 指引", "(示意)", "关键敏感项: 物理锚盘子", "hyperscaler 季报(示意)", "下修 >10% → 全链重算(示意)"),
        ],
    },
})

# ════════════ 全局格式 + 落盘 ════════════
K.finalize(wb, freeze={
    S_HIST: "B3", S_PX: "B4", S_CONS: "A2", S_HMULT: "B5", S_MULT: "B4", S_SW: "B3",
    S_ANCHOR: "B3", S_SEG: "B3", S_FUND: "B3", S_VAL: "B4", S_CMP: "B6", S_DASH: "B6",
    S_COVER: "A2",
})
_repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_output_dir = os.environ.get("VALUATION_OUTPUT_DIR", os.path.join(_repo_root, "out"))
out = os.path.join(_output_dir, "_template_demo.xlsx")
os.makedirs(_output_dir, exist_ok=True)
wb.save(out)
print("saved:", os.path.abspath(out))
print("sheets:", wb.sheetnames)
