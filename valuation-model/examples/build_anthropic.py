# -*- coding: utf-8 -*-
"""
build_anthropic.py — Anthropic 估值模型, 用 build_kit 引擎建为机构级 13-sheet 架构。

与 build_spcx 同属"私营 + 亏损刚转正 → EV/Sales 主镜头"范式, 但更进一步:
  Anthropic 完全私营(连 IPO 现价都没有)→ 估值口径是【整体 EV / $B】, 不是 per-share。
  "现锚" = Series H post-money $965B(2026-05); "隐含价" = 隐含整体估值($B), 直接对 $965B。
  无股本 / 无 fx / 无期权 / 无净现金, 比 spcx 更干净:隐含 EV = 收入 × EV/Sales。

估值方法(三镜头, 沿用 v3.1 物理锚模型的结论):
  主镜头 EV/Sales — 整体亏损刚转正, EV/EBITDA·P/E 主线会出负/NA → 用销售倍数出实数。
    逐年(2026E-2030E)forward: 隐含 EV(Y) = 总收入(Y) × 目标 EV/Sales(Y), 高增长期高、随成熟 normalize。
  交叉验证① P/E — 2030E 稳态净利 × 目标 P/E(28-30x); 亏损年标 N/M。
  交叉验证② DCF — UFCF 折现(WACC 12% / g 4%)≈ $575B, sanity check。
  三镜头加权目标 ≈ $883B(EV/Sales $1044B×0.4 + PE $975B×0.3 + DCF $575B×0.3)= 现锚 −8.5% → HOLD。

  通用 sheet 复用 kit: write_cover / write_price_chart / write_consensus
  / write_scenario_switch / write_anchor / write_segment_model / write_dashboard / finalize。
  历史财务 / 历史估值倍数 / 估值倍数假设 / 利润 / 情景估值 / 估值对比 六张按整体 EV/$B
  口径用 kit 低层 helper 手搓(私营无股价 → 不能用 write_history 的 per-share/PE/PB)。

跑:   cd examples && PYTHONUTF8=1 python build_anthropic.py
缓存: python ../scripts/recalc.py --backend auto <out>   (跨平台重算回写缓存值)
校验: python ../scripts/validate_valuation.py <out>   → verdict ≠ FAIL 方可交付
"""
import os
from openpyxl import Workbook
import build_kit as K

# ════════════ 0. 全局轴(美元单一币种, $B 整体口径; 无 fx / 无 per-share)════════════
ALLC = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
ALLY = ["2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
HC,  HY = ["B", "C", "D", "E", "F"], ["2021", "2022", "2023", "2024", "2025"]
FC      = ["F", "G", "H", "I", "J", "K"]    # 前瞻含基年 F=2025A
FCf     = FC[1:]                             # 纯前瞻 G-K = 2026E-2030E
FWY     = ["2026E", "2027E", "2028E", "2029E", "2030E"]
CASES   = ["Bear", "Base", "Bull"]

ANCHOR_EV = 965.0          # Series H post-money(现锚, $B)
IPO_MID   = 1100.0         # 媒体 IPO 首日中值预期

# ── 三案核心数据(v3.1 物理锚模型, 列序 = 2026E,2027E,2028E,2029E,2030E)────────
# token 量 YoY 增速(2027E-2030E; 2026E 共用锚 14.1 Q): 校准到三案 2027E 收入 70/110/137
TOK_G   = {"Bear": [1.20, 0.50, 0.35, 0.28], "Base": [2.25, 0.85, 0.49, 0.41], "Bull": [3.05, 1.00, 0.55, 0.45]}
# 目标 EV/Sales 逐年(高增长期高、随成熟 normalize): 校准 Base 2027E=9.5x → $1045B(v3.1 EV/Sales 镜头)
EVS     = {"Bear": [9, 8, 5.5, 4.2, 3.5], "Base": [14, 9.5, 6.5, 5.0, 4.2], "Bull": [22, 16, 10, 7, 5.5]}
# compute/收入比(关键 margin 驱动; Bear 价格战卡高位→稳态 OM 个位数)
CMP_R   = {"Bear": [0.66, 0.60, 0.58, 0.56, 0.60], "Base": [0.64, 0.58, 0.54, 0.51, 0.48], "Bull": [0.62, 0.54, 0.50, 0.46, 0.42]}
# 目标 P/E(2030E 稳态净利镜头)
PE_TGT  = {"Bear": 28, "Base": 29, "Bull": 30}
WEIGHTS = {"Bear": 0.20, "Base": 0.55, "Bull": 0.25}
# 三镜头加权目标($B, 各案; Base=EV/Sales 1044×.4 + PE 975×.3 + DCF 575×.3 ≈ 883)
TRILENS = {"Bear": 560, "Base": 883, "Bull": 1700}
DCF_EV  = 575.0           # DCF sanity check(WACC 12% / g 4%)

S_COVER, S_HIST, S_PX, S_CONS = "封面", "历史财务与估值", "股价走势", "卖方研报共识"
S_HMULT, S_MULT, S_SW = "历史估值倍数", "估值倍数假设", "情景切换"
S_ANCHOR, S_SEG, S_FUND = "物理锚", "分部测算", "利润与收入假设"
S_VAL, S_CMP, S_DASH = "情景估值", "估值对比", "综合判断仪表盘"

wb = Workbook()
wb.remove(wb["Sheet"])

# ════════════ 1. 封面 ════════════
K.write_cover(wb.create_sheet(S_COVER), {
    "title": "Anthropic (Private) 估值模型 — 物理锚 token 经济 × 整体 EV/Sales 三镜头",
    "span": 6,
    "meta": [
        ("报告日期", "2026-06-30"),
        ("数据截止", "Series H 融资公告(2026-05-28)+ S-1 递交(2026-06-01)+ 卖方共识(HSBC 2026-06-09 / Barclays 2026-03-11)"),
        ("现锚(整体估值)", f"${ANCHOR_EV:,.0f}B — Series H post-money(2026-05-28, 全球估值最高私营公司)"),
        ("口径说明", "完全私营、无公开股价 → 全程整体 EV/$B 口径, 不折每股; 历史'股价'= 历次融资轮 post-money 估值带; "
                     "对外 run-rate(当月×12)≠ GAAP 全年, 本模型走 GAAP。"),
        ("时效声明", "S-1 公开版(预计 2026-Q3)将首次披露审计 GAAP 收入 / gross-net 确认 / 算力承诺表, 是口径重锚日; 建议届时全面复核。"),
        ("方法一句话", "物理锚 = token 量 × 有效混合单价 → GAAP 收入 → compute/收入比驱动利润; "
                       "整体亏损刚转正 → 主镜头 EV/Sales(逐年 forward 出实数), PE(2030E 稳态)+ DCF 交叉验证 → 隐含整体估值 vs 现锚 $965B。"),
    ],
    "takeaways": [
        ("① 评级 / 目标估值", "HOLD(持有), 三镜头加权目标 ≈ $883B — 较现锚 $965B 约 −8.5%; 用更尊重商品化的稳态 margin, 当前一级定价 Base 下略偏贵。"),
        ("② 主镜头 EV/Sales(Base)", "隐含 EV 2026E≈$630B → 2027E≈$1,045B → 2030E≈$1,008B; 仅 EV/Sales 单镜头偏高, 被 DCF($575B)拉回到加权 $883B。"),
        ("③ 两个加权口径的分歧 = 押注 margin", "三镜头加权 $883B(−8.5%)vs 情景概率加权 $1,024B(+6%); 分歧本身 = 对'稳态 EBIT 能否高位'的押注大小(Base 稳态仅 ~17%)。"),
        ("④ 收支平衡的三个时点", "单季经调整经营利润 2026-Q2 已现(橱窗、不持续); 年度会计盈亏平衡 ~2027(净利 −$2.4B→+$4.6B); 正 FCF / 覆盖现金投入 ~2029(公司内部口径)。"),
        ("⑤ 主要风险(双向)", "下行: S-1 gross→net 口径折扣 15-27% / 价格战实质化 / GPT-6 反超 / 企业份额腰斩(Bear $560B)。上行: agent 垂直复制 + IPO 流动性溢价(Bull $1,700B)。"),
    ],
})

# ════════════ 2. 历史财务与估值(私营 EV/$B 口径, 手搓; 无 per-share)════════════
wh = wb.create_sheet(S_HIST)
K.hdr(wh, 1, "Anthropic 历史财务与估值 ($B) — 2021(成立)→2025A + 当下锚", 9)
hrr = K.mtext(wh, 2, ("私营公司、无公开股价 → 用整体 $B 口径, 不算 per-share/PE/PB。历史'股价'= 历次融资轮 post-money 估值"
                      "(见『股价走势』), 配当年 run-rate 得 EV/run-rate 估值带(de-rating)。三阶段: 烧钱换模型(2021-24)→ 企业 API 起量(2025)→ agent 变现(2026-)。"), "I", 2)
for col, w in zip("ABCDEFGHI", [22, 9, 9, 9, 9, 9, 11, 11, 11]):
    wh.column_dimensions[col].width = w
wh.column_dimensions["I"].width = 58
# 年份头
K.lab(wh, f"A{hrr}", "($B)", b=True)
for col, y in zip(HC, HY):
    wh[f"{col}{hrr}"] = y; wh[f"{col}{hrr}"].font = K.BF; wh[f"{col}{hrr}"].fill = K.CH
wh["G" + str(hrr)] = "当下(2026-06)"; wh[f"G{hrr}"].font = K.BF; wh[f"G{hrr}"].fill = K.CUR
wh["I" + str(hrr)] = "备注/来源"; wh[f"I{hrr}"].font = K.BF; wh[f"I{hrr}"].fill = K.CH
hrr += 1

K.band(wh, hrr, "经营(GAAP 确认, 私营期为媒体/卖方拼接估计, 非审计)", 9); hrr += 1
HREV = hrr
K.lab(wh, f"A{hrr}", "GAAP 营收", b=True)
K.introw(wh, hrr, HC, [0.0, 0.01, 0.1, 0.4, 4.5], None, K.N1)
K.logic(wh, f"I{hrr}", "GAAP 确认收入(非 run-rate); 2025 $4.5B 与物理锚 2025($4.98B)同量级、互为校验。Q1'26 单季已 $4.8B≈2025 全年。")
hrr += 1
K.lab(wh, f"A{hrr}", "  YoY", note=True)
for j in range(1, len(HC)):
    K.fml(wh, f"{HC[j]}{hrr}", f"=IFERROR({HC[j]}{HREV}/{HC[j-1]}{HREV}-1,\"n.m.\")", K.PCT)
hrr += 1
HRR = hrr
K.lab(wh, f"A{hrr}", "年末 run-rate(当月×12)")
K.introw(wh, hrr, HC, [0.0, 0.03, 0.3, 1.0, 9.0], None, K.N1)
K.logic(wh, f"I{hrr}", "公司对外口径 = run-rate(瞬时年化); 高增长下全年 GAAP ≪ 年末 exit run-rate(2025: GAAP $4.5B vs 年末 run-rate $9B)。")
hrr += 1
HNI = hrr
K.lab(wh, f"A{hrr}", "净亏损/利润", b=True)
K.introw(wh, hrr, HC, [-0.1, -0.5, -1.5, -2.7, -3.5], None, K.N1)
K.logic(wh, f"I{hrr}", "私营期为估; 烧钱换模型阶段算力开支是收入数倍, 靠 Amazon/Google 战略注资续命。")
hrr += 1
HCF = hrr
K.lab(wh, f"A{hrr}", "累计股权融资")
K.introw(wh, hrr, HC, [0.12, 0.7, 6.4, 14.4, 30.0], None, K.N1)
K.fml(wh, f"G{hrr}", "=95", K.N0); wh[f"G{hrr}"].fill = K.CUR
K.logic(wh, f"I{hrr}", "累计股权融资($B); 三大云(Amazon/Google/Microsoft)+ NVIDIA 全在股东名单。当下含 Series G+H ≈ $95B+。")
hrr += 1

K.band(wh, hrr, "估值锚(私营=融资轮 post-money; EV/run-rate 估值带 de-rating)", 9); hrr += 1
HPM = hrr
K.lab(wh, f"A{hrr}", "年末 post-money 估值", b=True)
K.introw(wh, hrr, HC, [0.85, 4.0, 18.0, 60.0, 183.0], None, K.N0)
K.inp(wh, f"G{hrr}", ANCHOR_EV, None, K.N0); wh[f"G{hrr}"].fill = K.CUR
K.logic(wh, f"I{hrr}", "历次融资轮 post-money 当'市值'时间序列(A $0.85B→H $965B); 当下 = Series H post-money $965B(现锚)。详见『股价走势』。")
hrr += 1
HEVR = hrr
K.lab(wh, f"A{hrr}", "EV / run-rate (x, 实际)", b=True); wh[f"A{hrr}"].fill = K.OUT
for col in HC:
    K.fml(wh, f"{col}{hrr}", f"=IFERROR({col}{HPM}/{col}{HRR},\"n.m.\")", K.MX)
K.fml(wh, f"G{hrr}", f"=G{HPM}/47", K.MX); wh[f"G{hrr}"].fill = K.CUR
K.logic(wh, f"I{hrr}", "= post-money ÷ 当年 run-rate; Series E ~41x → H 20.5x(规模越大倍数越低, de-rating); 当下 = $965B ÷ run-rate $47B = 20.5x。早年 run-rate≈0 标 n.m.。")
hrr += 1
K.mtext(wh, hrr, ("读法: EV/run-rate 从 Series E 41x 压到 H 20.5x, 符合上市公司'增速换挡→倍数收敛'; 当下 20.5x 仍高于成熟 SaaS"
                  "(EV/fwd-Rev 12-15x)、低于 OpenAI run-rate 28x。run-rate 是瞬时年化、非 GAAP, 仅作估值带刻度; 主估值走 GAAP 收入 × EV/Sales(后页)。"), "I", 3)

# ════════════ 3. 股价走势(融资轮 post-money 序列当'股价')════════════
ROUNDS = [("2021-05", 0.85), ("2022-04", 4), ("2023-05", 18), ("2023-09", 25), ("2024-11", 60),
          ("2025-03", 61.5), ("2025-09", 183), ("2026-02", 380), ("2026-05", 965)]
def round_phase(ym):
    if ym <= "2022-12":
        return "① 种子/早期(估值估)"
    if ym <= "2024-12":
        return "② Amazon/Google 战略注资"
    if ym <= "2025-09":
        return "③ 企业 API 起量"
    return "④ agent 变现 + IPO 冲刺"
px = K.write_price_chart(wb.create_sheet(S_PX), ROUNDS, {
    "fn": round_phase,
    "rows": [("① 种子/早期", "2021 A $0.85B → 2022 B $4B(FTX/SBF 领投, 估值估)。"),
             ("② 战略注资", "2023-09 Amazon 初投 $25B → 2024-11 追加至 $60B(累计 $8B); Google 跟投。"),
             ("③ 企业 API 起量", "2025-03 E $61.5B → 2025-09 F $183B; Bedrock 分销 + Sonnet 3.7 反超。"),
             ("④ agent 变现 + IPO", "2026-02 G $380B → 2026-05 H $965B(超 OpenAI 成全球估值最高私营公司)→ 06-01 递 S-1。")],
}, title="Anthropic 融资轮 post-money ($B) — 私营'股价'时间序列")

# ════════════ 4. 卖方研报共识 ════════════
K.write_consensus(wb.create_sheet(S_CONS), {
    "title": "卖方研报共识 — 覆盖 Anthropic 的卖方稀少(私营), 仅行业报告附带建模",
    "overview": ("一级市场最聪明的钱(Sequoia/Altimeter + 三大云)尽调后 $965B 进场, 员工拒在 $350B tender 卖股 → 内部信息优势方在多头一侧。"
                 "卖方在此标的系统性落后于现实: Barclays 3 月用公司 1 月指引($18B)两个月即被 run-rate 跑穿; HSBC(最新)反而比本模型激进。"),
    "assumptions": [
        ("2026E GAAP 收入",
         "HSBC B2B $54.7B(不含消费); Barclays $17.5B(3 月旧估, 已被现实碾过)。",
         "run-rate vs GAAP 口径差: 全年 GAAP ≪ 年末 exit run-rate。",
         "本模型 $45B(GAAP, 物理锚): Q1 4.8+Q2 10.9+H2 中性建桥; 另列 run-rate $47B 作 reality check。"),
        ("2027E GAAP 收入",
         "HSBC B2B $129.7B(上沿, 隐含通吃 CoPilot 工具市场一半); Barclays $50.5B(下沿, 旧估)。",
         "增速换挡幅度 + 工具层接力速度。",
         "本模型 $110B: 取 HSBC/Barclays 中偏保守, 显著偏 HSBC 一侧(同意'企业 AI 第一名')。"),
        ("现金流转正时点",
         "Barclays: 2030 年前持续为负(最悲观); HSBC 未给。",
         "$1T+ 算力 take-or-pay 算经营流出还是融资 capex, 决定 2027 还是 2029。",
         "本模型: 年度会计盈亏 2027(净利转正); 正 FCF/覆盖现金投入 2029(公司内部口径)。"),
        ("稳态 margin / 目标倍数",
         "卖方多用 EV/Revenue; frontier lab 组 run-rate 20-46x。",
         "稳态 EBIT 能否高位(定价权 vs 商品化)= 最大分歧。",
         "本模型稳态 EBIT ~17%(尊重商品化, 三巨头 token 份额 72%→33%); EV/Sales 2027E 9.5x, 见『估值倍数假设』。"),
    ],
    "divergences": [
        "① gross vs net 收入确认: S-1 若改 net, top-line 缩水 15-27% — 隐含估值最大重锚风险(OpenAI memo 指 Anthropic 高估 $8B)。",
        "② 稳态 margin: compute/收入比能否从 64% 降到 48%(Base)还是卡 60%(Bear)→ 决定 PE 镜头 $975B vs 失效。",
        "③ 估值口径: 一级 DCF 中性价($575B)vs IPO 流动性溢价(媒体首日中值 $1,100B), 重定价窗口本身是收益来源。",
    ],
    "stances": [
        "HSBC(2026-06-09, B2B up B2C down): Anthropic 2030E LLM 份额 52%、OpenAI 17%; 2027E B2B $129.7B(最激进)。",
        "Barclays(2026-03-11, AI Demand & Supply 框架): compute 开支 2030E $95B/年; 2030 年前现金流持续为负。",
        "UBS(2026-05-11, Enterprise AI Adoption): Claude Code 企业渗透 39%(半年前 24%), 通用 AI 工具 28%→50%。",
        "本模型 Base: 加权目标 $883B(HOLD), 落 HSBC 与 Barclays 之间偏 HSBC; 不同意一级只按 DCF 中性价定价。",
    ],
})

# ════════════ 5. 历史估值倍数(EV/run-rate 带 + 同业 EV 倍数光谱, 手搓)════════════
wm = wb.create_sheet(S_HMULT)
K.hdr(wm, 1, "历史估值倍数 — 融资轮 EV/run-rate 估值带 + 同业 EV 倍数对照(无公开股价)", 8)
mr = K.mtext(wm, 2, ("Anthropic 私营、无自身公开历史估值带 → 本页两件事: ① 历次融资轮 EV/run-rate 当'估值带'(de-rating 轨迹); "
                     "② 同业 EV 倍数光谱(frontier lab run-rate 倍数组 vs 成熟 SaaS forward-GAAP 倍数组, 两组不可跨比)。"
                     "主估值锚 2027E GAAP 收入 × EV/Sales(后页), 不锚 run-rate 倍数(瞬时年化失真)。"), "H", 2)
for col, w in zip("ABCDEFGH", [22, 11, 11, 11, 11, 11, 13, 13]):
    wm.column_dimensions[col].width = w
wm.column_dimensions["H"].width = 56

K.band(wm, mr, "① 融资轮 EV/run-rate 估值带(私营'估值带'; 规模越大倍数越低)", 8); mr += 1
for col, h in zip(["A", "B", "C", "D"], ["轮次/时点", "post-money$B", "当时run-rate$B", "EV/run-rate"]):
    wm[f"{col}{mr}"] = h; wm[f"{col}{mr}"].font = K.BF; wm[f"{col}{mr}"].fill = K.CH
wm.merge_cells(f"E{mr}:H{mr}"); wm[f"E{mr}"] = "备注"; wm[f"E{mr}"].font = K.BF; wm[f"E{mr}"].fill = K.CH
mr += 1
band_rows = [
    ("2023-05 Series C", 18, 0.15, "Google 首投 $300M; 估值跳升"),
    ("2025-03 Series E", 61.5, 1.5, "Lightspeed 领投"),
    ("2025-09 Series F", 183, 6, "ICONIQ 领投"),
    ("2026-02 Series G", 380, 14, "GIC+Coatue($350B term 上调)"),
    ("2026-05 Series H", 965, 47, "当前锚; 全球估值最高私营公司"),
]
for nm, pm, rr, note in band_rows:
    wm[f"A{mr}"] = nm; wm[f"A{mr}"].font = K.BF
    K.inp(wm, f"B{mr}", pm, None, K.N0)
    K.inp(wm, f"C{mr}", rr, None, K.N1)
    K.fml(wm, f"D{mr}", f"=B{mr}/C{mr}", K.MX)
    wm.merge_cells(f"E{mr}:H{mr}"); K.logic(wm, f"E{mr}", note)
    mr += 1
wm[f"A{mr}"] = "de-rating 读法"; wm[f"A{mr}"].font = K.BF; wm[f"A{mr}"].fill = K.OUT
wm.merge_cells(f"B{mr}:H{mr}")
K.logic(wm, f"B{mr}", "EV/run-rate 从 Series C 120x → E 41x → H 20.5x, 规模扩大倍数收敛(典型 de-rating); 当下 20.5x 高于成熟 SaaS(12-15x fwd)、低于 OpenAI(28x run-rate)。")
mr += 2

K.band(wm, mr, "② 同业 EV 倍数光谱(分两组, 不可跨组比大小)", 8); mr += 1
for col, h in zip(["A", "B", "C", "D"], ["公司/口径", "EV/run-rate", "EV/fwd-GAAP-Rev", "业务特征 / 取数"]):
    wm[f"{col}{mr}"] = h; wm[f"{col}{mr}"].font = K.BF; wm[f"{col}{mr}"].fill = K.CH
wm.merge_cells(f"D{mr}:H{mr}"); mr += 1
peers = [
    ("Anthropic(本案)", 20.5, 8.8, "现锚 $965B; run-rate $47B(20.5x)/ 2027E GAAP $110B(8.8x); 组内最便宜+最快"),
    ("OpenAI", 28.4, None, "run-rate $30B / $852B; 已递 S-1; 增速对方数倍但倍数更高"),
    ("xAI", 46.0, None, "run-rate ~$5B / ~$230B; 媒体口径波动大"),
    ("Palantir", None, 91.3, "上市 AI-SaaS 情绪锚; fwd-GAAP 口径 +40% 增速"),
    ("Snowflake / Datadog", None, 15.5, "成熟 SaaS 下沿 14-17x fwd-GAAP; +25% 增速"),
    ("NVIDIA", None, 18.8, "硬件层对照; fwd-GAAP ~18.8x / +55%"),
    ("Microsoft", None, 12.7, "hyperscaler 对照; fwd-GAAP ~12.7x / +15%"),
]
for nm, rr, fg, note in peers:
    wm[f"A{mr}"] = nm; wm[f"A{mr}"].font = K.BF
    if nm.startswith("Anthropic"):
        wm[f"A{mr}"].fill = K.CUR
    for col, v in zip("BC", [rr, fg]):
        if v is not None:
            K.inp(wm, f"{col}{mr}", v, None, K.MX)
        else:
            K.lab(wm, f"{col}{mr}", "—", note=True)
    wm.merge_cells(f"D{mr}:H{mr}"); K.logic(wm, f"D{mr}", note)
    mr += 1
mr += 1
K.band(wm, mr, "③ 读法 — 给『估值倍数假设』的输入", 8); mr += 1
K.mtext(wm, mr, ("A 组 run-rate 倍数(20.5x/28x)与 B 组 fwd-GAAP 倍数(8.8x)数量级不同, 禁跨组比贵贱。"
                 "Anthropic 在 B 组(fwd-GAAP)对成熟 SaaS(SNOW/DDOG 14-17x)溢价有限、但增速远高 → 主镜头取 EV/2027E GAAP 收入, "
                 "高增长期给高倍数、随增速换挡 normalize。下一页拍三案逐年 EV/Sales。"), "H", 3)

# ════════════ 6. 估值倍数假设(EV/Sales 三案逐年矩阵 + PE 目标, 手搓 spcx 模式)════════════
wu = wb.create_sheet(S_MULT)
K.hdr(wu, 1, "估值倍数假设 — 整体 EV/Sales 三案逐年(2026E-2030E)+ PE 目标(2030E 稳态)", 9)
ur = K.mtext(wu, 2, ("主镜头=EV/Sales(整体亏损刚转正, EV/EBITDA·P/E 主线会出负/NA, 故用销售倍数出实数)。"
                     "本页拍逐年 EV/Sales 三案(蓝字)+ PE 目标(2030E 净利镜头, 交叉验证); 『情景切换』link 并切换, 『情景估值』套用。"
                     "EV/Sales 据 B 组同业 fwd-GAAP 光谱 + 高增长期高、随增速换挡 normalize; Base 2027E 9.5x → $1,045B(= v3.1 EV/Sales 镜头)。"), "I", 2)
for col, w in zip("ABCDEFGHI", [22, 10, 10, 10, 10, 10, 11, 11, 11]):
    wu.column_dimensions[col].width = w
wu.column_dimensions["L"].width = 62

K.band(wu, ur, "整体 EV/Sales 逐年(x) — 三案", 9); ur += 1
wu[f"A{ur}"] = "案 / 年份"; wu[f"A{ur}"].font = K.BF
for col, y in zip(FCf, FWY):
    wu[f"{col}{ur}"] = y; wu[f"{col}{ur}"].font = K.BF; wu[f"{col}{ur}"].fill = K.CH
wu[f"L{ur}"] = "为什么这么给"; wu[f"L{ur}"].font = K.BF; wu[f"L{ur}"].fill = K.CH
ur += 1
EVS_ROWS = {}
evs_notes = {
    "Bear": "价格战 + 份额腰斩 → 倍数快速收敛: 9x 起步 normalize 到 3.5x(向成熟 SaaS 下沿之下, 反映增长被打断)。",
    "Base": "高增长换挡: 2026E 14x(刚 5 倍增长)→ 2027E 9.5x(对账 v3.1 EV/Sales 镜头 $1,045B)→ 2030E 4.2x(向成熟 SaaS 收敛, 增速降到 +19%)。",
    "Bull": "agent 垂直复制 + 定价权守住: 22x 起步, normalize 到 5.5x(全程高于 Base, 反映工具层 mix 全胜)。",
}
for cs in CASES:
    K.lab(wu, f"A{ur}", f"  {cs}")
    K.introw(wu, ur, FCf, EVS[cs], None, K.MX)
    K.logic(wu, f"L{ur}", evs_notes[cs])
    EVS_ROWS[cs] = ur
    ur += 1
ur += 1
K.band(wu, ur, "PE 目标(2030E 稳态净利镜头, 交叉验证)", 9); ur += 1
wu[f"A{ur}"] = "案"; wu[f"A{ur}"].font = K.BF
wu[f"B{ur}"] = "目标 P/E(x)"; wu[f"B{ur}"].font = K.BF; wu[f"B{ur}"].fill = K.CH
wu.merge_cells(f"C{ur}:L{ur}"); wu[f"C{ur}"] = "说明"; wu[f"C{ur}"].font = K.BF; wu[f"C{ur}"].fill = K.CH
ur += 1
PE_ROWS = {}
pe_notes = {"Bear": "稳态盈利薄 → PE 镜头近失效, 该案估值靠 EV/Sales 支撑(见三角)。",
            "Base": "2030E 净利 $33.6B × 29x ≈ $975B(接近现锚); 稳态 EBIT ~17% 下 PE 较 v1 激进 30x 显著下拉。",
            "Bull": "2030E 净利 $78B × 30x ≈ $2.34T; agent mix 全胜 + 定价权。"}
for cs in CASES:
    K.lab(wu, f"A{ur}", f"  {cs}")
    K.inp(wu, f"B{ur}", PE_TGT[cs], None, K.MX)
    wu.merge_cells(f"C{ur}:L{ur}"); K.logic(wu, f"C{ur}", pe_notes[cs])
    PE_ROWS[cs] = ur
    ur += 1
ur += 1
K.band(wu, ur, "卖方对账 + 数据源", 9); ur += 1
K.mtext(wu, ur, ("凭什么敢给非主流数: 卖方多用 EV/Revenue、frontier lab 组 20-46x run-rate; 本模型用 EV/fwd-GAAP-Sales, "
                 "据 B 组同业光谱(SNOW/DDOG 14-17x、PLTR 91x)+ 增长换挡逐年 normalize, 不靠拍脑袋。一致性: Base 2027E 9.5x×$110B=$1,045B 复现 v3.1 EV/Sales 镜头。"
                 "数据源: 同业 fwd-GAAP=公开市值÷NTM 收入(2026-06); PE 目标=2030E 稳态净利反推。"), "I", 3)

# ════════════ 7. 情景切换(业务杠杆直接拍 + 估值倍数 link 倍数假设)════════════
sw = K.write_scenario_switch(wb.create_sheet(S_SW), {
    "title": "情景切换 — 全模型唯一情景参数库 + 切换开关(默认 Base)",
    "usage": ("怎么用: B2 是唯一入口——下拉选案 → 案序号派生 → 各杠杆『当前案』行跟着切 → "
              "物理锚 token 量 / 利润 / 逐年 EV/Sales 估值整条链变档, 『情景估值』输出该案逐年隐含估值阶梯。"
              "三案对比不用切: 『估值对比』恒常三案并排(引本页矩阵行)。现锚/DCF/概率三案共用。"),
    "cases": CASES, "default": "Base",
    "triggers": [
        ("Bear", "S-1 gross→net 口径折扣 >30% + 价格战实质化(OpenAI/Google token 降价) + GPT-6 反超 → 企业 API 份额腰斩(40%→18%), compute/收入比卡 60%。"),
        ("Base", "API 增速换挡但 Enterprise/Cowork 工具层接力; 开源吃低端、Anthropic 守企业高端(份额→30%); 2027 年度盈亏平衡, compute/收入比 64%→48%。"),
        ("Bull", "agent 垂直复制(法律/医疗/金融)×3 + 价格战未实质发生 + coding 心智守住(份额 40%→38%); 单位毛利 65%, compute/收入比压到 42%。"),
    ],
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "levers": [
        {"key": "tokg", "name": "token 量 YoY 增速", "fmt": K.PCT, "cols": FC[2:],
         "vals": TOK_G,
         "desc": "物理锚: GAAP 收入 = token 量 × 单价, token 量都挂它。2026E 三案共用锚 14.1Q(校准卖方); 分歧在 2027 之后。改增速→token 量→收入→隐含估值全链动。",
         "stories": {"Bear": "份额腰斩 + 价格战压低用量增长 → 2027E 收入 ~$70B。", "Base": "agent token 弹性兑现 → 2027E $110B、2030E $240B。", "Bull": "垂直复制放大 token 消耗 → 2027E ~$137B。"},
         "hist": [None, None, None, None, None]},
        {"key": "cmpr", "name": "compute/收入比", "fmt": K.PCT, "cols": FCf,
         "vals": CMP_R,
         "desc": "关键 margin 驱动: EBIT margin ≈ 1 − compute/收入比 − 云分成% − 其他opex%。改它→EBIT→净利→PE 隐含随动。Base 64%→48%(稳态 EBIT ~17%)。",
         "stories": {"Bear": "价格战快于成本改善, 卡 60% → 稳态 OM 个位数。", "Base": "推理效率改善 vs agent 重负载抵消, 降到 48%。", "Bull": "效率 + 定价权 → 压到 42%, 稳态 EBIT 27%。"},
         "hist": [None, None, None, None, 1.05]},
    ],
    "linked": [
        {"key": "evs", "name": "EV/Sales(当前案)", "fmt": K.MX,
         "src_sheet": S_MULT, "src_row0": EVS_ROWS["Bear"], "note": "整体 EV/Sales 逐年, link『估值倍数假设』三案行(Bear 首行); 改去那页改蓝字。"},
    ],
})
SWB, SWACT = sw["SWB"], sw["SWACT"]
SC = wb[S_SW]
SWIDX = sw["idx_cell"]
_r = sw["next_row"]

# PE 目标(当前案)单值 link + 全局参数, 手搓 ──────
K.lab(SC, f"A{_r}", "PE 目标(当前案)", b=True)
for _c in ["A"] + list(ALLC):
    SC[f"{_c}{_r}"].fill = K.GREYF
K.logic(SC, f"L{_r}", "2030E 稳态净利的目标 P/E, link『估值倍数假设』; 当前案 INDEX 三案。")
_r += 1
_pe_base0 = _r
for cs in CASES:
    K.lab(SC, f"A{_r}", f"  {cs}")
    K.fml(SC, f"B{_r}", f"={K.R(S_MULT, 'B' + str(PE_ROWS[cs]))}", K.MX, link=True)
    _r += 1
K.lab(SC, f"A{_r}", "  当前案(链引此行)", note=True)
K.fml(SC, f"B{_r}", f"=INDEX(B{_pe_base0}:B{_pe_base0+2},{SWIDX})", K.MX)
PE_ACT = _r
_r += 2

K.band(SC, _r, "全局参数(三案共用)", 9); _r += 1
GLOB = {}
for nm, v, fmt, note in [
        ("现锚 post-money($B)", ANCHOR_EV, K.N0, "Series H post-money(2026-05-28); 隐含估值 vs 它 = 上行/下行空间。"),
        ("DCF 内在价值($B,Base)", DCF_EV, K.N0, "UFCF 折现(WACC 12%/g 4%); 三镜头之一, sanity check。"),
        ("IPO 媒体首日中值($B)", IPO_MID, K.N0, "Oct-2026 预期, +14% vs Series H; 流动性溢价参照。"),
        ("权重 Bear", WEIGHTS["Bear"], K.PCT, "情景概率(主观): 价格战+口径折扣尾部。"),
        ("权重 Base", WEIGHTS["Base"], K.PCT, "基线: API 换挡+工具层接力。"),
        ("权重 Bull", WEIGHTS["Bull"], K.PCT, "agent 垂直复制全胜。"),
        ("EV/Sales 镜头权重", 0.40, K.PCT, "三镜头加权: 主镜头 EV/2027E Sales。"),
        ("PE 镜头权重", 0.30, K.PCT, "三镜头加权: 2030E 稳态 PE。"),
        ("DCF 镜头权重", 0.30, K.PCT, "三镜头加权: DCF 内在价值。")]:
    K.lab(SC, f"A{_r}", nm, b=True)
    K.inp(SC, f"B{_r}", v, None, fmt)
    K.logic(SC, f"L{_r}", note)
    GLOB[nm] = _r
    _r += 1

# ════════════ 8. 物理锚 [ANCHOR] — token 经济 ════════════
anchor = K.write_anchor(wb.create_sheet(S_ANCHOR), {
    "title": "Anthropic 物理锚 — token 经济(GAAP 收入 = token 量 × 有效混合单价)",
    "all_cols": ALLC, "all_years": ALLY,
    "series": [
        ("有效混合单价 ($/M tokens)",
         [None, None, None, None, 6.0, 3.2, 2.4, 1.9, 1.6, 1.35],
         "输入: mix(Opus/Sonnet/Haiku 挂牌)×batch/缓存折扣推导; 商品化逐年下行(2025 $6→2030 $1.35)。", K.N2),
        ("token 处理量 (Q/yr, quadrillion)",
         [None, None, None, None, 0.83, 14.1, None, None, None, None],
         "输入 2025A/2026E 锚(校准卖方收入); 2027E+ = 上年 ×(1+情景切换 token 增速当前案)。", K.N1),
        ("⇒ GAAP 收入 ($B) = 量 × 价",
         [None, None, None, None, None, None, None, None, None, None],
         "= token 量 × 有效混合单价(物理锚出口); 改单价或 token 量 → 收入 → 隐含估值全链动。", K.N1),
    ],
    "source_note": ("口径: GAAP 收入 = token 量(quadrillion/yr)× 有效混合单价($/M)。来源: 单价=claude.com/pricing mix 推导; "
                    "token 量 2025/2026=卖方收入反推锚, 前瞻=锚×情景增速(『情景切换』)。Anthropic 企业付费 API 份额 40%(Menlo)。"),
    "role_note": "作用: token 量 → GAAP 收入(下行 = 量×价)→ 分部 / 利润 / 逐年 EV/Sales 隐含估值全链动。改单价或 token 量 → 隐含估值跟着动。",
})
A_PRC = anchor["row_of"]["有效混合单价 ($/M tokens)"]
A_TOK = anchor["row_of"]["token 处理量 (Q/yr, quadrillion)"]
REV_ANCHOR_ROW = anchor["row_of"]["⇒ GAAP 收入 ($B) = 量 × 价"]
# 前瞻 token 量 2027E+ = 上年 ×(1+情景切换 token 增速当前案)
TOKG_ACT = SWACT["tokg"]
for _i, _c in enumerate(FC[2:]):   # H-K = 2027E-2030E
    _prev = FC[1:][_i]            # G,H,I,J
    K.fml(wb[S_ANCHOR], f"{_c}{A_TOK}", f"={_prev}{A_TOK}*(1+{K.R(S_SW, _c + str(TOKG_ACT))})", K.N1, link=True)
# GAAP 收入 = 量 × 单价(出口行, F-K = 2025A-2030E)
warc = wb[S_ANCHOR]
warc[f"A{REV_ANCHOR_ROW}"].fill = K.OUT; warc[f"A{REV_ANCHOR_ROW}"].font = K.BF
for _c in FC:
    K.fml(warc, f"{_c}{REV_ANCHOR_ROW}", f"={_c}{A_TOK}*{_c}{A_PRC}", K.N1)

# ════════════ 9. 分部测算(总收入引物理锚 + 业务线占比拆分)════════════
seg = K.write_segment_model(wb.create_sheet(S_SEG), {
    "title": "分部测算 — 总收入(引物理锚 token 经济)× 业务线占比(API / Code / Enterprise / 消费)",
    "all_cols": ALLC, "all_years": ALLY, "logic_col": "M",
    "groups": [
        ("总收入(物理锚驱动)", [
            ("总收入 ($B)", None, K.N1, "= 引『物理锚』GAAP 收入(token 量×单价, 当前案); 改物理锚→这里→隐含估值全链动。"),
        ]),
        ("业务线占比(% 总收入; HSBC run-rate 拆分口径)", [
            ("模型 API %", [0.83, 0.80, 0.78, 0.75, 0.75, 0.575, 0.545, 0.528, 0.523, 0.521], K.PCT, "token-as-a-service; 商品化压力, 远期占比降至 52%。"),
            ("Claude Code %", [0.025, 0.03, 0.04, 0.05, 0.09, 0.16, 0.145, 0.137, 0.128, 0.121], K.PCT, "coding agent; 2026 放量后增速主动压到 +30% 下。"),
            ("Claude for Enterprise %", [0.05, 0.05, 0.06, 0.06, 0.06, 0.19, 0.245, 0.273, 0.284, 0.292], K.PCT, "Cowork/Excel; 最大增量引擎, 远期占比 29%。"),
            ("消费订阅 %", [0.095, 0.12, 0.12, 0.14, 0.10, 0.075, 0.064, 0.062, 0.064, 0.067], K.PCT, "Claude.ai; 防御性配置, 占比小。"),
        ]),
        ("业务线收入($B) = 总收入 × 占比", [
            ("模型 API 收入", None, K.N1, "= 总收入 × API%。"),
            ("Claude Code 收入", None, K.N1, "= 总收入 × Code%。"),
            ("Enterprise 收入", None, K.N1, "= 总收入 × Enterprise%。"),
            ("消费订阅 收入", None, K.N1, "= 总收入 × 消费%。"),
        ]),
    ],
})
sm = seg["m"]
for col in ALLC:
    src = REV_ANCHOR_ROW
    if col in FC:
        K.fml(wb[S_SEG], f"{col}{sm['总收入 ($B)']}", f"={K.R(S_ANCHOR, col + str(src))}", K.N1, link=True)
    else:
        # 2021A-2024E(B-E)历史收入: 物理锚未建早年, 直接引历史财务
        K.fml(wb[S_SEG], f"{col}{sm['总收入 ($B)']}", f"={K.R(S_HIST, col + str(HREV))}", K.N1, link=True)
seg_pct = [("模型 API 收入", "模型 API %"), ("Claude Code 收入", "Claude Code %"),
           ("Enterprise 收入", "Claude for Enterprise %"), ("消费订阅 收入", "消费订阅 %")]
for revname, pctname in seg_pct:
    for col in ALLC:
        K.fml(wb[S_SEG], f"{col}{sm[revname]}", f"={col}{sm['总收入 ($B)']}*{col}{sm[pctname]}", K.N1)
    for col in ALLC:
        wb[S_SEG][f"{col}{sm[revname]}"].fill = K.OUT

# ════════════ 10. 利润与收入假设(收入→compute比→EBIT→净利, 手搓)════════════
wf = wb.create_sheet(S_FUND)
K.hdr(wf, 1, "利润与收入假设 — 总收入 → compute/收入比 → EBIT → 净利(PE 镜头底座)", 9)
fr0 = K.mtext(wf, 2, ("粗颗粒利润: 总收入(引分部)× margin 链。EBIT = 收入 ×(1 − compute/收入比 − 云分成% − 其他opex%); "
                      "净利 =(EBIT + 利息净)×(1 − 税率)。compute/收入比是关键 margin 杠杆(引『情景切换』当前案)。净利喂 PE 镜头, 不做三表勾稽。"), "I", 2)
for col, w in zip("ABCDEFGHI", [22, 9, 9, 9, 9, 9, 9, 9, 9]):
    wf.column_dimensions[col].width = w
wf.column_dimensions["M"].width = 58
HR = fr0
K.lab(wf, f"A{HR}", "($B)", b=True)
for col, y in zip(ALLC, ALLY):
    wf[f"{col}{HR}"] = y; wf[f"{col}{HR}"].font = K.BF; wf[f"{col}{HR}"].fill = K.CH
wf[f"M{HR}"] = "逻辑/来源"; wf[f"M{HR}"].font = K.BF; wf[f"M{HR}"].fill = K.CH
fr = HR + 1
FU = {}
def fu_row(name, fmt=K.N1):
    global fr
    K.lab(wf, f"A{fr}", name)
    FU[name] = fr
    r = fr; fr += 1
    return r
# 总收入
RTOT = fu_row("总收入 ($B)", K.N1)
for col in ALLC:
    K.fml(wf, f"{col}{RTOT}", f"={K.R(S_SEG, col + str(sm['总收入 ($B)']))}", K.N1, link=True)
K.logic(wf, f"M{RTOT}", "= 引『分部测算』总收入(物理锚 token 经济驱动)。")
K.lab(wf, f"A{fr}", "  YoY", note=True)
for i in range(1, len(ALLC)):
    K.fml(wf, f"{ALLC[i]}{fr}", f"=IFERROR({ALLC[i]}{RTOT}/{ALLC[i-1]}{RTOT}-1,\"n.m.\")", K.PCT)
fr += 1
# compute/收入比(引情景切换当前案; 历史 2025=1.05)
CMPACT = SWACT["cmpr"]
RCMP = fu_row("compute/收入比 (%)", K.PCT)
K.fml(wf, f"F{RCMP}", f"={K.R(S_SW, 'F' + str(CMPACT))}", K.PCT, link=True)
for col in FCf:
    K.fml(wf, f"{col}{RCMP}", f"={K.R(S_SW, col + str(CMPACT))}", K.PCT, link=True)
K.logic(wf, f"M{RCMP}", "= 引『情景切换』当前案; 2025A 1.05x → Base 2030E 0.48。关键 margin 杠杆。")
# 云分成%
RCLD = fu_row("云分成 COGS (%)", K.PCT)
K.introw(wf, RCLD, FC, [0.05, 0.05, 0.05, 0.05, 0.05, 0.05], None, K.PCT)
K.logic(wf, f"M{RCLD}", "云 marketplace(Bedrock/Vertex/Azure)分成, 占收入 ~5%(输入)。")
# 其他opex%
ROPX = fu_row("其他 opex (%)", K.PCT)
K.introw(wf, ROPX, FC, [None, 0.39, 0.34, 0.33, 0.31, 0.30], None, K.PCT)
K.inp(wf, f"F{ROPX}", 1.10, None, K.PCT)
K.logic(wf, f"M{ROPX}", "S&M+G&A+非算力 R&D+D&A 占收入(输入); 2025A 110%(烧钱)→ Base 2030E 30%。")
# EBIT
REBIT = fu_row("EBIT ($B)", K.N1)
for col in FC:
    K.fml(wf, f"{col}{REBIT}", f"={col}{RTOT}*(1-{col}{RCMP}-{col}{RCLD}-{col}{ROPX})", K.N1)
K.logic(wf, f"M{REBIT}", "= 收入 ×(1 − compute比 − 云分成% − opex%)。Base 2026 −$3.6B → 2027 +$3.3B(年度盈亏平衡)→ 2030 +$40.8B。")
K.lab(wf, f"A{fr}", "  EBIT margin", note=True)
for col in FC:
    K.fml(wf, f"{col}{fr}", f"={col}{REBIT}/{col}{RTOT}", K.PCT)
fr += 1
# 利息净 + 税率
RINT = fu_row("利息净收入 ($B)", K.N1)
K.introw(wf, RINT, FC, [0.3, 1.2, 1.5, 1.2, 1.0, 1.2], None, K.N1)
K.logic(wf, f"M{RINT}", "净现金理财收益(输入); 一级募资 $95B 在手。")
RTAX = fu_row("有效税率 (%)", K.PCT)
K.introw(wf, RTAX, FC, [0.0, 0.0, 0.05, 0.12, 0.17, 0.20], None, K.PCT)
K.logic(wf, f"M{RTAX}", "亏损期 0, 盈利后逐步爬升至 20%(输入)。")
# 净利
RNI = fu_row("净利润 ($B)", K.N1)
for col in HC:
    K.fml(wf, f"{col}{RNI}", f"={K.R(S_HIST, col + str(HNI))}", K.N1, link=True)
for col in FCf:
    K.fml(wf, f"{col}{RNI}", f"=({col}{REBIT}+{col}{RINT})*(1-{col}{RTAX})", K.N1)
wf[f"A{RNI}"].fill = K.OUT; wf[f"A{RNI}"].font = K.BF
K.logic(wf, f"M{RNI}", "=(EBIT+利息净)×(1−税率); 历史引『历史财务』。Base: 2026 −$2.4B → 2027 +$4.6B → 2030 +$33.6B。喂 PE 镜头。")
K.lab(wf, f"A{fr}", "  净利率", note=True)
for col in FC:
    K.fml(wf, f"{col}{fr}", f"={col}{RNI}/{col}{RTOT}", K.PCT)
fr += 1
K.band(wf, fr, "口径说明", 9); fr += 1
K.mtext(wf, fr, ("粗颗粒(skill 海拔): 营收从物理锚 bottom-up, 利润用 margin 假设链, 不做三表勾稽 / 营运资本。"
                 "净利供 PE 交叉镜头; 主镜头走 EV/Sales(收入 × 倍数), 故利润精度非目标。"
                 "★ 收支平衡: 单季经调整经营利润 2026-Q2 已现(橱窗, 剔 SBC + 表外算力承诺, 不持续); 年度会计盈亏 2027; 正 FCF 2029(公司内部口径)。"), "I", 3)

# ════════════ 11. 情景估值(当前案逐年 EV/Sales 主镜头 + PE/DCF 交叉, 手搓)════════════
wv = wb.create_sheet(S_VAL)
K.hdr(wv, 1, "情景估值 — 当前案逐年 EV/Sales 隐含估值 + PE/DCF 三镜头交叉验证", 8)
K.lab(wv, "G1", "当前情景→", note=True)
K.fml(wv, "H1", f"={K.R(S_SW, 'B2')}", K.N0, link=True); wv["H1"].fill = K.CUR
vr = K.mtext(wv, 2, ("本表输出=『情景切换』当前案(默认 Base)。主镜头逐年(2026E-2030E): 隐含估值 EV = 总收入(引利润页, 链物理锚)× 目标 EV/Sales"
                     "(引『情景切换』当前案), forward 不折现, 直接对现锚 $965B。PE 交叉(2030E 稳态净利)+ DCF($575B)做三镜头三角, 加权目标见结论。"), "H", 3)
for col, w in zip("ABCDEFGH", [24, 12, 12, 12, 12, 12, 12, 12]):
    wv.column_dimensions[col].width = w
EVS_ACT = SWACT["evs"]
ANCHOR_REF = K.R(S_SW, "B" + str(GLOB["现锚 post-money($B)"]))
DCF_REF = K.R(S_SW, "B" + str(GLOB["DCF 内在价值($B,Base)"]))

K.band(wv, vr, "主镜头 EV/Sales(逐年 forward; 隐含估值 = 收入 × EV/Sales)", 8); vr += 1
wv[f"A{vr}"] = "项目 / 年份"; wv[f"A{vr}"].font = K.BF
for col, y in zip(FCf, FWY):
    wv[f"{col}{vr}"] = y; wv[f"{col}{vr}"].font = K.BF; wv[f"{col}{vr}"].fill = K.CH
vr += 1
VV = {}
def vrow(label, b=False, out=False):
    global vr
    K.lab(wv, f"A{vr}", label, b=b)
    if out:
        wv[f"A{vr}"].fill = K.OUT
    VV[label] = vr; r = vr; vr += 1
    return r
RV_REV = vrow("总收入 ($B)")
for col in FCf:
    K.fml(wv, f"{col}{RV_REV}", f"={K.R(S_FUND, col + str(RTOT))}", K.N1, link=True)
RV_EVS = vrow("目标 EV/Sales (x)")
for col in FCf:
    K.fml(wv, f"{col}{RV_EVS}", f"={K.R(S_SW, col + str(EVS_ACT))}", K.MX, link=True)
RV_EV = vrow("隐含估值 EV ($B)", b=True, out=True)
for col in FCf:
    K.fml(wv, f"{col}{RV_EV}", f"={col}{RV_REV}*{col}{RV_EVS}", K.N0)
RV_VS = vrow("隐含价 vs 现锚 $965B", b=True)
for col in FCf:
    K.fml(wv, f"{col}{RV_VS}", f"={col}{RV_EV}/{ANCHOR_REF}-1", K.PCT)
vr += 1
K.band(wv, vr, "PE 交叉验证(支线; 隐含 forward P/E = 隐含估值 ÷ 当年净利)", 8); vr += 1
wv[f"A{vr}"] = "项目 / 年份"; wv[f"A{vr}"].font = K.BF
for col, y in zip(FCf, FWY):
    wv[f"{col}{vr}"] = y; wv[f"{col}{vr}"].font = K.BF; wv[f"{col}{vr}"].fill = K.CH
vr += 1
RV_NI = vrow("净利润 ($B)")
for col in FCf:
    K.fml(wv, f"{col}{RV_NI}", f"={K.R(S_FUND, col + str(RNI))}", K.N1, link=True)
RV_IPE = vrow("隐含 forward P/E (x)")
for col in FCf:
    K.fml(wv, f"{col}{RV_IPE}", f'=IF({col}{RV_NI}<=0,"N/M",{col}{RV_EV}/{col}{RV_NI})', K.MX)
K.logic(wv, f"A{vr}", "读法: 亏损年(2026)净利<0 → 隐含 P/E = N/M, 这正是主镜头用 EV/Sales 的原因; 2030E 隐含 P/E 应落同业光谱内。")
wv.merge_cells(f"A{vr}:H{vr}")
vr += 2
K.band(wv, vr, "三镜头加权(当前案; EV/Sales 主 + PE 2030E + DCF)", 8); vr += 1
for col, h in zip(["A", "B", "C", "D"], ["镜头", "隐含估值$B", "权重", "加权贡献$B"]):
    wv[f"{col}{vr}"] = h; wv[f"{col}{vr}"].font = K.BF; wv[f"{col}{vr}"].fill = K.CH
wv.merge_cells(f"E{vr}:H{vr}"); wv[f"E{vr}"] = "说明"; wv[f"E{vr}"].font = K.BF; wv[f"E{vr}"].fill = K.CH
vr += 1
W_EVS = K.R(S_SW, "B" + str(GLOB["EV/Sales 镜头权重"]))
W_PE = K.R(S_SW, "B" + str(GLOB["PE 镜头权重"]))
W_DCF = K.R(S_SW, "B" + str(GLOB["DCF 镜头权重"]))
PE_ACT_REF = K.R(S_SW, "B" + str(PE_ACT))
# EV/Sales 镜头 = 2027E 隐含估值(H 列)
L1 = vr
K.lab(wv, f"A{vr}", "EV/2027E Sales")
K.fml(wv, f"B{vr}", f"=H{RV_EV}", K.N0, link=True)
K.fml(wv, f"C{vr}", f"={W_EVS}", K.PCT, link=True)
K.fml(wv, f"D{vr}", f"=B{vr}*C{vr}", K.N0)
wv.merge_cells(f"E{vr}:H{vr}"); K.logic(wv, f"E{vr}", "主镜头: 2027E 总收入 × 目标 EV/Sales(当前案); Base = $110B × 9.5x = $1,045B。")
vr += 1
L2 = vr
K.lab(wv, f"A{vr}", "PE 2030E 稳态")
K.fml(wv, f"B{vr}", f"=K{RV_NI}*{PE_ACT_REF}", K.N0, link=True)
K.fml(wv, f"C{vr}", f"={W_PE}", K.PCT, link=True)
K.fml(wv, f"D{vr}", f"=B{vr}*C{vr}", K.N0)
wv.merge_cells(f"E{vr}:H{vr}"); K.logic(wv, f"E{vr}", "交叉: 2030E 净利 × 目标 P/E(当前案); Base = $33.6B × 29x ≈ $975B。")
vr += 1
L3 = vr
K.lab(wv, f"A{vr}", "DCF 内在价值")
K.fml(wv, f"B{vr}", f"={DCF_REF}", K.N0, link=True)
K.fml(wv, f"C{vr}", f"={W_DCF}", K.PCT, link=True)
K.fml(wv, f"D{vr}", f"=B{vr}*C{vr}", K.N0)
wv.merge_cells(f"E{vr}:H{vr}"); K.logic(wv, f"E{vr}", "交叉: UFCF 折现(WACC 12%/g 4%)≈ $575B; 一级 DCF 中性价 ≈ Series H, 不为 Bull 期权付费。")
vr += 1
RV_WT = vr
K.lab(wv, f"A{vr}", "加权目标估值 ($B)", b=True); wv[f"A{vr}"].fill = K.OUT
K.fml(wv, f"D{vr}", f"=D{L1}+D{L2}+D{L3}", K.N0)
wv.merge_cells(f"E{vr}:H{vr}"); K.logic(wv, f"E{vr}", "= 三镜头加权; Base ≈ $883B。")
vr += 1
RV_WTVS = vr
K.lab(wv, f"A{vr}", "加权目标 vs 现锚 $965B", b=True)
K.fml(wv, f"D{vr}", f"=D{RV_WT}/{ANCHOR_REF}-1", K.PCT)
wv.merge_cells(f"E{vr}:H{vr}"); K.logic(wv, f"E{vr}", "Base ≈ −8.5%: 当前一级定价在 Base 三镜头下略偏贵, 上行须靠 Bull 兑现。")
vr += 2
K.band(wv, vr, "方法与结论", 8); vr += 1
K.mtext(wv, vr, ("方法: 主镜头逐年 EV/Sales(收入 × forward 倍数, 不折现); PE(2030E 稳态)+ DCF 三角。整体亏损刚转正故主镜头用销售倍数出实数。"
                 "结论(Base): EV/Sales 单镜头 2027E $1,045B 偏高, 被 DCF $575B 拉回 → 三镜头加权 $883B(−8.5% vs 现锚)→ HOLD。"
                 "情景概率加权见『估值对比』($1,023B, +6%); 两法分歧 = 对'稳态 EBIT 能否高位'的押注大小。S-1(2026-Q3)口径揭晓前不追高。"), "H", 4)

# ════════════ 12. 估值对比(三案逐年阶梯, 手搓; 防污染=引矩阵行/未翻档输入)════════════
wc = wb.create_sheet(S_CMP)
K.hdr(wc, 1, "估值对比 — Bear / Base / Bull 三案逐年(2026E-2030E)EV/Sales 隐含估值阶梯", 8)
K.mtext(wc, 2, ("三案各自完整推演: token 增速 → token 量 → 收入(×单价)→ EV/Sales → 隐含估值 EV → vs 现锚, 并出隐含 forward P/E 体检。"
                "三案永远并排, 不随『情景切换』开关变(只引矩阵各案行 + 物理锚单价 + 利润页未翻档输入, 防当前案污染)。"
                "底部双重加权: ① 三镜头加权(EV/Sales+PE+DCF)= 评级基准; ② 情景概率加权(0.20/0.55/0.25)。"), "H", 3)
for col, w in zip("ABCDEFGH", [24, 11, 11, 11, 11, 11, 13, 13]):
    wc.column_dimensions[col].width = w
PRC_REF = lambda c: K.R(S_ANCHOR, c + str(A_PRC))          # 单价(静态, 三案共用)
CLD_REF = lambda c: K.R(S_FUND, c + str(RCLD))            # 云分成%(Base 输入, 未翻档)
OPX_REF = lambda c: K.R(S_FUND, c + str(ROPX))            # opex%(Base 输入, 未翻档)
INT_REF = lambda c: K.R(S_FUND, c + str(RINT))
TAX_REF = lambda c: K.R(S_FUND, c + str(RTAX))
TOK2026 = 14.1
rows_def = [
    ("tokg",  "token YoY 增速", K.PCT),
    ("tok",   "token 量 (Q)", K.N1),
    ("prc",   "单价 ($/M)", K.N2),
    ("rev",   "总收入 ($B)", K.N1),
    ("evs",   "EV/Sales (x)", K.MX),
    ("ev",    "隐含估值 EV ($B)", K.N0),
    ("vs",    "vs 现锚 $965B", K.PCT),
    ("cmpr",  "compute/收入比", K.PCT),
    ("ebit",  "EBIT ($B)", K.N1),
    ("ni",    "净利 ($B)", K.N1),
    ("ipe",   "隐含 forward P/E", K.MX),
]
block_h = len(rows_def) + 3
CMP = {}
b0 = 6
TOKG_OFF = SWB["tokg"]; CMPR_OFF = SWB["cmpr"]
for ci, cname in enumerate(CASES):
    r0 = b0 + ci * block_h
    K.band(wc, r0, f"{cname} 案 — 逐年 EV/Sales 隐含估值阶梯", 8)
    yr = r0 + 1
    wc[f"A{yr}"] = "项目 / 年份"; wc[f"A{yr}"].font = K.BF
    for col, y in zip(FCf, FWY):
        wc[f"{col}{yr}"] = y; wc[f"{col}{yr}"].font = K.BF; wc[f"{col}{yr}"].fill = K.CH
    rr = yr + 1
    A = {}
    for key, label, fmt in rows_def:
        A[key] = rr
        K.lab(wc, f"A{rr}", label, b=(key in ("ev",)))
        if key == "ev":
            wc[f"A{rr}"].fill = K.OUT
        for j, col in enumerate(FCf):   # j=0→2026E ... 4→2030E
            if key == "tokg":
                if j == 0:
                    K.lab(wc, f"{col}{rr}", "锚", note=True); continue
                src = TOKG_OFF + ci
                f = f"={K.R(S_SW, col + str(src))}"
            elif key == "tok":
                if j == 0:
                    K.inp(wc, f"{col}{rr}", TOK2026, None, K.N1); continue
                prevcol = FCf[j-1]
                f = f"={prevcol}{A['tok']}*(1+{col}{A['tokg']})"
            elif key == "prc":
                f = f"={PRC_REF(col)}"
            elif key == "rev":
                f = f"={col}{A['tok']}*{col}{A['prc']}"
            elif key == "evs":
                src = EVS_ROWS[cname]
                f = f"={K.R(S_MULT, col + str(src))}"
            elif key == "ev":
                f = f"={col}{A['rev']}*{col}{A['evs']}"
            elif key == "vs":
                f = f"={col}{A['ev']}/{ANCHOR_REF}-1"
            elif key == "cmpr":
                src = CMPR_OFF + ci
                f = f"={K.R(S_SW, col + str(src))}"
            elif key == "ebit":
                f = f"={col}{A['rev']}*(1-{col}{A['cmpr']}-{CLD_REF(col)}-{OPX_REF(col)})"
            elif key == "ni":
                f = f"=({col}{A['ebit']}+{INT_REF(col)})*(1-{TAX_REF(col)})"
            elif key == "ipe":
                f = f'=IF({col}{A["ni"]}<=0,"N/M",{col}{A["ev"]}/{col}{A["ni"]})'
            link = ("'" in f)
            K.fml(wc, f"{col}{rr}", f, fmt, link=link)
        rr += 1
    CMP[cname] = A

# 底部: 双重加权 ──────
sr = b0 + len(CASES) * block_h + 1
K.band(wc, sr, "概率加权 + 三镜头加权 → 目标估值 vs 现锚", 8); sr += 1
for col, h in zip(["A", "B", "C", "D", "E"], ["项目", "Bear", "Base", "Bull", "加权"]):
    wc[f"{col}{sr}"] = h; wc[f"{col}{sr}"].font = K.BF; wc[f"{col}{sr}"].fill = K.CH
wc.merge_cells(f"F{sr}:H{sr}"); wc[f"F{sr}"] = "说明"; wc[f"F{sr}"].font = K.BF; wc[f"F{sr}"].fill = K.CH
sr += 1
WT_B = K.R(S_SW, "B" + str(GLOB["权重 Bear"]))
WT_BA = K.R(S_SW, "B" + str(GLOB["权重 Base"]))
WT_BU = K.R(S_SW, "B" + str(GLOB["权重 Bull"]))
# 概率权重
R_WT = sr
K.lab(wc, f"A{sr}", "情景概率")
K.fml(wc, f"B{sr}", f"={WT_B}", K.PCT, link=True)
K.fml(wc, f"C{sr}", f"={WT_BA}", K.PCT, link=True)
K.fml(wc, f"D{sr}", f"={WT_BU}", K.PCT, link=True)
wc.merge_cells(f"F{sr}:H{sr}"); K.logic(wc, f"F{sr}", "主观概率: 口径折扣 + 价格战尾部 → Bear 0.20 / Base 0.55 / Bull 0.25。")
sr += 1
# 主镜头 EV/Sales 2027E 隐含(12M)
R_EVS27 = sr
K.lab(wc, f"A{sr}", "EV/Sales 2027E 隐含($B)", b=True); wc[f"A{sr}"].fill = K.OUT
for col, cn in zip(["B", "C", "D"], CASES):
    K.fml(wc, f"{col}{sr}", f"=H{CMP[cn]['ev']}", K.N0, link=True)   # H 列 = 2027E(FCf[1])
K.fml(wc, f"E{sr}", f"=B{sr}*B{R_WT}+C{sr}*C{R_WT}+D{sr}*D{R_WT}", K.N0); wc[f"E{sr}"].fill = K.CUR
wc.merge_cells(f"F{sr}:H{sr}"); K.logic(wc, f"F{sr}", "主镜头 2027E 隐含估值(各案 收入×EV/Sales); 概率加权 ≈ $1,140B(单镜头偏高, 见三镜头)。")
sr += 1
# 三镜头加权(各案 headline; TRILENS 输入, 注明构成)
R_TRI = sr
K.lab(wc, f"A{sr}", "三镜头加权估值($B)", b=True); wc[f"A{sr}"].fill = K.OUT
for col, cn in zip(["B", "C", "D"], CASES):
    K.inp(wc, f"{col}{sr}", TRILENS[cn], None, K.N0)
K.fml(wc, f"E{sr}", f"=B{sr}*B{R_WT}+C{sr}*C{R_WT}+D{sr}*D{R_WT}", K.N0); wc[f"E{sr}"].fill = K.CUR
wc.merge_cells(f"F{sr}:H{sr}")
K.logic(wc, f"F{sr}", "各案 EV/Sales+PE+DCF 三镜头加权(Base $883B 构成见『情景估值』; Bear $560B/Bull $1,700B); 情景概率加权 ≈ $1,023B(+6%)。")
sr += 1
# 现锚 + vs
R_ANC = sr
K.lab(wc, f"A{sr}", "现锚 post-money($B)")
K.fml(wc, f"E{sr}", f"={ANCHOR_REF}", K.N0, link=True); wc[f"E{sr}"].fill = K.CUR
wc.merge_cells(f"F{sr}:H{sr}"); K.logic(wc, f"F{sr}", "Series H post-money $965B(2026-05-28)。")
sr += 1
R_TRIVS = sr
K.lab(wc, f"A{sr}", "vs 现锚(Base三镜头 / 概率加权)", b=True)
K.fml(wc, f"C{sr}", f"=C{R_TRI}/{ANCHOR_REF}-1", K.PCT)        # Base 三镜头 883 vs 965 = -8.5%
K.fml(wc, f"E{sr}", f"=E{R_TRI}/{ANCHOR_REF}-1", K.PCT)        # 情景概率加权 1023 vs 965 = +6%
wc.merge_cells(f"F{sr}:H{sr}"); K.logic(wc, f"F{sr}", "Base 三镜头加权 −8.5%(C列)、情景概率加权 +6%(E列): 区间在现锚附近震荡 → HOLD; 上行靠 IPO 流动性溢价(媒体首日中值 $1,100B)与 agent 垂直复制。")
sr += 2
K.mtext(wc, sr, ("关键洞察: Base 逐年隐含估值 2026E $630B → 2027E $1,045B → 2030E $1,008B(EV/Sales 主镜头); 但 DCF $575B 把三镜头加权拉到 $883B。"
                 "现锚 $965B 夹在'情景概率加权 $1,023B'与'三镜头加权 $883B'之间 → 一级定价已大致反映 Base 现金流, 未给 Bull 期权。"
                 "衰减扳机(S-1 口径折扣>30% / OpenRouter Claude token 份额连续 -10% / Claude Code 渗透回落<30%)触发 → 下调 Base 概率, 转 Bear($560B)重看。"), "H", 3)

# ════════════ 13. 综合判断仪表盘 ════════════
GAP_MKT = K.R(S_SW, "B" + str(GLOB["现锚 post-money($B)"]))
GAP_FAIR = K.R(S_CMP, "C" + str(R_TRI))      # Base 三镜头加权 $883B
dash = K.write_dashboard(wb.create_sheet(S_DASH), {
    "title": "综合判断仪表盘 — A 基本面拐点 · B 估值错位 · C 催化剂 · D 情绪确认",
    "usage": ("怎么用: 预测引擎 = B(估值错位)+ C(催化剂); 情绪 D 只做 timing + 过热刹车。"
              "私营标的无量价 → D 改读一级/二级转让与 IPO 预期。验收 = S-1 披露后回看本表读数。"),
    "blocks": [
        {"title": "A. 基本面拐点 — 业务在结构性变好吗?", "rows": [
            ("收入主体迁移", "✓ 已发生", "agent 工具层(Code+Enterprise)2026-05 占 run-rate 51%, 首次超过卖 token 的 API(46%)— 估值逻辑该从 API 商品切到 SaaS。"),
            ("企业份额第一", "✓ 已确立", "企业付费 API 份额 40%(Menlo, 2023→2026: 12%→40%)、编码 54%、Claude Code 企业渗透 39%(UBS)— 三个独立外部信号佐证。"),
            ("盈利能力", "△ 刚转正", "年度会计盈亏平衡 ~2027(净利 −$2.4B→+$4.6B); 正 FCF ~2029。稳态 EBIT 能否高位(17% vs Bear 个位数)是最大不确定。"),
            ("A 判断", "【中-强】", "收入主体迁移 + 企业第一已坐实; 盈利质量(稳态 margin)未证 → 中偏强。", True),
        ]},
        {"title": "B. 估值错位(预测引擎)— 市场给的 vs 三镜头该给的 → GAP", "rows": [
            ("市场现在给(现锚$B)", {"fml": f"={GAP_MKT}", "fmt": K.N0, "fill": True},
             "= Series H post-money $965B(一级最新成交)。"),
            ("基本面该给(Base 三镜头$B)", {"fml": f"={GAP_FAIR}", "fmt": K.N0},
             "= EV/Sales+PE+DCF 三镜头加权 $883B(引『估值对比』Base)。"),
            ("错位 GAP = 该给÷市场给 − 1",
             {"fml": lambda ro: f"=B{ro['基本面该给(Base 三镜头$B)']}/B{ro['市场现在给(现锚$B)']}-1", "fmt": K.PCT},
             "GAP ≈ −8.5%: 现锚略高于 Base 三镜头该给 → 基本面空间已被价格大致吃完, 上行靠 Bull 期权 + IPO 流动性, 非 Base 现金流。"),
            ("回测: Series G→H 的读数", "+154%(2026-02→05)", "G $380B → H $965B 三个月 +154%, 同期 run-rate $14B→$47B — 估值扩张由基本面(run-rate)驱动, 非纯倍数泡沫。"),
        ]},
        {"title": "C. 催化剂 — 什么会逼市场闭合 GAP", "rows": [
            ("S-1 公开版(2026-Q3)", "待", "首披审计 GAAP / gross-net 确认 / 算力承诺表 — 口径折扣 <15% 上修、>30% 重锚(双向最强催化)。"),
            ("IPO 定价与上市(最早 2026-10)", "待", "对标 OpenAI 私募→IPO 路径, 流动性溢价 + 指数纳入(媒体首日中值 $1,100B)。"),
            ("GPT-6 / Gemini 4 vs 下一代 Claude", "待", "coding/agentic benchmark 排位; 反超 = 目标 −20% 触发器。"),
            ("C 判断", "待兑现(双向)", "S-1 是分水岭; 口径干净 + IPO 顺利 → 上行兑现, 口径折扣 → 重锚。", True),
        ]},
        {"title": "D. 情绪确认 — 一级/二级 + IPO 预期(私营无量价)", "rows": [
            ("员工 tender 拒卖($350B)", "信心信号", "2026-04 员工集体拒在 $350B tender 卖股、等 IPO — 内部人信心的旁证。"),
            ("一级买方结构", "聪明钱进场", "Sequoia/Altimeter + 三大云尽调后 $965B 进场, 信息优势方在多头一侧。"),
            ("当前档位", "【启动-过热之间】", "估值带 de-rating(41x→20.5x)显示倍数已收敛, 非纯 FOMO; 但 IPO 前 6-12 月是重定价窗口, 情绪偏多。", True),
            ("衰减扳机", "3 条", "S-1 口径折扣>30% / OpenRouter Claude token 份额连续 -10% / Claude Code 渗透回落<30% — 任一触发 → 转 Bear。"),
        ]},
    ],
    "final": {"band": "★ 综合判断(A+B+C+D)",
              "text": ("收入主体迁移 + 企业第一已坐实(A 中-强), 但 B 估值错位仅 −8.5%(现锚 $965B ≈ Base 三镜头 $883B)— 一级定价已大致反映 Base 现金流。"
                       "上行(Bull $1,700B)靠 agent 垂直复制 + IPO 流动性溢价, 下行(Bear $560B)靠 S-1 口径折扣 + 价格战。风险报酬在现锚附近大致对称 → HOLD。"
                       "决策点 = S-1(2026-Q3): 口径折扣 <15% 转 BUY、>30% 转 SELL; 在此之前不追高、不杀跌, 等口径揭晓重定。")},
    "tracking": {
        "intro": "哪个指标恶化/兑现 → 哪个假设先动 → 触发什么动作(盯的优先级)。",
        "rows": [
            ("__band__", "一、口径(最大重锚风险)"),
            ("S-1 GAAP vs run-rate", "待", "命门: 收入绝对值 + EV/Sales 分母", "S-1 招股书(2026-Q3)", "折扣<15%→BUY; >30%→SELL"),
            ("__band__", "二、价格战 / 商品化"),
            ("OpenRouter Claude token 份额", "待跟踪", "命门: API 份额 → token 量 → 收入", "OpenRouter 月度", "连续 -10% → 下调 token 增速, 转 Bear"),
            ("compute/收入比", "Q1'26 0.71→Q2 0.56", "命门: 稳态 margin → PE 镜头", "季报 compute 披露", "卡 60% 不降 → PE 镜头失效, 估值靠 EV/Sales"),
            ("__band__", "三、工具层护城河"),
            ("Claude Code 企业渗透", "39%(UBS)", "命门: Enterprise 收入引擎", "UBS 半年度调研", "回落 <30% → 下调工具层, 转 Bear"),
        ],
    },
})

# ════════════ 全局格式 + 落盘 ════════════
K.finalize(wb, freeze={
    S_COVER: "A2", S_HIST: "B3", S_PX: "B4", S_CONS: "A2", S_HMULT: "B3",
    S_MULT: "B3", S_SW: "B3", S_ANCHOR: "B3", S_SEG: "B3",
    S_FUND: "B3", S_VAL: "B3", S_CMP: "B3", S_DASH: "B6",
})
_repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_output_dir = os.environ.get("VALUATION_OUTPUT_DIR", os.path.join(_repo_root, "out"))
out = os.path.join(_output_dir, "Anthropic_valuation_model.xlsx")
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("saved:", out)
print("sheets:", wb.sheetnames)
