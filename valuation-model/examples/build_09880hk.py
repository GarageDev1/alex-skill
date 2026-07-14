# -*- coding: utf-8 -*-
"""
build_09880hk.py — 优必选(UBTech Robotics, 09880.HK)物理锚估值模型。

用 build_kit v2 骨架。范式 = 未盈利高增长 → EV/Sales(P/S)主镜头 + 路径 P/E 交叉验证
(结构上是 build_2513hk 的孪生: 中国 AI/机器人、港股上市、RMB 财务 + HKD 股价)。

★ 单位口径(硬约束):
  - 财务量纲(营收/利润/权益/毛利)一律 USD $B。财务原始为 RMB 亿, 按 RMB/USD = 7.2 折 $B
    (亿RMB ÷ 72 = $B; 因 1亿 = 0.1 十亿, 且 十亿RMB/7.2 = 十亿USD)。
  - 股价 HKD; FX 行 = HKD/USD = 7.8, 只用于 每股(USD)↔HKD 隐含价 与 P/E/P/S 换算。
  - 隐含价(HKD) = 目标 P/S × 每股营收(USD) × 7.8。

★ 物理锚 = 全球人形机器人出货量(台, TAM); 优必选出货拆 工业线 + 消费线, 各 units × ASP。
  传统业务合计一条按增速外推(三案共用)。改底层出货量/ASP → 隐含价跟着动。

★ 三情景承重假设(工业/消费出货台数、工业/消费 ASP、归母净利率)+ 目标 P/S(6x/11x/16x)
  由主 agent 拍定, 本脚本只做工程实现, 不拟合现价、不改假设值。

跑:   cd examples && PYTHONUTF8=1 python build_09880hk.py
缓存: python ../scripts/recalc.py <out>
校验: python ../scripts/validate_valuation.py <out>  → verdict ≠ FAIL 方可交付
"""
from __future__ import annotations
import os, json, shutil
from openpyxl import Workbook
import build_kit as K

# ════════════ 0. 全局轴 ════════════
ALLC = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
ALLY = ["2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
HC, HY = ["B", "C", "D", "E", "F"], ["2021", "2022", "2023", "2024", "2025"]
FC = ["F", "G", "H", "I", "J", "K"]      # 含基年 F=2025A
FCf = FC[1:]                             # 纯前瞻 2026E-2030E
FWY = ["2026E", "2027E", "2028E", "2029E", "2030E"]
CASES = ["Bear", "Base", "Bull"]

RMBUSD = 7.2                             # RMB/USD, 财务折 $B
FX = 7.8                                 # HKD/USD, 只用于股价/每股换算
FX_HIST = [FX] * 5
PX_NOW = 88.7                            # 2026-07-10 收盘(HKD)
SH_NOW = 503.0                           # 现总股本(mn) ≈ 5.03 亿股


def ub(yi):
    """亿 RMB → USD $B。亿RMB ÷ 72。"""
    return None if yi is None else round(yi / (RMBUSD * 10), 4)


S_COVER, S_HIST, S_PX, S_CONS = "封面", "历史财务与估值", "股价走势", "卖方研报共识"
S_HMULT, S_MULT, S_SW = "历史估值倍数", "估值倍数假设", "情景切换"
S_ANCHOR, S_SEG, S_FUND = "物理锚-全球人形出货", "分部测算", "利润与收入假设"
S_VAL, S_CMP, S_DASH = "情景估值", "估值对比", "综合判断仪表盘"

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.environ.get("VALUATION_OUTPUT_DIR", os.path.join(REPO_ROOT, "out"))
VAULT = OUTPUT_DIR  # 向后兼容实例内部变量名
OUT_XLSX = os.path.join(OUTPUT_DIR, "09880_valuation_model.xlsx")
OUT_JSON = os.path.join(OUTPUT_DIR, "09880_input.json")
LOCAL_XLSX = os.path.join(os.path.dirname(__file__), "..", "out", "09880.HK_valuation_model.xlsx")

# ════════════ 公司数据(历史真实, RMB 亿 → $B)2021-2025A ════════════
HUM_H_YI  = [0.1, 0.2, 0.6, 2.5, 8.21]          # 人形机器人收入(亿RMB; 2021-2024 拆分为估, 2025A=全尺寸人形 8.21)
TRAD_H_YI = [8.07, 9.88, 9.96, 10.55, 11.80]    # 传统业务合计(亿RMB; = 总营收 − 人形)
TOT_H_YI  = [round(HUM_H_YI[i] + TRAD_H_YI[i], 2) for i in range(5)]  # 8.17/10.08/10.56/13.05/20.01 ✓
NI_H_YI   = [-9.20, -9.75, -12.34, -11.24, -7.03]   # 归母净亏(亿RMB)
EQ_H_YI   = [15, 12, 20, 20, 50]                    # 股东权益(亿RMB, 估)
SH_H      = [390, 390, 420, 460, 500]               # 期末股本(mn)
GM_H      = [0.313, 0.292, 0.315, 0.287, 0.377]     # 毛利率
NM_H      = [round(NI_H_YI[i] / TOT_H_YI[i], 4) for i in range(5)]   # 归母净利率(实际)

HUM_H  = [ub(x) for x in HUM_H_YI]
TRAD_H = [ub(x) for x in TRAD_H_YI]
NI_H   = [ub(x) for x in NI_H_YI]
EQ_H   = [ub(x) for x in EQ_H_YI]

# 年末股价(HKD; IPO 2023-12-29 @90, 2021-2022 无市价)
PX_END = [None, None, 90, 85, 120]
PX_AVG = [None, None, 90, 88, 105]

# 前瞻: 传统业务(三案共用, 亿RMB → $B) + 股本(mn, 三案共用, +3%/yr) + 毛利率(Base 路径)
TRAD_F_YI = [12.2, 12.6, 13.0, 13.4, 13.8]
TRAD_F = [ub(x) for x in TRAD_F_YI]
SH_F = [515, 530, 546, 563, 580]
GM_F = [0.42, 0.44, 0.45, 0.45, 0.46]

# 全球人形出货 TAM(台; 2021-2023 n.m.)
TAM = [None, None, None, 18000, 20000, 60000, 120000, 200000, 300000, 400000]

# ════════════ 三情景承重假设(主 agent 拍定, 勿改)════════════
# 工业线出货(台) — 前瞻 2026E-2030E
IND_U = {"Bear": [1800, 3000, 4500, 6500, 9000],
         "Base": [2200, 4000, 6500, 10000, 16000],
         "Bull": [3000, 6000, 10000, 18000, 28000]}
# 消费线(U1)出货(台)
CON_U = {"Bear": [500, 2000, 3500, 5500, 9000],
         "Base": [1300, 5500, 8500, 15000, 24000],
         "Bull": [3000, 8000, 14000, 27000, 42000]}
# 工业 ASP(RMB 万/台; 2025A 基年 = 76 三案共用)
IND_ASP = {"Bear": [58, 48, 42, 37, 33],
           "Base": [60, 52, 46, 41, 37],
           "Bull": [64, 57, 51, 46, 42]}
# 消费 ASP(RMB 万/台)
CON_ASP = {"Bear": [13, 12, 11, 10, 9],
           "Base": [15, 14, 13, 12, 11],
           "Bull": [16, 15, 14, 13, 12]}
# 归母净利率(归母净利/总营收) — 决定盈利拐点
NM = {"Bear": [-0.30, -0.18, -0.08, -0.02, 0.02],
      "Base": [-0.22, -0.05, 0.04, 0.08, 0.11],
      "Bull": [-0.12, 0.05, 0.12, 0.15, 0.17]}
# 历史列(实际, 三案同值)
IND_U_H  = [None, None, None, None, 1079]     # 2025A 全尺寸人形 1079 台(全部计工业线)
CON_U_H  = [None, None, None, None, 0]        # 消费线 U1 = 2026 起
IND_ASP_H = [None, None, None, None, 76]      # 2025A 工业 ASP 76 万
CON_ASP_H = [None, None, None, None, None]    # 无消费线 → n.m.

# 终局目标 P/S(主镜头, 应用于前瞻销售额) → 三层分解: 基准锚 11x × 溢价 1.0 × 情绪值
PS_ANCHOR, PS_PREM = 11.0, 1.0
PS_TARGET = {"Bear": 6.0, "Base": 11.0, "Bull": 16.0}
PS_SENT = {c: [round(PS_TARGET[c] / PS_ANCHOR, 4)] * 5 for c in CASES}  # 0.5455 / 1.0 / 1.4545


def hg(vals):
    return [None] + [round(vals[i] / vals[i - 1] - 1, 4) if vals[i - 1] else None
                     for i in range(1, len(vals))]


def build():
    wb = Workbook(); wb.remove(wb["Sheet"])

    # ═══ 1 封面 ═══
    K.write_cover(wb.create_sheet(S_COVER), {
        "title": "优必选 (09880.HK) 估值模型 — 物理锚 × EV/Sales(P/S) 主镜头",
        "meta": [
            ("报告日期", "2026-07-10"),
            ("数据截止", "FY2021-2025 业绩公告 + 2026-07-10 行情(现价 88.7 HKD / 市值约 446.5 亿 HKD)"),
            ("现价 / 市值", "88.7 HKD / 446.5 亿 HKD(≈$5.72B);总股本约 5.03 亿股;PE(TTM)负;52 周 75.5-161.0 HKD"),
            ("口径说明", "财务量纲一律 USD $B(RMB 亿 ÷ 72 折算, RMB/USD=7.2);股价 HKD, FX=HKD/USD=7.8 只用于每股换算。"),
            ("主线镜头", "EV/Sales(P/S)主镜头(未盈利高增长 → P/E·P/B 会出 N/M);路径 P/E 交叉验证(盈利转正后体检)。"),
            ("方法一句话", "全球人形出货 TAM → 优必选工业线 + 消费线(各 units × ASP) + 传统业务 → 总营收 → 目标 P/S × 前瞻每股营收 × FX → 隐含价(HKD);归母净利率驱动 P/E 交叉验证。"),
        ],
        "takeaways": [
            ("① 物理锚", "全球人形出货 TAM 2025 约 2 万台 → 2030E 40 万台;优必选出货拆 工业线 + 消费线(U1), 各 units × ASP。改出货量/ASP → 隐含价全链动。"),
            ("② 收入结构", "2025A 总营收 20.01 亿 RMB(≈$0.28B), 全尺寸人形 8.21 亿(1079 台、ASP 76 万)+ 传统业务 11.80 亿;人形从 2026 起靠工业交付放量。"),
            ("③ 差异化判断(审慎)", "Base 2026E 总营收约 27.4 亿 / 2027E 约 41.1 亿 RMB, 刻意低于卖方共识(2026E 36.9 / 2027E 55.2 亿)约 25% — 订单≠交付、消费 ASP 稀释、执行风险。"),
            ("④ 三情景 2027E 隐含价", "Bear 约 36 HKD(-59%) / Base 约 92 HKD(+4%) / Bull 约 192 HKD(+117%) vs 现价 88.7。P/S 6x/11x/16x 相乘出货放量 → 区间宽。"),
            ("⑤ 盈利拐点", "Base 2028 归母净利转正、Bull 2027 转正、Bear 期内基本不转正;主线 EV/Sales 出实数, P/E 转正后做体检。"),
        ],
    })

    # ═══ 2 历史财务与估值($B;股价 HKD)═══
    ha = K.write_history(wb.create_sheet(S_HIST), {
        "title": "优必选 历史财务与估值 ($B) — 2021-2025A + 当下(IPO 2023-12);财务 RMB 亿 ÷ 72",
        "hist_cols": HC, "hist_years": HY,
        "fx_hist": FX_HIST, "fx_now": FX,
        "vals_in_usd": True,                 # 值已折为 $B, 不再换算; FX 行仅供股价换算
        "fx_label": "FX (HKD/USD, 股价换算)",
        "cur_label": "当下(2025A/最新)",
        "segments": [
            ("人形机器人", HUM_H, True),
            ("传统业务", TRAD_H, True),
        ],
        "total_now": ub(20.01),
        "gm_pct": GM_H, "gm_now": 0.377,
        "ni": NI_H, "ni_now": ub(-7.03),
        "eq": EQ_H, "eq_now": ub(50),
        "shares": SH_H, "shares_now": SH_NOW,
        "px_end": PX_END, "px_now": PX_NOW,
        "px_avg": PX_AVG,
        "band_note": "IPO 2023-12-29 @90 HKD;2021-2022 无公开市价 → P/E·P/B·市值 n.m.。公司持续亏损 → 历史 P/E 全 N/M, 历史估值带用 P/S(下方 P/S 行)。",
        "notes": [
            ("人形机器人", "全尺寸人形(Walker 系列)。2025A 8.21 亿 RMB(1079 台 × ASP 76 万);2021-2024 为拆分估计(全尺寸人形 2024 前体量极小)。原始 RMB 亿 ÷ 72 折 $B。"),
            ("传统业务", "教育 + 物流 + 消费级及硬件 + 其他(2025A 合计 11.80 亿 RMB = 教育4.13+物流2.74+消费及硬件5.23−非具身0.48等)。= 总营收 − 人形。"),
            ("HGMP", "毛利率:2021-2025 = 31.3/29.2/31.5/28.7/37.7%;2025 回升主因人形与高毛利产品占比上升。来源:业绩公告。"),
            ("HNI", "归母净亏(亿RMB):-9.20/-9.75/-12.34/-11.24/-7.03。2025 亏损收窄。折 $B。人形机器人无历史正利润 → P/E 无意义。"),
            ("HEQ", "股东权益为估(招股书/年报口径不一);2025 约 50 亿 RMB(含配售)。主线 EV/Sales 不依赖账面, P/B 仅参照。"),
            ("HSH", "期末股本(mn):390/390/420(IPO)/460/500;当下 503。"),
            ("HPX", "年末股价(HKD):IPO 2023-12-29 @90 → 2024 末约 85 → 2025 末约 120;2025-10 高见 161。当下现价 88.7(2026-07-10)。2021-2022 无市价。"),
        ],
    })
    # 上市前 P/E·P/B·市值 n.m.(2021-2022 无市价)+ 全年份 P/E n.m.(持续亏损)
    ws_h = wb[S_HIST]
    for col in ["B", "C"]:
        K.lab(ws_h, f"{col}{ha['HPX']}", "n.m.", note=True)
        K.lab(ws_h, f"{col}{ha['HMC']}", "n.m.", note=True)
        K.lab(ws_h, f"{col}{ha['HPB']}", "n.m.", note=True)

    # 追加 每股营收 SPS(USD) + P/S(实际) — EV/Sales 主线的历史底座
    HSPS = ws_h.max_row + 1
    K.lab(ws_h, f"A{HSPS}", "SPS 销售额/股(USD)", note=True)
    for c in HC:
        K.fml(ws_h, f"{c}{HSPS}", f"={c}{ha['HREV']}*1000/{c}{ha['HSH']}", K.N2)
    K.fml(ws_h, f"G{HSPS}", f"=G{ha['HREV']}*1000/G{ha['HSH']}", K.N2)
    HPS = HSPS + 1
    K.lab(ws_h, f"A{HPS}", "P/S (实际, 主线底座)", b=True); ws_h[f"A{HPS}"].fill = K.OUT
    for c in HC:
        if c in ("B", "C"):
            K.lab(ws_h, f"{c}{HPS}", "n.m.", note=True)
        else:
            K.fml(ws_h, f"{c}{HPS}", f"={c}{ha['HPX']}/({c}{HSPS}*{c}{ha['HFX']})", K.MX)
    K.fml(ws_h, f"G{HPS}", f"=G{ha['HPX']}/(G{HSPS}*G{ha['HFX']})", K.MX); ws_h[f"G{HPS}"].fill = K.CUR
    K.logic(ws_h, f"H{HPS}", "= 年末股价 ÷ (每股营收USD × FX)。2023 约 33x → 2024/2025 约 28x → 当下 TTM 约 20.7x(随收入放大递减)。持续亏损故用 P/S 作历史估值带, P/E 全 N/M。")

    # ═══ 3 股价走势(月末收盘)═══
    MONTHLY = [
        ("2023-12", 90), ("2024-06", 95), ("2024-12", 85), ("2025-06", 76),
        ("2025-10", 161), ("2025-12", 120), ("2026-03", 115), ("2026-06", 98), ("2026-07", 88.7),
    ]

    def phase_fn(ym):
        if ym <= "2024-12":
            return "① IPO 与磨底"
        if ym <= "2025-09":
            return "② 人形叙事发酵"
        if ym <= "2025-12":
            return "③ 冲高 161"
        return "④ 回撤消化"
    px = K.write_price_chart(wb.create_sheet(S_PX), MONTHLY, {
        "fn": phase_fn,
        "rows": [("① IPO 与磨底", "2023-12-29 IPO @90 HKD;2024 全年在 85-95 区间磨底。"),
                 ("② 人形叙事发酵", "2025 全球人形机器人主题升温, Walker S 工业交付订单落地, 股价从 76 反弹。"),
                 ("③ 冲高 161", "2025-10 人形 FOMO 见顶 161 HKD;2025-12 回落约 120。"),
                 ("④ 回撤消化", "2026 上半年在 98-115 震荡, 2026-07-10 报 88.7;等交付兑现。")],
    }, title="优必选 09880.HK 月末收盘 (HKD, IPO 2023-12)")
    px["yhigh"].update({2025: 161, 2026: 115})
    px["ylow"].update({2025: 76, 2026: 88.7})

    # ═══ 4 卖方研报共识 ═══
    K.write_consensus(wb.create_sheet(S_CONS), {
        "title": "卖方研报共识 — 10 家强推;目标均价 153 HKD(区间 125-175)",
        "overview": "10 家 strong_buy, 目标均价 153 HKD(125-175, vs 现价 88.7 约 +72%)。分歧集中在两点:给 2027E 多少 P/S, 以及信不信 2027 转正。本模型 Base 目标 P/S 11x、盈利拐点 2028, 均比中位共识审慎;收入端 Base 2026E/2027E 刻意低于共识约 25%(订单≠交付)。",
        "assumptions": [
            ("2026E 收入", "卖方共识 2026E 约 36.9 亿 RMB(人形订单快速转交付)。",
             "在手订单转交付节奏 + 消费线 U1 上量时点。",
             "本模型 Base 约 27.4 亿(低约 25%):全尺寸人形爬产+良率+回款周期, 消费 ASP 稀释。差异化判断, 不上调贴共识。"),
            ("2027E 收入", "卖方共识 2027E 约 55.2 亿 RMB(同比翻倍以上)。",
             "工业线放量斜率 + 消费线渗透 + ASP 能否守住。",
             "本模型 Base 约 41.1 亿(低约 25%):Base 只放量不假设 ASP 上行;激进量价同涨情形放进 Bull。"),
            ("2027 是否转正", "部分卖方假设 2027 归母净利转正支撑高估值。",
             "规模效应 vs 研发/产能投入的赛跑。",
             "本模型 Base 拐点 2028(2027 净利率 -5%)、Bull 2027 转正、Bear 期内基本不转正。"),
            ("目标 P/S(2027E)", "HSBC 126@9.7x 2027e P/S;中金 138@18x;国泰海通 185@22x。",
             "给 2027E 多少 P/S = 分歧核心;区间 9.7-22x。",
             "本模型 Base 目标 P/S 11x(近 HSBC 一侧, 远低于中金/国泰海通), 只计入交付放量、不为量价同涨溢价。"),
        ],
        "divergences": [
            "① 2027E P/S 给多少:HSBC 9.7x vs 国泰海通 22x, 相差一倍多 —— 决定目标价上下限的最大分歧。",
            "② 2027 转正与否:信 2027 转正者敢给高 P/S;本模型判断拐点在 2028, 故 Base P/S 只给 11x。",
            "③ 订单转交付:共识把在手订单近似线性转为收入, 本模型对交付节奏打约 25% 折扣。",
        ],
        "stances": [
            "HSBC:目标 126 HKD, 9.7x 2027e P/S(共识内最审慎的倍数口径)。",
            "中金:目标 138 HKD, 18x 2027e P/S。",
            "国泰海通:目标 185 HKD, 22x 2027e P/S(共识上沿)。",
            "覆盖总体:10 家 strong_buy, 目标均价 153 HKD, 区间 125-175 —— 一致看多, 分歧在倍数与拐点。",
            "本模型:Base 目标 P/S 11x、拐点 2028, 收入低共识约 25% —— 明确审慎于中位共识, 差异化在'订单≠交付+消费 ASP 稀释+执行风险'。",
        ],
    })

    # ═══ 5 历史估值倍数(P/S 口径数据底座, 手写)═══
    ws_hm = wb.create_sheet(S_HMULT)
    K.hdr(ws_hm, 1, "历史估值倍数 — EV/Sales(P/S)口径: 自身历史带 + 当下 + 人形/高成长可比", 11)
    _r = K.mtext(ws_hm, 2, "公司持续亏损 → P/E 全 N/M, 主线用 EV/Sales(P/S)。本页给数据底座: ① 自身逐年实际 P/S 带(2023 IPO 后) + 当下 TTM;② 人形/机器人/高成长硬件可比 P/S 光谱。看完本页再去下一页拍三案目标 P/S。", "K", 2)
    K.band(ws_hm, _r, "① 自身: 逐年实际 P/S(链自『历史财务与估值』)+ 当下 TTM", 11); _r += 1
    ws_hm[f"A{_r}"] = "指标"; ws_hm[f"A{_r}"].font = K.BF
    for col, y in zip(HC, HY):
        ws_hm[f"{col}{_r}"] = y; ws_hm[f"{col}{_r}"].font = K.BF; ws_hm[f"{col}{_r}"].fill = K.CH
    ws_hm[f"G{_r}"] = "当下"; ws_hm[f"G{_r}"].font = K.BF; ws_hm[f"G{_r}"].fill = K.CUR
    ws_hm[f"L{_r}"] = "说明"; ws_hm[f"L{_r}"].font = K.BF; ws_hm[f"L{_r}"].fill = K.CH
    _r += 1
    K.lab(ws_hm, f"A{_r}", "年末股价(HKD)")
    for c in HC:
        K.fml(ws_hm, f"{c}{_r}", f"={K.R(S_HIST, c + str(ha['HPX']))}", K.PX, link=True)
    K.fml(ws_hm, f"G{_r}", f"={K.R(S_HIST, 'G' + str(ha['HPX']))}", K.PX, link=True); ws_hm[f"G{_r}"].fill = K.CUR
    _r += 1
    K.lab(ws_hm, f"A{_r}", "P/S(实际, 主线底座)", b=True); ws_hm[f"A{_r}"].fill = K.OUT
    for c in HC:
        K.fml(ws_hm, f"{c}{_r}", f"={K.R(S_HIST, c + str(HPS))}", K.MX, link=True)
    K.fml(ws_hm, f"G{_r}", f"={K.R(S_HIST, 'G' + str(HPS))}", K.MX, link=True); ws_hm[f"G{_r}"].fill = K.CUR
    K.logic(ws_hm, f"L{_r}", "自身 P/S 带: 2023 约 33x → 2024/2025 约 28x → 当下 TTM 约 20.7x。趋势下行=收入放大摊薄倍数。前瞻目标 P/S 应显著低于当下 TTM(前瞻销售额已含放量)。")
    _r += 2
    K.band(ws_hm, _r, "② 可比 P/S 光谱(人形/机器人/高成长硬件;公开口径)", 11); _r += 1
    for col, h in zip("ABCDEF", ["公司/类型", "P/S", "口径", "", "", "业务特征/来源"]):
        ws_hm[f"{col}{_r}"] = h; ws_hm[f"{col}{_r}"].font = K.BF; ws_hm[f"{col}{_r}"].fill = K.CH
    ws_hm.merge_cells(f"F{_r}:K{_r}")
    _r += 1
    _peers = [
        ("优必选 09880.HK", "20.7x / 11x", "TTM / 2027E Base",
         "本模型标的;当下 TTM 约 20.7x(现价÷2025 收入), 前瞻 Base 目标 11x(对 2027E 前瞻销售)。"),
        ("特斯拉 TSLA(整体)", "8-10x", "fwd",
         "含 Optimus 人形期权的整车+能源龙头;人形纯期权无独立 P/S, 作高成长硬件参照上沿。"),
        ("工业机器人(发那科/ABB 等)", "2-4x", "TTM",
         "成熟工业机器人 P/S 区间 = 无人形叙事时的估值地板。"),
        ("A 股人形概念(整机/零部件)", "8-20x", "fwd/主题",
         "国内人形主题标的主题溢价区间;波动大、随情绪。"),
        ("高成长 AI 硬件(未盈利)", "10-16x", "fwd",
         "未盈利高增长硬件在放量期市场愿给的前瞻销售倍数区间 = 本模型 Base/Bull 参照。"),
    ]
    for i, (nm, ps, kou, note) in enumerate(_peers):
        rr = _r + i
        K.lab(ws_hm, f"A{rr}", nm, b=(i == 0))
        if i == 0:
            ws_hm[f"A{rr}"].fill = K.CUR
        K.lab(ws_hm, f"B{rr}", ps); K.lab(ws_hm, f"C{rr}", kou, note=True)
        ws_hm.merge_cells(f"F{rr}:K{rr}"); K.logic(ws_hm, f"F{rr}", note)
        ws_hm.row_dimensions[rr].height = 28
    _r += len(_peers) + 1
    K.band(ws_hm, _r, "③ 读法 — 给『估值倍数假设』的输入", 11); _r += 1
    K.mtext(ws_hm, _r, "① 自身 TTM 约 20.7x 是对已发生(低)收入的倍数, 前瞻销售额放量后目标倍数自然压缩;② 成熟工业机器人 2-4x 是地板, 高成长未盈利硬件 10-16x 是放量期区间, 人形主题溢价可上探 20x;③ 结论: 基准锚取 11x(高成长硬件放量期中枢), 三案 = 基准锚 × 情绪值 → Bear 6x(叙事退潮/交付不及)/ Base 11x / Bull 16x(量价同涨+主题溢价)。均低于当下 TTM 20.7x。", "K", 4)
    for col, w in zip("ABCDEFGHIJK", [22, 11, 11, 9, 9, 9, 9, 9, 9, 9, 9]):
        ws_hm.column_dimensions[col].width = w
    ws_hm.column_dimensions["L"].width = 58

    # ═══ 6 估值倍数假设(EV/Sales 主线三层 → 三案目标 P/S)═══
    ma = K.write_multiple_assumptions(wb.create_sheet(S_MULT), {
        "title": "估值倍数假设 — EV/Sales(P/S)主线: 目标 P/S = 基准锚 × 结构溢价 × 情绪值(三案)",
        "intro": "数据底座在上一页。本页拍三案目标 P/S(对前瞻销售额): 目标 P/S = 基准锚 11x × 结构溢价 1.0 × 情绪值。『情景切换』引用切换, 『情景估值』套用当前案。P/S 主镜头 + 路径 P/E 交叉验证(≥2 镜头)。",
        "why_text": ("镜头选择是业务判断: 优必选 2027 前(Base)无盈利, P/E 失效、账面被巨亏消耗 P/B 失真、FCF 深负 DCF 全靠终值 —— 穿越当下持续存在并增长的是'收入及其背后的人形出货量/ASP', 故 EV/Sales(P/S, 净现金公司≈P/S)做主线, 强制配 path to profitability(Base 2028 归母净利转正, 见『利润与收入假设』)。"
                     "P/E 做支线交叉验证: 隐含价 ÷ 前瞻 EPS, 转正前标 N/M, 2028E+ 起用 30-40x 路径 P/E 体检隐含倍数是否荒谬。"
                     "镜头迁移触发: 当归母净利连续两年为正, 主线应迁往 P/E —— 届时市场也会换镜头(re-rating)。"),
        "why_rows": 6,
        "method_text": "三层分解: ①基准锚 11x = 高成长未盈利硬件放量期 forward P/S 中枢(工业机器人 2-4x 地板之上、人形主题 20x 之下), 也是当下 TTM 20.7x 随放量压缩后的合理前瞻位置;②结构溢价 1.0x(不额外加溢价, 人形主题溢价已含在情绪值层);③情绪值(三案): 表达市场为人形放量确定性愿付多少。一致性: 基准锚×溢价×情绪 = Base 11x / Bear 6x / Bull 16x, 恒定应用于前瞻销售额(放量→隐含价, 非倍数扩张)。",
        "peak": PS_ANCHOR, "peak_note": "基准锚 11x: 高成长未盈利硬件放量期 forward P/S 中枢;= 当下 TTM 20.7x 在前瞻销售放量后压缩到的合理位置, 高于工业机器人 2-4x 地板、低于人形主题 20x。",
        "premium": PS_PREM, "premium_note": "结构溢价 1.0x: 不在此层额外加溢价(人形主题/稀缺溢价放进情绪值), 避免双算。",
        "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
        "s_hist": S_HIST, "hpb_row": HPS,
        "cases": [
            ("Bear", PS_SENT["Bear"], "情绪值 0.55 → 目标 P/S 6x: 人形叙事退潮 + 交付不及 + 消费线 ASP 战 → 倍数向工业机器人区间收敛。恒定应用于各前瞻年销售额。"),
            ("Base", PS_SENT["Base"], "情绪值 1.0 → 目标 P/S 11x: 付足'高成长硬件放量期'基准, 靠交付放量而非倍数扩张长进估值。低于当下 TTM 20.7x 与中金/国泰海通 18-22x。"),
            ("Bull", PS_SENT["Bull"], "情绪值 1.45 → 目标 P/S 16x: 量价同涨 + 人形主题溢价 + 主权/工业大单 → 市场按主题龙头给 16x(仍低于 2025-10 情绪峰值隐含)。"),
        ],
        "sent_note": "情绪值 = 目标 P/S ÷ 基准锚(11x)。1.0=付足基准;>1=主题溢价;<1=退潮折价。三案 = 0.55/1.0/1.45 → 目标 P/S 6/11/16x, 逐年恒定(承重假设为'终局'倍数, 应用于每个前瞻年销售额)。历史列=实际 P/S ÷ 11 反推(参考)。",
        "target_note": "同一三层公式套三案情绪。历史列 = 实际 P/S(回看, 三案同值)。前瞻 = 11x × 1.0 × 情绪值 = 6/11/16x。",
        "reconcile_text": "卖方 2027E P/S: HSBC 9.7x / 中金 18x / 国泰海通 22x。本模型 Base 11x —— 近 HSBC 一侧、远低于中金/国泰海通。凭什么审慎: ①当下自身 TTM 20.7x 对应的是低基数收入, 前瞻销售放量后 20x+ 隐含'倍数与收入同高', 双重乐观;②2027 本模型判断未转正(Base 净利率 -5%), 未盈利公司给 18-22x 前瞻 P/S 需要 2027 就转正+高增速同时成立;③收入端已低共识约 25%, 若再叠高倍数=贴现两次乐观。Bull 16x 已容纳量价同涨情形。",
        "source_text": "基准锚/可比 P/S 来自上一页(公开口径 + 人形/机器人光谱);三案情绪值为承重假设(主 agent 拍定);卖方 P/S(HSBC 9.7x/中金 18x/国泰海通 22x)见『卖方研报共识』。",
    })

    # ═══ 7 情景切换 ═══
    sw = K.write_scenario_switch(wb.create_sheet(S_SW), {
        "title": "情景切换 — 全模型唯一情景参数库 + 切换开关(默认 Base)",
        "usage": "B2 是唯一开关。工业/消费出货台数、工业/消费 ASP、归母净利率按当前案联动;目标 P/S 情绪值 link『估值倍数假设』。整条明细链(物理锚→分部→利润→倍数→估值)变档。传统业务、股本、毛利率三案共用(非 load-bearing)。『估值对比』引三案矩阵行不被开关污染。",
        "cases": CASES, "default": "Base",
        "triggers": [
            ("Bear", "全球人形出货不及预期 + 优必选工业交付爬产慢(良率/回款)、消费线 U1 上量迟;价格战压 ASP;2027 仍亏、期内基本不转正;目标 P/S 收敛至 6x。"),
            ("Base", "工业线按在手订单稳步交付(2027E 4000 台)、消费线渐上量;ASP 随规模温和下行(工业 60→37 万);2028 归母净利转正;目标 P/S 11x。"),
            ("Bull", "全球人形产业爆发 + 优必选工业/消费双线放量(2030E 工业 2.8 万 + 消费 4.2 万台)、ASP 守住;2027 转正、2030 净利率 17%;市场按主题龙头给 16x。"),
        ],
        "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
        "levers": [
            {"key": "ind", "name": "工业线出货(台)", "fmt": K.N0, "cols": FCf,
             "vals": IND_U, "desc": "人形主引擎(工业)。工业收入 = 出货 × 工业ASP。改这行 → 人形收入 → 总营收 → 隐含价全链动。2025A 全尺寸人形 1079 台。",
             "stories": {"Bear": "爬产/良率/回款拖累, 2030E 仅 0.9 万台。", "Base": "在手订单稳步交付, 2027E 4000 台 → 2030E 1.6 万台。", "Bull": "工业场景爆发, 2030E 2.8 万台。"},
             "hist": IND_U_H},
            {"key": "con", "name": "消费线(U1)出货(台)", "fmt": K.N0, "cols": FCf,
             "vals": CON_U, "desc": "人形第二引擎(消费 U1)。消费收入 = 出货 × 消费ASP。2026 起从 0 起量, 弹性大。",
             "stories": {"Bear": "消费场景培育慢, 2030E 0.9 万台。", "Base": "渐进渗透, 2027E 5500 → 2030E 2.4 万台。", "Bull": "消费级放量, 2030E 4.2 万台。"},
             "hist": CON_U_H},
            {"key": "indasp", "name": "工业线ASP(万元/台)", "fmt": K.N1, "cols": FCf,
             "vals": IND_ASP, "desc": "工业人形单价(RMB 万元/台)。随规模效应/竞争逐年下行。2025A = 76 万(三案基年)。",
             "stories": {"Bear": "价格战, 2030E 33 万。", "Base": "规模温和降本, 2030E 37 万。", "Bull": "高端守价, 2030E 42 万。"},
             "hist": IND_ASP_H},
            {"key": "conasp", "name": "消费线ASP(万元/台)", "fmt": K.N1, "cols": FCf,
             "vals": CON_ASP, "desc": "消费人形单价(RMB 万元/台)。消费级本就低价且稀释。2026E 起。",
             "stories": {"Bear": "低价走量, 2030E 9 万。", "Base": "2030E 11 万。", "Bull": "溢价能力较强, 2030E 12 万。"},
             "hist": CON_ASP_H},
            {"key": "nm", "name": "归母净利率(%)", "fmt": K.PCT, "cols": FCf,
             "vals": NM, "desc": "决定盈利拐点。归母净利 = 总营收 × 本行。喂 P/E 交叉验证。历史 -35% 至 -117%(巨亏收窄)。",
             "stories": {"Bear": "规模上不去+投入刚性, 2030 仅 +2%。", "Base": "2028 转正(+4%), 2030 +11%。", "Bull": "2027 转正(+5%), 2030 +17%。"},
             "hist": NM_H},
        ],
        "linked": [
            {"key": "sent", "name": "情绪值(目标 P/S 第三层)", "fmt": K.N2,
             "src_sheet": S_MULT, "src_row0": ma["sent_row0"],
             "note": "三案取值与依据见『估值倍数假设』;目标 P/S = 基准锚 11x × 溢价 1.0 × 本行。"},
        ],
    })
    SWACT, SWB = sw["SWACT"], sw["SWB"]
    # 目标 P/S(当前案) = 锚 × 溢价 × 当前案情绪
    _pk = f"'{S_MULT}'!{ma['pk_cell']}"
    _pr = f"'{S_MULT}'!{ma['pr_cell']}"
    SWPS = sw["next_row"]
    K.lab(wb[S_SW], f"A{SWPS}", "目标 P/S(当前案)", b=True)
    for c in HC:
        K.lab(wb[S_SW], f"{c}{SWPS}", "n.m.", note=True)
    for c in FCf:
        K.fml(wb[S_SW], f"{c}{SWPS}", f"={_pk}*{_pr}*{c}{SWACT['sent']}", K.MX, link=True)
    K.logic(wb[S_SW], f"L{SWPS}", "= 基准锚 11x × 结构溢价 1.0x × 当前案情绪值 → 喂『情景估值』前瞻目标 P/S。")

    # ═══ 8 物理锚: 全球人形出货 TAM + 优必选出货 ═══
    anchor = K.write_anchor(wb.create_sheet(S_ANCHOR), {
        "title": "全球人形机器人出货量(台, TAM)+ 优必选工业/消费出货 — 需求物理盘子",
        "all_cols": ALLC, "all_years": ALLY,
        "series": [
            ("全球人形出货 TAM(台)", TAM,
             "GGII/IDC/MS 出货口径中枢: 2024 约 1.8 万 → 2025 约 2 万 → 2026E 6 万 → 2027E 12 万 → 2028E 20 万 → 2029E 30 万 → 2030E 40 万台。三案共用盘子。", K.N0),
            ("优必选工业线出货(台)", IND_U_H + [None] * 5,
             "2025A 全尺寸人形 1079 台;前瞻 = 『情景切换』当前案。", K.N0),
            ("优必选消费线出货(台)", CON_U_H + [None] * 5,
             "消费 U1 2026 起量;前瞻 = 『情景切换』当前案。", K.N0),
            ("优必选合计出货(台)", [None] * 10, "= 工业 + 消费。", K.N0),
            ("优必选全球出货占比(%)", [None] * 10, "= 优必选合计 ÷ 全球 TAM(隐含市占, 读数)。", K.PCT),
        ],
        "yoy_row": "全球人形出货 TAM(台)",
        "source_note": "口径 = 全球人形机器人年出货量(台)。来源: GGII/IDC/MS 等出货口径中枢。优必选出货为本公司工业+消费两线, 前瞻由『情景切换』承重假设驱动。",
        "role_note": "作用: 优必选工业/消费出货 × ASP = 人形收入 → 总营收 → 隐含价(改出货量/ASP → 隐含价全链动)。全球 TAM 为需求盘子背景, 份额行为隐含市占读数。",
    })
    aro = anchor["row_of"]
    TAM_R = aro["全球人形出货 TAM(台)"]
    IND_R = aro["优必选工业线出货(台)"]
    CON_R = aro["优必选消费线出货(台)"]
    TOT_R = aro["优必选合计出货(台)"]
    SHR_R = aro["优必选全球出货占比(%)"]
    wa = wb[S_ANCHOR]
    # 前瞻优必选出货 = 情景切换当前案
    for c in FCf:
        K.fml(wa, f"{c}{IND_R}", f"={K.R(S_SW, c + str(SWACT['ind']))}", K.N0, link=True)
        K.fml(wa, f"{c}{CON_R}", f"={K.R(S_SW, c + str(SWACT['con']))}", K.N0, link=True)
    # 合计 + 份额(2025A 起, 有 TAM 的列)
    for c in ["F"] + list(FCf):
        K.fml(wa, f"{c}{TOT_R}", f"={c}{IND_R}+{c}{CON_R}", K.N0)
        K.fml(wa, f"{c}{SHR_R}", f"={c}{TOT_R}/{c}{TAM_R}", K.PCT)

    # ═══ 9 分部测算 ═══
    seg = K.write_segment_model(wb.create_sheet(S_SEG), {
        "title": "分部测算 — 人形(工业+消费, units×ASP) + 传统业务",
        "all_cols": ALLC, "all_years": ALLY, "logic_col": "N",
        "groups": [
            ("物理锚引入(引『物理锚』出货)", [
                ("工业线出货(台)", None, K.N0, "引『物理锚』优必选工业出货。"),
                ("消费线出货(台)", None, K.N0, "引『物理锚』优必选消费出货。"),
            ]),
            ("人形-工业线 = 出货 × ASP", [
                ("工业ASP(万元/台)", None, K.N1, "历史 2025A=76(引情景切换 hist);前瞻=『情景切换』当前案。"),
                ("工业收入($B)", None, K.N1, "= 出货 × ASP(万元) ÷ 720000(万元→$B: ÷10000 得亿 ÷72 得$B)。历史=引『历史财务』人形(全部计工业)。"),
            ]),
            ("人形-消费线 = 出货 × ASP", [
                ("消费ASP(万元/台)", None, K.N1, "前瞻=『情景切换』当前案;2025A 前无消费线。"),
                ("消费收入($B)", None, K.N1, "= 出货 × ASP ÷ 720000。2025A 前 = 0。"),
            ]),
            ("汇总", [
                ("人形机器人收入($B)", None, K.N1, "= 工业收入 + 消费收入。"),
                ("传统业务收入($B)", None, K.N1, "历史=引『历史财务』传统;前瞻=输入(三案共用): 2026 12.2→2030 13.8 亿RMB ÷72。"),
                ("总营收($B)", None, K.N1, "= 人形合计 + 传统。喂利润表。"),
            ]),
        ],
    })
    m = seg["m"]
    ws_s = wb[S_SEG]
    HUM_HR = ha["seg_rows"]["人形机器人"]
    TRAD_HR = ha["seg_rows"]["传统业务"]
    # 出货: 引锚(全列)
    for c in ["F"] + list(FCf):
        K.fml(ws_s, f"{c}{m['工业线出货(台)']}", f"={K.R(S_ANCHOR, c + str(IND_R))}", K.N0, link=True)
        K.fml(ws_s, f"{c}{m['消费线出货(台)']}", f"={K.R(S_ANCHOR, c + str(CON_R))}", K.N0, link=True)
    for c in ["B", "C", "D", "E"]:
        K.lab(ws_s, f"{c}{m['工业线出货(台)']}", "n.m.", note=True)
        K.lab(ws_s, f"{c}{m['消费线出货(台)']}", "n.m.", note=True)
    # ASP
    for c in FCf:
        K.fml(ws_s, f"{c}{m['工业ASP(万元/台)']}", f"={K.R(S_SW, c + str(SWACT['indasp']))}", K.N1, link=True)
        K.fml(ws_s, f"{c}{m['消费ASP(万元/台)']}", f"={K.R(S_SW, c + str(SWACT['conasp']))}", K.N1, link=True)
    K.inp(ws_s, f"F{m['工业ASP(万元/台)']}", 76, None, K.N1)
    for c in ["B", "C", "D", "E"]:
        K.lab(ws_s, f"{c}{m['工业ASP(万元/台)']}", "n.m.", note=True)
        K.lab(ws_s, f"{c}{m['消费ASP(万元/台)']}", "n.m.", note=True)
    K.lab(ws_s, f"F{m['消费ASP(万元/台)']}", "n.m.", note=True)
    # 工业收入: 历史引 history 人形(全部计工业), 前瞻 = 出货×ASP/720000
    for c in HC:
        K.fml(ws_s, f"{c}{m['工业收入($B)']}", f"={K.R(S_HIST, c + str(HUM_HR))}", K.N1, link=True)
        K.inp(ws_s, f"{c}{m['消费收入($B)']}", 0, None, K.N1)
    for c in FCf:
        K.fml(ws_s, f"{c}{m['工业收入($B)']}", f"={c}{m['工业线出货(台)']}*{c}{m['工业ASP(万元/台)']}/720000", K.N1)
        K.fml(ws_s, f"{c}{m['消费收入($B)']}", f"={c}{m['消费线出货(台)']}*{c}{m['消费ASP(万元/台)']}/720000", K.N1)
    # 人形合计 / 传统 / 总营收
    for c in ALLC:
        K.fml(ws_s, f"{c}{m['人形机器人收入($B)']}", f"={c}{m['工业收入($B)']}+{c}{m['消费收入($B)']}", K.N1)
        K.fml(ws_s, f"{c}{m['总营收($B)']}", f"={c}{m['人形机器人收入($B)']}+{c}{m['传统业务收入($B)']}", K.N1)
    for c in HC:
        K.fml(ws_s, f"{c}{m['传统业务收入($B)']}", f"={K.R(S_HIST, c + str(TRAD_HR))}", K.N1, link=True)
    for i, c in enumerate(FCf):
        K.inp(ws_s, f"{c}{m['传统业务收入($B)']}", TRAD_F[i], None, K.N1)
    for c in FCf:
        ws_s[f"{c}{m['总营收($B)']}"].fill = K.OUT

    # ═══ 10 利润与收入假设(手写: 总营收 → 归母净利率 → 净利 → EPS/BPS/SPS)═══
    fw = wb.create_sheet(S_FUND)
    K.hdr(fw, 1, "利润与收入假设 — 总营收 → 归母净利率 → 归母净利 → EPS/BPS/SPS(粗颗粒, 不做三表勾稽)", 12)
    fw["A2"] = "假设/口径($B; 每股 USD)"; fw["A2"].font = K.BF
    for col, y in zip(ALLC, ALLY):
        fw[f"{col}2"] = y; fw[f"{col}2"].font = K.BF; fw[f"{col}2"].fill = K.CH
    fw["N2"] = "逻辑/来源"; fw["N2"].font = K.BF; fw["N2"].fill = K.CH
    r = 3
    fund = {}

    K.band(fw, r, "假设(历史实际锚 + 前瞻链情景切换)", 12); r += 1
    fund["NM"] = r
    K.lab(fw, f"A{r}", "归母净利率(%)")
    K.introw(fw, r, HC, NM_H, None, K.PCT)
    for c in FCf:
        K.fml(fw, f"{c}{r}", f"={K.R(S_SW, c + str(SWACT['nm']))}", K.PCT, link=True)
    K.logic(fw, f"N{r}", "历史 = 归母净亏 ÷ 总营收(-35% 至 -117%, 巨亏收窄);前瞻 = 『情景切换』当前案。Base 2028 转正。")
    r += 1
    fund["GM"] = r
    K.lab(fw, f"A{r}", "毛利率(%)", note=True)
    K.introw(fw, r, HC, GM_H, None, K.PCT)
    K.introw(fw, r, FCf, GM_F, None, K.PCT)
    K.logic(fw, f"N{r}", "辅助/展示行(Base 路径, 三案共用): 历史 31-38%, 前瞻 42-46%(人形+高毛利占比升)。主线 P/S 不依赖毛利, 净利率是 P/E 交叉验证的驱动。")
    r += 1
    fund["RET"] = r
    K.lab(fw, f"A{r}", "留存率(%)", note=True)
    K.introw(fw, r, FCf, [1.0] * 5, None, K.PCT)
    for c in HC:
        K.lab(fw, f"{c}{r}", "n.m.", note=True)
    K.logic(fw, f"N{r}", "不分红全额留存;仅用于权益递推/BPS 体检, 主估值(P/S)不依赖 P/B。")
    r += 1
    fund["SH"] = r
    K.lab(fw, f"A{r}", "股本(mn股)", note=True)
    for c in HC:
        K.fml(fw, f"{c}{r}", f"={K.R(S_HIST, c + str(ha['HSH']))}", K.N0, link=True)
    K.introw(fw, r, FCf, SH_F, None, K.N0)
    K.logic(fw, f"N{r}", "历史引『历史财务』;前瞻配售摊薄 +3%/yr(2026 515 → 2030 580 mn, 三案共用)。")
    r += 1

    K.band(fw, r, "收入 → 利润 → 每股(链)", 12); r += 1
    fund["HUM"] = r
    K.lab(fw, f"A{r}", "人形机器人收入($B)")
    for c in ALLC:
        K.fml(fw, f"{c}{r}", f"={K.R(S_SEG, c + str(m['人形机器人收入($B)']))}", K.N1, link=True)
    K.logic(fw, f"N{r}", "引『分部测算』(工业+消费, units×ASP)。")
    r += 1
    fund["TRAD"] = r
    K.lab(fw, f"A{r}", "传统业务收入($B)")
    for c in ALLC:
        K.fml(fw, f"{c}{r}", f"={K.R(S_SEG, c + str(m['传统业务收入($B)']))}", K.N1, link=True)
    K.logic(fw, f"N{r}", "引『分部测算』(三案共用增速外推)。")
    r += 1
    fund["REV"] = r
    K.lab(fw, f"A{r}", "总营收($B)", b=True)
    for c in ALLC:
        K.fml(fw, f"{c}{r}", f"={c}{fund['HUM']}+{c}{fund['TRAD']}", K.N1)
    fw[f"A{r}"].border = K.BORD
    r += 1
    K.lab(fw, f"A{r}", "  YoY", note=True)
    for i in range(1, len(ALLC)):
        K.fml(fw, f"{ALLC[i]}{r}", f'=IFERROR({ALLC[i]}{fund["REV"]}/{ALLC[i-1]}{fund["REV"]}-1,"n.m.")', K.PCT)
    r += 1
    fund["NI"] = r
    K.lab(fw, f"A{r}", "归母净利($B)", b=True); fw[f"A{r}"].fill = K.OUT
    for c in HC:
        K.fml(fw, f"{c}{r}", f"={K.R(S_HIST, c + str(ha['HNI']))}", K.N1, link=True)
    for c in FCf:
        K.fml(fw, f"{c}{r}", f"={c}{fund['REV']}*{c}{fund['NM']}", K.N1)
    K.logic(fw, f"N{r}", "历史引『历史财务』;前瞻 = 总营收 × 归母净利率(情景当前案)。避免用外部卖方净利截断物理锚链。")
    r += 1
    K.lab(fw, f"A{r}", "  净利率", note=True)
    for c in ALLC:
        K.fml(fw, f"{c}{r}", f"={c}{fund['NI']}/{c}{fund['REV']}", K.PCT)
    r += 1
    fund["EQ"] = r
    K.lab(fw, f"A{r}", "期末权益($B)")
    for c in HC:
        K.fml(fw, f"{c}{r}", f"={K.R(S_HIST, c + str(ha['HEQ']))}", K.N1, link=True)
    prevs = [HC[-1]] + list(FCf[:-1])
    for p, c in zip(prevs, FCf):
        K.fml(fw, f"{c}{r}", f"={p}{r}+{c}{fund['NI']}*{c}{fund['RET']}", K.N1)
    K.logic(fw, f"N{r}", "= 上年 + 归母净利 × 留存率(简式, 未计配售现金流入)。仅供 P/B 体检。")
    r += 1
    fund["EPS"] = r
    K.lab(fw, f"A{r}", "EPS (USD)")
    for c in ALLC:
        K.fml(fw, f"{c}{r}", f"={c}{fund['NI']}*1000/{c}{fund['SH']}", K.N2)
    K.logic(fw, f"N{r}", "= 归母净利($B) × 1000 ÷ 股本(mn) = USD/股。转正前为负 → P/E 交叉验证 N/M。")
    r += 1
    fund["BPS"] = r
    K.lab(fw, f"A{r}", "BPS (USD)", note=True)
    for c in ALLC:
        K.fml(fw, f"{c}{r}", f"={c}{fund['EQ']}*1000/{c}{fund['SH']}", K.N2)
    r += 1
    fund["SPS"] = r
    K.lab(fw, f"A{r}", "SPS 销售额/股(USD)", b=True); fw[f"A{r}"].fill = K.OUT
    for c in ALLC:
        K.fml(fw, f"{c}{r}", f"={c}{fund['REV']}*1000/{c}{fund['SH']}", K.N2)
    K.logic(fw, f"N{r}", "= 总营收($B) × 1000 ÷ 股本(mn) = USD/股销售额。『情景估值』主线 = 目标 P/S × 本行 × FX(HKD)。")
    r += 1
    fund["ROE"] = r
    K.lab(fw, f"A{r}", "ROE", note=True)
    for i, c in enumerate(ALLC):
        f = (f"={c}{fund['NI']}/{c}{fund['EQ']}" if i == 0
             else f"={c}{fund['NI']}/AVERAGE({ALLC[i-1]}{fund['EQ']},{c}{fund['EQ']})")
        K.fml(fw, f"{c}{r}", f, K.PCT)
    r += 1
    K.band(fw, r, "口径说明", 12); r += 1
    K.mtext(fw, r, "链: 分部收入(物理锚驱动)→ 总营收 → ×归母净利率(情景刀)→ 归母净利 → 权益/EPS/BPS/SPS。财务量纲一律 $B(RMB 亿 ÷ 72);每股 USD;隐含价经 FX(HKD/USD=7.8)换 HKD。前瞻净利由总营收×净利率算出, 不抄卖方净利。", "K", 3)
    K.set_widths(fw, 26, ALLC, 8, logic_col="N", logic_width=58)

    # fr for write_scenario_valuation(主线镜头分母 = SPS)
    fr = {"EPS": fund["EPS"], "BPS": fund["SPS"], "NI": fund["NI"], "REV": fund["REV"]}

    # ═══ 11 情景估值(EV/Sales 主线 + P/E 交叉验证)═══
    K.write_scenario_valuation(wb.create_sheet(S_VAL), {
        "title": "情景估值 — 当前案逐年隐含价 (EV/Sales(P/S) 主线; P/E 交叉验证)",
        "intro": "本表输出=『情景切换』当前案(默认 Base)。隐含价(HKD) = 目标 P/S(当前案) × 前瞻每股营收 SPS(USD) × FX(7.8)。历史列=实际 P/S 反推(事实);前瞻=预测, 不拟合现价。三案并排见『估值对比』。",
        "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf, "hist_years": HY,
        "fx_fwd": FX, "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
        "s_fund": S_FUND, "fr": fr,
        "s_switch": S_SW, "target_row": SWPS, "sw_cell": "B2",
        "yend": px["yend"], "yavg": px["yavg"],
        "mcap_div": 100, "mcap_usd_skip": True,
        "mcap_hist_label": "市值 实际年末(亿HKD, 历史)", "mcap_fwd_label": "市值 前瞻·主线(亿HKD)",
        "reading": "P/E 交叉验证读法: 转正前(Bear 全程 / Base 至 2027)净利<0 → 隐含 P/E = N/M, 这正是主镜头用 EV/Sales(P/S)出实数的原因;2028E+ 转正后用 30-40x 路径 P/E 体检隐含倍数(刚转正当年 EPS 极薄 → 隐含 P/E 偏高属正常, 应看更成熟年份)。",
        "method": "方法: 整体公司、EV/Sales(P/S)主线逐年估(公司净现金, EV≈市值)。基本面在『利润与收入假设』;目标 P/S 在『估值倍数假设』(三层);本表套用: 目标 P/S × SPS(USD) × FX → 隐含价(HKD)。",
        "concl": "结论(方向性, Base): 2026E 约 63 HKD(-29%) → 2027E 约 92 HKD(+4%) → 2028E 约 118 HKD → 2030E 约 204 HKD。现价 88.7 大致对应 Base 2027E → 市场已 price-in 交付放量起步。三情景与阶梯见『估值对比』。",
    })

    # ═══ 12 估值对比(三案恒常并排)═══
    ws_v = wb[S_VAL]
    SH_G = K.R(S_FUND, "$F$" + str(fund["SH"]))  # 用于回测口径的历史股本(2025A)
    TAM_REF = lambda c: K.R(S_ANCHOR, c + str(TAM_R))
    SH_FUND = lambda c: K.R(S_FUND, c + str(fund["SH"]))

    def sw_case(key, ci):
        return SWB[key] + ci

    cmp_rows = [
        {"key": "ind_u", "label": "工业线出货(台)", "fmt": K.N0,
         "hist": lambda c, ci, A: (f"={K.R(S_ANCHOR, c + str(IND_R))}" if c == "F" else None),
         "fwd": lambda c, j, ci, A: f"={K.R(S_SW, c + str(sw_case('ind', ci)))}"},
        {"key": "con_u", "label": "消费线出货(台)", "fmt": K.N0,
         "hist": lambda c, ci, A: (f"={K.R(S_ANCHOR, c + str(CON_R))}" if c == "F" else None),
         "fwd": lambda c, j, ci, A: f"={K.R(S_SW, c + str(sw_case('con', ci)))}"},
        {"key": "hum", "label": "人形机器人收入($B)", "fmt": K.N1,
         "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(HUM_HR))}",
         "fwd": lambda c, j, ci, A: (f"=({c}{A['ind_u']}*{K.R(S_SW, c + str(sw_case('indasp', ci)))}"
                                     f"+{c}{A['con_u']}*{K.R(S_SW, c + str(sw_case('conasp', ci)))})/720000")},
        {"key": "trad", "label": "传统业务收入($B)", "fmt": K.N1,
         "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(TRAD_HR))}",
         "fwd": lambda c, j, ci, A: f"={K.R(S_SEG, c + str(m['传统业务收入($B)']))}"},
        {"key": "rev", "label": "总营收($B)", "fmt": K.N1, "bold": True,
         "hist": lambda c, ci, A: f"={c}{A['hum']}+{c}{A['trad']}",
         "fwd": lambda c, j, ci, A: f"={c}{A['hum']}+{c}{A['trad']}"},
        {"key": "ni", "label": "归母净利($B)", "fmt": K.N1, "bold": True,
         "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HNI']))}",
         "fwd": lambda c, j, ci, A: f"={c}{A['rev']}*{K.R(S_SW, c + str(sw_case('nm', ci)))}"},
        {"key": "sps", "label": "SPS 销售额/股(USD)", "fmt": K.N2,
         "hist": lambda c, ci, A: f"={c}{A['rev']}*1000/{K.R(S_HIST, c + str(ha['HSH']))}",
         "fwd": lambda c, j, ci, A: f"={c}{A['rev']}*1000/{SH_FUND(c)}"},
        {"key": "ps", "label": "目标 P/S(历史=实际)", "fmt": K.MX,
         "hist": lambda c, ci, A: (f"={K.R(S_HIST, c + str(HPS))}" if c in ("D", "E", "F") else None),
         "fwd": lambda c, j, ci, A: f"={K.R(S_MULT, c + str(ma['target_row0'] + ci))}"},
        {"key": "px", "label": "隐含价 (HKD)", "fmt": K.PX, "bold": True, "out": True,
         "hist": lambda c, ci, A: (f"={c}{A['ps']}*{c}{A['sps']}*{K.R(S_HIST, c + str(ha['HFX']))}" if c in ("D", "E", "F") else None),
         "fwd": lambda c, j, ci, A: f"={c}{A['ps']}*{c}{A['sps']}*{FX}+0*{TAM_REF(c)}"},
        {"key": "ipe", "label": "隐含 forward P/E(体检)", "fmt": K.MX,
         "hist": lambda c, ci, A: (f'=IF({c}{A["ni"]}<=0,"N/M",{c}{A["px"]}/({c}{A["ni"]}*1000/{K.R(S_HIST, c + str(ha["HSH"]))}*{K.R(S_HIST, c + str(ha["HFX"]))}))' if c in ("D", "E", "F") else None),
         "fwd": lambda c, j, ci, A: f'=IF({c}{A["ni"]}<=0,"N/M",{c}{A["px"]}/({c}{A["ni"]}*1000/{SH_FUND(c)}*{FX}))'},
        {"key": "up", "label": "历史=回测(vs实际年末价≈0) / 前瞻=vs现价", "fmt": K.PCT,
         "hist": lambda c, ci, A: (f"={c}{A['px']}/{K.R(S_HIST, c + str(ha['HPX']))}-1" if c in ("D", "E", "F") else None),
         "fwd": lambda c, j, ci, A: f"={c}{A['px']}/{PX_NOW}-1"},
    ]
    cm = K.write_comparison(wb.create_sheet(S_CMP), {
        "title": "估值对比 — Bear / Base / Bull 三情景逐年隐含价并排",
        "intro": "三案各自完整推演: 工业/消费出货 × ASP + 传统 → 总营收 → 归母净利 → 目标 P/S(锚×溢价×该案情绪)→ 逐年隐含价(HKD)。恒常并排不随开关变;调假设去『情景切换』。历史列用同一条链填实际值: 隐含价历史列 ≈ 实际年末价(回测行 ≈0%)= 链的内置回测。",
        "case_names": CASES,
        "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
        "block_start": 16,
        "rows": cmp_rows,
        "summary": {
            "band": "三案汇总 (目标年 2027E;触发条件见『情景切换』)",
            "target_col": "H",
            "rows": [
                ("总营收($B)", "rev", K.N1, "= 人形(工业+消费) + 传统(该案)", False),
                ("归母净利($B)", "ni", K.N1, "= 总营收 × 该案归母净利率(2027: Bear/Base 亏, Bull 转正)", False),
                ("目标 P/S", "ps", K.MX, "= 基准锚 11x × 溢价 1.0x × 该案情绪(6/11/16x)", False),
                ("隐含价(HKD)", "px", K.PX, "= 目标 P/S × 2027E 每股营收(USD) × FX", True),
                ("隐含 forward P/E", "ipe", K.MX, "体检: 2027 Bear/Base 亏 → N/M;Bull 已转正", False),
                ("vs 现价", "up", K.PCT, "对照现价 88.7 HKD 的上行/下行空间", True),
            ],
            "mcap": {"label": "隐含市值(亿HKD)", "key": "px", "expr": f"*{SH_G}/100",
                     "note": "= 隐含价 × 股本(mn) ÷ 100"},
            "concl": "risk-reward(2027E vs 现价 88.7): Bear 约 36(-59%) / Base 约 92(+4%) / Bull 约 192(+117%)。Base 2027E 大致等于现价 → 市场已 price-in 交付放量起步;上行看工业/消费双线放量兑现 + 主题溢价, 下行看交付不及 + ASP 战。逐年阶梯见各案 block(2026E 起 Base 63→204 HKD)。",
        },
    })

    # ═══ 13 综合判断仪表盘 ═══
    REV26_B = K.R(S_CMP, "G" + str(cm["CMPA"]["Base"]["rev"]))
    dash = K.write_dashboard(wb.create_sheet(S_DASH), {
        "title": "综合判断仪表盘 — A 基本面拐点 · B 估值错位 · C 催化剂 · D 情绪确认",
        "usage": "把模型压成投后跟踪语言: 哪些指标验证 Base(交付放量+ASP 守住+2028 转正), 哪些把模型推向 Bear/Bull。B 列公式直接引模型输出。",
        "blocks": [
            {"title": "A. 基本面拐点 — 人形交付在放量吗?", "rows": [
                ("全尺寸人形交付", "2025A 1079 台", "工业线交付台数是收入引擎;Base 2027E 4000 台是承重假设。", True),
                ("毛利率", "2025A 37.7%(2021 以来最高)", "人形+高毛利占比上升 → 毛利率回升, 是盈利拐点先行指标。"),
                ("Base 2027E 归母净利", {"fml": f"={K.R(S_CMP, 'H' + str(cm['CMPA']['Base']['ni']))}", "fmt": K.N1, "fill": True}, "Base 2027 仍小亏(-5%), 2028 转正;若提前=上修 Bull。"),
            ]},
            {"title": "B. 估值错位(预测引擎)", "rows": [
                ("当前股价(HKD)", {"inp": PX_NOW, "fmt": K.PX, "fill": True}, "reality check, 不反向拟合。"),
                ("Base 2027E 隐含价", {"fml": f"={K.R(S_CMP, 'H' + str(cm['CMPA']['Base']['px']))}", "fmt": K.PX, "fill": True}, "主判断输出, ≈现价(+4%)。"),
                ("Base 2027E vs 现价", {"fml": f"={K.R(S_CMP, 'H' + str(cm['CMPA']['Base']['up']))}", "fmt": K.PCT, "fill": True}, "接近 0 = 现价已 price-in Base 交付起步。"),
                ("Bear 2027E 隐含价", {"fml": f"={K.R(S_CMP, 'H' + str(cm['CMPA']['Bear']['px']))}", "fmt": K.PX}, "下行: 交付不及+ASP 战+倍数收敛至 6x。"),
                ("Bull 2027E 隐含价", {"fml": f"={K.R(S_CMP, 'H' + str(cm['CMPA']['Bull']['px']))}", "fmt": K.PX}, "上行: 双线放量+ASP 守住+主题溢价 16x。"),
            ]},
            {"title": "C. 催化剂", "rows": [
                ("Bear 触发", "工业交付爬产/良率/回款慢;消费 U1 迟迟不上量;人形叙事退潮;ASP 价格战。", "先砍出货/ASP, 再把目标 P/S 压到 6x。"),
                ("Bull 触发", "全球人形产业爆发;优必选工业+消费双线放量;ASP 守住;主权/工业大单;2027 转正。", "先抬出货/守 ASP, 再抬目标 P/S 到 16x。"),
                ("最易错处", "把在手订单线性当收入(订单≠交付);把当下 TTM 20.7x 当前瞻 P/S(倍数与收入双重乐观)。", "本模型收入低共识约 25% + Base 前瞻 P/S 11x, 规避两个陷阱。"),
            ]},
            {"title": "D. 情绪确认", "rows": [
                ("量价温度计", "2025-10 冲高 161 → 2026-07 88.7, 回撤约 45%", "人形 FOMO 见顶后退潮;现价在回撤消化区, 非启动段。"),
                ("现价 P/S vs Base 目标", "现价 TTM 20.7x vs Base 前瞻 11x", "现价倍数高于 Base 前瞻锚, 但对应低基数收入;放量后倍数自然压缩。"),
                ("当前档位", "【退潮后消化】", "等交付兑现的观望区;基本面(交付)是下一波方向的裁判。", True),
            ]},
        ],
        "final": {"band": "★ 综合判断",
                  "text": "优必选是典型的'物理量放量决定价值'标的: 隐含价对工业/消费出货台数与 ASP 高度敏感。Base(目标 P/S 11x、拐点 2028、收入低共识约 25%)下 2027E 隐含价约 92 HKD ≈ 现价 → 市场已 price-in 交付放量起步。三情景 2027E 区间 36-192 HKD(Bear-Bull), 上行靠双线放量兑现+主题溢价, 下行靠交付不及+ASP 战。关键跟踪 = 全尺寸人形季度交付台数与 ASP 走势;交付超预期 → 上修 Bull 概率, 不及 → 转 Bear。"},
        "tracking": {
            "intro": "哪个指标恶化/兑现 → 哪个假设先动 → 触发什么动作。",
            "rows": [
                ("__band__", "一、物理量(出货)"),
                ("全尺寸人形季度交付", "2025A 1079 台/年", "命门: 工业线收入引擎", "季报/业绩公告 + 中标公告", "连续不及 Base 路径 → 下调 ind 出货转 Bear"),
                ("消费线 U1 上量", "2026 起量", "命门: 第二增长曲线", "产品发布/交付披露", "迟迟不上量 → 下调 con 出货"),
                ("__band__", "二、价(ASP)"),
                ("工业/消费 ASP 走势", "2025A 工业 76 万", "命门: 人形收入 = 量×价", "订单单价/年报 ASP", "价格战破位 → 下调 ASP 全链重算"),
                ("__band__", "三、盈利拐点"),
                ("归母净利率", "2025A -35%", "命门: P/E 交叉验证 + 拐点", "季报/年报", "2028 未转正 → 转 Bear;2027 转正 → 上修 Bull"),
                ("__band__", "四、估值倍数/情绪"),
                ("前瞻 P/S vs 共识", "Base 11x vs 共识 9.7-22x", "命门: 目标倍数", "卖方更新 + 主题热度", "主题降温 → 情绪值下调;大单/爆发 → 上调"),
            ],
        },
    })

    # ═══ input.json ═══
    payload = {
        "ticker": "09880.HK", "company": "UBTech Robotics (优必选)",
        "built_at": "2026-07-10", "currency_financials": "RMB (folded to USD $B at RMB/USD=7.2)",
        "currency_price": "HKD (HKD/USD=7.8 for per-share conversion)",
        "current_price_hkd": PX_NOW, "price_date": "2026-07-10", "market_cap_hkd_b": 44.65,
        "shares_m": SH_NOW, "pe_ttm": "negative (loss-making)",
        "method": "EV/Sales(P/S) main lens (target P/S × forward SPS × FX) + path P/E cross-check; physical anchor = global humanoid shipments TAM, UBTech split into industrial + consumer lines (units × ASP) + legacy business.",
        "target_price_note": "12M 目标价 ≈ Base 2027E 隐含价 ≈ 92 HKD(+4% vs 88.7);三案 2027E Bear/Base/Bull ≈ 36 / 92 / 192 HKD。",
        "anchor": {"type": "humanoid-shipments", "global_tam_units": TAM,
                   "ubtech_industrial_units_hist_2025A": 1079, "ubtech_consumer_units_hist_2025A": 0},
        "historical_rmb_yi": {"revenue_total": TOT_H_YI, "humanoid": HUM_H_YI, "legacy": TRAD_H_YI,
                              "gross_margin": GM_H, "net_income_attrib": NI_H_YI, "equity_est": EQ_H_YI,
                              "shares_m": SH_H, "net_margin": NM_H},
        "historical_usd_b": {"revenue_total": [ub(x) for x in TOT_H_YI], "humanoid": HUM_H, "legacy": TRAD_H,
                             "net_income_attrib": NI_H, "equity_est": EQ_H},
        "fx": {"rmb_usd": RMBUSD, "hkd_usd": FX},
        "price_end_hkd": PX_END, "price_now_hkd": PX_NOW,
        "scenario_assumptions": {
            "industrial_units": IND_U, "consumer_units": CON_U,
            "industrial_asp_rmb_wan": IND_ASP, "consumer_asp_rmb_wan": CON_ASP,
            "net_margin": NM, "legacy_rmb_yi_shared": TRAD_F_YI, "shares_m_shared": SH_F,
            "gross_margin_base_path": GM_F,
            "target_ps": PS_TARGET, "ps_anchor": PS_ANCHOR, "ps_premium": PS_PREM, "ps_sentiment": PS_SENT,
        },
        "consensus": {"n_strong_buy": 10, "tp_mean_hkd": 153, "tp_range_hkd": [125, 175],
                      "sellside_ps_2027e": {"HSBC": 9.7, "CICC_中金": 18, "GuotaiHaitong_国泰海通": 22},
                      "sellside_tp_hkd": {"HSBC": 126, "CICC_中金": 138, "GuotaiHaitong_国泰海通": 185},
                      "consensus_rev_rmb_yi": {"2026E": 36.9, "2027E": 55.2},
                      "model_rev_rmb_yi_base": {"2026E": 27.4, "2027E": 41.1},
                      "note": "本模型 Base 收入刻意低于共识约 25%(订单≠交付、消费 ASP 稀释、执行风险);目标 P/S 11x 近 HSBC、审慎于中金/国泰海通。"},
        "sources": {
            "financials": "FY2021-2025 业绩公告(总营收 8.17/10.08/10.56/13.05/20.01 亿RMB;归母净亏 -9.20/-9.75/-12.34/-11.24/-7.03;毛利率 31.3/29.2/31.5/28.7/37.7%)",
            "segment_2025A": "全尺寸人形 8.21 亿(1079 台, ASP 76 万) / 传统业务合计 11.80 亿(教育4.13+物流2.74+消费及硬件5.23-非具身0.48等)",
            "market": "2026-07-10 现价 88.7 HKD / 市值 446.5 亿 HKD / 总股本约 5.03 亿股 / 52周 75.5-161.0",
            "industry": "全球人形出货 TAM: GGII/IDC/MS 出货口径中枢(2025 约 2 万 → 2030E 40 万台)",
        },
    }
    os.makedirs(VAULT, exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    K.finalize(wb, freeze={
        S_COVER: "A2", S_HIST: "B3", S_PX: "B4", S_CONS: "A2", S_HMULT: "B3",
        S_MULT: "B3", S_SW: "B3", S_ANCHOR: "B3", S_SEG: "B3",
        S_FUND: "B3", S_VAL: "B4", S_CMP: "B6", S_DASH: "B6",
    })
    os.makedirs(os.path.dirname(LOCAL_XLSX), exist_ok=True)
    wb.save(LOCAL_XLSX)
    wb.save(OUT_XLSX)
    print("saved:", OUT_XLSX)
    print("local:", LOCAL_XLSX)
    print("sheets:", wb.sheetnames)
    print("rows:", {"HPS": HPS, "SWPS": SWPS, "SPS": fund["SPS"], "NI": fund["NI"]})


if __name__ == "__main__":
    build()
