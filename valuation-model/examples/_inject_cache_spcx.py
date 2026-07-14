# -*- coding: utf-8 -*-
"""把 formulas 库算出的公式结果作为缓存值 (<v>) 注入到 openpyxl 生成的带样式 xlsx,
不破坏任何样式/字色/冻结(本机无 LibreOffice; formulas 自带 writer 会抹掉样式, 故走 XML 注入)。

做法: 读 formulas 解出的 {(sheet,coord):value} → 用 openpyxl 重新打开带样式文件 →
对每个公式格, 把其 value 暂存、写成普通值再恢复? 不行(会丢公式)。
改为: 直接编辑 xlsx(zip)内各 sheet XML, 给含 <f> 的 <c> 补 <v>缓存</v>。
"""
import warnings, os, re, zipfile, shutil
warnings.filterwarnings("ignore")
import formulas
import openpyxl
from openpyxl.utils import get_column_letter

_repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_output_dir = os.environ.get("VALUATION_OUTPUT_DIR", os.path.join(_repo_root, "out"))
P = os.environ.get("VALUATION_MODEL_PATH", os.path.join(_output_dir, "SPCX_valuation_model.xlsx"))

# 1. 用 formulas 解出所有单元格值
xl = formulas.ExcelModel().loads(P).finish()
sol = xl.calculate()
vals = {}   # (SHEET_UPPER, COORD) -> python value
ref_re = re.compile(r"'\[[^\]]+\](.+?)'!(.+)$")
for k, v in sol.items():
    m = ref_re.match(k)
    if not m:
        continue
    sheet = m.group(1).upper()
    coord = m.group(2).replace("$", "")
    if not re.fullmatch(r"[A-Z]+\d+", coord):
        continue
    try:
        val = v.value[0, 0]
    except Exception:
        try:
            val = v.value
        except Exception:
            val = None
    vals[(sheet, coord)] = val

# 2. 取 openpyxl 的 sheet 名 → XML 文件映射(workbook.xml.rels + sheet 顺序)
wb = openpyxl.load_workbook(P)
sheet_titles = wb.sheetnames   # 物理顺序 = sheetN.xml 顺序(openpyxl 写出时按此)

# openpyxl 写出的 worksheet xml 命名为 sheet1.xml ... 按 workbook 顺序
# 用 zip 读 xl/workbook.xml 的 sheet r:id → 再读 rels → 确定文件; 简化: openpyxl 顺序即 sheetN
def num_to_xml(i):
    return f"xl/worksheets/sheet{i+1}.xml"

# 3. 编辑每个 sheet XML, 给含 <f> 的 <c> 注入 <v>
tmp = P + ".tmp"
shutil.copy(P, tmp)
with zipfile.ZipFile(tmp, "r") as zin:
    names = zin.namelist()
    data = {n: zin.read(n) for n in names}

cell_re = re.compile(rb'(<c r="([A-Z]+\d+)"[^>]*>)(.*?)(</c>)', re.S)
ftag_re = re.compile(rb'<f[ >].*?</f>|<f[^>]*/>', re.S)
vtag_re = re.compile(rb'<v>.*?</v>', re.S)

def esc(s):
    return str(s)

injected = 0
for i, title in enumerate(sheet_titles):
    xmlname = num_to_xml(i)
    if xmlname not in data:
        continue
    sheet_up = title.upper()
    content = data[xmlname]

    def repl(mc):
        global injected
        head, coord, body, tail = mc.group(1), mc.group(2).decode(), mc.group(3), mc.group(4)
        if b"<f" not in body:
            return mc.group(0)
        val = vals.get((sheet_up, coord))
        if val is None or isinstance(val, str):
            # 字符串结果或无值: 跳过(保持无缓存; 多为 N/M 文本本就是 inline str 不在公式格)
            if isinstance(val, str):
                return mc.group(0)
            return mc.group(0)
        if isinstance(val, bool):
            return mc.group(0)
        # 数值: 去掉旧 <v>, 在 <f> 后补 <v>
        body2 = vtag_re.sub(b"", body)
        try:
            num = float(val)
        except Exception:
            return mc.group(0)
        vbytes = (f"<v>{num!r}</v>").encode()
        body2 = body2 + vbytes
        # 去掉 head 里的 t="..."(数值不需要 type)
        head2 = re.sub(rb'\s+t="[^"]*"', b"", head)
        injected += 1
        return head2 + body2 + tail

    new_content = cell_re.sub(repl, content)
    data[xmlname] = new_content

# 4. 重新打包
out = P
with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
    for n in names:
        zout.writestr(n, data[n])
os.remove(tmp)
print("injected cached values:", injected)
