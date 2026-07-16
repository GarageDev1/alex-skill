# -*- coding: utf-8 -*-
"""
build_700hk.py — 腾讯控股(Tencent, 00700.HK)估值模型实例。
用 build_kit v2 骨架;单位:基本面 RMB 亿;股价 HKD;FX = HKD/RMB ≈ 1.08。
口径:估值用 Non-IFRS(经调整)归母净利(街上主看口径,剔除股权激励/投资收益/SSV)。
主线镜头 整体 Non-IFRS forward P/E(质量锚×情绪值三层);交叉镜头 SOTP(分部 P/E + 投资组合 30% 折价);P/E 历史带做体检。
物理锚:微信合并 MAU(亿) × 广告 ARPU(元/用户/年, = 加载率 × eCPM 的合成)→ 广告收入(平台型直驱段);游戏/社交/FBS 走周期增长。
"""
import os, json
from openpyxl import Workbook
import build_kit as K

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
VAULT = os.environ.get("VALUATION_OUTPUT_DIR", os.path.join(REPO_ROOT, "out"))

# ════════════ 0. 全局轴 ════════════
ALLC = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
ALLY = ["2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
HC, HY = ["B", "C", "D", "E", "F"], ["2021", "2022", "2023", "2024", "2025"]
FC = ["F", "G", "H", "I", "J", "K"]      # 含基年 F=2025A
FCf = FC[1:]                              # 2026E-2030E
FX = 1.08                                 # HKD per RMB(2026-06 口径)
FX_HIST = [1.21, 1.16, 1.10, 1.08, 1.09]
CASES = ["Bear", "Base", "Bull"]
PX_NOW = 417.20                           # 2026-06-23 盘中(Research Data API)
SH_NOW = 9080.0                           # mn 股(2025末口径,持续回购缩股;当下市值÷现价≈9,110mn)
PEAK_PE, PREMIUM = 24.0, 1.0              # 质量锚(平台龙头上沿)× 结构溢价(已含在锚内)

S_COVER, S_HIST, S_PX, S_CONS = "封面", "历史财务与估值", "股价走势", "卖方研报共识"
S_HMULT, S_MULT, S_SW = "历史估值倍数", "估值倍数假设", "情景切换"
S_ANCHOR, S_SEG, S_FUND = "物理锚-微信流量盘", "分部测算", "利润与收入假设"
S_VAL, S_CMP, S_DASH = "情景估值", "估值对比", "综合判断仪表盘"

GAME, SNS, AD, FBS, OTH = "游戏(本土+国际)", "社交网络", "营销服务(广告)", "金融科技及企业服务", "其他"

# ════════════ 公司数据(亿 RMB; 一手披露 2021-2025 年报 + 1Q26)════════════
REV_SEG = {
    GAME: [1743, 1707, 1799, 1977, 2416],
    SNS:  [1173, 1169, 1185, 1215, 1277],
    AD:   [887, 827, 1015, 1214, 1450],
    FBS:  [1722, 1771, 2038, 2120, 2294],
    OTH:  [78, 72, 54, 78, 81],
}
GM   = [0.44, 0.43, 0.48, 0.53, 0.56]
NI_ADJ = [1238, 1156, 1577, 2227, 2596]                 # Non-IFRS 归母净利(估值口径)
NI_REP = [2248, 1882, 1152, 1941, 2248]                 # IFRS 报表归母净利(参照)
EQ   = [7700, 8000, 8086, 9735, 11541]                  # 归母权益(2021/2022 为估)
SHARES = [9610, 9590, 9430, 9180, 9080]                 # mn 股(持续回购缩股)
WX_MAU = [12.68, 13.13, 13.43, 13.85, 14.10]            # 微信合并 MAU(亿)

# 月末收盘(复权 HKD; Research Data API hk-lake)
MONTHLY = [
    ("2021-06", 552), ("2021-12", 421),
    ("2022-03", 374), ("2022-06", 360), ("2022-10", 200), ("2022-12", 317),
    ("2023-06", 330), ("2023-09", 305), ("2023-12", 294),
    ("2024-03", 290), ("2024-06", 367), ("2024-09", 430), ("2024-12", 420),
    ("2025-03", 510), ("2025-06", 504), ("2025-09", 560), ("2025-12", 599),
    ("2026-02", 494), ("2026-03", 470), ("2026-05", 455), ("2026-06", 417),
]
PX_END = {"2021": 421, "2022": 317, "2023": 294, "2024": 420, "2025": 599}
PX_AVG = {"2021": 480, "2022": 320, "2023": 330, "2024": 360, "2025": 470}

# ════════════ 情景参数 ════════════
# 目标 P/E(对 Non-IFRS EPS)→ 情绪值 = 目标P/E / 质量锚
TARGET_PE = {"Bear": [13, 13, 12, 12, 11], "Base": [16, 17, 17, 16, 15], "Bull": [20, 21, 21, 20, 19]}
PE_SENT = {c: [round(v / PEAK_PE, 3) for v in vals] for c, vals in TARGET_PE.items()}

AD_ARPU = {"Bear": [110, 119, 128, 135, 141], "Base": [117, 131, 146, 158, 168], "Bull": [124, 143, 163, 180, 196]}
GAME_G  = {"Bear": [0.06, 0.05, 0.04, 0.04, 0.03], "Base": [0.10, 0.09, 0.08, 0.07, 0.06], "Bull": [0.14, 0.12, 0.10, 0.09, 0.08]}
FBS_G   = {"Bear": [0.06, 0.06, 0.05, 0.05, 0.04], "Base": [0.09, 0.10, 0.10, 0.09, 0.08], "Bull": [0.12, 0.13, 0.13, 0.12, 0.11]}
OPM     = {"Bear": [0.35, 0.34, 0.34, 0.33, 0.33], "Base": [0.37, 0.38, 0.38, 0.39, 0.39], "Bull": [0.38, 0.40, 0.41, 0.42, 0.42]}
OPM_H   = [0.31, 0.27, 0.30, 0.360, 0.373]              # 历史 Non-IFRS OPM(2021-2023 估)
# 非情景杠杆(三案共用)
SNS_G = [0.03, 0.02, 0.02, 0.02, 0.02]
OTH_G = [0.05, 0.05, 0.05, 0.05, 0.05]
MAU_FWD = [14.40, 14.50, 14.55, 14.60, 14.65]          # 微信 MAU 前瞻(接近天花板,平缓)
NETCONV = 0.92                                          # Non-IFRS 净利/Non-IFRS OP
RETENT = 0.85                                           # 留存率(派息+回购后)

# SOTP(2027E, Base)分部倍数与投资组合
SOTP = {
    "game_margin": 0.40, "game_pe": 15.0,
    "ad_margin": 0.35, "ad_pe": 22.0,
    "sns_margin": 0.25, "sns_pe": 16.0,
    "fbs_margin": 0.13, "fbs_pe": 15.0,
    "listed_inv": 6727, "unlisted_inv": 3631, "inv_discount": 0.30, "net_cash": 1071,
}


def hist_growth(vals):
    return [None] + [round(vals[i] / vals[i - 1] - 1, 4) for i in range(1, 5)]


def ad_arpu_hist():
    return [round(REV_SEG[AD][i] / WX_MAU[i], 1) for i in range(5)]


def build():
    wb = Workbook(); wb.remove(wb["Sheet"])

    # 1 封面
    K.write_cover(wb.create_sheet(S_COVER), {
        "title": "腾讯控股 (00700.HK) 估值模型 — 物理锚链(微信流量盘 → 分部收入 → Non-IFRS 利润 → P/E + SOTP)",
        "meta": [
            ("报告日期", "2026-06-23"),
            ("数据截止", "FY2025 年报(2026-03-18)+ 1Q26(2026-05-13)+ 2026-06-23 盘中行情 + 9 家外资行一致预期"),
            ("现价 / 市值", f"HK${PX_NOW:,.2f} / 市值约 HK$3.80 万亿(今日跌破 4 万亿,年内 -30%);PE(TTM,IFRS)≈15.3x"),
            ("估值口径", "估值用 Non-IFRS(经调整)归母净利——街上主看口径,剔除股权激励/投资公允价值收益/SSV;报表 IFRS 净利仅作参照。"),
            ("主线镜头", "整体 Non-IFRS forward P/E(质量锚 24x × 情绪值三层);交叉 SOTP(游戏 15x / 广告 22x / 社交 16x / FBS 15x P/E + 投资组合 30% 折价);P/E 历史带做体检。"),
            ("方法一句话", "微信 MAU × 广告 ARPU(加载率×eCPM)→ 广告收入;游戏/社交/FBS 周期增长 → 分部收入 → Non-IFRS OPM → Non-IFRS 净利/EPS → 目标 P/E → 隐含股价。"),
        ],
        "takeaways": [
            ("① 物理锚", "广告(增长核心)挂微信流量盘:MAU 14.1 亿 × 广告 ARPU(2025=103 元/用户)。视频号广告加载率仅 3-4% vs 抖音 20%,是 ARPU 提升的物理空间。改 MAU/ARPU 穿透到广告→净利→隐含价。"),
            ("② 基本面", "2025 收入 +14%、Non-IFRS 净利 +18%(利润增速持续快于收入);广告 +19% 连 11 季双位数;毛利率 44%→56%(高毛利结构升级);1Q26 收入 +9%(降速)、AI 新产品单季拖累经营利润 88 亿。"),
            ("③ 市场争议", "市场 2026 年内把腾讯从 5 万亿+ 杀到 3.8 万亿,叙事 = AI 落后(元宝 C 端第四)+ AI 投入翻倍压利润 + 回购缩水 + 港股外资流出。错配:AI 在广告(CTR 1%→3%)/云(服务 90% 大模型厂商)/微信入口已兑现,非赌独立 C 端产品。"),
            ("④ 估值判断", "现价 417 对应 Non-IFRS forward P/E ~14x,处 2015 年来 10-48x 带下沿(~10-15 分位);卖方 9 家全员买入、目标价中枢 HK$732。Base 取温和修复(16-17x),隐含价显著高于现价。"),
            ("⑤ 评级", "BUY(买入):风险收益不对称——下行有投资组合(1.04 万亿,约市值 1/4,打 30% 折价仍是安全垫)+ 持续回购托底,上行靠视频号加载率放量 + AI 货币化 + 估值修复。目标价见情景估值/估值对比页。"),
        ],
    })

    # 2 历史财务与估值
    ha = K.write_history(wb.create_sheet(S_HIST), {
        "title": "腾讯 历史财务与估值 (亿 RMB; 股价 HKD) — 2021-2025A + 当下(TTM/1Q26)",
        "hist_cols": HC, "hist_years": HY,
        "fx_hist": FX_HIST, "fx_now": FX,
        "vals_in_usd": True, "ps_scale": 1000, "mcap_div": 1000,
        "unit_label": "(十亿 RMB)", "mcap_label": "市值(十亿 RMB)", "fx_label": "FX (HKD/RMB)",
        "cur_label": "当下(TTM/1Q26)",
        "eps_label": "Non-IFRS EPS (RMB)", "bps_label": "BPS (RMB)",
        "segments": [(GAME, REV_SEG[GAME], True), (SNS, REV_SEG[SNS], True),
                     (AD, REV_SEG[AD], True), (FBS, REV_SEG[FBS], True), (OTH, REV_SEG[OTH], False)],
        "total_now": 768.3,
        "gm_pct": GM, "gm_now": 0.57,
        "ni": NI_ADJ, "ni_now": 266.2,
        "eq": EQ, "eq_now": 1140.0,
        "shares": SHARES, "shares_now": SH_NOW,
        "px_end": [PX_END[y] for y in HY], "px_now": PX_NOW,
        "px_avg": [PX_AVG[y] for y in HY],
        "band_note": "历史 Non-IFRS P/E:2021 流动性高点约 27x → 2023-2024 约 16x → 2025 年末 19x(对应当年价 599);当下现价 417 对应 Non-IFRS forward P/E ~14x,处 2015 年来 10-48x 历史带下沿。",
        "quarter": {
            "col": "H", "label": "1Q26(单季)",
            "segs": {GAME: (64.2, 0.07), SNS: (31.9, 0.04), AD: (38.2, 0.20), FBS: (59.9, 0.09), OTH: (2.3, None)},
            "ni": 67.9, "eq": 1127.7, "shares": SH_NOW, "fx": FX,
            "note": "1Q26 实际:总收入 1,965 亿(+9%,近六季最低)、毛利率 57%、Non-IFRS 归母净利 679 亿(+11%);剔除新 AI 产品经营利润 844 亿(+17%,margin 43%),AI 新产品单季拖累经营利润 88 亿。游戏=本土 454+国际 188。",
        },
        "notes": [
            (GAME, "本土+国际游戏。2025=2,416 亿(+22%),国际首破百亿美元(+33%);三角洲行动 2025 收入暴涨约 29 倍、DAU 5,000 万;王者荣耀/和平精英常青。前瞻按流水周期增长(情景刀)。"),
            (AD, "营销服务(2024Q3 前称网络广告)。2025=1,450 亿(+19%),连 11 季双位数;毛利率 55%→58%。前瞻挂物理锚:微信 MAU × 广告 ARPU(加载率×eCPM,见物理锚页)。视频号加载率 3-4% vs 抖音 20%。"),
            (FBS, "金融科技及企业服务(2019Q3 单列)。2025=2,294 亿(+8%);分部毛利率 40%→47%→51%;腾讯云 2025 首次规模化盈利;企业服务(含云)+20%、AI 智算驱动。"),
            ("HNI", "Non-IFRS 归母净利(估值口径,剔除股权激励/投资公允价值收益/SSV):1,238/1,156/1,577/2,227/2,596 亿。当下=TTM(2025−1Q25+1Q26)=2,662 亿。IFRS 报表净利 2,248/1,882/1,152/1,941/2,248 亿(见决策 memo)。"),
            ("HEQ", "归母权益;2021/2022 为估(数据 API 接口只回溯到 2023)。腾讯轻资产平台,P/B 仅参照,主线 P/E/SOTP。当下=1Q26 末约 11,400 亿。"),
            ("HSH", "总股本(mn),持续回购注销缩股:2024 大回购 1,120 亿港元、2025 回购 800 亿港元(注销 153.4mn 股)。前瞻保守按 9,080mn 固定(不预设未来回购增厚 EPS)。"),
            ("HFX", "FX = HKD per RMB(年均):人民币兑港币 2021 约 1.21 → 2024-2025 约 1.08-1.09。财报本位币 RMB,股价 HKD,P/E 用 HKD 价 ÷ (RMB EPS × FX)。"),
            ("HPX", "年末复权收盘(HKD,Research Data API hk-lake);当下=2026-06-23 盘中 417.20。"),
        ],
    })

    # 3 股价走势
    def phase_fn(ym):
        if ym <= "2022-12":
            return "① 监管+反垄断深跌"
        if ym <= "2024-12":
            return "② 回购+基本面修复"
        if ym <= "2025-12":
            return "③ AI 叙事+创新高"
        return "④ AI 落后叙事估值倍数下修"
    px = K.write_price_chart(wb.create_sheet(S_PX), MONTHLY, {
        "fn": phase_fn,
        "rows": [("① 监管+反垄断深跌", "2021-2022 教培/游戏版号/反垄断 + 大股东减持,股价从 700+ 跌至 2022-10 低点约 190。"),
                 ("② 回购+基本面修复", "2023-2024 千亿港元回购 + 利润修复(EPS 增速>净利>毛利>收入高质量范式),股价回到 420。"),
                 ("③ AI 叙事+创新高", "2025 视频号+AI 广告兑现 + 接入 DeepSeek,全年走强,2025 年末 599、阶段高点 683。"),
                 ("④ AI 落后叙事估值倍数下修", "2026 年内:DeepSeek V4 发布、元宝 C 端第四 → AI 落后叙事;AI 投入翻倍压利润+回购缩水+港股外资流出,跌破 4 万亿至 417。")],
    }, title="腾讯 00700.HK 月末收盘 (HKD, 复权)")
    px["yhigh"].update({2021: 683, 2022: 470, 2023: 388, 2024: 480, 2025: 645, 2026: 599})
    px["ylow"].update({2021: 412, 2022: 190, 2023: 280, 2024: 280, 2025: 400, 2026: 417})

    # 4 卖方研报共识
    K.write_consensus(wb.create_sheet(S_CONS), {
        "title": "卖方研报共识 — 9 家外资行(2026 Q1-Q2);全员买入,目标价中枢 HK$732,区间 650-800",
        "overview": "总览:GS/MS/UBS/JPM/Nomura/Bernstein/DB 等 9 家全员买入(Buy/OW/Outperform),目标价中枢约 HK$732(区间 650-800),vs 现价 417 隐含 +56~92%。主流方法 SOTP(游戏 13-18x / 广告 20-25x / 金融 15x P/E / 云 4-5x P/S / 投资组合打 20-30% 折价,核心业务约占 84%)。一致 2026E 营收 +10-12%、Non-IFRS 净利 ~2,910 亿、EPS ~30.3 RMB。当前 forward P/E ~13x 处 2015 年来 10-48x 带下沿。",
        "assumptions": [
            ("广告(营销服务)增速 2026E", "GS +18%;1Q26 +20%(视频号/搜索/AIM+ 驱动,AIM+ 约占广告投放 30%)。",
             "AI 货币化兑现速度 vs 宏观广告预算疲软;视频号加载率放开节奏由管理层对体验的权衡决定。", "Base 广告 ARPU 2026E +14%(MAU×ARPU),反映加载率温和提升 + AI eCPM;不假设激进放量(留给 Bull)。"),
            ("游戏增速 2026E", "GS +11%(国内+14%/海外+10%);递延收入 RMB141bn(+28% qoq)支撑 2H 加速。",
             "1Q26 国内游戏 +6% 软(春节时点),递延收入兑现节奏;新游(三角洲)生命周期。", "Base 游戏 +10% 逐年缓降,反映常青游戏稳态 + 国际延续 + 新游贡献,不押单一爆款。"),
            ("AI 投入与利润率", "2026 AI 新产品投入翻倍至 >360 亿;GS 把 FY26-27 capex 上修到 1,510/1,650 亿。",
             "增量 AI 成本是否无限期压制盈利(margin 拖累约 2ppt)vs 核心利润足以自融。", "Base Non-IFRS OPM 37%→38%:AI 投入压制被广告/云高毛利对冲,利润维持双位数增长。"),
            ("目标倍数 / 估值方法", "SOTP 为主:游戏 13-18x、广告 20-25x、FBS 15x P/E、云 4-5x P/S、投资组合 20-30% 折价;或整体 21x(JPM)/20x(Bernstein)P/E。",
             "AI 落后折价该给多深 vs 估值已在历史下沿(~13x)的修复空间。", "三层 P/E:质量锚 24x × 情绪值;Base 2026-27 给 16-17x(温和修复,落卖方下沿),Bull 到 20-21x(卖方上沿)。SOTP 交叉验证。"),
        ],
        "divergences": [
            "① AI 落后是结构问题还是市场过度反应:空头看元宝 C 端第四 + 混元语言模型非头部;多头看 AI 已在广告(CTR 3%)/云(90% 大模型厂商)/微信 14.3 亿 MAU 入口兑现。本模型判定折价部分过度,Base 偏修复。",
            "② AI 投入翻倍(>360 亿)对近期盈利的拖累 vs 货币化(视频号/搜索广告 +20%、企业服务 +20%)的兑现速度——这是 Bear/Base 的核心分野。",
            "③ 回购缩水:2026 为腾 AI 资金减少回购,失去回购支撑后空头反扑下股价脆弱;但 1Q26 仍连续 24 日回购 112 亿港元 + 投资组合安全垫。",
        ],
        "stances": [
            "Goldman Sachs(2026-05-14):Buy,TP HK$700,SOTP。",
            "Morgan Stanley(2026-06-02):OW,TP HK$650,SOTP(Core HK$570 DCF + Associate HK$80,30% 折价),≈18x 2026E Non-IFRS EPS。",
            "UBS(2026-05-14):Buy,TP HK$780,SOTP(核心 84% + 投资组合 16%)。",
            "J.P. Morgan(2026-05-14):OW,TP HK$690,21x 2026E P/E。",
            "Nomura(2026-05-15):Buy,TP HK$727,SOTP(游戏 18x/广告 20x/金融 15x P/E/云 5x P/S,投资组合 20% 折价)。",
            "Bernstein(2026-06-09):Outperform,TP HK$780,20x FY+1 P/E;核心运营业务仅按 2027E ~10x 交易,'非常超卖'。",
            "Deutsche Bank(2026-05-14):Buy,TP HK$800,DCF(WACC 8.0%,TGR 2%)。",
        ],
    })

    # 5 历史估值倍数
    hm = K.write_hist_multiples(wb.create_sheet(S_HMULT), {
        "title": "历史估值倍数 — 自身 Non-IFRS P/E 带 + 当下 + 同业(互联网平台)对照",
        "intro": "腾讯 Non-IFRS P/E 历史大致 13-30x(2021 流动性高点约 27x,2022-2024 监管/修复期 16-23x,2025 年末 19x)。MS 给的更长口径:2015 年来 forward P/E 区间 10-48x。当下现价 417 对应 Non-IFRS forward P/E ~14x,处带下沿(~10-15 分位)。主线以 Non-IFRS EPS×P/E 资本化盈利 + SOTP,P/B 仅体检(轻资产平台)。",
        "s_hist": S_HIST, "ha": ha, "hist_cols": HC, "hist_years": HY,
        "yhigh": {y: px["yhigh"][int(y)] for y in HY}, "ylow": {y: px["ylow"][int(y)] for y in HY},
        "fwd_note": "现价 417 ÷ Base 2026E Non-IFRS EPS(模型)≈14x;落在自身历史带与全球平台同业之下。",
        "self_name": "腾讯", "self_fwd_pe_label": "~14x(现价/2026E) / 16-17x(Base 目标)",
        "self_note": "P/E 主线:穿越周期的耐用资产是微信生态(14.3 亿 MAU)+ 游戏 IP + 高毛利货币化能力,不是账面净资产。",
        "peers": [
            {"name": "阿里巴巴", "yearly": None, "cur_pb": None, "cur_pe": 14.0, "fwd_pe": 12.0,
             "note": "中国电商+云,同样 AI capex 大增;广告(CMR)与云直接可比,倍数区间相近。"},
            {"name": "网易", "yearly": None, "cur_pb": None, "cur_pe": 14.0, "fwd_pe": 13.0,
             "note": "游戏纯玩家,腾讯游戏分部最直接同业;现金流稳、倍数中枢相近。"},
            {"name": "Meta Platforms", "yearly": None, "cur_pb": None, "cur_pe": 25.0, "fwd_pe": 22.0,
             "note": "全球社交广告龙头;盈利耐用性可比,享更高质量倍数(美股 + 无中国折价)。"},
            {"name": "恒生科技指数", "yearly": None, "cur_pb": None, "cur_pe": 18.0, "fwd_pe": 15.0,
             "note": "港股科技大盘参照;自 2025/10 高点回撤近 30%,腾讯随 beta 承压。"},
        ],
        "ratio": None,
        "reading": "① 自身:当下 ~14x 非 2022 监管底,但远低于 2021 高点(27x)与 2025 年末(19x);MS 长口径 10-48x 带下沿。② 同业:腾讯 ~14x ≈ 阿里/网易(中国互联网折价),显著低于 Meta 22x(中国折价 + AI 落后叙事双杀)。③ 结论:AI 落后折价部分过度——Base 给 16-17x(回到自身 2023-2024 中枢、卖方下沿)合理;若视频号加载率放量 + AI 货币化证实可上探 20-21x(Bull,卖方上沿);若 AI 投入持续吞利润 + 港股流动性恶化则停在 12-13x(Bear)。",
    })

    # 6 估值倍数假设
    ma = K.write_multiple_assumptions(wb.create_sheet(S_MULT), {
        "title": "估值倍数假设 — 主线 Non-IFRS P/E 三层拆解(质量锚 × 结构溢价 × 情绪值)",
        "intro": "目标 P/E 不是拍一个数:质量锚 24x × 结构溢价 1.0x(已含在锚内) × 情景情绪值。情景切换只引这里的三案情绪值;所有目标价由同一条业务链推导。SOTP 在情景估值页做交叉验证。",
        "why_text": ("腾讯穿越周期的耐用资产是『微信生态(14.3 亿 MAU + 关系链 + 支付闭环)+ 游戏 IP 矩阵 + 高毛利货币化能力』,不是账面净资产——它轻资产(相对盈利)、有定价权(游戏/广告)、生态锁定。"
                     "因此主估值分母选 Non-IFRS(经调整)归母净利,P/B 仅体检。Non-IFRS 口径剔除股权激励、投资公允价值收益(可逆、不可资本化)、SSV——这是街上看腾讯的标准口径,比对报表 IFRS 净利套倍数更严谨(IFRS 含大额投资损益,逐年剧烈波动:2023 仅 1,152 亿 vs Non-IFRS 1,577 亿)。"
                     "镜头迁移触发:若 AI 替代冲击微信入口、或游戏/广告定价权被侵蚀,镜头才应向更保守倍数下移。"),
        "why_rows": 6,
        "method_text": "目标 P/E = 质量锚(24x)× 结构溢价(1.0x)× 情景情绪值。Bear 回到 11-13x(AI 投入吞利润 + 港股流动性折价 + 回购缩水);Base 2026-27 修复到 16-17x(回自身 2023-2024 中枢、卖方下沿)后随成熟回落到 15x;Bull 在视频号加载率放量 + AI 货币化证实下到 19-21x(卖方上沿 JPM 21x/Bernstein 20x)。",
        "peak": PEAK_PE, "peak_note": "质量锚 24x:腾讯历史 Non-IFRS P/E 高点约 27x(2021 流动性泡沫,不可重现) + 卖方目标倍数 20-21x(JPM/Bernstein)显示平台龙头合理上沿;取 24x 作质量锚(介于历史高点与卖方目标之间),非当下 TTM 拟合。",
        "premium": PREMIUM, "premium_note": "结构溢价 1.0x:已含在 24x 质量锚内,避免双重计数;三案差异放情绪值层(港股流动性/AI 叙事/回购节奏)。",
        "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
        "s_hist": S_HIST, "hpb_row": ha["HPE"],
        "cases": [
            ("Bear", PE_SENT["Bear"], "AI 投入(>360 亿)持续吞利润 + 回购缩水 + 港股外资持续流出/恒科带 beta 下杀 + 视频号不敢放量 → 估值停在 11-13x(2022 监管底之上、但 AI 落后叙事坐实)。"),
            ("Base", PE_SENT["Base"], "广告/云 AI 货币化对冲 AI 投入,Non-IFRS 利润维持双位数 + 回购恢复 + 港股情绪企稳 → 16-17x(回自身 2023-2024 中枢、卖方下沿),随成熟回落到 15x。"),
            ("Bull", PE_SENT["Bull"], "视频号加载率向抖音一半放量 + AI 广告/agentic commerce 货币化证实 + 元宝/混元追平叙事反转 → 市场按平台龙头给 19-21x(卖方上沿)。"),
        ],
        "reconcile_text": "卖方主流 SOTP(隐含整体 ~18-21x 2026E)或直接 21x(JPM)/20x(Bernstein)。本模型 Base 2026E 16x 落在卖方下沿——刻意保守因为:① AI 投入翻倍对近期利润的压制是真实逆风(1Q26 已拖累 88 亿);② 港股外资流出/流动性折价短期难逆转。但 16x 仍是从现价 ~14x 的修复,叠加投资组合安全垫 → 风险收益向上不对称。Bull 20-21x 对齐卖方,需视频号加载率放量 + AI 叙事反转。",
        "source_text": "历史倍数来自复权股价与 Non-IFRS 净利;卖方口径来自 9 家外资行研报库(GS/MS/UBS/JPM/Nomura/Bernstein/DB);目标倍数与情景触发写在本页与情景切换页。质量锚 24x 介于历史高点 27x(2021)与卖方目标 20-21x 之间。",
    })

    # 7 情景切换
    sw = K.write_scenario_switch(wb.create_sheet(S_SW), {
        "title": "情景切换 — 全模型唯一情景参数库(默认 Base)",
        "usage": "B2 是唯一开关。广告 ARPU(×MAU)、游戏增速、FBS 增速、Non-IFRS OPM、目标 P/E 都按当前案联动;估值对比页直接引三案矩阵,不被开关污染。社交网络/其他增速、MAU、净利转换、留存三案共用 Base(非 load-bearing)。",
        "cases": CASES, "default": "Base",
        "triggers": [
            ("Bear", "AI 投入(>360 亿)无货币化对冲、margin 持续被压;视频号加载率因体验顾虑维持低位;港股外资持续流出、恒科带 beta 下杀;回购进一步缩水。任一兑现叠加即翻 Bear。"),
            ("Base", "广告 ARPU 沿加载率温和提升 + AI eCPM(+14%/年);游戏常青 +10%;FBS +9-10%(云/AI 驱动);Non-IFRS OPM 37%→39%;目标倍数修复到 16-17x;回购恢复。"),
            ("Bull", "视频号加载率向抖音一半放量(广告 ARPU +20%+);AI agentic commerce 货币化证实;混元/元宝追平、AI 落后叙事反转;市场给 19-21x。"),
        ],
        "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
        "levers": [
            {"key": "ad_arpu", "name": "广告 ARPU(元/用户/年)", "fmt": K.N1, "cols": FCf,
             "vals": AD_ARPU, "desc": "广告段物理刀。广告收入 = 微信 MAU(物理锚,共用) × 广告 ARPU。ARPU = 加载率 × eCPM 的合成:视频号加载率 3-4% vs 抖音 20%(量空间)+ AI 把 CTR 1%→3%、妙思渗透 65%(价空间)。",
             "stories": {"Bear": "宏观广告预算疲软 + 视频号因体验顾虑不敢放量 → ARPU 仅小幅提升。", "Base": "加载率温和提升 + AI eCPM,ARPU +14%/年。", "Bull": "加载率向抖音一半放量 + AIM+ 全面铺开,ARPU +20%+。"},
             "hist": ad_arpu_hist()},
            {"key": "game_g", "name": "游戏收入增速", "fmt": K.PCT, "cols": FCf,
             "vals": GAME_G, "desc": "游戏段周期刀。本土常青(王者/和平)+ 新游(三角洲 DAU 5,000 万)+ 国际(首破百亿美元)。递延收入是流水先行指标。",
             "stories": {"Bear": "国内宏观拖累 + 新游生命周期回落 + 缺爆款,增速掉到个位数。", "Base": "常青稳态 + 国际延续 + 三角洲贡献,+10% 逐年缓降。", "Bull": "新游持续放量 + 国际加速 + 小游戏抽成,维持双位数。"},
             "hist": hist_growth(REV_SEG[GAME])},
            {"key": "fbs_g", "name": "FBS 收入增速", "fmt": K.PCT, "cols": FCf,
             "vals": FBS_G, "desc": "金融科技及企业服务周期刀。支付(商业支付逐季回暖)+ 理财/分付(分付余额一年 8 倍)+ 云(2025 首次规模化盈利、企业服务 +20%、AI 智算驱动)。",
             "stories": {"Bear": "消费支付疲软 + 云竞争(被运营商云挤压)+ GPU 供给受限,增速回落。", "Base": "商业支付回暖 + 云 AI 智算驱动,+9-10%。", "Bull": "支付加速 + 云国产 ASIC 放量 + 出海,+12-13%。"},
             "hist": hist_growth(REV_SEG[FBS])},
            {"key": "opm", "name": "Non-IFRS 经营利润率", "fmt": K.PCT, "cols": FCf,
             "vals": OPM, "desc": "整体 Non-IFRS 经营利润率(剔除股权激励等)。2024=36%、2025=37%,高毛利结构升级(视频号/小游戏/AI 广告/云盈利)驱动;AI 新产品投入(2026 翻倍至 >360 亿)是反向拖累。",
             "stories": {"Bear": "AI 投入无货币化对冲 + 宏观拖累,margin 回落到 33-35%。", "Base": "广告/云高毛利对冲 AI 投入,margin 缓升到 38-39%。", "Bull": "经营杠杆充分释放 + AI 提效,margin 到 42%。"},
             "hist": OPM_H},
        ],
        "linked": [
            {"key": "sent", "name": "情绪值(P/E 第三层)", "fmt": K.N2,
             "src_sheet": S_MULT, "src_row0": ma["sent_row0"],
             "note": "来自估值倍数假设页。目标 P/E = 质量锚 24x × 1.0 × 本行。"},
        ],
    })
    pk = f"'{S_MULT}'!{ma['pk_cell']}"; pr = f"'{S_MULT}'!{ma['pr_cell']}"
    swpe = sw["next_row"]
    K.lab(wb[S_SW], f"A{swpe}", "目标P/E(当前案)", b=True)
    for col in ALLC:
        K.fml(wb[S_SW], f"{col}{swpe}", f"={pk}*{pr}*{col}{sw['SWACT']['sent']}", K.MX, link=True)
    K.logic(wb[S_SW], f"L{swpe}", "目标 P/E = 质量锚 × 结构溢价 × 当前案情绪值;喂情景估值。")

    # 8 物理锚
    anchor = K.write_anchor(wb.create_sheet(S_ANCHOR), {
        "title": "微信流量盘 — 广告物理锚(MAU × 广告 ARPU)",
        "all_cols": ALLC, "all_years": ALLY,
        "series": [
            ("微信合并 MAU(亿)", WX_MAU + MAU_FWD, "微信合并月活(亿):2025=14.10、1Q26=14.32,接近中国网民天花板,前瞻平缓(14.4→14.65)。三案共用——分歧放广告 ARPU(加载率×eCPM),不在用户盘。", K.N2),
        ],
        "yoy_row": "微信合并 MAU(亿)",
        "source_note": "口径=微信及 WeChat 合并 MAU(腾讯财报)。广告收入 = MAU × 广告 ARPU(情景切换当前案)。广告 ARPU = 加载率 × eCPM 的合成:视频号加载率 3-4% vs 抖音 20% 是量空间,AI(CTR 1%→3%、妙思 65% 渗透)是价空间。",
        "role_note": "作用:广告收入 = 微信 MAU × 广告 ARPU。改 MAU(用户盘)或 ARPU(货币化)→ 广告收入 → 总营收 → Non-IFRS 净利 → 隐含价(连通性测试入口)。游戏/社交/FBS 走周期增长,物理依据在分部测算逻辑列。",
    })
    mau_row = anchor["row_of"]["微信合并 MAU(亿)"]

    # 9 分部测算
    seg = K.write_segment_model(wb.create_sheet(S_SEG), {
        "title": "分部测算 — 广告挂物理锚(MAU×ARPU),游戏/社交/FBS/其他走周期增长",
        "all_cols": ALLC, "all_years": ALLY, "logic_col": "N",
        "groups": [
            ("微信流量盘(物理锚直驱)", [
                ("微信 MAU(亿)", None, K.N2, "引自物理锚页。改 MAU 穿透到广告收入→隐含价。"),
                ("广告 ARPU(元/用户/年)", None, K.N1, "历史=广告收入÷MAU 反推(2025≈103);前瞻=情景切换当前案(加载率×eCPM)。"),
                ("营销服务(广告)收入(十亿)", None, K.N1, "前瞻 = MAU × 广告 ARPU。校验:2026E Base ≈1,685 亿(+16%)≈ 卖方 +18%。"),
            ]),
            ("游戏/社交/FBS/其他(周期增长)", [
                ("游戏增速", None, K.PCT, "前瞻=情景切换当前案。"),
                ("游戏(本土+国际)收入(十亿)", None, K.N1, "前瞻 = 上年 × (1+增速)。"),
                ("社交增速", [None] + SNS_G, K.PCT, "成熟段,三案共用 Base(直播/音乐/会员稳态)。"),
                ("社交网络收入(十亿)", None, K.N1, "前瞻 = 上年 × (1+增速)。"),
                ("FBS增速", None, K.PCT, "前瞻=情景切换当前案。"),
                ("金融科技及企业服务收入(十亿)", None, K.N1, "前瞻 = 上年 × (1+增速)。"),
                ("其他增速", [None] + OTH_G, K.PCT, "小项,三案共用 Base。"),
                ("其他收入(十亿)", None, K.N1, "前瞻 = 上年 × (1+增速)。"),
            ]),
        ],
    })
    m = seg["m"]
    # MAU 引物理锚
    for col in ALLC:
        K.fml(wb[S_SEG], f"{col}{m['微信 MAU(亿)']}", f"={K.R(S_ANCHOR, col + str(mau_row))}", K.N2, link=True)
    # 历史: 收入取实数
    seg_hist_map = {"营销服务(广告)收入(十亿)": AD, "游戏(本土+国际)收入(十亿)": GAME,
                    "社交网络收入(十亿)": SNS, "金融科技及企业服务收入(十亿)": FBS, "其他收入(十亿)": OTH}
    for col in HC:
        for mname, sname in seg_hist_map.items():
            K.fml(wb[S_SEG], f"{col}{m[mname]}", f"={K.R(S_HIST, col + str(ha['seg_rows'][sname]))}", K.N1, link=True)
        K.fml(wb[S_SEG], f"{col}{m['广告 ARPU(元/用户/年)']}", f"={col}{m['营销服务(广告)收入(十亿)']}*1/{col}{m['微信 MAU(亿)']}", K.N1)
    # 历史增速行
    for grow_m, rev_m, vals in [("游戏增速", "游戏(本土+国际)收入(十亿)", REV_SEG[GAME]),
                                 ("社交增速", "社交网络收入(十亿)", REV_SEG[SNS]),
                                 ("FBS增速", "金融科技及企业服务收入(十亿)", REV_SEG[FBS]),
                                 ("其他增速", "其他收入(十亿)", REV_SEG[OTH])]:
        K.lab(wb[S_SEG], f"B{m[grow_m]}", "n.m.", note=True)
        for p, c in zip(HC[:-1], HC[1:]):
            K.fml(wb[S_SEG], f"{c}{m[grow_m]}", f"={c}{m[rev_m]}/{p}{m[rev_m]}-1", K.PCT)
    # 前瞻
    for prev, col in zip(["F"] + FCf[:-1], FCf):
        # 广告: MAU × ARPU
        K.fml(wb[S_SEG], f"{col}{m['广告 ARPU(元/用户/年)']}", f"={K.R(S_SW, col + str(sw['SWACT']['ad_arpu']))}", K.N1, link=True)
        K.fml(wb[S_SEG], f"{col}{m['营销服务(广告)收入(十亿)']}", f"={col}{m['微信 MAU(亿)']}*{col}{m['广告 ARPU(元/用户/年)']}", K.N1)
        wb[S_SEG][f"{col}{m['营销服务(广告)收入(十亿)']}"].fill = K.OUT
        # 游戏
        K.fml(wb[S_SEG], f"{col}{m['游戏增速']}", f"={K.R(S_SW, col + str(sw['SWACT']['game_g']))}", K.PCT, link=True)
        K.fml(wb[S_SEG], f"{col}{m['游戏(本土+国际)收入(十亿)']}", f"={prev}{m['游戏(本土+国际)收入(十亿)']}*(1+{col}{m['游戏增速']})", K.N1)
        # 社交
        K.fml(wb[S_SEG], f"{col}{m['社交网络收入(十亿)']}", f"={prev}{m['社交网络收入(十亿)']}*(1+{col}{m['社交增速']})", K.N1)
        # FBS
        K.fml(wb[S_SEG], f"{col}{m['FBS增速']}", f"={K.R(S_SW, col + str(sw['SWACT']['fbs_g']))}", K.PCT, link=True)
        K.fml(wb[S_SEG], f"{col}{m['金融科技及企业服务收入(十亿)']}", f"={prev}{m['金融科技及企业服务收入(十亿)']}*(1+{col}{m['FBS增速']})", K.N1)
        # 其他
        K.fml(wb[S_SEG], f"{col}{m['其他收入(十亿)']}", f"={prev}{m['其他收入(十亿)']}*(1+{col}{m['其他增速']})", K.N1)

    # 10 利润与收入假设
    fr = K.write_fundamentals(wb.create_sheet(S_FUND), {
        "title": "利润与收入假设 — 分部收入 → Non-IFRS OPM → Non-IFRS 净利/EPS/BPS",
        "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf, "logic_col": "N",
        "s_hist": S_HIST, "ha": ha, "share_fix_col": "F", "ps_scale": 1000,
        "eps_label": "Non-IFRS EPS (RMB)", "bps_label": "BPS (RMB)",
        "assum_groups": [
            ("利润率与转换(历史实际锚 + 前瞻)", [
                {"name": "Non-IFRS 经营利润率", "vals": OPM_H + [None] * 5, "fmt": K.PCT,
                 "logic": "整体 Non-IFRS OPM。2024=36%、2025=37%(高毛利结构升级);前瞻链情景切换(AI 投入压制 vs 广告/云对冲)。2021-2023 为估。",
                 "link": {"sheet": S_SW, "row": sw["SWACT"]["opm"]}},
                {"name": "净利转换率", "vals": [0.91, 0.92, 0.93, 0.937, 0.925] + [None] * 5, "fmt": K.PCT,
                 "logic": "Non-IFRS 净利/Non-IFRS 经营利润(含税+净利息+少数股东)。历史约 0.92-0.94;前瞻 0.92 共用。",
                 "vals_fwd_const": NETCONV},
                {"name": "留存率", "vals": [0.88, 0.87, 0.86, 0.85, 0.85] + [None] * 5, "fmt": K.PCT,
                 "logic": "派息+回购后留存(2025 回购 800 亿港元+派息 410 亿);用于 BPS/ROE,主估值不依赖 P/B。",
                 "vals_fwd_const": RETENT},
            ]),
        ],
        "segments": [
            {"name": GAME, "hist_row": GAME, "fwd": {"sheet": S_SEG, "row": m["游戏(本土+国际)收入(十亿)"]}},
            {"name": SNS, "hist_row": SNS, "fwd": {"sheet": S_SEG, "row": m["社交网络收入(十亿)"]}},
            {"name": AD, "hist_row": AD, "fwd": {"sheet": S_SEG, "row": m["营销服务(广告)收入(十亿)"]}},
            {"name": FBS, "hist_row": FBS, "fwd": {"sheet": S_SEG, "row": m["金融科技及企业服务收入(十亿)"]}},
            {"name": OTH, "hist_row": OTH, "fwd": {"sheet": S_SEG, "row": m["其他收入(十亿)"]}},
        ],
        "profit_terms": [
            ([GAME, SNS, AD, FBS, OTH], "Non-IFRS 经营利润率", False),
        ],
        "conv_assum": "净利转换率", "retention_assum": "留存率",
        "note_text": "前瞻 Non-IFRS 净利 = 总营收 × Non-IFRS OPM(情景刀)× 净利转换率,避免用外部卖方净利截断链条。历史净利=Non-IFRS 归母实际(链历史财务页 HNI)。EPS=Non-IFRS 净利;前瞻股本保守按 9,080mn 固定(不计未来回购增厚,属保守处理)。",
    })
    # 前瞻常数补齐(净利转换/留存)
    for col in FCf:
        K.fml(wb[S_FUND], f"{col}{fr['am']['净利转换率']}", f"={NETCONV}", K.PCT)
        K.fml(wb[S_FUND], f"{col}{fr['am']['留存率']}", f"={RETENT}", K.PCT)

    # 11 情景估值(P/E 主线 + SOTP 交叉)
    fr_pe = dict(fr); fr_pe["BPS"] = fr["EPS"]            # P/E 主线: 分母换成 EPS
    sv = K.write_scenario_valuation(wb.create_sheet(S_VAL), {
        "title": "情景估值 — 当前案逐年隐含价 (Non-IFRS P/E 主线; SOTP 交叉验证)",
        "intro": "本表输出=情景切换当前案(默认 Base)。隐含价 = 目标 P/E(当前案) × Non-IFRS EPS × FX(HKD)。历史列=实际年末价反推(事实);前瞻是预测、不拟合现价。三案并排见估值对比。",
        "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf, "hist_years": HY,
        "fx_fwd": FX, "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
        "s_fund": S_FUND, "fr": fr_pe,
        "s_switch": S_SW, "target_row": swpe, "sw_cell": "B2",
        "yend": PX_END, "yavg": PX_AVG,
        "mcap_div": 100, "mcap_usd_skip": True,
        "mcap_hist_label": "市值 实际年末(亿 HKD, 历史)", "mcap_fwd_label": "市值 前瞻·主线(亿 HKD)",
        "reading": "P/E 主线读法:历史列=实际年末价÷Non-IFRS EPS(回测,模型倍数贴合度);前瞻=目标 P/E×EPS×FX。隐含 forward P/E 体检行 = 隐含价÷该年 EPS,应等于目标 P/E(主线 P/E 镜头下二者一致)。下方 SOTP 块从分部加总独立交叉验证。",
        "method": "方法:整体公司、Non-IFRS P/E 主线逐年估。基本面在利润与收入假设;目标 P/E 在估值倍数假设(三层);本表套用 目标P/E×EPS×FX→隐含价(HKD)。SOTP 从分部收入×分部 margin×分部倍数 + 投资组合(30% 折价)独立加总交叉。",
        "concl": "结论(方向性):Base 2026E 隐含价(12 个月目标)较现价 +28% 上行;2027E 进一步上行。Bear 接近现价(下行有限,投资组合+回购托底);Bull 对齐卖方上沿。风险收益向上不对称 → BUY。三情景与概率加权见估值对比与决策 memo。",
    })

    # SOTP 交叉镜头(2027E, 列 H)
    write_tencent_sotp(wb[S_VAL], {
        "s_fund": S_FUND, "fr": fr, "s_seg": S_SEG, "m": m, "s_hist": S_HIST, "ha": ha,
        "px_now": PX_NOW, "fx": FX, "sh_now": SH_NOW,
    })

    # 12 估值对比
    SWB = sw["SWB"]
    SH_F = K.R(S_HIST, f"$F${ha['HSH']}")
    AI_H = {sn: ha["seg_rows"][sn] for sn in [GAME, SNS, AD, FBS, OTH]}

    def hist_seg(sn): return lambda col, ci, A: f"={K.R(S_HIST, col + str(AI_H[sn]))}"

    def prevcol(col): return ALLC[ALLC.index(col) - 1]

    rows = [
        {"key": "ad", "label": "广告收入(十亿)", "fmt": K.N1, "hist": hist_seg(AD),
         "fwd": lambda c, j, ci, A: f"={K.R(S_ANCHOR, c + str(mau_row))}*{K.R(S_SW, c + str(SWB['ad_arpu'] + ci))}"},
        {"key": "game", "label": "游戏收入(十亿)", "fmt": K.N1, "hist": hist_seg(GAME),
         "fwd": lambda c, j, ci, A: (f"={K.R(S_HIST, 'F' + str(AI_H[GAME]))}*(1+{K.R(S_SW, c + str(SWB['game_g'] + ci))})" if j == 0 else f"={prevcol(c)}{A['game']}*(1+{K.R(S_SW, c + str(SWB['game_g'] + ci))})")},
        {"key": "sns", "label": "社交收入(十亿)", "fmt": K.N1, "hist": hist_seg(SNS),
         "fwd": lambda c, j, ci, A: f"={K.R(S_SEG, c + str(m['社交网络收入(十亿)']))}"},
        {"key": "fbs", "label": "FBS收入(十亿)", "fmt": K.N1, "hist": hist_seg(FBS),
         "fwd": lambda c, j, ci, A: (f"={K.R(S_HIST, 'F' + str(AI_H[FBS]))}*(1+{K.R(S_SW, c + str(SWB['fbs_g'] + ci))})" if j == 0 else f"={prevcol(c)}{A['fbs']}*(1+{K.R(S_SW, c + str(SWB['fbs_g'] + ci))})")},
        {"key": "oth", "label": "其他收入(十亿)", "fmt": K.N1, "hist": hist_seg(OTH),
         "fwd": lambda c, j, ci, A: f"={K.R(S_SEG, c + str(m['其他收入(十亿)']))}"},
        {"key": "rev", "label": "总营收(十亿)", "fmt": K.N1, "bold": True,
         "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HREV']))}",
         "fwd": lambda c, j, ci, A: f"={c}{A['ad']}+{c}{A['game']}+{c}{A['sns']}+{c}{A['fbs']}+{c}{A['oth']}"},
        {"key": "ni", "label": "Non-IFRS 净利(十亿)", "fmt": K.N1, "bold": True,
         "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HNI']))}",
         "fwd": lambda c, j, ci, A: f"={c}{A['rev']}*{K.R(S_SW, c + str(SWB['opm'] + ci))}*{NETCONV}"},
        {"key": "eps", "label": "Non-IFRS EPS(RMB)", "fmt": K.N2,
         "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HNI']))}*100/{K.R(S_HIST, c + str(ha['HSH']))}",
         "fwd": lambda c, j, ci, A: f"={c}{A['ni']}*100/{SH_F.replace('$', '')}" if False else f"={c}{A['ni']}*100/{K.R(S_HIST, '$F$' + str(ha['HSH']))}"},
        {"key": "tpe", "label": "目标 P/E", "fmt": K.MX,
         "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HPE']))}",
         "fwd": lambda c, j, ci, A: f"={K.R(S_MULT, c + str(ma['target_row0'] + ci))}"},
        {"key": "px", "label": "隐含股价(HKD)", "fmt": K.PX, "bold": True, "out": True,
         "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HPX']))}",
         "fwd": lambda c, j, ci, A: f"={c}{A['tpe']}*{c}{A['eps']}*{FX}"},
        {"key": "ipe", "label": "隐含 forward P/E(体检)", "fmt": K.MX,
         "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HPE']))}",
         "fwd": lambda c, j, ci, A: f"={c}{A['px']}/({c}{A['eps']}*{FX})"},
        {"key": "up", "label": "历史回测 / 前瞻 vs 现价", "fmt": K.PCT,
         "hist": lambda c, ci, A: f"={c}{A['px']}/{K.R(S_HIST, c + str(ha['HPX']))}-1",
         "fwd": lambda c, j, ci, A: f"={c}{A['px']}/{PX_NOW}-1"},
    ]
    cmp = K.write_comparison(wb.create_sheet(S_CMP), {
        "title": "估值对比 — Bear / Base / Bull 三案并排",
        "intro": "三案从广告 ARPU、游戏/FBS 增速、Non-IFRS OPM、目标 P/E 同一条链推导。主判断年取 2026E(12 个月目标)+ 2027E。隐含价用 Non-IFRS EPS×目标 P/E×FX。历史列同链填实际、隐含价≈实际年末价=内置回测。",
        "case_names": CASES, "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
        "block_start": 24, "rows": rows,
        "summary": {
            "band": "2026E / 2027E 三案摘要(12 个月目标 = 2026E)", "target_col": "G",
            "rows": [
                ("总营收(十亿)", "rev", K.N1, "分部合计。", False),
                ("Non-IFRS 净利(十亿)", "ni", K.N1, "总营收 × Non-IFRS OPM × 净利转换。", False),
                ("Non-IFRS EPS(RMB)", "eps", K.N2, "主估值分母。", False),
                ("目标 P/E", "tpe", K.MX, "质量锚 24x × 情绪值。", False),
                ("隐含股价 2026E(HKD)", "px", K.PX, "P/E 主线 12 个月目标。", True),
                ("vs 现价", "up", K.PCT, f"现价 reality check(HK${PX_NOW:.0f})。", True),
            ],
            "mcap": {"label": "隐含市值 2026E(亿 HKD)", "key": "px", "expr": f"*{SH_NOW}/100", "note": "隐含股价 × 股本。"},
            "concl": "Base 2026E 隐含价较现价显著上行(12 个月目标);Bear 接近现价(下行有限,投资组合+回购托底);Bull 对齐卖方上沿。风险收益向上不对称 → BUY。精确隐含价见各案 block 与 2027E 行(切换开关查 2027E)。",
        },
    })

    # 13 仪表盘
    K.write_dashboard(wb.create_sheet(S_DASH), {
        "title": "综合判断仪表盘 — 基本面拐点 / 估值错位 / 催化剂 / 情绪",
        "usage": "把模型压成投后跟踪语言:哪些指标验证 Base,哪些把模型推向 Bear/Bull。B 列公式直接引模型输出(随情景切换变)。",
        "blocks": [
            {"title": "A. 基本面拐点", "rows": [
                ("1Q26 收入/Non-IFRS 净利", "+9% / +11%", "收入降速到近六季最低,但 Non-IFRS 利润仍双位数;剔除新 AI 产品经营利润 +17%(margin 43%)——底层业务比表观健康。", True),
                ("广告增速/毛利率", "+20% / 58%", "广告是 AI 货币化最确定兑现:连 11 季双位数、毛利率升;视频号加载率 3-4% 仍有 2.5-5x 空间。"),
                ("Base 2027E Non-IFRS 净利", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Base']['ni']))}", "fmt": K.N1, "fill": True}, "若 2027 净利不及该路径,Base 估值先降。"),
            ]},
            {"title": "B. 估值错位(预测引擎)", "rows": [
                ("当前股价", {"inp": PX_NOW, "fmt": K.PX, "fill": True}, "reality check,不反向拟合。"),
                ("Base 2026E 隐含价(12M 目标)", {"fml": f"={K.R(S_CMP, 'G' + str(cmp['CMPA']['Base']['px']))}", "fmt": K.PX, "fill": True}, "主判断输出。"),
                ("Base 2026E vs 现价", {"fml": f"={K.R(S_CMP, 'G' + str(cmp['CMPA']['Base']['up']))}", "fmt": K.PCT, "fill": True}, "GAP 正且大 = 重估空间。"),
                ("Bear 2026E 隐含价", {"fml": f"={K.R(S_CMP, 'G' + str(cmp['CMPA']['Bear']['px']))}", "fmt": K.PX}, "下行风险价:AI 投入吞利润+港股流动性折价。"),
                ("Bull 2026E 隐含价", {"fml": f"={K.R(S_CMP, 'G' + str(cmp['CMPA']['Bull']['px']))}", "fmt": K.PX}, "上行:视频号放量+AI 货币化+估值修复。"),
            ]},
            {"title": "C. 催化剂", "rows": [
                ("Bull 触发", "视频号加载率放量(广告 ARPU 加速);微信 AI agentic commerce 货币化;混元/元宝追平、AI 落后叙事反转;回购恢复。", "先抬 ad_arpu/opm,再抬目标 P/E。"),
                ("Bear 触发", "AI 投入持续吞利润且无货币化对冲;港股外资持续流出/恒科带 beta 下杀;游戏增长失速。", "先砍 ad_arpu/opm,再把目标 P/E 压到 12x 以下。"),
                ("最易错处", "用 IFRS 报表净利(含大额投资损益,逐年剧烈波动)套倍数会失真。", "本模型用 Non-IFRS 经调整净利,避免这个陷阱。"),
            ]},
            {"title": "D. 综合判断", "rows": [
                ("一句话结论", "BUY:市场对 AI 落后的折价部分过度,估值杀到历史下沿(~14x);AI 在广告/云/微信入口已兑现,投资组合(约市值 1/4)+ 回购提供下行保护。", "上行看视频号放量+AI 货币化,下行看 AI 投入与港股流动性。", True),
            ]},
        ],
        "final": {"band": "最终判断",
                  "text": "腾讯的 alpha 问题:2026 年市场把它从中国互联网核心资产重估为『AI 落后者 + 投入吞利润 + 回购缩水』,叠加港股系统性折价,把 Non-IFRS forward P/E 杀到 ~14x(2015 年来 10-48x 带下沿)。本模型判定折价部分过度——AI 已在广告(CTR 1%→3%)/云(90% 大模型厂商)/微信 14.3 亿 MAU 入口兑现,非赌独立 C 端产品;近期逆风(AI 投入压利润、增长降速)真实但可控。Base 取温和修复(16-17x),隐含价显著高于现价;下行有投资组合(1.04 万亿,30% 折价仍是安全垫)+ 持续回购托底。评级 BUY。"},
        "tracking": {
            "intro": "投后跟踪按季报滚动更新。",
            "rows": [
                ("__band__", "一、增长核心(广告)"),
                ("广告 ARPU / 视频号加载率", "广告 +20%(1Q26) / 加载率 3-4%", "关键敏感项:加载率放量节奏 + AI eCPM", "季报营销服务 + 加载率口径", "加载率停滞/广告增速回落 → 转 Bear 看广告"),
                ("__band__", "二、利润(AI 投入)"),
                ("Non-IFRS OPM / AI 投入", "37%(2025) / AI 投入翻倍 >360 亿", "关键敏感项:AI 投入是否被货币化对冲", "季报 Non-IFRS 经营利润 + AI 拖累口径", "margin 持续回落 → 下调 opm/目标 P/E"),
                ("__band__", "三、估值与资金"),
                ("Non-IFRS forward P/E / 回购", "~14x / 连续 24 日回购", "关键敏感项:估值修复 + 回购支撑", "现价÷模型 EPS + 回购公告", "回购停 + 倍数破 12x → 转 Bear"),
                ("__band__", "四、安全垫"),
                ("投资组合公允价值", "1.04 万亿(2025末)", "关键敏感项:SOTP 下行保护", "季报上市/非上市投资公允价值", "组合大幅缩水 → 下调 SOTP 安全垫"),
            ],
        },
    })

    # input.json
    payload = {
        "ticker": "00700.HK",
        "company": "腾讯控股 Tencent Holdings",
        "built_at": "2026-06-23",
        "currency": "RMB(基本面) / HKD(股价)", "fx_hkd_per_rmb": FX,
        "current_price_hkd": PX_NOW, "market_cap_hkd_bn": 38000, "shares_m": SH_NOW,
        "rating": "BUY(买入)", "target_price": None,
        "method": "Non-IFRS forward P/E 主线(质量锚 24x × 情绪值)+ SOTP(分部 P/E + 投资组合 30% 折价)交叉;物理锚=微信 MAU × 广告 ARPU → 广告收入。",
        "eps_note": "估值用 Non-IFRS 经调整归母净利(剔除股权激励/投资公允价值收益/SSV);报表 IFRS 净利仅参照。",
        "historical_financials_rmb_bn": {
            "revenue": [5601, 5546, 6090, 6603, 7518],
            "segments_fy2025": {k: v[-1] for k, v in REV_SEG.items()},
            "non_ifrs_ni": NI_ADJ, "ifrs_ni": NI_REP,
            "gross_margin": GM, "capex_2025": 792, "fcf_2025": 1826,
        },
        "anchor": {"type": "platform", "driver": "微信 MAU × 广告 ARPU", "wx_mau_2025": 14.10},
        "investment_portfolio_rmb_bn": {"listed": 6727, "unlisted": 3631, "discount": 0.30, "net_cash": 1071},
        "scenario": {"target_pe": TARGET_PE, "peak_pe": PEAK_PE, "pe_sentiment": PE_SENT,
                     "ad_arpu": AD_ARPU, "game_growth": GAME_G, "fbs_growth": FBS_G, "non_ifrs_opm": OPM},
        "consensus": {"target_mean_hkd": 732, "target_high_hkd": 800, "target_low_hkd": 650,
                      "rating": "buy_all_9_brokers",
                      "fy26e_revenue_rmb_bn": 8327, "fy26e_non_ifrs_ni_rmb_bn": 2910, "fy26e_non_ifrs_eps_rmb": 30.3,
                      "sellside_tp_hkd": {"GS": 700, "MS": 650, "UBS": 780, "JPM": 690, "Nomura": 727, "Bernstein": 780, "DB": 800}},
        "sources": {
            "official": ["腾讯 2025 年报(2026-03-18)/ 1Q26(2026-05-13)/ 2024 年报(2025-03-19)PDF"],
            "data_api": "可追溯行情/三表/K线/一致预期数据快照 2026-06-23",
            "research_kb": ["GS/MS/UBS/JPM/Nomura/Bernstein/DB 研报库 2026 Q1-Q2"],
        },
    }
    os.makedirs(VAULT, exist_ok=True)
    with open(os.path.join(VAULT, "00700.HK_input.json"), "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    K.finalize(wb)
    out = os.path.join(VAULT, "00700.HK_valuation_model.xlsx")
    wb.save(out)
    print("saved:", out)
    print("sheets:", wb.sheetnames)
    return cmp


def write_tencent_sotp(ws, d):
    """SOTP 交叉镜头(2027E, 列 H)— 分部收入×分部 margin×分部 P/E + 投资组合 30% 折价。"""
    fr, m = d["fr"], d["m"]
    S_FUND, S_SEG, S_HIST, ha = d["s_fund"], d["s_seg"], d["s_hist"], d["ha"]
    px_now, fx, sh_now = d["px_now"], d["fx"], d["sh_now"]
    r = ws.max_row + 2
    K.band(ws, r, "交叉镜头 SOTP(2027E 分部加总, 亿 RMB)— 游戏/广告/社交/FBS 各给 P/E + 投资组合 30% 折价", 11); r += 1
    H = "H"
    segrow = fr["seg_rows"]

    def seg_block(label, seg_key, margin, pe):
        nonlocal r
        K.lab(ws, f"A{r}", f"{label} 2027E 收入 × margin × P/E")
        K.fml(ws, f"B{r}", f"={K.R(S_FUND, H + str(segrow[seg_key]))}", K.N1, link=True)
        K.inp(ws, f"C{r}", margin, None, K.PCT)
        K.inp(ws, f"D{r}", pe, None, K.MX)
        K.fml(ws, f"E{r}", f"=B{r}*C{r}*D{r}", K.N0)
        val_row = r
        r += 1
        return val_row

    g = seg_block("游戏", GAME, SOTP["game_margin"], SOTP["game_pe"])
    a = seg_block("广告", AD, SOTP["ad_margin"], SOTP["ad_pe"])
    s = seg_block("社交", SNS, SOTP["sns_margin"], SOTP["sns_pe"])
    f = seg_block("FBS", FBS, SOTP["fbs_margin"], SOTP["fbs_pe"])
    K.lab(ws, f"A{r}", "核心业务估值合计(亿 RMB)", b=True)
    K.fml(ws, f"E{r}", f"=E{g}+E{a}+E{s}+E{f}", K.N0); core = r; r += 1
    K.lab(ws, f"A{r}", "投资组合(上市+非上市)× (1−30% 折价)")
    K.inp(ws, f"B{r}", SOTP["listed_inv"] + SOTP["unlisted_inv"], None, K.N0)
    K.inp(ws, f"C{r}", 1 - SOTP["inv_discount"], None, K.PCT)
    K.fml(ws, f"E{r}", f"=B{r}*C{r}", K.N0); inv = r; r += 1
    K.lab(ws, f"A{r}", "净现金(亿 RMB)")
    K.inp(ws, f"E{r}", SOTP["net_cash"], None, K.N0); nc = r; r += 1
    K.lab(ws, f"A{r}", "SOTP 隐含市值(亿 RMB)", b=True)
    K.fml(ws, f"E{r}", f"=E{core}+E{inv}+E{nc}", K.N0); mc = r; r += 1
    K.lab(ws, f"A{r}", "SOTP 隐含股价(HKD)", b=True); ws[f"A{r}"].fill = K.OUT
    K.fml(ws, f"E{r}", f"=E{mc}/{sh_now}*100*{fx}", K.PX); sp = r; r += 1
    K.lab(ws, f"A{r}", "SOTP vs 现价")
    K.fml(ws, f"E{r}", f"=E{sp}/{px_now}-1", K.PCT); r += 1
    K.logic(ws, f"G{g}", "SOTP 倍数(卖方下沿~中端):游戏 15x(成熟现金牛)、广告 22x(AI 货币化成长溢价)、社交 16x、FBS 15x P/E;投资组合(上市 6,727+非上市 3,631 亿)打 30% 控股折价;净现金 1,071 亿。2027E 分部收入引利润与收入假设(Base)。")
    K.lab(ws, f"A{r}", "SOTP 方法注", note=True)
    K.fml(ws, f"B{r}", "＝SOTP 隐含价是 P/E 主线的独立交叉:把高 margin/高成长分部(广告)单独给溢价倍数、投资组合单独折价计价——这是 P/E 主线(整体一个倍数)可能低估之处。两镜头收敛即验证;分歧即市场赌点。", K.N1)
    ws.merge_cells(f"B{r}:K{r}")
    ws[f"B{r}"].alignment = __import__("openpyxl").styles.Alignment(wrap_text=True, vertical="top")


if __name__ == "__main__":
    build()
