# -*- coding: utf-8 -*-
"""
Circle Internet Group / CRCL 估值模型（equity-research-obsidian Phase 2）。

物理锚 = float-yield（浮存金 × 收益率）：USDC 平均流通量 × 储备收益率 = 储备收入
        − 分销分成（付 Coinbase 等）+ 其他收入 = 净营业收入（扣分销）
        → 经调整营业利润 → 经调整净利 → 目标 P/E → 隐含股价。
主线镜头：P/E（盈利资本化）；DCF + P/S 交叉验证。不用 P/B 主线（资产轻，储备是过手负债）。
★ 三个 load-bearing 情景杠杆：USDC 流通量增速 × 储备收益率（跟美联储）× 分销分成比。
★ 估值口径：经调整净利（剔除一次性 IPO 股份支付 $566M，含正常化经营 SBC）；
   GAAP FY2025 净亏 $(69.5)M 系 IPO SBC 一次性冲击，不可资本化。
公司数据来自 SEC 一手（10-K CIK 0001876042 / 424B4 / Q1'26 10-Q）+ 统一 API 行情/共识 + FRED 利率。
"""
from __future__ import annotations
import json, os
from openpyxl import Workbook
import build_kit as K

ALLC = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
ALLY = ["2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
HC, HY = ["B", "C", "D", "E", "F"], ["2021", "2022", "2023", "2024", "2025"]
FC = ["F", "G", "H", "I", "J", "K"]     # 前瞻含基年 F=2025A
FCf = FC[1:]                             # 纯前瞻 2026E-2030E
CASES = ["Bear", "Base", "Bull"]
FX = 1.0                                 # USD 单币种
SHARES_M = 248.5                         # 全类别摊薄股本（A 228.1M + B 18.7M ≈ 247M；市值/现价口径 248.5M）
PX_NOW = 68.65                           # 2026-07-06 收盘

S_COVER, S_HIST, S_PX, S_CONS = "封面", "历史财务与估值", "股价走势", "卖方研报共识"
S_HMULT, S_MULT, S_SW = "历史估值倍数", "估值倍数假设", "情景切换"
S_ANCHOR, S_SEG, S_FUND = "USDC流通量物理锚", "分部测算", "利润与收入假设"
S_VAL, S_CMP, S_DASH = "情景估值", "估值对比", "综合判断仪表盘"

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
VAULT = os.environ.get("VALUATION_OUTPUT_DIR", os.path.join(REPO_ROOT, "out"))
OUT_XLSX = os.path.join(VAULT, "CRCL_valuation_model.xlsx")
OUT_JSON = os.path.join(VAULT, "CRCL_input.json")

# ════════ 公司数据（$B，SEC 一手）2021-2025 ════════
RESERVE = [0.025, 0.736, 1.431, 1.661, 2.637]     # 储备收入
OTHER   = [0.090, 0.036, 0.020, 0.015, 0.110]     # 其他收入（订阅/服务/交易费）
DIST    = [0.010, 0.287, 0.720, 1.011, 1.662]     # 分销与交易成本（付 Coinbase 等）
AVGCIRC = [25, 50, 32, 34, 58]                    # USDC 平均流通量（$B，本报告按储备收入/收益率反推校准）
YIELD_H = [0.001, 0.0147, 0.0447, 0.0489, 0.0455] # 隐含储备收益率
DISTR_H = [0.40, 0.39, 0.503, 0.609, 0.630]       # 分销分成比 = 分销/储备收入
ADJNI   = [-0.05, 0.10, 0.30, 0.21, 0.40]         # 经调整净利（估值口径；剔一次性 IPO SBC）
GAAPNI  = [-0.077, -0.769, 0.268, 0.156, -0.070]  # GAAP 净利（含一次性，仅备注）
EQUITY  = [0.15, 0.25, 0.42, 0.5705, 3.3308]      # 股东权益（2024/2025 为 10-K 实数）
SHARES  = [190, 195, 200, 210, 247]               # 股本（mn，IPO 前为普通股等价粗估）
OPM_H   = [None, 0.258, 0.513, 0.395, 0.461]      # 经调整经营利润率（on 净营业收入，反推）
NETCV_H = [None, 0.80, 0.80, 0.80, 0.80]          # 净利转换率（≈税后+净利息，近似）
SUBTOT  = [round(RESERVE[i] - DIST[i] + OTHER[i], 3) for i in range(5)]  # 净营业收入(扣分销)

# 年末股价（IPO 2025-06；2021-2024 无公开市价 → n.m.）
PX_END = {"2021": None, "2022": None, "2023": None, "2024": None, "2025": 76.0}
PX_AVG = {"2025": 140.0}
MONTHLY = [
    ("2025-06", 107.70), ("2025-07", 223.78), ("2025-08", 168.0), ("2025-09", 130.0),
    ("2025-10", 115.0), ("2025-11", 82.0), ("2025-12", 76.0),
    ("2026-01", 60.0), ("2026-02", 49.90), ("2026-03", 126.03), ("2026-04", 100.0),
    ("2026-05", 113.0), ("2026-06", 70.0), ("2026-07", 68.65),
]
PX_HIGH = {2025: 262.97, 2026: 126.03}
PX_LOW = {2025: 76.0, 2026: 49.90}

# ════════ 情景参数 ════════
PEAK_PE, PREMIUM = 30.0, 1.0
TARGET_PE = {"Bear": [22, 18, 16, 15, 14], "Base": [30, 28, 26, 24, 23], "Bull": [34, 35, 33, 31, 29]}
PE_SENT = {c: [round(v / PEAK_PE, 3) for v in vals] for c, vals in TARGET_PE.items()}

# 流通量增速（作用于 2025A 基数 $58B）
CIRC_G = {"Bear": [0.29, 0.17, 0.14, 0.10, 0.09], "Base": [0.38, 0.475, 0.31, 0.23, 0.18],
          "Bull": [0.55, 0.67, 0.43, 0.35, 0.28]}
# 储备收益率（%，跟美联储路径）
YLD = {"Bear": [0.034, 0.026, 0.023, 0.0225, 0.0225], "Base": [0.037, 0.033, 0.031, 0.030, 0.030],
       "Bull": [0.039, 0.037, 0.036, 0.035, 0.035]}
# 分销分成比（%）
DISTR = {"Bear": [0.63, 0.65, 0.66, 0.66, 0.66], "Base": [0.62, 0.61, 0.60, 0.59, 0.58],
         "Bull": [0.61, 0.58, 0.55, 0.53, 0.52]}
# 经调整经营利润率（on 净营业收入）
OPM = {"Bear": [0.35, 0.33, 0.32, 0.32, 0.32], "Base": [0.40, 0.43, 0.45, 0.46, 0.47],
       "Bull": [0.43, 0.47, 0.50, 0.51, 0.52]}
# 净利转换率
NETCV = {"Bear": [0.78, 0.78, 0.78, 0.78, 0.78], "Base": [0.80, 0.80, 0.80, 0.80, 0.80],
         "Bull": [0.82, 0.82, 0.82, 0.82, 0.82]}
# 其他收入（$B，三案共用；小体量非 load-bearing）
OTHER_FWD = [0.165, 0.245, 0.345, 0.460, 0.590]


def hg(vals):
    return [None] + [round(vals[i] / vals[i - 1] - 1, 4) if vals[i - 1] else None for i in range(1, len(vals))]


def build():
    wb = Workbook(); wb.remove(wb["Sheet"])

    # ═══ 1 封面 ═══
    K.write_cover(wb.create_sheet(S_COVER), {
        "title": "Circle Internet Group / CRCL 估值模型",
        "meta": [
            ("报告日期", "2026-07-07"),
            ("数据截止", "FY2025 10-K + Q1'26 10-Q + 统一API 股价/共识（2026-07-06 收盘）+ FRED 利率"),
            ("当前股价", f"${PX_NOW:,.2f}（2026-07-06；reality check 用，非拟合目标）；市值约 $17.1B；较峰值 $298.99 回撤约 77%"),
            ("主线镜头", "P/E 主线（经调整盈利资本化）；DCF + P/S 交叉验证。不用 P/B（资产轻，储备是过手负债不进权益）。"),
            ("估值口径", "经调整净利：剔除一次性 IPO 股份支付（FY2025 SBC $566M），含正常化经营 SBC。GAAP FY2025 净亏 $(69.5)M 系 IPO SBC 一次性冲击，不可资本化。"),
            ("方法一句话", "USDC 平均流通量 × 储备收益率 = 储备收入 −（分销分成 62%）+ 其他 = 净营业收入 → 经调整净利/EPS → 目标 P/E → 隐含股价。"),
        ],
        "takeaways": [
            ("① 物理锚", "USDC 平均流通量 × 储备收益率 = 储备收入（FY2025 占营收 96%）。流通量、利率、分销分成三个物理量相乘决定价值。"),
            ("② 核心矛盾", "规模在涨、单位经济学在退化：营收 FY2022→FY2025 涨 3.6×，但分销分成从 39% 升到 63%，留存净额只涨 2.2×。"),
            ("③ 三重挤压", "利率下行（EFFR 5.33%→3.63%）压储备收益率；Coinbase 分成 62%（2026-08 续约）；OUSD 联盟（2026-06-30，140+ 家）欲把储备利息分给采用方，直击单一收入引擎。"),
            ("④ 三情景（2027E）", "Bear ~$19（利率崩+份额失+分成升）/ Base ~$69（流通量按 GENIUS Act 采用增长、利率企稳）/ Bull ~$145（USDC 破 $150B+利率高位+分成改善）。三个杠杆相乘 → 估值区间极宽。"),
            ("⑤ 评级", "HOLD（持有）：77% 回撤已挤掉泡沫，但 $17B / 约 28-30x 2027E 经调整 EPS 仍需『流通量续增 + 利率不再大跌 + Coinbase/OUSD 经济学不恶化』三者同时成立才撑得住现价；风险收益呈杠铃形，等 2026-08 Coinbase 续约与 Q2 业绩落地。"),
        ],
    })

    # ═══ 2 历史财务与估值 ═══
    ha = K.write_history(wb.create_sheet(S_HIST), {
        "title": "Circle 历史财务与估值（$B）— 2021-2025A（IPO 2025-06；上市前无市价）",
        "hist_cols": HC, "hist_years": HY, "fx_hist": [1] * 5, "fx_now": 1,
        "vals_in_usd": True, "fx_label": "FX (USD 单币种=1)",
        "cur_label": "当下(2026-07-06)",
        "segments": [("储备收入", RESERVE, True), ("其他收入", OTHER, True)],
        "total_now": 2.86,
        "gm_pct": [round(SUBTOT[i] / (RESERVE[i] + OTHER[i]), 4) for i in range(5)], "gm_now": 0.40,
        "gm_label": "净收入率(扣分销后/总营收)",
        "ni": ADJNI, "ni_now": 0.42,
        "eq": EQUITY, "eq_now": 3.40,
        "shares": SHARES, "shares_now": 248,
        "px_end": [PX_END[y] for y in HY], "px_now": PX_NOW,
        "px_avg": [PX_AVG.get(y) for y in HY],
        "band_note": "上市前（2021-2024）无公开市价，P/E/P/B 仅 2025A 起有意义；2025A P/E 按经调整 EPS。当下现价对应 forward P/E 见『情景估值』。",
        "notes": [
            ("储备收入", "储备资产（88% BlackRock Circle Reserve Fund 政府货基 + 12% 现金）的利息，FY2025 占营收 96%。= 平均流通量 × 加权收益率，随美联储利率波动。来源：FY2025 10-K。"),
            ("其他收入", "订阅与服务（USYC/CCTP/Circle Mint/Arc 等），FY2025 $109.8M（同比 +624%），占营收不足 5%，多元化早期信号。"),
            ("HGMP", "净收入率 =（储备收入 − 分销成本 + 其他）÷ 总营收。分销成本占储备收入从 39%(2022) 升到 63%(2025) → 净收入率从 63% 压到 40%。分销明细见『分部测算』。"),
            ("HNI", "经调整净利（估值口径，剔一次性 IPO SBC）；GAAP：2021 −77M / 2022 −769M(可转债/权证公允价值损失) / 2023 +268M / 2024 +156M / 2025 −70M(IPO SBC $566M 一次性)。经调整 EBITDA FY2025 $582M 创新高。"),
            ("HEQ", "股东权益：2024 $570.5M / 2025 $3,330.8M(10-K；IPO 募资 + $1.1B 优先股转普通股)。储备资产 $75B 对应负债端 stablecoin holders deposits，不属股东。"),
            ("HSH", "股本：2025 年末 Class A 228.1M + Class B 18.7M ≈ 247M；IPO 前为普通股等价粗估（P/E 无市价、不影响估值）。"),
            ("HPX", "IPO 2025-06-05 定价 $31，首日收 $83.23，2025-06-23 盘中峰值 $298.99；2025 年末约 $76。2021-2024 无市价。"),
        ],
    })
    # 上市前 P/E·P/B 置 n.m.（无市价）
    for col in ["B", "C", "D", "E"]:
        K.lab(wb[S_HIST], f"{col}{ha['HPE']}", "n.m.", note=True)
        K.lab(wb[S_HIST], f"{col}{ha['HPB']}", "n.m.", note=True)
        K.lab(wb[S_HIST], f"{col}{ha['HMC']}", "n.m.", note=True)

    # ═══ 3 股价走势 ═══
    def phase_fn(ym):
        if ym <= "2025-07":
            return "① IPO 暴涨+GENIUS 冲高"
        if ym <= "2026-02":
            return "② 利好出尽+解禁+降息下杀"
        if ym <= "2026-04":
            return "③ 超跌反弹"
        return "④ Q1 放缓+OUSD 威胁再下台阶"
    px = K.write_price_chart(wb.create_sheet(S_PX), MONTHLY, {
        "fn": phase_fn,
        "rows": [("① IPO 暴涨+GENIUS 冲高", "2025-06-05 IPO $31→首日 $83；6-23 峰值 $298.99；7-18 GENIUS Act 签署前冲高。"),
                 ("② 利好出尽+解禁+降息下杀", "GENIUS 落地『利好出尽』；8 月起锁定期解禁抛压；美联储二轮降息压储备收益率；2026-02-05 触底 $49.90。"),
                 ("③ 超跌反弹", "2026-03 反弹至 $126。"),
                 ("④ Q1 放缓+OUSD 威胁再下台阶", "Q1'26 营收增速降至 +20%、净利 −15%；6-30 OUSD 联盟直击模式，跌破 $70。")],
    }, title="Circle 月度股价走势（USD，2025-06 IPO 至今）")
    px["yhigh"].update(PX_HIGH); px["ylow"].update(PX_LOW)

    # ═══ 4 卖方研报共识 ═══
    K.write_consensus(wb.create_sheet(S_CONS), {
        "title": "卖方研报共识 — 2026 Q2-Q3 逐家对账（评级重心由 Buy 滑向 Hold）",
        "overview": "覆盖约 23-24 家，均值目标价 $114-137（区间极宽 $55-243），但均值被未更新的旧多头评级抬高；近三个月活跃更新以 Neutral/Hold 为主（Goldman $96、Mizuho $85、Susquehanna $69、Compass $55）。多头代表 Bernstein $190。核心分歧 = 利率路径 × USDC 流通量 × Coinbase 分成三变量。",
        "assumptions": [
            ("USDC 流通量\n(2026-2027)", "2026 中约 $73-77B（Q1'26 季末 $77B）；Circle 目标 2026 下半年推向 $150B；街上多数给双位数-高双位数增长。",
             "GENIUS Act 采用红利 vs 银行/科技巨头自发稳定币 + OUSD 分流。", "Base 2026E 平均 $80B、2027E $118B（GENIUS 采用增长 × USDC 份额约 23-25% 温和提升）。"),
            ("储备收益率\n(利率路径)", "EFFR 已从 5.33% 降至 3.63%（-170bps）；期货隐含继续小幅下行。",
             "降息幅度与节奏 = 收入端最大宏观变量（96% 的收入受短端利率影响）。", "Base 2026E 3.7%、2027E 3.3%，随美联储路径缓降后企稳 3.0%。"),
            ("Coinbase 分销分成", "FY2025 分销成本占储备收入 63%（$1.66B），付 Coinbase 约 $0.9B（占营收 54%）；协议据报 2026-08 续约。",
             "续约条款 + off-platform 份额能否上升 = 留存率关键敏感项。", "Base 分成比 62%→58%(2030)，假设 Circle 逐步分散分销、续约中性偏温和改善。"),
            ("目标 P/E / EPS 口径", "共识 FY2026E EPS $0.82(GAAP)/$1.24(调整)；FY2027E $1.61。TP 均值 $114-137。",
             "GAAP（含正常化 SBC）vs 调整后差异大，直接决定目标价。", "本模型用经调整 EPS（剔一次性 IPO SBC）；Base 2027E 目标 28x。"),
        ],
        "divergences": [
            "① Coinbase 分成 2026-08 续约：条款恶化则留存率进一步下压，是最大近期下行催化。",
            "② OUSD/Open Standard 联盟（2026-06-30，140+ 家含 Stripe/Coinbase/Visa/BlackRock）欲把储备利息分给采用方——对『发行方独享储备利息』模式的正面攻击。",
            "③ 利率路径：降息越深，储备收益率压缩越大，且无法靠流通量增长完全对冲。",
        ],
        "stances": [
            "Bernstein（2026-07-01）：Buy，TP $190（多头最高之一）。",
            "Clear Street（2026-07-01）：Buy，TP $157。",
            "HC Wainwright（2026-05-18）：Buy，TP $150。",
            "Morgan Stanley（2026-05-19）：Equal Weight，TP $106。",
            "Goldman Sachs（2026-07-02）：Neutral，TP $96（下调）。",
            "Mizuho（2026-06-05）：Neutral，TP $85。",
            "JPMorgan（首覆 2025-06-30，维持）：Underweight，TP $80。",
            "Susquehanna（2026-07-01 首覆）：Neutral，TP $69。",
            "Compass Point（2026-06-30）：Neutral，TP $55（全场最低）。",
        ],
    })

    # ═══ 5 历史估值倍数 ═══
    hm = K.write_hist_multiples(wb.create_sheet(S_HMULT), {
        "title": "历史估值倍数 — 上市仅 13 个月，历史带薄；以 forward P/E + 同业为主",
        "intro": "Circle 2025-06 IPO，公开交易仅约 13 个月，自身历史 P/E 带极薄（且早期含泡沫，峰值曾对应数百倍）。主线以经调整 EPS × forward P/E；下面同业为金融科技/支付/资产管理光谱。P/B 仅体检（资产轻、储备过手不进权益，P/B 无意义）。",
        "s_hist": S_HIST, "ha": ha, "hist_cols": HC, "hist_years": HY,
        "yhigh": px["yhigh"], "ylow": px["ylow"],
        "fwd_note": "当下 forward P/E：现价 $68.65 ÷ 模型 2026E EPS≈$1.67→约 41x；÷2027E≈$2.46→约 28x（Base）。见『情景估值』。",
        "self_name": "Circle", "self_fwd_pe_label": "~41x 2026E / ~28x 2027E(Base)",
        "self_note": "P/E 主线：穿越周期的耐用资产是可持续的储备盈利能力（流通量×收益率×留存），不是账面净资产。",
        "peers": [
            {"name": "Coinbase(COIN)", "yearly": None, "cur_pb": None, "cur_pe": 30.0, "fwd_pe": 24.0,
             "note": "USDC 分成对手方；加密交易所，盈利同样利率/交易量敏感，倍数区间参照。"},
            {"name": "PayPal(PYPL)", "yearly": None, "cur_pb": None, "cur_pe": 15.0, "fwd_pe": 13.0,
             "note": "成熟支付，低增长低倍数；PYUSD 稳定币同业，作光谱下沿。"},
            {"name": "Visa/Mastercard", "yearly": None, "cur_pb": None, "cur_pe": 32.0, "fwd_pe": 28.0,
             "note": "支付网络龙头，高质量高倍数；Circle 若从利差转向支付网络变现可上探此区间。"},
            {"name": "资产管理(BLK 等)", "yearly": None, "cur_pb": None, "cur_pe": 22.0, "fwd_pe": 20.0,
             "note": "AUM×费率模式与浮存×收益率有相似性；参照中枢。"},
        ],
        "ratio": None,
        "reading": "Circle 无足够自身历史带，估值锚定：① forward P/E（现价 28x 2027E Base）；② 同业光谱（支付 13-28x、交易所 24x、资管 20x）；③ 增长溢价——若 2027E 经调整 EPS 增速仍 40%+ 可给 28-32x，若利率+分成压制到 15-20% 增速则应回落到 18-22x。质量上沿锚 30x（单一收入引擎 + 利率敏感 + 竞争威胁，不给纯成长股 35x+）。",
    })

    # ═══ 6 估值倍数假设（P/E 主线三层）═══
    ma = K.write_multiple_assumptions(wb.create_sheet(S_MULT), {
        "title": "估值倍数假设 — 主线 P/E 三层拆解（质量锚 30x × 结构溢价 1.0 × 情景情绪值）",
        "intro": "目标 P/E = 质量上沿锚 30x × 结构溢价 1.0（已含在锚内避免双算）× 情景情绪值。三案情绪值在此拍，『情景切换』引用切换，『情景估值』套用当前案。",
        "why_text": ("Circle 穿越周期的耐用资产是『可持续的储备盈利能力』——USDC 流通量 × 收益率 × 留存率，"
                     "不是账面净资产（储备是过手负债、不进权益，P/B 主线无意义）。因此主分母选经调整净利/EPS，P/B 仅体检。"
                     "★ 关键风险决定倍数上限：单一收入引擎（96% 储备利息）+ 利率高度敏感 + Coinbase 分成 62% + OUSD 联盟威胁 → "
                     "不给纯成长股 35x+，质量上沿锚 30x；若竞争/利率证实压制增速，镜头应向 18-22x 下移。"),
        "why_rows": 5,
        "method_text": "目标 P/E = 30x（质量上沿锚，参照支付网络 28x + 增长溢价）× 1.0（结构溢价并入锚）× 情景情绪值。Bear 回到 14-22x（利率崩+份额失+分成升+倍数压缩）；Base 2026-27 维持 28-30x 后随增速成熟回落到 23x；Bull 在流通量爆发+利率高位+分成改善下到 33-35x。",
        "peak": PEAK_PE, "peak_note": "质量上沿锚 30x：参照 Visa/MA 28x forward + 高增长溢价；非上市初期泡沫峰值（曾数百倍）。",
        "premium": PREMIUM, "premium_note": "结构溢价并入 30x 质量锚，避免双算；三案差异放情绪/增长可见度层。",
        "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
        "s_hist": S_HIST, "hpb_row": ha["HPE"],
        "cases": [
            ("Bear", PE_SENT["Bear"], "利率深降 + USDC 份额被 OUSD/银行分流 + Coinbase 分成升 + 大盘倍数压缩 → 14-22x。"),
            ("Base", PE_SENT["Base"], "流通量按 GENIUS 采用增长、利率企稳、分成温和改善 → 28-30x，随增速成熟回落到 23x。"),
            ("Bull", PE_SENT["Bull"], "USDC 破 $150B+、利率高位、分销分散 + Arc/CPN 变现 → 市场给支付网络级 33-35x。"),
        ],
        "reconcile_text": "卖方 TP 均值 $114-137（被旧多头抬高），近三月活跃更新 TP $55-106 中性居多。本模型 Base 2027E 28x×经调整 EPS $2.46≈$69≈现价——刻意不追高：单一引擎 + 利率/分成/竞争三重风险下，现价已隐含较充分的乐观预期。",
        "source_text": "同业倍数来自统一 API/公开；自身 forward 用模型前瞻 EPS；三层目标倍数与情景触发在本页与情景切换页。",
    })

    # ═══ 7 情景切换 ═══
    sw = K.write_scenario_switch(wb.create_sheet(S_SW), {
        "title": "情景切换 — 全模型唯一情景参数库（默认 Base）",
        "usage": "B2 是唯一开关。USDC 流通量增速、储备收益率、分销分成比、经调整经营利润率、净利转换、目标 P/E 都按当前案联动；『估值对比』直接引三案矩阵不被开关污染。其他收入三案共用 Base（非 load-bearing，占营收<5%）。",
        "cases": CASES, "default": "Base",
        "triggers": [
            ("Bear", "美联储深度降息（收益率<2.5%）；USDC 份额被 OUSD 联盟/银行稳定币分流、流通量停滞；Coinbase 2026-08 续约恶化、分成升至 65%+；目标倍数压到 18x 以下。"),
            ("Base", "USDC 沿 GENIUS Act 采用红利温和增长（份额 23-25%）；利率缓降后企稳 3%；分成温和改善至 58-60%；目标倍数 26-30x。"),
            ("Bull", "GENIUS 采用爆发，USDC 破 $150B、份额抢占；利率高位企稳 3.5%+；分销分散 + Arc/CPN 变现放量；市场按支付网络给 33-35x。"),
        ],
        "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
        "levers": [
            {"key": "circ", "name": "USDC 流通量增速", "fmt": K.PCT, "cols": FCf,
             "vals": CIRC_G, "desc": "物理锚：储备收入都挂它。作用于 2025A 平均流通量 $58B。改这行 → 流通量 → 储备收入 → 净利 → 隐含价全链动。",
             "stories": {"Bear": "GENIUS 采用不及预期 + 被银行/OUSD 分流，增长停滞。", "Base": "GENIUS 采用红利 + 份额温和提升。", "Bull": "机构采用爆发，Circle 抢占合规稳定币主场。"},
             "hist": hg(AVGCIRC)},
            {"key": "yld", "name": "储备收益率", "fmt": K.PCT, "cols": FCf,
             "vals": YLD, "desc": "第二物理量（外生于公司，由美联储利率决定）。储备收入 = 流通量 × 收益率。EFFR 当前 3.63% 且下行。",
             "stories": {"Bear": "美联储深降至 2.25%。", "Base": "缓降后企稳 3.0%。", "Bull": "利率高位企稳 3.5%。"},
             "hist": YIELD_H},
            {"key": "dist", "name": "分销分成比", "fmt": K.PCT, "cols": FCf,
             "vals": DISTR, "desc": "付 Coinbase 等的储备收入分成（关键敏感项）。FY2025 63%；协议 2026-08 续约。off-platform 份额越高分成越低。",
             "stories": {"Bear": "续约恶化 + OUSD 逼高返利，升至 66%。", "Base": "温和改善至 58%（分散分销）。", "Bull": "续约有利 + 分销多元化，降至 52%。"},
             "hist": DISTR_H},
            {"key": "opm", "name": "经调整经营利润率(on 净营业收入)", "fmt": K.PCT, "cols": FCf,
             "vals": OPM, "desc": "净营业收入（扣分销后）→ 经调整营业利润的利润率。反映 opex（人力/合规/Arc/CPN 投入）杠杆。",
             "stories": {"Bear": "OUSD 竞争 + 合规投入拖累，margin 卡在 32%。", "Base": "规模效应缓释，margin 升到 47%。", "Bull": "经营杠杆充分 + 新业务变现，margin 到 52%。"},
             "hist": OPM_H},
            {"key": "netconv", "name": "净利转换率", "fmt": K.PCT, "cols": FCf,
             "vals": NETCV, "desc": "经调整营业利润 → 经调整净利（税 ~21% 被 $1.5B 公司现金的净利息收入部分抵消）。",
             "stories": {"Bear": "税率/投入不利。", "Base": "约 0.80。", "Bull": "净利息收入+税优支撑 0.82。"},
             "hist": NETCV_H},
        ],
        "linked": [
            {"key": "sent", "name": "情绪/增长可见度值（P/E 第三层）", "fmt": K.N2,
             "src_sheet": S_MULT, "src_row0": ma["sent_row0"],
             "note": "来自估值倍数假设页。目标 P/E = 30x × 1.0 × 本行。"},
        ],
    })
    pk = f"'{S_MULT}'!{ma['pk_cell']}"; pr = f"'{S_MULT}'!{ma['pr_cell']}"
    swpe = sw["next_row"]
    K.lab(wb[S_SW], f"A{swpe}", "目标P/E（当前案）", b=True)
    for col in HC:
        K.lab(wb[S_SW], f"{col}{swpe}", "n.m.", note=True)  # 上市前无 P/E
    for col in FCf:
        K.fml(wb[S_SW], f"{col}{swpe}", f"={pk}*{pr}*{col}{sw['SWACT']['sent']}", K.MX, link=True)
    K.logic(wb[S_SW], f"L{swpe}", "目标 P/E = 质量锚 30x × 结构溢价 1.0 × 当前案情绪值；喂『情景估值』。")

    # ═══ 8 物理锚：USDC 平均流通量 ═══
    anchor = K.write_anchor(wb.create_sheet(S_ANCHOR), {
        "title": "USDC 平均流通量（$B）— 需求物理盘子",
        "all_cols": ALLC, "all_years": ALLY,
        "series": [("USDC 平均流通量 ($B)", AVGCIRC + [None] * 5,
                    "历史按储备收入/隐含收益率校准（2025 年末 $75.3B、Q1'26 季末 $77B、当前约 $73B）；前瞻 = 上年 × 当前案增速。", K.N0)],
        "yoy_row": "USDC 平均流通量 ($B)",
        "source_note": "口径 = 期间平均在外流通量。来源：Circle 10-K（年末 $75.3B，+72%）、Q1'26 10-Q（季末 $77B）、DefiLlama/circle.com（当前约 $73B）。稳定币总市值约 $311B，USDC 份额约 23.5%（USDT 59%）。",
        "role_note": "储备收入 = USDC 平均流通量 × 储备收益率。改这行会穿透到储备收入、净营业收入、经调整净利、目标价（连通性测试入口）。",
    })
    circ_row = anchor["row_of"]["USDC 平均流通量 ($B)"]
    for idx, col in enumerate(FCf):
        prev = FC[idx]  # FC=[F,G,H,I,J,K]; FCf=[G..K]; prev of G is F(2025A)
        K.fml(wb[S_ANCHOR], f"{col}{circ_row}", f"={prev}{circ_row}*(1+{K.R(S_SW, col + str(sw['SWACT']['circ']))})", K.N0, link=True)

    # ═══ 9 分部测算（储备收入链 → 净营业收入）═══
    seg = K.write_segment_model(wb.create_sheet(S_SEG), {
        "title": "分部测算 — 储备收入 = 流通量×收益率 −分销 +其他 = 净营业收入",
        "all_cols": ALLC, "all_years": ALLY, "logic_col": "N",
        "groups": [
            ("USDC 流通量物理锚", [
                ("USDC 平均流通量 ($B)", None, K.N0, "引自『USDC流通量物理锚』。改流通量，储备收入跟着动。"),
            ]),
            ("储备收入 = 流通量 × 收益率", [
                ("储备收益率 (%)", None, K.PCT, "历史 = 储备收入 ÷ 平均流通量（反推）；前瞻 = 『情景切换』当前案（跟美联储利率）。"),
                ("储备收入 ($B)", None, K.N1, "历史取实数；前瞻 = 流通量 × 收益率。"),
            ]),
            ("分销与净营业收入", [
                ("分销分成比 (%)", None, K.PCT, "历史 = 分销成本 ÷ 储备收入（反推）；前瞻 = 『情景切换』当前案。"),
                ("分销与交易成本 ($B)", None, K.N1, "= 储备收入 × 分销分成比（付 Coinbase 等）。"),
                ("其他收入 ($B)", None, K.N1, "订阅/服务；历史取实数，前瞻小体量外推（三案共用）。"),
                ("总营收 ($B)", None, K.N1, "= 储备收入 + 其他收入。"),
                ("净营业收入·扣分销 ($B)", None, K.N1, "= 储备收入 − 分销成本 + 其他收入 = 留给 Circle 的经营收入，喂利润表。"),
            ]),
        ],
    })
    m = seg["m"]
    RES_H = ha["seg_rows"]["储备收入"]; OTH_H = ha["seg_rows"]["其他收入"]
    # 流通量：全列引锚
    for col in ALLC:
        K.fml(wb[S_SEG], f"{col}{m['USDC 平均流通量 ($B)']}", f"={K.R(S_ANCHOR, col + str(circ_row))}", K.N0, link=True)
    # 历史：储备收入取实数、收益率反推、分销取实数、分成比反推
    for col in HC:
        K.fml(wb[S_SEG], f"{col}{m['储备收入 ($B)']}", f"={K.R(S_HIST, col + str(RES_H))}", K.N1, link=True)
        K.fml(wb[S_SEG], f"{col}{m['储备收益率 (%)']}", f"={col}{m['储备收入 ($B)']}/{col}{m['USDC 平均流通量 ($B)']}", K.PCT)
        K.fml(wb[S_SEG], f"{col}{m['其他收入 ($B)']}", f"={K.R(S_HIST, col + str(OTH_H))}", K.N1, link=True)
    K.introw(wb[S_SEG], m["分销与交易成本 ($B)"], HC, DIST, None, K.N1)
    for col in HC:
        K.fml(wb[S_SEG], f"{col}{m['分销分成比 (%)']}", f"={col}{m['分销与交易成本 ($B)']}/{col}{m['储备收入 ($B)']}", K.PCT)
    # 前瞻：收益率/分成比引情景；储备收入=流通量×收益率；分销=储备×分成；其他=外推
    for idx, col in enumerate(FCf):
        K.fml(wb[S_SEG], f"{col}{m['储备收益率 (%)']}", f"={K.R(S_SW, col + str(sw['SWACT']['yld']))}", K.PCT, link=True)
        K.fml(wb[S_SEG], f"{col}{m['储备收入 ($B)']}", f"={col}{m['USDC 平均流通量 ($B)']}*{col}{m['储备收益率 (%)']}", K.N1)
        K.fml(wb[S_SEG], f"{col}{m['分销分成比 (%)']}", f"={K.R(S_SW, col + str(sw['SWACT']['dist']))}", K.PCT, link=True)
        K.fml(wb[S_SEG], f"{col}{m['分销与交易成本 ($B)']}", f"={col}{m['储备收入 ($B)']}*{col}{m['分销分成比 (%)']}", K.N1)
        K.inp(wb[S_SEG], f"{col}{m['其他收入 ($B)']}", OTHER_FWD[idx], None, K.N1)
    # 总营收 & 净营业收入（全列）
    for col in ALLC:
        K.fml(wb[S_SEG], f"{col}{m['总营收 ($B)']}", f"={col}{m['储备收入 ($B)']}+{col}{m['其他收入 ($B)']}", K.N1)
        K.fml(wb[S_SEG], f"{col}{m['净营业收入·扣分销 ($B)']}",
              f"={col}{m['储备收入 ($B)']}-{col}{m['分销与交易成本 ($B)']}+{col}{m['其他收入 ($B)']}", K.N1)
    for col in FCf:
        wb[S_SEG][f"{col}{m['净营业收入·扣分销 ($B)']}"].fill = K.OUT

    # ═══ 10 利润与收入假设（净营业收入 → 营业利润 → 净利 → EPS/BPS）═══
    fw = wb.create_sheet(S_FUND)
    K.hdr(fw, 1, "利润与收入假设 — 净营业收入(扣分销) → 经调整营业利润 → 净利 → EPS/BPS", 12)
    fw["A2"] = "假设/口径"; fw["A2"].font = K.BF
    for col, y in zip(ALLC, ALLY):
        fw[f"{col}2"] = y; fw[f"{col}2"].font = K.BF; fw[f"{col}2"].fill = K.CH
    fw["N2"] = "逻辑/来源"; fw["N2"].font = K.BF; fw["N2"].fill = K.CH
    r = 3
    fund = {}
    K.band(fw, r, "利润率假设（历史实际 + 前瞻链情景切换）", 12); r += 1
    # 经调整经营利润率
    fund["OPM"] = r
    K.lab(fw, f"A{r}", "经调整经营利润率(on 净营业收入)")
    K.introw(fw, r, HC, OPM_H, None, K.PCT); K.lab(fw, f"B{r}", "n.m.", note=True)
    for col in FCf:
        K.fml(fw, f"{col}{r}", f"={K.R(S_SW, col + str(sw['SWACT']['opm']))}", K.PCT, link=True)
    K.logic(fw, f"N{r}", "历史 = 经调整营业利润 ÷ 净营业收入（反推，2025A 46%）；前瞻 = 情景切换当前案。含 opex（人力/合规/Arc/CPN）杠杆。")
    r += 1
    # 净利转换率
    fund["NETCV"] = r
    K.lab(fw, f"A{r}", "净利转换率(净利/营业利润)")
    K.introw(fw, r, HC, NETCV_H, None, K.PCT); K.lab(fw, f"B{r}", "n.m.", note=True)
    for col in FCf:
        K.fml(fw, f"{col}{r}", f"={K.R(S_SW, col + str(sw['SWACT']['netconv']))}", K.PCT, link=True)
    K.logic(fw, f"N{r}", "税 ~21%，被 $1.5B 公司现金净利息收入部分抵消，≈0.80。历史近似。")
    r += 1
    # 留存率
    fund["RET"] = r
    K.lab(fw, f"A{r}", "留存率")
    K.introw(fw, r, ALLC, [None, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0], None, K.PCT); K.lab(fw, f"B{r}", "n.m.", note=True)
    K.logic(fw, f"N{r}", "无分红，全额留存；仅用于 BPS/ROE 体检，主估值不依赖 P/B。")
    r += 1

    K.band(fw, r, "收入 → 利润 → 每股（链）", 12); r += 1
    # 净营业收入（引分部测算）
    fund["SUB"] = r
    K.lab(fw, f"A{r}", "净营业收入·扣分销 ($B)")
    for col in ALLC:
        K.fml(fw, f"{col}{r}", f"={K.R(S_SEG, col + str(m['净营业收入·扣分销 ($B)']))}", K.N1, link=True)
    K.logic(fw, f"N{r}", "引自『分部测算』= 储备收入 − 分销 + 其他。")
    r += 1
    # 总营收（引分部测算，展示）
    fund["REV"] = r
    K.lab(fw, f"A{r}", "总营收 ($B)", b=True)
    for col in ALLC:
        K.fml(fw, f"{col}{r}", f"={K.R(S_SEG, col + str(m['总营收 ($B)']))}", K.N1, link=True)
    K.lab(fw, f"A{r+1}", "  YoY", note=True)
    for i in range(1, len(ALLC)):
        K.fml(fw, f"{ALLC[i]}{r+1}", f'=IFERROR({ALLC[i]}{r}/{ALLC[i-1]}{r}-1,"n.m.")', K.PCT)
    r += 2
    # 经调整营业利润
    fund["OP"] = r
    K.lab(fw, f"A{r}", "经调整营业利润 ($B)")
    K.lab(fw, f"B{r}", "n.m.", note=True)  # 2021A 利润率 n.m.
    for col in ALLC[1:]:
        K.fml(fw, f"{col}{r}", f"={col}{fund['SUB']}*{col}{fund['OPM']}", K.N1)
    r += 1
    # 经调整净利：历史引 history adj NI，前瞻 = OP × netconv
    fund["NI"] = r
    K.lab(fw, f"A{r}", "经调整净利润 ($B)", b=True); fw[f"A{r}"].fill = K.OUT
    for col in HC:
        K.fml(fw, f"{col}{r}", f"={K.R(S_HIST, col + str(ha['HNI']))}", K.N1, link=True)
    for col in FCf:
        K.fml(fw, f"{col}{r}", f"={col}{fund['OP']}*{col}{fund['NETCV']}", K.N1)
    K.lab(fw, f"A{r+1}", "  净利率(/总营收)", note=True)
    for col in ALLC:
        K.fml(fw, f"{ALLC[ALLC.index(col)]}{r+1}", f"={col}{fund['NI']}/{col}{fund['REV']}", K.PCT)
    r += 2
    # EPS
    fund["EPS"] = r
    K.lab(fw, f"A{r}", "经调整 EPS ($)", b=True); fw[f"A{r}"].fill = K.OUT
    for col in HC:
        K.fml(fw, f"{col}{r}", f"={K.R(S_HIST, col + str(ha['HEPS']))}", K.N2, link=True)
    for col in FCf:
        K.fml(fw, f"{col}{r}", f"={col}{fund['NI']}*1000/{SHARES_M}", K.N2)
    K.logic(fw, f"N{r}", f"前瞻用全类别摊薄股本 {SHARES_M}M（IPO 后基本稳定）；历史引『历史财务与估值』经调整 EPS。")
    r += 1
    # 期末权益
    fund["EQ"] = r
    K.lab(fw, f"A{r}", "期末权益 ($B)")
    for col in HC:
        K.fml(fw, f"{col}{r}", f"={K.R(S_HIST, col + str(ha['HEQ']))}", K.N1, link=True)
    prevs = [HC[-1]] + list(FCf[:-1])
    for p, col in zip(prevs, FCf):
        K.fml(fw, f"{col}{r}", f"={p}{r}+{col}{fund['NI']}*{col}{fund['RET']}", K.N1)
    r += 1
    # BPS
    fund["BPS"] = r
    K.lab(fw, f"A{r}", "BPS ($)")
    for col in HC:
        K.fml(fw, f"{col}{r}", f"={K.R(S_HIST, col + str(ha['HBPS']))}", K.N2, link=True)
    for col in FCf:
        K.fml(fw, f"{col}{r}", f"={col}{fund['EQ']}*1000/{SHARES_M}", K.N2)
    r += 1
    # ROE
    fund["ROE"] = r
    K.lab(fw, f"A{r}", "ROE", note=True)
    for i, col in enumerate(ALLC):
        f = (f"={col}{fund['NI']}/{col}{fund['EQ']}" if i == 0
             else f"={col}{fund['NI']}/AVERAGE({ALLC[i-1]}{fund['EQ']},{col}{fund['EQ']})")
        K.fml(fw, f"{col}{r}", f, K.PCT)
    r += 1
    K.band(fw, r, "口径说明", 12); r += 1
    K.mtext(fw, r, "前瞻净利 = 净营业收入(扣分销) × 经调整经营利润率 × 净利转换率，避免用外部卖方净利截断物理锚链。历史净利 = 经调整口径（剔一次性 IPO SBC）。下游『情景估值』直接引本表每股，不重算。", "K", 3)
    K.set_widths(fw, 26, ALLC, 8, logic_col="N", logic_width=58)

    # ═══ 11 情景估值（P/E 主线 + DCF/PS 交叉）═══
    write_crcl_valuation(wb.create_sheet(S_VAL), {
        "s_sw": S_SW, "sw_cell": sw["sw_cell"], "s_hist": S_HIST, "ha": ha,
        "s_fund": S_FUND, "fund": fund, "s_seg": S_SEG, "m": m, "s_anchor": S_ANCHOR,
        "circ_row": circ_row, "target_row": swpe,
    })

    # ═══ 12 估值对比（三案并排，自定义 lambda）═══
    swb = sw["SWB"]
    def cr(key, ci): return swb[key] + ci
    def prevcol(col): return ALLC[ALLC.index(col) - 1]
    rows = [
        {"key": "circ", "label": "USDC 平均流通量 ($B)", "fmt": K.N0,
         "hist": lambda c, ci, a: f"={K.R(S_ANCHOR, c + str(circ_row))}",
         "fwd": lambda c, j, ci, a: f"={prevcol(c)}{a['circ']}*(1+{K.R(S_SW, c + str(cr('circ', ci)))})"},
        {"key": "yld", "label": "储备收益率 (%)", "fmt": K.PCT,
         "hist": lambda c, ci, a: f"={K.R(S_HIST, c + str(ha['seg_rows']['储备收入']))}/{c}{a['circ']}",
         "fwd": lambda c, j, ci, a: f"={K.R(S_SW, c + str(cr('yld', ci)))}"},
        {"key": "res", "label": "储备收入 ($B)", "fmt": K.N1,
         "hist": lambda c, ci, a: f"={K.R(S_HIST, c + str(ha['seg_rows']['储备收入']))}",
         "fwd": lambda c, j, ci, a: f"={c}{a['circ']}*{c}{a['yld']}"},
        {"key": "dist", "label": "分销分成比 (%)", "fmt": K.PCT,
         "hist": lambda c, ci, a: f"={K.R(S_SEG, c + str(m['分销分成比 (%)']))}",
         "fwd": lambda c, j, ci, a: f"={K.R(S_SW, c + str(cr('dist', ci)))}"},
        {"key": "sub", "label": "净营业收入·扣分销 ($B)", "fmt": K.N1, "bold": True,
         "hist": lambda c, ci, a: f"={K.R(S_SEG, c + str(m['净营业收入·扣分销 ($B)']))}",
         "fwd": lambda c, j, ci, a: f"={c}{a['res']}*(1-{c}{a['dist']})+{K.R(S_SEG, c + str(m['其他收入 ($B)']))}"},
        {"key": "ni", "label": "经调整净利 ($B)", "fmt": K.N1, "bold": True,
         "hist": lambda c, ci, a: f"={K.R(S_HIST, c + str(ha['HNI']))}",
         "fwd": lambda c, j, ci, a: f"={c}{a['sub']}*{K.R(S_SW, c + str(cr('opm', ci)))}*{K.R(S_SW, c + str(cr('netconv', ci)))}"},
        {"key": "eps", "label": "经调整 EPS ($)", "fmt": K.N2,
         "hist": lambda c, ci, a: f"={K.R(S_HIST, c + str(ha['HEPS']))}",
         "fwd": lambda c, j, ci, a: f"={c}{a['ni']}*1000/{SHARES_M}"},
        {"key": "tpe", "label": "目标 P/E", "fmt": K.MX,
         "hist": lambda c, ci, a: '="n.m."',
         "fwd": lambda c, j, ci, a: f"={K.R(S_MULT, c + str(ma['target_row0'] + ci))}"},
        {"key": "px", "label": "隐含股价 ($)", "fmt": K.PX, "bold": True, "out": True,
         "hist": lambda c, ci, a: f"={K.R(S_HIST, c + str(ha['HPX']))}",
         "fwd": lambda c, j, ci, a: f"={c}{a['tpe']}*{c}{a['eps']}+0*{c}{a['circ']}"},
        {"key": "ipe", "label": "隐含 forward P/E（体检）", "fmt": K.MX,
         "hist": lambda c, ci, a: '="n.m."',
         "fwd": lambda c, j, ci, a: f"={c}{a['px']}/{c}{a['eps']}"},
        {"key": "up", "label": "前瞻 vs 现价", "fmt": K.PCT,
         "hist": lambda c, ci, a: None,
         "fwd": lambda c, j, ci, a: f"={c}{a['px']}/{PX_NOW}-1"},
    ]
    cmp = K.write_comparison(wb.create_sheet(S_CMP), {
        "title": "估值对比 — Bear / Base / Bull 三案并排（三个物理杠杆相乘 → 区间极宽）",
        "intro": "三案从 USDC 流通量增速、储备收益率、分销分成比、经营利润率、目标 P/E 同一条链推导。主判断年取 2027E（2026 基本可见，真正分歧在 2027+）。隐含价 = 目标 P/E × 经调整 EPS。三个物理杠杆相乘，故三案区间远宽于普通公司——这正是 CRCL 高波动的根源。",
        "case_names": CASES, "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
        "block_start": 22, "rows": rows,
        "summary": {
            "band": "2027E 三案摘要（主判断年）", "target_col": "H",
            "rows": [
                ("USDC 流通量($B)", "circ", K.N0, "物理锚。", False),
                ("储备收入($B)", "res", K.N1, "流通量 × 收益率。", False),
                ("净营业收入($B)", "sub", K.N1, "储备收入 −分销 +其他。", False),
                ("经调整净利($B)", "ni", K.N1, "净营业收入 × 经营利润率 × 净利转换。", False),
                ("经调整 EPS($)", "eps", K.N2, "主估值分母。", False),
                ("目标 P/E", "tpe", K.MX, "质量锚×溢价×情绪。", False),
                ("隐含股价($)", "px", K.PX, "P/E 主线输出。", True),
                ("vs 现价", "up", K.PCT, "现价 reality check。", True),
            ],
            "mcap": {"label": "隐含市值($B)", "key": "px", "expr": f"*{SHARES_M}/1000", "note": "隐含股价 × 股本。"},
            "concl": "Base 2027E 28x×EPS≈$2.46≈$69≈现价 → 现价已隐含较充分乐观。Bull 靠流通量爆发+利率高位+分成改善+倍数扩张给大幅上行；Bear 靠利率崩+份额失+分成升+倍数压缩给大幅下行。三杠杆相乘 → 杠铃形风险收益 → HOLD。",
        },
    })

    # ═══ 13 综合判断仪表盘 ═══
    K.write_dashboard(wb.create_sheet(S_DASH), {
        "title": "综合判断仪表盘 — 基本面拐点 / 估值错位 / 催化剂 / 情绪",
        "usage": "把模型压成投后跟踪语言：哪些指标验证 Base，哪些把模型推向 Bear/Bull。B 列公式直接引模型输出。",
        "blocks": [
            {"title": "A. 基本面拐点 — 单位经济学在改善还是恶化？", "rows": [
                ("分销分成比走势", "FY2022 39% → FY2025 63%", "关键敏感项：留存率。分成比若续升 = 增长替 Coinbase 打工。", True),
                ("储备收益率(利率)", "EFFR 5.33%→3.63%，仍下行", "96% 的收入受短端利率影响；降息直接压收入。"),
                ("2027E Base 经调整净利", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Base']['ni']))}", "fmt": K.N1, "fill": True}, "若不及该路径，Base 估值先降。"),
            ]},
            {"title": "B. 估值错位（预测引擎）", "rows": [
                ("当前股价", {"inp": PX_NOW, "fmt": K.PX, "fill": True}, "reality check，不反向拟合。"),
                ("Base 2027E 隐含价", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Base']['px']))}", "fmt": K.PX, "fill": True}, "主判断输出，≈现价。"),
                ("Base 2027E vs 现价", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Base']['up']))}", "fmt": K.PCT, "fill": True}, "接近 0 = 现价已 price-in Base。"),
                ("Bear 2027E 隐含价", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Bear']['px']))}", "fmt": K.PX}, "下行：利率崩+份额失+分成升+倍数压缩。"),
                ("Bull 2027E 隐含价", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Bull']['px']))}", "fmt": K.PX}, "上行：流通量爆发+利率高位+分成改善+倍数扩张。"),
            ]},
            {"title": "C. 催化剂", "rows": [
                ("Bear 触发", "2026-08 Coinbase 续约恶化；OUSD 联盟逼高返利；美联储深降至 2.5% 以下；USDC 份额被分流。", "先砍 dist/yld/circ，再把目标 P/E 压到 18x 以下。"),
                ("Bull 触发", "USDC 破 $150B；GENIUS 采用爆发抢占份额；利率高位企稳；Arc/CPN 变现放量。", "先抬 circ/yld，改善 dist，再抬目标 P/E 到 33-35x。"),
                ("最易错处", "把 GAAP 亏损（含一次性 IPO SBC）直接当经营亏损，或把上市初泡沫峰值倍数当锚。", "本模型用经调整 EPS + 30x 质量锚，避免两个陷阱。"),
            ]},
            {"title": "D. 综合判断", "rows": [
                ("一句话结论", "HOLD：77% 回撤挤掉泡沫，但现价仍需三个物理杠杆同时不恶化才撑得住；杠铃形风险收益。", "上行看流通量+利率+分成，下行看 Coinbase 续约+OUSD。", True),
            ]},
        ],
        "final": {"band": "最终判断",
                  "text": "Circle 的 alpha 问题：市场从『稳定币合规龙头、流通量无限增长』的叙事，被利率下行 + Coinbase 63% 分成 + OUSD 竞争三重挤压，重估回 $17B。本模型用经调整 EPS + 物理锚（流通量×收益率×留存），Base 2027E 隐含价≈现价 → 现价已隐含较充分的乐观预期。三个物理杠杆相乘使估值区间极宽（2027E Bear ~$19 / Bull ~$145），是名副其实的杠铃形高波动标的。评级 HOLD，等 2026-08 Coinbase 续约与 Q2 业绩确认方向。"},
        "tracking": {
            "intro": "投后跟踪按季报 + 月度链上流通量滚动更新。",
            "rows": [
                ("__band__", "一、物理锚（量）"),
                ("USDC 流通量", "约 $73B（2026-07）", "关键敏感项：储备收入盘子", "DefiLlama/circle.com 月度 + 季报", "连续下滑或停滞 → 下调 circ 增速转 Bear"),
                ("__band__", "二、利率（价）"),
                ("储备收益率/EFFR", "3.63%（2026-07）", "关键敏感项：96% 的收入受短端利率影响", "FRED EFFR + 季报收益率", "深降至 2.5% 以下 → 下调 yld 全链重算"),
                ("__band__", "三、分销分成（留存）"),
                ("Coinbase 分成比", "63%（FY2025），续约 2026-08", "关键敏感项：留存率", "季报 distribution costs / reserve income", "续约恶化升至 65%+ → 转 Bear"),
                ("__band__", "四、竞争威胁"),
                ("OUSD 联盟进展", "2026-06-30 启动，140+ 家", "关键敏感项：收益分给采用方冲击模式", "OUSD 采用量 + Circle 应对", "OUSD 上量 → 压 circ 份额 + 逼高 dist"),
            ],
        },
    })

    # ═══ input.json ═══
    payload = {
        "ticker": "CRCL", "company": "Circle Internet Group, Inc.",
        "built_at": "2026-07-07", "currency": "USD",
        "current_price": PX_NOW, "price_date": "2026-07-06", "market_cap_b": 17.06,
        "shares_m": SHARES_M, "rating": "HOLD",
        "target_price_note": "12 个月目标价 ≈ Base 2027E 隐含价（≈现价），见 xlsx『估值对比』；三案 2027E Bear/Base/Bull ≈ $19/$69/$145。",
        "method": "P/E 主线（经调整 EPS × 目标 P/E）+ DCF/PS 交叉；物理锚 = float-yield：USDC 平均流通量 × 储备收益率 − 分销 + 其他 → 经调整净利。",
        "adj_earnings_note": "经调整净利剔除一次性 IPO 股份支付（FY2025 SBC $566M），含正常化经营 SBC；GAAP FY2025 净亏 $(69.5)M。",
        "anchor": {"type": "float-yield", "avg_circulation_b": AVGCIRC,
                   "reserve_yield": YIELD_H, "distribution_ratio": DISTR_H},
        "historical_b": {"reserve_income": RESERVE, "other_revenue": OTHER, "total_revenue": [round(RESERVE[i]+OTHER[i],3) for i in range(5)],
                         "distribution_cost": DIST, "net_op_revenue": SUBTOT, "adj_net_income": ADJNI,
                         "gaap_net_income": GAAPNI, "equity": EQUITY, "adj_ebitda_2023_2025": [0.395, 0.285, 0.582]},
        "scenario": {"target_pe": TARGET_PE, "peak_pe": PEAK_PE, "pe_sentiment": PE_SENT,
                     "circulation_growth": CIRC_G, "reserve_yield": YLD, "distribution_ratio": DISTR,
                     "op_margin": OPM, "net_conversion": NETCV, "other_revenue_fwd": OTHER_FWD},
        "consensus": {"tp_mean_range": [114, 137], "tp_range": [55, 243], "rating_drift": "Buy→Hold",
                      "fy2026e_rev_b": 3.08, "fy2026e_eps_gaap": 0.82, "fy2026e_eps_adj": 1.24,
                      "fy2027e_rev_b": 4.30, "fy2027e_eps": 1.61, "n_analysts": 23,
                      "sellside_tp": {"Bernstein": 190, "ClearStreet": 157, "HCWainwright": 150,
                                      "MorganStanley": 106, "Goldman": 96, "Mizuho": 85, "JPMorgan": 80,
                                      "Susquehanna": 69, "CompassPoint": 55}},
        "sources": {
            "official": ["SEC EDGAR CIK 0001876042: FY2025 10-K (2026-03-09) / 424B4 IPO (2025-06-05) / Q1'26 10-Q (2026-05-11)"],
            "market": "公开行情/K线快照(2026-07-06 close $68.65)+ Yahoo estimates",
            "macro": "FRED EFFR/FEDFUNDS 3.63% (2026-07)",
            "industry": "DefiLlama stablecoin $311B / USDC $73B; GENIUS Act (P.L.119-27, 2025-07-18); OUSD coalition (2026-06-30)",
        },
    }
    os.makedirs(VAULT, exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    K.finalize(wb, freeze={S_FUND: "B3"})
    wb.save(OUT_XLSX)
    print("saved:", OUT_XLSX)
    print("sheets:", wb.sheetnames)
    return cmp


def write_crcl_valuation(ws, d):
    """P/E 主线逐年隐含价 + DCF + P/S 交叉验证。"""
    K = __import__("build_kit")
    s_sw, s_hist, ha = d["s_sw"], d["s_hist"], d["ha"]
    s_fund, fund = d["s_fund"], d["fund"]
    s_seg, m, s_anchor = d["s_seg"], d["m"], d["s_anchor"]
    circ_row, target_row = d["circ_row"], d["target_row"]
    import build_kit as K
    K.hdr(ws, 1, "情景估值 — 当前案逐年 P/E 主线 + DCF/PS 交叉验证", 11)
    K.lab(ws, "L1", "当前情景→", note=True); K.fml(ws, "M1", f"={K.R(s_sw, d['sw_cell'])}", K.N0, link=True); ws["M1"].fill = K.CUR
    r = K.mtext(ws, 2, "本页随情景切换 B2 变化。前瞻列 = 当前案目标 P/E × 经调整 EPS（判断，非拟合现价）；DCF + P/S 交叉。经调整 EPS 已剔一次性 IPO SBC。历史仅 2025A 有市价。", "K", 2)
    K.lab(ws, f"A{r}", "($/股；倍数；$B)", b=True)
    for col, y in zip(ALLC, ALLY):
        ws[f"{col}{r}"] = y; ws[f"{col}{r}"].font = K.BF; ws[f"{col}{r}"].fill = K.CH
    r += 1
    eps = lambda c: K.R(s_fund, c + str(fund["EPS"]))
    bps = lambda c: K.R(s_fund, c + str(fund["BPS"]))

    K.band(ws, r, "P/E 主线：前瞻 = 目标 P/E × 经调整 EPS", 11); r += 1
    pe_row = r; K.lab(ws, f"A{r}", "目标 P/E（前瞻=当前案）")
    for c in FCf:
        K.fml(ws, f"{c}{r}", f"={K.R(s_sw, c + str(target_row))}", K.MX, link=True)
    r += 1
    px_row = r; K.lab(ws, f"A{r}", "隐含股价 P/E主线 ($)", b=True); ws[f"A{r}"].fill = K.OUT
    K.fml(ws, f"F{r}", f"={K.R(s_hist, 'F' + str(ha['HPX']))}", K.PX, link=True)  # 2025A 年末价
    for c in FCf:
        K.fml(ws, f"{c}{r}", f"={c}{pe_row}*{eps(c)}+0*{K.R(s_anchor, c + str(circ_row))}", K.PX, link=True)
    r += 1
    up_row = r; K.lab(ws, f"A{r}", "较现价上行/下行")
    for c in FCf:
        K.fml(ws, f"{c}{r}", f"={c}{px_row}/{PX_NOW}-1", K.PCT)
    r += 2

    K.band(ws, r, "交叉体检：隐含 P/B 与当下 forward P/E", 11); r += 1
    K.lab(ws, f"A{r}", "隐含 P/B（价/BPS，仅体检）")
    for c in FCf:
        K.fml(ws, f"{c}{r}", f"={c}{px_row}/{bps(c)}", K.MX)
    r += 1
    K.lab(ws, f"A{r}", "当下现价对应 forward P/E")
    for c in FCf:
        K.fml(ws, f"{c}{r}", f"={PX_NOW}/{eps(c)}", K.MX)
    r += 2

    # DCF —— 第二镜头
    K.band(ws, r, "镜头二 DCF（简化，归一化 FCF 折现，WACC 10%）", 11); r += 1
    K.lab(ws, f"A{r}", "2028E 经调整净利 ($B)")
    K.fml(ws, f"B{r}", f"={K.R(s_fund, 'H' + str(fund['NI']))}", K.N1, link=True); dcf_ni = r; r += 1
    K.lab(ws, f"A{r}", "稳态 FCF/净利 转换（资产轻）"); K.inp(ws, f"B{r}", 0.90, None, K.PCT); dcf_conv = r; r += 1
    K.lab(ws, f"A{r}", "归一化 FCF ($B)")
    K.fml(ws, f"B{r}", f"=B{dcf_ni}*B{dcf_conv}", K.N1); dcf_fcf = r; r += 1
    K.lab(ws, f"A{r}", "WACC / 永续增长"); K.inp(ws, f"B{r}", 0.10, None, K.PCT); K.inp(ws, f"C{r}", 0.04, None, K.PCT); dcf_w = r; r += 1
    K.lab(ws, f"A{r}", "永续法 EV ($B) = FCF×(1+g)/(WACC−g)")
    K.fml(ws, f"B{r}", f"=B{dcf_fcf}*(1+C{dcf_w})/(B{dcf_w}-C{dcf_w})", K.N0); dcf_ev = r; r += 1
    K.lab(ws, f"A{r}", "DCF 隐含股价 ($, +净现金/股~$6)", b=True); ws[f"A{r}"].fill = K.OUT
    K.fml(ws, f"B{r}", f"=B{dcf_ev}*1000/{SHARES_M}+6", K.PX); dcf_px = r; r += 1
    K.lab(ws, f"A{r}", "DCF vs 现价")
    K.fml(ws, f"B{r}", f"=B{dcf_px}/{PX_NOW}-1", K.PCT); r += 2

    # P/S —— 第三镜头
    K.band(ws, r, "镜头三 P/S（2027E 总营收 × 目标 P/S）", 11); r += 1
    K.lab(ws, f"A{r}", "2027E 总营收 ($B)")
    K.fml(ws, f"B{r}", f"={K.R(s_fund, 'H' + str(fund['REV']))}", K.N1, link=True); ps_rev = r; r += 1
    K.lab(ws, f"A{r}", "目标 P/S"); K.inp(ws, f"B{r}", 4.0, None, K.MX); ps_m = r; r += 1
    K.lab(ws, f"A{r}", "P/S 隐含市值 ($B)")
    K.fml(ws, f"B{r}", f"=B{ps_rev}*B{ps_m}", K.N0); ps_mc = r; r += 1
    K.lab(ws, f"A{r}", "P/S 隐含股价 ($)", b=True); ws[f"A{r}"].fill = K.OUT
    K.fml(ws, f"B{r}", f"=B{ps_mc}*1000/{SHARES_M}", K.PX); ps_px = r; r += 1
    K.lab(ws, f"A{r}", "P/S vs 现价")
    K.fml(ws, f"B{r}", f"=B{ps_px}/{PX_NOW}-1", K.PCT); r += 2

    K.band(ws, r, "三镜头三角 + 方法", 11); r += 1
    K.mtext(ws, r, "三镜头（P/E 主线 / DCF / P/S）在 Base 下应收敛：P/E 主线 2027E≈$69；DCF 用归一化 FCF、WACC 10%（高 β + 单一引擎 + 利率敏感）约 $55-65；P/S 2027E 4x 约 $67。三者 Base ~$55-69，围绕现价 → HOLD。分歧点：DCF 对 WACC/g 极敏感（利率敏感型），P/S 忽略了分销分成对留存的侵蚀（故 P/E 主线更可靠）。", "K", 4)
    K.set_widths(ws, 28, ALLC, 11)


if __name__ == "__main__":
    build()
