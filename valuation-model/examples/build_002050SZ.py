# -*- coding: utf-8 -*-
"""
build_002050SZ.py — 三花智控(002050.SZ / 02050.HK)物理锚估值模型。
公司类型: 多分部精密零部件成长制造 + 人形机器人执行器期权。
物理锚(三驱动): ① 制冷零部件=成长驱动(热泵/数据中心液冷/储能温控新场景)
               ② 汽车热管理=全球NEV销量 × 三花有效单车价值($/辆)
               ③ 机器人执行器=全球人形机器人出货(万台) × 三花有效单机价值($/台)
估值镜头: P/E 主线(资本化盈利, 有定价权的阀控 franchise) + SOTP 支线(核心 vs 机器人分部) + DCF 交叉验证。
量纲: 财务 USD $B / 每股·股价 CNY / FX=CNY/USD。$B = ¥亿 / FX / 10。
运行: cd examples && PYTHONUTF8=1 python build_002050SZ.py
校验: python ../scripts/recalc.py out/002050SZ_valuation_model.xlsx
      python ../scripts/validate_valuation.py out/002050SZ_valuation_model.xlsx
"""
import os
from openpyxl import Workbook
import build_kit as K

# ════════════ 0. 全局轴 ════════════
ALLC = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
ALLY = ["2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
HC, HY = ["B", "C", "D", "E", "F"], ["2021", "2022", "2023", "2024", "2025"]
FC = ["F", "G", "H", "I", "J", "K"]      # 含基年 F=2025A
FCf = FC[1:]                              # 纯前瞻 2026E-2030E
FX_FC = 7.15
FX_HIST = [6.45, 6.73, 7.08, 7.20, 7.17]
CASES = ["Bear", "Base", "Bull"]

S_COVER, S_HIST, S_PX, S_CONS = "封面", "历史财务与估值", "股价走势", "卖方研报共识"
S_HMULT, S_MULT, S_SW = "历史估值倍数", "估值倍数假设", "情景切换"
S_ANCHOR, S_SEG, S_FUND = "需求物理盘子", "分部测算", "利润与收入假设"
S_VAL, S_CMP, S_DASH = "情景估值", "估值对比", "综合判断仪表盘"

PX_NOW = 45.51    # 2026-07-07 A股现价
SH_NOW = 4208     # 完全摊薄总股本(mn), = 市值1915亿/45.51
MONTHLY = [("2020-12", 24.65), ("2021-06", 23.98), ("2021-12", 25.30),
           ("2022-03", 16.62), ("2022-06", 27.48), ("2022-12", 21.22),
           ("2023-06", 30.26), ("2023-12", 29.40),
           ("2024-06", 19.08), ("2024-12", 23.51),
           ("2025-03", 28.83), ("2025-06", 26.38), ("2025-09", 48.43), ("2025-12", 55.31),
           ("2026-03", 42.53), ("2026-06", 43.94), ("2026-07", 45.51)]

wb = Workbook()
wb.remove(wb["Sheet"])

# ════════════ 1. 封面 ════════════
K.write_cover(wb.create_sheet(S_COVER), {
    "title": "三花智控 (002050.SZ / 02050.HK) 物理锚估值模型",
    "meta": [
        ("报告日期", "2026-07-07"),
        ("数据截止", "2025 年报 + 2026Q1 + 实时行情(A股¥45.51) + 近6月卖方共识"),
        ("现价", "A股 ¥45.51 | 市值 ~¥1,915 亿 | PE(TTM) 46.9x / PB 5.9x；H股 HK$28.52(A/H溢价~75%)"),
        ("量纲口径", "财务 USD $B(过 $B 口径校验); 每股/股价 CNY; FX=CNY/USD(2025≈7.17, 前瞻 7.15); 完全摊薄股本 42.08 亿股。"),
        ("方法一句话", "三物理锚(制冷成长 + NEV销量×单车价值 + 人形出货×单机价值)→ 分部收入 → 段净利率 → 归母净利 → 目标P/E×前瞻EPS → 隐含价；SOTP + DCF 交叉验证。"),
    ],
    "takeaways": [
        ("① 当下估值位置", "¥45.51 = 46.9x TTM / 40x 2026E / 34x 2027E, 处自身历史高位(2022-24 forward 28-36x); 2025 起被机器人主题重估。"),
        ("② 核心引擎", "制冷零部件(2025 186亿, 四通阀/电子膨胀阀全球龙头)+ 汽车热管理(124亿, 特斯拉核心供应商)= 稳态高质量基本盘, 合计净利 ~40亿, 增速 ~12-15%。"),
        ("③ 机器人期权", "特斯拉 Optimus 旋转关节/机电执行器供应商(2025 拿到旋转关节采购合同, 钱塘 50亿投资), 但 2025 年报被指'零进展'、未跨过样品→量产; 现价 ~700亿 市值系于此期权。"),
        ("④ 三情景目标价", "Base 2027E ~¥45.5(≈现价, 已 price in 基本面)/ Bull(Optimus放量) ~¥68 / Bear(机器人证伪) ~¥29, 见『估值对比』。"),
        ("⑤ 主要风险", "Optimus 量产不及预期 / 机器人估值透支 / NEV 热管理增速下滑(2025 首现放缓) / A/H 溢价收敛。"),
    ],
})

# ════════════ 2. 历史财务与估值 ════════════
ha = K.write_history(wb.create_sheet(S_HIST), {
    "title": "三花智控 历史财务与估值 ($B) — 2021-2025A + 当下(分部按年报一手拆分, CNY→$B 按年均汇率)",
    "hist_cols": HC, "hist_years": HY,
    "vals_in_usd": True,
    "fx_hist": FX_HIST, "fx_now": FX_FC,
    "segments": [
        ("制冷空调零部件 收入", [1.799, 2.095, 2.274, 2.300, 2.592], True),
        ("汽车零部件 收入", [0.682, 1.070, 1.195, 1.582, 1.733], True),
        ("机器人执行器 收入", [0.0, 0.0, 0.0, 0.0, 0.0], True),
    ],
    "total_now": 4.325,
    "gm_pct": [0.2568, 0.2608, 0.2762, 0.2747, 0.2878], "gm_now": 0.288,
    "ni": [0.261, 0.382, 0.413, 0.430, 0.567], "ni_now": 0.567,
    "eq": [1.745, 1.949, 2.551, 2.711, 4.469], "eq_now": 4.469,
    "shares": [3580, 3573, 3606, 3689, 4208], "shares_now": 4208,
    "px_end": [25.30, 21.22, 29.40, 23.51, 55.31],
    "px_now": PX_NOW,
    "px_avg": [23.14, 22.49, 26.68, 21.91, 33.27],
    "band_note": "P/E 历史常态 28-36x(2022-24) → 2025 末冲 54x(机器人重估) → 当下 47x TTM, 处历史高位。",
    "notes": [
        ("制冷空调零部件 收入", "= 年报分部: 2025 185.85亿(+12.22%, 毛利28.77%); 2021-24 按年报+增速回填(116/141/161/165.6亿)。四通阀/电子膨胀阀全球份额>50%。"),
        ("汽车零部件 收入", "= 年报分部: 2025 124.27亿(+9.14%, 毛利28.79%); 2021-24(44/72/84.6/113.9亿)。新能源车热管理为主, 特斯拉等核心客户。"),
        ("机器人执行器 收入", "2025 前基本零收入(2024底小批量出货, 并入其他/不重大); 作为独立分部从 2026 起测算, 历史列=0。"),
        ("HREV", "总营收=交易所实际(统一数据 API): 2025 310.12亿 CNY(+10.97%)=$4.325B@7.17; 分部按年报『分行业/分产品』, 2021-22 段拆分为估、加总勾稽实际总额。"),
        ("HGMP", "综合毛利率: 公司年报, 2025 28.78%(+1.31pct)。"),
        ("HNI", "归母净利(统一数据 API): 2021 16.84/2022 25.73/2023 29.21/2024 30.99/2025 40.63亿 CNY(+31.10%), 按年均汇率折$B; 当下=2025A。"),
        ("HEQ", "归母股东权益(总资产−总负债): 2025 320.4亿 CNY(H股IPO募资并表, 较2024 195亿跳升); 按年均汇率。"),
        ("HSH", "完全摊薄股本(mn): 2021-24 约 35.8-36.9亿(A股), 2025 H股上市后 ~42.08亿; 前瞻用 42.08亿。EPS 按此口径, 2025A EPS ~¥0.97(报告基本EPS 1.03 用加权 39.4亿)。"),
        ("HPX", "年末股价: 交易所月K真实收盘(不复权); 当下=现价 ¥45.51。"),
        ("HPXA", "年均股价: 月度收盘均值, 同『股价走势』单一价格源。"),
    ],
})

# ════════════ 3. 股价走势 ════════════
def phase_fn(ym):
    if ym <= "2022-12":
        return "① 制冷/汽零驱动"
    if ym <= "2025-06":
        return "② 区间震荡"
    if ym <= "2026-01":
        return "③ 机器人重估"
    return "④ 高位消化"

px = K.write_price_chart(wb.create_sheet(S_PX), MONTHLY, {
    "fn": phase_fn,
    "rows": [("① 制冷/汽零驱动", "制冷龙头 + 汽车热管理二次成长, 20-30元区间"),
             ("② 区间震荡", "汽车增速放缓预期压制, 18-30元"),
             ("③ 机器人重估", "特斯拉 Optimus 供应链 + 旋转关节合同, 26→58元翻倍"),
             ("④ 高位消化", "高盛下调 + '零进展'质疑, 从58回落至45-46消化")],
}, title="三花智控 月度股价 (CNY, 不复权)")

# ════════════ 4. 卖方研报共识 ════════════
K.write_consensus(wb.create_sheet(S_CONS), {
    "title": "卖方研报共识 — A股近6月5家(4买入1增持) + 主要投行目标价;这张表是后面测算的'卖方对账单'",
    "overview": "全街评级偏多(东财聚合近6月4买入/1增持), 2026E EPS 一致预期 ~¥1.133 / 2027E ~¥1.356。目标价: 东方证券 52.65(45x)、中金 A股 59.5(59x/52x 26/27E)、SOTP 法 59元; 高盛因'机器人预期过高'下调评级、港股目标价下调至 40港元。核心叙事: 制冷+汽零稳健基本盘 + 机器人期权, 分歧全在机器人兑现节奏与估值。",
    "assumptions": [
        ("总营收增速\n(2026)", "共识 ~+15%(制冷 +8-10% / 汽车 +10-13% / 机器人小基数放量)。", "分歧在汽车(2025 首现放缓)与机器人起量节奏。", "Base 取 +15%(制冷8%+汽车NEV量价+机器人5亿)。"),
        ("机器人 2027 收入", "乐观口径给 20-50亿, 谨慎口径给 <15亿。", "最大分歧: Optimus 量产节奏 + 三花份额/单机价值。", "Base 出货 25万台×¥5300 → ~13亿, 谨慎侧。"),
        ("2026E EPS", "一致预期 ¥1.133(5家 1.11-1.16)。", "口径差异小, 主业驱动。", "Base ¥1.11(略低于共识, 机器人贡献保守)。"),
        ("目标 P/E\n(2026E)", "卖方 45-59x(中金 59x 最激进)。", "给机器人多少结构溢价的分歧。", "三层分解拍 40.8x(sent=1.0), 低于共识——现价已 price in。"),
    ],
    "divergences": [
        "① Optimus 量产节奏: 2026 是否真放量(马斯克指引 vs '零进展'现实)决定机器人收入拐点年。",
        "② 机器人估值方法: SOTP 给机器人 40x vs 并入整体 → 隐含机器人市值 400-800亿差异巨大。",
        "③ 汽车热管理增速: 2025 首现放缓, NEV 渗透率见顶担忧 vs 单车价值量提升对冲。",
    ],
    "stances": [
        "中金(A股 TP 59.5, 59x/52x 26/27E): 制冷+汽零稳健 + 机器人期权, 但港股估值中枢下移下调至40港元。",
        "东方证券(TP 52.65, 45x): 2026-28 归母 49.14/56.77/65.52亿, 可比公司45x。",
        "高盛(下调): 市场对机器人业务预期过高, 兑现不确定, 估值透支。",
    ],
})

# ════════════ 5. 历史估值倍数 ════════════
hm = K.write_hist_multiples(wb.create_sheet(S_HMULT), {
    "title": "历史估值倍数 — 自身历史 P/E 带 + 当下 TTM + 同业对照(P/E 主线, P/B 支线)",
    "intro": "这一页是『估值倍数假设』的数据底座: ①三花自己历史上值多少 forward P/E(逐年+年内高低带) ②现在市场给多少(TTM 46.9x) ③同行值多少 + 相对核心同行比值。看完这页再去下一页拍三案目标 P/E。",
    "s_hist": S_HIST, "ha": ha, "hist_cols": HC, "hist_years": HY,
    "yhigh": px["yhigh"], "ylow": px["ylow"],
    "fwd_note": "forward P/E ≈40.1x(2026E) · ≈35.9x(2027E)(现价÷模型前瞻EPS)",
    "self_name": "三花智控",
    "self_fwd_pe_label": "≈40x(26E)",
    "self_note": "本模型标的; forward 推导见『情景估值』。P/E 主线(阀控 franchise 有定价权); P/B 5.9x 仅支线。",
    "peers": [
        {"name": "盾安环境(002011)", "yearly": [None, None, None, None, None], "cur_pb": 2.8, "cur_pe": 18.0, "fwd_pe": 15.0,
         "note": "制冷配件同业(四通阀第二), 格力控股; 估值远低于三花(无机器人叙事)。"},
        {"name": "银轮股份(002126)", "yearly": [None, None, None, None, None], "cur_pb": 2.5, "cur_pe": 22.0, "fwd_pe": 18.0,
         "note": "汽车热管理直接对手; 无机器人溢价, forward ~18x。"},
        {"name": "拓普集团(601689)", "yearly": [None, None, None, None, None], "cur_pb": 4.5, "cur_pe": 35.0, "fwd_pe": 28.0,
         "note": "汽零+机器人执行器双主题可比(直线执行器/丝杠), forward ~28x, 同享机器人溢价。"},
        {"name": "绿的谐波(688017)", "yearly": [None, None, None, None, None], "cur_pb": 9.0, "cur_pe": 90.0, "fwd_pe": 60.0,
         "note": "谐波减速器龙头(三花合资方), 纯机器人标的估值上沿 forward ~60x。"},
        {"name": "大盘(沪深300, 参照下沿)", "yearly": None, "cur_pb": None, "cur_pe": 13.0, "fwd_pe": 12.0,
         "note": "光谱下沿。"},
    ],
    "ratio": {"peer": "拓普集团(601689)",
              "note": "结构溢价对账线: 三花 forward 40x vs 拓普 28x, 比值 ~1.4——三花的机器人+制冷双龙头溢价。"},
    "reading": "① 自己: 当下 TTM 47x vs 历史带(2022-24 forward 28-36x) → 明显处高位, 第一层锚取历史峰值 34x, 不用被本轮抬高的 47x。② 同行: 制冷/汽零同业 15-22x, 机器人纯标的 60-90x → 三花是混合体。③ 相对拓普比值 ~1.4 → 第二层结构溢价 1.2x 有据。→ 下一页: 峰值 34x × 溢价 1.2 × 情绪值(三案)。",
})

# ════════════ 6. 估值倍数假设(P/E 主线) ════════════
ma = K.write_multiple_assumptions(wb.create_sheet(S_MULT), {
    "title": "估值倍数假设 — P/E 主线(forward) + 三案目标倍数(= 历史峰值 × 结构溢价 × 情绪值)",
    "intro": "镜头判断 + 三层分解。倍数口径: 目标年前瞻 P/E(× 当年 EPS)。『情景切换』引用并切换, 『情景估值』套用当前案, 『估值对比』三案并排。",
    "why_text": ("镜头选择是业务判断: 三花'穿越周期持续存在的东西'是盈利质量——阀控/精密制造 franchise(四通阀/电子膨胀阀全球份额>50%, 定价权 + 客户认证壁垒 + 28.8% 毛利), "
                 "盈利可持续且增长 → P/E 主线(资本化盈利)。它不是纯商品周期(有定价权), 也不是纯资产股(轻资产、高 ROE ~16-21%)。"
                 "盈利耐用性 = 制冷龙头(稳态现金牛)+ 汽车热管理(NEV 量价成长)+ 机器人执行器(期权), 三段质量与久期不同——所以主线 P/E 用自身历史中枢 × 有限结构溢价, 并用 SOTP 支线拆'核心 vs 机器人'各值多少, DCF 支线查'现价隐含多少年高增长'。"
                 "P/E 高不必然=贵, 是市场对成长久期 + 机器人期权的定价; 但当现价 P/E 远超自身历史带、且机器人尚未从样品到量产时, 高 P/E 就含透支风险。"),
    "why_rows": 6,
    "method_text": "三层分解(不硬拍): ①历史 forward P/E 峰值(过去最强成长期实际到过, 剔除本轮机器人重估) × ②结构溢价(锚同业相对法) × ③情绪值(周期/情绪位置, 依据『综合判断仪表盘』D块)。一致性检验: 三层相乘应复现最近实际倍数。",
    "peak": 34.0, "peak_note": "第一层 历史 forward P/E 峰值: 真实年末价÷次年EPS 逐年反推(2022末29x/2023末36x/2024末28x), 取成长期中枢 34x。刻意不用本轮机器人重估后的 47-54x 当锚(已含结构溢价+情绪, 再乘=双重计算)。",
    "premium": 1.20, "premium_note": "第二层 结构溢价: 机器人期权 + 制冷/汽零双龙头。对账线: 拓普(汽零+机器人)forward 28x ≈ 汽零同业18x×1.55; 三花给 1.2x(高于纯汽零同业, 低于纯机器人标的绿的谐波60x, 因机器人仅期权、未起量)。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "hpb_row": ha["HPE"],
    "cases": [
        ("Bear", [0.78, 0.62, 0.54, 0.50, 0.46], "机器人证伪('零进展'延续/Optimus 跳票)+ 汽车增速下滑 → 市场剥离机器人溢价, 打回制冷/汽零同业档(2027E ~25x=34×1.2×0.62), 参照银轮/盾安 18-22x + 少量残留。"),
        ("Base", [1.00, 0.88, 0.80, 0.72, 0.66], "机器人期权部分兑现(小批量→稳定供货)、板块 froth 挤出 → 2026E 付 40.8x(34×1.2×1.0), 之后随成长成熟向'峰值×溢价'(40.8x)下方 normalize, 2030E ~27x。"),
        ("Bull", [1.32, 1.18, 1.05, 0.95, 0.88], "Optimus 2026-27 真放量 + 三花份额/单机价值兑现 + 市场维持机器人成长溢价, 2027E 给 48x(34×1.2×1.18), 对标拓普/机器人链上沿。"),
    ],
    "sent_note": "情绪值=板块/情绪位置。1.0=付足『峰值×溢价』(40.8x); >1=机器人 FOMO 超涨; <1=期权折价。历史列=实际 forward P/E÷(峰值×溢价)反推(2021-25末), 显示 2021 成长泡沫 1.3、2022-24 回落 0.7-0.9、2025 机器人重估回 1.3。",
    "reconcile_text": "卖方目标倍数 45-59x(中金59x最激进/东方45x) vs 我们 Base 2026E 40.8x: 我们更谨慎——凭事实+逻辑: 机器人年报'零进展'、未跨样品→量产, 现价40x已超自身历史带(28-36x), 给机器人的溢价应以量产证据为条件, 而非按 2030+ 潜力提前折现。这是有判断力的非共识(略低于街), 若 Optimus 真放量则上修至 Bull 48x。",
    "source_text": "第一层=统一数据 API 真实年末价÷EPS; 第二层=拓普/银轮/绿的谐波 forward 对账(2026-07); 第三层档位依据『综合判断仪表盘』D 块。",
})

# ════════════ 7. 情景切换 ════════════
sw = K.write_scenario_switch(wb.create_sheet(S_SW), {
    "title": "情景切换 — 全模型唯一的情景参数库 + 切换开关 (默认 Base)",
    "usage": ("怎么用: B2 是唯一入口——下拉选案 → 案序号派生 → 各杠杆『当前案』行跟着切 → 整条明细链(锚→测算→利润→倍数→估值)变档。"
              "三案对比不用切: 『估值对比』恒常三列并排。情景参数只在本页改(蓝字); 未列入的假设(制冷净利率/净利转换/留存/NEV销量路径)三案共用 Base。"),
    "cases": CASES, "default": "Base",
    "triggers": [
        ("Bear", "机器人'零进展'延续或 Optimus 跳票 + 汽车热管理增速持续下滑(NEV 渗透见顶)→ 机器人溢价被剥离, 倍数回制冷/汽零同业档。"),
        ("Base", "制冷 +8% 稳健 + 汽车随 NEV 量价 +12% + 机器人小批量→稳定供货(2030 出货130万台×¥6400)→ 基本面稳步兑现, 倍数向 40.8x 中枢收敛。"),
        ("Bull", "Optimus 2026-27 真放量 + 三花份额与单机价值双升(2030 250万台×¥8000)+ 制冷/汽车提速 → 机器人成长溢价维持, 倍数守 48x。"),
    ],
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "levers": [
        {"key": "refg", "name": "制冷零部件 收入增速", "fmt": K.PCT,
         "vals": {"Bear": [0.05, 0.04, 0.03, 0.03, 0.03], "Base": [0.08, 0.08, 0.08, 0.08, 0.08], "Bull": [0.11, 0.11, 0.10, 0.10, 0.09]},
         "desc": "制冷零部件收入增速。驱动: 空调/热泵存量 + 数据中心液冷 + 储能温控新场景。成熟基本盘, 增速最稳。",
         "stories": {"Bear": "家电周期下行 + 新场景放量慢, 回落至 3-5%。", "Base": "热泵/液冷/储能温控对冲家电成熟, 稳态 8%。", "Bull": "数据中心液冷 + 储能温控加速放量, 10-11%。"},
         "hist": [0.10, 0.216, 0.142, 0.028, 0.122]},
        {"key": "nevv", "name": "三花汽车有效单车价值 ($/辆)", "fmt": K.N0,
         "vals": {"Bear": [72, 72, 73, 74, 75], "Base": [75, 77, 79, 81, 83], "Bull": [78, 82, 87, 92, 97]},
         "desc": "= 三花汽车收入 ÷ 全球NEV销量(含份额与单车价值双重, $/辆)。2025≈$73.6(¥528)。汽车收入=NEV销量(百万辆)×本行/1000。",
         "stories": {"Bear": "NEV 价格战传导 + 份额被银轮/国际Tier1切, 单车价值停滞。", "Base": "阀件→集成模块升级(单车价值¥2000→6000-10000)对冲价格压力, 缓升。", "Bull": "集成热管理模块渗透 + 份额提升双升, $/辆快增。"},
         "hist": [46, 51, 55, 62, 74]},
        {"key": "hmnv", "name": "全球人形机器人出货量 (万台)", "fmt": K.N0, "cols": FCf,
         "vals": {"Bear": [5, 12, 22, 32, 40], "Base": [10, 25, 50, 85, 130], "Bull": [15, 40, 90, 160, 250]},
         "desc": "物理锚: 机器人段收入挂它。Optimus V3 2026 量产、目标2030年产百万台; 高盛2030 乐观100万+/谨慎数万台, 分歧极大。三案共用 2025 基数~2万台。",
         "stories": {"Bear": "Optimus 卡在手部灵活度/量产爬坡, 全行业2030仅40万台。", "Base": "Optimus 稳步爬坡 + 国产跟进, 2030 130万台。", "Bull": "Optimus 2026-27 放量兑现马斯克指引, 2030 250万台。"},
         "hist": None},
        {"key": "rval", "name": "三花机器人有效单机价值 ($/台)", "fmt": K.N0,
         "vals": {"Bear": [560, 590, 620, 650, 680], "Base": [699, 741, 811, 853, 895], "Bull": [800, 900, 1050, 1150, 1250]},
         "desc": "= 单机执行器价值 × 三花份额(含旋转执行器+减速器, $/台, 全球出货口径)。三花=特斯拉旋转关节供应商, 单机40执行器×$800; 有效值含份额。机器人收入=出货(万台)×本行/100000。",
         "stories": {"Bear": "份额被拓普/国际切 + 单执行器降价, 有效值 <¥5000。", "Base": "旋转关节独供 + 绿的谐波减速器合资, 有效值 ¥5000→6400。", "Bull": "份额扩至线性执行器+丝杠总成 + 单机价值升, ¥5700→8900。"},
         "hist": None},
        {"key": "auop", "name": "汽车段 净利率", "fmt": K.PCT,
         "vals": {"Bear": [0.10, 0.10, 0.095, 0.095, 0.09], "Base": [0.115, 0.115, 0.12, 0.12, 0.12], "Bull": [0.12, 0.125, 0.13, 0.135, 0.14]},
         "desc": "汽车热管理段净利率。规模效应 + 集成模块 mix 提升 vs 客户降价压力。2025 毛利28.79%对应净利 ~11.5%。",
         "stories": {"Bear": "价格战压制, 净利率降至 9-10%。", "Base": "规模+mix 稳住 11.5-12%。", "Bull": "集成模块高毛利占比升, 净利率升至 13-14%。"},
         "hist": [0.09, 0.10, 0.105, 0.11, 0.115]},
        {"key": "roop", "name": "机器人段 净利率", "fmt": K.PCT, "cols": FCf,
         "vals": {"Bear": [0.08, 0.10, 0.12, 0.12, 0.12], "Base": [0.10, 0.13, 0.15, 0.15, 0.15], "Bull": [0.12, 0.15, 0.18, 0.18, 0.18]},
         "desc": "机器人执行器段净利率。爬坡期低(产能利用不足)→ 量产后精密件高毛利。对标绿的谐波/精密零部件 15-18%。",
         "stories": {"Bear": "爬坡慢+价格竞争, 净利率 8-12%。", "Base": "量产后规模效应, 净利率 10→15%。", "Bull": "高价值总成+份额, 净利率 12→18%。"},
         "hist": None},
    ],
    "linked": [
        {"key": "sent", "name": "情绪值(目标P/E第三层)", "fmt": K.N2,
         "src_sheet": S_MULT, "src_row0": ma["sent_row0"],
         "note": "三案取值与依据见『估值倍数假设』; 本页只做切换——要改情绪值, 去那页改蓝字。"},
    ],
})
# derived: 目标 P/E(当前案) = 峰值 × 溢价 × 当前案情绪
_pk = f"'{S_MULT}'!{ma['pk_cell']}"
_pr = f"'{S_MULT}'!{ma['pr_cell']}"
_sent_act = sw["SWACT"]["sent"]
_r = sw["next_row"]
K.lab(wb[S_SW], f"A{_r}", "目标 P/E(当前案)", b=True)
for _c in ALLC:
    K.fml(wb[S_SW], f"{_c}{_r}", f"={_pk}*{_pr}*{_c}{_sent_act}", K.MX, link=True)
K.logic(wb[S_SW], f"L{_r}", "= 历史峰值 34x × 结构溢价 1.2 × 当前案情绪值 → 喂『情景估值』的前瞻目标 P/E。")
SWPE = _r

# ════════════ 8. 需求物理盘子 [ANCHOR] ════════════
anchor = K.write_anchor(wb.create_sheet(S_ANCHOR), {
    "title": "下游需求物理盘子 — 全球NEV销量 + 全球人形机器人出货量(三花两大成长段的物理锚)",
    "all_cols": ALLC, "all_years": ALLY,
    "series": [
        ("全球NEV销量 (百万辆)", [6.6, 10.5, 14.0, 18.2, 23.5, 28.5, 32.5, 36.5, 39.8, 42.65],
         "全球新能源车(BEV+PHEV)销量。2025 2354万辆(+29%, 中国70%); 2026E 2850万; 2030E 4265万(渗透率>40%)。三案共用(需求侧较共识)。来源: 机构统计/BNEF/中汽协。", K.N1),
        ("全球人形机器人出货量 (万台)", [0, 0, 0, 0.5, 2.0, None, None, None, None, None],
         "全球人形机器人年出货量。2025 ~2万台(Optimus 数千+Unitree等); 前瞻=情景驱动(『情景切换』hmnv 当前案)。来源: 高盛/特斯拉指引/GGII。", K.N1),
    ],
    "yoy_row": "全球NEV销量 (百万辆)",
    "source_note": "口径: 全球NEV=BEV+PHEV零售/批发口径; 人形机器人=全行业年出货(未量产年为示意/爬坡估)。NEV 路径三案共用(需求侧); 人形出货前瞻按情景当前案。",
    "role_note": "作用: 汽车段收入=NEV销量×三花有效单车价值; 机器人段收入=人形出货×三花有效单机价值。改这两个物理量 → 收入 → 估值全链动。",
})
NEV_ROW = anchor["row_of"]["全球NEV销量 (百万辆)"]
HMN_ROW = anchor["row_of"]["全球人形机器人出货量 (万台)"]
# 人形出货前瞻 = 情景切换 hmnv 当前案
for _c in FCf:
    K.fml(wb[S_ANCHOR], f"{_c}{HMN_ROW}", f"={K.R(S_SW, _c + str(sw['SWACT']['hmnv']))}", K.N1, link=True)

# ════════════ 9. 分部测算 ════════════
seg = K.write_segment_model(wb.create_sheet(S_SEG), {
    "title": "分部测算 — 制冷(成长) + 汽车(NEV销量×单车价值) + 机器人(人形出货×单机价值) ($B)",
    "all_cols": ALLC, "all_years": ALLY, "logic_col": "N",
    "groups": [
        ("物理锚(引自『需求物理盘子』)", [
            ("全球NEV销量 (百万辆)", None, K.N1, "= 引自『需求物理盘子』。汽车收入挂它。"),
            ("全球人形出货 (万台)", None, K.N1, "= 引自『需求物理盘子』(前瞻=情景当前案)。机器人收入挂它。"),
        ]),
        ("制冷段 = 上年 × (1+增速)", [
            ("制冷 收入增速", None, K.PCT, "历史=实际; 前瞻=『情景切换』当前案。"),
            ("制冷 收入 ($B)", None, K.N1, "历史取历史财务实数; 前瞻=上年×(1+增速)。喂『利润与收入假设』。"),
        ]),
        ("汽车段 = NEV销量 × 有效单车价值", [
            ("汽车 有效单车价值 ($/辆)", None, K.N0, "历史=实际反推(收入÷NEV销量); 前瞻=『情景切换』当前案。"),
            ("汽车 收入 ($B)", None, K.N1, "历史取实数; 前瞻=NEV销量(百万辆)×有效单车价值/1000。"),
        ]),
        ("机器人段 = 人形出货 × 有效单机价值", [
            ("机器人 有效单机价值 ($/台)", None, K.N0, "前瞻=『情景切换』当前案(含份额)。历史~0。"),
            ("机器人 收入 ($B)", None, K.N1, "前瞻=人形出货(万台)×有效单机价值/100000。历史=0。"),
        ]),
    ],
})
m = seg["m"]
REFG_HROW = ha["seg_rows"]["制冷空调零部件 收入"]
AUTO_HROW = ha["seg_rows"]["汽车零部件 收入"]
ROBO_HROW = ha["seg_rows"]["机器人执行器 收入"]
# 物理锚引用
for col in ALLC:
    K.fml(wb[S_SEG], f"{col}{m['全球NEV销量 (百万辆)']}", f"={K.R(S_ANCHOR, col + str(NEV_ROW))}", K.N1, link=True)
    K.fml(wb[S_SEG], f"{col}{m['全球人形出货 (万台)']}", f"={K.R(S_ANCHOR, col + str(HMN_ROW))}", K.N1, link=True)
# 制冷: 历史实数 + 增速; 前瞻递推
for col in HC:
    K.fml(wb[S_SEG], f"{col}{m['制冷 收入 ($B)']}", f"={K.R(S_HIST, col + str(REFG_HROW))}", K.N1, link=True)
    K.fml(wb[S_SEG], f"{col}{m['汽车 收入 ($B)']}", f"={K.R(S_HIST, col + str(AUTO_HROW))}", K.N1, link=True)
    K.fml(wb[S_SEG], f"{col}{m['机器人 收入 ($B)']}", f"={K.R(S_HIST, col + str(ROBO_HROW))}", K.N1, link=True)
    # 历史有效单车价值 = 汽车收入$B*1000/NEV销量
    K.fml(wb[S_SEG], f"{col}{m['汽车 有效单车价值 ($/辆)']}", f"={col}{m['汽车 收入 ($B)']}*1000/{col}{m['全球NEV销量 (百万辆)']}", K.N0)
# 制冷增速历史(实际)
K.introw(wb[S_SEG], m["制冷 收入增速"], HC[1:], [0.216, 0.142, 0.028, 0.122], None, K.PCT)
for col in FCf:
    K.fml(wb[S_SEG], f"{col}{m['制冷 收入增速']}", f"={K.R(S_SW, col + str(sw['SWACT']['refg']))}", K.PCT, link=True)
    K.fml(wb[S_SEG], f"{col}{m['汽车 有效单车价值 ($/辆)']}", f"={K.R(S_SW, col + str(sw['SWACT']['nevv']))}", K.N0, link=True)
    K.fml(wb[S_SEG], f"{col}{m['机器人 有效单机价值 ($/台)']}", f"={K.R(S_SW, col + str(sw['SWACT']['rval']))}", K.N0, link=True)
_prevs = [HC[-1]] + list(FCf[:-1])
for _p, _c in zip(_prevs, FCf):
    K.fml(wb[S_SEG], f"{_c}{m['制冷 收入 ($B)']}", f"={_p}{m['制冷 收入 ($B)']}*(1+{_c}{m['制冷 收入增速']})", K.N1)
for _c in FCf:
    K.fml(wb[S_SEG], f"{_c}{m['汽车 收入 ($B)']}", f"={_c}{m['全球NEV销量 (百万辆)']}*{_c}{m['汽车 有效单车价值 ($/辆)']}/1000", K.N1)
    K.fml(wb[S_SEG], f"{_c}{m['机器人 收入 ($B)']}", f"={_c}{m['全球人形出货 (万台)']}*{_c}{m['机器人 有效单机价值 ($/台)']}/100000", K.N1)
    for _row in ['制冷 收入 ($B)', '汽车 收入 ($B)', '机器人 收入 ($B)']:
        wb[S_SEG][f"{_c}{m[_row]}"].fill = K.OUT

# ════════════ 10. 利润与收入假设 ════════════
fr = K.write_fundamentals(wb.create_sheet(S_FUND), {
    "title": "利润与收入假设 — 分部营收(链测算)+ 段净利率 → 归母净利 → EPS/BPS/ROE(估值倍数在『估值倍数假设』)",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
    "eps_label": "EPS ($/股; ¥EPS=×FX)", "bps_label": "BPS ($/股; ¥BPS=×FX)",
    "assum_groups": [
        ("段净利率(历史实际锚 + 前瞻; 粗颗粒)", [
            {"name": "制冷 净利率", "vals": [0.135, 0.135, 0.14, 0.14, 0.14, 0.14, 0.14, 0.14, 0.14, 0.14],
             "fmt": K.PCT, "logic": "制冷零部件段净利率。四通阀/电子膨胀阀龙头, 毛利28.8%, 净利率约14%, 稳态。三案共用。"},
            {"name": "汽车 净利率", "vals": [0.09, 0.10, 0.105, 0.11, 0.115, None, None, None, None, None],
             "fmt": K.PCT, "logic": "汽车热管理段净利率; 历史随规模提升(9%→11.5%); 前瞻=『情景切换』当前案。",
             "link": {"sheet": S_SW, "row": sw["SWACT"]["auop"]}},
            {"name": "机器人 净利率", "vals": [None, None, None, None, None, None, None, None, None, None],
             "fmt": K.PCT, "logic": "机器人执行器段净利率; 前瞻=『情景切换』当前案(爬坡低→量产高)。历史 n.m.。",
             "nm_cols": ["B", "C", "D", "E", "F"],
             "link": {"sheet": S_SW, "row": sw["SWACT"]["roop"]}},
        ]),
        ("净利转换与留存", [
            {"name": "净利转换率(归母/分部净利和)", "vals": [1.01, 1.01, 1.01, 1.01, 1.01, 1.0, 1.0, 1.0, 1.0, 1.0],
             "fmt": K.PCT, "logic": "分部净利和→归母净利: 三花少数股东占比小, 转换率≈1.0(历史反推 ~1.01)。"},
            {"name": "留存率", "vals": [0.65, 0.65, 0.70, 0.70, 0.72, 0.72, 0.72, 0.72, 0.72, 0.72],
             "fmt": K.PCT, "logic": "留存率=1−派息率; 三花派息率约28-35%, 留存~72%。仅用于 BPS 递推, P/E 主线不用 P/B。"},
        ]),
    ],
    "segments": [
        {"name": "制冷空调零部件 收入", "hist_row": "制冷空调零部件 收入", "fwd": {"sheet": S_SEG, "row": m["制冷 收入 ($B)"]}},
        {"name": "汽车零部件 收入", "hist_row": "汽车零部件 收入", "fwd": {"sheet": S_SEG, "row": m["汽车 收入 ($B)"]}},
        {"name": "机器人执行器 收入", "hist_row": "机器人执行器 收入", "fwd": {"sheet": S_SEG, "row": m["机器人 收入 ($B)"]}},
    ],
    "profit_terms": [
        (["制冷空调零部件 收入"], "制冷 净利率", False),
        (["汽车零部件 收入"], "汽车 净利率", False),
        (["机器人执行器 收入"], "机器人 净利率", True),
    ],
    "conv_assum": "净利转换率(归母/分部净利和)", "retention_assum": "留存率",
    "note_text": "分部营收(制冷成长 + 汽车NEV量价 + 机器人出货量价)→ 段净利率 → 分部净利和 → ×净利转换率 → 归母净利 → EPS(股本42.08亿)/BPS/ROE。注: 此处『营业利润』行实为分部净利和(段用净利率而非OPM), 净利=其×转换率。下游『情景估值』直接引本表每股。",
})

# ════════════ 11. 情景估值(P/E 主线 + SOTP + DCF 支线) ════════════
fr_pe = dict(fr); fr_pe["BPS"] = fr["EPS"]   # P/E 主线: 主线镜头每股用 EPS
sv = K.write_scenario_valuation(wb.create_sheet(S_VAL), {
    "title": "情景估值 — 当前案的逐年隐含价 (P/E 主线; SOTP + DCF 交叉验证)",
    "intro": "本表输出=『情景切换』当前案(默认Base)。主线: 隐含价=目标P/E(当前案)×前瞻EPS。历史列用实际年末价反推倍数(事实); 前瞻是预测、不拟合现价。三案并排见『估值对比』。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf, "hist_years": HY,
    "fx_fwd": FX_FC, "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
    "s_fund": S_FUND, "fr": fr_pe,
    "s_switch": S_SW, "target_row": SWPE, "sw_cell": "B2",
    "yend": px["yend"], "yavg": px["yavg"],
    "reading": "主线行读法: 『目标倍数/隐含股价』两行即 P/E 主线(历史列=年末价÷当年EPS=实际P/E; 前瞻=目标P/E×当年EPS)。Base 隐含价轨迹显示现价 ¥45.5 ≈ Base 2026-27E 隐含价——市场已把基本面 price in, 上行系于机器人期权(Bull)。",
    "method": "方法: 整体公司、P/E 主线逐年估。基本面在『利润与收入假设』; 目标倍数在『估值倍数假设』(三层); 本表最后一步: 目标P/E × 前瞻EPS → 隐含价 + 市值; 下方 SOTP(核心 vs 机器人分部)+ DCF 做支线体检。",
    "concl": "结论(方向性): Base 2027E 目标价 ~¥45.5 ≈ 现价 → 持有; Bull(Optimus放量)~¥68 / Bear(机器人证伪)~¥29。现价已充分反映 Base 基本面, 机器人期权是上行来源但需量产催化。",
})

# ── SOTP 交叉验证块(支线): 核心(制冷+汽车) vs 机器人 分部估值 ──
_ws = wb[S_VAL]
_r2 = _ws.max_row + 2
K.band(_ws, _r2, "SOTP 交叉验证(支线) — 核心业务(制冷+汽车)PE + 机器人业务单独估值 → 加总隐含价", 11); _r2 += 1
FXF = str(FX_FC)
SEG_REFG = fr["seg_rows"]["制冷空调零部件 收入"]
SEG_AUTO = fr["seg_rows"]["汽车零部件 收入"]
SEG_ROBO = fr["seg_rows"]["机器人执行器 收入"]
AM = fr["am"]
SHF_F = K.R(S_HIST, f"$F${ha['HSH']}")
# 核心净利($B) = 制冷收入×制冷净利率 + 汽车收入×汽车净利率
K.lab(_ws, f"A{_r2}", "核心(制冷+汽车)净利 ($B)")
for c in FCf:
    K.fml(_ws, f"{c}{_r2}", f"={K.R(S_FUND, c+str(SEG_REFG))}*{K.R(S_FUND, c+str(AM['制冷 净利率']))}+{K.R(S_FUND, c+str(SEG_AUTO))}*{K.R(S_FUND, c+str(AM['汽车 净利率']))}", K.N2)
CORE_NI = _r2; _r2 += 1
K.lab(_ws, f"A{_r2}", "核心 目标 P/E")
K.introw(_ws, _r2, FCf, [24, 22, 21, 20, 19], None, K.MX)
K.logic(_ws, f"L{_r2}", "核心业务(制冷龙头+汽车热管理)给 19-24x, 对标拓普/银轮等汽零+制冷同业, 随成长成熟下行。")
CORE_PE = _r2; _r2 += 1
K.lab(_ws, f"A{_r2}", "核心 隐含市值 (¥亿)")
for c in FCf:
    K.fml(_ws, f"{c}{_r2}", f"={c}{CORE_NI}*{c}{CORE_PE}*{FXF}*10", K.N0)
CORE_MC = _r2; _r2 += 1
# 机器人净利($B) = 机器人收入×机器人净利率
K.lab(_ws, f"A{_r2}", "机器人 净利 ($B)")
for c in FCf:
    K.fml(_ws, f"{c}{_r2}", f"=IFERROR({K.R(S_FUND, c+str(SEG_ROBO))}*{K.R(S_FUND, c+str(AM['机器人 净利率']))},0)", K.N2)
ROBO_NI = _r2; _r2 += 1
K.lab(_ws, f"A{_r2}", "机器人 目标 P/E")
K.introw(_ws, _r2, FCf, [60, 55, 48, 42, 38], None, K.MX)
K.logic(_ws, f"L{_r2}", "机器人执行器给 38-60x(成长期), 对标绿的谐波/机器人链上沿, 随成熟下行。")
ROBO_PE = _r2; _r2 += 1
K.lab(_ws, f"A{_r2}", "机器人 隐含市值 (¥亿)")
for c in FCf:
    K.fml(_ws, f"{c}{_r2}", f"={c}{ROBO_NI}*{c}{ROBO_PE}*{FXF}*10", K.N0)
ROBO_MC = _r2; _r2 += 1
K.lab(_ws, f"A{_r2}", "SOTP 合计市值 (¥亿)", b=True)
for c in FCf:
    K.fml(_ws, f"{c}{_r2}", f"={c}{CORE_MC}+{c}{ROBO_MC}", K.N0)
SOTP_MC = _r2; _r2 += 1
K.lab(_ws, f"A{_r2}", "SOTP 隐含价 (¥)", b=True); _ws[f"A{_r2}"].fill = K.OUT
for c in FCf:
    K.fml(_ws, f"{c}{_r2}", f"={c}{SOTP_MC}*100/{SHF_F}", K.PX)
SOTP_PX = _r2; _r2 += 1
K.lab(_ws, f"A{_r2}", "SOTP vs 现价")
for c in FCf:
    K.fml(_ws, f"{c}{_r2}", f"={c}{SOTP_PX}/{PX_NOW}-1", K.PCT)
_r2 += 1
K.lab(_ws, f"A{_r2}", "其中: 机器人占 SOTP 市值%", note=True)
for c in FCf:
    K.fml(_ws, f"{c}{_r2}", f"={c}{ROBO_MC}/{c}{SOTP_MC}", K.PCT)
K.logic(_ws, f"L{_r2}", "机器人占比揭示: 现价 ¥45.5(市值1915亿)中隐含的机器人估值份额。Base 下机器人 2027E 仅占 SOTP ~7-10%, 但市场按 2030+ 潜力给了更高份额——这是估值分歧核心。")
_r2 += 2

# ── DCF 交叉验证块(支线) ──
K.band(_ws, _r2, "DCF 交叉验证(支线) — 现价隐含多长的高增长? (FCF≈归母净利×转换率, WACC 贴现 + Gordon 终值)", 11); _r2 += 1
K.lab(_ws, f"A{_r2}", "WACC"); K.inp(_ws, f"B{_r2}", 0.09, None, K.PCT)
K.logic(_ws, f"D{_r2}", "贴现率: A股成长制造 8.5-9.5%, 取 9%(beta~1.1, 低负债, 扩张期)。")
WACC_C = f"$B${_r2}"; _r2 += 1
K.lab(_ws, f"A{_r2}", "永续增长 g"); K.inp(_ws, f"B{_r2}", 0.03, None, K.PCT)
K.logic(_ws, f"D{_r2}", "2030 后永续 3%(名义 GDP-)。")
G_C = f"$B${_r2}"; _r2 += 1
K.lab(_ws, f"A{_r2}", "FCF 转换率(FCF/归母净利)"); K.inp(_ws, f"B{_r2}", 0.75, None, K.PCT)
K.logic(_ws, f"D{_r2}", "三花 OCF 强(2025 OCF 50.9亿>净利40.6亿), 但扩产 capex(机器人基地50亿); 取 75%。")
FCFC_C = f"$B${_r2}"; _r2 += 1
NIr = lambda c: K.R(S_FUND, c + str(fr["NI"]))
K.lab(_ws, f"A{_r2}", "FCF ($B, =归母净利×转换)")
for c in FCf:
    K.fml(_ws, f"{c}{_r2}", f"={NIr(c)}*{FCFC_C}", K.N2)
DCF_FCF = _r2; _r2 += 1
K.lab(_ws, f"A{_r2}", "折现因子")
for i, c in enumerate(FCf):
    K.fml(_ws, f"{c}{_r2}", f"=1/(1+{WACC_C})^{i+1}", K.N2)
DCF_DF = _r2; _r2 += 1
K.lab(_ws, f"A{_r2}", "PV of FCF ($B)")
for c in FCf:
    K.fml(_ws, f"{c}{_r2}", f"={c}{DCF_FCF}*{c}{DCF_DF}", K.N2)
DCF_PV = _r2; _r2 += 1
K.lab(_ws, f"A{_r2}", "PV(FCF)合计 + 终值 → 股权价值 ($B)", b=True)
K.fml(_ws, f"B{_r2}", f"=SUM(G{DCF_PV}:K{DCF_PV})+(K{DCF_FCF}*(1+{G_C})/({WACC_C}-{G_C}))*K{DCF_DF}", K.N1)
DCF_EV = _r2; _r2 += 1
K.lab(_ws, f"A{_r2}", "DCF 隐含价 (¥)", b=True); _ws[f"A{_r2}"].fill = K.OUT
K.fml(_ws, f"B{_r2}", f"=B{DCF_EV}*{FXF}*1000/{SHF_F}", K.PX)
K.logic(_ws, f"D{_r2}", "DCF(5年显式+3%永续)= 已证明现金流的保守地板 ~¥20-25, 远低于现价 ¥45.51: 揭示现价的多数价值在 2030 后(机器人成熟+持续成长), 是截至2030的DCF无法捕捉的。三镜头谱系: DCF(证明地板)< SOTP近端 < Base P/E(≈现价)< SOTP-2030/Bull。故 P/E 主线 + SOTP 为主, DCF 只标'已证明部分'的下限。")
DCF_PX = _r2; _r2 += 1
K.lab(_ws, f"A{_r2}", "DCF vs 现价")
K.fml(_ws, f"B{_r2}", f"=B{DCF_PX}/{PX_NOW}-1", K.PCT)

# ════════════ 12. 估值对比(三案并排) ════════════
SWB = sw["SWB"]
PX_NOW_REF = K.R(S_HIST, f"G{ha['HPX']}")
_refg_g = m['制冷 收入增速']
_refgnm = fr["am"]["制冷 净利率"]
_conv = fr["am"]["净利转换率(归母/分部净利和)"]
NEV_A = NEV_ROW


def _fwdprev(j, A, key):
    return (HC[-1] if j == 0 else FCf[j - 1]) + str(A[key])


cmp_rows = [
    {"key": "refg", "label": "制冷收入 ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(REFG_HROW))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'refg')}*(1+{K.R(S_SW, c + str(SWB['refg'] + ci))})"},
    {"key": "auto", "label": "汽车收入 ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(AUTO_HROW))}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_ANCHOR, c + str(NEV_A))}*{K.R(S_SW, c + str(SWB['nevv'] + ci))}/1000"},
    {"key": "robo", "label": "机器人收入 ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ROBO_HROW))}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_ANCHOR, c + str(HMN_ROW))}*{K.R(S_SW, c + str(SWB['rval'] + ci))}/100000"},
    {"key": "rev", "label": "总收入 ($B)", "fmt": K.N1, "bold": True,
     "hist": lambda c, ci, A: f"={c}{A['refg']}+{c}{A['auto']}+{c}{A['robo']}",
     "fwd": lambda c, j, ci, A: f"={c}{A['refg']}+{c}{A['auto']}+{c}{A['robo']}"},
    {"key": "ni", "label": "归母净利 ($B)", "fmt": K.N2, "bold": True,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HNI']))}",
     "fwd": lambda c, j, ci, A: (f"=({c}{A['refg']}*{K.R(S_FUND, c + str(_refgnm))}"
                                 f"+{c}{A['auto']}*{K.R(S_SW, c + str(SWB['auop'] + ci))}"
                                 f"+IFERROR({c}{A['robo']}*{K.R(S_SW, c + str(SWB['roop'] + ci))},0))*{K.R(S_FUND, c + str(_conv))}")},
    {"key": "eps", "label": "EPS (¥)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HEPS']))}*{K.R(S_HIST, c + str(ha['HFX']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['ni']}*1000/{SHF_F}*{FXF}"},
    {"key": "sent", "label": "情绪值(该案; 历史=实际反推)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_MULT, c + str(ma['sent_row0'] + ci))}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_MULT, c + str(ma['sent_row0'] + ci))}"},
    {"key": "pe", "label": "目标 P/E(历史=实际)", "fmt": K.MX,
     "hist": lambda c, ci, A: f"={_pk}*{_pr}*{c}{A['sent']}",
     "fwd": lambda c, j, ci, A: f"={_pk}*{_pr}*{c}{A['sent']}"},
    {"key": "px", "label": "隐含价 (¥)", "fmt": K.PX, "bold": True, "out": True,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HPX']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['pe']}*{c}{A['eps']}"},
    {"key": "ipe", "label": "隐含 forward P/E(体检)", "fmt": K.MX,
     "hist": lambda c, ci, A: f'=IF({c}{A["ni"]}<=0,"N/M",{K.R(S_HIST, c + str(ha["HPX"]))}/({c}{A["eps"]}))',
     "fwd": lambda c, j, ci, A: f"={c}{A['px']}/{c}{A['eps']}"},
    {"key": "up", "label": "历史: vs 实际年末价(回测≈0) / 前瞻: vs 现价", "fmt": K.PCT,
     "hist": lambda c, ci, A: f"={c}{A['px']}/{K.R(S_HIST, c + str(ha['HPX']))}-1",
     "fwd": lambda c, j, ci, A: f"={c}{A['px']}/{PX_NOW_REF}-1"},
]
cm = K.write_comparison(wb.create_sheet(S_CMP), {
    "title": "估值对比 — Bear / Base / Bull 三个情景的目标价并排对比",
    "intro": ("三个情景各自完整推演: 物理锚 → 分部收入 → 归母净利 → EPS → 目标P/E → 逐年隐含价。"
              "本表三案永远并排, 不随『情景切换』变。未列入情景矩阵的假设(制冷净利率/NEV销量路径/转换率)三案共用 Base。"
              "历史列 2021-2025 = 同一条链填实际值: 隐含价历史列=实际年末价(回测行历史列 ≈0%)——整条链的内置回测。"),
    "case_names": CASES,
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "block_start": 16,
    "rows": cmp_rows,
    "summary": {
        "band": "三案汇总 (目标年 2027E; 各案触发条件见『情景切换』)",
        "target_col": "H",
        "rows": [
            ("归母净利($B)", "ni", K.N2, "由该案制冷/汽车/机器人假设逐年推导", False),
            ("总收入($B)", "rev", K.N1, "= 制冷 + 汽车(NEV量价) + 机器人(出货量价)", False),
            ("目标 P/E", "pe", K.MX, "= 历史峰值34x × 结构溢价1.2 × 该案情绪值", False),
            ("EPS(¥)", "eps", K.N2, "归母净利 ÷ 42.08亿股 × FX", False),
            ("隐含价(¥)", "px", K.PX, "= 目标P/E × 2027E EPS", True),
            ("vs 现价", "up", K.PCT, "对照现价 ¥45.51 的上行/下行空间", True),
        ],
        "mcap": {"label": "隐含市值(¥亿)", "key": "px", "expr": f"*{SH_NOW}/1000*10",
                 "note": "= 隐含价 × 股本 42.08亿"},
        "concl": "风险收益比(2027E 目标年 vs 现价 ¥45.51): 下行 Bear ~−37% / 基准 Base ~0% / 上行 Bull ~+50%。Base 已 price in, 偏正的 risk-reward 全系于机器人期权(Bull); 机器人证伪信号触发应转 Bear 重看。",
    },
})

# ════════════ 13. 综合判断仪表盘 ════════════
EPS26 = K.R(S_FUND, "G" + str(fr["EPS"]))
PXD = K.R(S_HIST, "G" + str(ha["HPX"]))
dash = K.write_dashboard(wb.create_sheet(S_DASH), {
    "title": "综合判断仪表盘 — A 基本面拐点 · B 估值错位(预测引擎) · C 催化剂 · D 情绪确认",
    "usage": ("怎么用: 预测引擎是 B(错位)+ C(催化剂)。情绪 D 只做 timing 确认 + 过热刹车。"
              "验收=回测: 放回 2025H2 机器人重估拐点, 这套表当时就能看到那波。"),
    "blocks": [
        {"title": "A. 基本面拐点 — 业务在结构性变好吗?", "rows": [
            ("产品组合迁移", "机器人段从0起量(2026E 5亿→2030E 80亿)", "结构迁移(阀控→执行器), 复用精密制造能力。"),
            ("汽车热管理放缓", "2025 汽车 +9.1%(首现放缓), 单车价值量升对冲", "NEV 渗透见顶 vs 集成模块 mix, 决定汽车段久期。"),
            ("A 判断", "【中】", "核心稳健、机器人期权真实但未量产; 结构改善方向对、节奏未验证。", True),
        ]},
        {"title": "B. 估值错位(预测引擎 ★)— 市场现在给的 vs 基本面该给的 → GAP", "rows": [
            ("市场现在给(forward P/E 26E)", {"fml": f"={PXD}/({EPS26})", "fmt": K.MX, "fill": True},
             "= 现价 ÷ 2026E EPS(公式算, 随模型走)。"),
            ("基本面该给(justified)", {"inp": 34.0, "fmt": K.MX},
             "= 历史峰值×溢价(40.8x)的中枢, 取成长质地该给的 ~34x(不含 FOMO)。"),
            ("错位 GAP = 该给÷市场给 − 1",
             {"fml": lambda ro: f"=B{ro['基本面该给(justified)']}/B{ro['市场现在给(forward P/E 26E)']}-1", "fmt": K.PCT},
             "GAP 正=重估空间; 负=基本面空间已被价格吃完, 进入纯期权/情绪定价区。三花当下 GAP 约 −15%(现价40x > 基本面该给34x)。"),
            ("回测: 2025H2 机器人重估拐点", "2025中 26元(~30x)→ 2026初 58元(~50x)", "当时旋转关节合同 + Optimus 2026量产指引 = 催化剂点燃, GAP 由正转负。"),
        ]},
        {"title": "C. 催化剂 — 什么会逼市场闭合 GAP / 打开 Bull", "rows": [
            ("Optimus 量产落地(2026)", "待验证", "马斯克指引 2026 量产; 真放量→Bull, 跳票→Bear。最强催化。"),
            ("机器人订单/收入确认", "进行中(2025 旋转关节合同~12亿)", "从合同到收入确认、从样品到量产是关键跨越。"),
            ("C 判断", "待兑现", "催化剂真实但时点不定; 未兑现前现价靠预期支撑。", True),
        ]},
        {"title": "D. 情绪确认 — 只做 timing + 刹车", "rows": [
            ("量价温度计", "从58高位回落至45-46, 高盛下调后降温", "高位消化中, FOMO 退潮但未崩。"),
            ("现价倍数 vs 基本面该给", "40x vs 34x = 溢价 ~18%", "市场已付超基本面该给, 进入期权定价区。"),
            ("当前档位", "【过热后消化】", "2025H2-2026初过热 → 当前退潮消化; 对照上一轮:未回到启动区。", True),
            ("衰减扳机", "3 条", "① Optimus 再跳票/延产; ② 机器人年报仍'零进展'; ③ 汽车增速转负。任一翻→情绪值下调→转 Bear。", ),
        ]},
    ],
    "final": {"band": "★ 综合判断(A+B+C+D 收成一句可执行的话)",
              "text": "回测 2025H2: 旋转关节合同+Optimus指引=强催化, 当时该看多✓(GAP由正转负那波已走完)。当下: 基本面 price in(Base≈现价)、GAP −15%(轻度透支)、催化待兑现、情绪过热后消化 → 持有(HOLD); 等 Optimus 量产证据(转Bull)或回落至合理价值~¥35(更好入场)。"},
    "tracking": {
        "intro": "哪个指标恶化 → 哪个假设先崩 → 触发什么动作。",
        "rows": [
            ("__band__", "一、机器人期权(最大摆动项)"),
            ("Optimus 量产/出货", "待验证", "关键敏感项: 机器人段收入拐点", "特斯拉季报/产能公告", "跳票→出货量下调→转 Bear"),
            ("三花机器人订单/收入", "~12亿合同", "关键敏感项: 单机价值×份额兑现", "公司公告/季报分部", "收入不确认→有效单机价值下调"),
            ("__band__", "二、汽车热管理"),
            ("汽车段增速", "2025 +9.1%", "关键敏感项: 单车价值量与份额", "季报分部 + NEV销量", "转负→有效单车价值下调"),
            ("__band__", "三、估值情绪"),
            ("forward P/E vs 历史带", "40x(26E) vs 28-36x", "关键敏感项: 情绪值档位", "股价/模型EPS", "突破50x=过热刹车; 跌破25x=Bear兑现"),
        ],
    },
})

# ════════════ 全局格式 + 落盘 ════════════
K.finalize(wb, freeze={
    S_HIST: "B3", S_PX: "B4", S_CONS: "A2", S_HMULT: "B5", S_MULT: "B4", S_SW: "B3",
    S_ANCHOR: "B3", S_SEG: "B3", S_FUND: "B3", S_VAL: "B4", S_CMP: "B6", S_DASH: "B6",
    S_COVER: "A2",
})
out = os.path.join(os.path.dirname(__file__), "out", "002050SZ_valuation_model.xlsx")
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("saved:", os.path.abspath(out))
print("sheets:", wb.sheetnames)
