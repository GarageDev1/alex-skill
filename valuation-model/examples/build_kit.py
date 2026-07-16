# -*- coding: utf-8 -*-
"""
build_kit.py — valuation-model 可复用 Excel 建模引擎(公司无关)。v2: 对齐 v47 架构。

目的:把"每个公司从零写 ~1100 行 openpyxl"抽成一套引擎。新公司只写一层薄薄的
data dict(分部 / 历史财务 / 逐年汇率 / 假设 / 情景杠杆)+ 调本文件的 sheet 骨架
函数,就能产出与海力士 v47 同档次、且能过 scripts/validate_valuation.py 的机构级
Excel,而不必重写格式、helper、跨表引用、情景切换机制。

标准 sheet 架构(05-excel-format §1, 只向前引用的 DAG):
  封面 → 历史财务与估值 → 股价走势 → 卖方研报共识 → 历史估值倍数 → 估值倍数假设
  → 情景切换 → 物理锚[ANCHOR] → 分部测算 → 利润与收入假设 → 情景估值 → 估值对比
  → 综合判断仪表盘

工程规范(固化的坑, 调用方不需要再操心):
  - 全面禁用单元格批注(红三角): inp/introw 的 src 仅作代码内文档, 不生成 Comment。
    解释一律走可见列(逻辑列 / 备注列), 见 06-sourcing。
  - 文字开头 "=" 自动转全角 "＝"(lab/logic/txt 统一守卫), 否则 Excel 当公式解析失败。
  - 情景切换案序号用嵌套 IF(nested_if_switch), 禁 MATCH+数组常量(WPS/locale 下报错)。
  - 逻辑列正常黑色、非斜体, 去 markdown 粗体标记, 句末自动分行。
  - 行高不锁死 wrap 文字行(让 Excel 打开自适应), 只对色带行保留小高度。
  - finalize(): 年份表头右对齐 + 每张 sheet 冻结窗格。

本文件不含任何公司专有数字。build_hynix.py 是不依赖本 kit 的 v47 实现(抽取源);
build_template.py 是演示如何用本 kit 的瘦实例。
"""
import re as _re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

# ════════════════════════════ 样式常量 ════════════════════════════
BLUE  = Font(color="0000FF")                       # 硬编码输入
BLACK = Font(color="000000")                       # 公式
GREEN = Font(color="008000")                       # 跨表链接
GREY  = Font(color="808080")                       # 次要行名(不用斜体)
HF    = Font(color="FFFFFF", bold=True)            # 表头白字
BF    = Font(bold=True)                             # 加粗

HFILL = PatternFill("solid", fgColor="1F4E79")     # 主表头(深蓝)
BAND  = PatternFill("solid", fgColor="305496")     # 分节色带(中蓝)
CH    = PatternFill("solid", fgColor="D9E1F2")     # 列头(浅蓝)
OUT   = PatternFill("solid", fgColor="BDD7EE")      # 关键产出行(蓝灰)
CUR   = PatternFill("solid", fgColor="FCE4D6")      # 当下/现价(橙)
GREYF = PatternFill("solid", fgColor="F2F2F2")     # 灰底(组标题行)

PCT = '0.0%'
N1  = '#,##0.0;(#,##0.0);-'
N0  = '#,##0;(#,##0);-'
N2  = '0.00'
MX  = '0.0"x"'
PX  = '#,##0'                                       # 股价(整数)

_thin = Side(style="thin", color="BFBFBF")
BORD  = Border(bottom=_thin)


# ════════════════════════════ helper ════════════════════════════
def _esc(t):
    """文字开头 '=' 转全角(否则 Excel 当公式解析、整格渲染失败)。"""
    if isinstance(t, str) and t.startswith("="):
        return "＝" + t[1:]
    return t


def R(sheet, cell):
    """跨表引用,sheet 名统一加引号(渲染稳)。"""
    return f"'{sheet}'!{cell}"


def uy(v, fx):
    """逐年汇率换算:兆(本币) → $B。fx = 当年均值汇率(本币/USD)。"""
    return None if v is None else round(v * 1000 / fx, 2)


def hdr(ws, row, text, span=12):
    ws.cell(row, 1, text).font = HF
    for c in range(1, span + 1):
        ws.cell(row, c).fill = HFILL
    ws.row_dimensions[row].height = 18


def band(ws, row, text, span=12):
    """分节色带(隔断行,中蓝底白粗字)。"""
    ws.cell(row, 1, text).font = Font(color="FFFFFF", bold=True)
    for c in range(1, span + 1):
        ws.cell(row, c).fill = BAND
    ws.row_dimensions[row].height = 15


def inp(ws, c, v, src=None, fmt=N1):
    """硬编码输入(蓝字)。★ 不生成单元格批注(src 仅作代码内文档)——解释走可见列。"""
    if fmt is None:
        fmt = N1
    x = ws[c]
    x.value = v
    x.font = BLUE
    x.number_format = fmt
    return x


def introw(ws, row, cols, vals, src=None, fmt=N1):
    """整行输入(蓝字, 无批注)。"""
    for col, v in zip(cols, vals):
        if v is not None:
            inp(ws, f"{col}{row}", v, None, fmt)


def fml(ws, c, f, fmt=N1, link=False):
    """公式格。link=True → 绿字(跨表链接);否则黑字(本表公式)。"""
    x = ws[c]
    x.value = f
    x.font = GREEN if link else BLACK
    x.number_format = fmt
    return x


def lab(ws, c, t, b=False, note=False):
    """文字标签。b=加粗;note=灰色(次要/派生行名, 非斜体)。开头 '=' 自动转义。"""
    x = ws[c]
    x.value = _esc(t)
    x.font = BF if b else (GREY if note else BLACK)
    return x


def logic(ws, c, t):
    """可见逻辑/来源列:正常黑色、wrap;去 markdown 粗体;句末自动分行;'=' 转义。"""
    if t:
        t = _esc(t).replace("**", "")
        t = _re.sub("。(?=\\S)", "。\n", t)
    ws[c].value = t
    ws[c].font = BLACK
    ws[c].alignment = Alignment(wrap_text=True, vertical="top")


def txt(ws, r, t, span_cols, h=30, bold=False):
    """wrap 文本块:横向 merge 到 span_cols 最后一列。"""
    last = span_cols[-1]
    ws.merge_cells(f"A{r}:{last}{r}")
    c = ws[f"A{r}"]
    c.value = _esc(t)
    c.font = BF if bold else BLACK
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if h:
        ws.row_dimensions[r].height = h


def mtext(ws, r, t, last_col, nrows=2):
    """merge 一段说明文字(A{r} 跨到 last_col, 占 nrows 行), 行高不锁。返回下一可用行。"""
    logic(ws, f"A{r}", t)
    ws.merge_cells(f"A{r}:{last_col}{r + nrows - 1}")
    ws[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    return r + nrows


def set_widths(ws, a_width, year_cols, year_width, logic_col=None, logic_width=56,
               cur_col=None, cur_width=14):
    ws.column_dimensions["A"].width = a_width
    for col in year_cols:
        ws.column_dimensions[col].width = year_width
    if logic_col:
        ws.column_dimensions[logic_col].width = logic_width
    if cur_col:
        ws.column_dimensions[cur_col].width = cur_width


def nested_if_switch(sw_cell, case_names):
    """案序号派生公式:嵌套 IF(WPS/locale 兼容)。禁用 MATCH+数组常量 {"..."}.
    e.g. =IF(B2="Bear",1,IF(B2="Base",2,3))"""
    n = len(case_names)
    f = str(n)
    for i in range(n - 2, -1, -1):
        f = f'IF({sw_cell}="{case_names[i]}",{i + 1},{f})'
    return "=" + f


_YR = _re.compile(r"^20\d\d[AE]?$")


def finalize(wb, freeze=None, default_freeze="B3"):
    """全局格式 pass:年份表头右对齐 + 每张 sheet 冻结窗格(表头行+最左列)。
    freeze: {sheet名: "B4", ...} 覆盖默认。"""
    freeze = freeze or {}
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=6):
            for cell in row:
                if isinstance(cell.value, str) and _YR.match(cell.value.strip()):
                    cell.alignment = Alignment(horizontal="right")
        ws.freeze_panes = freeze.get(ws.title, default_freeze)


# ════════════════════════════ sheet 骨架函数 ════════════════════════════
# 约定:所有 write_* 吃一个已 create 好的 worksheet + data dict,就地写入,
# 返回"行号锚 dict"(供下游 sheet 跨表引用)。调用顺序 = DAG 顺序(见文件头)。

def write_cover(ws, data):
    """sheet 0 — 封面(报告日期 + 时效性 + Key Takeaways)。05 §0。

    data: title / meta=[(k,v),...] / takeaways=[(k,desc),...] / span(默认6)
    """
    span = data.get("span", 6)
    ws["A1"] = data["title"]
    ws["A1"].font = Font(bold=True, size=18, color="305496")
    ws.merge_cells(f"A1:{chr(64 + span)}1")
    ws.row_dimensions[1].height = 28

    def cov_row(k, v, h=None):
        rr = ws.max_row + 1
        ws[f"A{rr}"] = k
        ws[f"A{rr}"].font = BF
        ws[f"A{rr}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"B{rr}"] = _esc(v)
        ws[f"B{rr}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f"B{rr}:{chr(64 + span)}{rr}")
        if h:
            ws.row_dimensions[rr].height = h

    for k, v in data["meta"]:
        cov_row(k, v)
    rr = ws.max_row + 2
    ws[f"A{rr}"] = "Key Takeaways"
    ws[f"A{rr}"].font = HF
    for c in range(1, span + 1):
        ws.cell(rr, c).fill = HFILL
    for kt, desc in data["takeaways"]:
        cov_row(kt, desc, h=46)
    ws.column_dimensions["A"].width = 22
    for c in range(2, span + 1):
        ws.column_dimensions[chr(64 + c)].width = 22


def write_history(ws, data):
    """sheet 1 — 历史财务与估值(分部+盈利+权益+逐年股本/FX+年末/年均价+实际P/E·P/B+市值)。

    data 关键字段:
      title / hist_cols / hist_years / fx_hist(逐年均值汇率) / fx_now
      cur_label(G 列头, 默认"当下(TTM/最新)")
      segments [(名称, 兆本币逐年, 是否算YoY), ...]   (vals_in_usd=True 则视为 $B 不换算)
      total_now($B) / gm_pct(毛利率逐年, 可选) / gm_now
      ni / ni_now / ni_now_annualize(单季×4)
      eq / eq_now
      shares(逐年 mn股 list) / shares_now
      px_end(年末价逐年, 本币) / px_now / px_now_label
      px_avg(年均价逐年, 可选)
      quarter(可选, 最新季报列): {col,label,segs:{名称:(usd值,yoy)},ni,eq,shares,fx,note}
      band_note / notes [(行锚名 or 行号, 文本), ...](备注/来源可见列)
      note_col(默认 hist_cols 后第3列起算; 显式给如 "I")

    返回 anchor dict: seg_rows/HREV/HGM/HGMP/HNI/HEQ/HSH/HFX/HPX/HPXA/HEPS/HBPS/HPE/HPB/HMC
    """
    HC, HY = data["hist_cols"], data["hist_years"]
    fx_hist = data["fx_hist"]
    in_usd = data.get("vals_in_usd", False)
    q = data.get("quarter")
    span = len(HC) + (4 if q else 3)
    NOTE = data.get("note_col", "I" if q else "H")
    # 单位/币种扩展(通用): ps_scale = 基本面单位→每股的换算因子
    #   $B + mn股 → 1000;  亿(本币) + mn股 → 100(此时 fx 全 1, 单币种模型)
    PS = data.get("ps_scale", 1000)
    MD = data.get("mcap_div", 1000)          # 市值 = px*sh/fx/MD ($B→1000; 亿→100)
    UNIT = data.get("unit_label", "($B)")
    MCL = data.get("mcap_label", "市值($B)")
    FXL = data.get("fx_label", "FX (本币/USD,年均)")

    def cv(vals):
        return vals if in_usd else [uy(v, fx_hist[j]) for j, v in enumerate(vals)]

    hdr(ws, 1, data["title"], span)
    ws["A2"] = UNIT; ws["A2"].font = BF
    for col, y in zip(HC, HY):
        ws[f"{col}2"] = y; ws[f"{col}2"].font = BF; ws[f"{col}2"].fill = CH
    ws["G2"] = data.get("cur_label", "当下(TTM/最新)")
    ws["G2"].font = BF; ws["G2"].fill = CUR
    if q:
        ws[f"{q['col']}2"] = q["label"]; ws[f"{q['col']}2"].font = BF; ws[f"{q['col']}2"].fill = CUR
    ws[f"{NOTE}2"] = "备注/来源"; ws[f"{NOTE}2"].font = BF; ws[f"{NOTE}2"].fill = CH

    A = {}
    band(ws, 3, "分部营收(估计)", span)
    r = 4
    seg_rows = {}
    total_terms = []
    for name, vals, do_yoy in data["segments"]:
        lab(ws, f"A{r}", name)
        introw(ws, r, HC, cv(vals))
        seg_rows[name] = r
        total_terms.append(r)
        if q and name in q.get("segs", {}):
            sv, syoy = q["segs"][name]
            inp(ws, f"{q['col']}{r}", sv, None, N1)
            if do_yoy and syoy is not None:
                inp(ws, f"{q['col']}{r+1}", syoy, None, PCT)
        if do_yoy:
            lab(ws, f"A{r+1}", "  YoY", note=True)
            for j in range(1, len(HC)):
                fml(ws, f"{HC[j]}{r+1}", f'=IFERROR({HC[j]}{r}/{HC[j-1]}{r}-1,"n.m.")', PCT)
            r += 2
        else:
            r += 1
    HREV = r
    lab(ws, f"A{HREV}", "总营收", b=True)
    for col in HC:
        fml(ws, f"{col}{HREV}", "=" + "+".join(f"{col}{rr}" for rr in total_terms), N1)
    ws[f"A{HREV}"].border = BORD
    if data.get("total_now") is not None:
        inp(ws, f"G{HREV}", data["total_now"], None, N1)
    if q:
        fml(ws, f"{q['col']}{HREV}", "=" + "+".join(f"{q['col']}{rr}" for rr in total_terms), N1)
    r += 1

    band(ws, r, "分部占比 (% 总营收)", span); r += 1
    for name in seg_rows:
        lab(ws, f"A{r}", f"{name} 占比", note=True)
        for col in HC:
            fml(ws, f"{col}{r}", f"={col}{seg_rows[name]}/{col}{HREV}", PCT)
        r += 1

    band(ws, r, "盈利与权益", span); r += 1
    HGM = HGMP = None
    if data.get("gm_pct"):
        HGMP = r
        lab(ws, f"A{HGMP}", "毛利率(%)")
        introw(ws, HGMP, HC, data["gm_pct"], None, PCT)
        if data.get("gm_now") is not None:
            inp(ws, f"G{HGMP}", data["gm_now"], None, PCT)
        r += 1
        HGM = r
        lab(ws, f"A{HGM}", "毛利($B)")
        for col in HC:
            fml(ws, f"{col}{HGM}", f"={col}{HREV}*{col}{HGMP}", N1)
        fml(ws, f"G{HGM}", f"=G{HREV}*G{HGMP}", N1)
        r += 1
    HNI = r
    introw(ws, HNI, HC, cv(data["ni"]))
    lab(ws, f"A{HNI}", "净利润")
    if data.get("ni_now") is not None:
        inp(ws, f"G{HNI}", data["ni_now"], None, N1)
    if q:
        inp(ws, f"{q['col']}{HNI}", q["ni"], None, N1)
    r += 1
    lab(ws, f"A{r}", "净利率", note=True)
    for col in HC:
        fml(ws, f"{col}{r}", f"={col}{HNI}/{col}{HREV}", PCT)
    fml(ws, f"G{r}", f"=G{HNI}/G{HREV}", PCT)
    if q:
        fml(ws, f"{q['col']}{r}", f"={q['col']}{HNI}/{q['col']}{HREV}", PCT)
    r += 1
    HEQ = r
    introw(ws, HEQ, HC, cv(data["eq"]))
    lab(ws, f"A{HEQ}", "股东权益")
    if data.get("eq_now") is not None:
        inp(ws, f"G{HEQ}", data["eq_now"], None, N1)
    if q:
        inp(ws, f"{q['col']}{HEQ}", q["eq"], None, N1)
    r += 1
    lab(ws, f"A{r}", "ROE", note=True)
    for i, col in enumerate(HC):
        f = (f"={col}{HNI}/{col}{HEQ}" if i == 0
             else f"={col}{HNI}/AVERAGE({HC[i-1]}{HEQ},{col}{HEQ})")
        fml(ws, f"{col}{r}", f, PCT)
    r += 1
    HSH = r
    introw(ws, HSH, HC, data["shares"], None, N0)
    lab(ws, f"A{HSH}", "股本(mn股)")
    if data.get("shares_now") is not None:
        inp(ws, f"G{HSH}", data["shares_now"], None, N0)
    if q:
        inp(ws, f"{q['col']}{HSH}", q.get("shares", data["shares"][-1]), None, N0)
    r += 1
    HFX = r
    introw(ws, HFX, HC, fx_hist, None, N0)
    lab(ws, f"A{HFX}", FXL)
    inp(ws, f"G{HFX}", data["fx_now"], None, N0)
    if q:
        inp(ws, f"{q['col']}{HFX}", q.get("fx", data["fx_now"]), None, N0)
    r += 1

    band(ws, r, "市场估值 (历史=年末价→实际P/E·P/B; 当下列=TTM锚)", span); r += 1
    HPX = r
    introw(ws, HPX, HC, data["px_end"], None, PX)
    lab(ws, f"A{HPX}", "年末股价(本币)")
    inp(ws, f"G{HPX}", data["px_now"], None, PX); ws[f"G{HPX}"].fill = CUR
    r += 1
    HPXA = None
    if data.get("px_avg"):
        HPXA = r
        introw(ws, HPXA, HC, data["px_avg"], None, PX)
        lab(ws, f"A{HPXA}", "年均股价(本币)")
        r += 1
    HEPS = r
    lab(ws, f"A{HEPS}", data.get("eps_label", "EPS ($)"), note=True)
    for col in HC:
        fml(ws, f"{col}{HEPS}", f"={col}{HNI}*{PS}/{col}{HSH}", N2)
    annualize = "*4" if data.get("ni_now_annualize") else ""
    fml(ws, f"G{HEPS}", f"=G{HNI}{annualize}*{PS}/G{HSH}", N2)
    r += 1
    HBPS = r
    lab(ws, f"A{HBPS}", data.get("bps_label", "BPS ($)"), note=True)
    for col in HC:
        fml(ws, f"{col}{HBPS}", f"={col}{HEQ}*{PS}/{col}{HSH}", N2)
    fml(ws, f"G{HBPS}", f"=G{HEQ}*{PS}/G{HSH}", N2)
    r += 1
    HPE = r
    lab(ws, f"A{HPE}", "P/E (实际)", b=True)
    for col in HC:
        fml(ws, f"{col}{HPE}", f'=IF({col}{HNI}<=0,"N/M",{col}{HPX}/({col}{HEPS}*{col}{HFX}))', MX)
    fml(ws, f"G{HPE}", f"=G{HPX}/(G{HEPS}*G{HFX})", MX)
    r += 1
    HPB = r
    lab(ws, f"A{HPB}", "P/B (实际)", b=True); ws[f"A{HPB}"].fill = OUT
    for col in HC:
        fml(ws, f"{col}{HPB}", f"={col}{HPX}/({col}{HBPS}*{col}{HFX})", MX)
    fml(ws, f"G{HPB}", f"=G{HPX}/(G{HBPS}*G{HFX})", MX); ws[f"G{HPB}"].fill = CUR
    r += 1
    HMC = r
    lab(ws, f"A{HMC}", MCL, b=True)
    for col in HC:
        fml(ws, f"{col}{HMC}", f"={col}{HPX}*{col}{HSH}/{col}{HFX}/{MD}", N0)
    fml(ws, f"G{HMC}", f"=G{HPX}*G{HSH}/G{HFX}/{MD}", N0); ws[f"G{HMC}"].fill = CUR
    r += 1
    if data.get("band_note"):
        lab(ws, f"A{r}", "历史带读法", note=True)
        lab(ws, f"B{r}", data["band_note"], note=True)
        r += 1

    A = {"seg_rows": seg_rows, "HREV": HREV, "HGM": HGM, "HGMP": HGMP, "HNI": HNI,
         "HEQ": HEQ, "HSH": HSH, "HFX": HFX, "HPX": HPX, "HPXA": HPXA,
         "HEPS": HEPS, "HBPS": HBPS, "HPE": HPE, "HPB": HPB, "HMC": HMC}
    for key, note in data.get("notes", []):
        rr = A.get(key) if isinstance(key, str) else key
        if isinstance(rr, dict):
            continue
        if rr is None and isinstance(key, str):
            rr = seg_rows.get(key)
        if rr:
            logic(ws, f"{NOTE}{rr}", note)
    set_widths(ws, 16, HC, 9, logic_col=NOTE, logic_width=60, cur_col="G", cur_width=14)
    if q:
        ws.column_dimensions[q["col"]].width = 13
    if q and q.get("note"):
        logic(ws, f"{NOTE}{HREV}", q["note"])
    return A


def write_price_chart(ws, monthly_px, phases, title="月度股价"):
    """sheet 2 — 股价走势(月度收盘 + 阶段 + LineChart)。

    monthly_px : [(ym, price), ...] 真实月度收盘序列(单一价格来源, 上游统一定义)
    phases     : {"fn": ym->label, "rows": [(label, desc), ...]}
    返回 {yend, yavg, yhigh, ylow}(逐年末/均/高/低价, 供历史财务与估值倍数页)。
    """
    phase_fn = phases.get("fn") if isinstance(phases, dict) else None
    phase_rows = phases.get("rows", []) if isinstance(phases, dict) else phases
    hdr(ws, 1, title, 6)
    ws["A3"] = "月份"; ws["B3"] = "收盘价"; ws["C3"] = "阶段"
    for c in ["A3", "B3", "C3"]:
        ws[c].font = BF; ws[c].fill = CH
    for i, (ym, p) in enumerate(monthly_px, start=4):
        ws[f"A{i}"] = ym
        x = ws[f"B{i}"]; x.value = p; x.number_format = PX
        if phase_fn:
            lab(ws, f"C{i}", phase_fn(ym), note=True)
    n = len(monthly_px)
    ch = LineChart(); ch.title = title; ch.height = 9; ch.width = 24
    ch.add_data(Reference(ws, min_col=2, min_row=3, max_row=3 + n), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=4, max_row=3 + n))
    ch.y_axis.numFmt = PX
    ws.add_chart(ch, "E3")
    if phase_rows:
        ws["E26"] = "阶段说明"; ws["E26"].font = BF
        for i, (a, b) in enumerate(phase_rows, start=27):
            lab(ws, f"E{i}", a, b=True); lab(ws, f"F{i}", b, note=True)
    ws.column_dimensions["A"].width = 9; ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["E"].width = 24; ws.column_dimensions["F"].width = 58

    yend, ysum, ycnt, yhigh, ylow = {}, {}, {}, {}, {}
    for ym, p in monthly_px:
        y = ym[:4]
        yend[y] = p
        ysum[y] = ysum.get(y, 0) + p; ycnt[y] = ycnt.get(y, 0) + 1
        yhigh[y] = max(yhigh.get(y, p), p); ylow[y] = min(ylow.get(y, p), p)
    yavg = {y: round(ysum[y] / ycnt[y]) for y in ysum}
    return {"yend": yend, "yavg": yavg, "yhigh": yhigh, "ylow": ylow}


def write_consensus(ws, data):
    """sheet 3 — 卖方研报共识:说人话 + 按"模型假设"组织(卖方对账单, 05 §4d)。

    data: title / overview(一句话总览) /
          assumptions [(模型假设, 街上共识, 主要分歧, 我们取了什么+为什么), ...] /
          divergences [决定上下限的最大分歧 text, ...] /
          stances [各家一句话立场 text, ...]
    """
    span_cols = ["B", "C", "D", "E", "F", "G", "H", "I"]
    hdr(ws, 1, data["title"], 9)
    r = 2
    txt(ws, r, data["overview"], span_cols, 56); r += 1
    band(ws, r, "★ 模型每个关键假设 × 卖方怎么看(本表核心 = 支撑后面的测算)", 9); r += 1
    for col, h in zip("ABCD", ["模型假设", "街上共识(说人话)", "主要分歧", "我们模型取了什么 + 为什么"]):
        ws[f"{col}{r}"] = h
        ws[f"{col}{r}"].font = BF; ws[f"{col}{r}"].fill = CH
        ws[f"{col}{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
    for a, b, c, d in data["assumptions"]:
        for col, val in zip("ABCD", [a, b, c, d]):
            ws[f"{col}{r}"] = _esc(val)
            ws[f"{col}{r}"].alignment = Alignment(wrap_text=True, vertical="top")
            ws[f"{col}{r}"].font = BF if col == "A" else BLACK
        ws.row_dimensions[r].height = 58
        r += 1
    band(ws, r, "★ 最大的分歧(决定估值上下限)", 9); r += 1
    for t in data["divergences"]:
        txt(ws, r, t, span_cols, 40); r += 1
    band(ws, r, "各家一句话立场(评级 | 目标价 | 核心观点)", 9); r += 1
    for t in data["stances"]:
        txt(ws, r, t, span_cols, 30); r += 1
    for col, w in zip("ABCDEFGHI", [16, 42, 30, 44, 8, 8, 8, 8, 8]):
        ws.column_dimensions[col].width = w


def write_hist_multiples(ws, data):
    """sheet 4 — 历史估值倍数(数据底座):自身逐年带+年内高低+当下 + 同业对照 + 相对核心同行比值。

    data: title / intro / s_hist / ha(write_history 返回) / hist_cols / hist_years /
          yhigh, ylow({year:price}, 真实月度序列取的年内高低) /
          fwd_note(当下 forward 一句结论, 公式在下游表) /
          self_name / self_note /
          peers [{name, yearly(list, None→"n.a."), cur_pb, cur_pe, fwd_pe, note}, ...]
              yearly=None 整行 "—"; 数值缺标 n.a. 不硬补
          ratio {peer(名), note}   相对核心同行 P/B 比值行(结构溢价对账线)
          reading(③读法 — 给估值倍数假设的输入)
    返回 {own_pb, peer_rows{name:row}, ratio_row}
    """
    HC, HY = data["hist_cols"], data["hist_years"]
    S_HIST, ha = data["s_hist"], data["ha"]
    L = data.get("logic_col", "L")
    hdr(ws, 1, data["title"], 11)
    hr = mtext(ws, 2, data["intro"], "K", 2)
    for col, w in zip("ABCDEFGHIJK", [24, 10, 10, 10, 10, 10, 11, 11, 11, 9, 9]):
        ws.column_dimensions[col].width = w
    ws.column_dimensions[L].width = 58

    band(ws, hr, "① 自身: 逐年实际倍数 + 年内高低带 + 当下", 11); hr += 1
    ws[f"A{hr}"] = "指标"; ws[f"A{hr}"].font = BF
    for col, y in zip(HC, HY):
        ws[f"{col}{hr}"] = y; ws[f"{col}{hr}"].font = BF; ws[f"{col}{hr}"].fill = CH
    ws[f"G{hr}"] = "当下"; ws[f"G{hr}"].font = BF; ws[f"G{hr}"].fill = CUR
    ws[f"{L}{hr}"] = "说明"; ws[f"{L}{hr}"].font = BF; ws[f"{L}{hr}"].fill = CH
    hr += 1
    rows_link = [("年末股价(本币)", ha["HPX"], PX), ("年末 P/E(实际)", ha["HPE"], MX),
                 ("年末 P/B(实际)", ha["HPB"], MX)]
    OWN_PB = None
    for i, (nm, srow, fmt) in enumerate(rows_link):
        lab(ws, f"A{hr}", nm, b=(i == 2))
        for col in HC:
            fml(ws, f"{col}{hr}", f"={R(S_HIST, col + str(srow))}", fmt, link=True)
        fml(ws, f"G{hr}", f"={R(S_HIST, 'G' + str(srow))}", fmt, link=True)
        ws[f"G{hr}"].fill = CUR
        if i == 2:
            OWN_PB = hr
        if data.get("self_band_notes"):
            logic(ws, f"{L}{hr}", data["self_band_notes"][i])
        hr += 1
    HIGH = hr
    lab(ws, f"A{hr}", "年内最高价(月末)")
    for col, y in zip(HC, HY):
        if y in data["yhigh"]:
            inp(ws, f"{col}{hr}", data["yhigh"][y], None, PX)
    logic(ws, f"{L}{hr}", "由真实月度收盘序列取年内高/低(月末收盘口径, 非盘中极值)。")
    hr += 1
    LOW = hr
    lab(ws, f"A{hr}", "年内最低价(月末)")
    for col, y in zip(HC, HY):
        if y in data["ylow"]:
            inp(ws, f"{col}{hr}", data["ylow"][y], None, PX)
    hr += 1
    lab(ws, f"A{hr}", "P/B 年内高")
    for col in HC:
        fml(ws, f"{col}{hr}", f"={col}{HIGH}/({R(S_HIST, col + str(ha['HBPS']))}*{R(S_HIST, col + str(ha['HFX']))})", MX, link=True)
    logic(ws, f"{L}{hr}", "= 年内最高价 ÷ (当年每股净资产 × 当年汇率) → 历史估值带的上沿。")
    hr += 1
    lab(ws, f"A{hr}", "P/B 年内低")
    for col in HC:
        fml(ws, f"{col}{hr}", f"={col}{LOW}/({R(S_HIST, col + str(ha['HBPS']))}*{R(S_HIST, col + str(ha['HFX']))})", MX, link=True)
    logic(ws, f"{L}{hr}", "历史估值带的下沿——目标倍数应落在带内, 突破要写明理由(见下一页三层)。")
    hr += 1
    if data.get("fwd_note"):
        lab(ws, f"A{hr}", "当下 forward(指引)")
        ws.merge_cells(f"B{hr}:G{hr}")
        lab(ws, f"B{hr}", data["fwd_note"], note=True)
        logic(ws, f"{L}{hr}", "forward 要用模型前瞻每股(在后面的表), 公式放『情景估值』; 此处只放结论。")
        hr += 1
    hr += 1

    band(ws, hr, "② 同业对照: 年末 P/B 逐年 + 当下 TTM + forward P/E (实拉)", 11); hr += 1
    ws[f"A{hr}"] = "公司"; ws[f"A{hr}"].font = BF
    for col, y in zip(HC, HY):
        ws[f"{col}{hr}"] = y; ws[f"{col}{hr}"].font = BF; ws[f"{col}{hr}"].fill = CH
    for col, h in zip("GHI", ["当下P/B(TTM)", "当下P/E(TTM)", "forward P/E"]):
        ws[f"{col}{hr}"] = h
        ws[f"{col}{hr}"].font = BF; ws[f"{col}{hr}"].fill = CH
        ws[f"{col}{hr}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws[f"{L}{hr}"] = "业务特征 / 来源"; ws[f"{L}{hr}"].font = BF; ws[f"{L}{hr}"].fill = CH
    hr += 1
    SELF = hr
    ws[f"A{hr}"] = data["self_name"]; ws[f"A{hr}"].font = BF; ws[f"A{hr}"].fill = CUR
    for col in HC:
        fml(ws, f"{col}{hr}", f"={col}{OWN_PB}", MX, link=True)
    fml(ws, f"G{hr}", f"={R(S_HIST, 'G' + str(ha['HPB']))}", MX, link=True)
    fml(ws, f"H{hr}", f"={R(S_HIST, 'G' + str(ha['HPE']))}", MX, link=True)
    if data.get("self_fwd_pe_label"):
        lab(ws, f"I{hr}", data["self_fwd_pe_label"], note=True)
    logic(ws, f"{L}{hr}", data.get("self_note", ""))
    hr += 1
    peer_rows = {}
    for p in data["peers"]:
        ws[f"A{hr}"] = p["name"]; ws[f"A{hr}"].font = BF
        if p.get("yearly") is not None:
            for col, v in zip(HC, p["yearly"]):
                if v is not None:
                    inp(ws, f"{col}{hr}", v, None, MX)
                else:
                    lab(ws, f"{col}{hr}", "n.a.", note=True)
        else:
            for col in HC:
                lab(ws, f"{col}{hr}", "—", note=True)
        for col, key in zip("GHI", ["cur_pb", "cur_pe", "fwd_pe"]):
            v = p.get(key)
            if v is not None:
                inp(ws, f"{col}{hr}", v, None, MX)
            else:
                lab(ws, f"{col}{hr}", "—", note=True)
        logic(ws, f"{L}{hr}", p.get("note", ""))
        peer_rows[p["name"]] = hr
        hr += 1
    RATIO = None
    if data.get("ratio"):
        RATIO = hr
        rp = data["ratio"]
        prow = peer_rows[rp["peer"]]
        lab(ws, f"A{hr}", f"{data['self_name']} / {rp['peer'].split(' ')[0]} P/B 比值", b=True)
        ws[f"A{hr}"].fill = OUT
        peer_yearly = next(p.get("yearly") for p in data["peers"] if p["name"] == rp["peer"])
        for col, v in zip(HC, peer_yearly or []):
            if v is not None:
                fml(ws, f"{col}{hr}", f"={col}{SELF}/{col}{prow}", N2, link=True)
            else:
                lab(ws, f"{col}{hr}", "n.a.", note=True)
        fml(ws, f"G{hr}", f"=G{SELF}/G{prow}", N2, link=True)
        logic(ws, f"{L}{hr}", rp.get("note", ""))
        hr += 1
    hr += 1
    band(ws, hr, "③ 读法 — 给『估值倍数假设』的输入", 11); hr += 1
    mtext(ws, hr, data["reading"], "K", 4)
    return {"own_pb": OWN_PB, "peer_rows": peer_rows, "ratio_row": RATIO}


def write_multiple_assumptions(ws, data):
    """sheet 5 — 估值倍数假设(前置·自包含, 不引任何下游表):
    ①为什么主线镜头(业务判断长文) ②三层分解(峰值×结构溢价×情绪值)→三案目标倍数矩阵
    ③卖方对账 ④数据源。三案倍数在此拍;『情景切换』引用并切换, 『情景估值』套用。

    data: title / intro / why_text(镜头选择业务判断, 04 §镜头选择) / why_rows(merge 行数, 默认8)
          method_text(三层方法论 + 一致性检验) /
          peak(第一层常数) / peak_note / premium(第二层) / premium_note /
          all_cols / all_years / hist_cols / fwd_cols(不含基年) /
          s_hist / hpb_row(历史实际 P/B 行, 情绪值历史列反推用) /
          cases [(案名, 前瞻情绪值 list, 故事), ...] / sent_note / target_note /
          reconcile_text / source_text
    返回 {pk_cell, pr_cell, sent_row0, target_row0}(行号供情景切换/估值对比引用)
    """
    ALLC, ALLY = data["all_cols"], data["all_years"]
    HC, FCf = data["hist_cols"], data["fwd_cols"]
    S_HIST, HPB = data["s_hist"], data["hpb_row"]
    L = data.get("logic_col", "L")
    hdr(ws, 1, data["title"], 11)
    r = mtext(ws, 2, data["intro"], "K", 2)
    for col, w in zip("ABCDEFGHIJK", [24, 11, 10, 9, 9, 9, 9, 9, 9, 9, 9]):
        ws.column_dimensions[col].width = w
    ws.column_dimensions[L].width = 55

    band(ws, r, "① 为什么这个镜头做主线 — 镜头选择是业务判断(任何情景下支线镜头都要 double check)", 11); r += 1
    r = mtext(ws, r, data["why_text"], "K", data.get("why_rows", 8))

    band(ws, r, "② 三层分解 → 三案目标倍数(蓝字=情绪值; 三案在此拍, 『情景切换』负责切换)", 11); r += 1
    r = mtext(ws, r, data["method_text"], "K", 3)
    lab(ws, f"A{r}", "① 历史周期峰值")
    inp(ws, f"B{r}", data["peak"], None, MX)
    PK = f"$B${r}"
    ws.merge_cells(f"D{r}:K{r}"); logic(ws, f"D{r}", data["peak_note"]); r += 1
    lab(ws, f"A{r}", "② 结构溢价")
    inp(ws, f"B{r}", data["premium"], None, '0.00"x"')
    PR = f"$B${r}"
    ws.merge_cells(f"D{r}:K{r}"); logic(ws, f"D{r}", data["premium_note"]); r += 1
    ws[f"A{r}"] = "案 / 年份"; ws[f"A{r}"].font = BF
    for col, y in zip(ALLC, ALLY):
        ws[f"{col}{r}"] = y; ws[f"{col}{r}"].font = BF; ws[f"{col}{r}"].fill = CH
    ws[f"{L}{r}"] = "说明"; ws[f"{L}{r}"].font = BF; ws[f"{L}{r}"].fill = CH
    r += 1
    lab(ws, f"A{r}", "③ 情绪值(三案)", b=True)
    for _c in ["A"] + list(FCf):
        ws[f"{_c}{r}"].fill = GREYF
    logic(ws, f"{L}{r}", data.get("sent_note", "情绪值=周期/情绪位置。1.0=付足『峰值×溢价』; >1=FOMO 超涨; <1=下行折价。历史列=实际倍数÷(峰值×溢价)反推, 三案同值(一致性检验可见)。"))
    r += 1
    MSENT = r
    for cs, vals, story in data["cases"]:
        lab(ws, f"A{r}", f"  {cs}")
        for col in HC:
            fml(ws, f"{col}{r}", f'=IFERROR({R(S_HIST, col + str(HPB))}/({PK}*{PR}),"n.m.")', N2, link=True)
        for col, v in zip(FCf, vals):
            inp(ws, f"{col}{r}", v, None, N2)
        logic(ws, f"{L}{r}", story)
        r += 1
    lab(ws, f"A{r}", "三案目标倍数 = ①×②×③", b=True)
    for _c in ["A"] + list(FCf):
        ws[f"{_c}{r}"].fill = GREYF
    logic(ws, f"{L}{r}", data.get("target_note", "同一个三层公式套三案情绪。历史列 = 实际倍数(回看, 三案同值)。"))
    r += 1
    MPB3 = r
    for i, (cs, _v, _s) in enumerate(data["cases"]):
        lab(ws, f"A{r}", f"  {cs}")
        for col in HC:
            fml(ws, f"{col}{r}", f'=IFERROR({R(S_HIST, col + str(HPB))},"n.m.")', MX, link=True)
        for col in FCf:
            fml(ws, f"{col}{r}", f"={PK}*{PR}*{col}{MSENT + i}", MX, link=True)
        r += 1
    band(ws, r, "目标倍数: 卖方 vs 我们 — 凭什么敢给非主流数", 11); r += 1
    r = mtext(ws, r, data["reconcile_text"], "K", 4)
    band(ws, r, "数据源与口径", 11); r += 1
    mtext(ws, r, data["source_text"], "K", 4)
    return {"pk_cell": PK, "pr_cell": PR, "sent_row0": MSENT, "target_row0": MPB3}


def write_scenario_switch(ws, data):
    """sheet 6 — 情景切换(全模型唯一情景参数库 + 切换开关; 默认 Base)。04 §三情景。

    机制: B2 下拉(唯一入口) → 案序号(嵌套 IF 派生) → 各杠杆『当前案』行(INDEX) →
    整条明细链变档。矩阵布局: 杠杆名独立组标题行(指标本体描述) + 缩进三案行(各案故事,
    含历史实际列) + 当前案行。

    data: title / usage / cases(["Bear","Base","Bull"]) / default("Base") /
          triggers [(案, 触发条件), ...] /
          all_cols / all_years / hist_cols / fwd_cols(不含基年) /
          levers [{key, name, fmt, vals:{案:[...]}, desc(本体描述), stories:{案:故事},
                   cols(可选, 默认 fwd_cols), hist(可选, 历史实际 list, None项标 n.m.)}, ...]
          linked [{key, name, fmt, src_sheet, src_row0(三案首行), note}, ...]
              三案行=跨表链接(如情绪值链『估值倍数假设』), 当前案行照常 INDEX
          derived [{name, fmt, note, fml(col)->公式str}, ...](如 目标P/B(当前案))
    返回 {sw_cell, idx_cell, SWB{key:三案首行}, SWACT{key:当前案行}, derived_rows{name:row}, next_row}
    """
    ALLC, ALLY = data["all_cols"], data["all_years"]
    HC, FCf = data["hist_cols"], data["fwd_cols"]
    cases = data["cases"]
    L = data.get("logic_col", "L")
    hdr(ws, 1, data["title"], 11)
    lab(ws, "A2", "当前情景(下拉切换)", b=True)
    sw = ws["B2"]; sw.value = data.get("default", cases[1] if len(cases) > 1 else cases[0])
    sw.font = BLUE; sw.fill = CUR
    dv = DataValidation(type="list", formula1=f'"{",".join(cases)}"', allow_blank=False)
    ws.add_data_validation(dv); dv.add("B2")
    lab(ws, "C2", "案序号(由 B2 派生)", note=True)
    fml(ws, "D2", nested_if_switch("B2", cases), N0)
    SWIDX = "$D$2"
    swr = mtext(ws, 3, data["usage"], "K", 3)
    band(ws, swr, "三案触发条件(什么发生 = 落进这一档)", 11); swr += 1
    for cs, trig in data["triggers"]:
        lab(ws, f"A{swr}", cs, b=True)
        ws.merge_cells(f"B{swr}:K{swr}")
        logic(ws, f"B{swr}", trig)
        swr += 1
    band(ws, swr, "参数矩阵 — 可翻档杠杆 × 三案(蓝字在此改; 每杠杆末行 = 当前案, 明细链引它)", 11); swr += 1
    ws[f"A{swr}"] = "杠杆 / 案"; ws[f"A{swr}"].font = BF
    for col, y in zip(ALLC, ALLY):
        ws[f"{col}{swr}"] = y; ws[f"{col}{swr}"].font = BF; ws[f"{col}{swr}"].fill = CH
    ws[f"{L}{swr}"] = "说明"; ws[f"{L}{swr}"].font = BF; ws[f"{L}{swr}"].fill = CH
    swr += 1

    SWB, SWACT = {}, {}

    def _lever(lv):
        nonlocal swr
        cols = lv.get("cols") or FCf
        hist = lv.get("hist")
        desc = lv["desc"]
        if hist is not None:
            desc += " 历史列 = 实际值(历史只有一个, 三案同值; 仅供对照与回测)。"
        lab(ws, f"A{swr}", lv["name"], b=True)
        for _c in ["A"] + list(ALLC):
            ws[f"{_c}{swr}"].fill = GREYF
        logic(ws, f"{L}{swr}", desc)
        swr += 1
        for i, cs in enumerate(cases):
            lab(ws, f"A{swr}", f"  {cs}")
            if hist is not None:
                for col, v in zip(HC, hist):
                    if v is None:
                        lab(ws, f"{col}{swr}", "n.m.", note=True)
                    else:
                        inp(ws, f"{col}{swr}", v, None, lv["fmt"])
            for col, v in zip(cols, lv["vals"][cs]):
                if v is not None:
                    inp(ws, f"{col}{swr}", v, None, lv["fmt"])
            if i == 0:
                SWB[lv["key"]] = swr
            logic(ws, f"{L}{swr}", lv["stories"][cs])
            swr += 1
        lab(ws, f"A{swr}", "  当前案(链引此行)", note=True)
        allcols = ([c for c, v in zip(HC, hist) if v is not None] if hist is not None else []) + list(cols)
        for col in allcols:
            fml(ws, f"{col}{swr}", f"=INDEX({col}{SWB[lv['key']]}:{col}{SWB[lv['key']] + len(cases) - 1},{SWIDX})", lv["fmt"])
        SWACT[lv["key"]] = swr
        swr += 2

    for lv in data["levers"]:
        _lever(lv)
    for lk in data.get("linked", []):
        lab(ws, f"A{swr}", lk["name"], b=True)
        for _c in ["A"] + list(ALLC):
            ws[f"{_c}{swr}"].fill = GREYF
        logic(ws, f"{L}{swr}", lk["note"])
        swr += 1
        for i, cs in enumerate(cases):
            lab(ws, f"A{swr}", f"  {cs}")
            for col in ALLC:
                fml(ws, f"{col}{swr}", f"={R(lk['src_sheet'], col + str(lk['src_row0'] + i))}", lk["fmt"], link=True)
            if i == 0:
                SWB[lk["key"]] = swr
            swr += 1
        lab(ws, f"A{swr}", "  当前案(链引此行)", note=True)
        for col in ALLC:
            fml(ws, f"{col}{swr}", f"=INDEX({col}{SWB[lk['key']]}:{col}{SWB[lk['key']] + len(cases) - 1},{SWIDX})", lk["fmt"])
        SWACT[lk["key"]] = swr
        swr += 1
    derived_rows = {}
    for d in data.get("derived", []):
        lab(ws, f"A{swr}", d["name"], b=True)
        for col in ALLC:
            fml(ws, f"{col}{swr}", d["fml"](col), d["fmt"], link=True)
        logic(ws, f"{L}{swr}", d.get("note", ""))
        derived_rows[d["name"]] = swr
        swr += 1
    ws.column_dimensions["A"].width = 26
    for col in ALLC:
        ws.column_dimensions[col].width = 9
    ws.column_dimensions[L].width = 60
    return {"sw_cell": "B2", "idx_cell": SWIDX, "SWB": SWB, "SWACT": SWACT,
            "derived_rows": derived_rows, "next_row": swr}


def write_anchor(ws, data):
    """sheet 7 — 物理锚基座([ANCHOR] 标记必在 A1)。

    data: title / all_cols / all_years /
          series [(行名, 逐年值列表(None 项留给调用方回填公式, 如情景驱动的前瞻列),
                   来源说明, 数字格式), ...]
          yoy_row(给行名则其下加 YoY 行) / source_note / role_note
    返回 {row_of: {行名:row}}
    """
    AC, AY = data["all_cols"], data["all_years"]
    span = len(AC) + 1
    title = data["title"]
    if "[ANCHOR]" not in title:
        title = "[ANCHOR] " + title
    hdr(ws, 1, title, span)
    ws["A2"] = "年份"; ws["A2"].font = BF
    for col, y in zip(AC, AY):
        ws[f"{col}2"] = y; ws[f"{col}2"].font = BF; ws[f"{col}2"].fill = CH
    r = 3
    row_of = {}
    for name, vals, src, fmt in data["series"]:
        lab(ws, f"A{r}", name)
        introw(ws, r, AC, vals, None, fmt or N0)
        row_of[name] = r
        r += 1
        if data.get("yoy_row") == name:
            lab(ws, f"A{r}", "YoY", note=True)
            for i in range(1, len(AC)):
                fml(ws, f"{AC[i]}{r}", f'=IFERROR({AC[i]}{r-1}/{AC[i-1]}{r-1}-1,"n.m.")', PCT)
            r += 1
    r += 1
    band(ws, r, "口径与来源", span); r += 1
    if data.get("source_note"):
        r = mtext(ws, r, data["source_note"], AC[-1], 3)
    if data.get("role_note"):
        r = mtext(ws, r, data["role_note"], AC[-1], 2)
    set_widths(ws, 18, AC, 8)
    return {"row_of": row_of}


def write_segment_model(ws, data):
    """sheet 8 — 分部测算(capex×收入强度 / 周期段等物理传导)+ 可见逻辑列。

    通用机制: 本 helper 只铺 硬编码行 + 逻辑列, 值为 None 的行 = 公式行由调用方
    用返回的 row map 回填(各公司传导链不同, 不强行模板化)。
    ★ 建模方法见 anchors/<类型>.md(算力链 = AI 直驱段挂 capex×强度 + 消费段走周期;
      强度锚最近实际年反推、历史列公式反推、早年失真标 n.m.)。

    data: title / all_cols / all_years / logic_col /
          groups [(色带标题, [(行名, 逐年值或None, 数字格式, 逻辑句), ...]), ...]
    返回 {m: {行名:row}, next_row}
    """
    AC, AY = data["all_cols"], data["all_years"]
    LCOL = data.get("logic_col", "N")
    span = len(AC) + 1
    hdr(ws, 1, data.get("title", "分部测算"), span)
    ws["A2"] = "变量"; ws["A2"].font = BF
    for col, y in zip(AC, AY):
        ws[f"{col}2"] = y; ws[f"{col}2"].font = BF; ws[f"{col}2"].fill = CH
    ws[f"{LCOL}2"] = "逻辑/来源(整句)"; ws[f"{LCOL}2"].font = BF; ws[f"{LCOL}2"].fill = CH
    r = 3
    m = {}
    for band_title, rows in data["groups"]:
        band(ws, r, band_title, span); r += 1
        for name, vals, fmt, lg in rows:
            lab(ws, f"A{r}", name)
            if vals is not None:
                introw(ws, r, AC, vals, None, fmt or N1)
            m[name] = r
            logic(ws, f"{LCOL}{r}", lg)
            r += 1
    set_widths(ws, 24, AC, 8, logic_col=LCOL, logic_width=60)
    return {"m": m, "next_row": r}


def write_fundamentals(ws, data):
    """sheet 9 — 利润与收入假设:假设(其他增速/段OPM/净利转换/留存, 历史实际列打底)
    + 分部营收(链测算表) + 段驱动营业利润 → 净利 → 权益 → EPS/BPS/ROE。
    本表只到 EPS/BPS 为止, 不管估值倍数(那在『估值倍数假设』)。

    data: title / all_cols / all_years / hist_cols / fwd_cols(不含基年) / logic_col /
          s_hist / ha(write_history 返回, 需 HNI/HEQ/HSH + seg_rows) / share_fix_col(前瞻股本取该历史列)
          assum_groups [(色带, [{name, vals(与 all_cols 对齐, None跳过), fmt, logic,
                                 nm_cols(标 n.m. 的列), link:{sheet,row}(前瞻列改跨表链)}, ...]), ...]
          segments [{name, hist_row(历史财务 seg_rows 键), fwd:{'sheet','row'} 或 {'growth':假设名}}]
          profit_terms [([段名,...], OPM假设名, iferror), ...]  营业利润 = Σ 段收入×段OPM
          conv_assum(净利转换假设名) / retention_assum(留存率假设名)
          note_text(基本面口径说明)
    返回 {am, seg_rows, REV, OP, NI, EQ, EPS, BPS, ROE}
    """
    ALLC, ALLY = data["all_cols"], data["all_years"]
    HC, FCf = data["hist_cols"], data["fwd_cols"]
    LCOL = data.get("logic_col", "N")
    S_HIST, ha = data["s_hist"], data["ha"]
    span = len(ALLC) + 1
    fixc = data.get("share_fix_col", HC[-1])
    PS = data.get("ps_scale", 1000)

    def SHc(c):
        return R(S_HIST, (c if c in HC else fixc) + str(ha["HSH"]))

    hdr(ws, 1, data["title"], span)
    ws["A2"] = "假设"; ws["A2"].font = BF
    for col, y in zip(ALLC, ALLY):
        ws[f"{col}2"] = y; ws[f"{col}2"].font = BF; ws[f"{col}2"].fill = CH
    ws[f"{LCOL}2"] = "逻辑/来源(整句)"; ws[f"{LCOL}2"].font = BF; ws[f"{LCOL}2"].fill = CH
    r = 3
    am = {}
    for band_title, rows in data["assum_groups"]:
        band(ws, r, band_title, span); r += 1
        for row in rows:
            lab(ws, f"A{r}", row["name"])
            if row.get("vals") is not None:
                introw(ws, r, ALLC, row["vals"], None, row.get("fmt") or PCT)
            for cc in row.get("nm_cols", []):
                lab(ws, f"{cc}{r}", "n.m.", note=True)
            if row.get("link"):
                for cc in FCf:
                    fml(ws, f"{cc}{r}", f"={R(row['link']['sheet'], cc + str(row['link']['row']))}",
                        row.get("fmt") or PCT, link=True)
            am[row["name"]] = r
            logic(ws, f"{LCOL}{r}", row.get("logic", ""))
            r += 1

    def _yoy(row, dst):
        lab(ws, f"A{dst}", "  YoY", note=True)
        for i in range(1, len(ALLC)):
            # IFERROR 防零基年(新兴分部早年=0)0/0 除零 → "n.m."
            fml(ws, f"{ALLC[i]}{dst}", f'=IFERROR({ALLC[i]}{row}/{ALLC[i-1]}{row}-1,"n.m.")', PCT)

    band(ws, r, "分部营收预测(估计)", span); r += 1
    seg_rows = {}
    seg_pct_pending = []
    for seg in data["segments"]:
        rr = r
        lab(ws, f"A{rr}", seg["name"])
        hrow = ha["seg_rows"][seg["hist_row"]]
        for c in HC:
            fml(ws, f"{c}{rr}", f"={R(S_HIST, c + str(hrow))}", N1, link=True)
        fwd = seg["fwd"]
        if "sheet" in fwd:
            for c in FCf:
                fml(ws, f"{c}{rr}", f"={R(fwd['sheet'], c + str(fwd['row']))}", N1, link=True)
        else:
            grow = am[fwd["growth"]]
            prevs = [HC[-1]] + list(FCf[:-1])
            for p, c in zip(prevs, FCf):
                fml(ws, f"{c}{rr}", f"={p}{rr}*(1+{c}{grow})", N1)
        seg_rows[seg["name"]] = rr
        _yoy(rr, rr + 1)
        seg_pct_pending.append((seg["name"], rr, rr + 2))
        r += 3
    REV = r
    lab(ws, f"A{REV}", "总营收", b=True)
    for col in ALLC:
        fml(ws, f"{col}{REV}", "=" + "+".join(f"{col}{rr}" for _n, rr, _p in seg_pct_pending), N1)
    ws[f"A{REV}"].border = BORD
    _yoy(REV, REV + 1)
    r += 2
    for name, rev_r, pct_r in seg_pct_pending:
        lab(ws, f"A{pct_r}", "  占总营收%", note=True)
        for c in ALLC:
            fml(ws, f"{c}{pct_r}", f"={c}{rev_r}/{c}{REV}", PCT)

    band(ws, r, "盈利与账面", span); r += 1
    OP, OPM, NI, NM, EQ, EPS, BPS, ROE = r, r + 1, r + 2, r + 3, r + 4, r + 5, r + 6, r + 7
    r += 8
    lab(ws, f"A{OP}", "营业利润($B)")
    for c in ALLC:
        terms = []
        for segs, opm_key, iferr in data["profit_terms"]:
            sum_expr = "+".join(f"{c}{seg_rows[sn]}" for sn in segs)
            opm_ref = f"{c}{am[opm_key]}"
            t = f"({sum_expr})*{opm_ref}" if len(segs) > 1 else f"{c}{seg_rows[segs[0]]}*{opm_ref}"
            if iferr:
                t = f"IFERROR({t},0)"
            terms.append(t)
        fml(ws, f"{c}{OP}", "=" + "+".join(terms), N1)
    lab(ws, f"A{OPM}", "营业利润率(%)", note=True)
    for col in ALLC:
        fml(ws, f"{col}{OPM}", f"={col}{OP}/{col}{REV}", PCT)
    lab(ws, f"A{NI}", "净利润")
    for c in HC:
        fml(ws, f"{c}{NI}", f"={R(S_HIST, c + str(ha['HNI']))}", N1, link=True)
    for c in FCf:
        fml(ws, f"{c}{NI}", f"={c}{OP}*{c}{am[data['conv_assum']]}", N1)
    lab(ws, f"A{NM}", "净利率", note=True)
    for col in ALLC:
        fml(ws, f"{col}{NM}", f"={col}{NI}/{col}{REV}", PCT)
    lab(ws, f"A{EQ}", "期末权益")
    for c in HC:
        fml(ws, f"{c}{EQ}", f"={R(S_HIST, c + str(ha['HEQ']))}", N1, link=True)
    prevs = [HC[-1]] + list(FCf[:-1])
    for p, c in zip(prevs, FCf):
        fml(ws, f"{c}{EQ}", f"={p}{EQ}+{c}{NI}*{c}{am[data['retention_assum']]}", N1)
    lab(ws, f"A{EPS}", data.get("eps_label", "EPS ($)"))
    for col in ALLC:
        fml(ws, f"{col}{EPS}", f"={col}{NI}*{PS}/{SHc(col)}", N2)
    lab(ws, f"A{BPS}", data.get("bps_label", "BPS ($)"))
    for col in ALLC:
        fml(ws, f"{col}{BPS}", f"={col}{EQ}*{PS}/{SHc(col)}", N2)
    lab(ws, f"A{ROE}", "ROE", note=True)
    for i, col in enumerate(ALLC):
        fml(ws, f"{col}{ROE}",
            (f"={col}{NI}/{col}{EQ}" if i == 0 else f"={col}{NI}/AVERAGE({ALLC[i-1]}{EQ},{col}{EQ})"), PCT)
    band(ws, r, "基本面口径说明", span); r += 1
    if data.get("note_text"):
        r = mtext(ws, r, data["note_text"], "K", 3)
    set_widths(ws, 22, ALLC, 9, logic_col=LCOL, logic_width=56)
    return {"am": am, "seg_rows": seg_rows, "REV": REV, "OP": OP, "NI": NI,
            "EQ": EQ, "EPS": EPS, "BPS": BPS, "ROE": ROE}


def write_scenario_valuation(ws, data):
    """sheet 10 — 情景估值(当前案逐年):主线镜头(目标倍数×前瞻每股→隐含价) +
    P/E 交叉验证(隐含价÷前瞻EPS=隐含 forward PE + 当下 forward PE 证据行) +
    代表股价/市值(双币) + 方法与结论。历史列=实际价反推(事实), 前瞻=判断。

    data: title / intro / all_cols / all_years / hist_cols / fwd_cols(不含基年) / hist_years /
          fx_fwd / s_hist / ha(HPX/HFX/HSH) / s_fund / fr(write_fundamentals 返回) /
          s_switch / target_row(情景切换『目标P/B(当前案)』行) / sw_cell /
          yend, yavg({year:price}, 仅历史) / reading / method / concl
    返回 {pb, pb_px, ipe, fpe, mcap_usd, vs_now}
    """
    ALLC, ALLY = data["all_cols"], data["all_years"]
    HC, FCf, HY = data["hist_cols"], data["fwd_cols"], data["hist_years"]
    S_HIST, ha = data["s_hist"], data["ha"]
    S_FUND, fr = data["s_fund"], data["fr"]
    S_SW = data["s_switch"]
    fx_fwd = str(data["fx_fwd"])
    hdr(ws, 1, data["title"], 11)
    lab(ws, "L1", "当前情景→", note=True)
    fml(ws, "M1", f"={R(S_SW, data.get('sw_cell', 'B2'))}", N0, link=True)
    ws["M1"].fill = CUR
    r = mtext(ws, 2, data["intro"], "K", 1)
    ws[f"A{r}"] = "(本币/股; 市值双币)"; ws[f"A{r}"].font = BF
    for col, y in zip(ALLC, ALLY):
        ws[f"{col}{r}"] = y; ws[f"{col}{r}"].font = BF; ws[f"{col}{r}"].fill = CH
    r += 1

    def FXc(c):
        return R(S_HIST, c + str(ha["HFX"])) if c in HC else fx_fwd

    EPSr = lambda c: R(S_FUND, c + str(fr["EPS"]))
    BPSr = lambda c: R(S_FUND, c + str(fr["BPS"]))
    NIr = lambda c: R(S_FUND, c + str(fr["NI"]))

    band(ws, r, "主线镜头(历史=实际年末价反推; 前瞻=目标倍数(当前案)×前瞻每股→隐含价)", 11); r += 1
    VPB = r
    lab(ws, f"A{r}", "目标倍数 (历史=实际/前瞻=当前案)")
    for c in HC:
        fml(ws, f"{c}{r}", f'=IFERROR({R(S_HIST, c + str(ha["HPX"]))}/({BPSr(c)}*{FXc(c)}),"N/M")', MX, link=True)
    for c in FCf:
        fml(ws, f"{c}{r}", f"={R(S_SW, c + str(data['target_row']))}", MX, link=True)
    r += 1
    VPBP = r
    lab(ws, f"A{r}", "隐含股价 主线镜头(本币)", b=True); ws[f"A{r}"].fill = OUT
    for c in FCf:
        fml(ws, f"{c}{r}", f"={c}{VPB}*{BPSr(c)}*{FXc(c)}", PX)
    r += 1
    band(ws, r, "P/E 交叉验证(支线)(历史=实际; 前瞻=隐含价÷前瞻EPS=隐含 forward PE)", 11); r += 1
    VPE = r
    lab(ws, f"A{r}", "P/E (历史=实际 / 前瞻=隐含)")
    for c in HC:
        fml(ws, f"{c}{r}", f'=IF({NIr(c)}<=0,"N/M",{R(S_HIST, c + str(ha["HPX"]))}/({EPSr(c)}*{FXc(c)}))', MX, link=True)
    for c in FCf:
        fml(ws, f"{c}{r}", f"={c}{VPBP}/({EPSr(c)}*{FXc(c)})", MX, link=True)
    r += 1
    VFPE = r
    lab(ws, f"A{r}", "当下 forward P/E = 现价 ÷ 该年每股收益")
    for c in FCf:
        fml(ws, f"{c}{r}", f"={R(S_HIST, 'G' + str(ha['HPX']))}/({EPSr(c)}*{FXc(c)})", MX, link=True)
    r += 1
    if data.get("reading"):
        r = mtext(ws, r, data["reading"], "K", 2)
    band(ws, r, "代表股价 + 市值(主线)", 11); r += 1
    VNOW = r
    lab(ws, f"A{r}", "当下股价(本币)", b=True)
    g = fml(ws, f"G{r}", f"={R(S_HIST, 'G' + str(ha['HPX']))}", PX, link=True)
    ws[f"G{r}"].fill = CUR
    r += 1
    lab(ws, f"A{r}", "年末股价(本币, 历史)")
    for col, y in zip(HC, HY):
        if y in data["yend"]:
            inp(ws, f"{col}{r}", data["yend"][y], None, PX)
    r += 1
    lab(ws, f"A{r}", "年均股价(本币, 历史)")
    for col, y in zip(HC, HY):
        if y in data["yavg"]:
            inp(ws, f"{col}{r}", data["yavg"][y], None, PX)
    r += 1
    SHF = R(S_HIST, f"${data.get('share_fix_col', HC[-1])}${ha['HSH']}")
    MD1 = data.get("mcap_div", 1000000)
    VMCK = r
    lab(ws, f"A{r}", data.get("mcap_hist_label", "市值 实际年末(本币兆, 历史)"))
    for c in HC:
        fml(ws, f"{c}{r}", f'=IFERROR({R(S_HIST, c + str(ha["HPX"]))}*{R(S_HIST, c + str(ha["HSH"]))}/{MD1},"N/M")', N0)
    r += 1
    VMCKB = r
    lab(ws, f"A{r}", data.get("mcap_fwd_label", "市值 前瞻·主线(本币兆)"))
    for c in FCf:
        fml(ws, f"{c}{r}", f"={c}{VPBP}*{SHF}/{MD1}", N0)
    r += 1
    if data.get("mcap_usd_skip"):
        VMCU = VMCKB
    else:
        VMCU = r
        lab(ws, f"A{r}", "市值 ($B; 历史=实际, 前瞻=主线)", b=True)
        for c in HC:
            fml(ws, f"{c}{r}", f"={c}{VMCK}*1000/{FXc(c)}", N0)
        for c in FCf:
            fml(ws, f"{c}{r}", f"={c}{VMCKB}*1000/{fx_fwd}", N0)
        r += 1
    VUP = r
    lab(ws, f"A{r}", "前瞻隐含价(主线) vs 当下股价", b=True)
    for c in FCf:
        fml(ws, f"{c}{r}", f"={c}{VPBP}/$G${VNOW}-1", PCT)
    r += 1
    PEPX = None
    pl = data.get("pe_lens")
    if pl:
        band(ws, r, "P/E 平行镜头(双镜头标的: 隐含价 = 目标P/E(当前案) × 前瞻EPS, 与主线并列)", 11); r += 1
        PETR = r
        lab(ws, f"A{r}", "目标 P/E(当前案)")
        for c in FCf:
            fml(ws, f"{c}{r}", f"={R(data['s_switch'], c + str(pl['target_row']))}", MX, link=True)
        r += 1
        PEPX = r
        lab(ws, f"A{r}", "隐含股价 P/E镜头(本币)", b=True); ws[f"A{r}"].fill = OUT
        for c in FCf:
            fml(ws, f"{c}{r}", f"={c}{PETR}*{EPSr(c)}*{FXc(c)}", PX)
        r += 1
        lab(ws, f"A{r}", pl.get("mcap_label", "市值 前瞻·P/E镜头"))
        for c in FCf:
            fml(ws, f"{c}{r}", f"={c}{PEPX}*{SHF}/{MD1}", N0)
        r += 1
        lab(ws, f"A{r}", "前瞻隐含价(P/E镜头) vs 当下股价", b=True)
        for c in FCf:
            fml(ws, f"{c}{r}", f"={c}{PEPX}/$G${VNOW}-1", PCT)
        r += 1
    r += 1
    band(ws, r, "方法与结论", 11); r += 1
    if data.get("method"):
        r = mtext(ws, r, data["method"], "K", 3)
    if data.get("concl"):
        r = mtext(ws, r, data["concl"], "K", 3)
    ws.column_dimensions["A"].width = 20
    for col in ALLC:
        ws.column_dimensions[col].width = 11.5
    return {"pb": VPB, "pb_px": VPBP, "ipe": VPE, "fpe": VFPE,
            "mcap_usd": VMCU, "vs_now": VUP, "now_row": VNOW, "pe_px": PEPX}


def write_comparison(ws, data):
    """sheet 11 — 估值对比(三案恒常并排, 不随开关变):顶部三案汇总 + 每案逐年压缩链 block。
    ★ 防污染铁律(04): case 列只引『情景切换』矩阵行(各案行) + 链上未翻档行 + 静态历史锚,
      绝不引明细链上会随开关变的格(含『当前案』行/开关格)。Base 列与明细链同构(开关=Base 时一致)。
    ★ 每案 block 必有: 隐含价行(key='px') + 隐含 forward P/E 行(key='ipe', P/E 体检) +
      vs 行(key='up', 历史列=回测≈0 / 前瞻=vs 现价)。

    data: title / intro / case_names / all_cols / all_years / hist_cols / fwd_cols(不含基年) /
          block_start(默认16) /
          rows [{key, label, fmt, bold, out(高亮), hist(fn(col,ci,A)->'=..'|('inp',v)|None),
                 fwd(fn(col,j,ci,A)->'=..')}, ...]   A=该案已写行锚(写行前已含本行 key)
          summary {band, target_col, rows:[(label,key,fmt,note,bold)], mcap:{label,fml(case)->'=..',note},
                   concl}
    返回 {CMPA: {案名: {key:row}}}
    """
    ALLC, ALLY = data["all_cols"], data["all_years"]
    HC, FCf = data["hist_cols"], data["fwd_cols"]
    cases = data["case_names"]
    hdr(ws, 1, data["title"], 11)
    mtext(ws, 2, data["intro"], "K", 3)
    b0 = data.get("block_start", 16)
    rows = data["rows"]
    block_h = len(rows) + 3
    CMPA = {}
    for ci, cname in enumerate(cases):
        r0 = b0 + ci * block_h
        band(ws, r0, f"{cname} 案 — 逐年推演 (历史回测 + 前瞻)", 11)
        hrow = r0 + 1
        for col, y in zip(ALLC, ALLY):
            ws[f"{col}{hrow}"] = y; ws[f"{col}{hrow}"].font = BF; ws[f"{col}{hrow}"].fill = CH
        rr = hrow + 1
        A = {}
        for spec in rows:
            A[spec["key"]] = rr
            lab(ws, f"A{rr}", spec["label"], b=spec.get("bold", False))
            if spec.get("out"):
                ws[f"A{rr}"].fill = OUT
            if spec.get("hist"):
                for col in HC:
                    res = spec["hist"](col, ci, A)
                    if res is None:
                        continue
                    if isinstance(res, tuple) and res[0] == "inp":
                        inp(ws, f"{col}{rr}", res[1], None, spec["fmt"])
                    else:
                        fml(ws, f"{col}{rr}", res, spec["fmt"], link=res.count("'") > 0)
            if spec.get("fwd"):
                for j, col in enumerate(FCf):
                    res = spec["fwd"](col, j, ci, A)
                    if res is not None:
                        fml(ws, f"{col}{rr}", res, spec["fmt"], link=res.count("'") > 0)
            rr += 1
        CMPA[cname] = A
    # 顶部汇总
    sm = data["summary"]
    band(ws, 5, sm["band"], 11)
    hdr_cols = "ABCD"[:len(cases) + 1]
    for col, h in zip(hdr_cols, ["项目"] + list(cases)):
        ws[f"{col}6"] = h; ws[f"{col}6"].font = BF; ws[f"{col}6"].fill = CH
    ws.merge_cells("E6:K6"); ws["E6"] = "说明"; ws["E6"].font = BF; ws["E6"].fill = CH
    sr = 7
    tc = sm["target_col"]
    for label, key, fmt, note, bold in sm["rows"]:
        lab(ws, f"A{sr}", label, b=bold)
        for col, cname in zip(hdr_cols[1:], cases):
            fml(ws, f"{col}{sr}", f"={tc}{CMPA[cname][key]}", fmt)
        ws.merge_cells(f"E{sr}:K{sr}")
        lab(ws, f"E{sr}", note, note=True)
        sr += 1
    if sm.get("mcap"):
        mc = sm["mcap"]  # {label, key(引各案 block 行), expr(如 "*股本/FX/1000"), note}
        lab(ws, f"A{sr}", mc["label"])
        for col, cname in zip(hdr_cols[1:], cases):
            fml(ws, f"{col}{sr}", f"={tc}{CMPA[cname][mc['key']]}{mc['expr']}", N0)
        ws.merge_cells(f"E{sr}:K{sr}")
        lab(ws, f"E{sr}", mc.get("note", ""), note=True)
        sr += 1
    if sm.get("concl"):
        mtext(ws, sr, sm["concl"], "K", 2)
    ws.column_dimensions["A"].width = 24
    for col in ALLC:
        ws.column_dimensions[col].width = 9.5
    return {"CMPA": CMPA}


def write_dashboard(ws, data):
    """sheet 12(末张)— 综合判断仪表盘:A 基本面拐点 / B 估值错位(预测引擎) /
    C 催化剂 / D 情绪确认 + 综合判断一行 + 投后跟踪明细。05 §4c。
    验收 = 回测拐点:放回上一轮行情拐点, 这套表当时就能看到那波。

    data: title / usage /
          blocks [{title, rows:[(name, read, judge)]}, ...]
              read 可为 str | {"fml":公式,"fmt":fmt,"fill":bool} | {"inp":值,"fmt":fmt};
              verdict 行用 (name, read, judge, True) 四元组(加粗)
          final {band, text}
          tracking {intro, rows:[("__band__",标题) | (指标,当前值,为什么重要,怎么跟踪,触发动作)]}
    返回 {row_of:{name:row}}(B 块公式行可被引用)
    """
    hdr(ws, 1, data["title"], 5)
    dc = mtext(ws, 2, data["usage"], "E", 3)

    def dband(row, text):
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"].value = text
        ws[f"A{row}"].font = Font(color="FFFFFF", bold=True)
        for c in "ABCDE":
            ws[f"{c}{row}"].fill = BAND
        ws.row_dimensions[row].height = 15

    row_of = {}
    for blk in data["blocks"]:
        dband(dc, blk["title"]); dc += 1
        for item in blk["rows"]:
            name, read, judge = item[0], item[1], item[2]
            bold = len(item) > 3 and item[3]
            ws[f"A{dc}"] = name; ws[f"A{dc}"].font = BF
            if isinstance(read, dict):
                if "fml" in read:
                    f = read["fml"]
                    if callable(f):          # f(row_of) → 公式(可引用本表已写行, 如 GAP 行)
                        f = f(row_of)
                    fml(ws, f"B{dc}", f, read.get("fmt", MX), link=True)
                else:
                    inp(ws, f"B{dc}", read["inp"], None, read.get("fmt", MX))
                if read.get("fill"):
                    ws[f"B{dc}"].fill = CUR
            else:
                ws[f"B{dc}"] = _esc(read)
                ws[f"B{dc}"].font = BF if bold else BLACK
            ws.merge_cells(f"C{dc}:E{dc}")
            logic(ws, f"C{dc}", judge)
            for c in "AB":
                ws[f"{c}{dc}"].alignment = Alignment(wrap_text=True, vertical="top")
            row_of[name] = dc
            dc += 1
        dc += 1
    dband(dc, data["final"]["band"]); dc += 1
    dc = mtext(ws, dc, data["final"]["text"], "E", 3) + 1
    trk = data.get("tracking")
    if trk:
        dband(dc, "投后跟踪明细 — 哪个指标恶化 → 哪个假设先崩"); dc += 1
        if trk.get("intro"):
            dc = mtext(ws, dc, trk["intro"], "E", 3)
        heads = ["指标", "当前值/状态", "为什么重要(影响的关键假设)", "怎么跟踪(数据源/频率)", "触发什么动作"]
        for col, h in zip("ABCDE", heads):
            ws[f"{col}{dc}"] = h
            ws[f"{col}{dc}"].font = BF; ws[f"{col}{dc}"].fill = CH
            ws[f"{col}{dc}"].alignment = Alignment(wrap_text=True, vertical="top")
        dc += 1
        for item in trk["rows"]:
            if item[0] == "__band__":
                dband(dc, item[1]); dc += 1
                continue
            name, cur, why, how, act = item
            ws[f"A{dc}"] = name; ws[f"A{dc}"].font = BF
            for col, val in zip("BCDE", [cur, why, how, act]):
                ws[f"{col}{dc}"] = _esc(val)
                ws[f"{col}{dc}"].font = GREY
            for col in "ABCDE":
                ws[f"{col}{dc}"].alignment = Alignment(wrap_text=True, vertical="top")
            dc += 1
    ws.column_dimensions["A"].width = 20; ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 40; ws.column_dimensions["D"].width = 34
    ws.column_dimensions["E"].width = 34
    return {"row_of": row_of}
