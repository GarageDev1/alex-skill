# -*- coding: utf-8 -*-
"""
build_002436SZ.py — 兴森科技 (002436.SZ) 估值模型(kit v2 架构)。

单位约定(套用 kit 的 $B/百万股公式结构): 金额=亿元(B CNY), 股本=百万股, FX=1。
EPS = NI×1000/SH(元), BPS 同理; 股价=元。
物理锚: FCBGA 出货量(亿颗/年) × ASP(元/颗); 需求侧(中国 AIDC capex)仅作旁注+渗透率闭环。
数据 SOT: `VALUATION_INPUT_DIR/002436.SZ_input.json` 对应的历史快照。
"""
import os
from openpyxl import Workbook
import build_kit as K

ALLC = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
ALLY = ["2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
HC, HY = ["B", "C", "D", "E", "F"], ["2021", "2022", "2023", "2024", "2025"]
FC = ["F", "G", "H", "I", "J", "K"]
FCf = FC[1:]
FX_FC = 1
FX_HIST = [1, 1, 1, 1, 1]
CASES = ["Bear", "Base", "Bull"]

S_COVER, S_HIST, S_PX, S_CONS = "封面", "历史财务与估值", "股价走势", "卖方研报共识"
S_HMULT, S_MULT, S_SW = "历史估值倍数", "估值倍数假设", "情景切换"
S_ANCHOR, S_SEG, S_FUND = "FCBGA产能爬坡", "分部测算", "利润与收入假设"
S_VAL, S_CMP, S_DASH = "情景估值", "估值对比", "综合判断仪表盘"

MONTHLY = [("2021-06", 11.00), ("2021-09", 12.09), ("2021-12", 14.00),
           ("2022-03", 10.02), ("2022-06", 11.27), ("2022-09", 9.09), ("2022-12", 9.68),
           ("2023-03", 13.80), ("2023-06", 13.50), ("2023-09", 13.20), ("2023-12", 14.73),
           ("2024-03", 10.50), ("2024-06", 11.00), ("2024-09", 10.50), ("2024-12", 11.11),
           ("2025-03", 11.80), ("2025-06", 12.20), ("2025-09", 18.50), ("2025-12", 21.17),
           ("2026-03", 26.50), ("2026-06", 38.17)]
PX_NOW = 38.17

wb = Workbook()
wb.remove(wb["Sheet"])

# ════════════ 1. 封面 ════════════
K.write_cover(wb.create_sheet(S_COVER), {
    "title": "兴森科技 (002436.SZ) 估值模型 — FCBGA 产能爬坡 × P/B-ROE",
    "meta": [
        ("报告日期", "2026-06-12"),
        ("数据截止", "2025 年报 + 2026 一季报 + 卖方共识(2026-05) + 实时股价 2026-06-12"),
        ("现价", "38.17 元(总股本 17.0 亿股, 市值约 649 亿元)"),
        ("时效声明", "载板爬坡与情绪假设变化快, 每季财报后必须更新。"),
        ("方法一句话", "物理锚(FCBGA 出货×ASP 爬坡)+ BT/CSP 周期 + PCB 增速 → 段驱动利润 → 三层目标 P/B × 前瞻 BPS → 隐含价; P/E 支线体检。"),
    ],
    "takeaways": [
        ("① 当下估值位置", "P/B 12.3x = 5 年带 99% 分位; P/E TTM 449x。三层分解反推当下情绪值 1.69(过热档)。"),
        ("② 核心引擎", "载板收入 = FCBGA 出货(亿颗/年)×ASP + BT/CSP 周期; ABF 短缺下需求不约束, 收入由自家爬坡决定。"),
        ("③ 关键拐点", "Base: 载板段 2027 盈亏平衡、2028 转正至 +10% OPM; 共识更乐观(2026 即大幅减亏)。"),
        ("④ 三情景目标价", "目标年 2027E: Bear ~17 / Base ~32 / Bull ~44 元, 概率 25/50/25, 加权约 31 元 vs 现价 38.17。"),
        ("⑤ 主要风险", "FCBGA 客户放量证伪 / 折旧+利息压制 / 情绪退潮杀估值 / 实控人减持与再融资摊薄。"),
    ],
})

# ════════════ 2. 历史财务与估值 ════════════
ha = K.write_history(wb.create_sheet(S_HIST), {
    "title": "兴森科技 历史财务与估值 (亿元) — 2021-2025A + 当下",
    "vals_in_usd": True, "ps_scale": 100, "mcap_div": 100,
    "unit_label": "(亿元)", "mcap_label": "市值(亿元)", "eps_label": "EPS (元)", "bps_label": "BPS (元)",
    "fx_label": "FX(单币模型, 恒为 1)",
    "hist_cols": HC, "hist_years": HY,
    "fx_hist": FX_HIST, "fx_now": FX_FC,
    "segments": [
        ("PCB 收入", [37.94, 40.30, 40.91, 43.00, 48.97], True),
        ("IC载板 收入", [6.67, 6.90, 8.21, 11.16, 16.70], True),
        ("半导体测试板 收入", [4.17, 4.59, 2.65, 1.69, 2.39], False),
        ("其他 收入", [1.62, 1.75, 1.83, 2.33, 3.88], False),
    ],
    "total_now": 74.3,
    "gm_pct": [0.3217, 0.2866, 0.2332, 0.1587, 0.1957], "gm_now": 0.196,
    "ni": [6.21, 5.26, 2.11, -1.98, 1.35], "ni_now": 1.45,
    "eq": [37.62, 65.39, 53.34, 49.35, 53.48], "eq_now": 53.37,
    "shares": [1499, 1698, 1699, 1700, 1700], "shares_now": 1700,
    "px_end": [14.00, 9.68, 14.73, 11.11, 21.17],
    "px_now": PX_NOW,
    "px_avg": [13.29, 10.87, 13.59, 11.12, 16.56],
    "band_note": "P/B 5年带 2.3-13.4x(中位 4.0x); 当下 12.3x 处 99% 分位, 远超 2021 峰值 5.6x",
    "notes": [
        ("PCB 收入", "样板/小批量+Fineline+北京兴斐HDI+宜兴硅谷; 公司年报分部实际。"),
        ("IC载板 收入", "BT/CSP(珠海)+FCBGA(广州); 2025 毛利率 -16.1%(2024 -43.9%)。"),
        ("半导体测试板 收入", "高毛利小段(2025 GM 38%), 2023-24 受半导体下行收缩。"),
        ("其他 收入", "物业租赁/服务等, 毛利率高(89.8%)但体量小。"),
        ("HREV", "总营收=年报实际; 当下=TTM(2025A+2026Q1 增量)。"),
        ("HGMP", "整体毛利率: 2024 被载板爬坡折旧压到 15.9%, 2025 修复至 19.6%。"),
        ("HNI", "归母净利(亿元): 2024 上市以来首亏(FCBGA 费用 7.34 亿); 当下=TTM。"),
        ("HEQ", "归母净资产: 2022 跳升来自定增+载板子公司战投增资。"),
        ("HSH", "总股本(百万股): 2022 定增后约 17.0 亿股, 此后基本未稀释; 兴森转债存续转股中。"),
        ("HPX", "年末收盘价(东财月K); 当下=2026-06-12 实时 38.17。"),
        ("HPXA", "年均价=月度收盘均值, 同『股价走势』单一价格源。"),
    ],
})

# ════════════ 3. 股价走势 ════════════
def phase_fn(ym):
    if ym <= "2022-12":
        return "① 载板故事第一波退潮"
    if ym <= "2024-12":
        return "② 下行/首亏出清"
    if ym <= "2025-12":
        return "③ FCBGA 量产+扭亏"
    return "④ ABF 短缺重估"

px = K.write_price_chart(wb.create_sheet(S_PX), MONTHLY, {
    "fn": phase_fn,
    "rows": [("① 载板故事第一波退潮", "2021 高点 P/B~5.6x 后随盈利下行回落"),
             ("② 下行/首亏出清", "FCBGA 投入高峰+行业下行, 2024 首亏, 股价 9-15 区间"),
             ("③ FCBGA 量产+扭亏", "小批量量产+BT 涨价+年报扭亏, 12→21 元"),
             ("④ ABF 短缺重估", "ABF 缺口 10-42%+涨价+20层量产能力, 21→38 元创历史新高")],
}, title="兴森科技 月度股价 (元)")

# ════════════ 4. 卖方研报共识 ════════════
K.write_consensus(wb.create_sheet(S_CONS), {
    "title": "卖方研报共识 — 2026-05 口径; 这张表是后面测算的'卖方对账单'",
    "overview": "全街买入(中泰/中信/中银/招商/国海等); 2026E 归母 3.79-4.56 亿、2027E 6.81-8.31 亿; 13 家综合目标价约 26 元(主流 25-35), 已被现价 38 元穿越——市场实际交易的是 Bull 叙事。",
    "assumptions": [
        ("载板减亏节奏\n(2026)", "共识: 2026 载板大幅减亏, 全年归母 4.2 亿上下。", "2025 开发目标未完成+转债利息是否被低估。", "本模型 2026E 约 3.1 亿, 低于共识 ~25%: 载板 OPM -14%(共识隐含约 -8%), 净利转换率仅 0.65。"),
        ("FCBGA 放量\n(2027-28)", "共识: 2027 起放量, 2027E 归母 7 亿+。", "客户认证(两位数)何时变成公开大客户量产单。", "Base 取 2027 出货 0.8 亿颗(一期稼动 ~65%), 2027E 归母约 7 亿, 与共识一致。"),
        ("BT/存储载板", "受益存储涨价, 价格上行(高盛调研口径)。", "涨价持续性 vs 存储周期见顶。", "Base 取 2026 +15% 后逐年回落。"),
        ("目标倍数", "卖方 TP 26 元隐含 2027E P/B ~7x; 现价隐含 12.3x。", "稀缺产能给多高结构溢价、FOMO 撑多久。", "三层分解: 峰值 5.6 × 溢价 1.30 × 情绪(三案), 见『估值倍数假设』。"),
    ],
    "divergences": [
        "① 2026 减亏斜率: 共识快、本模型慢(-25%), 裁决指标=载板分部毛利率逐季变化。",
        "② 估值: 共识 TP 26 元 vs 现价 38 元——分歧已不在盈利而在倍数; 本模型用三层分解显性化情绪敞口。",
    ],
    "stances": [
        "中泰(买入): 2026E 4.56 / 2027E 7.48 亿, FCBGA 国产替代核心标的。",
        "中信(买入): 3.81 / 7.14 亿, AI PCB+载板双轮。",
        "招商(买入): 4.35 / 8.31 亿, 载板拐点+样板现金牛。",
        "UBS(2026-04): 盈利拐点确认但弹性弱于预期(Not Rated 调研口径)。",
    ],
})

# ════════════ 5. 历史估值倍数 ════════════
hm = K.write_hist_multiples(wb.create_sheet(S_HMULT), {
    "title": "历史估值倍数 — 自身历史带 + 当下倍数 + A 股算力链/载板同业对照",
    "intro": "①自己历史值多少(逐年+年内高低带) ②现在市场给多少 ③同行值多少 + 相对深南电路的比值(结构溢价对账线)。看完这页再去下一页拍三案倍数。",
    "s_hist": S_HIST, "ha": ha, "hist_cols": HC, "hist_years": HY,
    "yhigh": px["yhigh"], "ylow": px["ylow"],
    "fwd_note": "现价对模型 2027E EPS(约0.41元) ≈ 92x; 对 2028E(0.65元) ≈ 59x",
    "self_name": "兴森科技",
    "self_fwd_pe_label": "≈92x(27E)",
    "self_note": "本模型标的; forward 推导见『情景估值』。",
    "peers": [
        {"name": "深南电路(大陆载板一哥)", "yearly": [4.6, 2.9, 2.7, 2.6, 5.8], "cur_pb": 6.5, "cur_pe": 45.0, "fwd_pe": 32.0,
         "note": "BT 满产+已盈利, FCBGA 同处早期; 结构溢价对账核心可比。"},
        {"name": "沪电股份(AI PCB)", "yearly": [5.0, 3.5, 5.5, 7.5, 9.0], "cur_pb": 10.0, "cur_pe": 38.0, "fwd_pe": 26.0,
         "note": "AI 服务器 PCB 主力, 盈利已兑现——市场对'已兑现'给 26x forward。"},
        {"name": "Ibiden(全球 ABF 龙头)", "yearly": None, "cur_pb": 3.2, "cur_pe": 35.0, "fwd_pe": 22.0,
         "note": "全球龙头 forward 仅 22x: 海外市场对同一短缺叙事给的倍数远低于 A 股。"},
        {"name": "A股半导体(申万)中枢", "yearly": None, "cur_pb": None, "cur_pe": 60.0, "fwd_pe": 40.0,
         "note": "光谱中档: A 股半导体板块整体溢价。"},
    ],
    "ratio": {"peer": "深南电路(大陆载板一哥)",
              "note": "兴森/深南 P/B 比值 2021-24 约 0.9-1.2 → 当下 ~1.9: 本轮兴森相对深南额外重估近一倍, 结构溢价 1.30 已是偏厚口径, 超出部分计入情绪。"},
    "reading": "① 自己: 当下 12.3x 远超 2021 峰值 5.6x → 第一层锚用 5.6, 不用被本轮抬高的当下值。② 同行: 大陆载板/AI PCB 全板块 re-rating, 但兴森相对深南比值也翻倍 → 独贵成分存在。③ 比值走到 1.9 → 第二层结构溢价取 1.30, 其余归第三层情绪。",
})

# ════════════ 6. 估值倍数假设 ════════════
ma = K.write_multiple_assumptions(wb.create_sheet(S_MULT), {
    "title": "估值倍数假设 — P/B 主线方法论 + 三案目标倍数(= 历史周期峰值 × 结构溢价 × 情绪值)",
    "intro": "这一页只做判断: ①为什么 P/B 做主线 ②三层分解出三案目标倍数。『情景切换』引用并切换, 『情景估值』套用当前案, 『估值对比』三案并排。",
    "why_text": ("镜头选择是业务判断: 兴森穿越周期持续存在的是『产能与认证资产』而非盈利——载板是重资产爬坡期(2024 深亏、2025 微利, 盈利是爬坡状态量), "
                 "PCB 样板是稳定现金流但体量被载板投入吞噬。盈利在 -2 亿~+15 亿之间大摆, 资本化盈利无意义; 给『产能+其转正后的挣钱能力』定价才有意义 → P/B 主线。"
                 "P/E 永远做支线体检(每个情景都做, 对照 A 股算力链 forward PE 光谱); 爬坡亏损期另设 EV/Sales 诊断。"
                 "若 FCBGA 大客户量产单落地、载板段 OPM 稳定 15%+, 业务向'稳态盈利'迁移, 镜头应切 P/E 主线——这是触发条件, 不是永久标签。"),
    "why_rows": 5,
    "method_text": "三层分解(不硬拍): ①历史周期峰值 5.6x(2021 年末实际, 上一轮载板叙事+ROE 17.7% 峰值时市场真付过的价) × ②结构溢价 1.30(本轮 FCBGA 从 PPT 变实物产能+ABF 短缺稀缺度; 对账线=兴森/深南比值) × ③情绪值(记分卡定档)。一致性检验: 5.6×1.30×1.69 ≈ 12.3x = 当下实际 P/B ✓(当下处过热档)。",
    "peak": 5.6, "peak_note": "2021 年末实际 P/B(14.00 元 ÷ BPS 2.51); 不用本轮已重估的 12x 当锚(双重计算)。",
    "premium": 1.30, "premium_note": "锚兴森/深南历史比值 0.9-1.2 → 给到上沿再加稀缺产能溢价; 当下比值 1.9 中超出 1.3 的部分全部归情绪层。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "hpb_row": ha["HPB"],
    "cases": [
        ("Bear", [0.90, 0.70, 0.55, 0.50, 0.45],
         "载板减亏证伪(2026H2 分部毛利率仍 <-10%)或大客户认证流产 → FOMO 快速退潮; 类比 2022 年情绪 0.35-0.5 区间。"),
        ("Base", [1.35, 1.20, 1.00, 0.85, 0.75],
         "ABF 短缺可见度撑住 2026(情绪略低于当下 1.69 的过热), 2027 放量兑现后市场转向'看业绩', 情绪逐年向 1.0 以下退坡; 内存/载板类 FOMO 历史持续 <4-6 季。"),
        ("Bull", [1.70, 1.50, 1.25, 1.05, 0.90],
         "大客户量产单公告+二期启动, 短缺叙事撑到 2028, 当下过热档(1.69)维持一年后缓退; 对标台系载板厂 2021 ABF 缺货期的持续重估。"),
    ],
    "reconcile_text": "卖方 TP 26 元隐含 2027E P/B ~7x ≈ 本模型 Base(5.6×1.30×1.20=8.7x)略低; 现价 38 元隐含 2027E P/B ~10.3x ≈ 介于 Base 与 Bull——市场已支付 Bull 概率的大半。",
    "source_text": "第一层=东财月K×年报 BPS 实算; 第二层=兴森/深南 P/B 比值序列(历史估值倍数页); 第三层依据『综合判断仪表盘』D 块记分卡。",
})

# ════════════ 7. 情景切换 ════════════
sw = K.write_scenario_switch(wb.create_sheet(S_SW), {
    "title": "情景切换 — 全模型唯一的情景参数库 + 切换开关 (默认 Base)",
    "usage": ("怎么用: B2 是唯一入口——下拉选案 → 案序号派生 → 各杠杆『当前案』行跟着切 → "
              "整条明细链(锚→测算→利润→倍数→估值)变档, 『情景估值』输出该案逐年隐含价。"
              "三案对比不用切: 『估值对比』恒常三列并排。未列入的假设三案共用(跟 Base)。"),
    "cases": CASES, "default": "Base",
    "triggers": [
        ("Bear", "载板分部毛利率 2026H2 仍 <-10% / 大客户认证无量产转化 / ABF 短缺缓解叠加日台扩产; 实控人继续减持。"),
        ("Base", "FCBGA 2026 出货爬坡至 0.4 亿颗、2027 0.8 亿颗(一期稼动 ~65%), 载板段 2027 盈亏平衡、2028 OPM +10%; PCB 现金牛 +15%/+10%。"),
        ("Bull", "国产 AI 芯片大客户量产单公告 + 二期 2027 启动; ABF 涨价传导至 ASP; 载板 2026 即接近盈亏平衡。"),
    ],
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "levers": [
        {"key": "fcv", "name": "FCBGA 出货量(亿颗/年)", "fmt": K.N2,
         "vals": {"Bear": [0.25, 0.40, 0.55, 0.65, 0.70], "Base": [0.40, 0.80, 1.10, 1.20, 1.30], "Bull": [0.45, 0.95, 1.20, 1.55, 1.90]},
         "desc": "物理锚: 一期满产 1.2 亿颗/年(1,000 万颗/月), 两期合计 2.4 亿颗。载板估值故事全部挂在这条爬坡曲线上。",
         "stories": {"Bear": "认证客户不转化为量产单, 稼动率徘徊 50% 以下, 2030 仍未满产。",
                     "Base": "2027 稼动 ~65%、2028 接近一期满产; 二期不启动。",
                     "Bull": "2027 一期接近满产+二期启动, 2029-30 二期贡献出货。"},
         "hist": [0.0, 0.0, 0.01, 0.05, 0.12]},
        {"key": "asp", "name": "FCBGA ASP(元/颗)", "fmt": K.N1,
         "vals": {"Bear": [20, 19, 18, 17, 16], "Base": [23, 23, 22, 21, 20], "Bull": [25, 26, 25, 24, 23]},
         "desc": "满产产值 56 亿 ÷ 2.4 亿颗 ≈ 23 元/颗反推; 层数 mix 决定单价。",
         "stories": {"Bear": "只拿到低层数(<14 层)低价单, 单价逐年走低。",
                     "Base": "20 层及以下 mix 稳定, ASP 缓降(国产替代惯例)。",
                     "Bull": "ABF 涨价传导+高层数 mix 提升, ASP 先升后稳。"},
         "hist": [None, None, 20, 21, 21]},
        {"key": "btg", "name": "BT/CSP 收入增速", "fmt": K.PCT,
         "vals": {"Bear": [0.05, 0.03, 0.02, 0.02, 0.02], "Base": [0.15, 0.10, 0.08, 0.05, 0.05], "Bull": [0.22, 0.18, 0.12, 0.08, 0.06]},
         "desc": "BT/CSP(珠海+广州)2025 约 14.2 亿; 存储涨价周期+新扩 1.5 万㎡/月。",
         "stories": {"Bear": "存储周期 2026 见顶回落, 涨价红利消失。",
                     "Base": "DDR5 BOC 涨价延续到 2026, 之后回归量增。",
                     "Bull": "存储超级周期延长, 价量齐升。"},
         "hist": [None, 0.03, 0.19, 0.36, 0.50]},
        {"key": "pcbg", "name": "PCB 收入增速", "fmt": K.PCT,
         "vals": {"Bear": [0.08, 0.04, 0.03, 0.03, 0.03], "Base": [0.15, 0.10, 0.08, 0.07, 0.06], "Bull": [0.20, 0.15, 0.12, 0.10, 0.08]},
         "desc": "样板现金牛+Fineline+北京兴斐 AI HDI+1.6T 光模块板; 2026Q1 整体营收 +15.1%。",
         "stories": {"Bear": "研发景气回落, AI HDI 放量不及预期。",
                     "Base": "AI HDI/光模块板放量延续 Q1 动能, 之后回归高个位数。",
                     "Bull": "北美大厂订单放量+宜兴硅谷扭亏。"},
         "hist": [None, 0.062, 0.015, 0.051, 0.139]},
        {"key": "sopm", "name": "载板段 营业利润率", "fmt": K.PCT,
         "vals": {"Bear": [-0.22, -0.12, -0.05, 0.00, 0.03], "Base": [-0.14, 0.02, 0.10, 0.14, 0.16], "Bull": [-0.08, 0.08, 0.16, 0.20, 0.22]},
         "desc": "含 FCBGA 爬坡折旧(35 亿+投入转固)。2024 实际约 -55%, 2025 -26%; 转正时点是全模型最关键假设。",
         "stories": {"Bear": "稼动率不足, 折旧吞噬毛利, 2029 才盈亏平衡。",
                     "Base": "2027 盈亏平衡、2028 +10%(慢于共识一年)。",
                     "Bull": "规模效应+ABF 涨价, 2026 接近平衡、2027 即转正。"},
         "hist": [0.10, -0.02, -0.25, -0.55, -0.26]},
        {"key": "popm", "name": "PCB 段 营业利润率", "fmt": K.PCT,
         "vals": {"Bear": [0.08, 0.08, 0.085, 0.09, 0.09], "Base": [0.10, 0.115, 0.12, 0.12, 0.12], "Bull": [0.115, 0.13, 0.135, 0.135, 0.135]},
         "desc": "2025 GM 25.3% 扣段内费用约 15pct → OPM ~10%; CCL 涨价传导与 AI 占比决定弹性。",
         "stories": {"Bear": "覆铜板涨价传导不畅, 宜兴硅谷持续亏损。",
                     "Base": "AI HDI 占比提升对冲材料成本, OPM 缓升至 12%。",
                     "Bull": "高毛利 AI 板放量, OPM 升至 13.5%。"},
         "hist": [0.16, 0.14, 0.13, 0.11, 0.10]},
    ],
    "linked": [
        {"key": "sent", "name": "情绪值(倍数第三层)", "fmt": K.N2,
         "src_sheet": S_MULT, "src_row0": ma["sent_row0"],
         "note": "三案取值与依据见『估值倍数假设』; 本页只做切换。"},
    ],
})
_pk = f"'{S_MULT}'!{ma['pk_cell']}"
_pr = f"'{S_MULT}'!{ma['pr_cell']}"
_sent_act = sw["SWACT"]["sent"]
_r = sw["next_row"]
K.lab(wb[S_SW], f"A{_r}", "目标倍数(当前案)", b=True)
for _c in ALLC:
    K.fml(wb[S_SW], f"{_c}{_r}", f"={_pk}*{_pr}*{_c}{_sent_act}", K.MX, link=True)
K.logic(wb[S_SW], f"L{_r}", "= 历史周期峰值 5.6 × 结构溢价 1.30 × 当前案情绪值 → 喂『情景估值』的前瞻倍数。")
SWPB = _r

# ════════════ 8. 物理锚 [ANCHOR] ════════════
anchor = K.write_anchor(wb.create_sheet(S_ANCHOR), {
    "title": "FCBGA 产能爬坡 [ANCHOR] — 出货量 × ASP(供给侧物理锚) + 需求充分性旁注",
    "all_cols": ALLC, "all_years": ALLY,
    "series": [
        ("FCBGA 出货量(亿颗/年)",
         [0.0, 0.0, 0.01, 0.05, 0.12, None, None, None, None, None],
         "历史=小批量/样品估; 前瞻=『情景切换』当前案爬坡曲线。一期满产 1.2 亿颗/年。", K.N2),
        ("FCBGA ASP(元/颗)",
         [None, None, 20, 21, 21, None, None, None, None, None],
         "满产产值 56 亿 ÷ 2.4 亿颗 ≈ 23 元反推; 前瞻=当前案。", K.N1),
        ("中国 AIDC CapEx ($B)(需求旁注)",
         [None, None, None, 20, 45, 125, 160, 190, 210, 230],
         "共享基座 compute-aidc-base.json china 列; 仅作需求充分性参照, 不挂收入主链。", K.N0),
    ],
    "source_note": "锚=供给爬坡(公司公告/调研: 良率 85-92%、20 层量产能力、认证客户两位数); 需求=国产 AI 芯片载板盘子, ABF 缺口 10-42%(GS 2026-02)下需求不构成约束。",
    "role_note": "作用: 载板收入 = 出货×ASP + BT/CSP 周期。改出货爬坡 → 收入 → 利润 → 估值全链动。",
})
VOL_ROW = anchor["row_of"]["FCBGA 出货量(亿颗/年)"]
ASP_ROW = anchor["row_of"]["FCBGA ASP(元/颗)"]
CAPEX_ROW = anchor["row_of"]["中国 AIDC CapEx ($B)(需求旁注)"]
for _c in FCf:
    K.fml(wb[S_ANCHOR], f"{_c}{VOL_ROW}", f"={K.R(S_SW, _c + str(sw['SWACT']['fcv']))}", K.N2, link=True)
    K.fml(wb[S_ANCHOR], f"{_c}{ASP_ROW}", f"={K.R(S_SW, _c + str(sw['SWACT']['asp']))}", K.N1, link=True)

# ════════════ 9. 分部测算 ════════════
seg = K.write_segment_model(wb.create_sheet(S_SEG), {
    "title": "分部测算 — FCBGA(出货×ASP) + BT/CSP(周期) + PCB(增速) + 需求闭环检查",
    "all_cols": ALLC, "all_years": ALLY, "logic_col": "N",
    "groups": [
        ("FCBGA = 出货 × ASP(物理锚直驱)", [
            ("FCBGA 出货(亿颗/年)", None, K.N2, "= 引自『FCBGA产能爬坡』锚。"),
            ("FCBGA ASP(元/颗)", None, K.N1, "= 引自锚页; 前瞻=当前案。"),
            ("FCBGA 收入(亿元)", None, K.N2, "= 出货(亿颗) × ASP(元/颗) = 亿元。历史=小批量估(0.2/1.0/2.5 亿元)。"),
        ]),
        ("BT/CSP 载板 = 周期+扩产", [
            ("BT/CSP 收入(亿元)", None, K.N2, "历史 = 载板分部实际 − FCBGA 估; 前瞻 = 上年×(1+当前案增速)。"),
        ]),
        ("载板合计", [
            ("IC载板 收入(亿元)", None, K.N2, "= FCBGA + BT/CSP; 历史列自动等于年报分部实际(BT 为差额构造)。喂『利润与收入假设』。"),
        ]),
        ("PCB = 上年×(1+增速)", [
            ("PCB 收入(亿元)", None, K.N2, "历史=年报分部实际; 前瞻 = 上年×(1+当前案增速)。喂『利润与收入假设』。"),
        ]),
        ("需求充分性(旁注+国产替代闭环检查)", [
            ("国产载板需求盘子(亿元)", None, K.N1, "= 中国 AIDC capex($B) × GPU含量51% × 载板占GPU BOM 3% × 汇率7.2 × 10(折亿元); 旁注, 不挂主链。"),
            ("兴森 FCBGA 占需求盘子 %", None, K.PCT, "闭环检查: 国产替代渗透率应逐年上升; Base 2026 ~7% → 2030 ~14%, 与 thesis 同向且未荒谬(<20%)。"),
        ]),
    ],
})
m = seg["m"]
PCB_HROW = ha["seg_rows"]["PCB 收入"]
SUB_HROW = ha["seg_rows"]["IC载板 收入"]
FCBGA_HIST = [0.0, 0.0, 0.2, 1.0, 2.5]   # 亿元(估)
for col in ALLC:
    K.fml(wb[S_SEG], f"{col}{m['FCBGA 出货(亿颗/年)']}", f"={K.R(S_ANCHOR, col + str(VOL_ROW))}", K.N2, link=True)
    K.fml(wb[S_SEG], f"{col}{m['FCBGA ASP(元/颗)']}", f"={K.R(S_ANCHOR, col + str(ASP_ROW))}", K.N1, link=True)
for i, col in enumerate(HC):
    K.inp(wb[S_SEG], f"{col}{m['FCBGA 收入(亿元)']}", FCBGA_HIST[i], None, K.N2)
    K.fml(wb[S_SEG], f"{col}{m['BT/CSP 收入(亿元)']}", f"={K.R(S_HIST, col + str(SUB_HROW))}-{col}{m['FCBGA 收入(亿元)']}", K.N2, link=True)
    K.fml(wb[S_SEG], f"{col}{m['PCB 收入(亿元)']}", f"={K.R(S_HIST, col + str(PCB_HROW))}", K.N2, link=True)
for col in FCf:
    K.fml(wb[S_SEG], f"{col}{m['FCBGA 收入(亿元)']}", f"={col}{m['FCBGA 出货(亿颗/年)']}*{col}{m['FCBGA ASP(元/颗)']}", K.N2)
_prevs = [HC[-1]] + list(FCf[:-1])
for _p, _c in zip(_prevs, FCf):
    K.fml(wb[S_SEG], f"{_c}{m['BT/CSP 收入(亿元)']}", f"={_p}{m['BT/CSP 收入(亿元)']}*(1+{K.R(S_SW, _c + str(sw['SWACT']['btg']))})", K.N2)
    K.fml(wb[S_SEG], f"{_c}{m['PCB 收入(亿元)']}", f"={_p}{m['PCB 收入(亿元)']}*(1+{K.R(S_SW, _c + str(sw['SWACT']['pcbg']))})", K.N2)
for col in ALLC:
    K.fml(wb[S_SEG], f"{col}{m['IC载板 收入(亿元)']}", f"={col}{m['FCBGA 收入(亿元)']}+{col}{m['BT/CSP 收入(亿元)']}", K.N2)
for col in ALLC[3:]:
    K.fml(wb[S_SEG], f"{col}{m['国产载板需求盘子(亿元)']}", f"={K.R(S_ANCHOR, col + str(CAPEX_ROW))}*0.51*0.03*7.2*10", K.N1, link=True)
    K.fml(wb[S_SEG], f"{col}{m['兴森 FCBGA 占需求盘子 %']}", f"={col}{m['FCBGA 收入(亿元)']}/{col}{m['国产载板需求盘子(亿元)']}", K.PCT)
for col in FCf:
    wb[S_SEG][f"{col}{m['IC载板 收入(亿元)']}"].fill = K.OUT

# ════════════ 10. 利润与收入假设 ════════════
fr = K.write_fundamentals(wb.create_sheet(S_FUND), {
    "title": "利润与收入假设 — 段OPM(链切换) + 净利转换 + 留存 + 分部营收→利润→EPS/BPS",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "ha": ha, "share_fix_col": "F", "ps_scale": 100,
    "eps_label": "EPS (元)", "bps_label": "BPS (元)",
    "assum_groups": [
        ("营收驱动(PCB/载板已在分部测算; 此处仅小段)", [
            {"name": "测试板 增速", "vals": [None, 0.10, -0.42, -0.36, 0.41, 0.35, 0.30, 0.25, 0.20, 0.15],
             "fmt": K.PCT, "logic": "2022-25=实际 YoY; 前瞻=AI 芯片测试需求+北美大厂份额, 三案共用。"},
            {"name": "其他 增速", "vals": [None, 0.08, 0.05, 0.27, 0.67, 0.03, 0.03, 0.03, 0.03, 0.03],
             "fmt": K.PCT, "logic": "租赁/服务小段, 前瞻低速外推, 三案共用。"},
        ]),
        ("利润率假设(历史实际锚 + 前瞻; 粗颗粒)", [
            {"name": "PCB 营业利润率", "vals": [0.16, 0.14, 0.13, 0.11, 0.10, None, None, None, None, None],
             "fmt": K.PCT, "logic": "历史=GM−段内费用估; 前瞻=『情景切换』当前案。",
             "link": {"sheet": S_SW, "row": sw["SWACT"]["popm"]}},
            {"name": "载板 营业利润率", "vals": [0.10, -0.02, -0.25, -0.55, -0.26, None, None, None, None, None],
             "fmt": K.PCT, "logic": "含爬坡折旧, 2024 深亏 -55%; 前瞻=当前案(转正时点=最关键假设)。",
             "link": {"sheet": S_SW, "row": sw["SWACT"]["sopm"]}},
            {"name": "测试板 营业利润率", "vals": [0.06, 0.06, 0.00, 0.20, 0.22, 0.22, 0.22, 0.22, 0.22, 0.22],
             "fmt": K.PCT, "logic": "2025 GM 38% 扣费用; 高毛利小段, 三案共用。"},
            {"name": "其他 营业利润率", "vals": [0.50, 0.50, 0.45, 0.50, 0.50, 0.50, 0.50, 0.50, 0.50, 0.50],
             "fmt": K.PCT, "logic": "高毛利(GM 89.8%)非主业段, 三案共用。"},
        ]),
        ("净利转换与留存", [
            {"name": "净利转换率(净利/营业利润)", "vals": [0.85, 0.78, 0.55, None, 0.45, 0.65, 0.70, 0.73, 0.74, 0.75],
             "fmt": K.PCT, "logic": "吸收利息(有息负债 52.3 亿)/税/少数股东; 2024 亏损年 n.m.; 随载板减亏与转债转股回升。", "nm_cols": ["E"]},
            {"name": "留存率", "vals": [0.90, 0.92, 0.85, None, 1.00, 0.85, 0.85, 0.85, 0.85, 0.85],
             "fmt": K.PCT, "logic": "分红率低(扩产优先); 2024 亏损年 n.m.; 前瞻 0.85。", "nm_cols": ["E"]},
        ]),
    ],
    "segments": [
        {"name": "PCB 收入", "hist_row": "PCB 收入", "fwd": {"sheet": S_SEG, "row": m["PCB 收入(亿元)"]}},
        {"name": "IC载板 收入", "hist_row": "IC载板 收入", "fwd": {"sheet": S_SEG, "row": m["IC载板 收入(亿元)"]}},
        {"name": "半导体测试板 收入", "hist_row": "半导体测试板 收入", "fwd": {"growth": "测试板 增速"}},
        {"name": "其他 收入", "hist_row": "其他 收入", "fwd": {"growth": "其他 增速"}},
    ],
    "profit_terms": [
        (["PCB 收入"], "PCB 营业利润率", True),
        (["IC载板 收入"], "载板 营业利润率", False),
        (["半导体测试板 收入"], "测试板 营业利润率", False),
        (["其他 收入"], "其他 营业利润率", False),
    ],
    "conv_assum": "净利转换率(净利/营业利润)", "retention_assum": "留存率",
    "note_text": "分部营收(FCBGA 出货×ASP + BT 周期 + PCB 增速)→ 段驱动营业利润 → 净利(×净利转换)→ 权益(留存递推)→ EPS/BPS/ROE。下游『情景估值』直接引本表每股。",
})

# ════════════ 11. 情景估值 ════════════
sv = K.write_scenario_valuation(wb.create_sheet(S_VAL), {
    "title": "情景估值 — 当前案的逐年隐含价 (P/B 主线; P/E 交叉验证)",
    "intro": "本表输出=『情景切换』当前案(默认 Base)。隐含价 = 目标倍数(当前案) × 前瞻 BPS; P/E 仅交叉验证。历史列用实际年末价反推倍数(事实); 前瞻是预测、不拟合现价。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf, "hist_years": HY,
    "fx_fwd": FX_FC, "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
    "mcap_div": 100, "mcap_usd_skip": True,
    "mcap_hist_label": "市值 实际年末(亿元, 历史)", "mcap_fwd_label": "市值 前瞻·主线(亿元)",
    "s_fund": S_FUND, "fr": fr,
    "s_switch": S_SW, "target_row": SWPB, "sw_cell": "B2",
    "yend": px["yend"], "yavg": px["yavg"],
    "reading": "P/E 交叉验证读法: 隐含价÷前瞻 EPS, 对照光谱(沪电 26x / 深南 32x / A股半导体 40x / Ibiden 22x)。Base 2027 隐含 ~78x、2028 ~50x: 高于已兑现的 AI PCB 同行——只有载板转正兑现才能消化, 这正是 P/E 体检暴露的脆弱点。爬坡亏损期另看 EV/Sales: 现市值 649 亿 ÷ 2026E 收入 ~83 亿 ≈ 7.8x, 对照沪电 ~6x、深南 ~3.5x。",
    "method": "整体公司、P/B 主线逐年估。基本面在『利润与收入假设』; 目标倍数在『估值倍数假设』(三层); 本表只做最后一步: 目标倍数 × 前瞻 BPS → 隐含价 + 市值。",
    "concl": "方向性结论: Base 隐含价 2027E 约 32 元 < 现价 38 元——现价已支付 Bull 概率的大半; risk-reward 偏下行(Bear -55% vs Bull +15%)。",
})

# ════════════ 12. 估值对比 ════════════
SWB = sw["SWB"]
SH_F = K.R(S_HIST, f"$F${ha['HSH']}")
PX_NOW_REF = K.R(S_HIST, f"G{ha['HPX']}")
_tbg = fr["am"]["测试板 增速"]
_obg = fr["am"]["其他 增速"]
_topm = fr["am"]["测试板 营业利润率"]
_oopm = fr["am"]["其他 营业利润率"]
_conv = fr["am"]["净利转换率(净利/营业利润)"]
_ret = fr["am"]["留存率"]


def _fwdprev(j, A, key):
    return (HC[-1] if j == 0 else FCf[j - 1]) + str(A[key])


cmp_rows = [
    {"key": "vol", "label": "FCBGA 出货(亿颗/年)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_ANCHOR, c + str(VOL_ROW))}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_SW, c + str(SWB['fcv'] + ci))}"},
    {"key": "fcb", "label": "FCBGA 收入(亿元)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_SEG, c + str(m['FCBGA 收入(亿元)']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['vol']}*{K.R(S_SW, c + str(SWB['asp'] + ci))}"},
    {"key": "bt", "label": "BT/CSP 收入(亿元)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_SEG, c + str(m['BT/CSP 收入(亿元)']))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'bt')}*(1+{K.R(S_SW, c + str(SWB['btg'] + ci))})"},
    {"key": "pcb", "label": "PCB 收入(亿元)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(PCB_HROW))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'pcb')}*(1+{K.R(S_SW, c + str(SWB['pcbg'] + ci))})"},
    {"key": "tst", "label": "测试板 收入(亿元)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['seg_rows']['半导体测试板 收入']))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'tst')}*(1+{K.R(S_FUND, c + str(_tbg))})"},
    {"key": "oth", "label": "其他 收入(亿元)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['seg_rows']['其他 收入']))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'oth')}*(1+{K.R(S_FUND, c + str(_obg))})"},
    {"key": "rev", "label": "总收入(亿元)", "fmt": K.N2, "bold": True,
     "hist": lambda c, ci, A: f"={c}{A['fcb']}+{c}{A['bt']}+{c}{A['pcb']}+{c}{A['tst']}+{c}{A['oth']}",
     "fwd": lambda c, j, ci, A: f"={c}{A['fcb']}+{c}{A['bt']}+{c}{A['pcb']}+{c}{A['tst']}+{c}{A['oth']}"},
    {"key": "ni", "label": "净利(亿元)", "fmt": K.N2, "bold": True,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HNI']))}",
     "fwd": lambda c, j, ci, A: (f"=({c}{A['pcb']}*{K.R(S_SW, c + str(SWB['popm'] + ci))}"
                                 f"+({c}{A['fcb']}+{c}{A['bt']})*{K.R(S_SW, c + str(SWB['sopm'] + ci))}"
                                 f"+{c}{A['tst']}*{K.R(S_FUND, c + str(_topm))}"
                                 f"+{c}{A['oth']}*{K.R(S_FUND, c + str(_oopm))})*{K.R(S_FUND, c + str(_conv))}")},
    {"key": "eq", "label": "期末权益(亿元)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HEQ']))}",
     "fwd": lambda c, j, ci, A: (f"={K.R(S_HIST, 'F' + str(ha['HEQ']))}+{c}{A['ni']}*{K.R(S_FUND, c + str(_ret))}" if j == 0
                                 else f"={FCf[j-1]}{A['eq']}+{c}{A['ni']}*{K.R(S_FUND, c + str(_ret))}")},
    {"key": "bps", "label": "BPS (元)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={c}{A['eq']}*100/{K.R(S_HIST, c + str(ha['HSH']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['eq']}*100/{SH_F}"},
    {"key": "sent", "label": "情绪值(该案; 历史=实际反推)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_MULT, c + str(ma['sent_row0'] + ci))}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_MULT, c + str(ma['sent_row0'] + ci))}"},
    {"key": "pb", "label": "目标倍数(历史=实际)", "fmt": K.MX,
     "hist": lambda c, ci, A: f"={_pk}*{_pr}*{c}{A['sent']}",
     "fwd": lambda c, j, ci, A: f"={_pk}*{_pr}*{c}{A['sent']}"},
    {"key": "px", "label": "隐含价 (元)", "fmt": K.PX, "bold": True, "out": True,
     "hist": lambda c, ci, A: f"={c}{A['pb']}*{c}{A['bps']}",
     "fwd": lambda c, j, ci, A: f"={c}{A['pb']}*{c}{A['bps']}"},
    {"key": "ipe", "label": "隐含 forward P/E(交叉验证)", "fmt": K.MX,
     "hist": lambda c, ci, A: f'=IF({c}{A["ni"]}<=0,"N/M",{c}{A["px"]}/({c}{A["ni"]}*100/{K.R(S_HIST, c + str(ha["HSH"]))}))',
     "fwd": lambda c, j, ci, A: f'=IF({c}{A["ni"]}<=0,"N/M",{c}{A["px"]}/({c}{A["ni"]}*100/{SH_F}))'},
    {"key": "up", "label": "历史: vs 实际年末价(回测≈0) / 前瞻: vs 现价", "fmt": K.PCT,
     "hist": lambda c, ci, A: f"={c}{A['px']}/{K.R(S_HIST, c + str(ha['HPX']))}-1",
     "fwd": lambda c, j, ci, A: f"={c}{A['px']}/{PX_NOW_REF}-1"},
]
cm = K.write_comparison(wb.create_sheet(S_CMP), {
    "title": "估值对比 — Bear / Base / Bull 三个情景的目标价并排对比",
    "intro": ("三个情景各自完整推演: FCBGA 爬坡 → 分部收入 → 净利 → BPS → 目标倍数(三层) → 逐年隐含价。"
              "本表三案永远并排可见, 不随『情景切换』开关变化。未列入情景矩阵的假设三案共用 Base 取值。"
              "历史列 2021-2025 = 同一条链填实际值: 隐含价历史列应基本等于实际年末价(回测行 ≈0%)。"),
    "case_names": CASES,
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "block_start": 16,
    "rows": cmp_rows,
    "summary": {
        "band": "三案汇总 (目标年 2027E; 各案触发条件见『情景切换』)",
        "target_col": "H",
        "rows": [
            ("净利(亿元)", "ni", K.N2, "由该案的爬坡/价格/利润率假设逐年推导", False),
            ("总收入(亿元)", "rev", K.N2, "= FCBGA(出货×ASP) + BT/CSP + PCB + 测试板 + 其他", False),
            ("目标倍数", "pb", K.MX, "= 历史周期峰值 5.6 × 结构溢价 1.30 × 该案情绪值", False),
            ("隐含价(元)", "px", K.PX, "= 目标倍数 × 2027E BPS", True),
            ("隐含 forward P/E", "ipe", K.MX, "P/E 体检: 该案隐含价 ÷ 该案 EPS, 对照同业光谱(沪电26x/深南32x/Ibiden22x)", False),
            ("vs 现价", "up", K.PCT, "对照现价 38.17 元的上行/下行空间", True),
        ],
        "mcap": {"label": "隐含市值(亿元)", "key": "px", "expr": f"*{SH_F}/100",
                 "note": "= 隐含价 × 股本(17.0 亿股)"},
        "concl": "风险收益比(2027E vs 现价 38.17): Bear -55% / Base -16% / Bull +15%, 概率加权(25/50/25)约 -18%——现价已支付 Bull 概率的大半; 任一衰减信号触发应转向 Bear 重看。",
    },
})

# ════════════ 13. 综合判断仪表盘 ════════════
BPS27 = K.R(S_FUND, "H" + str(fr["BPS"]))
PXD = K.R(S_HIST, "G" + str(ha["HPX"]))
dash = K.write_dashboard(wb.create_sheet(S_DASH), {
    "title": "综合判断仪表盘 — A 基本面拐点 · B 估值错位 · C 催化剂 · D 情绪确认",
    "usage": ("预测引擎是 B(错位)+ C(催化剂); 情绪 D 只做 timing 确认 + 过热刹车。"
              "回测验收: 2025H2(12 元, FCBGA 进量产+BT 涨价, A 强/B GAP 正/C 兑现中/D 启动)→ 当时该看多 ✓; 当下读数已翻转。"),
    "blocks": [
        {"title": "A. 基本面拐点 — 业务在结构性变好吗?", "rows": [
            ("载板分部毛利率", "-43.9%(2024) → -16.1%(2025), 修复中", "方向为正, 但仍深度为负; 拐点=转正季。"),
            ("FCBGA 量产能力", "20 层及以下量产 + 良率 85-92% + 认证客户两位数", "工艺死亡谷已过, 但无公开大客户量产单。"),
            ("现金流", "2025 经营现金流 -0.63 亿, 有息负债 52.3 亿", "输血泵失血是 A 块最大减分项。"),
            ("A 判断", "【中】", "拐点真实但兑现在 2027-28, 且现金流逆风。", True),
        ]},
        {"title": "B. 估值错位 — 市场现在给的 vs 基本面该给的 → GAP", "rows": [
            ("市场现在给(forward P/B)", {"fml": f"={PXD}/{BPS27}", "fmt": K.MX, "fill": True},
             "= 现价 ÷ 2027E BPS(公式算, 随模型走)。"),
            ("基本面该给(justified)", {"inp": 2.6, "fmt": K.MX},
             "=(ROE−g)/(COE−g)=(16%−6%)/(9.8%−6%); 取载板转正后稳态 ROE 16%。注: 目标倍数=该给×稀缺/情绪溢价。"),
            ("错位 GAP = 该给÷市场给 − 1",
             {"fml": lambda ro: f"=B{ro['基本面该给(justified)']}/B{ro['市场现在给(forward P/B)']}-1", "fmt": K.PCT},
             "GAP 深度为负 = 基本面空间已被价格吃完, 当前完全处于情绪/稀缺定价区。"),
            ("回测: 2025H2 拐点读数", "当时 P/B ~3.8x vs 该给(含转正预期)~3x, GAP 接近 0 且催化剂密集", "那波的预测依据: 错位不大但催化剂强+情绪刚启动。"),
        ]},
        {"title": "C. 催化剂 — 什么会逼市场闭合 GAP", "rows": [
            ("大客户量产单公告", "待", "Bull 的核心扳机; 落地则结构溢价上修。"),
            ("载板分部毛利率转正", "进行中(市场预期 2026H2-2027)", "Base 路径的硬验证点, 逐季可跟踪。"),
            ("二期启动公告", "待", "管理层对需求的诚实投票; 启动=Bull, 持续observe=Base。"),
            ("ABF 涨价/缺口数据", "进行中(味之素涨价, 缺口 10-42%)", "行业顺风, 但利好已大半 price-in。"),
            ("C 判断", "已兑现大半", "剩余催化剂(大客户单/转正)兑现前, 超涨靠情绪维持。", True),
        ]},
        {"title": "D. 情绪确认 — 只做 timing + 刹车", "rows": [
            ("量价温度计", "5/29 涨停+机构单日净买 3.64 亿; 12 个月 +200%", "量价 ≫ 基本面进度 = FOMO 特征, 只当温度计。"),
            ("现价倍数 vs 基本面该给", "12.3x vs 2.6x(justified)", "市场已付出远超基本面该给, 深入情绪定价区。"),
            ("当前档位", "【过热】(情绪值 1.69)", "对照 2021 峰值 0.77、2025 末 0.92——本轮情绪为历史最高。", True),
            ("衰减扳机", "4 条", "①载板毛利率连续两季改善停滞 ②ABF 缺口收窄/日台扩产落地 ③实控人再减持 ④存储载板价格见顶; 任一翻 → 情绪降档。"),
        ]},
    ],
    "final": {"band": "★ 综合判断(A+B+C+D 收成一句可执行的话)",
              "text": "A 中(拐点真实但后置)+ B 深度负 GAP + C 已兑现大半 + D 过热 → 现价 38.17 已支付 Bull 大半概率, 概率加权目标 ~31 元: 持有不追高, 回调至 Base 隐含价(~32 元)下方再评估; 大客户量产单落地前不加仓。"},
    "tracking": {
        "intro": "哪个指标恶化 → 哪个假设先崩 → 触发什么动作。",
        "rows": [
            ("__band__", "一、载板爬坡主链"),
            ("载板分部毛利率(逐季)", "-16.1%(2025FY)", "命门: sopm 转正时点", "季报分部数据", "连续两季无改善 → 切 Bear 重算"),
            ("FCBGA 出货/稼动率", "小批量(~0.12 亿颗/年)", "命门: 物理锚爬坡曲线", "公告/调研纪要", "2026 末仍 <0.3 亿颗年化 → 下调 fcv"),
            ("大客户量产订单", "无公开订单", "Bull 扳机", "公告", "落地 → 切 Bull; 认证流产 → 切 Bear"),
            ("__band__", "二、行业供需"),
            ("ABF 缺口/味之素供应", "缺口 10-42%, 膜涨价", "稀缺溢价的行业前提", "GS/TrendForce tracker", "缺口收窄 → 结构溢价下调"),
            ("存储载板(BT)价格", "上行", "BT 段现金流", "行业价格 tracker", "见顶回落 → btg 切 Bear"),
            ("__band__", "三、资金与治理"),
            ("经营现金流/有息负债", "-0.63 亿 / 52.3 亿", "再融资摊薄风险", "季报", "现金流持续为负+增发公告 → BPS 稀释重算"),
            ("实控人增减持", "2025 高位累计减持", "治理信号", "公告", "再减持 → 情绪值下调"),
        ],
    },
})

# ════════════ 全局格式 + 落盘 ════════════
K.finalize(wb, freeze={
    S_HIST: "B3", S_PX: "B4", S_CONS: "A2", S_HMULT: "B5", S_MULT: "B4", S_SW: "B3",
    S_ANCHOR: "B3", S_SEG: "B3", S_FUND: "B3", S_VAL: "B4", S_CMP: "B6", S_DASH: "B6",
    S_COVER: "A2",
})
out = os.path.join(os.path.dirname(__file__), "..", "out", "002436.SZ_valuation_model.xlsx")
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("saved:", os.path.abspath(out))
print("sheets:", wb.sheetnames)
