# -*- coding: utf-8 -*-
"""
build_002409SZ.py — 雅克科技 (002409.SZ) 估值模型, 基于 build_kit v2。
物理锚: 全球 DRAM/HBM 前道 capex × 前驱体收入强度 (compute-semiconductor 锚变体: 存储材料挂存储 capex)。
A股 CNY 公司: 财务量纲 USD $B(过 $B 口径校验), 每股/股价 CNY, FX=CNY/USD。
镜头: P/E 主线(盈利型平台, 前驱体绑 HBM 高毛利 franchise + 稼动率经营杠杆) + DCF 支线。
数据 SOT: 用户提供的研究稿(2026-07-01)。
          (巨潮 2023/2024/2025 年报一手分部拆分 + 统一数据 API 行情/三表/共识, 刷新至 2026-07-01)。
"""
import os
from openpyxl import Workbook
import build_kit as K

# ════════════ 0. 全局轴 ════════════
ALLC = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
ALLY = ["2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
HC, HY = ["B", "C", "D", "E", "F"], ["2021", "2022", "2023", "2024", "2025"]
FC = ["F", "G", "H", "I", "J", "K"]     # 前瞻含基年 F=2025A
FCf = FC[1:]                            # 2026E..2030E
FX_FC = 7.15
FX_HIST = [6.45, 6.73, 7.08, 7.20, 7.17]
CASES = ["Bear", "Base", "Bull"]

S_COVER, S_HIST, S_PX, S_CONS = "封面", "历史财务与估值", "股价走势", "卖方研报共识"
S_HMULT, S_MULT, S_SW = "历史估值倍数", "估值倍数假设", "情景切换"
S_ANCHOR, S_SEG, S_FUND = "DRAM_HBM_Capex预测", "分部测算", "利润与收入假设"
S_VAL, S_CMP, S_DASH = "情景估值", "估值对比", "综合判断仪表盘"

# 真实月度收盘(统一数据 API /price 湖, 不复权; 无近期送转)
MONTHLY = [
    ("2021-06", 90.0), ("2021-12", 81.2), ("2022-06", 60.0), ("2022-12", 50.4),
    ("2023-06", 72.88), ("2023-09", 64.6), ("2023-12", 55.73),
    ("2024-03", 55.82), ("2024-06", 62.91), ("2024-09", 61.38), ("2024-12", 57.95),
    ("2025-03", 61.97), ("2025-06", 54.69), ("2025-09", 73.75), ("2025-12", 74.1),
    ("2026-01", 96.29), ("2026-02", 95.97), ("2026-03", 78.74), ("2026-04", 84.71),
    ("2026-05", 107.67), ("2026-06", 181.03), ("2026-07", 236.09),
]
PX_NOW = 236.09
YEND = {"2021": 81.2, "2022": 50.4, "2023": 55.7, "2024": 58.0, "2025": 74.1}
YAVG = {"2021": 73.1, "2022": 57.3, "2023": 62.1, "2024": 57.0, "2025": 61.9}
YHIGH = {"2021": 106.5, "2022": 81.3, "2023": 87.1, "2024": 74.4, "2025": 88.1}
YLOW = {"2021": 50.7, "2022": 37.4, "2023": 49.0, "2024": 33.2, "2025": 48.7}

wb = Workbook()
wb.remove(wb["Sheet"])

# ════════════ 1. 封面 ════════════
K.write_cover(wb.create_sheet(S_COVER), {
    "title": "雅克科技 (002409.SZ) 估值模型",
    "meta": [
        ("报告日期", "2026-07-01"),
        ("数据截止", "2025 年报(巨潮一手分部) + 2026Q1 + 卖方共识 + 现价 ¥236.09(2026-07-01, 当日 +5.16% 创历史新高)"),
        ("现价", "¥236.09 | 市值 ~¥1,124 亿 (~$157亿) | PE(TTM) 111.6x / PB 14.1x"),
        ("量纲口径", "财务量纲 USD $B(过 $B 口径校验); 每股/股价 CNY; FX=CNY/USD(2025≈7.17, 前瞻 7.15)。无近期送转, 股本 476mn。"),
        ("时效声明", "基于建模时点共识/指引/股价; HBM4 前驱体导入节奏与板块情绪变化快, 建议每季财报后更新。"),
        ("方法一句话", "物理锚(DRAM/HBM 前道 capex)→ 前驱体收入=capex×强度 + 其余5段(LNG周期/光刻胶特气硅微粉成长/化工稳态)→ 段净利率 → 归母净利 → P/E主线×2027E EPS → 隐含价; DCF 交叉验证。"),
    ],
    "takeaways": [
        ("① 当下估值位置", "现价 ¥236 ÷ 2026E EPS ~2.71 = 87x; ÷ 2027E ~3.39 = 70x——一家历史 forward P/E 长期 28-40x 的材料平台, 现价对近端盈利付 70-87x。2020-25 横盘六年(¥34-107), 涨幅集中在 2026 上半年(+200%)。"),
        ("② 核心引擎", "半导体前驱体(2025 收入 21.11亿/毛利 44.79%/稼动率仅 59.31%): 收入=DRAM/HBM capex × 强度(2025A 0.55% ← 锚); HBM4 High-K + 以钼代钨 + 国产化份额 + 稼动率 59%→80% 四因子推强度; Base 2027E 0.66%。前驱体是利润引擎, 但仅占营收 24.5%。"),
        ("③ 平台含金量分化", "剥离非半导体: LNG 保温板材 27.5%(在手订单 62亿, 独立造船周期) + 化工 13.5% 合计约 41%; 光刻胶 22.8% 中绝大部分是显示/面板胶(毛利仅 16.72%, 非半导体前道)。真正绑 HBM 的只有前驱体一段。"),
        ("④ 三情景目标价", "2027E 目标年: Bear ¥XX / Base ¥XX / Bull ¥XX(建模后从『估值对比』读入)。前驱体强度×净利率 + 板块情绪值 三案翻档, 沿同一条链算出。"),
        ("⑤ 主要争议", "看空: 前驱体 2025 仅 +8%(韩国 UP -41%), 高绝对 PE 是国产替代板块 beta, 板块回调难独善; 看多: 稼动率 59% 经营杠杆 + HBM4 国产化 + 江苏先科并表, Bull 情景比多数 AI 票更 live。"),
    ],
})

# ════════════ 2. 历史财务与估值 ════════════
ha = K.write_history(wb.create_sheet(S_HIST), {
    "title": "雅克科技 历史财务与估值 ($B) — 2021-2025A + 当下(分部按年报一手拆分, CNY→$B 按年均汇率)",
    "hist_cols": HC, "hist_years": HY,
    "fx_hist": FX_HIST, "fx_now": FX_FC,
    "vals_in_usd": True,
    "segments": [
        ("前驱体 收入", [0.109, 0.134, 0.161, 0.271, 0.294], True),
        ("LNG保温 收入", [0.047, 0.082, 0.122, 0.227, 0.330], False),
        ("光刻胶及配套 收入", [0.171, 0.186, 0.184, 0.213, 0.273], False),
        ("电子特气 收入", [0.054, 0.059, 0.062, 0.065, 0.058], False),
        ("硅微粉及LDS 收入", [0.039, 0.052, 0.046, 0.076, 0.083], False),
        ("化工及其他 收入", [0.168, 0.120, 0.094, 0.100, 0.162], False),
    ],
    "total_now": 1.201,
    "gm_pct": [0.258, 0.312, 0.313, 0.316, 0.310], "gm_now": 0.310,
    "ni": [0.0519, 0.0780, 0.0818, 0.1211, 0.1395], "ni_now": 0.1395,
    "eq": [0.744, 0.773, 0.847, 1.000, 1.107], "eq_now": 1.107,
    "shares": [467, 476, 476, 476, 476], "shares_now": 476,
    "px_end": [81.2, 50.4, 55.7, 58.0, 74.1],
    "px_now": PX_NOW,
    "px_avg": [73.1, 57.3, 62.1, 57.0, 61.9],
    "band_note": "forward P/E 历史带 28-40x(2022-25末价÷次年EPS), 2026 上半年重估至 70-87x; 当下 = 历史极值 2-3 倍。",
    "notes": [
        ("前驱体 收入", "= 年报『营业收入构成·分产品·半导体化学材料』: 2023 11.38亿/2024 19.54亿/2025 21.11亿(毛利 44.79%); 2021-22 为估(段拆分早年年报未下载, 按趋势回填)。AI/HBM 直驱段, 挂 DRAM/HBM capex。"),
        ("LNG保温 收入", "= 年报『LNG 保温绝热材料』: 2023 8.65/2024 16.35/2025 23.70亿(+44.9%)。独立造船周期, 在手订单>62亿, 不挂半导体。2021-22 估。"),
        ("光刻胶及配套 收入", "= 年报『光刻胶及配套试剂』: 2025 19.60亿(+27.7%, 毛利仅 16.72%)——绝大部分为显示/面板彩色胶(全球第一), 半导体前道 ArF/EUV 未量产。国产替代期权, 非半导体胶。2021-22 估。"),
        ("电子特气 收入", "= 年报『电子特种气体』: 2025 4.17亿(-11.1%), 仅 SF6/CF4(成都科美特), 无 WF6。边缘配角。"),
        ("硅微粉及LDS 收入", "= 年报『硅微粉』(2025 2.88亿, 华飞) + 『LDS 设备』(2025 3.04亿); 合并为半导体配套/先进封装段。"),
        ("化工及其他 收入", "= 年报『阻燃剂』(2.88) + 『租赁及工程服务』(6.82, 含 LNG 安装) + 『其他』(1.91), 2025 合计 11.61亿。稳态/杂项段。"),
        ("HREV", "总营收=交易所实际(统一数据 API): 2025 86.11亿 CNY(+25.5%)=$1.201B@7.17; 分部拆分按年报『营业收入构成』(2021-22 段拆分为估, 加总勾稽实际总额)。"),
        ("HGMP", "综合毛利率(年报): 2025 31.0%(2021 因光刻胶/LNG 低毛利放量降至 25.8%后回升); 前驱体 44.79% 拉动 mix。"),
        ("HNI", "归母净利(统一数据 API): 2021 3.35/2022 5.25/2023 5.79/2024 8.72/2025 10.00亿 CNY, 按年均汇率折$B; 当下=2025A。"),
        ("HEQ", "归母股东权益(估, 由 ROE 与净利反推 + 总权益扣少数股东): 2025 ≈79.4亿 CNY。少数股东主要为江苏先科 33.07% 直接持股。"),
        ("HSH", "总股本(年报): 476mn 股, 无近期送转; 2021 略低(467mn)。"),
        ("HPX", "年末收盘价(统一数据 API 湖, 不复权); 当下=2026-07-01 ¥236.09。"),
        ("HPXA", "年内月度收盘均价(真实月K)。"),
        ("HPE", "P/E=年末价÷当年 EPS。2025 74.1/2.10=35.3x; 2021-24 在 30-46x。"),
    ],
})

# ════════════ 3. 股价走势 ════════════
def phase_fn(ym):
    if ym <= "2023-12":
        return "① 材料平台横盘"
    if ym <= "2025-08":
        return "② 低位磨底"
    if ym <= "2025-12":
        return "③ 国产替代预期启动"
    return "④ HBM前驱体主题爆发"

px = K.write_price_chart(wb.create_sheet(S_PX), MONTHLY, {
    "fn": phase_fn,
    "rows": [("① 材料平台横盘", "2021-23 ¥50-107 区间震荡, 市场按'并购拼起来的化工材料平台'给 30-40x"),
             ("② 低位磨底", "2024-25H1 ¥33-74, 前驱体+光刻胶放量但盈利被低毛利业务摊薄, 估值不动"),
             ("③ 国产替代预期启动", "2025H2 ¥55→74, 大基金三期 + HBM 材料国产替代叙事发酵"),
             ("④ HBM前驱体主题爆发", "2026H1 ¥77→236(+200%), 前驱体绑 HBM4 + 以钼代钨 + 龙虎榜/大基金概念, 6-7月加速创历史新高")],
}, title="雅克科技 月度股价 (CNY, 不复权)")

# ════════════ 4. 卖方研报共识 ════════════
K.write_consensus(wb.create_sheet(S_CONS), {
    "title": "卖方研报共识 — A股覆盖(近端 2 家窄样本 / 东财聚合 8 家); 全买入, A股无目标价聚合",
    "overview": "近两月 2 家(华鑫 6/29 买入 EPS 2.81/3.48; 中银 5/8 买入 EPS 2.60/3.30); 东财聚合 8 家(6买入2增持)2026E EPS ~2.84 / 2027E ~3.45 / 2028E ~4.11。A股不出目标价聚合, 只有 EPS 与评级。共识核心叙事: 半导体高景气下前驱体收入持续提升、平台盈利兑现; 隐性接受现价 70-87x forward。方正证券给 2026-28 归母 13.51/17.35/22.14亿(强烈推荐)。",
    "assumptions": [
        ("2026E EPS", "共识 2.71-2.84(近端窄样本 2.71, 东财 8 家 2.84)。", "分歧小: 前驱体+光刻胶放量确定, 争议在 LNG 交付节奏与综合毛利率。", "Base 链算出 2026E EPS ~2.56, 略低于共识(前驱体强度给保守 0.58%, LNG 增速给 25%)。"),
        ("2027E EPS", "共识 3.39-3.45。", "最大分歧: 前驱体能否借 HBM4 High-K + 国产化再加速(2025 仅 +8%), 江苏先科并表节奏。", "Base 取前驱体强度 0.66% → EPS ~3.12, 略低于共识(不假设 HBM 国产化一步到位)。"),
        ("前驱体增速", "券商: HBM 高景气驱动前驱体持续提升, 稼动率 59%→80% 弹性 + 宜兴 2000吨 + 江苏先科回购。", "2025 前驱体仅 +8%(韩国 UP -41%): 一次性去库存 vs 结构性放缓。", "前驱体=DRAM/HBM capex×强度; 强度锚 2025A 0.55%, Base 缓升至 0.66%(2027)。"),
        ("目标倍数", "A股无目标价; 现价隐含 forward 70x(2027E)/87x(2026E)。", "国产替代主题给多少 vs 材料公司历史 28-40x。", "三层分解: 历史中枢 28x × HBM franchise 结构溢价 1.5x × 情绪值, Base 2027E ~46x, 见『估值倍数假设』。"),
    ],
    "divergences": [
        "① 前驱体强度路径: Bull 按 HBM4 High-K 用量倍增 + 以钼代钨新品类 + 国产化份额突破 + 稼动率 85% 给强度 0.9%+; Bear 按韩国 UP 份额被 DNF/Soulbrain 切 + 稼动率停滞给 0.48%——决定前驱体 2027 是 33亿还是 47亿。",
        "② 倍数 vs 现价: 现价 70x forward(2027E)远高于自身历史 28-40x; 本模型不跟随现价, 按历史带 × HBM 结构溢价拍, Base 46x 低于现价隐含。这是与'板块 beta 定价'的分歧。",
        "③ 平台剥离: 卖方常按集团一个 PE 估, 本模型强调 LNG(27.5%)+化工(13.5%)+显示光刻胶应折价, 只有前驱体享 HBM 溢价——SOTP 视角下集团合理倍数低于'纯半导体材料'认知。",
    ],
    "stances": [
        "华鑫证券(买入, 2026-06-29): 半导体高景气度下, 主线景气转化平台盈利; 点名钼前驱体放量(三星 Mo 4→10→80吨)。",
        "中银证券(买入, 2026-05-08): 业绩稳健增长, 前驱体收入持续提升。",
        "方正证券(强烈推荐): 2026-28 归母 13.51/17.35/22.14亿, 前驱体+平台化双驱动。",
    ],
})

# ════════════ 5. 历史估值倍数 ════════════
hm = K.write_hist_multiples(wb.create_sheet(S_HMULT), {
    "title": "历史估值倍数 — 自身历史带 + 当下 + 同业对照(forward P/E 口径)",
    "intro": "①雅克自己历史值多少(forward P/E 带 28-40x, 2026H1 重估至 70-87x) ②现在市场给多少 ③A股半导体材料同业光谱。看完再去下一页拍三案倍数。",
    "s_hist": S_HIST, "ha": ha, "hist_cols": HC, "hist_years": HY,
    "yhigh": YHIGH, "ylow": YLOW,
    "fwd_note": "forward P/E ≈87x(现价÷2026E EPS 2.71) / ≈70x(÷2027E 3.39); 历史年末 forward(年末价÷次年EPS): 2021末74x/2022末41x/2023末30x/2024末28x/2025末27x → 剔 2021 泡沫取中枢 ~28-30x",
    "self_name": "雅克科技",
    "self_fwd_pe_label": "≈70x (2027E)",
    "self_note": "本模型标的; forward 推导见『情景估值』。P/B 14.1x 由并购商誉(21.1亿)+重资产驱动, 仅支线, P/E 主线。",
    "peers": [
        {"name": "南大光电 (300346)", "yearly": None, "cur_pb": 18.0, "cur_pe": 176.0, "fwd_pe": 138.7,
         "note": "最直接可比(特气+光刻胶+前驱体); forward 138.7x, 前驱体国内追赶者(客户局限国内)。2026-07-01。"},
        {"name": "安集科技 (688019)", "yearly": None, "cur_pb": 16.3, "cur_pe": 89.2, "fwd_pe": 53.6,
         "note": "CMP抛光液龙头, 优质耗材标杆(毛利 56.7%/净利+47%), forward 53.6x = 优质半导体材料档下沿。"},
        {"name": "鼎龙股份 (300054)", "yearly": None, "cur_pb": 17.8, "cur_pe": 115.1, "fwd_pe": 89.3,
         "note": "CMP垫+显示材料平台, forward 89.3x, 平台型可比。"},
        {"name": "彤程新材 (603650)", "yearly": None, "cur_pb": 14.2, "cur_pe": 97.5, "fwd_pe": 79.6,
         "note": "KrF光刻胶国内龙头, forward 79.6x。"},
        {"name": "国际前驱体龙头(默克/法液空/Entegris)", "yearly": None, "cur_pb": None, "cur_pe": None, "fwd_pe": 22.0,
         "note": "全球前驱体 CR3, forward 15-30x(成熟大盘, 个位数增长)——光谱绝对下沿。约数 2026-06。"},
    ],
    "reading": "① 自己: 历史 forward 中枢 28-30x(材料平台), 2026H1 重估到 70-87x = 历史 2.5-3 倍。② A股光谱: 雅克 forward 70x 在优质耗材子集(安集54/彤程80/鼎龙89)中低位, 但整组被国产替代主题抬高; 对国际前驱体龙头(22x)则是 3 倍。③ 结论: 高绝对 PE 是板块 beta 不是个股独贵, 但目标倍数应锚自身历史中枢 × HBM 结构溢价(前驱体真实、但仅占营收 1/4), 而非接受当下 70x 作新常态。→ 下一页三层分解。",
})

# ════════════ 6. 估值倍数假设 ════════════
ma = K.write_multiple_assumptions(wb.create_sheet(S_MULT), {
    "title": "估值倍数假设 — P/E 主线(forward) + 三案目标倍数",
    "intro": "镜头判断+三层分解。『情景切换』引用并切换, 『情景估值』套用当前案, 『估值对比』三案并排。倍数口径: 目标年 2027E 的 forward P/E(× 2027E EPS)。",
    "why_text": ("镜头选择是业务判断: 雅克'穿越周期持续存在的东西'是盈利质量(前驱体 HBM franchise: UP Chemical 海外客户认证 + 44.8% 高毛利 + 稼动率经营杠杆), 不是单纯资产(虽重资产, 但 P/B 14x 被并购商誉扭曲)→ P/E 主线(资本化盈利)。"
                 "它不是纯商品周期(无定价权)——前驱体认证壁垒给定价权; 但也不是全生态垄断, 且集团 3/4 营收(LNG/光刻胶/化工)是较低质量的周期/成长业务。盈利耐用性 = 优质半导体材料 franchise(前驱体)+ 周期底盘(LNG/化工)的混合, 所以倍数锚自己历史中枢 × 有限 HBM 结构溢价, 而非国产替代主题上沿。"
                 "支线用 DCF: 检验'现价隐含多少年高增长', 防 P/E 倍数与 EPS 双乐观。"),
    "why_rows": 6,
    "method_text": "三层分解: ①历史 forward 中枢 28x(2022-25末价÷次年EPS = 41/30/28/27x, 剔 2021 泡沫取 28x) × ②HBM franchise 结构溢价 1.5x(前驱体从'国产替代概念'升格为'绑定 HBM 三巨头的高毛利卡位'; 但仅占营收 1/4, 故给 1.5x 而非纯半导体材料的 2x+; 对账: 安集 forward 54x≈材料中枢×2, 雅克给 1.5 因平台含金量稀释) × ③情绪值(三案)。一致性检验: 28×1.5×2.0≈84x ≈ 当下市场对 2026E 实付的 87x, 量级复现 ✓ → 当下情绪 ≈2.0, 过热档。",
    "peak": 28.0, "peak_note": "第一层 历史 forward 中枢: 真实年末价 ÷ 次年 EPS 逐年反推(2021末74x/2022末41x/2023末30x/2024末28x/2025末27x), 剔 2021 泡沫值, 取 2022-25 中枢 28x。刻意不用本轮 re-rating 后的 70-87x 当锚(那已含结构溢价+情绪, 再乘=双重计算)。",
    "premium": 1.5, "premium_note": "第二层 HBM franchise 结构溢价: 前驱体绑定 SK海力士/三星/美光 HBM 供应链(UP Chemical 认证壁垒 + 44.8% 毛利 + 稼动率 59% 经营杠杆), 从周期材料升格。对账线: 优质半导体耗材(安集)forward ~54x ≈ 材料中枢 28x × 1.9; 雅克给 1.5x(低于安集, 因前驱体仅占营收 24.5%, LNG/光刻胶/化工稀释平台质量)。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "hpb_row": ha["HPE"],
    "cases": [
        ("Bear", [1.05, 0.85, 0.78, 0.75, 0.72], "国产替代板块退潮 + 前驱体份额被韩系(DNF/Soulbrain)切 + 稼动率停滞 → 市场把雅克打回'材料平台'折价档(2027E P/E ~36x, 28×1.5×0.85), 参照自身 2024 年 forward 28x + 部分 HBM 溢价残留。"),
        ("Base", [1.60, 1.10, 0.95, 0.90, 0.88], "前驱体借 HBM4 缓步放量、板块 froth 挤出 → 2027E 付 46x(28×1.5×1.10), 之后随成长成熟向'中枢×溢价'(42x)收敛。"),
        ("Bull", [2.00, 1.60, 1.35, 1.20, 1.10], "HBM4 High-K 用量倍增 + 以钼代钨兑现 + 国产化份额突破 + 稼动率 85% + 江苏先科并表, 市场维持国产替代溢价, 2027E 给 67x(28×1.5×1.60)。"),
    ],
    "sent_note": "情绪值=板块/情绪位置。1.0=付足『中枢×溢价』(42x); >1=国产替代 FOMO 超涨; <1=板块折价。历史列=实际 forward P/E÷(中枢×溢价)反推(2021-25末: 1.76/0.98/0.72/0.67/0.65), 显示 2026 重估前市场只给 <1(无 HBM 溢价, 2021 泡沫除外), 当下升到 ~2.0。",
    "reconcile_text": "A股无目标价聚合; 现价隐含 forward 70x(2027E)/87x(2026E)。本模型三案 36-67x 覆盖并低于现价隐含, Base 46x——分歧不在业务(前驱体绑 HBM 真实)而在倍数: 现价对 2027E 付 70x 已超'中枢×溢价'(42x)的 67%, 这部分是 2026H1 国产替代+龙虎榜情绪涨出来的。敢给低于现价的 Base: 事实是自身历史 28-40x + 有限 HBM 溢价, 逻辑链是 70x 把 2029-30 盈利提前定价到今天, 且前驱体仅占营收 1/4。",
    "source_text": "第一层=统一数据 API 真实年末价 ÷ 年报 EPS; 第二层=安集/南大 forward 对账(2026-07-01); 第三层档位依据『综合判断仪表盘』D 块。",
})

# ════════════ 7. 情景切换 ════════════
sw = K.write_scenario_switch(wb.create_sheet(S_SW), {
    "title": "情景切换 — 全模型唯一的情景参数库 + 切换开关 (默认 Base)",
    "usage": ("怎么用: B2 是唯一入口——下拉选案 → 各杠杆『当前案』行跟着切 → 整条链(锚→测算→利润→倍数→估值)变档。"
              "三案对比不用切: 『估值对比』恒常三列并排。情景参数只在本页改(蓝字); 未列入的假设(光刻胶/特气/硅微粉/化工增速与净利率)三案共用 Base。"),
    "cases": CASES, "default": "Base",
    "triggers": [
        ("Bear", "国产替代板块退潮 / 前驱体强度回落(韩系切份额, 稼动率停滞) / LNG 交付放缓 → HBM 叙事降温, 雅克重回'材料平台'定价, 倍数向 36x 收敛。"),
        ("Base", "前驱体强度 0.55%→0.66% 缓升(HBM4 + 国产化温和推进) + LNG 订单交付兑现; froth 从 70x 回归中枢×溢价 46x。"),
        ("Bull", "HBM4 High-K 倍增 + 以钼代钨放量 + 国产化份额突破 + 稼动率 85% + 江苏先科并表 → 前驱体强度突破 0.9%, 市场维持国产替代溢价 67x。"),
    ],
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "levers": [
        {"key": "capex", "name": "DRAM/HBM capex 增速", "fmt": K.PCT, "cols": FC[2:],
         "vals": {"Bear": [0.05, 0.03, 0.02, 0.02], "Base": [0.12, 0.08, 0.05, 0.04], "Bull": [0.18, 0.12, 0.08, 0.06]},
         "desc": "物理锚: 前驱体收入挂它。2026E=$61.3B 三案共用(TrendForce, DRAM capex 2025 $53.7B→2026 +14%); 分歧在 2027+: Base=HBM 驱动温和扩张, Bull=HBM4/DDR5 超级周期, Bear=存储 capex 见顶。",
         "stories": {"Bear": "存储涨价周期见顶, 三星/海力士/美光收缩扩产。", "Base": "HBM 拉动 DRAM capex 温和续增(+12/8/5/4%)。", "Bull": "HBM4 + AI DRAM 需求超预期, capex 高增。"},
         "hist": [None, 1.14, 0.66, 1.90, 1.34]},
        {"key": "qint", "name": "前驱体收入强度 (%)", "fmt": K.PCT,
         "vals": {"Bear": [0.0053, 0.0052, 0.0050, 0.0049, 0.0048],
                  "Base": [0.0058, 0.0066, 0.0072, 0.0075, 0.0077],
                  "Bull": [0.0062, 0.0078, 0.0092, 0.0100, 0.0105]},
         "desc": "前驱体收入 ÷ DRAM/HBM capex(=前驱体占存储材料份额 × 雅克份额 × HBM/High-K 单位用量), 锚 2025A=0.294/53.7=0.55%。前瞻四因子推升: HBM4 High-K 用量↑ + 以钼代钨新品类 + 国产化份额↑ + 稼动率 59%→80%。2023 capex 低谷÷分母 使强度虚高(0.77%), 读作 n.m.。",
         "stories": {"Bear": "韩系(DNF/Soulbrain)切份额 + 稼动率停滞, 强度回落 0.48% → 前驱体 2027 ~33亿。", "Base": "HBM4 温和放量 + 国产化推进, 强度 0.55%→0.72% → 前驱体 2027 ~34亿(+62% vs 2025)。", "Bull": "HBM4 High-K 倍增 + 钼放量 + 国产化突破 + 稼动率 85%, 强度 0.92% → 前驱体 2027 ~44亿。"},
         "hist": [None, None, 0.0077, 0.0068, 0.0055]},
        {"key": "qopm", "name": "前驱体 净利率", "fmt": K.PCT,
         "vals": {"Bear": [0.26, 0.25, 0.24, 0.24, 0.24],
                  "Base": [0.28, 0.29, 0.30, 0.31, 0.31],
                  "Bull": [0.30, 0.32, 0.34, 0.35, 0.36]},
         "desc": "前驱体段净利率(由毛利 44.79% 扣段费用估)。历史: 2023 ~20%/2024 ~24%/2025 ~28%。稼动率 59%→80% 经营杠杆 + HBM 高值 mix 推升。",
         "stories": {"Bear": "韩系价格战 + 稼动率停滞, 卡 24%。", "Base": "稼动率爬坡: 28%→31%。", "Bull": "HBM High-K/钼高值品类 mix + 稼动率 85%, 36%。"},
         "hist": [None, None, 0.20, 0.24, 0.28]},
        {"key": "lngg", "name": "LNG保温 增速", "fmt": K.PCT,
         "vals": {"Bear": [0.10, 0.05, 0.00, -0.05, -0.05],
                  "Base": [0.25, 0.15, 0.08, 0.05, 0.03],
                  "Bull": [0.35, 0.25, 0.15, 0.08, 0.05]},
         "desc": "LNG 保温板材收入(独立造船周期), 在手订单 >62亿(5年>120亿)。2025 23.70亿(+44.9%)。收入靠存量订单交付节奏, 与半导体无关。",
         "stories": {"Bear": "联洋新材价格竞争 + 交付放缓, 见顶回落。", "Base": "在手订单稳定交付, 增速从 45% 逐步 normalize。", "Bull": "卡塔尔百船 + 中国 LNG 船份额破 30%, 订单持续。"},
         "hist": [None, 0.83, 0.57, 0.89, 0.45]},
    ],
    "linked": [
        {"key": "sent", "name": "情绪值(倍数第三层)", "fmt": K.N2,
         "src_sheet": S_MULT, "src_row0": ma["sent_row0"],
         "note": "三案取值与依据见『估值倍数假设』; 本页只做切换。"},
    ],
})
_pk = f"'{S_MULT}'!{ma['pk_cell']}"
_pr = f"'{S_MULT}'!{ma['pr_cell']}"
_sent_act = sw["SWACT"]["sent"]
_r = sw["next_row"]
K.lab(wb[S_SW], f"A{_r}", "目标P/E(当前案)", b=True)
for _c in ALLC:
    K.fml(wb[S_SW], f"{_c}{_r}", f"={_pk}*{_pr}*{_c}{_sent_act}", K.MX, link=True)
K.logic(wb[S_SW], f"L{_r}", "= 历史 forward 中枢 28x × HBM 结构溢价 1.5x × 当前案情绪值 → 喂『情景估值』的前瞻 P/E。")
SWPE = _r

# ════════════ 8. 物理锚 ════════════
anchor = K.write_anchor(wb.create_sheet(S_ANCHOR), {
    "title": "全球 DRAM/HBM 前道 CapEx ($B) — 前驱体需求物理盘子",
    "all_cols": ALLC, "all_years": ALLY,
    "series": [("DRAM/HBM capex ($B)",
                [28, 32, 21, 40, 53.7, 61.3, None, None, None, None],
                "2021-24=行业估(存储下行周期 2023 低谷$21B); 2025A=$53.7B / 2026E=$61.3B(+14%, TrendForce; 三案共用); 之后=锚×当前案增速", K.N1)],
    "yoy_row": "DRAM/HBM capex ($B)",
    "source_note": "口径=全球 DRAM(含 HBM)前道设备/产能 capex, 三星/SK海力士/美光合计。来源: 2025-26 TrendForce; 前瞻=锚×当前案增速(『情景切换』)。前驱体服务 DRAM/NAND 前道, 用 DRAM/HBM capex 作物理盘子。",
    "role_note": "作用: 前驱体收入 = capex × 收入强度挂在它上面。改 capex → 前驱体收入 → 净利 → 隐含价全链动。",
})
CAPEX_ROW = anchor["row_of"]["DRAM/HBM capex ($B)"]
for _i, _c in enumerate(FC[2:]):
    K.fml(wb[S_ANCHOR], f"{_c}{CAPEX_ROW}", f"={FC[1:][_i]}{CAPEX_ROW}*(1+{K.R(S_SW, _c + str(sw['SWACT']['capex']))})", K.N1, link=True)

# ════════════ 9. 分部测算(前驱体=capex×强度; 其余段增速在利润表) ════════════
seg = K.write_segment_model(wb.create_sheet(S_SEG), {
    "title": "分部测算 — 前驱体(capex×强度, AI/HBM直驱) ($B)",
    "all_cols": ALLC, "all_years": ALLY, "logic_col": "N",
    "groups": [
        ("DRAM/HBM capex 物理锚", [
            ("DRAM/HBM capex ($B)", None, K.N1, "= 引自『DRAM_HBM_Capex预测』。改 capex, 前驱体收入跟着动。"),
        ]),
        ("前驱体 = capex × 收入强度", [
            ("前驱体 收入强度 (%)", None, K.PCT,
             "历史=实际前驱体收入÷当年 capex(公式反推): 2025A 0.55% ← 锚; 2023 capex低谷使强度虚高(0.77%) n.m.。前瞻=『情景切换』当前案。"),
            ("前驱体 收入 ($B)", None, K.N2, "历史取实数; 前瞻=capex×强度。喂『利润与收入假设』。"),
        ]),
    ],
})
m = seg["m"]
Q_HROW = ha["seg_rows"]["前驱体 收入"]
for col in ALLC:
    K.fml(wb[S_SEG], f"{col}{m['DRAM/HBM capex ($B)']}", f"={K.R(S_ANCHOR, col + str(CAPEX_ROW))}", K.N1, link=True)
for col in HC:
    K.fml(wb[S_SEG], f"{col}{m['前驱体 收入强度 (%)']}", f"={K.R(S_HIST, col + str(Q_HROW))}/{col}{m['DRAM/HBM capex ($B)']}", K.PCT, link=True)
    K.fml(wb[S_SEG], f"{col}{m['前驱体 收入 ($B)']}", f"={K.R(S_HIST, col + str(Q_HROW))}", K.N2, link=True)
for col in FCf:
    K.fml(wb[S_SEG], f"{col}{m['前驱体 收入强度 (%)']}", f"={K.R(S_SW, col + str(sw['SWACT']['qint']))}", K.PCT, link=True)
    K.fml(wb[S_SEG], f"{col}{m['前驱体 收入 ($B)']}", f"={col}{m['DRAM/HBM capex ($B)']}*{col}{m['前驱体 收入强度 (%)']}", K.N2)
for col in FCf:
    wb[S_SEG][f"{col}{m['前驱体 收入 ($B)']}"].fill = K.OUT

# ════════════ 10. 利润与收入假设 ════════════
fr = K.write_fundamentals(wb.create_sheet(S_FUND), {
    "title": "利润与收入假设 — 5段增速 + 段净利率 + 净利转换 + 分部营收→归母净利→EPS/BPS",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
    "assum_groups": [
        ("营收驱动(前驱体在分部测算; 此处 LNG(情景联动) + 其余4段增速)", [
            {"name": "LNG保温 增速", "vals": [None, 0.83, 0.57, 0.89, 0.45, None, None, None, None, None],
             "fmt": K.PCT, "logic": "LNG 保温板材增速(独立造船周期); 前瞻=『情景切换』当前案(在手订单 62亿交付节奏)。", "nm_cols": ["B"],
             "link": {"sheet": S_SW, "row": sw["SWACT"]["lngg"]}},
            {"name": "光刻胶 增速", "vals": [None, 0.09, -0.01, 0.16, 0.28, 0.18, 0.15, 0.12, 0.10, 0.08],
             "fmt": K.PCT, "logic": "显示胶为主+半导体胶国产替代期权; 前瞻中高增速回落。三案共用。"},
            {"name": "电子特气 增速", "vals": [None, 0.09, 0.05, 0.05, -0.11, 0.10, 0.10, 0.08, 0.06, 0.05],
             "fmt": K.PCT, "logic": "SF6/CF4, 2025 -11%; 内蒙 1.5万吨 2026Q3 投产驱动恢复。三案共用。"},
            {"name": "硅微粉LDS 增速", "vals": [None, 0.31, -0.12, 0.66, 0.08, 0.20, 0.18, 0.15, 0.12, 0.10],
             "fmt": K.PCT, "logic": "华飞球硅 + LDS 设备, 先进封装拉动; 前瞻中速增长。三案共用。"},
            {"name": "化工其他 增速", "vals": [None, -0.25, -0.18, 0.08, 0.62, 0.05, 0.05, 0.05, 0.05, 0.05],
             "fmt": K.PCT, "logic": "阻燃剂+租赁工程+其他, 2025 因 LNG 安装工程 +122% 抬高; 前瞻稳态 5%。三案共用。"},
        ]),
        ("段净利率(历史实际锚 + 前瞻)", [
            {"name": "前驱体 净利率", "vals": [None, None, 0.20, 0.24, 0.28, None, None, None, None, None],
             "fmt": K.PCT, "logic": "前驱体段净利率(毛利 44.79% 扣段费用估); 历史 2023~20%/2024~24%/2025~28%; 前瞻=『情景切换』当前案(稼动率 59%→80% 经营杠杆)。", "nm_cols": ["B", "C"],
             "link": {"sheet": S_SW, "row": sw["SWACT"]["qopm"]}},
            {"name": "LNG保温 净利率", "vals": [0.15, 0.15, 0.15, 0.14, 0.13, 0.13, 0.12, 0.11, 0.11, 0.10],
             "fmt": K.PCT, "logic": "毛利率连续下滑(35.5%→25.3%)→净利率随之降; 联洋新材价格竞争 + 产能爬坡摊薄。三案共用。"},
            {"name": "光刻胶 净利率", "vals": [0.08, 0.08, 0.09, 0.08, 0.07, 0.07, 0.075, 0.08, 0.085, 0.09],
             "fmt": K.PCT, "logic": "毛利仅 16.72%(显示胶为主), 净利率低; 半导体胶 mix 提升缓慢改善。"},
            {"name": "电子特气 净利率", "vals": [0.12, 0.12, 0.12, 0.12, 0.12, 0.12, 0.12, 0.12, 0.12, 0.12],
             "fmt": K.PCT, "logic": "毛利 30.22%, 段净利率约 12%, 稳态。"},
            {"name": "硅微粉LDS 净利率", "vals": [0.09, 0.09, 0.09, 0.10, 0.10, 0.10, 0.11, 0.11, 0.12, 0.12],
             "fmt": K.PCT, "logic": "华飞球硅中低端 mix, 净利率约 10%, 高端化缓升。"},
            {"name": "化工其他 净利率", "vals": [0.06, 0.07, 0.08, 0.08, 0.08, 0.08, 0.08, 0.08, 0.08, 0.08],
             "fmt": K.PCT, "logic": "阻燃剂+工程, 稳态约 8%。"},
        ]),
        ("净利转换与留存", [
            {"name": "净利转换率(归母净利/分部净利和)", "vals": [None, None, 0.80, 0.80, 0.808, 0.81, 0.81, 0.81, 0.81, 0.81],
             "fmt": K.PCT, "logic": "分部净利和 → 归母净利: 扣未分配公司费用/利息/少数股东损益(江苏先科 33.07%)。历史反推 ~0.81。", "nm_cols": ["B", "C"]},
            {"name": "留存率", "vals": [0.90, 0.90, 0.90, 0.88, 0.85, 0.85, 0.85, 0.85, 0.85, 0.85],
             "fmt": K.PCT, "logic": "低派息(2025 每10股派3.10元)扩产优先; 留存率 ~85%。仅用于 BPS 递推, P/E 主线不用 P/B。"},
        ]),
    ],
    "segments": [
        {"name": "前驱体 收入", "hist_row": "前驱体 收入", "fwd": {"sheet": S_SEG, "row": m["前驱体 收入 ($B)"]}},
        {"name": "LNG保温 收入", "hist_row": "LNG保温 收入", "fwd": {"growth": "LNG保温 增速"}},
        {"name": "光刻胶及配套 收入", "hist_row": "光刻胶及配套 收入", "fwd": {"growth": "光刻胶 增速"}},
        {"name": "电子特气 收入", "hist_row": "电子特气 收入", "fwd": {"growth": "电子特气 增速"}},
        {"name": "硅微粉及LDS 收入", "hist_row": "硅微粉及LDS 收入", "fwd": {"growth": "硅微粉LDS 增速"}},
        {"name": "化工及其他 收入", "hist_row": "化工及其他 收入", "fwd": {"growth": "化工其他 增速"}},
    ],
    "profit_terms": [
        (["前驱体 收入"], "前驱体 净利率", True),
        (["LNG保温 收入"], "LNG保温 净利率", False),
        (["光刻胶及配套 收入"], "光刻胶 净利率", False),
        (["电子特气 收入"], "电子特气 净利率", False),
        (["硅微粉及LDS 收入"], "硅微粉LDS 净利率", False),
        (["化工及其他 收入"], "化工其他 净利率", False),
    ],
    "conv_assum": "净利转换率(归母净利/分部净利和)", "retention_assum": "留存率",
    "note_text": "分部营收(前驱体=capex×强度 + LNG周期 + 4段增速)→ 段净利率 → 分部净利和 → ×净利转换率(扣未分配费用+少数股东)→ 归母净利 → EPS(股本476mn)/BPS/ROE。闭环检查: 前驱体占营收% + 净利率应逐年上行(mix 改善), 与 thesis 同向。下游『情景估值』直接引本表每股。",
})

# ════════════ 11. 情景估值(P/E 主线 + DCF 支线) ════════════
fr_pe = dict(fr); fr_pe["BPS"] = fr["EPS"]
sv = K.write_scenario_valuation(wb.create_sheet(S_VAL), {
    "title": "情景估值 — 当前案的逐年隐含价 (P/E 主线; DCF 交叉验证)",
    "intro": "本表输出=『情景切换』当前案(默认Base)。主线: 隐含价=目标P/E(当前案)×前瞻 EPS。历史列用实际年末价反推倍数(事实); 前瞻是预测、不拟合现价。三案并排见『估值对比』。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf, "hist_years": HY,
    "fx_fwd": FX_FC, "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
    "s_fund": S_FUND, "fr": fr_pe,
    "s_switch": S_SW, "target_row": SWPE, "sw_cell": "B2",
    "yend": YEND, "yavg": YAVG,
    "reading": "主线行读法: 本表『目标倍数/隐含股价』两行即 P/E 主线(历史列=年末价÷当年EPS=实际P/E; 前瞻=目标P/E×当年EPS)。Base 隐含价轨迹显示现价 ¥236 在 Base 各年隐含价之上——市场把数年后的盈利提前定价到今天。",
    "method": "方法: 整体公司、P/E 主线逐年估。基本面在『利润与收入假设』; 目标倍数在『估值倍数假设』(三层分解); 本表只做最后一步: 目标P/E × 前瞻 EPS → 隐含价 + 市值; 下方 DCF 块做支线体检。",
    "concl": "结论(方向性): Base 目标价(2027E 目标年)vs 现价 ¥236 见『估值对比』。现价隐含情绪 ~2.0(过热), 评级基于价格而非质地——前驱体 franchise 真实且在国产替代主线上。",
})
# ── DCF 交叉验证块(支线) ──
_ws = wb[S_VAL]
_r2 = _ws.max_row + 2
K.band(_ws, _r2, "DCF 交叉验证(支线) — 现价隐含多长的高增长? (FCF≈归母净利×转换率, WACC 贴现 + Gordon 终值)", 11); _r2 += 1
K.lab(_ws, f"A{_r2}", "WACC"); K.inp(_ws, f"B{_r2}", 0.095, None, K.PCT)
K.logic(_ws, f"D{_r2}", "贴现率: A股材料成长股 9-10%, 取 9.5%(beta~1.3, 有息负债 43亿, 扩张期)。")
WACC_C = f"$B${_r2}"; _r2 += 1
K.lab(_ws, f"A{_r2}", "永续增长 g"); K.inp(_ws, f"B{_r2}", 0.03, None, K.PCT)
K.logic(_ws, f"D{_r2}", "2030 后永续 3%(名义 GDP-), 假设 2030 后高增长结束——支线保守锚。")
G_C = f"$B${_r2}"; _r2 += 1
K.lab(_ws, f"A{_r2}", "FCF 转换率(FCF/归母净利)"); K.inp(_ws, f"B{_r2}", 0.50, None, K.PCT)
K.logic(_ws, f"D{_r2}", "重资产扩张期: 2025 OCF 10.3亿 vs capex ~15亿(在建10.66亿); 归母净利→FCF 取 50%(扣产能扩张 capex + NWC)。")
FCFC_C = f"$B${_r2}"; _r2 += 1
K.lab(_ws, f"A{_r2}", "净债务 ($B)"); K.inp(_ws, f"B{_r2}", 0.24, None, K.N2)
K.logic(_ws, f"D{_r2}", "2025末: 有息负债 43.3亿(短8.89+长34.46) − 货币资金 25.89亿 = 净债 17.4亿 CNY ≈ $0.24B@7.15。")
ND_C = f"$B${_r2}"; _r2 += 1
NIr2 = lambda c: K.R(S_FUND, c + str(fr["NI"]))
K.lab(_ws, f"A{_r2}", "FCF ($B, 当前案)", note=True)
for _j, _c in enumerate(FCf):
    K.fml(_ws, f"{_c}{_r2}", f"={NIr2(_c)}*{FCFC_C}", K.N2, link=True)
FCF_R = _r2; _r2 += 1
K.lab(_ws, f"A{_r2}", "贴现因子", note=True)
for _j, _c in enumerate(FCf):
    K.fml(_ws, f"{_c}{_r2}", f"=1/(1+{WACC_C})^{_j+1}", K.N2)
DF_R = _r2; _r2 += 1
K.lab(_ws, f"A{_r2}", "DCF 隐含价 (¥/股, 支线)", b=True); _ws[f"A{_r2}"].fill = K.OUT
_pv_terms = "+".join(f"{c}{FCF_R}*{c}{DF_R}" for c in FCf)
_tv = f"{FCf[-1]}{FCF_R}*(1+{G_C})/({WACC_C}-{G_C})*{FCf[-1]}{DF_R}"
SH_F = K.R(S_HIST, f"$F${ha['HSH']}")
K.fml(_ws, f"G{_r2}", f"=(({_pv_terms})+{_tv}-{ND_C})*{FX_FC}*1000/{SH_F}", K.PX)
K.logic(_ws, f"I{_r2}", "=(Σ显性期 FCF 贴现 + 终值贴现 − 净债务)×汇率÷股本。若 2030 后只剩 3% 永续, 整盘生意 DCF 地板值即为此——现价 ¥236 与之的差额, 是市场对'2030 后前驱体仍高速复合'的预期。三角验证: 现价 vs DCF 地板 vs P/E 主线隐含价。")
DCF_R = _r2

# ════════════ 12. 估值对比 ════════════
SWB = sw["SWB"]
PX_NOW_REF = K.R(S_HIST, f"G{ha['HPX']}")
_conv = fr["am"]["净利转换率(归母净利/分部净利和)"]
_lithg = fr["am"]["光刻胶 增速"]
_gasg = fr["am"]["电子特气 增速"]
_powg = fr["am"]["硅微粉LDS 增速"]
_chmg = fr["am"]["化工其他 增速"]
_lngm = fr["am"]["LNG保温 净利率"]
_lithm = fr["am"]["光刻胶 净利率"]
_gasm = fr["am"]["电子特气 净利率"]
_powm = fr["am"]["硅微粉LDS 净利率"]
_chmm = fr["am"]["化工其他 净利率"]
Q_H = ha["seg_rows"]["前驱体 收入"]
LNG_H = ha["seg_rows"]["LNG保温 收入"]
LITH_H = ha["seg_rows"]["光刻胶及配套 收入"]
GAS_H = ha["seg_rows"]["电子特气 收入"]
POW_H = ha["seg_rows"]["硅微粉及LDS 收入"]
CHM_H = ha["seg_rows"]["化工及其他 收入"]


def _fwdprev(j, A, key):
    return (HC[-1] if j == 0 else FCf[j - 1]) + str(A[key])


cmp_rows = [
    {"key": "cap", "label": "DRAM/HBM capex ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_ANCHOR, c + str(CAPEX_ROW))}",
     "fwd": lambda c, j, ci, A: (f"={K.R(S_ANCHOR, 'G' + str(CAPEX_ROW))}" if j == 0
                                 else f"={FCf[j-1]}{A['cap']}*(1+{K.R(S_SW, c + str(SWB['capex'] + ci))})")},
    {"key": "q", "label": "前驱体 ($B)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(Q_H))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['cap']}*{K.R(S_SW, c + str(SWB['qint'] + ci))}"},
    {"key": "lng", "label": "LNG保温 ($B)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(LNG_H))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'lng')}*(1+{K.R(S_SW, c + str(SWB['lngg'] + ci))})"},
    {"key": "lith", "label": "光刻胶 ($B)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(LITH_H))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'lith')}*(1+{K.R(S_FUND, c + str(_lithg))})"},
    {"key": "gas", "label": "电子特气 ($B)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(GAS_H))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'gas')}*(1+{K.R(S_FUND, c + str(_gasg))})"},
    {"key": "pow", "label": "硅微粉LDS ($B)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(POW_H))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'pow')}*(1+{K.R(S_FUND, c + str(_powg))})"},
    {"key": "chm", "label": "化工其他 ($B)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(CHM_H))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'chm')}*(1+{K.R(S_FUND, c + str(_chmg))})"},
    {"key": "rev", "label": "总收入 ($B)", "fmt": K.N2, "bold": True,
     "hist": lambda c, ci, A: f"={c}{A['q']}+{c}{A['lng']}+{c}{A['lith']}+{c}{A['gas']}+{c}{A['pow']}+{c}{A['chm']}",
     "fwd": lambda c, j, ci, A: f"={c}{A['q']}+{c}{A['lng']}+{c}{A['lith']}+{c}{A['gas']}+{c}{A['pow']}+{c}{A['chm']}"},
    {"key": "ni", "label": "归母净利 ($B)", "fmt": K.N2, "bold": True,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HNI']))}",
     "fwd": lambda c, j, ci, A: (f"=({c}{A['q']}*{K.R(S_SW, c + str(SWB['qopm'] + ci))}"
                                 f"+{c}{A['lng']}*{K.R(S_FUND, c + str(_lngm))}"
                                 f"+{c}{A['lith']}*{K.R(S_FUND, c + str(_lithm))}"
                                 f"+{c}{A['gas']}*{K.R(S_FUND, c + str(_gasm))}"
                                 f"+{c}{A['pow']}*{K.R(S_FUND, c + str(_powm))}"
                                 f"+{c}{A['chm']}*{K.R(S_FUND, c + str(_chmm))})*{K.R(S_FUND, c + str(_conv))}")},
    {"key": "eps", "label": "EPS (¥)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={c}{A['ni']}*{FX_FC}*1000/{K.R(S_HIST, c + str(ha['HSH']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['ni']}*{FX_FC}*1000/{SH_F}"},
    {"key": "sent", "label": "情绪值(该案)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_MULT, c + str(ma['sent_row0'] + ci))}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_MULT, c + str(ma['sent_row0'] + ci))}"},
    {"key": "pe", "label": "目标P/E(该案)", "fmt": K.MX,
     "hist": lambda c, ci, A: f"={_pk}*{_pr}*{c}{A['sent']}",
     "fwd": lambda c, j, ci, A: f"={_pk}*{_pr}*{c}{A['sent']}"},
    {"key": "px", "label": "隐含价 (¥)", "fmt": K.PX, "bold": True, "out": True,
     "hist": lambda c, ci, A: f"={c}{A['pe']}*{c}{A['eps']}",
     "fwd": lambda c, j, ci, A: f"={c}{A['pe']}*{c}{A['eps']}"},
    {"key": "ipe", "label": "隐含 forward P/E(现价÷该案EPS 体检)", "fmt": K.MX,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HPX']))}/{c}{A['eps']}",
     "fwd": lambda c, j, ci, A: f"={PX_NOW_REF}/{c}{A['eps']}"},
    {"key": "up", "label": "历史: vs 实际年末价 / 前瞻: vs 现价", "fmt": K.PCT,
     "hist": lambda c, ci, A: f"={c}{A['px']}/{K.R(S_HIST, c + str(ha['HPX']))}-1",
     "fwd": lambda c, j, ci, A: f"={c}{A['px']}/{PX_NOW_REF}-1"},
]
cm_sheet = K.write_comparison(wb.create_sheet(S_CMP), {
    "title": "估值对比 — Bear / Base / Bull 三情景目标价并排",
    "intro": ("三情景各自完整推演: 物理锚 → 6段收入 → 归母净利 → EPS → 目标P/E → 逐年隐含价。"
              "本表三案恒常并排, 不随『情景切换』变化; case 列只引『情景切换』矩阵行(各案行)+未翻档共用行(光刻胶/特气/硅微粉/化工增速与净利率)+静态历史锚。"
              "历史列=同一条链填实际值, 隐含价历史列对照实际年末价(内置回测)。"),
    "case_names": CASES,
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "block_start": 16,
    "rows": cmp_rows,
    "summary": {
        "band": "三案汇总 (目标年 2027E; 各案触发条件见『情景切换』)",
        "target_col": "H",
        "rows": [
            ("总收入($B)", "rev", K.N2, "= 前驱体(capex×该案强度) + LNG(该案增速) + 光刻胶 + 特气 + 硅微粉LDS + 化工", False),
            ("归母净利($B)", "ni", K.N2, "= Σ(段收入×段净利率) × 净利转换率", False),
            ("EPS(¥)", "eps", K.N2, "= 归母净利 × 汇率 ÷ 股本 476M", False),
            ("目标P/E", "pe", K.MX, "= 历史中枢 28x × HBM结构溢价 1.5x × 该案情绪值", False),
            ("隐含价(¥)", "px", K.PX, "= 目标P/E × 2027E EPS", True),
            ("vs 现价", "up", K.PCT, "对照现价 ¥236.09 的上行/下行空间", True),
        ],
        "mcap": {"label": "隐含市值(亿¥)", "key": "px", "expr": f"*{SH_F}/1000/100*{FX_FC}",
                 "note": "= 隐含价 × 股本 476M"},
        "concl": "风险收益比(2027 目标年 vs 现价 ¥236)一行收口: Bear/Base/Bull 下行上行空间从『估值对比』读入; 概率加权。评级基于价格, 证伪触发器见『综合判断仪表盘』。",
    },
})

# ════════════ 13. 综合判断仪表盘 ════════════
EPS27 = K.R(S_FUND, "H" + str(fr["EPS"]))
PXD = K.R(S_HIST, "G" + str(ha["HPX"]))
dash = K.write_dashboard(wb.create_sheet(S_DASH), {
    "title": "综合判断仪表盘 — A 基本面拐点 · B 估值错位 · C 催化剂 · D 情绪确认",
    "usage": ("预测引擎是 B(错位)+C(催化剂); 情绪 D 只做 timing 确认+过热刹车。"
              "回测: 2025-08 板块启动前(¥60), A块(前驱体毛利跳到 44.8% + 稼动率 59% 弹性)+B块(forward ~28x vs 该给 42x, GAP 正)指向买入。"),
    "blocks": [
        {"title": "A. 基本面拐点 — 业务在结构性变好吗?", "rows": [
            ("产品组合迁移", "前驱体占营收: 2023 24% → 2025 24.5%(收入平) 但毛利率 35%→44.8%(+9.4pct); 综合毛利率结构改善", "量缓价升: 高毛利前驱体占利润比重上行, 2026Q1 '收入-6.9%/盈利+2.5%' 即此切换。真实但温和。"),
            ("HBM franchise", "UP Chemical 绑 SK海力士/三星/美光 HBM 前驱体; 稼动率仅 59.31%(盈亏平衡 80%); 以钼代钨新品类(三星 4→80吨)", "前驱体是绑定 HBM 的高毛利卡位 + 巨大经营杠杆——结构溢价 1.5x 的事实依据。"),
            ("增量催化", "宜兴国产化 2000吨 + 内蒙特气 1.5万吨(2026Q3) + 江苏先科 33.07%→并表(2027前回购)", "多个产能/股权增量在 2026-27 落地, 支撑前驱体量增 + 利润全额并入归母。"),
            ("A 判断", "【中-强】", "前驱体拐点真实但 2025 收入仅+8%(韩国 UP -41%); 量增靠 HBM4/国产化 兑现, 非已发生。", True),
        ]},
        {"title": "B. 估值错位(预测引擎 ★)— 市场现在给的 vs 基本面该给的", "rows": [
            ("市场现在给(forward P/E vs 2027E)", {"fml": f"={PXD}/{EPS27}", "fmt": K.MX, "fill": True},
             "= 现价 ÷ 2027E 模型 EPS(公式算, 随模型走)。"),
            ("基本面该给(justified P/E)", {"inp": 42.0, "fmt": K.MX},
             "= 历史中枢 28x × HBM 结构溢价 1.5x(三层前两层, 情绪中性)。"),
            ("错位 GAP = 该给÷市场给 − 1",
             {"fml": lambda ro: f"=B{ro['基本面该给(justified P/E)']}/B{ro['市场现在给(forward P/E vs 2027E)']}-1", "fmt": K.PCT},
             "GAP 为负 = 市场给的已超基本面该给 = 情绪定价区; 国产替代+龙虎榜叙事已把重估空间提前吃掉。"),
            ("回测: 2025-08 启动前的读数", "市场给 forward ~28x vs 该给 42x, GAP +50%", "当时错位为正 → 该买; 2026H1 涨 200% 后反过来。"),
        ]},
        {"title": "C. 催化剂 — 什么会逼市场闭合 GAP", "rows": [
            ("HBM4 前驱体导入 + 以钼代钨放量", "待; 2026H2 HBM4 量产, 三星 Mo 采购 4→10吨(2026)", "兑现=Bull 扳机(强度突破 0.9%); 导入慢=Bear。"),
            ("江苏先科回购并表", "预期 2026H2-2027 完成回购", "前驱体利润全额并入归母, EPS 增厚。"),
            ("2026 中报/三季报前驱体收入", "待; 验证 2025 +8% 是去库存还是结构放缓", "前驱体重回高增=confirm Base/Bull; 继续平=Bear。"),
            ("C 判断", "利多待兑现, 但已被 price-in", "国产替代+HBM 叙事已在 2026H1 反映到 70x 估值; 剩余催化有增量但市场已预付。", True),
        ]},
        {"title": "D. 情绪确认 — 只做 timing + 刹车", "rows": [
            ("量价温度计", "2026H1 ¥77→236(+200%); 7/1 创历史新高 +5.16%; 上龙虎榜、大基金持股概念领涨", "半年 3 倍 + 高位新高 + 龙虎榜 = 典型情绪顶形态; 仅作温度计, 不进倍数。"),
            ("现价倍数 vs 基本面该给", "forward P/E 70x(2027E) vs 该给 42x", "市场已付出超基本面该给 ~67% = 情绪定价区(当前隐含情绪值 ~2.0)。"),
            ("当前档位", "【过热】", "国产替代主题 + 动量资金驱动; PE(TTM) 111.6x 为历史极值。", True),
            ("衰减扳机", "5 条", "板块国产替代退潮 / 中报前驱体仍平淡 / 韩系切份额 / LNG 交付放缓 / 大盘半导体回调。任一翻 → 下调情绪值重算。"),
        ]},
    ],
    "final": {"band": "★ 综合判断(A+B+C+D 收成一句可执行的话)",
              "text": "A 中-强(前驱体拐点真实但 2025 量增仅+8%, 靠 HBM4/国产化 未来兑现)+B 负(forward 70x vs 该给 42x)+C 利多已 price-in +D 过热(半年3倍+龙虎榜赶顶) → 现价 ¥236 已把 Bull 剧本的大半提前定价。评级基于价格; 前驱体 franchise 在国产替代主线上, Bull 情景比多数 AI 票更 live(稼动率 59% 弹性 + 国产化空间), 但现价无安全边际。回调至 Base 目标区或 HBM4 国产化超预期改写 Bull 概率时重估。"},
    "tracking": {
        "intro": "哪个指标恶化 → 哪个假设先崩 → 触发什么动作。",
        "rows": [
            ("__band__", "一、前驱体主驱动"),
            ("前驱体收入强度", "2025A 0.55% → 2027E 0.66%(Base)", "命门: 前驱体 = capex × 强度", "季报分部收入 + HBM 导入进度", "强度回落<0.53% → 切 Bear, 重算"),
            ("前驱体稼动率", "2025 59.31%", "经营杠杆 = 净利率弹性", "年报/调研纪要", "稼动率停滞<65% → 下调净利率"),
            ("__band__", "二、需求盘子"),
            ("DRAM/HBM capex 指引", "2026E $61.3B(+14%)", "物理锚盘子", "TrendForce + 三星/海力士/美光 capex", "增速<5% → 2027E capex 下调全链重算"),
            ("__band__", "三、周期底盘"),
            ("LNG 保温订单交付", "在手 62亿, 5年>120亿", "非半导体基本盘节奏", "季报 LNG 分部收入", "交付放缓/联洋价格战 → 下调 LNG 增速"),
            ("__band__", "四、估值情绪"),
            ("板块国产替代 beta", "PE(TTM) 111.6x, 板块高位", "高 PE 与板块强相关", "半导体材料板块指数 + 龙虎榜", "板块回调 → 情绪值下调重算"),
        ],
    },
})

# ════════════ 全局格式 + 落盘 ════════════
K.finalize(wb, freeze={
    S_HIST: "B3", S_PX: "B4", S_CONS: "A2", S_HMULT: "B5", S_MULT: "B4", S_SW: "B3",
    S_ANCHOR: "B3", S_SEG: "B3", S_FUND: "B3", S_VAL: "B4", S_CMP: "B6", S_DASH: "B6",
    S_COVER: "A2",
})
out = os.path.join(os.path.dirname(__file__), "..", "out", "002409SZ_valuation_model.xlsx")
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("saved:", os.path.abspath(out))
print("sheets:", wb.sheetnames)
