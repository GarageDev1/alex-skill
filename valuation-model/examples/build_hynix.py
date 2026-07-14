# -*- coding: utf-8 -*-
"""
SK Hynix (000660.KS) 估值模型 v12 — 整体公司估值, Base case。
v12 改动(内存收入拆法重构: capex×收入强度, 锚 2025 实际):
  内存收入拆成 "AI 直驱段(挂 AIDC capex)+ 消费段(周期)"。AI 段 收入 = AIDC capex × 收入强度,
  强度锚 2025 实际(= 该段 2025 收入$B ÷ 2025 AIDC capex$B), 前瞻逐年外推。砍掉旧的供给/需求取min、
  后端产能/单片产出/ASP/份额、独立的「商品供需」sheet(四数连乘失真 + 产能=capex 概念错)。
 (1) 重建「HBM测算」: capex × HBM收入强度。行=AIDC capex / HBM收入强度(=HBM收入÷capex,锚2025) /
     HBM收入($B) / HBM收入(兆KRW)。保留"HBM 收入占 AIDC capex %"演化的精神(强度本身即占比)。
 (2) 新增「DRAM_NAND测算」(替代商品供需), 5 段细拆:
     DRAM 3 段 = HBM(链 HBM测算) + 服务器常规DRAM(capex×服务器DRAM强度) + 消费DRAM(周期 上年×(1+bit)×(1+价));
     NAND 2 段 = eSSD(capex×eSSD强度) + 消费NAND(周期)。
     传统DRAM 合计 = 服务器+消费; NAND = eSSD+消费。汇总喂「预测与估值」分部营收。
 (3) 行高不再手动锁死 wrap 文字行(让 Excel 打开自动适配), 只对色带 band 行保留小高度。
v11 改动: 在历史财务与预测两表的盈利段各加「毛利($B)」+「毛利率(%)」两行。
v10 改动: 历史加市值行; 2026 用市场价反推倍数; (已废)新增商品供需 sheet。
渲染: sheet名无特殊字符 + 跨表引用全引号 + 逻辑列写 build-up 推导;全簿禁单元格批注。
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference

BLUE=Font(color="0000FF"); BLACK=Font(color="000000"); GREEN=Font(color="008000"); GREY=Font(color="808080")  # 去斜体(全表不用斜体)
HF=Font(color="FFFFFF",bold=True); BF=Font(bold=True)
HFILL=PatternFill("solid",fgColor="1F4E79"); BAND=PatternFill("solid",fgColor="305496"); CH=PatternFill("solid",fgColor="D9E1F2")
OUT=PatternFill("solid",fgColor="BDD7EE"); CUR=PatternFill("solid",fgColor="FCE4D6"); GREYF=PatternFill("solid",fgColor="F2F2F2")
PCT='0.0%'; N1='#,##0.0;(#,##0.0);-'; N0='#,##0;(#,##0);-'; MX='0.0"x"'; KRW='#,##0;(#,##0);-'
thin=Side(style="thin",color="BFBFBF"); BORD=Border(bottom=thin)
FX=1410   # 当下/前瞻汇率 KRW/USD
# ★ 逐年均值汇率(KRW/USD, Wind/BBG 利润表口径):历史用当年均值, 前瞻用现汇 1410(无 FX 预测观点)
FXY={"2021A":1144,"2022A":1292,"2023A":1306,"2024A":1364,"2025A":1397,
     "2026E":1410,"2027E":1410,"2028E":1410,"2029E":1410,"2030E":1410}
FXH=[FXY[y] for y in ["2021A","2022A","2023A","2024A","2025A"]]   # B-F
# ★ 月度收盘序列(单一价格来源)= tvremix get_ohlcv(KRX:000660, 1W) resample 月末真实收盘(抓取 2026-06), 非手填。
#   历史财务「年末/年均股价」、股价走势 sheet、预测表「年末/年均价」全部引这一条, 避免多处手填漂移。
PX=[("2020-06",85400),("2020-07",82800),("2020-08",78700),("2020-09",84000),("2020-10",79900),("2020-11",115000),("2020-12",118500),("2021-01",122500),("2021-02",141500),("2021-03",141000),("2021-04",128000),("2021-05",128500),("2021-06",122500),("2021-07",112500),("2021-08",107000),("2021-09",100000),("2021-10",103000),("2021-11",118000),("2021-12",131000),("2022-01",120500),("2022-02",124500),("2022-03",116000),("2022-04",112500),("2022-05",107000),("2022-06",87500),("2022-07",97900),("2022-08",91700),("2022-09",83100),("2022-10",84500),("2022-11",81900),("2022-12",75000),("2023-01",92200),("2023-02",87300),("2023-03",88600),("2023-04",89500),("2023-05",110300),("2023-06",115200),("2023-07",120100),("2023-08",120000),("2023-09",114700),("2023-10",125800),("2023-11",132600),("2023-12",141500),("2024-01",134900),("2024-02",156200),("2024-03",183000),("2024-04",173200),("2024-05",189200),("2024-06",236500),("2024-07",173200),("2024-08",173700),("2024-09",174100),("2024-10",182200),("2024-11",159900),("2024-12",181900),("2025-01",199200),("2025-02",190200),("2025-03",182200),("2025-04",186000),("2025-05",204500),("2025-06",270500),("2025-07",258000),("2025-08",269000),("2025-09",395500),("2025-10",559000),("2025-11",530000),("2025-12",677000),("2026-01",909000),("2026-02",1061000),("2026-03",876000),("2026-04",1286000),("2026-05",2333000),("2026-06",2129000)]
yend={}; ysum={}; ycnt={}
for _ym,_p in PX:
    _y=_ym[:4]; yend[_y]=_p; ysum[_y]=ysum.get(_y,0)+_p; ycnt[_y]=ycnt.get(_y,0)+1
yavg={_y:round(ysum[_y]/ycnt[_y]) for _y in ysum}   # 年均=当年月度收盘均值
yhigh={_y:max(p for m,p in PX if m[:4]==_y) for _y in ysum}   # 年内最高(月末收盘口径)
ylow={_y:min(p for m,p in PX if m[:4]==_y) for _y in ysum}    # 年内最低(月末收盘口径)
def u(v): return None if v is None else round(v*1000/FX,2)        # 单一(仅当下列 G/前瞻用)
def uy(v,fx): return None if v is None else round(v*1000/fx,2)    # 兆KRW → $B (按年汇率)

def R(sheet,cell): return f"'{sheet}'!{cell}"
def hdr(ws,row,text,span=12):
    ws.cell(row,1,text).font=HF
    for c in range(1,span+1): ws.cell(row,c).fill=HFILL
    ws.row_dimensions[row].height=18
def band(ws,row,text,span=12):
    """分节色带(隔断行)"""
    ws.cell(row,1,text).font=Font(color="FFFFFF",bold=True)
    for c in range(1,span+1): ws.cell(row,c).fill=BAND
    ws.row_dimensions[row].height=15
def inp(ws,c,v,src=None,fmt=N1,extra=None):
    # ★ 全面禁用单元格批注(红三角): src 仅作代码内文档, 不生成 cell comment; 解释一律走可见列。
    if extra is not None: fmt=extra
    if fmt is None: fmt=N1
    x=ws[c]; x.value=v; x.font=BLUE; x.number_format=fmt
    return x
def introw(ws,row,cols,vals,src,fmt=N1):
    """整行输入:仅写蓝字输入,解释走可见逻辑/备注列。"""
    for j,(col,v) in enumerate(zip(cols,vals)):
        if v is not None: inp(ws,f"{col}{row}",v,src if j==0 else None,fmt)
def fml(ws,c,f,fmt=N1,link=False):
    x=ws[c]; x.value=f; x.font=GREEN if link else BLACK; x.number_format=fmt; return x
def lab(ws,c,t,b=False,note=False):
    # ★ 文字开头的 "=" 必须转义成全角 "＝", 否则 Excel 当公式解析、整格渲染失败(logic() 同款守卫)
    if isinstance(t,str) and t.startswith("="): t="＝"+t[1:]
    x=ws[c]; x.value=t; x.font=BF if b else (GREY if note else BLACK); return x
import re as _re
def logic(ws,c,t):
    # 逻辑/来源列: 正常黑色(非灰非斜)、去 markdown 粗体标记、句末自动分行(一句一行, 不挤成一段)
    if t and t.startswith("="): t="＝"+t[1:]
    t=t.replace("**","")
    t=_re.sub("。(?=\\S)","。\n",t)
    ws[c].value=t; ws[c].font=BLACK; ws[c].alignment=Alignment(wrap_text=True,vertical="top")

wb=Workbook()
ALLC=["B","C","D","E","F","G","H","I","J","K"]; ALLY=["2021A","2022A","2023A","2024A","2025A","2026E","2027E","2028E","2029E","2030E"]
HC=["B","C","D","E","F"]; HY=["2021","2022","2023","2024","2025"]
FC=["F","G","H","I","J","K"]; FY=["2025A","2026E","2027E","2028E","2029E","2030E"]
S_HIST="历史财务与估值"; S_PX="股价走势"; S_LOGIC="估值逻辑"; S_CAPEX="AIDC Capex预测"; S_HBM="HBM测算"; S_DN="DRAM_NAND测算"; S_ASM="利润与收入假设"; S_FC="预测与估值"; S_VAL="SOTP汇总"

# ═══════════ 0. 封面(报告日期 + 时效性 + Key Takeaways) ═══════════
cov=wb.create_sheet("封面"); wb.remove(wb["Sheet"])
cov["A1"]="SK Hynix (000660.KS) 估值模型"; cov["A1"].font=Font(bold=True,size=18,color="305496"); cov.merge_cells("A1:F1"); cov.row_dimensions[1].height=28
def cov_row(k,v,h=None):
    rr=cov.max_row+1
    cov[f"A{rr}"]=k; cov[f"A{rr}"].font=BF; cov[f"A{rr}"].alignment=Alignment(wrap_text=True,vertical="top")
    cov[f"B{rr}"]=v; cov[f"B{rr}"].alignment=Alignment(wrap_text=True,vertical="top"); cov.merge_cells(f"B{rr}:F{rr}")
    if h: cov.row_dimensions[rr].height=h
for k,v in [
 ("报告日期","2026-06-09"),
 ("数据截止","最新季报 2026 Q1(4/23 报:营收 52.6兆、OPM 72%、净利 40.35兆)=核心硬锚;卖方共识 2026-02~05(GS/JPM/Citi/Nomura/UBS/HSBC 12 篇);股价/财务 tvremix 实时 2026-06"),
 ("现价","2,129,000 KRW(tvremix 最新周线收盘, 2026-06)"),
 ("时效声明","本模型基于 2026-06 时点的卖方共识、capex 指引与真实股价。内存周期/HBM 份额/fab 投产时点变化快——建议每季财报后更新。"),
 ("方法一句话","物理锚(全球 AIDC capex)→ 分部收入(AI 段=capex×强度, 消费段=周期)→ 段驱动营业利润→净利 → P/E·P/B 双镜头, 整体公司估值。"),
]: cov_row(k,v)
rr=cov.max_row+2; cov[f"A{rr}"]="Key Takeaways"; cov[f"A{rr}"].font=HF;
for c in range(1,7): cov.cell(rr,c).fill=HFILL
for kt,desc in [
 ("① 当下估值:历史极值","现价对应 TTM P/E ≈19.8x、P/B ≈12.4x;历史 P/B 常态 1-2x、峰值 ~3.9x(2025末),现价已是峰值的 ~3x。市场把多年后的账面用峰值倍数提前定价。"),
 ("② 核心引擎:HBM 收入强度","HBM 收入 = 全球 AIDC capex × 收入强度。强度=HBM 占 GPU BOM × GPU 占 capex,结构性 ~4%(2025 实际反推 4.1%)。这条链决定营收上限——改 capex,HBM 收入按比例动。"),
 ("③ 内存超级周期:为什么 2026-27 缺、2028 才松","2026-27 缺口峰值(GS DRAM -4.9%@26, 15 年最严):AI 需求已到 + HBM 挤占产能(1 片 HBM=3 片 DDR5)+ 扩产 lead time 2-2.5 年。2028 才 normalize:龙仁 Y1 2H28 才规模出 bit。**时点不确定**——管理层/会长认为缺口或拖到 2029-30。"),
 ("④ 利润:峰值后回中周期","2026-27 营业利润率 ~70% 峰值(锚 Q1'26 实际 OPM 72%;商品价暴涨+经营杠杆);2028 价 normalize 后营业利润缓降(已平滑、非悬崖)。"),
 ("⑤ 逻辑闭环:AI 吃 bit 增量","服务器占传统DRAM% 50→64%↑、eSSD占NAND% 42→57%↑,逐年上升=AI 吃增量自洽。注:HBM 收入占营收% 涨价年反降(19%)是价格扭曲、非矛盾(GS 同)。"),
 ("⑥ 三情景目标价(2027E)","三档从业务杠杆翻起(capex增速/HBM强度/价格路径/段OPM/情绪值, 盈利沿链算出非手拍), P/B 主线:Bear ~1,260k(-41%)/ Base ~2,590k(+22%)/ Bull ~3,400k(+60%),risk-reward 偏上行(前提=信缺货持续)。三案恒常并排见『估值对比』;全链可在『情景切换』下拉变档(默认 Base)。"),
 ("⑦ 主要风险","内存价周期见顶 / HBM 三供(Samsung/Micron 2H26 过 HBM4 认证)分走份额 70%→50-60% / NVDA 单一客户依赖 / normalize 时点不确定。"),
]: cov_row(kt,desc,h=46)
cov.column_dimensions["A"].width=22
for c in "BCDEF": cov.column_dimensions[c].width=22

# ═══════════ 1. 历史财务与估值 ═══════════
s=wb.create_sheet(S_HIST)
hdr(s,1,"SK Hynix 历史财务与估值 ($B) — 2021-2025A + 最新季报1Q26 + 当下TTM",9)
s["A2"]="($B)"; s["A2"].font=BF
for col,y in zip(HC,HY): s[f"{col}2"]=y; s[f"{col}2"].font=BF; s[f"{col}2"].fill=CH
s["G2"]="当下(TTM/最新,TV口径)"; s["G2"].font=BF; s["G2"].fill=CUR
# 行号锚(供预测表引用)
HHBM,HDRAM,HNAND,HOTH,HREV=4,6,8,10,11
HGM,HGMP=18,19           # 毛利($B)/毛利率(%) 行锚(盈利段, 营收与净利之间)
HNI,HEQ,HSH,HFX,HPX=20,22,24,25,27
HPXA=28                  # 年均股价(KRW) 行锚(年末价 HPX 之下, 同一条实拉序列)
HEPS,HBPS,HPE,HPB,HMC=29,30,31,32,33   # EPS/BPS/P/E/P/B/市值($B) 行锚(供预测表反推 2026 倍数)
band(s,3,"分部营收(估计)",8)
src_seg="分部=GS 2026-03-11/Citi 2026-02-24 逐年 segment 原表(2023-25 精确): DRAM(含HBM) 20.85/45.15/74.93兆, NAND 9.72/18.92/20.41兆; 2021-22 卖方不覆盖按 DRAM~70%/NAND~25% 粗估。HBM 2024起+4.5x爆发"
# HBM/传统DRAM(ex-HBM)/NAND/其他, 兆KRW; 2023-25 对齐 GS 原表(总营收逐年 tie 实际 43.0/44.6/32.77/66.19/97.15)
seg=[("HBM 收入",HHBM,[0.5,1.0,2.0,11.0,28.0]),("传统DRAM 收入",HDRAM,[29.0,30.0,18.85,34.15,46.93]),
     ("NAND 收入",HNAND,[11.0,11.0,9.72,18.92,20.41]),("其他 收入",HOTH,[2.5,2.6,2.2,2.12,1.81])]
for name,row,vals in seg:
    lab(s,f"A{row}",name); introw(s,row,HC,[uy(v,FXH[i]) for i,v in enumerate(vals)],src_seg if row==HHBM else None)
    if row in (HHBM,HDRAM,HNAND):  # YoY 行(紧跟其下)
        lab(s,f"A{row+1}","  YoY",note=True)
        for i in range(1,5): fml(s,f"{HC[i]}{row+1}",f"={HC[i]}{row}/{HC[i-1]}{row}-1",PCT)
lab(s,f"A{HREV}","总营收",b=True)
for col in HC: fml(s,f"{col}{HREV}",f"={col}{HHBM}+{col}{HDRAM}+{col}{HNAND}+{col}{HOTH}",N1)
s[f"A{HREV}"].border=BORD; inp(s,f"G{HREV}",u(132.08),"营收(TTM 滚动12M, TV live 2026-06)",None,N1)
band(s,12,"分部占比 (% 总营收)",8)
for kk,row,sr in [("HBM 占比",13,HHBM),("传统DRAM 占比",14,HDRAM),("NAND 占比",15,HNAND),("其他 占比",16,HOTH)]:
    lab(s,f"A{row}",kk,note=True)
    for col in HC: fml(s,f"{col}{row}",f"={col}{sr}/{col}{HREV}",PCT)
band(s,17,"盈利与权益",8)
# 毛利率(历史实际, 估计): 2021 45%/2022 36%/2023 -2%/2024 48%/2025 60% (Citi blended GM)
lab(s,f"A{HGMP}","毛利率(%)"); introw(s,HGMP,HC,[0.45,0.36,-0.02,0.48,0.60],"Citi blended 毛利率, 早年粗估",PCT); inp(s,f"G{HGMP}",0.677,"TTM 毛利率(TV live)",None,PCT)
lab(s,f"A{HGM}","毛利($B)")  # 毛利 = 营收 × 毛利率
for col in HC: fml(s,f"{col}{HGM}",f"={col}{HREV}*{col}{HGMP}",N1)
fml(s,f"G{HGM}",f"=G{HREV}*G{HGMP}",N1)
introw(s,HNI,HC,[uy(v,FXH[i]) for i,v in enumerate([9.6,2.4,-9.1,19.8,42.95])],"净利润: Yahoo/公司实际(按当年均值汇率折$B)"); lab(s,f"A{HNI}","净利润"); inp(s,f"G{HNI}",u(75.14),"净利(TTM 滚动12M, TV live 2026-06)",None,N1)
lab(s,"A21","净利率",note=True)
for col in HC: fml(s,f"{col}21",f"={col}{HNI}/{col}{HREV}",PCT)
introw(s,HEQ,HC,[uy(v,FXH[i]) for i,v in enumerate([62.2,63.3,53.5,73.9,120.7])],"股东权益: 公司年报实际(按当年均值汇率)"); lab(s,f"A{HEQ}","股东权益"); inp(s,f"G{HEQ}",u(120.7),"2025年报权益(P/B同历史列口径); TV统计页用1Q26季度权益~155兆→P/B 10.2",None,N1)
lab(s,"A23","ROE",note=True)
for i,col in enumerate(HC):
    fml(s,f"{col}23",(f"={col}{HNI}/{col}{HEQ}" if i==0 else f"={col}{HNI}/AVERAGE({HC[i-1]}{HEQ},{col}{HEQ})"),PCT)
introw(s,HSH,HC,[710,710,710,705,700],"基本流通股(mn,扣库存股); 发行728M恒定→无增发稀释, 2024-25小幅回购致流通股微降; TV现~700M",N0); lab(s,f"A{HSH}","股本(mn股)"); inp(s,f"G{HSH}",700,"现~700M基本股(TV market cap/价反推)",None,N0)
introw(s,HFX,HC,FXH,"逐年均值汇率 KRW/USD(Wind/BBG 利润表口径)",N0); lab(s,f"A{HFX}","FX (KRW/USD,年均)"); inp(s,f"G{HFX}",1410,"现汇",None,N0)
band(s,26,"市场估值 (历史=年末价→实际P/E·P/B; 当下列=TTM锚, 对齐TV)",8)
introw(s,HPX,HC,[yend[y] for y in HY],"年末股价: tvremix 1W 真实收盘(KRW)","#,##0"); lab(s,f"A{HPX}","年末股价(KRW)"); inp(s,f"G{HPX}",2129000,"现价: tvremix 最新周线收盘 2026-06",None,"#,##0"); s[f"G{HPX}"].fill=CUR
# 年均股价(KRW): 同一条月度序列的当年均值, 平滑单点时点噪声(年末价之外的第二口径)
introw(s,HPXA,HC,[yavg[y] for y in HY],"年均股价: 月度收盘均值(同『股价走势』, tvremix 1W resample)","#,##0"); lab(s,f"A{HPXA}","年均股价(KRW)")
lab(s,f"A{HEPS}","EPS ($)",note=True)
for col in HC: fml(s,f"{col}{HEPS}",f"={col}{HNI}*1000/{col}{HSH}","0.00")
fml(s,f"G{HEPS}",f"=G{HNI}*1000/G{HSH}","0.00")   # TTM 净利已是滚动12M, 不×4
lab(s,f"A{HBPS}","BPS ($)",note=True)
for col in HC: fml(s,f"{col}{HBPS}",f"={col}{HEQ}*1000/{col}{HSH}","0.00")
fml(s,f"G{HBPS}",f"=G{HEQ}*1000/G{HSH}","0.00")
lab(s,f"A{HPE}","P/E (实际)",b=True)
for col in HC: fml(s,f"{col}{HPE}",f'=IF({col}{HNI}<=0,"N/M",{col}{HPX}/({col}{HEPS}*{col}{HFX}))',MX)
fml(s,f"G{HPE}",f"=G{HPX}/(G{HEPS}*G{HFX})",MX)
lab(s,f"A{HPB}","P/B (实际)",b=True); s[f"A{HPB}"].fill=OUT
for col in HC: fml(s,f"{col}{HPB}",f"={col}{HPX}/({col}{HBPS}*{col}{HFX})",MX)
fml(s,f"G{HPB}",f"=G{HPX}/(G{HBPS}*G{HFX})",MX); s[f"G{HPB}"].fill=CUR
# 市值($B): 年末价(KRW)×股本(mn)/当年FX/1000 → $B
lab(s,f"A{HMC}","市值($B)",b=True)
for col in HC: fml(s,f"{col}{HMC}",f"={col}{HPX}*{col}{HSH}/{col}{HFX}/1000",N0)
fml(s,f"G{HMC}",f"=G{HPX}*G{HSH}/G{HFX}/1000",N0); s[f"G{HMC}"].fill=CUR
# ═══ 最新季报 1Q26 列(H): 实际硬锚; YoY=同比(vs 1Q25); 单期不算 P/E·P/B(留空) ═══
s["H2"]="最新季报 1Q26"; s["H2"].font=BF; s["H2"].fill=CUR
for row,vq,tb in [(HHBM,10.0,2.33),(HDRAM,28.0,2.50),(HNAND,13.0,1.60),(HOTH,1.6,0.0)]:
    inp(s,f"H{row}",u(vq),None,N1)                                   # 分部(1Q26, 估: 公司季度未拆)
    if row in (HHBM,HDRAM,HNAND): inp(s,f"H{row+1}",tb,None,PCT)     # YoY 同比(vs 1Q25, 估)
fml(s,f"H{HREV}",f"=H{HHBM}+H{HDRAM}+H{HNAND}+H{HOTH}",N1)           # 总营收(=52.58兆 实际)
inp(s,f"H{HNI}",u(40.35),None,N1)                                    # 净利 1Q26 实际 40.35兆
fml(s,"G21",f"=G{HNI}/G{HREV}",PCT); fml(s,"H21",f"=H{HNI}/H{HREV}",PCT)  # 净利率(TTM/季报)
inp(s,f"H{HEQ}",u(160.0),None,N1)                                    # 期末权益(估: 2025末120.7+1Q26留存)
inp(s,f"H{HSH}",700,None,N0); inp(s,f"H{HFX}",1410,None,N0)
s.column_dimensions["H"].width=13
logic(s,f"I{HREV}","1Q26(H列)实际:营收 52.58兆(+198% 同比)、净利 40.35兆、OPM 72%、净利率 77%(史上最高单季)。公司季度未单拆分部,H 列分部及同比为估;单期不算 P/E·P/B(留空)。这一期是 base 上修的硬锚。")
lab(s,"A34","历史带读法",note=True); lab(s,"B34","P/E: 历史用真实年末价反推(2021 9.6x/2024 6.4x/2025末 11x), 与 TV 年度列基本一致; 当下=TTM锚 现价2,129k÷TTM EPS≈19.8x。P/B: 1.5→0.8(2022底)→1.9→1.7→3.9(2025末)→当下≈12.4x; 均远超历史常态1-2x",note=True)
# 备注/来源列(I): 替代单元格批注, 可见
s["I2"]="备注/来源"; s["I2"].font=BF; s["I2"].fill=CH
for rr,note in [(HHBM,"总营收实际(Yahoo/公司);HBM/DRAM/NAND 拆分对齐 GS/Citi 逐年 segment 原表,2021-22 卖方不覆盖按 DRAM~70%/NAND~25% 粗估"),
 (HGMP,"毛利率:Citi blended,早年粗估;当下=TTM(TV live)"),
 (HNI,"净利:Yahoo/公司实际,按当年均值汇率折$B;当下=TTM 滚动12M"),
 (HEQ,"股东权益:公司年报实际;当下=2025年报口径(TV统计页改用1Q26季度权益→P/B 10.2)"),
 (HSH,"基本流通股(扣库存):发行 728M 恒定、无增发稀释,2024-25 小幅回购致流通股微降;TV 现~700M"),
 (HFX,"逐年均值汇率(Wind/BBG 利润表口径);当下/前瞻用现汇 1410"),
 (HPX,"年末股价:tvremix 1W 真实收盘 resample;当下=现价 2,129,000(tvremix 最新周线, 2026-06)")]:
    logic(s,f"I{rr}",note)
s.column_dimensions["A"].width=16
for col in HC: s.column_dimensions[col].width=9
s.column_dimensions["G"].width=14
s.column_dimensions["I"].width=60

# ═══════════ 2. 股价走势 ═══════════
s=wb.create_sheet(S_PX)
hdr(s,1,"股价走势 (月度收盘 KRW) — 几时开始涨 + 阶段",6)
# 月度收盘序列 PX 已在文件顶部统一定义(单一价格来源), 此处直接用。
s["A3"]="月份"; s["B3"]="收盘价"; s["C3"]="阶段"
for c in ["A3","B3","C3"]: s[c].font=BF; s[c].fill=CH
def phase(ym):
    if ym<="2023-03": return "① 周期低迷"
    if ym<="2024-12": return "② 复苏+HBM3E"
    if ym<="2025-12": return "③ HBM爆发"
    return "④ AI超级周期"
for i,(ym,p) in enumerate(PX,start=4):
    s[f"A{i}"]=ym; x=s[f"B{i}"]; x.value=p; x.number_format="#,##0"; lab(s,f"C{i}",phase(ym),note=True)
n=len(PX)
ch=LineChart(); ch.title="SK Hynix 月度股价 (KRW)"; ch.height=9; ch.width=24
ch.add_data(Reference(s,min_col=2,min_row=3,max_row=3+n),titles_from_data=True)
ch.set_categories(Reference(s,min_col=1,min_row=4,max_row=3+n)); ch.y_axis.numFmt='#,##0'
s.add_chart(ch,"E3")
band(s,26,"阶段说明",10) if False else None
s["E26"]="阶段说明"; s["E26"].font=BF
for i,(a,b) in enumerate([("① 2020-2023Q1 周期低迷","2022下行+2023行业巨亏,股价75k-141k区间震荡"),
 ("② 2023Q2-2024 复苏+HBM3E","HBM3E量产+AI浪潮,90k→199k"),
 ("③ 2025 HBM爆发","HBM全面放量,199k年初→677k年末,H2加速(9-12月 395k→677k)"),
 ("④ 2026 AI超级周期","909k年初→5月冲2,333k峰值→现2,129k,年内约+1.3x")],start=27):
    lab(s,f"E{i}",a,b=True); lab(s,f"F{i}",b,note=True)
s.column_dimensions["A"].width=9; s.column_dimensions["B"].width=11; s.column_dimensions["C"].width=14; s.column_dimensions["E"].width=24; s.column_dimensions["F"].width=58

# ═══════════ 卖方研报共识 (12家, 2026-02~05) — 说人话 + 按"模型假设"组织, 支撑后面的测算 ═══════════
s=wb.create_sheet("卖方研报共识")
hdr(s,1,"卖方研报共识 — 12 家 (2026-02~05);这张表是后面测算的'卖方对账单':每个假设街上怎么看、我们取了谁",9)
def txt(s,r,t,h=None,bold=False):
    s.merge_cells(f"A{r}:I{r}"); c=s[f"A{r}"]; c.value=t
    c.font=BF if bold else BLACK; c.alignment=Alignment(wrap_text=True,vertical="top")
    if h: s.row_dimensions[r].height=h
r=2
txt(s,r,"一句话总览:全街都给买入。共识是内存正处在 15 年来最强的超级周期、2026 是史上最缺的一年,海力士是 HBM 龙头(NVDA 里份额近 70%)。目标价区间 1,250k–2,900k KRW(多数 1,250–1,560k,HSBC 最激进 2,900k)。",56); r+=1
band(s,r,"★ 模型每个关键假设 × 卖方怎么看(本表核心 = 支撑后面的测算)",9); r+=1
for col,h in zip("ABCD",["模型假设","街上共识(说人话)","主要分歧","我们模型取了什么 + 为什么"]):
    s[f"{col}{r}"]=h; s[f"{col}{r}"].font=BF; s[f"{col}{r}"].fill=CH; s[f"{col}{r}"].alignment=Alignment(wrap_text=True,vertical="top")
r+=1
asm_rows=[
 ("HBM 收入增速\n(2026)","分歧大。JPM 原表给 +53%($20B→$31B);GS/UBS 更乐观,接近翻倍(到 ~$40B)。","翻倍要靠把 HBM 单价拍得很高(UBS 拍到 $2.7/Gb,押 HBM4 高溢价);JPM 不给这个溢价。","base 取 JPM 的 +53%(它原表可读、不押 HBM4 高溢价);翻倍那档列为 bull 上沿。"),
 ("服务器 DRAM 增长\n(2026)","GS:AI 服务器把 server DRAM 需求拉 +39%。原因是一台 GPU 服务器的 DDR5 用量是普通服务器的 8-12 倍。","幅度各家略有差异,但方向一致(都看 server 高增)。","挂 AIDC capex × 强度;涨价年强度自然冲高,结果与 GS 的 +39% 量级一致。"),
 ("企业级 eSSD\n(2026)","GS:AI 推理要把上下文(KV cache)从内存下沉到 NAND 做缓存,把 eSSD 需求拉 +58%。","—(方向一致)。","挂 capex × 强度,反映 AI 机架存储暴增。"),
 ("DRAM 价格\n(2026)","共识是暴涨,但口径不同:Citi 用'混合'口径 +112%;GS 用'纯普通 DRAM'口径 +176~184%。","混合 vs 纯商品——纯普通 DRAM 涨得更猛,混进 server/HBM 后被稀释。","取两者中段 +130%。"),
 ("NAND 价格\n(2026)","Citi +137%。企业级 SSD 供不应求,外溢拉高整个 NAND。","—。","取 +137%。"),
 ("供需缺口\n(2026)","GS:DRAM 缺 4.9%、NAND 缺 4.2%——15 年来最紧。这是价格暴涨的根。","—。","作为价格暴涨假设的来源(本模型没单独建供给模型,价格直接拍)。"),
 ("缺口何时转松\n(normalize)","最大分歧。GS 偏早:2027 缺口就收窄。JPM/管理层/SK 会长偏晚:可能拖到 2028,甚至 2029-30。","'周期何时见顶'——这是整个估值最大的不确定。","base 取 2028(海力士龙仁 Y1 新厂 2H28 才规模出 bit 的物理时点);情景里可前后调。"),
 ("目标 P/B","GS 2.8x < JPM 3.0x < UBS 3.2x < Nomura 3.5x;HSBC 最激进。","给多高的 AI 溢价——差出近一倍的目标价。","2027 起 4.0→1.5x 逐年回落(跟 ROE 走,峰值后 normalize)。"),
 ("ROE / 利润率","共识:2026 ROE 飙到 ~80%(史上最高),之后逐年回落(GS:46%→34%)。","normalize 的斜率快慢。","跟模型的盈利/账面逐年递推走,不单独拍。"),
]
for a,b,c,d in asm_rows:
    for col,val in zip("ABCD",[a,b,c,d]):
        s[f"{col}{r}"]=val; s[f"{col}{r}"].alignment=Alignment(wrap_text=True,vertical="top"); s[f"{col}{r}"].font=BF if col=="A" else BLACK
    s.row_dimensions[r].height=58; r+=1
band(s,r,"★ 三个最大的分歧(决定估值上下限)",9); r+=1
for t in [
 "① 周期何时见顶:GS 偏早(2027 缺口就收窄)vs JPM/管理层/会长偏晚(拖到 2028-30)。这决定我们 normalize 时点拍在哪年。",
 "② Samsung 会不会抢走 HBM 份额:GS 更担心(把它列为海力士关键下行风险)vs JPM 认为海力士守得住 NVDA 里 HBM4 ≥60% 的份额。",
 "③ 目标价能给多高:HSBC 2,900k(最激进,赌涨价持续更久)vs 其余 1,250–1,560k。差就差在给多高的目标倍数。",
]: txt(s,r,t,40); r+=1
band(s,r,"各家一句话立场(评级 | 目标价 | 核心观点)",9); r+=1
for t in [
 "GS(Buy,TP 1,200–1,350k):内存 15 年最紧;普通 DRAM 2026 涨 +184%,但 2027 温和回调(逼近客户能承受的成本天花板)。",
 "JPM(OW,TP 1,250k):缺口拖到 2028/29;龙仁新厂 2H28 才出 bit;海力士守住 NVDA 里 HBM4 ≥60% 份额。",
 "Citi(Buy,TP 1,550k):server DRAM 2026 涨 +290%;用分部加总(SOTP)估值;64GB 服务器内存条 $450→$1,310。",
 "Nomura(Buy,TP 1,560k):长期 ROE 接近三星的 2 倍,所以敢给高 P/B 3.5x。",
 "UBS(Buy,TP 1,280k;bull 1,500k):预测一直做到 2030;DRAM 上行延到 2027 底;HBM 单价升到 $2.7/Gb。",
 "HSBC(Buy,TP 2,900k,最激进):赌服务器需求超预期、涨价持续更久;HBM 市场 2026 看到 $99B。",
]: txt(s,r,t,30); r+=1
for col,w in zip("ABCDEFGHI",[16,42,30,44,8,8,8,8,8]): s.column_dimensions[col].width=w

# ═══════════ 3a. 历史估值倍数 (数据底座: 自身历史带 + 当下 TTM/forward + 同业对照 + 相对三星比值) ═══════════
# 先看数据再做假设; 同业数据实拉(companiesmarketcap/investing 2026-06 抓取 + tvremix 实时), 缺口径标 n.a. 不硬补。
S_HMULT="历史估值倍数"
s=wb.create_sheet(S_HMULT)
hdr(s,1,"历史估值倍数 — 先看数据再做假设: 自身历史带 + 当下倍数 + 同业对照 (2026-06 实拉)",11)
logic(s,"A2",("这一页是『估值倍数假设』的数据底座: ① 海力士自己历史上值多少(逐年 + 年内高低带); ② 现在市场给多少(TTM / forward); ③ 同行值多少(同业逐年 P/B + 当下), 以及相对三星的溢价走到哪了(结构溢价的对账线)。看完这页, 再去下一页拍三案倍数。"))
s.merge_cells("A2:K3"); s["A2"].alignment=Alignment(wrap_text=True,vertical="top")
for col,w in zip("ABCDEFGHIJK",[24,10,10,10,10,10,11,11,11,9,9]): s.column_dimensions[col].width=w
s.column_dimensions["L"].width=58
hr=4
band(s,hr,"① 海力士自身: 逐年实际倍数 + 年内高低带 + 当下",11); hr+=1
s[f"A{hr}"]="指标"; s[f"A{hr}"].font=BF
for col,y in zip(HC,HY): s[f"{col}{hr}"]=y; s[f"{col}{hr}"].font=BF; s[f"{col}{hr}"].fill=CH
s[f"G{hr}"]="当下"; s[f"G{hr}"].font=BF; s[f"G{hr}"].fill=CUR
s[f"L{hr}"]="说明"; s[f"L{hr}"].font=BF; s[f"L{hr}"].fill=CH
hr+=1
lab(s,f"A{hr}","年末股价(KRW)")
for col in HC: fml(s,f"{col}{hr}",f"={R(S_HIST,col+str(HPX))}","#,##0",link=True)
fml(s,f"G{hr}",f"={R(S_HIST,'G'+str(HPX))}","#,##0",link=True); s[f"G{hr}"].fill=CUR
logic(s,f"L{hr}","tvremix 真实周线收盘; 当下 = 现价 2,129,000(2026-06)。"); hr+=1
lab(s,f"A{hr}","年末 P/E(实际)")
for col in HC: fml(s,f"{col}{hr}",f"={R(S_HIST,col+str(HPE))}",MX,link=True)
fml(s,f"G{hr}",f"={R(S_HIST,'G'+str(HPE))}",MX,link=True); s[f"G{hr}"].fill=CUR
logic(s,f"L{hr}","历史 = 年末价 ÷ 当年每股收益(2023 亏损为 N/M); 当下 = TTM ≈19.8x。"); hr+=1
HMPB=hr; lab(s,f"A{hr}","年末 P/B(实际)",b=True)
for col in HC: fml(s,f"{col}{hr}",f"={R(S_HIST,col+str(HPB))}",MX,link=True)
fml(s,f"G{hr}",f"={R(S_HIST,'G'+str(HPB))}",MX,link=True); s[f"G{hr}"].fill=CUR
logic(s,f"L{hr}","历史常态 1-2x: 2022 底 0.8x → 2025 末 3.9x(历史峰)。当下 TTM ≈12.4x(年报权益口径; 用 1Q26 季度权益 ≈10.2x); 2026/03 曾见 ~6.8x(companiesmarketcap 口径)。与 companiesmarketcap 年末序列 1.45/0.81/1.87/1.62/3.86 交叉一致。"); hr+=1
lab(s,f"A{hr}","年内最高价(月末)")
for col,y in zip(HC,HY): inp(s,f"{col}{hr}",yhigh[y],None,"#,##0")
logic(s,f"L{hr}","由真实月度收盘序列取年内高/低(月末收盘口径, 非盘中极值)。"); hr+=1
lab(s,f"A{hr}","年内最低价(月末)")
for col,y in zip(HC,HY): inp(s,f"{col}{hr}",ylow[y],None,"#,##0")
hr+=1
lab(s,f"A{hr}","P/B 年内高")
for col in HC: fml(s,f"{col}{hr}",f"={col}{hr-2}/({R(S_HIST,col+str(HBPS))}*{R(S_HIST,col+str(HFX))})",MX,link=True)
logic(s,f"L{hr}","= 年内最高价 ÷ (当年每股净资产 × 当年汇率) → 历史估值带的上沿。"); hr+=1
lab(s,f"A{hr}","P/B 年内低")
for col in HC: fml(s,f"{col}{hr}",f"={col}{hr-2}/({R(S_HIST,col+str(HBPS))}*{R(S_HIST,col+str(HFX))})",MX,link=True)
logic(s,f"L{hr}","历史估值带的下沿——目标倍数应落在带内, 突破要写明理由(见下一页三层)。"); hr+=1
lab(s,f"A{hr}","当下 forward(指引)")
s.merge_cells(f"B{hr}:G{hr}"); lab(s,f"B{hr}","forward P/E ≈10x · forward P/B ≈4.9x(现价 ÷ 模型 2026E 每股)",note=True)
logic(s,f"L{hr}","forward 要用模型前瞻每股(在后面的表), 故公式放『情景估值』与『综合判断仪表盘』; 此处只放结论。"); hr+=2
band(s,hr,"② 同业对照: 年末 P/B 逐年 + 当下 TTM + forward P/E (实拉, 2026-06)",11); hr+=1
s[f"A{hr}"]="公司"; s[f"A{hr}"].font=BF
for col,y in zip(HC,HY): s[f"{col}{hr}"]=y; s[f"{col}{hr}"].font=BF; s[f"{col}{hr}"].fill=CH
for col,h in zip("GHI",["当下P/B(TTM)","当下P/E(TTM)","forward P/E"]):
    s[f"{col}{hr}"]=h; s[f"{col}{hr}"].font=BF; s[f"{col}{hr}"].fill=CH; s[f"{col}{hr}"].alignment=Alignment(wrap_text=True,vertical="top")
s[f"L{hr}"]="业务特征 / 来源"; s[f"L{hr}"].font=BF; s[f"L{hr}"].fill=CH
hr+=1
HMHX=hr; s[f"A{hr}"]="SK Hynix"; s[f"A{hr}"].font=BF; s[f"A{hr}"].fill=CUR
for col in HC: fml(s,f"{col}{hr}",f"={col}{HMPB}",MX,link=True)
fml(s,f"G{hr}",f"={R(S_HIST,'G'+str(HPB))}",MX,link=True); fml(s,f"H{hr}",f"={R(S_HIST,'G'+str(HPE))}",MX,link=True)
lab(s,f"I{hr}","≈10x",note=True)
logic(s,f"L{hr}","本模型标的; forward 推导见『情景估值』。HBM 占营收 ~41%, 却长期被按纯内存定价。"); hr+=1
HMSS=hr; s[f"A{hr}"]="Samsung (005930)"; s[f"A{hr}"].font=BF
for col,v in zip(HC,[1.8,1.1,1.5,0.9,None]):
    if v is not None: inp(s,f"{col}{hr}",v,None,MX)
    else: lab(s,f"{col}{hr}","n.a.",note=True)
inp(s,f"G{hr}",5.03,None,MX); inp(s,f"H{hr}",25.8,None,MX); inp(s,f"I{hr}",13.0,None,MX)
logic(s,f"L{hr}","内存+代工+消费电子综合体。逐年 P/B = investing.com(2026-06 抓取; 2025 年末缺公开口径, 标 n.a. 不硬补); 当下 = tvremix 实时。"); hr+=1
s[f"A{hr}"]="Micron (MU)"; s[f"A{hr}"].font=BF
for col,v in zip(HC,[2.27,1.11,2.20,2.11,6.04]): inp(s,f"{col}{hr}",v,None,MX)
inp(s,f"G{hr}",19.7,None,MX); inp(s,f"H{hr}",44.8,None,MX); inp(s,f"I{hr}",11.0,None,MX)
logic(s,f"L{hr}","纯内存同业(最直接可比)。逐年 = companiesmarketcap; 注意不同数据商的当下 P/B 口径差大(6.5-19.7, 账面更新时点不同), 当下取 tvremix。年末 P/B 2.1 → 6.0 → 当下更高: 同业也在重估。"); hr+=1
for nm,ttmpe,fpe,note in [
 ("NVIDIA (NVDA)",32.0,28.0,"AI GPU 龙头(forward PE 参照系上沿)。"),
 ("TSMC (TSM)",35.5,22.0,"AI 代工垄断。"),
 ("Broadcom (AVGO)",66.0,33.0,"AI ASIC; GAAP 摊销致 TTM 虚高。"),
 ("S&P 500",None,22.0,"美股大盘均值。"),
 ("KOSPI",None,10.0,"韩股大盘均值(参照系下沿)。")]:
    s[f"A{hr}"]=nm; s[f"A{hr}"].font=BF
    for col in HC: lab(s,f"{col}{hr}","—",note=True)
    lab(s,f"G{hr}","—",note=True)
    if ttmpe is not None: inp(s,f"H{hr}",ttmpe,None,MX)
    else: lab(s,f"H{hr}","—",note=True)
    inp(s,f"I{hr}",fpe,None,MX)
    logic(s,f"L{hr}",note); hr+=1
HMRT=hr; lab(s,f"A{hr}","海力士 / 三星 P/B 比值",b=True); s[f"A{hr}"].fill=OUT
for col in "BCDE": fml(s,f"{col}{hr}",f"={col}{HMHX}/{col}{HMSS}","0.00",link=True)
lab(s,f"F{hr}","n.a.",note=True)
fml(s,f"G{hr}",f"=G{HMHX}/G{HMSS}","0.00",link=True)
logic(s,f"L{hr}","结构溢价的对账线: 0.8x(2021-22, 海力士比三星便宜)→ 1.25x(2023)→ 1.8x(2024)→ 当下 ~2.5x。GS 的『30% 溢价』是过去一年平均口径——年末口径已远超, 说明下一页结构溢价取 1.30 偏保守, 超出部分由情绪层吸收。"); hr+=2
band(s,hr,"③ 读法 — 给『估值倍数假设』的三个输入",11); hr+=1
logic(s,f"A{hr}",("① 自己: 当下 TTM P/B ~12.4x 已是历史带(0.8-3.9x)上沿的 3 倍, 但 TTM 失真(盈利暴涨、账面滞后), forward 口径 ~4.9x——所以第一层锚取历史峰值 2.7x(MS 口径), 不用被本轮抬高的当下倍数。"
 "② 同行: Micron 年末 P/B 2.1 → 6.0、当下更高, 这轮是行业性 re-rating, 不是海力士独贵; 但海力士 forward P/E ~10x 仍被定在 Micron/KOSPI 档, 几乎没给 HBM 溢价——错位即重估空间(两条独立上行腿: 盈利兑现 + 倍数重估)。"
 "③ 相对三星: P/B 比值从 0.8x 一路扩到 ~2.5x = HBM 领导力溢价持续扩张 → 第二层结构溢价 1.30 有据且偏保守。→ 下一页: 峰值 2.7 × 溢价 1.30 × 情绪值(三案)。"))
s.merge_cells(f"A{hr}:K{hr+3}"); s[f"A{hr}"].alignment=Alignment(wrap_text=True,vertical="top")

# ═══════════ 3b. 估值倍数假设 (前置: 三案倍数在此拍; 自包含, 不引任何下游表) ═══════════
# 表序纪律(04): 本页只引『历史财务与估值』等上游。标的自身 forward PE 要用模型 EPS(下游)→ 证据行放『情景估值』。
PB_PEAK=2.7      # ① 历史商品周期峰值 P/B(MS 口径; ★ 不用本轮已重估的当下倍数, 否则双重计算)
AI_PREM=1.30     # ② 结构溢价(GS: 海力士相对三星过去1年平均估值溢价 30% → ×1.30)
SENT3=[("Bear",[0.90,0.70,0.50,0.40,0.35]),("Base",[1.20,1.14,0.73,0.51,0.43]),("Bull",[1.35,1.40,1.00,0.70,0.55])]
SENT3_logic={
 "Bear":"任一衰减扳机触发(2027 不再提前售罄 / HBM 合约价首次环比跌 / 三星过 NVDA 认证 / capex 指引下修)→ 情绪快速退潮; 历史上内存股的 P/B 峰值撑不过 4 个季度(0.90 → 0.35)。",
 "Base":"产能售罄 + 长约把超涨撑 ~18-24 个月(FY26-27 情绪 1.20/1.14), FY28 起退潮——Cisco 教训: FOMO 溢价的退潮不可逆(→ 2030 0.43)。",
 "Bull":"市场接受『AI franchise 而非内存股』的重估叙事(UBS 4/8 'Re-rating likely ahead' 即此方向), 超涨延长到 FY28(1.35 → 0.55)。"}
S_MULT="估值倍数假设"
s=wb.create_sheet(S_MULT)
hdr(s,1,"估值倍数假设 — P/B 主线方法论 + 三案目标 P/B(= 历史周期峰值 × 结构溢价 × 情绪值)",11)
logic(s,"A2",("这一页只做判断(数据底座在上一页『历史估值倍数』): ① 为什么 P/B 做主线、P/E 只做交叉验证; ② 三层分解出 Bear/Base/Bull 三套目标 P/B。"
 "『情景切换』引用并切换三案, 『情景估值』套用当前案, 『估值对比』三案并排。"))
s.merge_cells("A2:K3"); s["A2"].alignment=Alignment(wrap_text=True,vertical="top")
for col,w in zip("ABCDEFGHIJK",[24,11,10,9,9,9,9,9,9,9,9]): s.column_dimensions[col].width=w
s.column_dimensions["L"].width=55
def mlg(row,t):
    s.merge_cells(f"D{row}:K{row}"); logic(s,f"D{row}",t)   # 说明列(D:K 合并)
r=4
# ───────── ① 为什么 P/B 主线、P/E 支线 ─────────
band(s,r,"① 为什么 P/B 做主线、P/E 做支线 — 镜头选择是业务判断(任何情景下 P/E 都要 double check)",11); r+=1
logic(s,f"A{r}",("估值镜头的选择是业务判断, 不是会计偏好——核心问题只有一个: 这家公司『穿越周期持续存在的东西』是什么, 盈利还是资产? 用那个东西做分母。"
 "为什么 NVIDIA 看 P/E、而且 P/E 很高: 它是 fabless 轻资产, 真正的资产是 CUDA 生态对客户的锁定和由此而来的定价权——盈利可持续、还在增长, 盈利本身就是耐用资产, 市场敢把今年的盈利外推很多年。P/E 高不是『贵』的同义词, 是市场对盈利耐用性 + 增长的定价。"
 "为什么台积电同样极重资产、P/B 却很高: 先进制程垄断 + 客户深度锁定, 让它的资产穿越周期都挣高且稳的回报——资产是印钞机不是包袱。P/B 的高低从来不由资产轻重决定, 由『资产挣钱能力及其稳定性』决定(justified P/B=(ROE-g)/(COE-g) 的业务含义就在这)。"
 "为什么海力士(商品内存)只能 P/B 主线: 内存是无差异化商品, 价格由行业供需定, 公司对自己的盈利没有定价权——盈利是周期的函数(5 年内净利从约 +$13B 摆到约 -$7B, 2023 年直接转亏), 产能和账面才是穿越周期持续存在的结构存量。"
 "给波动的盈利资本化没有意义, 财务表现就是 P/E 系统性反向: 2019 年(反转起点)P/E 高达 32.6x、2021 年周期小顶只有 8.9x、2024 年盈利翻正只有 6.0x; 低 P/E 陷阱: 以 FY27 峰值盈利算约 12x 表观便宜, FY28 盈利回落约 37% 后同一价格跳到约 18x。给『产能 + 其穿越周期的挣钱能力』定价才有意义 → P/B 主线(分母=股东权益, 逐年留存累积、单调增厚; 海力士股息率约 0.2%、利润几乎全留存扩产, 账面增厚特别快)。"
 "★ HBM 正在改变这个答案: 定制 + 12-18 月认证壁垒 + 预付长约 = 海力士第一次对部分产品拥有定价权, 业务正从『商品』向『准代工』迁移——这正是下方三层分解里『结构溢价』那一层的业务来源; 若 HBM franchise 完全证实, 市场会逐步换 P/E 镜头看它(re-rating 的本质 = 市场换镜头)。"
 "★ 但无论哪个情景, P/E 的 double check 都不可省略: 『情景估值』把当前案隐含价反算成隐含 forward P/E; 『估值对比』对 Bear/Base/Bull 三案各反算一行, 对照『历史估值倍数』同业光谱看是否荒谬——P/B 给答案, P/E 验答案。"))
s.merge_cells(f"A{r}:K{r+7}"); s[f"A{r}"].alignment=Alignment(wrap_text=True,vertical="top"); r+=8
# (forward PE 光谱 / 自身倍数纵览 已移入『历史估值倍数』数据页)
# ───────── ③ 三层分解 → 三案目标 P/B ─────────
band(s,r,"② 三层分解 → 三案目标 P/B(蓝字=情绪值; 三案在此拍, 『情景切换』负责切换)",11); r+=1
logic(s,f"A{r}",("不硬拍倍数,GS 三层分解: ① 历史周期峰值 P/B(过去最强商品周期【实际】峰值, MS 口径)× ② 结构溢价(本轮 AI/HBM 结构性重估, GS: 相对三星 30%)× ③ 情绪值(周期/情绪位置, 定性档位, 依据见『综合判断仪表盘』D块)。"
 "一致性检验: 2.7×1.30×1.11 ≈ 3.9x = 2025末实际 P/B ✓。情绪值 1.0=付足峰值×溢价, >1.0=FOMO 超涨, <1.0=下行折价。"))
s.merge_cells(f"A{r}:K{r+2}"); s[f"A{r}"].alignment=Alignment(wrap_text=True,vertical="top"); r+=3
lab(s,f"A{r}","① 历史周期峰值 P/B"); inp(s,f"B{r}",PB_PEAK,None,MX); PKc=f"$B${r}"
mlg(r,"过去最强商品周期【实际】到过的 P/B 峰值(MS 口径 2.7x)。★ 不用本轮已重估的当下倍数当锚, 否则双重计算。"); r+=1
lab(s,f"A{r}","② 结构溢价(AI/HBM)"); inp(s,f"B{r}",AI_PREM,None,'0.00"x"'); PRc=f"$B${r}"
mlg(r,"GS 原文: 海力士相对三星过去1年平均估值溢价 30% → ×1.30。半硬、慢变, 盯相对三星 P/B 比值。"); r+=1
s[f"A{r}"]="案 / 年份"; s[f"A{r}"].font=BF
for col,y in zip(ALLC,ALLY): s[f"{col}{r}"]=y; s[f"{col}{r}"].font=BF; s[f"{col}{r}"].fill=CH
s[f"L{r}"]="说明"; s[f"L{r}"].font=BF; s[f"L{r}"].fill=CH
r+=1
lab(s,f"A{r}","③ 情绪值(三案)",b=True)
for _c in ["A"]+list(FC[1:]): s[f"{_c}{r}"].fill=GREYF
logic(s,f"L{r}","情绪值 = 周期/情绪位置(倍数的第三层)。1.0 = 市场付足『峰值×溢价』; >1.0 = FOMO 超涨; <1.0 = 下行折价。依据 = 『综合判断仪表盘』D 块档位 + 衰减扳机。历史列 2021-2025 = 实际 P/B ÷(峰值×溢价)反推、三案同值——2025 反推出 1.11, 即一致性检验。"); r+=1
MSENT=r
for cs,vals in SENT3:
    lab(s,f"A{r}",f"  {cs}")
    for col in HC: fml(s,f"{col}{r}",f"={R(S_HIST,col+str(HPB))}/({PKc}*{PRc})","0.00",link=True)   # 历史=实际P/B反推(三案同值)
    for col,v in zip(FC[1:],vals): inp(s,f"{col}{r}",v,None,"0.00")
    logic(s,f"L{r}",SENT3_logic[cs]); r+=1
lab(s,f"A{r}","三案目标 P/B = ①×②×③",b=True)
for _c in ["A"]+list(FC[1:]): s[f"{_c}{r}"].fill=GREYF
logic(s,f"L{r}","同一个三层公式套三案情绪。2027E: Bear ~2.46x / Base ~4.00x / Bull ~4.91x(对标卖方: GS 2.9 / JPM 3.0 / UBS 3.2 / Nomura 3.5)。历史列 = 实际 P/B(回看, 三案同值)。"); r+=1
MPB3=r
for i,(cs,_v) in enumerate(SENT3):
    lab(s,f"A{r}",f"  {cs}")
    for col in HC: fml(s,f"{col}{r}",f"={R(S_HIST,col+str(HPB))}",MX,link=True)   # 历史=实际P/B
    for col in FC[1:]: fml(s,f"{col}{r}",f"={PKc}*{PRc}*{col}{MSENT+i}",MX,link=True)
    r+=1
# ───────── 卖方对账 + 数据源 ─────────
band(s,r,"目标倍数: 卖方 vs 我们 — 凭什么敢给非主流数",11); r+=1
logic(s,f"A{r}",("卖方目标 P/B 区间:GS 2.8x / JPM 3.0x / UBS 3.2x / Nomura 3.5x / HSBC 更高。我们 2027E Base:P/B 4.0x(上方③)。"
 "凭什么比多数卖方 P/B 高:① Nomura 论据——Hynix 长期 ROE 接近三星 2 倍,高 ROE 配高 P/B;② HBM 结构性溢价(定制+认证壁垒+定价权)不该按商品内存倍数;③ 但仍低于 AI 半导体(TSMC 22x/NVDA 28x)——反映『41% HBM + 59% 周期性内存』混合身份。"
 "这是有判断力的非主流数:不贴现价、不抄共识,在事实(41% HBM、ROE、份额)与逻辑(混合身份)支撑范围内拍。"))
s.merge_cells(f"A{r}:K{r+3}"); s[f"A{r}"].alignment=Alignment(wrap_text=True,vertical="top"); r+=4
band(s,r,"数据源与口径",11); r+=1
logic(s,f"A{r}",("历史峰值 2.7x = MS『历史商品周期峰值』口径; 结构溢价 1.30 = GS『海力士相对三星过去一年平均估值溢价 30%』原文(对账线见『历史估值倍数』的比值行); 情绪值依据 = 『综合判断仪表盘』D 块。"
 "倍数数据底座(自身历史带 / 当下 TTM / 同业 / 相对三星比值)全在上一页『历史估值倍数』, 含来源与抓取日。"
 "三案目标 P/B 由本页②三层公式得出; 切换在『情景切换』, 套用在『情景估值』, 三案并排在『估值对比』。"))
s.merge_cells(f"A{r}:K{r+3}"); s[f"A{r}"].alignment=Alignment(wrap_text=True,vertical="top")

# ═══════════ 3c. 情景切换 (全模型唯一情景参数库 + 切换开关; 默认 Base) ═══════════
# 架构(见 04「三情景」): 下拉开关 → 各杠杆『当前案』行(=INDEX) → 全链(AIDC/HBM/DRAM_NAND/利润/倍数)变档。
# 『估值对比』恒常三列并排(引矩阵行, 不引当前案行, 防污染)。
from openpyxl.worksheet.datavalidation import DataValidation
S_SW="情景切换"
s=wb.create_sheet(S_SW)
hdr(s,1,"情景切换 — 全模型唯一的情景参数库 + 切换开关 (默认 Base)",11)
lab(s,"A2","当前情景(下拉切换)",b=True)
_sw=s["B2"]; _sw.value="Base"; _sw.font=BLUE; _sw.fill=CUR
_dv=DataValidation(type="list",formula1='"Bear,Base,Bull"',allow_blank=False)
s.add_data_validation(_dv); _dv.add("B2")
lab(s,"C2","案序号(由 B2 派生)",note=True); fml(s,"D2",'=IF(B2="Bear",1,IF(B2="Base",2,3))',N0); SWIDX="$D$2"   # 嵌套IF, 不用数组常量MATCH(WPS/locale兼容)
logic(s,"A3",("怎么用: B2 是唯一入口——下拉选 Bear/Base/Bull → 案序号(D2, 派生)→ 各杠杆『当前案』行(=INDEX)跟着切 → AIDC Capex/HBM测算/DRAM_NAND/利润假设/估值倍数 整条明细链变档, 『情景估值』输出该案逐年隐含价。默认 Base = 主线。"
 "三案对比不用切: 『估值对比』恒常三列并排(引本页矩阵行)。情景参数只在本页改(蓝字); 未列入的假设三案共用(跟 Base)。"))
s.merge_cells("A3:K5"); s["A3"].alignment=Alignment(wrap_text=True,vertical="top")
band(s,6,"三案触发条件(什么发生 = 落进这一档)",11)
swr=7
for _cs,_trig in [("Bear","三星 HBM4 执行改善并过 NVDA 认证、切走海力士份额(UBS 2/9 已见三星执行改善迹象)+ 商品内存价 2026 底/2027 初见顶(Bernstein 口径: 高价破坏需求 + 三星 P5/M15X 新产能放量)+ 任一衰减扳机触发 → 情绪退潮。"),
 ("Base","JPM 口径: 供需缺口 2028 才松——物理锚是龙仁 Y1 要 2H28 才规模出 bit, 2027 新供给只是涓流(UBS 也认为 DRAM 至少缺到 1Q27); 2026 产能售罄 + 12-18 月预付长约把盈利可见度与情绪撑到 FY27。"),
 ("Bull","HBM4 代际溢价兑现(UBS bull: ASP 升到 $2.7/Gb)+ 缺货拖到 2029-30(SK 会长口径)+ 市场接受『AI franchise』重估叙事(UBS 4/8 'Re-rating likely ahead' 即此方向)。")]:
    lab(s,f"A{swr}",_cs,b=True); s.merge_cells(f"B{swr}:K{swr}"); logic(s,f"B{swr}",_trig); swr+=1
band(s,swr,"参数矩阵 — 可翻档杠杆 × 三案(蓝字在此改; 每杠杆第 4 行 = 当前案, 明细链引它)",11); swr+=1
s[f"A{swr}"]="杠杆 / 案"; s[f"A{swr}"].font=BF
for col,y in zip(ALLC,ALLY): s[f"{col}{swr}"]=y; s[f"{col}{swr}"].font=BF; s[f"{col}{swr}"].fill=CH
s[f"L{swr}"]="说明"; s[f"L{swr}"].font=BF; s[f"L{swr}"].fill=CH
swr+=1
SWB={}; SWACT={}
def sw_lever(key,name,fmt,bear,base,bull,desc,notes,cols=None,hist=None):
    """一个杠杆 5 行: 组标题(杠杆名+指标本体描述) + Bear/Base/Bull(各行带各自案的故事) + 当前案(=INDEX, 链引此行)
    hist: 2021-2025 实际值(三案同值填 B-F, 不留空; None 项标 n.m.)"""
    global swr
    cols=cols or FC[1:]
    if hist is not None: desc=desc+" 历史列 2021-2025 = 实际值(历史只有一个, 三案同值; 仅供对照与回测)。"
    lab(s,f"A{swr}",name,b=True)                                   # 组标题行: 杠杆名独立一行
    for _c in ["A"]+list(ALLC): s[f"{_c}{swr}"].fill=GREYF
    logic(s,f"L{swr}",desc)                                        # 指标本体描述(是什么/为什么重要/共用锚)
    swr+=1
    for i,(cs,vals) in enumerate(zip(["Bear","Base","Bull"],[bear,base,bull])):
        lab(s,f"A{swr}",f"  {cs}")
        if hist is not None:
            for col,v in zip(HC,hist):
                if v is None: lab(s,f"{col}{swr}","n.m.",note=True)
                else: inp(s,f"{col}{swr}",v,None,fmt)
        for col,v in zip(cols,vals):
            if v is not None: inp(s,f"{col}{swr}",v,None,fmt)
        if i==0: SWB[key]=swr
        logic(s,f"L{swr}",notes[i])                                # 该案的具体故事(事件→时点→数字→出处), 各放各行
        swr+=1
    lab(s,f"A{swr}","  当前案(链引此行)",note=True)
    _allcols=(([c for c,v in zip(HC,hist) if v is not None]) if hist is not None else [])+list(cols)
    for col in _allcols: fml(s,f"{col}{swr}",f"=INDEX({col}{SWB[key]}:{col}{SWB[key]+2},{SWIDX})",fmt)
    SWACT[key]=swr; swr+=2
sw_lever('capex',"AIDC capex 增速",PCT,[0.10,0.05,0.03,0.02],[0.265,0.181,0.113,0.087],[0.35,0.25,0.15,0.10],
 "全球 AI 数据中心 capex 是整个模型的物理锚: HBM/服务器DRAM/eSSD 三段收入都挂它。2026 年三案共用同一个锚 $830B(四大云厂已给指引、且海力士 2026 产能已售罄, 这一年分歧很小; 改『AIDC Capex预测』G3 三案全动)。三案分歧在 2027 之后云厂还砸不砸钱。",
 ["AI 投资回报被证伪, 云厂集体踩刹车, 增速骤降(2027 +10% → 2030 +2%)。JPM 提醒过: capex 增速放缓本身就是情绪杀手。",
  "按 GS 'Tracking Trillions' 基线路径, capex 从 $830B 走到 2030 年 ~$1.5T(2027 +26% → 2030 +9%)。",
  "推理需求(agent/视频)爆发 + 主权 AI 加入, 增速几年内不减速(2027 +35% → 2030 +10%)。"],cols=FC[2:],hist=[None,1.00,1.333,1.857,1.44])
sw_lever('hbmi',"HBM 收入强度",PCT,[0.035,0.038,0.035,0.032,0.030],[0.037,0.044,0.042,0.040,0.038],[0.039,0.048,0.047,0.045,0.043],
 "HBM 收入 ÷ AIDC capex, 即每 $1 数据中心投资里海力士能拿到的 HBM 收入——拆开 = 『HBM 占 capex 的份额』×『海力士占 HBM 的份额』, 2025 实际反推 ~4.1%。",
 ["三星 HBM4 执行改善(UBS 2/9 已见迹象)+ 美光扩产, 海力士 HBM 份额从 ~57% 滑向 ~45%, 合约价同时被压(2026 3.5% → 2030 3.0%)。",
  "份额缓降但 HBM4 单价更高对冲; UBS 1/6: HBM4 在 NVDA 里海力士出货份额或仍 >70%(3.7% → 3.8%)。",
  "HBM4/4E 高溢价兑现(UBS bull: ASP 升到 $2.7/Gb), NVDA 份额守住 ≥60%(3.9% → 4.3%)。"],hist=[0.029,0.026,0.022,0.040,0.041])
sw_lever('pd',"消费DRAM 价格涨幅",PCT,[1.10,-0.10,-0.25,-0.15,-0.05],[1.30,0.10,-0.15,-0.12,-0.05],[1.40,0.30,0.05,-0.10,-0.10],
 "本轮超级周期里弹性最大的变量。2026 年涨翻倍是共识: GS(4/24 季报后)把海力士全年混合 DRAM 均价上修到 +182%, Nomura +166%; 我们 Base 取 +130% 偏保守(混合口径)。三案分歧全在『何时见顶』。",
 ["Bernstein 口径——2026 底/2027 初见顶: 高价破坏终端需求 + 三星 P5/海力士 M15X 新产能放量, 2027 年价格就转跌(2026 +110%, 2027 -10%)。",
  "缺口 2028 才松——物理锚是龙仁 Y1 要 2H28 才规模出 bit(JPM), 2027 新供给只是涓流; UBS 也认为 DRAM 至少缺到 1Q27(+130% → +10% → 2028 起回落)。",
  "SK 会长口径『缺货到 2030』——HBM 持续挤占晶圆产能, 涨价拖到 2029-30(+140% → +30% → 2028 +5%)。"],hist=[0.10,-0.18,-0.37,0.77,0.28])
sw_lever('pn',"消费NAND 价格涨幅",PCT,[1.10,-0.10,-0.25,-0.15,-0.05],[1.37,0.10,-0.15,-0.12,-0.05],[1.45,0.30,0.05,-0.10,-0.10],
 "机制同消费DRAM, 弹性略小(GS 缺口口径: NAND -4.2% vs DRAM -4.9%)。这轮 NAND 复苏的引擎是企业级 eSSD(AI 推理把 KV-cache 下沉到 NAND + HDD 缺货转单), 消费端跟涨; GS 5/16 Kioxia read-across 也指向 NAND 价格与利润率还有上行。",
 ["见顶时点跟随消费DRAM 的 Bear 故事: 2026 底/2027 初见顶, 2027 价格转跌。",
  "跟随 Base 故事: 缺口 2028 才松, 2026 大涨后 2027 持平略升、2028 起回落。",
  "跟随 Bull 故事: 缺货拖到 2029-30, 涨价持续更久。"],hist=[0.05,-0.25,-0.43,0.90,-0.08])
sw_lever('hopm',"HBM 营业利润率",PCT,[0.65,0.55,0.48,0.45,0.42],[0.72,0.70,0.66,0.62,0.60],[0.75,0.74,0.70,0.66,0.64],
 "HBM 是定制品(12-18 月认证壁垒 + 预付长约), 营业利润率结构性高于商品内存; 锚 = 2026 一季报公司整体营业利润率(OPM)实际 72%。",
 ["三星/美光都过认证后三供成型, NVDA 反过来压价, HBM 从『卖方市场』回到普通零部件定价(2026 65% → 2030 42%)。",
  "定价权随竞争入局缓慢稀释(72% → 60%)。",
  "HBM4 代际溢价 + 定制 base die(与台积电合作)把客户绑得更深, 定价权维持更久(75% → 64%)。"],hist=[None,None,None,0.55,0.62])
sw_lever('copm',"商品 营业利润率",PCT,[0.60,0.42,0.30,0.25,0.25],[0.72,0.66,0.48,0.41,0.40],[0.74,0.70,0.55,0.48,0.45],
 "商品内存(传统DRAM+NAND)是纯周期段——2023 年行业低谷时营业利润率一度 -24%。它的利润率几乎完全由上面『消费价格』那条路径决定。",
 ["价格 2027 转跌, 利润率快速坍回中周期之下(2026 60% → 2030 25%)。",
  "跟随 Base 价格路径, 2026-27 峰值后回中周期(72% → 40%)。",
  "高价多吃 1-2 年, 峰值利润率拖后回落(74% → 45%)。"],hist=[0.28,0.10,-0.24,0.35,0.42])
# 情绪值杠杆: 三案取值与依据在『估值倍数假设』③拍(那里有完整三层方法论), 本页只引用并切换
lab(s,f"A{swr}","情绪值(倍数第三层)",b=True)
for _c in ["A"]+list(FC[1:]): s[f"{_c}{swr}"].fill=GREYF
logic(s,f"L{swr}","三案取值与依据见『估值倍数假设』③(历史峰值 × 结构溢价 × 情绪值 的完整方法论在那页); 本页只做切换——要改情绪值, 去那页改蓝字。")
swr+=1
for _i,_cs in enumerate(["Bear","Base","Bull"]):
    lab(s,f"A{swr}",f"  {_cs}")
    for col in ALLC: fml(s,f"{col}{swr}",f"={R(S_MULT,col+str(MSENT+_i))}","0.00",link=True)
    if _i==0: SWB['sent']=swr
    swr+=1
lab(s,f"A{swr}","  当前案(链引此行)",note=True)
for col in ALLC: fml(s,f"{col}{swr}",f"=INDEX({col}{SWB['sent']}:{col}{SWB['sent']+2},{SWIDX})","0.00")
SWACT['sent']=swr; swr+=1
SWPB=swr; lab(s,f"A{swr}","  目标 P/B(当前案)",b=True)
for col in ALLC: fml(s,f"{col}{swr}",f"='{S_MULT}'!{PKc}*'{S_MULT}'!{PRc}*{col}{SWACT['sent']}",MX,link=True)
logic(s,f"L{swr}","= 历史周期峰值 × 结构溢价 × 当前案情绪值 → 喂『情景估值』的前瞻 P/B。")
swr+=2
s.column_dimensions["A"].width=26
for col in FC[1:]: s.column_dimensions[col].width=9
s.column_dimensions["L"].width=60

# ═══════════ (估值逻辑不再单开页 —— 逻辑分散进各 sheet 的可见逻辑列/备注列;方法+风险放预测表底部)═══════════

# ═══════════ 4. AIDC Capex预测 ═══════════
s=wb.create_sheet(S_CAPEX)
hdr(s,1,"[ANCHOR] 全球 AI 数据中心 CapEx ($B) — HBM 需求物理盘子",11)
s["A2"]="年份"; s["A2"].font=BF
for col,y in zip(ALLC,ALLY): s[f"{col}2"]=y; s[f"{col}2"].font=BF; s[f"{col}2"].fill=CH
lab(s,"A3","AI 数据中心 capex ($B)")
introw(s,3,HC,[15,30,70,200,488],None,N0)         # 历史(实际/粗估)
inp(s,"G3",830,None,N0)                            # 2026E 锚(公司指引, 三案共用; 改这格三案全动)
for _i,_cc in enumerate(FC[2:]):                   # 2027-2030 = 上年 × (1 + 当前案增速) ← 『情景切换』
    fml(s,f"{_cc}3",f"={FC[1:][_i]}3*(1+{R(S_SW,_cc+str(SWACT['capex']))})",N0,link=True)
lab(s,"A4","YoY",note=True)
for i in range(1,10): fml(s,f"{ALLC[i]}4",f"={ALLC[i]}3/{ALLC[i-1]}3-1",PCT)
band(s,6,"口径与来源",11)
logic(s,"A7",("口径=全球 AI 数据中心专项 capex(严格锁 AI 维度,非 hyperscaler 总额)。"
 "2021-2024(15/30/70/200)为 AI 起步期粗估——AI capex 2023 才起量,2021-22 极小;2025A 起为 AI 规模化期,是本模型重点。"
 "来源:2025=488B CreditSights/公司财报;2026=830B 公司指引合计(TrendForce 2026/5, 三案共用锚);2027-2030 = 锚 × 当前案增速(『情景切换』; Base=Goldman 'Tracking Trillions' 基线路径);2021-2024 粗估。"))
s.merge_cells("A7:K9")
s["A7"].alignment=Alignment(wrap_text=True,vertical="top")
logic(s,"A11","作用:HBM/服务器DRAM/eSSD 三个 AI 直驱段都按『收入 = capex × 收入强度』挂在它上面(见 HBM测算 / DRAM_NAND测算)。改 capex → 三段收入同动 → 营收→估值全链动。")
s.merge_cells("A11:K12"); s["A11"].alignment=Alignment(wrap_text=True,vertical="top")
s.column_dimensions["A"].width=15
for col in ALLC: s.column_dimensions[col].width=8
# 行高不锁: 让 Excel 打开自动适配 merge 区 wrap 文字。

# ═══════════ 5. HBM测算 (capex × 收入强度; 强度锚 2025 实际) ═══════════
# 方法: HBM 收入 = AIDC capex × HBM收入强度。强度 = Hynix 每 $1 AIDC capex 捕获的 HBM 收入,
#   锚 2025 实际(= 历史 HBM 收入$B ÷ 2025 capex$B = $20B/$488B ≈ 4.1%; 28兆KRW=$20B@1397), 前瞻从 4.1% 平滑续上、随含量升小幅上行后随竞争 normalize。
# 砍掉旧的 供给/需求取min、后端产能/单片产出/ASP/份额 四数连乘。
s=wb.create_sheet(S_HBM)
hdr(s,1,"HBM 收入测算 — AIDC capex × 收入强度(强度锚 2025 实际)",14)
s["A2"]="变量"; s["A2"].font=BF
for col,y in zip(ALLC,ALLY): s[f"{col}2"]=y; s[f"{col}2"].font=BF; s[f"{col}2"].fill=CH
s["N2"]="逻辑/来源(整句)"; s["N2"].font=BF; s["N2"].fill=CH
m={}; r=3
band(s,r,"AIDC capex 物理锚",14); r+=1
HCAP=r; lab(s,f"A{r}","AIDC capex ($B)")
for col in ALLC: fml(s,f"{col}{r}",f"={R(S_CAPEX,col+'3')}",N0,link=True)
logic(s,f"N{r}","= 引自『AIDC Capex预测』物理锚。HBM 收入挂它 → 改 capex, HBM 收入跟着动(数值灵敏)。"); r+=1
band(s,r,"HBM 收入强度 (= HBM收入$B ÷ capex$B; 即 HBM 收入占 AIDC capex %)",14); r+=1
# 历史强度(2021-2025)= 历史财务 HBM收入$B ÷ 当年 capex(公式, 2025 即锚点); 前瞻(2026-2030)蓝字输入。
HINT=r; lab(s,f"A{r}","HBM 收入强度 (%)")
for col in HC:  # 历史: 由实数反推 = 历史HBM收入 ÷ capex
    fml(s,f"{col}{r}",f"={R(S_HIST,col+str(HHBM))}/{col}{HCAP}",PCT,link=True)
for _cc in FC[1:]: fml(s,f"{_cc}{r}",f"={R(S_SW,_cc+str(SWACT['hbmi']))}",PCT,link=True)   # 前瞻强度 = 『情景切换』当前案
logic(s,f"N{r}",("假设 HBM 收入强度(HBM收入÷AIDC capex)在一段时间内大致不变——因为这个比值 =『HBM 占 capex 的份额』×『Hynix 占 HBM 的份额』, 两块短期都相对稳。按 2025 实际除出来 ≈4.1%, 往前几年也除一遍, 大致 2.9%~4.1%。"
 "它不是完全恒定, 之所以会先降后升再降, 原因是:"
 "① 2026 略降——这一年全球疯狂铺数据中心, 钱大量砸在厂房/供电/网络/GPU 整机上, HBM 只占其中一小块; 加上 Hynix 的 HBM 产能还在 ramp、基本卖光扩不出更多, 所以它能拿到的 HBM 收入一时追不上 capex 的爆发;"
 "② 2027 回升——龙仁新厂投产、HBM4 起量且单价更高, Hynix 供给跟上, 比值修复;"
 "③ 2028 后结构性下行——三供(Samsung/Micron)陆续过 NVDA 认证、把 Hynix 的 HBM 份额切走, 同时 capex 盘子已巨大, 同样的增量除以更大的分母, 比值自然往下。数值与 JPM 研报基本相符。")); r+=1
band(s,r,"HBM 收入",14); r+=1
HUSD=r; lab(s,f"A{r}","HBM 收入 ($B)",b=True); s[f"A{r}"].fill=OUT
for col in HC: fml(s,f"{col}{r}",f"={R(S_HIST,col+str(HHBM))}",N1,link=True)  # 历史: 直接取历史实数
for col in FC[1:]: fml(s,f"{col}{r}",f"={col}{HCAP}*{col}{HINT}",N1,link=True)  # 前瞻: capex × 强度
logic(s,f"N{r}","前瞻 = AIDC capex × 上方强度;历史取历史财务实数。喂『利润与收入假设』分部营收(改 capex,这条跟着动)。"); r+=1
HKRW=r; lab(s,f"A{r}","HBM 收入 (兆KRW)")
for col in ALLC:
    fxr=R(S_HIST,(col if col in HC else 'G')+str(HFX))   # 引用『历史财务与估值』逐年汇率行(前瞻列引现汇 G=1410), 不在本表拍
    fml(s,f"{col}{r}",f"={col}{HUSD}*{fxr}/1000",N1,link=True)
logic(s,f"N{r}","=$B × 『历史财务与估值』逐年均值汇率(B-F当年实际/前瞻列引现汇 G=1410), 汇率不在本表硬编码。仅展示用, 预测分部营收走 $B 行。"); r+=1
m['HBM 收入 ($B)']=HUSD; m['HBM 收入 (兆KRW)']=HKRW
s.column_dimensions["A"].width=24
for col in ALLC: s.column_dimensions[col].width=8
s.column_dimensions["N"].width=64
# (删"读法"段: 解释已全部进上方各行 N 列, 不再重复)

# ═══════════ 5b. DRAM_NAND测算 (替代商品供需; 5 段细拆: AI 直驱挂 capex + 消费走周期) ═══════════
# 方法: 内存收入拆成 "AI 直驱段(挂 AIDC capex)+ 消费段(周期)"。
#   DRAM 3 段: HBM(链 HBM测算, 不重算) + 服务器常规DRAM(=capex×服务器DRAM强度, 锚2025) + 消费DRAM(=上年×(1+bit)×(1+价))。
#   NAND 2 段: eSSD(=capex×eSSD强度, 锚2025) + 消费NAND(周期)。
#   传统DRAM 合计 = 服务器+消费; NAND 合计 = eSSD+消费 → 喂『利润与收入假设』分部营收。
# 强度锚 2025 实际(=该段 2025 收入$B ÷ 2025 capex$B); 消费段 2025 base = 历史段合计 − AI 段。
s=wb.create_sheet(S_DN)
hdr(s,1,"DRAM / NAND 收入测算 — AI 直驱段(挂 capex)+ 消费段(周期); 5 段细拆",14)
s["A2"]="变量"; s["A2"].font=BF
for col,y in zip(ALLC,ALLY): s[f"{col}2"]=y; s[f"{col}2"].font=BF; s[f"{col}2"].fill=CH
s["N2"]="逻辑/来源(整句)"; s["N2"].font=BF; s["N2"].fill=CH
dn={}; r=3
GS="GS 2/8"; CITI="Citi 2/24"; CYC="本模型周期判断(非卖方)"
band(s,r,"AIDC capex 物理锚",14); r+=1
DCAP=r; lab(s,f"A{r}","AIDC capex ($B)")
for col in ALLC: fml(s,f"{col}{r}",f"={R(S_CAPEX,col+'3')}",N0,link=True)
logic(s,f"N{r}","= 引自『AIDC Capex预测』物理锚。服务器DRAM/eSSD 两 AI 直驱段挂它 → 改 capex 两段同动。"); r+=1
# ── DRAM 三段 ──
band(s,r,"一、DRAM — (a) HBM (链 HBM测算, 不重复算)",14); r+=1
lab(s,f"A{r}","HBM 收入 ($B)")
for col in ALLC: fml(s,f"{col}{r}",f"={R(S_HBM,col+str(m['HBM 收入 ($B)']))}",N1,link=True)
dn['HBM']=r; logic(s,f"N{r}","= 引自『HBM测算』(capex×HBM强度)。DRAM 总盘含 HBM, 但 HBM 单列、不与传统DRAM重复。"); r+=1
band(s,r,"一、DRAM — (b) 服务器常规DRAM = capex × 服务器DRAM强度(强度锚 2025 实数; 含云基底早年反推失真)",14); r+=1
# 服务器占传统DRAM %(历史 estimate): 用于把历史传统DRAM实数拆出服务器段→反推历史强度
SSHR=r; lab(s,f"A{r}","服务器占传统DRAM %(估)")
introw(s,r,HC,[0.30,0.30,0.30,0.40,0.50],None,PCT)
logic(s,f"N{r}",("用服务器占传统DRAM 的比例,把传统DRAM 拆成服务器+消费(以便反推服务器强度)。"
 "**30/40/50% 怎么估的**:TrendForce 给的是 server 占总DRAM 的 bit 口径(2023=37.6%、2024≈50%);我要的是占传统DRAM(剔HBM)的收入口径——剔掉 HBM、再考虑 server RDIMM 单价高于 mobile(收入占比≥bit占比)→ 估早年 ~30%、2024 40%、2025 ~50%。"
 "**为什么升**:AI 服务器单台 DDR5 用量是通用服务器的 8-12 倍,且 AI 服务器出货高增,把 DRAM 增量几乎全吃走、挤压手机/PC(GS:server 含HBM 占全球DRAM 需求 53%@26)。前瞻列由模型反算(=服务器÷传统DRAM合计)。")); r+=1
# 服务器常规DRAM($B): 历史=传统DRAM实数×占比(反推); 前瞻=capex×强度(强度行后回填)
SUSD=r; lab(s,f"A{r}","服务器常规DRAM ($B)")
for col in HC: fml(s,f"{col}{r}",f"={R(S_HIST,col+str(HDRAM))}*{col}{SSHR}",N1,link=True)
dn['服务器DRAM']=r
logic(s,f"N{r}","历史=传统DRAM×服务器占比(反推);前瞻=AIDC capex×服务器强度。挂 capex(改 capex 跟着动)。"); r+=1
# 服务器DRAM强度(%): 历史 2024-25 公式反推(2021-23 标 n.m. 失真), 前瞻从 2025 锚平滑续
SINT=r; lab(s,f"A{r}","服务器DRAM强度 (%)")
for col in ["E","F"]: fml(s,f"{col}{r}",f"={col}{SUSD}/{col}{DCAP}",PCT,link=True)  # 2024-25 实数反推
for col in ["B","C","D"]: lab(s,f"{col}{r}","n.m.",note=True)                      # 2021-23: AI capex 前云基底÷极小capex→失真
introw(s,r,FC[1:],[0.063,0.060,0.045,0.038,0.036],None,PCT)  # 2028+ 平滑下滑(配合价格分年normalize, 不悬崖)
dn['服务器DRAM强度']=r
logic(s,f"N{r}",("服务器 DRAM 收入强度(服务器DRAM÷capex)。注意它**不像 HBM 那样稳定**——因为 = 收入÷capex,而收入里含了内存价,所以涨价年(2026-27)强度被价推高到 6%+,2028 价 normalize 后回落。2024≈5.0%/2025≈3.4%(实数反推);2021-23 n.m.(AI capex 前服务器 DRAM 还主要是云基底,比值失真)。"
 "为什么服务器 DRAM 增长这么猛:AI 服务器单台 DDR5 用量是通用服务器的 8-12 倍(一台要喂 8 张 GPU 的数据流),再叠加 AI 服务器台数本身高增和这轮涨价——所以服务器收入增速远超 capex 增速,强度才冲高。真正稳定、能做闭环检查的是下方『服务器占传统DRAM 份额』(逐年上升)。")); r+=1
# 回填服务器DRAM收入前瞻列 = capex × 强度
for col in FC[1:]: fml(s,f"{col}{SUSD}",f"={col}{DCAP}*{col}{SINT}",N1,link=True)
band(s,r,"一、DRAM — (c) 消费DRAM = 手机(LPDDR)+PC/笔电(DDR); 周期波动最大; 前瞻=上年×(1+bit)×(1+价)",14); r+=1
CBIT_D=r; lab(s,f"A{r}","消费DRAM bit增速")
introw(s,r,HC,[0.20,0.08,0.14,0.17,0.24],None,PCT)
introw(s,r,FC[1:],[-0.10,-0.08,0.05,0.05,0.05],None,PCT)
dn['消费DRAM bit']=r; logic(s,f"N{r}","消费 DRAM = 手机+PC 用的内存。**为什么缺货年 bit 是负的**:2026-27 内存严重缺货,厂商把有限的产能优先分给高价值的服务器/HBM,留给手机/PC 的晶圆被挤掉;加上涨价本身打掉一部分消费需求(需求破坏)→ 消费 bit 不增反降。2028 价格 normalize 后恢复温和增长。历史列放行业总 DRAM bit 作周期参照。"); r+=1
CPRC_D=r; lab(s,f"A{r}","消费DRAM 价格变化")
introw(s,r,HC,[0.10,-0.18,-0.37,0.77,0.28],None,PCT)
for _cc in FC[1:]: fml(s,f"{_cc}{r}",f"={R(S_SW,_cc+str(SWACT['pd']))}",PCT,link=True)  # 前瞻价格 = 『情景切换』当前案(Base: normalize 摊到 2028-30)
dn['消费DRAM 价格']=r; logic(s,f"N{r}",(f"**为什么 2026-27 缺到价格暴涨、2028 才松(挂物理事件,不是日历假设)**:"
 f"① 需求:AI(HBM+服务器DDR5+eSSD)2025-26 井喷;② 供给被 HBM 物理挤占——1 片 HBM 晶圆吃掉约 3 片 DDR5 的产能(HBM die 大2-3x + TSV 良率损耗),2026 HBM 已吃掉约 23% 的 DRAM 晶圆;"
 f"③ 2023-24 下行期厂商压 capex 没扩产;④ 扩产物理 lead time 2-2.5 年(洁净室→装机→量产)→ 新供给最早 2H27 才涓流。所以缺口峰值卡在 2026(GS -4.9%, 15 年最严)、2027 收窄(-2.5%)→ 价格 2026 +130%、2027 +10%。"
 f"**2028 才 normalize 的物理锚**:SK Hynix 龙仁 Y1 2H28 才规模出 bit、Samsung P5 2028 下半年量产;2027 那些厂只在装机/试产、bit 没规模(M15X 2026末投产但主要喂 HBM、不补普通DRAM)。**时点是敏感变量**:管理层/会长认为缺口或拖到 2029-30,故 normalize 时点在情景里可调({CYC})。"
 f"历史列同一机制反向,解释了 2022-23 的价崩(前期扩产+需求塌→过剩→价崩-37%)。涨幅与 Citi/GS 基本相符。")); r+=1
CUSD_D=r; lab(s,f"A{r}","消费DRAM ($B)")
for col in HC: fml(s,f"{col}{r}",f"={R(S_HIST,col+str(HDRAM))}-{col}{SUSD}",N1,link=True)  # 历史=传统DRAM实数−服务器(派生)
for i in range(1,6): c=FC[i]; p=FC[i-1]; fml(s,f"{c}{r}",f"={p}{r}*(1+{c}{CBIT_D})*(1+{c}{CPRC_D})",N1)
dn['消费DRAM']=r; logic(s,f"N{r}","历史=传统DRAM实数−服务器(派生);前瞻=上年×(1+bit)×(1+价)。消费段走周期, 不挂 capex。"); r+=1
band(s,r,"一、DRAM — 传统DRAM 合计 (服务器 + 消费; 不含 HBM)",14); r+=1
DSUM=r; lab(s,f"A{r}","传统DRAM 合计 ($B)",b=True); s[f"A{r}"].fill=OUT
for col in HC: fml(s,f"{col}{r}",f"={R(S_HIST,col+str(HDRAM))}",N1,link=True)  # 历史取实数
for col in FC[1:]: fml(s,f"{col}{r}",f"={col}{SUSD}+{col}{CUSD_D}",N1)
dn['传统DRAM 合计']=r; logic(s,f"N{r}","= 服务器常规DRAM + 消费DRAM(历史取实数)。喂『利润与收入假设』传统DRAM收入。不含 HBM(HBM 单列)。"); r+=1
# ── NAND 两段 ──
band(s,r,"二、NAND — (a) eSSD = capex × eSSD强度(强度锚 2025 实数; 含消费基底早年反推失真)",14); r+=1
# eSSD占NAND %(历史 estimate): 用于把历史NAND实数拆出 eSSD 段→反推历史强度
ESHR=r; lab(s,f"A{r}","eSSD占NAND %(估)")
introw(s,r,HC,[0.20,0.20,0.22,0.28,0.42],None,PCT)
logic(s,f"N{r}",("用企业级 eSSD 占 NAND 的比例, 把 NAND 拆成 eSSD + 消费两块(以便反推 eSSD 强度)。这比例 2024-25 从 ~28% 冲到 ~42%, "
 "因为 AI 推理要把上下文(KV-cache)下沉到 NAND 做近存、加上 HDD 短缺订单转 NAND。与 TrendForce 基本相符。前瞻列由模型反算(=eSSD÷NAND合计)。")); r+=1
# eSSD($B): 历史=NAND实数×占比(反推); 前瞻=capex×强度(强度行后回填)
EUSD=r; lab(s,f"A{r}","eSSD ($B)")
for col in HC: fml(s,f"{col}{r}",f"={R(S_HIST,col+str(HNAND))}*{col}{ESHR}",N1,link=True)
dn['eSSD']=r
logic(s,f"N{r}","历史=NAND×eSSD占比(反推);前瞻=AIDC capex×eSSD强度。挂 capex(改 capex 跟着动)。"); r+=1
# eSSD强度(%): 历史 2024-25 公式反推(2021-23 标 n.m. 失真), 前瞻从 2025 锚平滑续
EINT=r; lab(s,f"A{r}","eSSD强度 (%)")
for col in ["E","F"]: fml(s,f"{col}{r}",f"={col}{EUSD}/{col}{DCAP}",PCT,link=True)  # 2024-25 实数反推
for col in ["B","C","D"]: lab(s,f"{col}{r}","n.m.",note=True)                      # 2021-23: AI capex 前消费基底÷极小capex→失真
introw(s,r,FC[1:],[0.025,0.026,0.019,0.016,0.015],None,PCT)  # 2028+ 平滑下滑(配合价格分年normalize)
dn['eSSD强度']=r
logic(s,f"N{r}",("eSSD 收入强度(eSSD÷capex)。同服务器 DRAM,它**不像 HBM 那样稳定**——收入含 NAND 价,涨价年(2026-27)强度被推高、2028 价 normalize 后回落。2024≈1.9%/2025≈1.3%(实数反推);2021-23 n.m.(AI capex 前 eSSD 还主要是消费/企业基底,比值失真)。"
 "为什么 AI 拉爆 eSSD:大模型推理的上下文(KV-cache)越长越占空间,HBM/DRAM 装不下就下沉到 NAND 做近存,再加 checkpoint 存储和 HDD 短缺转单。稳定、可做闭环检查的是下方『eSSD 占 NAND 份额』(逐年上升)。")); r+=1
# 回填 eSSD收入前瞻列 = capex × 强度
for col in FC[1:]: fml(s,f"{col}{EUSD}",f"={col}{DCAP}*{col}{EINT}",N1,link=True)
band(s,r,"二、NAND — (b) 消费NAND = 手机(eMMC/UFS)+PC client SSD; 周期; 前瞻=上年×(1+bit)×(1+价)",14); r+=1
CBIT_N=r; lab(s,f"A{r}","消费NAND bit增速")
introw(s,r,HC,[0.30,0.10,0.18,-0.02,0.12],None,PCT)
introw(s,r,FC[1:],[-0.05,-0.03,0.06,0.06,0.06],None,PCT)
dn['消费NAND bit']=r; logic(s,f"N{r}","消费 NAND = 手机+PC 用的闪存。同消费 DRAM,缺货年 bit 被压(产能优先给企业级 eSSD + 涨价的需求破坏),但比 DRAM 端略缓(NAND 缺口稍小)。2028 价 normalize 后恢复增长。历史列放行业总 NAND bit 作参照。"); r+=1
CPRC_N=r; lab(s,f"A{r}","消费NAND 价格变化")
introw(s,r,HC,[0.05,-0.25,-0.43,0.90,-0.08],None,PCT)
for _cc in FC[1:]: fml(s,f"{_cc}{r}",f"={R(S_SW,_cc+str(SWACT['pn']))}",PCT,link=True)  # 前瞻价格 = 『情景切换』当前案
dn['消费NAND 价格']=r; logic(s,f"N{r}",(f"价格机制同消费 DRAM(供给刚性→缺口→弹性大)。这轮 eSSD 供不应求外溢到消费 NAND, 2026 大涨、之后随产能放量见顶回落({CYC})。"
 f"历史列同一机制反向, 解释了 2022-23 价崩;注意 2025 消费 NAND 价还是负的, 这轮复苏几乎全靠 eSSD。与 Citi/GS 基本相符。")); r+=1
CUSD_N=r; lab(s,f"A{r}","消费NAND ($B)")
for col in HC: fml(s,f"{col}{r}",f"={R(S_HIST,col+str(HNAND))}-{col}{EUSD}",N1,link=True)  # 历史=NAND实数−eSSD(派生)
for i in range(1,6): c=FC[i]; p=FC[i-1]; fml(s,f"{c}{r}",f"={p}{r}*(1+{c}{CBIT_N})*(1+{c}{CPRC_N})",N1)
dn['消费NAND']=r; logic(s,f"N{r}","历史=NAND实数−eSSD(派生);前瞻=上年×(1+bit)×(1+价)。消费段走周期, 不挂 capex。"); r+=1
band(s,r,"二、NAND — NAND 合计 (eSSD + 消费)",14); r+=1
NSUM=r; lab(s,f"A{r}","NAND 合计 ($B)",b=True); s[f"A{r}"].fill=OUT
for col in HC: fml(s,f"{col}{r}",f"={R(S_HIST,col+str(HNAND))}",N1,link=True)
for col in FC[1:]: fml(s,f"{col}{r}",f"={col}{EUSD}+{col}{CUSD_N}",N1)
dn['NAND 合计']=r; logic(s,f"N{r}","= eSSD + 消费NAND(历史取实数)。喂『利润与收入假设』NAND收入。"); r+=1
# 占比前瞻列回填(派生, 让两行 2021-2030 都填满, 不只历史): 服务器占比=服务器÷传统DRAM合计; eSSD占比=eSSD÷NAND合计
for col in FC[1:]:
    fml(s,f"{col}{SSHR}",f"={col}{SUSD}/{col}{DSUM}",PCT,link=True)
    fml(s,f"{col}{ESHR}",f"={col}{EUSD}/{col}{NSUM}",PCT,link=True)
# ★ 逻辑闭环检查
band(s,r,"★ 逻辑闭环检查 (thesis: AI 吃掉内存增量 → 这两个份额必须逐年上升)",14); r+=1
SCK=r; lab(s,f"A{r}","服务器占传统DRAM %",b=True)
for col in ALLC: fml(s,f"{col}{r}",f"={col}{SSHR}",PCT,link=True)
logic(s,f"N{r}","闭环检查①:AI 服务器抢走 DRAM 的 bit 增量、挤压消费,所以服务器份额必须逐年升。模型跑出来 50%→约 60%→64% 确实在升 → 自洽。若反向下降(旧版曾掉到 41%)= consumer 拍太松/服务器太紧,模型内部矛盾,要回去改。"); r+=1
ECK=r; lab(s,f"A{r}","eSSD 占 NAND %",b=True)
for col in ALLC: fml(s,f"{col}{r}",f"={col}{ESHR}",PCT,link=True)
logic(s,f"N{r}","闭环检查②:AI 机架存储抢走 NAND 增量,eSSD 份额必须逐年升。模型 42%→约 52%→57% 在升 → 自洽。"); r+=1
sm=dn  # 兼容下游若有引用
s.column_dimensions["A"].width=22
for col in ALLC: s.column_dimensions[col].width=8
s.column_dimensions["N"].width=64

# ═══════════ 6. 假设 ═══════════
s=wb.create_sheet(S_ASM)
hdr(s,1,"利润与收入假设 (Base) — 其他增速 + 段利润率 + 净利转换 + 留存 + 分部营收→利润→EPS/BPS(估值倍数在『估值倍数假设』)",13)
s["A2"]="假设"; s["A2"].font=BF
for col,y in zip(["B","C","D","E"],["2021A","2022A","2023A","2024A"]): s[f"{col}2"]=y; s[f"{col}2"].font=BF; s[f"{col}2"].fill=CH  # 利润率/留存行用历史列做锚
for col,y in zip(FC,FY): s[f"{col}2"]=y; s[f"{col}2"].font=BF; s[f"{col}2"].fill=CH
s["N2"]="逻辑/来源(整句)"; s["N2"].font=BF; s["N2"].fill=CH
band(s,3,"营收驱动 (HBM/DRAM/NAND 已在专表细拆; 此处仅其他段)",13)
# DRAM/NAND 收入改在『DRAM_NAND测算』(AI 段挂 capex + 消费段周期), 不再走旧 bit×价 假设行。
r=4; am={}
# 其他 增速(CIS/代工, 无 capex 直驱, 直接外推)
lab(s,f"A{r}","其他 增速"); introw(s,r,ALLC,[None,0.04,-0.154,-0.036,-0.146,0.08,0.08,0.08,0.08,0.08],None,PCT)  # 2022-25=其他收入实际YoY; 2026+ 外推+8%
am['其他 增速']=r; logic(s,f"N{r}","2022-25 为其他(CIS/代工)收入的实际 YoY(随消费电子周期波动);2026+ 小体量、简单外推 +8%/yr(无 capex 直驱)。"); r+=1
band(s,r,"利润率假设 (历史实际锚 2021-2025 + 前瞻; 粗颗粒)",13); r+=1
# 历史列(B-F 2021-2025)=实际/卖方估, 做前瞻的锚+回测(见 03 假设行配历史实际列); 前瞻(G-K)=假设。
# 段驱动利润(选项2): 营业利润 = HBM×HBM率 + 商品×商品率 + 其他×其他率; 净利 = 营业利润×净利转换。段利润率成为载荷驱动(不再孤儿)。
rows2=[
 ("HBM 营业利润率",[None,None,None,0.55,0.62,0.72,0.70,0.66,0.62,0.60],PCT,
   "为什么 HBM 利润率远高于商品内存:HBM 不是标准品,而是按客户(NVDA)定制、要过 1-2 年认证才能供货,加上供不应求 → Hynix 握有定价权。2024-25 实际约 55-62%。前瞻 2026-27 升到 72/70%(锚最新 Q1'26 实际整体 OPM 72%),之后随对手认证入局、定价权稀释而缓降。"),
 ("商品 营业利润率",[0.28,0.10,-0.24,0.35,0.42,0.72,0.66,0.48,0.41,0.40],PCT,
   "商品内存(DRAM+NAND 除 HBM)是强周期:2023 一度 -24%(深亏)。前瞻 2026-27 冲到 72/66% 峰值——锚最新 Q1'26 实际:整体 OPM 72%、DRAM ASP 单季就 +mid-60% QoQ,涨价把商品段利润率也推到史高;2028 起价回落而回中周期。"),
 ("其他 营业利润率",[0.10,0.08,0.05,0.10,0.10,0.10,0.10,0.10,0.10,0.10],PCT,
   "CIS/代工小体量,低个位~10% 简单外推。"),
]
for name,vals,fmt,lg in rows2:
    lab(s,f"A{r}",name); introw(s,r,ALLC,vals,None,fmt)
    if name=="HBM 营业利润率":
        for cc in ["B","C","D"]: lab(s,f"{cc}{r}","n.m.",note=True)
    if name in ("HBM 营业利润率","商品 营业利润率"):   # 前瞻列 = 『情景切换』当前案(历史列保留实际)
        _k='hopm' if name.startswith("HBM") else 'copm'
        for cc in FC[1:]: fml(s,f"{cc}{r}",f"={R(S_SW,cc+str(SWACT[_k]))}",PCT,link=True)
    am[name]=r; logic(s,f"N{r}",lg); r+=1
lab(s,f"A{r}","净利转换率(净利/营业利润)"); introw(s,r,ALLC,[None,None,None,0.86,0.90,0.93,0.91,0.89,0.89,0.89],None,PCT)
for cc in ["B","C","D"]: lab(s,f"{cc}{r}","n.m.",note=True)
am['净利转换']=r; logic(s,f"N{r}","营业利润扣掉税和利息等到净利的比例。Hynix 盈利年税负低、且有非经营收益:Q1'26 实际净利率 77% > OPM 72%(转换>1)。前瞻 2026 取 0.93、之后 0.89(2023 亏损年 n.m.)。"); r+=1
lab(s,f"A{r}","留存率"); introw(s,r,ALLC,[0.91,0.64,1.00,0.94,0.95,0.90,0.90,0.90,0.90,0.90],None,PCT)
lab(s,f"D{r}","n.m.",note=True)
am['留存率']=r; logic(s,f"N{r}","留存率=1−派息率。Hynix 历来低派息、扩产优先(盈利年派息率仅 5-9%),2023 亏损年还照派固定股利。前瞻取留存 90%(派息对估值影响很小,绝大部分现金回投扩产)。"); r+=1
# ── 基本面预测(从旧『预测与估值』前半移入: P/E 镜头之上的全部基本面数据)──
# 分部营收 → 段驱动营业利润 → 净利 → 权益 → EPS/BPS/ROE。下游『估值倍数假设』『估值』直接引本表 EPS/BPS/净利, 不重算。
# ★ 估值倍数(目标 P/E·P/B)在『估值倍数假设』sheet, 隐含价在『估值』sheet; 本表只到 EPS/BPS 为止。
HBM_USD=m['HBM 收入 ($B)']
def SHc(c): return R(S_HIST,(c if c in HC else 'F')+str(HSH))   # 历史用当年股本; 前瞻=2025股本(无增发)
def _yoy(row,dst):
    lab(s,f"A{dst}","  YoY",note=True)
    for i in range(1,10): fml(s,f"{ALLC[i]}{dst}",f"={ALLC[i]}{row}/{ALLC[i-1]}{row}-1",PCT)
band(s,r,"分部营收预测(估计)",13); r+=1
RHBM,RHBMY,RHBMP=r,r+1,r+2; r+=3
RDRAM,RDRAMY,RDRAMP=r,r+1,r+2; r+=3
RNAND,RNANDY,RNANDP=r,r+1,r+2; r+=3
ROTH,ROTHY,ROTHP=r,r+1,r+2; r+=3
RTOT,RTOTY=r,r+1; r+=2
def _pct(rev,dst):
    lab(s,f"A{dst}","  占总营收%",note=True)
    for c in ALLC: fml(s,f"{c}{dst}",f"={c}{rev}/{c}{RTOT}",PCT)
lab(s,f"A{RHBM}","HBM 收入")
for c in HC: fml(s,f"{c}{RHBM}",f"={R(S_HIST,c+str(HHBM))}",N1,link=True)
for c in FC[1:]: fml(s,f"{c}{RHBM}",f"={R(S_HBM,c+str(HBM_USD))}",N1,link=True)
_yoy(RHBM,RHBMY); _pct(RHBM,RHBMP)
for name,rev,yv,pv,hrow,dnrow in [("传统DRAM 收入",RDRAM,RDRAMY,RDRAMP,HDRAM,dn['传统DRAM 合计']),("NAND 收入",RNAND,RNANDY,RNANDP,HNAND,dn['NAND 合计'])]:
    lab(s,f"A{rev}",name)
    for c in HC: fml(s,f"{c}{rev}",f"={R(S_HIST,c+str(hrow))}",N1,link=True)
    for c in FC[1:]: fml(s,f"{c}{rev}",f"={R(S_DN,c+str(dnrow))}",N1,link=True)
    _yoy(rev,yv); _pct(rev,pv)
lab(s,f"A{ROTH}","其他 收入")
for c in HC: fml(s,f"{c}{ROTH}",f"={R(S_HIST,c+str(HOTH))}",N1,link=True)
for i in range(1,6): c=FC[i]; p=FC[i-1]; fml(s,f"{c}{ROTH}",f"={p}{ROTH}*(1+{R(S_ASM,c+str(am['其他 增速']))})",N1)
_yoy(ROTH,ROTHY); _pct(ROTH,ROTHP)
lab(s,f"A{RTOT}","总营收",b=True)
for col in ALLC: fml(s,f"{col}{RTOT}",f"={col}{RHBM}+{col}{RDRAM}+{col}{RNAND}+{col}{ROTH}",N1)
s[f"A{RTOT}"].border=BORD; _yoy(RTOT,RTOTY)
logic(s,f"N{RHBMP}","HBM 占营收 2025→2026 从 ~29% 降到 ~19% 不是 bug:2026 商品内存价超级周期(DRAM +130%/NAND +137%)把传统DRAM/NAND 基数灌爆(+159%/+173%), HBM(base)只 +53% → 被价格摊薄(GS/Citi 同)。反映 thesis(AI 吃 bit 增量)的是『DRAM_NAND测算』服务器占比/eSSD占比两行(逐年升)。")
band(s,r,"盈利与账面",13); r+=1
FGM,FGMP,FNI,FNM,FEQ,FEPS,FBPS,FROE=r,r+1,r+2,r+3,r+4,r+5,r+6,r+7; r+=8
HOPM=lambda c:R(S_ASM,c+str(am['HBM 营业利润率'])); COPM=lambda c:R(S_ASM,c+str(am['商品 营业利润率'])); OOPM=lambda c:R(S_ASM,c+str(am['其他 营业利润率']))
lab(s,f"A{FGM}","营业利润($B)")
for c in ALLC: fml(s,f"{c}{FGM}",f"=IFERROR({c}{RHBM}*{HOPM(c)},0)+({c}{RDRAM}+{c}{RNAND})*{COPM(c)}+{c}{ROTH}*{OOPM(c)}",N1,link=True)
lab(s,f"A{FGMP}","营业利润率(%)",note=True)
for col in ALLC: fml(s,f"{col}{FGMP}",f"={col}{FGM}/{col}{RTOT}",PCT)
lab(s,f"A{FNI}","净利润")
for c in HC: fml(s,f"{c}{FNI}",f"={R(S_HIST,c+str(HNI))}",N1,link=True)
for i in range(1,6): c=FC[i]; fml(s,f"{c}{FNI}",f"={c}{FGM}*{R(S_ASM,c+str(am['净利转换']))}",N1)
lab(s,f"A{FNM}","净利率",note=True)
for col in ALLC: fml(s,f"{col}{FNM}",f"={col}{FNI}/{col}{RTOT}",PCT)
lab(s,f"A{FEQ}","期末权益")
for c in HC: fml(s,f"{c}{FEQ}",f"={R(S_HIST,c+str(HEQ))}",N1,link=True)
for i in range(1,6): c=FC[i]; p=FC[i-1]; fml(s,f"{c}{FEQ}",f"={p}{FEQ}+{c}{FNI}*{R(S_ASM,c+str(am['留存率']))}",N1)
lab(s,f"A{FEPS}","EPS ($)")
for col in ALLC: fml(s,f"{col}{FEPS}",f"={col}{FNI}*1000/{SHc(col)}","0.00")
lab(s,f"A{FBPS}","BPS ($)")
for col in ALLC: fml(s,f"{col}{FBPS}",f"={col}{FEQ}*1000/{SHc(col)}","0.00")
lab(s,f"A{FROE}","ROE",note=True)
for i,col in enumerate(ALLC):
    fml(s,f"{col}{FROE}",(f"={col}{FNI}/{col}{FEQ}" if i==0 else f"={col}{FNI}/AVERAGE({ALLC[i-1]}{FEQ},{col}{FEQ})"),PCT)
band(s,r,"基本面口径说明",13); r+=1
logic(s,f"A{r}",("本表 = 旧『预测与估值』的基本面前半:分部营收(AI 直驱段=capex×强度 + 消费段周期)→ 段驱动营业利润 → 净利(×净利转换)→ 权益(留存递推)→ EPS/BPS/ROE。"
 "历史 2021-2025 列取实际(引『历史财务与估值』);前瞻 2026-2030 由上方假设驱动。下游『估值倍数假设』(forward PE 用本表 EPS)与『估值』(隐含价=目标倍数×本表每股)直接引本表, 不重算。"))
s.merge_cells(f"A{r}:K{r+2}"); s[f"A{r}"].alignment=Alignment(wrap_text=True,vertical="top"); r+=3
s.column_dimensions["A"].width=22
for col in ALLC: s.column_dimensions[col].width=9
s.column_dimensions["N"].width=56
# 行高不锁: wrap 文字行让 Excel 打开自动适配。

# (估值倍数假设已前置到『卖方研报共识』之后; 三案倍数在那页拍, 此处不再建表)

# ═══════════ 10. 估值 (P/E·P/B 双镜头 → 隐含价 + 市值; 引『利润与收入假设』每股 × 『估值倍数假设』目标倍数) ═══════════
# 年末价 yend / 年均价 yavg 已在文件顶部由 PX 统一计算(单一价格来源)。
# ★ 本表只做最后一步: 目标倍数(引『估值倍数假设』) × 前瞻每股(引『利润与收入假设』)→ 隐含价 + 市值。
#   依赖只向前(两个来源表都在本表之前建)。
S_VAL2="情景估值"
s=wb.create_sheet(S_VAL2)
hdr(s,1,"情景估值 — 当前案的逐年隐含价 (P/B 主线; P/E 交叉验证; 2021A-2030E)",11)
lab(s,"L1","当前情景→",note=True); fml(s,"M1",f"={R(S_SW,'B2')}",N0,link=True); s["M1"].fill=CUR
lab(s,"A2","本表输出 = 『情景切换』当前案(默认 Base)。P/B 主线: 隐含价 = 目标P/B × 前瞻BPS(目标P/B 引『估值倍数假设』三层; BPS 引『利润与收入假设』——两者都随开关变档)。P/E 仅交叉验证。历史 2021-2025 用实际年末价反推倍数(事实);前瞻是预测、不拟合现价。三案并排见『估值对比』。",note=True)
s.merge_cells("A2:K2")
s["A3"]="(KRW/股; 市值 兆KRW / $B)"; s["A3"].font=BF
for col,y in zip(ALLC,ALLY): s[f"{col}3"]=y; s[f"{col}3"].font=BF; s[f"{col}3"].fill=CH
def FXc(c): return R(S_HIST,c+str(HFX)) if c in HC else "1410"   # 历史链历史财务逐年, 前瞻=1410
EPSr=lambda c:R(S_ASM,c+str(FEPS)); BPSr=lambda c:R(S_ASM,c+str(FBPS)); NIr=lambda c:R(S_ASM,c+str(FNI))  # 每股/净利引『利润与收入假设』
r=4
band(s,r,"P/B 镜头(主线)(2021-2025 实际年末价反推; 2026-2030 = 目标P/B × 前瞻BPS → 隐含价)",11); r+=1
VPB=r; lab(s,f"A{r}","P/B (历史=实际/前瞻=目标)")
for c in HC: fml(s,f"{c}{r}",f"={R(S_HIST,c+str(HPX))}/({BPSr(c)}*{FXc(c)})",MX,link=True)
for c in FC[1:]: fml(s,f"{c}{r}",f"={R(S_SW,c+str(SWPB))}",MX,link=True)  # 前瞻 = 目标P/B(当前案) ← 『情景切换』(= 倍数页三层 × 当前案情绪)
r+=1
VPBP=r; lab(s,f"A{r}","隐含股价 P/B镜头(KRW)",b=True); s[f"A{r}"].fill=OUT
for i,c in enumerate(FC[1:]): fml(s,f"{c}{r}",f"={c}{VPB}*{BPSr(c)}*{FXc(c)}","#,##0")
r+=1
band(s,r,"P/E 交叉验证(支线)(历史=实际; 前瞻 = P/B 隐含价 ÷ 前瞻EPS = 隐含 forward PE, 对照同业光谱)",11); r+=1
VPE=r; lab(s,f"A{r}","P/E (历史=实际 / 前瞻=P/B隐含)")
for c in HC: fml(s,f"{c}{r}",f'=IF({NIr(c)}<=0,"N/M",{R(S_HIST,c+str(HPX))}/({EPSr(c)}*{FXc(c)}))',MX,link=True)  # 历史实际
for i,c in enumerate(FC[1:]): fml(s,f"{c}{r}",f"={c}{VPBP}/({EPSr(c)}*{FXc(c)})",MX,link=True)  # = P/B隐含价 ÷ 前瞻EPS = 隐含 forward PE
r+=1
VFPE=r; lab(s,f"A{r}","当下 forward P/E = 现价 ÷ 该年每股收益")
for c in FC[1:]: fml(s,f"{c}{r}",f"={R(S_HIST,'G'+str(HPX))}/({EPSr(c)}*{FXc(c)})",MX,link=True)  # 标的 forward PE 证据(倍数页光谱里 Hynix ≈10x 的出处)
r+=1
logic(s,f"A{r}","P/E 交叉验证读法: 『当下 forward P/E』行 = 现价 ÷ 模型各年每股收益——这就是『估值倍数假设』光谱里 Hynix ≈10x 的出处(2026E ~10x、2027E ~9x, 被定在 Micron/KOSPI 档)。『P/E 前瞻』行 = P/B 隐含价 ÷ 前瞻每股收益, 对照同业光谱(NVDA 28 / TSMC 22 / Micron 11)看是否合理; FY28 盈利回落 → 同价 P/E 跳升(周期顶低 P/E 陷阱), 这正是 P/E 只作支线、不当主线的原因。")
s.merge_cells(f"A{r}:K{r+1}"); s[f"A{r}"].alignment=Alignment(wrap_text=True,vertical="top"); r+=2
band(s,r,"代表股价 + 市值(P/B 主线)",11); r+=1
VNOW=r; lab(s,f"A{r}","当下股价(KRW)",b=True)
g=fml(s,f"G{r}",f"={R(S_HIST,'G'+str(HPX))}",MX,link=True); g.number_format="#,##0"; s[f"G{r}"].fill=CUR  # 现价(观测点), 引历史财务
r+=1
VPXE=r; lab(s,f"A{r}","年末股价(KRW, 历史)")
for col,y in zip(HC,HY):
    if y in yend: inp(s,f"{col}{r}",yend[y],None,"#,##0")   # 仅历史(2026 年未结束, 不写年末价)
r+=1
VPXA=r; lab(s,f"A{r}","年均股价(KRW, 历史)")
for col,y in zip(HC,HY):
    if y in yavg: inp(s,f"{col}{r}",yavg[y],None,"#,##0")
r+=1
SHF=R(S_HIST,'$F$'+str(HSH))
VMCK=r; lab(s,f"A{r}","市值 实际年末(兆KRW, 历史)")
for c in HC: fml(s,f"{c}{r}",f"={R(S_HIST,c+str(HPX))}*{R(S_HIST,c+str(HSH))}/1000000",N0)
r+=1
VMCKB=r; lab(s,f"A{r}","市值 前瞻·P/B主线(兆KRW)")
for i,c in enumerate(FC[1:]): fml(s,f"{c}{r}",f"={c}{VPBP}*{SHF}/1000000",N0)
r+=1
VMCU=r; lab(s,f"A{r}","市值 ($B; 历史=实际, 前瞻=P/B主线)",b=True)
for c in HC: fml(s,f"{c}{r}",f"={c}{VMCK}*1000/{FXc(c)}",N0)
for i,c in enumerate(FC[1:]): fml(s,f"{c}{r}",f"={c}{VMCKB}*1000/1410",N0)
r+=1
VUPB=r; lab(s,f"A{r}","前瞻隐含价(P/B) vs 当下股价",b=True)
for i,c in enumerate(FC[1:]): fml(s,f"{c}{r}",f"={c}{VPBP}/$G${VNOW}-1",PCT)
r+=1
band(s,r,"方法与结论",11); r+=1
logic(s,f"A{r}",("方法:把 Hynix 当整体公司,**P/B 主线**逐年估(强周期重资产→P/B,见『估值倍数假设』①)。基本面(分部营收→营业利润→净利→权益→EPS/BPS)在『利润与收入假设』;目标 P/B 在『估值倍数假设』(三层=历史峰值×结构溢价×情绪值);本表只做最后一步:目标P/B × 前瞻BPS → 隐含价 + 市值。P/E 仅交叉验证(隐含价÷前瞻EPS=隐含 forward PE)。"
 "倍数处理:历史 2021-2025 = 当年实际价→倒推 P/B(事实);当下 = TTM 锚(现价÷TTM EPS≈19.8x、÷当前BPS≈12.4x, 见『历史财务与估值』, 只作观测)。前瞻 2026-2030 = 目标P/B × 前瞻BPS → 隐含价(判断)。这是预测、不是拟合现价。"))
s.merge_cells(f"A{r}:K{r+2}"); s[f"A{r}"].alignment=Alignment(wrap_text=True,vertical="top"); r+=3
logic(s,f"A{r}",("结论(方向性):现价 2,129k KRW;锚最新 Q1'26 实际(OPM 72%)上修 base 后,隐含价上行——2027E Base 均值≈2,570k = 比现价 +21%。模型给的是上行,不是拟合现价。"
 "三情景(见『估值对比』, 业务杠杆驱动+P/B 主线):Bear ~1,260k(-41%)/ Base ~2,590k(+22%)/ Bull ~3,400k(+60%),risk-reward 偏上行(前提=信缺货持续)。风险:内存价见顶、HBM 三供份额下滑、NVDA 单一客户、normalize 时点。"))
s.merge_cells(f"A{r}:K{r+2}"); s[f"A{r}"].alignment=Alignment(wrap_text=True,vertical="top"); r+=3
s.column_dimensions["A"].width=20
for col in ALLC: s.column_dimensions[col].width=11.5   # 加宽: KRW 隐含价/年末价(百万级)不再显示为 ####

# ═══════════ 7c. 估值对比 (三案恒常并排, 不随开关变; 顶部汇总 + 每案逐年压缩链 block) ═══════════
# 防污染铁律(04): case 列只引『情景切换』矩阵行(Bear/Base/Bull 各自的行) + 链上未翻档的行(强度/bit/转换率) + 静态历史锚,
#   绝不引明细链上会随开关变的格。Base block 与明细链同公式同输入 → 开关=Base 时完全一致(同构自检)。
S_CMP="估值对比"
s=wb.create_sheet(S_CMP)
hdr(s,1,"估值对比 — Bear / Base / Bull 三个情景的目标价并排对比",11)
logic(s,"A2",("三个情景各自完整推演一遍: AIDC capex → 分部收入 → 净利 → 每股净资产(BPS) → 目标 P/B → 逐年隐含价。"
 "本表三案永远并排可见, 不随『情景切换』的开关变化; 想调某个情景的假设, 去『情景切换』页改对应案的参数即可。"
 "未列入情景矩阵的假设(服务器DRAM/eSSD 收入强度、消费端出货量、净利转换率等)三案共用 Base 取值。"
 "历史列 2021-2025 = 同一条链填实际值: 隐含价历史列应基本等于实际年末价(最后一行历史列 ≈0%)——这是整条估值链的内置回测, 公式放回过去能复现现实, 前瞻的数才可信。"))
s.merge_cells("A2:K4"); s["A2"].alignment=Alignment(wrap_text=True,vertical="top")
SH3=R(S_HIST,'$F$'+str(HSH)); PXN3=R(S_HIST,'G'+str(HPX))
_PK=f"'{S_MULT}'!{PKc}"; _PR=f"'{S_MULT}'!{PRc}"   # 三层第一、二层常数(历史峰值/结构溢价, 三案共用)
CMPA={}   # case → 行锚
_b0=16    # 第一个 block 的 band 行
for ci,cname in enumerate(["Bear","Base","Bull"]):
    r0=_b0+ci*18
    band(s,r0,f"{cname} 案 — 逐年推演 (2021-2025 历史回测 + 2026E-2030E 前瞻)",11)
    hrow=r0+1
    for col,y in zip(ALLC,ALLY): s[f"{col}{hrow}"]=y; s[f"{col}{hrow}"].font=BF; s[f"{col}{hrow}"].fill=CH
    rr=hrow+1; A={}
    A['cap']=rr; lab(s,f"A{rr}","AIDC capex ($B)")
    for _cc in HC: fml(s,f"{_cc}{rr}",f"={R(S_CAPEX,_cc+'3')}",N0,link=True)   # 历史=实际
    fml(s,f"G{rr}",f"={R(S_CAPEX,'G3')}",N0,link=True)   # 2026 锚(三案共用, 蓝字在 AIDC 页)
    for _i,_cc in enumerate(FC[2:]):
        fml(s,f"{_cc}{rr}",f"={FC[1:][_i]}{rr}*(1+{R(S_SW,_cc+str(SWB['capex']+ci))})",N0,link=True)
    rr+=1
    A['hbm']=rr; lab(s,f"A{rr}","HBM 收入 ($B)")
    for _cc in HC: fml(s,f"{_cc}{rr}",f"={R(S_HIST,_cc+str(HHBM))}",N1,link=True)   # 历史=实际
    for _cc in FC[1:]: fml(s,f"{_cc}{rr}",f"={_cc}{A['cap']}*{R(S_SW,_cc+str(SWB['hbmi']+ci))}",N1,link=True)
    rr+=1
    A['srv']=rr; lab(s,f"A{rr}","服务器DRAM+eSSD ($B)")
    for _cc in HC: fml(s,f"{_cc}{rr}",f"={R(S_DN,_cc+str(SUSD))}+{R(S_DN,_cc+str(EUSD))}",N1,link=True)   # 历史=实际拆分
    for _cc in FC[1:]: fml(s,f"{_cc}{rr}",f"={_cc}{A['cap']}*({R(S_DN,_cc+str(SINT))}+{R(S_DN,_cc+str(EINT))})",N1,link=True)
    rr+=1
    A['cd']=rr; lab(s,f"A{rr}","消费DRAM ($B)")
    for _cc in HC: fml(s,f"{_cc}{rr}",f"={R(S_DN,_cc+str(CUSD_D))}",N1,link=True)   # 历史=实际派生
    fml(s,f"G{rr}",f"={R(S_DN,'F'+str(CUSD_D))}*(1+{R(S_DN,'G'+str(CBIT_D))})*(1+{R(S_SW,'G'+str(SWB['pd']+ci))})",N1,link=True)
    for _i,_cc in enumerate(FC[2:]):
        fml(s,f"{_cc}{rr}",f"={FC[1:][_i]}{rr}*(1+{R(S_DN,_cc+str(CBIT_D))})*(1+{R(S_SW,_cc+str(SWB['pd']+ci))})",N1,link=True)
    rr+=1
    A['cn']=rr; lab(s,f"A{rr}","消费NAND ($B)")
    for _cc in HC: fml(s,f"{_cc}{rr}",f"={R(S_DN,_cc+str(CUSD_N))}",N1,link=True)   # 历史=实际派生
    fml(s,f"G{rr}",f"={R(S_DN,'F'+str(CUSD_N))}*(1+{R(S_DN,'G'+str(CBIT_N))})*(1+{R(S_SW,'G'+str(SWB['pn']+ci))})",N1,link=True)
    for _i,_cc in enumerate(FC[2:]):
        fml(s,f"{_cc}{rr}",f"={FC[1:][_i]}{rr}*(1+{R(S_DN,_cc+str(CBIT_N))})*(1+{R(S_SW,_cc+str(SWB['pn']+ci))})",N1,link=True)
    rr+=1
    A['oth']=rr; lab(s,f"A{rr}","其他 ($B)")
    for _cc in HC: fml(s,f"{_cc}{rr}",f"={R(S_HIST,_cc+str(HOTH))}",N1,link=True)   # 历史=实际
    fml(s,f"G{rr}",f"={R(S_HIST,'F'+str(HOTH))}*(1+{R(S_ASM,'G'+str(am['其他 增速']))})",N1,link=True)
    for _i,_cc in enumerate(FC[2:]):
        fml(s,f"{_cc}{rr}",f"={FC[1:][_i]}{rr}*(1+{R(S_ASM,_cc+str(am['其他 增速']))})",N1,link=True)
    rr+=1
    A['rev']=rr; lab(s,f"A{rr}","总收入 ($B)",b=True)
    for _cc in ALLC: fml(s,f"{_cc}{rr}",f"={_cc}{A['hbm']}+{_cc}{A['srv']}+{_cc}{A['cd']}+{_cc}{A['cn']}+{_cc}{A['oth']}",N1)
    rr+=1
    A['ni']=rr; lab(s,f"A{rr}","净利 ($B)",b=True)
    for _cc in HC: fml(s,f"{_cc}{rr}",f"={R(S_HIST,_cc+str(HNI))}",N0,link=True)   # 历史=实际净利
    for _cc in FC[1:]:
        fml(s,f"{_cc}{rr}",(f"=({_cc}{A['hbm']}*{R(S_SW,_cc+str(SWB['hopm']+ci))}"
            f"+({_cc}{A['srv']}+{_cc}{A['cd']}+{_cc}{A['cn']})*{R(S_SW,_cc+str(SWB['copm']+ci))}"
            f"+{_cc}{A['oth']}*{R(S_ASM,_cc+str(am['其他 营业利润率']))})*{R(S_ASM,_cc+str(am['净利转换']))}"),N0)
    rr+=1
    A['eq']=rr; lab(s,f"A{rr}","期末权益 ($B)")
    for _cc in HC: fml(s,f"{_cc}{rr}",f"={R(S_HIST,_cc+str(HEQ))}",N0,link=True)   # 历史=实际权益
    fml(s,f"G{rr}",f"={R(S_HIST,'F'+str(HEQ))}+G{A['ni']}*0.9",N0,link=True)
    for _i,_cc in enumerate(FC[2:]): fml(s,f"{_cc}{rr}",f"={FC[1:][_i]}{rr}+{_cc}{A['ni']}*0.9",N0)
    rr+=1
    A['bps']=rr; lab(s,f"A{rr}","BPS ($)")
    for _cc in HC: fml(s,f"{_cc}{rr}",f"={_cc}{A['eq']}*1000/{R(S_HIST,_cc+str(HSH))}","0.00")   # 历史用当年股本
    for _cc in FC[1:]: fml(s,f"{_cc}{rr}",f"={_cc}{A['eq']}*1000/{SH3}","0.00")
    rr+=1
    A['sent']=rr; lab(s,f"A{rr}","情绪值(该案; 历史=实际反推)")
    for _cc in ALLC: fml(s,f"{_cc}{rr}",f"={R(S_MULT,_cc+str(MSENT+ci))}","0.00",link=True)
    rr+=1
    A['pb']=rr; lab(s,f"A{rr}","目标 P/B(历史=实际)")
    for _cc in ALLC: fml(s,f"{_cc}{rr}",f"={_PK}*{_PR}*{_cc}{A['sent']}",MX,link=True)
    rr+=1
    A['px']=rr; lab(s,f"A{rr}","隐含价 (KRW)",b=True); s[f"A{rr}"].fill=OUT
    for _cc in HC: fml(s,f"{_cc}{rr}",f"={_cc}{A['pb']}*{_cc}{A['bps']}*{R(S_HIST,_cc+str(HFX))}","#,##0")   # 历史用当年汇率 → 应≈实际年末价
    for _cc in FC[1:]: fml(s,f"{_cc}{rr}",f"={_cc}{A['pb']}*{_cc}{A['bps']}*1410","#,##0")
    rr+=1
    A['ipe']=rr; lab(s,f"A{rr}","隐含 forward P/E(交叉验证)")
    for _cc in HC: fml(s,f"{_cc}{rr}",f'=IF({_cc}{A["ni"]}<=0,"N/M",{_cc}{A["px"]}/({_cc}{A["ni"]}*1000/{R(S_HIST,_cc+str(HSH))}*{R(S_HIST,_cc+str(HFX))}))',MX)   # 历史=实际 P/E
    for _cc in FC[1:]: fml(s,f"{_cc}{rr}",f"={_cc}{A['px']}/({_cc}{A['ni']}*1000/{SH3}*1410)",MX)   # P/E 体检: 该案隐含价 ÷ 该案每股收益
    rr+=1
    A['up']=rr; lab(s,f"A{rr}","历史: vs 实际年末价(回测≈0) / 前瞻: vs 现价",note=True)
    for _cc in HC: fml(s,f"{_cc}{rr}",f"={_cc}{A['px']}/{R(S_HIST,_cc+str(HPX))}-1",PCT)   # 回测自检: 应≈0%
    for _cc in FC[1:]: fml(s,f"{_cc}{rr}",f"={_cc}{A['px']}/{PXN3}-1",PCT)
    CMPA[cname]=A
# ── 顶部三案汇总(目标年 2027E = H 列; 引下方各 block)──
band(s,5,"三案汇总 (目标年 2027E; 各案触发条件见『情景切换』)",11)
for col,h in zip("ABCD",["项目(2027E)","Bear","Base","Bull"]):
    s[f"{col}6"]=h; s[f"{col}6"].font=BF; s[f"{col}6"].fill=CH
s.merge_cells("E6:K6"); s["E6"]="说明"; s["E6"].font=BF; s["E6"].fill=CH
_sumrows=[("净利($B)",'ni',N0,"由该案的 capex / 价格 / 利润率假设逐步推导得出"),
 ("总收入($B)",'rev',N0,"= AI 直驱段(capex × 收入强度) + 消费段(按该案价格路径)"),
 ("目标 P/B",'pb',MX,"= 历史周期峰值 2.7 × 结构溢价 1.30 × 该案情绪值"),
 ("隐含价(KRW)",'px',"#,##0","= 目标 P/B × 2027E 每股净资产(BPS) × 汇率"),
 ("隐含 forward P/E",'ipe',MX,"P/E 体检(任何情景都做): 该案隐含价 ÷ 该案每股收益, 对照『历史估值倍数』同业光谱(Micron ~11x / NVDA 28x)看是否荒谬"),
 ("vs 现价",'up',PCT,"对照现价 2,129,000 韩元的上行 / 下行空间"),]
_sr=7
for nm,key,fmt,note in _sumrows:
    lab(s,f"A{_sr}",nm,b=(key in('px','up')))
    for col,cname in zip("BCD",["Bear","Base","Bull"]):
        fml(s,f"{col}{_sr}",f"=H{CMPA[cname][key]}",fmt)
    s.merge_cells(f"E{_sr}:K{_sr}"); lab(s,f"E{_sr}",note,note=True)
    _sr+=1
lab(s,f"A{_sr}","隐含市值($B)")
for col,cname in zip("BCD",["Bear","Base","Bull"]):
    fml(s,f"{col}{_sr}",f"=H{CMPA[cname]['px']}*{SH3}/1410/1000",N0)
s.merge_cells(f"E{_sr}:K{_sr}"); lab(s,f"E{_sr}","= 隐含价 × 股本, 折回美元",note=True)
_sr+=1
logic(s,f"A{_sr}",("风险收益比(2027E 目标价 vs 现价): Bear 约 -41%、Base 约 +22%、Bull 约 +60%——下行一档、上行两档, 整体偏上行。"
 "成立的前提是『缺货持续』: 最新季报(2026 一季度)的营业利润率(OPM)已实际做到 72%, SK 集团会长也公开表示缺货可能持续到 2030 年——这两点支撑 Base 和 Bull。"
 "若『综合判断仪表盘』里的衰减信号触发(如 2027 产能不再提前售罄、HBM 合约价首次环比下跌、三星通过 NVIDIA 认证), 应转向 Bear 案重看估值。"))
s.merge_cells(f"A{_sr}:K{_sr+1}"); s[f"A{_sr}"].alignment=Alignment(wrap_text=True,vertical="top")
s.column_dimensions["A"].width=24
for col in ALLC: s.column_dimensions[col].width=9.5

# ═══════════ 8. 综合判断仪表盘 (决策页: A基本面拐点 + B估值错位(预测引擎) + C催化剂 + D情绪确认) + 投后跟踪明细 ═══════════
# 设计(见 04/05): 预测引擎 = B 错位 + C 催化剂(当下可观测); 情绪 D 只做 timing 确认 + 过热刹车(定性档位, 不精确量化)。
# 验收 = 回测拐点: 放回 2025 年中, 这套表必须能让你当时就看到那波。
s=wb.create_sheet("综合判断仪表盘")
hdr(s,1,"综合判断仪表盘 — 基于当下信息的综合判断: A 基本面拐点 · B 估值错位(预测引擎) · C 催化剂 · D 情绪确认",5)
logic(s,"A2",("怎么用: 预测引擎是 B(错位)+ C(催化剂)——两样都当下可观测、可量化; 情绪 D 只做 timing 确认 + 过热刹车, 定性档位即可(FOMO 的精确幅度本就不可测, 2025 年没人能算出会冲 P/B 6.8x)。"
 "验收=回测: 把这张表放回 2025 年中——A 强 + B 错位 +78% + C 催化剂密集兑现 + D 刚启动 → 当时就该强烈看多。这套表看得到那波, 才算装上了预测引擎。"))
s.merge_cells("A2:E4"); s["A2"].alignment=Alignment(wrap_text=True,vertical="top")
def dband(row,text):
    s.merge_cells(f"A{row}:E{row}"); s[f"A{row}"].value=text; s[f"A{row}"].font=Font(color="FFFFFF",bold=True)
    for c in "ABCDE": s[f"{c}{row}"].fill=BAND
    s.row_dimensions[row].height=15
def drow(row,name,read,judge):
    s[f"A{row}"]=name; s[f"A{row}"].font=BF; s[f"B{row}"]=read
    s.merge_cells(f"C{row}:E{row}"); logic(s,f"C{row}",judge)
    for c in "AB": s[f"{c}{row}"].alignment=Alignment(wrap_text=True,vertical="top")
dc=5
# ── A 基本面拐点 ──
dband(dc,"A. 基本面拐点 — 业务在结构性变好吗?(可建模、可观测)"); dc+=1
for name,read,judge in [
 ("HBM 占营收","14%(2023)→28%(2025)→~41%(2026E)","结构迁移、不是周期反弹——产品组合在根本性重构。"),
 ("可持续 ROE","周期常态 ~15% → HBM franchise 拉向 ~30%+","盈利质量升级(定制+认证壁垒+长约定价权), 不是周期顶的虚高。"),
 ("盈利可见度","2026 产能全售罄 + 12-18 月预付长约","内存史上首次有 4-6 个季度盈利可见度(传统周期只有 1-2 季)。"),
]:
    drow(dc,name,read,judge); dc+=1
drow(dc,"A 判断","【强】","从『周期股』变『AI 内存 franchise』——这正是估值方法需要切换、市场认知会重估的根本原因。"); s[f"B{dc}"].font=BF; dc+=2
# ── B 估值错位(预测引擎) ──
dband(dc,"B. 估值错位(预测引擎 ★)— 市场现在给的 vs 基本面该给的 → GAP"); dc+=1
BPS26=R(S_ASM,'G'+str(FBPS)); PXD=R(S_HIST,'G'+str(HPX))
BMKT=dc; drow(dc,"市场现在给(forward P/B)","","= 现价 ÷ FY26E 前瞻 BPS × FX(公式算, 随模型走)。")
fml(s,f"B{dc}",f"={PXD}/({BPS26}*1410)",MX,link=True); s[f"B{dc}"].fill=CUR; dc+=1
BJUST=dc; drow(dc,"基本面该给(justified P/B)","","=(ROE−g)/(COE−g): 可持续 ROE 30%(结构性乐观上限)/ COE 11.5% / g 3% → 3.2x。注: 估值倍数页目标 4.2x = 该给 3.2x × 情绪溢价(D 块判断 FY26-27 超涨可持续); 本块 GAP 衡量的是纯基本面错位。")
inp(s,f"B{dc}",3.2,None,MX); dc+=1
drow(dc,"错位 GAP = 该给÷市场给 − 1","","GAP 正且大 = 重估空间(预测引擎点火, 该买); GAP 转负 = 基本面空间已被价格吃完、进入纯情绪定价区(只剩 D 块在撑)。")
fml(s,f"B{dc}",f"=B{BJUST}/B{BMKT}-1",PCT,link=True); s[f"B{dc}"].font=BF; dc+=1
drow(dc,"回测: 2025 年中的读数","1.8x vs 3.2x = +78%","当时市场按『又一轮周期顶』给 1.8x, 基本面已该给 3.2x——光错位就 +78%, 还没算盈利增长。这(而不是算准情绪)就是那波的预测依据。"); dc+=2
# ── C 催化剂 ──
dband(dc,"C. 催化剂 — 什么会逼市场闭合 GAP(逐条列状态)"); dc+=1
for name,read,judge in [
 ("HBM4 全球首发量产准备","2025/9 ✓","代际领先实锤, 锁定 NVDA Rubin 主供——盈利可见度延到 2027-28。"),
 ("『2026 sold out』官宣","2025 Q3 业绩会 ✓","叙事颠覆: 从『扩产→跌价』变『扩产都不够卖』——周期折价切换成长溢价的扳机。"),
 ("Microsoft 独家 HBM 供应","2026/1 ✓","客户维度打开: 不只 NVDA, 所有自研 AI 芯片(Maia/TPU/MTIA/Trainium)都是客户。"),
 ("库存股注销 ₩12.2兆 + 派息上调","2026/1 ✓","首次把 AI 红利分给股东(value-up)——韩国折价收窄(COE↓)的证据。"),
 ("下一批: HBM4 放量 / 龙仁投产 / 2027 售罄官宣","待","兑现 → 撑 FY27 情绪; 落空 → 衰减扳机触发。"),
]:
    drow(dc,name,read,judge); dc+=1
drow(dc,"C 判断","已密集兑现","重估已被点燃并走完大半; 剩余催化剂决定超涨还能不能撑到 FY27。"); s[f"B{dc}"].font=BF; dc+=2
# ── D 情绪确认 ──
dband(dc,"D. 情绪确认 — 只做 timing + 刹车(定性档位: 萌芽/启动/过热/退潮; 不精确量化)"); dc+=1
for name,read,judge in [
 ("量价温度计(只当温度计, 不进倍数)","2026/5 冲 2,333k 后回 2,129k; 3 月曾见 P/B 6.8x","量价从现价来, 若拟合进目标倍数 → 隐含价追着现价跑(循环论证, 信号归零)。只用来看背离: 量价 ≫ 基本面 = FOMO 脆弱该收; 量价 ≪ 基本面 = 过度悲观买点。"),
 ("现价倍数 vs 基本面该给","forward ~4.9x ≫ justified 3.2x","市场已付出超过基本面该给的价 = 进入情绪定价区(Baker-Wurgler: 高情绪预示低后续回报——量价情绪天生是反向指标)。"),
 ("散户/资金参与","KOSPI 散户 FOMO + 外资回流(KRX 可逐日盯散户净买入)","散户追涨是超涨的燃料, 也是退潮时最快的抛压。"),
]:
    drow(dc,name,read,judge); dc+=1
drow(dc,"当前档位","【过热】","新钱不追; 持仓持有但盯衰减扳机, 任一触发 → 逐步兑现。(对照 2025 年中: 档位=【启动】突破+放量+散户刚进 → 上车窗口)"); s[f"B{dc}"].font=BF; dc+=1
drow(dc,"衰减扳机(情绪还能撑多久)","四条","① 2027 是否还提前售罄(最强先行) ② HBM ASP 首次环比跌 ③ 三星 HBM4 过 NVDA 认证 ④ hyperscaler capex 指引下修——任一翻: D 档位降档 + 估值倍数页情绪值下调。"); dc+=2
# ── 综合判断 ──
dband(dc,"★ 综合判断(A+B+C+D 收成一句可执行的话)"); dc+=1
logic(s,f"A{dc}",("回测 2025 年中: A 强 + B 错位 +78% + C 点火 + D 刚启动 → 强烈看多 ✓(这套表当时就看得到这波)。"
 "当下(2026-06): A 仍强 + B GAP 已转负(基本面空间被价格吃完) + C 兑现大半 + D 过热 → 基本面强但已被计入: 新钱不追; 持仓持有、盯四条衰减扳机, 任一触发 → 兑现。"))
s.merge_cells(f"A{dc}:E{dc+2}"); s[f"A{dc}"].alignment=Alignment(wrap_text=True,vertical="top"); dc+=4
# ───── 投后跟踪明细 ─────
dband(dc,"投后跟踪明细 — 哪个指标恶化 → 哪个假设先崩"); dc+=1
logic(s,f"A{dc}",("① HBM 收入强度回落(份额被三供蚕食/含量升放缓/合约价跌)→ HBM 收入链先崩 → 营收与峰值 ROE 掉 → 目标 P/B 路径塌、隐含价腰斩。"
 "② 传统DRAM/NAND 合约价见顶 + 库存周数回升 → 消费段周期假设(2026 +112%/+137%)证伪、净利率从峰值急坠。"
 "③ 全球 AI capex 指引下修 → 物理锚盘子缩水 → 三个 AI 直驱段同缩。盯的优先级: HBM 收入强度&份额 ≈ AI capex 指引 > 商品价&库存 > 认证进度。"))
s.merge_cells(f"A{dc}:E{dc+2}"); s[f"A{dc}"].alignment=Alignment(wrap_text=True,vertical="top"); dc+=3
heads=["指标","当前值/状态","为什么重要(哪个假设的命门)","怎么跟踪(数据源/频率)","触发动作"]
hr=dc
for col,h in zip(["A","B","C","D","E"],heads): s[f"{col}{hr}"]=h; s[f"{col}{hr}"].font=BF; s[f"{col}{hr}"].fill=CH; s[f"{col}{hr}"].alignment=Alignment(wrap_text=True,vertical="top")
def trk_band(row,text):
    s.merge_cells(f"A{row}:E{row}"); s[f"A{row}"].value=text; s[f"A{row}"].font=Font(color="FFFFFF",bold=True)
    for c in ["A","B","C","D","E"]: s[f"{c}{row}"].fill=BAND
    s.row_dimensions[row].height=15
rows_trk=[
 ("__band__","一、HBM 链(估值核心驱动)"),
 ("HBM 收入强度 (收入÷capex)","2025 锚≈4.1%(实数反推)→前瞻 base 3.7/4.4/4.2/4.0/3.8%(JPM HBM $31B/$46B@26/27; bull UBS 4.7%)",
  "「HBM测算」核心命门:HBM收入=AIDC capex×HBM收入强度。强度回落(份额/含量/价)→HBM收入与营收直接掉。",
  "公司季报 capex 明细 + TEL/AMAT 设备订单(季度);龙仁/清州 fab 进度(公司 IR、行业新闻)。",
  "扩产指引连续两季下修 → 下调产能强度假设、重算 HBM 供给 bit。"),
 ("HBM 合约价 + Hynix 份额","ASP ~$1.9/Gb(2026E);bit 份额 ~62%(25峰)→假设逐年下台阶至48%(30E)",
  "「HBM测算」需求侧/收入命门:ASP 和份额直接乘出 HBM 收入,也是峰值 ROE→目标 P/B 3.0x 的根。份额被 Samsung/Micron 蚕食→收入与倍数双杀。",
  "TrendForce HBM 月度价格与份额报告;Counterpoint;NVIDIA 供应链 channel check(月度)。",
  "份额单季掉 >3pct 或 ASP 环比转跌 → 下调份额/ASP 假设,重测隐含价。"),
 ("Samsung/Micron HBM 认证进度","Samsung HBM3E 一度未过 NVIDIA 认证(故 Hynix 25 份额冲 62%);HBM4 认证进行中",
  "份额假设的命门:三供认证一旦全部通过,Hynix 份额下台阶会更快,直接打掉 HBM 份额路径(60→48%)。",
  "NVIDIA 供应链/财报电话会 commentary;三星/美光季报认证披露(季度);行业媒体(DigiTimes 等)。",
  "竞品通过 NVIDIA HBM4 认证 → 把份额下滑斜率调陡,重算需求 bit。"),
 ("__band__","二、商品内存周期(P/E 反转的扳机)"),
 ("传统DRAM/NAND 合约价 + 库存周数","2026 价格 DRAM +112%/NAND +137%;2028 起 normalize(-30%)",
  "「DRAM_NAND测算」消费段价格变化的命门:决定净利率峰值与回落时点,也决定 P/E 周期反转('峰值年给低 PE')何时触发。价见顶+库存回升=周期拐点。",
  "Bernstein 月度合约价 tracker;BofA 周度 spot/contract tracker;DRAMeXchange/TrendForce 月度合约价。",
  "合约价连续两月转跌或库存周数回到 >8 周 → 把价格见顶年前移,重排 P/E 路径。"),
 ("Server DRAM 需求","2026 共识 +39%(GS, 剔HBM);驱动=AI 服务器 + eSSD(KV cache)",
  "「DRAM_NAND测算」服务器DRAM强度(2.8%→3.3%)的命门:server 需求撑住服务器常规DRAM 的量,挂 AIDC capex;若被涨价'需求破坏'反噬→强度下修。",
  "GS/Citi 季度 server DRAM 需求模型;hyperscaler server 出货(IDC/Counterpoint, 季度)。",
  "server DRAM 需求增速下修 >10pct → 下调服务器DRAM强度。"),
 ("内存供需(S/D)缺口","2026 DRAM 缺口 -4.9%、NAND -4.2%(GS, 15 年最紧)",
  "整个消费段价格涨幅(+112%/+137%)假设的总命门:缺口收窄=涨价逻辑松动,峰值盈利与 ROE 提前回落。",
  "GS/Citi 月度或季度 S/D balance 模型;行业产能扩张跟踪。",
  "缺口转为过剩(供给增速反超需求)→ 把商品价格 normalize 提前一年,重算净利率与 P/E。"),
 ("__band__","三、需求总盘子(物理锚)"),
 ("全球 AI capex 指引","2026 ~830B(hyperscaler 指引合计, TrendForce);2027-30 GS 'Tracking Trillions' 基线",
  "「AIDC Capex预测」物理锚的命门:HBM/服务器DRAM/eSSD 三个 AI 直驱段都按 capex×强度挂它。capex 下修→三段收入同缩。",
  "四大 hyperscaler(MSFT/GOOGL/AMZN/META)季报 capex 指引;TrendForce/GS capex 汇总(季度)。",
  "hyperscaler 合计 capex 指引下修 >10% → 直接改 AIDC Capex 行,全链重算。"),
]
rr=hr+1
for item in rows_trk:
    if item[0]=="__band__":
        trk_band(rr,item[1]); rr+=1; continue
    name,cur,why,how,act=item
    s[f"A{rr}"]=name; s[f"A{rr}"].font=BF
    for col,val in zip(["B","C","D","E"],[cur,why,how,act]):
        s[f"{col}{rr}"]=val; s[f"{col}{rr}"].font=GREY
    for col in ["A","B","C","D","E"]: s[f"{col}{rr}"].alignment=Alignment(wrap_text=True,vertical="top")
    rr+=1  # 行高不锁: wrap 文字行让 Excel 打开自动适配
s.column_dimensions["A"].width=20; s.column_dimensions["B"].width=26
s.column_dimensions["C"].width=40; s.column_dimensions["D"].width=34; s.column_dimensions["E"].width=34

# ═══════════ 全局格式 pass: 年份表头右对齐 + 冻结窗格(每张 sheet 首行+最左列)═══════════
import re
_yr=re.compile(r'^20\d\d[AE]?$')          # 2021A / 2026E / 2025 等年份表头
for ws in wb.worksheets:
    for row in ws.iter_rows(min_row=1,max_row=6):
        for cell in row:
            if isinstance(cell.value,str) and _yr.match(cell.value.strip()):
                cell.alignment=Alignment(horizontal="right")   # 年份与下方数字列右对齐
# 冻结: 值=数据起始格(其左+其上全冻结)。各表表头占行数不同。
_freeze={S_HIST:"B3", S_PX:"B4", "卖方研报共识":"A2", "历史估值倍数":"B5", S_CAPEX:"B3", S_HBM:"B3",
         S_DN:"B3", S_ASM:"B3", "情景切换":"B3", "估值倍数假设":"B4", "情景估值":"B4", "估值对比":"B6", "综合判断仪表盘":"B5"}
for _nm,_fp in _freeze.items():
    if _nm in wb.sheetnames: wb[_nm].freeze_panes=_fp

out=os.path.join(os.path.dirname(__file__),"..","out","000660.KS_valuation_model_v47.xlsx")
os.makedirs(os.path.dirname(out),exist_ok=True); wb.save(out)
print("saved; sheets:",wb.sheetnames)
