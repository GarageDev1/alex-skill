# -*- coding: utf-8 -*-
"""
build_6880hk.py — Momenta Global / 06880.HK 估值模型（equity-research-obsidian Phase 2）。

范式：per-share（已上市，HK$ 报价，235.5M 股）+ 亏损高增长 → EV/Sales 主镜头（Anthropic 式）
+ RMB 记账 / HK$ 联动 / US$ IPO 锚（三币种）。通用 sheet 复用 build_kit，亏损 EV/Sales 的
历史/倍数/利润/估值四张手搓。

物理锚链（端到端公式联动）：
  年度装车量(k) = 中国+海外 NOA 新车销量 × 独立商份额 × Momenta 占独立
  × 单车 License 费(RMB) = License 收入
  + 技术开发服务费(定点/SOP 驱动)
  = 总营收 → 毛利 → 经营利润(R&D/费用杠杆) → 经调整净利/EPS
  → EV/Sales × 前瞻营收 + 净现金 → 每股权益(RMB) → ×HKD/CNY → 隐含 HK$ 股价
  Robotaxi = 期权层（Base 不计入主链，仅 Bull 显式给值 + SOTP 分开估）

镜头：主 EV/Sales（目标=稳态 EBITDA margin × 目标 EV/EBITDA 反推 + 同业光谱）；
      DCF（path to profitability 现金流折现）交叉；SOTP（量产 License 业务 vs Robotaxi 分开）。
      P/E·P/B 退诊断行（亏损 + 优先股致会计负账面，转股后翻正）。

跑:   cd examples && PYTHONUTF8=1 python build_6880hk.py
缓存: python ../scripts/recalc.py <out>
校验: python ../scripts/validate_valuation.py <out>  → verdict ≠ FAIL 方可交付
"""
from __future__ import annotations
import json, os
from openpyxl import Workbook
import build_kit as K

# ════════ 0. 全局轴（RMB 十亿=RMB'B 记账；估值转 US$/HK$）════════
ALLC = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
ALLY = ["2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
# Momenta 追踪期仅 2023-2025（2021/22 无数据）→ 历史块用 3 年，但轴保持 5 列以复用 kit（21/22 标 n.a.）
HC, HY = ["B", "C", "D", "E", "F"], ["2021", "2022", "2023", "2024", "2025"]
FC = ["F", "G", "H", "I", "J", "K"]          # 前瞻含基年 F=2025A
FCf = FC[1:]                                  # 纯前瞻 G-K = 2026E-2030E
FWY = ["2026E", "2027E", "2028E", "2029E", "2030E"]
CASES = ["Bear", "Base", "Bull"]

# 币种
FX_CNY_USD = 6.80          # RMB per USD（2026-07 即期）
FX_HKD_USD = 7.839         # HKD per USD（联系汇率）
HKD_PER_CNY = FX_HKD_USD / FX_CNY_USD   # ≈1.1528 HK$/RMB
PX_NOW = 296.2             # 现价 HK$（2026-07-10）
SHARES_M = 235.538         # 上市后总股本（官方，未行绿鞋、不含期权）
NETCASH_RMB_B = 14.9       # 净现金（现金1.31+流动理财8.70+定期0.06+IPO净额4.86，RMB'B；≈US$2.19bn）
MKTCAP_HKD_B = round(PX_NOW * SHARES_M / 1000, 2)   # HK$'B

S_COVER, S_HIST, S_PX, S_CONS = "封面", "历史财务与估值", "股价走势", "卖方研报共识"
S_HMULT, S_MULT, S_SW = "历史估值倍数", "估值倍数假设", "情景切换"
S_ANCHOR, S_SEG, S_FUND = "物理锚_装车量单车费", "分部测算", "利润与收入假设"
S_VAL, S_CMP, S_DASH = "情景估值", "估值对比", "综合判断仪表盘"

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.environ.get("VALUATION_OUTPUT_DIR", os.path.join(REPO_ROOT, "out"))
VAULT = OUTPUT_DIR
OUT_XLSX = os.path.join(OUTPUT_DIR, "6880HK_valuation_model.xlsx")
OUT_JSON = os.path.join(OUTPUT_DIR, "6880HK_input.json")

LIC, DEV, ROB = "License 授权费", "技术开发服务费", "Robotaxi"

# ════════ 公司历史数据（RMB'B，招股书 BDO 会计师报告 IFRS；2023-2025A）════════
# 分部收入（RMB'B）
REV_LIC = [None, None, 0.023, 0.292, 0.968]      # Licensing service
REV_DEV = [None, None, 0.719, 1.032, 1.445]      # Technical development service
REV_ROB = [None, None, 0.0, 0.0, 0.0]            # Robotaxi（Track Record immaterial）
REV_TOTAL = [None, None, 0.743, 1.325, 2.413]
GP = [None, None, 0.130, 0.649, 1.727]           # 毛利
GM = [None, None, 0.175, 0.490, 0.716]           # 毛利率
OP = [None, None, -1.469, -1.291, -0.642]        # 经营亏损（推算：GP+其他收入-三费，剔利息/优先股/财务）
NI_ADJ = [None, None, -1.093, -0.959, -0.303]    # 经调整净亏损（non-IFRS，还原股份薪酬+优先股FV+上市费）
NI_REP = [None, None, -2.570, -3.206, -3.458]    # 报表净亏损（含优先股FV -1.19/-1.97/-2.84，非现金）
RD = [None, None, 1.281, 1.508, 1.869]           # R&D 费用
SGA = [None, None, 0.408, 0.476, 0.629]          # S&M + G&A
FCF_H = [None, None, -1.133, -0.966, -0.372]     # 自由现金流（OCF-capex）
OCF_H = [None, None, -1.069, -0.836, -0.281]     # 经营现金流
EQ_EXPREF = [None, None, None, None, 11.919]     # 剔除优先股后净资产（2025 末，RMB'B）

# 物理锚历史（年度装车量 k；单车 License 费 RMB，从 License收入/装车量 反推校准）
EQUIP_K = [None, None, 12.0, 150.0, 500.0]       # 年度 Momenta 装车量（千辆，估算，校准至 License 收入）
FEE_RMB = [None, None, 1933, 1947, 1936]         # 单车 License 费（RMB，= License收入/装车量）

# 上市前融资轮 post-money 估值（US$B，当"股价"序列用）
ROUNDS = [("2017-A", 0.03), ("2021-B", 0.55), ("2023-C1", 1.02), ("2024-C", 3.61),
          ("2025-C12", 5.37), ("2025-C13", 6.19), ("2026-07 IPO", 8.885)]

# 上市后日 K（HK$，2026-07-08 挂牌）
MONTHLY = [("2026-07-08", 295.6), ("2026-07-09", 295.6), ("2026-07-10", 296.2)]

# ════════ 情景参数（三案）════════
# 物理锚：年度装车量（k）三案路径
EQUIP = {"Bear": [700, 950, 1250, 1550, 1800],
         "Base": [850, 1300, 1850, 2500, 3200],
         "Bull": [1000, 1650, 2600, 3700, 4800]}
# 单车 License 费（RMB）三案路径（随 L3/Urban mix 上行）
FEE = {"Bear": [1950, 1980, 2000, 2050, 2100],
       "Base": [2000, 2100, 2250, 2400, 2550],
       "Bull": [2050, 2250, 2500, 2750, 3000]}
# 技术开发服务费（RMB'B）三案路径（定点/SOP 驱动，随 License 接力而减速）
DEVREV = {"Bear": [1.70, 1.95, 2.15, 2.30, 2.40],
          "Base": [1.85, 2.30, 2.75, 3.15, 3.50],
          "Bull": [2.00, 2.60, 3.20, 3.80, 4.30]}
# Robotaxi 收入（RMB'B）——Base=0（期权，不进主链）；Bull 2028+ 显式计入
ROBREV = {"Bear": [0.0, 0.0, 0.0, 0.0, 0.0],
          "Base": [0.0, 0.0, 0.0, 0.0, 0.0],
          "Bull": [0.0, 0.0, 0.30, 1.20, 3.00]}
# 毛利率路径（licensing mix 上升 → 边际近 100%）
GMF = {"Bear": [0.72, 0.73, 0.74, 0.75, 0.75],
       "Base": [0.74, 0.76, 0.77, 0.78, 0.79],
       "Bull": [0.75, 0.78, 0.80, 0.81, 0.82]}
# 经营利润率路径（R&D/费用杠杆；经营利润=营收×OPM）
OPMF = {"Bear": [-0.16, -0.05, 0.04, 0.10, 0.15],
        "Base": [-0.11, 0.05, 0.17, 0.27, 0.35],
        "Bull": [-0.07, 0.10, 0.24, 0.34, 0.42]}
# 净利转换率（经调整净利/经营利润；含利息收入、税；经营转正后 ~1.05 含利息，成熟后税拖累~0.85）
NETCONV = {"Bear": [1.10, 1.05, 0.95, 0.88, 0.85],
           "Base": [1.10, 1.10, 0.98, 0.90, 0.87],
           "Bull": [1.10, 1.08, 1.00, 0.92, 0.88]}
# FCF 转换率（FCF/经营利润；轻资产，营运资本随 License 放量有拖累）
FCFCONV = {"Bear": [0.70, 0.75, 0.80, 0.83, 0.85],
           "Base": [0.75, 0.82, 0.88, 0.90, 0.92],
           "Bull": [0.80, 0.88, 0.92, 0.95, 0.96]}
# 目标 EV/Sales 逐年（高增长期高、随增速换挡 normalize）
EVS = {"Bear": [8, 6, 4.5, 3.5, 3.0],
       "Base": [12, 9, 7, 5.5, 4.5],
       "Bull": [16, 13, 10, 7.5, 6.0]}
EVS_SENT = {c: [round(EVS[c][i] / EVS["Bull"][i], 3) for i in range(5)] for c in CASES}  # 相对 Bull 上沿


def rev_total(cs, i):
    """三案某年总营收（RMB'B）= License + Dev + Robotaxi。"""
    lic = EQUIP[cs][i] * FEE[cs][i] / 1e6
    return round(lic + DEVREV[cs][i] + ROBREV[cs][i], 3)


def build():
    wb = Workbook(); wb.remove(wb["Sheet"])

    # ══ 1 封面 ══
    K.write_cover(wb.create_sheet(S_COVER), {
        "title": "Momenta Global / 06880.HK 估值模型 — 物理锚(装车量×单车License费) × EV/Sales 三镜头",
        "meta": [
            ("报告日期", "2026-07-10"),
            ("数据截止", "港交所招股书(2026-06-29 聆讯后资料集, BDO 审计 2023-2025) + 上市后行情(2026-07-10) + CIC 行业"),
            ("现价", f"HK${PX_NOW}（≈US${PX_NOW/FX_HKD_USD:.1f} / ≈RMB{PX_NOW/HKD_PER_CNY:.0f}）；总市值 HK${MKTCAP_HKD_B}B ≈ US${MKTCAP_HKD_B/FX_HKD_USD:.1f}B（reality check 用，非拟合目标）"),
            ("主线镜头", "主 EV/Sales（亏损高增长，P/E·P/B 因负盈利/优先股负账面失效退诊断行）+ DCF 交叉 + SOTP（量产 License 业务 vs Robotaxi 分开）。"),
            ("口径说明", "财务 RMB，估值转 US$/HK$；净现金 RMB14.9B(≈US$2.19bn，含 IPO 净额)；账面净负债系优先股列为负债的会计假象，转股后翻为净资产 RMB11.9B。"),
            ("方法一句话", "年度装车量 × 单车 License 费 = License 收入；+ 技术开发服务费 → 总营收 → 毛利/经营利润(费用杠杆) → EV/Sales × 前瞻营收 + 净现金 → 隐含 HK$ 股价。Robotaxi 作期权(Base 不计入)。"),
        ],
        "takeaways": [
            ("① 评级/目标价", "HOLD（持有），Base 2027E 隐含 ≈ HK$295（≈现价）。量产 License 业务已被 IPO 定价充分计入；上行主要来自 Robotaxi 期权(Bull)与份额韧性。"),
            ("② 主镜头 EV/Sales(Base)", "2026E 12x→2027E 9x→2030E 4.5x（随 40%+ 增速换挡 normalize）。2027E 营收 RMB5.0B × 9x + 净现金 → 每股 ≈ HK$295，2030E ≈ HK$330(+12%)。"),
            ("③ 物理锚", "年度装车量 Base 500k(2025)→3,200k(2030)；单车 License 费 ~RMB1,940→2,550。中国 NOA 渗透率 30%→94%(2030)、Momenta 独立 Urban NOA 份额 64.5% 是量的底气，份额normalize 是关键变量。"),
            ("④ 拐点", "毛利率 17.5%→71.6%(2025)、经营现金流出 -1.07B→-0.28B、经调整净亏损率 -147%→-12.6% → Base 经营利润 2027E 转正。软件规模效应(成本刚性、License 边际近 100%)是拐点机理。"),
            ("⑤ 风险(双向)", "下行(Bear HK$~190)：份额被华为/地平线压缩、单车费停滞、2027-01 基石解禁、客户与股东重叠。上行(Bull HK$~450)：Robotaxi 2028 商业化兑现 + 份额韧性 + 单车费随 L3 上行。"),
        ],
    })

    # ══ 2 历史财务与估值（RMB'B，亏损 + per-share，手搓）══
    wh = wb.create_sheet(S_HIST)
    K.hdr(wh, 1, "Momenta 历史财务与估值（RMB'B / HK$每股）— 2023-2025A（追踪期）", 8)
    hrr = K.mtext(wh, 2, ("追踪期仅 2023-2025（2021/22 无披露，标 n.a.）。亏损 + 优先股致会计负账面 → P/E·P/B 无意义(标 N/M)，"
                          "主估值走 EV/Sales。账面净负债系优先股列为流动负债的会计假象，转股后净资产 RMB11.9B。单期股价上市后仅 3 日，估值带以 IPO 与融资轮锚。"), "H", 2)
    for col, w in zip("ABCDEFGH", [24, 9, 9, 10, 10, 10, 12, 12]):
        wh.column_dimensions[col].width = w
    wh.column_dimensions["N"].width = 60
    K.lab(wh, f"A{hrr}", "(RMB'B)", b=True)
    for col, y in zip(HC, HY):
        wh[f"{col}{hrr}"] = y; wh[f"{col}{hrr}"].font = K.BF; wh[f"{col}{hrr}"].fill = K.CH
    wh[f"G{hrr}"] = "当下(2026-07)"; wh[f"G{hrr}"].font = K.BF; wh[f"G{hrr}"].fill = K.CUR
    wh[f"N{hrr}"] = "备注/来源"; wh[f"N{hrr}"].font = K.BF; wh[f"N{hrr}"].fill = K.CH
    hrr += 1

    ROWS = {}
    def hist_row(key, label, vals, fmt=K.N1, bold=False, note="", nowval=None, nowfmt=None):
        nonlocal hrr
        K.lab(wh, f"A{hrr}", label, b=bold)
        for i, col in enumerate(HC):
            v = vals[i]
            if v is None:
                K.lab(wh, f"{col}{hrr}", "n.a.", note=True)
            else:
                K.inp(wh, f"{col}{hrr}", v, None, fmt)
        if nowval is not None:
            K.inp(wh, f"G{hrr}", nowval, None, nowfmt or fmt); wh[f"G{hrr}"].fill = K.CUR
        if note:
            K.logic(wh, f"N{hrr}", note)
        ROWS[key] = hrr; hrr += 1

    K.band(wh, hrr, "分部收入（RMB'B）", 8); hrr += 1
    hist_row("lic", "  License 授权费", REV_LIC, note="按每台量产车一次性收费(point in time)；占比 3.1%→22.1%→40.1%，成长引擎。")
    hist_row("dev", "  技术开发服务费", REV_DEV, note="SOP 前项目费(over time)；占比 96.8%→59.9%，随 License 接力而降。")
    hist_row("rob", "  Robotaxi", REV_ROB, note="Track Record immaterial(并入其他)；2028 才有意义商业化。")
    hist_row("rev", "总营收", REV_TOTAL, bold=True, note="YoY N/A / +78.4% / +82.1%。", nowval=None)
    K.lab(wh, f"A{hrr}", "  YoY", note=True)
    for j in range(1, len(HC)):
        if REV_TOTAL[j] and REV_TOTAL[j-1]:
            K.fml(wh, f"{HC[j]}{hrr}", f"={HC[j]}{ROWS['rev']}/{HC[j-1]}{ROWS['rev']}-1", K.PCT)
        else:
            K.lab(wh, f"{HC[j]}{hrr}", "n.a.", note=True)
    hrr += 1

    K.band(wh, hrr, "盈利与现金流（RMB'B）", 8); hrr += 1
    hist_row("gp", "毛利", GP, note="成本刚性(0.61→0.69B)，收入 3.2 倍。")
    hist_row("gm", "  毛利率", GM, fmt=K.PCT, bold=True, note="17.5%→71.6%：licensing 高毛利 mix 上升 + 软件规模效应。")
    hist_row("op", "经营利润(推算)", OP, note="GP + 其他收入 − 三费，剔利息/优先股/财务；亏损率 -198%→-27% 收窄。")
    hist_row("niadj", "经调整净利(non-IFRS)", NI_ADJ, bold=True, note="还原股份薪酬+优先股FV+上市费；-147%→-12.6%，逼近盈亏平衡。")
    hist_row("nirep", "报表净利", NI_REP, note="含优先股公允价值变动 -1.19/-1.97/-2.84B(非现金，转股后消失)。")
    hist_row("ocf", "经营现金流", OCF_H, note="-1.07→-0.28B，快速收窄，逼近现金平衡。")
    hist_row("fcf", "自由现金流", FCF_H, note="OCF − capex(≈0.09B/年，资本极轻，无存货)。")

    K.band(wh, hrr, "股本 / 权益 / 股价（估值口径）", 8); hrr += 1
    K.lab(wh, f"A{hrr}", "总股本(M股)")
    for col in HC:
        K.lab(wh, f"{col}{hrr}", "n.a.", note=True)
    K.inp(wh, f"G{hrr}", SHARES_M, None, K.N1); wh[f"G{hrr}"].fill = K.CUR
    K.logic(wh, f"N{hrr}", "上市后 235.5M 股(官方，未行绿鞋/不含期权)；绿鞋全行 238.5M，含期权最严摊薄 ~263M(+16.85%)。")
    ROWS["sh"] = hrr; hrr += 1
    hist_row("eq", "净资产(剔优先股，RMB'B)", EQ_EXPREF, bold=True,
             note="2025 末剔除优先股后净资产 RMB11.9B(含约 US$1.6bn 净现金/理财)；报表净负债 -17.9B 是优先股列负债的会计假象，转股后翻正。",
             nowval=round(11.919 + 4.86, 1), nowfmt=K.N1)
    K.lab(wh, f"A{hrr}", "股价 HK$(年末/现价)", b=True)
    for col in HC:
        K.lab(wh, f"{col}{hrr}", "n.a.", note=True)
    K.inp(wh, f"G{hrr}", PX_NOW, None, K.PX); wh[f"G{hrr}"].fill = K.CUR
    K.logic(wh, f"N{hrr}", "2026-07-08 IPO HK$295.6，现价 HK$296.2(几乎平价)；上市前无公开股价，估值带以融资轮/IPO 锚(见『股价走势』)。")
    ROWS["px"] = hrr; hrr += 1
    K.lab(wh, f"A{hrr}", "EV/Sales(当下 TTM，诊断)", b=True); wh[f"A{hrr}"].fill = K.OUT
    K.fml(wh, f"G{hrr}", f"=(G{ROWS['px']}*{SHARES_M}/1000/{HKD_PER_CNY}-{NETCASH_RMB_B})/F{ROWS['rev']}", K.MX)
    K.logic(wh, f"N{hrr}", f"=(市值HK$÷HKD/CNY−净现金RMB)÷2025营收；市值RMB{round(PX_NOW*SHARES_M/1000/HKD_PER_CNY,1)}B−净现金{NETCASH_RMB_B}=EV，÷2.413≈现价对应 TTM EV/Sales。P/S(含现金)≈{round(PX_NOW*SHARES_M/1000/HKD_PER_CNY/2.413,1)}x。")
    ROWS["evs_now"] = hrr; hrr += 1
    ha = {"rev": ROWS["rev"], "px": ROWS["px"], "gp": ROWS["gp"], "niadj": ROWS["niadj"],
          "lic": ROWS["lic"], "dev": ROWS["dev"], "eq": ROWS["eq"]}
    K.set_widths(wh, 24, HC + ["G"], 10, logic_col="N", logic_width=60)

    # ══ 3 股价走势（融资轮 post-money + IPO 当"股价"序列）══
    def round_phase(ym):
        if ym <= "2022-12":
            return "① 早期融资(估值估)"
        if ym <= "2024-12":
            return "② 战略资本(车企)入场"
        if ym <= "2025-12":
            return "③ 量产放量估值抬升"
        return "④ IPO 上市"
    K.write_price_chart(wb.create_sheet(S_PX), ROUNDS, {
        "fn": round_phase,
        "rows": [("① 早期融资(估值估)", "2017 A $0.03B → 2021 B $0.55B；曹旭东团队从商汤出走创立。"),
                 ("② 战略资本(车企)入场", "2023-2024 C 系列估值升至 $3.6B；上汽/通用/奔驰/丰田等车企战略入股。"),
                 ("③ 量产放量估值抬升", "2025 C12/C13 $5.4B→$6.19B(2025-12，每股 US$25.28)；SOP 车型 26→68、装机破 68 万。"),
                 ("④ IPO 上市", "2026-07-08 港股挂牌，招股价 HK$295.60、估值 US$8.885B(较末轮 +44%)；首日平价、无溢价。")],
    }, title="Momenta 融资轮 post-money + IPO 估值（US$B）")

    # ══ 4 卖方研报共识（无首覆，共识空白 + comps 参照）══
    K.write_consensus(wb.create_sheet(S_CONS), {
        "title": "卖方研报共识 — 截至 2026-07-10 无卖方首覆（承销静默期），共识空白",
        "overview": ("Momenta 2026-07-08 刚上市，主流券商仍在承销静默期，无首次覆盖、无目标价、无盈利预测。"
                     "本地研报库仅在 Bernstein 智驾行业报告中被提及(别克 Electra L7 搭载城市 NOA)。估值只能靠可比公司(WeRide/地平线/robotaxi 期权)+ 招股书财务自建，无街道一致预期可对标——这正是共识空白期的 alpha 窗口。"),
        "assumptions": [
            ("2026E 营收增速",
             "无卖方预测。招股书 2025 +82.1%；管理层未给 2026 指引。",
             "License 放量节奏 vs 定点转化速度。",
             "本模型 Base 2026E +47%(装车量 500k→850k、单车费升 + 技术开发服务费减速)。"),
            ("单车 License 费",
             "招股书未直接披露。2025 反推 ≈RMB1,936(License 968M ÷ 装车量 ~500k)。",
             "L3/Urban mix 抬价 vs 竞争压价、下沉车型稀释。",
             "本模型 Base ~RMB1,940→2,550(2030)，随 L3/Urban 上行、竞争部分对冲。"),
            ("盈亏平衡时点",
             "无卖方预测。经调整净亏损率 -147%→-12.6%(2025)。",
             "R&D(占收入 77.5%)增速能否持续低于收入增速(费用杠杆)。",
             "本模型 Base 经营利润 2027E 转正、经调整净利 2027E 微正。"),
            ("目标估值倍数",
             "无卖方。可比：WeRide(UBS +24.5% 回报预期)、地平线、Waymo 期权。",
             "亏损期用 EV/Sales，稳态 EBITDA margin 决定 normalize 水平。",
             "本模型 Base EV/Sales 2027E 9x(高增长换挡)，稳态经营利润率 ~35%。"),
        ],
        "divergences": [
            "① 无卖方锚 → 估值全靠可比 + 招股书自建，市场处于价格发现期(首日平价、二级观望)。",
            "② 单车 License 费与年度装车量未直接披露 → 物理锚需估算+校准(License 收入反推)，是最大不确定。",
            "③ Robotaxi 期权定价：市场为一个 2028 才商业化、2025 仅 31 台在运的业务付了多少？(SOTP 分开看)",
        ],
        "stances": [
            "承销静默期：CICC / Deutsche Bank / J.P. Morgan(联席保荐) 首覆预计上市后 ~40 天(2026-08 中下旬)。",
            "基石(占发行 49.95%)：GIC/Fidelity 各 $100M、BlackRock/Mercedes 各 $25M、BYD $15M 等 → 一级最聪明的钱背书，6 个月锁定。",
            "可比 robotaxi(WeRide/文远)：UBS 2026-01 给 WeRide forecast return +24.5%；库里唯一带估值口径的同赛道标的。",
            "本模型 Base：HOLD，量产 License 业务已被 IPO 充分定价，Robotaxi 是未定价期权 → 上行看 Bull。",
        ],
    })

    # ══ 5 历史估值倍数（EV/Sales 带 + 同业光谱，手搓）══
    wm = wb.create_sheet(S_HMULT)
    K.hdr(wm, 1, "历史估值倍数 — EV/Sales 锚(无自身历史带) + 同业 EV/Sales 光谱", 8)
    mr = K.mtext(wm, 2, ("Momenta 次新股、无自身历史估值带 → 本页两件事：① IPO/融资轮 EV/Sales 当估值锚；"
                         "② 同业 EV/Sales 光谱(智驾/robotaxi 高增长组 vs 成熟软件组，不可跨组比)。"
                         "亏损 → P/E·P/B 失效，主镜头取 EV/前瞻营收，高增长期高、随增速换挡 normalize。"), "H", 2)
    for col, w in zip("ABCDEFGH", [24, 12, 12, 12, 12, 12, 12, 12]):
        wm.column_dimensions[col].width = w
    wm.column_dimensions["N"].width = 56

    K.band(wm, mr, "① 自身估值锚（IPO / 末轮 / 现价 EV/Sales）", 8); mr += 1
    for col, h in zip(["A", "B", "C", "D"], ["时点/口径", "估值US$B", "对应营收US$B", "EV/Sales"]):
        wm[f"{col}{mr}"] = h; wm[f"{col}{mr}"].font = K.BF; wm[f"{col}{mr}"].fill = K.CH
    wm.merge_cells(f"E{mr}:N{mr}"); wm[f"E{mr}"] = "备注"; wm[f"E{mr}"].font = K.BF; wm[f"E{mr}"].fill = K.CH
    mr += 1
    rev25_usd = round(2.413 / FX_CNY_USD, 3)
    band_rows = [
        ("2025-12 末轮(C13)", 6.19, rev25_usd, "每股 US$25.28；EV/Sales 按 2025 营收 US$0.355B(含现金口径 P/S)。"),
        ("2026-07 IPO", 8.885, rev25_usd, "招股价 HK$295.60；较末轮 6 个月 +44%。"),
        ("2026-07-10 现价", round(PX_NOW*SHARES_M/1000/FX_HKD_USD, 3), rev25_usd, "现价 HK$296.2；EV=市值−净现金 US$2.19B。"),
    ]
    for nm, val, rev, note in band_rows:
        wm[f"A{mr}"] = nm; wm[f"A{mr}"].font = K.BF
        K.inp(wm, f"B{mr}", val, None, K.N1)
        K.inp(wm, f"C{mr}", rev, None, K.N2)
        if "现价" in nm:
            K.fml(wm, f"D{mr}", f"=(B{mr}-{round(NETCASH_RMB_B/FX_CNY_USD,2)})/C{mr}", K.MX)
        else:
            K.fml(wm, f"D{mr}", f"=B{mr}/C{mr}", K.MX)
        wm.merge_cells(f"E{mr}:N{mr}"); K.logic(wm, f"E{mr}", note)
        mr += 1
    mr += 1

    K.band(wm, mr, "② 同业 EV/Sales 光谱（两组，不可跨组比）", 8); mr += 1
    for col, h in zip(["A", "B", "C", "D"], ["公司/口径", "EV/Sales(fwd)", "增速", "业务特征"]):
        wm[f"{col}{mr}"] = h; wm[f"{col}{mr}"].font = K.BF; wm[f"{col}{mr}"].fill = K.CH
    wm.merge_cells(f"D{mr}:N{mr}"); mr += 1
    peers = [
        ("Momenta(本案 2027E)", 9.0, "+42%", "本模型 Base 2027E 目标；亏损转正、毛利 76%、装车量高增长"),
        ("WeRide 文远(robotaxi)", 15.0, "高增速", "纯 robotaxi，未盈利，情绪锚；UBS 给 forecast return +24.5%"),
        ("地平线 Horizon(9660.HK)", 12.0, "+50%+", "智驾芯片+方案，港股，高增速高倍数"),
        ("Palantir(AI-SaaS 情绪锚)", 40.0, "+40%", "上市 AI-SaaS 情绪上沿"),
        ("成熟软件(SAP/Adobe 组)", 8.0, "+12%", "成熟 SaaS 下沿 7-10x fwd，增速远低"),
        ("英伟达对照", 15.0, "+50%", "硬件层高增速对照"),
    ]
    for nm, evs_v, g, note in peers:
        wm[f"A{mr}"] = nm; wm[f"A{mr}"].font = K.BF
        if nm.startswith("Momenta"):
            wm[f"A{mr}"].fill = K.CUR
        K.inp(wm, f"B{mr}", evs_v, None, K.MX)
        K.lab(wm, f"C{mr}", g)
        wm.merge_cells(f"D{mr}:N{mr}"); K.logic(wm, f"D{mr}", note)
        mr += 1
    mr += 1
    K.band(wm, mr, "③ 读法 — 给『估值倍数假设』的输入", 8); mr += 1
    K.mtext(wm, mr, ("Momenta 在高增长智驾组，倍数应高于成熟软件(8x)、可对标地平线/WeRide(12-15x)但低于纯情绪锚(Palantir 40x)。"
                     "现价对应 TTM EV/Sales ~18x(2025 营收)、~9x(2026E)——市场已给高增长溢价。主镜头取 EV/前瞻营收，"
                     "Base 2027E 9x(增速换挡)，向成熟软件 normalize；下一页拍三案逐年 EV/Sales。"), "N", 3)

    # ══ 6 估值倍数假设（EV/Sales 三案逐年 + 稳态反推，手搓）══
    wu = wb.create_sheet(S_MULT)
    K.hdr(wu, 1, "估值倍数假设 — EV/Sales 三案逐年(2026E-2030E) + 稳态 margin 反推", 8)
    ur = K.mtext(wu, 2, ("主镜头=EV/Sales(亏损，P/E·P/B 主线出负/NA)。目标 EV/Sales 两条腿定："
                         "① 稳态 EBITDA margin × 目标 EV/EBITDA 反推(自洽)；② 同业 fwd 光谱(智驾组 12-15x、成熟软件 8x)。"
                         "高增长期给高倍数、随增速换挡 normalize。『情景切换』link 本页三案行并切换，『情景估值』套用。"), "H", 2)
    for col, w in zip("ABCDEFGH", [24, 11, 11, 11, 11, 11, 12, 12]):
        wu.column_dimensions[col].width = w
    wu.column_dimensions["N"].width = 62

    K.band(wu, ur, "稳态反推（自洽锚）", 8); ur += 1
    K.lab(wu, f"A{ur}", "稳态 EBITDA margin"); K.inp(wu, f"B{ur}", 0.37, None, K.PCT)
    K.logic(wu, f"N{ur}", "2030E Base 经营利润率 ~35% + D&A → EBITDA margin ~37%(软件规模效应，R&D 仍重)。")
    ss_m = ur; ur += 1
    K.lab(wu, f"A{ur}", "稳态目标 EV/EBITDA"); K.inp(wu, f"B{ur}", 12.0, None, K.MX)
    K.logic(wu, f"N{ur}", "成熟高质量软件 EV/EBITDA 12x(增速换挡后)。")
    ss_ee = ur; ur += 1
    K.lab(wu, f"A{ur}", "反推稳态 EV/Sales", b=True); wu[f"A{ur}"].fill = K.OUT
    K.fml(wu, f"B{ur}", f"=B{ss_m}*B{ss_ee}", K.MX)
    K.logic(wu, f"N{ur}", "= 稳态 EBITDA margin × 目标 EV/EBITDA ≈ 4.4x，与 Base 2030E 4.5x 自洽。")
    ur += 2

    K.band(wu, ur, "EV/Sales 逐年(x) — 三案", 8); ur += 1
    wu[f"A{ur}"] = "案 / 年份"; wu[f"A{ur}"].font = K.BF
    for col, y in zip(FCf, FWY):
        wu[f"{col}{ur}"] = y; wu[f"{col}{ur}"].font = K.BF; wu[f"{col}{ur}"].fill = K.CH
    wu[f"N{ur}"] = "为什么这么给"; wu[f"N{ur}"].font = K.BF; wu[f"N{ur}"].fill = K.CH
    ur += 1
    EVS_ROWS = {}
    evs_notes = {
        "Bear": "份额被华为/地平线压缩 + 增速打断 → 8x 起快速收敛到 3x(成熟软件下沿之下)。",
        "Base": "高增长换挡：2026E 12x → 2027E 9x(对标地平线/成熟软件之间) → 2030E 4.5x(向稳态反推 4.4x 收敛)。",
        "Bull": "份额韧性 + Robotaxi 期权兑现 + 单车费随 L3 上行：16x 起，normalize 到 6x(全程高于 Base)。",
    }
    for cs in CASES:
        K.lab(wu, f"A{ur}", f"  {cs}")
        K.introw(wu, ur, FCf, EVS[cs], None, K.MX)
        K.logic(wu, f"N{ur}", evs_notes[cs])
        EVS_ROWS[cs] = ur; ur += 1
    ur += 1
    K.mtext(wu, ur, ("EV/Sales(本页三案行)喂『情景切换』当前案；估值对比页直接引三案行(防当前案污染)。"
                     "Base 2027E 9x：现价对应 2026E ~9x、2027E ~6x(以模型营收算)——市场已 price-in 高增长；9x 是『增速换挡中段的合理倍数』判断，非现价拟合。"), "N", 3)
    ev_sales_rows = {cs: EVS_ROWS[cs] for cs in CASES}

    # ══ 7 情景切换 ══
    sw = K.write_scenario_switch(wb.create_sheet(S_SW), {
        "title": "情景切换 — 全模型唯一情景参数库（默认 Base）",
        "usage": "B2 是唯一开关。装车量增速、单车费、技术开发服务费、毛利率、经营利润率、净利转换、EV/Sales 都按当前案联动；估值对比页直接引三案矩阵，不被开关污染。Robotaxi 仅 Bull 计入主链。",
        "cases": CASES, "default": "Base",
        "triggers": [
            ("Bear", "华为/地平线抢占份额、Momenta 独立 Urban NOA 份额从 64.5% 快速回落；单车费被下沉车型与竞争压制；L3/Robotaxi 延后；客户(与股东重叠)收缩定点。"),
            ("Base", "份额温和 normalize(64.5%→blended ~25-30%)但绝对装车量随 NOA 渗透率 30%→94% 高增长；单车费随 L3/Urban mix 温和上行；经营利润 2027E 转正；Robotaxi 作期权不计入。"),
            ("Bull", "份额韧性(数据积累正反馈+交付效率壁垒守住)；单车费随 L3 显著上行；Robotaxi 2028 商业化兑现并显式计入；EV/Sales 维持高位。"),
        ],
        "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
        "levers": [
            {"key": "equip", "name": "年度装车量(千辆)", "fmt": K.N0, "cols": FCf,
             "vals": EQUIP, "desc": "物理锚第一层：年度 Momenta 装车量 = 中国+海外 NOA 新车销量 × 独立商份额 × Momenta 占独立。2025 ≈500k(累计破 68 万，2026 中破 100 万)。",
             "stories": {"Bear": "份额被压缩，装车量增速腰斩，2030 仅 1,800k。",
                         "Base": "NOA 渗透率 30%→94% 拉动，份额 normalize 但绝对量高增长，2030 3,200k。",
                         "Bull": "份额韧性 + 出海放量，2030 4,800k。"},
             "hist": [None, None, None, None, None]},
            {"key": "fee", "name": "单车License费(RMB)", "fmt": K.N0, "cols": FCf,
             "vals": FEE, "desc": "物理锚第二层：每台量产车一次性 License 费。2025 反推 ≈RMB1,936(License 968M÷装车量 ~500k)。随 L3/Urban mix 上行。",
             "stories": {"Bear": "竞争压价 + 下沉车型稀释，单车费停在 RMB1,950-2,100。",
                         "Base": "L3/Urban mix 温和抬价，RMB1,940→2,550。",
                         "Bull": "L3 显著抬价 + 高端车型，RMB2,050→3,000。"},
             "hist": [None, None, None, None, None]},
            {"key": "dev", "name": "技术开发服务费(RMB'B)", "fmt": K.N1, "cols": FCf,
             "vals": DEVREV, "desc": "定点/SOP 驱动的项目费(over time)。2025 RMB1.445B，随 License 接力而减速增长。",
             "stories": {"Bear": "新增定点放缓(2025 已 52<2024 的 69)，2030 仅 2.4B。",
                         "Base": "定点稳增，2030 3.5B。",
                         "Bull": "定点加速 + 海外 OEM，2030 4.3B。"},
             "hist": [None, None, None, None, None]},
            {"key": "robrev", "name": "Robotaxi收入(RMB'B)", "fmt": K.N1, "cols": FCf,
             "vals": ROBREV, "desc": "期权层。Base=0(不进主链，2025 仅 31 台在运)；Bull 2028 商业化后计入。",
             "stories": {"Bear": "商业化持续延后，0。",
                         "Base": "作期权，不计入 Base 主链(SOTP 单独看)。",
                         "Bull": "2028 上海/Abu Dhabi/Munich 车队起量，2030 RMB3.0B。"},
             "hist": [None, None, None, None, None]},
            {"key": "gm", "name": "毛利率", "fmt": K.PCT, "cols": FCf,
             "vals": GMF, "desc": "licensing 边际近 100% + 规模效应。2025 实际 71.6%。",
             "stories": {"Bear": "价格战侵蚀，72-75%。", "Base": "mix 上升，74→79%。", "Bull": "licensing 主导，75→82%。"},
             "hist": [None, None, 0.175, 0.490, 0.716]},
            {"key": "opm", "name": "经营利润率", "fmt": K.PCT, "cols": FCf,
             "vals": OPMF, "desc": "R&D(占收入 77.5%)与费用杠杆。2025 经营亏损率 -27%。经营利润=营收×本行。",
             "stories": {"Bear": "R&D 降不下来，2030 仅 15%。", "Base": "费用杠杆释放，2027 转正、2030 35%。", "Bull": "规模效应充分 + Robotaxi 摊薄，2030 42%。"},
             "hist": [None, None, -1.469/0.743, -1.291/1.325, -0.642/2.413]},
            {"key": "netconv", "name": "净利转换率", "fmt": K.N2, "cols": FCf,
             "vals": NETCONV, "desc": "经调整净利/经营利润(含利息收入~0.15B/年、税)。经营转正后 ~1.05(利息加成)，成熟后税拖累~0.87。",
             "stories": {"Bear": "税+SBC 拖累，0.85-1.10。", "Base": "利息加成后税归一，0.87-1.10。", "Bull": "税优化，0.88-1.10。"},
             "hist": [None, None, None, None, None]},
            {"key": "fcfconv", "name": "FCF转换率", "fmt": K.N2, "cols": FCf,
             "vals": FCFCONV, "desc": "FCF/经营利润(轻资产，营运资本随 License 放量有拖累)。喂 DCF。",
             "stories": {"Bear": "应收拖累，0.70-0.85。", "Base": "营运资本改善，0.75-0.92。", "Bull": "现金转化强，0.80-0.96。"},
             "hist": [None, None, None, None, None]},
        ],
        "linked": [
            {"key": "evs", "name": "目标 EV/Sales(当前案)", "fmt": K.MX,
             "src_sheet": S_MULT, "src_row0": ev_sales_rows["Bear"],
             "note": "来自估值倍数假设页三案行(Bear/Base/Bull 连续行)，按当前案 INDEX。"},
        ],
    })
    # 目标 EV/Sales 当前案行（引倍数页三案行，用 CHOOSE/INDEX 按案序号）——手动补，因三案行连续
    swact = sw["SWACT"]
    # linked 已由 write_scenario_switch 处理 evs 当前案行；取其行号
    evs_act_row = swact.get("evs")

    # ══ 8 物理锚（装车量 × 单车费；[ANCHOR] 标记）══
    anchor = K.write_anchor(wb.create_sheet(S_ANCHOR), {
        "title": "[ANCHOR] 物理锚 — 年度装车量(千辆) × 单车 License 费(RMB) → License 收入",
        "all_cols": ALLC, "all_years": ALLY,
        "series": [
            ("年度装车量 (千辆)", [None, None, 12.0, 150.0, 500.0] + [None]*5,
             "物理锚第一层。= 中国+海外 NOA 新车销量 × 独立商份额 × Momenta 占独立。2025 ≈500k(累计破 68 万)。前瞻=情景切换当前案。", K.N0),
            ("单车 License 费 (RMB)", [None, None, 1933.0, 1947.0, 1936.0] + [None]*5,
             "物理锚第二层。= License 收入 ÷ 装车量(反推校准，招股书未直接披露)。前瞻=情景切换当前案。", K.N0),
            ("License 收入 (RMB'B)", [None, None, 0.023, 0.292, 0.968] + [None]*5,
             "= 装车量 × 单车费 / 1e6。历史与招股书 License 收入对账一致。", K.N2),
        ],
        "yoy_row": "年度装车量 (千辆)",
        "source_note": "装车量/单车费=招股书装机数 + CIC 渗透率/份额 + License 收入反推(标估算)；License 收入=招股书 Note 5。",
        "role_note": "License 收入 = 装车量 × 单车费。改装车量或单车费(情景切换) → 穿透到 License 收入、总营收、经营利润、EV/Sales 估值、隐含价(连通性测试入口)。",
    })
    R_EQ = anchor["row_of"]["年度装车量 (千辆)"]
    R_FEE = anchor["row_of"]["单车 License 费 (RMB)"]
    R_LICR = anchor["row_of"]["License 收入 (RMB'B)"]
    for col in FCf:
        K.fml(wb[S_ANCHOR], f"{col}{R_EQ}", f"={K.R(S_SW, col + str(swact['equip']))}", K.N0, link=True)
        K.fml(wb[S_ANCHOR], f"{col}{R_FEE}", f"={K.R(S_SW, col + str(swact['fee']))}", K.N0, link=True)
    for col in FC:
        K.fml(wb[S_ANCHOR], f"{col}{R_LICR}", f"={col}{R_EQ}*{col}{R_FEE}/1000000", K.N2)
    # 物理锚 YoY 行(=R_EQ+1)早期列(2021/2022 无装车量)会除空值 → 标 n.a. 防 #DIV/0!
    yoy_r = R_EQ + 1
    for col in ["B", "C", "D"]:
        wb[S_ANCHOR][f"{col}{yoy_r}"] = None
        K.lab(wb[S_ANCHOR], f"{col}{yoy_r}", "n.a.", note=True)

    # ══ 9 分部测算（License + Dev + Robotaxi → 总营收）══
    seg = K.write_segment_model(wb.create_sheet(S_SEG), {
        "title": "分部测算 — License(物理锚) + 技术开发服务费 + Robotaxi → 总营收",
        "all_cols": ALLC, "all_years": ALLY, "logic_col": "N",
        "groups": [
            ("收入分部（RMB'B）", [
                ("License 授权费", None, K.N2, "引自 [物理锚] License 收入(装车量×单车费)。"),
                ("技术开发服务费", None, K.N1, "引自 [情景切换] 当前案(定点/SOP 驱动)。历史=实际。"),
                ("Robotaxi", None, K.N1, "引自 [情景切换] 当前案(Base=0 期权层，仅 Bull 计入)。"),
                ("总营收", None, K.N1, "= License + 技术开发服务费 + Robotaxi。"),
            ]),
        ],
    })
    m = seg["m"]
    for i, col in enumerate(HC):
        # 历史（2023-2025 在 D/E/F）
        if REV_LIC[i] is not None:
            K.inp(wb[S_SEG], f"{col}{m['License 授权费']}", REV_LIC[i], None, K.N2)
            K.inp(wb[S_SEG], f"{col}{m['技术开发服务费']}", REV_DEV[i], None, K.N1)
            K.inp(wb[S_SEG], f"{col}{m['Robotaxi']}", 0.0, None, K.N1)
        else:
            for kk in ['License 授权费', '技术开发服务费', 'Robotaxi']:
                K.lab(wb[S_SEG], f"{col}{m[kk]}", "n.a.", note=True)
    for col in FCf:
        K.fml(wb[S_SEG], f"{col}{m['License 授权费']}", f"={K.R(S_ANCHOR, col + str(R_LICR))}", K.N2, link=True)
        K.fml(wb[S_SEG], f"{col}{m['技术开发服务费']}", f"={K.R(S_SW, col + str(swact['dev']))}", K.N1, link=True)
        K.fml(wb[S_SEG], f"{col}{m['Robotaxi']}", f"={K.R(S_SW, col + str(swact['robrev']))}", K.N1, link=True)
    for col in FC:
        K.fml(wb[S_SEG], f"{col}{m['总营收']}", f"={col}{m['License 授权费']}+{col}{m['技术开发服务费']}+{col}{m['Robotaxi']}", K.N1)
    for col in HC[:2]:
        K.lab(wb[S_SEG], f"{col}{m['总营收']}", "n.a.", note=True)
    for i, col in enumerate(HC[2:], start=2):
        K.inp(wb[S_SEG], f"{col}{m['总营收']}", REV_TOTAL[i], None, K.N1)

    # ══ 10 利润与收入假设（营收 → 毛利 → 经营利润 → 经调整净利/EPS/FCF）══
    wf = wb.create_sheet(S_FUND)
    K.hdr(wf, 1, "利润与收入假设 — 总营收 → 毛利 → 经营利润(费用杠杆) → 经调整净利 / EPS / FCF", 8)
    fr0 = K.mtext(wf, 2, ("营收接分部测算；毛利率/经营利润率/净利转换/FCF转换=情景切换当前案(历史实际列在左)。"
                          "亏损期主口径=经调整净利(非报表，剔优先股 FV/股份薪酬)；EPS=经调整净利/股本；BPS/ROE 亏损期无意义(标 n.m.)。"), "H", 2)
    for col, w in zip("ABCDEFGH", [24, 9, 9, 10, 10, 10, 10, 10]):
        wf.column_dimensions[col].width = w
    wf.column_dimensions["N"].width = 60
    K.lab(wf, f"A{fr0}", "(RMB'B / RMB每股)", b=True)
    for col, y in zip(ALLC, ALLY):
        wf[f"{col}{fr0}"] = y; wf[f"{col}{fr0}"].font = K.BF; wf[f"{col}{fr0}"].fill = K.CH
    fr0 += 1
    FR = {}
    def frow(key, label, fmt=K.N1, bold=False, note=""):
        nonlocal fr0
        K.lab(wf, f"A{fr0}", label, b=bold)
        if note:
            K.logic(wf, f"N{fr0}", note)
        FR[key] = fr0; fr0 += 1

    rev = lambda c: f"{c}{FR['rev']}"
    K.band(wf, fr0, "收入与毛利", 8); fr0 += 1
    frow("rev", "总营收", bold=True, note="引自分部测算总营收。")
    for col in ALLC:
        if col in HC[:2]:
            K.lab(wf, f"{col}{FR['rev']}", "n.a.", note=True)
        else:
            K.fml(wf, f"{col}{FR['rev']}", f"={K.R(S_SEG, col + str(m['总营收']))}", K.N1, link=True)
    frow("gmr", "毛利率", fmt=K.PCT, note="历史实际；前瞻=情景切换当前案。")
    for i, col in enumerate(HC):
        v = GM[i]
        K.lab(wf, f"{col}{FR['gmr']}", "n.a." if v is None else None, note=True) if v is None else K.inp(wf, f"{col}{FR['gmr']}", v, None, K.PCT)
    for col in FCf:
        K.fml(wf, f"{col}{FR['gmr']}", f"={K.R(S_SW, col + str(swact['gm']))}", K.PCT, link=True)
    frow("gp", "毛利", bold=True, note="= 营收 × 毛利率。")
    for col in ALLC:
        if col in HC[:2]:
            K.lab(wf, f"{col}{FR['gp']}", "n.a.", note=True)
        else:
            K.fml(wf, f"{col}{FR['gp']}", f"={rev(col)}*{col}{FR['gmr']}", K.N1)

    K.band(wf, fr0, "经营利润与净利（费用杠杆 → path to profitability）", 8); fr0 += 1
    frow("opm", "经营利润率", fmt=K.PCT, note="历史实际(亏损)；前瞻=情景切换当前案，反映 R&D/费用杠杆。")
    hist_opm = [None, None, -1.469/0.743, -1.291/1.325, -0.642/2.413]
    for i, col in enumerate(HC):
        if hist_opm[i] is None:
            K.lab(wf, f"{col}{FR['opm']}", "n.a.", note=True)
        else:
            K.inp(wf, f"{col}{FR['opm']}", round(hist_opm[i], 4), None, K.PCT)
    for col in FCf:
        K.fml(wf, f"{col}{FR['opm']}", f"={K.R(S_SW, col + str(swact['opm']))}", K.PCT, link=True)
    frow("op", "经营利润", bold=True, note="= 营收 × 经营利润率。2027E 转正。")
    for col in ALLC:
        if col in HC[:2]:
            K.lab(wf, f"{col}{FR['op']}", "n.a.", note=True)
        else:
            K.fml(wf, f"{col}{FR['op']}", f"={rev(col)}*{col}{FR['opm']}", K.N1)
    frow("nc", "净利转换率", fmt=K.N2, note="经调整净利/经营利润(含利息收入~0.15B、税)；历史 n.m.(经营亏损)。")
    for col in HC:
        K.lab(wf, f"{col}{FR['nc']}", "n.m.", note=True)
    for col in FCf:
        K.fml(wf, f"{col}{FR['nc']}", f"={K.R(S_SW, col + str(swact['netconv']))}", K.N2, link=True)
    frow("ni", "经调整净利", bold=True, note="历史=non-IFRS 实际；前瞻=经营利润 × 净利转换。")
    for i, col in enumerate(HC):
        if NI_ADJ[i] is None:
            K.lab(wf, f"{col}{FR['ni']}", "n.a.", note=True)
        else:
            K.inp(wf, f"{col}{FR['ni']}", NI_ADJ[i], None, K.N1)
    for col in FCf:
        K.fml(wf, f"{col}{FR['ni']}", f"={col}{FR['op']}*{col}{FR['nc']}", K.N1)
    frow("nim", "  经调整净利率", fmt=K.PCT, note="-147%→-12.6%(2025)→转正。")
    for col in ALLC:
        if col in HC[:2]:
            K.lab(wf, f"{col}{FR['nim']}", "n.a.", note=True)
        else:
            K.fml(wf, f"{col}{FR['nim']}", f"={col}{FR['ni']}/{rev(col)}", K.PCT)
    frow("eps", "经调整 EPS (RMB)", fmt=K.N2, note=f"= 经调整净利 × 1000 / {SHARES_M}M 股。")
    for col in ALLC:
        if col in HC[:2]:
            K.lab(wf, f"{col}{FR['eps']}", "n.a.", note=True)
        else:
            K.fml(wf, f"{col}{FR['eps']}", f"={col}{FR['ni']}*1000/{SHARES_M}", K.N2)

    K.band(wf, fr0, "现金盈利锚（喂 DCF）", 8); fr0 += 1
    frow("fcfc", "FCF转换率", fmt=K.N2, note="FCF/经营利润；历史 n.m.(经营亏损)。")
    for col in HC:
        K.lab(wf, f"{col}{FR['fcfc']}", "n.m.", note=True)
    for col in FCf:
        K.fml(wf, f"{col}{FR['fcfc']}", f"={K.R(S_SW, col + str(swact['fcfconv']))}", K.N2, link=True)
    frow("fcf", "自由现金流 FCF", bold=True, note="历史实际；前瞻=经营利润 × FCF转换率(经营转正后)。")
    for i, col in enumerate(HC):
        if FCF_H[i] is None:
            K.lab(wf, f"{col}{FR['fcf']}", "n.a.", note=True)
        else:
            K.inp(wf, f"{col}{FR['fcf']}", FCF_H[i], None, K.N1)
    for col in FCf:
        K.fml(wf, f"{col}{FR['fcf']}", f"={col}{FR['op']}*{col}{FR['fcfc']}", K.N1)

    K.band(wf, fr0, f"USD 换算（供估值/对标；FX=RMB{FX_CNY_USD}/US$）", 8); fr0 += 1
    frow("revusd", "总营收 (US$B)", note=f"= RMB 总营收 ÷ {FX_CNY_USD}；估值/对标口径。2025 ≈ US$0.36B。")
    for col in ALLC:
        if col in HC[:2]:
            K.lab(wf, f"{col}{FR['revusd']}", "n.a.", note=True)
        else:
            K.fml(wf, f"{col}{FR['revusd']}", f"={col}{FR['rev']}/{FX_CNY_USD}", K.N2)
    frow("niusd", "经调整净利 (US$B)", note=f"= RMB 经调整净利 ÷ {FX_CNY_USD}。")
    for col in ALLC:
        if col in HC[:2]:
            K.lab(wf, f"{col}{FR['niusd']}", "n.a.", note=True)
        else:
            K.fml(wf, f"{col}{FR['niusd']}", f"={col}{FR['ni']}/{FX_CNY_USD}", K.N2)
    K.set_widths(wf, 24, ALLC, 10, logic_col="N", logic_width=60)
    fr = {"rev": FR["rev"], "ni": FR["ni"], "eps": FR["eps"], "op": FR["op"], "fcf": FR["fcf"], "gp": FR["gp"]}

    # ══ 11 情景估值（EV/Sales 主 + DCF + SOTP，per-share HK$，手搓）══
    val = write_momenta_valuation(wb.create_sheet(S_VAL), {
        "s_fund": S_FUND, "fr": fr, "s_sw": S_SW, "sw_cell": sw["sw_cell"], "swact": swact,
        "s_mult": S_MULT, "evs_rows": ev_sales_rows,
    })

    # ══ 12 估值对比（三案并排）══
    swb = sw["SWB"]
    def cr(key, ci): return swb[key] + ci
    def prevcol(col): return ALLC[ALLC.index(col) - 1]

    def rev_hist(c, ci, a):
        i = HC.index(c)
        return "=\"n.a.\"" if REV_TOTAL[i] is None else f"={REV_TOTAL[i]}"

    rows = [
        {"key": "equip", "label": "年度装车量(千辆)", "fmt": K.N0,
         "hist": lambda c, ci, a: (f"={EQUIP_K[HC.index(c)]}" if EQUIP_K[HC.index(c)] else "=\"n.a.\""),
         "fwd": lambda c, j, ci, a: f"={K.R(S_SW, c + str(cr('equip', ci)))}"},
        {"key": "fee", "label": "单车License费(RMB)", "fmt": K.N0,
         "hist": lambda c, ci, a: (f"={FEE_RMB[HC.index(c)]}" if FEE_RMB[HC.index(c)] else "=\"n.a.\""),
         "fwd": lambda c, j, ci, a: f"={K.R(S_SW, c + str(cr('fee', ci)))}"},
        {"key": "licr", "label": "License 收入(RMB'B)", "fmt": K.N2,
         "hist": lambda c, ci, a: (f"={REV_LIC[HC.index(c)]}" if REV_LIC[HC.index(c)] is not None else "=\"n.a.\""),
         "fwd": lambda c, j, ci, a: f"={c}{a['equip']}*{c}{a['fee']}/1000000"},
        {"key": "devr", "label": "技术开发服务费(RMB'B)", "fmt": K.N1,
         "hist": lambda c, ci, a: (f"={REV_DEV[HC.index(c)]}" if REV_DEV[HC.index(c)] is not None else "=\"n.a.\""),
         "fwd": lambda c, j, ci, a: f"={K.R(S_SW, c + str(cr('dev', ci)))}"},
        {"key": "robr", "label": "Robotaxi(RMB'B)", "fmt": K.N1,
         "hist": lambda c, ci, a: "=0",
         "fwd": lambda c, j, ci, a: f"={K.R(S_SW, c + str(cr('robrev', ci)))}"},
        {"key": "rev", "label": "总营收(RMB'B)", "fmt": K.N1, "bold": True,
         "hist": rev_hist,
         "fwd": lambda c, j, ci, a: f"={c}{a['licr']}+{c}{a['devr']}+{c}{a['robr']}"},
        {"key": "opm", "label": "经营利润率", "fmt": K.PCT,
         "hist": lambda c, ci, a: "=\"n.a.\"" if HC.index(c) < 2 else f"={[None,None,-1.469/0.743,-1.291/1.325,-0.642/2.413][HC.index(c)]:.4f}",
         "fwd": lambda c, j, ci, a: f"={K.R(S_SW, c + str(cr('opm', ci)))}"},
        {"key": "op", "label": "经营利润(RMB'B)", "fmt": K.N1,
         "hist": lambda c, ci, a: "=\"n.a.\"" if HC.index(c) < 2 else f"={OP[HC.index(c)]}",
         "fwd": lambda c, j, ci, a: f"={c}{a['rev']}*{c}{a['opm']}"},
        {"key": "ni", "label": "经调整净利(RMB'B)", "fmt": K.N1, "bold": True,
         "hist": lambda c, ci, a: "=\"n.a.\"" if NI_ADJ[HC.index(c)] is None else f"={NI_ADJ[HC.index(c)]}",
         "fwd": lambda c, j, ci, a: f"={c}{a['op']}*{K.R(S_SW, c + str(cr('netconv', ci)))}"},
        {"key": "evs", "label": "目标 EV/Sales(x)", "fmt": K.MX,
         "hist": lambda c, ci, a: "=\"—\"",
         "fwd": lambda c, j, ci, a: f"={K.R(S_MULT, c + str(ev_sales_rows[CASES[ci]]))}"},
        {"key": "ev", "label": "隐含 EV(RMB'B)", "fmt": K.N1,
         "hist": lambda c, ci, a: "=\"—\"",
         "fwd": lambda c, j, ci, a: f"={c}{a['rev']}*{c}{a['evs']}"},
        {"key": "eq", "label": "隐含权益(RMB'B)", "fmt": K.N1,
         "hist": lambda c, ci, a: "=\"—\"",
         "fwd": lambda c, j, ci, a: f"={c}{a['ev']}+{NETCASH_RMB_B}"},
        {"key": "px", "label": "隐含股价(HK$)", "fmt": K.PX, "bold": True, "out": True,
         "hist": lambda c, ci, a: f"={PX_NOW}" if HC.index(c) == 4 else "=\"n.a.\"",
         "fwd": lambda c, j, ci, a: f"={c}{a['eq']}*1000/{SHARES_M}*{HKD_PER_CNY:.4f}"},
        {"key": "up", "label": "较现价上行/下行", "fmt": K.PCT,
         "hist": lambda c, ci, a: "=\"—\"",
         "fwd": lambda c, j, ci, a: f"={c}{a['px']}/{PX_NOW}-1"},
        {"key": "ps", "label": "隐含 forward P/S(诊断)", "fmt": K.MX,
         "hist": lambda c, ci, a: "=\"—\"",
         "fwd": lambda c, j, ci, a: f"={c}{a['eq']}/{c}{a['rev']}"},
        {"key": "ipe", "label": "隐含 forward P/E(诊断，亏损年 N/M)", "fmt": K.MX,
         "hist": lambda c, ci, a: "=\"—\"",
         "fwd": lambda c, j, ci, a: f"=IF({c}{a['ni']}<=0,\"N/M\",{c}{a['px']}/({c}{a['ni']}*1000/{SHARES_M}*{HKD_PER_CNY:.4f}))"},
    ]
    cmp = K.write_comparison(wb.create_sheet(S_CMP), {
        "title": "估值对比 — Bear / Base / Bull 三案并排",
        "intro": ("三案从装车量、单车费、技术开发服务费、经营利润率、目标 EV/Sales 同一条链推导。主判断年取 2027E。"
                  "隐含价=(营收×EV/Sales+净现金)/股本×HKD/CNY。Base 不含 Robotaxi(期权，见 SOTP)；Bull 含。历史列填实际、2025 隐含=现价近似=回测参照。"),
        "case_names": CASES, "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
        "block_start": 22, "rows": rows,
        "summary": {
            "band": "2027E 三案摘要（主判断年）", "target_col": "H",
            "rows": [
                ("总营收(RMB'B)", "rev", K.N1, "License + 技术开发服务费 + Robotaxi。", False),
                ("经营利润(RMB'B)", "op", K.N1, "营收 × 经营利润率(费用杠杆)。", False),
                ("经调整净利(RMB'B)", "ni", K.N1, "经营利润 × 净利转换。", False),
                ("目标 EV/Sales", "evs", K.MX, "高增长换挡中段倍数。", False),
                ("隐含 EV(RMB'B)", "ev", K.N1, "营收 × EV/Sales。", False),
                ("隐含股价(HK$)", "px", K.PX, "EV/Sales 主线输出。", True),
                ("vs 现价", "up", K.PCT, "现价 reality check。", True),
            ],
            "mcap": {"label": "隐含市值(HK$'B)", "key": "px", "expr": f"*{SHARES_M}/1000", "note": "隐含股价 × 股本。"},
            "concl": ("Base 2027E 隐含 ≈ HK$295(≈现价) → 量产 License 业务已被 IPO 定价充分计入，市场未对其大幅折让或溢价。"
                      "Bull 靠 Robotaxi 2028 商业化 + 份额韧性 + 单车费上行给 +50% 上行；Bear 靠份额被压缩 + 单车费停滞给 -35%。"
                      "风险收益接近对称、Robotaxi 期权免费 → HOLD，等待 Robotaxi 里程碑或回调加仓。"),
        },
    })

    # ══ 13 综合判断仪表盘 ══
    K.write_dashboard(wb.create_sheet(S_DASH), {
        "title": "综合判断仪表盘 — 基本面拐点 / 估值错位 / 催化剂 / 情绪",
        "usage": "把模型压成投后跟踪语言：哪些指标验证 Base、哪些推向 Bear/Bull。B 列公式引模型输出。",
        "blocks": [
            {"title": "A. 基本面拐点", "rows": [
                ("毛利率/经营现金流", "71.6% / OCF -0.28B(2025)", "软件规模效应兑现，逼近现金平衡；经营利润 2027E 转正。", True),
                ("装机/份额", "累计破 100 万台；独立 Urban NOA 份额 64.5%", "物理锚底气；份额 normalize 速度是关键变量。"),
                ("2027E Base 经营利润", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Base']['op']))}", "fmt": K.N1, "fill": True}, "若不及该路径，Base 估值先降。"),
            ]},
            {"title": "B. 估值错位", "rows": [
                ("当前股价 HK$", {"inp": PX_NOW, "fmt": K.PX, "fill": True}, "reality check，不反向拟合。"),
                ("Base 2027E 隐含价 HK$", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Base']['px']))}", "fmt": K.PX, "fill": True}, "主判断输出。"),
                ("Base 2027E vs 现价", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Base']['up']))}", "fmt": K.PCT, "fill": True}, "≈0 = 量产业务已被充分定价。"),
                ("Bear 2027E 隐含价 HK$", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Bear']['px']))}", "fmt": K.PX}, "下行：份额压缩 + 单车费停滞。"),
                ("Bull 2027E 隐含价 HK$", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Bull']['px']))}", "fmt": K.PX}, "上行：Robotaxi 兑现 + 份额韧性。"),
            ]},
            {"title": "C. 催化剂", "rows": [
                ("Bull 触发", "Robotaxi 上海/Abu Dhabi/Munich 商业运营落地；新增大额 OEM 定点；单车费随 L3 上行获验证。", "先抬装车量/单车费/Robotaxi，再抬 EV/Sales。"),
                ("Bear 触发", "华为/地平线拿下核心 OEM 定点；单车费环比下滑;新增定点持续 <2024 水平。", "先砍装车量/份额，再压 EV/Sales 到 6x 以下。"),
                ("供给冲击", "2027-01 基石(占发行 49.95%)6 个月锁定到期解禁。", "中期供给压力，风险章标注。"),
            ]},
            {"title": "D. 综合判断", "rows": [
                ("一句话结论", "HOLD：量产 License 业务撑起现价、Robotaxi 是未定价免费期权。等 Robotaxi 里程碑或回调至 Bear 区间(HK$~190-220)加仓。", "上行看 Robotaxi + 份额韧性；下行看份额被压缩。", True),
            ]},
        ],
        "final": {"band": "最终判断",
                  "text": ("Momenta 的 alpha 问题：市场为一家 2025 仍亏 RMB3.5B(会计口径)、Robotaxi 仅 31 台在运的公司付了 US$8.9B。"
                           "本模型把公司拆成两块看——量产 License 业务(可证伪、有装机数据支撑)在 Base 下 2027E 隐含 ≈ 现价 HK$295，说明市场已充分(未过度)定价这块；"
                           "Robotaxi(2028 商业化、中国 2030 TAM US$38bn)是叠加在上面的免费期权，只在 Bull 计入。软件规模效应(毛利 71.6%、经营现金流出快速收窄)是真拐点。"
                           "最大未证伪下行=独立 Urban NOA 份额 64.5% 被华为/地平线压缩(License 是一次性费、非订阅，份额一丢收入即回落)。评级 HOLD，回调或 Robotaxi 兑现加仓。")},
        "tracking": {
            "intro": "投后跟踪按季报/月度交付数据滚动更新。",
            "rows": [
                ("__band__", "一、物理锚"),
                ("年度装车量 / 累计装机", "累计破 100 万台(2026 中)", "关键敏感项：NOA 渗透 × Momenta 份额", "CIC / 公司月度交付", "装车量增速掉档 → 转 Bear"),
                ("独立 Urban NOA 份额", "64.5%(2026-02 前 12 月)", "关键敏感项：数据积累正反馈/交付壁垒能否守住份额", "CIC 季度", "份额跌破 50% → 砍装车量路径"),
                ("__band__", "二、单车经济"),
                ("单车 License 费(反推)", "~RMB1,940(2025)", "关键敏感项：L3/Urban mix 抬价 vs 竞争压价", "License 收入÷装车量", "环比下滑 → 下调 fee 路径"),
                ("__band__", "三、盈利拐点"),
                ("经营利润 / 现金流", "经营亏损 -0.64B / OCF -0.28B(2025)", "关键敏感项：R&D 增速 < 收入增速", "季报", "R&D 增速反超收入 → 推迟转正"),
                ("__band__", "四、Robotaxi 期权"),
                ("Robotaxi 商业运营", "上海获无人化许可(2026-01)；仅 31 台", "关键敏感项：2028 商业化能否兑现", "公司/监管公告", "上海/Abu Dhabi 车队起量 → 计入 Base"),
                ("__band__", "五、供给"),
                ("基石解禁", "占发行 49.95%，6 个月锁定", "关键敏感项：2027-01 解禁供给冲击", "招股书锁定期", "解禁前后波动 → 交易层面留意"),
            ],
        },
    })

    # input.json
    payload = {
        "ticker": "06880.HK", "name": "Momenta Global Limited", "built_at": "2026-07-10",
        "currency_financials": "RMB", "currency_price": "HKD",
        "current_price_hkd": PX_NOW, "market_cap_hkd_b": MKTCAP_HKD_B,
        "market_cap_usd_b": round(MKTCAP_HKD_B / FX_HKD_USD, 2),
        "shares_m": SHARES_M, "net_cash_rmb_b": NETCASH_RMB_B,
        "fx": {"cny_usd": FX_CNY_USD, "hkd_usd": FX_HKD_USD, "hkd_per_cny": round(HKD_PER_CNY, 4)},
        "rating": "HOLD", "target_price_hkd": None,  # 由模型 Base 2027E 隐含价决定，见 xlsx
        "method": "EV/Sales 主镜头(亏损高增长)+ DCF 交叉 + SOTP(量产 License vs Robotaxi)；物理锚=年度装车量×单车License费+技术开发服务费。",
        "anchor": {"type": "auto-license (装车量×单车费)",
                   "equip_k_hist": {"2023": 12, "2024": 150, "2025": 500},
                   "fee_rmb_hist": {"2023": 1933, "2024": 1947, "2025": 1936},
                   "note": "装车量为估算(校准至 License 收入)，单车费=License收入÷装车量反推；招股书未直接披露。"},
        "historical_financials_rmb_b": {
            "revenue": {"2023": 0.743, "2024": 1.325, "2025": 2.413},
            "revenue_segments_2025": {"license": 0.968, "dev_service": 1.445, "robotaxi": 0.0},
            "gross_margin": {"2023": 0.175, "2024": 0.490, "2025": 0.716},
            "operating_loss": {"2023": -1.469, "2024": -1.291, "2025": -0.642},
            "adj_net_loss": {"2023": -1.093, "2024": -0.959, "2025": -0.303},
            "reported_net_loss": {"2023": -2.570, "2024": -3.206, "2025": -3.458},
            "ocf": {"2023": -1.069, "2024": -0.836, "2025": -0.281},
            "net_assets_ex_preferred_2025": 11.919,
            "liquidity_2025": 10.899,
        },
        "scenario": {"equip_k": EQUIP, "fee_rmb": FEE, "dev_rev_rmb_b": DEVREV, "robotaxi_rmb_b": ROBREV,
                     "gross_margin": GMF, "operating_margin": OPMF, "net_conversion": NETCONV,
                     "fcf_conversion": FCFCONV, "target_ev_sales": EVS},
        "consensus": {"coverage": "none (post-IPO quiet period)", "target_mean": None,
                      "comps": {"WeRide": "UBS forecast return +24.5%", "Horizon 9660.HK": "EV/Sales ~12x", "Palantir": "EV/Sales ~40x"}},
        "sources": {
            "official": ["港交所 hkexnews 招股书(2026-06-29 聆讯后资料集/全球发售) BDO 审计 2023-2025",
                         "https://www1.hkexnews.hk/listedco/listconews/sehk/2026/0629/2026062900057.pdf"],
            "market": "tvremix/TradingView 行情 2026-07-10; 现价 HK$296.2",
            "industry": "CIC(灼识咨询) 招股书 Industry Overview: 中国量产 NOA SAM 2030E US$38.6bn/CAGR 33.8%",
        },
    }
    os.makedirs(VAULT, exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    K.finalize(wb)
    wb.save(OUT_XLSX)
    print("saved:", OUT_XLSX)
    print("sheets:", wb.sheetnames)
    return cmp


def write_momenta_valuation(ws, d):
    """EV/Sales 主线逐年隐含价(HK$) + DCF 交叉 + SOTP。亏损 → P/E·P/B 退诊断。"""
    s_fund, fr = d["s_fund"], d["fr"]
    s_sw = d["s_sw"]; swact = d["swact"]
    s_mult, evs_rows = d["s_mult"], d["evs_rows"]
    K.hdr(ws, 1, "情景估值 — 当前案逐年 EV/Sales 主线(HK$) + DCF 交叉 + SOTP", 8)
    K.lab(ws, "L1", "当前情景→", note=True); K.fml(ws, "M1", f"={K.R(s_sw, d['sw_cell'])}", K.N0, link=True); ws["M1"].fill = K.CUR
    r = K.mtext(ws, 2, ("本页随情景切换 B2 变化。主镜头 EV/Sales：EV=前瞻营收×目标倍数，+净现金→权益→/股本→RMB每股→×HKD/CNY→隐含 HK$ 股价。"
                        "DCF 交叉(现金盈利折现)。SOTP 把量产 License 业务与 Robotaxi 期权分开。P/E·P/B 亏损失效退诊断。"), "K", 2)
    K.lab(ws, f"A{r}", "(RMB'B / HK$每股 / 倍数)", b=True)
    for col, y in zip(ALLC, ALLY):
        ws[f"{col}{r}"] = y; ws[f"{col}{r}"].font = K.BF; ws[f"{col}{r}"].fill = K.CH
    r += 1
    revc = lambda c: K.R(s_fund, c + str(fr["rev"]))
    fcfc = lambda c: K.R(s_fund, c + str(fr["fcf"]))

    K.band(ws, r, "EV/Sales 主线：EV = 前瞻营收 × 目标倍数 + 净现金 → 隐含 HK$ 股价", 8); r += 1
    evs_row = r; K.lab(ws, f"A{r}", "目标 EV/Sales(当前案)")
    for c in HC:
        K.lab(ws, f"{c}{r}", "—", note=True)
    # 当前案 EV/Sales：优先引『情景切换』已算好的当前案行(使其非孤儿)；否则嵌套 IF 兜底
    sw_cell = d["sw_cell"]
    evs_act = swact.get("evs")
    for c in FCf:
        if evs_act:
            K.fml(ws, f"{c}{r}", f"={K.R(s_sw, c + str(evs_act))}", K.MX, link=True)
        else:
            b, ba, bu = evs_rows["Bear"], evs_rows["Base"], evs_rows["Bull"]
            f = (f"=IF({K.R(s_sw, sw_cell)}=1,{K.R(s_mult, c+str(b))},"
                 f"IF({K.R(s_sw, sw_cell)}=2,{K.R(s_mult, c+str(ba))},{K.R(s_mult, c+str(bu))}))")
            K.fml(ws, f"{c}{r}", f, K.MX, link=True)
    r += 1
    ev_row = r; K.lab(ws, f"A{r}", "隐含 EV (RMB'B)")
    for c in FCf:
        K.fml(ws, f"{c}{r}", f"={revc(c)}*{c}{evs_row}", K.N1, link=True)
    r += 1
    eq_row = r; K.lab(ws, f"A{r}", f"隐含权益 (RMB'B) = EV + 净现金 {NETCASH_RMB_B}")
    for c in FCf:
        K.fml(ws, f"{c}{r}", f"={c}{ev_row}+{NETCASH_RMB_B}", K.N1)
    r += 1
    px_row = r; K.lab(ws, f"A{r}", "隐含股价 EV/Sales主线 (HK$)", b=True); ws[f"A{r}"].fill = K.OUT
    for c in FCf:
        K.fml(ws, f"{c}{r}", f"={c}{eq_row}*1000/{SHARES_M}*{HKD_PER_CNY:.4f}", K.PX)
    r += 1
    up_row = r; K.lab(ws, f"A{r}", "较现价上行/下行")
    for c in FCf:
        K.fml(ws, f"{c}{r}", f"={c}{px_row}/{PX_NOW}-1", K.PCT)
    r += 1
    K.lab(ws, f"A{r}", "隐含 forward P/S(诊断)")
    for c in FCf:
        K.fml(ws, f"{c}{r}", f"={c}{eq_row}/{revc(c)}", K.MX)
    r += 2

    # DCF 交叉
    K.band(ws, r, "镜头二 DCF（现金盈利折现，WACC 13% / 永续 g 4%）", 8); r += 1
    K.lab(ws, f"A{r}", "2030E FCF (RMB'B)")
    K.fml(ws, f"B{r}", f"={K.R(s_fund, 'K' + str(fr['fcf']))}", K.N1, link=True); dcf_fcf = r; r += 1
    K.lab(ws, f"A{r}", "WACC / 永续增长 g")
    K.inp(ws, f"B{r}", 0.13, None, K.PCT); K.inp(ws, f"C{r}", 0.04, None, K.PCT); dcf_w = r; r += 1
    K.lab(ws, f"A{r}", "永续法终值 EV (RMB'B) = FCF×(1+g)/(WACC−g)")
    K.fml(ws, f"B{r}", f"=B{dcf_fcf}*(1+C{dcf_w})/(B{dcf_w}-C{dcf_w})", K.N1); dcf_tv = r; r += 1
    K.lab(ws, f"A{r}", "折现到今(÷(1+WACC)^5)+ 净现金")
    K.fml(ws, f"B{r}", f"=B{dcf_tv}/(1+B{dcf_w})^5+{NETCASH_RMB_B}", K.N1); dcf_eq = r; r += 1
    K.lab(ws, f"A{r}", "DCF 隐含股价 (HK$)", b=True); ws[f"A{r}"].fill = K.OUT
    K.fml(ws, f"B{r}", f"=B{dcf_eq}*1000/{SHARES_M}*{HKD_PER_CNY:.4f}", K.PX); dcf_px = r; r += 1
    K.lab(ws, f"A{r}", "DCF vs 现价")
    K.fml(ws, f"B{r}", f"=B{dcf_px}/{PX_NOW}-1", K.PCT); r += 2

    # SOTP
    K.band(ws, r, "镜头三 SOTP（量产 License 业务 + Robotaxi 期权，分开估）", 8); r += 1
    K.lab(ws, f"A{r}", "量产业务 EV (2027E营收×9x, RMB'B)")
    K.fml(ws, f"B{r}", f"=({K.R(s_fund, 'H'+str(fr['rev']))})*9", K.N1); sotp_core = r; r += 1
    K.lab(ws, f"A{r}", "Robotaxi 期权价值 (RMB'B, 风险调整)")
    K.inp(ws, f"B{r}", 8.0, None, K.N1); sotp_rob = r; r += 1
    K.logic(ws, f"N{sotp_rob}", "中国+全球 Robotaxi 2030 TAM ~US$120bn；Momenta 特斯拉式量产数据派，假设成熟期取 3-5% 份额、软件经济、按 30% 概率与折现风险调整后 ≈RMB8B(≈US$1.2bn)。Base 不计入(期权)，此处仅 SOTP 展示上行。")
    K.lab(ws, f"A{r}", "SOTP 权益 (RMB'B) = 量产EV + Robotaxi + 净现金")
    K.fml(ws, f"B{r}", f"=B{sotp_core}+B{sotp_rob}+{NETCASH_RMB_B}", K.N1); sotp_eq = r; r += 1
    K.lab(ws, f"A{r}", "SOTP 隐含股价 (HK$)", b=True); ws[f"A{r}"].fill = K.OUT
    K.fml(ws, f"B{r}", f"=B{sotp_eq}*1000/{SHARES_M}*{HKD_PER_CNY:.4f}", K.PX); sotp_px = r; r += 1
    K.lab(ws, f"A{r}", "SOTP vs 现价")
    K.fml(ws, f"B{r}", f"=B{sotp_px}/{PX_NOW}-1", K.PCT); r += 2

    K.band(ws, r, "三镜头三角 + 方法", 8); r += 1
    K.mtext(ws, r, ("三镜头(EV/Sales 主线 / DCF / SOTP)交叉。EV/Sales 是主锚(亏损期出实数)；DCF 用 2030E 现金盈利折现做长久期 sanity；"
                    "SOTP 把量产 License 业务(≈现价)与 Robotaxi 期权(上行)分开——回答『市场为 Robotaxi 付了多少』。"
                    "Base 三镜头收敛在现价附近 → 量产业务充分定价、Robotaxi 免费期权 → HOLD。P/E·P/B 不做(亏损+优先股负账面)。"), "K", 4)
    K.set_widths(ws, 30, ALLC, 10, logic_col="N", logic_width=58)
    return {"evs": evs_row, "px": px_row, "dcf_px": dcf_px, "sotp_px": sotp_px}


if __name__ == "__main__":
    build()
