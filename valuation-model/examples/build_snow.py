# -*- coding: utf-8 -*-
"""
build_snow.py — Snowflake (SNOW) 估值模型, 用 build_kit 引擎 + 手搓估值三页(EV/Sales 每股)。

标的性质: 消费型数据云 SaaS。GAAP 亏损(几乎全来自 SBC)、non-GAAP + adj FCF 已转正。
物理锚: 客户数 × ARR/客户 × NRR —— 产品收入 = 客户数 × 每客户产品收入;
  客户数增速(logo) + ARR/客户增速(NRR 存量扩张) 两杠杆决定产品收入。
估值镜头(≥2, 三角):
  ① EV/Sales(前瞻总营收) —— 主镜头(市场对高增长数据 SaaS 的实际定价方式, 当前 ~15x forward)
  ② EV/FCF(adj FCF 已正) —— 第二镜头(实数, 不是 n.m.), 交叉验证"用 FCF 看贵不贵"
  DCF 作 sanity(备注, 不单列页)。
财年: 1/31 结账。历史 FY2022A-FY2026A, 前瞻 FY2027E-FY2031E。美元单一币种(fx=1)。

跑: cd examples && PYTHONUTF8=1 python build_snow.py
校验: python ../scripts/recalc.py <out> && python ../scripts/validate_valuation.py <out>
"""
import os
from openpyxl import Workbook
import build_kit as K

# ════════════ 0. 全局轴 ════════════
ALLC = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
ALLY = ["FY22A", "FY23A", "FY24A", "FY25A", "FY26A", "FY27E", "FY28E", "FY29E", "FY30E", "FY31E"]
HC, HY = ["B", "C", "D", "E", "F"], ["FY22", "FY23", "FY24", "FY25", "FY26"]
FC = ["F", "G", "H", "I", "J", "K"]     # 含基年 F=FY26A
FCf = FC[1:]                             # 纯前瞻 FY27E-FY31E (G H I J K)
FWY = ["FY27E", "FY28E", "FY29E", "FY30E", "FY31E"]
FX = 1.0
CASES = ["Bear", "Base", "Bull"]

PRICE = 267.49          # 2026-07-10 实时
SHARES_M = 347          # 当前流通股(市值 267.49×347/1000 ≈ $92.8B, 对齐实时市值)
NET_CASH = 2.5          # 现金+投资 $4.79B − 零息可转债 $2.3B ≈ 净现金 $2.5B

# 逐年三案数据(列顺序 = FY27E,FY28E,FY29E,FY30E,FY31E)──────────────
# 客户数增速(logo 增长)
G_CUST = {"Bear": [0.12, 0.10, 0.08, 0.07, 0.06], "Base": [0.15, 0.13, 0.11, 0.10, 0.09], "Bull": [0.18, 0.16, 0.14, 0.12, 0.10]}
# ARR/客户增速(NRR 存量扩张 + 大客户 mix)
G_ARR  = {"Bear": [0.10, 0.08, 0.07, 0.06, 0.05], "Base": [0.13, 0.11, 0.09, 0.08, 0.07], "Bull": [0.16, 0.14, 0.12, 0.10, 0.09]}
# 专业服务增速(小额, 低敏感)
G_SVC  = {"Bear": [0.10, 0.08, 0.06, 0.05, 0.04], "Base": [0.15, 0.12, 0.10, 0.08, 0.06], "Bull": [0.18, 0.15, 0.12, 0.10, 0.08]}
# adj FCF margin(占总营收)
FCFM   = {"Bear": [0.22, 0.22, 0.23, 0.24, 0.25], "Base": [0.24, 0.25, 0.27, 0.28, 0.30], "Bull": [0.25, 0.27, 0.30, 0.32, 0.34]}
# 目标 EV/Sales(总营收)
EVS    = {"Bear": [9.0, 7.5, 6.5, 6.0, 5.5], "Base": [14.0, 12.0, 10.5, 9.5, 8.5], "Bull": [17.0, 15.0, 13.5, 12.0, 11.0]}
# 目标 EV/FCF
EVF    = {"Bear": [30, 28, 25, 23, 20], "Base": [45, 40, 36, 32, 28], "Bull": [55, 50, 45, 40, 35]}
WEIGHTS = {"Bear": 0.30, "Base": 0.45, "Bull": 0.25}

S_COVER, S_HIST, S_PX, S_CONS = "封面", "历史财务与估值", "股价走势", "卖方研报共识"
S_HMULT, S_MULT, S_SW = "历史估值倍数", "估值倍数假设", "情景切换"
S_ANCHOR, S_SEG, S_FUND = "物理锚", "分部测算", "利润与收入假设"
S_VAL, S_CMP, S_DASH = "情景估值", "估值对比", "综合判断仪表盘"

# 月度股价(3 年, 来自 rdata /price)
import json
_repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_input_dir = os.environ.get("VALUATION_INPUT_DIR", os.path.join(_repo_root, "out"))
_pxfile = os.path.join(_input_dir, "snow_monthly.json")
if os.path.exists(_pxfile):
    with open(_pxfile, "r", encoding="utf-8") as _f:
        MONTHLY = [(x[0], x[1]) for x in json.load(_f)]
else:
    MONTHLY = None

wb = Workbook(); wb.remove(wb["Sheet"])

# ════════════ 1. 封面 ════════════
K.write_cover(wb.create_sheet(S_COVER), {
    "title": "Snowflake (SNOW) 估值模型 — 物理锚(客户×ARR×NRR)→ EV/Sales + EV/FCF 双镜头",
    "span": 6,
    "meta": [
        ("报告日期", "2026-07-10"),
        ("数据截止", "FY2026 10-K + Q1 FY2027 10-Q(2026-05-29) + 2026-07-10 实时股价"),
        ("现价", f"${PRICE:.2f}(2026-07-10; 4月 $153 → 3 个月 +75%)"),
        ("市值 / 股本", f"约 $92.7B  |  股本 ~347M  |  净现金 ~$2.5B → EV ~$90B"),
        ("时效声明", "SNOW 消费型 SaaS + AI(Cortex)变现节奏变化快, 建议每季财报(下次 Q2 FY27 约 2026-08 底)后更新。"),
        ("方法一句话", "物理锚(客户数×ARR/客户×NRR)→ 产品收入 + 专业服务 → 总营收/FCF → "
                    "EV/Sales(主, 前瞻营收)+ EV/FCF(第二, FCF 已正)双镜头逐年 → ÷股本 = 隐含价。三案概率加权得 12M 目标价。"),
    ],
    "takeaways": [
        ("① 当下估值位置", f"EV ~$90B, 对应 FY27E 前瞻 EV/Sales ~14.8x、EV/FCF ~61x — 营收倍数在高增长数据 SaaS 合理区, FCF 倍数偏贵(FCF margin 仍在爬坡)。"),
        ("② 核心引擎", "产品收入 = 客户数(13.9K)×ARR/客户($336K)×NRR(126%); Q1 FY27 产品收入 +34% 重新加速, NRR 触底回升(125→126%)。"),
        ("③ AI 变现判断", "Cortex Code(CoCo)9 周 50% 客户使用、早期采用者消费 +11% — AI 是消费飞轮催化, 尚未成独立收入线, 是 Bull 期权。"),
        ("④ 三情景 12M 目标价", "Bear/Base/Bull 从 客户增速 + ARR 扩张 + FCF margin + 目标倍数 四杠杆翻档沿同一条链算出, 见『估值对比』。"),
        ("⑤ 主要风险", "Databricks 增速约 SNOW 2.4x 且 ARR 已反超 + Iceberg 开放格式削弱数据锁定 + 消费型 FinOps 降本 + GAAP 仍亏损。"),
    ],
})

# ════════════ 2. 历史财务与估值 ════════════
ha = K.write_history(wb.create_sheet(S_HIST), {
    "title": "Snowflake 历史财务与估值 ($B) — FY2022-FY2026A + 当下(Q1 FY27)",
    "hist_cols": HC, "hist_years": HY,
    "fx_hist": [1, 1, 1, 1, 1], "fx_now": 1.0,
    "vals_in_usd": True, "ps_scale": 1000, "mcap_div": 1000,
    "unit_label": "($B)", "mcap_label": "市值($B)", "fx_label": "FX (USD 单币种=1)",
    "cur_label": "当下(2026-07)",
    "segments": [
        ("产品收入 (Product)", [1.140, 1.939, 2.667, 3.462, 4.472], True),
        ("专业服务及其他", [0.079, 0.127, 0.140, 0.164, 0.212], True),
    ],
    "total_now": 5.55,   # Q1FY27 总营收 $1.391B 年化约 $5.55B(仅锚, 非精确 TTM)
    "gm_pct": [0.624, 0.652, 0.680, 0.665, 0.672], "gm_now": 0.666,
    "ni": [-0.680, -0.797, -0.836, -1.286, -1.332], "ni_now": -0.296,
    "eq": [5.049, 5.456, 5.180, 3.000, 1.924], "eq_now": 1.90,
    "shares": [300.3, 318.7, 328.0, 332.7, 337.5], "shares_now": SHARES_M,
    "px_end": [290, 148, 191, 158, 217],
    "px_now": PRICE,
    "px_avg": [300, 150, 180, 150, 185],
    "band_note": "SNOW 历年 GAAP 亏损 → P/E 恒 N/M; 估值主镜头看 EV/Sales(下方专段)。P/B 意义有限(轻资产 SaaS)。",
    "notes": [
        ("产品收入 (Product)", "按消费用量(存储+计算 credit)计费, 占总收入约 95%; FY26 $4.47B(+29%), Q1FY27 +34% 重新加速。"),
        ("专业服务及其他", "咨询/培训/实施, 占约 5%, 低毛利, 战略上服务产品消费。"),
        ("HGMP", "GAAP 总毛利率: FY24 68% 峰值后 FY25-26 回落至 67%(AI 基础设施+云成本); non-GAAP 产品毛利率约 75-76%。"),
        ("HNI", "GAAP 净亏损: FY26 -$1.33B, 几乎全部来自 SBC($1.60B); adj FCF 同期 +$1.19B — 现金流与账面利润背离。"),
        ("HEQ", "股东权益 FY22 $5.0B → FY26 $1.9B: 累计亏损 + $4.5B 回购授权消耗所致。"),
        ("HSH", "GAAP 加权摊薄股数(反摊薄, 百万); 2025-07 注销 Class B 简化为单层股权。当下列=当前流通 ~347M。"),
        ("HPX", "年末=各财年 1/31 前后股价(近似); FY26A(2026-01)约 $217, 当下(2026-07)$267.49。"),
    ],
})
PROD_HROW = ha["seg_rows"]["产品收入 (Product)"]
SVC_HROW = ha["seg_rows"]["专业服务及其他"]

# 历史财务尾部补 EV/Sales 实数底座(亏损标的主镜头必须有实数底座)──────
wh = wb[S_HIST]
NOTE_H = "H"
er = wh.max_row + 2
K.band(wh, er, "EV / Sales 底座(亏损标的主镜头; P/E 亏损年 N/M 不可用)", len(HC) + 3); er += 1
R_HMCAP = er
K.lab(wh, f"A{er}", "市值($B, 当下)", b=True)
K.fml(wh, f"G{er}", f"=G{ha['HPX']}*G{ha['HSH']}/1000", K.N0); wh[f"G{er}"].fill = K.CUR
K.logic(wh, f"{NOTE_H}{er}", "= 现价 $267.49 × 347M 股 ≈ $92.8B。")
er += 1
R_HNC = er
K.lab(wh, f"A{er}", "净现金($B)")
K.inp(wh, f"G{er}", NET_CASH, None, K.N1)
K.logic(wh, f"{NOTE_H}{er}", "= 现金+短投+长投 $4.79B − 零息可转债 $2.3B ≈ $2.5B。")
er += 1
R_HEV = er
K.lab(wh, f"A{er}", "企业价值 EV($B, 当下)", b=True); wh[f"A{er}"].fill = K.OUT
K.fml(wh, f"G{er}", f"=G{R_HMCAP}-G{R_HNC}", K.N0)
K.logic(wh, f"{NOTE_H}{er}", "= 市值 − 净现金 ≈ $90B。")
er += 1
R_HEVS = er
K.lab(wh, f"A{er}", "EV / Sales(前瞻, EV÷FY27E营收)", b=True); wh[f"A{er}"].fill = K.OUT
for col in HC:
    K.lab(wh, f"{col}{er}", "n.m.", note=True)
K.fml(wh, f"G{er}", f"=G{R_HEV}/6.06", K.MX); wh[f"G{er}"].fill = K.CUR
K.logic(wh, f"{NOTE_H}{er}", "当下前瞻 EV/Sales = EV ÷ FY27E 总营收(~$6.06B) ≈ 14.8x — 高增长数据 SaaS 合理区(Databricks 私募隐含 ~19x, MDB ~8x)。"
                            "历史 EV/Sales 曾 2021-22 泡沫期 30-60x, 2024 低点约 10x, 现 ~15x 处历史中低位。")

# ════════════ 3. 股价走势 ════════════
def phase_fn(ym):
    if ym <= "2024-09": return "① 绝望期(Slootman 卸任/NRR 下滑)"
    if ym <= "2025-12": return "② AI 转型修复"
    if ym <= "2026-04": return "③ 软件杀估值回落"
    return "④ Q1 加速再重估"
if MONTHLY:
    px = K.write_price_chart(wb.create_sheet(S_PX), MONTHLY, {
        "fn": phase_fn,
        "rows": [("① 绝望期", "2024-09 触底 $108: Slootman 卸任+NRR 一路下滑+增速换挡。"),
                 ("② AI 转型修复", "2025 全年 $149→$230: Ramaswamy AI 转型叙事 + 产品收入重新加速。"),
                 ("③ 软件杀估值回落", "2026-04 回落 $153: 软件板块/宏观杀估值。"),
                 ("④ Q1 加速再重估", "Q1 FY27 产品 +34% + 指引上调 → 3 个月飙至峰值 $280, 现 $267。")],
    }, title="Snowflake 月度股价 ($)")
else:
    px = K.write_price_chart(wb.create_sheet(S_PX), [("2026-01", 217), ("2026-07", PRICE)], {
        "rows": [("近况", "FY26A 约 $217 → 2026-07 $267.49。")]}, title="Snowflake 股价")

# ════════════ 4. 卖方研报共识 ════════════
K.write_consensus(wb.create_sheet(S_CONS), {
    "title": "卖方研报共识 — 一致目标价 ~$292(48 分析师), Strong Buy",
    "overview": ("Q1 FY27(2026-05-27)重估分水岭: 产品收入 +34% 加速、指引上调 +27%→+31%, 财报次日 +34-37%。"
                 "多数投行上调至 $280-300(MS/BofA $300, JPM/KeyBanc/RBC $284-285); Bernstein $250(Market-Perform)、UBS $210(滞后)偏空两端。"
                 "我们的分歧: 3 个月 +75% 后, EV/Sales ~15x 已把'AI 赢家'叙事定价充分, 双镜头看隐含价约在现价附近。"),
    "assumptions": [
        ("产品收入增速\n(FY27E)", "街上共识 +30-31%(指引 $5.84B/+31%)。", "Q1 加速是否可持续 vs 易比基数一次性。",
         "Base 取客户 +15% × ARR +13% = 产品 +30%(对齐指引)。"),
        ("NRR", "Q1FY27 126%, 环比 +1pt 触底回升。", "回升 1pt 是否确立趋势 vs FY27 起无易比基数。",
         "Base ARR/客户增速 13%→7% 递减(NRR 长期下行), 反映成熟客户占比升。"),
        ("adj FCF margin", "指引 FY27 23%, 街上 FY28 ~24-25%。", "AI 推理成本 vs 规模效应。",
         "Base 24%→30%(FY31E), 规模效应释放。"),
        ("目标 EV/Sales", "卖方隐含 forward EV/Sales ~14-15x(BofA 从 10.3x 上调至 14.7x)。", "给高增长数据 SaaS 多少溢价。",
         "Base FY27 14x, 随增速换挡 normalize 到 8.5x(FY31E)。"),
    ],
    "divergences": [
        "① Cortex/AI 变现: 是消费飞轮催化(已现)还是独立收入线(未现)——决定 Bull 概率。",
        "② vs Databricks 份额: Databricks ARR $6.9B(+80%)已反超 SNOW ~$5.6B(+30%), 份额天平向对手。",
        "③ 估值: 3 个月 +75% 后, EV/Sales 15x 是合理还是已 price-in 完美执行。",
    ],
    "stances": [
        "Morgan Stanley(增持, TP $300): 软件业少数 AI 赢家, AI 变现进更高档。",
        "BofA(买入, TP $300): AI Data Cloud 受益者被验证, CoCo 消费飞轮。",
        "JPMorgan(增持, TP $285): CoCo 是 FY27 指引上调最大单一驱动。",
        "Bernstein(Market-Perform, TP $250): 认可动能, 担心竞争加剧 + AI 变现落地。",
        "UBS(买入, TP $210, 4/20 滞后): AI 颠覆风险可控但逐项排查。",
    ],
})

# ════════════ 5. 历史估值倍数(EV/Sales 带 + 同业, 手搓)════════════
wsm = wb.create_sheet(S_HMULT)
K.hdr(wsm, 1, "历史估值倍数 — EV/Sales 带(SNOW 亏损, 主镜头非 P/E)+ 同业对照", 8)
hr = K.mtext(wsm, 2, ("SNOW 历年 GAAP 亏损 → 无 P/E 带, 主镜头看 EV/Sales。本页: "
                      "① 现价对应的当下前瞻 EV/Sales / EV/FCF(市场现在给多少); "
                      "② 同业 EV/Sales 光谱(数据 SaaS / 私募 Databricks), 看现价倍数落在什么位置。"
                      "估值锚逐年(FY27E-FY31E)EV/Sales + EV/FCF(后页)。"), "H", 3)
for col, w in zip("ABCDEFGH", [24, 13, 13, 13, 13, 11, 11, 11]):
    wsm.column_dimensions[col].width = w
wsm.column_dimensions["H"].width = 56

K.band(wsm, hr, "① SNOW 当下定价倍数(现价 $267.49 → EV ~$90B)", 8); hr += 1
for col, h in zip(["A", "B", "C", "D"], ["项目", "数值", "口径", "说明"]):
    wsm[f"{col}{hr}"] = h; wsm[f"{col}{hr}"].font = K.BF; wsm[f"{col}{hr}"].fill = K.CH
wsm.merge_cells(f"D{hr}:H{hr}"); hr += 1
R_MCAP = hr
K.lab(wsm, f"A{hr}", "市值($B)"); K.inp(wsm, f"B{hr}", 92.8, None, K.N0)
K.lab(wsm, f"C{hr}", "现价×股本", note=True)
wsm.merge_cells(f"D{hr}:H{hr}"); K.logic(wsm, f"D{hr}", "= $267.49 × 347M 股。"); hr += 1
R_NC2 = hr
K.lab(wsm, f"A{hr}", "净现金($B)"); K.inp(wsm, f"B{hr}", NET_CASH, None, K.N1)
K.lab(wsm, f"C{hr}", "现金+投资−可转债", note=True)
wsm.merge_cells(f"D{hr}:H{hr}"); K.logic(wsm, f"D{hr}", "= $4.79B − $2.3B 零息可转债。"); hr += 1
R_EV2 = hr
K.lab(wsm, f"A{hr}", "企业价值 EV($B)", b=True); wsm[f"A{hr}"].fill = K.OUT
K.fml(wsm, f"B{hr}", f"=B{R_MCAP}-B{R_NC2}", K.N0)
K.lab(wsm, f"C{hr}", "市值−净现金", note=True)
wsm.merge_cells(f"D{hr}:H{hr}"); K.logic(wsm, f"D{hr}", "= 市值 − 净现金 ≈ $90B。"); hr += 1
K.lab(wsm, f"A{hr}", "EV/Sales(前瞻 FY27E)", b=True)
K.fml(wsm, f"B{hr}", f"=B{R_EV2}/6.06", K.MX)
K.lab(wsm, f"C{hr}", "EV÷FY27E营收", note=True)
wsm.merge_cells(f"D{hr}:H{hr}"); K.logic(wsm, f"D{hr}", "≈ 14.8x — 高增长数据 SaaS 合理区。历史: 2021-22 泡沫 30-60x, 2024 低 ~10x, 现中低位。"); hr += 1
K.lab(wsm, f"A{hr}", "EV/FCF(前瞻 FY27E)", b=True)
K.fml(wsm, f"B{hr}", f"=B{R_EV2}/1.45", K.MX)
K.lab(wsm, f"C{hr}", "EV÷FY27E adj FCF", note=True)
wsm.merge_cells(f"D{hr}:H{hr}"); K.logic(wsm, f"D{hr}", "≈ 62x — FCF 倍数偏贵(FCF margin 24% 仍在爬坡, 随规模化 normalize)。"); hr += 1
hr += 1

K.band(wsm, hr, "② 同业 EV/Sales 对照(给『估值倍数假设』定锚)", 8); hr += 1
for col, h in zip(["A", "B", "C", "D"], ["可比公司/板块", "EV/NTM Sales", "增速", "业务特征 / 取数"]):
    wsm[f"{col}{hr}"] = h; wsm[f"{col}{hr}"].font = K.BF; wsm[f"{col}{hr}"].fill = K.CH
wsm.merge_cells(f"D{hr}:H{hr}"); hr += 1
peers = [
    ("Snowflake (现价)", 14.8, "+30%", "本标的; 消费型数据云, 产品收入 +30%。"),
    ("Databricks (私募)", 19.0, "+80%", "最直接对手, Series L $134B 隐含 ~19x run-rate; 增速约 SNOW 2.4x。"),
    ("Palantir (PLTR)", 36.0, "+85%", "AI 高成长溢价上沿, 已盈利。"),
    ("Datadog (DDOG)", 17.0, "+32%", "可观测性 SaaS, 增速/倍数与 SNOW 接近。"),
    ("MongoDB (MDB)", 8.0, "+25%", "文档数据库, 增速调整后与 SNOW parity, 但 GAAP 亏损更小。"),
    ("成熟 SaaS 中位", 6.0, "+15%", "低增长 SaaS forward EV/Sales 中位(去溢价终点参照)。"),
]
for name, evs, g, note in peers:
    wsm[f"A{hr}"] = name; wsm[f"A{hr}"].font = K.BF
    K.inp(wsm, f"B{hr}", evs, None, K.MX)
    K.lab(wsm, f"C{hr}", g, note=True)
    wsm.merge_cells(f"D{hr}:H{hr}"); K.logic(wsm, f"D{hr}", note); hr += 1
hr += 1
K.band(wsm, hr, "③ 读法 — 给『估值倍数假设』的输入", 8); hr += 1
K.mtext(wsm, hr, ("SNOW 现价 EV/Sales 14.8x 在 DDOG(17)与 MDB(8)之间, 低于 Databricks(19)/PLTR(36)。"
                  "三案逐年 EV/Sales 据此拍且高增长期高、随增速换挡 normalize: "
                  "Base 14→8.5x(向成熟数据 SaaS 收敛)、Bear 9→5.5x(Databricks 抢份额去溢价)、Bull 17→11x(Cortex 变现撑高溢价)。"
                  "EV/FCF 第二镜头: Base 45→28x(FCF margin 扩张摊薄倍数)。下页拍三案逐年倍数。"), "H", 4)

# ════════════ 6. 估值倍数假设(逐年 EV/Sales + EV/FCF 三案矩阵, 手搓)════════════
wmu = wb.create_sheet(S_MULT)
K.hdr(wmu, 1, "估值倍数假设 — 逐年(FY27E-FY31E)EV/Sales(主)+ EV/FCF(第二)三案矩阵", 9)
mr = K.mtext(wmu, 2, ("主镜头=EV/Sales(GAAP 亏损, EV/EBITDA/PE 出负; 用销售倍数出实数); 第二镜头=EV/FCF(adj FCF 已正, 实数)。"
                      "本页只拍逐年三案倍数(蓝字)+ 理由; 『情景切换』引用切换, 『情景估值』套用, 『估值对比』三案并排。"
                      "倍数据同业光谱(上页)+ 高增长期高、随增速换挡 normalize; 不锚自身 P/E(亏损)。"), "L", 3)
for col, w in zip("ABCDEFGHI", [20, 9, 9, 9, 9, 9, 11, 11, 11]):
    wmu.column_dimensions[col].width = w
wmu.column_dimensions["L"].width = 62

MULT_ROWS = {}
def write_mult_block(mkey, mname, data_dict, fmt, case_notes):
    global mr
    K.band(wmu, mr, f"{mname} 逐年三案", 9); mr += 1
    wmu[f"A{mr}"] = "案 / 年份"; wmu[f"A{mr}"].font = K.BF
    for col, y in zip(FCf, FWY):
        wmu[f"{col}{mr}"] = y; wmu[f"{col}{mr}"].font = K.BF; wmu[f"{col}{mr}"].fill = K.CH
    wmu[f"L{mr}"] = "为什么这么给"; wmu[f"L{mr}"].font = K.BF; wmu[f"L{mr}"].fill = K.CH
    mr += 1
    rows = {}
    for cs in CASES:
        K.lab(wmu, f"A{mr}", f"  {cs}")
        K.introw(wmu, mr, FCf, data_dict[cs], None, fmt)
        K.logic(wmu, f"L{mr}", case_notes[cs])
        rows[cs] = mr
        mr += 1
    MULT_ROWS[mkey] = rows
    mr += 1

write_mult_block("evs", "目标 EV/Sales (x)", EVS, K.MX, {
    "Bear": "Databricks 抢份额 + 增速换挡 → 去溢价至成熟数据 SaaS: 9x 起步 normalize 到 5.5x(MDB 下方)。",
    "Base": "增速优雅换挡、保住高增长数据 SaaS 溢价: 14x(≈现价)起步, 随增速从 30% 降到中双位数, normalize 到 8.5x。",
    "Bull": "Cortex 变现 + 增速守 25%+ → 保住 AI 数据云溢价: 17x 起步 normalize 到 11x(仍低于 Databricks 私募 19x)。"})
write_mult_block("evf", "目标 EV/FCF (x)", EVF, K.MX, {
    "Bear": "FCF margin 停滞 + 竞争压价 → 向成熟软件下沿: 30x 起步 normalize 到 20x。",
    "Base": "FCF margin 24%→30% 扩张、倍数随之摊薄: 45x 起步 normalize 到 28x(成熟高质软件区)。",
    "Bull": "FCF margin 破 34% + 高增长维持: 55x 起步 normalize 到 35x。"})

mr += 1
K.band(wmu, mr, "卖方对账 + 数据源", 9); mr += 1
K.mtext(wmu, mr, ("凭什么敢给: 卖方财报后把 EV/Sales 从 10.3x 上调至 14.7x(BofA)仍给买入; 我们 Base 14x 与之一致, "
                  "但逐年 normalize(承认增速终将换挡 + Databricks 份额压力), 不把当前高倍数外推。"
                  "EV/FCF 第二镜头独立三角: 当前前瞻 62x 偏贵, Base 假设 45x 起步已隐含现价上方需 FCF 放量。"
                  "数据源: EV/Sales/EV/FCF 同业光谱(上页)+ 增速换挡逻辑; 三案倍数=本研究判断。"), "L", 4)

# ════════════ 7. 情景切换(全模型唯一情景参数库; 逐年杠杆)════════════
sw = K.write_scenario_switch(wb.create_sheet(S_SW), {
    "title": "情景切换 — 客户增速/ARR增速/服务增速/FCF margin + 目标倍数(默认 Base)",
    "usage": ("B2 下拉切案 → 各杠杆『当前案』行跟切 → 客户数→产品收入→FCF→逐年 EV/Sales·EV/FCF 隐含价 全链变档。"
              "三案并排见『估值对比』。核心杠杆: ①客户数增速(logo) ②ARR/客户增速(NRR 扩张) ③FCF margin ④目标 EV/Sales ⑤目标 EV/FCF。"),
    "cases": CASES, "default": "Base",
    "triggers": [
        ("Bear", "Databricks 抢份额加剧(DBSQL 反攻)+ Iceberg 削弱锁定 + Cortex 变现不及 + NRR 跌破 120% → 产品增速降至 ~20%、EV/Sales 去溢价至 5.5-9x。"),
        ("Base", "Q1 加速部分延续、NRR 守 120%+、Cortex 消费飞轮渐起、FCF margin 稳步扩张 → 产品增速 30%→中双位数、EV/Sales 温和 normalize 到 8.5x。"),
        ("Bull", "Cortex/Snowflake Intelligence 成第二曲线 + NRR 回升至 130%+ + 增速守 25%+ + FCF margin 破 34% → 保住 AI 数据云溢价(EV/Sales 11-17x)。"),
    ],
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "levers": [
        {"key": "g_cust", "name": "客户数增速(logo)", "fmt": K.PCT, "cols": FCf,
         "vals": G_CUST,
         "desc": "总客户数净增速(新增 logo)。物理锚下半, 随基数扩大递减。",
         "stories": {"Bear": "大客户渗透见顶+竞争分流新签。", "Base": "稳健扩张, G2000 渗透continue。", "Bull": "AI 用例拉动新签加速。"},
         "hist": [None, 0.317, 0.199, 0.172, 0.209]},
        {"key": "g_arr", "name": "ARR/客户增速(NRR扩张)", "fmt": K.PCT, "cols": FCf,
         "vals": G_ARR,
         "desc": "每客户产品收入增速 = NRR 存量扩张 + 大客户 mix 上移。NRR>100% 时存量自然扩张, 估值最敏感变量。",
         "stories": {"Bear": "NRR 跌破 120%, 消费降本压制扩张。", "Base": "NRR 守 120%+, 大客户消费深化。", "Bull": "Cortex 抬客单价, NRR 回升 130%+。"},
         "hist": [None, 0.293, 0.146, 0.107, 0.067]},
        {"key": "g_svc", "name": "专业服务增速", "fmt": K.PCT, "cols": FCf,
         "vals": G_SVC,
         "desc": "专业服务收入增速(占约 5%, 低敏感)。",
         "stories": {"Bear": "实施需求随消费放缓。", "Base": "跟随产品温和增长。", "Bull": "AI 实施带动服务。"},
         "hist": [None, 0.608, 0.102, 0.171, 0.293]},
        {"key": "fcfm", "name": "adj FCF margin", "fmt": K.PCT, "cols": FCf,
         "vals": FCFM,
         "desc": "adj FCF / 总营收。指引 FY27 23%; 规模效应释放路径。",
         "stories": {"Bear": "AI 推理成本+竞争压价, margin 停滞。", "Base": "规模效应, 24%→30%。", "Bull": "运营杠杆充分, 破 34%。"},
         "hist": [None, None, 0.28, 0.24, 0.24]},
    ],
    "linked": [
        {"key": "evs", "name": "目标 EV/Sales(链引估值倍数假设)", "fmt": K.MX,
         "src_sheet": S_MULT, "src_row0": MULT_ROWS["evs"]["Bear"], "note": "逐年 EV/Sales, 三案链引『估值倍数假设』; 改去那页改蓝字。仅前瞻列(FCf)有值。"},
        {"key": "evf", "name": "目标 EV/FCF(链引估值倍数假设)", "fmt": K.MX,
         "src_sheet": S_MULT, "src_row0": MULT_ROWS["evf"]["Bear"], "note": "逐年 EV/FCF, 三案链引『估值倍数假设』。仅前瞻列(FCf)有值。"},
    ],
})
# linked 用 ALLC 全列引用, 但 MULT 只有 FCf 列有值 → 历史列会引到空, 需覆盖为仅 FCf
# 修正: linked 三案行 + 当前案行的历史列清空(MULT 无历史列)
SC_ws = wb[S_SW]
for lk_key in ["evs", "evf"]:
    b0 = sw["SWB"][lk_key]
    for rr in range(b0, b0 + len(CASES) + 1):
        for col in HC:
            SC_ws[f"{col}{rr}"].value = None

# 全局参数(三案共用)──────
_r = sw["next_row"]
K.band(SC_ws, _r, "全局参数(三案共用; 逐年 forward, 不折现)", 11); _r += 1
gp = [("净现金($B)", NET_CASH, K.N1, "现金+投资 $4.79B − 可转债 $2.3B。"),
      ("股本(mn股)", SHARES_M, K.N0, "当前流通 ~347M(SBC 稀释与回购大致对冲, 前瞻近似不变)。"),
      ("现价($)", PRICE, K.PX, "2026-07-10。")]
GLOBAL_ROWS = {}
for nm, v, fmt, note in gp:
    K.lab(SC_ws, f"A{_r}", nm, b=True)
    K.inp(SC_ws, f"B{_r}", v, None, fmt)
    K.logic(SC_ws, f"L{_r}", note)
    GLOBAL_ROWS[nm] = _r
    _r += 1

SWB = sw["SWB"]; SWACT = sw["SWACT"]

# ════════════ 8. 物理锚 [ANCHOR] ════════════
anchor = K.write_anchor(wb.create_sheet(S_ANCHOR), {
    "title": "Snowflake 物理锚 — 客户数 × ARR/客户 × NRR",
    "all_cols": ALLC, "all_years": ALLY,
    "series": [
        ("总客户数(K)",
         [5.944, 7.828, 9.384, 10.996, 13.296, None, None, None, None, None],
         "总客户数(千); 前瞻由『情景切换』客户增速驱动。FY26 13,296, Q1FY27 13,912。", K.N1),
        ("$1M+ 产品收入客户数",
         [184, 330, 449, 576, 733, None, None, None, None, None],
         "TTM 产品收入>$1M 客户; 贡献约 68% 产品收入。FY26 733, Q1FY27 779。", K.N0),
        ("NRR (%)",
         [1.78, 1.51, 1.33, 1.26, 1.25, None, None, None, None, None],
         "净收入留存率; FY23 158% 峰值 → 一路降至 FY26 125%, Q1FY27 触底回升 126%。估值最敏感变量。", K.PCT),
        ("RPO ($B)",
         [None, 2.60, 5.17, 6.87, 9.77, None, None, None, None, None],
         "剩余履约义务; Q1FY27 $9.21B(+38%), 未来 12 月确认约 50%。订单动能先行指标。", K.N1),
        ("ARR/客户 ($K)",
         [192, 248, 284, 315, 336, None, None, None, None, None],
         "= 产品收入 ÷ 总客户数(千); 反映客单价, 由 NRR 扩张 + 大客户 mix 驱动。前瞻由『情景切换』ARR 增速驱动。", K.N0),
    ],
    "source_note": ("口径: 客户数/[1M+]/NRR/RPO 来自 10-K/10-Q 一手披露(最新重述序列); ARR/客户=产品收入÷客户数(反推)。"
                    "早年客户数(FY22-23)为公司披露值。"),
    "role_note": ("作用: 产品收入 = 客户数 × ARR/客户 —— 两个可观测物理量。客户数增速(logo)+ ARR/客户增速(NRR 扩张)是驱动产品收入的物理量, "
                  "NRR/[1M+]/RPO 是它们的证据行。改『情景切换』客户增速或 ARR 增速 → 客户数/ARr → 产品收入 → EV → 隐含价 全链动(连通性铁律落点)。"),
})
A_CUST = anchor["row_of"]["总客户数(K)"]
A_ARR = anchor["row_of"]["ARR/客户 ($K)"]

# ★ 连通性: 前瞻客户数 = 上年 × (1+客户增速当前案); 前瞻 ARR/客户 = 上年 × (1+ARR增速当前案)
_anc = wb[S_ANCHOR]
_prevs = [HC[-1]] + list(FCf[:-1])
for _p, _c in zip(_prevs, FCf):
    K.fml(_anc, f"{_c}{A_CUST}", f"={_p}{A_CUST}*(1+{K.R(S_SW, _c + str(SWACT['g_cust']))})", K.N1, link=True)
    K.fml(_anc, f"{_c}{A_ARR}", f"={_p}{A_ARR}*(1+{K.R(S_SW, _c + str(SWACT['g_arr']))})", K.N0, link=True)

# ════════════ 9. 分部测算 ════════════
seg = K.write_segment_model(wb.create_sheet(S_SEG), {
    "title": "分部测算 — 产品收入(客户×ARR) + 专业服务",
    "all_cols": ALLC, "all_years": ALLY, "logic_col": "N",
    "groups": [
        ("产品收入: 客户数 × ARR/客户(物理锚驱动)", [
            ("客户数(K)", None, K.N1, "= 引『物理锚』总客户数。"),
            ("ARR/客户($K)", None, K.N0, "= 引『物理锚』ARR/客户。"),
            ("产品收入($B)", None, K.N1, "= 客户数(K) × ARR/客户($K) / 1000。喂利润表。"),
        ]),
        ("专业服务及其他(增速驱动)", [
            ("服务增速", None, K.PCT, "历史=实际 YoY; 前瞻=『情景切换』服务增速。"),
            ("专业服务($B)", None, K.N1, "历史实数; 前瞻=上年×(1+增速)。"),
        ]),
    ],
})
m = seg["m"]
# 客户数 / ARR 引物理锚(全列)
for col in ALLC:
    K.fml(wb[S_SEG], f"{col}{m['客户数(K)']}", f"={K.R(S_ANCHOR, col + str(A_CUST))}", K.N1, link=True)
    K.fml(wb[S_SEG], f"{col}{m['ARR/客户($K)']}", f"={K.R(S_ANCHOR, col + str(A_ARR))}", K.N0, link=True)
    K.fml(wb[S_SEG], f"{col}{m['产品收入($B)']}", f"={col}{m['客户数(K)']}*{col}{m['ARR/客户($K)']}/1000", K.N1)
    wb[S_SEG][f"{col}{m['产品收入($B)']}"].fill = K.OUT
# 专业服务: 历史引历史财务, 前瞻增速驱动
for col in HC:
    K.fml(wb[S_SEG], f"{col}{m['专业服务($B)']}", f"={K.R(S_HIST, col + str(SVC_HROW))}", K.N1, link=True)
K.lab(wb[S_SEG], f"B{m['服务增速']}", "n.m.", note=True)
for _i in range(1, len(HC)):
    _c, _p = HC[_i], HC[_i-1]
    K.fml(wb[S_SEG], f"{_c}{m['服务增速']}", f'=IFERROR({_c}{m["专业服务($B)"]}/{_p}{m["专业服务($B)"]}-1,"n.m.")', K.PCT)
for col in FCf:
    K.fml(wb[S_SEG], f"{col}{m['服务增速']}", f"={K.R(S_SW, col + str(SWACT['g_svc']))}", K.PCT, link=True)
_prevs = [HC[-1]] + list(FCf[:-1])
for _p, _c in zip(_prevs, FCf):
    K.fml(wb[S_SEG], f"{_c}{m['专业服务($B)']}", f"={_p}{m['专业服务($B)']}*(1+{_c}{m['服务增速']})", K.N1)

# ════════════ 10. 利润与收入假设(营收 roll-up + FCF, 手搓)════════════
wfu = wb.create_sheet(S_FUND)
K.hdr(wfu, 1, "利润与收入假设 — 产品+服务 → 总营收 → adj FCF(EV/FCF 镜头底座)", 9)
fr0 = K.mtext(wfu, 2, ("总营收 = 产品收入(物理锚驱动)+ 专业服务。adj FCF = 总营收 × FCF margin(情景), 是 EV/FCF 镜头底座。"
                       "GAAP 亏损(SBC 拖累)不进估值主链——主镜头 EV/Sales, 第二镜头 EV/FCF, 均绕开 GAAP 净利。"
                       "非 GAAP 营业利润率仅作盈利质量参照。"), "K", 2)
for col, w in zip("ABCDEFGHIJK", [22, 9, 9, 9, 9, 9, 9, 9, 9, 9, 9]):
    wfu.column_dimensions[col].width = w
wfu.column_dimensions["N"].width = 56
HDR_ROW = fr0
for col, y in zip(ALLC, ALLY):
    wfu[f"{col}{HDR_ROW}"] = y; wfu[f"{col}{HDR_ROW}"].font = K.BF; wfu[f"{col}{HDR_ROW}"].fill = K.CH
wfu[f"A{HDR_ROW}"] = "项目"; wfu[f"A{HDR_ROW}"].font = K.BF
wfu[f"N{HDR_ROW}"] = "逻辑/来源"; wfu[f"N{HDR_ROW}"].font = K.BF; wfu[f"N{HDR_ROW}"].fill = K.CH
fr = fr0 + 1
K.band(wfu, fr, "营收(引分部测算; EV/Sales 镜头底座)", 9); fr += 1
FU = {}
seg_link = [("产品收入($B)", m["产品收入($B)"]), ("专业服务($B)", m["专业服务($B)"])]
for name, srow in seg_link:
    K.lab(wfu, f"A{fr}", name)
    for col in ALLC:
        K.fml(wfu, f"{col}{fr}", f"={K.R(S_SEG, col + str(srow))}", K.N1, link=True)
    K.logic(wfu, f"N{fr}", f"= 引『{S_SEG}』{name}。")
    FU[name] = fr; fr += 1
R_TOTREV = fr
K.lab(wfu, f"A{fr}", "总营收($B)", b=True); wfu[f"A{fr}"].border = K.BORD; wfu[f"A{fr}"].fill = K.OUT
for col in ALLC:
    K.fml(wfu, f"{col}{fr}", f"={col}{FU['产品收入($B)']}+{col}{FU['专业服务($B)']}", K.N1)
K.logic(wfu, f"N{fr}", "= 产品 + 服务。")
fr += 1
R_REVYOY = fr
K.lab(wfu, f"A{fr}", "  总营收 YoY", note=True)
for i in range(1, len(ALLC)):
    K.fml(wfu, f"{ALLC[i]}{fr}", f"={ALLC[i]}{R_TOTREV}/{ALLC[i-1]}{R_TOTREV}-1", K.PCT)
fr += 1
R_PRODYOY = fr
K.lab(wfu, f"A{fr}", "  产品收入 YoY", note=True)
for i in range(1, len(ALLC)):
    K.fml(wfu, f"{ALLC[i]}{fr}", f"={ALLC[i]}{FU['产品收入($B)']}/{ALLC[i-1]}{FU['产品收入($B)']}-1", K.PCT)
fr += 1

K.band(wfu, fr, "adj FCF(EV/FCF 镜头底座)", 9); fr += 1
R_FCFM = fr
K.lab(wfu, f"A{fr}", "adj FCF margin(%)")
# 历史 FCF margin (adj FCF / 总营收): FY22 n.m., FY23 ~7%, FY24 28%, FY25 24%, FY26 24%
K.introw(wfu, fr, HC, [0.0, 0.07, 0.28, 0.24, 0.24], None, K.PCT)
for col in FCf:
    K.fml(wfu, f"{col}{fr}", f"={K.R(S_SW, col + str(SWACT['fcfm']))}", K.PCT, link=True)
K.logic(wfu, f"N{fr}", "历史=公司披露 adj FCF/营收(FY24 峰值 28% 后 FY25-26 24%); 前瞻=『情景切换』FCF margin。")
fr += 1
R_FCF = fr
K.lab(wfu, f"A{fr}", "adj FCF($B)", b=True); wfu[f"A{fr}"].fill = K.OUT
for col in ALLC:
    K.fml(wfu, f"{col}{fr}", f"={col}{R_TOTREV}*{col}{R_FCFM}", K.N1)
K.logic(wfu, f"N{fr}", "= 总营收 × adj FCF margin。FY26 ≈ $1.19B(实际)。")
fr += 1

K.band(wfu, fr, "非 GAAP 营业利润率(盈利质量参照, 不进估值链)", 9); fr += 1
R_NGOPM = fr
K.lab(wfu, f"A{fr}", "non-GAAP 营业利润率")
K.introw(wfu, fr, HC, [None, None, 0.08, 0.06, 0.10], None, K.PCT)
K.introw(wfu, fr, FCf, [0.135, 0.15, 0.17, 0.19, 0.21], None, K.PCT)
K.lab(wfu, f"B{fr}", "n.m.", note=True); K.lab(wfu, f"C{fr}", "n.m.", note=True)
K.logic(wfu, f"N{fr}", "指引 FY27 13.5%; 前瞻运营杠杆扩张。仅参照, 估值不用。")
fr += 1
K.band(wfu, fr, "口径说明", 9); fr += 1
K.mtext(wfu, fr, ("估值主链: 产品收入(客户×ARR 物理锚)+ 服务 = 总营收 → EV/Sales 镜头; adj FCF(营收×margin)→ EV/FCF 镜头。"
                  "GAAP 净利亏损(SBC 拖累)绕开, 不进估值。GAAP 转盈公司指引 Q4 FY28。"), "K", 2)
PROD_REVROW = FU["产品收入($B)"]
SVC_REVROW = FU["专业服务($B)"]

# ════════════ 11. 情景估值(逐年双镜头隐含价阶梯, 当前案, 手搓)════════════
wva = wb.create_sheet(S_VAL)
K.hdr(wva, 1, "情景估值 — 当前案逐年(FY27E-FY31E)EV/Sales + EV/FCF 双镜头隐含价", 8)
K.lab(wva, "G1", "当前情景→", note=True)
K.fml(wva, "H1", f"={K.R(S_SW, 'B2')}", K.N0, link=True); wva["H1"].fill = K.CUR
vr = K.mtext(wva, 2, ("本表=『情景切换』当前案(默认 Base)逐年隐含价。① EV/Sales 镜头: 总营收×EV/Sales=EV +净现金=权益 ÷股本=隐含价; "
                      "② EV/FCF 镜头: adj FCF×EV/FCF=EV +净现金=权益 ÷股本=隐含价。两镜头取均值=三角隐含价。不折现(逐年 forward 自立)。三案见『估值对比』。"), "H", 3)
for col, w in zip("ABCDEFGH", [24, 12, 12, 12, 12, 12, 12, 12]):
    wva.column_dimensions[col].width = w

NC_REF = K.R(S_SW, "B" + str(GLOBAL_ROWS["净现金($B)"]))
SH_REF = K.R(S_SW, "B" + str(GLOBAL_ROWS["股本(mn股)"]))
PX_REF = K.R(S_SW, "B" + str(GLOBAL_ROWS["现价($)"]))

K.band(wva, vr, "① EV/Sales 镜头(主)", 8); vr += 1
wva[f"A{vr}"] = "项目 / 年份"; wva[f"A{vr}"].font = K.BF
for col, y in zip(FCf, FWY):
    wva[f"{col}{vr}"] = y; wva[f"{col}{vr}"].font = K.BF; wva[f"{col}{vr}"].fill = K.CH
vr += 1
VR = {}
def val_row(label, b=False, out=False):
    global vr
    K.lab(wva, f"A{vr}", label, b=b)
    if out:
        wva[f"A{vr}"].fill = K.OUT
    rr = vr; vr += 1
    return rr
R_REV = val_row("总营收($B)")
for col in FCf:
    K.fml(wva, f"{col}{R_REV}", f"={K.R(S_FUND, col + str(R_TOTREV))}", K.N1, link=True)
R_EVS = val_row("EV/Sales(x)")
for col in FCf:
    K.fml(wva, f"{col}{R_EVS}", f"={K.R(S_SW, col + str(SWACT['evs']))}", K.MX, link=True)
R_EV_S = val_row("EV(销售镜头, $B)")
for col in FCf:
    K.fml(wva, f"{col}{R_EV_S}", f"={col}{R_REV}*{col}{R_EVS}", K.N0)
R_PX_S = val_row("隐含价 EV/Sales($)", b=True, out=True)
for col in FCf:
    K.fml(wva, f"{col}{R_PX_S}", f"=({col}{R_EV_S}+{NC_REF})/{SH_REF}*1000", K.PX, link=True)
vr += 1
K.band(wva, vr, "② EV/FCF 镜头(第二, adj FCF 已正)", 8); vr += 1
wva[f"A{vr}"] = "项目 / 年份"; wva[f"A{vr}"].font = K.BF
for col, y in zip(FCf, FWY):
    wva[f"{col}{vr}"] = y; wva[f"{col}{vr}"].font = K.BF; wva[f"{col}{vr}"].fill = K.CH
vr += 1
R_FCFV = val_row("adj FCF($B)")
for col in FCf:
    K.fml(wva, f"{col}{R_FCFV}", f"={K.R(S_FUND, col + str(R_FCF))}", K.N1, link=True)
R_EVF = val_row("EV/FCF(x)")
for col in FCf:
    K.fml(wva, f"{col}{R_EVF}", f"={K.R(S_SW, col + str(SWACT['evf']))}", K.MX, link=True)
R_EV_F = val_row("EV(FCF镜头, $B)")
for col in FCf:
    K.fml(wva, f"{col}{R_EV_F}", f"={col}{R_FCFV}*{col}{R_EVF}", K.N0)
R_PX_F = val_row("隐含价 EV/FCF($)", b=True, out=True)
for col in FCf:
    K.fml(wva, f"{col}{R_PX_F}", f"=({col}{R_EV_F}+{NC_REF})/{SH_REF}*1000", K.PX, link=True)
vr += 1
K.band(wva, vr, "三角: 双镜头均值隐含价 vs 现价", 8); vr += 1
wva[f"A{vr}"] = "项目 / 年份"; wva[f"A{vr}"].font = K.BF
for col, y in zip(FCf, FWY):
    wva[f"{col}{vr}"] = y; wva[f"{col}{vr}"].font = K.BF; wva[f"{col}{vr}"].fill = K.CH
vr += 1
R_PX_AVG = val_row("三角隐含价(双镜头均值, $)", b=True, out=True)
for col in FCf:
    K.fml(wva, f"{col}{R_PX_AVG}", f"=AVERAGE({col}{R_PX_S},{col}{R_PX_F})", K.PX)
R_VSNOW = val_row("vs 现价", b=True)
for col in FCf:
    K.fml(wva, f"{col}{R_VSNOW}", f"={col}{R_PX_AVG}/{PX_REF}-1", K.PCT, link=True)
vr += 2
K.band(wva, vr, "方法与结论", 8); vr += 1
K.mtext(wva, vr, ("方法: 逐年(FY27E-FY31E)双镜头 —— EV/Sales(总营收×倍数)与 EV/FCF(adj FCF×倍数)各出隐含价, 取均值=三角隐含价, 不折现。"
                  "GAAP 亏损故绕开 P/E; DCF 作 sanity(FCF 已正, 内在价值与倍数法同量级)。"
                  "结论(Base): 三角隐含价 FY27E→FY31E 逐年抬升; 近端(FY27-28)EV/FCF 镜头偏低(FCF 倍数贵)、EV/Sales 镜头接近现价, 均值≈现价附近 —— "
                  "现价已把'AI 赢家'定价充分, 上行需 Cortex 变现兑现(Bull), 下行来自 Databricks 份额(Bear)。12M 目标取 FY28E 概率加权, 见『估值对比』。"), "H", 4)

# ════════════ 12. 估值对比(三案逐年 block, 手搓; 防污染=只引矩阵行/全局)════════════
wcp = wb.create_sheet(S_CMP)
K.hdr(wcp, 1, "估值对比 — Bear/Base/Bull 三案逐年(FY27E-FY31E)双镜头隐含价", 8)
K.mtext(wcp, 2, ("三案各自推演逐年双镜头: 产品(客户×ARR)+服务=总营收 → EV/Sales; adj FCF → EV/FCF; 取均值=三角隐含价。"
                 "三案永远并排, 不随开关变(只引『情景切换』矩阵各案行 + 全局参数, 防当前案污染)。"
                 "底部: 12M 目标价 = 各案 FY28E 隐含价概率加权(0.30/0.45/0.25)。"), "H", 3)
for col, w in zip("ABCDEFGH", [24, 11, 11, 11, 11, 11, 13, 13]):
    wcp.column_dimensions[col].width = w

# 三案矩阵源行
def bcol(col): return ["G", "H", "I", "J", "K"].index(col)  # FCf 内偏移
rows_def = [
    ("cust", "客户数(K)", K.N1, False, False, "cust"),
    ("arr", "ARR/客户($K)", K.N0, False, False, "arr"),
    ("prod", "产品收入($B)", K.N1, False, False, "prod"),
    ("svc", "专业服务($B)", K.N1, False, False, "svc"),
    ("rev", "总营收($B)", K.N1, True, False, "rev"),
    ("evs", "EV/Sales(x)", K.MX, False, False, "evs"),
    ("ev_s", "EV(销售, $B)", K.N0, False, False, "ev_s"),
    ("px_s", "隐含价 EV/Sales($)", K.PX, False, False, "px_s"),
    ("fcf", "adj FCF($B)", K.N1, False, False, "fcf"),
    ("evf", "EV/FCF(x)", K.MX, False, False, "evf"),
    ("ev_f", "EV(FCF, $B)", K.N0, False, False, "ev_f"),
    ("px_f", "隐含价 EV/FCF($)", K.PX, False, False, "px_f"),
    ("px", "三角隐含价($)", K.PX, True, True, "px"),
    ("vsnow", "vs 现价", K.PCT, False, False, "vsnow"),
    ("ipe", "隐含 forward P/E(GAAP亏损 N/M, 见 EV/Sales)", K.MX, False, False, "nm_pe"),
]
block_h = len(rows_def) + 3
CMP_BLOCK = {}
b0 = 6
NC = NET_CASH; SH = SHARES_M
for ci, cname in enumerate(CASES):
    r0 = b0 + ci * block_h
    K.band(wcp, r0, f"{cname} 案 — 逐年双镜头隐含价", 8)
    yr = r0 + 1
    wcp[f"A{yr}"] = "项目 / 年份"; wcp[f"A{yr}"].font = K.BF
    for col, y in zip(FCf, FWY):
        wcp[f"{col}{yr}"] = y; wcp[f"{col}{yr}"].font = K.BF; wcp[f"{col}{yr}"].fill = K.CH
    rr = yr + 1
    A = {}
    for key, label, fmt, bold, out, kind in rows_def:
        A[key] = rr
        K.lab(wcp, f"A{rr}", label, b=bold)
        if out:
            wcp[f"A{rr}"].fill = K.OUT
        if kind == "nm_pe":
            for col in FCf:
                K.lab(wcp, f"{col}{rr}", "N/M", note=True)
            rr += 1
            continue
        for j, col in enumerate(FCf):
            prev = FWY[j-1] if j > 0 else None
            pcol = FCf[j-1] if j > 0 else None
            if kind == "cust":
                # 客户数逐年 = 上年 × (1+客户增速案行); FY26 基年引『物理锚』F 列(静态历史锚, 防污染允许)
                gsrc = SWB["g_cust"] + ci
                if j == 0:
                    f = f"={K.R(S_ANCHOR, 'F' + str(A_CUST))}*(1+{K.R(S_SW, col + str(gsrc))})"
                else:
                    f = f"={pcol}{A['cust']}*(1+{K.R(S_SW, col + str(gsrc))})"
            elif kind == "arr":
                gsrc = SWB["g_arr"] + ci
                if j == 0:
                    f = f"={K.R(S_ANCHOR, 'F' + str(A_ARR))}*(1+{K.R(S_SW, col + str(gsrc))})"
                else:
                    f = f"={pcol}{A['arr']}*(1+{K.R(S_SW, col + str(gsrc))})"
            elif kind == "prod":
                f = f"={col}{A['cust']}*{col}{A['arr']}/1000"
            elif kind == "svc":
                gsrc = SWB["g_svc"] + ci
                if j == 0:
                    f = f"={K.R(S_HIST, 'F' + str(SVC_HROW))}*(1+{K.R(S_SW, col + str(gsrc))})"
                else:
                    f = f"={pcol}{A['svc']}*(1+{K.R(S_SW, col + str(gsrc))})"
            elif kind == "rev":
                f = f"={col}{A['prod']}+{col}{A['svc']}"
            elif kind == "evs":
                src = MULT_ROWS["evs"][cname]
                f = f"={K.R(S_MULT, col + str(src))}"
            elif kind == "ev_s":
                f = f"={col}{A['rev']}*{col}{A['evs']}"
            elif kind == "px_s":
                f = f"=({col}{A['ev_s']}+{NC})/{SH}*1000"
            elif kind == "fcf":
                fsrc = SWB["fcfm"] + ci
                f = f"={col}{A['rev']}*{K.R(S_SW, col + str(fsrc))}"
            elif kind == "evf":
                src = MULT_ROWS["evf"][cname]
                f = f"={K.R(S_MULT, col + str(src))}"
            elif kind == "ev_f":
                f = f"={col}{A['fcf']}*{col}{A['evf']}"
            elif kind == "px_f":
                f = f"=({col}{A['ev_f']}+{NC})/{SH}*1000"
            elif kind == "px":
                f = f"=AVERAGE({col}{A['px_s']},{col}{A['px_f']})"
            elif kind == "vsnow":
                f = f"={col}{A['px']}/{PRICE}-1"
            link = ("'" in f)
            K.fml(wcp, f"{col}{rr}", f, fmt, link=link)
        rr += 1
    CMP_BLOCK[cname] = A

# 底部汇总: FY28E 加权(12M TP)──────
sum_r0 = b0 + len(CASES) * block_h + 1
K.band(wcp, sum_r0, "概率加权 → 12M 目标价(FY28E) + FY31E 终点", 8); sum_r0 += 1
for col, h in zip(["A", "B", "C", "D", "E"], ["项目", "Bear", "Base", "Bull", "概率加权"]):
    wcp[f"{col}{sum_r0}"] = h; wcp[f"{col}{sum_r0}"].font = K.BF; wcp[f"{col}{sum_r0}"].fill = K.CH
wcp.merge_cells(f"F{sum_r0}:H{sum_r0}"); wcp[f"F{sum_r0}"] = "说明"
wcp[f"F{sum_r0}"].font = K.BF; wcp[f"F{sum_r0}"].fill = K.CH
sum_r0 += 1
R_WT = sum_r0
K.lab(wcp, f"A{sum_r0}", "概率权重")
for col, cn in zip(["B", "C", "D"], CASES):
    K.inp(wcp, f"{col}{sum_r0}", WEIGHTS[cn], None, K.PCT)
wcp.merge_cells(f"F{sum_r0}:H{sum_r0}")
K.logic(wcp, f"F{sum_r0}", "主观概率: 承认 Databricks 份额压力 + Q1 加速可持续性未定 → Bear 0.30 / Base 0.45 / Bull 0.25。")
sum_r0 += 1
PX_KEY = {cn: CMP_BLOCK[cn]["px"] for cn in CASES}
R_TP28 = sum_r0
K.lab(wcp, f"A{sum_r0}", "FY28E 三角隐含价(12M 目标)", b=True); wcp[f"A{sum_r0}"].fill = K.OUT
for col, cn in zip(["B", "C", "D"], CASES):
    K.fml(wcp, f"{col}{sum_r0}", f"=H{PX_KEY[cn]}", K.PX, link=True)   # H 列=FY28E
K.fml(wcp, f"E{sum_r0}", f"=B{sum_r0}*B{R_WT}+C{sum_r0}*C{R_WT}+D{sum_r0}*D{R_WT}", K.PX)
wcp[f"E{sum_r0}"].fill = K.CUR
wcp.merge_cells(f"F{sum_r0}:H{sum_r0}")
K.logic(wcp, f"F{sum_r0}", "= 各案 FY28E 三角隐含价 × 概率 → 12M 目标价。")
sum_r0 += 1
R_TP31 = sum_r0
K.lab(wcp, f"A{sum_r0}", "FY31E 三角隐含价(成长终点)", b=True); wcp[f"A{sum_r0}"].fill = K.OUT
for col, cn in zip(["B", "C", "D"], CASES):
    K.fml(wcp, f"{col}{sum_r0}", f"=K{PX_KEY[cn]}", K.PX, link=True)   # K 列=FY31E
K.fml(wcp, f"E{sum_r0}", f"=B{sum_r0}*B{R_WT}+C{sum_r0}*C{R_WT}+D{sum_r0}*D{R_WT}", K.PX)
wcp.merge_cells(f"F{sum_r0}:H{sum_r0}")
K.logic(wcp, f"F{sum_r0}", "= 各案 FY31E 三角隐含价 × 概率 → 成长进价终点。")
sum_r0 += 1
K.lab(wcp, f"A{sum_r0}", "现价对照($)")
K.inp(wcp, f"E{sum_r0}", PRICE, None, K.PX); wcp[f"E{sum_r0}"].fill = K.CUR
wcp.merge_cells(f"F{sum_r0}:H{sum_r0}")
K.logic(wcp, f"F{sum_r0}", "现价 $267.49; 街上一致目标 $292(Strong Buy)。")
sum_r0 += 1
K.lab(wcp, f"A{sum_r0}", "对照: 街上一致目标")
K.inp(wcp, f"C{sum_r0}", 292, None, K.PX)
wcp.merge_cells(f"F{sum_r0}:H{sum_r0}")
K.logic(wcp, f"F{sum_r0}", "48 分析师均值 $292(高 $500/低 $110); 我们的 Base 与之对比在'分歧根源'讨论。")
sum_r0 += 2
K.mtext(wcp, sum_r0, ("关键洞察: 3 个月 +75% 后, Base 三角隐含价近端≈现价, 12M 概率加权目标接近现价 —— "
                      "现价已把 Q1 加速 + AI 赢家叙事定价充分。Bull(Cortex 成第二曲线)才把估值推向显著上行; "
                      "Bear(Databricks 抢份额 + 去溢价)是对称下行。风险收益偏均衡, 非追高点。"), "H", 3)

# ════════════ 13. 综合判断仪表盘 ════════════
IMPL_28 = K.R(S_VAL, "H" + str(R_PX_AVG))   # FY28E(当前案 Base)
IMPL_31 = K.R(S_VAL, "K" + str(R_PX_AVG))   # FY31E
PX_D = K.R(S_SW, "B" + str(GLOBAL_ROWS["现价($)"]))
K.write_dashboard(wb.create_sheet(S_DASH), {
    "title": "综合判断仪表盘 — A 基本面拐点 · B 估值错位 · C 催化剂 · D 情绪确认",
    "usage": ("预测引擎=B(估值错位)+C(催化剂); D 情绪只做 timing。估值用逐年双镜头(EV/Sales+EV/FCF)三角隐含价。"),
    "blocks": [
        {"title": "A. 基本面拐点 — 业务在结构性变好吗?", "rows": [
            ("产品收入重新加速", "✓ 已现", "Q1FY27 产品 +34%(环比 +30% 加速 4pp), 史上最强单季环比 +$108M; 指引上调 +27%→+31%。"),
            ("NRR 触底回升", "✓ 初现", "NRR 连跌 10+ 季后 Q1FY27 回升至 126%(+1pt); 回升 1pt 未确立趋势, 需持续。"),
            ("AI 变现(Cortex)", "✗ 待验证", "CoCo 9 周 50% 客户用、早期采用者消费 +11%; 是消费飞轮催化, 尚未成独立收入线。"),
            ("A 判断", "【中强】", "加速真实且被指引确认, 但可持续性(易比基数)与 AI 变现体量未坐实。", True),
        ]},
        {"title": "B. 估值错位(预测引擎)— 市场给 vs 双镜头该给 → GAP", "rows": [
            ("市场现在给(现价)", {"fml": f"={PX_D}", "fmt": K.PX, "fill": True}, "= 现价 $267.49。"),
            ("双镜头该给(Base FY28E, 12M)", {"fml": f"={IMPL_28}", "fmt": K.PX}, "= Base FY28E 三角隐含价(引『情景估值』当前案)。"),
            ("双镜头该给(Base FY31E 终点)", {"fml": f"={IMPL_31}", "fmt": K.PX}, "= Base FY31E 三角隐含价(成长终点)。"),
            ("错位 GAP = FY28E该给÷市场给 − 1",
             {"fml": lambda ro: f"=B{ro['双镜头该给(Base FY28E, 12M)']}/B{ro['市场现在给(现价)']}-1", "fmt": K.PCT},
             "GAP 近 0 = 现价≈Base 双镜头该给 → 定价充分, 上行靠 Cortex(Bull)、下行靠份额(Bear)。"),
            ("回测: 2024-09 $108 低点读数", "GAP 当时深正", "增速换挡+NRR 下滑绝望期, 双镜头该给远高于 $108 → 那波修复的预测依据。"),
        ]},
        {"title": "C. 催化剂 — 什么改变 GAP", "rows": [
            ("Q2 FY27 财报(约 2026-08)", "进行中", "产品增速能否守 30%+ 与 NRR 是否续升; 指引 Q2 产品 +30%。"),
            ("Cortex 变现里程碑", "待", "CoCo 渗透曲线 + 是否被后续季度增量计入指引 → AI 变现体量直接验证。"),
            ("Databricks 动态 / IPO", "待", "Databricks ARR +80% 反超, 若加速抢份额 → 下调 SNOW; 其 IPO 重定价影响板块。"),
            ("C 判断", "近端明确", "Q2 财报是最近 catalyst; AI 变现与份额是中期胜负手。", True),
        ]},
        {"title": "D. 情绪确认 — timing + 刹车(定性档位)", "rows": [
            ("3 个月 +75%", "过热信号", "4 月 $153 → 7 月 $267, 财报后 +37% 单日; 情绪与基本面共振。"),
            ("现价 vs 双镜头该给", "≈ Base 该给", "现价≈Base 12M 三角隐含价, 非极端高估也非低估。"),
            ("当前档位", "【重估兑现/偏热】", "AI 赢家叙事已 price-in; 追高性价比低, 回调更好。", True),
            ("衰减扳机", "NRR 跌破 120% / 产品增速跌破 28% / Databricks 加速抢单", "任一翻→降档+EV/Sales 下调。", ),
        ]},
    ],
    "final": {"band": "★ 综合判断(A+B+C+D)",
              "text": ("基本面中强(A: 加速真实但可持续性待验)+ 估值错位近 0(B: 现价≈Base 双镜头该给)+ 催化剂近端明确(C)+ 情绪偏热(D)"
                       "→ 现价已把 Q1 加速 + AI 赢家叙事定价充分。12M 概率加权目标≈现价附近, 街上 $292 略偏乐观。"
                       "评级中性/持有; 上行期权=Cortex 成第二曲线(Bull), 下行=Databricks 抢份额+去溢价(Bear), 风险收益偏均衡。"
                       "操作: 不追高, 回调至 Bear 支撑区分批。")},
    "tracking": {
        "intro": "哪个指标恶化/兑现 → 哪个假设先动 → 触发什么动作。",
        "rows": [
            ("__band__", "一、核心驱动(NRR + 产品增速)"),
            ("NRR", "126% 回升", "命门: ARR/客户扩张乘数, 估值最敏感", "季报披露", "跌破 120% → 下调 ARR 增速、转 Bear"),
            ("产品收入增速", "+34%(Q1)", "命门: 客户数×ARR 主链", "季报", "跌破 28% → 下调 g_cust/g_arr"),
            ("__band__", "二、AI 变现"),
            ("Cortex/CoCo 消费占比", "早期(+11% 采用者)", "命门: Bull 第二曲线期权", "季报/投资者日", "计入独立收入线 → 上调 Bull 概率"),
            ("__band__", "三、竞争"),
            ("Databricks ARR/增速", "$6.9B/+80%", "命门: 份额天平", "Databricks 披露/IPO", "加速抢单/DBSQL 反攻 → 下调 EV/Sales"),
            ("__band__", "四、估值"),
            ("EV/Sales(前瞻)", "~14.8x", "命门: 倍数对增速敏感", "股价/一致预期", "增速换挡未兑现 → 目标倍数下调"),
        ],
    },
})

# ════════════ 全局格式 + 落盘 ════════════
K.finalize(wb, freeze={
    S_COVER: "A2", S_HIST: "B3", S_PX: "B4", S_CONS: "A2", S_HMULT: "B3",
    S_MULT: "B3", S_SW: "B3", S_ANCHOR: "B3", S_SEG: "B3",
    S_FUND: "B5", S_VAL: "B3", S_CMP: "B3", S_DASH: "B6",
})
_output_dir = os.environ.get("VALUATION_OUTPUT_DIR", os.path.join(_repo_root, "out"))
os.makedirs(_output_dir, exist_ok=True)
out = os.path.join(_output_dir, "SNOW_valuation_model.xlsx")
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("saved:", out)
print("sheets:", wb.sheetnames)
print("PRICE:", PRICE, "SHARES_M:", SHARES_M)
