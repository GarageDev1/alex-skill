# -*- coding: utf-8 -*-
"""
build_adbe.py — Adobe (ADBE) 估值模型，基于 build_kit(v2)。P/E-primary SaaS 变体。

SaaS 锚：Customer Group 订阅收入(C&MP / BP&C) + Product&Services 残值；ARR/RPO/AI-ARR 为物理证据行。
分部：C&MP(创意+营销云，~68%) + BP&C(Acrobat/Express/消费者，~27%) + 产品与服务(~4%，衰减)。
估值镜头：P/E 平行镜头=主结论(GAAP EPS；Adobe 回购致 BPS 萎缩、P/B 失真) + P/B 主线(kit 合规，参考)。
关键 Adobe 特色：① 回购致股本逐年下降(501M→399M)，forward EPS 用递减股本 override；
              ② 留存率为负(回购>净利)，权益逐年萎缩→P/B 失真，故 P/E 做主线；
              ③ FY26 分部口径合并(DM/DX/P&A→Customer Group)+收购 Semrush(+$480M ARR)。
口径：USD 单币种；财年截至每年 11 月最后周五(非自然年)；股本稀释百万股。
"""
import os, json
from openpyxl import Workbook
import build_kit as K

# ════════════ 0. 全局轴 ════════════
ALLC = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
ALLY = ["2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
HC, HY = ["B", "C", "D", "E", "F"], ["2021", "2022", "2023", "2024", "2025"]
FC = ["F", "G", "H", "I", "J", "K"]
FCf = FC[1:]
FX = 1.0
CASES = ["Bear", "Base", "Bull"]

S_COVER, S_HIST, S_PX, S_CONS = "封面", "历史财务与估值", "股价走势", "卖方研报共识"
S_HMULT, S_MULT, S_SW = "历史估值倍数", "估值倍数假设", "情景切换"
S_ANCHOR, S_SEG, S_FUND = "SaaS物理锚", "分部测算", "利润与收入假设"
S_VAL, S_CMP, S_DASH = "情景估值", "估值对比", "综合判断仪表盘"

SEG_CMP, SEG_BPC, SEG_PS = "C&MP 收入", "BP&C 收入", "产品与服务 收入"

# 月度股价（calendar 月末，来自 tvremix 周K重采样）
_repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_input_dir = os.environ.get("VALUATION_INPUT_DIR", os.path.join(_repo_root, "out"))
with open(os.path.join(_input_dir, "adobe_monthly.json"), encoding="utf-8") as _f:
    _px = json.load(_f)
MONTHLY = [(ym, round(c, 2)) for ym, c in _px["monthly"]]
YE, YA, YH, YL = _px["yend"], _px["yavg"], _px["yhigh"], _px["ylow"]
PX_NOW = 205.02

wb = Workbook(); wb.remove(wb["Sheet"])

# ════════════ 1. 封面 ════════════
K.write_cover(wb.create_sheet(S_COVER), {
    "title": "Adobe (ADBE) 估值模型",
    "meta": [
        ("报告日期", "2026-07-01"),
        ("数据截止", "FY25 10-K + Q2 FY26 10-Q/8-K + 2026-07-01 收盘"),
        ("现价", f"${PX_NOW:.2f}  |  市值 $81.5B  |  股本 ~399M(基本)"),
        ("估值口径", "TTM P/E 11.7x(GAAP EPS $17.50) / FY26E 11.4x(GAAP $17.95) / 8.4x(非GAAP $24.40)"),
        ("时效声明", "基于 2026-07-01 时点的财报/共识/股价；AI 竞争与治理(CEO/CFO 双离任)假设变化快，建议每季财报后更新。"),
        ("方法一句话", "物理锚(Customer Group 订阅ARR/RPO/AI-ARR)→ 分部收入(C&MP/BP&C/P&S 增速驱动)→ 段驱动利润 → EPS(递减股本)→ 目标P/E×前瞻每股 → 隐含价。"),
    ],
    "takeaways": [
        ("① 当下估值位置", "现价 $205 对应 TTM P/E 11.7x(GAAP)，处自身至少近十年低位(常态 35-50x，FY25 已降至 20x)。"),
        ("② 核心引擎", "Total Adobe ARR $27.1B(Q2 FY26，+12.5% YoY，含 Semrush；有机 ~+10-11%)；AI-first ARR >$500M(同比×3)。"),
        ("③ 核心分歧", "11.7x 是价值陷阱(AI 抹平创意技能溢价 + Canva/Figma 蚕食 + 治理真空)还是过度恐惧(ARR 未减速、Adobe 嵌 AI 而非被替代)。"),
        ("④ 三情景目标价", "Bear/Base/Bull 从 C&MP/BP&C 增速 + 利润率 + 目标P/E 三杠杆翻档，沿同一条链算出，见『估值对比』。"),
        ("⑤ 主要风险", "Creative Cloud ARR 有机增速跌破高单位数 / 治理(CEO 继任未定+CFO 换人) / 利润端 Q2 FY26 净利仅 +1.2% / 板块系统性估值倍数下修。"),
    ],
})

# ════════════ 2. 历史财务与估值 ════════════
ha = K.write_history(wb.create_sheet(S_HIST), {
    "title": "Adobe 历史财务与估值 ($B) — FY21-25A + 当下(TTM/Q2 FY26)",
    "hist_cols": HC, "hist_years": HY,
    "fx_hist": [1, 1, 1, 1, 1], "fx_now": 1.0,
    "vals_in_usd": True, "fx_label": "FX (USD 单币种=1)",
    "segments": [
        # Customer Group 口径(FY24-25 披露；FY21-23 按 mix 估，标"估")
        (SEG_CMP, [10.5, 11.7, 13.0, 14.7, 16.3], True),
        (SEG_BPC, [4.0, 4.5, 5.1, 5.7, 6.5], True),
        (SEG_PS, [1.3, 1.4, 1.3, 1.1, 1.0], False),
    ],
    "total_now": 25.2,
    "gm_pct": [0.882, 0.877, 0.879, 0.890, 0.893], "gm_now": 0.888,
    "ni": [4.822, 4.756, 5.428, 5.560, 7.130], "ni_now": 7.229,
    "eq": [13.3, 14.8, 16.5, 14.1, 11.6], "eq_now": 11.5,
    "shares": [481, 471, 459, 450, 427], "shares_now": 413,  # 稀释百万股；当下=TTM 均值(对齐 TTM EPS $17.50)
    "px_end": [YE["2021"], YE["2022"], YE["2023"], YE["2024"], YE["2025"]],
    "px_now": PX_NOW,
    "px_avg": [YA["2021"], YA["2022"], YA["2023"], YA["2024"], YA["2025"]],
    "band_note": "P/E 历史：FY21 57x / FY22 33x / FY23 50x / FY24 35x / FY25 20x / 当下 11.7x——处至少近十年低位。P/B 因回购致 BPS 萎缩而失真(7.3x 且走高)，估值看 P/E。",
    "notes": [
        (SEG_CMP, "Creative & Marketing Professionals：Creative Cloud(创意) + Document Cloud(专业) + Experience Cloud(营销云)。FY24 $14.75B / FY25 $16.30B(10-K 披露)；FY21-23 按 Customer Group mix 估。"),
        (SEG_BPC, "Business Professionals & Consumers：Acrobat/Express/消费者 + 文档协作。FY24 $5.66B / FY25 $6.50B(披露，+15%)；增速最快。"),
        (SEG_PS, "产品与服务(非订阅残值，许可/培训/服务)，逐年衰减。FY25 $0.97B。"),
        ("HREV", "总营收=10-K/10-Q 实际(EDGAR XBRL 对账)；分部 FY24-25 披露、FY21-23 估。TTM $25.2B。"),
        ("HGMP", "毛利率：公司实际，订阅模式 88-89% 稳定。"),
        ("HNI", "净利：GAAP 实际(EDGAR)；TTM $7.23B。"),
        ("HEQ", "股东权益：FY23-25 实际(16.5/14.1/11.6 $B)；FY21-22 估。回购致权益逐年萎缩。"),
        ("HSH", "稀释股本(百万)：实际(EDGAR)，501M(FY17)→399M(Q2 FY26 基本)，8 年回购压减 20%。"),
        ("HPX", "年末股价：calendar 月末收盘(tvremix 周K重采样)；当下=现价 $205。"),
    ],
})

# ════════════ 3. 股价走势 ════════════
def phase_fn(ym):
    if ym <= "2022-09": return "① 利率估值倍数下修+Figma 收购"
    if ym <= "2023-11": return "② Figma 终止+AI 叙事反弹"
    if ym <= "2024-11": return "③ 高位回落(增长减速)"
    if ym <= "2025-11": return "④ 持续下行(AI 竞争)"
    if ym <= "2026-03": return "⑤ CEO 卸任"
    return "⑥ CFO 离任+板块估值倍数下修"
px = K.write_price_chart(wb.create_sheet(S_PX), MONTHLY, {
    "fn": phase_fn,
    "rows": [("① 利率估值倍数下修+Figma 收购", "2022 加息 + $200 亿收购 Figma 引发垄断与整合担忧，股价从 $670 跌至 $275"),
             ("② Figma 终止+AI 叙事反弹", "2023-12 监管阻力下 Figma 收购终止(付 $10 亿分手费)，AI 叙事推动反弹至 $600"),
             ("③ 高位回落(增长减速)", "2024 增长从 20%+ 减速至 ~11%，估值从 35x 回落"),
             ("④ 持续下行(AI 竞争)", "2025 生成式 AI(Midjourney/Canva/Sora)+ Figma IPO 压制创意软件护城河叙事，全年下行"),
             ("⑤ CEO 卸任", "2026-03 Narayen 宣布卸任(继任未定)，单日跌约 23%"),
             ("⑥ CFO 离任+板块估值倍数下修", "2026-06 Q2 beat&raise 但 CFO 离任 + 全板块 AI agent 估值倍数下修，触及 7 年新低 $190")],
}, title="Adobe 月度股价 ($)")

# ════════════ 4. 卖方研报共识 ════════════
K.write_consensus(wb.create_sheet(S_CONS), {
    "title": "卖方研报共识 — 38 位分析师；目标价均值 $267.75(vs 现价 $205，+30.6%)",
    "overview": "评级 Buy(评分 1.78)：10 买/1 增持/25 持有/0 减持/2 卖。目标价区间 $190-$460，均值 $267.75。核心叙事分裂：多方(AI-ARR 加速+估值低位) vs 空方(AI 商品化 Creative Cloud+治理动荡)。",
    "assumptions": [
        ("收入增速\n(FY26E)", "街上共识 ~+11%(rev $26.5B)，与 Adobe 上调后指引 $26.5-26.6B 一致。", "分歧在 AI 对 Creative Cloud 量价 ARPU 的实质影响何时显现。", "base 取 +10-11%(有机)，与指引 ARR +10.2% 对齐；FY27+ 温和回落。"),
        ("Total ARR 增速", "Q2 FY26 +12.5%(含 Semrush)，有机 ~+10-11%；街上预期维持高单位数-低双位数。", "有机 ARR 能否守住 ~10% 是估值关键敏感项——跌破高单位数 = AI 替代风险兑现。", "base 有机 +10%、Semrush 并表贡献 ~$0.5B；bull AI-first ARR 成第二曲线。"),
        ("GAAP 营业利润率", "FY25 36.6%；共识 FY26E ~35-36%(Semrush 摊销+AI 投入小幅拖累)。", "AI 投入对毛利/获客成本的拖累程度。", "base 持稳 ~35.5%；bear 压缩至 30%(价格战+整合)。"),
        ("目标 P/E(GAAP)", "卖方隐含 forward P/E 11-20x(目标价÷FY26E EPS)，中位 ~15x。", "给一个正经历 AI 替代风险威胁 + 治理真空的 10% 增长者多少倍数。", "base 17x(GAAP)——10% 增长者理论 justified P/E ~17x，处自身历史下沿之下；bear 11x(价值陷阱)、bull 24x( fallen angel 复位)。"),
    ],
    "divergences": [
        "① AI 是否结构性替代 Creative Cloud：决定 ARR 有机增速能否守住 ~10%——这是 11.7x 是陷阱还是礼物的唯一裁决。",
        "② 治理风险定价：CEO 继任未定 + CFO 换人，市场要多少折扣；继任落地=催化剂。",
        "③ 板块 systemic 估值倍数下修 vs Adobe idiosyncratic：NOW -48%/CRM -9%/ADBE -37%(YTD)，Adobe 的折价里多少是板块、多少是自身。",
    ],
    "stances": [
        "Goldman(Sell，TP $220)：生成式 AI 侵蚀 Creative Cloud 护城河，评级里仅有的两家 Sell 之一。",
        "DA Davidson(Buy，TP $250)：Figma S-1 反证 Adobe 护城河，按 FY26 EPS 给 10x。",
        "RBC(Outperform，TP $350)：AI-first ARR 翻三倍说明 Firefly 在变现而非被替代。",
        "Stifel(Buy→Hold，TP $350→$200)：CFO 当日砍目标价 43%。",
    ],
})

# ════════════ 5. 历史估值倍数 ════════════
hm = K.write_hist_multiples(wb.create_sheet(S_HMULT), {
    "title": "历史估值倍数 — 自身 P/E·P/B 带 + 当下 + 同业对照",
    "intro": "数据底座：①自己历史上值多少(P/E 常态 33-57x，FY25 降至 20x，当下 11.7x——至少近十年低位)②现在市场给多少 ③同行(CRM/MSFT/NOW/INTU)+ 相对比值。看完再拍三案 P/E。注：Adobe P/B 因回购致 BPS 萎缩而失真，估值主线看 P/E。",
    "s_hist": S_HIST, "ha": ha, "hist_cols": HC, "hist_years": HY,
    "yhigh": YH, "ylow": YL,
    "fwd_note": "forward P/E ≈11x(GAAP $17.95)/8x(非GAAP $24.40)；forward P/B ≈7x(失真)",
    "self_name": "Adobe",
    "self_fwd_pe_label": "≈11x(GAAP)",
    "self_note": "TTM P/E 11.7x(GAAP)，处自身至少近十年低位——板块 AI 估值倍数下修 + CEO/CFO 双离任 + Creative Cloud AI 替代风险担忧叠加。",
    "peers": [
        {"name": "Microsoft (MSFT)", "yearly": [None, None, 36.0, 35.0, 33.0], "cur_pb": 11.0, "cur_pe": 33.0, "fwd_pe": 30.0,
         "note": "综合龙头，Copilot 全家桶切入文档/设计/营销；估值远高于 Adobe。"},
        {"name": "Salesforce (CRM)", "yearly": [None, None, 30.0, 28.0, 22.0], "cur_pb": 4.2, "cur_pe": 45.0, "fwd_pe": 22.0,
         "note": "CRM 龙头，增速放缓(8-10%)，营销云对位 Experience Cloud；2026 与 Adobe 同跌。"},
        {"name": "ServiceNow (NOW)", "yearly": [None, None, 84.0, 154.0, 59.0], "cur_pb": 7.5, "cur_pe": 59.0, "fwd_pe": 48.0,
         "note": "企业工作流 SaaS，增速 21% 远高于 Adobe，估值溢价最大。"},
        {"name": "Intuit (INTU)", "yearly": [None, None, 30.0, 28.0, 25.0], "cur_pb": 7.0, "cur_pe": 30.0, "fwd_pe": 25.0,
         "note": "大型应用软件，2026 与 Adobe 一同被抛售；财税/小企业 SaaS。"},
        {"name": "大盘应用软件中位", "yearly": None, "cur_pb": None, "cur_pe": None, "fwd_pe": 26.0,
         "note": "forward P/E 光谱中位 ~26x(参照)；Adobe 11x 显著偏低。"},
    ],
    "ratio": {"peer": "Intuit (INTU)",
              "note": "Adobe/INTU forward P/E 比 ~0.45x(INTU 同为大型应用软件、增速相近)——历史常态 ~1.0x，当前深度折价。修复路径取决于 AI 替代风险叙事能否被证伪。"},
    "reading": "① 自己：当前 11.7x 显著低于 FY22 低谷 33x 与 5 年均值 ~35x → 非周期底部可解释，含结构性折价。② 同行：Adobe forward 11x 远低于 MSFT(30)/NOW(48)/INTU(25)/CRM(22)，是可比里最便宜——市场已把 AI 替代风险 + 治理折价大部分定价。③ 相对 INTU 比值 0.45x(历史 ~1.0x)→ 若 AI 担忧证伪，重估空间大。→ 下一页 P/E 三案 11-24x。",
})

# ════════════ 6. 估值倍数假设（P/B 主线=kit 合规参考；P/E 平行镜头=主结论）════════════
ma = K.write_multiple_assumptions(wb.create_sheet(S_MULT), {
    "title": "估值倍数假设 — P/E 平行镜头(主结论) + P/B 主线(kit 合规，Adobe 失真仅参考)",
    "intro": "Adobe 是轻资产高利润 SaaS(净利率 30%、FCF margin ~40%)，市场用 P/E 定价。P/B 对 Adobe 失真：回购致 BPS 逐年萎缩(FY23 36.5→FY25 27.3)，P/B 被动抬升、不反映盈利能力。主结论走 P/E 平行镜头(目标 P/E × GAAP EPS → 隐含价)；P/B 主线仅作 kit 合规与资产视角参考。P/E 三案目标在『情景切换』页直接拍。",
    "why_text": ("镜头选择：Adobe 已盈利稳态(GM 89%、OpM 36%、FCF $9.9B)，且回购是 EPS 增长主驱动(BPS 萎缩)——"
                 "盈利视角(P/E)才是市场对 Adobe 的实际定价方式。P/B 对 Adobe 无意义：账面权益被回购冲减至 $11.5B，"
                 "P/B 7.3x 且将持续走高，反映的是资本回报政策而非估值高低。主结论走 P/E：目标 P/E × 前瞻 GAAP EPS → 隐含价。"
                 "P/B 镜头同步给出做形式上的三角(两个镜头差距极大 = 对 BPS 含义的认知分歧，轻资产+回购 SaaS 的正常现象)。"),
    "why_rows": 5,
    "method_text": "P/B 三层(kit 主线，仅参考)：①历史峰值 P/B ~22x(FY21)× ②结构溢价 0.35x(轻资产+回购致 BPS 萎缩)× ③情绪值。P/E 三案在『情景切换』页直接拍目标(不经三层，因 Adobe P/E 由增速+利润率+AI 替代风险风险+治理折价综合决定，非周期情绪单一驱动)。",
    "peak": 22.0, "peak_note": "FY21 P/B 峰值 ~22x(BPS $27.7 × 低价权益)。",
    "premium": 0.35, "premium_note": "轻资产 SaaS + 回购致 BPS 萎缩，结构溢价含义弱。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "hpb_row": ha["HPB"],
    "cases": [
        ("Bear", [0.32, 0.30, 0.28, 0.26, 0.25], "情绪退潮 + BPS 萎缩，P/B 继续走低。"),
        ("Base", [0.38, 0.36, 0.34, 0.32, 0.30], "温和，P/B 随 BPS 萎缩被动走高。"),
        ("Bull", [0.45, 0.42, 0.40, 0.38, 0.36], "AI 兑现 + 减速回购，P/B 企稳。"),
    ],
    "reconcile_text": "卖方目标价 $268 隐含 FY26E P/E ~15x(GAAP)。我们的 base FY27E P/E 为 17x(GAAP),高于卖方,因为市场对 AI 替代 Creative Cloud 的风险定价过高。Q2 FY26 ARR +12.5%、AI-first ARR 翻三倍、收入增速从 10% 升至 13%,均不支持'已出现明显替代'的判断。若 ARR 有机增速维持在 ~10%,justified P/E 约为 17x(10% 增长者理论值)。",
    "source_text": "P/B 历史峰值=『历史财务与估值』实际 P/B；同业=Yahoo/卖方；P/E 三案=本研究判断(增速+利润率+AI 风险+治理折价综合)。",
})

# ════════════ 7. 情景切换（SaaS 杠杆：分部增速 + OPM + 目标P/E）════════════
sw = K.write_scenario_switch(wb.create_sheet(S_SW), {
    "title": "情景切换 — SaaS 参数库(C&MP/BP&C 增速 + OPM + 目标P/E) + 开关(默认 Base)",
    "usage": ("B2 下拉切案 → 各杠杆『当前案』行跟切 → 分部收入→利润→EPS→隐含价 全链变档。"
              "三案并排见『估值对比』。Adobe 核心杠杆：①C&MP 增速(创意+营销云，~68%) ②BP&C 增速(Acrobat/Express，~27%) "
              "③综合营业利润率(Semrush 摊销+AI 投入 vs 运营杠杆) ④目标 P/E(主镜头，AI 替代风险风险+治理折价定价)。"),
    "cases": CASES, "default": "Base",
    "triggers": [
        ("Bear", "Creative Cloud 有机 ARR 增速跌破 ~6% + Canva/Figma 实质蚕食 + 治理真空延续 + 利润率压缩 → 估值停留在 11-12x P/E 价值陷阱区。"),
        ("Base", "AI 增量而未形成替代(ARR 有机守住 ~10%) + 利润率持稳 ~35% + 治理继任落地 → 估值从恐慌 11.7x 重估至 ~17x P_E(justified 区间)。"),
        ("Bull", "AI-first ARR 成第二曲线(占比→10%+) + Firefly/GenStudio 货币化兑现 + 新 CEO 重塑叙事 → 估值复位至 ~24x P/E。"),
    ],
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "levers": [
        {"key": "g_cmp", "name": "C&MP 增速(创意+营销云)", "fmt": K.PCT, "cols": FCf,
         "vals": {"Bear": [0.05, 0.04, 0.03, 0.02, 0.02], "Base": [0.10, 0.09, 0.085, 0.08, 0.08], "Bull": [0.12, 0.11, 0.10, 0.09, 0.09]},
         "desc": "Creative & Marketing Professionals：Creative Cloud + Document Cloud(专业) + Experience Cloud。占收入 ~68%，最大段。AI/Canva/Figma 竞争主战场。",
         "stories": {"Bear": "AI 抹平技能溢价 + Canva(免费 Affinity)蚕食 + Figma 吃协作，增速腰斩至近停滞。", "Base": "Adobe 嵌 AI(Firefly)守住工作流，有机 ~9-10%。", "Bull": "Firefly/GenStudio 提客单价，增速维持双位数。"},
         "hist": [None, 0.114, 0.111, 0.131, 0.109]},
        {"key": "g_bpc", "name": "BP&C 增速(Acrobat/Express)", "fmt": K.PCT, "cols": FCf,
         "vals": {"Bear": [0.06, 0.05, 0.04, 0.03, 0.03], "Base": [0.13, 0.12, 0.11, 0.10, 0.10], "Bull": [0.16, 0.15, 0.14, 0.13, 0.12]},
         "desc": "Business Professionals & Consumers：Acrobat AI Assistant + Express + 文档协作。占 ~27%，增速最快(Acrobat AI 货币化)。FY25 +15%(披露)。",
         "stories": {"Bear": "ChatGPT/Copilot 通用助手蚕食 PDF 场景，Acrobat AI 粘性不足，增速大幅回落。", "Base": "Acrobat AI Assistant 稳步渗透，+12-13%。", "Bull": "企业 PDF 工作流锁定 + AI 提价，+15%+。"},
         "hist": [None, 0.125, 0.133, 0.118, 0.140]},
        {"key": "opm", "name": "综合营业利润率(GAAP)", "fmt": K.PCT, "cols": FCf,
         "vals": {"Bear": [0.34, 0.31, 0.29, 0.28, 0.27], "Base": [0.355, 0.355, 0.355, 0.355, 0.355], "Bull": [0.37, 0.375, 0.38, 0.385, 0.39]},
         "desc": "GAAP 营业利润率。FY25 实际 36.6%。Semrush 摩销 + AI 投入 vs 运营杠杆 + 规模效应的平衡。Q2 FY26 净利仅 +1.2%(警示)。",
         "stories": {"Bear": "AI 成本侵蚀 + 价格战 + 整合成本，压缩至 27-29%。", "Base": "运营杠杆抵消 AI 投入，持稳 ~35.5%。", "Bull": "AI 提效 + 规模效应，扩张至 39%。"},
         "hist": [0.368, 0.346, 0.343, 0.313, 0.366]},
        {"key": "tpe", "name": "目标 P/E(GAAP，主镜头)", "fmt": K.MX, "cols": FCf,
         "vals": {"Bear": [11, 10.5, 10, 9.5, 9.5], "Base": [17, 17, 17, 17, 17], "Bull": [22, 23, 24, 25, 25]},
         "desc": "★ 主结论镜头：目标 P/E × 前瞻 GAAP EPS → 隐含价。10% 增长者理论 justified P/E ~17x；Bear=价值陷阱区(11x)、Bull=fallen angel 复位(24x，仍低于历史 35-50x)。当前 TTM 11.7x。",
         "stories": {"Bear": "AI 替代风险兑现 + 治理真空，永久去溢价至 11x。", "Base": "AI 增量未形成替代，重估至 justified 17x。", "Bull": "AI 货币化 + 治理复位，回到 24x(仍低于历史)。"},
         "hist": [57, 33, 50, 35, 20]},
    ],
    "linked": [
        {"key": "sent", "name": "情绪值(P/B 第三层)", "fmt": K.N2,
         "src_sheet": S_MULT, "src_row0": ma["sent_row0"],
         "note": "P/B 镜头情绪值(kit 主线用，Adobe P/B 失真仅参考)；P/E 主镜头不经此，P/E 三案直接在上方杠杆行。"},
    ],
})
# derived: 目标 P/B(当前案, kit 主线参考)
_pk = f"'{S_MULT}'!{ma['pk_cell']}"; _pr = f"'{S_MULT}'!{ma['pr_cell']}"
_sent_act = sw["SWACT"]["sent"]; _r = sw["next_row"]
K.lab(wb[S_SW], f"A{_r}", "目标 P/B(当前案, kit 主线参考)", b=True)
for _c in ALLC:
    K.fml(wb[S_SW], f"{_c}{_r}", f"={_pk}*{_pr}*{_c}{_sent_act}", K.MX, link=True)
K.logic(wb[S_SW], f"L{_r}", "= P/B 峰值 × 结构溢价 × 当前案情绪 → 喂情景估值 P/B 主线(参考镜头，Adobe 失真)。")
SWPB = _r

# ════════════ 8. 物理锚 [ANCHOR] — SaaS 订阅盘子 ════════════
anchor = K.write_anchor(wb.create_sheet(S_ANCHOR), {
    "title": "SaaS 物理锚 — Total Adobe ARR / AI-first ARR / RPO / DM-ARR(订阅盘子证据)",
    "all_cols": ALLC, "all_years": ALLY,
    "series": [
        ("Digital Media ARR ($B)", [12.5, 13.7, 15.0, 17.2, 19.2, None, None, None, None, None],
         "DM ARR(旧口径 KPI)：FY24 $17.22B / FY25 $19.20B(+11.5%)；FY21-23 估。FY26 起并入 Total Adobe ARR。", K.N1),
        ("Total Adobe ARR ($B)", [None, None, None, 22.6, 25.2, None, None, None, None, None],
         "Total Adobe ARR(新口径 KPI，FY26 起)：FY24 $22.61B / FY25 $25.20B(+11.5%) / Q2 FY26 $27.10B(+12.5%，含 Semrush $0.48B)。", K.N1),
        ("AI-first ARR ($B)", [0.0, 0.0, 0.0, 0.05, 0.17, None, None, None, None, None],
         "AI-first ARR(Firefly/GenStudio/Acrobat AI)：Q2 FY26 >$0.5B(同比×3)；FY25 ~$0.17B 估。AI 货币化证据。", K.N2),
        ("RPO ($B)", [None, None, 17.2, 20.0, 22.5, None, None, None, None, None],
         "Remaining Performance Obligation(已签未确认)：FY23 $17.22B / FY24 $19.96B / FY25 $22.52B / Q2 FY26 $22.27B。订单动能。", K.N1),
        ("大客户数(>$1M ARR) (千)", [1.30, 1.45, 1.60, 1.80, 2.00, None, None, None, None, None],
         "ARR>$1M 客户数(千，估)：FY25 ~2k+。大企业客户驱动高 ACV 与高留存。", K.N1),
    ],
    "source_note": "口径：ARR/RPO 来自 Adobe 10-K FY25 + Q2 FY26 10-Q/8-K 一手披露(EDGAR)；FY21-23 部分项不披露按趋势估。AI-first ARR、Total Adobe ARR 为 FY26 新增 KPI(分部合并后)。",
    "role_note": "作用：SaaS 订阅盘子证据行。C&MP 收入增速 = Total ARR 有机扩张(~10%) + Semrush 并表 + AI-first ARR 增量——驱动收入主力(68% 占比)的物理量。改『C&MP ARR 增速(当前案)』或情景 → C&MP 收入 → 净利 → EPS → 隐含价 全链动(连通性铁律落点)。",
})
# 连通性锚点：C&MP ARR 增速(当前案) 落在物理锚 sheet，前瞻由情景驱动；分部测算 C&MP 收入引此行
ANC_GCMP = anchor["row_of"]["大客户数(>$1M ARR) (千)"] + 2
_anc = wb[S_ANCHOR]
K.lab(_anc, f"A{ANC_GCMP}", "C&MP ARR 增速(当前案)", b=True)
for _c in FCf:
    K.fml(_anc, f"{_c}{ANC_GCMP}", f"={K.R(S_SW, _c + str(sw['SWACT']['g_cmp']))}", K.PCT, link=True)
for _i in range(1, len(HC)):
    _c, _p = HC[_i], HC[_i-1]
    K.fml(_anc, f"{_c}{ANC_GCMP}", f'=IFERROR({K.R(S_HIST, _c + str(ha["seg_rows"][SEG_CMP]))}/{K.R(S_HIST, _p + str(ha["seg_rows"][SEG_CMP]))}-1,"n.m.")', K.PCT, link=True)
K.logic(_anc, f"L{ANC_GCMP}", "C&MP 收入净增速 = Total ARR 有机扩张 + Semrush 并表 + 新增 logo(前瞻按情景，历史=实际反推)。分部测算 C&MP 收入引此行 → 物理锚进主链。")

# ════════════ 9. 分部测算（3 组，增速驱动）════════════
seg = K.write_segment_model(wb.create_sheet(S_SEG), {
    "title": "分部测算 — C&MP / BP&C / 产品与服务 三组(增速驱动)",
    "all_cols": ALLC, "all_years": ALLY, "logic_col": "N",
    "groups": [
        ("Creative & Marketing Professionals — 基本盘(创意+营销云)", [
            ("C&MP 增速", None, K.PCT, "历史=实际 YoY；前瞻=『物理锚』C&MP ARR 增速(当前案，引自情景切换 g_cmp)。占收入 ~68%。"),
            ("C&MP 收入 ($B)", None, K.N1, "历史取实数；前瞻=上年×(1+增速)。喂利润表。"),
        ]),
        ("Business Professionals & Consumers — 高增长段(Acrobat AI)", [
            ("BP&C 增速", None, K.PCT, "前瞻=『情景切换』g_bpc。占 ~27%，Acrobat AI Assistant 货币化驱动。"),
            ("BP&C 收入 ($B)", None, K.N1, "历史实数；前瞻=上年×(1+增速)。"),
        ]),
        ("产品与服务 — 衰减残值", [
            ("产品与服务 增速", [-0.10, -0.07, -0.05, -0.05, -0.04, -0.10, -0.08, -0.06, -0.05, -0.05], K.PCT,
             "非订阅(许可/培训/服务)，逐年衰减；三案共用(体量小 ~4%)。"),
            ("产品与服务 收入 ($B)", None, K.N1, "历史实数；前瞻=上年×(1+增速)。"),
        ]),
    ],
})
m = seg["m"]
CMP_HROW = ha["seg_rows"][SEG_CMP]
BPC_HROW = ha["seg_rows"][SEG_BPC]
PS_HROW = ha["seg_rows"][SEG_PS]
for col in HC:
    K.fml(wb[S_SEG], f"{col}{m['C&MP 收入 ($B)']}", f"={K.R(S_HIST, col + str(CMP_HROW))}", K.N1, link=True)
    K.fml(wb[S_SEG], f"{col}{m['BP&C 收入 ($B)']}", f"={K.R(S_HIST, col + str(BPC_HROW))}", K.N1, link=True)
    K.fml(wb[S_SEG], f"{col}{m['产品与服务 收入 ($B)']}", f"={K.R(S_HIST, col + str(PS_HROW))}", K.N1, link=True)
# 历史增速(YoY 反推；B 列无前值标 n.m.)
for _rev, _grow in [("C&MP 收入 ($B)", "C&MP 增速"), ("BP&C 收入 ($B)", "BP&C 增速"), ("产品与服务 收入 ($B)", "产品与服务 增速")]:
    K.lab(wb[S_SEG], f"B{m[_grow]}", "n.m.", note=True)
    for _i in range(1, len(HC)):
        _c, _p = HC[_i], HC[_i-1]
        K.fml(wb[S_SEG], f"{_c}{m[_grow]}", f'=IFERROR({_c}{m[_rev]}/{_p}{m[_rev]}-1,"n.m.")', K.PCT)
# 前瞻：C&MP 增速引物理锚；BP&C 引情景；产品与服务已在 series 给定值(覆盖前瞻列)
for col in FCf:
    K.fml(wb[S_SEG], f"{col}{m['C&MP 增速']}", f"={K.R(S_ANCHOR, col + str(ANC_GCMP))}", K.PCT, link=True)
    K.fml(wb[S_SEG], f"{col}{m['BP&C 增速']}", f"={K.R(S_SW, col + str(sw['SWACT']['g_bpc']))}", K.PCT, link=True)
    # 产品与服务 增速 已由 series 整行写入(含前瞻)；收入 = 上年×(1+增速)
for key_g, key_r in [("C&MP 增速", "C&MP 收入 ($B)"), ("BP&C 增速", "BP&C 收入 ($B)"), ("产品与服务 增速", "产品与服务 收入 ($B)")]:
    _prevs = [HC[-1]] + list(FCf[:-1])
    for _p, _c in zip(_prevs, FCf):
        K.fml(wb[S_SEG], f"{_c}{m[key_r]}", f"={_p}{m[key_r]}*(1+{_c}{m[key_g]})", K.N1)
for col in FCf:
    wb[S_SEG][f"{col}{m['C&MP 收入 ($B)']}"].fill = K.OUT
    wb[S_SEG][f"{col}{m['BP&C 收入 ($B)']}"].fill = K.OUT

# ════════════ 10. 利润与收入假设 ════════════
fr = K.write_fundamentals(wb.create_sheet(S_FUND), {
    "title": "利润与收入假设 — 分部营收(3组) → 综合OPM → 净利 → EPS(递减股本)/BPS",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
    "assum_groups": [
        ("利润率假设", [
            {"name": "综合营业利润率(GAAP)", "vals": [0.368, 0.346, 0.343, 0.313, 0.366, None, None, None, None, None],
             "fmt": K.PCT, "logic": "历史=10-K 实际 GAAP OP(EDGAR)；前瞻=『情景切换』当前案 opm。",
             "link": {"sheet": S_SW, "row": sw["SWACT"]["opm"]}},
            {"name": "净利转换率(净利/营业利润)", "vals": [0.831, 0.780, 0.816, 0.825, 0.819, 0.82, 0.82, 0.82, 0.82, 0.82],
             "fmt": K.PCT, "logic": "营业利润→净利(扣税~20%+净利息)；历史实际锚，前瞻 normalize 0.82。"},
            {"name": "留存率(权益增量/净利)", "vals": [0.32, 0.31, -0.43, -0.35, 0.10, 0.10, 0.10, 0.10, 0.10],
             "fmt": K.PCT, "logic": "Adobe 不派息、回购是资本回报主渠道。FY24-25 回购($9.5-11.3B)>净利致权益萎缩(负留存)；前瞻回购放缓至 ~$7B、权益微增(留存 ~0.10)。", "nm_cols": []},
            {"name": "稀释股本(mn，前瞻递减)", "vals": [481, 471, 459, 450, 427, 410, 403, 396, 389, 382],
             "fmt": K.N0, "logic": "★ 回购致股本逐年下降：历史实际(EDGAR 稀释)；前瞻 ~$7-8B/年回购(2026-04 新授权 $25B)→ 年降 ~1.7%。EPS 增长的重要驱动。"},
        ]),
    ],
    "segments": [
        {"name": SEG_CMP, "hist_row": SEG_CMP, "fwd": {"sheet": S_SEG, "row": m["C&MP 收入 ($B)"]}},
        {"name": SEG_BPC, "hist_row": SEG_BPC, "fwd": {"sheet": S_SEG, "row": m["BP&C 收入 ($B)"]}},
        {"name": SEG_PS, "hist_row": SEG_PS, "fwd": {"sheet": S_SEG, "row": m["产品与服务 收入 ($B)"]}},
    ],
    "profit_terms": [
        ([SEG_CMP, SEG_BPC, SEG_PS], "综合营业利润率(GAAP)", False),
    ],
    "conv_assum": "净利转换率(净利/营业利润)", "retention_assum": "留存率(权益增量/净利)",
    "note_text": "分部营收(3组增速驱动)→ 综合营业利润(总收入×GAAP OPM)→ 净利(×转换率 0.82)→ 权益(留存递推，回购致微增)→ EPS(净利÷递减股本)/BPS。历史列取 10-K 实际；前瞻 EPS 用『稀释股本(前瞻递减)』行 override(捕捉回购驱动的 EPS 杠杆)。P/B 因 BPS 萎缩失真，估值主线看 P/E。",
})
# ★ override 前瞻 EPS/BPS 用递减股本(捕捉回购杠杆)
SHR_ROW = fr["am"]["稀释股本(mn，前瞻递减)"]
for _c in FCf:
    K.fml(wb[S_FUND], f"{_c}{fr['EPS']}", f"={_c}{fr['NI']}*1000/{_c}{SHR_ROW}", K.N2)
    K.fml(wb[S_FUND], f"{_c}{fr['BPS']}", f"={_c}{fr['EQ']}*1000/{_c}{SHR_ROW}", K.N2)

# ════════════ 11. 情景估值（P/E 平行镜头=主结论；P/B 主线=参考）════════════
sv = K.write_scenario_valuation(wb.create_sheet(S_VAL), {
    "title": "情景估值 — 当前案逐年隐含价(P/E 平行镜头=主结论；P/B 主线=参考)",
    "intro": "本表=『情景切换』当前案(默认 Base)。P/B 主线(kit)=目标P/B×BPS→隐含价(Adobe 回购致 BPS 萎缩、P/B 失真，仅参考)；P/E 平行镜头=目标P/E×GAAP EPS→隐含价(主结论，SaaS 市场定价方式，EPS 含回购递减股本杠杆)。历史列=实际价反推倍数；前瞻不拟合现价。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf, "hist_years": HY,
    "fx_fwd": FX, "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
    "s_fund": S_FUND, "fr": fr,
    "s_switch": S_SW, "target_row": SWPB, "sw_cell": "B2",
    "yend": YE, "yavg": YA,
    "pe_lens": {"target_row": sw["SWACT"]["tpe"], "mcap_label": "市值 前瞻·P/E镜头($B)"},
    "reading": "P/E 镜头读法：『当下 forward P/E』= 现价÷模型各年 GAAP EPS(光谱里 Adobe 的出处，11x)；『P/E 前瞻隐含』= 目标P/E×EPS。P/B 与 P/E 两镜头差距极大 = 对 BPS 含义的认知分歧(轻资产+回购 SaaS 的正常现象，Adobe P/B 无意义)。",
    "method": "方法：P/E 平行镜头逐年估(主结论)。目标 P/E(情景切换当前案) × 前瞻 GAAP EPS(含递减股本) → 隐含价。基本面在『利润与收入假设』；目标 P/E 在『情景切换』。",
    "concl": "结论(方向性)：三情景见『估值对比』；当前 $205 处自身历史 P/E 低位，base 隐含价显著高于现价=市场过度定价 AI 替代风险风险，但需 ARR 守住 + 治理继任兑现。",
})

# ════════════ 12. 估值对比（三案 P/E 链为主）════════════
SWB = sw["SWB"]
SH_F = K.R(S_HIST, f"$F${ha['HSH']}")
PX_NOW_REF = K.R(S_HIST, f"G{ha['HPX']}")
_opm = fr["am"]["综合营业利润率(GAAP)"]
_conv = fr["am"]["净利转换率(净利/营业利润)"]
_ret = fr["am"]["留存率(权益增量/净利)"]

def _fwdprev(j, A, key):
    return (HC[-1] if j == 0 else FCf[j - 1]) + str(A[key])

cmp_rows = [
    {"key": "cmp", "label": "C&MP ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(CMP_HROW))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'cmp')}*(1+{K.R(S_SW, c + str(SWB['g_cmp'] + ci))})"},
    {"key": "bpc", "label": "BP&C ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(BPC_HROW))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'bpc')}*(1+{K.R(S_SW, c + str(SWB['g_bpc'] + ci))})"},
    {"key": "ps", "label": "产品与服务 ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(PS_HROW))}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_SEG, c + str(m['产品与服务 收入 ($B)']))}"},
    {"key": "rev", "label": "总收入 ($B)", "fmt": K.N1, "bold": True,
     "hist": lambda c, ci, A: f"={c}{A['cmp']}+{c}{A['bpc']}+{c}{A['ps']}",
     "fwd": lambda c, j, ci, A: f"={c}{A['cmp']}+{c}{A['bpc']}+{c}{A['ps']}"},
    {"key": "op", "label": "营业利润 ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={c}{A['rev']}*{K.R(S_FUND, c + str(_opm))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['rev']}*{K.R(S_SW, c + str(SWB['opm'] + ci))}"},
    {"key": "ni", "label": "净利 ($B)", "fmt": K.N1, "bold": True,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HNI']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['op']}*{K.R(S_FUND, c + str(_conv))}"},
    {"key": "eq", "label": "期末权益 ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HEQ']))}",
     "fwd": lambda c, j, ci, A: (f"={K.R(S_HIST, 'F' + str(ha['HEQ']))}+{c}{A['ni']}*{K.R(S_FUND, c + str(_ret))}" if j == 0
                                 else f"={FCf[j-1]}{A['eq']}+{c}{A['ni']}*{K.R(S_FUND, c + str(_ret))}")},
    {"key": "eps", "label": "EPS GAAP ($)", "fmt": K.N2, "bold": True,
     "hist": lambda c, ci, A: f"={c}{A['ni']}*1000/{K.R(S_HIST, c + str(ha['HSH']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['ni']}*1000/{K.R(S_FUND, c + str(SHR_ROW))}"},
    {"key": "tpe", "label": "目标 P/E(GAAP，该案)", "fmt": K.MX,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HPE']))}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_SW, c + str(SWB['tpe'] + ci))}"},
    {"key": "px", "label": "隐含价 P/E镜头 ($)", "fmt": K.PX, "bold": True, "out": True,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HPX']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['tpe']}*{c}{A['eps']}"},
    {"key": "ipe", "label": "隐含 forward P/E(交叉)", "fmt": K.MX,
     "hist": lambda c, ci, A: f'=IF({c}{A["ni"]}<=0,"N/M",{K.R(S_HIST, c + str(ha["HPX"]))}/{c}{A["eps"]})',
     "fwd": lambda c, j, ci, A: f"={c}{A['px']}/{c}{A['eps']}"},
    {"key": "up", "label": "历史:vs年末价(回测≈0)/前瞻:vs现价", "fmt": K.PCT,
     "hist": lambda c, ci, A: f"={c}{A['px']}/{K.R(S_HIST, c + str(ha['HPX']))}-1",
     "fwd": lambda c, j, ci, A: f"={c}{A['px']}/{PX_NOW_REF}-1"},
]
cm = K.write_comparison(wb.create_sheet(S_CMP), {
    "title": "估值对比 — Bear/Base/Bull 三案目标价并排(P/E 主镜头，GAAP)",
    "intro": ("三案各自完整推演：分部增速(3组)→总收入→营业利润(×OPM)→净利(×转换率)→EPS(÷递减股本)→目标P/E→隐含价。"
              "三案恒常并排。历史列 2021-2025 = 实际值，隐含价历史列=实际年末价(回测行历史列≈0%，链的内置回测)。"
              "主结论=P/E 镜头隐含价；P/B 镜头见『情景估值』作参考(Adobe 失真)。"),
    "case_names": CASES,
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "block_start": 18,
    "rows": cmp_rows,
    "summary": {
        "band": "三案汇总 (目标年 2027E；触发条件见『情景切换』)",
        "target_col": "H",
        "rows": [
            ("总收入($B)", "rev", K.N1, "= C&MP+BP&C+产品与服务 三组增速驱动", False),
            ("净利($B)", "ni", K.N1, "= 总收入×GAAP OPM×净利转换率", False),
            ("EPS GAAP ($)", "eps", K.N2, "= 净利÷递减股本(回购杠杆)", False),
            ("目标 P/E(GAAP)", "tpe", K.MX, "= 该案 P/E(AI 替代风险风险+治理折价)", False),
            ("隐含价 P/E镜头 ($)", "px", K.PX, "= 目标P/E × EPS(主结论)", True),
            ("隐含 forward P/E", "ipe", K.MX, "= 隐含价÷EPS，交叉验证", False),
            ("vs 现价 $205", "up", K.PCT, "= 隐含价÷$205.02−1", True),
        ],
        "mcap": {"label": "隐含市值($B)", "key": "px", "expr": f"*{SH_F}/1000",
                 "note": "= 隐含价 × 股本(前瞻递减)"},
        "concl": "风险收益比(2027E vs $205)：base 隐含价上行空间取决于 ARR 守住 + 治理继任；Bear 下行风险=AI 实质蚕食+估值陷阱；Bull=AI 货币化兑现+fallen angel 复位。详见决策 memo。",
    },
})

# ════════════ 13. 综合判断仪表盘 ════════════
EPS27 = K.R(S_FUND, "H" + str(fr["EPS"]))
PXD = K.R(S_HIST, "G" + str(ha["HPX"]))
dash = K.write_dashboard(wb.create_sheet(S_DASH), {
    "title": "综合判断仪表盘 — A 基本面拐点 · B 估值错位(预测引擎) · C 催化剂 · D 情绪确认",
    "usage": "预测引擎=B(错位)+C(催化剂)，当下可观测；D 情绪只做 timing 确认。验收=回测：放回 2022 低点($275)或 2025 高点($610)，这套表当时能看到。",
    "blocks": [
        {"title": "A. 基本面拐点 — AI 是形成替代还是带来增量?", "rows": [
            ("ARR 有机增速", "Total ARR Q2 FY26 +12.5%(含 Semrush)；有机 ~+10-11%，未减速", "关键敏感项：有机 ARR 守住 ~10% = 尚未出现明显替代"),
            ("AI-first ARR", ">$0.5B(同比×3)，Firefly 240 亿张图/MAU 600 万+", "AI 货币化证据，但占总 ARR <2%(基数小)"),
            ("收入增速", "季度加速：10%→11%→11%→12%→13%(Q1FY24→Q2FY26)", "减速未在收入端显现；Q2 净利仅 +1.2% 是利润端警示"),
            ("A 判断", "【中】", "基本面未崩(ARR 加速)，但利润端压力 + 治理真空是悬顶之剑。", True),
        ]},
        {"title": "B. 估值错位(预测引擎 ★)— 市场给 vs 基本面该给 → GAP", "rows": [
            ("市场现在给(forward P/E GAAP)", {"fml": f"={PXD}/{EPS27}", "fmt": K.MX, "fill": True},
             "= 现价 ÷ 2027E GAAP EPS(公式算)。"),
            ("基本面该给(justified forward P/E)", {"inp": 17.0, "fmt": K.MX},
             "= 10% 增长者理论 justified P_E ~17x(r=10%,g=4%；Adobe FCF margin 40%+ 支撑)。"),
            ("错位 GAP = 该给÷市场给 − 1",
             {"fml": lambda ro: f"=B{ro['基本面该给(justified forward P/E)']}/B{ro['市场现在给(forward P/E GAAP)']}-1", "fmt": K.PCT},
             "GAP 正且大=重估空间(该买)；转负=已透支。当前 forward ~11x < 该给 17x = 显著上行空间(若 ARR 守住)。"),
            ("回测: 2021 高点读数", "当时 forward P/E ~50x vs 该给 ~35x → GAP 转负 → 那波从 $670 回落的预测依据", "现价已深度反向错位。"),
        ]},
        {"title": "C. 催化剂 — 什么逼市场闭合 GAP", "rows": [
            ("CEO 继任落地", "待(最高优先级)", "新 CEO 公布 + AI 战略清晰化 = 最大单一催化剂，解除治理折价。"),
            ("Q3/Q4 FY26 财报", "✓ 进行中", "Total ARR 有机增速 + AI-first ARR 披露：守住 ~10% = 证伪替代风险判断。"),
            ("板块估值回暖", "待", "NOW/CRM 等止跌 + 利率下行带动 SaaS 估值修复。"),
            ("C 判断", "部分兑现", "财报是近 catalyst；CEO 继任是行情级别催化剂。", True),
        ]},
        {"title": "D. 情绪确认 — timing + 刹车（定性档位）", "rows": [
            ("恐慌位置", "从 2024 高 $635 跌 68% 至 $190(7 年新低)", "板块 AI 估值倍数下修 + 双高管离任，恐慌释放中。"),
            ("现价 vs 基本面该给", "forward 11x 远低于该给 17x", "深度反向错位，非合理偏低而是恐慌定价。"),
            ("当前档位", "【退潮末期/恐慌底部】", "下行恐慌接近尾声，反转需 CEO 继任 + ARR 守住 catalyst。", True),
            ("衰减扳机", "Total ARR 有机增速跌破 ~8% / 营业利润率环比转跌 / CEO 继任再拖延", "任一翻→降档+P/E 目标下调至 Bear。"),
        ]},
    ],
    "final": {"band": "★ 综合判断(A+B+C+D)",
              "text": "基本面未崩(A，ARR 加速)+估值深度反向错位(B，forward 11x vs 该给 17x)+催化剂部分兑现(C，财报+CEO 继任)+情绪恐慌底部(D)→ 风险收益比偏多但非无脑重仓；base 隐含价显著高于现价，建议逢 catalyst(CEO 继任落地/Q3 ARR 守住)分批建仓，Bear 档作止损参照。核心赌注：ARR 有机增速守住 ~10%(证伪 AI 替代风险叙事)。"},
    "tracking": {
        "intro": "哪个指标恶化 → 哪个假设先崩 → 触发什么动作（盯的优先级）。",
        "rows": [
            ("__band__", "一、核心驱动链（ARR 有机增速 = 估值关键敏感项）"),
            ("Total ARR 有机增速", "~10-11%(Q2 FY26)", "关键敏感项:AI 是否正在替代 Creative Cloud", "季报 8-K 披露 + Semrush 剥离", "跌破 ~8% → 全链下调、转 Bear"),
            ("AI-first ARR", ">$0.5B(同比×3)", "关键敏感项：AI 货币化能否成第二曲线", "季报 segment 披露", "增速跌破 ~30% → 下调 g_cmp/Bull 权重"),
            ("__band__", "二、利润端（Q2 FY26 净利仅 +1.2% 警示）"),
            ("GAAP 营业利润率", "36.6%(FY25)", "关键敏感项：Semrush 摊销+AI 投入 vs 运营杠杆", "季报", "环比转跌 → 下调 opm 路径"),
            ("__band__", "三、治理与估值"),
            ("CEO 继任", "Narayen 3/12 宣布卸任，未定", "关键敏感项：治理折价 + AI 战略清晰度", "公司公告/8-K", "继任落地 = 上调情绪值；拖延 = 维持折价"),
            ("forward P/E(GAAP)", "~11x", "关键敏感项：板块 systemic 估值倍数下修", "股价/一致预期", "NOW/CRM 继续破位 → P/E 目标下调"),
        ],
    },
})

# ════════════ 全局格式 + 落盘 ════════════
K.finalize(wb, freeze={
    S_HIST: "B3", S_PX: "B4", S_CONS: "A2", S_HMULT: "B5", S_MULT: "B4", S_SW: "B3",
    S_ANCHOR: "B3", S_SEG: "B3", S_FUND: "B3", S_VAL: "B4", S_CMP: "B6", S_DASH: "B6",
    S_COVER: "A2",
})
out = os.path.join(os.path.dirname(__file__), "..", "out", "ADBE_valuation_model.xlsx")
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("saved:", os.path.abspath(out))
print("sheets:", wb.sheetnames)
print("PX_NOW:", PX_NOW)
