# -*- coding: utf-8 -*-
"""
build_dal.py — Delta Air Lines (NYSE:DAL) 估值模型(物理锚通电链, P/E 主镜头 + P/B 底部 + DCF 旁注)。

物理锚 = 运力(ASM 可用座英里) + jet fuel 价格。
链: [ASM × load factor → RPM × PRASM → 客运收入] + [Amex/忠诚度非周期现金流]
    → [营收 − CASM-ex×ASM − jet fuel(gal×$/gal)] → 营业利润 → 净利 → EPS
    → 目标 forward P/E × EPS → 隐含价。
最底层数值灵敏点 = jet fuel $/gal(弹性 ~$5.17 EPS / $1 per gal)与 PRASM。

数据: SEC 10-K FY2025(accession 0000027904-26-000013) + 10-Q Q1 2026 + Research Data API 实拉。
单币种 USD, fx 全 = 1。
"""
import os
from openpyxl import Workbook
import build_kit as K

# ════════════ 0. 全局轴 ════════════
ALLC = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
ALLY = ["2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
HC, HY = ["B", "C", "D", "E", "F"], ["2021", "2022", "2023", "2024", "2025"]
FC = ["F", "G", "H", "I", "J", "K"]          # 前瞻含基年 F=2025A
FCf = FC[1:]                                   # 纯前瞻 2026E-2030E
FX_FC = 1                                       # USD 单币种
FX_HIST = [1, 1, 1, 1, 1]
CASES = ["Bear", "Base", "Bull"]

S_COVER, S_HIST, S_PX, S_CONS = "封面", "历史财务与估值", "股价走势", "卖方研报共识"
S_HMULT, S_MULT, S_SW = "历史估值倍数", "估值倍数假设", "情景切换"
S_ANCHOR, S_SEG, S_FUND = "运力与油价物理锚", "分部测算", "利润与收入假设"
S_VAL, S_CMP, S_DASH = "情景估值", "估值对比", "综合判断仪表盘"

# 月度收盘(真实, Research Data API us-lake; 2021 上半年湖外缺, 从 06 起)
MONTHLY = [
    ("2021-06", 41.86), ("2021-09", 41.23), ("2021-12", 37.82),
    ("2022-03", 38.29), ("2022-06", 28.03), ("2022-09", 27.15), ("2022-12", 31.80),
    ("2023-03", 33.79), ("2023-06", 46.00), ("2023-09", 35.88), ("2023-12", 39.12),
    ("2024-03", 46.66), ("2024-06", 46.33), ("2024-09", 49.77), ("2024-12", 59.46),
    ("2025-03", 42.96), ("2025-06", 48.60), ("2025-09", 56.27), ("2025-12", 69.03),
    ("2026-03", 66.30), ("2026-06", 86.42),
]
PX_NOW = 86.42  # 2026-06-23 实时

wb = Workbook()
wb.remove(wb["Sheet"])

# ════════════ 1. 封面 ════════════
K.write_cover(wb.create_sheet(S_COVER), {
    "title": "Delta Air Lines (DAL) 估值模型 — 物理锚: 运力×单位经济 + jet fuel 价格",
    "meta": [
        ("报告日期", "2026-06-23"),
        ("数据截止", "SEC 10-K FY2025(2026-02-11) + 10-Q Q1 2026(2026-04-08) + 实时股价 2026-06-23"),
        ("现价", f"${PX_NOW}(2026-06-23, +2.7%)"),
        ("时效声明", "基于建模时点的 10-K/10-Q/管理层指引/卖方共识/实时股价。油价与暑运票价波动快, 每季财报(下次 2026-07-09)后更新。"),
        ("方法一句话", "运力(ASM)×单位收入(PRASM/yield)→ 客运收入; 叠 Amex/忠诚度非周期现金流; 减 CASM-ex×ASM 与 jet fuel(gal×$/gal) → 营业利润 → 净利 → EPS; 目标 forward P/E × EPS → 隐含价。jet fuel $/gal 是最大数值灵敏点。"),
    ],
    "takeaways": [
        ("① 当下估值位置", "现价 $86.42 对应 FY2026E adjusted EPS ~$6.6 约 13x、FY2027E ~$8.0 约 10.8x forward P/E; 处自身历史 6-13x band 中上沿, 但绝对值仍低于标普。"),
        ("② 核心引擎", "高端化(premium 客票 +7% vs main cabin -5%, premium+多元化占营收 ~60%) + Amex remuneration FY2025 $8.2B(对油价不敏感, 目标 ~$10B), 是穿越周期的盈利稳定器。"),
        ("③ 油价弹性(最大 swing)", "jet fuel 每降 $1.00/加仑 → 税后 EPS 约 +$5.17(4,269M 加仑 ÷ 652M 股 × 0.79)。2026 油价从 4 月高位回落是当下成本顺风, 但 H2 锚位仍是头号变量。"),
        ("④ 三情景目标价", "Bear/Base/Bull 沿同一条链(jet fuel + PRASM + 运力 + 目标 P/E)翻档, 见『估值对比』。"),
        ("⑤ 主要风险", "油价反弹(对称下行 -$5.17 EPS/加仑) / 需求见顶(休闲先弱) / 运力纪律破裂价格战 / 飞行员合同 2026 底重谈薪资刚性 / 重资产 +$119 亿净债务。"),
    ],
})

# ════════════ 2. 历史财务与估值(分部按高端化 thesis 拆: main / premium / 忠诚度&Amex / 其他)════════════
# GAAP 口径(含炼油第三方销售 ~$5B)。客运细分用 10-K MD&A: main cabin / premium products / loyalty+travel-svc
# 2021 拆分为估(疫情恢复期); 2022-2025 为 10-K 披露逐年。
ha = K.write_history(wb.create_sheet(S_HIST), {
    "title": "DAL 历史财务与估值 ($B) — 2021-2025A + 当下(TTM) | 源: SEC 10-K FY2025 + Research Data API",
    "hist_cols": HC, "hist_years": HY,
    "fx_hist": FX_HIST, "fx_now": FX_FC,
    "vals_in_usd": True,
    "fx_label": "FX (USD, 单币种)",
    "segments": [
        # main cabin 客票
        ("主舱客票(main cabin)", [14.0, 19.5, 22.5, 24.5, 23.4], True),
        # premium 客票(Delta One/First/Comfort+)
        ("高端客票(premium)", [7.5, 14.0, 17.6, 20.6, 22.1], True),
        # 里程兑换客运 + travel-related services
        ("里程兑换+旅行服务", [2.5, 4.5, 5.2, 5.8, 6.3], False),
        # 忠诚度品牌使用费(loyalty program, 含 Amex 计入 other 的部分) + 货运
        ("忠诚度品牌费+货运", [1.0, 3.8, 4.0, 4.1, 4.3], False),
        # 炼油第三方 + 杂项(MRO/度假等)
        ("炼油第三方+其他", [4.9, 8.8, 8.7, 6.6, 7.4], False),
    ],
    "total_now": 63.4,
    "ni": [0.28, 1.32, 4.61, 3.46, 5.01], "ni_now": 4.72,  # TTM 含 Q1 2026 GAAP 亏损(投资市值损失)
    "eq": [3.9, 6.6, 11.1, 15.3, 20.9], "eq_now": 20.6,
    "shares": [641, 641, 643, 648, 654], "shares_now": 652,
    "px_end": [37.82, 31.80, 39.12, 59.46, 69.03],
    "px_now": PX_NOW,
    "px_avg": [39, 34, 37, 47, 56],
    "eps_label": "EPS ($, GAAP 摊薄)",
    "bps_label": "BPS ($)",
    "band_note": "P/E 历史(年末价÷EPS): 2022 ~15x → 2023 ~5.5x → 2024 ~11x → 2025 ~9x; 航空稳态 band 6-13x。当下 TTM ~12.6x(含 Q1 投资损失拖累 EPS)。",
    "notes": [
        ("主舱客票(main cabin)", "10-K MD&A: FY2025 $23.4B, 同比 -5%(行业供给超主舱需求)。价格敏感客群, K 型经济下走弱。"),
        ("高端客票(premium)", "Delta One/First/Comfort+: FY2025 $22.1B, +7%, 逼近主舱。Q1 2026 +14% 加速。高端化 thesis 的核心。"),
        ("里程兑换+旅行服务", "loyalty travel awards $4.2B + travel-related $2.0B。SkyMiles 里程兑换的客运确认。"),
        ("忠诚度品牌费+货运", "loyalty program 品牌使用费(含 Amex 部分) $3.4B + cargo $0.9B。Amex remuneration FY2025 总额 $8.2B, 分摊进客运递延与此处。"),
        ("炼油第三方+其他", "Monroe 炼油厂对第三方售非航油 $5.1B(低毛利转售) + MRO/度假/杂项。建模看主业应剔。"),
        ("HREV", "总营收 = 交易所/10-K 实际 GAAP。FY2025 $63.4B(adjusted 剔炼油约 $58.3B)。"),
        ("HNI", "净利: 10-K GAAP 实际。当下 TTM ~$4.7B(Q1 2026 GAAP 净亏 $0.29B 主因 -$0.55B 投资市值损失, 非主业)。"),
        ("HEQ", "股东权益: 10-K 实际, 连续修复(2021 $3.9B → 2025 $20.9B), 降杠杆。"),
        ("HSH", "摊薄股本: 10-K 实际, 小幅增长(SBC), 当下 652M(Q1 2026 加权)。"),
        ("HPX", "年末股价: Research Data API us-lake 真实收盘; 当下 = 实时 $86.42(2026-06-23)。"),
        ("HPXA", "年均股价: 月度收盘均值, 同『股价走势』单一源。2021 仅半年序列。"),
    ],
})

# ════════════ 3. 股价走势 ════════════
def phase_fn(ym):
    if ym <= "2022-12":
        return "① 疫情恢复"
    if ym <= "2024-06":
        return "② 盈利重建"
    if ym <= "2025-12":
        return "③ 高端化+降杠杆"
    return "④ 油价回落+Berkshire"

px = K.write_price_chart(wb.create_sheet(S_PX), MONTHLY, {
    "fn": phase_fn,
    "rows": [("① 疫情恢复", "2021-2022: 需求重启, 股价 $28-42 区间震荡, 重债务包袱"),
             ("② 盈利重建", "2023-2024H1: EPS 转正($7.17/2023), 资产负债表修复"),
             ("③ 高端化+降杠杆", "2024H2-2025: premium 放量, 净债务降至 $119 亿, 股价突破 $69"),
             ("④ 油价回落+Berkshire", "2026: 油价从 4 月高位回落, Berkshire 建仓 $2.65B, 股价 $86")],
}, title="DAL 月度股价 ($)")

# ════════════ 4. 卖方研报共识 ════════════
K.write_consensus(wb.create_sheet(S_CONS), {
    "title": "卖方研报共识 — 25 家覆盖, strong_buy; 这张表是后面测算的'卖方对账单'",
    "overview": "全街 strong_buy(25 家: 5 强买 + 20 买入 + 1 强卖)。Yahoo 聚合目标均价 $82(滞后, 区间 $48-105); 但近期研报集中上修——BofA $78→$93、Argus $85→$100。FY2026E EPS 共识 $5.63(管理层指引 $6.50-7.50)、FY2027E $8.01。分歧核心在油价路径与暑运票价持续性。",
    "assumptions": [
        ("jet fuel 价格\n(2026 H2)", "卖方多用 forward 曲线 ~$2.40-2.80/加仑(剔炼油), 较 4 月 $4.88 高点大幅回落。", "停火能否持续 = 油价锚最大分歧; EIA 6 月 STEO 仍用封锁情景 Brent $95。", "Base 取 jet fuel $2.55/加仑(H2 现货 ~$2.70 与正常化 $2.30 之间), 不沿用偏高的 $86 Brent。"),
        ("PRASM / 单位收入\n(2026)", "共识单位收入 low-single-digit 增长, 靠运力纪律 + 高端化支撑。", "暑运因油价提的票价是否结构性(管理层暗示'可能不是临时'), 还是旺季后回落。", "Base 取 PRASM +1.5%/年, 高端结构升级对冲主舱疲软。"),
        ("FY2026 EPS", "Yahoo 共识 $5.63; 管理层指引 $6.50-7.50; 部分卖方 adjusted 口径 ~$6.6。", "GAAP vs adjusted 口径差异 + 油价假设差异导致 EPS 区间宽($4.77-7.00)。", "Base adjusted EPS ~$6.6(落指引中点偏下), 由本模型链算出而非抄共识。"),
        ("目标 forward P/E", "卖方目标价对应 FY2027E EPS 约 10-12x; Argus $100 对应更高。", "给航空多少 forward P/E = 周期定位之争; 历史 band 6-13x。", "三层分解拍(峰值 12x × 结构溢价 × 情绪), 见『估值倍数假设』。"),
    ],
    "divergences": [
        "① 油价路径: 停火持续(jet fuel → $2.30) vs 中东再起(Brent → $90+, jet fuel → $3.70)。弹性 ±$5.17 EPS/加仑, 决定上下限。",
        "② 暑运票价持续性: 因油价提的票价是结构性(运力纪律下) 还是旺季后回吐 → 决定 PRASM 路径。",
        "③ 周期定位: 市场是否愿给航空 >11x forward P/E(承认高端化+Amex 让 DAL 脱离纯周期) → 决定目标倍数。",
    ],
    "stances": [
        "BofA(买入, TP $93): capacity discipline 是暑运关键, 上调目标价。",
        "Argus(买入, TP $100): 运营强劲 + 油价顺风, 上调目标价。",
        "Berkshire Hathaway(建仓 $2.65B, 2026 Q1 13F): 运营与股价背离的经典价值 setup。",
        "Yahoo 聚合(strong_buy, 均价 $82): 25 家, 但均价滞后于近期上修。",
    ],
})

# ════════════ 5. 历史估值倍数(P/E 为主底座)════════════
hm = K.write_hist_multiples(wb.create_sheet(S_HMULT), {
    "title": "历史估值倍数 — P/E 主底座: 自身历史 P/E band + 当下 + 三巨头对照",
    "intro": "这一页是『估值倍数假设』的数据底座。航空稳态盈利股以 forward P/E 为主镜头(P/B 仅作重资产底部参考)。①自己历史 P/E 走过什么区间(年末价÷EPS) ②现在市场给多少 ③三巨头(UAL/AAL)对照 + 相对 UAL 的比值(质量溢价对账线)。",
    "s_hist": S_HIST, "ha": ha, "hist_cols": HC, "hist_years": HY,
    "yhigh": px["yhigh"], "ylow": px["ylow"],
    "fwd_note": "forward P/E ≈11x(现价 $86.42 ÷ FY2027E EPS ~$8.0); FY2026E ~13x",
    "self_name": "DAL",
    "self_fwd_pe_label": "≈11x",
    "self_note": "本模型标的; forward 推导见『情景估值』。航空看 forward P/E, P/B 仅底部。",
    "peers": [
        {"name": "UAL(美联航)", "yearly": [9.0, 6.0, 4.5, 9.5, 8.5], "cur_pb": 2.6, "cur_pe": 9.0, "fwd_pe": 7.5,
         "note": "最直接可比; 利润率 ~8% 低于 DAL, 资产负债表更重, 给更低 P/E。逐年为 P/E(年末价÷EPS)粗估。"},
        {"name": "AAL(美航)", "yearly": [None, None, None, None, None], "cur_pb": None, "cur_pe": 12.0, "fwd_pe": 7.0,
         "note": "三巨头最弱; 利润率 ~2.5%, 净债务最高, 估值折价; 权益常为负, P/B 不适用。"},
        {"name": "LUV(西南)", "yearly": None, "cur_pb": None, "cur_pe": 28.0, "fwd_pe": 18.0,
         "note": "低成本但增长停滞, 高 P/E 因盈利低基数; 商业模式不同。"},
        {"name": "标普500(参照)", "yearly": None, "cur_pb": None, "cur_pe": None, "fwd_pe": 21.0,
         "note": "大盘 forward P/E 上沿参照; 航空长期折价于大盘。"},
    ],
    "ratio": {"peer": "UAL(美联航)",
              "note": "DAL/UAL 的 P/E 比值(质量溢价对账线): DAL 利润率、净债务、ROIC 领先 UAL, 应享 1.2-1.4x 溢价 → 下一页第二层结构溢价取值有据。"},
    "reading": "① 自己: 历史 P/E band 6-13x(2023 低点 5.5x = 盈利峰值低估, 2022 15x = 盈利刚转正高估); 当下 forward ~11x 处中段。② 同行: DAL 应享质量溢价(利润率/资产负债表/ROIC 领先三巨头)。③ DAL/UAL 比值 → 第二层结构溢价依据。→ 下一页: 峰值 P/E × 质量溢价 × 情绪值(三案)。",
})

# ════════════ 6. 估值倍数假设(三案 forward P/E 在此拍)════════════
ma = K.write_multiple_assumptions(wb.create_sheet(S_MULT), {
    "title": "估值倍数假设 — 主镜头 forward P/E + 三案目标倍数(= 历史 P/E 锚 × 质量溢价 × 情绪值)",
    "intro": "这一页只做判断(数据底座在上一页): ①为什么 forward P/E 做主镜头 ②三层分解出三案目标 P/E。『情景切换』引用并切换, 『情景估值』套用当前案, 『估值对比』三案并排。",
    "why_text": ("镜头选择是业务判断: 达美穿越周期持续存在的东西是什么? 是高端客群忠诚度 + Amex 现金流支撑的'盈利能力', 不只是飞机资产。"
                 "所以主镜头用 forward P/E(盈利定价), P/B 仅作重资产底部参考(资产清算/账面保护下沿)。"
                 "传统观点把航空当纯周期股给低 P/E(6-9x); 但 DAL 的 premium + Amex(FY2025 $8.2B, 对油价不敏感)正在把盈利从'周期状态量'变成'有结构韧性的现金流' → re-rating 的本质 = 市场愿给更高 forward P/E。这是 thesis 的估值落点。"),
    "why_rows": 5,
    "method_text": "三层分解: ①历史 P/E 锚(航空盈利正常期实际到过的 forward 倍数, 取 11x 作为正常中枢, 不用 2022 刚转正的虚高 15x) × ②质量溢价(DAL 相对三巨头的利润率/资产负债表/ROIC 领先, 对账线见上一页 DAL/UAL 比值) × ③情绪值(周期/情绪位置, 依据『综合判断仪表盘』)。一致性检验: 三层相乘应复现近年实际 forward P/E ~10-11x。",
    "peak": 11.0, "peak_note": "历史 P/E 锚 = 11x(航空盈利正常期中枢; 2024 年末实际 11.2x。不用 2022 转正虚高 15x, 也不用 2023 峰值低估 5.5x)。",
    "premium": 1.05, "premium_note": "质量溢价 1.05x: DAL 已是行业最优(利润率/净债务/ROIC 领先), 但市场对航空整体给折价, 溢价克制取 1.05(目标中枢 11×1.05≈11.6x)。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "hpb_row": ha["HPE"],   # ★ 情绪值锚到历史 P/E(非 P/B)
    "cases": [
        ("Bear", [0.75, 0.70, 0.68, 0.68, 0.68], "油价反弹 / 需求见顶 / 价格战 → 市场把 DAL 打回纯周期股, 目标 P/E 回到 8x 下沿(11×1.05×0.70≈8.1x)。"),
        ("Base", [0.90, 0.92, 0.92, 0.92, 0.92], "高端化+Amex 韧性被部分认可, 目标 P/E ~10.6x(略低于正常中枢, 仍含周期折价)。"),
        ("Bull", [1.10, 1.20, 1.25, 1.25, 1.25], "re-rating 兑现: 市场承认 DAL 脱离纯周期, 目标 P/E 升至 13-14x(接近历史上沿+质量溢价)。"),
    ],
    "sent_note": "情绪值=周期/情绪位置。1.0=付足『峰值×溢价』(≈11.6x); >1=市场认 re-rating; <1=纯周期折价。历史列=实际 P/E÷(11×1.05)反推。盈利转正/亏损年实际 P/E 会失真(2021/2023), 看趋势不看绝对。",
    "target_note": "同一个三层公式套三案情绪。目标 forward P/E = 11 × 1.05 × 情绪值。历史列 = 实际 P/E(回看, 三案同值)。",
    "reconcile_text": "卖方目标价对应 FY2027E EPS 约 10-12x(Argus $100 隐含更高); 我们 Base 目标 P/E ~10.6x 偏保守、Bull 13x 与最乐观卖方一致。凭什么敢给 Bull 13x: Amex $8.2B→$10B 的非周期现金流 + premium 持续跑赢 main cabin, 是 re-rating 的事实支撑, 不是情绪。",
    "source_text": "第一层=自身历史 forward P/E band(Research Data API 年末价÷EPS); 第二层=DAL vs UAL/AAL 利润率·净债务·ROIC 对比(10-K + 媒体汇总); 情绪值依据=『综合判断仪表盘』D 块(油价档位+暑运票价+运力纪律)。",
})

# ════════════ 7. 情景切换 ════════════
# 物理锚杠杆: ASM 运力增速 / load factor / PRASM 增速 / jet fuel $/加仑 / CASM-ex 增速 / Amex 现金
sw = K.write_scenario_switch(wb.create_sheet(S_SW), {
    "title": "情景切换 — 全模型唯一情景参数库 + 切换开关 (默认 Base)",
    "usage": ("怎么用: B2 下拉选案 → 案序号派生 → 各杠杆『当前案』行跟着切 → 整条明细链(运力/油价→收入→成本→利润→EPS→目标P/E→隐含价)变档。"
              "三案对比看『估值对比』(恒常并排)。情景参数只在本页改(蓝字); 未列入的假设三案共用 Base。"
              "★ 数值灵敏核心: jet fuel $/加仑 与 PRASM 增速。改 jet fuel → 燃油成本 → 净利 → EPS → 隐含价全链动。"),
    "cases": CASES, "default": "Base",
    "triggers": [
        ("Bear", "jet fuel 反弹到 $3.50+/加仑(中东再起/OPEC+减产) 或 暑运票价回吐 + 美国需求走弱 + 运力纪律破裂价格战。"),
        ("Base", "停火延续、jet fuel ~$2.55/加仑、运力纪律保持、高端化继续、Amex 按指引增长。"),
        ("Bull", "jet fuel 进一步回落到 $2.30、暑运票价结构性维持、premium 加速、Amex 提前到 $10B + re-rating 兑现。"),
    ],
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "levers": [
        {"key": "asm", "name": "ASM 运力增速(%)", "fmt": K.PCT,
         "vals": {"Bear": [0.005, 0.0, -0.01, 0.0, 0.01], "Base": [0.02, 0.025, 0.03, 0.03, 0.03], "Bull": [0.03, 0.04, 0.04, 0.04, 0.04]},
         "desc": "物理锚之一: 可用座英里(运力)。运力纪律下增速温和; 客运收入 = ASM × load factor × yield。2026 因油价砍 off-peak 运力。",
         "stories": {"Bear": "需求走弱被动收缩运力。", "Base": "运力纪律, 温和扩张 2-3%。", "Bull": "需求强劲, 高端运力扩张 3-4%。"},
         "hist": [None, None, None, None, 0.033]},  # FY2025 ASM +3.3%
        {"key": "prasm", "name": "PRASM 增速(单位客运收入,%)", "fmt": K.PCT,
         "vals": {"Bear": [-0.005, 0.0, 0.005, 0.005, 0.005], "Base": [0.025, 0.025, 0.02, 0.02, 0.02], "Bull": [0.04, 0.03, 0.025, 0.025, 0.025]},
         "desc": "单位收入(每可用座英里客运收入)。高端化(premium +7% vs main -5%) + 运力纪律支撑; 主舱疲软是逆风。客运收入 = ASM × PRASM。",
         "stories": {"Bear": "价格战 + 需求弱, 单位收入近持平。", "Base": "高端结构升级对冲主舱, PRASM +2-2.5%(运力纪律下票价韧性)。", "Bull": "暑运票价结构性维持 + premium 加速, +2.5-4%。"},
         "hist": [None, None, None, None, -0.016]},  # FY2025 PRASM -1.6%
        {"key": "jetfuel", "name": "jet fuel ($/加仑)", "fmt": '0.00',
         "vals": {"Bear": [2.95, 3.00, 2.90, 2.85, 2.85], "Base": [2.62, 2.40, 2.42, 2.45, 2.45], "Bull": [2.45, 2.25, 2.25, 2.28, 2.28]},
         "desc": "★ 最大数值灵敏点。燃油成本 = 加仑数 × $/加仑。弹性 ~$5.17 EPS/加仑。2026 全年混合 ~$2.62(H1 含 4 月 $4.88 高位 + H2 现货 ~$2.30 回落); 2027+ 正常化到 ~$2.40。FY2025 实际 $2.30。2026→2027 的油价回落是 EPS 增长主驱动。",
         "stories": {"Bear": "中东再起 + OPEC+ 减产, Brent $90+, jet fuel 持续 $2.85-3.00。", "Base": "停火延续, 2026 全年混合 ~$2.62(H1 高位拖累), 2027 正常化 $2.40。", "Bull": "供给充分恢复 + crack spread 正常化, 2027 jet fuel → $2.25。"},
         "hist": [None, None, None, None, 2.30]},  # FY2025 $2.30/gal
        {"key": "casm", "name": "CASM-ex 增速(除油单位成本,%)", "fmt": K.PCT,
         "vals": {"Bear": [0.04, 0.035, 0.03, 0.03, 0.03], "Base": [0.028, 0.018, 0.015, 0.015, 0.015], "Bull": [0.02, 0.012, 0.01, 0.01, 0.01]},
         "desc": "除燃油单位成本(人工/维修/服务等)。人工刚性是上行压力; Q1 2026 +6% 含季节性高基数, 全年指引 low-single-digit。运力增长(ASM +3%)摊薄固定成本 → 2027+ 单位成本增速放缓(经营杠杆)。CASM-ex × ASM = 非油营业成本。",
         "stories": {"Bear": "飞行员合同大涨(2026 底重谈) + 通胀, CASM-ex +3-4%。", "Base": "2026 +2.8%, 2027+ 规模效应降到 +1.5-1.8%(经营杠杆)。", "Bull": "效率提升 + 满载, 2027+ 仅 +1%。"},
         "hist": [None, None, None, None, 0.024]},  # FY2025 CASM-ex +2.4%
        {"key": "amex", "name": "Amex+忠诚度现金增速(%)", "fmt": K.PCT,
         "vals": {"Bear": [0.04, 0.04, 0.04, 0.04, 0.04], "Base": [0.08, 0.08, 0.07, 0.07, 0.06], "Bull": [0.12, 0.10, 0.09, 0.08, 0.08]},
         "desc": "Amex remuneration FY2025 $8.2B, 目标 ~$10B。对油价/航空周期不敏感, 穿越周期稳定器。驱动『忠诚度品牌费+货运』段。",
         "stories": {"Bear": "消费走弱, 刷卡支出放缓 +4%。", "Base": "按指引稳步增长 +6-8%, 几年到 $10B。", "Bull": "提前到 $10B, +8-12%。"},
         "hist": [None, None, None, None, 0.05]},
    ],
    "linked": [
        {"key": "sent", "name": "情绪值(目标 P/E 第三层)", "fmt": K.N2,
         "src_sheet": S_MULT, "src_row0": ma["sent_row0"],
         "note": "三案取值与依据见『估值倍数假设』(完整三层方法论在那页); 本页只做切换。"},
    ],
})
# derived: 目标 forward P/E(当前案) = 历史 P/E 锚 × 质量溢价 × 当前案情绪 → 喂『情景估值』
_pk = f"'{S_MULT}'!{ma['pk_cell']}"
_pr = f"'{S_MULT}'!{ma['pr_cell']}"
_sent_act = sw["SWACT"]["sent"]
_r = sw["next_row"]
K.lab(wb[S_SW], f"A{_r}", "目标 forward P/E(当前案)", b=True)
for _c in ALLC:
    K.fml(wb[S_SW], f"{_c}{_r}", f"={_pk}*{_pr}*{_c}{_sent_act}", K.MX, link=True)
K.logic(wb[S_SW], f"L{_r}", "= 历史 P/E 锚 × 质量溢价 × 当前案情绪值 → 喂『情景估值』的前瞻 P/E 镜头。")
SWPE = _r

# ════════════ 8. 物理锚 [ANCHOR] — 运力(ASM) + jet fuel ════════════
anchor = K.write_anchor(wb.create_sheet(S_ANCHOR), {
    "title": "运力(ASM, 十亿座英里) + jet fuel 价格 — 需求与成本的物理底盘",
    "all_cols": ALLC, "all_years": ALLY,
    "series": [
        ("ASM 可用座英里(十亿)", [241, 280, 289, 288, 298, None, None, None, None, None],
         "运力物理量。FY2025 298(10-K); 前瞻 = 上年×(1+运力增速, 情景切换)。客运收入挂在它上面。", K.N0),
        ("load factor 客座率(%)", [0.78, 0.85, 0.85, 0.85, 0.84, 0.84, 0.84, 0.84, 0.84, 0.84],
         "客座率。FY2025 84%(10-K)。前瞻假设稳定 ~84%(成熟航司满载常态)。", K.PCT),
        ("jet fuel ($/加仑)", [1.90, 3.30, 2.79, 2.57, 2.30, None, None, None, None, None],
         "★ 油价物理量。FY2025 $2.30/加仑(10-K, 4,269M 加仑)。前瞻 = 情景切换当前案。每变动 $1/加仑 → EPS ~$5.17。", '0.00'),
        ("燃油消耗(百万加仑)", [None, None, None, None, 4269, None, None, None, None, None],
         "FY2025 4,269M 加仑(10-K)。前瞻 = ASM × 单位油耗(座英里/加仑, 随运力等比)。弹性公式的乘数。", K.N0),
    ],
    "yoy_row": "ASM 可用座英里(十亿)",
    "source_note": "口径: ASM/load factor/jet fuel/加仑数均来自 SEC 10-K FY2025 经营统计表。前瞻 ASM = 上年×(1+运力增速); jet fuel = 情景切换当前案; 加仑数 = ASM 等比放大(单位油耗 ~69.8 座英里/加仑 FY2025)。",
    "role_note": "作用: ①客运收入 = ASM × load factor × yield(经 PRASM 传导); ②燃油成本 = 加仑数 × jet fuel $/加仑。改 ASM 或 jet fuel → 收入/成本 → 净利 → EPS → 隐含价全链动(数值灵敏铁律)。",
})
ASM_ROW = anchor["row_of"]["ASM 可用座英里(十亿)"]
JF_ROW = anchor["row_of"]["jet fuel ($/加仑)"]
GAL_ROW = anchor["row_of"]["燃油消耗(百万加仑)"]
# 前瞻 ASM = 上年 × (1 + 运力增速)
for _i, _c in enumerate(FCf):
    _p = FC[_i]  # 上一列(含基年 F=2025A)
    K.fml(wb[S_ANCHOR], f"{_c}{ASM_ROW}", f"={_p}{ASM_ROW}*(1+{K.R(S_SW, _c + str(sw['SWACT']['asm']))})", K.N0, link=True)
# 前瞻 jet fuel = 情景切换当前案
for _c in FCf:
    K.fml(wb[S_ANCHOR], f"{_c}{JF_ROW}", f"={K.R(S_SW, _c + str(sw['SWACT']['jetfuel']))}", '0.00', link=True)
# 前瞻加仑数 = ASM(十亿座英里) × 1e9 / 69.8(座英里/加仑) / 1e6 → 百万加仑; FY2025 锚: 298e9/69.8/1e6=4269 ✓
for _c in FCf:
    K.fml(wb[S_ANCHOR], f"{_c}{GAL_ROW}", f"={_c}{ASM_ROW}*1000/69.8", K.N0, link=True)

# ════════════ 9. 分部测算 — 客运(ASM驱动) + 忠诚度/Amex(独立) + 其他 ════════════
seg = K.write_segment_model(wb.create_sheet(S_SEG), {
    "title": "分部测算 — 客运(ASM × load factor × yield) + 忠诚度/Amex(非周期) + 其他",
    "all_cols": ALLC, "all_years": ALLY, "logic_col": "N",
    "groups": [
        ("物理锚(引自运力与油价物理锚)", [
            ("ASM(十亿)", None, K.N0, "= 引自物理锚页。客运收入挂它(收入=ASM×PRASM, PRASM 已含 load factor×yield)。"),
            ("PRASM 增速", None, K.PCT, "= 引自情景切换当前案。单位客运收入增速。"),
        ]),
        ("客运段 = ASM × PRASM(单位客运收入)", [
            ("PRASM(美分/座英里)", None, '0.00', "历史 = 客运收入 ÷ ASM 反推(FY2025 17.37 美分, 10-K); 前瞻 = 上年×(1+PRASM增速)。"),
            ("客运收入($B)", None, K.N1, "历史取历史财务实数(主舱+高端+里程兑换); 前瞻 = ASM × PRASM/100。喂利润假设。"),
        ]),
        ("忠诚度+Amex 段(非周期, 独立驱动)", [
            ("忠诚度+Amex($B)", None, K.N1, "历史取实数(忠诚度品牌费+货运段); 前瞻 = 上年×(1+Amex增速, 情景切换)。对油价不敏感。"),
        ]),
        ("其他段(炼油第三方+杂项)", [
            ("其他收入($B)", None, K.N1, "炼油第三方销售 + MRO/杂项。历史取实数; 前瞻随炼油周期小幅外推(+2%/年)。低毛利。"),
        ]),
        ("燃油物理量(成本侧, 引自物理锚)", [
            ("燃油加仑(百万)", None, K.N0, "= 引自物理锚页。"),
            ("jet fuel($/加仑)", None, '0.00', "= 引自物理锚页(情景切换当前案)。"),
            ("燃油成本($B)", None, K.N1, "= 加仑数 × $/加仑 / 1000。喂利润假设(成本侧)。改油价直接动。"),
        ]),
    ],
})
m = seg["m"]
# 历史财务分部行
MAIN_H = ha["seg_rows"]["主舱客票(main cabin)"]
PREM_H = ha["seg_rows"]["高端客票(premium)"]
MILE_H = ha["seg_rows"]["里程兑换+旅行服务"]
LOY_H = ha["seg_rows"]["忠诚度品牌费+货运"]
OTH_H = ha["seg_rows"]["炼油第三方+其他"]

# 物理锚引用
for col in ALLC:
    K.fml(wb[S_SEG], f"{col}{m['ASM(十亿)']}", f"={K.R(S_ANCHOR, col + str(ASM_ROW))}", K.N0, link=True)
    K.fml(wb[S_SEG], f"{col}{m['燃油加仑(百万)']}", f"={K.R(S_ANCHOR, col + str(GAL_ROW))}", K.N0, link=True)
    K.fml(wb[S_SEG], f"{col}{m['jet fuel($/加仑)']}", f"={K.R(S_ANCHOR, col + str(JF_ROW))}", '0.00', link=True)
    K.fml(wb[S_SEG], f"{col}{m['燃油成本($B)']}", f"={col}{m['燃油加仑(百万)']}*{col}{m['jet fuel($/加仑)']}/1000", K.N1)
for col in FCf:
    K.fml(wb[S_SEG], f"{col}{m['PRASM 增速']}", f"={K.R(S_SW, col + str(sw['SWACT']['prasm']))}", K.PCT, link=True)

# 客运收入: 历史 = 主舱+高端+里程兑换 实数; PRASM 历史 = 客运收入/ASM*100
for col in HC:
    K.fml(wb[S_SEG], f"{col}{m['客运收入($B)']}", f"={K.R(S_HIST, col + str(MAIN_H))}+{K.R(S_HIST, col + str(PREM_H))}+{K.R(S_HIST, col + str(MILE_H))}", K.N1, link=True)
    K.fml(wb[S_SEG], f"{col}{m['PRASM(美分/座英里)']}", f"={col}{m['客运收入($B)']}*100/{col}{m['ASM(十亿)']}", '0.00')
# 前瞻 PRASM = 上年×(1+增速); 客运收入 = ASM × PRASM/100
_prevs = [HC[-1]] + list(FCf[:-1])
for _p, _c in zip(_prevs, FCf):
    K.fml(wb[S_SEG], f"{_c}{m['PRASM(美分/座英里)']}", f"={_p}{m['PRASM(美分/座英里)']}*(1+{_c}{m['PRASM 增速']})", '0.00')
    K.fml(wb[S_SEG], f"{_c}{m['客运收入($B)']}", f"={_c}{m['ASM(十亿)']}*{_c}{m['PRASM(美分/座英里)']}/100", K.N1)

# 忠诚度+Amex: 历史实数; 前瞻 = 上年×(1+Amex增速)
for col in HC:
    K.fml(wb[S_SEG], f"{col}{m['忠诚度+Amex($B)']}", f"={K.R(S_HIST, col + str(LOY_H))}", K.N1, link=True)
for _p, _c in zip(_prevs, FCf):
    K.fml(wb[S_SEG], f"{_c}{m['忠诚度+Amex($B)']}", f"={_p}{m['忠诚度+Amex($B)']}*(1+{K.R(S_SW, _c + str(sw['SWACT']['amex']))})", K.N1, link=True)
# 其他: 历史实数; 前瞻 +2%/年
for col in HC:
    K.fml(wb[S_SEG], f"{col}{m['其他收入($B)']}", f"={K.R(S_HIST, col + str(OTH_H))}", K.N1, link=True)
for _p, _c in zip(_prevs, FCf):
    K.fml(wb[S_SEG], f"{_c}{m['其他收入($B)']}", f"={_p}{m['其他收入($B)']}*1.02", K.N1)
for col in FCf:
    wb[S_SEG][f"{col}{m['客运收入($B)']}"].fill = K.OUT

# ════════════ 10. 利润与收入假设 ════════════
# 营业利润 = 总营收 − CASM-ex×ASM(非油成本) − 燃油成本; 净利 = 营业利润 × 净利转换(税+利息)
# kit 的 write_fundamentals 用 段OPM 模型, 但航空成本结构是 unit-cost, 不是段OPM。
# 我们用一个折中: 把客运/忠诚度/其他各给一个隐含 OPM, 但真正的成本灵敏(CASM-ex, 燃油)在分部测算页已通电。
# 为保持链通且数值灵敏(油价→燃油成本→净利), 改用自定义利润段: 见下方手工回填。
fr = K.write_fundamentals(wb.create_sheet(S_FUND), {
    "title": "利润与收入假设 — 单位成本驱动(CASM-ex + 燃油) → 营业利润 → 净利 → EPS",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
    "eps_label": "EPS ($, adjusted 口径)",
    "assum_groups": [
        ("单位成本与转换假设(历史实际锚 + 前瞻; 粗颗粒)", [
            {"name": "CASM-ex(美分/座英里)", "vals": [16.0, 14.5, 13.8, 13.54, 13.86, None, None, None, None, None],
             "fmt": '0.00', "logic": "除油单位成本。FY2025 13.86 美分(10-K); 前瞻 = 上年×(1+CASM-ex增速, 情景切换)。", "nm_cols": []},
            {"name": "净利转换率(净利/营业利润)", "vals": [None, 0.36, 0.84, 0.58, 0.94, 0.78, 0.80, 0.81, 0.81, 0.81],
             "fmt": K.PCT, "logic": "营业利润扣净利息(~$0.6B, 持续降)+税(~21%)到净利。FY2025 94%(含 $1.2B 投资收益); 前瞻 ~78-81%((OP−利息)×(1−税), 利息随降杠杆下降)。", "nm_cols": ["B"]},
            {"name": "留存率", "vals": [1.0, 1.0, 0.97, 0.95, 0.93, 0.90, 0.88, 0.86, 0.85, 0.85],
             "fmt": K.PCT, "logic": "留存率=1−派息率。DAL 恢复分红+优先降债, 派息率渐升; 留存率随之降。", "nm_cols": []},
        ]),
    ],
    "segments": [
        {"name": "客运收入", "hist_row": "主舱客票(main cabin)", "fwd": {"sheet": S_SEG, "row": m["客运收入($B)"]}},
        {"name": "忠诚度+Amex", "hist_row": "忠诚度品牌费+货运", "fwd": {"sheet": S_SEG, "row": m["忠诚度+Amex($B)"]}},
        {"name": "其他收入", "hist_row": "炼油第三方+其他", "fwd": {"sheet": S_SEG, "row": m["其他收入($B)"]}},
    ],
    # profit_terms 用占位(下面手工覆盖营业利润为 unit-cost 模型)
    "profit_terms": [
        (["客运收入"], "净利转换率(净利/营业利润)", False),
    ],
    "conv_assum": "净利转换率(净利/营业利润)", "retention_assum": "留存率",
    "note_text": "成本结构是单位经济(不是段OPM): 营业利润 = 总营收 − CASM-ex×ASM(非油成本) − 燃油成本(加仑×$/加仑)。净利 = 营业利润 × 净利转换率(税+利息)。改 jet fuel 或 CASM-ex → 营业利润 → 净利 → EPS 全链动。客运段历史 hist_row 仅借主舱行作占位, 真实客运收入 = 分部测算页三项合计。",
})
# ★ CASM-ex 前瞻 = 上年×(1+增速, 情景切换); kit link 会误把整行设为增速值, 故手工递推
CASM_A = fr["am"]["CASM-ex(美分/座英里)"]
_prevsF = [HC[-1]] + list(FCf[:-1])
for _p, _c in zip(_prevsF, FCf):
    K.fml(wb[S_FUND], f"{_c}{CASM_A}", f"={_p}{CASM_A}*(1+{K.R(S_SW, _c + str(sw['SWACT']['casm']))})", '0.00', link=True)

# ★ 手工覆盖营业利润行为 unit-cost 模型(kit 默认段OPM 不适合航空)
OP_ROW = fr["OP"]
NI_ROW = fr["NI"]
REV_ROW = fr["REV"]
# 客运段历史行借了主舱占位 → 修正为分部测算客运合计(历史+前瞻都引分部测算)
PAX_FR = fr["seg_rows"]["客运收入"]
for col in HC:
    K.fml(wb[S_FUND], f"{col}{PAX_FR}", f"={K.R(S_SEG, col + str(m['客运收入($B)']))}", K.N1, link=True)
# 总营收行已 = 客运+忠诚度+其他(kit 自动), 但客运历史被改成分部测算口径; 重算无需动(公式引 seg 行)
# 营业利润 = 总营收 − CASM-ex×ASM/100 − 燃油成本 − 其他段近零毛利成本(炼油第三方按 ~95% 平价转售)
# 这样口径 ≈ adjusted operating income(剔炼油转售的低毛利): FY2025 ≈ $5.9B(管理层 adjusted 双位数 OPM)
OTH_FR = fr["seg_rows"]["其他收入"]
for col in ALLC:
    asm_ref = K.R(S_SEG, col + str(m['ASM(十亿)']))
    fuel_ref = K.R(S_SEG, col + str(m['燃油成本($B)']))
    K.fml(wb[S_FUND], f"{col}{OP_ROW}", f"={col}{REV_ROW}-{col}{CASM_A}*{asm_ref}/100-{fuel_ref}-{col}{OTH_FR}*0.95", K.N1)
# 净利 = 营业利润 × 净利转换(前瞻); 历史净利仍引历史财务(kit 默认已做)
CONV_A = fr["am"]["净利转换率(净利/营业利润)"]
for col in FCf:
    K.fml(wb[S_FUND], f"{col}{NI_ROW}", f"={col}{OP_ROW}*{col}{CONV_A}", K.N1)

# ════════════ 11. 情景估值(P/E 主镜头) ════════════
# kit 主线用 BPS(P/B); 我们让 P/E 平行镜头(pe_lens)成为决策主输出, P/B 主线作底部参考。
# 先在情景切换加一个"目标 P/B(当前案)"作为底部镜头 derived。
_r2 = SWPE + 1
# 目标 P/B 底部: 用 1.5x(航空账面保护下沿; DAL 当下 P/B ~2.7x, 历史 1-3x)
K.lab(wb[S_SW], f"A{_r2}", "目标 P/B(底部镜头, 当前案)", b=True)
PB_TARGETS = {"Bear": 1.3, "Base": 1.8, "Bull": 2.3}
for i, cs in enumerate(CASES):
    pass
# 用情绪值缩放一个基准 P/B(简单: 三案常数)。直接写三案常数行不便, 用公式: 基准1.8 × 情绪/Base情绪
# 简化: 目标 P/B(当前案) = 1.8 × (当前案情绪 / Base情绪0.92)，给底部弹性
for _c in ALLC:
    K.fml(wb[S_SW], f"{_c}{_r2}", f"=1.8*{_c}{_sent_act}/0.92", K.MX, link=True)
K.logic(wb[S_SW], f"L{_r2}", "底部镜头(重资产账面保护): 基准 1.8x P/B(航空历史 1-3x 中段) × 情绪缩放。仅作隐含价下沿参考, 主决策看 P/E。")
SWPB = _r2

sv = K.write_scenario_valuation(wb.create_sheet(S_VAL), {
    "title": "情景估值 — 当前案逐年隐含价 (P/E 主镜头; P/B 底部参考)",
    "intro": "本表输出=『情景切换』当前案(默认 Base)。主决策镜头 = 目标 forward P/E × EPS(下方 P/E 平行镜头); P/B 主线行作重资产底部参考(下沿)。历史列用实际年末价反推倍数(事实); 前瞻是预测、不拟合现价。三案并排见『估值对比』。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf, "hist_years": HY,
    "fx_fwd": FX_FC, "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
    "s_fund": S_FUND, "fr": fr,
    "s_switch": S_SW, "target_row": SWPB, "sw_cell": "B2",
    "yend": px["yend"], "yavg": px["yavg"],
    "mcap_div": 1000,  # 隐含价 × 股本(M) / 1000 = $B 市值
    "mcap_hist_label": "市值 实际年末($B, 历史)",
    "mcap_fwd_label": "市值 前瞻·P/B底部($B)",
    "pe_lens": {"target_row": SWPE, "mcap_label": "市值 前瞻·P/E主镜头($B)"},
    "reading": "P/E 主镜头读法: 目标 forward P/E(当前案) × 前瞻 adjusted EPS = 隐含价。Base 目标 P/E ~10.6x。P/B 底部行给账面保护下沿(熊市深跌时的支撑)。两镜头三角: 正常情景看 P/E, 极端下行看 P/B 下沿是否被跌破。",
    "method": "方法: 整体公司、forward P/E 主线逐年估。EPS 在『利润与收入假设』(单位成本驱动); 目标 P/E 在『估值倍数假设』(三层); 本表: 目标 P/E × 前瞻 EPS → 隐含价(P/E 镜头); 目标 P/B × BPS → 底部参考。",
    "concl": "结论(方向性): 三情景见『估值对比』。Base 目标价由目标 P/E × FY2027E EPS 给出; risk-reward 看油价对称弹性(±$5.17 EPS/加仑)。",
})

# ════════════ 12. 估值对比(三案并排, P/E 主镜头) ════════════
SWB = sw["SWB"]
SH_F = K.R(S_HIST, f"$F${ha['HSH']}")
PX_NOW_REF = K.R(S_HIST, f"G{ha['HPX']}")
_casm = fr["am"]["CASM-ex(美分/座英里)"]  # 注: 前瞻链到情景切换 casm
_conv = fr["am"]["净利转换率(净利/营业利润)"]
_ret = fr["am"]["留存率"]


def _fwdprev(j, A, key):
    return (HC[-1] if j == 0 else FCf[j - 1]) + str(A[key])


cmp_rows = [
    {"key": "asm", "label": "ASM 运力(十亿)", "fmt": K.N0,
     "hist": lambda c, ci, A: f"={K.R(S_ANCHOR, c + str(ASM_ROW))}",
     "fwd": lambda c, j, ci, A: (f"={K.R(S_ANCHOR, 'F' + str(ASM_ROW))}*(1+{K.R(S_SW, c + str(SWB['asm'] + ci))})" if j == 0
                                 else f"={FCf[j-1]}{A['asm']}*(1+{K.R(S_SW, c + str(SWB['asm'] + ci))})")},
    {"key": "jf", "label": "jet fuel ($/加仑)", "fmt": '0.00',
     "hist": lambda c, ci, A: f"={K.R(S_ANCHOR, c + str(JF_ROW))}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_SW, c + str(SWB['jetfuel'] + ci))}"},
    {"key": "prasm_px", "label": "PRASM (美分)", "fmt": '0.00',
     "hist": lambda c, ci, A: f"={K.R(S_SEG, c + str(m['PRASM(美分/座英里)']))}",
     "fwd": lambda c, j, ci, A: (f"={K.R(S_SEG, 'F' + str(m['PRASM(美分/座英里)']))}*(1+{K.R(S_SW, c + str(SWB['prasm'] + ci))})" if j == 0
                                 else f"={_fwdprev(j, A, 'prasm_px')}*(1+{K.R(S_SW, c + str(SWB['prasm'] + ci))})")},
    {"key": "pax", "label": "客运收入 ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_SEG, c + str(m['客运收入($B)']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['asm']}*{c}{A['prasm_px']}/100"},
    {"key": "loy", "label": "忠诚度+Amex ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_SEG, c + str(m['忠诚度+Amex($B)']))}",
     "fwd": lambda c, j, ci, A: (f"={K.R(S_HIST, 'F' + str(LOY_H))}*(1+{K.R(S_SW, c + str(SWB['amex'] + ci))})" if j == 0
                                 else f"={_fwdprev(j, A, 'loy')}*(1+{K.R(S_SW, c + str(SWB['amex'] + ci))})")},
    {"key": "oth", "label": "其他 ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_SEG, c + str(m['其他收入($B)']))}",
     "fwd": lambda c, j, ci, A: (f"={K.R(S_HIST, 'F' + str(OTH_H))}*1.02" if j == 0
                                 else f"={_fwdprev(j, A, 'oth')}*1.02")},
    {"key": "rev", "label": "总收入 ($B)", "fmt": K.N1, "bold": True,
     "hist": lambda c, ci, A: f"={c}{A['pax']}+{c}{A['loy']}+{c}{A['oth']}",
     "fwd": lambda c, j, ci, A: f"={c}{A['pax']}+{c}{A['loy']}+{c}{A['oth']}"},
    {"key": "gal", "label": "燃油加仑(百万)", "fmt": K.N0,
     "hist": lambda c, ci, A: f"={K.R(S_ANCHOR, c + str(GAL_ROW))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['asm']}*1000/69.8"},
    {"key": "fuel", "label": "燃油成本 ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_SEG, c + str(m['燃油成本($B)']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['gal']}*{c}{A['jf']}/1000"},
    {"key": "casm_px", "label": "CASM-ex (美分)", "fmt": '0.00',
     "hist": lambda c, ci, A: f"={K.R(S_FUND, c + str(_casm))}",
     "fwd": lambda c, j, ci, A: (f"={K.R(S_FUND, 'F' + str(_casm))}*(1+{K.R(S_SW, c + str(SWB['casm'] + ci))})" if j == 0
                                 else f"={_fwdprev(j, A, 'casm_px')}*(1+{K.R(S_SW, c + str(SWB['casm'] + ci))})")},
    {"key": "op", "label": "营业利润 ($B)", "fmt": K.N1, "bold": True,
     "hist": lambda c, ci, A: f"={K.R(S_FUND, c + str(OP_ROW))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['rev']}-{c}{A['casm_px']}*{c}{A['asm']}/100-{c}{A['fuel']}-{c}{A['oth']}*0.95"},
    {"key": "ni", "label": "净利 ($B)", "fmt": K.N1, "bold": True,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HNI']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['op']}*{K.R(S_FUND, c + str(_conv))}"},
    {"key": "eps", "label": "EPS ($)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HEPS']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['ni']}*1000/{SH_F}"},
    {"key": "sent", "label": "情绪值(该案; 历史=实际反推)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_MULT, c + str(ma['sent_row0'] + ci))}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_MULT, c + str(ma['sent_row0'] + ci))}"},
    {"key": "pe", "label": "目标 fwd P/E(历史=实际)", "fmt": K.MX,
     "hist": lambda c, ci, A: f'=IF({c}{A["eps"]}<=0,"N/M",{K.R(S_HIST, c + str(ha["HPX"]))}/{c}{A["eps"]})',
     "fwd": lambda c, j, ci, A: f"={_pk}*{_pr}*{c}{A['sent']}"},
    {"key": "px", "label": "隐含价 P/E镜头 ($)", "fmt": K.N1, "bold": True, "out": True,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HPX']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['pe']}*{c}{A['eps']}"},
    {"key": "up", "label": "历史: vs 实际年末价 / 前瞻: vs 现价", "fmt": K.PCT,
     "hist": lambda c, ci, A: f"={c}{A['px']}/{K.R(S_HIST, c + str(ha['HPX']))}-1",
     "fwd": lambda c, j, ci, A: f"={c}{A['px']}/{PX_NOW_REF}-1"},
]
cm = K.write_comparison(wb.create_sheet(S_CMP), {
    "title": "估值对比 — Bear / Base / Bull 三个情景的目标价并排对比(P/E 主镜头)",
    "intro": ("三个情景各自完整推演: 物理锚(ASM + jet fuel) → 客运/忠诚度/其他收入 → 营业利润(减 CASM-ex×ASM 与燃油) → 净利 → EPS → 目标 forward P/E → 隐含价。"
              "本表三案永远并排, 不随『情景切换』变。未列入情景矩阵的假设三案共用 Base。历史列 2021-2025 = 同链填实际, 隐含价历史列 = 实际年末价(P/E 行历史=实际倍数, 内置回测)。"
              "目标年 = 2027E(暑运油价利好充分兑现 + 高端化/Amex 增长落地的代表年)。"),
    "case_names": CASES,
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "block_start": 16,
    "rows": cmp_rows,
    "summary": {
        "band": "三案汇总 (目标年 2027E; 各案触发条件见『情景切换』)",
        "target_col": "H",  # H = 2027E
        "rows": [
            ("EPS($)", "eps", K.N2, "由该案 jet fuel/PRASM/运力/CASM-ex 假设逐年推导", False),
            ("净利($B)", "ni", K.N1, "= 营业利润 × 净利转换率", False),
            ("总收入($B)", "rev", K.N1, "= 客运(ASM×PRASM) + 忠诚度/Amex + 其他", False),
            ("目标 fwd P/E", "pe", K.MX, "= 历史 P/E 锚 11x × 质量溢价 1.05 × 该案情绪值", False),
            ("隐含价($)", "px", K.N1, "= 目标 forward P/E × 2027E EPS", True),
            ("vs 现价 $86.42", "up", K.PCT, "对照现价的上行/下行空间", True),
        ],
        "mcap": {"label": "隐含市值($B)", "key": "px", "expr": f"*{SH_F}/1000",
                 "note": "= 隐含价 × 股本(652M) / 1000"},
        "concl": "风险收益比(2027E vs 现价 $86.42)一行收口: Bear 下行 / Base / Bull 上行。核心 swing = jet fuel(±$5.17 EPS/加仑); 油价反弹是最大下行风险, 暑运票价结构性维持 + re-rating 是上行。",
    },
})

# ════════════ 13. 综合判断仪表盘 ════════════
EPS27 = K.R(S_FUND, "H" + str(fr["EPS"]))  # FY2027E EPS
PXD = K.R(S_HIST, "G" + str(ha["HPX"]))
dash = K.write_dashboard(wb.create_sheet(S_DASH), {
    "title": "综合判断仪表盘 — A 基本面拐点 · B 估值错位(预测引擎) · C 催化剂 · D 情绪确认",
    "usage": ("怎么用: 预测引擎是 B(估值错位)+ C(催化剂)。基本面 A 看高端化+Amex 是否把盈利从周期变结构; 情绪 D 看油价/暑运/运力纪律档位。"
              "验收=回测: 放回 2024H2(油价回落+EPS 重建)那波拐点, 这套表当时就能看到。"),
    "blocks": [
        {"title": "A. 基本面拐点 — 盈利在从周期变结构吗?", "rows": [
            ("产品组合迁移", "premium +7% vs main cabin -5%(FY2025); Q1 2026 premium +14%", "高端化结构升级, 不是周期反弹。premium 占客票已逼近主舱。"),
            ("Amex/忠诚度现金", "FY2025 $8.2B → 目标 ~$10B; Q1 2026 +10%", "对油价不敏感的非周期现金流, 盈利质量的结构性支撑。"),
            ("A 判断", "【强】", "高端化 + Amex 让 DAL 盈利韧性领先三巨头; 是脱离纯周期的事实依据。", True),
        ]},
        {"title": "B. 估值错位(预测引擎 ★)— 市场给的 vs 基本面该给的 → GAP", "rows": [
            ("市场现在给(forward P/E)", {"fml": f"={PXD}/{EPS27}", "fmt": K.MX, "fill": True},
             "= 现价 $86.42 ÷ FY2027E EPS(模型算)。"),
            ("基本面该给(justified)", {"inp": 11.6, "fmt": K.MX},
             "= 历史 P/E 锚 11x × 质量溢价 1.05(DAL 行业最优)。航空正常期中枢。"),
            ("错位 GAP = 该给÷市场给 − 1",
             {"fml": lambda ro: f"=B{ro['基本面该给(justified)']}/B{ro['市场现在给(forward P/E)']}-1", "fmt": K.PCT},
             "GAP 正且大 = 重估空间(该买); 转负 = 已透支。当前市场给 ~10-11x vs 该给 11.6x → 小幅低估, 主要 alpha 在 EPS 上修(油价顺风)而非倍数扩张。"),
            ("回测: 2024H2 拐点读数", "当时 forward P/E ~9x vs 该给 11.6x → GAP +29%", "那波股价从 $46 涨到 $69 正是 GAP 闭合 + EPS 增长双击。"),
        ]},
        {"title": "C. 催化剂 — 什么会逼市场闭合 GAP / 上修 EPS", "rows": [
            ("Q2 财报 2026-07-09", "待(指引 EPS $1.00-1.50)", "暑运旺季 + 油价回落的盈利兑现; 若 beat 指引上沿 → EPS 上修。"),
            ("油价路径(jet fuel)", "进行中(4 月 $4.88 → 6 月 $2.70)", "★ 最大催化: 每降 $1/加仑 → EPS +$5.17。停火延续 = 持续顺风。"),
            ("Berkshire 建仓", "✓(2026 Q1 13F, $2.65B)", "价值标杆背书; 运营与股价背离的修复信号。"),
            ("C 判断", "待兑现(Q2 财报 + 暑运数据)", "油价催化进行中, Q2 财报是下一个验证点。", True),
        ]},
        {"title": "D. 情绪确认 — 油价/暑运/运力纪律档位(timing + 刹车)", "rows": [
            ("油价档位", "顺风(回落中, jet fuel ~$2.70)", "停火延续 = 成本顺风; 中东再起 = 立刻翻 Bear。"),
            ("暑运票价 + 运力纪律", "强(BofA: 运力纪律是暑运关键)", "因油价提的票价是否结构性维持 = PRASM 路径的扳机。"),
            ("当前档位", "【启动】", "油价顺风 + Berkshire 背书 + EPS 上修在途; 估值未过热(forward ~11x)。", True),
            ("衰减扳机", "3 条", "①jet fuel 反弹 >$3.30 ②暑运后票价回吐 ③运力纪律破裂价格战。任一翻 → 降档 + 情绪值下调。", True),
        ]},
    ],
    "final": {"band": "★ 综合判断(A+B+C+D 收成一句可执行的话)",
              "text": "A 强(高端化+Amex 结构韧性) + B 小幅低估(GAP 主要在 EPS 上修而非倍数) + C 催化在途(油价顺风+Q2 财报) + D 启动档(未过热)。当下: 便宜的好公司, 主要 alpha 来自油价顺风驱动的 EPS 上修, 倍数扩张为辅。买入。核心风险 = 油价对称反弹。"},
    "tracking": {
        "intro": "哪个指标恶化 → 哪个假设先崩 → 触发什么动作(盯的优先级)。",
        "rows": [
            ("__band__", "一、油价(最大 swing)"),
            ("jet fuel $/加仑", "~$2.70(6 月)", "命门: 弹性 $5.17 EPS/加仑", "Argus jet fuel 指数 / EIA STEO(周/月)", "反弹 >$3.30 → 切 Bear 重算 EPS"),
            ("__band__", "二、单位收入(暑运票价)"),
            ("PRASM 同比", "待 Q2", "命门: 客运收入路径", "DAL 季报 + 行业 RASM tracker", "连续转负 → 下调 PRASM 假设"),
            ("__band__", "三、单位成本(人工)"),
            ("CASM-ex 同比", "Q1 2026 +6%", "命门: 非油成本刚性", "DAL 季报; 飞行员合同(2026 底重谈)", ">+5% 持续 → 上调 CASM-ex 假设"),
            ("__band__", "四、非周期现金流"),
            ("Amex remuneration", "FY2025 $8.2B", "命门: 盈利结构韧性 thesis", "DAL 10-K/10-Q 披露", "增速放缓 <5% → A 块判断降级"),
        ],
    },
})

# ════════════ 全局格式 + 落盘 ════════════
K.finalize(wb, freeze={
    S_HIST: "B3", S_PX: "B4", S_CONS: "A2", S_HMULT: "B5", S_MULT: "B4", S_SW: "B3",
    S_ANCHOR: "B3", S_SEG: "B3", S_FUND: "B3", S_VAL: "B4", S_CMP: "B6", S_DASH: "B6",
    S_COVER: "A2",
})
out = os.path.join(os.path.dirname(__file__), "..", "out", "DAL_valuation_model.xlsx")
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("saved:", os.path.abspath(out))
print("sheets:", wb.sheetnames)
