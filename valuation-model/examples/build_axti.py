# -*- coding: utf-8 -*-
"""
build_axti.py — AXT Inc. (NASDAQ: AXTI) 物理锚估值模型。
锚: AI 光模块出货 × InP 激光器颗数 → InP 衬底片数 × 片价 → 行业 TAM → AXT 切片(份额∧产能)。
主线镜头: EV/Sales(P/S 近似); 支线: P/E 交叉验证。单位: $M; 股本=千股; fx=1。
数据 SOT: `VALUATION_INPUT_DIR/AXTI_input.json` 对应的历史快照(2026-06-12)。
"""
import datetime
import os
from openpyxl import Workbook
from build_kit import (R, hdr, band, inp, introw, fml, lab, logic, txt, mtext,
                       set_widths, finalize, write_cover, write_history,
                       write_price_chart, write_consensus, write_multiple_assumptions,
                       write_scenario_switch, write_anchor, write_segment_model,
                       write_fundamentals, write_scenario_valuation, write_comparison,
                       write_dashboard, PCT, N1, N0, N2, MX, PX, BF, CH, CUR, OUT)

_repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_output_dir = os.environ.get("VALUATION_OUTPUT_DIR", os.path.join(_repo_root, "out"))
OUT_PATH = os.path.join(_output_dir, "AXTI_valuation_model.xlsx")
os.makedirs(_output_dir, exist_ok=True)

# ── 周收盘(t,c) TradingView 实拉 2020-11→2026-06 ──
WEEKLY = [(1606141800,9.97),(1606746600,9.99),(1607351400,9.12),(1607956200,9.84),(1608561000,9.64),(1609165800,9.57),(1609770600,11.19),(1610375400,11.59),(1611066600,11.94),(1611585000,10.33),(1612189800,11.23),(1612794600,12.49),(1613485800,15.37),(1614004200,12.93),(1614609000,11.29),(1615213800,13.87),(1615815000,13.11),(1616419800,11.36),(1617024600,12.31),(1617629400,12.41),(1618234200,10.935),(1618839000,11.08),(1619443800,9.86),(1620048600,9.66),(1620653400,8.95),(1621258200,9.95),(1621863000,10.18),(1622554200,10.58),(1623072600,10.2),(1623677400,11.1),(1624282200,10.93),(1624887000,10.69),(1625578200,10.4),(1626096600,9.37),(1626701400,9.82),(1627306200,10.2),(1627911000,9.78),(1628515800,9.59),(1629120600,8.61),(1629725400,9.33),(1630330200,9.81),(1631021400,9.23),(1631539800,7.92),(1632144600,8.6),(1632749400,8.28),(1633354200,8.06),(1633959000,8.17),(1634563800,8.54),(1635168600,8.17),(1635773400,9.22),(1636381800,9.11),(1636986600,8.81),(1637591400,8.15),(1638196200,7.86),(1638801000,8.29),(1639405800,8.2),(1640010600,8.8),(1640615400,8.81),(1641220200,8.41),(1641825000,8.54),(1642516200,7.54),(1643034600,7.06),(1643639400,7.05),(1644244200,7.31),(1644849000,7.45),(1645540200,7.34),(1646058600,6.97),(1646663400,6.84),(1647264600,7.25),(1647869400,7.22),(1648474200,6.9),(1649079000,6.33),(1649683800,6.34),(1650288600,6.11),(1650893400,5.9),(1651498200,5.73),(1652103000,5.47),(1652707800,5.2),(1653312600,5.79),(1654003800,5.86),(1654522200,5.74),(1655127000,5.61),(1655818200,6.19),(1656336600,5.73),(1657027800,6.03),(1657546200,6.52),(1658151000,6.67),(1658755800,8.77),(1659360600,8.77),(1659965400,9.5),(1660570200,9.13),(1661175000,9.14),(1661779800,8.33),(1662471000,8.65),(1662989400,7.78),(1663594200,6.75),(1664199000,6.7),(1664803800,5.13),(1665408600,4.3),(1666013400,4.4),(1666618200,4.6),(1667223000,4.91),(1667831400,5.29),(1668436200,5.31),(1669041000,5.38),(1669645800,5.36),(1670250600,4.97),(1670855400,4.72),(1671460200,4.4),(1672151400,4.38),(1672756200,4.84),(1673274600,5.28),(1673965800,5.48),(1674484200,5.84),(1675089000,6.48),(1675693800,5.92),(1676298600,4.88),(1676989800,4.5),(1677508200,4.34),(1678113000,3.76),(1678714200,3.76),(1679319000,3.77),(1679923800,3.98),(1680528600,3.85),(1681133400,3.5),(1681738200,3.1),(1682343000,2.67),(1682947800,2.89),(1683552600,3.0),(1684157400,3.4),(1684762200,3.66),(1685453400,3.48),(1685971800,3.74),(1686576600,3.59),(1687267800,3.28),(1687786200,3.44),(1688391000,2.9),(1688995800,3.0),(1689600600,3.03),(1690205400,2.99),(1690810200,2.92),(1691415000,2.6),(1692019800,2.34),(1692624600,2.28),(1693229400,2.64),(1693920600,2.37),(1694439000,2.32),(1695043800,2.35),(1695648600,2.4),(1696253400,2.49),(1696858200,2.33),(1697463000,2.24),(1698067800,2.07),(1698672600,2.25),(1699281000,1.98),(1699885800,2.04),(1700490600,1.99),(1701095400,2.02),(1701700200,2.33),(1702305000,2.7),(1702909800,2.48),(1703601000,2.4),(1704205800,2.5),(1704724200,2.63),(1705415400,2.65),(1705933800,2.575),(1706538600,2.52),(1707143400,2.52),(1707748200,2.4),(1708439400,3.86),(1708957800,4.27),(1709562600,4.58),(1710163800,4.87),(1710768600,4.95),(1711373400,4.59),(1711978200,3.11),(1712583000,3.43),(1713187800,2.77),(1713792600,2.97),(1714397400,3.66),(1715002200,3.61),(1715607000,3.28),(1716211800,3.46),(1716903000,3.81),(1717421400,3.56),(1718026200,3.63),(1718631000,3.6),(1719235800,3.38),(1719840600,3.29),(1720445400,3.75),(1721050200,3.35),(1721655000,3.25),(1722259800,2.68),(1722864600,2.43),(1723469400,2.87),(1724074200,2.94),(1724679000,2.65),(1725370200,2.32),(1725888600,2.35),(1726493400,2.1),(1727098200,2.55),(1727703000,2.53),(1728307800,2.58),(1728912600,2.61),(1729517400,2.72),(1730122200,2.11),(1730730600,2.05),(1731335400,1.84),(1731940200,1.97),(1732545000,2.14),(1733149800,2.19),(1733754600,2.31),(1734359400,2.18),(1734964200,2.28),(1735569000,2.27),(1736173800,2.13),(1736778600,2.26),(1737469800,2.3),(1737988200,2.09),(1738593000,2.1),(1739197800,2.23),(1739889000,1.58),(1740407400,1.6),(1741012200,1.62),(1741613400,1.8),(1742218200,1.68),(1742823000,1.48),(1743427800,1.24),(1744032600,1.16),(1744637400,1.19),(1745242200,1.42),(1745847000,1.39),(1746451800,1.25),(1747056600,1.49),(1747661400,1.46),(1748352600,1.51),(1748871000,1.78),(1749475800,2.16),(1750080600,1.84),(1750685400,2.03),(1751290200,2.14),(1751895000,2.2),(1752499800,2.44),(1753104600,2.35),(1753709400,1.92),(1754314200,2.14),(1754919000,2.05),(1755523800,2.82),(1756128600,2.9),(1756819800,3.11),(1757338200,3.66),(1757943000,4.34),(1758547800,4.66),(1759152600,5.13),(1759757400,4.03),(1760362200,4.51),(1760967000,6.06),(1761571800,7.95),(1762180200,9.46),(1762785000,10.44),(1763389800,8.93),(1763994600,10.7),(1764599400,11.58),(1765204200,14.81),(1765809000,14.65),(1766413800,15.37),(1767018600,16.76),(1767623400,22.99),(1768228200,22.09),(1768919400,17.4),(1769437800,18.54),(1770042600,24.06),(1770647400,24.24),(1771338600,29.68),(1771857000,37.9),(1772461800,32.37),(1773063000,48.86),(1773667800,54.24),(1774272600,60.63),(1774877400,52.84),(1775482200,64.18),(1776087000,82.56),(1776691800,76.16),(1777296600,96.0),(1777901400,116.36),(1778506200,123.78),(1779111000,140.83),(1779802200,103.16),(1780925400,88.34)]

# 周 → 月度收盘(每月最后一周)
_m = {}
for t, c in WEEKLY:
    ym = datetime.datetime.utcfromtimestamp(t).strftime("%Y-%m")
    _m[ym] = c
MONTHLY = sorted(_m.items())
MONTHLY = [(ym, c) for ym, c in MONTHLY if ym >= "2021-01"]

PX_NOW, SH_NOW = 88.34, 65400          # $; 千股(2026-04 增发后 ~65.4M 股)
HC, HY = list("BCDEF"), ["2021A","2022A","2023A","2024A","2025A"]
FCf, FY = list("GHIJK"), ["2026E","2027E","2028E","2029E","2030E"]
ALLC, ALLY = HC + FCf, HY + FY
S_HIST, S_SW, S_ANC, S_SEG, S_FUND, S_MA = "历史财务与估值","情景切换","物理锚","分部测算","利润与收入假设","估值倍数假设"
CASES = ["Bear","Base","Bull"]

wb = Workbook(); wb.remove(wb.active)

# ════ 0 封面 ════
write_cover(wb.create_sheet("封面"), {
    "title": "AXT (AXTI) 物理锚估值模型 — InP 衬底 × AI 光互连",
    "meta": [("报告日期","2026-06-12"),("现价 / 市值","$88.34 / ~$5.78B (2026-06-11 收盘, 股本 ~65.4M)"),
             ("数据时效","FY2025 10-K(2026-03-17) + 26Q1 10-Q(2026-05-14) + 26Q1 电话会(2026-04-30); 行情 TradingView 实拉"),
             ("物理锚","800G/1.6T 光模块出货 × InP 激光器颗数 → 衬底片数 × 片价 → 行业 TAM → AXT 切片(份额∧产能取小); 挂 AIDC capex 共享基座 v1.0.0"),
             ("估值镜头","主线 EV/Sales(P/S 近似, 净现金 ~$9.2/股未计入作安全垫); 支线 P/E 交叉验证; P/B 地板参考")],
    "takeaways": [
        ("结论(卖出)","三情景 2028E 隐含价: Bear ~$9 / Base ~$41 / Bull ~$85; 概率加权(20/50/30) ~$48, 较现价 $88.34 下行 ~45%。现价已 price-in Bull(公司产能计划全兑现+涨价持续)。"),
        ("链路","锚一动价就动: 光模块出货/片价/产能任一变, 隐含价同步重算。Base 的关键判断: 2027 起需求约束(TAM×份额)先于产能约束生效, 2028 后片价温和回归。"),
        ("与街分歧","卖方均值 $96.5(Northland $125/CH $95/Wedbush $93)外推涨价+产能满产满销; 本模型按行业 TAM 封顶, 公司 2029-30 产能存在过剩风险。"),
        ("证伪触发","对美 InP 许可获批/单季 InP<$15M/2吋现货价跌破$1,500/backlog 环比降 → 回模型改数重算(见仪表盘)。")],
})

# ════ 1 历史财务与估值 ════
ws_h = wb.create_sheet(S_HIST)
ha = write_history(ws_h, {
    "title": "历史财务与估值 ($M; 股本=千股; fx=1)", "hist_cols": HC, "hist_years": HY,
    "fx_hist": [1,1,1,1,1], "fx_now": 1, "vals_in_usd": True,
    "cur_label": "当下(TTM)",
    "segments": [("InP 衬底",[30.0,37.0,20.0,33.0,28.5],True),
                 ("非InP(GaAs/Ge/原材料)",[107.4,104.1,55.8,66.4,59.8],True)],
    "total_now": 95.9, "gm_pct": [0.345,0.369,0.176,0.240,0.127], "gm_now": 0.213,
    "ni": [16.5,18.7,-17.9,-11.6,-21.3], "ni_now": -14.0,
    "eq": [205,222,208,200,276], "eq_now": 274.9,
    "shares": [41500,43000,43300,43800,52900], "shares_now": SH_NOW,
    "px_end": [8.81,4.38,2.40,2.27,16.76], "px_now": PX_NOW,
    "quarter": {"col":"H","label":"26Q1A",
                "segs":{"InP 衬底":(13.6,2.58),"非InP(GaAs/Ge/原材料)":(13.3,None)},
                "ni":-1.62,"eq":274.9,"shares":53320,"fx":1,
                "note":"26Q1 实际(10-Q 2026-05-14): 收入 $26.9M(+39% YoY), InP $13.6M(+258%), GAAP 毛利率 29.6%, 净亏 $1.62M; Q2 指引 ~$34M、EPS 转正。股本现值 65,400 千股为 2026-04 增发($632.5M @ $64.25)后。"},
    "notes": [("HREV","FY2025 收入 $88.3M(-11%): InP 出口管制停摆上半年。InP 逐年拆分为电话会披露拼接估算。"),
              ("HNI","连亏三年: -17.9/-11.6/-21.3; 26Q2 指引 GAAP EPS $0.05-0.07 转正。"),
              ("HEQ","2021-24 权益为按净利/增发回推估算; 2025A=276 由 26Q1 实际 274.9 回推; 现金 pro forma ~$720M(含 4 月增发净 $600M), 净现金 ~$646M(~$9.2/股)。"),
              ("HSH","2025-12 增发 8.16M 股 @$12.25; 2026-04 增发 9.84M 股 @$64.25 → 股本一年 +48%。"),
              ("HPX","年末价取月末收盘(TradingView); 52 周区间 $1.80-143.16(高点 2026-05)。")],
})
ws_h["A2"] = "($M)"
ws_h[f"A{ha['HSH']}"] = "股本(千股)"
ws_h[f"A{ha['HMC']}"] = "市值($M)"
# 追加 SPS / P/S 行(主线镜头的历史底座)
r = ws_h.max_row + 1
band(ws_h, r, "主线镜头底座: 每股销售与 P/S(实际)", 9); r += 1
HSPS = r
lab(ws_h, f"A{HSPS}", "SPS 每股销售($)", note=True)
for c in HC:
    fml(ws_h, f"{c}{HSPS}", f"={c}{ha['HREV']}*1000/{c}{ha['HSH']}", N2)
fml(ws_h, f"G{HSPS}", f"=G{ha['HREV']}*1000/G{ha['HSH']}", N2)
r += 1
HPS = r
lab(ws_h, f"A{HPS}", "P/S (实际)", b=True); ws_h[f"A{HPS}"].fill = OUT
for c in HC:
    fml(ws_h, f"{c}{HPS}", f"={c}{ha['HPX']}/{c}{HSPS}", MX)
fml(ws_h, f"G{HPS}", f"=G{ha['HPX']}/G{HSPS}", MX); ws_h[f"G{HPS}"].fill = CUR
logic(ws_h, f"I{HPS}", "EV/Sales 主线的历史带: 2021 高点 ~2.6x → 2024 谷底 ~1.0x → 2025 末 10x → 当下 TTM ~60x(regime 突变)。")

# ════ 2 股价走势 ════
def _phase(ym):
    if ym <= "2025-01": return "管制前/低迷期"
    if ym <= "2025-05": return "InP 管制停摆"
    if ym <= "2025-12": return "许可恢复+AI 叙事"
    if ym == "2026-01": return "Q4 下修闪崩"
    return "产能扩张重估"
pxinfo = write_price_chart(wb.create_sheet("股价走势"), MONTHLY, {
    "fn": _phase,
    "rows": [("~2025-01 管制前", "衬底商品周期下行+连亏, $2 一线; 2025-02 铟纳入出口管制后最低 $1.13"),
             ("2025-02~05 停摆", "InP 出口逐单许可, 25Q1 毛利率 -6.4%"),
             ("2025-06~12 反转", "首批许可(6/11 8-K)→25Q3 $28M 大超→backlog $49M→$60M; 12 月 $12.25 增发"),
             ("2026-01 闪崩", "25Q4 许可签发不及预期, 下修至 $23M, 单日 -22~31%"),
             ("2026-02~06 重估", "26Q1 双超+backlog>$100M+Q2 指引 $34M+4 月 $64.25 增发 $632.5M; 5 月高点 $143 后回落 $88")]},
    title="AXTI 月度收盘(2021-01 ~ 2026-06)")
YEND = {f"{y}A" if len(y)==4 else y: v for y, v in pxinfo["yend"].items()}
YAVG = {f"{y}A": v for y, v in pxinfo["yavg"].items()}
YHI  = {f"{y}A": v for y, v in pxinfo["yhigh"].items()}
YLO  = {f"{y}A": v for y, v in pxinfo["ylow"].items()}

# ════ 3 卖方研报共识 ════
write_consensus(wb.create_sheet("卖方研报共识"), {
    "title": "卖方共识对账(2026-06)",
    "overview": "覆盖 ~5 家。共识 2026E 收入 $143.1M(+62%)、EPS $0.30; 目标价均值 ~$96.5(高 Northland $125 / Craig-Hallum $95 / Wedbush $93; Needham 2026-01 高位降至 Hold, B.Riley 降 Neutral)。市场叙事 = InP 衬底缺货 70%+、产能 2026/2027 两次翻倍、backlog $100M。",
    "assumptions": [
        ("2026E 收入", "共识 $143M(+62%)", "许可签发节奏决定季度落地", "模型 Base ~$152M: Q1 26.9 实际+Q2 指引 34+产能爬坡, 略高于共识"),
        ("InP 产能兑现", "按公司指引满产满销外推(2028 run-rate $260-280M)", "产能=收入? 需求侧是否接得住没人测算", "Base 按行业 TAM×份额封顶: 2028 InP ~$230M < 产能 $270M, 需求约束先 bind"),
        ("衬底价格", "外推当前 $2,300-2,500/片高价", "2027-28 住友+40%/JX×3/云南锗业×3/Coherent×5 集中投产", "Base 2028 起温和回归($2,200→$2,000); Bear 腰斩(商品周期均值回归)"),
        ("估值倍数", "目标价对应 2026E P/S 30-40x, 用 2027-28 高增长 justify", "AI 材料股给多少倍没有锚", "主线 2028E P/S: Base 9.1x(=历史峰值 2.6x × 结构溢价 3.5x), 对应隐含 fwd P/E ~45-50x"),
        ("出口许可", "默认持续放量, 对美获批是 upside", "10-K 自述'最大挑战'; 25Q4 已演示下修 -31%", "Base 假设欧/日/中国内需持续、对美不通; 触发器跟踪")],
    "divergences": [
        "最大分歧①: 2028 后片价走势——街隐含'缺货常态化', 本模型按扩产周期 18-24 个月推演 2028 起供给追上, 价格回归是 Base 不是 Bear。",
        "最大分歧②: 现价 $88 已 price-in Bull——即便公司产能计划全兑现+价格坚挺(本模型 Bull), 2028E 隐含价也只有 ~$85。"],
    "stances": [
        "Northland: Outperform $125 (2026-06-03, $90→$125) — InP 稀缺性+产能翻倍",
        "Craig-Hallum: Buy $95 (2026-05-01, $29→$95) — backlog 与提价",
        "Wedbush: Outperform $93 — AI 光互连上游材料卡点",
        "Needham (Charles Shi): Buy→Hold (2026-01-20) — 估值已透支",
        "B. Riley: Buy→Neutral (2026-01) — 估值; 2 月目标 $21 已严重过时"],
})

# ════ 4 历史估值倍数(custom: P/S 带 + 同业) ════
ws_m = wb.create_sheet("历史估值倍数")
hdr(ws_m, 1, "历史估值倍数 — P/S 带 + 同业对照 (主线 EV/Sales 的数据底座)", 11)
mr = mtext(ws_m, 2, "AXT 净现金 ~$646M(~11% 市值), P/S≈EV/Sales 略偏保守。读法: 历史带 1.0-2.6x(商品衬底 regime) vs 当下 TTM ~60x(AI 重估 regime)——目标倍数必须用『峰值×结构溢价×情绪』三层显式 justify 突破历史带。", "K", 2)
band(ws_m, mr, "① 自身 P/S 带(年末/年内高低)", 11); mr += 1
ws_m[f"A{mr}"] = "指标"; ws_m[f"A{mr}"].font = BF
for col, y in zip(HC, HY):
    ws_m[f"{col}{mr}"] = y; ws_m[f"{col}{mr}"].font = BF; ws_m[f"{col}{mr}"].fill = CH
ws_m[f"G{mr}"] = "当下"; ws_m[f"G{mr}"].font = BF; ws_m[f"G{mr}"].fill = CUR
mr += 1
lab(ws_m, f"A{mr}", "P/S 年末(实际)", b=True)
for c in HC:
    fml(ws_m, f"{c}{mr}", f"={R(S_HIST, c+str(HPS))}", MX, link=True)
fml(ws_m, f"G{mr}", f"={R(S_HIST, 'G'+str(HPS))}", MX, link=True); ws_m[f"G{mr}"].fill = CUR
mr += 1
RHI = mr
lab(ws_m, f"A{mr}", "年内最高价(月末)")
for col, y in zip(HC, HY):
    if y in YHI: inp(ws_m, f"{col}{mr}", YHI[y], None, N2)
mr += 1
RLO = mr
lab(ws_m, f"A{mr}", "年内最低价(月末)")
for col, y in zip(HC, HY):
    if y in YLO: inp(ws_m, f"{col}{mr}", YLO[y], None, N2)
mr += 1
lab(ws_m, f"A{mr}", "P/S 年内高")
for c in HC:
    fml(ws_m, f"{c}{mr}", f"={c}{RHI}/{R(S_HIST, c+str(HSPS))}", MX, link=True)
mr += 1
lab(ws_m, f"A{mr}", "P/S 年内低")
for c in HC:
    fml(ws_m, f"{c}{mr}", f"={c}{RLO}/{R(S_HIST, c+str(HSPS))}", MX, link=True)
mr += 2
band(ws_m, mr, "② 同业对照(2026-06 估算口径, P/S 或 EV/Sales TTM)", 11); mr += 1
for col, h in zip("ABC", ["公司", "倍数", "业务特征"]):
    ws_m[f"{col}{mr}"] = h; ws_m[f"{col}{mr}"].font = BF; ws_m[f"{col}{mr}"].fill = CH
mr += 1
for nm, v, note in [("住友电工(5802.T)", 1.1, "InP 老大但衬底只占集团个位数% — 不可比, 仅示意"),
                    ("中际旭创(300308)", 8.0, "800G/1.6T 模块龙头, A 股 AI 光通信锚"),
                    ("新易盛(300502)", 10.0, "同上, 弹性更大"),
                    ("Lumentum(LITE)", 8.5, "EML 双寡头之一, NVDA 注资锁产能"),
                    ("Coherent(COHR)", 4.0, "光器件+InP 自有 fab"),
                    ("云南锗业(002428)", 15.0, "A 股 InP/Ge 衬底国产替代标的, 情绪溢价"),
                    ("IQE(IQE.L)", 2.0, "外延代工, 链上下一站")]:
    ws_m[f"A{mr}"] = nm; ws_m[f"A{mr}"].font = BF
    inp(ws_m, f"B{mr}", v, None, MX)
    logic(ws_m, f"C{mr}", note)
    mr += 1
mr += 1
band(ws_m, mr, "③ 读法 — 给『估值倍数假设』的输入", 11); mr += 1
mtext(ws_m, mr, "AI 光链可比公司 P/S 集中在 4-10x(已盈利、增速 40-60%)。AXT 当下 TTM 60x 远超同链所有环节, 隐含的是'2027-28 收入×3 + 净利率 25%'的完全兑现。目标倍数取法: Base 9.1x(可比上沿, 因 AXT 增速更高但有许可/周期双风险), Bull 13.6x(稀缺卡点溢价), Bear 4.1x(回到可比下沿)。", "K", 3)
for col, w in zip("ABC", [20, 10, 60]):
    ws_m.column_dimensions[col].width = w

# ════ 5 估值倍数假设 ════
ws_ma = wb.create_sheet(S_MA)
ma = write_multiple_assumptions(ws_ma, {
    "title": "估值倍数假设 — 主线 EV/Sales(P/S 近似) 三层分解",
    "intro": "三案目标 P/S 在此拍定; 『情景切换』引用切换, 『情景估值』套用到前瞻 SPS。",
    "why_text": ("为什么 EV/Sales(P/S) 做主线: ①AXT 连亏三年、26Q2 才指引转正, 盈利基数不可靠, P/E 在拐点年失真(周期底 EPS≈0); "
                 "②收入是产能×价格的直接映射, 与物理锚链同构, 倍数假设的含义清晰('每 1 美元 InP 收入市场愿付几美元'); "
                 "③净现金 ~$646M(~$9.2/股)未加回, P/S 比 EV/Sales 略保守, 作为安全垫。"
                 "支线 P/E: 对 2028E 起的盈利做交叉验证——任何情景下隐含 forward P/E 超过 60x 即视为倍数过度乐观的红灯。"
                 "P/B 地板: BPS 现 ~$13.4(增发后权益 ~$875M/65.4M 股), 1.5x P/B ≈ $20 是熊市地板参考。"),
    "method_text": "目标 P/S = ①历史周期峰值 × ②结构溢价 × ③情绪值。历史列由实际 P/S 反推情绪值(一致性检验)。",
    "peak": 2.6, "peak_note": "2021 年末实际 P/S 2.66x = 上一轮景气(半导体超级周期)市场愿付的峰值倍数。",
    "premium": 3.5, "premium_note": "regime change: 商品衬底厂→AI 光互连卡点材料。理由: 全球前二份额+200% 提价权+backlog $100M+收入 3 年 ×3 的可见性; 对标 AI 光链可比(4-10x)上沿之上。突破历史带的全部理由在此, 不另设。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "hpb_row": HPS,
    "cases": [("Bear", [0.45]*5, "AI capex 减速+供给释放, 衬底回归商品定价, 倍数回可比下沿(~4.1x)"),
              ("Base", [1.0]*5, "付足峰值×溢价 = 9.1x, AI 光链可比上沿, 对应高增长+待验证盈利"),
              ("Bull", [1.5]*5, "稀缺卡点 FOMO 延续(对美放开/CPO 放量), 13.6x, 高于全链可比")],
    "sent_note": "情绪值: 1.0=付足峰值×溢价; 历史列=实际 P/S÷9.1 反推(2021 ~0.29 → 当下 TTM ~6.6, 可见当前情绪远超 Bull)。",
    "reconcile_text": ("卖方目标价均值 $96.5 对应 2026E P/S ~40x、2028E P/S ~20x——他们实质上把'缺货+产能翻倍'当成永续状态定价。"
                       "本模型敢给低于街 50% 的数, 依据: ①扩产周期 18-24 个月意味着 2027H2-2028 全行业供给集中落地(住友+40%/JX×3/Coherent×5/云南锗业×3 全部公告在案); "
                       "②AXT 自身 2022→2023 收入 -46% 证明衬底是周期品; ③需求侧 TAM(光模块颗数×衬底价值量)封顶后, 公司产能雄心(2029-30 $320M+)无处安放。"),
    "source_text": "片价: BigGo/财联社 2026-04; 出货: TrendForce 2026-02; 可比倍数: 2026-06 各市场行情估算。历史 P/S: 本模型『历史财务与估值』实拉实算。",
})

# ════ 6 情景切换 ════
ws_sw = wb.create_sheet(S_SW)
sw = write_scenario_switch(ws_sw, {
    "title": "情景切换 — 全模型唯一情景参数库", "cases": CASES, "default": "Base",
    "usage": "B2 下拉切换 Bear/Base/Bull → 全链(TAM→InP 收入→利润→隐含价)整体变档。各杠杆取值与故事见行内说明; 概率: Bear 20% / Base 50% / Bull 30%。",
    "triggers": [("Bear", "出口许可再度收紧或对美持续不通+任一季 InP<$15M+2吋现货价跌破 $1,500+同业扩产提前——任两条成立即落 Bear。"),
                 ("Base", "许可常态化(欧/日/中国内), 产能按计划爬坡但 2027 起需求约束 bind, 片价 2028 起温和回归。"),
                 ("Bull", "对美 InP 许可获批+CPO/1.6T 超预期放量+寡头价格纪律维持, 公司产能计划全兑现满产满销。")],
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "levers": [
        {"key":"mod","name":"光模块出货乘数(×Base)","fmt":N2,
         "vals":{"Bear":[0.85,0.80,0.75,0.70,0.70],"Base":[1,1,1,1,1],"Bull":[1.10,1.20,1.25,1.30,1.30]},
         "desc":"对物理锚 Base 出货量(63/95/125/150/170 百万只)的乘数 = AIDC capex 兑现度。",
         "stories":{"Bear":"AI capex 2027 后减速, 光模块重演 2022-23 去库存","Base":"TrendForce/LightCounting 路径","Bull":"CPO+scale-up 网络放量, 1.6T 提前"}},
        {"key":"asp","name":"衬底片价乘数(×Base)","fmt":N2,
         "vals":{"Bear":[0.90,0.75,0.60,0.50,0.45],"Base":[1,1,1,1,1],"Bull":[1.05,1.15,1.25,1.20,1.15]},
         "desc":"对物理锚 Base 片价($2,300→$2,000)的乘数 = 供给释放 vs 寡头纪律。",
         "stories":{"Bear":"2027-28 扩产集中落地, 价格腰斩回 $1,000 量级(商品周期均值回归)","Base":"CR3 95% 寡头维持纪律, 温和回落","Bull":"缺货延续到 2028, 再提价"}},
        {"key":"cap","name":"AXT InP 产能年化($M)","fmt":N0,
         "vals":{"Bear":[70,110,150,170,180],"Base":[90,190,270,300,320],"Bull":[95,220,330,400,450]},
         "desc":"公司指引: 2026 底产能翻倍(run-rate ~$35M/季), 2027底-2028初 $65-70M/季。",
         "stories":{"Bear":"许可+爬坡+greenfield 延期, 打六折","Base":"按公司计划兑现","Bull":"brownfield 超额+greenfield 提前"}},
        {"key":"share","name":"AXT 可获份额上限","fmt":PCT,
         "vals":{"Bear":[0.28]*5,"Base":[0.32,0.40,0.45,0.46,0.46],"Bull":[0.32,0.42,0.48,0.50,0.50]},
         "desc":"AXT 占行业 InP 衬底 TAM 的上限(当前 ~30-40%, CEO 自称 40%)。",
         "stories":{"Bear":"住友/JX/国产同步扩产, 份额守不住","Base":"AXT 扩产快于日系, 份额升到 ~45%","Bull":"对美放开+6吋先发, 份额 50%"}},
        {"key":"opm","name":"衬底分部 OPM","fmt":PCT,
         "vals":{"Bear":[0.05,0.08,0.08,0.05,0.03],"Base":[0.12,0.22,0.28,0.27,0.25],"Bull":[0.15,0.28,0.34,0.33,0.32]},
         "desc":"InP 分部营业利润率(含分摊费用)。26Q1 公司毛利率 29.6% 且提价传导中。",
         "stories":{"Bear":"价格战+稼动率不足, 回到微利","Base":"涨价+满产, 2028 峰值 28% 后随价格回落","Bull":"量价齐升+6吋摊薄成本"}},
        {"key":"noninp","name":"非InP 营收($M)","fmt":N0,
         "vals":{"Bear":[58,56,54,52,50],"Base":[62,66,70,73,76],"Bull":[65,72,80,86,92]},
         "desc":"GaAs+Ge+原材料合计(GaAs LiDAR 温和增长+原材料随金属价)。",
         "stories":{"Bear":"GaAs 价格竞争+金属价回落","Base":"低个位数增长","Bull":"LiDAR VCSEL+1.6T 多模放量"}}],
    "linked": [{"key":"sent","name":"P/S 情绪值(链『估值倍数假设』)","fmt":N2,
                "src_sheet":S_MA,"src_row0":ma["sent_row0"],
                "note":"三案行=跨表链接估值倍数假设; 当前案行 INDEX 切换。"}],
    "derived": [{"name":"目标 P/S(当前案)","fmt":MX,
                 "fml":lambda col: f"={R(S_MA, ma['pk_cell'])}*{R(S_MA, ma['pr_cell'])}*{col}{{SENT}}",
                 "note":"= 峰值 2.6x × 结构溢价 3.5x × 情绪值(当前案)。历史列=实际 P/S 复现。"}],
})
# derived 公式占位替换(需要 SWACT['sent'] 行号)
TPS_ROW = sw["derived_rows"]["目标 P/S(当前案)"]
for col in ALLC:
    ws_sw[f"{col}{TPS_ROW}"] = f"={R(S_MA, ma['pk_cell'])}*{R(S_MA, ma['pr_cell'])}*{col}{sw['SWACT']['sent']}"

# ════ 7 物理锚 ════
ws_a = wb.create_sheet(S_ANC)
anc = write_anchor(ws_a, {
    "title": "AI 光互连 → InP 衬底 行业基座 (Base 路径)",
    "all_cols": ALLC, "all_years": ALLY,
    "series": [
        ("AIDC capex ($B)", [20,30,70,200,488,830,1050,1240,1380,1500], "", N0),
        ("800G+1.6T 光模块出货(百万只)", [0,0,2,10,24,63,95,125,150,170], "", N0),
        ("单模块 InP 激光器颗数(加权)", [None,None,8.0,7.0,6.5,6.0,5.5,5.0,4.5,4.0], "", N1),
        ("AI InP 衬底需求(万片, 折2吋)", [None]*10, "", N1),
        ("2吋等效片价($/片)", [700,700,700,800,1200,2300,2200,2200,2100,2000], "", N0),
        ("AI InP 衬底 TAM($M)", [None]*10, "", N0),
        ("非AI InP 基本盘($M)", [120,125,115,130,140,160,170,185,195,205], "", N0),
        ("行业 InP 衬底 TAM($M)", [None]*10, "", N0)],
    "yoy_row": "800G+1.6T 光模块出货(百万只)",
    "source_note": ("AIDC capex = 共享基座 compute-aidc-base.json v1.0.0 (CreditSights 2025A $488B / TrendForce 2026E $830B / GS 基线)。"
                    "出货: TrendForce 2026-02(2025 24M→2026 ~63M, +2.6x); 2027-30 本研究外推(50%→32%→20%→13%)。"
                    "颗数: 可插拔 8×EML → 硅光 2-4×CW → CPO ~1颗/端口, 按渗透率加权递减(LightCounting: SiPh 2026>50%)。"
                    "片数 = 出货×颗数×1e6 ÷ 7,000(die/2吋片, Wukong Research) ÷ 0.6(良率×利用率)。"
                    "片价: 2025 初 ~$800 → 2026-04 $2,300-2,500(财联社/BigGo); Base 假设寡头纪律下 2028 起温和回落。"
                    "非AI 基本盘: 电信相干/传感/RF(InP 纯衬底市场 2024 ~$175M, QYResearch, 减 AI 切片回推)。"),
    "role_note": "本表是全模型最底层输入(Base 路径); 情景乘数在『情景切换』, 调节在『分部测算』。验收: 改本表任意出货/片价 → 隐含股价必须动。",
})
RA = anc["row_of"]
R_MOD, R_LPM = RA["800G+1.6T 光模块出货(百万只)"], RA["单模块 InP 激光器颗数(加权)"]
R_WAF, R_PRC = RA["AI InP 衬底需求(万片, 折2吋)"], RA["2吋等效片价($/片)"]
R_AIT, R_NAI, R_TAM = RA["AI InP 衬底 TAM($M)"], RA["非AI InP 基本盘($M)"], RA["行业 InP 衬底 TAM($M)"]
for i, c in enumerate(ALLC):
    if i >= 2:  # 2023 起有颗数
        fml(ws_a, f"{c}{R_WAF}", f"={c}{R_MOD}*{c}{R_LPM}/42", N1)
        fml(ws_a, f"{c}{R_AIT}", f"={c}{R_WAF}*{c}{R_PRC}/100", N0)
    fml(ws_a, f"{c}{R_TAM}", f"=SUM({c}{R_AIT},{c}{R_NAI})", N0)
logic(ws_a, f"L{R_WAF}", "= 出货(百万只)×颗数 ÷42 (派生: ×1e6÷7000 die/片÷0.6 良率÷1e4 换万片)。")
logic(ws_a, f"L{R_AIT}", "= 万片 × 片价 ÷100 (换 $M)。2026E ≈ 9.0 万片 × $2,300 ≈ $207M。")
ws_a.column_dimensions["L"].width = 56

# ════ 8 分部测算 ════
ws_s = wb.create_sheet(S_SEG)
seg = write_segment_model(ws_s, {
    "title": "分部测算 — 行业 TAM(情景调节) → AXT InP 切片(份额∧产能取小)",
    "all_cols": ALLC, "all_years": ALLY, "logic_col": "N",
    "groups": [
        ("行业盘子(当前情景)", [
            ("TAM·情景调节后($M)", None, N0, "=(锚 AI TAM × 出货乘数 + 非AI 基本盘) × 片价乘数(当前案)。历史列=锚表实际。"),
        ]),
        ("AXT InP 切片", [
            ("需求侧切片: TAM×份额($M)", None, N0, "= TAM(调节后) × AXT 份额上限(当前案)。"),
            ("供给侧上限: 产能年化($M)", None, N0, "= 链『情景切换』产能行(当前案)。"),
            ("AXT InP 营收($M)", None, N0, "= MIN(需求切片, 产能上限) ★ 全模型核心行: Base 2027 起需求约束 bind(低于产能), 这是与卖方的最大分歧。历史列=历史财务实际。"),
            ("隐含行业份额", None, PCT, "= AXT InP ÷ TAM(调节后), 自检: 不应超过 55%(住友不会坐视)。"),
        ])],
})
M = seg["m"]
R_TAMQ, R_DEM, R_SUP, R_INP, R_SHR = (M["TAM·情景调节后($M)"], M["需求侧切片: TAM×份额($M)"],
    M["供给侧上限: 产能年化($M)"], M["AXT InP 营收($M)"], M["隐含行业份额"])
for c in HC:
    fml(ws_s, f"{c}{R_TAMQ}", f"={R(S_ANC, c+str(R_TAM))}", N0, link=True)
    fml(ws_s, f"{c}{R_INP}", f"={R(S_HIST, c+str(ha['seg_rows']['InP 衬底']))}", N0, link=True)
    fml(ws_s, f"{c}{R_SHR}", f"={c}{R_INP}/{c}{R_TAMQ}", PCT)
for c in FCf:
    fml(ws_s, f"{c}{R_TAMQ}", f"=({R(S_ANC, c+str(R_AIT))}*{c}{sw['SWACT']['mod']}+{R(S_ANC, c+str(R_NAI))})*{c}{sw['SWACT']['asp']}".replace(f"{c}{sw['SWACT']['mod']}", f"{R(S_SW, c+str(sw['SWACT']['mod']))}").replace(f"{c}{sw['SWACT']['asp']}", f"{R(S_SW, c+str(sw['SWACT']['asp']))}"), N0, link=True)
    fml(ws_s, f"{c}{R_DEM}", f"={c}{R_TAMQ}*{R(S_SW, c+str(sw['SWACT']['share']))}", N0, link=True)
    fml(ws_s, f"{c}{R_SUP}", f"={R(S_SW, c+str(sw['SWACT']['cap']))}", N0, link=True)
    fml(ws_s, f"{c}{R_INP}", f"=MIN({c}{R_DEM},{c}{R_SUP})", N0)
    fml(ws_s, f"{c}{R_SHR}", f"={c}{R_INP}/{c}{R_TAMQ}", PCT)

# ════ 9 利润与收入假设 ════
ws_f = wb.create_sheet(S_FUND)
fr = write_fundamentals(ws_f, {
    "title": "利润与收入假设 ($M → EPS/SPS/BPS)",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "logic_col": "N", "s_hist": S_HIST, "ha": ha, "share_fix_col": "G",
    "assum_groups": [("利润假设(历史列=公司整体实际 OPM 打底)", [
        {"name":"衬底分部 OPM","vals":[0.094,0.089,-0.285,-0.149,-0.249,None,None,None,None,None],
         "fmt":PCT,"logic":"前瞻=链『情景切换』当前案; 历史列=公司整体实际营业利润率(分部口径不可得, 打底对照)。",
         "link":{"sheet":S_SW,"row":sw["SWACT"]["opm"]}},
        {"name":"非InP OPM","vals":[0.094,0.089,-0.285,-0.149,-0.249,0.08,0.08,0.08,0.08,0.08],
         "fmt":PCT,"logic":"GaAs 稳态+原材料低双位数毛利, 合计 8% 固定假设。历史列同上打底。"},
        {"name":"净利转换系数","vals":[None]*5+[0.85]*5,"fmt":N2,
         "logic":"OP→NI: 税+通美 14.5% 少数股东拖累, $700M 现金利息部分对冲 → 0.85。","nm_cols":HC},
        {"name":"留存率","vals":[None]*5+[1.0]*5,"fmt":N2,"logic":"不分红, 全留存。","nm_cols":HC}])],
    "segments": [
        {"name":"InP 衬底","hist_row":"InP 衬底","fwd":{"sheet":S_SEG,"row":R_INP}},
        {"name":"非InP(GaAs/Ge/原材料)","hist_row":"非InP(GaAs/Ge/原材料)","fwd":{"sheet":S_SW,"row":sw["SWACT"]["noninp"]}}],
    "profit_terms": [(["InP 衬底"],"衬底分部 OPM",True),
                     (["非InP(GaAs/Ge/原材料)"],"非InP OPM",True)],
    "conv_assum":"净利转换系数","retention_assum":"留存率",
    "note_text": ("前瞻股本固定取当下 65,400 千股(2026-04 增发后), 不再假设增发——greenfield 若再融资是下修触发器。"
                  "历史 NI/权益=实际(2021-24 权益为回推估算); 前瞻权益=期初+NI×留存。粗颗粒利润线(营收×OPM), 不做三表勾稽。"),
})
# 追加 SPS 行(主线镜头的前瞻每股)
spr = ws_f.max_row + 1
SH_FIX = R(S_HIST, f"$G${ha['HSH']}")
lab(ws_f, f"A{spr}", "SPS 每股销售($)", b=True); ws_f[f"A{spr}"].fill = OUT
for c in HC:
    fml(ws_f, f"{c}{spr}", f"={c}{fr['REV']}*1000/{R(S_HIST, c+str(ha['HSH']))}", N2)
for c in FCf:
    fml(ws_f, f"{c}{spr}", f"={c}{fr['REV']}*1000/{SH_FIX}", N2)
logic(ws_f, f"N{spr}", "= 总营收×1000 ÷ 股本(前瞻固定 65,400 千股)。主线镜头的'每股'输入。")
fr_ps = dict(fr); fr_ps["BPS"] = spr   # 主线镜头用 SPS 替代 BPS

# ════ 10 情景估值 ════
ws_v = wb.create_sheet("情景估值")
sv = write_scenario_valuation(ws_v, {
    "title": "情景估值(当前案) — 主线 EV/Sales(P/S): 目标倍数 × SPS → 隐含价",
    "intro": "历史列=实际价反推(回测); 前瞻=目标 P/S(当前案)×SPS。支线 P/E 交叉验证: 隐含价÷前瞻 EPS。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf, "hist_years": HY,
    "fx_fwd": 1, "s_hist": S_HIST, "ha": ha, "s_fund": S_FUND, "fr": fr_ps,
    "s_switch": S_SW, "target_row": TPS_ROW, "share_fix_col": "G",
    "yend": YEND, "yavg": YAVG,
    "reading": "读法: 隐含 forward P/E 行是体检线——Base 2028E ~45-50x 对应'30%+ 增速+25% 净利率'尚可自洽; 若任何情景算出 >60x, 说明 P/S 倍数透支, 回『估值倍数假设』下调情绪值。",
    "method": "主线 = 目标 P/S(峰值 2.6×溢价 3.5×情绪)× 前瞻 SPS。净现金 ~$9.2/股未加回(安全垫)。P/B 地板: 增发后 BPS ~$13.4, 1.5x ≈ $20。",
    "concl": "Base(2028E 列) 隐含价 ~$41, 较现价 -53%; 加权(20/50/30) ~$48, -45%。现价 $88.34 ≈ Bull 2028E 隐含价(~$85)——市场已按公司产能计划全兑现+价格不回归定价。结论: 卖出/回避, 等待价格或基本面追上。",
})

# ════ 11 估值对比(三案并排, 防污染: 只引情景切换案行) ════
SWB = sw["SWB"]
def _swc(key, ci):
    return SWB[key] + ci
ws_c = wb.create_sheet("估值对比")
cmp_rows = [
    {"key":"tam","label":"TAM·情景调节($M)","fmt":N0,
     "hist":lambda c,ci,A: f"={R(S_ANC, c+str(R_TAM))}",
     "fwd":lambda c,j,ci,A: f"=({R(S_ANC, c+str(R_AIT))}*{R(S_SW, c+str(_swc('mod',ci)))}+{R(S_ANC, c+str(R_NAI))})*{R(S_SW, c+str(_swc('asp',ci)))}"},
    {"key":"inp","label":"AXT InP 营收($M)","fmt":N0,
     "hist":lambda c,ci,A: f"={R(S_HIST, c+str(ha['seg_rows']['InP 衬底']))}",
     "fwd":lambda c,j,ci,A: f"=MIN({R(S_SW, c+str(_swc('cap',ci)))},{c}{A['tam']}*{R(S_SW, c+str(_swc('share',ci)))})"},
    {"key":"rev","label":"总营收($M)","fmt":N0,"bold":True,
     "hist":lambda c,ci,A: f"={R(S_HIST, c+str(ha['HREV']))}",
     "fwd":lambda c,j,ci,A: f"={c}{A['inp']}+{R(S_SW, c+str(_swc('noninp',ci)))}"},
    {"key":"ni","label":"净利润($M)","fmt":N1,
     "hist":lambda c,ci,A: f"={R(S_HIST, c+str(ha['HNI']))}",
     "fwd":lambda c,j,ci,A: f"=({c}{A['inp']}*{R(S_SW, c+str(_swc('opm',ci)))}+{R(S_SW, c+str(_swc('noninp',ci)))}*0.08)*0.85"},
    {"key":"eps","label":"EPS($)","fmt":N2,
     "hist":lambda c,ci,A: f"={c}{A['ni']}*1000/{R(S_HIST, c+str(ha['HSH']))}",
     "fwd":lambda c,j,ci,A: f"={c}{A['ni']}*1000/{R(S_HIST, '$G$'+str(ha['HSH']))}"},
    {"key":"sps","label":"SPS($)","fmt":N2,
     "hist":lambda c,ci,A: f"={c}{A['rev']}*1000/{R(S_HIST, c+str(ha['HSH']))}",
     "fwd":lambda c,j,ci,A: f"={c}{A['rev']}*1000/{R(S_HIST, '$G$'+str(ha['HSH']))}"},
    {"key":"ps","label":"目标 P/S(该案)","fmt":MX,
     "hist":lambda c,ci,A: f"={R(S_HIST, c+str(HPS))}",
     "fwd":lambda c,j,ci,A: f"={R(S_MA, ma['pk_cell'])}*{R(S_MA, ma['pr_cell'])}*{R(S_MA, c+str(ma['sent_row0']+ci))}"},
    {"key":"px","label":"隐含股价($)","fmt":N2,"bold":True,"out":True,
     "hist":lambda c,ci,A: f"={c}{A['ps']}*{c}{A['sps']}",
     "fwd":lambda c,j,ci,A: f"={c}{A['ps']}*{c}{A['sps']}"},
    {"key":"ipe","label":"隐含 forward P/E","fmt":MX,
     "hist":lambda c,ci,A: f'=IF({c}{A["eps"]}<=0,"N/M",{c}{A["px"]}/{c}{A["eps"]})',
     "fwd":lambda c,j,ci,A: f'=IF({c}{A["eps"]}<=0,"N/M",{c}{A["px"]}/{c}{A["eps"]})'},
    {"key":"up","label":"vs 现价(历史列=回测)","fmt":PCT,
     "hist":lambda c,ci,A: f"={c}{A['px']}/{R(S_HIST, c+str(ha['HPX']))}-1",
     "fwd":lambda c,j,ci,A: f"={c}{A['px']}/{R(S_HIST, '$G$'+str(ha['HPX']))}-1"},
]
cmpres = write_comparison(ws_c, {
    "title": "估值对比 — 三案并排(不随开关变)",
    "intro": ("三案完整复算(只引『情景切换』案行+『估值倍数假设』案行+静态锚, 不引任何当前案行)。"
              "历史列回测: 隐含价=实际 P/S×实际 SPS=实际年末价, vs 行应 ≈0%。"
              "目标年 2028E(产能兑现年): Bear ~$9 / Base ~$41 / Bull ~$85; 概率加权 = 0.2×Bear+0.5×Base+0.3×Bull ≈ $48。"),
    "case_names": CASES, "all_cols": ALLC, "all_years": ALLY,
    "hist_cols": HC, "fwd_cols": FCf, "block_start": 16,
    "rows": cmp_rows,
    "summary": {"band": "三案汇总(2028E 目标年)", "target_col": "I",
        "rows": [("隐含股价 2028E($)","px",N2,"主线 EV/Sales 镜头, 2028E 列",True),
                 ("总营收 2028E($M)","rev",N0,"InP(份额∧产能)+非InP",False),
                 ("EPS 2028E($)","eps",N2,"净利×1000÷65,400 千股",False),
                 ("隐含 forward P/E","ipe",MX,"体检线: >60x = 倍数透支",False),
                 ("vs 现价 $88.34","up",PCT,"上行/下行空间",True)],
        "mcap": {"label":"隐含市值 2028E($M)","key":"px",
                 "expr": f"*{R(S_HIST, '$G$'+str(ha['HSH']))}/1000","note":"= 隐含价 × 65,400 千股"},
        "concl": "概率加权目标 ~$48(-45%); 现价位于 Base 与 Bull 隐含价之间且贴近 Bull → 评级: 卖出。市场为'产能计划全兑现+片价不回归'付了全价, 而这在本模型里是 Bull 不是 Base。"},
})

# ════ 12 综合判断仪表盘 ════
CM = cmpres["CMPA"]
ws_d = wb.create_sheet("综合判断仪表盘")
write_dashboard(ws_d, {
    "title": "综合判断仪表盘 — 基本面拐点 × 估值错位 × 催化剂 × 情绪",
    "usage": "活模型用法: 下列任一读数触发 → 回『情景切换』/『物理锚』改数重算, 不要拍脑袋改结论。",
    "blocks": [
        {"title":"A. 基本面拐点(InP 量价)","rows":[
            ("季度 InP 收入","26Q1 $13.6M(+258% YoY)","<$15M(26Q3 起) = 许可/需求恶化 → 滑向 Bear; 连续 >$22M = Base 产能兑现确认。"),
            ("InP backlog",">$100M(26Q1, 连创纪录)","环比下降 = 下游下单热度退潮的最早信号(领先收入 2-3 个季度)。"),
            ("2吋现货价","$2,300-2,500/片(2026-04)","跌破 $1,500 = asp 走 Bear 路径; 再涨 = Bull。频率: 季度(财联社/行业纪要)。"),
            ("产能里程碑","2026 底翻倍(~$35M/季)目标","26Q4 电话会未确认翻倍 = cap 杠杆下修 20%+。")]},
        {"title":"B. 估值错位(模型 vs 市场)","rows":[
            ("Base 隐含价(2028E)",{"fml":f"='估值对比'!I{CM['Base']['px']}","fmt":N2,"fill":True},"主线镜头 Base 案。"),
            ("Bull 隐含价(2028E)",{"fml":f"='估值对比'!I{CM['Bull']['px']}","fmt":N2},"现价 ≈ 此值 → 市场在按 Bull 定价。"),
            ("概率加权目标价",{"fml":f"=0.2*'估值对比'!I{CM['Bear']['px']}+0.5*'估值对比'!I{CM['Base']['px']}+0.3*'估值对比'!I{CM['Bull']['px']}","fmt":N2,"fill":True},"Bear 20% / Base 50% / Bull 30%。"),
            ("GAP vs 现价",{"fml":lambda ro: f"=B{ro['概率加权目标价']}/{R(S_HIST,'G'+str(ha['HPX']))}-1","fmt":PCT},"<-30% = 卖出区; -10%~+10% = 持有区; >+30% = 买入区。")]},
        {"title":"C. 催化剂日历","rows":[
            ("26Q2 财报(2026-07 末)","指引 $34M / EPS $0.05-0.07","实际 <$30M 或未转正 = Base 产能假设砍 20%。"),
            ("对美 InP 许可","未获批(商务部要求补材料)","获批 = Bull 概率上调(北美 TAM 打开); 本模型 Base 未计入。"),
            ("镓锗对美禁运暂停到期","2026-11-27","不续 = 政策风险重估, Bear 概率上调。"),
            ("通美科创板注册","中止(待更新财务资料)4 年","终止公告 = $49M 回购义务 + 中国融资渠道关闭。")]},
        {"title":"D. 情绪确认","rows":[
            ("当下 TTM P/S",{"fml":f"={R(S_HIST,'G'+str(HPS))}","fmt":MX,"fill":True},"~60x vs 历史带 1-2.6x vs 目标 Base 9.1x(2028E SPS 口径) — 情绪远超 Bull。"),
            ("卖方动作","Northland $125 / 共识 $96.5; Needham/B.Riley 已离场","上调潮 = 情绪顶部特征; 第二家降级 = 拐点确认。"),
            ("内部人","2026-03 Young 家族 $51 减持 3.8 万股","加速减持 = 红灯。")]}],
    "final": {"band": "综合判断",
              "text": ("评级: 卖出(回避)。目标价 = Base 2028E 隐含价 ~$41; 概率加权 ~$48(-45%)。"
                       "业务本身是真受益者(InP 卡点+产能翻倍+backlog 创纪录), 但 $5.8B 市值已为 Bull 剧本付全价; "
                       "且商业模式自带'许可证 beta'(单季 -31% 下修已演示)。等待: ①价格回到 $50 下方(加权目标附近) 或 ②对美许可+片价持稳把 Bull 变成 Base, 再重估。")},
    "tracking": {"intro": "证伪触发器(与决策 memo 同源):",
        "rows": [("季度 InP 收入/backlog","$13.6M / >$100M","Base 产能兑现的关键敏感项","季报+电话会","InP<$15M 或 backlog 环比降 → 切 Bear 重算"),
                 ("2吋现货价","$2,300-2,500","asp 杠杆的关键敏感项","行业纪要/财联社, 季度","跌破 $1,500 → asp 走 Bear 路径"),
                 ("同业扩产投产","住友+40%/JX×3/Coherent×5 公告在案","2028 价格回归时点","公司公告, 半年","投产提前 → 份额+价格双下修"),
                 ("出口许可(对美)","未获批","Bull 的入场券","8-K/商务部公告","获批 → Bull 概率 30%→45%"),
                 ("26Q2 实际","指引 $34M","近端执行力","2026-07 末财报","miss → cap 杠杆 -20%")]},
})

finalize(wb, freeze={"封面": "A2", "综合判断仪表盘": "A2"})
wb.save(OUT_PATH)
print("saved:", OUT_PATH)
