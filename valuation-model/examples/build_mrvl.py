# -*- coding: utf-8 -*-
"""
build_mrvl.py — Marvell Technology (MRVL) 估值模型, 基于 build_kit v2。
物理锚: AIDC CapEx × Data Center 收入强度 (compute-semiconductor 锚, 引用 shared-base)。
USD 公司: FX 全部 = 1。盈利口径 = non-GAAP(街口径; GAAP 含无形摊销/SBC, FY22-25 为账面亏损)。
财年: 止 1 月底; FY20XX ≈ 日历 20XX-1 年(capex 锚按此错位对齐)。
数据 SOT: `VALUATION_INPUT_DIR/MRVL_input.json` 对应的历史快照(2026-06-12)。
"""
import os
from openpyxl import Workbook
import build_kit as K

# ════════════ 0. 全局轴 ════════════
ALLC = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
ALLY = ["FY2022A", "FY2023A", "FY2024A", "FY2025A", "FY2026A",
        "FY2027E", "FY2028E", "FY2029E", "FY2030E", "FY2031E"]
HC, HY = ["B", "C", "D", "E", "F"], ["FY2022", "FY2023", "FY2024", "FY2025", "FY2026"]
FC = ["F", "G", "H", "I", "J", "K"]
FCf = FC[1:]                       # FY2027E..FY2031E
FX_FC = 1
FX_HIST = [1, 1, 1, 1, 1]
CASES = ["Bear", "Base", "Bull"]

S_COVER, S_HIST, S_PX, S_CONS = "封面", "历史财务与估值", "股价走势", "卖方研报共识"
S_HMULT, S_MULT, S_SW = "历史估值倍数", "估值倍数假设", "情景切换"
S_ANCHOR, S_SEG, S_FUND = "AIDC Capex预测", "分部测算", "利润与收入假设"
S_VAL, S_CMP, S_DASH = "情景估值", "估值对比", "综合判断仪表盘"

# 真实月度收盘(Research Data API /price, 2026-06-12 抓取, 月末收盘)
MONTHLY = [("2021-06", 57), ("2021-09", 60), ("2021-12", 87), ("2022-03", 72),
           ("2022-06", 44), ("2022-09", 43), ("2022-12", 37), ("2023-03", 43),
           ("2023-06", 60), ("2023-09", 54), ("2023-12", 60), ("2024-03", 71),
           ("2024-06", 70), ("2024-09", 72), ("2024-12", 110), ("2025-03", 63),
           ("2025-06", 77), ("2025-09", 84), ("2025-12", 85), ("2026-01", 79),
           ("2026-02", 82), ("2026-03", 99), ("2026-04", 165), ("2026-05", 205),
           ("2026-06", 281)]
PX_NOW = 280.71
# 财年口径(止1月底)的年末/年均/高低, 由真实日线算出(脚本 mrvl_px.json)
YEND = {"FY2022": 70, "FY2023": 43, "FY2024": 67, "FY2025": 112, "FY2026": 79}
YAVG = {"FY2022": 66, "FY2023": 51, "FY2024": 53, "FY2025": 79, "FY2026": 77}
YHIGH = {"FY2022": 90, "FY2023": 76, "FY2024": 72, "FY2025": 126, "FY2026": 119}
YLOW = {"FY2022": 53, "FY2023": 34, "FY2024": 37, "FY2025": 56, "FY2026": 49}

wb = Workbook()
wb.remove(wb["Sheet"])

# ════════════ 1. 封面 ════════════
K.write_cover(wb.create_sheet(S_COVER), {
    "title": "Marvell Technology (MRVL) 估值模型",
    "meta": [
        ("报告日期", "2026-06-12"),
        ("数据截止", "FY2026 10-K(2026-03-11) + FY27Q1 10-Q(2026-05-28) + 卖方共识(41家) + 现价 $280.71"),
        ("现价", f"${PX_NOW} | 市值 ~$245.6B (2026-06-11 收盘)"),
        ("时效声明", "FY27Q2 财报预计 2026-08 下旬发布(指引 $2.7B), 财报后必须更新本模型。"),
        ("方法一句话", "物理锚(AIDC capex)→ Data Center 收入=capex×收入强度 + Comm&Other 周期 → 非GAAP利润 → P/E主线×FY2028前瞻EPS → 隐含价; DCF 交叉验证。"),
        ("盈利口径", "全模型用 non-GAAP(街口径)。GAAP 含无形摊销+SBC, FY22-25 账面亏损, P/E 失效; FY26 GAAP 净利含卖汽车以太网给 Infineon 的一次性收益 $1.8B。"),
    ],
    "takeaways": [
        ("① 当下估值位置", "现价 $281 ÷ FY2027E 共识 EPS $4.05 = 69x; ÷ FY2028E $6.17 = 45.5x——市场已按 FY2028 的盈利全额定价, 且付的倍数高于历史 forward 中枢(~32x)。"),
        ("② 核心引擎", "DC 收入强度(DC收入÷AIDC capex): FY2026A 实际 1.25% → FY2027E 1.08%(capex +70% 跑赢 DC +48%) → FY2028E 1.26%(custom 翻倍 + FY28 新客户)。"),
        ("③ 三情景目标价", "FY2028 目标年: Bear $104 / Base $213 / Bull $309, 概率加权(25/50/25) ~$210 vs 现价 $281 = -25%。"),
        ("④ 与共识的位置", "卖方 41 家 strong_buy 但平均目标价 $233 低于现价(追赶式上调); MS $195 / GS $125。本模型 Base $213 落在共识目标价之下、现价之下。"),
        ("⑤ 主要风险", "看空主险=过早下车(Google 签约/新 tier-1 客户落地会把 Bull 变 Base); 看多主险=custom all-or-nothing(单点丢单 2025 年曾把股价打到 $62)。"),
    ],
})

# ════════════ 2. 历史财务与估值 ════════════
ha = K.write_history(wb.create_sheet(S_HIST), {
    "title": "Marvell 历史财务与估值 ($B) — FY2022-2026A + 当下 + FY27Q1",
    "hist_cols": HC, "hist_years": HY,
    "fx_hist": FX_HIST, "fx_now": FX_FC,
    "vals_in_usd": True,
    "segments": [
        ("Data Center 收入", [1.78, 2.59, 2.20, 4.16, 6.10], True),
        ("Communications & Other 收入", [2.68, 3.33, 3.31, 1.61, 2.09], True),
    ],
    "total_now": 8.78,
    "gm_pct": [0.462, 0.505, 0.416, 0.413, 0.510], "gm_now": 0.512,
    "ni": [1.332, 1.823, 1.307, 1.377, 2.466], "ni_now": 2.64,
    "eq": [15.70, 15.64, 14.83, 13.43, 14.31], "eq_now": 14.6,
    "shares": [848, 860, 866, 877, 868], "shares_now": 893,
    "px_end": [70, 43, 67, 112, 79],
    "px_now": PX_NOW,
    "px_avg": [66, 51, 53, 79, 77],
    "quarter": {"col": "H", "label": "FY27Q1(实际)",
                "segs": {"Data Center 收入": (1.833, 0.27),
                         "Communications & Other 收入": (0.585, 0.29)},
                "ni": 0.714, "eq": 14.6, "shares": 893, "fx": 1,
                "note": "FY27Q1(止2026-05-02, 10-Q): 营收 $2,417.8M(+28% YoY), DC $1,832.7M(76%), non-GAAP EPS $0.80(净利≈$714M)。YoY 为同比(vs FY26Q1); 单季不算 P/E·P/B。"},
    "band_note": "非GAAP P/E 历史带 20-71x(71x 为 FY2025 盈利低谷年的虚高); 当下 TTM ~95x = 历史极值, forward 口径见『历史估值倍数』。",
    "notes": [
        ("Data Center 收入", "FY24-26=10-K 披露(FY2026 $6,100.3M, 占74%; FY2025 按季度披露加总≈$4.16B); FY22-23 为旧五分部口径下 data center 端市场估算, 标'估'。含定制 XPU + 电光互联 + 交换/存储。"),
        ("Communications & Other 收入", "= 总营收 − DC。企业网络/运营商/消费残余, FY2025 为去库存深坑, FY26 起周期复苏。"),
        ("HREV", "总营收=10-K/10-Q 实际(FY 止1月底): FY26 $8,194.6M(+42.1%)。当下列=TTM(FY26 − FY26Q1 + FY27Q1 ≈ $8.78B)。"),
        ("HGMP", "GAAP 毛利率(10-K)。non-GAAP 毛利率更高(FY27Q1 58.9%), 此行仅作历史对照。"),
        ("HNI", "净利=non-GAAP(公司 8-K 调节表): FY22 $1,332M/FY23 $1,823M/FY24 $1,307M/FY25 $1,377M/FY26 $2,466M。GAAP FY22-25 为亏损(摊销+SBC), FY26 GAAP $2,670M 含一次性 $1.8B 出售收益, 均不可比。当下=TTM non-GAAP ≈$2.64B。"),
        ("HEQ", "股东权益(10-K GAAP); 回购+派息 ≈ 盈利, 权益基本走平。"),
        ("HSH", "non-GAAP 稀释股本(= non-GAAP 净利 ÷ 公司披露 non-GAAP EPS 反推); 当下 893M(FY27Q1)。SBC 致股本每年 +1.5-2.5%。"),
        ("HPX", "财年末(1月底)收盘价, 由 Research Data API 真实日线取得; 当下=2026-06-11 收盘 $280.71。"),
        ("HPXA", "财年内日收盘均价(真实日线算得)。"),
        ("HPE", "P/E=年末价÷当年 non-GAAP EPS(TTM 口径)。FY2025 71x 为盈利低谷年虚高, FY2023 20x 为 2022 杀估值底。"),
    ],
})

# ════════════ 3. 股价走势 ════════════
def phase_fn(ym):
    if ym <= "2025-09":
        return "① 怀疑期"
    if ym <= "2026-03":
        return "② 预期修复"
    if ym <= "2026-05":
        return "③ 催化引爆"
    return "④ 万亿叙事"

px = K.write_price_chart(wb.create_sheet(S_PX), MONTHLY, {
    "fn": phase_fn,
    "rows": [("① 怀疑期", "市场担心 Maia 丢单/Trainium 归属, 'custom 故事讲不下去', 股价 $62-100 横盘近两年"),
             ("② 预期修复", "2026-04-01 NVIDIA 宣布投资 $2B + NVLink Fusion 接入, $99→$165"),
             ("③ 催化引爆", "2026-05-27 FY27Q1 财报: FY27 指引 ~$11.5B(+40%)/FY28 ~$16.5B(+45%), 互联指引上调至 >70%, $165→$205"),
             ("④ 万亿叙事", "2026-06-02 黄仁勋 Computex 称 MRVL 是'下一个万亿美元公司', 当日 +32.5%; 6/4 摸高 $316 后回落至 $281")],
}, title="Marvell 月度股价 (USD)")

# ════════════ 4. 卖方研报共识 ════════════
K.write_consensus(wb.create_sheet(S_CONS), {
    "title": "卖方研报共识 — 41家; 评级 strong_buy 但平均目标价低于现价",
    "overview": "Yahoo 聚合 41 家: 评级 strong_buy, 平均目标 $233(低于现价 $281!), 区间 $110-321。共识 FY2027 营收 $11.52B / EPS $4.05; FY2028 $16.68B / EPS $6.17 —— 与管理层指引($11.5B/$16.5B)几乎重合, 卖方在'抄指引', 真正的分歧在给多少倍数。",
    "assumptions": [
        ("FY2027 营收/EPS", "共识 $11.52B / EPS $4.05(38家, 区间 $3.81-4.21)。", "分歧小: 指引 $11.5B 已锁, Q2 指引 $2.7B 是下一个验证点。", "Base 链算出 FY27 营收 $11.4B / EPS ~$4.0, 对齐共识(指引可信度高)。"),
        ("FY2028 营收/EPS", "共识 $16.68B / EPS $6.17(40家, 区间 $5.34-7.13)。", "最大分歧: custom 翻倍(Maia 300 0.8M颗 + FY28 新 tier-1 客户)能否全额兑现。", "Base 取 DC 强度 1.26% → 营收 $15.9B / EPS $5.79, 略低于共识——给 custom 兑现折一点风险。"),
        ("互联增速", "管理层 FY27 指引 >70%(MS 建模 ~80%, FY28 ~45%)。", "Broadcom 1.6T DSP 抢份额 vs '供不应求下份额之争不重要'(MS)。", "不单独拆互联, 包含在 DC 强度路径里; 1.6T 份额是仪表盘跟踪项。"),
        ("目标倍数", "MS PT $195(隐含 ~40x CY27)/GS $125(28x normalized $4.50)/最高 $321。", "成长股 forward 30-45x vs 'custom 周期性+客户集中'压倍数。", "三层分解: 历史 forward 中枢 32x × 结构溢价 1.15x × 情绪值, Base FY28 = 36.8x, 见『估值倍数假设』。"),
    ],
    "divergences": [
        "① custom 的 all-or-nothing: Bear(GS $125)按'每代重新竞标、随时可丢单'给商品折价; Bull 按'NVLink Fusion 生态锁定+18 个 design win 组合分散'给平台溢价 —— 决定倍数 28x 还是 45x。",
        "② 评级 vs 目标价矛盾: 41 家 strong_buy 但平均目标 $233 < 现价 $281 —— 卖方在追赶式上调(6/2 单日+32.5% 后 PT 还没跟上), 本模型不跟随现价、按链算。",
    ],
    "stances": [
        "Morgan Stanley(EW, PT $195, 2026-05-28): 长期展望上调但'与 NVDA 股价相近而明年 EPS 差一倍'; 期权隐含跌破 $130 概率 53%。",
        "Goldman Sachs(Neutral, PT $125, 2026-05-13): 28x normalized EPS $4.50, 对 custom 持续性最保守。",
        "JPM(CoWoS 跟踪, 2026-04): MAIA 300 由 Marvell 承做, 2026→2027 单量 0.1M→0.8M 颗。",
        "UBS(2026-05-03): Maia 3 单量近 50 万颗 → 明年 $2B+ 收入弹性。",
    ],
})

# ════════════ 5. 历史估值倍数 ════════════
hm = K.write_hist_multiples(wb.create_sheet(S_HMULT), {
    "title": "历史估值倍数 — 自身历史带 + 当下 + 同业对照(non-GAAP 口径)",
    "intro": "①MRVL 自己历史值多少(non-GAAP TTM P/E 带 20-71x; forward P/E 在财年末口径 20-43x, 中枢 ~32x) ②现在市场给多少(TTM ~95x / forward FY28 45.5x) ③同业光谱。看完再去下一页拍三案倍数。",
    "s_hist": S_HIST, "ha": ha, "hist_cols": HC, "hist_years": HY,
    "yhigh": YHIGH, "ylow": YLOW,
    "fwd_note": "forward P/E ≈69x(现价÷FY2027E 共识 EPS $4.05) / ≈45.5x(÷FY2028E $6.17); 历史财年末 forward(年末价÷下年实际 EPS): FY22末 33x / FY23末 28x / FY24末 43x / FY25末 39x / FY26末 20x → 中枢 ~32x",
    "self_name": "Marvell (MRVL)",
    "self_fwd_pe_label": "≈45.5x (FY28E)",
    "self_note": "本模型标的; forward 推导见『情景估值』。注意 P/B 行对 fabless 轻资产公司无估值含义(权益 $14B 多为并购商誉), 仅作展示。",
    "peers": [
        {"name": "Broadcom (AVGO)", "yearly": None, "cur_pb": None, "cur_pe": 64.2, "fwd_pe": 33.2,
         "note": "定制 ASIC 双寡头老大(Google TPU/OpenAI 在手)+网络垄断; 体量与确定性高一档, forward 33x 是 MRVL 倍数最直接的天花板参照。现价 $385.57, TTM/forward 按 Yahoo 共识 FY26 EPS $11.62。"},
        {"name": "NVIDIA (NVDA)", "yearly": None, "cur_pb": None, "cur_pe": 31.4, "fwd_pe": 22.9,
         "note": "AI 平台垄断者, forward 仅 22.9x(FY27 共识 EPS $8.96)——MS 的核心质疑: MRVL 凭什么比 NVDA 贵一倍。现价 $204.87。"},
        {"name": "Credo (CRDO)", "yearly": None, "cur_pb": None, "cur_pe": 106.7, "fwd_pe": 43.6,
         "note": "互联细分小巨头(AEC); forward 43.6x(FY27 共识 EPS $6.07), 代表互联纯标的的情绪上沿。现价 $264.76。"},
        {"name": "Astera Labs (ALAB)", "yearly": None, "cur_pb": None, "cur_pe": 248.3, "fwd_pe": 122.4,
         "note": "PCIe retimer 新贵, 高倍数小票(forward 122x), 仅作光谱上极端参照。现价 $367.47。"},
        {"name": "标普500(参照下沿)", "yearly": None, "cur_pb": None, "cur_pe": 24.0, "fwd_pe": 21.0,
         "note": "大盘光谱下沿(约数, 2026-06)。"},
    ],
    "reading": "① 自己: TTM 95x 不可作锚(FY26 EPS 还没反映指引爆发), 市场实际按 forward 定价——对 FY28 付 45.5x。② 光谱: MRVL 45.5x(FY28) 高于 AVGO 33x(NTM)、两倍于 NVDA 23x——市场把它当'高确定性成长档'顶格定价, 而它的 custom 业务每代重新竞标、客户集中度(Distributor A 45%)都高于 AVGO。③ 结论: 目标倍数应锚历史 forward 中枢 32x × 有限结构溢价(NVDA 投资+生态位确实质变), 而不是接受当下 45x 作为新常态。→ 下一页三层分解。",
})

# ════════════ 6. 估值倍数假设 ════════════
ma = K.write_multiple_assumptions(wb.create_sheet(S_MULT), {
    "title": "估值倍数假设 — P/E 主线(non-GAAP forward) + 三案目标倍数",
    "intro": "镜头判断+三层分解。『情景切换』引用并切换, 『情景估值』套用当前案, 『估值对比』三案并排。倍数口径: 目标年 FY2028 的 forward P/E(× FY2028E non-GAAP EPS)。",
    "why_text": ("镜头选择: MRVL 是 fabless 轻资产成长公司, '穿越周期持续存在的东西'是盈利(签名 IP + 客户联合开发锁定 + 多年期定制项目), 不是资产(权益 $14B 大半是 Inphi/Cavium 并购商誉, P/B ~17x 无业务含义)→ P/E 主线(资本化盈利)。"
                 "但它不是 NVDA 那种全生态定价权: custom 每代重新竞标(all-or-nothing)、买方是议价能力极强的云巨头、互联面临 1.6T 代际份额战 —— 盈利的'耐用性'介于平台与项目制之间, 所以倍数锚自己历史 forward 中枢而非 NVDA/AVGO 上沿。"
                 "支线用 DCF(在『情景估值』): 检验'现价隐含多长的高增长', 防 P/E 倍数与 EPS 双乐观。GAAP P/E 不可用(FY22-25 摊销+SBC 致账面亏损)。"),
    "why_rows": 6,
    "method_text": "三层分解: ①历史 forward P/E 中枢 32x(FY22-26 各财年末 年末价÷下一年实际 non-GAAP EPS = 33/28/43/39/20x, 取中枢 32x; 不用 TTM 95x 极值) × ②结构溢价 1.15x(NVDA $2B 投资 + NVLink Fusion 生态接入 + 互联 #1 卡位 CPO, 生态位确实质变, 但 custom 竞标属性未变, 只给 15%) × ③情绪值(三案)。一致性检验: 32×1.15×1.32 ≈ 48.6x ≈ 当下市场对 FY28 实付的 45.5x(现价÷FY28共识EPS), 量级复现 ✓ → 当下情绪 ≈1.25-1.32, 处于过热档。",
    "peak": 32.0, "peak_note": "第一层 历史 forward 中枢: 由真实财年末价 ÷ 下一财年实际 non-GAAP EPS 逐年反推(33/28/43/39/20x), 取 32x。刻意不用本轮 re-rating 后的倍数当锚(那已含结构溢价+情绪, 再乘=双重计算)。",
    "premium": 1.15, "premium_note": "第二层 结构溢价: 2026-04 NVDA $2B 投资 + 定制 XPU/scale-up 接入 NVLink Fusion + 与 NVDA 共研硅光 —— 从'二线定制商'升格为'NVDA 钦点互联资产'是结构性变化; 对账线: AVGO forward 33x ≈ MRVL 中枢 32x × 1.03, 给满 1.15x 已隐含 MRVL 倍数略超 AVGO, 不再多给。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "hpb_row": ha["HPE"],
    "cases": [
        ("Bear", [0.95, 0.80, 0.80, 0.80, 0.80], "custom 丢单证伪(Maia 300 单量腰斩或 Trainium4 旁落)→ 市场把它打回'项目制定制商'折价档(P/E ~29x), 参照 2025 年怀疑期 forward 仅 ~25x。"),
        ("Base", [1.10, 1.00, 0.95, 0.90, 0.90], "FY27 万亿叙事余温仍偏热(1.10)→ FY28 付足'中枢×溢价'(1.00, P/E 36.8x)→ 之后随成长成熟 normalize(0.90, ~33x)。"),
        ("Bull", [1.30, 1.15, 1.10, 1.05, 1.00], "Google 签约 + FY28 新 tier-1 客户落地 + 互联持续超指引, '万亿公司'叙事被基本面接住, FY28 给 42.3x(=32×1.15×1.15), 类比 AVGO 2024-25 的 AI re-rating。"),
    ],
    "sent_note": "情绪值=周期/情绪位置。1.0=付足『中枢×溢价』; >1=FOMO 超涨; <1=折价。历史列=实际 TTM P/E÷(中枢×溢价) 反推, 三案同值(注意历史列是 TTM 口径、与前瞻 forward 口径有别, 仅作量级对照: FY25 的 1.9 是盈利低谷年 EPS 失真所致)。",
    "reconcile_text": "卖方隐含倍数: GS 28x(normalized) / MS ~40x(CY27) / 共识目标 $233÷FY28 $6.17≈38x。本模型三案 29-42x 覆盖该区间, Base 36.8x 与 MS/共识量级一致——分歧不在倍数而在 EPS: 我们 Base FY28 EPS $5.79 比共识 $6.17 低 6%(给 custom 兑现折风险)。敢给低于现价的 Base: 现价对 FY28 付 45.5x 已超过'中枢×溢价'(36.8x)的 24%, 这 24% 是黄仁勋评论后两周内涨出来的纯情绪层。",
    "source_text": "第一层=Research Data API 真实财年末价 ÷ 公司 8-K non-GAAP EPS; 第二层=AVGO/NVDA forward 对账(Yahoo 共识, 2026-06-12 实拉); 第三层档位依据『综合判断仪表盘』D 块。MS/GS PT 来自本地研报库(2026-05-28/05-13)。",
})

# ════════════ 7. 情景切换 ════════════
sw = K.write_scenario_switch(wb.create_sheet(S_SW), {
    "title": "情景切换 — 全模型唯一的情景参数库 + 切换开关 (默认 Base)",
    "usage": ("怎么用: B2 是唯一入口——下拉选案 → 各杠杆『当前案』行跟着切 → 整条链(锚→测算→利润→倍数→估值)变档。"
              "三案对比不用切: 『估值对比』恒常三列并排。情景参数只在本页改(蓝字)。"),
    "cases": CASES, "default": "Base",
    "triggers": [
        ("Bear", "JPM CoWoS 跟踪显示 MAIA 300 单量较 0.8M 颗腰斩 / Trainium4 网络侧旁落 / 云 capex 指引集体下修至 <20% → custom 故事二次证伪, 重演 2025 年 $62 的定价逻辑(但有互联资产托底)。"),
        ("Base", "FY27 $11.5B / FY28 $16.5B 指引大体兑现(custom 翻倍 + 互联 >70% 增速), 无新增 tier-1 客户惊喜; 倍数随成长成熟从过热回归中枢×溢价。"),
        ("Bull", "Google 两颗芯片签约落地 + FY28 新 tier-1 XPU 客户 ramp + scale-up optics/CPO 超预期 → FY28 营收超 $19B, 市场接受'准平台'定价。"),
    ],
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "levers": [
        {"key": "capex", "name": "AIDC capex 增速", "fmt": K.PCT, "cols": FC[2:],
         "vals": {"Bear": [0.08, 0.02, 0.05, 0.07], "Base": [0.27, 0.18, 0.11, 0.09], "Bull": [0.39, 0.28, 0.15, 0.10]},
         "desc": "物理锚。FY2027E 对应日历2026 capex $830B 三案共用(四大云厂指引已锁); 分歧在 FY2028E(日历2027)起: Bear=AI 回报质疑发酵踩刹车(+8%), Base=GS 'Tracking Trillions' 基线(+27%), Bull=Barclays 上修(+39%)。",
         "stories": {"Bear": "折旧悬崖+AI 投资回报质疑, hyperscaler 集体踩刹车。", "Base": "GS 基线: 2027 +27% / 2028 +18% / 2029 +11% / 2030 +9%。", "Bull": "推理需求爆发, 企业级+主权 AI 接力。"},
         "hist": [None, 1.00, 1.33, 1.86, 1.44]},
        {"key": "dcint", "name": "DC 收入强度", "fmt": K.PCT,
         "vals": {"Bear": [0.0100, 0.0100, 0.0092, 0.0088, 0.0085],
                  "Base": [0.0108, 0.0126, 0.0130, 0.0128, 0.0125],
                  "Bull": [0.0112, 0.0142, 0.0155, 0.0155, 0.0150]},
         "desc": "DC 收入 ÷ AIDC capex(财年对上一日历年)。历史反推: FY2025A 2.08%(含非AI云基底, 偏高), FY2026A 1.25% ← 锚。FY27E 1.08% 续降是机械结果: capex +70% 而 DC 指引 +48%(GPU 吃掉 capex 增量大头)。FY28 起分歧=custom 翻倍兑现度。",
         "stories": {"Bear": "Maia 300 单量减半+无新客户, 强度跌破 1%——对应 DC FY28 仅 ~$9B。", "Base": "Maia 0.8M 颗+Trainium3 配套+FY28 新客户如期 → 1.26%, 对应 DC ~$13.2B(=指引 $16.5B×80%)。", "Bull": "Google 签约+scale-up optics 翻倍 → 1.42%, DC ~$16.4B。"},
         "hist": [None, None, None, 0.0208, 0.0125]},
        {"key": "opm", "name": "非GAAP 营业利润率", "fmt": K.PCT,
         "vals": {"Bear": [0.330, 0.330, 0.320, 0.320, 0.320],
                  "Base": [0.360, 0.385, 0.390, 0.390, 0.390],
                  "Bull": [0.375, 0.405, 0.410, 0.410, 0.410]},
         "desc": "一根线挂两段(custom 摊薄毛利率但增厚利润额, 经营杠杆随规模释放)。历史(估, 8-K non-GAAP): FY22 33%/FY23 35%/FY24 28%/FY25 28%/FY26 34.5%; FY27Q1 ~35%。共识隐含 FY28 ~38-39%。",
         "stories": {"Bear": "custom 占比升+竞标压价, OPM 卡在 32-33%。", "Base": "规模杠杆: FY27 36% → FY28+ 39%(对齐共识隐含)。", "Bull": "互联(高毛利)超预期放量, 41%。"},
         "hist": [0.330, 0.350, 0.280, 0.280, 0.345]},
    ],
    "linked": [
        {"key": "sent", "name": "情绪值(倍数第三层)", "fmt": K.N2,
         "src_sheet": S_MULT, "src_row0": ma["sent_row0"],
         "note": "三案取值与依据见『估值倍数假设』; 本页只做切换。"},
    ],
})
_pk = f"'{S_MULT}'!{ma['pk_cell']}"
_pr = f"'{S_MULT}'!{ma['pr_cell']}"
_sent_act = sw["SWACT"]["sent"]
_r = sw["next_row"]
K.lab(wb[S_SW], f"A{_r}", "目标P/E(当前案)", b=True)
for _c in ALLC:
    K.fml(wb[S_SW], f"{_c}{_r}", f"={_pk}*{_pr}*{_c}{_sent_act}", K.MX, link=True)
K.logic(wb[S_SW], f"L{_r}", "= 历史 forward 中枢 32x × 结构溢价 1.15x × 当前案情绪值 → 喂『情景估值』的前瞻 P/E。")
SWPE = _r

# ════════════ 8. 物理锚 ════════════
anchor = K.write_anchor(wb.create_sheet(S_ANCHOR), {
    "title": "全球 AI 数据中心 CapEx ($B) — 需求物理盘子",
    "all_cols": ALLC, "all_years": ALLY,
    "series": [("AI 数据中心 capex ($B)",
                [15, 30, 70, 200, 488, 830, None, None, None, None],
                "财年错位: MRVL FY20XX(止1月) 对应日历 20XX-1 年 capex。2021-25=实际(CreditSights/财报); 日历2026=$830B(TrendForce+四大云厂指引>$700B, 三案共用); 之后=锚×当前案增速", K.N0)],
    "yoy_row": "AI 数据中心 capex ($B)",
    "source_note": "口径=全球 AI 数据中心专项 capex(非 hyperscaler 总额), 美 Top5+中国 BAT字节。引自 shared-base/compute-aidc-base.json v1.0.0 (2026-06-03), 与 Hynix/NVDA/TSMC/MU 模型共用同一基座。列标 FY20XXA/E = MRVL 财年, 值 = 对应上一日历年的 capex。",
    "role_note": "作用: DC 收入 = capex × DC 收入强度挂在它上面。改 capex → DC 收入 → 利润 → 隐含价全链动。",
})
CAPEX_ROW = anchor["row_of"]["AI 数据中心 capex ($B)"]
for _i, _c in enumerate(FC[2:]):
    K.fml(wb[S_ANCHOR], f"{_c}{CAPEX_ROW}", f"={FC[1:][_i]}{CAPEX_ROW}*(1+{K.R(S_SW, _c + str(sw['SWACT']['capex']))})", K.N0, link=True)

# ════════════ 9. 分部测算 ════════════
seg = K.write_segment_model(wb.create_sheet(S_SEG), {
    "title": "分部测算 — Data Center(capex×强度) + Comm&Other(周期) ($B)",
    "all_cols": ALLC, "all_years": ALLY, "logic_col": "N",
    "groups": [
        ("AIDC capex 物理锚", [
            ("AIDC capex ($B)", None, K.N0, "= 引自『AIDC Capex预测』。改 capex, DC 收入跟着动。"),
        ]),
        ("Data Center = capex × DC 收入强度", [
            ("DC 收入强度 (%)", None, K.PCT,
             "历史=实际 DC 收入÷对应日历年 capex(公式反推): FY25A 2.08%(含非AI云基底偏高), FY26A 1.25% ← 锚。FY22-24 列大基底÷极小 AI capex 失真, 读作 n.m.。前瞻=『情景切换』当前案: FY27E 1.08%(capex+70% 跑赢 DC+48%)→FY28E 1.26%(custom 翻倍+新客户)。"),
            ("Data Center 收入 ($B)", None, K.N1, "历史取实数; 前瞻=capex×强度。FY27E base $9.0B(=830×1.08%), 对齐'FY27 指引 $11.5B × DC 占比 ~78%'。"),
        ]),
        ("Comm & Other = 周期(上年×(1+增速))", [
            ("Comm&Other 增速", [None, 0.243, -0.006, -0.514, 0.298, 0.18, 0.08, 0.05, 0.04, 0.03], K.PCT,
             "历史=实际(FY25 -51% 为企业网络/运营商去库存深坑, FY26 +30% 复苏); 前瞻: FY27 +18%(FY27Q1 已 +29%, 全年减速), 之后回到 GDP+ 低个位数。非 AI 主线, 三案共用。"),
            ("Comm&Other 收入 ($B)", None, K.N1, "历史取实数; 前瞻=上年×(1+增速)。不挂 capex, 走周期复苏。"),
        ]),
    ],
})
m = seg["m"]
DC_HROW = ha["seg_rows"]["Data Center 收入"]
CM_HROW = ha["seg_rows"]["Communications & Other 收入"]
for col in ALLC:
    K.fml(wb[S_SEG], f"{col}{m['AIDC capex ($B)']}", f"={K.R(S_ANCHOR, col + str(CAPEX_ROW))}", K.N0, link=True)
for col in HC:
    K.fml(wb[S_SEG], f"{col}{m['DC 收入强度 (%)']}", f"={K.R(S_HIST, col + str(DC_HROW))}/{col}{m['AIDC capex ($B)']}", K.PCT, link=True)
    K.fml(wb[S_SEG], f"{col}{m['Data Center 收入 ($B)']}", f"={K.R(S_HIST, col + str(DC_HROW))}", K.N1, link=True)
    K.fml(wb[S_SEG], f"{col}{m['Comm&Other 收入 ($B)']}", f"={K.R(S_HIST, col + str(CM_HROW))}", K.N1, link=True)
for col in FCf:
    K.fml(wb[S_SEG], f"{col}{m['DC 收入强度 (%)']}", f"={K.R(S_SW, col + str(sw['SWACT']['dcint']))}", K.PCT, link=True)
    K.fml(wb[S_SEG], f"{col}{m['Data Center 收入 ($B)']}", f"={col}{m['AIDC capex ($B)']}*{col}{m['DC 收入强度 (%)']}", K.N1)
_prevs = [HC[-1]] + list(FCf[:-1])
for _p, _c in zip(_prevs, FCf):
    K.fml(wb[S_SEG], f"{_c}{m['Comm&Other 收入 ($B)']}",
          f"={_p}{m['Comm&Other 收入 ($B)']}*(1+{_c}{m['Comm&Other 增速']})", K.N1)
for col in FCf:
    wb[S_SEG][f"{col}{m['Data Center 收入 ($B)']}"].fill = K.OUT

# ════════════ 10. 利润与收入假设 ════════════
fr = K.write_fundamentals(wb.create_sheet(S_FUND), {
    "title": "利润与收入假设 — 段OPM(一根线) + 净利转换 + 稀释股本 + 分部营收→利润→EPS/BPS",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
    "assum_groups": [
        ("利润率假设(non-GAAP, 历史实际锚 + 前瞻; 粗颗粒)", [
            {"name": "非GAAP 营业利润率", "vals": [0.330, 0.350, 0.280, 0.280, 0.345, None, None, None, None, None],
             "fmt": K.PCT, "logic": "历史=8-K non-GAAP 调节表(估); 前瞻=『情景切换』当前案。",
             "link": {"sheet": S_SW, "row": sw["SWACT"]["opm"]}},
            {"name": "净利转换率(净利/营业利润)", "vals": [0.880, 0.880, 0.850, 0.850, 0.870, 0.875, 0.875, 0.875, 0.875, 0.875],
             "fmt": K.PCT, "logic": "non-GAAP 营业利润 → 净利: 税(non-GAAP 税率 ~10-13%)+净利息。历史反推 0.85-0.88, 前瞻取 0.875 常数。"},
        ]),
        ("股本与留存", [
            {"name": "稀释股本(mn)", "vals": [848, 860, 866, 877, 868, 910, 925, 940, 955, 970],
             "fmt": K.N0, "logic": "历史=non-GAAP 稀释股本; FY27Q1 已 893M(一季 +22M)。SBC 致每年 +1.5-2%: 前瞻 910→970M。高 SBC 是卖方(MS)点名压制估值的因素, EPS 已按逐年股本摊薄。"},
            {"name": "留存率", "vals": [None, None, None, None, 0.10, 0.40, 0.45, 0.45, 0.45, 0.45],
             "fmt": K.PCT, "logic": "FY26 回购 $2.0B+派息 $205M ≈ 净利 90% 返还 → 权益走平; 前瞻假设返还率降至 ~55%(回购授权余额 $5.5B)。仅用于 BPS 递推, P/B 不作镜头。", "nm_cols": ["B", "C", "D", "E"]},
        ]),
    ],
    "segments": [
        {"name": "Data Center 收入", "hist_row": "Data Center 收入", "fwd": {"sheet": S_SEG, "row": m["Data Center 收入 ($B)"]}},
        {"name": "Communications & Other 收入", "hist_row": "Communications & Other 收入", "fwd": {"sheet": S_SEG, "row": m["Comm&Other 收入 ($B)"]}},
    ],
    "profit_terms": [
        (["Data Center 收入", "Communications & Other 收入"], "非GAAP 营业利润率", False),
    ],
    "conv_assum": "净利转换率(净利/营业利润)", "retention_assum": "留存率",
    "note_text": "分部营收(DC=capex×强度 + Comm 周期)→ 一根 OPM 线 → 净利 → EPS(逐年稀释股本)/BPS/ROE。闭环检查: DC 占总营收 % 行应逐年上行(FY26 74% → FY31 ~86%), 与 thesis '业务重心向 AI 数据中心迁移'同向。下游『情景估值』直接引本表每股。",
})
# ── 补丁: EPS/BPS 前瞻列改用「稀释股本(mn)」逐年行(kit 默认用固定历史列股本, 会少算 SBC 稀释) ──
_shrow = fr["am"]["稀释股本(mn)"]
for _c in FCf:
    K.fml(wb[S_FUND], f"{_c}{fr['EPS']}", f"={_c}{fr['NI']}*1000/{_c}{_shrow}", K.N2)
    K.fml(wb[S_FUND], f"{_c}{fr['BPS']}", f"={_c}{fr['EQ']}*1000/{_c}{_shrow}", K.N2)

# ════════════ 11. 情景估值(P/E 主线 + DCF 支线) ════════════
fr_pe = dict(fr); fr_pe["BPS"] = fr["EPS"]   # 主线镜头按 P/E: 倍数×EPS → 隐含价
sv = K.write_scenario_valuation(wb.create_sheet(S_VAL), {
    "title": "情景估值 — 当前案的逐年隐含价 (P/E 主线; DCF 交叉验证)",
    "intro": "本表输出=『情景切换』当前案(默认Base)。主线: 隐含价=目标P/E(当前案)×前瞻 non-GAAP EPS。历史列用实际财年末价反推倍数(事实); 前瞻是预测、不拟合现价。三案并排见『估值对比』。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf, "hist_years": HY,
    "fx_fwd": FX_FC, "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
    "s_fund": S_FUND, "fr": fr_pe,
    "s_switch": S_SW, "target_row": SWPE, "sw_cell": "B2",
    "yend": YEND, "yavg": YAVG,
    "reading": "主线行读法: 本表『目标倍数/隐含股价』两行即 P/E 主线(历史列=年末价÷当年EPS=实际P/E; 前瞻=目标P/E×当年EPS)。Base 隐含价轨迹 FY27 $160 → FY28 $213(目标年) → FY31 ~$253: FY27 列低于现价不是看空 FY27, 而是说明现价已把 FY28 的盈利全额定价(对 FY28 付 45.5x)。",
    "method": "方法: 整体公司、P/E 主线逐年估。基本面在『利润与收入假设』; 目标倍数在『估值倍数假设』(三层分解); 本表只做最后一步: 目标P/E × 前瞻EPS → 隐含价 + 市值; 下方 DCF 块做支线体检。",
    "concl": "结论(方向性): Base 目标价 $213(FY2028 目标年), vs 现价 $281 = -24%。三情景概率加权(25/50/25) ~$210 = -25%。现价已透支 Bull 情景的大半(Bull 也仅 +10%), 评级: 卖出(基于价格, 非基于基本面)。",
})
# ── DCF 交叉验证块(支线): FCF≈净利×转换率, 5年显性 + Gordon 终值, WACC 贴现 ──
_ws = wb[S_VAL]
_r2 = _ws.max_row + 2
K.band(_ws, _r2, "DCF 交叉验证(支线) — 现价隐含多长的高增长? (FCF≈净利×转换率, WACC 贴现 + Gordon 终值)", 11); _r2 += 1
K.lab(_ws, f"A{_r2}", "WACC"); K.inp(_ws, f"B{_r2}", 0.10, None, K.PCT)
K.logic(_ws, f"D{_r2}", "贴现率: 大盘半导体成长股 9.5-11%, 取 10%(beta~1.5, 无显著债务风险)。")
WACC_C = f"$B${_r2}"; _r2 += 1
K.lab(_ws, f"A{_r2}", "永续增长 g"); K.inp(_ws, f"B{_r2}", 0.04, None, K.PCT)
K.logic(_ws, f"D{_r2}", "FY2031 后永续 4%(名义 GDP+), 即假设 FY31 后高增长结束——这正是支线要检验的保守锚。")
G_C = f"$B${_r2}"; _r2 += 1
K.lab(_ws, f"A{_r2}", "FCF 转换率(FCF/净利)"); K.inp(_ws, f"B{_r2}", 0.95, None, K.PCT)
K.logic(_ws, f"D{_r2}", "fabless 轻资产: FY26 OCF $1.75B vs capex $0.35B; non-GAAP 净利→FCF 取 95%(扣少量 NWC 增量)。")
FCFC_C = f"$B${_r2}"; _r2 += 1
K.lab(_ws, f"A{_r2}", "净债务 ($B)"); K.inp(_ws, f"B{_r2}", 1.33, None, K.N2)
K.logic(_ws, f"D{_r2}", "FY2026 末: 总债务 $3.97B − 现金 $2.64B(10-K)。")
ND_C = f"$B${_r2}"; _r2 += 1
NIr2 = lambda c: K.R(S_FUND, c + str(fr["NI"]))
K.lab(_ws, f"A{_r2}", "FCF ($B, 当前案)", note=True)
for _j, _c in enumerate(FCf):
    K.fml(_ws, f"{_c}{_r2}", f"={NIr2(_c)}*{FCFC_C}", K.N1, link=True)
FCF_R = _r2; _r2 += 1
K.lab(_ws, f"A{_r2}", "贴现因子", note=True)
for _j, _c in enumerate(FCf):
    K.fml(_ws, f"{_c}{_r2}", f"=1/(1+{WACC_C})^{_j+1}", K.N2)
DF_R = _r2; _r2 += 1
K.lab(_ws, f"A{_r2}", "DCF 隐含价 ($/股, 支线)", b=True); _ws[f"A{_r2}"].fill = K.OUT
_pv_terms = "+".join(f"{c}{FCF_R}*{c}{DF_R}" for c in FCf)
_tv = f"{FCf[-1]}{FCF_R}*(1+{G_C})/({WACC_C}-{G_C})*{FCf[-1]}{DF_R}"
K.fml(_ws, f"G{_r2}", f"=(({_pv_terms})+{_tv}-{ND_C})*1000/{K.R(S_FUND, 'G' + str(_shrow))}", K.PX)
K.logic(_ws, f"I{_r2}", "=(Σ显性期 FCF 贴现 + 终值贴现 − 净债务)÷股本。Base ≈$105: 若 FY31 后只剩 4% 永续, 整盘生意只值 ~$105/股 —— 现价 $281 与之的差额, 全部是市场对'FY31 之后仍两位数复合'的预期。P/E 主线 Base $213 介于两者之间 = 隐含 FY28 后增速温和放缓。三角验证结论: 现价处在 DCF 地板($105)与 P/E 主线($213)之上, 透支明确。")
DCF_R = _r2

# ════════════ 12. 估值对比 ════════════
SWB = sw["SWB"]
PX_NOW_REF = K.R(S_HIST, f"G{ha['HPX']}")
_conv = fr["am"]["净利转换率(净利/营业利润)"]
_cmg = m["Comm&Other 增速"]


def _fwdprev(j, A, key):
    return (HC[-1] if j == 0 else FCf[j - 1]) + str(A[key])


cmp_rows = [
    {"key": "cap", "label": "AIDC capex ($B)", "fmt": K.N0,
     "hist": lambda c, ci, A: f"={K.R(S_ANCHOR, c + str(CAPEX_ROW))}",
     "fwd": lambda c, j, ci, A: (f"={K.R(S_ANCHOR, 'G' + str(CAPEX_ROW))}" if j == 0
                                 else f"={FCf[j-1]}{A['cap']}*(1+{K.R(S_SW, c + str(SWB['capex'] + ci))})")},
    {"key": "dc", "label": "Data Center 收入 ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(DC_HROW))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['cap']}*{K.R(S_SW, c + str(SWB['dcint'] + ci))}"},
    {"key": "cm", "label": "Comm&Other 收入 ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(CM_HROW))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'cm')}*(1+{K.R(S_SEG, c + str(_cmg))})"},
    {"key": "rev", "label": "总收入 ($B)", "fmt": K.N1, "bold": True,
     "hist": lambda c, ci, A: f"={c}{A['dc']}+{c}{A['cm']}",
     "fwd": lambda c, j, ci, A: f"={c}{A['dc']}+{c}{A['cm']}"},
    {"key": "ni", "label": "净利 non-GAAP ($B)", "fmt": K.N1, "bold": True,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HNI']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['rev']}*{K.R(S_SW, c + str(SWB['opm'] + ci))}*{K.R(S_FUND, c + str(_conv))}"},
    {"key": "eps", "label": "EPS ($, non-GAAP)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={c}{A['ni']}*1000/{K.R(S_HIST, c + str(ha['HSH']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['ni']}*1000/{K.R(S_FUND, c + str(_shrow))}"},
    {"key": "sent", "label": "情绪值(该案)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_MULT, c + str(ma['sent_row0'] + ci))}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_MULT, c + str(ma['sent_row0'] + ci))}"},
    {"key": "pe", "label": "目标P/E(该案)", "fmt": K.MX,
     "hist": lambda c, ci, A: f"={_pk}*{_pr}*{c}{A['sent']}",
     "fwd": lambda c, j, ci, A: f"={_pk}*{_pr}*{c}{A['sent']}"},
    {"key": "px", "label": "隐含价 ($)", "fmt": K.PX, "bold": True, "out": True,
     "hist": lambda c, ci, A: f"={c}{A['pe']}*{c}{A['eps']}",
     "fwd": lambda c, j, ci, A: f"={c}{A['pe']}*{c}{A['eps']}"},
    {"key": "ipe", "label": "隐含 forward P/E(现价÷该案EPS 体检)", "fmt": K.MX,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HPX']))}/{c}{A['eps']}",
     "fwd": lambda c, j, ci, A: f"={PX_NOW_REF}/{c}{A['eps']}"},
    {"key": "up", "label": "历史: vs 实际年末价 / 前瞻: vs 现价", "fmt": K.PCT,
     "hist": lambda c, ci, A: f"={c}{A['px']}/{K.R(S_HIST, c + str(ha['HPX']))}-1",
     "fwd": lambda c, j, ci, A: f"={c}{A['px']}/{PX_NOW_REF}-1"},
]
cm_sheet = K.write_comparison(wb.create_sheet(S_CMP), {
    "title": "估值对比 — Bear / Base / Bull 三情景目标价并排",
    "intro": ("三情景各自完整推演: 物理锚 → DC/Comm 收入 → 净利 → EPS → 目标P/E → 逐年隐含价。"
              "本表三案恒常并排, 不随『情景切换』变化; case 列只引『情景切换』矩阵行(各案行)+未翻档共用行+静态历史锚。"
              "历史列=同一条链填实际值, 隐含价历史列对照实际财年末价(内置回测; 历史列倍数为 TTM 实际口径×当年情绪反推, 量级对照用)。"),
    "case_names": CASES,
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "block_start": 16,
    "rows": cmp_rows,
    "summary": {
        "band": "三案汇总 (目标年 FY2028E ≈ 12个月维度; 各案触发条件见『情景切换』)",
        "target_col": "H",
        "rows": [
            ("总收入($B)", "rev", K.N1, "= DC(capex×该案强度) + Comm&Other(周期复苏)", False),
            ("净利 non-GAAP($B)", "ni", K.N1, "= 总收入 × 该案OPM × 净利转换 0.875", False),
            ("EPS($)", "eps", K.N2, "= 净利 ÷ FY2028E 稀释股本 925M(SBC 逐年摊薄)", False),
            ("目标P/E", "pe", K.MX, "= 历史 forward 中枢 32x × 结构溢价 1.15x × 该案情绪值", False),
            ("隐含价($)", "px", K.PX, "= 目标P/E × FY2028E EPS", True),
            ("vs 现价", "up", K.PCT, "对照现价 $280.71 的上行/下行空间", True),
        ],
        "mcap": {"label": "隐含市值($B)", "key": "px", "expr": f"*{K.R(S_FUND, 'H' + str(_shrow))}/1000",
                 "note": "= 隐含价 × FY2028E 稀释股本 925M"},
        "concl": "风险收益比(FY2028 目标年 vs 现价 $281): Bear -63% / Base -24% / Bull +10%。概率加权(25/50/25) ≈ $210, -25%。Bull 情景才勉强打平现价 = 现价已基本定价完美剧本; 下行不对称(custom all-or-nothing 一旦证伪重演 2025 年 $62-100 的折价机制)。评级: 卖出(价格), 证伪触发器见『综合判断仪表盘』。",
    },
})

# ════════════ 13. 综合判断仪表盘 ════════════
EPS28 = K.R(S_FUND, "H" + str(fr["EPS"]))
PXD = K.R(S_HIST, "G" + str(ha["HPX"]))
dash = K.write_dashboard(wb.create_sheet(S_DASH), {
    "title": "综合判断仪表盘 — A 基本面拐点 · B 估值错位 · C 催化剂 · D 情绪确认",
    "usage": ("预测引擎是 B(错位)+C(催化剂); 情绪 D 只做 timing 确认+过热刹车。"
              "回测验收: 2025-09 怀疑期底部($62), A块(MAIA 300 回归在途)+B块(forward ~20x vs 该给 36.8x, GAP 大幅为正)指向强烈买入——这套表当时能看到那波 4 倍。"),
    "blocks": [
        {"title": "A. 基本面拐点 — 业务在结构性变好吗?", "rows": [
            ("产品组合迁移", "DC 占收入: FY2024 40% → FY2026 74% → FY27Q1 76% → FY2031E ~86%", "从多元化网络芯片商变形为 AI 数据中心公司, 结构迁移真实(闭环检查: 模型 DC 占比逐年上行 ✓)。"),
            ("生态位质变", "NVDA $2B 投资 + NVLink Fusion 接入 + 共研硅光(2026-04)", "从'被怀疑丢单的二线定制商'升格为 NVDA 钦点互联资产——结构溢价 1.15x 的事实依据。"),
            ("订单可见度", "FY27 ~$11.5B(+40%) / FY28 ~$16.5B(+45%) 指引; 互联 >70%", "管理层连续上修; custom $1.5B 年化在手 + 18 个 design win。"),
            ("A 判断", "【强】", "基本面拐点真实且在加速; 问题不在业务, 在价格。", True),
        ]},
        {"title": "B. 估值错位(预测引擎 ★)— 市场现在给的 vs 基本面该给的", "rows": [
            ("市场现在给(forward P/E vs FY28)", {"fml": f"={PXD}/{EPS28}", "fmt": K.MX, "fill": True},
             "= 现价 ÷ FY2028E 模型 EPS(公式算, 随模型走), 当前 ≈48x。"),
            ("基本面该给(justified P/E)", {"inp": 36.8, "fmt": K.MX},
             "= 历史 forward 中枢 32x × 结构溢价 1.15x(三层分解前两层, 情绪中性)。"),
            ("错位 GAP = 该给÷市场给 − 1",
             {"fml": lambda ro: f"=B{ro['基本面该给(justified P/E)']}/B{ro['市场现在给(forward P/E vs FY28)']}-1", "fmt": K.PCT},
             "GAP 为负 = 市场给的已超过基本面该给 = 情绪定价区; 重估空间已被 6/2 单日 +32.5% 一次性吃完。"),
            ("回测: 2025-09 怀疑期底的读数", "市场给 forward ~20x vs 该给 32x+, GAP +60%", "当时错位为正且大 → 该买; 现在反过来了。"),
        ]},
        {"title": "C. 催化剂 — 什么会逼市场闭合 GAP", "rows": [
            ("FY27Q2 财报(2026-08 下旬)", "待; 指引 $2.7B", "高位 beat 才有增量, miss 杀伤大——赔率不对称。"),
            ("Google 两颗芯片签约", "在谈, 未签(媒体口径)", "签约 = Bull 扳机(custom 第三大客户), 也是本模型从卖出转持有的首要观察项。"),
            ("MAIA 300 单量爬坡", "JPM CoWoS: 2026 0.1M → 2027 0.8M 颗", "Base 的 custom 翻倍主驱动; 月度 CoWoS 分配是最硬的前瞻数据。"),
            ("C 判断", "利好大部分已兑现", "黄仁勋背书+指引上修+NVDA 投资三连发已 price-in; 剩余正催化(Google)有真增量但市场已预付一半。", True),
        ]},
        {"title": "D. 情绪确认 — 只做 timing + 刹车", "rows": [
            ("量价温度计", "14个月 $62→$316(5倍); 6/2 单日+32.5% 创纪录; 6/4 见 $316 后回落 12%", "单日历史最大涨幅+随后高位宽幅震荡 = 典型情绪顶部形态; 仅作温度计, 不进倍数。"),
            ("现价倍数 vs 基本面该给", "forward P/E 45.5x(vs FY28) vs 该给 36.8x", "市场已付出超过基本面该给 ~24% = 情绪定价区(当前隐含情绪值 ~1.25-1.32)。"),
            ("当前档位", "【过热】", "'万亿美元公司'类比 = 4 倍空间叙事先行; 期权市场隐含跌破 $130 概率 53%(MS)显示双向波动预期极大。", True),
            ("衰减扳机", "5 条", "FY27Q2 miss $2.7B / MAIA 300 CoWoS 单量下修 / Google 谈判破裂 / 1.6T DSP 份额丢给 AVGO / 云 capex 指引增速<20%。任一翻 → 下调情绪值重算。"),
        ]},
    ],
    "final": {"band": "★ 综合判断(A+B+C+D 收成一句可执行的话)",
              "text": "A 强(拐点真实)但 B 负(forward 48x vs 该给 36.8x)+C 不对称(利好兑现/利空待验)+D 过热(单日+32.5% 赶顶形态) → 现价 $281 已透支 Bull 大半, 概率加权目标 ~$210(-25%)。评级: 卖出/已持仓减仓——这不是看空公司, 是看空价格。若回调至 $190-215(Base 目标价区)或 Google 签约改写 Bull 概率, 重新评估。"},
    "tracking": {
        "intro": "哪个指标恶化 → 哪个假设先崩 → 触发什么动作。",
        "rows": [
            ("__band__", "一、custom 主驱动(all-or-nothing)"),
            ("MAIA 300 CoWoS 单量", "2026 0.1M → 2027E 0.8M 颗(JPM)", "命门: DC 强度 FY28 1.26% 的最大单一支柱", "JPM TSMC CoWoS 跟踪(月度)", "单量下修>30% → 切 Bear, 重算"),
            ("Google 签约进展", "在谈未签", "Bull 情景扳机(强度 1.42%)", "媒体+管理层电话会", "签约 → Bull 概率上调, 重出 memo"),
            ("__band__", "二、互联与份额"),
            ("1.6T DSP 份额", "MRVL 第一(~6-7成)", "互联是 DC 强度里最稳的一块", "产业链调研/光模块厂月度", "AVGO 份额明显上行 → 下调强度路径"),
            ("互联增速 vs >70% 指引", "FY27 指引 >70%", "管理层最激进的一条指引", "季报分部 commentary", "连续两季 <50% → 下调 Base"),
            ("__band__", "三、需求总盘子"),
            ("AIDC capex 指引", "日历2026 $830B(+70%)", "物理锚盘子", "四大云厂季报 capex 指引", "合计增速 <20% → FY28E capex 下调全链重算"),
            ("__band__", "四、财务与估值"),
            ("FY27Q2 实际 vs $2.7B 指引", "2026-08 下旬", "指引可信度 = Base 的地基", "8-K", "miss → 切 Bear; 大 beat+上修 → 复核 Bull"),
            ("稀释股数", "893M, 一季 +22M", "高 SBC 持续摊薄 EPS", "10-Q 股本表", "年化稀释 >3% → 下调前瞻 EPS"),
        ],
    },
})

# ════════════ 全局格式 + 落盘 ════════════
K.finalize(wb, freeze={
    S_HIST: "B3", S_PX: "B4", S_CONS: "A2", S_HMULT: "B5", S_MULT: "B4", S_SW: "B3",
    S_ANCHOR: "B3", S_SEG: "B3", S_FUND: "B3", S_VAL: "B4", S_CMP: "B6", S_DASH: "B6",
    S_COVER: "A2",
})
out = os.path.join(os.path.dirname(__file__), "..", "out", "MRVL_valuation_model.xlsx")
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("saved:", os.path.abspath(out))
print("sheets:", wb.sheetnames)
