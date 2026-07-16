# -*- coding: utf-8 -*-
"""
build_2513hk.py — 智谱(Knowledge Atlas / Zhipu AI, 02513.HK)估值模型实例。
用 build_kit v2 骨架;单位:基本面 RMB 亿;股价 HKD;FX = HKD/RMB ≈ 1.09。
主线镜头 EV/Sales(净现金公司 ≈ P/S, 目标 P/S = 基准锚 × 结构溢价 × 情绪值);
支线 = P/E 体检(转正前 N/M)+ 远期盈利贴现(2030E NI × 稳态 P/E 贴现回看)。
历史估值带替代:上市仅 5 个月 → 用一级融资估值序列(¥150亿→¥243.77亿→IPO HK$518亿)
的隐含 P/S 作"估值带",叠 2026 上市后真实日线(_price_2513.json, 实拉落盘)。
"""
import os, json
from openpyxl import Workbook
import build_kit as K

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUT_DIR = os.environ.get("VALUATION_INPUT_DIR", os.path.join(REPO_ROOT, "out"))
OUTPUT_DIR = os.environ.get("VALUATION_OUTPUT_DIR", os.path.join(REPO_ROOT, "out"))

# ════════════ 0. 全局轴 ════════════
ALLC = ["B", "C", "D", "E", "F", "G", "H", "I", "J"]
ALLY = ["2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
HC, HY = ["B", "C", "D", "E"], ["2022", "2023", "2024", "2025"]
FC = ["E", "F", "G", "H", "I", "J"]      # 含基年 E=2025A
FCf = FC[1:]                              # 2026E-2030E
FX = 1.09                                 # HKD per RMB(2026-06 中间价口径, 1 RMB≈1.09 HKD)
FX_HIST = [1.09, 1.09, 1.09, 1.09]
CASES = ["Bear", "Base", "Bull"]
PX_NOW = 1087.0                           # 2026-06-12 实时报价(tvremix get_quotes_batch)
SH_NOW = 435.2                            # mn 股 = 市值 HK$4,730 亿 ÷ 现价(全部 H+内资股口径)
SH_PRE = 397.8                            # IPO 前股本(435.2 - 发行 37.4mn)

S_COVER, S_HIST, S_PX, S_CONS = "封面", "历史财务与估值", "股价走势", "卖方研报共识"
S_HMULT, S_MULT, S_SW = "历史估值倍数", "估值倍数假设", "情景切换"
S_ANCHOR, S_SEG, S_FUND = "物理锚-Token盘与政企渗透", "分部测算", "利润与收入假设"
S_VAL, S_CMP, S_DASH = "情景估值", "估值对比", "综合判断仪表盘"

# 真实日线(实拉落盘) → 月末收盘
with open(os.path.join(INPUT_DIR, "_price_2513.json"), encoding="utf-8") as f:
    _bars = json.load(f)["bars"]
_m = {}
for b in _bars:
    _m[b["date"][:7]] = b["close"]      # 同月后值覆盖前值 → 月末收盘
MONTHLY = sorted(_m.items())

wb = Workbook(); wb.remove(wb["Sheet"])

# ════════════ 1. 封面 ════════════
K.write_cover(wb.create_sheet(S_COVER), {
    "title": "智谱 (02513.HK) 估值模型 — EV/Sales 物理锚链",
    "meta": [
        ("报告日期", "2026-06-12"),
        ("数据截止", "FY2025 年报(2026-04 披露)+ 2026-06-12 实时行情 + 12 家卖方一致预期"),
        ("现价 / 市值", "HK$1,087 / HK$4,730 亿(≈¥4,340 亿);2026-01-08 上市, 发行价 HK$116.20"),
        ("时效声明", "解禁(2026-07-08)与 A 股发行定价是未来 6 个月最大变量;中报(2026-08)后必须更新。"),
        ("方法一句话", "物理锚(中国大模型 token 调用盘 × 智谱份额 × 单价 + 政企客户数 × 客单价)→ 分部收入 → 经营利润率路径 → 目标 P/S(基准锚×结构溢价×情绪值)× 前瞻每股营收 → 隐含价;P/E 与远期盈利贴现做体检。"),
    ],
    "takeaways": [
        ("① 当下估值位置", "现价对应 P/S(FY25)≈600x、P/S(2026E 模型)≈148x;一级融资带隐含 P/S 34-120x、IPO 发行 72x——现价远在带外, 属'叙事+筹码'定价 regime。"),
        ("② 核心引擎", "MaaS 收入 = 行业 token 盘(2026-03 已 140 万亿/日)× 智谱份额(~3%)× 单价(三次提价 = 定价权证据);政企 = 客户数 × 客单价(国产化采购底仓)。"),
        ("③ 现价反推", "持有现价要求 2030E 收入 ¥560-1,120 亿(2025 的 77-155 倍, CAGR 139%+), 超出本模型 Bull 案(¥463 亿)上界——现价已透支最乐观路径。"),
        ("④ 三情景目标价(2027E)", "Bear HK$116 / Base HK$345 / Bull HK$666;概率加权(25/50/25)≈ HK$370, 较现价 -66%。"),
        ("⑤ 主要风险与触发", "2026-07-08 解禁、A 股发行定价锚(≥¥1,875 亿 vs 港股 ¥4,340 亿)、DeepSeek V4/Qwen 追平 coding 能力、中报 ARR 兑现度。"),
    ],
})

# ════════════ 2. 历史财务与估值(单位: 亿 RMB; 股价 HKD)════════════
ha = K.write_history(wb.create_sheet(S_HIST), {
    "title": "智谱 历史财务与估值 (亿 RMB) — 2022-2025A + 当下;历史'股价'=一级融资估值隐含价",
    "hist_cols": HC, "hist_years": HY,
    "fx_hist": FX_HIST, "fx_now": FX,
    "vals_in_usd": True,                 # 单币种: 值即亿 RMB, 不做汇率换算
    "ps_scale": 100, "mcap_div": 100,
    "unit_label": "(亿 RMB)", "mcap_label": "市值(亿 RMB)", "fx_label": "FX (HKD/RMB)",
    "cur_label": "当下(FY25+实时)",
    "segments": [
        ("MaaS 云端 API 收入", [0.05, 0.12, 0.475, 1.90], True),
        ("政企本地化及其他 收入", [0.524, 1.125, 2.649, 5.34], True),
    ],
    "total_now": 7.24,
    "gm_pct": [0.546, 0.646, 0.563, 0.41], "gm_now": 0.41,
    "ni": [-0.97, -6.21, -24.66, -47.18], "ni_now": -47.18,
    "eq": [12.0, 18.0, 10.0, 20.0], "eq_now": 58.0,
    "shares": [SH_PRE, SH_PRE, SH_PRE, SH_PRE], "shares_now": SH_NOW,
    "px_end": [32.9, 41.1, 54.8, 66.8],
    "px_now": PX_NOW,
    "band_note": "历史列='一级融资估值带'(投后估值÷当时股本×FX, 非交易价): 2023 ¥150亿 / 2024 ¥200亿 / 2025-05 ¥243.77亿; IPO 发行 HK$518亿 → HK$119/股; 现价 HK$1,087 = 发行价 9.3 倍。",
    "notes": [
        ("MaaS 云端 API 收入", "FY25 ¥1.90亿(+292.6%, 年报); 2024=云端占比 15.2%×总收入; 2022-23 按云端占比 8%/10% 估。MaaS ARR ¥17亿(2025 年报)是先行指标, 与确认收入差值反映 25Q4-26 初爆发。"),
        ("政企本地化及其他 收入", "= 总收入 − 云端; 含企业级智能体 ¥1.66亿(+248.8%)与本地化部署。FY25 占 73.7%。"),
        ("HREV", "总收入 公司年报实际: 0.574/1.245/3.124/7.24 亿 RMB(2025 +131.9%)。"),
        ("HGMP", "毛利率 招股书/年报实际; 2025 降至 41% 因 MaaS 占比上升+算力成本。"),
        ("HNI", "2022-24 为经调整净亏损(招股书); 2025 为报表净亏 ¥47.18亿(含研发费用化与上市开支)。口径混用已标明。"),
        ("HEQ", "股东权益为粗估(招股书未披露口径统一的逐年净资产): 按累计融资−累计亏损推; 当下=2025末估值+IPO净额¥38亿。本模型主线 EV/Sales 不依赖账面, P/B 行仅供参照。"),
        ("HSH", "上市前股本=435.2−发行37.4=397.8mn(口径推算); 当下 435.2mn = 实时市值÷现价。"),
        ("HPX", "历史列=一级融资投后估值 ÷ 当时股本 × FX(HKD), 即'一级估值带隐含价'; 2022 按 ~¥120亿 投后估。当下=2026-06-12 实时价 HK$1,087(tvremix)。"),
    ],
})
# 追加 每股营收 SPS + P/S(实际) 行 — EV/Sales 主线的历史底座
ws_h = wb[S_HIST]
HSPS = ws_h.max_row + 1
K.lab(ws_h, f"A{HSPS}", "SPS 每股营收(RMB)", note=True)
for c in HC:
    K.fml(ws_h, f"{c}{HSPS}", f"={c}{ha['HREV']}*100/{c}{ha['HSH']}", K.N2)
K.fml(ws_h, f"G{HSPS}", f"=G{ha['HREV']}*100/G{ha['HSH']}", K.N2)
HPS = HSPS + 1
K.lab(ws_h, f"A{HPS}", "P/S (实际, 主线底座)", b=True); ws_h[f"A{HPS}"].fill = K.OUT
for c in HC:
    K.fml(ws_h, f"{c}{HPS}", f"={c}{ha['HPX']}/({c}{HSPS}*{c}{ha['HFX']})", K.MX)
K.fml(ws_h, f"G{HPS}", f"=G{ha['HPX']}/(G{HSPS}*G{ha['HFX']})", K.MX); ws_h[f"G{HPS}"].fill = K.CUR
K.logic(ws_h, f"H{HPS}", "历史列=一级融资估值隐含 P/S(2023 ~120x → 2024 ~64x → 2025 ~34x, 随收入放大递减); 当下=现价÷FY25 SPS ≈600x。一级带是本模型的'历史估值带'替代(上市仅 5 个月)。")

# ════════════ 3. 股价走势(2026 上市后真实日线 → 月末)════════════
def phase_fn(ym):
    if ym <= "2026-01":
        return "① 上市与定价"
    if ym <= "2026-02":
        return "② GLM-5 重估"
    if ym <= "2026-04":
        return "③ 提价验证"
    return "④ 解禁/A股博弈"

px = K.write_price_chart(wb.create_sheet(S_PX), MONTHLY, {
    "fn": phase_fn,
    "rows": [("① 上市与定价", "2026-01-08 发行 HK$116.2, 首日盘中破发后快速翻倍"),
             ("② GLM-5 重估", "2026-02-11 GLM-5 发布(昇腾全栈+MIT开源)+ Coding Plan 提价 30%, 股价 224→575"),
             ("③ 提价验证", "GLM-5.1 再提价 10%, MaaS ARR ¥17亿 披露, 股价冲 HK$1,028"),
             ("④ 解禁/A股博弈", "5/29 盘中 HK$1,993 见顶; A股 150亿 募资公告+7/8 解禁临近, 两周回撤 -33% 至 HK$1,074")],
}, title="智谱 02513.HK 月末收盘 (HKD, 2026-01 上市)")

# ════════════ 4. 卖方研报共识 ════════════
K.write_consensus(wb.create_sheet(S_CONS), {
    "title": "卖方研报共识 — 12 家覆盖; 目标价均值 HK$1,171, 区间 HK$900-2,001",
    "overview": "总览: 12 家覆盖以买入/跑赢为主, 目标价均值 HK$1,171(vs 现价 HK$1,087 仅 +8%), 区间 HK$900-2,001。共识叙事 = 'A+H 唯一纯大模型标的稀缺性 + GLM-5 定价权 + 国家冠军', 估值方法普遍为远期 P/S 或'对标 Anthropic 一级估值折算', 极少给出可证伪的收入-利润链。",
    "assumptions": [
        ("2026E 收入", "共识 ¥32.4 亿(区间 28-38 亿): MaaS ARR ¥17亿 兑现 + 政企增长。", "ARR→确认收入的转化率与确认节奏。", "我们取 ¥29.4 亿, 略低于共识: ARR 含订阅预付与用量承诺, 确认收入有滞后; 解禁/A股扰动不影响收入但影响政企回款节奏。"),
        ("2027E 收入", "共识 ¥71.7 亿(同比 +121%)。", "提价持续性: DeepSeek V4/Qwen 下一代若追平 coding, 单价回落。", "我们取 ¥63.7 亿: 份额与单价同涨的共识情形只放进 Bull; Base 假设单价持平、靠盘子扩张。"),
        ("目标倍数", "远期 P/S 15-40x(2027E)不等, 部分直接对标 Anthropic 一级隐含 P/S。", "稀缺溢价给多少、给多久。", "三层分解: 全球前沿模型公司一级 forward P/S 锚 18x × A+H 稀缺溢价 1.5x × 情绪值(三案), 见『估值倍数假设』。"),
        ("解禁与A股", "多数报告提示 7/8 解禁但不入模型; A股募 ¥150亿 视为利好。", "A股发行定价若显著低于 H 股, 会向下重锚。", "我们把解禁+A股定价作为情绪值路径的显性扳机(Bear 触发条件), 不只做风险提示。"),
    ],
    "divergences": [
        "① 现价合理性: 共识用'对标 Anthropic'正当化 HK$4,700亿 市值; 我们反推显示现价隐含 2030E 收入 ¥560-1,120亿, 超出 Bull 案——分歧本质是'稀缺/筹码溢价能否当作价值'。",
        "② ARR 质量: 共识把 ¥17亿 ARR 当确认收入的先行指标线性外推; 模型判断其中 Coding Plan 订阅对竞品能力追平高度敏感, 给 Bear 案 40% 折扣。",
    ],
    "stances": [
        "覆盖均值(12家): 买入/跑赢为主 | TP HK$1,171 | 'A+H 稀缺 + GLM-5 定价权 + 国产算力全栈'。",
        "最高 TP HK$2,001: 对标 Anthropic 一级估值映射, 给 2027E P/S ~40x。",
        "最低 TP HK$900: 同样看多基本面, 但对解禁与 A股定价折价做了显性扣减。",
        "本所 2026-05 报告: 跑赢, TP HK$1,450(当时市值 HK$5,100亿)——本次模型按物理锚链下修, 详见决策 memo。",
    ],
})

# ════════════ 5. 历史估值倍数(手写: P/S 口径数据底座)════════════
ws_hm = wb.create_sheet(S_HMULT)
K.hdr(ws_hm, 1, "历史估值倍数 — P/S 口径: 自身一级估值带 + 当下 + 全球/国内可比", 11)
_r = K.mtext(ws_hm, 2, "上市仅 5 个月, 无传统历史估值带 → 用一级融资估值序列的隐含 P/S 作带, 叠 IPO 定价与当下; 同业用已上市 AI 公司 P/S + 美股一级(OpenAI/Anthropic)隐含 forward P/S 折算。看完本页再去下一页拍三案目标 P/S。", "K", 2)
K.band(ws_hm, _r, "① 自身: 一级估值带 → IPO → 当下 (P/S, 分母=当期最近年收入)", 11); _r += 1
for col, h in zip("ABCDEF", ["事件", "时点", "估值", "对应收入(亿RMB)", "隐含 P/S", "说明"]):
    ws_hm[f"{col}{_r}"] = h; ws_hm[f"{col}{_r}"].font = K.BF; ws_hm[f"{col}{_r}"].fill = K.CH
_r += 1
_self_rows = [
    ("B+轮等(2023)", "2023 末", "¥150 亿", 1.245, "多轮累计 ¥25亿+; 美团/阿里/腾讯/启明等"),
    ("国资轮(2024)", "2024-12", "¥200 亿+", 3.124, "中关村科学城领投+30亿新轮; 国资接力开始"),
    ("IPO 前最后一轮", "2025-05", "¥243.77 亿", 7.24, "投后(招股书); 政策性资本定价"),
    ("港股 IPO 发行", "2026-01-08", "HK$518 亿(≈¥475亿)", 7.24, "发行价 HK$116.2; 基石占近 7 成"),
    ("当下", "2026-06-12", "HK$4,730 亿(≈¥4,340亿)", 7.24, "现价 HK$1,087; 较发行 +835%"),
]
_vals_ev = [150, 200, 243.77, 475, 4340]
for i, (ev_rmb, (nm, tp, evs, rev, note)) in enumerate(zip(_vals_ev, _self_rows)):
    rr = _r + i
    K.lab(ws_hm, f"A{rr}", nm, b=(i >= 3))
    K.lab(ws_hm, f"B{rr}", tp, note=True)
    K.lab(ws_hm, f"C{rr}", evs)
    K.inp(ws_hm, f"D{rr}", rev, None, K.N1)
    K.inp(ws_hm, f"E{rr}", round(ev_rmb / rev, 1), None, K.MX)
    K.logic(ws_hm, f"F{rr}", note)
_r += len(_self_rows)
_r += 1
K.band(ws_hm, _r, "② 可比: 已上市 AI 公司 + 美股一级隐含 forward P/S (2026-06 实拉/公开口径)", 11); _r += 1
for col, h in zip("ABCDEF", ["公司", "市值", "收入口径", "P/S", "口径", "业务特征/来源"]):
    ws_hm[f"{col}{_r}"] = h; ws_hm[f"{col}{_r}"].font = K.BF; ws_hm[f"{col}{_r}"].fill = K.CH
_r += 1
_peers = [
    ("智谱 02513.HK", "HK$4,730亿", "FY25 ¥7.24亿 / 2026E ¥29.4亿(模型)", "600x / 148x", "TTM / fwd",
     "本模型标的; TTM 与 forward 双口径均远超全部可比。市值=2026-06-12 实时(tvremix)。"),
    ("商汤 0020.HK", "HK$614亿", "FY25 ~¥45亿(公开口径)", "≈12x", "TTM",
     "国内已上市 AI 第一可比; 亏损收窄中。市值实拉(tvremix 2026-06-12), 收入为年报公开数。"),
    ("第四范式 6682.HK", "≈HK$200亿", "FY25 ~¥60亿(公开口径)", "≈3x", "TTM",
     "企业 AI 解决方案商('AI 外包'估值的下界参照); 数字为公开口径粗估。"),
    ("OpenAI(一级)", "≈$5,000亿(2025末轮)", "2026E ~$30B(媒体口径)", "≈17x", "fwd",
     "全球上沿; 一级估值÷下一年收入 = forward P/S 锚的核心来源。"),
    ("Anthropic(一级)", "≈$3,500亿(2026 初口径)", "2026E run-rate ~$20B(媒体口径)", "≈18x", "fwd",
     "coding/agent API 龙头 = 智谱叙事的对标; 即便它也只有 ~18x forward P/S。"),
]
for i, (nm, mc, rev, ps, kou, note) in enumerate(_peers):
    rr = _r + i
    K.lab(ws_hm, f"A{rr}", nm, b=(i == 0))
    if i == 0:
        ws_hm[f"A{rr}"].fill = K.CUR
    K.lab(ws_hm, f"B{rr}", mc); K.lab(ws_hm, f"C{rr}", rev); K.lab(ws_hm, f"D{rr}", ps)
    K.lab(ws_hm, f"E{rr}", kou, note=True); K.logic(ws_hm, f"F{rr}", note)
    ws_hm.row_dimensions[rr].height = 30
_r += len(_peers) + 1
K.band(ws_hm, _r, "③ 读法 — 给『估值倍数假设』的输入", 11); _r += 1
K.mtext(ws_hm, _r, "① 自己: 一级带隐含 P/S 随收入放大从 120x→34x 递减, IPO 发行 72x, 现价 600x(TTM)/148x(2026E)——现价不在任何带内。② 可比: 全球最强的两家(OpenAI/Anthropic)一级 forward P/S 也只有 17-18x, 国内已上市可比 3-12x。③ 结论: 第一层锚取 18x(全球前沿一级 forward P/S), 第二层 A+H 稀缺/国家冠军溢价 1.5x, 第三层情绪值表达'市场愿意为期权付多少'——现价对应情绪值 ≈5.5(2026E 口径), 即市场付了'锚×溢价'的 5 倍半。", "K", 4)
for col, w in zip("ABCDEF", [18, 16, 30, 14, 10, 56]):
    ws_hm.column_dimensions[col].width = w

# ════════════ 6. 估值倍数假设(三案目标 P/S)════════════
ma = K.write_multiple_assumptions(wb.create_sheet(S_MULT), {
    "title": "估值倍数假设 — EV/Sales 主线: 目标 P/S = 基准锚 × 结构溢价 × 情绪值(三案)",
    "intro": "数据底座在上一页。本页只做判断: ①为什么 EV/Sales 做主线 ②三层分解出三案目标 P/S(对当年收入)。『情景切换』引用并切换, 『情景估值』套用当前案。",
    "why_text": ("镜头选择是业务判断: 智谱 2029E 前无盈利(P/E 失效)、账面被巨亏快速消耗(P/B 失真)、FCF 深负(DCF 全靠终值假设≈空中楼阁)——穿越当下持续存在并增长的只有'收入与其背后的 token 用量/客户数', 所以 EV/Sales(净现金公司≈P/S)做主线, 且强制配 path to profitability(本模型: Base 2029 盈亏平衡、2030 净利率 ~10%, 见『利润与收入假设』)。"
                 "P/E 做支线体检: 隐含价 ÷ 前瞻 EPS, 转正前标 N/M, 2029-30 起检验隐含倍数是否荒谬。第三检验 = 远期盈利贴现(2030E NI × 稳态 P/E ÷ 1.15^4), 防 P/S 镜头把'永远不兑现的收入'资本化。"
                 "镜头迁移触发: 当公司连续两个财年经营利润为正, 主线应迁往 P/E——届时市场也会换镜头(re-rating/de-rating 的本质)。"),
    "why_rows": 6,
    "method_text": "三层分解: ①基准锚 = 全球前沿模型公司一级市场 forward P/S(OpenAI ~17x / Anthropic ~18x → 取 18x, 这是'世界上最强的同类资产'的市场化定价上沿) × ②结构溢价 = A+H 唯一纯大模型标的稀缺 + 国家冠军 + 南向被动配置 → 1.5x × ③情绪值(三案路径, 随收入兑现递减)。一致性检验: 现价 ÷ (锚×溢价) 在 2026E 口径 = 情绪值 ≈5.5 → 市场当前付了基准的 5 倍半, 这就是泡沫层的厚度。",
    "peak": 18.0, "peak_note": "基准锚 18x = OpenAI(≈$5,000亿/2026E $30B≈17x)与 Anthropic(≈$3,500亿/run-rate $20B≈18x)一级 forward P/S; 国内已上市可比 3-12x 在其下方。取上沿因为智谱增速更高(2026E +306%)。",
    "premium": 1.50, "premium_note": "结构溢价 1.5x: A+H 唯一可交易纯大模型标的(筹码稀缺, 基石+国资锁定 7 成)、国家冠军采购确定性、指数纳入与南向流动性。对账线: 商汤无此溢价 P/S 12x, 智谱 IPO 发行 72x——市场确实长期愿意给倍数级溢价。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "hpb_row": HPS,
    "cases": [
        ("Bear", [0.60, 0.45, 0.35, 0.30, 0.28], "解禁抛压兑现 + A股低价定锚 + DeepSeek V4/Qwen 2026H2 追平 coding → 提价回吐, 稀缺溢价坍缩; 情绪值向'锚×溢价'之下折价(0.6→0.28), 对应目标 P/S 16.2x→7.6x。"),
        ("Base", [1.00, 0.80, 0.65, 0.55, 0.50], "A股顺利募 ¥120亿+, 解禁冲击被南向承接; 能力代差维持 12 个月后被部分追平; 情绪值从付足基准(1.0)随收入兑现常态化递减至 0.5, 对应目标 P/S 27x→13.5x——仍高于 OpenAI 当下一级水平, 已计入高增长。"),
        ("Bull", [1.30, 1.10, 0.90, 0.75, 0.65], "GLM-5.2/6 续期能力代差 + 再提价成功 + 主权 AI 大单连发, '中国 Anthropic'叙事被全球接受; 情绪值 1.3 起步缓退, 对应目标 P/S 35.1x→17.6x(2030 仍≈当下 Anthropic 一级水平)。"),
    ],
    "sent_note": "情绪值 = 市场愿付倍数 ÷ (基准锚×结构溢价)。1.0=付足'全球上沿×稀缺溢价'; >1=超涨; <1=折价。历史列=一级估值隐含 P/S 反推(2023 ~4.4 反映一级市场为期权付价, 参考性弱)。",
    "reconcile_text": "卖方 2027E P/S 15-40x vs 我们三案 12.2-29.7x(2027): 区间重叠但中枢更低。凭什么: ①全球上沿(OpenAI/Anthropic 一级)只有 17-18x forward, 给智谱 Base 21.6x(2027)已含 1.5x 稀缺溢价+0.8 情绪, 再高就需要论证'智谱配得上比 OpenAI 更贵'; ②解禁+A股定价两个硬事件在 12 个月内必然落地, 共识把它们当脚注、我们当扳机。",
    "source_text": "第一层: OpenAI/Anthropic 一级估值与收入为 2026 上半年媒体广泛报道口径(量级可靠, 精确值有噪声); 第二层: 对账线=智谱 IPO 发行 P/S 72x vs 商汤 12x; 第三层: 情绪值依据『综合判断仪表盘』D 块(高点回撤 -33%、解禁临近=退潮初段)。",
})

# ════════════ 7. 情景切换 ════════════
sw = K.write_scenario_switch(wb.create_sheet(S_SW), {
    "title": "情景切换 — 全模型唯一的情景参数库 + 切换开关 (默认 Base)",
    "usage": "怎么用: B2 下拉选案 → 案序号派生 → 各杠杆『当前案』行跟着切 → 整条明细链(Token盘→分部→利润→倍数→估值)变档。三案对比不用切:『估值对比』恒常并排。情景参数只在本页改(蓝字); 未列入的假设三案共用。",
    "cases": CASES, "default": "Base",
    "triggers": [
        ("Bear", "①2026H2 DeepSeek V4 / Qwen 新代在 coding/agent 基准追平 GLM-5 且 API 单价显著低于智谱 → 提价回吐; ②7/8 解禁后 30 日减持公告密集 + A股发行估值 <¥2,500亿(较 H 股折价>40%); ③中报云端占比停滞(<28%)或 MaaS ARR 增速大幅回落。任一兑现即翻 Bear。"),
        ("Base", "A股 2026H2 过会、募资 ≥¥100亿; 解禁抛压被南向/被动资金承接(回撤但不崩); GLM-5.2 如期发布维持国内第一梯队; 中报云端占比 30%+、ARR keeps 兑现为确认收入。"),
        ("Bull", "GLM-6 续期对国内竞品的能力代差并再次提价成功; 主权 AI 落地 ≥2 个国家级大单; A股以接近 H 股估值发行(定价向上确认); OpenRouter 份额进入全球前五。"),
    ],
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "levers": [
        {"key": "mshare", "name": "智谱 token 份额(全行业)", "fmt": K.PCT,
         "vals": {"Bear": [0.026, 0.024, 0.022, 0.020, 0.019],
                  "Base": [0.030, 0.032, 0.033, 0.034, 0.035],
                  "Bull": [0.034, 0.038, 0.042, 0.045, 0.048]},
         "desc": "MaaS 收入的份额刀。行业盘(三案共用, 见物理锚页)×份额×单价=MaaS 收入。智谱当前公有云 token 份额第二梯队(火山 49.2%/阿里 27%/百度 17%, IDC 2025H1), 反推 2025 ≈3%。",
         "stories": {"Bear": "豆包/Qwen 价格战 + DeepSeek V4 开源追平 → 高价份额被切, 份额缓降至 1.9%。",
                     "Base": "coding/agent 高价值场景守住, 份额微升至 3.5%(走量市场让给字节/阿里)。",
                     "Bull": "GLM 代差续期 + 海外 z.ai 放量, 份额升至 4.8%(仍不挑战大厂走量盘)。"},
         "hist": [None, None, None, None]},
        {"key": "mprice", "name": "MaaS 实现单价(¥/M token)", "fmt": K.N2,
         "vals": {"Bear": [0.50, 0.42, 0.36, 0.33, 0.30],
                  "Base": [0.60, 0.60, 0.65, 0.72, 0.78],
                  "Bull": [0.68, 0.75, 0.85, 0.95, 1.05]},
         "desc": "blended 实现单价(含缓存/订阅折算)。2026 年三次提价(Coding Plan +30%、API +10%、海外近翻倍)是定价权的核心证据; GLM-5.1 缓存价 $0.475/M 已贴近 Claude Opus 4.5。2025 实现价 ≈¥0.50/M(由收入÷token 反推)。",
         "stories": {"Bear": "竞品追平 → 提价全数回吐并跌破 2025 水平(价格战重启)。",
                     "Base": "提价守住但不再上行, 2028 起靠结构(agent 长任务占比)缓升。",
                     "Bull": "能力代差续期, 沿 Anthropic 路径持续提价至 ¥1.05/M。"},
         "hist": [None, None, None, 0.50]},
        {"key": "gcust", "name": "政企付费客户数(家)", "fmt": K.N0,
         "vals": {"Bear": [900, 1150, 1450, 1700, 1900],
                  "Base": [1000, 1400, 1850, 2300, 2700],
                  "Bull": [1150, 1700, 2400, 3100, 3800]},
         "desc": "政企段的量刀。央企 AI 渗透率 2024 ~5% → 2026E ~20%(物理锚页), 全栈国产化是采购硬指标, 智谱适配 40+ 国产芯片近乎政企标配。2025 活跃付费政企/机构客户 ≈700 家(累计 1,000+ 政企客户口径折算)。",
         "stories": {"Bear": "地方财政紧张拉长采购, 新客增速腰斩——但国产化底仓仍在涨(这就是争议②'确定性底仓'的含义)。",
                     "Base": "渗透率沿 IDC 路径走, 客户数年增 35-40% 渐降。",
                     "Bull": "渗透加速 + 主权 AI 带动海外政企复制。"},
         "hist": [80, 150, 350, 700]},
        {"key": "gticket", "name": "政企客单价(¥百万)", "fmt": K.N2,
         "vals": {"Bear": [1.12, 1.55, 2.10, 2.72, 3.45],
                  "Base": [1.30, 1.80, 2.30, 2.60, 3.00],
                  "Bull": [1.30, 1.90, 2.40, 2.70, 3.00]},
         "desc": "政企段的价刀。历史客单 ¥65-76 万(项目制+订阅), 随大模型项目从试点转向全栈部署, 单项目金额上行(行业大单已到千万级)。",
         "stories": {"Bear": "客户数掉但活下来的都是大单央企 → 客单反而更高(组合效应), 总量仍最低。",
                     "Base": "试点→全栈, 客单稳步抬到 ¥300 万。",
                     "Bull": "与 Base 接近(政企价上行受预算约束, 弹性给量)。"},
         "hist": [0.65, 0.75, 0.76, 0.76]},
        {"key": "opm", "name": "综合经营利润率(含研发)", "fmt": K.PCT,
         "vals": {"Bear": [-2.20, -1.30, -0.75, -0.45, -0.25],
                  "Base": [-1.53, -0.55, -0.13, 0.03, 0.11],
                  "Bull": [-1.25, -0.40, 0.00, 0.10, 0.20]},
         "desc": "path to profitability 的核心刀: 研发(FY25 ¥31.8亿=收入 4.4 倍)近似固定成本, 收入放大即摊薄。历史 OPM: -169%/-499%/-789%/-652%(经调整/报表口径混合)。",
         "stories": {"Bear": "收入上不去 + 算力军备不能停 → 2030 仍亏 ¥26亿, 现金依赖持续融资(生存问题, 见仪表盘)。",
                     "Base": "2029 盈亏平衡、2030 净利率 ~10%——锚 Anthropic 路径(规模化推理毛利+研发摊薄)。",
                     "Bull": "提价+昇腾自有算力降本, 2028 打平、2030 净利率 ~20%。"},
         "hist": [-1.69, -4.99, -7.89, -6.52]},
    ],
    "linked": [
        {"key": "sent", "name": "情绪值(目标 P/S 第三层)", "fmt": K.N2,
         "src_sheet": S_MULT, "src_row0": ma["sent_row0"],
         "note": "三案取值与依据见『估值倍数假设』; 本页只做切换。"},
    ],
})
# 目标 P/S(当前案) = 锚 × 溢价 × 当前案情绪
_pk = f"'{S_MULT}'!{ma['pk_cell']}"
_pr = f"'{S_MULT}'!{ma['pr_cell']}"
_sent_act = sw["SWACT"]["sent"]
_r = sw["next_row"]
K.lab(wb[S_SW], f"A{_r}", "目标 P/S(当前案)", b=True)
for _c in ALLC:
    K.fml(wb[S_SW], f"{_c}{_r}", f"={_pk}*{_pr}*{_c}{_sent_act}", K.MX, link=True)
K.logic(wb[S_SW], f"L{_r}", "= 基准锚 18x × 结构溢价 1.5x × 当前案情绪值 → 喂『情景估值』的前瞻目标 P/S。")
SWPS = _r

# ════════════ 8. 物理锚 [ANCHOR] ════════════
anchor = K.write_anchor(wb.create_sheet(S_ANCHOR), {
    "title": "中国大模型 Token 调用盘 + 央企 AI 渗透 — 需求物理盘子(三案共用)",
    "all_cols": ALLC, "all_years": ALLY,
    "series": [
        ("行业日均 token 调用(万亿/日)", [None, None, 6, 35, 250, 550, 950, 1400, 1900],
         "2025-12 实测 63 万亿/日、2026-03 已 140 万亿+/日(行业数据); 2025 年均 ~35; 前瞻按 2026 年均 250(年中已 140+且指数增长)→ 增速 120%→73%→47%→36% 递减", K.N0),
        ("行业年 token 盘(万亿/年)", [None, None, None, None, None, None, None, None, None],
         "= 日均 × 365(公式)", K.N0),
        ("央企/政企 AI 渗透率", [0.01, 0.02, 0.05, 0.09, 0.14, 0.20, 0.26, 0.31, 0.35],
         "央企渗透率 2024 ~5% → 2026E ~20%(研究稿/行业口径); 驱动政企客户数假设的外生轨道", K.PCT),
    ],
    "yoy_row": "行业日均 token 调用(万亿/日)",
    "source_note": "口径: 中国全市场大模型日均 token 调用(2026-03 公开行业数据 140 万亿+/日, 2025-12 为 63 万亿)。三案共用同一盘子——分歧放在份额与单价(『情景切换』), 不在盘子。央企渗透率为 IDC/公司路演口径。",
    "role_note": "作用: MaaS 收入 = 年 token 盘 × 智谱份额 × 单价 × 0.01(单位换算: ¥/M token × 万亿 token = 百万元 ×0.01=亿)。政企收入挂渗透率驱动的客户数。改本页盘子 → 收入 → 隐含价全链动。",
})
TOK_D = anchor["row_of"]["行业日均 token 调用(万亿/日)"]
TOK_Y = anchor["row_of"]["行业年 token 盘(万亿/年)"]
for _c, _y in zip(ALLC, ALLY):
    if _y in ("2022A", "2023A"):
        K.lab(wb[S_ANCHOR], f"{_c}{TOK_Y}", "n.m.", note=True)
    else:
        K.fml(wb[S_ANCHOR], f"{_c}{TOK_Y}", f"={_c}{TOK_D}*365", K.N0)

# ════════════ 9. 分部测算 ════════════
seg = K.write_segment_model(wb.create_sheet(S_SEG), {
    "title": "分部测算 — MaaS(盘×份额×单价) + 政企(客户×客单价)",
    "all_cols": ALLC, "all_years": ALLY, "logic_col": "N",
    "groups": [
        ("物理锚引入", [
            ("行业年 token 盘(万亿)", None, K.N0, "= 引物理锚页。改盘子, MaaS 收入跟着动。"),
        ]),
        ("MaaS 段 = 盘 × 份额 × 单价 × 0.01", [
            ("智谱份额", None, K.PCT, "历史: 2025=收入反推(≈3.0%), 早年盘子口径缺失标 n.m.; 前瞻=『情景切换』当前案。"),
            ("实现单价(¥/M token)", None, K.N2, "2025=收入÷(盘×份额)反推 ≈¥0.50/M(价格战下 blended); 前瞻=『情景切换』当前案(三次提价=上行证据)。"),
            ("MaaS 收入(亿)", None, K.N1, "历史取年报实际; 前瞻 = 盘×份额×单价×0.01。校验: 2026E Base ≈¥16.4亿 ≈ MaaS ARR ¥17亿 的兑现口径。"),
        ]),
        ("政企段 = 客户数 × 客单价", [
            ("政企付费客户数(家)", None, K.N0, "历史=估(累计 8,000 机构客户中活跃付费政企口径); 前瞻=『情景切换』当前案, 外生轨道=央企渗透率(物理锚页)。"),
            ("客单价(¥百万)", None, K.N2, "历史=收入÷客户数反推(¥65-76万); 前瞻=『情景切换』当前案。"),
            ("政企收入(亿)", None, K.N1, "历史取年报实际(总收入−云端); 前瞻 = 客户数×客单价÷100。"),
        ]),
    ],
})
m = seg["m"]
MA_H = ha["seg_rows"]["MaaS 云端 API 收入"]
GO_H = ha["seg_rows"]["政企本地化及其他 收入"]
for col in ALLC:
    K.fml(wb[S_SEG], f"{col}{m['行业年 token 盘(万亿)']}", f"={K.R(S_ANCHOR, col + str(TOK_Y))}", K.N0, link=True)
# 历史列
for col in HC:
    K.fml(wb[S_SEG], f"{col}{m['MaaS 收入(亿)']}", f"={K.R(S_HIST, col + str(MA_H))}", K.N1, link=True)
    K.fml(wb[S_SEG], f"{col}{m['政企收入(亿)']}", f"={K.R(S_HIST, col + str(GO_H))}", K.N1, link=True)
    K.fml(wb[S_SEG], f"{col}{m['客单价(¥百万)']}", f"={K.R(S_HIST, col + str(GO_H))}*100/{K.R(S_SW, col + str(sw['SWB']['gcust'] - 1 + 1))}", K.N2, link=True)
for col, v in zip(HC, [80, 150, 350, 700]):
    pass  # 历史客户数已在情景切换 hist 列, 客单价公式引该行
# 修正: 历史客户数引情景切换 gcust 杠杆的 hist 列(三案同值) → 用 Bear 行(首行)hist
_gc_hist_row = sw["SWB"]["gcust"]
for col in HC:
    K.fml(wb[S_SEG], f"{col}{m['政企付费客户数(家)']}", f"={K.R(S_SW, col + str(_gc_hist_row))}", K.N0, link=True)
    # 客单价历史 = 收入/客户
    K.fml(wb[S_SEG], f"{col}{m['客单价(¥百万)']}", f"={K.R(S_HIST, col + str(GO_H))}*100/{col}{m['政企付费客户数(家)']}", K.N2)
# MaaS 份额/单价历史: 2022-24 n.m.; 2025 单价输入 0.50, 份额反推
for col in ["B", "C", "D"]:
    K.lab(wb[S_SEG], f"{col}{m['智谱份额']}", "n.m.", note=True)
    K.lab(wb[S_SEG], f"{col}{m['实现单价(¥/M token)']}", "n.m.", note=True)
K.inp(wb[S_SEG], f"E{m['实现单价(¥/M token)']}", 0.50, None, K.N2)
K.fml(wb[S_SEG], f"E{m['智谱份额']}", f"={K.R(S_HIST, 'E' + str(MA_H))}/(E{m['行业年 token 盘(万亿)']}*0.01*E{m['实现单价(¥/M token)']})", K.PCT)
# 前瞻列
for col in FCf:
    K.fml(wb[S_SEG], f"{col}{m['智谱份额']}", f"={K.R(S_SW, col + str(sw['SWACT']['mshare']))}", K.PCT, link=True)
    K.fml(wb[S_SEG], f"{col}{m['实现单价(¥/M token)']}", f"={K.R(S_SW, col + str(sw['SWACT']['mprice']))}", K.N2, link=True)
    K.fml(wb[S_SEG], f"{col}{m['MaaS 收入(亿)']}", f"={col}{m['行业年 token 盘(万亿)']}*{col}{m['智谱份额']}*{col}{m['实现单价(¥/M token)']}*0.01", K.N1)
    K.fml(wb[S_SEG], f"{col}{m['政企付费客户数(家)']}", f"={K.R(S_SW, col + str(sw['SWACT']['gcust']))}", K.N0, link=True)
    K.fml(wb[S_SEG], f"{col}{m['客单价(¥百万)']}", f"={K.R(S_SW, col + str(sw['SWACT']['gticket']))}", K.N2, link=True)
    K.fml(wb[S_SEG], f"{col}{m['政企收入(亿)']}", f"={col}{m['政企付费客户数(家)']}*{col}{m['客单价(¥百万)']}/100", K.N1)
    wb[S_SEG][f"{col}{m['MaaS 收入(亿)']}"].fill = K.OUT

# ════════════ 10. 利润与收入假设 ════════════
fr = K.write_fundamentals(wb.create_sheet(S_FUND), {
    "title": "利润与收入假设 — 分部营收→综合经营利润→净利→权益→每股(粗颗粒, 不做三表勾稽)",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "ha": ha, "share_fix_col": "G", "ps_scale": 100,
    "eps_label": "EPS (RMB)", "bps_label": "BPS (RMB)",
    "assum_groups": [
        ("利润率与转换(历史实际锚 + 前瞻)", [
            {"name": "综合经营利润率(含研发)", "vals": [-1.69, -4.99, -7.89, -6.52, None, None, None, None, None],
             "fmt": K.PCT, "logic": "历史=经调整亏损÷收入(2025 为报表口径); 前瞻=『情景切换』当前案——研发近固定成本, 收入放大即摊薄, Base 2029 打平。",
             "link": {"sheet": S_SW, "row": sw["SWACT"]["opm"]}},
            {"name": "净利转换率", "vals": [None, None, None, None, 1.00, 1.00, 1.00, 1.00, 0.90],
             "fmt": K.PCT, "logic": "亏损期 = 1.0(无税); 2030 盈利后按 10% 有效税负折(高新企业+研发加计, 实际税负低)。", "nm_cols": ["B", "C", "D", "E"]},
            {"name": "留存率", "vals": [None, None, None, None, 1.00, 1.00, 1.00, 1.00, 1.00],
             "fmt": K.PCT, "logic": "不分红(成长期全留存)。", "nm_cols": ["B", "C", "D", "E"]},
            {"name": "股权融资流入(亿)", "vals": [None, None, None, None, 38.0, 120.0, 0, 0, 0],
             "fmt": K.N1, "logic": "2026=港股 IPO 净额 ¥38亿(HK$41.7亿, 2026-01 已落账); 2027=A股科创板按 Base 募 ¥120亿(公告上限 150亿×8 折)。Bear 案下若 A股失败为最大生存风险(仪表盘跟踪), 此处三案共用 Base——保守化会进一步压低 Bear 净资产。", "nm_cols": ["B", "C", "D", "E"]},
        ]),
    ],
    "segments": [
        {"name": "MaaS 云端 API 收入", "hist_row": "MaaS 云端 API 收入", "fwd": {"sheet": S_SEG, "row": m["MaaS 收入(亿)"]}},
        {"name": "政企本地化及其他 收入", "hist_row": "政企本地化及其他 收入", "fwd": {"sheet": S_SEG, "row": m["政企收入(亿)"]}},
    ],
    "profit_terms": [
        (["MaaS 云端 API 收入", "政企本地化及其他 收入"], "综合经营利润率(含研发)", False),
    ],
    "conv_assum": "净利转换率", "retention_assum": "留存率",
    "note_text": "链: 分部收入(物理锚驱动)→ ×综合经营利润率(情景刀)→ 净利 → 权益(上年+净利+股权融资)→ EPS/BPS。每股营收 SPS 行在本表底部, 是 EV/Sales 主线的分母。单位: 亿 RMB; 每股 RMB。前瞻股本固定 435.2mn(A股增发 2-8% 的稀释在 memo 中单列敏感性, 不入主链)。",
})
# 权益前瞻公式覆写: 加股权融资流入
ws_f = wb[S_FUND]
_fin = fr["am"]["股权融资流入(亿)"]
_prevs = [HC[-1]] + list(FCf[:-1])
for p, c in zip(_prevs, FCf):
    K.fml(ws_f, f"{c}{fr['EQ']}", f"={p}{fr['EQ']}+{c}{fr['NI']}*{c}{fr['am']['留存率']}+{c}{_fin}", K.N1)
# SPS 行(EV/Sales 主线分母)
SPS = ws_f.max_row + 1
K.lab(ws_f, f"A{SPS}", "SPS 每股营收(RMB)", b=True); ws_f[f"A{SPS}"].fill = K.OUT
for c in HC:
    K.fml(ws_f, f"{c}{SPS}", f"={c}{fr['REV']}*100/{K.R(S_HIST, c + str(ha['HSH']))}", K.N2)
for c in FCf:
    K.fml(ws_f, f"{c}{SPS}", f"={c}{fr['REV']}*100/{K.R(S_HIST, '$G$' + str(ha['HSH']))}", K.N2)
K.logic(ws_f, f"N{SPS}", "= 总营收×100 ÷ 股本(mn)。历史用上市前股本 397.8mn; 前瞻用当下 435.2mn。『情景估值』主线 = 目标 P/S × 本行 × FX。")
fr_ps = dict(fr); fr_ps["BPS"] = SPS   # 主线镜头分母 = SPS

# ════════════ 11. 情景估值 ════════════
sv = K.write_scenario_valuation(wb.create_sheet(S_VAL), {
    "title": "情景估值 — 当前案逐年隐含价 (EV/Sales 主线; P/E 体检 + 远期盈利贴现)",
    "intro": "本表输出=『情景切换』当前案(默认 Base)。隐含价 = 目标 P/S(当前案) × 前瞻每股营收 SPS × FX(HKD)。历史列=一级估值隐含价反推(事实); 前瞻是预测、不拟合现价。三案并排见『估值对比』。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf, "hist_years": HY,
    "fx_fwd": FX, "s_hist": S_HIST, "ha": ha, "share_fix_col": "G",
    "s_fund": S_FUND, "fr": fr_ps,
    "s_switch": S_SW, "target_row": SWPS, "sw_cell": "B2",
    "yend": {}, "yavg": {},
    "mcap_div": 100, "mcap_usd_skip": True,
    "mcap_hist_label": "市值 实际(亿HKD, 历史=一级估值)", "mcap_fwd_label": "市值 前瞻·主线(亿HKD)",
    "reading": "P/E 体检读法: 2026-2028 净利为负 → 隐含 P/E 无意义(负值即 N/M); Base 2029 转正后, 2030 隐含 P/E ≈137x——偏高, 反映 P/S 镜头把 2031+ 利润率扩张当期权计价。第三检验(下方'远期盈利贴现')只付已见盈利路径, 给出更保守下界; 两镜头的差 = 为'尚未证实的利润率扩张'付的价。",
    "method": "方法: 整体公司、EV/Sales 主线逐年估(公司净现金, EV≈市值)。基本面在『利润与收入假设』; 目标 P/S 在『估值倍数假设』(三层); 本表套用: 目标 P/S × SPS × FX → 隐含价(HKD)。",
    "concl": "结论(方向性): Base 2027E 隐含价 HK$345, 较现价 -68%; 即便 Bull 2027E 也只有 HK$666(-39%)。现价唯有'2030E 收入 ¥560亿+'的路径才能 justify——超出 Bull。三情景与概率加权见『估值对比』与决策 memo。",
})
# 远期盈利贴现 交叉镜头(第三检验): 2030E NI × 稳态 P/E ÷ 1.15^4
ws_v = wb[S_VAL]
_r = ws_v.max_row + 2
K.band(ws_v, _r, "远期盈利贴现(第三检验, DCF 简式): 2030E 净利 × 稳态 P/E ÷ (1+15%)^4 → 2026 现值", 11); _r += 1
NI30 = K.R(S_FUND, "J" + str(fr["NI"]))
K.lab(ws_v, f"A{_r}", "稳态 P/E(2030, 当前案语境)")
K.inp(ws_v, f"B{_r}", 45.0, None, K.MX)
K.logic(ws_v, f"D{_r}", "45x = 中国 AI/高成长软件龙头稳态盈利的市场定价区间中枢(Bull 语境可看 55x, Bear 不适用——2030 仍亏损时该镜头退化为残值: 现金+牌照 ≈¥120亿 → 约 HK$30/股)。")
_PE_ST = f"B{_r}"; _r += 1
K.lab(ws_v, f"A{_r}", "隐含股价 远期盈利贴现(HKD)", b=True); ws_v[f"A{_r}"].fill = K.OUT
K.fml(ws_v, f"B{_r}", f"=IF({NI30}<=0,\"N/M(2030 仍亏损)\",{NI30}*{_PE_ST}*100/{K.R(S_HIST, 'G' + str(ha['HSH']))}*{FX}/1.15^4)", K.PX, link=True)
K.logic(ws_v, f"D{_r}", "= 2030E NI(亿) × 稳态P/E ÷ 股本 × FX ÷ 1.15^4。Base ≈HK$172 —— 低于 EV/Sales 主线的 HK$345: 两镜头的差就是'为利润率扩张期权付的价'。贴现率 15% = 高 beta 成长股权益成本。")
EARN_DISC_ROW = _r

# ════════════ 12. 估值对比(三案恒常并排)════════════
SWB = sw["SWB"]
SH_G = K.R(S_HIST, f"$G${ha['HSH']}")
PX_NOW_REF = K.R(S_HIST, f"G{ha['HPX']}")
_conv = fr["am"]["净利转换率"]
_ret = fr["am"]["留存率"]

def _fp(j, A, k):
    return (HC[-1] if j == 0 else FCf[j - 1]) + str(A[k])

cmp_rows = [
    {"key": "tok", "label": "行业年 token 盘(万亿)", "fmt": K.N0,
     "hist": lambda c, ci, A: f"={K.R(S_ANCHOR, c + str(TOK_Y))}" if c in ("D", "E") else None,
     "fwd": lambda c, j, ci, A: f"={K.R(S_ANCHOR, c + str(TOK_Y))}"},
    {"key": "maas", "label": "MaaS 收入(亿)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(MA_H))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['tok']}*{K.R(S_SW, c + str(SWB['mshare'] + ci))}*{K.R(S_SW, c + str(SWB['mprice'] + ci))}*0.01"},
    {"key": "gov", "label": "政企收入(亿)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(GO_H))}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_SW, c + str(SWB['gcust'] + ci))}*{K.R(S_SW, c + str(SWB['gticket'] + ci))}/100"},
    {"key": "rev", "label": "总收入(亿)", "fmt": K.N1, "bold": True,
     "hist": lambda c, ci, A: f"={c}{A['maas']}+{c}{A['gov']}",
     "fwd": lambda c, j, ci, A: f"={c}{A['maas']}+{c}{A['gov']}"},
    {"key": "ni", "label": "净利(亿)", "fmt": K.N1, "bold": True,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HNI']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['rev']}*{K.R(S_SW, c + str(SWB['opm'] + ci))}*{K.R(S_FUND, c + str(_conv))}"},
    {"key": "sps", "label": "SPS 每股营收(RMB)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={c}{A['rev']}*100/{K.R(S_HIST, c + str(ha['HSH']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['rev']}*100/{SH_G}"},
    {"key": "sent", "label": "情绪值(该案; 历史=实际反推)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_MULT, c + str(ma['sent_row0'] + ci))}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_MULT, c + str(ma['sent_row0'] + ci))}"},
    {"key": "ps", "label": "目标 P/S(历史=实际)", "fmt": K.MX,
     "hist": lambda c, ci, A: f"={_pk}*{_pr}*{c}{A['sent']}",
     "fwd": lambda c, j, ci, A: f"={_pk}*{_pr}*{c}{A['sent']}"},
    {"key": "px", "label": "隐含价 (HKD)", "fmt": K.PX, "bold": True, "out": True,
     "hist": lambda c, ci, A: f"={c}{A['ps']}*{c}{A['sps']}*{K.R(S_HIST, c + str(ha['HFX']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['ps']}*{c}{A['sps']}*{FX}"},
    {"key": "ipe", "label": "隐含 forward P/E(体检)", "fmt": K.MX,
     "hist": lambda c, ci, A: f'=IF({c}{A["ni"]}<=0,"N/M",{c}{A["px"]}/({c}{A["ni"]}*100/{K.R(S_HIST, c + str(ha["HSH"]))}*{K.R(S_HIST, c + str(ha["HFX"]))}))',
     "fwd": lambda c, j, ci, A: f'=IF({c}{A["ni"]}<=0,"N/M",{c}{A["px"]}/({c}{A["ni"]}*100/{SH_G}*{FX}))'},
    {"key": "up", "label": "历史: vs 一级估值隐含价(回测≈0) / 前瞻: vs 现价", "fmt": K.PCT,
     "hist": lambda c, ci, A: f"={c}{A['px']}/{K.R(S_HIST, c + str(ha['HPX']))}-1",
     "fwd": lambda c, j, ci, A: f"={c}{A['px']}/{PX_NOW_REF}-1"},
]
cm = K.write_comparison(wb.create_sheet(S_CMP), {
    "title": "估值对比 — Bear / Base / Bull 三情景目标价并排",
    "intro": "三案各自完整推演: token 盘 → 份额×单价 + 客户×客单 → 总收入 → 净利 → 目标 P/S(锚×溢价×该案情绪)→ 逐年隐含价(HKD)。恒常并排不随开关变; 调假设去『情景切换』。历史列用同一条链填实际值: 隐含价历史列 ≈ 一级估值隐含价(回测行 ≈0%)= 链的内置回测。",
    "case_names": CASES,
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "block_start": 16,
    "rows": cmp_rows,
    "summary": {
        "band": "三案汇总 (目标年 2027E; 触发条件见『情景切换』)",
        "target_col": "G",
        "rows": [
            ("总收入(亿)", "rev", K.N1, "= 盘×份额×单价 + 客户×客单(该案)", False),
            ("净利(亿)", "ni", K.N1, "= 收入 × 该案经营利润率(2027 三案均亏损)", False),
            ("目标 P/S", "ps", K.MX, "= 基准锚 18x × 结构溢价 1.5x × 该案情绪值", False),
            ("隐含价(HKD)", "px", K.PX, "= 目标 P/S × 2027E 每股营收 × FX", True),
            ("隐含 forward P/E", "ipe", K.MX, "体检: 2027 三案均亏损 → N/M; 盈利后年份见各案 block", False),
            ("vs 现价", "up", K.PCT, "对照现价 HK$1,087 的上行/下行空间", True),
        ],
        "mcap": {"label": "隐含市值(亿HKD)", "key": "px", "expr": f"*{SH_G}/100",
                 "note": "= 隐含价 × 股本 435.2mn ÷ 100"},
        "concl": "risk-reward(2027E vs 现价 HK$1,087): Bear -89% / Base -68% / Bull -39%——三案全部低于现价, 概率加权(25/50/25)目标价 ≈HK$370(-66%)。现价唯一的支撑是'稀缺筹码+叙事', 而 7/8 解禁与 A股定价正是拆筹码结构的两个日历事件。结论: 卖出。",
    },
})

# ════════════ 13. 综合判断仪表盘 ════════════
REV26 = K.R(S_FUND, "F" + str(fr["REV"]))
dash = K.write_dashboard(wb.create_sheet(S_DASH), {
    "title": "综合判断仪表盘 — A 基本面拐点 · B 估值错位 · C 催化剂 · D 情绪确认",
    "usage": "怎么用: 预测引擎是 B(错位)+C(催化剂); 情绪 D 只做 timing 确认。本标的的特殊性: 基本面在变好(A 强)而估值错位为深度负值(B 极差)——'好公司 ≠ 好股票'的典型分裂, 决策由 B+C 主导。",
    "blocks": [
        {"title": "A. 基本面拐点 — 业务在结构性变好吗?", "rows": [
            ("收入结构迁移", "云端占比 15.2%→26.3%; MaaS ARR ¥17亿(+60x)", "从'AI 外包'向'AI 平台'迁移是真的, 这是 A 股 150 亿故事的基本面内核。"),
            ("定价权", "2026 年三次提价且用量续涨", "DeepSeek 价格战背景下逆势提价 = coding/agent 能力代差的市场化证据。"),
            ("生存约束", "现金 ¥22.59亿 + IPO ¥38亿 vs 年烧 ¥30-40亿", "跑道 ~18 个月; A股 ¥150亿 是续命刚需而非锦上添花——基本面改善有一个融资前提。"),
            ("A 判断", "【强, 但带融资前提】", "收入质量与定价权确实在拐点上; 风险不在业务在资金。", True),
        ]},
        {"title": "B. 估值错位 — 市场给的 vs 基本面该给的", "rows": [
            ("市场现在给(2026E P/S)", {"fml": f"={PX_NOW_REF}*{SH_G}/100/{FX}/{REV26}", "fmt": K.MX, "fill": True},
             "= 现价市值(亿RMB) ÷ 模型 2026E 收入(Base)。随模型与开关变。"),
            ("基本面该给(锚×溢价)", {"fml": f"={_pk}*{_pr}", "fmt": K.MX},
             "= 全球前沿一级 forward P/S 18x × 稀缺溢价 1.5x = 27x——已经是'按世界最贵同类+溢价'给的。"),
            ("错位 GAP = 该给÷市场给 − 1",
             {"fml": lambda ro: f"=B{ro['基本面该给(锚×溢价)']}/B{ro['市场现在给(2026E P/S)']}-1", "fmt": K.PCT},
             "≈ -82%: 市场付了基本面该给的 5.5 倍。GAP 深负 = 基本面空间早被吃完, 纯情绪/筹码定价区。"),
            ("现价反推", "需 2030E 收入 ¥560-1,120亿", "持有现价(要求 15% 年化)需要 2030E 收入达 2025 的 77-155 倍——超出 Bull 案(¥463亿)。"),
        ]},
        {"title": "C. 催化剂 — 什么会逼市场闭合 GAP(本案为下行催化)", "rows": [
            ("2026-07-08 解禁", "26 天后", "早期 VC 成本对应估值 ¥80-244亿(账面 15-40 倍), 抛压方向确定; 6 月已抢跑(-33%)。"),
            ("A股发行定价", "2026H2 申报/过会", "公告隐含估值锚 ≥¥1,875亿 vs 港股 ¥4,340亿——若按科创板惯例折价发行, 直接向下重锚 H 股。"),
            ("2026 中报(8月)", "ARR 兑现 + 云端占比", "唯一可能的正向催化: ARR→确认收入超预期会延长叙事; 不及预期则三杀。"),
            ("C 判断", "下行催化密集且有日历", "未来 90 天三个硬事件, 两个方向向下。", True),
        ]},
        {"title": "D. 情绪确认 — timing + 刹车", "rows": [
            ("量价温度计", "5/29 盘中 HK$1,993 → 6/12 HK$1,074, 两周 -33%, 放量下跌", "高点放量滞涨 + 急跌 = 退潮初段而非回调结束; 只当温度计不进倍数。"),
            ("现价倍数 vs 该给", "148x vs 27x(2026E)", "市场付价 >> 基本面该给, 深度情绪定价区。"),
            ("当前档位", "【过热 → 退潮初段】", "对照 2026-02 GLM-5 启动段(情绪萌芽)——现在是镜像位置。", True),
            ("衰减扳机", "4 条", "①解禁后大宗/减持公告 ②A股定价折价>40% ③DeepSeek V4 coding 追平 ④中报 ARR 增速 <50%。任一翻 → 情绪值下调、Bear 概率上调。"),
        ]},
    ],
    "final": {"band": "★ 综合判断",
              "text": "A 强(业务真在变好)× B 深度负错位(付价=该给的 5.5 倍)× C 下行催化密集(解禁+A股定价 90 天内)× D 退潮初段 → 卖出/回避。概率加权目标价 HK$370(-66%)。正向证伪路径只有一条: 中报 ARR 超预期 + A股以接近 H 股估值发行(=市场用真金白银确认稀缺溢价), 届时回模型上调情绪值与 Bull 概率。"},
    "tracking": {
        "intro": "哪个指标恶化 → 哪个假设先崩 → 触发什么动作。",
        "rows": [
            ("__band__", "一、定价权与能力代差(MaaS 链关键敏感项)"),
            ("竞品 coding 基准", "GLM-5 国内领先", "关键敏感项: 单价假设(三次提价的根基)", "DeepSeek/Qwen 新版发布即测(SWE-bench 等)", "追平 → mprice 翻 Bear 路径"),
            ("MaaS ARR 增速", "¥17亿(+60x)", "关键敏感项: 份额×单价的总验证", "中报/年报披露", "<50% YoY → 收入链下修"),
            ("__band__", "二、筹码与资金事件"),
            ("解禁减持公告", "7/8 起", "关键敏感项: 情绪值路径", "联交所权益披露/大宗成交", "密集减持 → 情绪值下调 0.1-0.2"),
            ("A股发行隐含估值", "待申报", "关键敏感项: 估值重锚方向", "上交所审核公告", "<¥2,500亿 → 翻 Bear; 接近 H 股 → 上调 Bull 概率"),
            ("__band__", "三、生存线"),
            ("现金余额/烧钱", "¥22.59亿+38亿, 年烧 30-40亿", "关键敏感项: 持续经营(Bear 案 2027 见底)", "中报现金流量表", "A股 2027H1 前未落地 → 生存折价入模"),
            ("云端收入占比", "26.3%", "关键敏感项: '平台化'叙事的可验证指标", "中报分部", "停滞 <28% → 政企外包估值回归(P/S 3-12x)"),
        ],
    },
})

# ════════════ 全局格式 + 落盘 ════════════
K.finalize(wb, freeze={
    S_HIST: "B3", S_PX: "B4", S_CONS: "A2", S_HMULT: "B5", S_MULT: "B4", S_SW: "B3",
    S_ANCHOR: "B3", S_SEG: "B3", S_FUND: "B3", S_VAL: "B4", S_CMP: "B6", S_DASH: "B6",
    S_COVER: "A2",
})
os.makedirs(OUTPUT_DIR, exist_ok=True)
out = os.path.join(OUTPUT_DIR, "2513.HK_valuation_model.xlsx")
wb.save(out)
print("saved:", out)
print("sheets:", wb.sheetnames)
print("rows:", {"HPS": HPS, "SPS": SPS, "SWPS": SWPS, "EARN_DISC": EARN_DISC_ROW})
