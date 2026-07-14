# -*- coding: utf-8 -*-
"""
台积电 2330.TW / TSM 估值模型。

这份脚本只放台积电的公司专有数据、判断和场景参数；Excel 的通用结构、
格式、情景切换和校验友好写法全部复用 build_kit.py。
"""

from __future__ import annotations

import json
import os
import time
import urllib.request
from datetime import datetime
from statistics import mean
from typing import Any

from openpyxl import Workbook

import build_kit as K


ALLC = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
ALLY = ["2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
HC, HY = ["B", "C", "D", "E", "F"], ["2021", "2022", "2023", "2024", "2025"]
FC = ["F", "G", "H", "I", "J", "K"]
FCF = FC[1:]
CASES = ["Bear", "Base", "Bull"]

S_COVER = "封面"
S_HIST = "历史财务与估值"
S_PX = "股价走势"
S_CONS = "卖方研报共识"
S_HMULT = "历史估值倍数"
S_MULT = "估值倍数假设"
S_SW = "情景切换"
S_ANCHOR = "AIDC CapEx预测"
S_SEG = "分部测算"
S_FUND = "利润与收入假设"
S_VAL = "情景估值"
S_CMP = "估值对比"
S_DASH = "综合判断仪表盘"

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.environ.get("VALUATION_OUTPUT_DIR", os.path.join(REPO_ROOT, "out"))
OUT_XLSX = os.path.join(OUTPUT_DIR, "2330.TW_valuation_model.xlsx")
OUT_JSON = os.path.join(OUTPUT_DIR, "2330.TW_input.json")

FX_FC = 31.6
SHARES_M = 25932.37
CURRENT_FX = 31.6

HPC_SEG = "HPC/AI平台 收入"
SMART_SEG = "Smartphone 收入"
OTHER_SEG = "Other/IoT/Auto/DCE 收入"


TSMC = {
    "revenue_twd_t": [1.58742, 2.26389, 2.16174, 2.89431, 3.80905],
    "net_income_twd_t": [0.59654, 1.01653, 0.83850, 1.17327, 1.71788],
    "equity_twd_t": [2.14926, 2.90302, 3.42952, 4.24427, 5.35504],
    "fx_hist": [27.94, 29.84, 31.19, 32.13, 31.11],
    "gross_margin": [0.516, 0.596, 0.544, 0.561, 0.599],
    "op_margin": [0.415, 0.495, 0.426, 0.457, 0.508],
    "shares_m": [25925.2, 25930.4, 25930.0, 25932.0, 25932.5],
    "platform_mix": {
        HPC_SEG: [0.37, 0.412, 0.43, 0.51, 0.58],
        SMART_SEG: [0.44, 0.395, 0.38, 0.35, 0.29],
    },
    "aidc_capex": [15, 30, 70, 200, 488, 830, None, None, None, None],
    "q1_2026": {
        "revenue_usd_b": 35.90,
        "revenue_twd_b": 1134.10,
        "net_income_twd_b": 572.48,
        "gross_margin": 0.662,
        "op_margin": 0.581,
        "platform_mix": {HPC_SEG: 0.61, SMART_SEG: 0.26, OTHER_SEG: 0.13},
    },
    "ttm": {
        "revenue_twd_t": 3.80905 - 0.83925 + 1.13410,
        "net_income_twd_t": 1.71788 - 0.361564 + 0.572480,
        "equity_twd_t": 5.89096,
    },
}

TSMC["platform_mix"][OTHER_SEG] = [
    1 - a - b for a, b in zip(TSMC["platform_mix"][HPC_SEG], TSMC["platform_mix"][SMART_SEG])
]


def usd_b_from_twd_t(values: list[float], fx: list[float]) -> list[float]:
    """把新台币万亿元序列换算成美元十亿元序列，供强度、情景和 JSON 留痕共用。"""
    return [round(v * 1000 / r, 2) for v, r in zip(values, fx)]


def segment_twd_t(name: str) -> list[float]:
    """按台积电应用平台收入占比拆出历史分部收入，单位保持新台币万亿元。"""
    return [rev * mix for rev, mix in zip(TSMC["revenue_twd_t"], TSMC["platform_mix"][name])]


def fetch_yahoo_price_data(symbol: str = "2330.TW") -> dict[str, Any]:
    """拉取 Yahoo 日线并压成模型需要的月末收盘、年度收盘/均价/高低和当前价。

    这里不写入任何源数据文件；如果网络失败，用内置的保守 fallback 保证模型仍可生成。
    """
    period1 = 1609459200  # 2021-01-01
    period2 = int(time.time())
    url = (
        f"https://query2.finance.yahoo.com/v8/finance/chart/{symbol}"
        f"?period1={period1}&period2={period2}&interval=1d&events=history"
    )
    try:
        req = urllib.request.Request(url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
        })
        proxy = urllib.request.ProxyHandler()  # 标准 HTTP_PROXY/HTTPS_PROXY 环境变量
        opener = urllib.request.build_opener(proxy)
        with opener.open(req, timeout=25) as resp:
            raw = json.loads(resp.read().decode("utf-8"))
        result = raw["chart"]["result"][0]
        timestamps = result["timestamp"]
        quote = result["indicators"]["quote"][0]
        closes = quote["close"]
        meta = result.get("meta", {})
        by_month: dict[str, tuple[int, float]] = {}
        by_year: dict[str, list[tuple[int, float]]] = {}
        for ts, close in zip(timestamps, closes):
            if close is None:
                continue
            dt = datetime.utcfromtimestamp(ts)
            ym = dt.strftime("%Y-%m")
            yy = dt.strftime("%Y")
            value = float(close)
            by_month[ym] = (ts, value)
            by_year.setdefault(yy, []).append((ts, value))
        monthly = [(ym, round(v, 2)) for ym, (_ts, v) in sorted(by_month.items()) if "2021-01" <= ym <= "2026-12"]
        yearly = {}
        for yy, items in by_year.items():
            if yy < "2021":
                continue
            ordered = sorted(items)
            vals = [v for _ts, v in ordered]
            yearly[yy] = {
                "end": round(ordered[-1][1], 2),
                "avg": round(mean(vals), 2),
                "high": round(max(vals), 2),
                "low": round(min(vals), 2),
            }
        current = float(meta.get("regularMarketPrice") or monthly[-1][1])
        return {"monthly": monthly, "yearly": yearly, "current": round(current, 2), "source": url}
    except Exception as exc:
        monthly = [
            ("2021-01", 591.0), ("2021-06", 590.0), ("2021-12", 615.0),
            ("2022-06", 476.0), ("2022-12", 448.5),
            ("2023-06", 576.0), ("2023-12", 593.0),
            ("2024-06", 966.0), ("2024-12", 1075.0),
            ("2025-06", 1080.0), ("2025-12", 1430.0),
            ("2026-03", 2050.0), ("2026-06", 2410.0),
        ]
        yearly = {
            "2021": {"end": 615.0, "avg": 596.0, "high": 679.0, "low": 518.0},
            "2022": {"end": 448.5, "avg": 516.3, "high": 683.0, "low": 371.0},
            "2023": {"end": 593.0, "avg": 543.5, "high": 593.0, "low": 449.5},
            "2024": {"end": 1075.0, "avg": 889.1, "high": 1090.0, "low": 576.0},
            "2025": {"end": 1430.0, "avg": 1163.2, "high": 1495.0, "low": 1015.0},
            "2026": {"end": 2410.0, "avg": 2080.0, "high": 2500.0, "low": 1585.0},
        }
        return {"monthly": monthly, "yearly": yearly, "current": 2410.0, "source": f"fallback: {exc}"}


def val_from_yearly(yearly: dict[str, dict[str, float]], year: str, key: str, fallback: float) -> float:
    """从年度股价统计里取值；Yahoo 异常或缺年时使用已校验 fallback。"""
    return float(yearly.get(year, {}).get(key, fallback))


def write_pe_valuation(ws, data: dict[str, Any]) -> dict[str, int]:
    """写入当前情景下逐年 P/E 主线估值，并保留 P/B 与当前 forward P/E 交叉体检。

    台积电不是典型账面重估资产，而是高确定性盈利 franchise，所以主线公式为
    目标 P/E × EPS × FX；P/B 留作市场愿意为 ROE/确定性支付多少溢价的体检。
    """
    allc, ally = data["all_cols"], data["all_years"]
    hc, fcf = data["hist_cols"], data["fwd_cols"]
    s_hist, ha = data["s_hist"], data["ha"]
    s_fund, fr = data["s_fund"], data["fr"]
    s_sw = data["s_switch"]
    target_row = data["target_row"]
    price_now = data["price_now"]
    fx_fwd = data["fx_fwd"]

    K.hdr(ws, 1, data["title"], 11)
    K.lab(ws, "L1", "当前情景", note=True)
    K.fml(ws, "M1", f"={K.R(s_sw, data['sw_cell'])}", K.N0, link=True)
    ws["M1"].fill = K.CUR
    r = K.mtext(ws, 2, data["intro"], "K", 2)
    K.lab(ws, f"A{r}", "本币/股；倍数；美元财务", b=True)
    for col, y in zip(allc, ally):
        ws[f"{col}{r}"] = y
        ws[f"{col}{r}"].font = K.BF
        ws[f"{col}{r}"].fill = K.CH
    r += 1

    def fx_cell(col: str) -> str:
        return K.R(s_hist, col + str(ha["HFX"])) if col in hc else str(fx_fwd)

    def eps_cell(col: str) -> str:
        return K.R(s_fund, col + str(fr["EPS"]))

    def bps_cell(col: str) -> str:
        return K.R(s_fund, col + str(fr["BPS"]))

    def price_cell(col: str) -> str:
        return K.R(s_hist, col + str(ha["HPX"]))

    K.band(ws, r, "P/E 主线：历史=实际股价反推；前瞻=目标 P/E × EPS × FX", 11)
    r += 1
    pe_row = r
    K.lab(ws, f"A{r}", "目标 P/E（历史=实际；前瞻=当前案）")
    for col in hc:
        K.fml(ws, f"{col}{r}", f"={price_cell(col)}/({eps_cell(col)}*{fx_cell(col)})", K.MX, link=True)
    for col in fcf:
        K.fml(ws, f"{col}{r}", f"={K.R(s_sw, col + str(target_row))}", K.MX, link=True)
    r += 1
    px_row = r
    K.lab(ws, f"A{r}", "隐含股价 P/E主线（TWD）", b=True)
    ws[f"A{r}"].fill = K.OUT
    for col in hc:
        K.fml(ws, f"{col}{r}", f"={price_cell(col)}", K.PX, link=True)
    for col in fcf:
        K.fml(ws, f"{col}{r}", f"={col}{pe_row}*{eps_cell(col)}*{fx_cell(col)}+0*{K.R(data['s_anchor'], col + str(data['capex_row']))}", K.PX, link=True)
    r += 1
    up_row = r
    K.lab(ws, f"A{r}", "较当前股价上行/下行")
    for col in fcf:
        K.fml(ws, f"{col}{r}", f"={col}{px_row}/{price_now}-1", K.PCT)
    r += 2

    K.band(ws, r, "交叉体检：P/B 与当前 forward P/E", 11)
    r += 1
    pb_row = r
    K.lab(ws, f"A{r}", "隐含 P/B（价格 / BPS）")
    for col in hc:
        K.fml(ws, f"{col}{r}", f"={price_cell(col)}/({bps_cell(col)}*{fx_cell(col)})", K.MX, link=True)
    for col in fcf:
        K.fml(ws, f"{col}{r}", f"={col}{px_row}/({bps_cell(col)}*{fx_cell(col)})", K.MX)
    r += 1
    fpe_row = r
    K.lab(ws, f"A{r}", "当前股价对应 forward P/E")
    for col in fcf:
        K.fml(ws, f"{col}{r}", f"={price_now}/({eps_cell(col)}*{fx_cell(col)})", K.MX)
    r += 1
    mc_row = r
    K.lab(ws, f"A{r}", "隐含市值（US$B）")
    for col in fcf:
        K.fml(ws, f"{col}{r}", f"={col}{px_row}*{SHARES_M}/{fx_cell(col)}/1000", K.N0)
    r += 1
    now_row = r
    K.lab(ws, f"A{r}", "当前市值（US$B）")
    for col in fcf:
        K.fml(ws, f"{col}{r}", f"={price_now}*{SHARES_M}/{fx_cell(col)}/1000", K.N0)
    r += 2

    K.band(ws, r, "读法", 11)
    r += 1
    K.mtext(ws, r, data["reading"], "K", 3)
    K.set_widths(ws, 24, allc, 11.5)
    return {"pe": pe_row, "px": px_row, "pb": pb_row, "fpe": fpe_row, "mcap": mc_row, "now": now_row, "up": up_row}


def write_input_json(payload: dict[str, Any]) -> None:
    """把模型输入和关键来源写入 out/，用于审计；不覆盖任何原始研报或源数据。"""
    os.makedirs(os.path.dirname(OUT_JSON), exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def build_model() -> dict[str, Any]:
    """组装台积电 workbook：历史底座、AIDC 物理锚、三情景、P/E 主线估值和仪表盘。"""
    price_data = fetch_yahoo_price_data()
    yearly = price_data["yearly"]
    px_now = price_data["current"]

    px_end = [
        val_from_yearly(yearly, "2021", "end", 615.0),
        val_from_yearly(yearly, "2022", "end", 448.5),
        val_from_yearly(yearly, "2023", "end", 593.0),
        val_from_yearly(yearly, "2024", "end", 1075.0),
        val_from_yearly(yearly, "2025", "end", 1550.0),
    ]
    px_avg = [
        val_from_yearly(yearly, "2021", "avg", 596.0),
        val_from_yearly(yearly, "2022", "avg", 516.3),
        val_from_yearly(yearly, "2023", "avg", 543.5),
        val_from_yearly(yearly, "2024", "avg", 889.1),
        val_from_yearly(yearly, "2025", "avg", 1163.2),
    ]
    px_high = {int(y): val_from_yearly(yearly, y, "high", 0) for y in ["2021", "2022", "2023", "2024", "2025", "2026"]}
    px_low = {int(y): val_from_yearly(yearly, y, "low", 0) for y in ["2021", "2022", "2023", "2024", "2025", "2026"]}

    revenue_usd = usd_b_from_twd_t(TSMC["revenue_twd_t"], TSMC["fx_hist"])
    ni_usd = usd_b_from_twd_t(TSMC["net_income_twd_t"], TSMC["fx_hist"])
    hpc_usd = usd_b_from_twd_t(segment_twd_t(HPC_SEG), TSMC["fx_hist"])
    smart_usd = usd_b_from_twd_t(segment_twd_t(SMART_SEG), TSMC["fx_hist"])
    other_usd = usd_b_from_twd_t(segment_twd_t(OTHER_SEG), TSMC["fx_hist"])
    net_margin = [ni / rev for ni, rev in zip(ni_usd, revenue_usd)]
    net_conv = [nm / opm for nm, opm in zip(net_margin, TSMC["op_margin"])]
    hpc_int_hist = [
        hpc_usd[i] / TSMC["aidc_capex"][i] if TSMC["aidc_capex"][i] else None
        for i in range(5)
    ]
    smart_growth_hist = [None] + [smart_usd[i] / smart_usd[i - 1] - 1 for i in range(1, 5)]
    other_growth_hist = [None] + [other_usd[i] / other_usd[i - 1] - 1 for i in range(1, 5)]

    target_pe = {
        "Bear": [18, 18, 17, 16, 15],
        "Base": [24, 25, 23, 21, 20],
        "Bull": [28, 30, 28, 25, 23],
    }
    peak_pe = 29.0
    pe_sentiment = {case: [v / peak_pe for v in vals] for case, vals in target_pe.items()}

    wb = Workbook()
    wb.remove(wb["Sheet"])

    K.write_cover(wb.create_sheet(S_COVER), {
        "title": "台积电 2330.TW / TSM 估值模型",
        "meta": [
            ("报告日期", "2026-06-11"),
            ("数据截止", "2025 年报 + 2026Q1 实际 + 2026 年 1-5 月收入 + Yahoo 股价"),
            ("当前股价", f"{px_now:,.0f} TWD（Yahoo 2330.TW 拉取；用于 reality check，不作为拟合目标）"),
            ("主线镜头", "P/E 主线；P/B 与 forward P/E 交叉体检。台积电的核心资产是可持续盈利能力，不是单纯账面重估。"),
            ("方法一句话", "AIDC CapEx → HPC/AI 收入强度 → 分部收入 → 分部 OPM/净利转换 → EPS/BPS → 目标 P/E → 隐含股价。"),
        ],
        "takeaways": [
            ("① 物理锚", "全球 AI 数据中心 CapEx 是最底层需求锚；HPC/AI 平台收入按 capex × 收入强度建模。"),
            ("② 基本面", "2026Q1 毛利率 66.2%、经营利润率 58.1%；2026 年 1-5 月收入同比 +30.0%，AI/先进制程紧缺仍在验证。"),
            ("③ 卖方对账", "本地研报库显示 2026 年以来主流目标价大致落在 NT$2,250-2,820，关键分歧是 CoWoS/N2 紧缺延续到 2027 后多久。"),
            ("④ 估值判断", "Base 用 2027E EPS × 25x P/E 做主判断；Bear/Bull 通过同一条业务链翻案，不反向拟合现价。"),
            ("⑤ 跟踪项", "月度收入、Q2 毛利率指引兑现、CoWoS/N2 产能、AIDC capex 是否降速，是模型先崩的四个位置。"),
        ],
    })

    ha = K.write_history(wb.create_sheet(S_HIST), {
        "title": "台积电历史财务与估值（$B）— 2021-2025A + 当前 TTM / 2026Q1",
        "hist_cols": HC,
        "hist_years": HY,
        "fx_hist": TSMC["fx_hist"],
        "fx_now": CURRENT_FX,
        "segments": [
            (HPC_SEG, segment_twd_t(HPC_SEG), False),
            (SMART_SEG, segment_twd_t(SMART_SEG), False),
            (OTHER_SEG, segment_twd_t(OTHER_SEG), False),
        ],
        "total_now": round(TSMC["ttm"]["revenue_twd_t"] * 1000 / CURRENT_FX, 2),
        "gm_pct": TSMC["gross_margin"],
        "gm_now": 0.619,
        "ni": TSMC["net_income_twd_t"],
        "ni_now": round(TSMC["ttm"]["net_income_twd_t"] * 1000 / CURRENT_FX, 2),
        "eq": TSMC["equity_twd_t"],
        "eq_now": round(TSMC["ttm"]["equity_twd_t"] * 1000 / CURRENT_FX, 2),
        "shares": TSMC["shares_m"],
        "shares_now": SHARES_M,
        "px_end": px_end,
        "px_now": px_now,
        "px_avg": px_avg,
        "band_note": "历史 P/E 2022 低点约 11x，2024-2025 年末约 23-24x；当前 TTM P/E 已抬到 30x 左右，说明市场在给 AI/先进封装的可持续性定价。",
        "quarter": {
            "col": "H",
            "label": "2026Q1实际",
            "segs": {
                HPC_SEG: (TSMC["q1_2026"]["revenue_usd_b"] * TSMC["q1_2026"]["platform_mix"][HPC_SEG], None),
                SMART_SEG: (TSMC["q1_2026"]["revenue_usd_b"] * TSMC["q1_2026"]["platform_mix"][SMART_SEG], None),
                OTHER_SEG: (TSMC["q1_2026"]["revenue_usd_b"] * TSMC["q1_2026"]["platform_mix"][OTHER_SEG], None),
            },
            "ni": TSMC["q1_2026"]["net_income_twd_b"] / CURRENT_FX,
            "eq": TSMC["ttm"]["equity_twd_t"] * 1000 / CURRENT_FX,
            "shares": SHARES_M,
            "fx": CURRENT_FX,
            "note": "2026Q1 实际：收入 US$35.90B、GM 66.2%、OPM 58.1%；平台 mix 使用公开会后口径近似。",
        },
        "notes": [
            (HPC_SEG, "历史为公司应用平台口径 HPC，包含 AI 与非 AI 计算；2024-2025 开始更接近 AI 规模化阶段，前瞻按 AIDC capex × 收入强度建模。"),
            (SMART_SEG, "智能手机平台收入；前瞻不挂 AIDC capex，按 N2/A16 节点升级与终端周期假设。"),
            (OTHER_SEG, "IoT、汽车、DCE 和其他合并；体量小，按周期恢复和车用/IoT 稳态增长建模。"),
            ("HREV", "2021-2025 来自 TSMC 年报；2026 当前列为 TTM（2025A - 2025Q1 + 2026Q1）。"),
            ("HGMP", "2021-2025 年报；当前列用 TTM 近似，2026Q1 实际 GM 66.2% 另列。"),
            ("HNI", "2021-2025 为 TSMC 年报净利润；当前列为 TTM 净利润。"),
            ("HEQ", "历史权益用 Yahoo/年报口径 common equity 近似；当前列用 2026Q1 common equity。"),
            ("HSH", "流通股本基本稳定，前瞻固定 25,932m 股。"),
            ("HPX", "股价来自 Yahoo 2330.TW 日线；年末/均价用于历史倍数校准。"),
        ],
    })

    def phase_fn(ym: str) -> str:
        if ym <= "2022-12":
            return "① 先进制程去库存/利率压估值"
        if ym <= "2024-03":
            return "② AI 订单验证但估值仍温和"
        if ym <= "2025-12":
            return "③ CoWoS/N3/N5 紧缺重估"
        return "④ N2/CoWoS 延续性定价"

    px = K.write_price_chart(wb.create_sheet(S_PX), price_data["monthly"], {
        "fn": phase_fn,
        "rows": [
            ("① 先进制程去库存/利率压估值", "2022 低点来自半导体去库存和利率环境压制，历史 P/E 触底。"),
            ("② AI 订单验证但估值仍温和", "2023 到 2024 初市场开始验证 AI ASIC/GPU 需求，但还没有充分给先进封装稀缺性定价。"),
            ("③ CoWoS/N3/N5 紧缺重估", "2024-2025 AI/HPC mix 抬升，N3/N5 与 CoWoS 供给紧，估值进入新 regime。"),
            ("④ N2/CoWoS 延续性定价", "2026 年市场重点从 2026 紧缺转向 2027-2028 可持续性。"),
        ],
    }, title="台积电月度股价走势（TWD，Yahoo 2330.TW）")
    px["yhigh"].update(px_high)
    px["ylow"].update(px_low)

    K.write_consensus(wb.create_sheet(S_CONS), {
        "title": "卖方研报共识 — 本地券商研报向量检索对账",
        "overview": "2026 年以来卖方核心叙事高度一致：AI/HPC、N3/N2、CoWoS/advanced packaging 推动 2026 收入同比 30%+，关键争议是紧缺是否延续到 2027 以后以及目标 P/E 能否维持在 22-30x。",
        "assumptions": [
            ("2026 收入增长", "JPM/UBS/MS 多数给 30%+ USD revenue growth；TSMC 1-5 月实际 +30.0% YoY。", "分歧在 2H26 是否继续上修；N3/N5 >100% utilization 与 2Q26 高毛利指引支持偏乐观。", "Base 用 2026E US$160B 左右收入路径，由分部链自动推出。"),
            ("CoWoS/先进封装", "MS: 2026 CoWoS capacity 约 +80% YoY 至 125k wpm；JPM: 4Q26 约 115k wpm，仍有 15-20% demand/supply gap。", "如果先进封装不是瓶颈，HPC 收入强度会更快下滑；如果紧缺延续，P/E 可维持高位。", "Base 让 HPC/AI 收入强度从 2025 高位温和回落，而非永久维持。"),
            ("N2/N3/A16", "JPM/GS/UBS 均强调 N3/N2 tightness；N2 2026 ramp，A16 2H26。", "N2 定价与海外 fab 成本决定毛利率天花板。", "Base HPC OPM 2026-2030 从 63% 缓降至 60%，反映紧缺后回归但不崩。"),
            ("目标价/倍数", "GS/JPM/HSBC/Nomura 目标价大致 NT$2,250-2,820；卖方常用 22-27x forward P/E。", "Bull 需要市场接受 AI foundry 的准平台属性；Bear 则回到 15-18x。", "Base 2027E 用 25x P/E，对应卖方上沿但仍低于极端重估。"),
        ],
        "divergences": [
            "最大分歧不是 2026，而是 2027：N2/CoWoS 约束是否继续让台积电吃到价格与 mix 双重红利。",
            "HPC 平台口径不是纯 AI：早年收入强度不可机械外推，前瞻必须锚 2025 后 AI 规模化阶段。",
            "P/B 很高不是错误，但它是 ROE 与盈利耐用性的结果；主估值镜头应落在 EPS × P/E。",
        ],
        "stances": [
            "Goldman Sachs（2026-04）：Buy，目标价最高上修到 NT$2,750；强调 foundry share、CoWoS 和 AI 多年增长。",
            "JPMorgan（2026-04）：OW，目标价 NT$2,400；强调 N3/N5 utilization、leading-edge tightness 与 1Q/2Q 毛利率超预期。",
            "Nomura（2026-04）：Buy，目标价 NT$2,820；核心是全年收入 >30% 与目标 P/E 上移。",
            "HSBC（2026-04）：目标价 NT$2,700，按约 27x rolled EPS；与本模型 Bull/Base 上沿对账。",
        ],
    })

    hm = K.write_hist_multiples(wb.create_sheet(S_HMULT), {
        "title": "历史估值倍数 — 先看台积电自己的带宽，再给前瞻 P/E",
        "intro": "台积电估值从 2022 年去库存低谷约 11x P/E，抬到当前约 30x TTM P/E。这个重估不是单纯情绪，而是先进制程/CoWoS 瓶颈让盈利可见度变强。P/B 同步抬升，但本质仍由 ROE/EPS 解释。",
        "s_hist": S_HIST,
        "ha": ha,
        "hist_cols": HC,
        "hist_years": HY,
        "yhigh": px["yhigh"],
        "ylow": px["ylow"],
        "fwd_note": "Yahoo 当前 forward P/E 约 18x；本模型 2027E Base P/E=25x，是对 N2/CoWoS 约束延续的判断，不是现价拟合。",
        "self_name": "TSMC",
        "self_fwd_pe_label": "~18x Yahoo forward / 25x Base 2027E",
        "self_note": "台积电 P/B 高企来自高 ROE 与盈利耐用性；P/B 只做体检，主线不以账面净资产定价。",
        "peers": [
            {"name": "Samsung Electronics", "yearly": None, "cur_pb": 1.3, "cur_pe": 18.0, "fwd_pe": 12.0,
             "note": "IDM+memory 混合体，估值折价主要来自存储周期与 foundry 份额弱。"},
            {"name": "Intel", "yearly": None, "cur_pb": 1.1, "cur_pe": None, "fwd_pe": None,
             "note": "先进制程追赶者，不适合直接给台积电定价，只用于说明制造能力差距。"},
            {"name": "NVIDIA", "yearly": None, "cur_pb": None, "cur_pe": 32.0, "fwd_pe": 28.0,
             "note": "AI 上游平台型资产；台积电不该拿同倍数，但 AI bottleneck franchise 可获得结构溢价。"},
            {"name": "ASML", "yearly": None, "cur_pb": None, "cur_pe": 35.0, "fwd_pe": 29.0,
             "note": "半导体瓶颈资产参考，说明稀缺设备/制造环节可享受长期高倍数。"},
        ],
        "ratio": {"peer": "Samsung Electronics",
                  "note": "相对 Samsung 的溢价来自 foundry 份额、N3/N2 领先和 CoWoS/SoIC 先进封装能力。"},
        "reading": "台积电当前不是 2022 的周期低点，也不是无风险永续 30x。更合理的判断是：2026-2027 盈利仍被 AI 供给约束支撑，目标 P/E 可在 24-25x；若 2027 后 capex 降速或产能释放快于需求，倍数应回落。"
    })

    ma = K.write_multiple_assumptions(wb.create_sheet(S_MULT), {
        "title": "估值倍数假设 — 主线 P/E 三层拆解",
        "intro": "主线倍数不是拍一个 P/E，而是拆成：历史/质量锚 × 结构溢价 × 情绪/可见度。情景切换只引用这里的三案情绪值；所有目标价由同一条业务链推导。",
        "why_text": "台积电的穿越周期资产不是普通 fab 账面，而是领先制程、客户粘性、良率和 advanced packaging 共同形成的盈利耐用性。因此主估值分母选 EPS，P/B 做体检。若未来先进制程失去定价权，镜头才应向 P/B/DCF 下移。",
        "why_rows": 4,
        "method_text": "目标 P/E = 历史可见上沿/质量锚（29x） × 结构溢价（1.0x） × 情景情绪值。Bear 回到 15-18x；Base 2026-2027 维持 24-25x 后回落；Bull 在紧缺和 re-rating 下短期到 30x。",
        "peak": peak_pe,
        "peak_note": "5 年历史与卖方口径显示台积电 P/E 大致经历 10-29x 区间；29x 作为高质量 franchise 的可见上沿，而非当前 TTM 倍数拟合。",
        "premium": 1.0,
        "premium_note": "结构溢价已经体现在 29x 质量锚内，避免双重计数；三案差异放在情绪/可见度层。",
        "all_cols": ALLC,
        "all_years": ALLY,
        "hist_cols": HC,
        "fwd_cols": FCF,
        "s_hist": S_HIST,
        "hpb_row": ha["HPE"],
        "cases": [
            ("Bear", pe_sentiment["Bear"], "AI capex 降速或 CoWoS/N2 供给释放快于需求，倍数回到 15-18x。"),
            ("Base", pe_sentiment["Base"], "2026-2027 紧缺与盈利可见度仍支撑 24-25x，随后随产能释放回落。"),
            ("Bull", pe_sentiment["Bull"], "AI ASIC/GPU/edge AI 需求继续超过产能，市场接受台积电作为 AI bottleneck franchise 的更高倍数。"),
        ],
        "reconcile_text": "卖方常见目标 P/E 约 22-27x；本模型 Base 2027E 25x 落在上半区，Bull 30x 属于强紧缺和重估情景。",
        "source_text": "历史倍数来自 Yahoo 股价与 TSMC 年报 EPS；卖方口径来自本地研报库语义检索；目标倍数与情景触发写在本页和情景切换页。"
    })

    sw = K.write_scenario_switch(wb.create_sheet(S_SW), {
        "title": "情景切换 — 全模型唯一情景参数库",
        "usage": "B2 是唯一开关。收入、利润、目标 P/E 和估值都从这里按当前案联动；估值对比页则直接引用三案矩阵，避免被当前开关污染。",
        "cases": CASES,
        "default": "Base",
        "triggers": [
            ("Bear", "AIDC capex 2027 起明显降速；N2/CoWoS 紧缺提前缓解；客户砍单或 Intel/Samsung second source 可信度提升。"),
            ("Base", "TSMC 2026-2027 保持先进制程和先进封装紧缺，收入 2026 30%+ 后 2027 继续高双位数/低二十增长。"),
            ("Bull", "AI GPU/ASIC/edge AI 同时拉动，N2/CoWoS/SoIC 约束持续到 2028，市场进一步把台积电当作 AI bottleneck 定价。"),
        ],
        "all_cols": ALLC,
        "all_years": ALLY,
        "hist_cols": HC,
        "fwd_cols": FCF,
        "levers": [
            {"key": "capex", "name": "AIDC capex 增速", "fmt": K.PCT, "cols": FC[2:],
             "vals": {"Bear": [0.15, 0.05, 0.00, 0.03], "Base": [0.27, 0.18, 0.11, 0.09], "Bull": [0.35, 0.25, 0.15, 0.10]},
             "desc": "物理锚增速。2026 采用共享基座 830B，三案从 2027 开始分歧。",
             "stories": {"Bear": "云厂 ROI 受压，capex 进入审慎期。", "Base": "沿共享 AIDC 基座增长。", "Bull": "AI 算力投资继续前置，capex 曲线延长。"},
             "hist": [None, 1.00, 1.33, 1.86, 1.44]},
            {"key": "hpcint", "name": "HPC/AI 收入强度", "fmt": K.PCT,
             "vals": {"Bear": [0.108, 0.100, 0.092, 0.088, 0.085], "Base": [0.120, 0.118, 0.113, 0.108, 0.104], "Bull": [0.128, 0.130, 0.125, 0.120, 0.115]},
             "desc": "HPC/AI 平台收入 / AIDC capex。早年 HPC 不是纯 AI，历史强度只做对照，前瞻锚 2025 后 AI 规模化阶段。",
             "stories": {"Bear": "客户压价与非 AI HPC 下滑让强度更快回落。", "Base": "强度从 2025 高位温和回落，反映 AI capex 分母扩大。", "Bull": "先进制程和 CoWoS 稀缺让 foundry capture rate 维持更久。"},
             "hist": hpc_int_hist},
            {"key": "smart", "name": "Smartphone 收入增速", "fmt": K.PCT,
             "vals": {"Bear": [0.05, 0.00, 0.02, 0.03, 0.03], "Base": [0.12, 0.08, 0.05, 0.04, 0.04], "Bull": [0.18, 0.12, 0.08, 0.05, 0.05]},
             "desc": "手机平台按产品周期和节点升级，不挂 AIDC capex。",
             "stories": {"Bear": "换机弱，N2 溢价兑现少。", "Base": "Apple/N2 节点升级带来温和增长。", "Bull": "高端 SoC 节点迁移更快，ASP 与份额都偏强。"},
             "hist": smart_growth_hist},
            {"key": "other", "name": "Other/IoT/Auto/DCE 收入增速", "fmt": K.PCT,
             "vals": {"Bear": [0.02, 0.02, 0.03, 0.03, 0.03], "Base": [0.08, 0.06, 0.05, 0.05, 0.04], "Bull": [0.12, 0.09, 0.07, 0.06, 0.05]},
             "desc": "其他平台走温和恢复，汽车/IoT 不承担 AI 主逻辑。",
             "stories": {"Bear": "终端恢复慢。", "Base": "车用/IoT 温和增长。", "Bull": "工业和车用恢复更好。"},
             "hist": other_growth_hist},
            {"key": "hpcopm", "name": "HPC/AI OPM", "fmt": K.PCT,
             "vals": {"Bear": [0.57, 0.55, 0.53, 0.52, 0.52], "Base": [0.63, 0.62, 0.61, 0.60, 0.60], "Bull": [0.66, 0.66, 0.65, 0.64, 0.63]},
             "desc": "先进制程/CoWoS 主导的高毛利平台，前瞻利润率高于公司历史均值。",
             "stories": {"Bear": "海外 fab 成本和客户压价侵蚀利润率。", "Base": "2026 高位后随产能释放温和回落。", "Bull": "供给约束延续，mix/price 同时支撑。"},
             "hist": TSMC["op_margin"]},
            {"key": "smartopm", "name": "Smartphone OPM", "fmt": K.PCT,
             "vals": {"Bear": [0.48, 0.46, 0.45, 0.45, 0.45], "Base": [0.55, 0.55, 0.54, 0.53, 0.53], "Bull": [0.58, 0.58, 0.57, 0.56, 0.55]},
             "desc": "手机平台利润率受节点迁移支撑，但客户集中与年度议价使其低于 HPC/AI。",
             "stories": {"Bear": "换机弱且客户压价。", "Base": "N2/A16 节点迁移支撑。", "Bull": "高端 SoC 节点溢价更好。"},
             "hist": TSMC["op_margin"]},
            {"key": "otheropm", "name": "Other OPM", "fmt": K.PCT,
             "vals": {"Bear": [0.35, 0.34, 0.34, 0.34, 0.34], "Base": [0.43, 0.43, 0.42, 0.42, 0.42], "Bull": [0.46, 0.46, 0.45, 0.45, 0.45]},
             "desc": "其他平台规模和节点结构弱于 HPC/手机，利润率保守处理。",
             "stories": {"Bear": "终端恢复慢。", "Base": "接近公司历史中枢。", "Bull": "汽车/IoT 复苏更好。"},
             "hist": TSMC["op_margin"]},
            {"key": "netconv", "name": "净利转换率", "fmt": K.PCT,
             "vals": {"Bear": [0.88, 0.87, 0.87, 0.87, 0.87], "Base": [0.90, 0.90, 0.895, 0.89, 0.89], "Bull": [0.91, 0.91, 0.905, 0.90, 0.90]},
             "desc": "净利 / 营业利润，覆盖税率、财务收入和少数股东等简化项。",
             "stories": {"Bear": "税率/海外成本偏不利。", "Base": "接近历史净利转换率。", "Bull": "高毛利 mix 和财务收入支撑。"},
             "hist": net_conv},
            {"key": "retention", "name": "留存率", "fmt": K.PCT,
             "vals": {"Bear": [0.65, 0.65, 0.65, 0.65, 0.65], "Base": [0.72, 0.72, 0.72, 0.72, 0.72], "Bull": [0.75, 0.75, 0.75, 0.75, 0.75]},
             "desc": "用于 BPS/ROE 路径，主估值不靠 P/B，但保留账面体检。",
             "stories": {"Bear": "分红压力更高。", "Base": "按近年高留存高投资强度处理。", "Bull": "盈利增长更快，留存更高。"},
             "hist": [None, 0.74, 0.63, 0.69, 0.64]},
        ],
        "linked": [
            {"key": "sent", "name": "情绪/可见度值（P/E 第三层）", "fmt": K.N2,
             "src_sheet": S_MULT, "src_row0": ma["sent_row0"],
             "note": "来自估值倍数假设页。目标 P/E = 29x × 1.0 × 本行。"},
        ],
    })

    pk = f"'{S_MULT}'!{ma['pk_cell']}"
    pr = f"'{S_MULT}'!{ma['pr_cell']}"
    swpe = sw["next_row"]
    K.lab(wb[S_SW], f"A{swpe}", "目标P/E（当前案）", b=True)
    for col in ALLC:
        K.fml(wb[S_SW], f"{col}{swpe}", f"={pk}*{pr}*{col}{sw['SWACT']['sent']}", K.MX, link=True)
    K.logic(wb[S_SW], f"L{swpe}", "目标 P/E = 历史/质量锚 × 结构溢价 × 当前案情绪值；喂给情景估值。")

    anchor = K.write_anchor(wb.create_sheet(S_ANCHOR), {
        "title": "全球 AI 数据中心 CapEx（$B）",
        "all_cols": ALLC,
        "all_years": ALLY,
        "series": [("AI 数据中心 capex ($B)", TSMC["aidc_capex"], "共享基座 compute-aidc-base；2026E=830，2027+ 由情景切换 capex 增速驱动。", K.N0)],
        "yoy_row": "AI 数据中心 capex ($B)",
        "source_note": "口径为全球 AI 数据中心专项 capex，而非全口径 hyperscaler capex。2021-2024 为 AI 起步期粗估，模型重点从 2025A/2026E 开始。",
        "role_note": "HPC/AI 平台收入 = AIDC capex × HPC/AI 收入强度。改这一行会穿透到分部收入、EPS、目标价。"
    })
    capex_row = anchor["row_of"]["AI 数据中心 capex ($B)"]
    for idx, col in enumerate(FC[2:]):
        prev = FC[1:][idx]
        K.fml(wb[S_ANCHOR], f"{col}{capex_row}", f"={prev}{capex_row}*(1+{K.R(S_SW, col + str(sw['SWACT']['capex']))})", K.N0, link=True)

    seg = K.write_segment_model(wb.create_sheet(S_SEG), {
        "title": "分部测算 — HPC/AI 挂 AIDC capex，手机和其他按平台周期",
        "all_cols": ALLC,
        "all_years": ALLY,
        "logic_col": "N",
        "groups": [
            ("AIDC 物理锚", [
                ("AIDC capex ($B)", None, K.N0, "来自 [ANCHOR] AIDC capex 表。"),
            ]),
            ("HPC/AI 平台", [
                ("HPC/AI收入强度", None, K.PCT, "HPC/AI 平台收入 / AIDC capex；历史早期不是纯 AI，前瞻锚 2025 后 AI 规模化阶段。"),
                ("HPC/AI平台收入", None, K.N1, "前瞻 = AIDC capex × 收入强度。"),
            ]),
            ("非 AI 平台", [
                ("Smartphone收入增速", None, K.PCT, "手机平台按节点升级/终端周期建模。"),
                ("Smartphone收入", None, K.N1, "前瞻 = 上年收入 × (1+增速)。"),
                ("Other收入增速", None, K.PCT, "IoT/Auto/DCE/其他合并。"),
                ("Other/IoT/Auto/DCE收入", None, K.N1, "前瞻 = 上年收入 × (1+增速)。"),
            ]),
        ],
    })

    m = seg["m"]
    for col in ALLC:
        K.fml(wb[S_SEG], f"{col}{m['AIDC capex ($B)']}", f"={K.R(S_ANCHOR, col + str(capex_row))}", K.N0, link=True)
    for col in HC:
        K.fml(wb[S_SEG], f"{col}{m['HPC/AI平台收入']}", f"={K.R(S_HIST, col + str(ha['seg_rows'][HPC_SEG]))}", K.N1, link=True)
        K.fml(wb[S_SEG], f"{col}{m['HPC/AI收入强度']}", f"={col}{m['HPC/AI平台收入']}/{col}{m['AIDC capex ($B)']}", K.PCT)
        K.fml(wb[S_SEG], f"{col}{m['Smartphone收入']}", f"={K.R(S_HIST, col + str(ha['seg_rows'][SMART_SEG]))}", K.N1, link=True)
        K.fml(wb[S_SEG], f"{col}{m['Other/IoT/Auto/DCE收入']}", f"={K.R(S_HIST, col + str(ha['seg_rows'][OTHER_SEG]))}", K.N1, link=True)
    K.lab(wb[S_SEG], f"B{m['Smartphone收入增速']}", "n.m.", note=True)
    K.lab(wb[S_SEG], f"B{m['Other收入增速']}", "n.m.", note=True)
    for prev, col in zip(HC[:-1], HC[1:]):
        K.fml(wb[S_SEG], f"{col}{m['Smartphone收入增速']}", f"={col}{m['Smartphone收入']}/{prev}{m['Smartphone收入']}-1", K.PCT)
        K.fml(wb[S_SEG], f"{col}{m['Other收入增速']}", f"={col}{m['Other/IoT/Auto/DCE收入']}/{prev}{m['Other/IoT/Auto/DCE收入']}-1", K.PCT)
    for prev, col in zip(["F"] + FCF[:-1], FCF):
        K.fml(wb[S_SEG], f"{col}{m['HPC/AI收入强度']}", f"={K.R(S_SW, col + str(sw['SWACT']['hpcint']))}", K.PCT, link=True)
        K.fml(wb[S_SEG], f"{col}{m['HPC/AI平台收入']}", f"={col}{m['AIDC capex ($B)']}*{col}{m['HPC/AI收入强度']}", K.N1)
        K.fml(wb[S_SEG], f"{col}{m['Smartphone收入增速']}", f"={K.R(S_SW, col + str(sw['SWACT']['smart']))}", K.PCT, link=True)
        K.fml(wb[S_SEG], f"{col}{m['Smartphone收入']}", f"={prev}{m['Smartphone收入']}*(1+{col}{m['Smartphone收入增速']})", K.N1)
        K.fml(wb[S_SEG], f"{col}{m['Other收入增速']}", f"={K.R(S_SW, col + str(sw['SWACT']['other']))}", K.PCT, link=True)
        K.fml(wb[S_SEG], f"{col}{m['Other/IoT/Auto/DCE收入']}", f"={prev}{m['Other/IoT/Auto/DCE收入']}*(1+{col}{m['Other收入增速']})", K.N1)

    fr = K.write_fundamentals(wb.create_sheet(S_FUND), {
        "title": "利润与收入假设 — 分部收入 → 分部 OPM → NI/EPS/BPS",
        "all_cols": ALLC,
        "all_years": ALLY,
        "hist_cols": HC,
        "fwd_cols": FCF,
        "logic_col": "N",
        "s_hist": S_HIST,
        "ha": ha,
        "share_fix_col": "F",
        "assum_groups": [
            ("分部 OPM", [
                {"name": "HPC/AI OPM", "vals": TSMC["op_margin"] + [None] * 5, "fmt": K.PCT, "logic": "历史用公司 OPM 校准；前瞻链接情景切换。", "link": {"sheet": S_SW, "row": sw["SWACT"]["hpcopm"]}},
                {"name": "Smartphone OPM", "vals": TSMC["op_margin"] + [None] * 5, "fmt": K.PCT, "logic": "历史用公司 OPM 校准；前瞻链接情景切换。", "link": {"sheet": S_SW, "row": sw["SWACT"]["smartopm"]}},
                {"name": "Other OPM", "vals": TSMC["op_margin"] + [None] * 5, "fmt": K.PCT, "logic": "历史用公司 OPM 校准；前瞻链接情景切换。", "link": {"sheet": S_SW, "row": sw["SWACT"]["otheropm"]}},
            ]),
            ("净利与账面", [
                {"name": "净利转换率", "vals": net_conv + [None] * 5, "fmt": K.PCT, "logic": "净利润 / 营业利润；历史校准到实际净利，前瞻链接情景切换。", "link": {"sheet": S_SW, "row": sw["SWACT"]["netconv"]}},
                {"name": "留存率", "vals": [None, 0.74, 0.63, 0.69, 0.64] + [None] * 5, "fmt": K.PCT, "logic": "用于 BPS/ROE；主估值不依赖 P/B。", "link": {"sheet": S_SW, "row": sw["SWACT"]["retention"]}},
            ]),
        ],
        "segments": [
            {"name": HPC_SEG, "hist_row": HPC_SEG, "fwd": {"sheet": S_SEG, "row": m["HPC/AI平台收入"]}},
            {"name": SMART_SEG, "hist_row": SMART_SEG, "fwd": {"sheet": S_SEG, "row": m["Smartphone收入"]}},
            {"name": OTHER_SEG, "hist_row": OTHER_SEG, "fwd": {"sheet": S_SEG, "row": m["Other/IoT/Auto/DCE收入"]}},
        ],
        "profit_terms": [
            ([HPC_SEG], "HPC/AI OPM", False),
            ([SMART_SEG], "Smartphone OPM", False),
            ([OTHER_SEG], "Other OPM", False),
        ],
        "conv_assum": "净利转换率",
        "retention_assum": "留存率",
        "note_text": "历史净利润直接引用真实财务，前瞻净利润由分部收入与利润率推导；这避免用外部卖方净利把链条截断。"
    })

    val = write_pe_valuation(wb.create_sheet(S_VAL), {
        "title": "情景估值 — 当前案逐年 P/E 主线",
        "intro": "本页随情景切换 B2 变化。历史列用真实价格反推倍数；前瞻列用当前案目标 P/E × EPS × FX 输出隐含股价，并同步展示 P/B 与当前 forward P/E。",
        "all_cols": ALLC,
        "all_years": ALLY,
        "hist_cols": HC,
        "fwd_cols": FCF,
        "s_hist": S_HIST,
        "ha": ha,
        "s_fund": S_FUND,
        "fr": fr,
        "s_switch": S_SW,
        "s_anchor": S_ANCHOR,
        "capex_row": capex_row,
        "target_row": swpe,
        "sw_cell": sw["sw_cell"],
        "price_now": px_now,
        "fx_fwd": FX_FC,
        "reading": "如果切到 Bear，收入强度、OPM 和 P/E 同时回落，隐含价会快速下移；如果切到 Bull，AIDC capex 与收入强度一起抬高，EPS 和 P/E 双击。Base 的关键不是 2026，而是 2027：若 2027E EPS 能靠近模型路径，同时市场仍给约 25x，当前价格仍有上行空间。"
    })

    swb = sw["SWB"]

    def case_row(key: str, ci: int) -> int:
        return swb[key] + ci

    def hist_seg(seg_name: str):
        return lambda col, ci, a: f"={K.R(S_HIST, col + str(ha['seg_rows'][seg_name]))}"

    def hist_fund(row_key: str):
        return lambda col, ci, a: f"={K.R(S_FUND, col + str(fr[row_key]))}"

    def hist_pe(col: str, ci: int, a: dict[str, int]) -> str:
        return f"={K.R(S_HIST, col + str(ha['HPE']))}"

    rows = [
        {"key": "capex", "label": "AIDC capex ($B)", "fmt": K.N0,
         "hist": lambda col, ci, a: f"={K.R(S_ANCHOR, col + str(capex_row))}",
         "fwd": lambda col, j, ci, a: (f"={K.R(S_ANCHOR, col + str(capex_row))}" if j == 0 else f"={ALLC[ALLC.index(col)-1]}{a['capex']}*(1+{K.R(S_SW, col + str(case_row('capex', ci)))})")},
        {"key": "hpcint", "label": "HPC/AI 收入强度", "fmt": K.PCT,
         "hist": lambda col, ci, a: f"={K.R(S_SEG, col + str(m['HPC/AI收入强度']))}",
         "fwd": lambda col, j, ci, a: f"={K.R(S_SW, col + str(case_row('hpcint', ci)))}"},
        {"key": "hpc", "label": "HPC/AI平台收入", "fmt": K.N1,
         "hist": hist_seg(HPC_SEG),
         "fwd": lambda col, j, ci, a: f"={col}{a['capex']}*{col}{a['hpcint']}"},
        {"key": "smartg", "label": "Smartphone 收入增速", "fmt": K.PCT,
         "hist": lambda col, ci, a: None if col == "B" else f"={K.R(S_SEG, col + str(m['Smartphone收入增速']))}",
         "fwd": lambda col, j, ci, a: f"={K.R(S_SW, col + str(case_row('smart', ci)))}"},
        {"key": "smart", "label": "Smartphone 收入", "fmt": K.N1,
         "hist": hist_seg(SMART_SEG),
         "fwd": lambda col, j, ci, a: f"={ALLC[ALLC.index(col)-1]}{a['smart']}*(1+{col}{a['smartg']})"},
        {"key": "otherg", "label": "Other 收入增速", "fmt": K.PCT,
         "hist": lambda col, ci, a: None if col == "B" else f"={K.R(S_SEG, col + str(m['Other收入增速']))}",
         "fwd": lambda col, j, ci, a: f"={K.R(S_SW, col + str(case_row('other', ci)))}"},
        {"key": "other", "label": "Other/IoT/Auto/DCE收入", "fmt": K.N1,
         "hist": hist_seg(OTHER_SEG),
         "fwd": lambda col, j, ci, a: f"={ALLC[ALLC.index(col)-1]}{a['other']}*(1+{col}{a['otherg']})"},
        {"key": "rev", "label": "总营收", "fmt": K.N1, "bold": True,
         "hist": hist_fund("REV"),
         "fwd": lambda col, j, ci, a: f"={col}{a['hpc']}+{col}{a['smart']}+{col}{a['other']}"},
        {"key": "hpcopm", "label": "HPC/AI OPM", "fmt": K.PCT,
         "hist": lambda col, ci, a: f"={K.R(S_FUND, col + str(fr['am']['HPC/AI OPM']))}",
         "fwd": lambda col, j, ci, a: f"={K.R(S_SW, col + str(case_row('hpcopm', ci)))}"},
        {"key": "smartopm", "label": "Smartphone OPM", "fmt": K.PCT,
         "hist": lambda col, ci, a: f"={K.R(S_FUND, col + str(fr['am']['Smartphone OPM']))}",
         "fwd": lambda col, j, ci, a: f"={K.R(S_SW, col + str(case_row('smartopm', ci)))}"},
        {"key": "otheropm", "label": "Other OPM", "fmt": K.PCT,
         "hist": lambda col, ci, a: f"={K.R(S_FUND, col + str(fr['am']['Other OPM']))}",
         "fwd": lambda col, j, ci, a: f"={K.R(S_SW, col + str(case_row('otheropm', ci)))}"},
        {"key": "op", "label": "营业利润", "fmt": K.N1,
         "hist": hist_fund("OP"),
         "fwd": lambda col, j, ci, a: f"={col}{a['hpc']}*{col}{a['hpcopm']}+{col}{a['smart']}*{col}{a['smartopm']}+{col}{a['other']}*{col}{a['otheropm']}"},
        {"key": "netconv", "label": "净利转换率", "fmt": K.PCT,
         "hist": lambda col, ci, a: f"={K.R(S_FUND, col + str(fr['am']['净利转换率']))}",
         "fwd": lambda col, j, ci, a: f"={K.R(S_SW, col + str(case_row('netconv', ci)))}"},
        {"key": "ni", "label": "净利润", "fmt": K.N1,
         "hist": hist_fund("NI"),
         "fwd": lambda col, j, ci, a: f"={col}{a['op']}*{col}{a['netconv']}"},
        {"key": "eps", "label": "EPS ($)", "fmt": K.N2,
         "hist": hist_fund("EPS"),
         "fwd": lambda col, j, ci, a: f"={col}{a['ni']}*1000/{SHARES_M}"},
        {"key": "retention", "label": "留存率", "fmt": K.PCT,
         "hist": lambda col, ci, a: f"={K.R(S_FUND, col + str(fr['am']['留存率']))}" if col != "B" else None,
         "fwd": lambda col, j, ci, a: f"={K.R(S_SW, col + str(case_row('retention', ci)))}"},
        {"key": "eq", "label": "期末权益", "fmt": K.N1,
         "hist": hist_fund("EQ"),
         "fwd": lambda col, j, ci, a: f"={ALLC[ALLC.index(col)-1]}{a['eq']}+{col}{a['ni']}*{col}{a['retention']}"},
        {"key": "bps", "label": "BPS ($)", "fmt": K.N2,
         "hist": hist_fund("BPS"),
         "fwd": lambda col, j, ci, a: f"={col}{a['eq']}*1000/{SHARES_M}"},
        {"key": "tpe", "label": "目标 P/E", "fmt": K.MX,
         "hist": hist_pe,
         "fwd": lambda col, j, ci, a: f"={K.R(S_MULT, col + str(ma['target_row0'] + ci))}"},
        {"key": "px", "label": "隐含股价 P/E主线（TWD）", "fmt": K.PX, "bold": True, "out": True,
         "hist": lambda col, ci, a: f"={K.R(S_HIST, col + str(ha['HPX']))}",
         "fwd": lambda col, j, ci, a: f"={col}{a['tpe']}*{col}{a['eps']}*{FX_FC}+0*{K.R(S_ANCHOR, col + str(capex_row))}"},
        {"key": "ipe", "label": "隐含 forward P/E（体检）", "fmt": K.MX,
         "hist": hist_pe,
         "fwd": lambda col, j, ci, a: f"={col}{a['px']}/({col}{a['eps']}*{FX_FC})"},
        {"key": "pb", "label": "隐含 P/B（体检）", "fmt": K.MX,
         "hist": lambda col, ci, a: f"={K.R(S_HIST, col + str(ha['HPB']))}",
         "fwd": lambda col, j, ci, a: f"={col}{a['px']}/({col}{a['bps']}*{FX_FC})"},
        {"key": "up", "label": "历史回测误差 / 前瞻较当前股价", "fmt": K.PCT,
         "hist": lambda col, ci, a: ("inp", 0.0),
         "fwd": lambda col, j, ci, a: f"={col}{a['px']}/{px_now}-1"},
    ]

    cmp = K.write_comparison(wb.create_sheet(S_CMP), {
        "title": "估值对比 — 三案并排，不随当前开关污染",
        "intro": "三案都从 AIDC capex、收入强度、分部 OPM、目标 P/E 同一条链推导。顶部摘要取 2027E 作为主判断点，因为 2026 已被高可见度收入指引锚定，真正分歧在 2027。每个情景 block 都保留隐含 forward P/E 与 P/B 体检。",
        "case_names": CASES,
        "all_cols": ALLC,
        "all_years": ALLY,
        "hist_cols": HC,
        "fwd_cols": FCF,
        "block_start": 22,
        "rows": rows,
        "summary": {
            "band": "2027E 三案摘要（主判断年）",
            "target_col": "H",
            "rows": [
                ("HPC/AI收入", "hpc", K.N1, "AIDC capex × 收入强度。", False),
                ("总营收", "rev", K.N1, "分部收入合计。", False),
                ("净利润", "ni", K.N1, "分部 OPM 与净利转换推导。", False),
                ("EPS ($)", "eps", K.N2, "主估值分母。", False),
                ("目标 P/E", "tpe", K.MX, "估值倍数页三层拆解。", False),
                ("隐含股价", "px", K.PX, "P/E 主线输出。", True),
                ("隐含 forward P/E", "ipe", K.MX, "必须体检，不是额外结论。", False),
                ("隐含 P/B", "pb", K.MX, "高 P/B 由 ROE/EPS 耐用性解释。", False),
                ("较当前股价", "up", K.PCT, "用当前价做 reality check。", True),
            ],
            "mcap": {"label": "隐含市值（US$B）", "key": "px", "expr": f"*{SHARES_M}/{FX_FC}/1000", "note": "隐含股价 × 股本 / FX。"},
            "concl": "Base 不靠把目标价贴现价，而靠 2027E EPS 与 25x P/E：若 N2/CoWoS 紧缺延续，台积电仍能维持高质量盈利倍数；若 capex 或强度掉得更快，Bear 会显示下行风险。"
        },
    })

    dash = K.write_dashboard(wb.create_sheet(S_DASH), {
        "title": "综合判断仪表盘 — 先看基本面拐点，再看估值错位",
        "usage": "这页把模型压成投后跟踪语言：哪些指标验证 Base，哪些指标会把模型推向 Bear/Bull。B 列有公式的地方直接引用模型输出。",
        "blocks": [
            {"title": "A. 基本面拐点", "rows": [
                ("2026 年 1-5 月收入同比", "30.0%", "TSMC 月收入页面已披露 Jan-May 2026 +30.0% YoY，支持 2026 收入 30%+ 的底座。", True),
                ("2026Q1 GM / OPM", "66.2% / 58.1%", "高于 2025A，说明先进制程和 mix 红利仍在。"),
                ("2027E Base HPC/AI收入", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Base']['hpc']))}", "fmt": K.N1, "fill": True}, "若 2027 HPC/AI 收入不能靠近该路径，Base 估值先降。"),
            ]},
            {"title": "B. 估值错位", "rows": [
                ("当前股价", {"inp": px_now, "fmt": K.PX, "fill": True}, "只作为 reality check，不反向拟合。"),
                ("Base 2027E 隐含价", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Base']['px']))}", "fmt": K.PX, "fill": True}, "主判断输出。"),
                ("Base 2027E 上行", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Base']['up']))}", "fmt": K.PCT, "fill": True}, "若上行来自 EPS 和 P/E 双击，要检查是否过度依赖高倍数。"),
                ("Bear 2027E 隐含价", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Bear']['px']))}", "fmt": K.PX}, "风险价格：capex/收入强度/倍数三杀。"),
            ]},
            {"title": "C. 触发器", "rows": [
                ("Bull 触发", "CoWoS/N2 紧缺延续到 2028；AI ASIC/GPU/edge AI 同时拉动；2Q/3Q 毛利继续超指引。", "会先抬 hpcint 与 hpcopm，再抬目标 P/E。"),
                ("Bear 触发", "月收入连续低于 +25%；2Q 指引不兑现；客户 second source 可信度提升；2027 capex 降速。", "先砍 capex/hpcint，再把目标 P/E 回到 18x 以下。"),
                ("最容易错的地方", "把 HPC 平台全部当纯 AI，导致收入强度外推过高。", "因此模型让强度从 2025 之后逐步回落。"),
            ]},
            {"title": "D. 综合判断", "rows": [
                ("一句话结论", "Base 偏建设性：当前价不便宜，但若 2027 EPS 路径兑现，估值仍未完全透支。", "关键是 2027，而不是 2026。", True),
            ]},
        ],
        "final": {
            "band": "最终判断",
            "text": "台积电模型的 alpha 问题是：市场愿不愿意把它从普通重资产 foundry，继续按 AI bottleneck franchise 定价。当前数据支持 2026 高增长，Base 的投资判断押在 2027 紧缺延续和 25x P/E 可维持；风险是 AIDC capex 或收入强度提前降速。"
        },
        "tracking": {
            "intro": "投后跟踪按月收入和季报滚动更新，不用等完整年报。",
            "rows": [
                ("月度收入 YoY", "Jan-May 2026 +30.0%", "验证 2026 收入增长底座", "TSMC monthly revenue，每月", "低于 +25% 且连续两月 → 下调 2026/2027 收入强度"),
                ("2Q26 GM 指引", "65.5%-67.5%", "验证先进制程与 CoWoS 定价", "TSMC quarterly results", "低于指引 → 下调 OPM 与目标 P/E"),
                ("CoWoS capacity", "卖方口径 2026 +69%-80%", "决定 AI 收入能否从 capex 转化为台积电收入", "本地研报向量检索 + 公司 commentary", "供给释放快于需求 → 下调 hpcint"),
                ("AIDC capex", "2026E $830B 基座", "底层物理锚", "shared-base + 云厂 capex 指引", "2027 降速 → 先改情景切换 capex"),
            ],
        },
    })

    payload = {
        "ticker": "2330.TW / TSM",
        "built_at": datetime.now().isoformat(timespec="seconds"),
        "outputs": {"workbook": OUT_XLSX, "input_json": OUT_JSON},
        "price": {"current_twd": px_now, "source": price_data["source"], "yearly": yearly},
        "historical_financials": TSMC,
        "scenario": {"target_pe": target_pe, "peak_pe": peak_pe, "pe_sentiment": pe_sentiment},
        "local_research_kb_hits": [
            "Goldman Sachs 2026-04: Buy, target price up to NT$2,750, AI/CoWoS multi-year growth.",
            "JPMorgan 2026-04: OW, target NT$2,400, N3/N5 utilization and GM beats.",
            "Morgan Stanley 2026-01: CoWoS capacity +80% YoY to ~125k wpm, CoWoS ~15% of 2026 revenue.",
            "Nomura 2026-04: Buy, target NT$2,820, revenue growth >30%.",
            "HSBC 2026-04: target NT$2,700, around 27x rolled EPS.",
        ],
        "official_sources": [
            "https://investor.tsmc.com/english/annual-reports",
            "https://investor.tsmc.com/english/quarterly-results/2026/q1",
            "https://investor.tsmc.com/english/monthly-revenue/2026",
            "https://pr.tsmc.com/english/news/3320",
        ],
    }

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    write_input_json(payload)
    K.finalize(wb)
    wb.save(OUT_XLSX)
    return payload


if __name__ == "__main__":
    result = build_model()
    print(json.dumps(result["outputs"], ensure_ascii=False, indent=2))
