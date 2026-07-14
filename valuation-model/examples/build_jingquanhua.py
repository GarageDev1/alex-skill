# -*- coding: utf-8 -*-
"""
build_jingquanhua.py — 京泉华(002885.SZ)端到端通电估值模型(单币种 CNY·亿; P/E 主线 + P/B 平行镜头)。

公司画像:磁性元器件 + 特种变压器 + 电源 的电子制造商。报表主体是非 AI 业务(新能源/汽车/工业/海外电网),
2025 净利仅 0.84 亿、PE_TTM 131x。投资 thesis 全在一个"期权"上:给伊顿(Eaton)SST 系统供 MFT(中频隔离变压器),
随 NVIDIA 800V/SST 数据中心供电革命放量。年报口径下 SST 还在小试,真实 AI 收入 2025≈数千万(<1.5% 营收)。

链:[① 物理锚 = 全球/美国数据中心 SST 市场(亿元)] → [② 京泉华 SST 收入 = SST TAM × 公司份额(MFT 价值量切片)]
    + [基础业务 = 周期/增速式] → [③ 总营收] → [④ 段OPM→营业利润→净利→权益→EPS/BPS] → [⑤ P/E 主线 + P/B 验] → 隐含价

三情景全在 SST 这条期权上翻档:放量时点 × 份额(MFT 切到多少 TAM)× SST 段净利率 + 基础业务增速 + 目标 P/E 情绪。

单位约定(单币种,无 FX):财务量纲 = 亿元(人民币);股本 = mn股;ps_scale=100(EPS=NI亿×100/股本mn=元/股);
mcap_div=100(市值=价×股本mn/100=亿元);fx 全部=1(占位,不做币种换算)。

run: PYTHONUTF8=1 python examples/build_jingquanhua.py
     python scripts/recalc.py out/002885_valuation_model.xlsx
     python scripts/validate_valuation.py out/002885_valuation_model.xlsx
"""
import os
from openpyxl import Workbook
import build_kit as K

# ════════════ 0. 全局轴 ════════════
ALLC = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
ALLY = ["2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
HC, HY = ["B", "C", "D", "E", "F"], ["2021", "2022", "2023", "2024", "2025"]
FC = ["F", "G", "H", "I", "J", "K"]          # 前瞻含基年 F=2025A
FCf = FC[1:]                                   # 纯前瞻 2026E-2030E
FX = 1                                          # 单币种占位
FX_HIST = [1, 1, 1, 1, 1]
CASES = ["Bear", "Base", "Bull"]
USDCNY = 7.1                                    # 把 $B 口径的 SST TAM 折成亿元(1 $B = 71 亿元)用

S_COVER, S_HIST, S_PX, S_CONS = "封面", "历史财务与估值", "股价走势", "卖方研报共识"
S_HMULT, S_MULT, S_SW = "历史估值倍数", "估值倍数假设", "情景切换"
S_ANCHOR, S_SEG, S_FUND = "数据中心SST市场", "分部测算", "利润与收入假设"
S_VAL, S_CMP, S_DASH = "情景估值", "估值对比", "综合判断仪表盘"

# 月度收盘(不复权, 本币元; 来自 quant-lake/东财, 年内高低用月末口径)
MONTHLY = [("2021-06", 16.1), ("2021-12", 26.08),
           ("2022-06", 30.0), ("2022-12", 29.82),
           ("2023-06", 25.0), ("2023-12", 18.12),
           ("2024-06", 11.0), ("2024-12", 14.25),
           ("2025-03", 13.0), ("2025-06", 16.0), ("2025-09", 22.0), ("2025-12", 25.69),
           ("2026-03", 35.0), ("2026-06", 46.87)]
PX_NOW = 46.87

wb = Workbook()
wb.remove(wb["Sheet"])

# ════════════ 1. 封面 ════════════
K.write_cover(wb.create_sheet(S_COVER), {
    "title": "京泉华(002885.SZ)估值模型 — 单币种(人民币·亿) · P/E 主线 + P/B 验",
    "span": 6,
    "meta": [
        ("报告日期", "2026-06-22"),
        ("数据截止", "2025 年报(巨潮 2026-04-25)+ 2026Q1 + 实时股价 6/22"),
        ("现价", "46.87 元(6/22 涨停)"),
        ("时效声明", "SST/MFT 量产节奏与订单变化极快;伊顿 2026 年 4-5 月验证结果、每季磁件分部增速是关键更新点,建议每季财报后重算。"),
        ("方法一句话", "物理锚 = 全球/美国数据中心 SST 市场(亿元)→ 京泉华 SST 收入 = SST 市场 × 公司份额(MFT 价值量切片);基础业务走增速/周期。段OPM 驱动净利,P/E 主线 × 前瞻 EPS → 隐含价,P/B 平行验。"),
    ],
    "takeaways": [
        ("① 当下估值位置", "PE_TTM 131x、PB 8.3x;净利绝对值仅 0.84 亿。估值完全建立在 SST/AI 期权 2027+ 兑现的预期上,而非当下基本面。"),
        ("② 核心问题答案", "AI/数据中心直接收入 2025≈数千万(<1.5% 营收),年报未单独披露;唯一硬卡位是伊顿 SST 的 MFT,但年报口径 MFT 还在小试,未规模量产。"),
        ("③ 物理锚与切片", "全球数据中心 SST 市场(Bernstein:美国 2030E $3bn 基准/$10bn blue-sky;800V 2027 Rubin Ultra 必需)。京泉华切的是 MFT 这一刀,份额由「独供伊顿+扩客户」假设驱动。"),
        ("④ 三情景隐含价", "Bear/Base/Bull 全在 SST 放量时点×份额×段净利率 + 基础增速 + 目标 P/E 情绪上翻档,沿同一条链算出,见『估值对比』。"),
        ("⑤ 主要风险", "SST 量产推迟/验证不及预期(最大,直接证伪);单一大客户 21.7%;盈利质量差(经营现金流弱、应收高);两融杠杆驱动、0 机构覆盖、情绪退潮回撤大。"),
    ],
})

# ════════════ 2. 历史财务与估值(单币种·亿)════════════
# 分部(亿,实际): 磁性元器件 / 特种变压器 / 电源。vals_in_usd=True → 直接当"亿"不做 FX 换算。
ha = K.write_history(wb.create_sheet(S_HIST), {
    "title": "京泉华 历史财务与估值(人民币·亿) — 2021-2025A + 当下(2025年报/季报实际)",
    "hist_cols": HC, "hist_years": HY,
    "vals_in_usd": True,           # 值已是"亿", 不做汇率换算
    "fx_hist": FX_HIST, "fx_now": FX,
    "unit_label": "(人民币·亿)",
    "fx_label": "FX(单币种=1)",
    "mcap_label": "市值(亿元)",
    "eps_label": "EPS(元)", "bps_label": "BPS(元)",
    "ps_scale": 100, "mcap_div": 100,
    "cur_label": "当下(TTM/最新)",
    "segments": [
        ("磁性元器件 收入", [None, None, None, 11.81, 13.88, 0.0], False),   # 2021-23 年报未单独披露此口径, 留空; 不算YoY(早年缺基数)
        ("特种变压器 收入", [None, None, None, 11.20, 12.86, 0.0], False),
        ("电源 收入", [None, None, None, 6.77, 6.45, 0.0], False),
        ("其他/早年合并 收入", [19.10, 25.84, 25.94, 0.31, 0.28, 0.0], False),  # 2021-23 总营收实际, 24/25 为其他业务
    ],
    "total_now": 33.96,    # TTM 营收 ≈ 2025(33.47)+ (Q1'26 8.42 − Q1'25 7.93)
    "gm_pct": [0.1134, 0.1466, 0.1268, 0.1250, 0.1290], "gm_now": 0.1460,
    "ni": [0.199, 1.429, 0.342, 0.382, 0.842], "ni_now": 0.842,
    "eq": [8.131, 9.476, 14.070, 14.023, 15.111], "eq_now": 15.111,
    "shares": [240, 257, 271, 271, 271], "shares_now": 271,
    "px_end": [26.08, 29.82, 18.12, 14.25, 25.69],
    "px_now": PX_NOW,
    "px_avg": [16.1, 27.7, 25.1, 12.5, 17.2],
    "band_note": "P/B 历史常态 1.5-3.5x → 2026 现价冲到 8.3x, 远超历史带上沿(题材重估)。P/E 历史 20-80x 波动大(低净利)。",
    "notes": [
        ("磁性元器件 收入", "电感/变压器/隔离变压器/滤波器等; 2024-25 年报『分产品』口径实际(P32-33)。SST 用 MFT 归在此分部。"),
        ("特种变压器 收入", "中大功率定制变压器(海外电网/数据中心配电/新能源并网); 2024-25 年报实际。海外放量主角、毛利率最低(8.18%)。"),
        ("电源 收入", "开关电源/电源模块整机; 2024-25 年报实际。毛利率最高(17.24%)但 2025 唯一下滑分部。"),
        ("其他/早年合并 收入", "2021-2023 = 当年总营收实际(年报未按现口径拆分产品, 故合并入此行避免倒挤); 2024-25 = 年报其他业务。"),
        ("HREV", "总营收 = 交易所/年报实际: 2021=19.10 / 2022=25.84 / 2023=25.94 / 2024=30.12 / 2025=33.47 亿。"),
        ("HGMP", "综合毛利率: 年报实际(2025=12.90%); 当下列=2026Q1 14.60%(改善)。"),
        ("HNI", "归母净利: 年报实际(亿); 2025=0.842(+120% YoY, 主因海外高毛利+费用控制, 非 AI)。当下=TTM。"),
        ("HEQ", "归母股东权益: 年报实际(亿); 2023 跳升因资本公积变动。"),
        ("HSH", "总股本(mn股): 2023 起 271mn, 无重大增发。"),
        ("HPX", "年末收盘(不复权, 元): 东财/quant-lake; 当下=现价 46.87(6/22 涨停)。"),
        ("HPXA", "年均价(不复权, 元): 月度收盘均值。"),
    ],
})

# ════════════ 3. 股价走势 ════════════
def phase_fn(ym):
    if ym <= "2024-12":
        return "① 周期底"
    if ym <= "2025-09":
        return "② AI电源叙事萌芽"
    if ym <= "2025-12":
        return "③ 800V白皮书催化"
    return "④ SST题材FOMO"

px = K.write_price_chart(wb.create_sheet(S_PX), MONTHLY, {
    "fn": phase_fn,
    "rows": [("① 周期底", "2024 末 14.25 元, 平庸电子制造股估值"),
             ("② AI电源叙事萌芽", "2025 市场把京泉华与『伊顿 SST MFT 独供』挂钩"),
             ("③ 800V白皮书催化", "2025/10/15 NVIDIA 800V HVDC 白皮书, 数据中心电源集体涨停"),
             ("④ SST题材FOMO", "2026 H1 一路冲高, 年内最高 55.8; 6/22 放量涨停 46.87 突破前高, 两融余额一周 +35%")],
}, title="京泉华 月度股价(不复权, 元)")

# ════════════ 4. 卖方研报共识 ════════════
K.write_consensus(wb.create_sheet(S_CONS), {
    "title": "卖方研报共识 — 京泉华无任何券商正式覆盖(0 家),本表是『市场叙事 vs 我们』对账单",
    "overview": "关键事实:京泉华无卖方研报覆盖、无 2026/2027/2028E 一致预期(东财研报接口=0 篇、同花顺无机构覆盖)。市场叙事来自散户/财经号(东财财富号)+ 互动易,不是机构定价。下表把这些『市场口径』当软信号对账,我们模型自行拍数、标注与口径差异。",
    "assumptions": [
        ("SST/AI 收入\n放量时点", "市场叙事: 伊顿 2025 下单数千万→2026 H2 批量→2027 规模放量。", "无机构验证, 全是互动易/财经号转述; 年报口径 MFT 仍在小试(10/13.8kV 待批量, 35kV 样品阶段)。", "Base 取 2026 仍是验证期(贡献小)、2027 起放量、2028-29 上量。比最乐观叙事保守一年。"),
        ("SST 业务\n2027 体量", "财经号乐观口径: 2026 营收 1.5-2 亿、2027 营收 4-5 亿、毛利率 40-50%、甚至『2027 净利 8 亿』。", "口径无出处、跳跃极大(从数千万到 4-5 亿是 10 倍跃迁), 把份额×单价×毛利全拍满。", "Base 取 2027 SST 收入 ~2.5 亿(份额温和、单价中性); Bull 才给 4-5 亿。8 亿净利口径不采。"),
        ("基础业务\n增速", "市场不关注(都盯 SST); 实际 2025 营收 +11%、磁件+17.5%、特种变压器+14.8%、电源-4.7%。", "海外放量是周期性还是结构性有分歧; 海外缺口缓解后高毛利可能回落。", "Base 基础业务增速从 2025 的 +11% 逐步 normalize 到 7-9%。"),
        ("目标 P/E", "市场用『2027 故事 EPS』反推, 说『forward PE 不算贵』。", "用未兑现的预测 EPS 算 forward PE = 循环论证; 当下 TTM PE 131x 是事实。", "目标 P/E 走情绪三层(历史峰值 × 重估溢价 × 情绪值),不照搬市场叙事。"),
    ],
    "divergences": [
        "① SST 放量到底哪年、放多大: 决定 2027-2029 净利,是估值最大不确定。年报硬数据(小试)vs 财经号叙事(4-5 亿)差一个数量级。",
        "② 基础业务海外高毛利是结构性还是周期性: 决定 normalize 后的常态净利地板(0.4-0.5 亿 vs 1 亿+)。",
    ],
    "stances": [
        "无机构覆盖 — 不存在券商评级/目标价。本模型为该标的的首次独立估值,无街上锚点可对标。",
        "市场情绪(东财财富号/互动易): 一致看多 SST 重估, 但建立在未经原文核对的订单/收入数字上。",
    ],
})

# ════════════ 5. 历史估值倍数(数据底座)════════════
hm = K.write_hist_multiples(wb.create_sheet(S_HMULT), {
    "title": "历史估值倍数 — 自身历史带 + 当下 TTM + 同业对照(电源/磁件 A股可比)",
    "intro": "数据底座: ①京泉华自己历史上 P/E·P/B 走过什么带(逐年+年内高低) ②现在市场给多少(PE 131x/PB 8.3x) ③同业(电源/磁件/数据中心电源链 A股)给多少。看完这页再去下一页拍三案目标倍数。注:京泉华是 P/E 主线(盈利是投资人定价的对象, 资产不是穿越周期的锚), P/B 仅作过热验。",
    "s_hist": S_HIST, "ha": ha, "hist_cols": HC, "hist_years": HY,
    "yhigh": px["yhigh"], "ylow": px["ylow"],
    "fwd_note": "forward P/E 取决于 SST 放量: 若 2027 净利按 Base ~2.2 亿, 现价对应 2027E forward PE ≈ 58x(仍高); 详见『情景估值』。",
    "self_name": "京泉华",
    "self_fwd_pe_label": "≈58x(2027E,Base)",
    "self_note": "本模型标的; P/E 主线。当下 TTM PE 131x 远超历史与同业, 完全 price-in SST 期权。",
    "peers": [
        {"name": "麦格米特(002851, 电源/服务器PSU)", "yearly": [None, None, None, None, None], "cur_pb": 4.5, "cur_pe": 38.0, "fwd_pe": 28.0,
         "note": "更大的电源整机玩家, GS 点名服务器 PSU『第二供应商』; 业务更成熟、估值更低。可比性高。"},
        {"name": "可立克(002782, 磁性元器件)", "yearly": [None, None, None, None, None], "cur_pb": 5.0, "cur_pe": 45.0, "fwd_pe": 30.0,
         "note": "纯磁件同业, 也沾 AI 电源/数据中心题材; TTM PE 显著低于京泉华。最直接磁件可比。"},
        {"name": "数据中心电源链 A股(题材均值)", "yearly": None, "cur_pb": None, "cur_pe": 60.0, "fwd_pe": 40.0,
         "note": "四方股份/中恒电气/科士达等题材股 TTM PE 区间上沿(参照: 题材热度档)。"},
        {"name": "A股电子元器件(申万, 参照下沿)", "yearly": None, "cur_pb": None, "cur_pe": None, "fwd_pe": 25.0,
         "note": "行业中位 forward PE(光谱下沿, 无题材溢价时该值多少)。"},
    ],
    "ratio": {"peer": "可立克(002782, 磁性元器件)",
              "note": "对账线: 京泉华 / 可立克 的 TTM PE 比值 ≈ 131/45 ≈ 2.9x —— 京泉华相对纯磁件同业有 ~3x 的题材溢价, 这就是 SST 期权 price-in 的幅度。"},
    "reading": "① 自己: TTM PE 131x 远超历史(20-80x)与同业(38-45x)→ 第一层峰值取自身历史合理上沿(~50x, 非被本轮抬高的 131x)。② 同业: 麦格米特/可立克 38-45x 是『成熟电源磁件』档, 京泉华 131x 多出的部分=SST 期权溢价。③ 京泉华/可立克 PE 比值 ~2.9x = 当前题材溢价幅度, 但这是情绪不是结构(SST 未兑现)→ 第三层情绪值现在>1(过热), 兑现路径决定能否撑住。→ 下一页: 历史峰值 PE × 重估溢价 × 情绪值(三案)。",
})

# ════════════ 6. 估值倍数假设(三案目标 P/E + 平行 P/B; 三层分解)════════════
# 本模型 P/E 主线: 三层 = 历史峰值 PE × SST 重估溢价 × 情绪值。第一层 peak/premium 复用 kit 的 pk/pr 槽(语义改成 PE)。
ma = K.write_multiple_assumptions(wb.create_sheet(S_MULT), {
    "title": "估值倍数假设 — P/E 主线三案目标倍数(= 历史峰值 P/E × SST 重估溢价 × 情绪值)+ P/B 验",
    "intro": "本页只做判断(数据底座在上一页)。京泉华 P/E 主线: 投资人定价的对象是『SST 兑现后的盈利』, 不是账面资产。三层分解出三案目标 P/E; 『情景切换』引用并切换, 『情景估值』套用当前案 × 前瞻 EPS → 隐含价; P/B 作平行镜头与过热验。",
    "why_text": ("镜头选择是业务判断: 京泉华『穿越周期持续存在、且投资人愿意资本化的东西』是什么? "
                 "——它不是商品型重资产(资产不是印钞机, 主业毛利率仅 12.7%、ROE 长期个位数), 也没有 TSMC 那种高且稳的资产回报。"
                 "它的全部投资价值在一个『盈利期权』: SST/MFT 能不能从小试走到放量、把净利从 0.84 亿推到几亿。投资人买的就是这个未来盈利, 所以 P/E 主线(资本化预期盈利)。"
                 "P/B 不做主线但必须并行验: 当前 PB 8.3x 本身就远超历史(1.5-3.5x), 是过热的硬证据——P/B 镜头负责回答『就算 SST 兑现, 账面支撑得起这个价吗』。"
                 "触发迁移: 若 SST 完全证实、成为稳定高毛利主业, 市场会更敢资本化盈利(PE 中枢上移); 若证伪, 打回『成熟电源磁件』档(PE 25-40x)。"),
    "why_rows": 5,
    "method_text": "三层分解(P/E 口径, 不硬拍): ①历史峰值 P/E(京泉华历史合理上沿, 非本轮被题材抬到的 131x) × ②SST 重估溢价(相对成熟电源磁件同业的结构溢价, 锚同业相对法) × ③情绪值(题材/情绪位置, 依据仪表盘 D 块)。一致性检验: 三层相乘应能近似复现当下实际 PE。",
    "peak": 50.0, "peak_note": "第一层=历史峰值 P/E 50x: 京泉华历史盈利高峰年(2022)前后实际 PE 上沿区间; 不用被本轮题材抬到的 131x(那含溢价+情绪, 再乘=双重计算)。",
    "premium": 1.30, "premium_note": "第二层=SST 重估溢价 1.30x: 相对成熟电源磁件同业(麦格米特/可立克 ~38-45x)的结构溢价, 反映京泉华独供伊顿 MFT 的稀缺卡位(全球仅村田可竞争)。慢变、半硬。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "hpb_row": ha["HPE"],   # P/E 主线 → 情绪值历史列用实际 P/E 反推
    "cases": [
        ("Bear", [1.05, 0.75, 0.55, 0.45, 0.40],
         "SST 验证不及预期/推迟: 2026 H1 伊顿验证未通过或推迟、扩客户落空 → 题材证伪, 情绪快速退潮。情绪值从当下>1 跌到 0.4-0.55, 目标 PE 收敛到成熟电源磁件档(50×1.3×0.45≈29x)。事件: 互动易披露验证延期 / 磁件分部增速无跳升。"),
        ("Base", [1.30, 1.10, 0.80, 0.65, 0.55],
         "SST 按 Base 路径放量(2027 收入 ~2.5 亿): 伊顿验证通过、2027 起上量但未爆发。情绪值从启动期 1.3 随可见度兑现逐步 normalize 到 0.55(2030)。目标 PE 2026≈85x→2030≈36x(随成长成熟下行)。"),
        ("Bull", [1.45, 1.45, 1.10, 0.85, 0.70],
         "SST 大放量(2027 收入 4-5 亿、扩到台达/施耐德): 市场接受『AI 算力核心部件供应商』重估叙事, 超涨延长。情绪值高位维持更久(2026-27 在 1.45), 之后才回落。"),
    ],
    "sent_note": "情绪值=题材/情绪位置(P/E 口径)。1.0=付足『历史峰值×溢价』; >1=FOMO 超涨(当下题材就是>1); <1=证伪后折价。历史列=实际 PE ÷(峰值×溢价)反推, 三案同值(一致性检验可见; 亏损/低基数年 PE 失真处会偏高, 仅作参照)。",
    "target_note": "三案目标 PE = 50x × 1.30 × 当案情绪值。历史列=实际 PE(回看)。",
    "reconcile_text": "市场叙事用『2027 故事 EPS』反推说 forward PE 不贵, 是循环论证(拿未兑现预测证明便宜)。我们的目标 PE 从历史峰值+结构溢价+情绪三层拍, 不照搬叙事; 当下情绪值>1 反映题材过热, 兑现路径(伊顿验证)决定它向 1.0 收敛还是跌破。凭什么敢和『8 亿净利』叙事不同: 年报硬数据=MFT 小试、AI 收入数千万, 4-5 亿收入需要 10 倍跃迁+多客户落地, Base 不押这个、Bull 才给。",
    "source_text": "第一层(历史峰值 PE)=京泉华自身历史倍数带(本页上游『历史估值倍数』)。第二层(重估溢价)=相对麦格米特/可立克同业相对法。第三层(情绪值)=『综合判断仪表盘』D 块(伊顿验证进度/磁件分部增速/两融杠杆)。",
})

# 平行 P/B 镜头的目标倍数: 用一组独立三案 P/B 情绪值(简化: 直接拍三案目标 P/B, 走 derived)。
# 为保持 kit 结构, P/B 目标走『情景切换』的 linked + derived。这里先准备 P/B 三案常数。
PB_CASE = {"Bear": [3.5, 3.0, 2.5, 2.2, 2.0], "Base": [5.0, 4.5, 3.8, 3.3, 3.0], "Bull": [6.5, 6.0, 5.0, 4.2, 3.6]}

# ════════════ 7. 情景切换 ════════════
sw = K.write_scenario_switch(wb.create_sheet(S_SW), {
    "title": "情景切换 — 全模型唯一情景参数库 + 切换开关(默认 Base)",
    "usage": ("怎么用: B2 下拉选案 → 案序号派生 → 各杠杆『当前案』行跟着切 → 整条明细链(SST 收入→分部→利润→倍数→估值)变档, "
              "『情景估值』输出该案逐年隐含价。三案对比不用切: 『估值对比』恒常三列并排。情景参数只在本页改(蓝字); 未列入的假设三案共用(跟 Base)。"
              "核心可翻杠杆都在 SST 期权这条线上 + 基础业务增速 + 目标 P/E/PB 情绪。"),
    "cases": CASES, "default": "Base",
    "triggers": [
        ("Bear", "落进 Bear: 伊顿 2026 H1 验证未通过/推迟, 扩客户(台达/施耐德)落空, SST 题材证伪 → SST 收入卡在数千万, 情绪退潮、目标 PE 收敛成熟档。"),
        ("Base", "落进 Base: 伊顿验证通过、2027 起 SST 上量但未爆发(2027 SST 收入 ~2.5 亿), 基础业务 normalize 增长。可见度撑住但不超涨。"),
        ("Bull", "落进 Bull: SST 大放量(2027 SST 收入 4-5 亿)、扩到多客户, 市场接受『AI 算力核心部件』重估, 超涨延长。"),
    ],
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "levers": [
        # ① 京泉华 SST 份额(%): 物理锚的切片杠杆。SST 收入 = SST 市场(锚) × 本份额 → 锚载荷, 改锚价格动。
        {"key": "sstshr", "name": "京泉华 占 SST 市场份额(%)", "fmt": K.PCT,
         "vals": {"Bear": [0.075, 0.045, 0.030, 0.027, 0.025],
                  "Base": [0.105, 0.087, 0.090, 0.110, 0.130],
                  "Bull": [0.130, 0.130, 0.165, 0.230, 0.280]},
         "desc": "京泉华切全球数据中心 SST 市场的 MFT 价值量份额。SST 收入 = SST 市场(物理锚) × 本份额。2025 实际≈0.4亿/0.3$B盘≈19%(早期小盘份额高, 失真; 看绝对额)。三案=份额扩张快慢, 对应放量时点×客户数。",
         "stories": {"Bear": "卡在伊顿小批量、验证推迟/扩客户落空, 份额随大盘扩张被稀释到 1-1.5%(SST 收入 2030 仍<1 亿)。",
                     "Base": "伊顿上量+1-2 新客户, 份额随大盘成长稳在 6-14%(SST 收入 2027≈2.5亿/2030≈8.5亿)。",
                     "Bull": "多客户+MFT 价值量提升至 1-1.2 元/W, 份额扩到 16-29%(SST 收入 2027≈4.5亿/2030≈18亿)。"},
         "hist": [None, None, None, None, 0.19]},
        # ② SST 段净利率: 高毛利定制段
        {"key": "sstm", "name": "SST 段净利率", "fmt": K.PCT,
         "vals": {"Bear": [0.10, 0.12, 0.14, 0.15, 0.15], "Base": [0.12, 0.18, 0.20, 0.20, 0.20], "Bull": [0.15, 0.22, 0.24, 0.24, 0.24]},
         "desc": "SST/MFT 是高价值量定制件, 净利率结构性高于基础业务(基础业务净利率仅 2-3%)。媒体口径 SST 毛利率 40-50% → 折算净利率 15-24%。三案=定价权强弱。",
         "stories": {"Bear": "三供成型/客户压价, SST 净利率压到 10-15%。", "Base": "定价权随竞争缓降, 净利率 12-20%。", "Bull": "独供卡位深、价值量升, 净利率 22-24%。"},
         "hist": [None, None, None, None, 0.12]},
        # ③ 基础业务增速(磁件+特种变压器+电源 ex-SST 合并)
        {"key": "basg", "name": "基础业务 增速", "fmt": K.PCT,
         "vals": {"Bear": [0.03, 0.02, 0.02, 0.02, 0.02], "Base": [0.09, 0.08, 0.07, 0.07, 0.07], "Bull": [0.12, 0.11, 0.10, 0.09, 0.09]},
         "desc": "非 SST 主业(磁件+特种变压器+电源)合并增速。2025 实际 +11%(海外放量驱动)。Bear=海外缺口缓解、高毛利回落、增长熄火; Base=normalize 到 7-9%; Bull=海外份额持续扩。",
         "stories": {"Bear": "海外大功率变压器供需缺口缓解, 高溢价回落, 主业增长熄火到 2-3%。", "Base": "从 2025 的 +11% 平滑 normalize 到 7%。", "Bull": "海外份额持续扩张, 维持双位数。"},
         "hist": [0.20, 0.35, 0.004, 0.16, 0.11]},
        # ④ 基础业务净利率
        {"key": "basm", "name": "基础业务 净利率", "fmt": K.PCT,
         "vals": {"Bear": [0.022, 0.020, 0.018, 0.018, 0.018], "Base": [0.026, 0.028, 0.030, 0.030, 0.030], "Bull": [0.030, 0.033, 0.035, 0.036, 0.036]},
         "desc": "基础业务净利率(剔除 SST)。历史综合净利率 1-5.5%(2025=2.5%)。改善靠海外高毛利占比升+费用控制; Bear=海外溢价回落压回 1.8-2.2%。",
         "stories": {"Bear": "海外高毛利回落, 净利率压回历史中枢 1.8-2.2%。", "Base": "结构改善延续, 2.6-3.0%。", "Bull": "海外占比持续升, 3.0-3.6%。"},
         "hist": [0.010, 0.055, 0.013, 0.013, 0.025]},
    ],
    "linked": [
        {"key": "sent", "name": "P/E 情绪值(主线倍数第三层)", "fmt": K.N2,
         "src_sheet": S_MULT, "src_row0": ma["sent_row0"],
         "note": "三案取值与依据见『估值倍数假设』(完整三层方法论在那页); 本页只做切换。"},
    ],
})
# derived: 目标 P/E(当前案) = 历史峰值 × 重估溢价 × 当前案情绪 → 喂『情景估值』主线
_pk = f"'{S_MULT}'!{ma['pk_cell']}"
_pr = f"'{S_MULT}'!{ma['pr_cell']}"
_sent_act = sw["SWACT"]["sent"]
_r = sw["next_row"]
K.lab(wb[S_SW], f"A{_r}", "目标 P/E(当前案)", b=True)
for _c in ALLC:
    K.fml(wb[S_SW], f"{_c}{_r}", f"={_pk}*{_pr}*{_c}{_sent_act}", K.MX, link=True)
K.logic(wb[S_SW], f"L{_r}", "= 历史峰值 P/E × SST 重估溢价 × 当前案情绪值 → 喂『情景估值』主线前瞻 P/E。")
SWPE = _r
_r += 1
# 平行 P/B 目标(当前案): 三案 P/B 常数行 + INDEX 当前案
K.lab(wb[S_SW], f"A{_r}", "目标 P/B 三案 (Bear/Base/Bull)", b=True)
for _c in ["A"] + list(ALLC):
    wb[S_SW][f"{_c}{_r}"].fill = K.GREYF
K.logic(wb[S_SW], f"L{_r}", "平行 P/B 镜头的三案目标倍数(过热验): 当前 PB 8.3x 远超历史 1.5-3.5x, P/B 镜头回答『账面撑不撑得起』。")
PB3_0 = _r + 1
for _i, _cs in enumerate(CASES):
    _rr = _r + 1 + _i
    K.lab(wb[S_SW], f"A{_rr}", f"  {_cs}")
    K.introw(wb[S_SW], _rr, FCf, PB_CASE[_cs], None, K.MX)
PB_ACT = _r + 1 + len(CASES)
K.lab(wb[S_SW], f"A{PB_ACT}", "  目标 P/B(当前案)", note=True)
for _c in FCf:
    K.fml(wb[S_SW], f"{_c}{PB_ACT}", f"=INDEX({_c}{PB3_0}:{_c}{PB3_0+len(CASES)-1},{sw['idx_cell']})", K.MX)
SWPB = PB_ACT

# ════════════ 8. 物理锚: 数据中心 SST 市场 ════════════
anchor = K.write_anchor(wb.create_sheet(S_ANCHOR), {
    "title": "全球数据中心 SST(固态变压器)市场 — 需求物理盘子(亿元)",
    "all_cols": ALLC, "all_years": ALLY,
    "series": [
        ("全球数据中心 SST 市场 ($B)",
         [None, None, None, None, 0.3, 0.6, 1.2, 2.0, 2.6, 3.2],
         "口径: 美国数据中心 SST TAM(Bernstein 基准 2030E ~$3bn, 2025-30 渗透爬坡); 800V 2027 Rubin Ultra 首个必需节点。全球略大于美国, 此处取美国基准近似全球可及。", K.N1),
        ("全球数据中心 SST 市场(亿元)",
         [None, None, None, None, None, None, None, None, None, None],
         "= 上一行 $B × 7.1(USDCNY)。这是京泉华 SST 收入的物理盘子。", K.N0),
    ],
    "source_note": "来源: Bernstein《AI Data Centers to unlock Solid-State Transformation》(2026-03-25): 美国数据中心 SST TAM 2030E 基准 $3bn(对应 2025-30 新增 17.5% CAGR + 2030 800V 渗透 45%), blue-sky 近 $10bn。BofA: SST 整机/半导体 TAM CY30 $500mn-$1bn(口径更窄)。NVIDIA 800V 路线图: 2026 试点/2027 Rubin Ultra 必需/2028-30 工业化。",
    "role_note": "作用: 京泉华 SST 收入 = 本盘子 × 京泉华 MFT 价值量份额(见情景切换 sst 行的三案路径)。改盘子或份额 → SST 收入 → 净利 → 估值全链动。注: 本模型 SST 收入直接在情景切换三案拍(已隐含『盘子×份额』判断), 本锚页提供份额对账的分母。",
})
SST_USD_ROW = anchor["row_of"]["全球数据中心 SST 市场 ($B)"]
SST_CNY_ROW = anchor["row_of"]["全球数据中心 SST 市场(亿元)"]
for _c in ALLC:
    K.fml(wb[S_ANCHOR], f"{_c}{SST_CNY_ROW}", f"={_c}{SST_USD_ROW}*{USDCNY}", K.N0, link=False)

# ════════════ 9. 分部测算 ════════════
seg = K.write_segment_model(wb.create_sheet(S_SEG), {
    "title": "分部测算 — SST 段(物理锚切片) + 基础业务段(增速) + 份额对账",
    "all_cols": ALLC, "all_years": ALLY, "logic_col": "N",
    "groups": [
        ("数据中心 SST 市场(物理锚, 引自上页)", [
            ("SST 市场(亿元)", None, K.N0, "= 引自『数据中心SST市场』物理锚(亿元行)。京泉华 SST 收入的分母。"),
        ]),
        ("SST 段 = 市场 × 京泉华份额", [
            ("SST/AI 收入(亿元)", None, K.N1, "历史=年报小试口径(2025≈0.4 亿); 前瞻=『情景切换』当前案 SST 收入行。这是 thesis 的命门。"),
            ("京泉华 SST 份额(%)", None, K.PCT, "= SST 收入 ÷ SST 市场(对账行, 闭环检查): Base 路径下份额从~1% 缓升到~2.5%。份额>5% 需警惕(MFT 只是 SST 的一个部件, 不可能拿走整个 SST 盘子)。"),
        ]),
        ("基础业务段 = 上年 ×(1+增速)", [
            ("基础业务 收入(亿元)", None, K.N1, "磁件+特种变压器+电源(剔除 SST)。历史=总营收−SST; 前瞻=上年×(1+『情景切换』基础增速)。"),
        ]),
        ("闭环检查", [
            ("SST 占总营收(%)", None, K.PCT, "闭环: thesis=AI 重估, SST 占比应逐年升。Base 从 2025 的~1% 升到 2030 的~17%(若兑现)。"),
        ]),
    ],
})
m = seg["m"]
# SST 收入 = SST 市场(物理锚) × 京泉华份额(情景切换当前案) → 锚载荷(改锚, SST 收入→净利→隐含价全动)
HREV_ROW = ha["HREV"]
for col in ALLC:
    # SST 市场(亿元) = 引物理锚
    K.fml(wb[S_SEG], f"{col}{m['SST 市场(亿元)']}", f"={K.R(S_ANCHOR, col + str(SST_CNY_ROW))}", K.N0, link=True)
    # 份额 = 情景切换当前案(前瞻) / 2025=当前案历史列 / 2021-2024=0(无 SST)
    if col in HC and col != "F":
        K.inp(wb[S_SEG], f"{col}{m['京泉华 SST 份额(%)']}", 0.0, None, K.PCT)
    else:
        K.fml(wb[S_SEG], f"{col}{m['京泉华 SST 份额(%)']}", f"={K.R(S_SW, col + str(sw['SWACT']['sstshr']))}", K.PCT, link=True)
    # SST 收入 = 市场 × 份额(锚通电的关键一步)
    K.fml(wb[S_SEG], f"{col}{m['SST/AI 收入(亿元)']}", f"={col}{m['SST 市场(亿元)']}*{col}{m['京泉华 SST 份额(%)']}", K.N1)
# 基础业务: 历史 = 总营收 − SST; 前瞻 = 上年 ×(1+增速)
for col in HC:
    K.fml(wb[S_SEG], f"{col}{m['基础业务 收入(亿元)']}", f"={K.R(S_HIST, col + str(HREV_ROW))}-{col}{m['SST/AI 收入(亿元)']}", K.N1, link=True)
_prevs = [HC[-1]] + list(FCf[:-1])
for _p, col in zip(_prevs, FCf):
    K.fml(wb[S_SEG], f"{col}{m['基础业务 收入(亿元)']}", f"={_p}{m['基础业务 收入(亿元)']}*(1+{K.R(S_SW, col + str(sw['SWACT']['basg']))})", K.N1, link=True)
# 占比(全列)
for col in ALLC:
    K.fml(wb[S_SEG], f"{col}{m['SST 占总营收(%)']}", f"={col}{m['SST/AI 收入(亿元)']}/({col}{m['SST/AI 收入(亿元)']}+{col}{m['基础业务 收入(亿元)']})", K.PCT)
for col in FCf:
    wb[S_SEG][f"{col}{m['SST/AI 收入(亿元)']}"].fill = K.OUT

# ════════════ 10. 利润与收入假设(手工构建: 两段净利率驱动)════════════
# kit 的 write_fundamentals 走『段OPM→OP×净利转换』; 京泉华用『段收入×段净利率』直驱两段, 手工铺更直接。
fsh = wb.create_sheet(S_FUND)
K.hdr(fsh, 1, "利润与收入假设 — 两段净利率驱动(SST 高利润段 + 基础业务段)→ 净利 → 权益 → EPS/BPS/ROE", len(ALLC) + 1)
fsh["A2"] = "假设"; fsh["A2"].font = K.BF
for _c, _y in zip(ALLC, ALLY):
    fsh[f"{_c}2"] = _y; fsh[f"{_c}2"].font = K.BF; fsh[f"{_c}2"].fill = K.CH
fsh["N2"] = "逻辑/来源(整句)"; fsh["N2"].font = K.BF; fsh["N2"].fill = K.CH
_fr = 3
fam = {}
# 分部营收(引分部测算)
K.band(fsh, _fr, "分部营收(引『分部测算』)", len(ALLC) + 1); _fr += 1
FSEG_SST = _fr
K.lab(fsh, f"A{_fr}", "SST/AI 收入(亿元)")
for _c in ALLC:
    K.fml(fsh, f"{_c}{_fr}", f"={K.R(S_SEG, _c + str(m['SST/AI 收入(亿元)']))}", K.N1, link=True)
K.logic(fsh, f"N{_fr}", "引自『分部测算』SST 段。"); _fr += 1
FSEG_BAS = _fr
K.lab(fsh, f"A{_fr}", "基础业务 收入(亿元)")
for _c in ALLC:
    K.fml(fsh, f"{_c}{_fr}", f"={K.R(S_SEG, _c + str(m['基础业务 收入(亿元)']))}", K.N1, link=True)
K.logic(fsh, f"N{_fr}", "引自『分部测算』基础业务段。"); _fr += 1
FREV = _fr
K.lab(fsh, f"A{_fr}", "总营收(亿元)", b=True)
for _c in ALLC:
    K.fml(fsh, f"{_c}{_fr}", f"={_c}{FSEG_SST}+{_c}{FSEG_BAS}", K.N1)
fsh[f"A{_fr}"].border = K.BORD; _fr += 1
K.lab(fsh, f"A{_fr}", "  SST 占总营收%", note=True)
for _c in ALLC:
    K.fml(fsh, f"{_c}{_fr}", f"={_c}{FSEG_SST}/{_c}{FREV}", K.PCT)
_fr += 1
# 段净利率假设
K.band(fsh, _fr, "段净利率假设(历史实际锚 + 前瞻引『情景切换』当前案)", len(ALLC) + 1); _fr += 1
FSSTM = _fr
K.lab(fsh, f"A{_fr}", "SST 段净利率")
for _c in HC:
    if _c == "F":
        K.inp(fsh, f"{_c}{_fr}", 0.12, None, K.PCT)
    else:
        K.lab(fsh, f"{_c}{_fr}", "n.m.", note=True)
for _c in FCf:
    K.fml(fsh, f"{_c}{_fr}", f"={K.R(S_SW, _c + str(sw['SWACT']['sstm']))}", K.PCT, link=True)
K.logic(fsh, f"N{_fr}", "SST/MFT 高价值量定制件净利率; 2025≈12%(小批量); 前瞻=『情景切换』当前案。"); _fr += 1
FBASM = _fr
K.lab(fsh, f"A{_fr}", "基础业务 净利率")
K.introw(fsh, _fr, HC, [0.010, 0.055, 0.013, 0.013, 0.025], None, K.PCT)
for _c in FCf:
    K.fml(fsh, f"{_c}{_fr}", f"={K.R(S_SW, _c + str(sw['SWACT']['basm']))}", K.PCT, link=True)
K.logic(fsh, f"N{_fr}", "剔除 SST 的主业综合净利率; 历史实际(2025=2.5%); 前瞻=『情景切换』当前案。"); _fr += 1
# 净利 / 权益 / 每股
K.band(fsh, _fr, "盈利与账面(净利=两段相加; 历史=年报实际)", len(ALLC) + 1); _fr += 1
FNI = _fr
K.lab(fsh, f"A{_fr}", "净利润(亿元)", b=True)
for _c in HC:
    K.fml(fsh, f"{_c}{_fr}", f"={K.R(S_HIST, _c + str(ha['HNI']))}", K.N2, link=True)   # 历史=年报实际
for _c in FCf:
    K.fml(fsh, f"{_c}{_fr}", f"={_c}{FSEG_SST}*{_c}{FSSTM}+{_c}{FSEG_BAS}*{_c}{FBASM}", K.N2)
K.logic(fsh, f"N{_fr}", "历史=年报实际归母净利; 前瞻=SST 收入×SST净利率 + 基础业务收入×基础净利率(两段)。"); _fr += 1
K.lab(fsh, f"A{_fr}", "  净利率", note=True)
for _c in ALLC:
    K.fml(fsh, f"{_c}{_fr}", f"={_c}{FNI}/{_c}{FREV}", K.PCT)
_fr += 1
FRET = _fr
K.lab(fsh, f"A{_fr}", "留存率")
K.introw(fsh, _fr, ALLC, [0.95, 0.70, 0.90, 0.90, 0.90, 0.90, 0.90, 0.90, 0.90, 0.90], None, K.PCT)
K.logic(fsh, f"N{_fr}", "留存率=1−派息率; 京泉华低分红(扩产/补流优先), 留存约 90%。"); _fr += 1
FEQ = _fr
K.lab(fsh, f"A{_fr}", "期末权益(亿元)")
for _c in HC:
    K.fml(fsh, f"{_c}{_fr}", f"={K.R(S_HIST, _c + str(ha['HEQ']))}", K.N1, link=True)
_eqprev = [HC[-1]] + list(FCf[:-1])
for _p, _c in zip(_eqprev, FCf):
    K.fml(fsh, f"{_c}{_fr}", f"={_p}{FEQ}+{_c}{FNI}*{_c}{FRET}", K.N1)
K.logic(fsh, f"N{_fr}", "历史=年报实际权益; 前瞻=上年+当年净利×留存率。"); _fr += 1
SH_FIX = K.R(S_HIST, f"$F${ha['HSH']}")
FEPS = _fr
K.lab(fsh, f"A{_fr}", "EPS(元)")
for _c in HC:
    K.fml(fsh, f"{_c}{_fr}", f"={_c}{FNI}*100/{K.R(S_HIST, _c + str(ha['HSH']))}", K.N2)
for _c in FCf:
    K.fml(fsh, f"{_c}{_fr}", f"={_c}{FNI}*100/{SH_FIX}", K.N2)
K.logic(fsh, f"N{_fr}", "= 净利(亿)×100 / 总股本(mn股); 单币种, 元/股。"); _fr += 1
FBPS = _fr
K.lab(fsh, f"A{_fr}", "BPS(元)")
for _c in HC:
    K.fml(fsh, f"{_c}{_fr}", f"={_c}{FEQ}*100/{K.R(S_HIST, _c + str(ha['HSH']))}", K.N2)
for _c in FCf:
    K.fml(fsh, f"{_c}{_fr}", f"={_c}{FEQ}*100/{SH_FIX}", K.N2)
K.logic(fsh, f"N{_fr}", "= 权益(亿)×100 / 总股本(mn股)。"); _fr += 1
FROE = _fr
K.lab(fsh, f"A{_fr}", "ROE", note=True)
for _i, _c in enumerate(ALLC):
    _f = (f"={_c}{FNI}/{_c}{FEQ}" if _i == 0 else f"={_c}{FNI}/AVERAGE({ALLC[_i-1]}{FEQ},{_c}{FEQ})")
    K.fml(fsh, f"{_c}{_fr}", _f, K.PCT)
_fr += 1
K.band(fsh, _fr, "基本面口径说明", len(ALLC) + 1); _fr += 1
K.mtext(fsh, _fr, "净利 = SST 收入×SST净利率 + 基础业务收入×基础净利率(两段, 单币种亿元)。历史列净利取年报实际归母(与两段近似的小差异不影响前瞻); 前瞻沿两段链算。下游『情景估值』直接引本表 EPS/BPS。", "N", 2)
K.set_widths(fsh, 22, ALLC, 9, logic_col="N", logic_width=56)
fr = {"am": {"留存率": FRET}, "REV": FREV, "NI": FNI, "EQ": FEQ, "EPS": FEPS, "BPS": FBPS, "ROE": FROE,
      "SST_REV": FSEG_SST, "BAS_REV": FSEG_BAS}

# ════════════ 11. 情景估值(P/E 主线 + P/B 平行镜头)════════════
# 手工构建 P/E 主线 + P/B 平行镜头(kit 的 write_scenario_valuation 主线锁 P/B, 京泉华是 P/E 主线)
vsh = wb.create_sheet(S_VAL)
K.hdr(vsh, 1, "情景估值 — 当前案逐年隐含价(P/E 主线 + P/B 平行镜头)", 11)
K.lab(vsh, "L1", "当前情景→", note=True)
K.fml(vsh, "M1", f"={K.R(S_SW, 'B2')}", K.N0, link=True); vsh["M1"].fill = K.CUR
_vr = K.mtext(vsh, 2, "本表输出=『情景切换』当前案(默认 Base)。主线: 目标 P/E(当前案)× 前瞻 EPS → 隐含价; 平行: 目标 P/B × 前瞻 BPS。历史列=实际年末价反推(事实); 前瞻=预测、不拟合现价。三案并排见『估值对比』。", "K", 1)
vsh[f"A{_vr}"] = "(元/股; 市值亿元)"; vsh[f"A{_vr}"].font = K.BF
for _c, _y in zip(ALLC, ALLY):
    vsh[f"{_c}{_vr}"] = _y; vsh[f"{_c}{_vr}"].font = K.BF; vsh[f"{_c}{_vr}"].fill = K.CH
_vr += 1
EPSr = lambda c: K.R(S_FUND, c + str(fr["EPS"]))
BPSr = lambda c: K.R(S_FUND, c + str(fr["BPS"]))
NIr = lambda c: K.R(S_FUND, c + str(fr["NI"]))
HPXr = lambda c: K.R(S_HIST, c + str(ha["HPX"]))
# 主线 P/E
K.band(vsh, _vr, "P/E 主线(历史=实际年末价反推; 前瞻=目标 P/E(当前案)× 前瞻 EPS → 隐含价)", 11); _vr += 1
VPE = _vr
K.lab(vsh, f"A{_vr}", "目标 P/E(历史=实际/前瞻=当前案)")
for _c in HC:
    K.fml(vsh, f"{_c}{_vr}", f'=IF({NIr(_c)}<=0,"N/M",{HPXr(_c)}/{EPSr(_c)})', K.MX, link=True)
for _c in FCf:
    K.fml(vsh, f"{_c}{_vr}", f"={K.R(S_SW, _c + str(SWPE))}", K.MX, link=True)
_vr += 1
VPEPX = _vr
K.lab(vsh, f"A{_vr}", "隐含股价 P/E 主线(元)", b=True); vsh[f"A{_vr}"].fill = K.OUT
for _c in FCf:
    K.fml(vsh, f"{_c}{_vr}", f"={_c}{VPE}*{EPSr(_c)}", K.PX)
_vr += 1
# 平行 P/B
K.band(vsh, _vr, "P/B 平行镜头(过热验; 历史=实际年末价反推; 前瞻=目标 P/B × 前瞻 BPS)", 11); _vr += 1
VPB = _vr
K.lab(vsh, f"A{_vr}", "目标 P/B(历史=实际/前瞻=当前案)")
for _c in HC:
    K.fml(vsh, f"{_c}{_vr}", f"={HPXr(_c)}/{BPSr(_c)}", K.MX, link=True)
for _c in FCf:
    K.fml(vsh, f"{_c}{_vr}", f"={K.R(S_SW, _c + str(SWPB))}", K.MX, link=True)
_vr += 1
VPBPX = _vr
K.lab(vsh, f"A{_vr}", "隐含股价 P/B 镜头(元)", b=True); vsh[f"A{_vr}"].fill = K.OUT
for _c in FCf:
    K.fml(vsh, f"{_c}{_vr}", f"={_c}{VPB}*{BPSr(_c)}", K.PX)
_vr += 1
# 当下 forward PE 诊断
VFPE = _vr
K.lab(vsh, f"A{_vr}", "当下 forward P/E = 现价 ÷ 该年 EPS")
for _c in FCf:
    K.fml(vsh, f"{_c}{_vr}", f"={K.R(S_HIST, 'G' + str(ha['HPX']))}/{EPSr(_c)}", K.MX, link=True)
_vr += 1
K.mtext(vsh, _vr, "读法: 当下 forward P/E=现价 46.87÷各年模型 EPS(看以今日价在为哪年盈利付几倍)。Base 2027E EPS 若≈0.8 元, 对应 forward PE≈58x, 仍高于成熟电源磁件档(38-45x)。P/E 主线给隐含价, P/B 镜头验账面: 两镜头隐含价差距=盈利兑现 vs 账面支撑的张力。", "K", 2); _vr += 2
# 代表股价 + 市值 + vs 现价
K.band(vsh, _vr, "代表股价 + 市值(主线 P/E)", 11); _vr += 1
VNOW = _vr
K.lab(vsh, f"A{_vr}", "当下股价(元)", b=True)
K.fml(vsh, f"G{_vr}", f"={K.R(S_HIST, 'G' + str(ha['HPX']))}", K.PX, link=True); vsh[f"G{_vr}"].fill = K.CUR
_vr += 1
K.lab(vsh, f"A{_vr}", "年末股价(元, 历史)")
for _c, _y in zip(HC, HY):
    if _y in px["yend"]:
        K.inp(vsh, f"{_c}{_vr}", px["yend"][_y], None, K.PX)
_vr += 1
VMC = _vr
K.lab(vsh, f"A{_vr}", "市值 前瞻·主线(亿元)")
for _c in FCf:
    K.fml(vsh, f"{_c}{_vr}", f"={_c}{VPEPX}*{SH_FIX}/100", K.N0)
_vr += 1
VUP = _vr
K.lab(vsh, f"A{_vr}", "前瞻隐含价(主线) vs 现价 46.87", b=True)
for _c in FCf:
    K.fml(vsh, f"{_c}{_vr}", f"={_c}{VPEPX}/$G${VNOW}-1", K.PCT)
_vr += 2
K.band(vsh, _vr, "方法与结论", 11); _vr += 1
_vr = K.mtext(vsh, _vr, "方法: 整体公司、P/E 主线逐年估(京泉华盈利是投资人定价对象, 见『估值倍数假设』镜头判断)。基本面在『利润与收入假设』; 目标 P/E 在『估值倍数假设』(三层); 本表只做最后一步: 目标倍数 × 前瞻每股 → 隐含价 + 市值。P/B 作平行过热验。", "K", 2)
K.mtext(vsh, _vr, "结论(方向性): 三情景见『估值对比』。核心 risk-reward: 现价 46.87 已 price-in SST 大幅兑现; Base 2027E 隐含价上行有限甚至倒挂; Bull 才有显著上行; Bear 下行大。衰减信号(伊顿验证延期/磁件增速无跳升)触发转 Bear 重看。", "K", 2)
vsh.column_dimensions["A"].width = 24
for _c in ALLC:
    vsh.column_dimensions[_c].width = 10
sv = {"pe_px": VPEPX, "pb_px": VPBPX, "vs_now": VUP}

# ════════════ 12. 估值对比(三案并排; P/E 主线; 防污染)════════════
SWB = sw["SWB"]
SH_F = K.R(S_HIST, f"$F${ha['HSH']}")
PX_NOW_REF = K.R(S_HIST, f"G{ha['HPX']}")
_ret = fr["am"]["留存率"]
EQ_F = K.R(S_HIST, f"F{ha['HEQ']}")   # 2025 基年权益


def _fwdprev(j, A, key):
    return (HC[-1] if j == 0 else FCf[j - 1]) + str(A[key])


cmp_rows = [
    {"key": "sstmkt", "label": "SST 市场(亿元)", "fmt": K.N0,
     "hist": lambda c, ci, A: f"={K.R(S_ANCHOR, c + str(SST_CNY_ROW))}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_ANCHOR, c + str(SST_CNY_ROW))}"},
    {"key": "sst", "label": "SST/AI 收入(亿元)", "fmt": K.N1,
     "hist": lambda c, ci, A: (f"={c}{A['sstmkt']}*{K.R(S_SW, c + str(SWB['sstshr'] + ci))}" if c == "F" else ("inp", 0.0)),
     "fwd": lambda c, j, ci, A: f"={c}{A['sstmkt']}*{K.R(S_SW, c + str(SWB['sstshr'] + ci))}"},
    {"key": "bas", "label": "基础业务 收入(亿元)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HREV']))}-{c}{A['sst']}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'bas')}*(1+{K.R(S_SW, c + str(SWB['basg'] + ci))})"},
    {"key": "rev", "label": "总收入(亿元)", "fmt": K.N1, "bold": True,
     "hist": lambda c, ci, A: f"={c}{A['sst']}+{c}{A['bas']}",
     "fwd": lambda c, j, ci, A: f"={c}{A['sst']}+{c}{A['bas']}"},
    {"key": "ni", "label": "净利(亿元)", "fmt": K.N2, "bold": True,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HNI']))}",
     "fwd": lambda c, j, ci, A: (f"={c}{A['sst']}*{K.R(S_SW, c + str(SWB['sstm'] + ci))}"
                                 f"+{c}{A['bas']}*{K.R(S_SW, c + str(SWB['basm'] + ci))}")},
    {"key": "eps", "label": "EPS(元)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={c}{A['ni']}*100/{K.R(S_HIST, c + str(ha['HSH']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['ni']}*100/{SH_F}"},
    {"key": "eq", "label": "期末权益(亿元)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HEQ']))}",
     "fwd": lambda c, j, ci, A: (f"={EQ_F}+{c}{A['ni']}*{K.R(S_FUND, c + str(_ret))}" if j == 0
                                 else f"={FCf[j-1]}{A['eq']}+{c}{A['ni']}*{K.R(S_FUND, c + str(_ret))}")},
    {"key": "bps", "label": "BPS(元)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={c}{A['eq']}*100/{K.R(S_HIST, c + str(ha['HSH']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['eq']}*100/{SH_F}"},
    {"key": "sent", "label": "P/E 情绪值(该案; 历史=实际反推)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_MULT, c + str(ma['sent_row0'] + ci))}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_MULT, c + str(ma['sent_row0'] + ci))}"},
    {"key": "pe", "label": "目标 P/E(历史=实际)", "fmt": K.MX,
     "hist": lambda c, ci, A: f'=IF({c}{A["ni"]}<=0,"N/M",{K.R(S_HIST, c + str(ha["HPX"]))}/{c}{A["eps"]})',
     "fwd": lambda c, j, ci, A: f"={_pk}*{_pr}*{c}{A['sent']}"},
    {"key": "px", "label": "隐含价(元, P/E 主线)", "fmt": K.PX, "bold": True, "out": True,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HPX']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['pe']}*{c}{A['eps']}"},
    {"key": "pbpx", "label": "隐含价(元, P/B 验)", "fmt": K.PX,
     "hist": lambda c, ci, A: None,
     "fwd": lambda c, j, ci, A: f"={K.R(S_SW, c + str(PB3_0 + ci))}*{c}{A['bps']}"},
    {"key": "up", "label": "历史: vs 实际年末价(≈0) / 前瞻: vs 现价", "fmt": K.PCT,
     "hist": lambda c, ci, A: f"={c}{A['px']}/{K.R(S_HIST, c + str(ha['HPX']))}-1",
     "fwd": lambda c, j, ci, A: f"={c}{A['px']}/{PX_NOW_REF}-1"},
]
cm = K.write_comparison(wb.create_sheet(S_CMP), {
    "title": "估值对比 — Bear / Base / Bull 三情景目标价并排(P/E 主线)",
    "intro": ("三个情景各自完整推演: SST 市场(物理锚)→ SST 收入(份额切片)+ 基础业务 → 净利(两段净利率)→ EPS → 目标 P/E → 隐含价。"
              "本表三案永远并排, 不随开关变; 调假设去『情景切换』。未列入矩阵的假设三案共用 Base。"
              "历史列 2021-2025 隐含价=实际年末价(P/E 主线历史列直接取实际价, 回测行≈0)。目标年=2027E(SST 放量首年)。"),
    "case_names": CASES,
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "block_start": 16,
    "rows": cmp_rows,
    "summary": {
        "band": "三案汇总(目标年 2027E; 各案触发条件见『情景切换』)",
        "target_col": "H",
        "rows": [
            ("净利(亿元)", "ni", K.N2, "= SST 收入×SST净利率 + 基础业务收入×基础净利率", False),
            ("总收入(亿元)", "rev", K.N1, "= SST 段(市场×份额) + 基础业务段(该案增速)", False),
            ("EPS(元)", "eps", K.N2, "= 净利×100 / 总股本(271mn)", False),
            ("目标 P/E", "pe", K.MX, "= 历史峰值 50x × 重估溢价 1.30 × 该案情绪值", False),
            ("隐含价(元)", "px", K.PX, "= 目标 P/E × 2027E EPS", True),
            ("vs 现价 46.87", "up", K.PCT, "对照现价上行/下行空间", True),
        ],
        "mcap": {"label": "隐含市值(亿元)", "key": "px", "expr": f"*{SH_F}/100",
                 "note": "= 隐含价 × 股本(271mn) / 100"},
        "concl": "风险收益比(2027E vs 现价 46.87): Bear 下行大(SST 证伪打回成熟电源磁件档)/ Base 上行有限或倒挂(现价已 price-in)/ Bull 才有显著上行(SST 大放量)。偏空——现价隐含的兑现度高于 Base 路径。衰减信号触发应转 Bear 重看。",
    },
})

# ════════════ 13. 综合判断仪表盘 ════════════
EPS27 = K.R(S_FUND, "H" + str(fr["EPS"]))   # 2027E EPS (H列)
PXD = K.R(S_HIST, "G" + str(ha["HPX"]))
dash = K.write_dashboard(wb.create_sheet(S_DASH), {
    "title": "综合判断仪表盘 — A 基本面拐点 · B 估值错位(预测引擎) · C 催化剂 · D 情绪确认",
    "usage": ("怎么用: 预测引擎是 B(错位)+ C(催化剂)。京泉华特殊性: 基本面拐点(A)尚未发生(SST 还在小试), 当下纯靠 C(伊顿验证) 兑现预期 + D(题材情绪)。"
              "验收=回测: 放回 2025/10(800V 白皮书)拐点, 这套表当时能看到题材启动(C 催化剂兑现中 + D 情绪刚启动)。"),
    "blocks": [
        {"title": "A. 基本面拐点 — 业务在结构性变好吗?", "rows": [
            ("SST 占营收逐年升?", "Base 路径: 2025~1% → 2027~10% → 2030~17%(若兑现)", "目前还是『预期』而非『已发生』: 年报口径 MFT 小试, 占比仍<1.5%。"),
            ("盈利质量", "净利 0.84 亿、经营现金流弱、应收 10.7 亿", "2025 净利 +120% 主因海外高毛利(周期性), 非 AI 结构升级。"),
            ("A 判断", "【弱】", "基本面拐点尚未发生, SST 仍是期权; A 块当下不支持高估值, 全靠 C/D。", True),
        ]},
        {"title": "B. 估值错位(预测引擎 ★)— 市场现在给的 vs 基本面该给的 → GAP", "rows": [
            ("市场现在给(TTM P/E)", {"fml": f"={PXD}/({K.R(S_HIST,'G'+str(ha['HEPS']))})", "fmt": K.MX, "fill": True},
             "= 现价 ÷ TTM EPS(0.31 元)= 131x(公式算, 随模型走)。"),
            ("基本面该给(justified)", {"inp": 45.0, "fmt": K.MX},
             "成熟电源磁件同业 PE 38-45x(麦格米特/可立克)。若 SST 不兑现, 京泉华该给这个档。"),
            ("错位 GAP = 该给÷市场给 − 1",
             {"fml": lambda ro: f"=B{ro['基本面该给(justified)']}/B{ro['市场现在给(TTM P/E)']}-1", "fmt": K.PCT},
             "GAP 负且大(~ -66%)= 当下 TTM 倍数已远超基本面该给, 进入纯情绪/期权定价区(透支)。重估空间为负, 除非 SST 把 EPS 做大。"),
            ("回测: 2025/10 拐点读数", "800V 白皮书→题材启动", "当时市场刚开始 price-in SST, GAP 尚未拉到极端, 这是那波启动的依据。"),
        ]},
        {"title": "C. 催化剂 — 什么会逼市场闭合 GAP(把 EPS 做大)", "rows": [
            ("伊顿 2026 H1 验证", "待(2026 年 4-5 月)", "★最关键: 验证通过→SST 放量路径成立(向 Base/Bull); 推迟/失败→证伪(向 Bear)。"),
            ("磁件分部增速跳升", "待(2026 半年报/三季报)", "量产信号: 磁件分部增速若突然跳升=SST 上量的财报印证。"),
            ("新增 SST 客户批量", "待(台达/维谛/施耐德送样中)", "扩客户=从单一伊顿到多客户, 决定 Bull 能否成立。"),
            ("C 判断", "全待兑现", "当下催化剂全部『待』, 估值靠预期透支; 兑现与否未来 1-4 季见分晓。", True),
        ]},
        {"title": "D. 情绪确认 — 只做 timing + 刹车(定性档位)", "rows": [
            ("量价温度计", "两融余额一周 +35% 到 4.05 亿; 6/22 放量涨停; 股东户数 -5.25%", "杠杆资金驱动、0 机构覆盖、0 龙虎榜——典型情绪/题材, 脆弱。"),
            ("现价倍数 vs 基本面该给", "TTM 131x vs 该给 45x", "市场已付出远超基本面该给 = 深度情绪定价区。"),
            ("当前档位", "【过热】", "题材 FOMO 阶段, 情绪值>1; 需 C 催化剂兑现才能从『过热』转为『结构』, 否则退潮。", True),
            ("衰减扳机", "3 条", "①伊顿验证延期/失败 ②磁件增速无跳升 ③两融见顶回落——任一翻=情绪值下调、转 Bear。"),
        ]},
    ],
    "final": {"band": "★ 综合判断(A+B+C+D 收成一句可执行的话)",
              "text": "A 弱(基本面拐点未到)+ B 透支(TTM 131x vs 该给 45x, GAP -66%)+ C 全待兑现 + D 过热。当下=纯期权/情绪定价, 现价已 price-in SST 大幅兑现。结论: 评级 HOLD(持有/中性偏谨慎)——除非 Bull 路径(SST 大放量)兑现, 否则现价缺乏基本面安全边际; 伊顿验证是分水岭, 验证前不追、验证后按结果重判。"},
    "tracking": {
        "intro": "哪个指标恶化 → 哪个假设先崩 → 触发什么动作。",
        "rows": [
            ("__band__", "一、SST 期权(命门)"),
            ("伊顿验证 + SST 收入", "待(2026 H1)", "命门: 全部估值溢价靠它", "互动易/季报/公告", "验证失败→转 Bear 重算(打回 45x PE)"),
            ("磁件分部增速", "2026 半年报看", "命门: SST 上量的财报印证", "半年报/三季报分部数据", "无跳升→下调 SST 收入假设"),
            ("__band__", "二、基础业务"),
            ("海外业务毛利率", "2025=25.66%", "命门: 基础业务净利率地板", "年报/半年报分地区", "回落→下调基础净利率(向 Bear)"),
            ("__band__", "三、情绪/资金"),
            ("两融余额 + 换手", "4.05 亿(+35%/周)", "命门: 情绪档位/股价支撑", "东财融资融券日更", "见顶回落→情绪值下调"),
        ],
    },
})

# ════════════ 全局格式 + 落盘 ════════════
K.finalize(wb, freeze={
    S_HIST: "B3", S_PX: "B4", S_CONS: "A2", S_HMULT: "B5", S_MULT: "B4", S_SW: "B3",
    S_ANCHOR: "B3", S_SEG: "B3", S_FUND: "B3", S_VAL: "B4", S_CMP: "B6", S_DASH: "B6",
    S_COVER: "A2",
})
out = os.path.join(os.path.dirname(__file__), "..", "out", "002885_valuation_model.xlsx")
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("saved:", os.path.abspath(out))
print("sheets:", wb.sheetnames)
