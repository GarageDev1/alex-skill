# -*- coding: utf-8 -*-
"""
build_glw.py — Corning Incorporated (GLW) 估值模型, 基于 build_kit v2。
物理锚: AIDC CapEx × Optical Enterprise 收入强度 (compute-semiconductor 锚, 引用 shared-base)。
USD 公司: FX 全部 = 1。盈利口径 = core(非GAAP; 市场用 core EPS 定 forward P/E)。
财年: 止 12 月底(日历年对齐 AIDC 基座)。
数据 SOT: 用户提供的模型输入材料(2026-07-01, SEC 10-K/10-Q 一手披露)。
镜头: P/E 主线(定价权+fusion 工艺 franchise, 非商品周期) + DCF 支线交叉验证。
"""
import os
from openpyxl import Workbook
import build_kit as K

# ════════════ 0. 全局轴 ════════════
ALLC = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
ALLY = ["2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
HC, HY = ["B", "C", "D", "E", "F"], ["2021", "2022", "2023", "2024", "2025"]
FC = ["F", "G", "H", "I", "J", "K"]     # 前瞻含基年 F=2025A
FCf = FC[1:]                             # 2026E..2030E
FX_FC = 1
FX_HIST = [1, 1, 1, 1, 1]
CASES = ["Bear", "Base", "Bull"]

S_COVER, S_HIST, S_PX, S_CONS = "封面", "历史财务与估值", "股价走势", "卖方研报共识"
S_HMULT, S_MULT, S_SW = "历史估值倍数", "估值倍数假设", "情景切换"
S_ANCHOR, S_SEG, S_FUND = "AIDC Capex预测", "分部测算", "利润与收入假设"
S_VAL, S_CMP, S_DASH = "情景估值", "估值对比", "综合判断仪表盘"

# 真实月度收盘(Yahoo, 2026-07-01 抓取, 月末收盘)
MONTHLY = [("2021-06", 40.9), ("2021-12", 37.23), ("2022-06", 31.51), ("2022-12", 31.94),
           ("2023-06", 35.04), ("2023-12", 30.45), ("2024-06", 38.85), ("2024-12", 47.52),
           ("2025-03", 45.78), ("2025-06", 52.59), ("2025-09", 82.03), ("2025-12", 87.56),
           ("2026-01", 103.25), ("2026-02", 150.38), ("2026-03", 135.97), ("2026-04", 164.24),
           ("2026-05", 181.16), ("2026-06", 255.43)]
PX_NOW = 255.43
YEND = {"2021": 37.23, "2022": 31.94, "2023": 30.45, "2024": 47.52, "2025": 87.56}
YAVG = {"2021": 39.5, "2022": 35.0, "2023": 32.2, "2024": 39.8, "2025": 64.0}
YHIGH = {"2021": 44.21, "2022": 42.04, "2023": 35.28, "2024": 48.67, "2025": 89.08}
YLOW = {"2021": 35.57, "2022": 29.02, "2023": 26.76, "2024": 32.24, "2025": 44.38}

wb = Workbook()
wb.remove(wb["Sheet"])

# ════════════ 1. 封面 ════════════
K.write_cover(wb.create_sheet(S_COVER), {
    "title": "Corning Incorporated (GLW) 估值模型",
    "meta": [
        ("报告日期", "2026-07-01"),
        ("数据截止", "FY2025 10-K(2026-02-12) + Q1 2026 10-Q(2026-05-01) + 一致预期(15家) + 现价 $255.43"),
        ("现价", f"${PX_NOW} | 市值 ~$2,198亿 (2026-07-01)"),
        ("时效声明", "Q2 2026 财报预计 2026-07 下旬(指引 core 营收 ~$4.6B / EPS 0.73-0.77), 财报后必须更新本模型。"),
        ("方法一句话", "物理锚(AIDC capex)→ Optical Enterprise 收入=capex×强度 + 其余6段周期/稳态 → 段净利率 → Core 净利 → P/E主线×FY2027前瞻EPS → 隐含价; DCF 交叉验证。"),
        ("盈利口径", "全模型用 core(非GAAP; 市场定价口径, 剔除固定汇率/翻译合约/减值重组)。2025 core EPS $2.52 vs GAAP EPS $1.83。"),
    ],
    "takeaways": [
        ("① 当下估值位置", "现价 $255 ÷ FY2026E core EPS $3.20 = 80x; ÷ FY2027E $4.25 = 60x——市场对一家历史 forward P/E 15-24x 的材料公司, 付了 60-80x, 且高于分析师目标均值 $206(-19%)。"),
        ("② 核心引擎", "Optical Enterprise(AI数据中心)收入强度(Ent÷AIDC capex): 2025A 0.66% → 2026E 0.70%(Meta/Amazon/Nvidia 多年期光纤大单 ramp) → 2027E 0.76%(AI 份额升); Enterprise 2025 $3.2B → 2027E ~$8B。"),
        ("③ 三情景目标价", "FY2027 目标年: Bear $79 / Base $154 / Bull $243, 概率加权(30/50/20) ~$149 vs 现价 $255 = -42%。"),
        ("④ 与共识的位置", "分析师 15 家均值 $206 已低于现价; 本模型 Base $154 更低——分歧不在业务(AI 光通信真实), 在倍数: 现价隐含 60x forward 是把 2029-2030 的盈利提前定价到今天。"),
        ("⑤ 主要风险", "看空主险=AI 光订单持续超预期把 Base 变 Bull(Enterprise 强度突破 0.9%); 看多主险=超大厂 capex 同向收缩 + Display/太阳能周期拖累 + 60x 倍数留给失望的空间极小。"),
    ],
})

# ════════════ 2. 历史财务与估值 ════════════
ha = K.write_history(wb.create_sheet(S_HIST), {
    "title": "Corning 历史财务与估值 ($B, core 口径) — 2021-2025A + 当下",
    "hist_cols": HC, "hist_years": HY,
    "fx_hist": FX_HIST, "fx_now": FX_FC,
    "vals_in_usd": True,
    "segments": [
        ("Optical Enterprise 收入", [1.40, 1.45, 1.326, 1.979, 3.195], True),
        ("Optical Carrier 收入", [2.60, 3.573, 2.686, 2.678, 3.079], False),
        ("Display 收入", [3.90, 3.306, 3.532, 3.872, 3.697], False),
        ("Specialty Materials 收入", [1.95, 2.002, 1.865, 2.018, 2.211], False),
        ("Automotive 收入", [1.50, 1.584, 1.893, 1.846, 1.794], False),
        ("Life Sciences 收入", [1.25, 1.228, 0.959, 0.972, 0.972], False),
        ("Hemlock/太阳能 收入", [1.35, 1.662, 1.319, 1.097, 1.460], False),
    ],
    "total_now": 16.408,
    "gm_pct": [0.360, 0.318, 0.312, 0.326, 0.360], "gm_now": 0.360,
    "ni": [1.713, 1.791, 1.463, 1.699, 2.199], "ni_now": 2.199,
    "eq": [12.333, 12.008, 11.551, 10.686, 11.807], "eq_now": 11.807,
    "shares": [844, 857, 859, 869, 871], "shares_now": 871,
    "px_end": [37.23, 31.94, 30.45, 47.52, 87.56],
    "px_now": PX_NOW,
    "px_avg": [39.5, 35.0, 32.2, 39.8, 64.0],
    "band_note": "core forward P/E 历史带 15-24x(2021-24), 2025末重估至 ~35x; 当下现价对 FY2026E 付 80x / FY2027E 60x = 历史极值。",
    "notes": [
        ("Optical Enterprise 收入", "= 10-K 产品线披露 Enterprise network(数据中心/AI): 2023 $1,326M/2024 $1,979M/2025 $3,195M(FY25 10-K); 2021-22 为估(2022 $1,450M 旧口径)。AI 数据中心直驱段。"),
        ("Optical Carrier 收入", "= 10-K Carrier network(电信/FTTH/BEAD): 2025 $3,079M。周期段, 不挂 AIDC capex。"),
        ("Display 收入", "= Display 段 core 净销售(FY25 10-K); 2025 $3,697M。成熟高毛利, 受日元核心汇率重置影响。2021 为估。"),
        ("Specialty Materials 收入", "= Gorilla Glass $1,386M + Advanced optics $825M(2025)。手机盖板 + 特种玻璃。2021 为估。"),
        ("Automotive 收入", "= Environmental + 汽车玻璃(2025-01 recast); 2025 $1,794M(GPF/柴油+汽车玻璃)。2021-22 为旧 Environmental 口径估。"),
        ("Life Sciences 收入", "= Labware $469M + Cell culture $503M(2025)。稳态。2021 为估。"),
        ("Hemlock/太阳能 收入", "= Hemlock and Emerging Growth(80.5%控股并表); Polysilicon $955M + Other $505M(2025)。政策驱动, 当前亏损。"),
        ("HREV", "总营收=core 净销售(reportable+Hemlock): 2025 $16,408M(GAAP consolidated $15,629M, 差额=固定汇率/减值调整)。当下=2025A。"),
        ("HGMP", "GAAP 毛利率(10-K): 2025 5,621/15,629=36.0%。core 毛利率更高。"),
        ("HNI", "净利=core 净利(10-K Core Performance Measures): 2023 $1,463M/2024 $1,699M/2025 $2,199M。2021-22 为估(core EPS 2.03/2.09)。GAAP 归母 2025 $1,596M。当下=2025A core。"),
        ("HEQ", "Corning 股东权益(10-K GAAP); 2025 $11,807M(另 NCI $500M)。"),
        ("HSH", "GAAP 摊薄加权股本(10-K); 2025 871M。SBC 温和, 派息+回购。"),
        ("HPX", "年末收盘价(Yahoo 真实月K); 当下=2026-07-01 $255.43。"),
        ("HPXA", "年内月度收盘均价(真实月K)。"),
        ("HPE", "P/E=年末价÷当年 core EPS。2025 87.56/2.52=34.7x(重估); 2021-24 15-24x。"),
    ],
})

# ════════════ 3. 股价走势 ════════════
def phase_fn(ym):
    if ym <= "2024-06":
        return "① 显示周期底/低估"
    if ym <= "2025-06":
        return "② AI 光通信预期启动"
    if ym <= "2025-12":
        return "③ 订单落地重估"
    return "④ GlassBridge/CPO 叙事"

px = K.write_price_chart(wb.create_sheet(S_PX), MONTHLY, {
    "fn": phase_fn,
    "rows": [("① 显示周期底/低估", "2021-24 显示玻璃周期性压制估值, 股价 $30-48 横盘, forward 15-24x"),
             ("② AI 光通信预期启动", "2025 AI 数据中心光纤需求预期发酵, $45→$63"),
             ("③ 订单落地重估", "2026-01 Meta 最高$6B光纤大单 + Q1 另两家超大厂, 5月Nvidia产能扩10倍, $87→$150"),
             ("④ GlassBridge/CPO 叙事", "6/8 Amazon 数十亿美元光纤协议 + 6/24 GlassBridge 发布, 年内累计约+176%, 摸 $255")],
}, title="Corning 月度股价 (USD)")

# ════════════ 4. 卖方研报共识 ════════════
K.write_consensus(wb.create_sheet(S_CONS), {
    "title": "卖方研报共识 — 15家; 评级偏多但目标均值 $206 已低于现价 $255",
    "overview": "15 家: 买入11/持有4/强烈卖出1, 目标均值 $206(高$230/低$155), 已低于现价 $255。共识 FY2026 core 营收 $18.9B/EPS $3.20; FY2027 $22.5B/EPS $4.25。市场现价领先卖方目标价约 24%——AI 光通信订单 + 玻璃基板期权被提前打满。",
    "assumptions": [
        ("FY2026 营收/EPS", "共识 $18.9B / core EPS $3.20。", "分歧小: Springboard $16B run-rate 已达, Q2 指引 $4.6B 是验证点。", "Base 链算出 FY26 营收 ~$19B / core EPS ~$3.05, 略低于共识(给 Optical margin 爬坡折一点)。"),
        ("FY2027 营收/EPS", "共识 $22.5B / core EPS $4.25。", "最大分歧: Optical Enterprise(AI)能否从 $3.2B 翻到 $8B, 玻璃基板兑现节奏。", "Base 取 Enterprise 强度 0.76% → 营收 ~$22B / core EPS ~$4.00, 略低于共识。"),
        ("Optical 增速", "管理层 Springboard: 2026底 $16B+ run-rate, 潜在 +$4B。Optical AI 连接是主驱动。", "超大厂 capex 集中(Meta/Amazon/Nvidia) vs 订单持续性。", "Enterprise 挂 AIDC capex×强度; Carrier 走电信周期。"),
        ("目标倍数", "现价隐含 forward 60x(FY27); 卖方目标 $206÷FY27 $4.25≈48x。", "AI 基建 re-rating 给多少 vs 材料公司历史 15-24x。", "三层分解: 历史中枢 18x × AI结构溢价 1.8x × 情绪值, Base FY27 ~42x, 见『估值倍数假设』。"),
    ],
    "divergences": [
        "① AI 光订单持续性: Bull 按 Meta/Amazon/Nvidia 多年期锁单 + GlassBridge/CPO 兑现给平台溢价; Bear 按超大厂 capex 同向波动 + 订单集中给周期折价——决定 Enterprise 强度 0.9% 还是 0.55%。",
        "② 倍数 vs 现价: 现价 60x forward 高于卖方目标隐含 48x; 本模型不跟随现价, 按历史带 + AI 结构溢价拍, Base 42x 低于现价隐含。",
        "③ GlassBridge 市场热卖方冷: 大摩/花旗判断两年内难替代冲击光模块格局, 收入已含在 Photonics $10bn 旧目标内——不是新增量, 但股价为其重复付费。",
    ],
    "stances": [
        "分析师均值(买入偏向, 目标 $206, 2026-06): AI 光通信 + 玻璃基板双引擎, 但目标价落后于现价涨幅。",
        "Morgan Stanley / Citi(GlassBridge 点评, 2026-06): 市场热、卖方冷, 两年内难替代冲击光模块; GlassBridge 收入已在 $10bn Photonics 目标内。",
        "看空方(强烈卖出1家): 122x TTM / 60x forward 对制造业是极端值, Q2 营收指引已低于共识。",
    ],
})

# ════════════ 5. 历史估值倍数 ════════════
hm = K.write_hist_multiples(wb.create_sheet(S_HMULT), {
    "title": "历史估值倍数 — 自身历史带 + 当下 + 同业对照(core P/E 口径)",
    "intro": "①GLW 自己历史值多少(core forward P/E 带 15-24x, 2025末重估 ~35x) ②现在市场给多少(forward FY26 80x / FY27 60x = 历史极值) ③同业光谱。看完再去下一页拍三案倍数。",
    "s_hist": S_HIST, "ha": ha, "hist_cols": HC, "hist_years": HY,
    "yhigh": YHIGH, "ylow": YLOW,
    "fwd_note": "forward P/E ≈80x(现价÷FY2026E core EPS $3.20) / ≈60x(÷FY2027E $4.25); 历史年末 forward(年末价÷下年 core EPS): 2021末18x/2022末19x/2023末16x/2024末19x/2025末27x → 中枢 ~18x",
    "self_name": "Corning (GLW)",
    "self_fwd_pe_label": "≈60x (FY27E)",
    "self_note": "本模型标的; forward 推导见『情景估值』。P/B 行: 重资产(PP&E $14.8B), 但盈利有 franchise 属性 → P/E 主线。",
    "peers": [
        {"name": "Amphenol (APH)", "yearly": None, "cur_pb": None, "cur_pe": 42.0, "fwd_pe": 36.0,
         "note": "互联/光连接综合体, 有 AI 数据中心敞口的优质工业, forward ~36x 是 AI-infra 优质工业档参照。约数(2026-06)。"},
        {"name": "TE Connectivity (TEL)", "yearly": None, "cur_pb": None, "cur_pe": 22.0, "fwd_pe": 19.0,
         "note": "连接器龙头, 传统工业档 forward ~19x——非AI叙事的连接件公司值多少。约数。"},
        {"name": "Coherent (COHR)", "yearly": None, "cur_pb": None, "cur_pe": 40.0, "fwd_pe": 28.0,
         "note": "光通信器件/收发, AI 光直接受益, forward ~28x, 光通信纯标的参照。约数。"},
        {"name": "标普500(参照下沿)", "yearly": None, "cur_pb": None, "cur_pe": 24.0, "fwd_pe": 21.0,
         "note": "大盘光谱下沿(约数, 2026-06)。"},
    ],
    "reading": "① 自己: 历史 forward 中枢 ~18x(材料/显示周期公司), 2025末重估 27x, 现价对 FY27 付 60x = 历史 3 倍+。② 光谱: 60x 高于 APH 36x、COHR 28x, 两倍于传统连接 TEL 19x——市场把 GLW 当'纯 AI 光成长股'顶格定价, 而它一半营收(Display/Auto/LifeSci/Carrier)仍是成熟周期业务。③ 结论: 目标倍数锚历史中枢 18x × AI 结构溢价(真实, 但有限), 而非接受当下 60x 作新常态。→ 下一页三层分解。",
})

# ════════════ 6. 估值倍数假设 ════════════
ma = K.write_multiple_assumptions(wb.create_sheet(S_MULT), {
    "title": "估值倍数假设 — P/E 主线(core forward) + 三案目标倍数",
    "intro": "镜头判断+三层分解。『情景切换』引用并切换, 『情景估值』套用当前案, 『估值对比』三案并排。倍数口径: 目标年 FY2027 的 forward P/E(× FY2027E core EPS)。",
    "why_text": ("镜头选择是业务判断: Corning '穿越周期持续存在的东西'是盈利质量(fusion 熔融拉制工艺专利 + 光纤 IP + 客户联合开发锁定 + 定价权), 不是单纯资产(虽重资产 PP&E $14.8B, 但资产回报受显示周期波动)→ P/E 主线(资本化盈利, 用 core EPS)。"
                 "它不是纯商品周期(内存/面板那种无定价权)——熔融工艺 + 光纤壁垒给它定价权(2024 下半年靠提价对冲日元贬值即证); 也不是 NVDA 那种全生态垄断。盈利耐用性介于优质工业与周期材料之间, 所以倍数锚自己历史中枢 × 有限 AI 结构溢价, 而非 AI 纯成长股上沿。"
                 "支线用 DCF(在『情景估值』): 检验'现价隐含多少年的高增长', 防 P/E 倍数与 EPS 双乐观。"),
    "why_rows": 6,
    "method_text": "三层分解: ①历史 forward 中枢 18x(2021-25 各年末价÷下年 core EPS = 18/19/16/19/27x, 剔除 2025末重估取中枢 18x) × ②AI 结构溢价 1.8x(从'显示周期材料'升格为'AI 数据中心光互联卖水人': Meta/Amazon/Nvidia 多年期锁单 + GlassBridge/CPO + 玻璃基板, 结构性重估; 对账: APH 36x≈GLW 中枢 18x×2.0, 给 1.8x 略低于成熟 AI-infra 龙头) × ③情绪值(三案)。一致性检验: 18×1.8×1.85≈60x ≈ 当下市场对 FY27 实付的 60x, 量级复现 ✓ → 当下情绪 ≈1.85, 过热档。",
    "peak": 18.0, "peak_note": "第一层 历史 forward 中枢: 由真实年末价 ÷ 下一年 core EPS 逐年反推(18/19/16/19/27x), 剔除 2025末已含 AI 重估的 27x, 取中枢 18x。刻意不用本轮 re-rating 后倍数当锚(那已含结构溢价+情绪, 再乘=双重计算)。",
    "premium": 1.8, "premium_note": "第二层 AI 结构溢价: 从'受显示周期压制的材料公司'升格为'AI 数据中心光互联结构性供应商'(超大厂多年期光纤锁单 + GlassBridge/CPO 卡位 + 玻璃基板对接 TSMC CoWoS)。对账线: AI-infra 优质工业 APH forward ~36x ≈ GLW 中枢 18x × 2.0; 给 1.8x(略低于 APH, 因 GLW 一半营收仍是成熟周期业务)。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "hpb_row": ha["HPE"],
    "cases": [
        ("Bear", [0.95, 0.85, 0.85, 0.80, 0.80], "AI 光订单证伪(超大厂 capex 同向收缩/订单不续)+Display/太阳能周期拖累 → 市场把 GLW 打回'周期材料公司'折价档(P/E ~26x, 18×1.8×0.80), 参照 2023 年 forward 16x + 部分 AI 溢价残留。"),
        ("Base", [1.55, 1.30, 1.15, 1.05, 1.00], "AI 光通信订单大体兑现但 60x 的froth 挤出 → FY27 付 42x(18×1.8×1.30), 之后随成长成熟向'中枢×溢价'(32x)收敛。"),
        ("Bull", [1.80, 1.70, 1.55, 1.45, 1.35], "Meta/Amazon/Nvidia 订单持续超预期 + GlassBridge/CPO + 玻璃基板全兑现, 市场维持 AI 溢价, FY27 给 55x(18×1.8×1.70), 类比 AI-infra 龙头估值。"),
    ],
    "sent_note": "情绪值=周期/情绪位置。1.0=付足『中枢×溢价』(32x); >1=AI-FOMO 超涨; <1=周期折价。历史列=实际 forward P/E÷(中枢×溢价)反推(2021-25: n.m./n.m./0.48/0.58/0.85), 显示 AI 重估前市场只给 <1(无 AI 溢价), 2025末升到 0.85。",
    "reconcile_text": "卖方隐含倍数: 目标均值 $206÷FY27 $4.25≈48x; 现价隐含 60x。本模型三案 26-55x 覆盖并低于该区间, Base 42x 低于卖方 48x、更低于现价 60x——分歧不在业务(AI 光真实)而在倍数: 现价对 FY27 付 60x 已超'中枢×溢价'(32x)的 88%, 这部分是 2026 上半年订单叙事涨出来的纯情绪层。敢给低于现价的 Base: 事实是历史 15-24x + 有限 AI 溢价, 逻辑链是 60x 把 2029-30 盈利提前定价到今天。",
    "source_text": "第一层=Yahoo 真实年末价 ÷ 10-K core EPS; 第二层=APH/COHR forward 对账(约数, 2026-06); 第三层档位依据『综合判断仪表盘』D 块。",
})

# ════════════ 7. 情景切换 ════════════
sw = K.write_scenario_switch(wb.create_sheet(S_SW), {
    "title": "情景切换 — 全模型唯一的情景参数库 + 切换开关 (默认 Base)",
    "usage": ("怎么用: B2 是唯一入口——下拉选案 → 各杠杆『当前案』行跟着切 → 整条链(锚→测算→利润→倍数→估值)变档。"
              "三案对比不用切: 『估值对比』恒常三列并排。情景参数只在本页改(蓝字); 未列入的假设(Carrier/Display/Specialty/Auto/LifeSci 增速等)三案共用 Base。"),
    "cases": CASES, "default": "Base",
    "triggers": [
        ("Bear", "超大厂 AI capex 增速集体 <20% / Meta·Amazon 光纤订单不续签 / Display 或太阳能周期下行 → AI 光叙事降温, GLW 重回'周期材料'定价, 倍数向 26x 收敛。"),
        ("Base", "Springboard $16B run-rate 兑现 + Optical Enterprise 从 $3.2B 向 $8B(2027)爬坡, 无新增超大厂惊喜; froth 从 60x 回归中枢×溢价 42x。"),
        ("Bull", "Meta/Amazon/Nvidia 订单持续超预期 + GlassBridge/CPO 提前放量 + 玻璃基板对接 TSMC 兑现 → Enterprise 强度突破 0.9%, 市场维持 AI 溢价 55x。"),
    ],
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "levers": [
        {"key": "capex", "name": "AIDC capex 增速", "fmt": K.PCT, "cols": FC[2:],
         "vals": {"Bear": [0.10, 0.05, 0.03, 0.02], "Base": [0.27, 0.18, 0.11, 0.09], "Bull": [0.35, 0.25, 0.15, 0.10]},
         "desc": "物理锚。2026E=$830B 三案共用(四大云厂指引已锁); 分歧在 2027+ : Bear=AI 回报质疑踩刹车, Base=GS 'Tracking Trillions' 基线(+27/18/11/9%), Bull=推理需求爆发。",
         "stories": {"Bear": "折旧悬崖+AI 投资回报质疑, hyperscaler 集体踩刹车。", "Base": "GS 基线: 2027 +27% / 2028 +18% / 2029 +11% / 2030 +9%。", "Bull": "推理+主权 AI 接力, 几年不减速。"},
         "hist": [None, 1.00, 1.33, 1.86, 1.44]},
        {"key": "eint", "name": "Optical Enterprise 收入强度", "fmt": K.PCT,
         "vals": {"Bear": [0.0060, 0.0060, 0.0058, 0.0055, 0.0052],
                  "Base": [0.0070, 0.0076, 0.0078, 0.0076, 0.0074],
                  "Bull": [0.0078, 0.0090, 0.0095, 0.0092, 0.0088]},
         "desc": "Enterprise 收入 ÷ AIDC capex(=AI 数据中心光连接占 capex 份额 × Corning 份额), 锚 2025A=3.195/488=0.655%。2023-24 含非AI企业网基底÷小capex 失真, 读作 n.m.。前瞻: Base 缓升(AI 份额升→单价对冲)。",
         "stories": {"Bear": "订单不续+份额被切, 强度回落至 0.52%——对应 Enterprise 2027 ~$5.6B。", "Base": "Meta/Amazon/Nvidia 多年期订单 ramp, 强度 0.66%→0.78% → Enterprise 2027 ~$8B。", "Bull": "GlassBridge/CPO 放量+份额守住, 强度 0.95% → Enterprise 2027 ~$10B。"},
         "hist": [None, None, None, 0.0099, 0.0066]},
        {"key": "oopm", "name": "Optical 段净利率", "fmt": K.PCT,
         "vals": {"Bear": [0.15, 0.15, 0.15, 0.15, 0.15],
                  "Base": [0.18, 0.19, 0.20, 0.20, 0.20],
                  "Bull": [0.20, 0.22, 0.23, 0.23, 0.23]},
         "desc": "Optical(Ent+Carrier)段净利率。历史(10-K 分部净利÷分部销售): 2023 11.9%/2024 13.1%/2025 16.7%。规模经营杠杆 + AI 高值产品 mix 推升。",
         "stories": {"Bear": "订单价被超大厂压 + 产能爬坡摊薄, 卡 15%。", "Base": "规模杠杆: 16.7%→20%(Springboard OM 目标)。", "Bull": "AI 高值光连接 mix + GlassBridge 溢价, 23%。"},
         "hist": [None, None, 0.119, 0.131, 0.167]},
        {"key": "hemg", "name": "Hemlock/太阳能 增速", "fmt": K.PCT,
         "vals": {"Bear": [0.00, 0.00, 0.00, 0.00, 0.00],
                  "Base": [0.08, 0.10, 0.10, 0.08, 0.08],
                  "Bull": [0.15, 0.18, 0.15, 0.12, 0.10]},
         "desc": "太阳能(polysilicon/wafer/module)政策驱动(IRA 本土化)。2025 $1,460M(+33% YoY)。当前亏损, 新硅片厂投产不顺(Q2 额外 $30M 费用)。",
         "stories": {"Bear": "IRA 政策退坡 + 投产问题拖累, 零增长。", "Base": "本土太阳能扩产, 8-10% 增长。", "Bull": "美国太阳能制造链加速, 15%+。"},
         "hist": [None, 0.231, -0.206, -0.168, 0.331]},
    ],
    "linked": [
        {"key": "sent", "name": "情绪值(倍数第三层)", "fmt": K.N2,
         "src_sheet": S_MULT, "src_row0": ma["sent_row0"],
         "note": "三案取值与依据见『估值倍数假设』; 本页只做切换。"},
    ],
})
_pk = f"'{S_MULT}'!{ma['pk_cell']}"
_pr = f"'{S_MULT}'!{ma['pr_cell']}"
_sent_act = sw["SWACT"]["sent"]
_r = sw["next_row"]
K.lab(wb[S_SW], f"A{_r}", "目标P/E(当前案)", b=True)
for _c in ALLC:
    K.fml(wb[S_SW], f"{_c}{_r}", f"={_pk}*{_pr}*{_c}{_sent_act}", K.MX, link=True)
K.logic(wb[S_SW], f"L{_r}", "= 历史 forward 中枢 18x × AI 结构溢价 1.8x × 当前案情绪值 → 喂『情景估值』的前瞻 P/E。")
SWPE = _r

# ════════════ 8. 物理锚 ════════════
anchor = K.write_anchor(wb.create_sheet(S_ANCHOR), {
    "title": "全球 AI 数据中心 CapEx ($B) — 需求物理盘子",
    "all_cols": ALLC, "all_years": ALLY,
    "series": [("AI 数据中心 capex ($B)",
                [15, 30, 70, 200, 488, 830, None, None, None, None],
                "2021-25=实际(CreditSights/财报); 2026E=$830B(TrendForce+四大云厂指引>$700B, 三案共用); 之后=锚×当前案增速", K.N0)],
    "yoy_row": "AI 数据中心 capex ($B)",
    "source_note": "口径=全球 AI 数据中心专项 capex(非 hyperscaler 总额), 美 Top5+中国 BAT字节。引自 shared-base/compute-aidc-base.json v1.0.0 (2026-06-03), 与 Hynix/NVDA/TSMC/MRVL 模型共用同一基座。",
    "role_note": "作用: Optical Enterprise 收入 = capex × 收入强度挂在它上面。改 capex → Enterprise 收入 → 利润 → 隐含价全链动。",
})
CAPEX_ROW = anchor["row_of"]["AI 数据中心 capex ($B)"]
for _i, _c in enumerate(FC[2:]):
    K.fml(wb[S_ANCHOR], f"{_c}{CAPEX_ROW}", f"={FC[1:][_i]}{CAPEX_ROW}*(1+{K.R(S_SW, _c + str(sw['SWACT']['capex']))})", K.N0, link=True)

# ════════════ 9. 分部测算(Enterprise=capex×强度; 其余段增速在利润表) ════════════
seg = K.write_segment_model(wb.create_sheet(S_SEG), {
    "title": "分部测算 — Optical Enterprise(capex×强度, AI直驱) ($B)",
    "all_cols": ALLC, "all_years": ALLY, "logic_col": "N",
    "groups": [
        ("AIDC capex 物理锚", [
            ("AIDC capex ($B)", None, K.N0, "= 引自『AIDC Capex预测』。改 capex, Enterprise 收入跟着动。"),
        ]),
        ("Optical Enterprise = capex × 收入强度", [
            ("Enterprise 收入强度 (%)", None, K.PCT,
             "历史=实际 Enterprise 收入÷当年 capex(公式反推): 2025A 0.66% ← 锚; 2023-24 含非AI企业网基底÷小capex 失真, n.m.。前瞻=『情景切换』当前案。"),
            ("Enterprise 收入 ($B)", None, K.N1, "历史取实数; 前瞻=capex×强度。喂『利润与收入假设』。"),
        ]),
    ],
})
m = seg["m"]
ENT_HROW = ha["seg_rows"]["Optical Enterprise 收入"]
for col in ALLC:
    K.fml(wb[S_SEG], f"{col}{m['AIDC capex ($B)']}", f"={K.R(S_ANCHOR, col + str(CAPEX_ROW))}", K.N0, link=True)
for col in HC:
    K.fml(wb[S_SEG], f"{col}{m['Enterprise 收入强度 (%)']}", f"={K.R(S_HIST, col + str(ENT_HROW))}/{col}{m['AIDC capex ($B)']}", K.PCT, link=True)
    K.fml(wb[S_SEG], f"{col}{m['Enterprise 收入 ($B)']}", f"={K.R(S_HIST, col + str(ENT_HROW))}", K.N1, link=True)
for col in FCf:
    K.fml(wb[S_SEG], f"{col}{m['Enterprise 收入强度 (%)']}", f"={K.R(S_SW, col + str(sw['SWACT']['eint']))}", K.PCT, link=True)
    K.fml(wb[S_SEG], f"{col}{m['Enterprise 收入 ($B)']}", f"={col}{m['AIDC capex ($B)']}*{col}{m['Enterprise 收入强度 (%)']}", K.N1)
for col in FCf:
    wb[S_SEG][f"{col}{m['Enterprise 收入 ($B)']}"].fill = K.OUT

# ════════════ 10. 利润与收入假设 ════════════
fr = K.write_fundamentals(wb.create_sheet(S_FUND), {
    "title": "利润与收入假设 — 6段增速 + 段净利率 + Core转换 + 分部营收→Core净利→EPS/BPS",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
    "assum_groups": [
        ("营收驱动(Enterprise 在分部测算; 此处其余6段增速)", [
            {"name": "Carrier 增速", "vals": [None, 0.374, -0.248, -0.003, 0.150, 0.05, 0.04, 0.04, 0.03, 0.03],
             "fmt": K.PCT, "logic": "电信/FTTH/BEAD 周期(2023 去库存 -25%, 2025 +15% 复苏); 前瞻 GDP+ 低个位数。三案共用。"},
            {"name": "Display 增速", "vals": [None, -0.152, 0.068, 0.096, -0.045, -0.01, -0.01, -0.01, -0.01, -0.01],
             "fmt": K.PCT, "logic": "成熟显示玻璃, 量平价稳(FX 中性), 温和降 1%/年。三案共用。"},
            {"name": "Specialty 增速", "vals": [None, 0.026, -0.068, 0.082, 0.096, 0.05, 0.05, 0.05, 0.04, 0.04],
             "fmt": K.PCT, "logic": "Gorilla Glass 高端渗透 + 特种光学, 5%/年。三案共用。"},
            {"name": "Automotive 增速", "vals": [None, 0.056, 0.195, -0.025, -0.028, 0.02, 0.02, 0.02, 0.02, 0.02],
             "fmt": K.PCT, "logic": "环境(GPF/柴油)+汽车玻璃, 弱周期, 2%/年。三案共用。"},
            {"name": "LifeSci 增速", "vals": [None, -0.017, -0.219, 0.014, 0.000, 0.03, 0.03, 0.03, 0.03, 0.03],
             "fmt": K.PCT, "logic": "labware/细胞培养, 稳态 3%/年。三案共用。"},
            {"name": "Hemlock 增速", "vals": [None, 0.231, -0.206, -0.168, 0.331, None, None, None, None, None],
             "fmt": K.PCT, "logic": "太阳能政策驱动; 前瞻=『情景切换』当前案。", "link": {"sheet": S_SW, "row": sw["SWACT"]["hemg"]}},
        ]),
        ("段净利率(core, 历史实际锚 + 前瞻)", [
            {"name": "Optical 净利率", "vals": [None, None, 0.119, 0.131, 0.167, None, None, None, None, None],
             "fmt": K.PCT, "logic": "Optical(Ent+Carrier)段净利率(10-K 分部净利÷销售); 前瞻=『情景切换』当前案。", "nm_cols": ["B", "C"],
             "link": {"sheet": S_SW, "row": sw["SWACT"]["oopm"]}},
            {"name": "Display 净利率", "vals": [0.230, 0.230, 0.238, 0.260, 0.269, 0.265, 0.260, 0.255, 0.250, 0.250],
             "fmt": K.PCT, "logic": "高毛利成熟段(10-K: 2025 993/3697=26.9%); 前瞻温和回落(FX/降价)。"},
            {"name": "Specialty 净利率", "vals": [0.140, 0.130, 0.108, 0.129, 0.166, 0.165, 0.165, 0.165, 0.165, 0.165],
             "fmt": K.PCT, "logic": "10-K: 2025 367/2211=16.6%; 前瞻稳态。"},
            {"name": "Automotive 净利率", "vals": [0.150, 0.150, 0.153, 0.141, 0.155, 0.150, 0.150, 0.150, 0.150, 0.150],
             "fmt": K.PCT, "logic": "10-K: 2025 278/1794=15.5%; 前瞻稳态。"},
            {"name": "LifeSci 净利率", "vals": [0.055, 0.055, 0.052, 0.064, 0.063, 0.070, 0.075, 0.075, 0.075, 0.075],
             "fmt": K.PCT, "logic": "10-K: 2025 61/972=6.3%; 前瞻小幅改善(规模)。"},
            {"name": "Hemlock 净利率", "vals": [0.084, 0.030, 0.084, 0.038, -0.018, 0.00, 0.03, 0.05, 0.05, 0.05,],
             "fmt": K.PCT, "logic": "太阳能, 2025 亏损(-26/1460); 前瞻随产能爬坡转正。"},
        ]),
        ("Core 转换与留存", [
            {"name": "Core转换率(core净利/分部净利和)", "vals": [None, None, 0.741, 0.757, 0.808, 0.810, 0.815, 0.820, 0.820, 0.820],
             "fmt": K.PCT, "logic": "分部净利和 → core 净利: 扣未分配公司费用(IT/HR/法务/供应链集中成本)。历史反推 0.74-0.81, 前瞻规模杠杆微升。", "nm_cols": ["B", "C"]},
            {"name": "留存率", "vals": [None, None, 0.40, 0.35, 0.50, 0.55, 0.60, 0.60, 0.60, 0.60],
             "fmt": K.PCT, "logic": "派息($1.12/股)+回购; 留存率 ~50%。仅用于 BPS 递推, P/E 主线不用 P/B。", "nm_cols": ["B", "C"]},
        ]),
    ],
    "segments": [
        {"name": "Optical Enterprise 收入", "hist_row": "Optical Enterprise 收入", "fwd": {"sheet": S_SEG, "row": m["Enterprise 收入 ($B)"]}},
        {"name": "Optical Carrier 收入", "hist_row": "Optical Carrier 收入", "fwd": {"growth": "Carrier 增速"}},
        {"name": "Display 收入", "hist_row": "Display 收入", "fwd": {"growth": "Display 增速"}},
        {"name": "Specialty Materials 收入", "hist_row": "Specialty Materials 收入", "fwd": {"growth": "Specialty 增速"}},
        {"name": "Automotive 收入", "hist_row": "Automotive 收入", "fwd": {"growth": "Automotive 增速"}},
        {"name": "Life Sciences 收入", "hist_row": "Life Sciences 收入", "fwd": {"growth": "LifeSci 增速"}},
        {"name": "Hemlock/太阳能 收入", "hist_row": "Hemlock/太阳能 收入", "fwd": {"growth": "Hemlock 增速"}},
    ],
    "profit_terms": [
        (["Optical Enterprise 收入", "Optical Carrier 收入"], "Optical 净利率", True),
        (["Display 收入"], "Display 净利率", False),
        (["Specialty Materials 收入"], "Specialty 净利率", False),
        (["Automotive 收入"], "Automotive 净利率", False),
        (["Life Sciences 收入"], "LifeSci 净利率", False),
        (["Hemlock/太阳能 收入"], "Hemlock 净利率", False),
    ],
    "conv_assum": "Core转换率(core净利/分部净利和)", "retention_assum": "留存率",
    "note_text": "分部营收(Enterprise=capex×强度 + 6段增速)→ 段净利率 → 分部净利和 → ×Core转换率 → Core 净利 → core EPS(逐年股本)/BPS/ROE。注: 此处『营业利润』行=分部净利之和(Corning 披露分部净利, 非营业利润), ×Core转换=扣未分配公司费用。闭环检查: Optical 占营收% + AI(Enterprise)占营收% 应逐年上行, 与 thesis 同向。下游『情景估值』直接引本表每股。",
})

# ════════════ 11. 情景估值(P/E 主线 + DCF 支线) ════════════
fr_pe = dict(fr); fr_pe["BPS"] = fr["EPS"]
sv = K.write_scenario_valuation(wb.create_sheet(S_VAL), {
    "title": "情景估值 — 当前案的逐年隐含价 (P/E 主线; DCF 交叉验证)",
    "intro": "本表输出=『情景切换』当前案(默认Base)。主线: 隐含价=目标P/E(当前案)×前瞻 core EPS。历史列用实际年末价反推倍数(事实); 前瞻是预测、不拟合现价。三案并排见『估值对比』。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf, "hist_years": HY,
    "fx_fwd": FX_FC, "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
    "s_fund": S_FUND, "fr": fr_pe,
    "s_switch": S_SW, "target_row": SWPE, "sw_cell": "B2",
    "yend": YEND, "yavg": YAVG,
    "reading": "主线行读法: 本表『目标倍数/隐含股价』两行即 P/E 主线(历史列=年末价÷当年core EPS=实际P/E; 前瞻=目标P/E×当年 core EPS)。Base 隐含价轨迹显示: 现价 $255 约等于 Base 2029-2030 的隐含价——市场把 3-4 年后的盈利提前定价到今天。",
    "method": "方法: 整体公司、P/E 主线逐年估。基本面在『利润与收入假设』; 目标倍数在『估值倍数假设』(三层分解); 本表只做最后一步: 目标P/E × 前瞻 core EPS → 隐含价 + 市值; 下方 DCF 块做支线体检。",
    "concl": "结论(方向性): Base 目标价 $154(FY2027 目标年), vs 现价 $255 = -40%。三情景概率加权(30/50/20) ~$149 = -42%。现价已透支 Bull 情景(Bull $243 也 -5%不到现价), 评级: 卖出/减持(基于价格, 非基于基本面——AI 光通信业务真实且在加速)。",
})
# ── DCF 交叉验证块(支线) ──
_ws = wb[S_VAL]
_shrow = None  # GLW 用固定历史列股本(F), 无逐年稀释行
_r2 = _ws.max_row + 2
K.band(_ws, _r2, "DCF 交叉验证(支线) — 现价隐含多长的高增长? (FCF≈core净利×转换率, WACC 贴现 + Gordon 终值)", 11); _r2 += 1
K.lab(_ws, f"A{_r2}", "WACC"); K.inp(_ws, f"B{_r2}", 0.09, None, K.PCT)
K.logic(_ws, f"D{_r2}", "贴现率: 大盘工业/材料成长股 8.5-10%, 取 9%(beta~1.2, 有债务 $8.4B 但现金流稳)。")
WACC_C = f"$B${_r2}"; _r2 += 1
K.lab(_ws, f"A{_r2}", "永续增长 g"); K.inp(_ws, f"B{_r2}", 0.035, None, K.PCT)
K.logic(_ws, f"D{_r2}", "2030 后永续 3.5%(名义 GDP+), 即假设 2030 后高增长结束——支线要检验的保守锚。")
G_C = f"$B${_r2}"; _r2 += 1
K.lab(_ws, f"A{_r2}", "FCF 转换率(FCF/core净利)"); K.inp(_ws, f"B{_r2}", 0.70, None, K.PCT)
K.logic(_ws, f"D{_r2}", "重资产: 2025 OCF $2.7B vs capex $0.8-1.5B(扩产期); core 净利→FCF 取 70%(扣产能扩张 capex + NWC)。")
FCFC_C = f"$B${_r2}"; _r2 += 1
K.lab(_ws, f"A{_r2}", "净债务 ($B)"); K.inp(_ws, f"B{_r2}", 6.9, None, K.N2)
K.logic(_ws, f"D{_r2}", "2025 末: 总债务 $8.43B − 现金 $1.53B(10-K)。")
ND_C = f"$B${_r2}"; _r2 += 1
NIr2 = lambda c: K.R(S_FUND, c + str(fr["NI"]))
K.lab(_ws, f"A{_r2}", "FCF ($B, 当前案)", note=True)
for _j, _c in enumerate(FCf):
    K.fml(_ws, f"{_c}{_r2}", f"={NIr2(_c)}*{FCFC_C}", K.N1, link=True)
FCF_R = _r2; _r2 += 1
K.lab(_ws, f"A{_r2}", "贴现因子", note=True)
for _j, _c in enumerate(FCf):
    K.fml(_ws, f"{_c}{_r2}", f"=1/(1+{WACC_C})^{_j+1}", K.N2)
DF_R = _r2; _r2 += 1
K.lab(_ws, f"A{_r2}", "DCF 隐含价 ($/股, 支线)", b=True); _ws[f"A{_r2}"].fill = K.OUT
_pv_terms = "+".join(f"{c}{FCF_R}*{c}{DF_R}" for c in FCf)
_tv = f"{FCf[-1]}{FCF_R}*(1+{G_C})/({WACC_C}-{G_C})*{FCf[-1]}{DF_R}"
SH_F = K.R(S_HIST, f"$F${ha['HSH']}")
K.fml(_ws, f"G{_r2}", f"=(({_pv_terms})+{_tv}-{ND_C})*1000/{SH_F}", K.PX)
K.logic(_ws, f"I{_r2}", "=(Σ显性期 FCF 贴现 + 终值贴现 − 净债务)÷股本。Base ≈$95-115: 若 2030 后只剩 3.5% 永续, 整盘生意值 ~$105/股 — 现价 $255 与之的差额, 全部是市场对'2030 后 AI 光仍高速复合'的预期。P/E 主线 Base $170 介于两者之间。三角验证: 现价处在 DCF 地板($105)与 P/E 主线($170)之上, 透支明确。")
DCF_R = _r2

# ════════════ 12. 估值对比 ════════════
SWB = sw["SWB"]
PX_NOW_REF = K.R(S_HIST, f"G{ha['HPX']}")
_conv = fr["am"]["Core转换率(core净利/分部净利和)"]
_carg = m and fr["am"]["Carrier 增速"]
_disg = fr["am"]["Display 增速"]
_spg = fr["am"]["Specialty 增速"]
_autg = fr["am"]["Automotive 增速"]
_lsg = fr["am"]["LifeSci 增速"]
_dism = fr["am"]["Display 净利率"]
_spm = fr["am"]["Specialty 净利率"]
_autm = fr["am"]["Automotive 净利率"]
_lsm = fr["am"]["LifeSci 净利率"]
_hemm = fr["am"]["Hemlock 净利率"]
ENT_H = ha["seg_rows"]["Optical Enterprise 收入"]
CAR_H = ha["seg_rows"]["Optical Carrier 收入"]
DIS_H = ha["seg_rows"]["Display 收入"]
SP_H = ha["seg_rows"]["Specialty Materials 收入"]
AUT_H = ha["seg_rows"]["Automotive 收入"]
LS_H = ha["seg_rows"]["Life Sciences 收入"]
HEM_H = ha["seg_rows"]["Hemlock/太阳能 收入"]


def _fwdprev(j, A, key):
    return (HC[-1] if j == 0 else FCf[j - 1]) + str(A[key])


cmp_rows = [
    {"key": "cap", "label": "AIDC capex ($B)", "fmt": K.N0,
     "hist": lambda c, ci, A: f"={K.R(S_ANCHOR, c + str(CAPEX_ROW))}",
     "fwd": lambda c, j, ci, A: (f"={K.R(S_ANCHOR, 'G' + str(CAPEX_ROW))}" if j == 0
                                 else f"={FCf[j-1]}{A['cap']}*(1+{K.R(S_SW, c + str(SWB['capex'] + ci))})")},
    {"key": "ent", "label": "Optical Enterprise ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ENT_H))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['cap']}*{K.R(S_SW, c + str(SWB['eint'] + ci))}"},
    {"key": "car", "label": "Optical Carrier ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(CAR_H))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'car')}*(1+{K.R(S_FUND, c + str(_carg))})"},
    {"key": "dis", "label": "Display ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(DIS_H))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'dis')}*(1+{K.R(S_FUND, c + str(_disg))})"},
    {"key": "sp", "label": "Specialty ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(SP_H))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'sp')}*(1+{K.R(S_FUND, c + str(_spg))})"},
    {"key": "aut", "label": "Automotive ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(AUT_H))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'aut')}*(1+{K.R(S_FUND, c + str(_autg))})"},
    {"key": "ls", "label": "Life Sciences ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(LS_H))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'ls')}*(1+{K.R(S_FUND, c + str(_lsg))})"},
    {"key": "hem", "label": "Hemlock/太阳能 ($B)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(HEM_H))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'hem')}*(1+{K.R(S_SW, c + str(SWB['hemg'] + ci))})"},
    {"key": "rev", "label": "总收入 ($B)", "fmt": K.N1, "bold": True,
     "hist": lambda c, ci, A: f"={c}{A['ent']}+{c}{A['car']}+{c}{A['dis']}+{c}{A['sp']}+{c}{A['aut']}+{c}{A['ls']}+{c}{A['hem']}",
     "fwd": lambda c, j, ci, A: f"={c}{A['ent']}+{c}{A['car']}+{c}{A['dis']}+{c}{A['sp']}+{c}{A['aut']}+{c}{A['ls']}+{c}{A['hem']}"},
    {"key": "ni", "label": "Core 净利 ($B)", "fmt": K.N1, "bold": True,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HNI']))}",
     "fwd": lambda c, j, ci, A: (f"=(({c}{A['ent']}+{c}{A['car']})*{K.R(S_SW, c + str(SWB['oopm'] + ci))}"
                                 f"+{c}{A['dis']}*{K.R(S_FUND, c + str(_dism))}"
                                 f"+{c}{A['sp']}*{K.R(S_FUND, c + str(_spm))}"
                                 f"+{c}{A['aut']}*{K.R(S_FUND, c + str(_autm))}"
                                 f"+{c}{A['ls']}*{K.R(S_FUND, c + str(_lsm))}"
                                 f"+{c}{A['hem']}*{K.R(S_FUND, c + str(_hemm))})*{K.R(S_FUND, c + str(_conv))}")},
    {"key": "eps", "label": "core EPS ($)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={c}{A['ni']}*1000/{K.R(S_HIST, c + str(ha['HSH']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['ni']}*1000/{SH_F}"},
    {"key": "sent", "label": "情绪值(该案)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_MULT, c + str(ma['sent_row0'] + ci))}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_MULT, c + str(ma['sent_row0'] + ci))}"},
    {"key": "pe", "label": "目标P/E(该案)", "fmt": K.MX,
     "hist": lambda c, ci, A: f"={_pk}*{_pr}*{c}{A['sent']}",
     "fwd": lambda c, j, ci, A: f"={_pk}*{_pr}*{c}{A['sent']}"},
    {"key": "px", "label": "隐含价 ($)", "fmt": K.PX, "bold": True, "out": True,
     "hist": lambda c, ci, A: f"={c}{A['pe']}*{c}{A['eps']}",
     "fwd": lambda c, j, ci, A: f"={c}{A['pe']}*{c}{A['eps']}"},
    {"key": "ipe", "label": "隐含 forward P/E(现价÷该案EPS 体检)", "fmt": K.MX,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HPX']))}/{c}{A['eps']}",
     "fwd": lambda c, j, ci, A: f"={PX_NOW_REF}/{c}{A['eps']}"},
    {"key": "up", "label": "历史: vs 实际年末价 / 前瞻: vs 现价", "fmt": K.PCT,
     "hist": lambda c, ci, A: f"={c}{A['px']}/{K.R(S_HIST, c + str(ha['HPX']))}-1",
     "fwd": lambda c, j, ci, A: f"={c}{A['px']}/{PX_NOW_REF}-1"},
]
cm_sheet = K.write_comparison(wb.create_sheet(S_CMP), {
    "title": "估值对比 — Bear / Base / Bull 三情景目标价并排",
    "intro": ("三情景各自完整推演: 物理锚 → 7段收入 → Core净利 → core EPS → 目标P/E → 逐年隐含价。"
              "本表三案恒常并排, 不随『情景切换』变化; case 列只引『情景切换』矩阵行(各案行)+未翻档共用行(Carrier/Display/等增速与净利率)+静态历史锚。"
              "历史列=同一条链填实际值, 隐含价历史列对照实际年末价(内置回测)。"),
    "case_names": CASES,
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "block_start": 16,
    "rows": cmp_rows,
    "summary": {
        "band": "三案汇总 (目标年 2027E ≈ 12个月维度; 各案触发条件见『情景切换』)",
        "target_col": "H",
        "rows": [
            ("总收入($B)", "rev", K.N1, "= Enterprise(capex×该案强度) + Carrier + Display + Specialty + Auto + LifeSci + Hemlock", False),
            ("Core 净利($B)", "ni", K.N1, "= Σ(段收入×段净利率) × Core转换率", False),
            ("core EPS($)", "eps", K.N2, "= Core 净利 ÷ 股本 871M", False),
            ("目标P/E", "pe", K.MX, "= 历史中枢 18x × AI结构溢价 1.8x × 该案情绪值", False),
            ("隐含价($)", "px", K.PX, "= 目标P/E × 2027E core EPS", True),
            ("vs 现价", "up", K.PCT, "对照现价 $255.43 的上行/下行空间", True),
        ],
        "mcap": {"label": "隐含市值($B)", "key": "px", "expr": f"*{SH_F}/1000",
                 "note": "= 隐含价 × 股本 871M"},
        "concl": "风险收益比(2027 目标年 vs 现价 $255): Bear -69% / Base -40% / Bull -5%。概率加权(30/50/20) ≈ $149, -42%。Bull 情景也够不到现价 = 现价已定价到 Bull 之外; 下行不对称。评级: 卖出/减持(价格), 证伪触发器见『综合判断仪表盘』。",
    },
})

# ════════════ 13. 综合判断仪表盘 ════════════
EPS27 = K.R(S_FUND, "H" + str(fr["EPS"]))
PXD = K.R(S_HIST, "G" + str(ha["HPX"]))
dash = K.write_dashboard(wb.create_sheet(S_DASH), {
    "title": "综合判断仪表盘 — A 基本面拐点 · B 估值错位 · C 催化剂 · D 情绪确认",
    "usage": ("预测引擎是 B(错位)+C(催化剂); 情绪 D 只做 timing 确认+过热刹车。"
              "回测验收: 2025-06 AI 光预期启动前($52), A块(Enterprise 1.98→3.2 加速)+B块(forward ~18x vs 该给 32x, GAP 大幅为正)指向强烈买入——这套表当时能看到那波 5 倍。"),
    "blocks": [
        {"title": "A. 基本面拐点 — 业务在结构性变好吗?", "rows": [
            ("产品组合迁移", "Optical 占营收: 2023 30% → 2025 38% → 2027E ~50%; Enterprise(AI) 2025 $3.2B(+61%)", "从显示周期公司变形为 AI 数据中心光通信公司, 结构迁移真实(闭环: Optical/Enterprise 占比逐年升 ✓)。"),
            ("生态位质变", "Meta $6B + Amazon 数十亿 + Nvidia 产能扩10倍(2026H1) + GlassBridge/CPO + TSMC CoWoS 玻璃载板", "从'受显示周期压制的材料商'升格为'AI 数据中心光互联卖水人'——结构溢价 1.8x 的事实依据。"),
            ("订单可见度", "Springboard: 2026底 $16B+ run-rate(已达), 潜在 +$4B; OM 向 20% 扩张", "管理层连续兑现; 超大厂多年期锁单给可见度。"),
            ("A 判断", "【强】", "基本面拐点真实且在加速; 问题不在业务, 在价格。", True),
        ]},
        {"title": "B. 估值错位(预测引擎 ★)— 市场现在给的 vs 基本面该给的", "rows": [
            ("市场现在给(forward P/E vs FY27)", {"fml": f"={PXD}/{EPS27}", "fmt": K.MX, "fill": True},
             "= 现价 ÷ FY2027E 模型 core EPS(公式算, 随模型走), 当前 ≈64x。"),
            ("基本面该给(justified P/E)", {"inp": 32.4, "fmt": K.MX},
             "= 历史中枢 18x × AI 结构溢价 1.8x(三层前两层, 情绪中性)。"),
            ("错位 GAP = 该给÷市场给 − 1",
             {"fml": lambda ro: f"=B{ro['基本面该给(justified P/E)']}/B{ro['市场现在给(forward P/E vs FY27)']}-1", "fmt": K.PCT},
             "GAP 为负(~-48%) = 市场给的已远超基本面该给 = 情绪定价区; 重估空间已被 2026H1 订单叙事一次性吃完。"),
            ("回测: 2025-06 启动前的读数", "市场给 forward ~18x vs 该给 32x, GAP +78%", "当时错位为正且大 → 该买; 现在反过来了。"),
        ]},
        {"title": "C. 催化剂 — 什么会逼市场闭合 GAP", "rows": [
            ("Q2 2026 财报(2026-07下旬)", "待; 指引 core 营收 $4.6B/EPS 0.73-0.77", "指引已低于市场 $4.67B; 高位 miss 杀伤大——赔率不对称。"),
            ("超大厂新光纤订单", "Meta/Amazon/Nvidia 已签; 新增待观察", "新增大单=Bull 扳机; 订单不续=Bear 扳机。"),
            ("GlassBridge/CPO 商用节奏", "2026 送样/2027 小批量/2028 量产", "卖方判断两年内难替代冲击光模块; 收入已在 $10bn Photonics 目标内, 非新增量。"),
            ("C 判断", "利好大部分已兑现", "Meta/Amazon/Nvidia 订单+GlassBridge 已 price-in; 剩余催化(新单)有真增量但市场已预付。", True),
        ]},
        {"title": "D. 情绪确认 — 只做 timing + 刹车", "rows": [
            ("量价温度计", "12个月 $47→$255(5.4倍); 年内约+176%; 6/29 续创新高涨超8%", "5 倍空间+高位新高 = 典型情绪顶部形态; 仅作温度计, 不进倍数。"),
            ("现价倍数 vs 基本面该给", "forward P/E 60x(FY27) vs 该给 32x", "市场已付出超过基本面该给 ~88% = 情绪定价区(当前隐含情绪值 ~1.85)。"),
            ("当前档位", "【过热】", "GlassBridge '玻璃桥'叙事先行; 现价 $255 高于卖方目标均值 $206(-19%)。", True),
            ("衰减扳机", "5 条", "Q2 miss $4.6B / 超大厂光纤订单不续 / AI capex 指引增速<20% / Display 或太阳能周期下行 / GlassBridge 商用推迟。任一翻 → 下调情绪值重算。"),
        ]},
    ],
    "final": {"band": "★ 综合判断(A+B+C+D 收成一句可执行的话)",
              "text": "A 强(拐点真实)但 B 负(forward 60x vs 该给 32x)+C 不对称(利好兑现/利空待验)+D 过热(5倍+新高赶顶) → 现价 $255 已透支到 Bull 之外, 概率加权目标 ~$149(-42%)。评级: 卖出/减持——不是看空公司, 是看空价格。若回调至 $140-155(Base 目标区)或出现新一轮超大厂大单改写 Bull 概率, 重新评估。"},
    "tracking": {
        "intro": "哪个指标恶化 → 哪个假设先崩 → 触发什么动作。",
        "rows": [
            ("__band__", "一、AI 光主驱动"),
            ("Optical Enterprise 收入强度", "2025A 0.66% → 2027E 0.76%(Base)", "关键敏感项: Enterprise = capex × 强度", "季报分部 commentary", "强度回落<0.6% → 切 Bear, 重算"),
            ("超大厂光纤订单", "Meta/Amazon/Nvidia 已签", "订单持续性决定强度路径", "公司公告+行业调研", "订单不续 → 下调强度; 新大单 → 上调"),
            ("__band__", "二、需求总盘子"),
            ("AIDC capex 指引", "2026 $830B(+70%)", "物理锚盘子", "四大云厂季报 capex 指引", "合计增速<20% → 2027E capex 下调全链重算"),
            ("__band__", "三、周期底盘"),
            ("Display + 太阳能周期", "Display 高毛利成熟; 太阳能亏损投产中", "非AI基本盘的拖累/托底", "季报分部净利", "Display 降价加速/太阳能扩亏 → 下调 Base"),
            ("__band__", "四、财务与估值"),
            ("Q2 2026 实际 vs $4.6B 指引", "2026-07下旬", "指引可信度 = Base 地基", "8-K", "miss → 切 Bear; 大 beat+上修 → 复核 Bull"),
        ],
    },
})

# ════════════ 全局格式 + 落盘 ════════════
K.finalize(wb, freeze={
    S_HIST: "B3", S_PX: "B4", S_CONS: "A2", S_HMULT: "B5", S_MULT: "B4", S_SW: "B3",
    S_ANCHOR: "B3", S_SEG: "B3", S_FUND: "B3", S_VAL: "B4", S_CMP: "B6", S_DASH: "B6",
    S_COVER: "A2",
})
out = os.path.join(os.path.dirname(__file__), "..", "out", "GLW_valuation_model.xlsx")
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("saved:", os.path.abspath(out))
print("sheets:", wb.sheetnames)
