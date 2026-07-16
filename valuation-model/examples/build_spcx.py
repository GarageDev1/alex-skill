# -*- coding: utf-8 -*-
"""
build_spcx.py — SpaceX (SPCX) 估值模型, 用 build_kit 引擎重建为机构级 13-sheet 架构。

估值方法(v2, 落实 skill 方法论更新):
  SpaceX 整体 + AI 段亏损(P/E·EV/EBITDA 会出负/NA)→ 主镜头切 EV/Sales, 出实数。
  四事业部经济学差异极大 → SOTP。前瞻不再单拍 2030 终端年, 改逐年(2026E-2030E)阶梯:
    各分部 EV(Y) = 分部营收(Y) × 分部 EV/Sales(Y) → ΣEV(Y) + 期权(Y) + 净现金
    → ÷股本×1000 = 当年隐含价(Y)。逐年 forward 口径, 每年用当年 forward 倍数自立, 不折现。
  交叉验证: 加一行 EV/EBITDA(仅 EBITDA>0 的分部年份显示, 否则 n.m.), 满足 ≥2 镜头。
  12M 目标价 = 概率加权(0.30/0.45/0.25)的 2027E 隐含价(NTM+1); 2030E 加权看"成长进价"终点。

  通用 sheet 复用 kit: write_cover / write_history / write_price_chart / write_consensus
  / write_scenario_switch / write_anchor / write_segment_model / write_dashboard / finalize。
  估值三张(估值倍数假设 / 情景估值 / 估值对比)逐年 SOTP × EV/Sales 用 kit 低层 helper 手搓,
  视觉风格(蓝=输入 黑=公式 绿=跨表 + 色带 + 逻辑列 + 冻结)与 kit 一致。

跑: cd examples && PYTHONUTF8=1 python build_spcx.py
注缓存: python _inject_cache_spcx.py   (formulas 解算 → XML 注入缓存值, 不丢样式)
校验: python ../scripts/validate_valuation.py <out>
"""
import os
from openpyxl import Workbook
import build_kit as K

# ════════════ 0. 全局轴 ════════════
# 美元单一币种模型: fx 全 1, ps_scale=1000($B + mn股 → $/股), mcap_div=1000。
ALLC = ["B", "C", "D", "E", "F", "G", "H", "I"]
ALLY = ["2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
HC,  HY = ["B", "C", "D"], ["2023", "2024", "2025"]
FC      = ["D", "E", "F", "G", "H", "I"]   # 前瞻含基年 D=2025A
FCf     = FC[1:]                            # 纯前瞻 2026E-2030E (E F G H I)
FWY     = ["2026E", "2027E", "2028E", "2029E", "2030E"]
FX_HIST = [1, 1, 1]                         # 单一币种(美元), 汇率恒 1
FX_NOW  = 1
CASES   = ["Bear", "Base", "Bull"]

PRICE       = 160.95
SHARES_M    = 13076
NET_CASH    = 67.4

# 逐年三案数据(列顺序 = 2026E,2027E,2028E,2029E,2030E)──────────────
# 分部营收 ($B)
SL_REV = {"Bear": [15, 20, 26, 32, 38], "Base": [17, 25, 34, 43, 52], "Bull": [20, 30, 42, 55, 70]}
SP_REV = {"Bear": [4.3, 4.8, 5.5, 6, 7], "Base": [4.5, 5.5, 7, 9, 12], "Bull": [5, 8, 15, 22, 30]}
AI_REV = {"Bear": [12, 18, 25, 32, 40], "Base": [14, 28, 45, 65, 90], "Bull": [18, 40, 75, 120, 160]}
# 分部 EV/Sales (x)
SL_PS  = {"Bear": [7, 6, 6, 5, 5], "Base": [10, 9, 8, 7, 6], "Bull": [14, 13, 12, 11, 10]}
SP_PS  = {"Bear": [5, 5, 4, 4, 4], "Base": [7, 7, 6, 6, 5], "Bull": [10, 10, 9, 9, 8]}
AI_PS  = {"Bear": [6, 5, 5, 4, 4], "Base": [14, 12, 10, 8, 7], "Bull": [18, 16, 14, 12, 10]}
# 期权 PV ($B)
OPT    = {"Bear": [0, 0, 0, 0, 0], "Base": [40, 50, 60, 65, 70], "Bull": [80, 120, 160, 180, 200]}
# 分部 EBITDA margin(仅做 EV/EBITDA 交叉验证用; AI 早年为负 → 该段标 n.m.)
SL_MGN = 0.66
SP_MGN = 0.24
AI_MGN = {"Bear": [-0.05, 0.05, 0.12, 0.12, 0.12], "Base": [-0.05, 0.10, 0.20, 0.24, 0.28], "Bull": [0.05, 0.18, 0.28, 0.32, 0.35]}
WEIGHTS = {"Bear": 0.30, "Base": 0.45, "Bull": 0.25}

S_COVER, S_HIST, S_PX, S_CONS = "封面", "历史财务与估值", "股价走势", "卖方研报共识"
S_HMULT, S_MULT, S_SW = "历史估值倍数", "估值倍数假设", "情景切换"
S_ANCHOR, S_SEG, S_FUND = "物理锚", "分部测算", "利润与收入假设"
S_VAL, S_CMP, S_DASH = "情景估值", "估值对比", "综合判断仪表盘"

wb = Workbook()
wb.remove(wb["Sheet"])

# ════════════ 1. 封面 ════════════
K.write_cover(wb.create_sheet(S_COVER), {
    "title": "SpaceX (SPCX) 估值模型 — 逐年 SOTP × EV/Sales 阶梯",
    "span": 6,
    "meta": [
        ("报告日期", "2026-06-15"),
        ("数据截止", "S-1/A#2 财报(2023-2025) + 卖方共识(Oppenheimer 2026-06-11) + 实时股价(2026-06-14)"),
        ("现价", f"${PRICE:,.2f}(2026-06-14; 2026-06-12 IPO 定价 $135, 首日 +19%)"),
        ("股本 / 市值", f"{SHARES_M:,} mn 股(IPO 后摊薄); 市值约 $2,104B"),
        ("时效声明", "SpaceX 2026-06-12 才上市, 无历史估值带; 关键假设(轨道算力可行性 / Starship 商业化 / ARPU)演进极快, 建议每季财报后更新。"),
        ("方法一句话", "整体+AI 段亏损 → 主镜头 EV/Sales(出实数, 不留 NA); 四事业部经济学差异大 → SOTP。"
                    "前瞻逐年(2026E-2030E): 各分部当年营收 × 当年 EV/Sales → ΣEV + 期权 + 净现金 → ÷股本 = 当年隐含价, 不折现(逐年 forward 倍数自立)。"),
    ],
    "takeaways": [
        ("① 评级 / 12M 目标价", "卖出(UNDERPERFORM), 12M TP ≈ $56(2027E 概率加权)— 较现价 $161 下行约 65%; 与 Oppenheimer($190)两极。"),
        ("② 逐年阶梯(Base)", "Base 隐含价 2026E $39 → 2027E $55 → 2030E $87, 全程显著低于现价 $161 — 现价 price-in 的不是 Base 路径。"),
        ("③ 现价在为 Bull 的 2029-30 定价", "Bull 2029E $190 / 2030E $215 才超过现价 $161 — 即现价把 Bull 情景的 2029-2030 当成确定来定价, 几乎不留执行/物理风险折让。"),
        ("④ 最大变量 = AI 轨道算力", "AI 分部 2030E 营收 Bear/Base/Bull 跨度 $40B→$160B + 期权 PV $0→$200B — 轨道算力(太空散热)能否物理跑通是隐含价的最大摆幅来源。"),
        ("⑤ 主要风险(双向)", "下行: ARPU 跌破 $50 / Starship 再炸 / 租赁合同流失 / Terafab 跳票。上行: AI1 散热实测达标 / Starship Flight 13+ 成功 / AI 分部转正。"),
    ],
})

# ════════════ 2. 历史财务与估值 ════════════
ha = K.write_history(wb.create_sheet(S_HIST), {
    "title": "SpaceX 历史财务与估值 ($B) — 2023-2025A + 当下",
    "hist_cols": HC, "hist_years": HY,
    "fx_hist": FX_HIST, "fx_now": FX_NOW,
    "vals_in_usd": True, "ps_scale": 1000, "mcap_div": 1000,
    "unit_label": "($B)", "mcap_label": "市值($B)", "fx_label": "FX (单一币种=美元)",
    "cur_label": "当下(2026-06)",
    "segments": [
        ("Starlink(连接) 收入", [3.9, 7.6, 11.387], True),
        ("发射(Space) 收入",    [3.6, 3.8, 4.086], True),
        ("AI 收入",             [3.0, 2.6, 3.201], True),
    ],
    "total_now": None,
    "ni": [-4.628, 0.791, -4.937], "ni_now": None,
    "eq": [4.0, 4.863, 2.573], "eq_now": None,
    "shares": [SHARES_M, SHARES_M, SHARES_M], "shares_now": SHARES_M,
    "px_end": [None, None, None],          # 私营无历史股价
    "px_now": PRICE,
    "band_note": "SpaceX 2026-06-12 才上市; 历史年份为私营期, 无市场股价 → P/E·P/B 历史列不适用(亏损年 N/M); EV/Sales 行(下方)给实数底座。",
    "notes": [
        ("Starlink(连接) 收入", "连接(Connectivity)分部=Starlink 消费+企业+航空+海事+军用; 2025 营收 $11.4B(+50% YoY), Adj EBITDA $7.2B(margin 63%), 集团唯一现金牛。"),
        ("发射(Space) 收入", "发射(Space)分部=对外发射+政府长约; 2025 营收 $4.1B, 经营亏损约 $0.7B(含约 $3B Starship R&D)。"),
        ("AI 收入", "AI 分部=算力租赁+Grok+X 广告+Cursor; 2025 营收 $3.2B 但经营亏损约 $6.4B、吃集团约 60% capex, 是集团亏损主因。"),
        ("HREV", "总营收=S-1/A#2 损益表实际; 2025 合并体含 xAI+X(2026-02 共同控制下重组追溯并表)。"),
        ("HNI", "净利: S-1/A#2 实际; 2024 微利 $0.79B, 2023/2025 亏损(2025 亏 $4.9B 主因 AI R&D 与 SBC 跳升)。"),
        ("HEQ", "股东权益: S-1/A#2 资产负债表; 2025 普通股权益仅 $2.6B(可赎回优先股 $38.8B 不计入)。"),
        ("HSH", "股本: IPO 后摊薄口径 ~13,076 mn 股(2026-05 做 5 拆 1; A 类 7,380 + B 类 5,696)。"),
        ("HPX", "私营期无市场股价; 当下=IPO 后现价 $160.95。"),
    ],
})
SL_HROW = ha["seg_rows"]["Starlink(连接) 收入"]
SP_HROW = ha["seg_rows"]["发射(Space) 收入"]
AI_HROW = ha["seg_rows"]["AI 收入"]

# 历史财务尾部补一段 EV/Sales 实数行(亏损标的主镜头必须有实数底座)──────────
wh = wb[S_HIST]
NOTE_H = "H"
er = wh.max_row + 2
K.band(wh, er, "EV / Sales 底座(亏损标的主镜头; P/E·P/B 亏损年 N/M 不可用)", len(HC) + 3); er += 1
# 当下市值 / 净现金 / EV
R_HMCAP = er
K.lab(wh, f"A{er}", "市值($B, 当下)", b=True)
K.fml(wh, f"G{er}", f"=G{ha['HPX']}*G{ha['HSH']}/1000", K.N0); wh[f"G{er}"].fill = K.CUR
K.logic(wh, f"{NOTE_H}{er}", "= 现价 $160.95 × 13,076 mn 股 ≈ $2,104B(companiesmarketcap / IPO 后摊薄)。")
er += 1
R_HNC = er
K.lab(wh, f"A{er}", "净现金($B)")
K.inp(wh, f"G{er}", NET_CASH, None, K.N1)
K.logic(wh, f"{NOTE_H}{er}", "= 现金 $90.3B − 含租赁总负债 $22.9B = $67.4B(IPO 募资净额多为已承诺 2026E capex)。")
er += 1
R_HEV = er
K.lab(wh, f"A{er}", "企业价值 EV($B, 当下)", b=True); wh[f"A{er}"].fill = K.OUT
K.fml(wh, f"G{er}", f"=G{R_HMCAP}-G{R_HNC}", K.N0)
K.logic(wh, f"{NOTE_H}{er}", "= 市值 − 净现金(净现金为正 → EV < 市值)。")
er += 1
R_HEVS = er
K.lab(wh, f"A{er}", "EV / Sales(当下, TTM)", b=True); wh[f"A{er}"].fill = K.OUT
# 历史年 EV/Sales: 私营无市价 → 标 n.m.; 当下=EV ÷ 2025 营收
for col in HC:
    K.lab(wh, f"{col}{er}", "n.m.", note=True)
K.fml(wh, f"G{er}", f"=G{R_HEV}/D{ha['HREV']}", K.MX); wh[f"G{er}"].fill = K.CUR
K.logic(wh, f"{NOTE_H}{er}", "当下 EV/Sales = EV ÷ 2025 营收 ≈ 109x — 远高于电信(2-4x)、超大规模云(8-12x), 反映市场把远期增长全提前定价。"
                            "私营期无公开市价, 历史年 EV/Sales 标 n.m.。")

# ════════════ 3. 股价走势 ════════════
px = K.write_price_chart(wb.create_sheet(S_PX), [("2026-06", 135), ("2026-06", 161)], {
    "rows": [
        ("上市仅数日", "2026-06-12 IPO 定价 $135, 首日收涨约 +19%; 2026-06-14 报 $160.95。"),
        ("无历史估值带", "私营期(2002-2026)无公开股价, 无月度历史 → 不构建历史估值带; 估值锚定逐年 SOTP×EV/Sales 阶梯, 不依赖历史 multiple。"),
    ],
}, title="SpaceX 股价 — 上市初期(无历史带)")

# ════════════ 4. 卖方研报共识 ════════════
K.write_consensus(wb.create_sheet(S_CONS), {
    "title": "卖方研报共识 — 模型每个关键假设的卖方对账单",
    "overview": ("Oppenheimer 跑赢(OUTPERFORM)、目标价 $190; 街均值约 $164、区间 $63-$227, 极度分化。"
                 "多头叙事=优质资产 + 10 年 DCF 把 2035 远期愿景(2035E 营收 $885B / EBITDA $500B)近乎不折现地搬到今天; "
                 "我们的分歧=对 2026E-2030E 逐年用当年 forward EV/Sales 自立估值(不把远期硬贴现成一个点), 并给轨道算力散热极低物理概率。"),
    "assumptions": [
        ("Starlink 逐年营收 / EV/Sales",
         "多数认现金牛续强; Oppenheimer 路径偏陡。",
         "ARPU 是企稳还是继续下探(blended rev/sub 已从 2024 $189 跌到 2025 $143)。",
         "Base 2030E $52B(EV/Sales 6x, 由 66% margin × ~9x EV/EBITDA 反推); EV/Sales 逐年从 10x normalize 到 6x; 分歧最小的一段。"),
        ("AI 逐年营收 / 盈利时点",
         "Oppenheimer 给陡峭轨迹, 2027 前后 AI 超 Starlink。",
         "AI 分部 2025 仍亏 $6.4B; 转正时点与轨道算力可行性是最大不确定。",
         "Base 2030E $90B(EV/Sales 7x, 高增长期 14x 逐年 normalize); Bull $160B / Bear $40B, 跨度即估值主摆幅。"),
        ("Starship 商业化(发射段)",
         "牛市假设 2027 商用 + $100/kg。",
         "Starship V3 刚炸 + FAA 停飞 → 近期兑现概率被高估。",
         "Base 发射 2030E $12B(EV/Sales 5x); 下调近期兑现概率, Bull 才给 $30B / 8x。"),
        ("轨道算力(太空数据中心)可行性",
         "牛市叙事核心: 100GW 路径启动。",
         "太空散热(辐射散热 >1000 W/㎡ 稳定运行)是否物理可行=未证伪也未证实。",
         "给极低概率: 仅计入期权 PV(Base 2030E $70B), 不进 AI 主营 — 这是我们与多头的最大分歧。"),
        ("分部目标 EV/Sales(逐年)",
         "多头隐含倍数极高(把 2035 远期当现值)。",
         "给现金牛(Starlink)和亏损成长段(AI)同一高倍数是否合理。",
         "分段逐年给: Starlink 10→6x、发射 7→5x、AI 14→7x(高增长期高、随成熟 normalize); 见『估值倍数假设』。"),
    ],
    "divergences": [
        "① 轨道算力散热的物理可行性: 决定 AI 分部是 $40B 还是 $160B、期权 PV 是 $0 还是 $200B — 隐含价最大摆幅。",
        "② 估值口径: 市场把 2035 愿景近乎不折现搬到今天, 我们逐年(2026E-2030E)用当年 forward EV/Sales 自立, 看公司哪一年才把估值'长'到现价。",
        "③ Starship 近期兑现概率: V3 刚炸 + FAA 停飞 → 近期发射/期权兑现概率下调。",
    ],
    "stances": [
        "Oppenheimer(跑赢, TP $190): 10 年 DCF(WACC 10.3% / g 5%), 隐含公司价值约 $2.5T, 2035E 营收 $885B / EBITDA $500B。",
        "Morgan Stanley(2026 Space Outlook, 2026-01): 看好发射+Starlink 长期 TAM, 偏成长叙事。",
        "Bernstein(Future of Tech: Space Tech, 2026-03): 强调太空 TAM 但提示执行/物理风险。",
        "街均值 TP $164(高 $227 / 低 $63): 区间极宽=对 AI/轨道算力分歧巨大。",
    ],
})

# ════════════ 5. 历史估值倍数(上市定价倍数 + 同业对照, 手搓)════════════
wsm = wb.create_sheet(S_HMULT)
K.hdr(wsm, 1, "历史估值倍数 — 上市定价倍数 + 同业 EV 倍数对照(无历史带)", 8)
hr = K.mtext(wsm, 2, ("SpaceX 私营期无公开股价 → 无自身历史估值带。本页改为两件事: "
                      "① 现价对应的当下定价倍数(EV/Sales、EV/EBITDA), 作为'市场现在给多少'的观测点; "
                      "② 同业 EV 倍数光谱(电信 / AI 基建 / 卫星通信 / 超大规模云), 看现价倍数落在同业什么位置。"
                      "估值不锚历史 multiple, 锚逐年(2026E-2030E)分部 EV/Sales(后页)。"), "H", 2)
for col, w in zip("ABCDEFGH", [26, 13, 13, 13, 13, 11, 11, 11]):
    wsm.column_dimensions[col].width = w
wsm.column_dimensions["H"].width = 56

K.band(wsm, hr, "① SpaceX 当下定价倍数(现价 $160.95 → EV)", 8); hr += 1
for col, h in zip(["A", "B", "C", "D"], ["项目", "数值", "口径", "说明"]):
    wsm[f"{col}{hr}"] = h; wsm[f"{col}{hr}"].font = K.BF; wsm[f"{col}{hr}"].fill = K.CH
wsm.merge_cells(f"D{hr}:H{hr}"); hr += 1
R_MCAP = hr
K.lab(wsm, f"A{hr}", "市值($B)"); K.inp(wsm, f"B{hr}", 2104.0, None, K.N0)
K.lab(wsm, f"C{hr}", "现价 × 股本", note=True)
wsm.merge_cells(f"D{hr}:H{hr}"); K.logic(wsm, f"D{hr}", "= $160.95 × 13,076 mn 股 ≈ $2,104B(companiesmarketcap / IPO 后摊薄)。"); hr += 1
R_NETDEBT = hr
K.lab(wsm, f"A{hr}", "净现金($B)"); K.inp(wsm, f"B{hr}", NET_CASH, None, K.N1)
K.lab(wsm, f"C{hr}", "现金−含租赁负债", note=True)
wsm.merge_cells(f"D{hr}:H{hr}"); K.logic(wsm, f"D{hr}", "= 现金 $90.3B − 含租赁总负债 $22.9B = $67.4B(IPO 募资净额 $74.4B 多为已承诺 2026E capex)。"); hr += 1
R_EV = hr
K.lab(wsm, f"A{hr}", "企业价值 EV($B)", b=True); wsm[f"A{hr}"].fill = K.OUT
K.fml(wsm, f"B{hr}", f"=B{R_MCAP}-B{R_NETDEBT}", K.N0)
K.lab(wsm, f"C{hr}", "市值 − 净现金", note=True)
wsm.merge_cells(f"D{hr}:H{hr}"); K.logic(wsm, f"D{hr}", "= 市值 − 净现金(净现金为正 → EV < 市值)。"); hr += 1
R_REV25 = hr
K.lab(wsm, f"A{hr}", "2025A 营收($B)")
K.fml(wsm, f"B{hr}", f"={K.R(S_HIST, 'D' + str(ha['HREV']))}", K.N1, link=True)
K.lab(wsm, f"C{hr}", "引历史财务", note=True)
wsm.merge_cells(f"D{hr}:H{hr}"); K.logic(wsm, f"D{hr}", "2025 合并营收(连接+发射+AI), 引『历史财务与估值』。"); hr += 1
R_EBITDA25 = hr
K.lab(wsm, f"A{hr}", "2025A Adj EBITDA($B)"); K.inp(wsm, f"B{hr}", 6.584, None, K.N1)
K.lab(wsm, f"C{hr}", "S-1/A#2", note=True)
wsm.merge_cells(f"D{hr}:H{hr}"); K.logic(wsm, f"D{hr}", "集团 Adj EBITDA $6.6B(连接 $7.2B + 发射 $0.7B − AI $1.2B)。"); hr += 1
K.lab(wsm, f"A{hr}", "EV / Sales (当下, TTM)", b=True)
K.fml(wsm, f"B{hr}", f"=B{R_EV}/B{R_REV25}", K.MX)
K.lab(wsm, f"C{hr}", "EV ÷ 2025 营收", note=True)
wsm.merge_cells(f"D{hr}:H{hr}"); K.logic(wsm, f"D{hr}", "现价对应 EV/Sales ≈ 109x — 远高于电信(2-4x)、超大规模云(8-12x), 接近高增 AI 基建上沿, 反映市场把远期增长全提前定价。"); hr += 1
K.lab(wsm, f"A{hr}", "EV / EBITDA (当下, TTM)", b=True)
K.fml(wsm, f"B{hr}", f"=B{R_EV}/B{R_EBITDA25}", K.MX)
K.lab(wsm, f"C{hr}", "EV ÷ 2025 EBITDA", note=True)
wsm.merge_cells(f"D{hr}:H{hr}"); K.logic(wsm, f"D{hr}", "现价对应 EV/EBITDA ≈ 310x — 当下盈利无法支撑现价, 估值完全由 2030E+ 远期盈利兑现支撑。"); hr += 1
hr += 1

K.band(wsm, hr, "② 同业 EV 倍数对照(给『估值倍数假设』分段逐年 EV/Sales 定锚)", 8); hr += 1
for col, h in zip(["A", "B", "C", "D"], ["可比公司/板块", "EV/EBITDA", "EV/Sales", "业务特征 / 取数说明"]):
    wsm[f"{col}{hr}"] = h; wsm[f"{col}{hr}"].font = K.BF; wsm[f"{col}{hr}"].fill = K.CH
wsm.merge_cells(f"D{hr}:H{hr}"); hr += 1
peers = [
    ("电信运营商(AT&T/Verizon)", 7.0, 2.5, "成熟现金牛、低增长 → Starlink 现金牛 EV/Sales 的下沿锚(成熟态 5-6x 给溢价反映 50% 增速)。"),
    ("卫星通信(ASTS 等)", None, 25.0, "高增长、未盈利 → EV/Sales 高、EV/EBITDA 无意义; 对照 Starlink 移动/D2C 期权与发射成长段。"),
    ("超大规模云(AWS/Azure 隐含)", 14.0, 9.0, "AI 算力租赁可比 → AI 分部 EV/Sales(Base 高增长期 14x)在云基建与卫星成长之间。"),
    ("Nvidia(AI 算力光谱上沿)", 30.0, 22.0, "AI 基建估值上沿; Bull 给 AI 早年 18x 仍低于 Nvidia, 反映租赁模式确定性高于芯片周期。"),
]
for name, evb, evs, note in peers:
    wsm[f"A{hr}"] = name; wsm[f"A{hr}"].font = K.BF
    if evb is not None:
        K.inp(wsm, f"B{hr}", evb, None, K.MX)
    else:
        K.lab(wsm, f"B{hr}", "n.m.", note=True)
    K.inp(wsm, f"C{hr}", evs, None, K.MX)
    wsm.merge_cells(f"D{hr}:H{hr}"); K.logic(wsm, f"D{hr}", note); hr += 1
hr += 1
K.band(wsm, hr, "③ 读法 — 给『估值倍数假设』的输入", 8); hr += 1
K.mtext(wsm, hr, ("当下 EV/Sales ≈109x、EV/EBITDA ≈310x = 现价把 2030E+ 远期盈利近乎全额提前定价。"
                  "分段 EV/Sales 逐年据同业光谱拍, 且高增长期高、随成熟 normalize: "
                  "Starlink 10→6x(向电信成长上沿收敛)、发射 7→5x(成长+政府长约)、"
                  "AI 14→7x(高增长期靠云基建上沿, 成熟后向云收敛)。下一页用这些拍三案逐年倍数。"), "H", 4)

# ════════════ 6. 估值倍数假设(逐年 EV/Sales 三案矩阵, 手搓)════════════
wmu = wb.create_sheet(S_MULT)
K.hdr(wmu, 1, "估值倍数假设 — 三分部逐年(2026E-2030E)EV/Sales 三案矩阵", 9)
mr = K.mtext(wmu, 2, ("主镜头=EV/Sales(整体+AI 段亏损, EV/EBITDA 会出负/NA, 故用销售倍数出实数)。"
                      "本页只拍各分部逐年(2026E-2030E)EV/Sales 三案(蓝字)与理由; "
                      "『情景切换』引用并切换当前案, 『情景估值』逐年套用, 『估值对比』三案逐年并排。"
                      "倍数据上一页同业光谱 + 由目标 EV/EBITDA × 稳态 margin 反推; 不锚自身历史(SpaceX 无历史带)。"), "I", 2)
for col, w in zip("ABCDEFGHI", [22, 9, 9, 9, 9, 9, 11, 11, 11]):
    wmu.column_dimensions[col].width = w
wmu.column_dimensions["L"].width = 62

# 每分部: 组标题 + 三案行(E-I 逐年) ; 记录三案首行
MULT_ROWS = {}   # 分部key -> {"Bear":row,...}
def write_mult_block(seg_key, seg_name, data_dict, case_notes):
    global mr
    K.band(wmu, mr, f"{seg_name} 逐年 EV/Sales (x)", 9); mr += 1
    # 年份头
    wmu[f"A{mr}"] = "案 / 年份"; wmu[f"A{mr}"].font = K.BF
    for col, y in zip(FCf, FWY):
        wmu[f"{col}{mr}"] = y; wmu[f"{col}{mr}"].font = K.BF; wmu[f"{col}{mr}"].fill = K.CH
    wmu[f"L{mr}"] = "为什么这么给"; wmu[f"L{mr}"].font = K.BF; wmu[f"L{mr}"].fill = K.CH
    mr += 1
    rows = {}
    for cs in CASES:
        K.lab(wmu, f"A{mr}", f"  {cs}")
        K.introw(wmu, mr, FCf, data_dict[cs], None, K.MX)
        K.logic(wmu, f"L{mr}", case_notes[cs])
        rows[cs] = mr
        mr += 1
    MULT_ROWS[seg_key] = rows
    mr += 1

write_mult_block("sl", "Starlink", SL_PS, {
    "Bear": "现金牛但 ARPU 下探: 由成熟态目标 EV/EBITDA × 较低稳态 margin 反推, 7x 起步快速 normalize 到 5x(电信下沿)。",
    "Base": "由目标 EV/EBITDA(成熟约 9x)× 稳态 EBITDA margin 66% ≈ 6x 反推 2030E 终值; 高增长期(2026 50% 增速)给 10x, 随成熟逐年 normalize 到 6x。下沿锚电信成长段。",
    "Bull": "移动/D2C 打开 → 更高且更持久: 14x 起步 normalize 到 10x, 反映持续高增速 + 频谱期权进主营。"})
write_mult_block("sp", "发射", SP_PS, {
    "Bear": "Starship 推迟、仅 Falcon: 5x 起步快速 normalize 到 4x(发射量摊薄、降本未兑现)。",
    "Base": "成长 + 政府长约确定性: 由发射段稳态 margin 24% × EV/EBITDA 反推; 早年 7x(运力扩张), 随发射量摊薄/成熟 normalize 到 5x。",
    "Bull": "Starship $100/kg 兑现打开成本结构 → 倍数更高更持久: 10x 起步 normalize 到 8x。"})
write_mult_block("ai", "AI", AI_PS, {
    "Bear": "轨道算力证伪、只剩低利润租赁: 全程低位 6x→4x。",
    "Base": "云基建与卫星成长之间: 高增长期(2026 14x)对标超大规模云上沿/卫星成长, 随规模化 normalize 到 7x(向云基建 9x 下方收敛)。",
    "Bull": "陡峭轨迹 + 轨道算力规模化: 18x 起步(仍低于 Nvidia 22x, 反映租赁确定性), normalize 到 10x。"})

# 期权 PV 逐年三案
K.band(wmu, mr, "期权价值 PV(逐年, 三案; 不进分部营收 × 倍数, 单列加总)", 9); mr += 1
wmu[f"A{mr}"] = "案 / 年份"; wmu[f"A{mr}"].font = K.BF
for col, y in zip(FCf, FWY):
    wmu[f"{col}{mr}"] = y; wmu[f"{col}{mr}"].font = K.BF; wmu[f"{col}{mr}"].fill = K.CH
wmu[f"L{mr}"] = "说明"; wmu[f"L{mr}"].font = K.BF; wmu[f"L{mr}"].fill = K.CH
mr += 1
OPT_ROWS = {}
opt_notes = {
    "Bear": "轨道算力被物理证伪(散热不可行)→ 期权价值归零, 全程 $0。",
    "Base": "轨道算力(太空数据中心)+ Starship 深空 + 移动频谱期权, 小规模验证、随兑现进度逐年抬升 $40B→$70B。",
    "Bull": "100GW 路径启动 → 期权大幅放量 $80B→$200B(轨道算力规模化兑现)。",
}
for cs in CASES:
    K.lab(wmu, f"A{mr}", f"  {cs}")
    K.introw(wmu, mr, FCf, OPT[cs], None, K.N0)
    K.logic(wmu, f"L{mr}", opt_notes[cs])
    OPT_ROWS[cs] = mr
    mr += 1
mr += 1
K.band(wmu, mr, "卖方对账 + 数据源", 9); mr += 1
K.mtext(wmu, mr, ("凭什么敢给非主流数: Oppenheimer 把 2035 远期当现值、隐含倍数极高; 我们逐年给当年 forward EV/Sales, "
                  "每段倍数据 a)目标 EV/EBITDA × 稳态 margin 反推 + b)同业 EV/Sales 光谱(上一页)双重定锚, 不靠缺少依据, 且高增长期高、随成熟 normalize。"
                  "数据源: EV/Sales=同业光谱 + 盈利反推; 期权 PV=轨道算力可行性概率 × 潜在 TAM 的粗估。"), "I", 4)

# ════════════ 7. 情景切换(全模型唯一情景参数库; 逐年杠杆)════════════
# 杠杆全部逐年(cols=FCf, vals 给 5 值): 三分部营收 + 三分部 EV/Sales + 期权。
# EV/Sales 与期权三案行链引『估值倍数假设』矩阵 → 防重复硬编码; 营收三案直接在此拍。
sw = K.write_scenario_switch(wb.create_sheet(S_SW), {
    "title": "情景切换 — 全模型唯一情景参数库 + 切换开关(默认 Base)",
    "usage": ("怎么用: B2 是唯一入口——下拉选案 → 案序号派生 → 各杠杆『当前案』行跟着切 → "
              "分部测算/利润/逐年 SOTP 估值整条链变档, 『情景估值』输出该案逐年隐含价阶梯。"
              "三案对比不用切: 『估值对比』恒常三案并排(引本页矩阵行)。净现金/股本三案共用。"),
    "cases": CASES, "default": "Base",
    "triggers": [
        ("Bear", "轨道算力被物理证伪(散热不可行); Starship 商业化推迟到 2030+; AI 维持亏损只剩租赁; Starlink ARPU 继续下探、移动战略受阻。"),
        ("Base", "Starlink 续强但 ARPU 仅企稳; 发射稳健、Starship 2028 前后渐入商用; AI 靠租赁+Cursor 放量 2027 转正, 轨道算力小规模验证未规模化。"),
        ("Bull", "Starship 2027 商业化 + $100/kg 兑现; 轨道算力规模化(100GW 路径启动); AI 2027 超 Starlink; Starlink 靠 EchoStar 频谱拿下移动通信。"),
    ],
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "levers": [
        {"key": "sl_rev", "name": "Starlink 营收逐年($B)", "fmt": K.N0, "cols": FCf,
         "vals": SL_REV,
         "desc": "Starlink 逐年营收。物理锚=平均用户 × blended rev/sub; 分歧在 ARPU 企稳 vs 下探 + 移动/D2C 是否打开。",
         "stories": {"Bear": "ARPU 继续下探, 移动受阻 → 2030E $38B。", "Base": "用户续增、ARPU 企稳 → 2030E $52B。", "Bull": "EchoStar 频谱拿下移动 → 2030E $70B。"}},
        {"key": "sp_rev", "name": "发射 营收逐年($B)", "fmt": K.N0, "cols": FCf,
         "vals": SP_REV,
         "desc": "发射逐年营收。物理锚=对外发射次数 × 单次收入 + 政府长约; Bull 假设 Starship 商用打开运力与降本。",
         "stories": {"Bear": "Starship 推迟到 2030+, 仅 Falcon → 2030E $7B。", "Base": "Starship 2028 前后渐入商用 → 2030E $12B。", "Bull": "Starship 2027 商用 + $100/kg → 2030E $30B。"}},
        {"key": "ai_rev", "name": "AI 营收逐年($B)", "fmt": K.N0, "cols": FCf,
         "vals": AI_REV,
         "desc": "AI 逐年营收。物理锚=算力 GW → 租赁合同 + Grok + 广告 + Cursor; 跨度最大(估值主摆幅)。",
         "stories": {"Bear": "轨道算力证伪, 只剩租赁 → 2030E $40B。", "Base": "租赁+Cursor 放量 2027 转正 → 2030E $90B。", "Bull": "陡峭轨迹 2027 超 Starlink → 2030E $160B。"}},
    ],
    "linked": [
        {"key": "sl_ps", "name": "Starlink EV/Sales 逐年(x)", "fmt": K.MX,
         "src_sheet": S_MULT, "src_row0": None, "note": ""},  # 占位, 下面手搓(矩阵在 E-I 列, kit linked 走 ALLC 不匹配)
    ] if False else [],
})

SC_ws = wb[S_SW]
SWIDX = sw["idx_cell"]
SWB = sw["SWB"]; SWACT = sw["SWACT"]
_r = sw["next_row"]

# EV/Sales + 期权逐年 lever 手搓(三案行链引『估值倍数假设』E-I 列, 当前案 INDEX)──────
def add_linked_yearly(name, key, mult_rows_dict, fmt, note):
    """三案行逐年(E-I)链引『估值倍数假设』对应案行; 当前案 INDEX。"""
    global _r
    K.lab(SC_ws, f"A{_r}", name, b=True)
    for _c in ["A"] + list(ALLC):
        SC_ws[f"{_c}{_r}"].fill = K.GREYF
    K.logic(SC_ws, f"L{_r}", note)
    _r += 1
    base0 = _r
    for cs in CASES:
        K.lab(SC_ws, f"A{_r}", f"  {cs}")
        src_row = mult_rows_dict[cs]
        for col in FCf:
            K.fml(SC_ws, f"{col}{_r}", f"={K.R(S_MULT, col + str(src_row))}", fmt, link=True)
        _r += 1
    K.lab(SC_ws, f"A{_r}", "  当前案(链引此行)", note=True)
    for col in FCf:
        K.fml(SC_ws, f"{col}{_r}", f"=INDEX({col}{base0}:{col}{base0+2},{SWIDX})", fmt)
    act = _r
    SWB[key] = base0; SWACT[key] = act
    _r += 2

add_linked_yearly("Starlink EV/Sales 逐年(x)", "sl_ps", MULT_ROWS["sl"], K.MX,
    "Starlink 逐年 EV/Sales, 链引『估值倍数假设』; 改去那页改蓝字。高增长期高、随成熟 normalize。")
add_linked_yearly("发射 EV/Sales 逐年(x)", "sp_ps", MULT_ROWS["sp"], K.MX,
    "发射 逐年 EV/Sales, 链引『估值倍数假设』; 改去那页改蓝字。")
add_linked_yearly("AI EV/Sales 逐年(x)", "ai_ps", MULT_ROWS["ai"], K.MX,
    "AI 逐年 EV/Sales, 链引『估值倍数假设』; 改去那页改蓝字。AI 段亏损故用 EV/Sales 不用 EV/EBITDA。")
add_linked_yearly("期权 PV 逐年($B)", "opt", OPT_ROWS, K.N0,
    "轨道算力+Starship+移动频谱期权价值, 逐年链引『估值倍数假设』; 不进分部营收×倍数, 单列加总。")

# 全局参数(三案共用; 逐年 forward 不折现, 无 WACC/折现因子)──────
K.band(SC_ws, _r, "全局参数(三案共用; 逐年 forward 口径, 不折现)", 9); _r += 1
gp = [("净现金($B)", NET_CASH, K.N1, "现金 $90.3B − 含租赁负债 $22.9B; 加到各年 EV 求权益。"),
      ("股本(mn股)", SHARES_M, K.N0, "IPO 后摊薄。"),
      ("现价($)", PRICE, K.PX, "2026-06-14。")]
GLOBAL_ROWS = {}
for nm, v, fmt, note in gp:
    K.lab(SC_ws, f"A{_r}", nm, b=True)
    K.inp(SC_ws, f"B{_r}", v, None, fmt)
    K.logic(SC_ws, f"L{_r}", note)
    GLOBAL_ROWS[nm] = _r
    _r += 1

# 期末锚: 各营收 lever 当前案行(逐年 E-I 列)
SL_REV_ACT = SWACT["sl_rev"]; SP_REV_ACT = SWACT["sp_rev"]; AI_REV_ACT = SWACT["ai_rev"]
SL_PS_ACT = SWACT["sl_ps"]; SP_PS_ACT = SWACT["sp_ps"]; AI_PS_ACT = SWACT["ai_ps"]
OPT_ACT = SWACT["opt"]

# ════════════ 8. 物理锚 [ANCHOR] ════════════
anchor = K.write_anchor(wb.create_sheet(S_ANCHOR), {
    "title": "SpaceX 物理锚 — Starlink 用户 / Falcon 发射 / AI 算力",
    "all_cols": ALLC, "all_years": ALLY,
    "series": [
        ("Starlink 平均订阅用户(M)",
         [3.4, 6.7, 9.6, 12.5, 16.0, 20.0, 24.0, 28.0],
         "年内平均用户(年末 2023 2.3M→2024 4.4M→2025 8.9M→2026Q1 10.3M; 取年均); 前瞻按渗透曲线外推。", K.N1),
        ("Falcon 对外发射次数",
         [96, 132, 165, 185, 200, 215, 230, 245],
         "对外发射次数(可观测物理量, 每次有公开记录; Space.com/Payload 统计); 内部 Starlink 发射不计对外收入。", K.N0),
        ("AI 算力(GW)",
         [0.5, 1.0, 2.0, 4.0, 6.0, 9.0, 13.0, 18.0],
         "AI 数据中心装机算力(Colossus 555k GPU/2GW 为 2025 锚); 前瞻=采购/功率披露外推, Bull 路径指向 100GW。", K.N1),
    ],
    "yoy_row": "Starlink 平均订阅用户(M)",
    "source_note": ("口径: Starlink=平均订阅用户(变现直接挂用户, 优于卫星数); 发射=对外发射次数(运力=营收物理上限); "
                    "AI=装机算力 GW(供给驱动收入)。来源: S-1/A#2 KPI 表 + Quilty Space / Space.com / Payload + 公司功率披露。"),
    "role_note": ("作用: 三条物理锚各驱动一个分部——Starlink 用户 → 连接营收; 发射次数 → 发射营收; "
                  "算力 GW → AI 营收。改锚 → 分部营收 → 逐年 SOTP 隐含价全链动。"),
})
A_SL_SUB = anchor["row_of"]["Starlink 平均订阅用户(M)"]
A_LAUNCH = anchor["row_of"]["Falcon 对外发射次数"]
A_AIGW   = anchor["row_of"]["AI 算力(GW)"]

# ════════════ 9. 分部测算 ════════════
seg = K.write_segment_model(wb.create_sheet(S_SEG), {
    "title": "分部测算 — Starlink(用户×rev/sub) / 发射(次数×单价) / AI(算力×收入强度)",
    "all_cols": ALLC, "all_years": ALLY, "logic_col": "K",
    "groups": [
        ("Starlink(连接): 平均用户 × blended 月均 rev/sub", [
            ("Starlink 平均用户(M)", None, K.N1, "= 引『物理锚』Starlink 平均订阅用户。"),
            ("blended rev/sub ($/月)", [123, 189, 143, 140, 142, 145, 150, 155], K.N0,
             "总连接营收/平均用户(含硬件+企业+航空+海事+军用); 2024 $189→2025 $143; 前瞻列由当年情景营收 ÷ (用户×12) 反推隐含 rev/sub。"),
            ("Starlink 营收($B)", None, K.N1, "= 平均用户(M) × rev/sub($/月) × 12 / 1000; 喂『利润与收入假设』。"),
        ]),
        ("发射(Space): 对外发射次数 × 单次收入 + 政府长约", [
            ("对外发射次数", None, K.N0, "= 引『物理锚』Falcon 对外发射次数。"),
            ("单次综合收入($M)", [37.5, 28.8, 24.8, 27, 31, 36, 42, 49], K.N1,
             "= 发射营收/次数反推(含政府长约摊销); 前瞻列由当年情景营收 ÷ 次数 反推隐含单价。"),
            ("发射 营收($B)", None, K.N1, "= 对外发射次数 × 单次收入($M) / 1000。"),
        ]),
        ("AI: 装机算力 × 收入强度", [
            ("AI 算力(GW)", None, K.N1, "= 引『物理锚』AI 算力。"),
            ("AI 收入强度($B/GW)", [6.0, 2.6, 1.6, 2.0, 2.8, 3.5, 4.2, 5.0], K.N1,
             "= AI 营收/算力 GW(租赁+Grok+广告+Cursor 合计÷装机); 前瞻列由当年情景营收 ÷ 算力 GW 反推隐含强度。"),
            ("AI 营收($B)", None, K.N1, "= 算力(GW) × 收入强度($B/GW)。"),
        ]),
    ],
})
m = seg["m"]

# 物理锚引用行
for col in ALLC:
    K.fml(wb[S_SEG], f"{col}{m['Starlink 平均用户(M)']}", f"={K.R(S_ANCHOR, col + str(A_SL_SUB))}", K.N1, link=True)
    K.fml(wb[S_SEG], f"{col}{m['对外发射次数']}",       f"={K.R(S_ANCHOR, col + str(A_LAUNCH))}", K.N0, link=True)
    K.fml(wb[S_SEG], f"{col}{m['AI 算力(GW)']}",        f"={K.R(S_ANCHOR, col + str(A_AIGW))}", K.N1, link=True)

# 前瞻列(E-I)单位驱动 = 当年情景营收 ÷ 物理量(隐含强度), 使收入 = 物理量×隐含强度 = 情景营收;
# 公式同时含物理锚物理量项 + 当年情景营收项 → 逐年隐含价既挂物理锚, 又随情景每年营收蓝字联动。
for ci, col in enumerate(FCf):
    K.fml(wb[S_SEG], f"{col}{m['blended rev/sub ($/月)']}",
          f"={K.R(S_SW, col + str(SL_REV_ACT))}*1000/({col}{m['Starlink 平均用户(M)']}*12)", K.N0, link=True)
    K.fml(wb[S_SEG], f"{col}{m['单次综合收入($M)']}",
          f"={K.R(S_SW, col + str(SP_REV_ACT))}*1000/{col}{m['对外发射次数']}", K.N1, link=True)
    K.fml(wb[S_SEG], f"{col}{m['AI 收入强度($B/GW)']}",
          f"={K.R(S_SW, col + str(AI_REV_ACT))}/{col}{m['AI 算力(GW)']}", K.N1, link=True)

# 各分部营收(全年): 物理量 × 单位驱动
for col in ALLC:
    K.fml(wb[S_SEG], f"{col}{m['Starlink 营收($B)']}",
          f"={col}{m['Starlink 平均用户(M)']}*{col}{m['blended rev/sub ($/月)']}*12/1000", K.N1)
    K.fml(wb[S_SEG], f"{col}{m['发射 营收($B)']}",
          f"={col}{m['对外发射次数']}*{col}{m['单次综合收入($M)']}/1000", K.N1)
    K.fml(wb[S_SEG], f"{col}{m['AI 营收($B)']}",
          f"={col}{m['AI 算力(GW)']}*{col}{m['AI 收入强度($B/GW)']}", K.N1)
for key in ["Starlink 营收($B)", "发射 营收($B)", "AI 营收($B)"]:
    for col in ALLC:
        wb[S_SEG][f"{col}{m[key]}"].fill = K.OUT

# ════════════ 10. 利润与收入假设(分部营收 roll-up + EBITDA 交叉验证底座, 手搓)════════════
wfu = wb.create_sheet(S_FUND)
K.hdr(wfu, 1, "利润与收入假设 — 分部营收(SOTP 主底座)+ 分部 EBITDA(EV/EBITDA 交叉验证用)", 9)
fr0 = K.mtext(wfu, 2, ("SOTP 主线走 EV/Sales: 分部逐年营收(引『分部测算』)= 估值主底座。"
                       "另算分部 EBITDA(营收 × margin)供『估值对比』做 EV/EBITDA 交叉验证(仅 EBITDA>0 的分部年份显示, AI 早年亏损标 n.m.)。"
                       "整体净利不进 SOTP(逐年按分部营收×EV/Sales, 不按整体净利×P/E)。"), "I", 2)
for col, w in zip("ABCDEFGHI", [22, 9, 9, 9, 9, 9, 9, 9, 9]):
    wfu.column_dimensions[col].width = w
wfu.column_dimensions["K"].width = 56
HDR_ROW = fr0
for col, y in zip(ALLC, ALLY):
    wfu[f"{col}{HDR_ROW}"] = y; wfu[f"{col}{HDR_ROW}"].font = K.BF; wfu[f"{col}{HDR_ROW}"].fill = K.CH
wfu[f"A{HDR_ROW}"] = "项目"; wfu[f"A{HDR_ROW}"].font = K.BF
wfu[f"K{HDR_ROW}"] = "逻辑/来源"; wfu[f"K{HDR_ROW}"].font = K.BF; wfu[f"K{HDR_ROW}"].fill = K.CH
fr = fr0 + 1
K.band(wfu, fr, "分部营收(引分部测算; 逐年 SOTP 主底座)", 9); fr += 1
FU = {}
seg_link = [("Starlink 营收($B)", m["Starlink 营收($B)"]),
            ("发射 营收($B)",      m["发射 营收($B)"]),
            ("AI 营收($B)",        m["AI 营收($B)"])]
for name, srow in seg_link:
    K.lab(wfu, f"A{fr}", name)
    for col in ALLC:
        K.fml(wfu, f"{col}{fr}", f"={K.R(S_SEG, col + str(srow))}", K.N1, link=True)
    K.logic(wfu, f"K{fr}", f"= 引『{S_SEG}』{name}(物理锚自下而上; 前瞻列由物理量 × 当年情景隐含强度标定到当年情景营收)。")
    FU[name] = fr; fr += 1
R_TOTREV = fr
K.lab(wfu, f"A{fr}", "总营收($B)", b=True); wfu[f"A{fr}"].border = K.BORD
for col in ALLC:
    K.fml(wfu, f"{col}{fr}", f"={col}{FU['Starlink 营收($B)']}+{col}{FU['发射 营收($B)']}+{col}{FU['AI 营收($B)']}", K.N1)
K.logic(wfu, f"K{fr}", "= 三分部营收求和。")
fr += 1
K.lab(wfu, f"A{fr}", "  YoY", note=True)
for i in range(1, len(ALLC)):
    K.fml(wfu, f"{ALLC[i]}{fr}", f"={ALLC[i]}{R_TOTREV}/{ALLC[i-1]}{R_TOTREV}-1", K.PCT)
fr += 1

K.band(wfu, fr, "分部 EBITDA margin(历史实际 + 前瞻; 供 EV/EBITDA 交叉验证)", 9); fr += 1
margin_rows = {}
margin_specs = [
    ("Starlink EBITDA margin", [0.55, 0.60, 0.63, 0.64, 0.645, 0.65, 0.655, SL_MGN],
     "2025 实际约 63%; 前瞻小幅上行至稳态 66%。"),
    ("发射 EBITDA margin", [0.10, 0.12, 0.16, 0.18, 0.20, 0.21, 0.22, SP_MGN],
     "2025 约 16%(含 Starship R&D 拖累); 稳态 24%。"),
    # AI margin 逐年随情景(此处取 Base 路径作为 EBITDA 交叉验证基准; 早年为负 → EV/EBITDA 标 n.m.)
    ("AI EBITDA margin(Base)", [-0.50, -0.30, -0.39] + AI_MGN["Base"],
     "2025 仍亏(约 −39%); 2027 前后转正; 取 Base 路径供 EV/EBITDA 交叉验证(早年负值 → 该段 EV/EBITDA 标 n.m.)。"),
]
for name, vals, note in margin_specs:
    K.lab(wfu, f"A{fr}", name)
    K.introw(wfu, fr, ALLC, vals, None, K.PCT)
    K.logic(wfu, f"K{fr}", note)
    margin_rows[name] = fr; fr += 1

K.band(wfu, fr, "分部 EBITDA($B) = 营收 × margin(EV/EBITDA 交叉验证底座)", 9); fr += 1
ebitda_rows = {}
for seg_name, rev_row_key, mgn_name in [
        ("Starlink EBITDA($B)", "Starlink 营收($B)", "Starlink EBITDA margin"),
        ("发射 EBITDA($B)",      "发射 营收($B)",      "发射 EBITDA margin"),
        ("AI EBITDA($B,Base)",   "AI 营收($B)",        "AI EBITDA margin(Base)")]:
    K.lab(wfu, f"A{fr}", seg_name)
    for col in ALLC:
        K.fml(wfu, f"{col}{fr}", f"={col}{FU[rev_row_key]}*{col}{margin_rows[mgn_name]}", K.N1)
    K.logic(wfu, f"K{fr}", f"= {rev_row_key} × {mgn_name}。")
    ebitda_rows[seg_name] = fr; fr += 1
R_TOTEBITDA = fr
K.lab(wfu, f"A{fr}", "EBITDA 合计($B,Base)", b=True); wfu[f"A{fr}"].fill = K.OUT
for col in ALLC:
    K.fml(wfu, f"{col}{fr}", f"={col}{ebitda_rows['Starlink EBITDA($B)']}+{col}{ebitda_rows['发射 EBITDA($B)']}+{col}{ebitda_rows['AI EBITDA($B,Base)']}", K.N1)
K.logic(wfu, f"K{fr}", "= 三分部 EBITDA 求和; 供 EV/EBITDA 交叉验证(主估值走 EV/Sales)。")
fr += 1

K.band(wfu, fr, "Capex($B)", 9); fr += 1
R_CAPEX = fr
K.lab(wfu, f"A{fr}", "Capex($B)")
K.introw(wfu, fr, ALLC, [None, None, 20.7, 35, 42, 48, 52, 55], None, K.N0)
K.lab(wfu, f"B{fr}", "n.m.", note=True); K.lab(wfu, f"C{fr}", "n.m.", note=True)
K.logic(wfu, f"K{fr}", "2025 实际约 $20.7B(AI 吃约 60%); 前瞻随 AI 装机与 Starship 产能爬坡, IPO 募资多用于此。")
fr += 1
K.mtext(wfu, fr, ("基本面口径: 分部营收(物理锚驱动)= 逐年 SOTP×EV/Sales 主底座; 分部 EBITDA(margin)仅供 EV/EBITDA 交叉验证。"
                  "整体亏损 + AI 段亏损 → 主镜头用 EV/Sales 出实数, 不走整体净利×P/E。"), "I", 2)
SL_EBITDA_ROW = ebitda_rows["Starlink EBITDA($B)"]
SP_EBITDA_ROW = ebitda_rows["发射 EBITDA($B)"]
AI_EBITDA_ROW = ebitda_rows["AI EBITDA($B,Base)"]
# 供下游引用的逐年分部营收行
SL_REVROW = FU["Starlink 营收($B)"]
SP_REVROW = FU["发射 营收($B)"]
AI_REVROW = FU["AI 营收($B)"]

# ════════════ 11. 情景估值(逐年 SOTP×EV/Sales 阶梯, 当前案, 手搓)════════════
wva = wb.create_sheet(S_VAL)
K.hdr(wva, 1, "情景估值 — 当前案逐年(2026E-2030E)SOTP × EV/Sales 隐含价阶梯", 8)
K.lab(wva, "G1", "当前情景→", note=True)
K.fml(wva, "H1", f"={K.R(S_SW, 'B2')}", K.N0, link=True); wva["H1"].fill = K.CUR
vr = K.mtext(wva, 2, ("本表输出=『情景切换』当前案(默认 Base)的逐年隐含价阶梯。每年每分部: 当年营收(引『利润与收入假设』) × 当年 EV/Sales"
                      "(引『情景切换』当前案) = 当年分部 EV → Σ 三分部 EV + 当年期权 PV + 净现金 = 当年权益 → ÷股本×1000 = 当年隐含价 vs 现价。"
                      "不折现(逐年 forward 倍数自立)。下方加 EV/EBITDA 交叉验证(EBITDA>0 才显示)。三案见『估值对比』。"), "H", 3)
for col, w in zip("ABCDEFGH", [24, 12, 12, 12, 12, 12, 12, 12]):
    wva.column_dimensions[col].width = w

NC_REF = K.R(S_SW, "B" + str(GLOBAL_ROWS["净现金($B)"]))
SH_REF = K.R(S_SW, "B" + str(GLOBAL_ROWS["股本(mn股)"]))
PX_REF = K.R(S_SW, "B" + str(GLOBAL_ROWS["现价($)"]))

# 年份头
K.band(wva, vr, "逐年 SOTP(各分部营收 × 当年 EV/Sales = 当年 EV)", 8); vr += 1
wva[f"A{vr}"] = "项目 / 年份"; wva[f"A{vr}"].font = K.BF
for col, y in zip(FCf, FWY):
    wva[f"{col}{vr}"] = y; wva[f"{col}{vr}"].font = K.BF; wva[f"{col}{vr}"].fill = K.CH
vr += 1

VR = {}
def val_row(label, b=False, out=False):
    global vr
    K.lab(wva, f"A{vr}", label, b=b)
    if out:
        wva[f"A{vr}"].fill = K.OUT
    r = vr; vr += 1
    return r

# 三分部: 营收 / EV/Sales / EV
SEG_VAL = [("Starlink", SL_REVROW, SL_PS_ACT), ("发射", SP_REVROW, SP_PS_ACT), ("AI", AI_REVROW, AI_PS_ACT)]
seg_ev_rows = {}
for name, revrow, psact in SEG_VAL:
    rrev = val_row(f"{name} 营收($B)")
    for col in FCf:
        K.fml(wva, f"{col}{rrev}", f"={K.R(S_FUND, col + str(revrow))}", K.N1, link=True)
    rps = val_row(f"{name} EV/Sales(x)")
    for col in FCf:
        K.fml(wva, f"{col}{rps}", f"={K.R(S_SW, col + str(psact))}", K.MX, link=True)
    rev_ = val_row(f"{name} EV($B)")
    for col in FCf:
        K.fml(wva, f"{col}{rev_}", f"={col}{rrev}*{col}{rps}", K.N0)
    seg_ev_rows[name] = rev_

R_SUMEV = val_row("Σ 三分部 EV($B)", b=True)
for col in FCf:
    K.fml(wva, f"{col}{R_SUMEV}", f"={col}{seg_ev_rows['Starlink']}+{col}{seg_ev_rows['发射']}+{col}{seg_ev_rows['AI']}", K.N0)
wva[f"A{R_SUMEV}"].border = K.BORD
R_OPTPV = val_row("+ 期权 PV($B)")
for col in FCf:
    K.fml(wva, f"{col}{R_OPTPV}", f"={K.R(S_SW, col + str(OPT_ACT))}", K.N0, link=True)
R_NC = val_row("+ 净现金($B)")
for col in FCf:
    K.fml(wva, f"{col}{R_NC}", f"={NC_REF}", K.N0, link=True)
R_EQV = val_row("= 权益价值($B)", b=True, out=True)
for col in FCf:
    K.fml(wva, f"{col}{R_EQV}", f"={col}{R_SUMEV}+{col}{R_OPTPV}+{col}{R_NC}", K.N0)
R_IMPLPX = val_row("隐含价($/股)", b=True, out=True)
for col in FCf:
    K.fml(wva, f"{col}{R_IMPLPX}", f"={col}{R_EQV}/{SH_REF}*1000", K.PX, link=True)
R_VSNOW = val_row("vs 现价", b=True)
for col in FCf:
    K.fml(wva, f"{col}{R_VSNOW}", f"={col}{R_IMPLPX}/{PX_REF}-1", K.PCT, link=True)
vr += 1

# EV/EBITDA 交叉验证(仅 EBITDA>0 才显示; 用 Base 路径 EBITDA 行)──────
K.band(wva, vr, "EV/EBITDA 交叉验证(隐含; 仅 EBITDA>0 显示, 否则 n.m.)", 8); vr += 1
wva[f"A{vr}"] = "项目 / 年份"; wva[f"A{vr}"].font = K.BF
for col, y in zip(FCf, FWY):
    wva[f"{col}{vr}"] = y; wva[f"{col}{vr}"].font = K.BF; wva[f"{col}{vr}"].fill = K.CH
vr += 1
xv_segs = [("Starlink", seg_ev_rows["Starlink"], SL_EBITDA_ROW),
           ("发射", seg_ev_rows["发射"], SP_EBITDA_ROW),
           ("AI", seg_ev_rows["AI"], AI_EBITDA_ROW)]
for name, evrow, ebrow in xv_segs:
    K.lab(wva, f"A{vr}", f"{name} 隐含 EV/EBITDA")
    for col in FCf:
        # 仅当 EBITDA>0 才算, 否则 "n.m."
        K.fml(wva, f"{col}{vr}",
              f'=IF({K.R(S_FUND, col + str(ebrow))}>0,{col}{evrow}/{K.R(S_FUND, col + str(ebrow))},"n.m.")',
              K.MX, link=True)
    vr += 1
K.logic(wva, f"A{vr}", "读法: EV/EBITDA = 该分部当年 EV ÷ 当年 EBITDA(Base margin)。AI 早年 EBITDA<0 → n.m., 这正是主镜头用 EV/Sales 的原因。")
wva.merge_cells(f"A{vr}:H{vr}")
vr += 2
K.band(wva, vr, "方法与结论", 8); vr += 1
K.mtext(wva, vr, ("方法: 逐年(2026E-2030E)SOTP × EV/Sales — 各分部当年营收 × 当年 forward EV/Sales = 当年 EV, "
                  "Σ + 当年期权 PV + 净现金 = 权益, ÷股本 = 当年隐含价。不折现(每年用当年 forward 倍数自立)。"
                  "整体+AI 段亏损故主镜头 EV/Sales(出实数); EV/EBITDA 仅做交叉验证。"
                  "结论(Base): 隐含价 2026E≈$39 → 2027E≈$55 → 2030E≈$87, 全程低于现价 $161 — 现价 price-in 的不是 Base。"
                  "12M 目标价取 2027E 概率加权; 三情景见『估值对比』。"), "H", 4)

# ════════════ 12. 估值对比(三案各一逐年 block, 手搓; 防污染=只引矩阵行/全局)════════════
wcp = wb.create_sheet(S_CMP)
K.hdr(wcp, 1, "估值对比 — Bear / Base / Bull 三案逐年(2026E-2030E)隐含价阶梯", 8)
K.mtext(wcp, 2, ("三案各自完整推演逐年 SOTP×EV/Sales: 当年分部营收 × 当年 EV/Sales → ΣEV + 期权 + 净现金 → ÷股本 = 当年隐含价。"
                 "三案永远并排可见, 不随『情景切换』开关变化(只引矩阵各案行 + 全局参数, 防当前案污染)。"
                 "底部: ① 12M 目标价 = 各案 2027E 隐含价(NTM+1)概率加权(0.30/0.45/0.25); ② 2030E 概率加权看'成长进价'终点。"), "H", 3)
for col, w in zip("ABCDEFGH", [26, 11, 11, 11, 11, 11, 13, 13]):
    wcp.column_dimensions[col].width = w

case_cols_mult = {"Bear": 0, "Base": 1, "Bull": 2}  # 在 MULT_ROWS/SWB 里的偏移
# 三案矩阵源: 营收引『情景切换』SWB(三案在行, 逐年 E-I 列); EV/Sales+期权引『估值倍数假设』(三案行 E-I 列)
DF_NC = NC_REF

# block 行定义: (key, label, fmt, bold, out, kind)
#   kind: ("rev",seg)/("ps",seg)/("ev",a,b)/("sumev")/("opt")/("nc")/("eqv")/("px")/("vsnow")/("iev",seg)
rows_def = [
    ("sl_rev", "Starlink 营收($B)", K.N1, False, False, ("rev", "sl_rev")),
    ("sl_ps",  "Starlink EV/Sales(x)", K.MX, False, False, ("ps", "sl")),
    ("sl_ev",  "Starlink EV($B)", K.N0, False, False, ("ev", "sl_rev", "sl_ps")),
    ("sp_rev", "发射 营收($B)", K.N1, False, False, ("rev", "sp_rev")),
    ("sp_ps",  "发射 EV/Sales(x)", K.MX, False, False, ("ps", "sp")),
    ("sp_ev",  "发射 EV($B)", K.N0, False, False, ("ev", "sp_rev", "sp_ps")),
    ("ai_rev", "AI 营收($B)", K.N1, False, False, ("rev", "ai_rev")),
    ("ai_ps",  "AI EV/Sales(x)", K.MX, False, False, ("ps", "ai")),
    ("ai_ev",  "AI EV($B)", K.N0, False, False, ("ev", "ai_rev", "ai_ps")),
    ("sumev",  "Σ 三分部 EV($B)", K.N0, True, False, ("sumev",)),
    ("opt",    "+ 期权 PV($B)", K.N0, False, False, ("opt",)),
    ("nc",     "+ 净现金($B)", K.N0, False, False, ("nc",)),
    ("eqv",    "= 权益价值($B)", K.N0, True, False, ("eqv",)),
    ("px",     "隐含价($/股)", K.PX, True, True, ("px",)),
    ("vsnow",  "vs 现价", K.PCT, False, False, ("vsnow",)),
    ("iev_ai", "AI 隐含 EV/EBITDA(交叉验证)", K.MX, False, False, ("iev", "ai_ev")),
    ("ipe",    "隐含 forward P/E(整体亏损 N/M, 见 EV/Sales)", K.MX, False, False, ("nm_pe",)),
]
block_h = len(rows_def) + 3   # 色带 + 年份头 + 行 + 1空
CMP_BLOCK = {}
b0 = 6
for ci, cname in enumerate(CASES):
    r0 = b0 + ci * block_h
    K.band(wcp, r0, f"{cname} 案 — 逐年 SOTP×EV/Sales 隐含价阶梯", 8)
    yr = r0 + 1
    wcp[f"A{yr}"] = "项目 / 年份"; wcp[f"A{yr}"].font = K.BF
    for col, y in zip(FCf, FWY):
        wcp[f"{col}{yr}"] = y; wcp[f"{col}{yr}"].font = K.BF; wcp[f"{col}{yr}"].fill = K.CH
    rr = yr + 1
    A = {}
    for key, label, fmt, bold, out, kind in rows_def:
        A[key] = rr
        K.lab(wcp, f"A{rr}", label, b=bold)
        if out:
            wcp[f"A{rr}"].fill = K.OUT
        if kind[0] == "nm_pe":   # 整体亏损 → forward P/E 不可用, 整行 N/M(主镜头走 EV/Sales)
            for col in FCf:
                K.lab(wcp, f"{col}{rr}", "N/M", note=True)
            rr += 1
            continue
        for col in FCf:
            k0 = kind[0]
            if k0 == "rev":
                # 引情景切换 SWB[营收key] + ci (三案在行)
                src = SWB[kind[1]] + ci
                f = f"={K.R(S_SW, col + str(src))}"
            elif k0 == "ps":
                # 引估值倍数假设 MULT_ROWS[seg][case]
                src = MULT_ROWS[kind[1]][cname]
                f = f"={K.R(S_MULT, col + str(src))}"
            elif k0 == "ev":
                f = f"={col}{A[kind[1]]}*{col}{A[kind[2]]}"
            elif k0 == "sumev":
                f = f"={col}{A['sl_ev']}+{col}{A['sp_ev']}+{col}{A['ai_ev']}"
            elif k0 == "opt":
                src = OPT_ROWS[cname]
                f = f"={K.R(S_MULT, col + str(src))}"
            elif k0 == "nc":
                f = f"={DF_NC}"
            elif k0 == "eqv":
                f = f"={col}{A['sumev']}+{col}{A['opt']}+{col}{A['nc']}"
            elif k0 == "px":
                f = f"={col}{A['eqv']}/{SH_REF}*1000"
            elif k0 == "vsnow":
                f = f"={col}{A['px']}/{PX_REF}-1"
            elif k0 == "iev":
                # AI 隐含 EV/EBITDA: EV ÷ (营收 × AI margin); 仅 EBITDA>0 显示
                aimgn = AI_MGN[cname][["E", "F", "G", "H", "I"].index(col)]
                if aimgn > 0:
                    f = f"={col}{A['ai_ev']}/({col}{A['ai_rev']}*{aimgn})"
                else:
                    K.lab(wcp, f"{col}{rr}", "n.m.", note=True)
                    continue
            link = ("'" in f)
            K.fml(wcp, f"{col}{rr}", f, fmt, link=link)
        rr += 1
    CMP_BLOCK[cname] = A

# 底部汇总: 2027E 加权(12M TP) + 2030E 加权 ──────
sum_r0 = b0 + len(CASES) * block_h + 1
K.band(wcp, sum_r0, "概率加权 → 12M 目标价(2027E) + 2030E 终点", 8); sum_r0 += 1
for col, h in zip(["A", "B", "C", "D", "E"], ["项目", "Bear", "Base", "Bull", "概率加权"]):
    wcp[f"{col}{sum_r0}"] = h; wcp[f"{col}{sum_r0}"].font = K.BF; wcp[f"{col}{sum_r0}"].fill = K.CH
wcp.merge_cells(f"F{sum_r0}:H{sum_r0}"); wcp[f"F{sum_r0}"] = "说明"
wcp[f"F{sum_r0}"].font = K.BF; wcp[f"F{sum_r0}"].fill = K.CH
sum_r0 += 1
# 概率权重行
R_WT = sum_r0
K.lab(wcp, f"A{sum_r0}", "概率权重")
K.inp(wcp, f"B{sum_r0}", WEIGHTS["Bear"], None, K.PCT)
K.inp(wcp, f"C{sum_r0}", WEIGHTS["Base"], None, K.PCT)
K.inp(wcp, f"D{sum_r0}", WEIGHTS["Bull"], None, K.PCT)
wcp.merge_cells(f"F{sum_r0}:H{sum_r0}")
K.logic(wcp, f"F{sum_r0}", "主观概率: 轨道算力散热给低概率 → Bear 0.30 / Base 0.45 / Bull 0.25。")
sum_r0 += 1
# 2027E 隐含价 + 加权 (12M TP)
R_TP27 = sum_r0
PX_KEY_ROW = {cn: CMP_BLOCK[cn]["px"] for cn in CASES}
K.lab(wcp, f"A{sum_r0}", "2027E 隐含价(12M, NTM+1)", b=True); wcp[f"A{sum_r0}"].fill = K.OUT
for col, cname in zip(["B", "C", "D"], CASES):
    K.fml(wcp, f"{col}{sum_r0}", f"=F{PX_KEY_ROW[cname]}", K.PX, link=True)  # F 列 = 2027E
K.fml(wcp, f"E{sum_r0}", f"=B{sum_r0}*B{R_WT}+C{sum_r0}*C{R_WT}+D{sum_r0}*D{R_WT}", K.PX)
wcp[f"E{sum_r0}"].fill = K.CUR
wcp.merge_cells(f"F{sum_r0}:H{sum_r0}")
K.logic(wcp, f"F{sum_r0}", "= 各案 2027E 隐含价 × 概率 → 12M 目标价 ≈ $56(NTM+1 口径); 较现价 $161 下行约 65%。")
sum_r0 += 1
# 2030E 隐含价 + 加权
R_TP30 = sum_r0
K.lab(wcp, f"A{sum_r0}", "2030E 隐含价(成长进价终点)", b=True); wcp[f"A{sum_r0}"].fill = K.OUT
for col, cname in zip(["B", "C", "D"], CASES):
    K.fml(wcp, f"{col}{sum_r0}", f"=I{PX_KEY_ROW[cname]}", K.PX, link=True)  # I 列 = 2030E
K.fml(wcp, f"E{sum_r0}", f"=B{sum_r0}*B{R_WT}+C{sum_r0}*C{R_WT}+D{sum_r0}*D{R_WT}", K.PX)
wcp.merge_cells(f"F{sum_r0}:H{sum_r0}")
K.logic(wcp, f"F{sum_r0}", "= 各案 2030E 隐含价 × 概率 ≈ $103; Bull 2030E $215 单独才超现价 — 看公司'成长进价'到哪一年。")
sum_r0 += 1
# 现价 + 关键洞察
K.lab(wcp, f"A{sum_r0}", "现价对照($)")
K.fml(wcp, f"E{sum_r0}", f"={PX_REF}", K.PX, link=True); wcp[f"E{sum_r0}"].fill = K.CUR
wcp.merge_cells(f"F{sum_r0}:H{sum_r0}")
K.logic(wcp, f"F{sum_r0}", "现价 $160.95: 仅 Bull 的 2029E($190)/2030E($215)隐含价超过它 → 现价在把 Bull 的 2029-30 当确定来定价。")
sum_r0 += 1
K.lab(wcp, f"A{sum_r0}", "对照: Oppenheimer / 街均值")
K.inp(wcp, f"B{sum_r0}", 190, None, K.PX)
K.inp(wcp, f"C{sum_r0}", 164, None, K.PX)
wcp.merge_cells(f"F{sum_r0}:H{sum_r0}")
K.logic(wcp, f"F{sum_r0}", "Oppenheimer $190(跑赢)/ 街均值 $164(区间 $63-$227); 我们与多头分歧=轨道算力概率 + 逐年自立估值(不把 2035 硬贴现成一个点)。")
sum_r0 += 2
K.mtext(wcp, sum_r0, ("关键洞察: Base 逐年阶梯 $39→$55→$68→$77→$87 全程低于现价; 唯有 Bull 在 2029-2030 才把估值'长'到 $190-$215 超过现价。"
                      "即现价 $161 = 市场把 Bull 情景的 2029-30 当成确定来定价, 几乎不留执行/物理风险折让。"
                      "衰减信号(ARPU 跌破 $50 / Starship 再炸 / 租赁流失)触发应下调 Bull 概率; 上行扳机(AI1 散热达标 / Flight 13+ 成功 / AI 转正)触发上调 Bull 概率。"), "H", 3)

# ════════════ 13. 综合判断仪表盘 ════════════
# 当前案(Base)2027E 隐含价(F 列) 与 2030E(I 列) 作为估值错位锚
IMPL_27 = K.R(S_VAL, "F" + str(R_IMPLPX))   # 2027E
IMPL_30 = K.R(S_VAL, "I" + str(R_IMPLPX))   # 2030E
PX_D = K.R(S_SW, "B" + str(GLOBAL_ROWS["现价($)"]))
dash = K.write_dashboard(wb.create_sheet(S_DASH), {
    "title": "综合判断仪表盘 — A 基本面拐点 · B 估值错位 · C 催化剂 · D 情绪确认",
    "usage": ("怎么用: 预测引擎是 B(估值错位)+ C(催化剂); 情绪 D 只做 timing + 过热刹车(定性档位)。"
              "估值用逐年 SOTP×EV/Sales 阶梯: 现价仅在 Bull 2029-30 才被追上 → 现价已 price-in 牛市远期。"),
    "blocks": [
        {"title": "A. 基本面拐点 — 业务在结构性变好吗?", "rows": [
            ("Starlink 现金牛", "✓ 已确立", "2025 营收 $11.4B(+50%)、EBITDA margin 63%、全球 LEO 份额 75% — 现金牛地位扎实。"),
            ("AI 能否扭亏", "✗ 待验证", "AI 2025 仍亏 $6.4B、吃 60% capex; 转正(2027)取决于租赁+Cursor 放量与轨道算力, 是最大不确定。"),
            ("A 判断", "【中】", "现金牛确立但增长引擎(AI)未证盈利能力; 结构在变但拐点未坐实。", True),
        ]},
        {"title": "B. 估值错位(预测引擎)— 市场给的 vs 逐年 SOTP 该给的 → GAP", "rows": [
            ("市场现在给(现价)", {"fml": f"={PX_D}", "fmt": K.PX, "fill": True},
             "= 现价 $160.95(市场对全公司的定价)。"),
            ("SOTP 该给(Base 2027E, 12M)", {"fml": f"={IMPL_27}", "fmt": K.PX},
             "= 逐年 SOTP×EV/Sales 的 Base 2027E 隐含价(引『情景估值』当前案; NTM+1 口径)。"),
            ("SOTP 该给(Base 2030E 终点)", {"fml": f"={IMPL_30}", "fmt": K.PX},
             "= Base 2030E 隐含价(成长进价终点; 仍低于现价)。"),
            ("错位 GAP = 2027E该给÷市场给 − 1",
             {"fml": lambda ro: f"=B{ro['SOTP 该给(Base 2027E, 12M)']}/B{ro['市场现在给(现价)']}-1", "fmt": K.PCT},
             "GAP 约 −66% = 现价远高于 Base 2027E SOTP 该给 → 已透支牛市愿景, 进入纯情绪/期权定价区(该卖, 不是该买)。"),
            ("回测: IPO 定价 $135 的读数", "(同向偏负)", "即便按 IPO $135, Base 2027E $55 仍隐含大幅下行 — 上市即高估。"),
        ]},
        {"title": "C. 催化剂 — 什么会改变 GAP", "rows": [
            ("Starship Flight 13+", "待", "成功并演示上面级回收复用 → 上调发射/期权(向 Bull)。"),
            ("AI1 轨道卫星散热实测", "待", ">1000 W/㎡ 稳定运行 → 轨道算力从期权转实质, 大幅上调 AI(GAP 最大翻转项)。"),
            ("Starlink ARPU 企稳 / AI 转正", "进行中", "ARPU 在 $60 企稳 + AI 经营利润转正 → 上调 Starlink/AI。"),
            ("C 判断", "待兑现", "关键催化剂(散热实测)未兑现前, 现价的牛市定价缺乏证据支撑。", True),
        ]},
        {"title": "D. 情绪确认 — timing + 刹车(定性档位)", "rows": [
            ("IPO 首日 +19%", "过热信号", "首日大涨 + 散户追捧 + 指数纳入预期 → 情绪推动, 非基本面驱动。"),
            ("散户 + 指数纳入", "过热", "题材稀缺性(唯一可投太空+AI)放大 FOMO, 价格脱离基本面锚。"),
            ("当前档位", "【过热】", "情绪定价区: 现价远超 Base 逐年 SOTP, 仅 Bull 2029-30 追上 → timing 偏向回避/做空, 不追。", True),
            ("衰减扳机", "4 条", "ARPU 跌破 $50 / Starship 再炸 / 租赁合同流失 / Terafab 跳票 — 任一触发 → 转 Bear。"),
        ]},
    ],
    "final": {"band": "★ 综合判断(A+B+C+D)",
              "text": ("优质资产, 但 $161 现价已 price-in Bull 情景的 2029-2030 远期(逐年阶梯里仅 Bull 2029E $190/2030E $215 超现价), 几乎不折让执行/物理风险。"
                       "B 估值错位 −66%(Base 2027E 该给 ≈$55)+ D 情绪过热 + C 关键催化剂未兑现 → 评级卖出(UNDERPERFORM), 12M TP ≈ $56。"
                       "风险收益严重不对称偏下行; 上行需轨道算力散热实测达标这一低概率事件把概率从 Bull 0.25 上调。")},
    "tracking": {
        "intro": "哪个指标恶化/兑现 → 哪个假设先动 → 触发什么动作(盯的优先级)。",
        "rows": [
            ("__band__", "一、AI / 轨道算力(最大摆幅)"),
            ("AI1 散热实测", "待", "关键敏感项: 轨道算力期权(逐年 $0→$200B)与 AI 营收", "AI1 卫星在轨数据 / 公司披露", "达标 → 上调 Bull 概率 + AI 营收/倍数"),
            ("AI 经营利润", "亏损", "关键敏感项: AI 2027 转正假设", "季报分部附注", "转正 → 验证 Base; 持续亏 → 转 Bear"),
            ("__band__", "二、Starlink 现金牛"),
            ("blended ARPU", "$143 下行", "关键敏感项: Starlink 逐年营收/EV/Sales", "季度 KPI 表", "跌破 $50 → 下调 Starlink; 企稳 $60 → 维持"),
            ("__band__", "三、发射 / Starship"),
            ("Starship Flight", "V3 刚炸", "关键敏感项: 发射逐年营收 + $100/kg 期权", "SpaceX 发射记录 / FAA", "Flight 13+ 成功 → 上调; 再炸 → 下调"),
        ],
    },
})

# ════════════ 全局格式 + 落盘 ════════════
K.finalize(wb, freeze={
    S_COVER: "A2", S_HIST: "B3", S_PX: "B4", S_CONS: "A2", S_HMULT: "B3",
    S_MULT: "B3", S_SW: "B3", S_ANCHOR: "B3", S_SEG: "B3",
    S_FUND: "B5", S_VAL: "B3", S_CMP: "B3", S_DASH: "B6",
})
_repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_output_dir = os.environ.get("VALUATION_OUTPUT_DIR", os.path.join(_repo_root, "out"))
out = os.path.join(_output_dir, "SPCX_valuation_model.xlsx")
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("saved:", out)
print("sheets:", wb.sheetnames)
