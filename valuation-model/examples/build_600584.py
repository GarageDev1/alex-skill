# -*- coding: utf-8 -*-
"""
build_600584.py — 长电科技(600584.SS, JCET)估值模型。从 build_template.py 出发, 只调 build_kit。

锚: AIDC CapEx(shared-base/compute-aidc-base.json) → 运算电子(AI直驱, 强度式)
    + 通讯/消费(周期式) + 汽车/工医(增速假设) → 整体OPM → 净利 → 账面
镜头: P/B-ROE 主线 + P/E 平行镜头(A股市场惯用 PE 看封测, 双镜头并行输出)。
单位: 人民币亿元(reporting currency); AIDC capex $B 在锚页 ×USD/CNY×10 折亿元。

运行: PYTHONUTF8=1 python examples/build_600584.py
校验: python scripts/validate_valuation.py out/600584.SS_valuation_model.xlsx
"""
import os, csv, json
from openpyxl import Workbook
import build_kit as K

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "..", "out")

# ════════════ 0. 全局轴 ════════════
ALLC = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
ALLY = ["2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
HC, HY = ["B", "C", "D", "E", "F"], ["2021", "2022", "2023", "2024", "2025"]
FC = ["F", "G", "H", "I", "J", "K"]
FCf = FC[1:]
CASES = ["Bear", "Base", "Bull"]
FX1 = 1                                   # 单币种模型: 全表人民币, FX 行恒 1
PCT3 = '0.000%'                           # 收入强度(千分位级)专用格式

S_COVER, S_HIST, S_PX, S_CONS = "封面", "历史财务与估值", "股价走势", "卖方研报共识"
S_HMULT, S_MULT, S_SW = "历史估值倍数", "估值倍数假设", "情景切换"
S_ANCHOR, S_SEG, S_FUND = "AIDC Capex基座", "分部测算", "利润与收入假设"
S_VAL, S_CMP, S_DASH = "情景估值", "估值对比", "综合判断仪表盘"

# 实拉月度前复权收盘(腾讯 ifzq, 2003-06 → 2026-06-12), 单一价格来源
MONTHLY = []
with open(os.path.join(OUT, "600584_price_monthly.csv"), encoding="utf-8") as f:
    for row in csv.DictReader(f):
        MONTHLY.append((row["ym"], float(row["close"])))
PX_NOW = 69.79                            # 2026-06-12 腾讯实时收盘

# 历史财务(东财 emweb 三表接口, 2026-06-12 实拉; 人民币亿元)
SEG_HIST = {                              # 年报"按市场应用领域"披露拆分
    "运算电子": [40.3, 58.7, 42.1, 58.3, 82.8],
    "通讯电子": [122.0, 132.7, 130.2, 161.1, 141.5],
    "消费电子": [103.1, 98.9, 74.7, 86.7, 91.7],
    "汽车电子": [7.9, 14.9, 23.4, 28.4, 37.3],
    "工业及医疗": [31.4, 32.4, 26.1, 25.2, 35.4],
}
NI_H = [29.59, 32.31, 14.71, 16.10, 15.65]
OP_H = [31.70, 32.46, 15.20, 16.51, 17.39]
OPM_H = [0.104, 0.096, 0.051, 0.046, 0.045]
CONV_H = [0.933, 0.995, 0.968, 0.975, 0.900]
EQ_H = [209.91, 246.43, 260.66, 276.19, 286.71]
SH_H = [1779.6, 1779.6, 1788.8, 1789.4, 1789.4]   # mn 股
GM_H = [0.184, 0.170, 0.137, 0.131, 0.141]
PX_END = [30.37, 22.60, 29.61, 40.71, 36.78]
PX_AVG = [34.67, 24.10, 30.04, 31.59, 36.78]
NI_TTM = 16.52                            # = 2025FY 15.65 − 1Q25 2.03 + 1Q26 2.90
EQ_NOW = 287.66                           # 1Q26 归母权益
USDCNY_H = [6.45, 6.74, 7.05, 7.12, 7.10]
USDCNY_F = 6.76                           # 2026-06-12 腾讯 whUSDCNY 实拉

wb = Workbook()
wb.remove(wb["Sheet"])

# ════════════ 1. 封面 ════════════
K.write_cover(wb.create_sheet(S_COVER), {
    "title": "长电科技(600584.SS)估值模型 — AIDC CapEx 物理锚 × P/B-ROE + P/E 双镜头",
    "meta": [
        ("报告日期", "2026-06-12"),
        ("数据截止", "2025 年报 + 2026Q1 季报(东财实拉) + 2026-06 卖方共识 + 2026-06-12 实时股价"),
        ("现价", "69.79 元(2026-06-12); TTM P/E 75.6x / P/B 4.34x / 市值 1,249 亿元"),
        ("时效声明", "AIDC capex 路径取 shared-base 共享基座(2026-06 版, CreditSights/TrendForce/GS 口径); 股价与共识为 2026-06-12 时点, 每季财报后应更新。"),
        ("方法一句话", "AIDC CapEx($B, 共享基座)→ 运算电子收入(capex×收入强度) + 通讯/消费周期段 + 汽车/工医 → 整体营业利润率 → 净利/账面 → 目标 P/B(三层分解)与目标 P/E × 前瞻每股 → 隐含价。"),
    ],
    "takeaways": [
        ("① 当下估值位置", "P/B 4.34x 已高于 2019-2025 全部历史(年末带 1.6-2.6x, 2021 周期峰 3.4x, 2020 FOMO 顶约 4.2x); TTM P/E 75.6x、对 2026E 共识 EPS 65x——市场已在按 AI 先进封装重估定价, 不在历史带内。"),
        ("② 核心引擎", "运算电子是唯一 AI 直驱段: 收入 = AIDC capex(亿元) × 收入强度。2025 实际强度 0.239%(82.8亿÷34,648亿), 因 capex 增速远快于公司 AI 切片, 前瞻强度逐年回落(Base 0.205%→0.155%), 段收入仍 +39%/+17%/+9% 增长。"),
        ("③ 盈利判断", "净利率历史 4-10% 摆动, 2025 增收不增利(扣非 -11%), 2026Q1 年化仅 11.6 亿。Base 净利 2026E 20 亿(EPS 1.13, 略高于共识 1.07)→ 2027E 24 亿(EPS 1.36 ≈ 共识 1.32), 靠产能利用率 + 先进封装占比修复 OPM 至 5-6%。"),
        ("④ 三情景目标价(12M, 2027E 口径)", "P/B 镜头: Bear ~44 / Base ~68 / Bull ~84 元; P/E 镜头: Bear ~25 / Base ~54 / Bull ~85 元。现价 69.79 已贴 P/B Base 上限、显著高于 P/E Base——上行靠 Bull(盈利兑现超预期), 下行空间更深, risk-reward 偏不利。"),
        ("⑤ 主要风险", "AIDC capex 下修(物理锚)、运算电子强度回落快于预期(AI 封测订单流向台系 OSAT)、传统段价格战、净利率修复证伪(2026Q1 年化即 Bear 地板 11.6 亿)、情绪退潮(P/B 从 4.3x 回历史带)。"),
    ],
})

# ════════════ 2. 历史财务与估值 ════════════
ha = K.write_history(wb.create_sheet(S_HIST), {
    "title": "长电科技 历史财务与估值(人民币亿元)— 2021-2025A + 当下(TTM) + 1Q26",
    "hist_cols": HC, "hist_years": HY,
    "fx_hist": [FX1] * 5, "fx_now": FX1,
    "vals_in_usd": True,                 # 数值已是最终单位(亿元), 不做换算
    "ps_scale": 100, "mcap_div": 100,
    "unit_label": "(人民币 亿元)", "mcap_label": "市值(亿元)",
    "eps_label": "EPS (元)", "bps_label": "BPS (元)",
    "fx_label": "记账币种=人民币(估值同币种, 此行恒 1)",
    "cur_label": "当下(TTM)",
    "segments": [
        ("运算电子", SEG_HIST["运算电子"], True),
        ("通讯电子", SEG_HIST["通讯电子"], True),
        ("消费电子", SEG_HIST["消费电子"], True),
        ("汽车电子", SEG_HIST["汽车电子"], True),
        ("工业及医疗", SEG_HIST["工业及医疗"], False),
    ],
    "total_now": 388.7 - 93.35 + 91.71,   # TTM 营收 = 2025FY − 1Q25 + 1Q26
    "gm_pct": GM_H, "gm_now": 0.141,
    "ni": NI_H, "ni_now": NI_TTM,
    "eq": EQ_H, "eq_now": EQ_NOW,
    "shares": SH_H, "shares_now": 1789.4,
    "px_end": PX_END, "px_now": PX_NOW, "px_avg": PX_AVG,
    "quarter": {"col": "H", "label": "1Q26A", "segs": {},
                "ni": 2.90, "eq": EQ_NOW, "shares": 1789.4, "fx": FX1,
                "note": "1Q26: 营收 91.71亿(-1.8% YoY), 归母 2.90亿(+42.7%), 毛利率 14.55%。最新一期实际, 是 Bear 档的盈利地板锚(年化 11.6亿)。分部季度未披露, 留空。"},
    "band_note": "年末 P/B 带 1.63-2.64x(2021-25), 2021 月末峰 3.39x; 当下 4.34x 已破历史带 → AI 重估 regime",
    "notes": [
        ("运算电子", "唯一 AI 直驱段(服务器/HPC/存储芯片封测, 含 PC CPU 基底)。2025 +42.6% 创新高, 占比升至 21.3%。拆分=年报『按市场应用领域』披露(2021-2025 各年年报), 收入=占比×总营收换算。"),
        ("通讯电子", "手机 SoC/射频/基带封测, 最大段(2025 占 36.4%)。2025 -12%: 大客户手机订单回落。"),
        ("消费电子", "消费类 IC 封测, 随消费电子周期。"),
        ("汽车电子", "车规封测, 基数小增速高(2025 +31.7%); 含晟碟车规等产线。"),
        ("工业及医疗", "工业/医疗 IC 封测(2025 +40.6%)。"),
        ("HREV", "总营收=东财利润表实际(2026-06-12 拉取); 分部=年报应用领域占比换算。当下列=TTM(2025FY−1Q25+1Q26)。"),
        ("HGMP", "毛利率: 公司年报; 封测重资产, 毛利随稼动率/产品结构摆动 13-18%。"),
        ("HNI", "归母净利: 东财利润表实际。当下=TTM 16.52亿(=15.65−2.03+2.90)。2023-25 在 15-16亿平台徘徊, 增收不增利(折旧+ASP 年降)。"),
        ("HEQ", "归母权益: 东财资产负债表实际; 当下=1Q26 末 287.66亿 → BPS 16.08 元, 对齐终端 P/B 4.34x。"),
        ("HSH", "总股本(东财): 2023 年小幅增发后 17.894 亿股至今未变, 无稀释。"),
        ("HPX", "年末股价: 腾讯月K前复权真实收盘(2026-06-12 刷新); 当下=现价 69.79。"),
        ("HPXA", "年均股价: 同一条月度序列的当年均值, 平滑年末单点噪声。"),
    ],
})

# 1Q26 列总营收: 分部季度未披露(留空), 总营收直接填实际值, 避免求和为 0
K.inp(wb[S_HIST], f"H{ha['HREV']}", 91.71, None, K.N1)

# ════════════ 3. 股价走势 ════════════
def phase_fn(ym):
    if ym <= "2019-06":
        return "① 传统封测周期"
    if ym <= "2021-12":
        return "② 5G+国产替代"
    if ym <= "2023-12":
        return "③ 下行去库存"
    if ym <= "2025-12":
        return "④ AI 封测复苏"
    return "⑤ AI 重估"

px = K.write_price_chart(wb.create_sheet(S_PX), MONTHLY, {
    "fn": phase_fn,
    "rows": [
        ("① 传统封测周期", "2003-2019: 收购星科金朋消化期, 股价随半导体周期宽幅震荡"),
        ("② 5G+国产替代", "2020-2021: 5G+国产替代行情, 2020-07 见前复权高点 ~44 元(P/B ~4.2x), 2021 周期峰 P/B 3.4x"),
        ("③ 下行去库存", "2022-2023: 半导体下行, 净利从 32亿 跌到 15亿, 股价回 21-33 区间"),
        ("④ AI 封测复苏", "2024-2025: 收入创新高但增收不增利, 股价 32-44 区间"),
        ("⑤ AI 重估", "2026-05 起: AI 先进封装叙事点燃, 单月从 45.6 冲 82.1, 现回 69.8 — 市场切换到重估定价"),
    ],
}, title="长电科技 月度股价(元, 前复权)")

# ════════════ 4. 卖方研报共识 ════════════
K.write_consensus(wb.create_sheet(S_CONS), {
    "title": "卖方研报共识 — 东财近 6 个月研报 n=5(买入 2 / 增持 3); 本表是后面测算的卖方对账单",
    "overview": "全街共识(2026-06): 评级 买入2/增持3; 2026E EPS 均值 1.072 元(区间 0.95-1.19, n=5), 2027E 1.322 元(1.21-1.39)。核心叙事: 先进封装(XDFOI 2.5D/3D)龙头受益 AI 算力, 2025 先进封装收入 270 亿创新高; 分歧在净利率修复的斜率。",
    "assumptions": [
        ("运算电子增速\n(2026E)", "街上普遍引用 2025 +42.6% 的势头, 认为 AI 服务器/HPC 封测需求延续高增(国信: '2025 收入+8%, 运算电子+42.6%'; 华鑫: '先进封装龙头受益 AI 算力')。",
         "增速能否延续: 运算段含 PC CPU 基底, AI 占比不透明; 台系 OSAT(日月光/Amkor)拿走绝大部分 GPU/HBM 封测。",
         "我们不直接拍增速, 而是用 AIDC capex × 收入强度: 2025 实际强度 0.239%, Base 2026 取 0.205%(capex +70% 而强度回落)→ 段收入 +39%, 与街上方向一致但有物理锚可证伪。"),
        ("2026E 盈利\n(EPS)", "共识 2026E EPS 1.072(净利约 19.2亿), 隐含净利率从 4.0% 修复到 ~4.4%。",
         "区间 0.95-1.19 较宽: 分歧在稼动率回升斜率与 ASP 年降的对冲。2026Q1 年化仅 0.65 元, 全年要靠下半年放量。",
         "Base 净利 20.2亿(EPS 1.13), 略高于共识——运算段放量+产品结构改善给 OPM 5.0%(2025 4.5%); Bear 用 2026Q1 年化 11.6亿 做地板。"),
        ("资本开支", "开源证券: '百亿资本开支'——2026 年 capex 指引约百亿级, 投向晟碟存储封测与 XDFOI 先进封装产能。",
         "高 capex 短期压折旧/净利率, 长期决定能否吃到 AI 封测外溢。",
         "不进收入主链(产能=建厂周期, 不作收入驱动); 作为净利率假设的旁注: 折旧上行是 OPM 修复偏慢取 5-6%(而非回 10%)的原因之一。"),
        ("目标倍数", "卖方多用 PE 给目标价(A股封测惯例); 当前板块 2026E forward PE: 长电 65x / 通富 59x / 华天 44x(2026-06-12 实拉价格÷共识 EPS)。",
         "板块整体被 AI 行情抬到 forward 45-65x, 远高于 2025 末的 ~34x; 这部分是情绪还是结构, 是估值最大分歧。",
         "P/B 三层分解做主线(峰值 3.4x × 结构溢价 1.15x × 情绪值), 复现当下 4.34x 需情绪值 1.11(过热档); P/E 镜头目标倍数 Base 2026 45x(=历史年末带上沿), 低于市场当前 65x——不拟合现价。"),
    ],
    "divergences": [
        "① 净利率修复斜率: 2023-2025 净利卡在 15-16亿 平台(增收不增利), 共识赌 2026-27 修复到 19→24亿。若折旧+价格战继续吞掉收入增量, Bear 的 13-14亿 才是常态——这是 EPS 上下限的主分歧。",
        "② AI 含量成色: 运算电子 21.3% 占比里多少是真 AI(2.5D/3D 先进封装)、多少是 PC/通用服务器基底? 强度路径(0.24%→0.16% vs 守住 0.21%)决定运算段 2030 是 157亿 还是 177亿。",
        "③ 倍数 regime: P/B 4.34x 破历史带, 是'AI 先进封装稀缺标的'的新常态, 还是 2020 式 FOMO(当年 4.2x 顶后两年腰斩)? 三案情绪值 0.45-1.25 覆盖这个分歧。",
    ],
    "stances": [
        "国信证券(增持): 2025 收入 +8% 创新高, 运算电子 +42.6% 验证 AI 拉动; 关注净利率修复。",
        "开源证券(买入): 百亿资本开支卡位先进封装, 晟碟并表+XDFOI 放量是 2026-27 增长主线。",
        "华鑫证券(买入): 先进封装龙头受益 AI 算力, 国产算力链封测环节稀缺标的。",
        "共识均值: 2026E EPS 1.072 / 2027E 1.322(东财, 2026-06 口径, n=5)。",
    ],
})

# ════════════ 5. 历史估值倍数(数据底座)════════════
hm = K.write_hist_multiples(wb.create_sheet(S_HMULT), {
    "title": "历史估值倍数 — 自身历史带 + 当下 TTM + A股封测同业对照(2026-06-12 实拉)",
    "intro": "先看数据再做假设: ①长电自己历史上值多少(逐年年末 + 年内高低带) ②现在市场给多少(TTM 对齐终端) ③同业(通富微电/华天科技)值多少 + 相对华天的 P/B 比值(结构溢价对账线)。看完这页再去下一页拍三案倍数。",
    "s_hist": S_HIST, "ha": ha, "hist_cols": HC, "hist_years": HY,
    "yhigh": px["yhigh"], "ylow": px["ylow"],
    "fwd_note": "现价÷共识 2026E EPS 1.072 ≈ 65x forward P/E; 现价÷1Q26 BPS 16.08 = 4.34x P/B(对齐腾讯终端 75.59x/4.34x)",
    "self_name": "长电科技",
    "self_fwd_pe_label": "≈65x(共识口径)",
    "self_note": "本模型标的。2026-06-12 腾讯实时: 价 69.79 / PE_TTM 75.59 / PB 4.34 / 市值 1,249亿。模型 forward 推导见『情景估值』。",
    "peers": [
        {"name": "通富微电(002156)", "yearly": [4.6, 2.4, 3.3, 3.4, 3.9], "cur_pb": 5.62, "cur_pe": 60.4, "fwd_pe": 59.0,
         "note": "AMD 封测主力(AI GPU 含量最高的 A股封测), 本轮重估最猛。当下行 2026-06-12 腾讯实拉(价 57.55); forward = 价÷共识 2026E EPS ~0.97(机构预测净利 ~14.8亿)。历史年末 P/B 为东财口径约值。"},
        {"name": "华天科技(002185)", "yearly": [3.6, 1.7, 2.2, 2.2, 2.0], "cur_pb": 2.99, "cur_pe": 65.5, "fwd_pe": 43.5,
         "note": "传统封测占比高、AI 含量最低 → 板块下沿参照。当下 2026-06-12 实拉(价 16.08); forward = 价÷共识 2026E EPS 0.37。"},
        {"name": "A股半导体板块(参照)", "yearly": None, "cur_pb": None, "cur_pe": None, "fwd_pe": 45.0,
         "note": "申万半导体 forward 约 40-50x 区间(粗估, 仅作光谱中档参照)。"},
    ],
    "ratio": {"peer": "华天科技(002185)",
              "note": "长电/华天 P/B 比值 = 结构溢价对账线: 2022 前 ~1.0-1.3, 当下 4.34/2.99 = 1.45 — 市场已给长电相对传统封测 ~45% 溢价(先进封装+规模)。下一页第二层取 1.15x(给已验证的部分, 不给满)。"},
    "reading": "① 自己: 年末 P/B 带 1.63-2.64x, 2021 月末峰 3.39x, 2020 FOMO 顶约 4.2x(BPS 粗估); 当下 4.34x 已在历史最高位之上。② 同业: 板块整体重估(通富 5.6x/华天 3.0x), 不是长电独贵 — AI 行情是行业性的; 但 forward PE 45-65x 全板块都在历史上沿。③ 相对华天比值 1.45 → 第二层结构溢价 1.15x 有据且留有余量。→ 下一页: 峰值 3.4x × 溢价 1.15x × 情绪值(三案)。",
})

# ════════════ 6. 估值倍数假设(P/B 三层 + P/E 镜头三案)════════════
ma = K.write_multiple_assumptions(wb.create_sheet(S_MULT), {
    "title": "估值倍数假设 — P/B 三层分解(峰值×结构溢价×情绪值) + P/E 镜头三案目标倍数",
    "intro": "这一页只做判断(数据底座在上一页): ①为什么 P/B-ROE 做主线、P/E 做平行镜头 ②P/B 三层分解出三案目标倍数 ③P/E 镜头三案目标倍数。『情景切换』引用并切换, 『情景估值』套用当前案, 『估值对比』三案并排。",
    "why_text": ("镜头选择是业务判断: 封测(OSAT)是典型商品型重资产——产品无定价权(封装代工价由产能供需定, ASP 常年年降), 盈利是周期状态量(长电净利率历史 4-10% 摆动, 2023-25 增收不增利), "
                 "产能/资产才是穿越周期持续存在的结构存量 → P/B-ROE 做主线。"
                 "但 A股市场对长电的定价惯例是 P/E(卖方研报全部用 PE 给目标价, 散户/机构都看 PE)——市场用什么镜头定价本身是事实, 不能无视 → P/E 不降级为支线, 而是平行镜头: 两套隐含价并列输出, 分歧本身就是结论"
                 "(P/B 说市场在为资产+稀缺性付钱, P/E 说盈利兑现撑不起这个价)。"
                 "镜头迁移触发条件: 若 XDFOI 先进封装做成真 franchise(毛利率结构性上 20%+、净利率回 8%+), 市场会彻底切到 P/E 资本化盈利, 届时 P/E 镜头升为主线。"),
    "why_rows": 6,
    "method_text": "P/B 三层分解(不硬拍): ①历史周期峰值 3.4x(2021 国产替代周期月末实际峰值, 不用 2020 FOMO 顶 4.2x——BPS 粗估且属泡沫顶, 也不用已被本轮抬高的当下值) × ②结构溢价 1.15x(锚长电/华天比值) × ③情绪值(三案)。一致性检验: 3.4×1.15×1.11 = 4.34x = 当下实际 P/B ✓(情绪值 1.11 = 当下处过热启动段)。",
    "peak": 3.4, "peak_note": "2021 周期峰值 P/B(月末口径 3.39x, 由真实月度价÷当年 BPS 11.80 实算)。2020-07 曾见 ~4.2x, 但 2020 BPS 为粗估且属 5G FOMO 顶(其后两年腰斩), 不作锚。",
    "premium": 1.15, "premium_note": "本轮结构性重估: 先进封装(XDFOI/晟碟)+运算占比 13%→21% 的真实 mix 迁移。锚相对华天比值: 历史 ~1.0-1.3 → 当下 1.45; 取 1.15 = 给已被收入验证的部分, FOMO 部分留给情绪层。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "hpb_row": ha["HPB"],
    "cases": [
        ("Bear", [0.80, 0.65, 0.55, 0.50, 0.45],
         "AI capex 下修(2027 增速砍到 10%)+净利率修复证伪 → 情绪 2026 内退潮(0.80), 2027 起回历史带中下沿(3.4×1.15×0.55≈2.2x ≈ 2023 年末实际)。类比 2020 顶后路径: 4.2x→1.6x 用了 18 个月。"),
        ("Base", [1.10, 0.95, 0.82, 0.72, 0.65],
         "可见度撑 12-18 个月(2026 业绩兑现期情绪 1.10≈当下), 2027 起随盈利兑现但增速回落, 情绪向 1.0 之下缓退; 2030 余 0.65(=目标 P/B 2.54x, 仍高于历史中枢 2.3x, 给结构迁移留溢价)。"),
        ("Bull", [1.25, 1.15, 1.00, 0.88, 0.80],
         "2.5D/3D 先进封装外溢订单落地+国产算力链自主化加速 → 市场接受'中国版日月光'叙事, 情绪 2026 冲 1.25(P/B ~4.9x, 略破当下), 之后高位缓退、2028 仍付足峰值×溢价。"),
    ],
    "sent_note": "情绪值=周期/情绪位置。1.0=付足『峰值×溢价』(3.91x); >1=FOMO 超涨; <1=下行折价。历史列=实际 P/B÷3.91 公式反推, 三案同值(一致性检验可见): 2024 0.68 / 2025 0.59 → 当下 1.11, 重估发生在 2026-05 单月。",
    "target_note": "同一个三层公式套三案情绪。历史列=实际倍数(回看, 三案同值)。",
    "reconcile_text": "卖方 vs 我们: 卖方用 PE 给目标价(隐含 2026E 45-65x); 我们 P/B Base 2027 目标 3.71x(=3.4×1.15×0.95), 对应隐含价 ~68 元 ≈ 现价——即市场当前价已把 Base 情景基本付足; P/E 镜头 Base 45x→40x 低于市场当前 65x, 隐含价 ~54 元——两镜头的分歧(68 vs 54)正是'按资产稀缺定价 vs 按盈利兑现定价'的差, 是本模型的核心结论而非误差。",
    "source_text": "第一层: 真实月度价(腾讯 ifzq 实拉)÷东财年报 BPS 实算。第二层: 长电/华天 P/B 比值(上一页对账线, 2026-06-12 实拉)。第三层情绪值: 依据『综合判断仪表盘』D 块档位(当下=过热启动段 1.11), 持续性类比 2020 顶(4.2x→1.6x/18个月)与本轮 AI 可见度(云厂 2026 指引已锁)。",
})

# —— P/E 镜头三案目标倍数(追加在倍数假设页, 供『情景切换』linked 引用) ——
wsM = wb[S_MULT]
_r = wsM.max_row + 2
K.band(wsM, _r, "P/E 镜头 — 三案目标 P/E(历史列=实际年末 P/E; A股市场对封测的定价惯例镜头)", 11); _r += 1
K.lab(wsM, f"A{_r}", "年份");
for col, y in zip(ALLC, ALLY):
    wsM[f"{col}{_r}"] = y; wsM[f"{col}{_r}"].font = K.BF; wsM[f"{col}{_r}"].fill = K.CH
wsM[f"L{_r}"] = "说明"; wsM[f"L{_r}"].font = K.BF; wsM[f"L{_r}"].fill = K.CH
_r += 1
MPE3 = _r
_pe_cases = [
    ("Bear", [35, 35, 30, 28, 25],
     "盈利证伪年给的不是更低 PE 而是周期中性 PE(盈利低基数, 30-35x ≈ 2023-25 年末实际 36-45x 之下); 2028 后随板块退潮回 25-30x。"),
    ("Base", [45, 40, 34, 30, 28],
     "2026 取 45x = 自身历史年末带上沿(2024 年末实际 45.2x), 明显低于市场当前 forward 65x——判断: 板块 FOMO 部分(65x→45x)会在 12 个月内被盈利增长消化; 之后随增速回落向 28-34x 收敛(对应 PEG~1.5-2)。"),
    ("Bull", [55, 50, 45, 40, 35],
     "先进封装兑现 + 国产算力稀缺性 → 市场维持准当前倍数(55x vs 现 65x), 随 EPS 高增缓降; 2028 仍 45x(比照通富当下 59x 的 AI 含量定价)。"),
]
for cs, vals, story in _pe_cases:
    K.lab(wsM, f"A{MPE3 + _pe_cases.index((cs, vals, story))}", f"  {cs} 目标P/E")
    rr = MPE3 + [c[0] for c in _pe_cases].index(cs)
    for col in HC:
        K.fml(wsM, f"{col}{rr}", f"={K.R(S_HIST, col + str(ha['HPE']))}", K.MX, link=True)
    for col, v in zip(FCf, vals):
        K.inp(wsM, f"{col}{rr}", v, None, K.MX)
    K.logic(wsM, f"L{rr}", story)
_r = MPE3 + 3
K.mtext(wsM, _r, "P/E 目标倍数依据: 自身历史年末 P/E 带 12.4x(2022 盈利峰)— 45.2x(2024 盈利谷), 当下 TTM 75.6x / forward 65x; 同业 forward 通富 59x / 华天 44x(上一页实拉)。Base 45x 落在历史带上沿之内; Bull 55x 突破历史带, 理由=AI 重估 regime + 同业(通富 59x)已交易在该水平。", "K", 2)

# ════════════ 7. 情景切换 ════════════
sw = K.write_scenario_switch(wb.create_sheet(S_SW), {
    "title": "情景切换 — 全模型唯一的情景参数库 + 切换开关(默认 Base)",
    "usage": ("怎么用: B2 是唯一入口——下拉选案 → 案序号派生 → 各杠杆『当前案』行跟着切 → 整条明细链(锚→分部测算→利润→倍数→估值)变档, 『情景估值』输出该案逐年隐含价。"
              "三案对比不用切: 『估值对比』恒常三列并排(引本页矩阵行)。情景参数只在本页改(蓝字); 未列入的假设(汽车/工医增速、净利转换、留存率)三案共用(跟 Base)。"),
    "cases": CASES, "default": "Base",
    "triggers": [
        ("Bear", "落进 Bear 的信号: ①云厂 2027 capex 指引转下修(物理锚增速 27%→10%) ②运算电子单季增速跌破 +15%(强度回落快于 Base) ③传统段价格战重启(通富/华天扩产+稼动率不足) ④连续两季净利率低于 3.5%(修复证伪, 向 2026Q1 年化 11.6亿 地板收敛)。"),
        ("Base", "基线: AIDC capex 走共享基座路径(2026 830 → 2030 1500 $B, 云厂指引+GS 基线); 运算段随 capex 放量但强度自然摊薄; 传统段量微增价微降; OPM 靠稼动率+结构修复到 5-6%; 2026 净利 ~20亿 兑现共识。"),
        ("Bull", "落进 Bull 的信号: ①2.5D/3D(XDFOI)拿到国产 GPU/ASIC 大客户量产订单(强度守住 0.20%+) ②净利率单季破 5.5%(OPM 修复超预期) ③AI capex 2027 指引上修(>+30%) ④国产算力自主化政策加码, 封测环节本土化率强制提升。"),
    ],
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "levers": [
        {"key": "capex", "name": "AIDC capex 增速(物理锚)", "fmt": K.PCT, "cols": FC[2:],
         "vals": {"Bear": [0.10, 0.00, -0.05, 0.00], "Base": [0.27, 0.18, 0.11, 0.09], "Bull": [0.35, 0.25, 0.15, 0.10]},
         "desc": "物理锚: 运算电子收入挂它。2026=830$B 是云厂已给指引, 三案共用(改锚页 G 列三案全动); 分歧从 2027 起。共用锚=shared-base/compute-aidc-base.json(CreditSights 2025A 488 / TrendForce 2026E 830 / GS 基线 2027-30)。",
         "stories": {"Bear": "AI 投资回报证伪 → 云厂 2027 集体踩刹车(类比 2000 宽带 capex 见顶后 -35%): 2027 +10%, 2028 持平, 2029 -5%。",
                     "Base": "GS 'Tracking Trillions' 基线: 1050/1240/1380/1500 $B, 增速 27%→9% 自然减速。",
                     "Bull": "推理需求爆发 + 主权 AI 加码: 2027 +35%, 减速更慢。"},
         "hist": [None, 1.00, 1.33, 1.86, 1.44]},
        {"key": "aint", "name": "运算电子收入强度(收入÷capex)", "fmt": PCT3,
         "vals": {"Bear": [0.00185, 0.00165, 0.00150, 0.00140, 0.00135],
                  "Base": [0.00205, 0.00190, 0.00175, 0.00165, 0.00155],
                  "Bull": [0.00215, 0.00205, 0.00195, 0.00185, 0.00175]},
         "desc": "运算电子收入 ÷ AIDC capex(亿元) = 『封测后道占 capex 的价值切片』×『长电在其中的份额』。锚最近实际年公式反推: 2025 = 82.8÷34,648 = 0.239%。强度趋势性回落的机制: capex 增量大头在 GPU/HBM/电力, 其封测在台系 OSAT 与 IDM 自有产线, 长电吃到的中国侧切片增长慢于全球盘子; 2021-23 标 n.m.(运算段含 PC CPU 基底, ÷极小 AI capex 失真)。",
         "stories": {"Bear": "AI 封测订单留在台系/IDM, 长电只剩 PC/通用基底 → 强度 2026 0.185%、2030 滑到 0.135%(段收入增速跌到个位数)。",
                     "Base": "2026 0.205%(段收入 +39%, 顺接 2025 +42.6% 的减速轨道)→ 2030 0.155%: 中国侧 AI 封测份额守住但全球盘子摊薄。",
                     "Bull": "XDFOI 2.5D/3D 拿下国产 GPU/ASIC 量产单 → 强度 2026 0.215%、2030 仍 0.175%(段收入 2030 达 ~177亿)。"},
         "hist": [None, None, None, 0.0041, 0.0024]},
        {"key": "tvol", "name": "传统段(通讯+消费)量增速", "fmt": K.PCT,
         "vals": {"Bear": [0.00, -0.02, 0.00, 0.00, 0.01], "Base": [0.05, 0.04, 0.03, 0.03, 0.03], "Bull": [0.08, 0.06, 0.05, 0.04, 0.04]},
         "desc": "通讯+消费两段(2025 合计 233亿, 占 60%)的封测量增速, 随手机/消费电子出货周期。历史量价未单独披露, 只看段收入 YoY(2022 +2.9% / 2023 -11.5% / 2024 +20.9% / 2025 -5.9%)。",
         "stories": {"Bear": "手机大客户订单继续流失 + 消费电子疲软: 量零增长。",
                     "Base": "2025 低基数后温和复苏(端侧 AI 换机小周期): +5% 渐退到 +3%。",
                     "Bull": "端侧 AI 换机大周期启动: +8% 起步。"},
        },
        {"key": "tasp", "name": "传统段 ASP 变化", "fmt": K.PCT,
         "vals": {"Bear": [-0.05, -0.04, -0.03, -0.03, -0.02], "Base": [-0.02, -0.02, -0.02, -0.02, -0.02], "Bull": [0.00, -0.01, -0.01, -0.01, -0.01]},
         "desc": "封测代工价常年年降(成熟封装无差异化, 客户年度议价); 行业产能过剩时降幅放大。",
         "stories": {"Bear": "通富/华天扩产落地 + 稼动率不足 → 价格战, 年降 -5%/-4%。",
                     "Base": "常态年降 -2%(行业惯例水平)。",
                     "Bull": "先进封装外溢拉满成熟产能稼动率 → 年降收窄到 0~-1%。"},
        },
        {"key": "opm", "name": "整体营业利润率(OPM)", "fmt": K.PCT,
         "vals": {"Bear": [0.038, 0.035, 0.038, 0.040, 0.042], "Base": [0.050, 0.056, 0.060, 0.062, 0.062], "Bull": [0.056, 0.064, 0.068, 0.070, 0.070]},
         "desc": "营业利润÷营收。历史 10.4%(2021 周期峰)→ 4.5%(2025): 增收不增利 = 折旧爬坡(百亿 capex)+ASP 年降吞掉收入增量。修复机制: 稼动率回升的经营杠杆 + 先进封装/汽车占比提升的结构改善, 但被新产线折旧持续对冲 → 三案都回不到 2021 的 10%。",
         "stories": {"Bear": "修复证伪: 2026 3.8%(≈2026Q1 实际 3.2% 略升), 净利贴年化地板 11.6亿 上方。",
                     "Base": "2026 5.0% → 2028 6.0% 企稳: 对应净利 20→27亿, 2026 兑现共识 19.2亿 略上。",
                     "Bull": "先进封装毛利拉动 2026 5.6% → 2028 6.8%, 净利 2027 30亿。"},
         "hist": OPM_H},
    ],
    "linked": [
        {"key": "sent", "name": "情绪值(P/B 第三层)", "fmt": K.N2,
         "src_sheet": S_MULT, "src_row0": ma["sent_row0"],
         "note": "三案取值与依据见『估值倍数假设』; 本页只做切换——要改情绪值, 去那页改蓝字。"},
        {"key": "petgt", "name": "目标 P/E(P/E 镜头)", "fmt": K.MX,
         "src_sheet": S_MULT, "src_row0": MPE3,
         "note": "三案取值与依据见『估值倍数假设』P/E 镜头块; 历史列=实际年末 P/E。"},
    ],
})
# 目标 P/B(当前案) = 峰值 × 溢价 × 当前案情绪 → 喂『情景估值』
_pk = f"'{S_MULT}'!{ma['pk_cell']}"
_pr = f"'{S_MULT}'!{ma['pr_cell']}"
_sent_act = sw["SWACT"]["sent"]
_r = sw["next_row"]
K.lab(wb[S_SW], f"A{_r}", "目标 P/B(当前案)", b=True)
for _c in ALLC:
    K.fml(wb[S_SW], f"{_c}{_r}", f"={_pk}*{_pr}*{_c}{_sent_act}", K.MX, link=True)
K.logic(wb[S_SW], f"L{_r}", "＝ 历史周期峰值 3.4x × 结构溢价 1.15x × 当前案情绪值 → 喂『情景估值』的前瞻 P/B。历史列=实际 P/B(反推一致)。")
SWPB = _r

# ════════════ 8. 物理锚 [ANCHOR] ════════════
anchor = K.write_anchor(wb.create_sheet(S_ANCHOR), {
    "title": "全球 AI 数据中心 CapEx — 需求物理盘子(shared-base 共享基座)",
    "all_cols": ALLC, "all_years": ALLY,
    "series": [
        ("AIDC capex ($B)", [15, 30, 70, 200, 488, 830, None, None, None, None],
         "2022-25=实际(CreditSights/公司财报; 2021 起步期粗估 15); 2026=830 云厂指引(TrendForce 2026-05, 美四大云厂指引合计>$700B); 2027+ = 上年×(1+当前案增速)", K.N0),
        ("USD/CNY(年均)", USDCNY_H + [USDCNY_F] * 5,
         "历史=当年均值(外汇管理局口径约值); 前瞻=现汇 6.76(2026-06-12 腾讯 whUSDCNY 实拉, 无汇率预测观点)", K.N2),
        ("AIDC capex (亿元)", [None] * 10,
         "= $B × USD/CNY × 10。运算电子收入强度的分母, 与公司报表同币种", K.N0),
    ],
    "yoy_row": "AIDC capex ($B)",
    "source_note": "口径=全球 AI 数据中心专项 capex(非 hyperscaler 总额)。来源: shared-base/compute-aidc-base.json v1.0.0(2026-06-03)——2025A 488(CreditSights)/2026E 830(TrendForce+云厂指引)/2027-30E 1050/1240/1380/1500(GS 'Tracking Trillions' 基线)。与 NVDA/Hynix/TSMC 模型共用同一基座, 跨模型一致。",
    "role_note": "作用: 封测后道需求的最底层物理量 = AI 算力芯片的封装/测试量, 由 AIDC capex 拉动。运算电子收入 = capex(亿元) × 收入强度(见分部测算)。改本页 capex → 运算收入 → 净利 → 隐含价全链动。",
})
CAPEX_ROW = anchor["row_of"]["AIDC capex ($B)"]
FXROW = anchor["row_of"]["USD/CNY(年均)"]
CAPEXC_ROW = anchor["row_of"]["AIDC capex (亿元)"]
wsA = wb[S_ANCHOR]
for _i, _c in enumerate(FC[2:]):   # 2027E+ = 上年 × (1+当前案增速)
    K.fml(wsA, f"{_c}{CAPEX_ROW}", f"={FC[1:][_i]}{CAPEX_ROW}*(1+{K.R(S_SW, _c + str(sw['SWACT']['capex']))})", K.N0, link=True)
for _c in ALLC:                    # 亿元 = $B × FX × 10
    K.fml(wsA, f"{_c}{CAPEXC_ROW}", f"={_c}{CAPEX_ROW}*{_c}{FXROW}*10", K.N0)

# ════════════ 9. 分部测算 ════════════
seg = K.write_segment_model(wb.create_sheet(S_SEG), {
    "title": "分部测算 — 运算电子(AI 直驱: capex×收入强度) + 通讯/消费(周期式)",
    "all_cols": ALLC, "all_years": ALLY, "logic_col": "N",
    "groups": [
        ("AIDC capex 物理锚(亿元)", [
            ("AIDC capex (亿元)", None, K.N0, "引『AIDC Capex基座』($B×USD/CNY×10)。改 capex → 运算电子收入跟着动, 这是全模型的物理总开关。"),
        ]),
        ("运算电子 = capex × 收入强度(AI 直驱段)", [
            ("运算电子 收入强度", None, PCT3,
             "假设: 长电从每元 AIDC capex 里切到的封测价值比例短期稳定(封测约占芯片成本的固定小比例, 芯片又占 capex 的固定大比例)。历史=实际收入÷当年 capex 公式反推: 2024 0.41% → 2025 0.239%(锚)。2021-23 标 n.m.: 运算段当年主要是 PC CPU/通用服务器基底, 除以极小的 AI capex 比值失真无意义。前瞻=『情景切换』当前案: 趋势回落, 因为 capex 增量大头(GPU/HBM 封测)在台系 OSAT 与 IDM 自有产线, 长电的中国侧切片增长慢于全球盘子; 回落斜率三案分歧见情景页。"),
            ("运算电子 收入(亿元)", None, K.N1,
             "历史=年报实际(应用领域拆分); 前瞻 = capex(亿元) × 强度。2026E Base ≈115亿(+39%), 顺接 2025 实际 +42.6% 的自然减速, 数值与卖方'运算高增延续'的方向一致。喂『利润与收入假设』。"),
        ]),
        ("通讯/消费 = 周期式(上年 × (1+量) × (1+ASP))", [
            ("传统段 量增速", None, K.PCT,
             "前瞻=『情景切换』当前案(通讯+消费共用)。历史量价公司未拆分披露, 历史列留空, 周期参照=段收入 YoY(2023 -11.5% / 2024 +20.9% / 2025 -5.9%)。"),
            ("传统段 ASP 变化", None, K.PCT,
             "前瞻=『情景切换』当前案。封测代工价常年年降: 成熟封装无差异化, 大客户年度议价压价是行业惯例; 产能过剩年降幅放大。"),
            ("通讯电子 收入(亿元)", None, K.N1,
             "历史=年报实际; 前瞻 = 上年×(1+量)×(1+ASP)。最大段(2025 占 36.4%), 手机 SoC/射频封测, 不挂 capex、走消费周期。"),
            ("消费电子 收入(亿元)", None, K.N1,
             "历史=年报实际; 前瞻同通讯段公式(共用量/价杠杆): 同属消费侧周期, 驱动同源(手机/PC/消费 IC 出货)。"),
        ]),
    ],
})
m = seg["m"]
RH = {k: ha["seg_rows"][k] for k in SEG_HIST}
for col in ALLC:
    K.fml(wb[S_SEG], f"{col}{m['AIDC capex (亿元)']}", f"={K.R(S_ANCHOR, col + str(CAPEXC_ROW))}", K.N0, link=True)
for j, col in enumerate(HC):
    if j >= 3:   # 2024-2025 强度公式反推; 2021-23 n.m.
        K.fml(wb[S_SEG], f"{col}{m['运算电子 收入强度']}", f"={K.R(S_HIST, col + str(RH['运算电子']))}/{col}{m['AIDC capex (亿元)']}", PCT3, link=True)
    else:
        K.lab(wb[S_SEG], f"{col}{m['运算电子 收入强度']}", "n.m.", note=True)
    K.fml(wb[S_SEG], f"{col}{m['运算电子 收入(亿元)']}", f"={K.R(S_HIST, col + str(RH['运算电子']))}", K.N1, link=True)
    K.fml(wb[S_SEG], f"{col}{m['通讯电子 收入(亿元)']}", f"={K.R(S_HIST, col + str(RH['通讯电子']))}", K.N1, link=True)
    K.fml(wb[S_SEG], f"{col}{m['消费电子 收入(亿元)']}", f"={K.R(S_HIST, col + str(RH['消费电子']))}", K.N1, link=True)
for col in FCf:
    K.fml(wb[S_SEG], f"{col}{m['运算电子 收入强度']}", f"={K.R(S_SW, col + str(sw['SWACT']['aint']))}", PCT3, link=True)
    K.fml(wb[S_SEG], f"{col}{m['运算电子 收入(亿元)']}", f"={col}{m['AIDC capex (亿元)']}*{col}{m['运算电子 收入强度']}", K.N1)
    K.fml(wb[S_SEG], f"{col}{m['传统段 量增速']}", f"={K.R(S_SW, col + str(sw['SWACT']['tvol']))}", K.PCT, link=True)
    K.fml(wb[S_SEG], f"{col}{m['传统段 ASP 变化']}", f"={K.R(S_SW, col + str(sw['SWACT']['tasp']))}", K.PCT, link=True)
    wb[S_SEG][f"{col}{m['运算电子 收入(亿元)']}"].fill = K.OUT
_prevs = [HC[-1]] + list(FCf[:-1])
for _p, _c in zip(_prevs, FCf):
    for _seg in ("通讯电子 收入(亿元)", "消费电子 收入(亿元)"):
        K.fml(wb[S_SEG], f"{_c}{m[_seg]}",
              f"={_p}{m[_seg]}*(1+{_c}{m['传统段 量增速']})*(1+{_c}{m['传统段 ASP 变化']})", K.N1)

# ════════════ 10. 利润与收入假设 ════════════
fr = K.write_fundamentals(wb.create_sheet(S_FUND), {
    "title": "利润与收入假设 — 汽车/工医增速 + OPM + 净利转换 + 留存 + 分部营收→净利→EPS/BPS",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
    "ps_scale": 100, "eps_label": "EPS (元)", "bps_label": "BPS (元)",
    "assum_groups": [
        ("营收驱动(运算/通讯/消费在分部测算; 此处为汽车/工医, 三案共用)", [
            {"name": "汽车电子 增速", "vals": [None, 0.886, 0.570, 0.205, 0.317, 0.18, 0.15, 0.12, 0.10, 0.10],
             "fmt": K.PCT, "logic": "历史=实际 YoY(基数小、车规认证产线爬坡)。前瞻: 车规封测随汽车电动化/智能化渗透, 从 +18% 渐退到 +10%——增速逐年回落因基数变大, 方向参照公司车规产线扩产节奏与 2024 年报'增速继续高于市场平均'。三案共用(非 load-bearing)。"},
            {"name": "工业及医疗 增速", "vals": [None, 0.032, -0.194, -0.034, 0.406, 0.10, 0.08, 0.08, 0.06, 0.06],
             "fmt": K.PCT, "logic": "历史=实际 YoY(2025 +40.6% 为低基数反弹)。前瞻回到工业半导体常态个位数增速。三案共用。"},
        ]),
        ("利润率假设(历史实际锚 + 前瞻; 一根 OPM 线, 粗颗粒)", [
            {"name": "整体营业利润率(OPM)", "vals": OPM_H + [None] * 5,
             "fmt": K.PCT, "logic": "历史=东财利润表实际(营业利润÷营收): 10.4%(2021 周期峰)→4.5%(2025)。分部利润率公司不披露, 按模型海拔用一根整体 OPM 线; 前瞻=『情景切换』当前案(修复机制与三案故事见该页)。",
             "link": {"sheet": S_SW, "row": sw["SWACT"]["opm"]}},
        ]),
        ("净利转换与留存(三案共用)", [
            {"name": "净利转换率(归母净利/营业利润)", "vals": CONV_H + [0.92, 0.92, 0.92, 0.92, 0.92],
             "fmt": K.PCT, "logic": "历史实际 0.90-1.00(税率低 + 少数股东损益小 + 政府补助/投资收益对冲利息)。前瞻取 0.92 ≈ 近五年中位, 三案共用。"},
            {"name": "留存率(1−派息率)", "vals": [0.95, 0.93, 0.85, 0.87, 0.86, 0.87, 0.87, 0.87, 0.87, 0.87],
             "fmt": K.PCT, "logic": "公司低派息(2024 拟 10派1.2 元, 派息率 ~13%), 扩产优先。前瞻 0.87, 三案共用。期末权益 = 上年 + 净利×留存。"},
        ]),
    ],
    "segments": [
        {"name": "运算电子", "hist_row": "运算电子", "fwd": {"sheet": S_SEG, "row": m["运算电子 收入(亿元)"]}},
        {"name": "通讯电子", "hist_row": "通讯电子", "fwd": {"sheet": S_SEG, "row": m["通讯电子 收入(亿元)"]}},
        {"name": "消费电子", "hist_row": "消费电子", "fwd": {"sheet": S_SEG, "row": m["消费电子 收入(亿元)"]}},
        {"name": "汽车电子", "hist_row": "汽车电子", "fwd": {"growth": "汽车电子 增速"}},
        {"name": "工业及医疗", "hist_row": "工业及医疗", "fwd": {"growth": "工业及医疗 增速"}},
    ],
    "profit_terms": [
        (["运算电子", "通讯电子", "消费电子", "汽车电子", "工业及医疗"], "整体营业利润率(OPM)", False),
    ],
    "conv_assum": "净利转换率(归母净利/营业利润)", "retention_assum": "留存率(1−派息率)",
    "note_text": "口径: 人民币亿元 / EPS·BPS 为元。链: 分部营收(运算=capex×强度 + 通讯/消费周期 + 汽车/工医增速)→ 营业利润(=总营收×OPM)→ 归母净利(×净利转换)→ 期末权益(留存递推)→ EPS/BPS/ROE。历史列取东财实际; 下游『情景估值』直接引本表每股, 不重算。",
})

# ════════════ 11. 情景估值 ════════════
sv = K.write_scenario_valuation(wb.create_sheet(S_VAL), {
    "title": "情景估值 — 当前案逐年隐含价(P/B 主线 + P/E 平行镜头)",
    "intro": "本表输出=『情景切换』当前案(默认 Base)。P/B 镜头: 隐含价 = 目标P/B(当前案)×前瞻BPS; P/E 镜头平行输出: 隐含价 = 目标P/E(当前案)×前瞻EPS。历史列用实际年末价反推倍数(事实); 前瞻是预测、不拟合现价——与现价的差就是投资信号。三案并排见『估值对比』。",
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf, "hist_years": HY,
    "fx_fwd": FX1, "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
    "s_fund": S_FUND, "fr": fr,
    "s_switch": S_SW, "target_row": SWPB, "sw_cell": "B2",
    "mcap_div": 100, "mcap_usd_skip": True,
    "mcap_hist_label": "市值 实际年末(亿元, 历史)",
    "mcap_fwd_label": "市值 前瞻·P/B镜头(亿元)",
    "pe_lens": {"target_row": sw["SWACT"]["petgt"], "mcap_label": "市值 前瞻·P/E镜头(亿元)"},
    "yend": px["yend"], "yavg": px["yavg"],
    "reading": "读法: 『当下 forward P/E』= 现价÷模型各年 EPS(Base 2026E ≈62x、2027E ≈51x)——市场当前为 2026 年盈利付 60x+; 『P/E(前瞻=隐含)』= P/B 镜头隐含价÷EPS, 检验 P/B 答案在盈利维度是否荒谬。双镜头分歧(P/B 隐含 > P/E 隐含)= 市场按资产稀缺定价 vs 盈利兑现不足的张力, 是结论不是误差。",
    "method": "方法: 整体公司估值(不做 SOTP), 逐年排。基本面在『利润与收入假设』; 目标倍数在『估值倍数假设』(P/B 三层 + P/E 三案); 本表只做最后一步: 倍数 × 前瞻每股 → 隐含价 + 市值。",
    "concl": "结论(Base, 12M=2027E 口径): P/B 镜头 ~68 元(≈现价 69.79, 市场已付足 Base); P/E 镜头 ~54 元(-22%)。方向判断: 现价已透支 Base 盈利路径, 上行依赖 Bull 兑现(先进封装大单+净利率超修复); 三案与 risk-reward 见『估值对比』。",
})

# ════════════ 12. 估值对比(三案恒常并排) ════════════
SWB = sw["SWB"]
SH_F = K.R(S_HIST, f"$F${ha['HSH']}")
PX_NOW_REF = K.R(S_HIST, f"G{ha['HPX']}")
_agrow = fr["am"]["汽车电子 增速"]
_igrow = fr["am"]["工业及医疗 增速"]
_conv = fr["am"]["净利转换率(归母净利/营业利润)"]
_ret = fr["am"]["留存率(1−派息率)"]


def _fwdprev(j, A, key):
    return (HC[-1] if j == 0 else FCf[j - 1]) + str(A[key])


cmp_rows = [
    {"key": "capd", "label": "AIDC capex ($B)", "fmt": K.N0,
     "hist": lambda c, ci, A: f"={K.R(S_ANCHOR, c + str(CAPEX_ROW))}",
     "fwd": lambda c, j, ci, A: (f"={K.R(S_ANCHOR, 'G' + str(CAPEX_ROW))}" if j == 0
                                 else f"={FCf[j-1]}{A['capd']}*(1+{K.R(S_SW, c + str(SWB['capex'] + ci))})")},
    {"key": "cap", "label": "AIDC capex (亿元)", "fmt": K.N0,
     "hist": lambda c, ci, A: f"={c}{A['capd']}*{K.R(S_ANCHOR, c + str(FXROW))}*10",
     "fwd": lambda c, j, ci, A: f"={c}{A['capd']}*{K.R(S_ANCHOR, c + str(FXROW))}*10"},
    {"key": "comp", "label": "运算电子 (亿元)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(RH['运算电子']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['cap']}*{K.R(S_SW, c + str(SWB['aint'] + ci))}"},
    {"key": "tele", "label": "通讯电子 (亿元)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(RH['通讯电子']))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'tele')}*(1+{K.R(S_SW, c + str(SWB['tvol'] + ci))})*(1+{K.R(S_SW, c + str(SWB['tasp'] + ci))})"},
    {"key": "cons", "label": "消费电子 (亿元)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(RH['消费电子']))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'cons')}*(1+{K.R(S_SW, c + str(SWB['tvol'] + ci))})*(1+{K.R(S_SW, c + str(SWB['tasp'] + ci))})"},
    {"key": "auto", "label": "汽车电子 (亿元)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(RH['汽车电子']))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'auto')}*(1+{K.R(S_FUND, c + str(_agrow))})"},
    {"key": "ind", "label": "工业及医疗 (亿元)", "fmt": K.N1,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(RH['工业及医疗']))}",
     "fwd": lambda c, j, ci, A: f"={_fwdprev(j, A, 'ind')}*(1+{K.R(S_FUND, c + str(_igrow))})"},
    {"key": "rev", "label": "总营收 (亿元)", "fmt": K.N1, "bold": True,
     "hist": lambda c, ci, A: f"={c}{A['comp']}+{c}{A['tele']}+{c}{A['cons']}+{c}{A['auto']}+{c}{A['ind']}",
     "fwd": lambda c, j, ci, A: f"={c}{A['comp']}+{c}{A['tele']}+{c}{A['cons']}+{c}{A['auto']}+{c}{A['ind']}"},
    {"key": "ni", "label": "归母净利 (亿元)", "fmt": K.N1, "bold": True,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HNI']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['rev']}*{K.R(S_SW, c + str(SWB['opm'] + ci))}*{K.R(S_FUND, c + str(_conv))}"},
    {"key": "eq", "label": "期末归母权益 (亿元)", "fmt": K.N0,
     "hist": lambda c, ci, A: f"={K.R(S_HIST, c + str(ha['HEQ']))}",
     "fwd": lambda c, j, ci, A: (f"={K.R(S_HIST, 'F' + str(ha['HEQ']))}+{c}{A['ni']}*{K.R(S_FUND, c + str(_ret))}" if j == 0
                                 else f"={FCf[j-1]}{A['eq']}+{c}{A['ni']}*{K.R(S_FUND, c + str(_ret))}")},
    {"key": "bps", "label": "BPS (元)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={c}{A['eq']}*100/{K.R(S_HIST, c + str(ha['HSH']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['eq']}*100/{SH_F}"},
    {"key": "eps", "label": "EPS (元)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={c}{A['ni']}*100/{K.R(S_HIST, c + str(ha['HSH']))}",
     "fwd": lambda c, j, ci, A: f"={c}{A['ni']}*100/{SH_F}"},
    {"key": "sent", "label": "情绪值(该案; 历史=实际反推)", "fmt": K.N2,
     "hist": lambda c, ci, A: f"={K.R(S_MULT, c + str(ma['sent_row0'] + ci))}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_MULT, c + str(ma['sent_row0'] + ci))}"},
    {"key": "pb", "label": "目标 P/B(历史=实际)", "fmt": K.MX,
     "hist": lambda c, ci, A: f"={_pk}*{_pr}*{c}{A['sent']}",
     "fwd": lambda c, j, ci, A: f"={_pk}*{_pr}*{c}{A['sent']}"},
    {"key": "px", "label": "隐含价 P/B镜头 (元)", "fmt": K.PX, "bold": True, "out": True,
     "hist": lambda c, ci, A: f"={c}{A['pb']}*{c}{A['bps']}",
     "fwd": lambda c, j, ci, A: f"={c}{A['pb']}*{c}{A['bps']}"},
    {"key": "pe", "label": "目标 P/E(历史=实际)", "fmt": K.MX,
     "hist": lambda c, ci, A: f"={K.R(S_MULT, c + str(MPE3 + ci))}",
     "fwd": lambda c, j, ci, A: f"={K.R(S_MULT, c + str(MPE3 + ci))}"},
    {"key": "pxpe", "label": "隐含价 P/E镜头 (元)", "fmt": K.PX, "bold": True, "out": True,
     "hist": lambda c, ci, A: f"={c}{A['pe']}*{c}{A['eps']}",
     "fwd": lambda c, j, ci, A: f"={c}{A['pe']}*{c}{A['eps']}"},
    {"key": "ipe", "label": "隐含 forward P/E(P/B价÷EPS, 交叉验证)", "fmt": K.MX,
     "hist": lambda c, ci, A: f"={c}{A['px']}/{c}{A['eps']}",
     "fwd": lambda c, j, ci, A: f"={c}{A['px']}/{c}{A['eps']}"},
    {"key": "up", "label": "P/B镜头: 历史=vs实际年末价(回测≈0)/前瞻=vs现价", "fmt": K.PCT,
     "hist": lambda c, ci, A: f"={c}{A['px']}/{K.R(S_HIST, c + str(ha['HPX']))}-1",
     "fwd": lambda c, j, ci, A: f"={c}{A['px']}/{PX_NOW_REF}-1"},
    {"key": "uppe", "label": "P/E镜头: 前瞻 vs 现价", "fmt": K.PCT,
     "hist": lambda c, ci, A: None,
     "fwd": lambda c, j, ci, A: f"={c}{A['pxpe']}/{PX_NOW_REF}-1"},
]
cm = K.write_comparison(wb.create_sheet(S_CMP), {
    "title": "估值对比 — Bear / Base / Bull 三情景目标价并排(P/B 与 P/E 双镜头各一行)",
    "intro": ("三个情景各自完整推演: AIDC capex → 运算电子(capex×强度) + 通讯/消费(量×价) + 汽车/工医 → 净利(×OPM×转换)→ 权益/BPS → 目标 P/B 与目标 P/E → 双镜头逐年隐含价。"
              "本表三案永远并排可见, 不随『情景切换』开关变化; 要调假设去『情景切换』改对应案参数。未列入情景矩阵的假设(汽车/工医增速、净利转换、留存)三案共用 Base。"
              "历史列 2021-2025 = 同一条链填实际值: P/B 镜头隐含价历史列应基本等于实际年末价(回测行 ≈0%)——整条估值链的内置回测。"),
    "case_names": CASES,
    "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCf,
    "block_start": 18,
    "rows": cmp_rows,
    "summary": {
        "band": "三案汇总(12M 目标 = 2027E 列; 触发条件见『情景切换』; 概率参考 Bear25%/Base50%/Bull25%)",
        "target_col": "H",
        "rows": [
            ("归母净利(亿元)", "ni", K.N1, "由该案 capex/强度/量价/OPM 沿同一条链推出, 非手拍", False),
            ("总营收(亿元)", "rev", K.N0, "= 运算(capex×强度) + 通讯/消费(该案量价) + 汽车/工医", False),
            ("目标 P/B", "pb", K.MX, "= 峰值 3.4x × 结构溢价 1.15x × 该案情绪值", False),
            ("隐含价 P/B镜头(元)", "px", K.PX, "= 目标P/B × 2027E BPS; 主线镜头(资产/产能定价)", True),
            ("目标 P/E", "pe", K.MX, "三案目标 P/E 见『估值倍数假设』(A股惯用镜头)", False),
            ("隐含价 P/E镜头(元)", "pxpe", K.PX, "= 目标P/E × 2027E EPS; 平行镜头(盈利兑现定价)", True),
            ("隐含 forward P/E", "ipe", K.MX, "P/E 体检: P/B 镜头隐含价÷该案 EPS, 对照同业 forward(通富59x/华天44x)", False),
            ("P/B镜头 vs 现价", "up", K.PCT, "对照现价 69.79 的上行/下行空间", True),
            ("P/E镜头 vs 现价", "uppe", K.PCT, "同上, P/E 镜头口径", True),
        ],
        "mcap": {"label": "隐含市值 P/B镜头(亿元)", "key": "px", "expr": f"*{SH_F}/100",
                 "note": "= 隐含价 × 股本 17.89亿"},
        "concl": "Risk-reward(12M, vs 现价 69.79): P/B 镜头 Bear -37% / Base -3% / Bull +20%; P/E 镜头 Bear -64% / Base -22% / Bull +21%。两镜头概率加权(25/50/25)≈ 66 元 / 55 元, 均不高于现价——现价已付足 Base、上行依赖 Bull 兑现, 下行尾部更深。任一 Bear 触发信号出现应先减仓再验证。",
    },
})

# ════════════ 13. 综合判断仪表盘 ════════════
BPS26 = K.R(S_FUND, "G" + str(fr["BPS"]))
EPS26 = K.R(S_FUND, "G" + str(fr["EPS"]))
PXD = K.R(S_HIST, "G" + str(ha["HPX"]))
dash = K.write_dashboard(wb.create_sheet(S_DASH), {
    "title": "综合判断仪表盘 — A 基本面拐点 · B 估值错位(预测引擎) · C 催化剂 · D 情绪确认",
    "usage": ("怎么用: 预测引擎是 B(错位)+C(催化剂)——两样都当下可观测; D 情绪只做 timing 确认+过热刹车。"
              "回测锚: 2026-04(股价 45.6, P/B 2.9x)时 A 强(运算 +42.6% 年报刚出)+B 错位为正+C 催化剂(AI 算力国产化)在途 → 该看多; 一个月后 +80%。当下读数已翻向'过热'。"),
    "blocks": [
        {"title": "A. 基本面拐点 — 业务在结构性变好吗?", "rows": [
            ("产品组合迁移", "运算电子占比 13%(2021)→21%(2025)→Base 2030 ~30%", "结构迁移是真的(先进封装收入 270亿 创新高); 但净利率还没跟上——迁移在收入端兑现、利润端待验证。"),
            ("可持续 ROE", "2025 实际 5.6%; Base 2027E ~7.6%", "ROE 修复靠 OPM 从 4.5%→5.6%; 若修复证伪, 5-6% 的 ROE 撑不住 4x P/B。"),
            ("A 判断", "【中】", "收入结构在变好(强), 利润兑现待验证(弱)→ 综合中。", True),
        ]},
        {"title": "B. 估值错位(预测引擎 ★)— 市场现在给的 vs 基本面该给的 → GAP", "rows": [
            ("市场现在给(forward P/B)", {"fml": f"={PXD}/{BPS26}", "fmt": K.MX, "fill": True},
             "= 现价 ÷ 2026E BPS(公式算, 随模型走)。"),
            ("基本面该给", {"inp": 2.6, "fmt": K.MX},
             "= 历史 P/B 中枢 2.3x × 结构溢价 1.15x ≈ 2.6x(已验证部分)。注: 纯 justified P/B=(ROE 7.6%−g3%)/(COE 9.5%−g3%)≈0.7x, A股封测从不按它交易, 列作下限警示而非锚。"),
            ("错位 GAP = 该给÷市场给 − 1",
             {"fml": lambda ro: f"=B{ro['基本面该给']}/B{ro['市场现在给(forward P/B)']}-1", "fmt": K.PCT},
             "GAP 深负 = 基本面空间已被价格吃完, 进入纯情绪定价区——上行只能靠情绪续杠或基本面超预期。"),
            ("回测: 2026-04 拐点读数", "市场给 2.9x vs 该给 ~2.6x, GAP ≈ -10%", "当时错位接近闭合+催化剂在途 → 还有情绪段可赚; 现在 GAP ≈ -40%, 同一框架给出相反信号。"),
        ]},
        {"title": "C. 催化剂 — 什么会逼市场闭合/扩大 GAP", "rows": [
            ("XDFOI 2.5D/3D 国产大客户量产单", "进行中", "Bull 的核心扳机: 强度守住 0.20%+ 的唯一路径; 落地前涨幅全是预付。"),
            ("2026 中报净利率", "待(2026-08)", "OPM 修复 5%+ 是 Base 成立的第一个检查点; 低于 4% 即向 Bear 滑。"),
            ("云厂 2027 capex 指引", "待(2026Q4-2027Q1 财报季)", "物理锚的下一次重定价: 27%(Base)还是 10%(Bear)。"),
            ("C 判断", "关键催化剂均未兑现", "现价先于催化剂兑现 → 涨的是预期, 兑现不及即回撤。", True),
        ]},
        {"title": "D. 情绪确认 — timing + 刹车(定性档位)", "rows": [
            ("量价温度计(只当温度计, 不进倍数)", "2026-05 单月 +80%, 月度换手历史峰值后回落 15%", "典型情绪脉冲形态; 量价 ≫ 基本面变化速度 = FOMO 段脆弱。"),
            ("现价倍数 vs 基本面该给", "4.1x forward vs 2.6x", "市场已付出超过基本面该给 ~58% → 纯情绪定价区。"),
            ("当前档位", "【过热】", "对照 2020-07(P/B 4.2x 顶, 其后 18 个月 -50%): 同等海拔。", True),
            ("衰减扳机", "4 条", "①中报净利率 <4% ②运算电子单季增速 <+15% ③云厂 capex 指引下修 ④板块龙头(通富)倍数先行回落。任一翻 → 情绪值降档重算。"),
        ]},
    ],
    "final": {"band": "★ 综合判断(A+B+C+D 收成一句可执行的话)",
              "text": "A 中(结构迁移真、利润未兑现)+ B GAP 深负(市场给 4.1x vs 该给 2.6x)+ C 关键催化剂未兑现 + D 过热档 → 当下不是买点: 持有者沿衰减扳机设防减仓, 空仓者等中报净利率验证或回踩 P/B 3x 带(~50 元)再做; Bull 兑现(国产大单+OPM 5.5%+)则按 84 元重看。回测 2026-04 同框架给出'该买'✓, 引擎自洽。"},
    "tracking": {
        "intro": "哪个指标恶化 → 哪个假设先崩 → 触发什么动作。",
        "rows": [
            ("__band__", "一、物理锚"),
            ("云厂/AIDC capex 指引", "2026E $830B(锁定)", "命门: 运算电子收入 = capex × 强度", "四大云厂季报 capex 指引(季度)", "2027 指引 <+15% → 切 Bear 重算"),
            ("__band__", "二、AI 切片"),
            ("运算电子季度增速", "2025 +42.6% / 1Q26 未拆", "命门: 收入强度路径(0.205% vs 0.185%)", "公司季报/年报应用领域拆分(季/年)", "增速 <+15% → 强度下调一档"),
            ("XDFOI 先进封装订单", "客户认证/小批量", "命门: Bull 情景与结构溢价第二层", "公司公告/法说会(事件)", "量产大单落地 → 切 Bull; 一年无进展 → 溢价 1.15→1.05"),
            ("__band__", "三、盈利兑现"),
            ("单季净利率", "1Q26 3.2%", "命门: OPM 修复路径(Base 2026 5.0%)", "季报(季度)", "连续两季 <3.5% → 切 Bear(地板=年化 11.6亿)"),
            ("__band__", "四、估值/情绪"),
            ("P/B vs 3.91x(峰值×溢价)", "当下 4.34x(情绪 1.11)", "命门: 情绪值路径", "周度复核", "跌破 3.9x 且无基本面恶化 = 情绪正常化; 跌破 3x 带回历史区间 → 重新评估买点"),
        ],
    },
})

# ════════════ 全局格式 + 落盘 ════════════
K.finalize(wb, freeze={
    S_HIST: "B3", S_PX: "B4", S_CONS: "A2", S_HMULT: "B5", S_MULT: "B4", S_SW: "B3",
    S_ANCHOR: "B3", S_SEG: "B3", S_FUND: "B3", S_VAL: "B4", S_CMP: "B6", S_DASH: "B6",
    S_COVER: "A2",
})
xlsx_path = os.path.join(OUT, "600584.SS_valuation_model.xlsx")
wb.save(xlsx_path)
print("saved:", os.path.abspath(xlsx_path))
print("sheets:", wb.sheetnames)

# ════════════ input.json(数据 + source 元数据, 全研究假设 SOT) ════════════
SRC_EM = {"source": "东方财富 emweb 三表接口(实拉)", "date": "2026-06-12"}
SRC_AR = {"source": "公司年报『按市场应用领域』披露(2021-2025 各年年报, 占比×总营收换算)", "date": "2026-06-12"}
input_json = {
    "meta": {"ticker": "600584.SS", "name": "长电科技 JCET", "report_date": "2026-06-12",
             "unit": "人民币亿元(EPS/BPS 为元; AIDC capex 为 $B, 模型内 ×USD/CNY×10 折亿元)",
             "anchor": "AIDC CapEx(shared-base/compute-aidc-base.json v1.0.0)",
             "lenses": ["P/B-ROE(主线)", "P/E(平行镜头, A股市场惯用)"]},
    "market": {"price": 69.79, "price_src": {"source": "腾讯实时行情", "date": "2026-06-12"},
               "pe_ttm": 75.59, "pb": 4.34, "mcap_yi": 1248.83, "shares_mn": 1789.4,
               "fx_usdcny_spot": {"value": 6.76, "source": "腾讯 whUSDCNY 实拉", "date": "2026-06-12"},
               "fx_usdcny_hist_avg": dict(zip(HY, USDCNY_H))},
    "hist_financials": {y: {"revenue": sum(SEG_HIST[k][i] for k in SEG_HIST),
                            "segments": {k: SEG_HIST[k][i] for k in SEG_HIST},
                            "op": OP_H[i], "net_income": NI_H[i], "equity": EQ_H[i],
                            "shares_mn": SH_H[i], "gm": GM_H[i],
                            "px_end": PX_END[i], "px_avg": PX_AVG[i]}
                        for i, y in enumerate(HY)},
    "latest_quarter_1Q26": {"revenue": 91.71, "rev_yoy": -0.0176, "net_income": 2.90, "ni_yoy": 0.4274,
                            "gm": 0.1455, "equity": 287.66, "op": 2.95,
                            "note": "Bear 盈利地板 = 年化 11.6 亿", "src": SRC_EM},
    "ttm": {"net_income": NI_TTM, "eps": 0.923, "bps_1q26": 16.08,
            "note": "现价/TTM EPS=75.6x, 现价/1Q26 BPS=4.34x, 对齐腾讯终端"},
    "hist_valuation_bands": {
        "PB_yearend": dict(zip(HY, [2.57, 1.63, 2.03, 2.64, 2.30])),
        "PB_peak_monthly": {"2020": "约4.2x(BPS粗估, FOMO顶)", "2021": 3.39},
        "PE_yearend": dict(zip(HY, [18.3, 12.4, 36.0, 45.2, 42.0])),
        "current": {"PB": 4.34, "PE_TTM": 75.59, "PE_fwd_consensus": 65.1},
        "src": "腾讯月K前复权实拉 ÷ 东财年报 BPS/EPS 实算"},
    "peers": {"通富微电002156": {"price": 57.55, "pe_ttm": 60.4, "pb": 5.62, "fwd_pe_2026e": 59.0,
                                  "fwd_basis": "共识 2026E 净利 ~14.8亿 / EPS ~0.97"},
              "华天科技002185": {"price": 16.08, "pe_ttm": 65.5, "pb": 2.99, "fwd_pe_2026e": 43.5,
                                  "fwd_basis": "共识 2026E EPS 0.37(2 家机构)"},
              "src": {"source": "腾讯实时行情 + 同花顺/东财盈利预测(web)", "date": "2026-06-12"}},
    "consensus": {"eps_2026e": {"mean": 1.072, "range": [0.95, 1.19], "n": 5},
                  "eps_2027e": {"mean": 1.322, "range": [1.21, 1.39]},
                  "ratings": "买入2/增持3",
                  "src": {"source": "东方财富研报中心近6月研报(国信/开源/华鑫等)", "date": "2026-06"}},
    "anchor_inputs": {"aidc_capex_usd_b": {"2022A": 30, "2023A": 70, "2024A": 200, "2025A": 488,
                                           "2026E": 830, "2027E": 1050, "2028E": 1240, "2029E": 1380, "2030E": 1500,
                                           "source": "shared-base/compute-aidc-base.json v1.0.0(CreditSights/TrendForce/GS)",
                                           "note": "2027-30E 为 Base; 模型内由情景增速驱动"},
                      "intensity_2025_actual": {"value": 0.00239, "formula": "82.8亿 ÷ (488×7.10×10)亿",
                                                "note": "运算电子收入强度锚, Excel 内公式反推; 2021-23 n.m."}},
    "assumptions": {
        "levers": {
            "aidc_capex_yoy_2027_30": {"bear": [0.10, 0.00, -0.05, 0.00], "base": [0.27, 0.18, 0.11, 0.09], "bull": [0.35, 0.25, 0.15, 0.10]},
            "compute_intensity_2026_30": {"bear": [0.00185, 0.00165, 0.0015, 0.0014, 0.00135],
                                          "base": [0.00205, 0.0019, 0.00175, 0.00165, 0.00155],
                                          "bull": [0.00215, 0.00205, 0.00195, 0.00185, 0.00175]},
            "trad_volume": {"bear": [0.0, -0.02, 0.0, 0.0, 0.01], "base": [0.05, 0.04, 0.03, 0.03, 0.03], "bull": [0.08, 0.06, 0.05, 0.04, 0.04]},
            "trad_asp": {"bear": [-0.05, -0.04, -0.03, -0.03, -0.02], "base": [-0.02] * 5, "bull": [0.0, -0.01, -0.01, -0.01, -0.01]},
            "opm": {"bear": [0.038, 0.035, 0.038, 0.04, 0.042], "base": [0.05, 0.056, 0.06, 0.062, 0.062], "bull": [0.056, 0.064, 0.068, 0.07, 0.07]},
            "sentiment": {"bear": [0.8, 0.65, 0.55, 0.5, 0.45], "base": [1.1, 0.95, 0.82, 0.72, 0.65], "bull": [1.25, 1.15, 1.0, 0.88, 0.8]},
            "target_pe": {"bear": [35, 35, 30, 28, 25], "base": [45, 40, 34, 30, 28], "bull": [55, 50, 45, 40, 35]},
        },
        "common": {"auto_growth": [0.18, 0.15, 0.12, 0.10, 0.10], "indmed_growth": [0.10, 0.08, 0.08, 0.06, 0.06],
                   "ni_conversion": 0.92, "retention": 0.87},
        "pb_three_layer": {"peak": 3.4, "premium": 1.15,
                           "consistency_check": "3.4×1.15×1.11=4.34x=当下实际 ✓"},
    },
    "sources": {
        "hist_financials": SRC_EM, "segments": SRC_AR,
        "prices": {"source": "腾讯 ifzq 月K前复权(2003-06→2026-06-12)", "file": "600584_price_monthly.csv"},
        "aidc_base": {"source": "shared-base/compute-aidc-base.json", "version": "1.0.0", "date": "2026-06-03"},
        "consensus": {"source": "东财研报中心(国信/开源/华鑫等, n=5)", "date": "2026-06"},
        "segment_split_2024": {"source": "2024年报: 通讯44.8/消费24.1/运算16.2/汽车7.9/工医7.0%", "url": "paper.cnstock.com"},
        "segment_split_2023_2022_2021": {"source": "2023年报(43.9/25.2/14.2/7.9/8.8) 2022年报(39.3/29.3/17.4/4.4/9.6) 2021年报(40.0/33.8/13.2/2.6/10.3)"},
    },
}
jpath = os.path.join(OUT, "600584_input.json")
with open(jpath, "w", encoding="utf-8") as f:
    json.dump(input_json, f, ensure_ascii=False, indent=2)
print("saved:", os.path.abspath(jpath))
