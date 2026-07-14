# -*- coding: utf-8 -*-
"""
Uber Technologies / UBER 估值模型（equity-research-obsidian Phase 2）。

公司专有数据 + 判断 + 情景参数在本脚本；Excel 通用结构 / 格式 / 情景切换 / 校验友好写法
全部复用 build_kit.py。

主线镜头：P/E（平台 franchise 资本化盈利，核心经营 EPS）；DCF（核心 FCF）交叉验证。
P/B 不适用（轻资产、账面无意义、终端 pb=null），不做 P/B 主线。

物理锚链（端到端通电）：
  MAPC(月活, M) × 年人均频次(Trips/MAPC) = 总 Trips(M)
  → 单均 Gross Bookings($) → 总 Gross Bookings($B)
  → 各段 take rate → 各段收入 → 总收入
  → 各段 EBITDA margin(相对 GB) → 总 Adjusted EBITDA → 核心净利 → 核心 EPS；FCF = Adj EBITDA × 转化率
  → 目标 P/E × 核心 EPS → 隐含股价

★ 核心 EPS 口径：剔除大额一次性递延税资产释放（FY24 +$6.4B、FY25 +$5.0B）+
  股权投资公允价值重估（Aurora/Didi/Grab 等）。报表净利 FY24 $9.86B / FY25 $10.05B 严重失真，
  经营性核心净利 FY24 ~$1.6B / FY25 ~$5.2B。估值用核心口径，绝不对报表 EPS 套倍数。
"""
from __future__ import annotations
import json, os
from openpyxl import Workbook
import build_kit as K

ALLC = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
ALLY = ["2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
HC, HY = ["B", "C", "D", "E", "F"], ["2021", "2022", "2023", "2024", "2025"]
FC = ["F", "G", "H", "I", "J", "K"]
FCF = FC[1:]
CASES = ["Bear", "Base", "Bull"]
FX_FC = 1.0                       # USD 单币种
SHARES_M = 2120.0                 # 稀释股本（FY2025 2,119.7M；回购大致对冲 SBC，前瞻持平）
PX_NOW = 72.10

S_COVER, S_HIST, S_PX, S_CONS = "封面", "历史财务与估值", "股价走势", "卖方研报共识"
S_HMULT, S_MULT, S_SW = "历史估值倍数", "估值倍数假设", "情景切换"
S_ANCHOR, S_SEG, S_FUND = "物理锚_MAPC频次", "分部测算", "利润与收入假设"
S_VAL, S_CMP, S_DASH = "情景估值", "估值对比", "综合判断仪表盘"

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
VAULT = os.environ.get("VALUATION_OUTPUT_DIR", os.path.join(REPO_ROOT, "out"))
OUT_XLSX = os.path.join(VAULT, "UBER_valuation_model.xlsx")
OUT_JSON = os.path.join(VAULT, "UBER_input.json")

MOB, DEL, FRT = "Mobility 出行", "Delivery 外卖", "Freight 货运"

# ════════ 公司数据（$B，SEC 10-K / 8-K；2021-2025A）════════
# Gross Bookings 分部（$B）— FY21/22 取 10-K 历史披露，FY23/24/25 与一手对账
GB_SEG = {
    MOB: [38.65, 54.91, 68.90, 83.02, 97.50],
    DEL: [51.63, 59.36, 63.73, 74.61, 90.86],
    FRT: [2.13, 7.06, 5.24, 5.14, 5.09],
}
# 分部收入（$B）
REV_SEG = {
    MOB: [6.41, 14.03, 19.83, 25.09, 29.67],
    DEL: [8.36, 10.90, 12.20, 13.75, 17.25],
    FRT: [2.13, 6.95, 5.25, 5.14, 5.10],
}
# 分部 Adjusted EBITDA（$B）— 含 Corporate/Platform 在合并行体现
SEG_EBITDA = {
    MOB: [1.17, 1.71, 4.96, 6.50, 7.90],
    DEL: [-0.62, 0.21, 1.51, 2.47, 3.57],
    FRT: [-0.01, 0.01, -0.06, -0.07, -0.03],
}
CORP = [-2.10, -1.93, -2.35, -2.41, -2.71]      # Corporate G&A + Platform R&D（$B）
ADJ_EBITDA = [-1.83, 0.0, 4.05, 6.48, 8.73]     # 总 Adjusted EBITDA（FY22≈$0，对账值）
REV_TOTAL = [17.46, 31.88, 37.28, 43.98, 52.02]
NI_REP = [-0.50, -9.14, 1.89, 9.86, 10.05]      # 报表净利（含大额一次性，失真）
NI_CORE = [-1.20, -2.50, 0.30, 1.62, 5.15]      # 核心经营净利（剔 DTA 释放 + 投资重估）
EQ = [15.15, 8.07, 12.03, 22.38, 27.92]
SHARES = [1896, 1975, 2092, 2151, 2120]
FCF_H = [-0.74, 0.39, 3.36, 6.90, 9.76]         # 自由现金流 = OCF − capex
OCF = [-0.45, 0.64, 3.59, 7.14, 10.10]

# 物理锚历史（年度 = 当年 Q4 口径）
MAPC = [118, 131, 150, 171, 202]                # 月活平台消费者（M，年末 Q4）
TRIPS = [6371, 7637, 9446, 11273, 13567]        # 全年总 Trips（M）

# 月度股价（USD，前复权）
MONTHLY = [
    ("2021-06", 50.51), ("2021-09", 44.87), ("2021-12", 41.93),
    ("2022-03", 33.06), ("2022-06", 22.37), ("2022-09", 28.58), ("2022-12", 24.73),
    ("2023-03", 30.75), ("2023-06", 44.42), ("2023-09", 44.91), ("2023-12", 61.57),
    ("2024-03", 77.91), ("2024-06", 70.85), ("2024-09", 77.14), ("2024-12", 60.32),
    ("2025-03", 72.75), ("2025-06", 93.30), ("2025-09", 99.57), ("2025-12", 81.71),
    ("2026-03", 71.93), ("2026-05", 73.77), ("2026-06", 72.10),
]
PX_END = {"2021": 41.93, "2022": 24.73, "2023": 61.57, "2024": 60.32, "2025": 81.71}
PX_AVG = {"2021": 43.73, "2022": 29.35, "2023": 41.78, "2024": 70.77, "2025": 84.70}
PX_HIGH = {2021: 51.73, 2022: 44.42, 2023: 63.28, 2024: 86.34, 2025: 100.10, 2026: 87.59}
PX_LOW = {2021: 35.73, 2022: 20.46, 2023: 25.36, 2024: 57.58, 2025: 63.17, 2026: 68.61}

# ════════ 情景参数 ════════
# P/E 主线三层：质量锚 28x × 结构溢价 1.0 × 情景情绪值
PEAK_PE, PREMIUM = 28.0, 1.0
TARGET_PE = {"Bear": [16, 16, 15, 15, 14], "Base": [25, 24, 23, 22, 21], "Bull": [28, 28, 27, 26, 25]}
PE_SENT = {c: [round(v / PEAK_PE, 3) for v in vals] for c, vals in TARGET_PE.items()}

# 物理锚杠杆：MAPC 增速 / 人均频次增速 / 单均 GB 增速（三段共用单均增速，分段 take rate 单列）
MAPC_G = {"Bear": [0.12, 0.09, 0.07, 0.06, 0.05],
          "Base": [0.15, 0.13, 0.11, 0.10, 0.09],
          "Bull": [0.17, 0.16, 0.14, 0.12, 0.11]}
FREQ_G = {"Bear": [0.01, 0.01, 0.00, 0.00, 0.00],     # Trips/MAPC 增速（Uber One 拉动频次）
          "Base": [0.025, 0.02, 0.02, 0.015, 0.015],
          "Bull": [0.04, 0.035, 0.03, 0.025, 0.02]}
ASP_G = {"Bear": [-0.01, -0.01, 0.00, 0.00, 0.00],    # 单均 GB 增速（mix + 定价）
         "Base": [0.005, 0.005, 0.01, 0.01, 0.01],
         "Bull": [0.015, 0.015, 0.02, 0.02, 0.02]}
# 段 EBITDA margin（相对 GB）三案路径
MOB_EM = {"Bear": [0.078, 0.078, 0.079, 0.080, 0.080],
          "Base": [0.083, 0.085, 0.087, 0.088, 0.089],
          "Bull": [0.087, 0.090, 0.093, 0.095, 0.097]}
DEL_EM = {"Bear": [0.040, 0.043, 0.045, 0.047, 0.048],
          "Base": [0.044, 0.048, 0.052, 0.055, 0.058],
          "Bull": [0.048, 0.054, 0.060, 0.065, 0.070]}
# 核心净利转换率（核心净利 / Adj EBITDA，含 SBC/D&A/税/利息归一化）
NETCONV = {"Bear": [0.50, 0.52, 0.53, 0.54, 0.55],
           "Base": [0.56, 0.59, 0.61, 0.63, 0.64],
           "Bull": [0.60, 0.63, 0.66, 0.68, 0.70]}
# FCF / Adj EBITDA 转化率（轻资产 + 保险准备金时间差）
FCFCONV = {"Bear": [0.95, 0.92, 0.90, 0.88, 0.86],
           "Base": [1.05, 1.00, 0.97, 0.95, 0.93],
           "Bull": [1.10, 1.06, 1.02, 1.00, 0.98]}
# Freight 三案共用 Base（非 load-bearing，长期微亏、占比 < 3%）
FRT_G = [-0.01, 0.0, 0.02, 0.03, 0.03]


def hist_growth(vals):
    return [None] + [round(vals[i] / vals[i - 1] - 1, 4) for i in range(1, 5)]


def take_rate(seg):
    return [round(REV_SEG[seg][i] / GB_SEG[seg][i], 4) for i in range(5)]


def build():
    wb = Workbook(); wb.remove(wb["Sheet"])

    # 1 封面
    K.write_cover(wb.create_sheet(S_COVER), {
        "title": "Uber Technologies, Inc. / UBER 估值模型",
        "meta": [
            ("报告日期", "2026-06-23"),
            ("数据截止", "FY2025 10-K + 2026Q1 实际 + 统一API 股价/一致预期"),
            ("当前股价", f"${PX_NOW:,.2f}（reality check 用，非拟合目标）；市值约 $1,468 亿"),
            ("主线镜头", "P/E 主线（平台 franchise 资本化盈利，核心经营 EPS）；DCF（核心 FCF）交叉验证。P/B 不适用（轻资产、账面无意义）。"),
            ("核心 EPS 口径", "剔除一次性递延税资产释放（FY24 +$6.4B、FY25 +$5.0B）+ 股权投资公允价值重估；报表净利 FY24/25 $9.9B/$10.1B 失真，核心净利 FY24/25 ~$1.6B/~$5.2B。"),
            ("方法一句话", "MAPC × 人均频次 → 总 Trips → 单均 GB → Gross Bookings → take rate → 收入 → 段 EBITDA margin → 核心净利/EPS → 目标 P/E → 隐含股价。"),
        ],
        "takeaways": [
            ("① 物理锚", "月活平台消费者(MAPC) × 年人均出行频次 = 总 Trips；× 单均 Gross Bookings = 总 GB（$193.5B, FY25）。GB 增长由量(Trips +20%)驱动，单均稳定 ~$14.3。"),
            ("② 基本面", "Q1'26 GB $53.7B(+25%)、Adj EBITDA $2.5B(+33%/GB margin 4.6%)、MAPC 199M(+17%)、Trips 3.64B(+20%)、Uber One 50M、TTM FCF $9.8B(record)、AV 行程 +10x YoY。"),
            ("③ 卖方对账", "一致目标价均值 $104.48（区间 $70-$150），strong_buy（46 buy / 5-8 hold / 1 sell, 51 家）。多头：robotaxi 期权 + 平台盈利 + 广告/会员 + FCF/回购 + 估值便宜；空头(Melius $73/Wedbush $84)：AV 去中介化。"),
            ("④ 估值判断", "Base 用 2027E 核心 EPS × 23x P/E。核心 EPS 口径下隐含价显著高于现价——市场用『AV 颠覆空头情景』定价（forward P/E 已压到历史下沿），是预期被错杀的中期预期差。"),
            ("⑤ 评级", "BUY（买入）：robotaxi 被市场误读为颠覆威胁、年内 -10% 被错杀；实际 Uber 是 AV 公司的需求聚合分发渠道(30+ 伙伴)，robotaxi 是免费期权。平台已规模化盈利(FCF $9.8B/转化 112%)。下行风险=AV 收敛到 1-2 家自营赢家。"),
        ],
    })

    # 2 历史财务与估值
    ha = K.write_history(wb.create_sheet(S_HIST), {
        "title": "Uber 历史财务与估值（$B）— 2021-2025A + 当下 TTM / 2026Q1",
        "hist_cols": HC, "hist_years": HY,
        "fx_hist": [1, 1, 1, 1, 1], "fx_now": 1,
        "vals_in_usd": True, "fx_label": "FX (USD 单币种=1)",
        "segments": [(MOB, REV_SEG[MOB], True), (DEL, REV_SEG[DEL], True), (FRT, REV_SEG[FRT], False)],
        "total_now": 52.02,
        "gm_pct": [None] * 5, "gm_now": None,
        "ni": NI_REP, "ni_now": 10.05,
        "eq": EQ, "eq_now": 27.92,
        "shares": SHARES, "shares_now": 2120,
        "px_end": [PX_END[y] for y in HY], "px_now": PX_NOW,
        "px_avg": [PX_AVG[y] for y in HY],
        "band_note": "历史 P/E 早年无意义（2021/22 亏损、2023 转正）；2024-2025 报表 P/E 因一次性 DTA 释放虚低（报表 EPS $4.56/$4.73 含 ~$3/股一次性税收 benefit）。当下 TTM 报表 P/E ~17.9x，核心 EPS 口径下 ~30x；本模型对核心 EPS 套倍数。",
        "quarter": {
            "col": "H", "label": "2026Q1实际",
            "segs": {MOB: (7.46, 0.18), DEL: (4.33, 0.25), FRT: (1.30, -0.02)},
            "ni": 2.63, "eq": 28.5, "shares": 2110, "fx": 1,
            "note": "2026Q1 实际：GB $53.7B(+25%)、营收 $13.2B(+14%)、Adj EBITDA $2.5B(+33%)、经营利润 $1.9B(+57%)；GAAP 净利 $2.63B 被股权公允价值变动扰动，看经营利润更代表盈利能力。",
        },
        "notes": [
            (MOB, "出行撮合，高 take rate(~30%)、高 margin；FY25 GB $97.5B(+17%)、收入 $29.7B、Adj EBITDA $7.9B。take rate ~30% 含英国 VAT/保险毛额列报口径，净额约 20%。"),
            (DEL, "外卖配送，低 take rate(~19%)但增速更快、margin 改善最猛；FY25 GB $90.9B(+22%)、Adj EBITDA $3.57B(+45%)，广告(年化>$2B)是高毛利增量。"),
            (FRT, "货运承运（principal，毛额列报，GB≈收入），长期微亏、占总 GB <3%，估值上可忽略。"),
            ("HNI", "报表净利含大额一次性：FY24 +$6.4B 美国 DTA 释放、FY25 +$5.0B 荷兰 DTA 释放 + 股权投资重估。建模估值用核心经营净利。当下=FY2025 报表。"),
            ("HEQ", "股东权益；FY25 $27.9B。轻资产平台，账面不是估值分母（P/B 不适用）。"),
            ("HSH", "稀释股本；FY25 2,119.7M；2025 起 $6.5B 回购，授权剩余 $19.2B，前瞻按回购对冲 SBC 持平处理。"),
            ("HPX", "股价前复权（2019-05 IPO）；湖回溯至 2021-06，年末/均价用于历史倍数校准。"),
        ],
    })

    # 3 股价走势
    def phase_fn(ym):
        if ym <= "2022-12":
            return "① 成长股杀估值/盈利未证"
        if ym <= "2024-12":
            return "② 盈利转正/纳入标普500"
        if ym <= "2025-09":
            return "③ 平台盈利兑现创新高"
        return "④ robotaxi 颠覆恐慌错杀"
    px = K.write_price_chart(wb.create_sheet(S_PX), MONTHLY, {
        "fn": phase_fn,
        "rows": [("① 成长股杀估值/盈利未证", "2022 加息周期杀无盈利成长股，叠加 Uber 自身尚未稳定盈利，股价探底 $20。"),
                 ("② 盈利转正/纳入标普500", "2023 营业利润转正，2023-12 纳入标普 500 指数，盈利叙事确立，股价回升。"),
                 ("③ 平台盈利兑现创新高", "2024-2025 连续超三年框架，FCF/EBITDA 双增，2025-09 创新高 ~$99.6。"),
                 ("④ robotaxi 颠覆恐慌错杀", "2025-11 起 Waymo/Tesla robotaxi 叙事发酵，市场担心 AV 去中介化，从高点回撤约 28%、年内约 -10% 至 ~$72。")],
    }, title="Uber 月度股价走势（USD, 前复权）")
    px["yhigh"].update(PX_HIGH); px["ylow"].update(PX_LOW)

    # 4 卖方研报共识
    K.write_consensus(wb.create_sheet(S_CONS), {
        "title": "卖方研报共识 — 一致预期 + 各投行对账（2026 Q1-Q2）",
        "overview": "约 51 家覆盖，Buy/Overweight 为主（46 buy / 5-8 hold / 1 sell），一致目标价均值约 $104.48（区间 $70-$150），隐含上行约 +45%。核心叙事：robotaxi 是看涨期权而非颠覆、平台盈利兑现、广告/会员高毛利、强 FCF + 回购、估值便宜。最大分歧：robotaxi 对 Uber 是期权还是结构性杀手。",
        "assumptions": [
            ("Gross Bookings 增速\n(2026/2027)", "一致 2026E 营收 +36%、2027E +33%；GB Q1'26 +25%、Q2 指引 cc +18-22%；管理层三年框架 GB mid-to-high teens CAGR。",
             "分歧在成熟市场是否饱和：bull 维持 14-15% CAGR、bear 个位数饱和。", "Base GB 由 MAPC +13%/频次 +2%/单均 +0.5% 推，2026E +18% 后缓降到 +12%。"),
            ("take rate / margin", "Mobility take ~30%、Delivery ~19% 且回升（广告拉动）；Adj EBITDA/GB 从 2.94%→4.51%(FY23→25)，每年 ~+50bps。",
             "take rate 是否见顶 + robotaxi 主导后是否压缩 take rate。", "Base 段 EBITDA margin(相对GB) Mobility 8.3%→8.9%、Delivery 4.4%→5.8%，广告/规模驱动。"),
            ("核心盈利口径", "报表净利 FY24/25 $9.9B/$10.1B 含 $6.4B/$5.0B 一次性 DTA 释放；经营利润 FY25 $5.57B、Adj EBITDA $8.73B、核心净利 ~$5.2B。",
             "用报表 EPS 套倍数会严重高估目标价（headline P/E 17.9x 是假便宜）。", "本模型用核心经营净利（剔一次性），目标 P/E 对核心 EPS。"),
            ("robotaxi 影响", "30+ AV 伙伴(Waymo/Zoox/Lucid+Nuro 等)；Waymo 经 Uber 在 Austin/Atlanta 上线、利用率高 30%；AV 行程 +10x YoY。MS：AV 未来 12-24 月财务影响极小。",
             "bull(Evercore $150/MS)：聚合器赢家；bear(Melius $73/Wedbush $84)：Waymo/Tesla 自建绕过。", "Base 不把 AV 当主收入驱动，作免费期权处理；下行情景计入 take rate 压缩。"),
            ("目标 P/E / EPS", "卖方多用 20-29x 2027E 调整后 EPS；一致 2027E EPS $4.47（含部分一次性）；BofA ~12x 2027 EV/EBITDA。",
             "EPS 口径(含/不含一次性) + 目标倍数差异决定目标价。", "本模型 Base 2027E 核心 EPS × 23x；刻意落区间中段。"),
        ],
        "divergences": [
            "① robotaxi 是看涨期权还是去中介化威胁——全街最大多空分歧，决定 Uber 长期 take rate 与撮合价值（Evercore $150 vs Melius $73，差一倍）。",
            "② 核心盈利口径：报表净利含 $6.4B/$5.0B 一次性 DTA 释放，用 headline EPS 套倍数会造『假便宜』。",
            "③ 增长持续性：成熟市场是否饱和，bull 14-15% CAGR vs bear 个位数。",
        ],
        "stances": [
            "Evercore ISI（Mahaney）：Outperform，TP $150（全街最高）；AV 聚合器赢家『供给越分散、中间的网络越受益』。",
            "Goldman Sachs（Sheridan）：Conviction Buy，TP $115；下修利润率假设但长期仍 Conviction。",
            "J.P. Morgan（Anmuth）：Overweight，TP $105；21x 2026E FCF(~$9.8B)。",
            "BofA（Justin Post）：Buy，TP $103；『AV 恐慌制造的折价入场点』，~12x 2027 EV/EBITDA。",
            "Morgan Stanley（Nowak）：Overweight，TP $100；AV 未来 12-24 月财务影响极小、核心 Mobility 被低估。",
            "Wedbush（Devitt）：Hold，TP $84；对 AV 路径谨慎。",
            "Melius Research：Sell，TP $73（全街最空）；现价未充分计入竞争威胁。",
        ],
    })

    # 5 历史估值倍数
    hm = K.write_hist_multiples(wb.create_sheet(S_HMULT), {
        "title": "历史估值倍数 — 自身 P/E 带（核心 EPS）+ 当下 + 同业（平台）对照",
        "intro": "Uber 历史报表 P/E 早年无意义（2021/22 亏损）；2024-25 报表 P/E 因一次性 DTA 释放虚低。当下 TTM 报表 P/E ~17.9x（headline 假便宜），核心 EPS 口径下 ~30x。本轮折价由 robotaxi 颠覆恐慌驱动（forward P/E 压到历史下沿）。主线以核心 EPS × P/E 资本化盈利，P/B 不适用。",
        "s_hist": S_HIST, "ha": ha, "hist_cols": HC, "hist_years": HY,
        "yhigh": px["yhigh"], "ylow": px["ylow"],
        "fwd_note": "统一API 当下 TTM 报表 P/E≈17.9x（含一次性，失真）；本模型 Base 2027E 23x 是对核心经营 EPS 的目标倍数，反映平台盈利耐用性，非现价拟合。",
        "self_name": "Uber", "self_fwd_pe_label": "~30x TTM(核心) / 23x Base 2027E(核心)",
        "self_note": "P/E 主线：穿越周期的耐用资产是双边网络 + 多业务平台 + 会员锁定的可持续盈利，不是账面净资产（轻资产，P/B 无意义）。",
        "peers": [
            {"name": "DoorDash", "yearly": None, "cur_pb": None, "cur_pe": 60.0, "fwd_pe": 38.0,
             "note": "最直接外卖平台同业；增速更高（含并购），倍数显著高于 Uber。"},
            {"name": "Airbnb", "yearly": None, "cur_pb": None, "cur_pe": 30.0, "fwd_pe": 26.0,
             "note": "同属轻资产撮合平台，盈利质量可比，倍数 26-30x。"},
            {"name": "Booking Holdings", "yearly": None, "cur_pb": None, "cur_pe": 26.0, "fwd_pe": 22.0,
             "note": "成熟出行撮合平台、强 FCF + 回购，最可比的『成熟平台稳态』倍数锚。"},
            {"name": "Lyft", "yearly": None, "cur_pb": None, "cur_pe": 18.0, "fwd_pe": 14.0,
             "note": "纯网约车同业，规模/盈利远不及 Uber，倍数偏低。"},
            {"name": "S&P 500", "yearly": None, "cur_pb": None, "cur_pe": 23.0, "fwd_pe": 21.0,
             "note": "大盘参照；Uber 核心 EPS 口径前瞻倍数与大盘相近或略低。"},
        ],
        "ratio": None,
        "reading": "Uber 核心 EPS 口径下，合理倍数判断：成熟平台(Booking 22x / Airbnb 26x)与高增长(DoorDash 38x)之间。Base 23x 是『市场误读 AV 威胁、给了大盘平均倍数』的判断——若 robotaxi 证实为期权(平台增长延续)，可上探 26-28x；若 AV 收敛到自营赢家、take rate 压缩，应回落到 15-16x(Bear)。",
    })

    # 6 估值倍数假设
    ma = K.write_multiple_assumptions(wb.create_sheet(S_MULT), {
        "title": "估值倍数假设 — 主线 P/E 三层拆解（质量锚 × 结构溢价 × 情绪/可见度）",
        "intro": "目标 P/E 不是拍一个数，而是：质量锚（28x）× 结构溢价（1.0x，已含在锚内避免双算）× 情景情绪值。情景切换只引这里的三案情绪值；所有目标价由同一条业务链推导。",
        "why_text": ("Uber 穿越周期的耐用资产是『双边网络密度 + 多业务平台交叉(Mobility+Delivery 共享用户/数据) + Uber One 会员锁定』带来的可持续盈利，"
                     "不是账面净资产——它轻资产(FY25 capex 仅占收入 0.6%)、有撮合定价权、会员/网络锁定。"
                     "因此主估值分母选核心经营 EPS，P/B 不适用。若 robotaxi 收敛到 1-2 家自营赢家、Uber 撮合价值被削弱，镜头才应向更保守倍数下移。"
                     "★ 关键：核心 EPS 剔除一次性 DTA 释放($6.4B/$5.0B)与股权公允价值重估(可逆、不可资本化)，这是本模型相对『直接对报表 EPS 套倍数』更严谨之处——headline P/E 17.9x 是假便宜。"),
        "why_rows": 5,
        "method_text": "目标 P/E = 质量上沿锚（28x）× 结构溢价（1.0x）× 情景情绪值。Bear 回到 14-16x（AV 去中介化 + take rate 压缩 + 增长饱和）；Base 2026-27 维持 24-25x 后随成熟回落到 21x；Bull 在 robotaxi 期权证实 + 广告/会员加速下到 25-28x。",
        "peak": PEAK_PE, "peak_note": "成熟轻资产平台同业(Airbnb 26x、Booking 22x fwd、DoorDash 38x)显示前瞻 P/E 上沿约 26-28x；28x 作为质量上沿锚，非当下拟合。",
        "premium": PREMIUM, "premium_note": "结构溢价已含在 28x 质量锚内，避免双重计数；三案差异放情绪/可见度层。",
        "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCF,
        "s_hist": S_HIST, "hpb_row": ha["HPE"],
        "cases": [
            ("Bear", PE_SENT["Bear"], "robotaxi 收敛到 Waymo/Tesla 自营、take rate 被压缩 + 成熟市场饱和 + 大盘倍数压缩 → 回到 14-16x。"),
            ("Base", PE_SENT["Base"], "平台盈利兑现、robotaxi 维持期权属性、广告/会员延续 → 23-25x，随成熟逐步回落到 21x。"),
            ("Bull", PE_SENT["Bull"], "robotaxi 期权证实(Uber 成 AV 默认分发层)+ 广告/会员加速 + 倍数扩张 → 市场给成熟平台龙头 25-28x。"),
        ],
        "reconcile_text": "卖方多用 20-29x 2027E 调整后 EPS（含部分一次性，口径偏松）。本模型 Base 2027E 23x 落在区间中段——对核心 EPS（剔一次性）套倍数，比对 headline EPS 严格；同时现价 forward(核心) P/E 已压到历史下沿，Base 23x 反映『市场用 AV 空头情景定价、给了大盘平均倍数』的预期差。",
        "source_text": "历史倍数来自前复权股价与 SEC 财务；同业倍数来自统一API/Web；目标倍数与情景触发写在本页与情景切换页。",
    })

    # 7 情景切换
    sw = K.write_scenario_switch(wb.create_sheet(S_SW), {
        "title": "情景切换 — 全模型唯一情景参数库（默认 Base）",
        "usage": "B2 是唯一开关。MAPC 增速、人均频次增速、单均 GB 增速、段 EBITDA margin、核心净利转换、FCF 转换、目标 P/E 都按当前案联动；估值对比页直接引三案矩阵，不被开关污染。Freight 增速三案共用 Base（非 load-bearing）。",
        "cases": CASES, "default": "Base",
        "triggers": [
            ("Bear", "robotaxi 收敛到 Waymo/Tesla 自营网络、绕过 Uber、take rate 被压缩；成熟市场出行/外卖饱和、增速掉到个位数；欧盟平台工人指令抬高成本；大盘倍数压缩。"),
            ("Base", "MAPC 维持双位数增长、Uber One 拉动频次；robotaxi 维持期权属性（Uber 做 AV 默认分发层）；段 EBITDA margin 沿广告/规模兑现缓升；目标倍数 23-25x。"),
            ("Bull", "robotaxi 期权证实（Uber 聚合 30+ AV 伙伴、成默认需求层）；广告(>$2B/+50%)与会员(50M)加速；段 margin 加速扩张；市场按成熟平台龙头给 25-28x。"),
        ],
        "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCF,
        "levers": [
            {"key": "mapc", "name": "MAPC 增速", "fmt": K.PCT, "cols": FCF,
             "vals": MAPC_G, "desc": "物理锚第一层：月活平台消费者增速，接全球可触达用户渗透天花板。FY25 MAPC 202M、Q1'26 199M(+17%)。",
             "stories": {"Bear": "成熟市场饱和、新兴市场获客成本上升（Keeta 等高补贴对手），MAPC 增速降到 5-12%。",
                         "Base": "全球渗透仍低(<5%出行支出)，MAPC 维持 9-15% 双位数增长。",
                         "Bull": "新兴市场 + 多产品交叉 + robotaxi 把不开车人群拉进平台，MAPC 维持 11-17%。"},
             "hist": hist_growth(MAPC)},
            {"key": "freq", "name": "人均频次增速", "fmt": K.PCT, "cols": FCF,
             "vals": FREQ_G, "desc": "物理锚第二层：Trips/MAPC 年增速。Uber One(50M 会员)拉动老用户用更勤——Trips 增速(+20%)持续高于 MAPC 增速(+18%)的证据。",
             "stories": {"Bear": "频次提升见顶，老用户使用强度趋稳，增速 0-1%。",
                         "Base": "Uber One 渗透 + 多产品交叉持续拉频次，增速 1.5-2.5%。",
                         "Bull": "会员深化 + 新品类(杂货/零售)把频次拉满，增速 2-4%。"},
             "hist": [None, None, None, None, None]},
            {"key": "asp", "name": "单均 GB 增速", "fmt": K.PCT, "cols": FCF,
             "vals": ASP_G, "desc": "物理锚第三层：单均 Gross Bookings($) 增速。历史单均稳定 ~$14.3（GB 增长是量而非价驱动）。前瞻含 mix(出行单价高于外卖) + 温和定价。",
             "stories": {"Bear": "竞争/robotaxi 价格战压单价，单均零增长或微降。",
                         "Base": "mix 改善 + 温和定价，单均 +0.5-1%。",
                         "Bull": "高价值出行占比上升 + 定价权，单均 +1.5-2%。"},
             "hist": [None, None, None, None, None]},
            {"key": "mobem", "name": "Mobility EBITDA margin(/GB)", "fmt": K.PCT, "cols": FCF,
             "vals": MOB_EM, "desc": "Mobility Adj EBITDA / Mobility GB。FY25 实际 8.1%（$7.9B/$97.5B）。规模效应 + 保险成本优化驱动。",
             "stories": {"Bear": "司机成本/监管(欧盟工人指令)+ 价格战，margin 卡在 7.8-8.0%。",
                         "Base": "规模效应 + 保险优化，margin 缓升到 8.3-8.9%。",
                         "Bull": "经营杠杆充分释放 + robotaxi 高利用率，margin 到 9.7%。"},
             "hist": [round(SEG_EBITDA[MOB][i] / GB_SEG[MOB][i], 4) for i in range(5)]},
            {"key": "delem", "name": "Delivery EBITDA margin(/GB)", "fmt": K.PCT, "cols": FCF,
             "vals": DEL_EM, "desc": "Delivery Adj EBITDA / Delivery GB。FY25 实际 3.9%（$3.57B/$90.9B），从负转正、改善最猛。广告(年化>$2B)是高毛利增量。",
             "stories": {"Bear": "Keeta/DoorDash 价格战 + 新兴市场补贴，margin 卡在 4.0-4.8%。",
                         "Base": "广告渗透(Delivery GB 广告占比破 2%)+ 规模，margin 升到 4.4-5.8%。",
                         "Bull": "广告加速到 DoorDash 水平 + 杂货品类放量，margin 到 7.0%。"},
             "hist": [round(SEG_EBITDA[DEL][i] / GB_SEG[DEL][i], 4) for i in range(5)]},
            {"key": "netconv", "name": "核心净利转换率(/EBITDA)", "fmt": K.PCT, "cols": FCF,
             "vals": NETCONV, "desc": "核心经营净利 / 总 Adj EBITDA，含 SBC(~$1.8B)/D&A/正常化税(~21%)/净利息。历史列为核心口径反推（剔一次性）。FY25 核心 ~$5.2B / Adj EBITDA $8.73B ≈ 0.59。",
             "stories": {"Bear": "SBC 摊薄 + 税率上行 + 利息，转换率卡在 0.50-0.55。",
                         "Base": "经营杠杆 + SBC 占比下降，转换率升到 0.56-0.64。",
                         "Bull": "税收优化 + SBC 控制 + 利息收入，转换率到 0.70。"},
             "hist": [round(NI_CORE[i] / ADJ_EBITDA[i], 4) if ADJ_EBITDA[i] > 0 else None for i in range(5)]},
            {"key": "fcfconv", "name": "FCF / Adj EBITDA 转换率", "fmt": K.PCT, "cols": FCF,
             "vals": FCFCONV, "desc": "自由现金流 / 总 Adj EBITDA。FY25 实际 112%（$9.76B/$8.73B，轻资产 + 保险准备金时间差使 >100%）。前瞻随营运资本正贡献递减、capex 温和上升而向 ~95% 归一。",
             "stories": {"Bear": "保险准备金时间差消退 + capex 上升，转换率回落到 0.86-0.95。",
                         "Base": "轻资产持续，转换率 0.93-1.05。",
                         "Bull": "营运资本正贡献延续，转换率维持高位 0.98-1.10。"},
             "hist": [round(FCF_H[i] / ADJ_EBITDA[i], 4) if ADJ_EBITDA[i] > 0 else None for i in range(5)]},
        ],
        "linked": [
            {"key": "sent", "name": "情绪/可见度值（P/E 第三层）", "fmt": K.N2,
             "src_sheet": S_MULT, "src_row0": ma["sent_row0"],
             "note": "来自估值倍数假设页。目标 P/E = 28x × 1.0 × 本行。"},
        ],
    })
    pk = f"'{S_MULT}'!{ma['pk_cell']}"; pr = f"'{S_MULT}'!{ma['pr_cell']}"
    swpe = sw["next_row"]
    K.lab(wb[S_SW], f"A{swpe}", "目标P/E（当前案）", b=True)
    for col in HC:  # 历史年无目标 P/E（亏损年情绪值为 n.m.），标 n.m.
        K.lab(wb[S_SW], f"{col}{swpe}", "n.m.", note=True)
    for col in FCF:
        K.fml(wb[S_SW], f"{col}{swpe}", f"={pk}*{pr}*{col}{sw['SWACT']['sent']}", K.MX, link=True)
    K.logic(wb[S_SW], f"L{swpe}", "目标 P/E = 质量锚 × 结构溢价 × 当前案情绪值；喂情景估值。")

    # 8 物理锚（MAPC × 频次 → Trips；单均 GB）
    anchor = K.write_anchor(wb.create_sheet(S_ANCHOR), {
        "title": "物理锚 — MAPC × 人均频次 → 总 Trips；单均 Gross Bookings",
        "all_cols": ALLC, "all_years": ALLY,
        "series": [
            ("MAPC 月活平台消费者 (M)", MAPC + [None] * 5, "物理锚第一层。前瞻=上年×(1+情景切换 MAPC 增速)。接全球可触达用户渗透天花板。", K.N0),
            ("总 Trips (M)", TRIPS + [None] * 5, "= MAPC × 人均年频次。前瞻=上年×(1+MAPC增速)×(1+频次增速)。", K.N0),
            ("单均 Gross Bookings ($)", [round(sum(GB_SEG[s][i] for s in [MOB, DEL, FRT]) * 1000 / TRIPS[i], 2) for i in range(5)] + [None] * 5,
             "总 GB / 总 Trips。历史稳定 ~$14.3（量驱动非价驱动）。前瞻=上年×(1+单均增速)。", K.N2),
        ],
        "yoy_row": "MAPC 月活平台消费者 (M)",
        "source_note": "MAPC/Trips=SEC 10-K/10-Q（年度=Q4 口径）；单均 GB=总 GB÷总 Trips 反推。",
        "role_note": "总 GB = 总 Trips × 单均 GB；分部 GB 按段比例分配。改 MAPC 增速 → 穿透到 Trips、GB、分部收入、Adj EBITDA、核心净利、目标价（连通性测试入口）。",
    })
    R_MAPC = anchor["row_of"]["MAPC 月活平台消费者 (M)"]
    R_TRIPS = anchor["row_of"]["总 Trips (M)"]
    R_ASP = anchor["row_of"]["单均 Gross Bookings ($)"]
    # 前瞻：MAPC、Trips、单均 GB
    for idx, col in enumerate(FCF):
        prev = FC[idx]
        K.fml(wb[S_ANCHOR], f"{col}{R_MAPC}", f"={prev}{R_MAPC}*(1+{K.R(S_SW, col + str(sw['SWACT']['mapc']))})", K.N0, link=True)
        K.fml(wb[S_ANCHOR], f"{col}{R_TRIPS}", f"={prev}{R_TRIPS}*(1+{K.R(S_SW, col + str(sw['SWACT']['mapc']))})*(1+{K.R(S_SW, col + str(sw['SWACT']['freq']))})", K.N0, link=True)
        K.fml(wb[S_ANCHOR], f"{col}{R_ASP}", f"={prev}{R_ASP}*(1+{K.R(S_SW, col + str(sw['SWACT']['asp']))})", K.N2, link=True)

    # 9 分部测算（总 GB → 分部 GB → take rate → 分部收入）
    # 分部 GB 占比（前瞻沿用 FY25 占比缓变：Mobility 升、Freight 降）
    gb_share_fy25 = {s: GB_SEG[s][-1] / sum(GB_SEG[x][-1] for x in [MOB, DEL, FRT]) for s in [MOB, DEL, FRT]}
    seg = K.write_segment_model(wb.create_sheet(S_SEG), {
        "title": "分部测算 — 总 GB → 分部 GB → take rate → 分部收入",
        "all_cols": ALLC, "all_years": ALLY, "logic_col": "N",
        "groups": [
            ("物理锚 → 总 GB", [
                ("总 Trips (M)", None, K.N0, "引自 [物理锚] 总 Trips。"),
                ("单均 GB ($)", None, K.N2, "引自 [物理锚] 单均 GB。"),
                ("总 Gross Bookings ($B)", None, K.N1, "= 总 Trips × 单均 GB / 1000。"),
            ]),
            ("分部 GB（占比分配）", [
                ("Mobility GB 占比", None, K.PCT, "前瞻=上年占比 + 0.3pct/年（出行占比缓升）。历史=实际。"),
                ("Mobility GB ($B)", None, K.N1, "= 总 GB × Mobility 占比。"),
                ("Delivery GB 占比", None, K.PCT, "前瞻=1 − Mobility占比 − Freight占比。历史=实际。"),
                ("Delivery GB ($B)", None, K.N1, "= 总 GB × Delivery 占比。"),
                ("Freight GB ($B)", None, K.N1, "前瞻=上年×(1+Freight增速,三案共用Base)，占比<3%。"),
            ]),
            ("take rate → 分部收入", [
                ("Mobility take rate", take_rate(MOB) + [0.305, 0.307, 0.309, 0.310, 0.311], K.PCT, "FY25 30.4%（含英国VAT/保险毛额口径）；前瞻温和升。"),
                ("Mobility 收入 ($B)", None, K.N1, "= Mobility GB × take rate。"),
                ("Delivery take rate", take_rate(DEL) + [0.192, 0.194, 0.196, 0.198, 0.200], K.PCT, "FY25 19.0%；广告拉动前瞻回升至 20%。"),
                ("Delivery 收入 ($B)", None, K.N1, "= Delivery GB × take rate。"),
                ("Freight 收入 ($B)", None, K.N1, "= Freight GB（principal 毛额列报，take≈100%）。"),
            ]),
        ],
    })
    m = seg["m"]
    # 总 Trips / 单均 GB 全列引物理锚
    for col in ALLC:
        K.fml(wb[S_SEG], f"{col}{m['总 Trips (M)']}", f"={K.R(S_ANCHOR, col + str(R_TRIPS))}", K.N0, link=True)
        K.fml(wb[S_SEG], f"{col}{m['单均 GB ($)']}", f"={K.R(S_ANCHOR, col + str(R_ASP))}", K.N2, link=True)
        K.fml(wb[S_SEG], f"{col}{m['总 Gross Bookings ($B)']}", f"={col}{m['总 Trips (M)']}*{col}{m['单均 GB ($)']}/1000", K.N1)
    # 历史 Mobility/Delivery 占比 = 实际
    for i, col in enumerate(HC):
        tot = sum(GB_SEG[x][i] for x in [MOB, DEL, FRT])
        K.inp(wb[S_SEG], f"{col}{m['Mobility GB 占比']}", round(GB_SEG[MOB][i] / tot, 4), None, K.PCT)
        K.inp(wb[S_SEG], f"{col}{m['Delivery GB 占比']}", round(GB_SEG[DEL][i] / tot, 4), None, K.PCT)
        K.inp(wb[S_SEG], f"{col}{m['Mobility GB ($B)']}", GB_SEG[MOB][i], None, K.N1)
        K.inp(wb[S_SEG], f"{col}{m['Delivery GB ($B)']}", GB_SEG[DEL][i], None, K.N1)
        K.inp(wb[S_SEG], f"{col}{m['Freight GB ($B)']}", GB_SEG[FRT][i], None, K.N1)
    # 前瞻占比与 GB
    for idx, col in enumerate(FCF):
        prev = FC[idx]
        K.fml(wb[S_SEG], f"{col}{m['Mobility GB 占比']}", f"={prev}{m['Mobility GB 占比']}+0.003", K.PCT)
        K.fml(wb[S_SEG], f"{col}{m['Freight GB ($B)']}", f"={prev}{m['Freight GB ($B)']}*(1+{FRT_G[idx]})", K.N1)
        K.fml(wb[S_SEG], f"{col}{m['Delivery GB 占比']}", f"=1-{col}{m['Mobility GB 占比']}-{col}{m['Freight GB ($B)']}/{col}{m['总 Gross Bookings ($B)']}", K.PCT)
        K.fml(wb[S_SEG], f"{col}{m['Mobility GB ($B)']}", f"={col}{m['总 Gross Bookings ($B)']}*{col}{m['Mobility GB 占比']}", K.N1)
        K.fml(wb[S_SEG], f"{col}{m['Delivery GB ($B)']}", f"={col}{m['总 Gross Bookings ($B)']}*{col}{m['Delivery GB 占比']}", K.N1)
    # take rate → 收入（全列）
    for col in ALLC:
        K.fml(wb[S_SEG], f"{col}{m['Mobility 收入 ($B)']}", f"={col}{m['Mobility GB ($B)']}*{col}{m['Mobility take rate']}", K.N1)
        K.fml(wb[S_SEG], f"{col}{m['Delivery 收入 ($B)']}", f"={col}{m['Delivery GB ($B)']}*{col}{m['Delivery take rate']}", K.N1)
        K.fml(wb[S_SEG], f"{col}{m['Freight 收入 ($B)']}", f"={col}{m['Freight GB ($B)']}", K.N1)

    # 10 利润与收入假设（分部收入 → 段 EBITDA → 核心净利/EPS）
    # Uber 用段 EBITDA margin(相对 GB)，与 GOOGL 的段 OPM 不同 → 自定义利润段，不用 profit_terms
    fr = K.write_fundamentals(wb.create_sheet(S_FUND), {
        "title": "利润与收入假设 — 分部收入 → 段 Adj EBITDA → 核心净利/EPS/BPS",
        "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCF, "logic_col": "N",
        "s_hist": S_HIST, "ha": ha, "share_fix_col": "F",
        "assum_groups": [
            ("段 EBITDA margin（相对 GB）", [
                {"name": "Mobility EBITDA margin", "vals": [round(SEG_EBITDA[MOB][i] / GB_SEG[MOB][i], 4) for i in range(5)] + [None] * 5, "fmt": K.PCT,
                 "logic": "Mobility Adj EBITDA / Mobility GB；历史反推，前瞻链情景切换。",
                 "link": {"sheet": S_SW, "row": sw["SWACT"]["mobem"]}},
                {"name": "Delivery EBITDA margin", "vals": [round(SEG_EBITDA[DEL][i] / GB_SEG[DEL][i], 4) for i in range(5)] + [None] * 5, "fmt": K.PCT,
                 "logic": "Delivery Adj EBITDA / Delivery GB；历史反推（早年负），前瞻链情景切换。",
                 "link": {"sheet": S_SW, "row": sw["SWACT"]["delem"]}},
            ]),
            ("核心净利与 FCF 转换", [
                {"name": "核心净利转换率", "vals": [round(NI_CORE[i] / ADJ_EBITDA[i], 4) if ADJ_EBITDA[i] > 0 else None for i in range(5)] + [None] * 5, "fmt": K.PCT,
                 "logic": "核心经营净利 / 总 Adj EBITDA（剔一次性）；历史反推（早年 EBITDA≤0 标 n.m.），前瞻链情景切换。",
                 "nm_cols": ["B", "C"], "link": {"sheet": S_SW, "row": sw["SWACT"]["netconv"]}},
                {"name": "FCF转换率", "vals": [round(FCF_H[i] / ADJ_EBITDA[i], 4) if ADJ_EBITDA[i] > 0 else None for i in range(5)] + [None] * 5, "fmt": K.PCT,
                 "logic": "FCF / 总 Adj EBITDA；历史反推，前瞻链情景切换（喂 DCF 镜头）。",
                 "nm_cols": ["B", "C"], "link": {"sheet": S_SW, "row": sw["SWACT"]["fcfconv"]}},
                {"name": "留存率", "vals": [1.0, 1.0, 1.0, 1.0, 1.0] + [None] * 5, "fmt": K.PCT,
                 "logic": "无分红，净利全留存（用于 BPS，主估值不依赖 P/B）。", "vals_fwd_const": 1.0},
            ]),
        ],
        "segments": [
            {"name": MOB, "hist_row": MOB, "fwd": {"sheet": S_SEG, "row": m["Mobility 收入 ($B)"]}},
            {"name": DEL, "hist_row": DEL, "fwd": {"sheet": S_SEG, "row": m["Delivery 收入 ($B)"]}},
            {"name": FRT, "hist_row": FRT, "fwd": {"sheet": S_SEG, "row": m["Freight 收入 ($B)"]}},
        ],
        # profit_terms 用不上（Uber EBITDA 挂 GB 不挂收入），给个占位避免 OP 行报错：用收入×0 占位，真正 EBITDA/NI 下面手算
        "profit_terms": [([FRT], "留存率", False)],
        "conv_assum": "核心净利转换率", "retention_assum": "留存率",
        "note_text": "★ 净利由『分部 GB × 段 EBITDA margin = 段 Adj EBITDA → 总 Adj EBITDA → × 核心净利转换』推导。本表 OP 行为占位（Uber 主口径是 Adj EBITDA 不是营业利润），真正的 Adj EBITDA / 核心净利 / FCF 在下方补行。历史核心净利=剔一次性 DTA 释放($6.4B/$5.0B)与投资重估，前瞻=核心口径。",
    })
    # ★ 在 fundamentals 表底部补：总 Adj EBITDA / 核心净利 / FCF 行（覆盖 profit_terms 的占位 NI）
    # 找到 NI 行，改成：Adj EBITDA = Σ段GB×段margin + Corporate；核心净利 = Adj EBITDA × 核心转换
    fund_ws = wb[S_FUND]
    NI_row = fr["NI"]
    am = fr["am"]
    # GB 行引用（来自分部测算）
    gb_mob = lambda c: K.R(S_SEG, c + str(m["Mobility GB ($B)"]))
    gb_del = lambda c: K.R(S_SEG, c + str(m["Delivery GB ($B)"]))
    mobem_r, delem_r = am["Mobility EBITDA margin"], am["Delivery EBITDA margin"]
    # Corporate 行：插在 OP 行（占位）位置上方逻辑——直接重写 NI 行为核心净利，并新增 EBITDA/FCF 行到表末
    # 重写 OP 行 = 总 Adj EBITDA
    OP_row = fr["OP"]
    K.lab(fund_ws, f"A{OP_row}", "总 Adj EBITDA($B)", b=True)
    for i, col in enumerate(ALLC):
        if col in HC:
            K.fml(fund_ws, f"{col}{OP_row}", f"={ADJ_EBITDA[i]}", K.N1)
        else:
            corp = CORP[-1] * (1 + 0.06) ** (FCF.index(col))  # Corporate 温和增长
            K.fml(fund_ws, f"{col}{OP_row}",
                  f"={gb_mob(col)}*{col}{mobem_r}+{gb_del(col)}*{col}{delem_r}+({corp:.2f})", K.N1)
    # 重写 NI 行 = 核心净利 = Adj EBITDA × 核心净利转换
    K.lab(fund_ws, f"A{NI_row}", "核心净利($B)", b=True)
    for i, col in enumerate(ALLC):
        if col in HC:
            K.fml(fund_ws, f"{col}{NI_row}", f"={NI_CORE[i]}", K.N1)
        else:
            K.fml(fund_ws, f"{col}{NI_row}", f"={col}{OP_row}*{col}{am['核心净利转换率']}", K.N1)
    # 新增 FCF 行到表末（跳过 ROE 之后的 band+mtext 说明区，落到安全空行）
    fcf_row = fr["ROE"] + 8
    K.band(fund_ws, fcf_row - 1, "现金盈利锚（喂 DCF）", len(ALLC) + 1)
    K.lab(fund_ws, f"A{fcf_row}", "自由现金流 FCF($B)")
    for i, col in enumerate(ALLC):
        if col in HC:
            K.fml(fund_ws, f"{col}{fcf_row}", f"={FCF_H[i]}", K.N1)
        else:
            K.fml(fund_ws, f"{col}{fcf_row}", f"={col}{OP_row}*{col}{am['FCF转换率']}", K.N1)
    K.logic(fund_ws, f"N{fcf_row}", "FCF = 总 Adj EBITDA × FCF转换率；喂 DCF 镜头。")

    # 11 情景估值（自定义：P/E 主线 + DCF 第二镜头）
    val = write_uber_valuation(wb.create_sheet(S_VAL), {
        "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCF,
        "s_hist": S_HIST, "ha": ha, "s_fund": S_FUND, "fr": fr, "fcf_row": fcf_row,
        "s_switch": S_SW, "target_row": swpe, "sw_cell": sw["sw_cell"],
        "price_now": PX_NOW, "fx_fwd": FX_FC,
    })

    # 12 估值对比（三案并排）
    swb = sw["SWB"]
    def cr(key, ci): return swb[key] + ci
    def prevcol(col): return ALLC[ALLC.index(col) - 1]
    GB_H = {s: [GB_SEG[s][i] for i in range(5)] for s in [MOB, DEL, FRT]}
    corp_fwd = lambda c: f"({CORP[-1] * (1 + 0.06) ** (FCF.index(c)):.2f})" if c in FCF else "0"
    # 历史锚（静态）
    def hist_gb(s): return lambda c, ci, a: f"={GB_H[s][HC.index(c)]}"
    rows = [
        {"key": "mapc", "label": "MAPC (M)", "fmt": K.N0,
         "hist": lambda c, ci, a: f"={K.R(S_ANCHOR, c + str(R_MAPC))}",
         "fwd": lambda c, j, ci, a: (f"={K.R(S_ANCHOR, c + str(R_MAPC))}" if j == 0 else f"={prevcol(c)}{a['mapc']}*(1+{K.R(S_SW, c + str(cr('mapc', ci)))})")},
        {"key": "trips", "label": "总 Trips (M)", "fmt": K.N0,
         "hist": lambda c, ci, a: f"={K.R(S_ANCHOR, c + str(R_TRIPS))}",
         "fwd": lambda c, j, ci, a: (f"={K.R(S_ANCHOR, c + str(R_TRIPS))}" if j == 0 else f"={prevcol(c)}{a['trips']}*(1+{K.R(S_SW, c + str(cr('mapc', ci)))})*(1+{K.R(S_SW, c + str(cr('freq', ci)))})")},
        {"key": "asp", "label": "单均 GB ($)", "fmt": K.N2,
         "hist": lambda c, ci, a: f"={K.R(S_ANCHOR, c + str(R_ASP))}",
         "fwd": lambda c, j, ci, a: (f"={K.R(S_ANCHOR, c + str(R_ASP))}" if j == 0 else f"={prevcol(c)}{a['asp']}*(1+{K.R(S_SW, c + str(cr('asp', ci)))})")},
        {"key": "gb", "label": "总 GB ($B)", "fmt": K.N1, "bold": True,
         "hist": lambda c, ci, a: f"={sum(GB_SEG[s][HC.index(c)] for s in [MOB, DEL, FRT]):.1f}",
         "fwd": lambda c, j, ci, a: f"={c}{a['trips']}*{c}{a['asp']}/1000"},
        {"key": "mobgb", "label": "Mobility GB ($B)", "fmt": K.N1, "hist": hist_gb(MOB),
         "fwd": lambda c, j, ci, a: f"={c}{a['gb']}*{K.R(S_SEG, c + str(m['Mobility GB 占比']))}"},
        {"key": "delgb", "label": "Delivery GB ($B)", "fmt": K.N1, "hist": hist_gb(DEL),
         "fwd": lambda c, j, ci, a: f"={c}{a['gb']}*{K.R(S_SEG, c + str(m['Delivery GB 占比']))}"},
        {"key": "rev", "label": "总营收 ($B)", "fmt": K.N1, "bold": True,
         "hist": lambda c, ci, a: f"={REV_TOTAL[HC.index(c)]:.1f}",
         "fwd": lambda c, j, ci, a: f"={c}{a['mobgb']}*{K.R(S_SEG, c + str(m['Mobility take rate']))}+{c}{a['delgb']}*{K.R(S_SEG, c + str(m['Delivery take rate']))}+{K.R(S_SEG, c + str(m['Freight 收入 ($B)']))}"},
        {"key": "ebitda", "label": "Adj EBITDA ($B)", "fmt": K.N1,
         "hist": lambda c, ci, a: f"={ADJ_EBITDA[HC.index(c)]:.2f}",
         "fwd": lambda c, j, ci, a: f"={c}{a['mobgb']}*{K.R(S_SW, c + str(cr('mobem', ci)))}+{c}{a['delgb']}*{K.R(S_SW, c + str(cr('delem', ci)))}+{corp_fwd(c)}"},
        {"key": "ni", "label": "核心净利 ($B)", "fmt": K.N1, "bold": True,
         "hist": lambda c, ci, a: f"={NI_CORE[HC.index(c)]:.2f}",
         "fwd": lambda c, j, ci, a: f"={c}{a['ebitda']}*{K.R(S_SW, c + str(cr('netconv', ci)))}"},
        {"key": "eps", "label": "核心 EPS ($)", "fmt": K.N2,
         "hist": lambda c, ci, a: f"={NI_CORE[HC.index(c)] * 1000 / SHARES[HC.index(c)]:.2f}",
         "fwd": lambda c, j, ci, a: f"={c}{a['ni']}*1000/{SHARES_M}"},
        {"key": "tpe", "label": "目标 P/E", "fmt": K.MX,
         "hist": lambda c, ci, a: (f"={K.R(S_HIST, c + str(ha['HPX']))}/{NI_CORE[HC.index(c)] * 1000 / SHARES[HC.index(c)]:.2f}" if NI_CORE[HC.index(c)] > 0 else '="N/M"'),
         "fwd": lambda c, j, ci, a: f"={K.R(S_SW, c + str(cr('sent', ci)))}*{PEAK_PE}*{PREMIUM}"},
        {"key": "px", "label": "隐含股价 ($)", "fmt": K.PX, "bold": True, "out": True,
         "hist": lambda c, ci, a: f"={K.R(S_HIST, c + str(ha['HPX']))}",
         "fwd": lambda c, j, ci, a: f"={c}{a['tpe']}*{c}{a['eps']}+0*{c}{a['mapc']}"},
        {"key": "ipe", "label": "隐含 forward P/E（体检）", "fmt": K.MX,
         "hist": lambda c, ci, a: '="—"',
         "fwd": lambda c, j, ci, a: f"={c}{a['px']}/{c}{a['eps']}"},
        {"key": "up", "label": "历史回测 / 前瞻 vs 现价", "fmt": K.PCT,
         "hist": lambda c, ci, a: f"={c}{a['px']}/{K.R(S_HIST, c + str(ha['HPX']))}-1",
         "fwd": lambda c, j, ci, a: f"={c}{a['px']}/{PX_NOW}-1"},
    ]
    cmp = K.write_comparison(wb.create_sheet(S_CMP), {
        "title": "估值对比 — Bear / Base / Bull 三案并排",
        "intro": "三案从 MAPC、人均频次、单均 GB、段 EBITDA margin、核心净利转换、目标 P/E 同一条链推导。主判断年取 2027E（2026 已被高可见度 GB 指引锚定，真正分歧在 2027 之后）。隐含价用核心经营 EPS×目标 P/E。历史列同链填实际、隐含价≈实际年末价=内置回测。",
        "case_names": CASES, "all_cols": ALLC, "all_years": ALLY, "hist_cols": HC, "fwd_cols": FCF,
        "block_start": 22, "rows": rows,
        "summary": {
            "band": "2027E 三案摘要（主判断年）", "target_col": "H",
            "rows": [
                ("总 GB($B)", "gb", K.N1, "Trips × 单均 GB。", False),
                ("总营收($B)", "rev", K.N1, "分部 GB × take rate。", False),
                ("Adj EBITDA($B)", "ebitda", K.N1, "段 GB × 段 margin + Corporate。", False),
                ("核心净利($B)", "ni", K.N1, "Adj EBITDA × 核心净利转换（剔一次性）。", False),
                ("核心 EPS($)", "eps", K.N2, "主估值分母。", False),
                ("目标 P/E", "tpe", K.MX, "质量锚×溢价×情绪。", False),
                ("隐含股价($)", "px", K.PX, "P/E 主线输出。", True),
                ("隐含 forward P/E", "ipe", K.MX, "体检，非额外结论。", False),
                ("vs 现价", "up", K.PCT, "现价 reality check。", True),
            ],
            "mcap": {"label": "隐含市值($B)", "key": "px", "expr": f"*{SHARES_M}/1000", "note": "隐含股价 × 股本。"},
            "concl": "Base 2027E 核心 EPS×23x 隐含价显著高于现价 → 市场用『AV 颠覆空头情景』给 Uber 定价（forward P/E 压到历史下沿）。Bull 靠 robotaxi 期权证实 + 广告/会员加速 + 倍数扩张给更大上行；Bear 靠 AV 去中介化 + take rate 压缩给下行。风险收益偏多 → BUY。",
        },
    })

    # 13 仪表盘
    K.write_dashboard(wb.create_sheet(S_DASH), {
        "title": "综合判断仪表盘 — 基本面拐点 / 估值错位 / 催化剂 / 情绪",
        "usage": "把模型压成投后跟踪语言：哪些指标验证 Base，哪些把模型推向 Bear/Bull。B 列公式直接引模型输出。",
        "blocks": [
            {"title": "A. 基本面拐点", "rows": [
                ("Q1'26 GB/Adj EBITDA", "+25% / +33%", "经营加速且 margin 扩张；TTM FCF $9.8B record。", True),
                ("MAPC/Trips/Uber One", "199M(+17%) / 3.64B(+20%) / 50M", "物理锚强劲；Trips 增速>MAPC=频次提升（会员拉动）。"),
                ("2027E Base 核心净利", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Base']['ni']))}", "fmt": K.N1, "fill": True}, "若 2027 核心净利不及该路径，Base 估值先降。"),
            ]},
            {"title": "B. 估值错位", "rows": [
                ("当前股价", {"inp": PX_NOW, "fmt": K.PX, "fill": True}, "reality check，不反向拟合。"),
                ("Base 2027E 隐含价", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Base']['px']))}", "fmt": K.PX, "fill": True}, "主判断输出。"),
                ("Base 2027E vs 现价", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Base']['up']))}", "fmt": K.PCT, "fill": True}, "正且大 = 市场用 AV 空头情景错杀。"),
                ("Bear 2027E 隐含价", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Bear']['px']))}", "fmt": K.PX}, "下行风险价：AV 去中介化 + take rate 压缩 + 倍数压缩。"),
                ("Bull 2027E 隐含价", {"fml": f"={K.R(S_CMP, 'H' + str(cmp['CMPA']['Bull']['px']))}", "fmt": K.PX}, "上行：robotaxi 期权证实 + 广告/会员加速 + 倍数扩张。"),
            ]},
            {"title": "C. 催化剂", "rows": [
                ("Bull 触发", "Waymo 新城市继续选 Uber 分发；AV 行程持续 10x；广告破 $3B；Uber One 破 60M。", "先抬段 margin/频次，再抬目标 P/E。"),
                ("Bear 触发", "Waymo/Tesla 新城市全部自营绕过 Uber；take rate 首次环比下滑；欧盟工人指令重罚。", "先砍 MAPC/段 margin，再把目标 P/E 压到 16x 以下。"),
                ("最易错处", "用报表 EPS（含 $6.4B/$5.0B DTA 释放）直接套倍数会造『假便宜』（headline P/E 17.9x）。", "本模型用核心经营 EPS，避免这个陷阱。"),
            ]},
            {"title": "D. 综合判断", "rows": [
                ("一句话结论", "BUY：robotaxi 被市场误读为颠覆威胁、年内 -10% 错杀；实际 Uber 是 AV 需求聚合分发层，robotaxi 是免费期权。平台已规模化盈利。", "上行看 robotaxi 期权 + 广告/会员；下行看 AV 收敛到自营赢家。", True),
            ]},
        ],
        "final": {"band": "最终判断",
                  "text": "Uber 的 alpha 问题：市场把 robotaxi 当成 Uber 的颠覆者，2025-11 起从高点回撤约 28%、年内约 -10%。本模型用核心经营 EPS（剔除一次性 DTA 释放）下，现价对应的 forward P/E 已压到历史下沿——市场在用『AV 去中介化空头情景』给一家已规模化盈利、FCF $9.8B/转化 112%、与 30+ AV 公司合作的需求聚合平台定价。Base 2027E 隐含价显著高于现价 → 预期被错杀的中期预期差。最大未证伪下行是 robotaxi 收敛到 1-2 家自营赢家(Waymo/Tesla)绕过 Uber。评级 BUY。"},
        "tracking": {
            "intro": "投后跟踪按季报滚动更新。",
            "rows": [
                ("__band__", "一、物理锚"),
                ("MAPC/Trips 增速", "Q1'26 199M(+17%)/3.64B(+20%)", "命门：用户与频次增长持续性", "季报运营指标", "MAPC 增速掉到个位数 → 转 Bear"),
                ("__band__", "二、盈利质量"),
                ("段 EBITDA margin", "Mob 8.1% / Del 3.9%(FY25/GB)", "命门：经营杠杆 + 广告兑现", "季报分部 Adj EBITDA", "margin 停滞 → 下调段 margin 路径"),
                ("__band__", "三、robotaxi（核心争议）"),
                ("Waymo 新城市分发选择", "Austin/Atlanta 用 Uber；SF/DC/London 自营", "命门：AV 是期权还是去中介化", "Uber/Waymo 公告", "新城市全部自营绕过 Uber → 转 Bear 砍 take rate"),
                ("Tesla Cybercab 规模", "~20-25 辆(2026 中)", "命门：垂直整合能否跨城规模化", "Tesla 季报/NHTSA", "跨城无监督 + 单位经济碾压 → 威胁前置"),
                ("__band__", "四、资本回报"),
                ("FCF / 回购", "FY25 FCF $9.8B / 回购 $6.5B(剩 $19.2B)", "命门：现金返还持续性", "季报现金流", "FCF 转化跌破 90% → 下调 DCF"),
            ],
        },
    })

    # input.json
    payload = {
        "ticker": "UBER",
        "built_at": "2026-06-23",
        "currency": "USD",
        "current_price": PX_NOW, "market_cap_b": 1467.67, "shares_m": SHARES_M,
        "rating": "BUY", "target_price": None,  # 由模型 Base 2027E 隐含价决定，见 xlsx
        "method": "P/E 主线（核心经营 EPS×目标 P/E）+ DCF（核心 FCF）交叉；物理锚=MAPC×频次→Trips→GB→take rate→收入→段 EBITDA→核心净利。",
        "core_eps_note": "核心 EPS 剔除一次性 DTA 释放（FY24 +$6.4B、FY25 +$5.0B）+ 股权投资公允价值重估；报表净利 FY24/25 $9.86B/$10.05B 失真，核心净利 ~$1.6B/~$5.2B。",
        "historical_financials_b": {
            "gross_bookings": [90.4, 115.4, 137.87, 162.77, 193.45],
            "segments_gb_fy2025": {k: v[-1] for k, v in GB_SEG.items()},
            "revenue": REV_TOTAL,
            "adj_ebitda": ADJ_EBITDA,
            "net_income_reported": NI_REP, "net_income_core": NI_CORE,
            "fcf": FCF_H, "ocf_2025": 10.10, "capex_2025": 0.336,
            "mapc_m": MAPC, "trips_m": TRIPS,
        },
        "anchor": {"type": "platform-mapc", "mapc_m": MAPC, "trips_m": TRIPS,
                   "unit_gb_usd": [round(sum(GB_SEG[s][i] for s in [MOB, DEL, FRT]) * 1000 / TRIPS[i], 2) for i in range(5)]},
        "scenario": {"target_pe": TARGET_PE, "peak_pe": PEAK_PE, "pe_sentiment": PE_SENT,
                     "mapc_growth": MAPC_G, "freq_growth": FREQ_G, "asp_growth": ASP_G,
                     "mob_ebitda_margin": MOB_EM, "del_ebitda_margin": DEL_EM,
                     "net_conversion": NETCONV, "fcf_conversion": FCFCONV},
        "consensus": {"target_mean": 104.48, "target_high": 150, "target_low": 70,
                      "rating": "strong_buy", "n_analysts": 51,
                      "fy26e_eps_headline": 3.03, "fy27e_eps_headline": 4.47,
                      "fy26e_rev_b": 58.16, "fy27e_rev_b": 67.03,
                      "sellside_tp": {"Evercore": 150, "GS": 115, "JPM": 105, "BofA": 103,
                                      "MS": 100, "Wedbush": 84, "Melius": 73}},
        "sources": {
            "official": ["SEC EDGAR CIK 0001543151: FY2025 10-K (2026-02-13) / FY2024 10-K (2025-02-14) / Q1'26 10-Q (2026-05-06)"],
            "data_api": "SEC XBRL financials + Yahoo estimates/quote 数据快照 2026-06-23",
            "research_web": ["一致目标价 $104.48 (Yahoo 51 家)", "Evercore $150", "GS $115", "JPM $105", "BofA $103", "MS $100", "Melius $73"],
        },
    }
    os.makedirs(VAULT, exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    K.finalize(wb)
    wb.save(OUT_XLSX)
    print("saved:", OUT_XLSX)
    print("sheets:", wb.sheetnames)
    return cmp


def write_uber_valuation(ws, d):
    """P/E 主线逐年隐含价 + DCF 第二镜头交叉。P/B 不适用（轻资产）。"""
    allc, ally = d["all_cols"], d["all_years"]
    hc, fcf = d["hist_cols"], d["fwd_cols"]
    s_hist, ha = d["s_hist"], d["ha"]
    s_fund, fr = d["s_fund"], d["fr"]
    s_sw = d["s_switch"]
    target_row = d["target_row"]
    fcf_row = d["fcf_row"]
    price_now = d["price_now"]
    K.hdr(ws, 1, "情景估值 — 当前案逐年 P/E 主线 + DCF 交叉验证", 11)
    K.lab(ws, "L1", "当前情景→", note=True); K.fml(ws, "M1", f"={K.R(s_sw, d['sw_cell'])}", K.N0, link=True); ws["M1"].fill = K.CUR
    r = K.mtext(ws, 2, "本页随情景切换 B2 变化。历史列用真实价反推倍数（事实，亏损年标 N/M）；前瞻列=当前案目标 P/E × 核心 EPS（判断），并以 DCF（核心 FCF）交叉。核心 EPS 已剔除一次性 DTA 释放与投资重估。P/B 不适用（轻资产、账面无意义）。", "K", 2)
    K.lab(ws, f"A{r}", "($/股；倍数；$B 财务)", b=True)
    for col, y in zip(allc, ally):
        ws[f"{col}{r}"] = y; ws[f"{col}{r}"].font = K.BF; ws[f"{col}{r}"].fill = K.CH
    r += 1
    eps = lambda c: K.R(s_fund, c + str(fr["EPS"]))
    ni = lambda c: K.R(s_fund, c + str(fr["NI"]))
    pxc = lambda c: K.R(s_hist, c + str(ha["HPX"]))

    K.band(ws, r, "P/E 主线：历史=实际价反推（亏损标 N/M）；前瞻=目标 P/E × 核心 EPS", 11); r += 1
    pe_row = r; K.lab(ws, f"A{r}", "目标 P/E（历史=实际；前瞻=当前案）")
    for c in hc:
        K.fml(ws, f"{c}{r}", f'=IF({ni(c)}<=0,"N/M",{pxc(c)}/{eps(c)})', K.MX, link=True)
    for c in fcf:
        K.fml(ws, f"{c}{r}", f"={K.R(s_sw, c + str(target_row))}", K.MX, link=True)
    r += 1
    px_row = r; K.lab(ws, f"A{r}", "隐含股价 P/E主线 ($)", b=True); ws[f"A{r}"].fill = K.OUT
    for c in hc:
        K.fml(ws, f"{c}{r}", f"={pxc(c)}", K.PX, link=True)
    for c in fcf:
        K.fml(ws, f"{c}{r}", f"={c}{pe_row}*{eps(c)}", K.PX, link=True)
    r += 1
    up_row = r; K.lab(ws, f"A{r}", "较现价上行/下行")
    for c in fcf:
        K.fml(ws, f"{c}{r}", f"={c}{px_row}/{price_now}-1", K.PCT)
    r += 1
    fpe_row = r; K.lab(ws, f"A{r}", "当下现价对应 forward P/E（核心）")
    for c in fcf:
        K.fml(ws, f"{c}{r}", f"={price_now}/{eps(c)}", K.MX)
    r += 2

    # DCF（第二镜头，核心 FCF 折现）
    K.band(ws, r, "镜头二 DCF（核心 FCF 折现，WACC 8.5% / 永续 g 3.5%）", 11); r += 1
    K.lab(ws, f"A{r}", "2027E 核心 FCF ($B)")
    K.fml(ws, f"B{r}", f"={K.R(s_fund, 'H' + str(fcf_row))}", K.N1, link=True); dcf_fcf = r; r += 1
    K.lab(ws, f"A{r}", "WACC / 永续增长 g")
    K.inp(ws, f"B{r}", 0.085, None, K.PCT); K.inp(ws, f"C{r}", 0.035, None, K.PCT); dcf_w = r; r += 1
    K.lab(ws, f"A{r}", "永续法 EV ($B) = FCF×(1+g)/(WACC−g)")
    K.fml(ws, f"B{r}", f"=B{dcf_fcf}*(1+C{dcf_w})/(B{dcf_w}-C{dcf_w})", K.N0); dcf_ev = r; r += 1
    K.lab(ws, f"A{r}", "净现金/股调整 ($, FY25 现金$7.1B−债$10.5B≈−$1.6/股)")
    K.inp(ws, f"B{r}", -1.6, None, K.PX); dcf_nc = r; r += 1
    K.lab(ws, f"A{r}", "DCF 隐含股价 ($)", b=True); ws[f"A{r}"].fill = K.OUT
    K.fml(ws, f"B{r}", f"=B{dcf_ev}*1000/{SHARES_M}+B{dcf_nc}", K.PX); dcf_px = r; r += 1
    K.lab(ws, f"A{r}", "DCF vs 现价")
    K.fml(ws, f"B{r}", f"=B{dcf_px}/{price_now}-1", K.PCT); r += 2

    K.band(ws, r, "两镜头三角 + 方法", 11); r += 1
    K.mtext(ws, r, "两镜头（P/E 主线 / DCF）应收敛在一个区间，分歧即是市场赌点。P/E 主线随情景切换给三档（核心 EPS×目标倍数）；DCF 用核心 FCF（Uber 轻资产、FCF/EBITDA 转化 ~95-112%，FCF 是比 GAAP 净利干净的现金盈利锚）。两者 Base 收敛在显著高于现价 $72 的区间 → 验证市场用 AV 空头情景错杀、BUY。P/B 不做（轻资产平台账面非估值分母）。", "K", 4)
    K.set_widths(ws, 30, allc, 11)
    return {"pe": pe_row, "px": px_row, "dcf_px": dcf_px}


if __name__ == "__main__":
    build()
