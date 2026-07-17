#!/usr/bin/env python3
"""extract_model_tables.py — 从「已 recalc 的」估值模型 xlsx 抽取关键报表为 Markdown 表格。

用途:equity-research 报告组装阶段(Phase 4),把模型里的关键 sheet 以 Markdown 表格
形式嵌进报告『估值分析』章节:让报告的三情景数字与模型同源、自动一致、可随模型刷新。
单一真相源永远是模型 xlsx;报告里的表是它的投影,不手敲。

★ 前置硬要求:必须先对该 xlsx 跑 recalc.py --backend auto(跨平台重算回写缓存值)。
   否则 openpyxl data_only=True 读公式格得 None,本脚本会显式报 "未 recalc" 并拒绝产出空表。

抽取的标准表(跨公司通用,按 sheet 名定位,容忍行布局差异):
  1. 三情景估值汇总   ← 『估值对比』里 Bear/Base/Bull 表头下的汇总块(核心表)
  2. 历史财务与估值   ← 『历史财务与估值』整页(全行: 分部/毛利率/净利/权益/股本/FX/股价/P-E/P-B/市值; --with-history)
  3. 盈利预测表       ← 『利润与收入假设』Base 逐年明细(分部收入/占比/增速/利润率/净利/EPS/BPS/ROE; --forecast)
  4. 三情景逐年隐含价 ← 『估值对比』三案块各自的『隐含价』逐年行(--ladder)

用法:
  python extract_model_tables.py <model.xlsx>                 # 仅三情景汇总
  python extract_model_tables.py <model.xlsx> -o tables.md    # 写文件
  python extract_model_tables.py <model.xlsx> --with-history  # 附历史财务与估值整页
  python extract_model_tables.py <model.xlsx> --forecast      # 附盈利预测表(Base 逐年)
  python extract_model_tables.py <model.xlsx> --ladder        # 附三情景逐年隐含价
  python extract_model_tables.py <model.xlsx> --full          # 以上全部(研报组装默认用这个)

返回:Markdown 字符串。可被 import 复用:from extract_model_tables import scenario_table
"""
import sys, re, argparse
import openpyxl
from openpyxl.utils import get_column_letter

# Windows 控制台常默认为 GBK;Markdown 含中文时强制 UTF-8,避免 print 失败或乱码。
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


def _fmt(label, v):
    """按行标签推断数值格式(%/倍数/股价/$B/通用)。label 决定语义,而非死板猜。"""
    if v is None:
        return ""
    if not isinstance(v, (int, float)):
        return str(v).replace("\n", " ").strip()
    lab = label.lower()
    # 百分比:vs现价/增速/YoY/涨幅
    if any(k in lab for k in ("vs现价", "vs 现价", "upside", "增速", "涨幅", "yoy")):
        return f"{v*100:+.1f}%"
    # 任何 "率" 一律百分比(汇率除外:但汇率不在标准抽取的指标白名单里)
    if ("%" in label) or ("率" in label and "汇率" not in label) or \
       any(k in lab for k in ("margin", "占比", "强度", "roe", "roic")):
        return f"{v*100:.1f}%"
    # 倍数
    if any(k in lab for k in ("p/b", "p/e", "p/s", "倍", "ev/ebitda", "ev/sales", "multiple")):
        return f"{v:.1f}x"
    # 股价 / 隐含价(整数千分位)
    if any(k in lab for k in ("价", "股价", "krw", "implied", "/股")):
        return f"{v:,.0f}"
    # 金额($B / 收入 / 净利 / 市值 / 利润 / 权益 / 毛利)
    if any(k in lab for k in ("$b", "收入", "净利", "市值", "利润", "ebitda", "营收", "现金", "fcf", "权益", "毛利")):
        if abs(v) < 10:
            return f"{v:,.2f}"
        return f"{v:,.1f}" if abs(v) < 1000 else f"{v:,.0f}"
    # 通用
    if abs(v) < 10 and v != int(v):
        return f"{v:.2f}"
    return f"{v:,.0f}"


def _find_sheet(wb, kw):
    return next((n for n in wb.sheetnames if kw in n), None)


def _has_cache(ws):
    """判断 sheet 是否有缓存计算值(随便找一个数字格)。全 None 视为没 recalc。"""
    for row in ws.iter_rows(min_row=3, max_row=min(ws.max_row, 60)):
        for c in row:
            if isinstance(c.value, (int, float)):
                return True
    return False


def scenario_table(path):
    """抽『估值对比』Bear/Base/Bull 汇总块 → Markdown。核心表。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    sn = _find_sheet(wb, "估值对比")
    if not sn:
        return None, "无『估值对比』sheet"
    ws = wb[sn]
    if not _has_cache(ws):
        return None, "未 recalc:公式格无缓存值,先跑 recalc.py"
    # 定位表头行:同一行里出现 Bear/Base/Bull
    hdr_row = None
    for row in ws.iter_rows(max_row=ws.max_row):
        vals = [str(c.value).strip().lower() for c in row if isinstance(c.value, str)]
        if "bear" in vals and "base" in vals and "bull" in vals:
            hdr_row = row[0].row
            cols = {}
            for c in row:
                t = str(c.value).strip().lower() if isinstance(c.value, str) else ""
                if t in ("bear", "base", "bull"):
                    cols[t] = c.column
            break
    if hdr_row is None:
        return None, "『估值对比』未找到 Bear/Base/Bull 表头行"
    # 表头行第一格(A 列)的标题,如 "项目(2027E)"
    head_label = ws.cell(hdr_row, 1).value or "项目"
    lines = [f"| {head_label} | Bear | Base | Bull |", "|---|---|---|---|"]
    n = 0
    for r in range(hdr_row + 1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if not isinstance(label, str) or not label.strip():
            break  # 遇空行/非标签行 → 块结束
        # 新案逐年推演块的小标题(含 "案 —")也视为结束
        if re.search(r"案\s*[—-]", label):
            break
        vb = ws.cell(r, cols["bear"]).value
        vba = ws.cell(r, cols["base"]).value
        vbu = ws.cell(r, cols["bull"]).value
        if all(x is None for x in (vb, vba, vbu)):
            continue
        lines.append(f"| {label.strip()} | {_fmt(label, vb)} | {_fmt(label, vba)} | {_fmt(label, vbu)} |")
        n += 1
    if n == 0:
        return None, "汇总块无数据行(可能未 recalc)"
    return "\n".join(lines), None


def _locate_cols(ws, max_scan=14, min_years=3):
    """定位年份表头行 + 返回 [(col,label)]:年份列 + 其右侧任何带字符串表头的列(当下/TTM/Q1)。
    通用于『历史财务与估值』(含当下/季报列)、『利润与收入假设』(纯年列)等。"""
    year_re = re.compile(r"^20\d\d")
    for row in ws.iter_rows(max_row=min(ws.max_row, max_scan)):
        yc = [(c.column, str(c.value).strip()) for c in row
              if isinstance(c.value, str) and year_re.match(str(c.value).strip())]
        if len(yc) >= min_years:
            first = yc[0][0]
            yre = re.compile(r"^20\d\d")
            extra_ok = ("当下", "ttm", "实际", "最新", "q1", "q2", "q3", "q4")
            cols = []
            for c in row:
                if c.column < first:
                    continue
                t = c.value
                if not (isinstance(t, str) and t.strip()):
                    continue
                lab = t.strip()
                # 只保留年份列 + 当下/季报列;剔除尾部「逻辑/备注/来源/说明」文字列
                if yre.match(lab) or any(k in lab.lower() for k in extra_ok):
                    cols.append((c.column, lab))
            return row[0].row, cols
    return None, []


def _grid_md(ws, hdr_row, cols, stop_blank_after=3):
    """把 hdr_row 之下、A 列有标签且年列有≥1数值的行渲成 Markdown。
    对重复的子行(YoY/占比)用上一主行短名消歧。遇连续多空行收尾。"""
    hdr = "| 指标 | " + " | ".join(y for _, y in cols) + " |"
    sep = "|---" * (len(cols) + 1) + "|"
    lines, n, blanks, last_main = [hdr, sep], 0, 0, "项目"
    for r in range(hdr_row + 1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        cells = [_fmt(ws.cell(r, 1).value or "", ws.cell(r, col).value) for col, _ in cols]
        if not isinstance(label, str) or not label.strip():
            blanks += 1
            if blanks >= stop_blank_after and n > 0:
                break
            continue
        lab = label.strip()
        if not any(cells):
            # 纯文字行(组标题/说明):跳过, 不进表
            continue
        # 说明/读法行:某个"数值格"其实是长文本(band_note 等)→ 跳过
        if any(len(c) > 22 for c in cells):
            continue
        blanks = 0
        sub = lab in ("YoY", "占总营收%", "占比", "占营收%")
        if sub:
            disp = f"　{last_main[:8]} {lab}"
        else:
            last_main = lab
            disp = lab
        lines.append(f"| {disp} | " + " | ".join(cells) + " |")
        n += 1
    return ("\n".join(lines), None) if n else (None, "无数据行")


def history_table(path):
    """抽『历史财务与估值』整页(全行) → Markdown。含分部/毛利率/净利/权益/股本/FX/股价/P-E/P-B/市值。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    sn = _find_sheet(wb, "历史财务")
    if not sn:
        return None, "无『历史财务』sheet"
    ws = wb[sn]
    if not _has_cache(ws):
        return None, "未 recalc:公式格无缓存值,先跑 recalc.py"
    hdr_row, cols = _locate_cols(ws)
    if hdr_row is None:
        return None, "未找到年份表头行"
    return _grid_md(ws, hdr_row, cols)


def forecast_table(path):
    """抽『利润与收入假设』Base 逐年明细 → Markdown(分部收入/占比/增速/利润率/净利/EPS/BPS/ROE)。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    sn = _find_sheet(wb, "利润与收入假设") or _find_sheet(wb, "利润")
    if not sn:
        return None, "无『利润与收入假设』sheet"
    ws = wb[sn]
    if not _has_cache(ws):
        return None, "未 recalc"
    hdr_row, cols = _locate_cols(ws)
    if hdr_row is None:
        return None, "未找到年份表头行"
    return _grid_md(ws, hdr_row, cols)


def ladder_table(path):
    """抽『估值对比』三案块各自的『隐含价』逐年行 → 三情景逐年隐含价 Markdown。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    sn = _find_sheet(wb, "估值对比")
    if not sn:
        return None, "无『估值对比』sheet"
    ws = wb[sn]
    if not _has_cache(ws):
        return None, "未 recalc"
    hdr_row, cols = _locate_cols(ws, max_scan=ws.max_row, min_years=4)
    if hdr_row is None:
        return None, "未找到年份表头行"
    # 仅取前瞻年(2026E 起), 跳过历史回测列
    fwd = [(col, y) for col, y in cols if y.endswith("E")]
    if not fwd:
        fwd = cols
    cases = ["Bear", "Base", "Bull"]
    rows = []
    for r in range(1, ws.max_row + 1):
        lab = ws.cell(r, 1).value
        if isinstance(lab, str) and "隐含价" in lab:
            vals = [ws.cell(r, col).value for col, _ in fwd]
            if sum(1 for v in vals if isinstance(v, (int, float))) >= 3:
                rows.append(vals)
    if len(rows) < 3:
        return None, f"仅找到 {len(rows)} 行隐含价(需3案)"
    hdr = "| 情景 | " + " | ".join(y for _, y in fwd) + " |"
    sep = "|---" * (len(fwd) + 1) + "|"
    lines = [hdr, sep]
    for cn, vals in zip(cases, rows[:3]):
        lines.append(f"| {cn} | " + " | ".join(_fmt("隐含价", v) for v in vals) + " |")
    return "\n".join(lines), None


def build_markdown(path, with_history=False, forecast=False, ladder=False):
    out = []
    tbl, err = scenario_table(path)
    out.append("### 三情景估值汇总（模型输出）\n\n" + tbl if tbl else f"<!-- 三情景表抽取失败: {err} -->")
    if forecast:
        f, fe = forecast_table(path)
        out.append("### 盈利预测表（Base，模型输出）\n\n" + f if f else f"<!-- 盈利预测表抽取失败: {fe} -->")
    if ladder:
        l, le = ladder_table(path)
        out.append("### 三情景逐年隐含价（模型输出）\n\n" + l if l else f"<!-- 逐年隐含价抽取失败: {le} -->")
    if with_history:
        h, he = history_table(path)
        out.append("### 历史财务与估值（模型基础数据，整页）\n\n" + h if h else f"<!-- 历史财务表抽取失败: {he} -->")
    return "\n\n".join(out)


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("-o", "--out")
    ap.add_argument("--with-history", action="store_true")
    ap.add_argument("--forecast", action="store_true")
    ap.add_argument("--ladder", action="store_true")
    ap.add_argument("--full", action="store_true", help="三情景汇总 + 盈利预测 + 逐年隐含价 + 历史整页")
    a = ap.parse_args()
    wh, fc, ld = a.with_history, a.forecast, a.ladder
    if a.full:
        wh = fc = ld = True
    md = build_markdown(a.xlsx, with_history=wh, forecast=fc, ladder=ld)
    if a.out:
        with open(a.out, "w", encoding="utf-8") as f:
            f.write(md)
        print(f"written: {a.out}")
    else:
        print(md)
