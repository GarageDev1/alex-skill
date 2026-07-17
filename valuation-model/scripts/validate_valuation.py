#!/usr/bin/env python3
"""
validate_valuation.py — valuation-model 业务级校验(对应 SKILL.md 连通性要求 + 验收清单)。
Excel 由 valuation_kit + examples/build_template.py 创建,公式重算由 scripts/recalc.py 完成。

检查(prose 要求的机检化, 见 SKILL.md 沉淀协议):
  1 连通性(到业务驱动量)  2 孤儿  3 溯源覆盖  4 ≥2 估值方法  5 禁单元格批注
  6 封面存在(05 §0)  7 表序=只向前引用 DAG(05 §1)
  8 估值对比引用隔离(04 三情景: 不得引用『情景切换』当前情景行或开关单元格)
  9 每个情景区块包含隐含 forward P/E 行
  10 估值对比历史列回测 ≈0%(需先 recalc 生成缓存值)
不是公式重算(先走本仓 scripts/recalc.py)。

用法:  python validate_valuation.py <path.xlsx>
输出:  JSON {checks:[{name,status,detail}], summary, verdict}
       status ∈ pass / warn / fail。fail = 触碰要求,必须修。

已知局限(静态正则解析公式,做不到的会标 warn 让人复核,不误判 fail):
  - INDEX/OFFSET/INDIRECT 等间接引用、命名区间 无法静态追踪 → 连通性可能 under-detect。
  - 主题色(theme color)的蓝字识别不到 → 建模用显式 Font(color="0000FF")。
  - 回测检查依赖缓存计算值: openpyxl 刚生成的文件没有缓存值,先用 scripts/recalc.py。
"""
import sys, re, json
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# 跨表引用:'Sheet Name'!$A$1 或 Sheet!A1 或 A1;区间 A1:B9
SHEET = r"(?:('[^']+'|[\w一-鿿]+)!)?"
RANGE_RE = re.compile(SHEET + r"\$?([A-Z]{1,3})\$?(\d+):\$?([A-Z]{1,3})\$?(\d+)")
CELL_RE  = re.compile(SHEET + r"\$?([A-Z]{1,3})\$?(\d+)")

ANCHOR_MARKER = "[anchor]"           # 基座 sheet 的 A1 须含此标记(见 05-excel-format.md)
ANCHOR_KW = ("基座", "anchor", "aidc", "tam", "capex", "装机", "出货", "mau", "arpu", "供需", "arr")
OUTPUT_KW = ("隐含价", "隐含股价", "目标价", "per share", "每股", "implied", "公允价")
# 估值方法识别用「明确 token」,避免 'pe' 命中 capex 这类误判
LENS_PATTERNS = {
    "P/E":       [r"p\s*/\s*e\b", r"市盈"],
    "P/B":       [r"p\s*/\s*b\b", r"市净", r"justified\s*p/?b"],
    "DCF":       [r"\bdcf\b", r"\bwacc\b", r"贴现", r"折现现金流"],
    "EV/EBITDA": [r"ev\s*/\s*ebitda"],
    "EV/Sales":  [r"ev\s*/\s*sales", r"ev\s*/\s*销售"],
    "DDM":       [r"\bddm\b", r"股息贴现"],
    "SOTP":      [r"\bsotp\b", r"分部估值"],
}
# 末端叶子行白名单:这些本就该没有下游引用,不算孤儿
LEAF_KW = ("合计", "总营收", "总计", "隐含价", "目标价", "per share", "每股", "mix", "占比",
           "yoy", "同比", "roe", "eps", "bps", "margin", "利润率", "净利率", "率", "vs", "差异", "upside",
           "市值", "毛利", "隐含", "情绪", "回测", "比值", "股价", "forward", "gap", "p/e", "p/b", "判断",
           "bear", "base", "bull", "兆krw", "闭环", "占")
# 展示/输出型 sheet 整张豁免孤儿扫描(它们的行天生无下游引用; 孤儿检查针对中间 driver 表:
# 该检查用于发现未被后续公式引用的基准、测算和假设。展示表由连通性、引用隔离和回测检查覆盖。
ORPHAN_EXEMPT_SHEET_KW = ("封面", "共识", "股价走势", "历史估值倍数", "情景估值", "估值对比", "仪表盘")


def key(sheet, coord):
    return f"{sheet}\x00{coord}"


def parse_refs(sheet, formula, names):
    """抽公式依赖:先展开区间(矩形内全部 cell),再抽单格。表名缺省=本表。"""
    refs = set()
    def resolve(sh):
        return sh.strip("'") if sh else sheet
    for m in RANGE_RE.finditer(formula):
        sh = resolve(m.group(1))
        if sh not in names:
            continue
        c1, r1, c2, r2 = m.group(2), int(m.group(3)), m.group(4), int(m.group(5))
        a, b = column_index_from_string(c1), column_index_from_string(c2)
        for col in range(min(a, b), max(a, b) + 1):
            for row in range(min(r1, r2), max(r1, r2) + 1):
                refs.add(key(sh, f"{get_column_letter(col)}{row}"))
    # 抽单格:把区间从字符串抹掉再扫,避免把区间端点当单格重复处理
    residual = RANGE_RE.sub("  ", formula)
    for m in CELL_RE.finditer(residual):
        sh = resolve(m.group(1))
        if sh in names:
            refs.add(key(sh, f"{m.group(2)}{m.group(3)}"))
    return refs


def row_label(ws, row):
    """该行的文字标签(取本行第一个字符串 cell),用于白名单判断。"""
    for c in ws[row]:
        if isinstance(c.value, str) and c.value.strip():
            return c.value.lower()
    return ""


def main(path):
    wb = openpyxl.load_workbook(path, data_only=False)
    names = set(wb.sheetnames)
    checks = []

    deps, referenced = {}, set()
    formula_cells, input_cells = [], []
    text_blob_parts = []
    # 标记法识别 anchor sheet
    marked_anchor_sheets = set()

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                if isinstance(v, str):
                    text_blob_parts.append(v.lower())
                    if ANCHOR_MARKER in v.lower():
                        marked_anchor_sheets.add(ws.title)
                k = key(ws.title, cell.coordinate)
                if isinstance(v, str) and v.startswith("="):
                    formula_cells.append((k, ws.title, cell))
                    r = parse_refs(ws.title, v, names)
                    deps[k] = r
                    referenced |= r
                elif isinstance(v, (int, float)):
                    input_cells.append((k, ws.title, cell))
    text_blob = " ".join(text_blob_parts)

    # ── 1: ≥2 估值方法(明确 token,不用无边界子串)──────────────
    lenses = [name for name, pats in LENS_PATTERNS.items()
              if any(re.search(p, text_blob) for p in pats)]
    checks.append({"name": "多种估值方法(≥2)",
                   "status": "pass" if len(lenses) >= 2 else "fail",
                   "detail": f"检出估值方法: {lenses or '无'}"})

    # ── 2: 溯源覆盖(行级:每个蓝字输入行需有可见逻辑文本)────────
    # 新模型全面禁用单元格批注,所以溯源只认可见注释列/逻辑文本。
    blue_rows = set(); logic_rows = set(); comment_cells = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            row_has_logic = any(isinstance(c.value, str) and not c.value.startswith("=")
                                and len(c.value.strip()) >= 8 for c in row)
            for cell in row:
                rk = (ws.title, cell.row)
                col = cell.font and cell.font.color
                is_blue = bool(col and col.rgb and str(col.rgb).upper().endswith("0000FF"))
                if is_blue and isinstance(cell.value, (int, float)):
                    blue_rows.add(rk)
                if cell.comment:
                    comment_cells.append(f"{ws.title}!{cell.coordinate}")
            if row_has_logic:
                logic_rows.add((ws.title, row[0].row))
        # 合并单元格的可见逻辑覆盖其跨到的所有行(说明可合并写一次, 不必逐行重复)
        for mr in ws.merged_cells.ranges:
            tl = ws.cell(mr.min_row, mr.min_col).value
            if isinstance(tl, str) and not tl.startswith("=") and len(tl.strip()) >= 8:
                for rr in range(mr.min_row, mr.max_row + 1):
                    logic_rows.add((ws.title, rr))
    sourced = blue_rows & logic_rows
    if not blue_rows:
        cov_status, cov_detail = "warn", "未识别到蓝字硬编码输入行"
    else:
        rate = len(sourced) / len(blue_rows)
        cov_status = "pass" if rate >= 0.9 else ("warn" if rate >= 0.7 else "fail")
        miss = sorted({f"{s}!行{r}" for (s, r) in (blue_rows - sourced)})[:6]
        cov_detail = (f"蓝字输入行 {len(blue_rows)}, 有可见逻辑/备注 {len(sourced)} ({rate*100:.0f}%)"
                      + (f";缺溯源行抽样 {miss}" if miss else ""))
    checks.append({"name": "溯源覆盖(行级)", "status": cov_status, "detail": cov_detail})

    # ── 3: 禁单元格批注(红三角)────────────────────────
    checks.append({"name": "禁单元格批注",
                   "status": "pass" if not comment_cells else "fail",
                   "detail": "未发现单元格批注" if not comment_cells
                             else f"发现 {len(comment_cells)} 个单元格批注(抽样 {comment_cells[:6]})"})

    # ── 4: 连通性(隐含价 → 业务驱动量)─────────────────────
    anchor_sheets = marked_anchor_sheets or {n for n in names
                                             if any(kw in n.lower() for kw in ANCHOR_KW)}
    # 输出行识别: 只认行标签(第一个文字格), 跳过表头/标题行(行1-2);
    # 连通性按「行级」判定:隐含价行的历史列用于回测,依据历史实际值,不关联前瞻业务驱动量。
    # 只要该行有 ≥1 个公式格(前瞻列)能回溯到基准, 整行算通。
    out_rows = []   # [(sheet, row_idx, [formula keys]), ...]
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=3):
            rlab = row_label(ws, row[0].row)
            if any(kw in rlab for kw in OUTPUT_KW):
                cells = [key(ws.title, c.coordinate) for c in row
                         if isinstance(c.value, str) and c.value.startswith("=")]
                if cells:
                    out_rows.append((ws.title, row[0].row, cells))
    out_targets = [k for _s, _r, cs in out_rows for k in cs]
    if not anchor_sheets:
        checks.append({"name": "连通性(到业务驱动量)", "status": "warn",
                       "detail": "未找到 [ANCHOR] 标记或基准/基座 sheet;无法自动判定,请人工跑'改底层假设'测试"})
    elif not out_targets:
        checks.append({"name": "连通性(到业务驱动量)", "status": "warn",
                       "detail": "未识别到隐含价/目标价输出格(行内需含'隐含价/目标价'等字样)"})
    else:
        def reaches_anchor(t):
            seen, stack = set(), [t]
            while stack:
                cur = stack.pop()
                if cur in seen:
                    continue
                seen.add(cur)
                if cur.split("\x00")[0] in anchor_sheets:
                    return True
                stack.extend(deps.get(cur, ()))
            return False

        reached_rows, fail_rows = 0, []
        for sh, ridx, cells in out_rows:
            if any(reaches_anchor(t) for t in cells):
                reached_rows += 1
            else:
                fail_rows.append(f"{sh}!行{ridx}")
        status = "pass" if reached_rows == len(out_rows) else ("warn" if reached_rows else "fail")
        src = "标记" if marked_anchor_sheets else "名字关键词(建议改用 [ANCHOR] 标记更稳)"
        checks.append({"name": "连通性(到业务驱动量)", "status": status,
                       "detail": (f"{reached_rows}/{len(out_rows)} 个隐含价输出行能回溯到基准 sheet "
                                  f"{sorted(anchor_sheets)}(行级:历史回测列不要求关联业务驱动量;识别方式:{src})"
                                  + (f";未通行 {fail_rows[:6]}" if fail_rows else ""))})

    # ── 5: 孤儿扫描(中间 driver 算了没人引用)──────────
    out_set = set(out_targets)
    # 行级判定: 一行里只要有 ≥1 个公式格被下游引用, 整行算"有人用"
    # (如情景切换『当前案』行: 前瞻列被引用用、历史列只作展示对照, 不算孤儿)
    row_cells, row_used = {}, set()
    for k, sh, cell in formula_cells:
        if any(kw in sh for kw in ORPHAN_EXEMPT_SHEET_KW):
            continue  # 展示/输出表整张豁免
        rk = (sh, cell.row)
        row_cells.setdefault(rk, []).append(k)
        if k in referenced or k in out_set:
            row_used.add(rk)
    mid_orphans = []
    for rk, cells in row_cells.items():
        if rk in row_used:
            continue
        if any(kw in row_label(wb[rk[0]], rk[1]) for kw in LEAF_KW):
            continue  # 末端叶子行(合计/隐含价/占比…)本就无下游,放行
        mid_orphans.append(f"{rk[0]}!行{rk[1]}")
    if not mid_orphans:
        checks.append({"name": "孤儿扫描", "status": "pass",
                       "detail": "无中间 driver 孤儿(末端叶子行已白名单豁免)"})
    else:
        checks.append({"name": "孤儿扫描", "status": "warn",
                       "detail": f"疑似中间 driver 无下游引用 {len(mid_orphans)} 个(抽样 {mid_orphans[:8]})"
                                 " — 逐个核:真 driver 出现在此=链没接全,需接上"})

    # ── 6: 封面存在(第一张 sheet = 封面, 含 Key Takeaways; 05 §0)──
    first = wb.worksheets[0]
    first_text = " ".join(str(c.value).lower() for row in first.iter_rows()
                          for c in row if isinstance(c.value, str))
    has_cover = ("封面" in first.title) or ("key takeaway" in first_text)
    checks.append({"name": "封面(第一张+Key Takeaways)",
                   "status": "pass" if has_cover else "fail",
                   "detail": f"第一张 sheet '{first.title}'" + ("" if has_cover else " 不是封面/无 Key Takeaways")})

    # ── 7: 表序 = 只向前引用的 DAG(05 §1: 每张表只引它左侧的表)──
    sheet_idx = {n: i for i, n in enumerate(wb.sheetnames)}
    dag_viol = []
    for k, sh, cell in formula_cells:
        for ref in deps.get(k, ()):
            tgt = ref.split("\x00")[0]
            if tgt != sh and sheet_idx.get(tgt, -1) > sheet_idx[sh]:
                dag_viol.append(f"{sh}!{cell.coordinate}→{tgt}")
    checks.append({"name": "表序DAG(只向前引用)",
                   "status": "pass" if not dag_viol else "fail",
                   "detail": "无反向引用" if not dag_viol
                   else f"{len(dag_viol)} 处引用右侧表(抽样 {sorted(set(dag_viol))[:6]}) — 调整表序或拆分"})

    # ── 8-10: 三情景架构检查(有『情景切换』才查; 单情景模型跳过)──
    sw_sheet = next((n for n in wb.sheetnames if "情景切换" in n), None)
    cmp_sheet = next((n for n in wb.sheetnames if "估值对比" in n), None)
    if sw_sheet:
        if not cmp_sheet:
            checks.append({"name": "三情景架构", "status": "fail",
                           "detail": "有『情景切换』但缺『估值对比』(三案恒常并排是主输出)"})
        else:
            wcmp = wb[cmp_sheet]
            # 8: 引用隔离。对比表不得引用『情景切换』的当前情景行、开关或情景序号行。
            wsw = wb[sw_sheet]
            polluted_rows = set()
            for row in wsw.iter_rows():
                rtext = " ".join(str(c.value) for c in row if isinstance(c.value, str))
                if ("当前案" in rtext) or ("当前情景" in rtext) or ("案序号" in rtext):
                    polluted_rows.add(row[0].row)
            bad = []
            for k, sh, cell in formula_cells:
                if sh != cmp_sheet:
                    continue
                for ref in deps.get(k, ()):
                    tgt, coord = ref.split("\x00")
                    if tgt == sw_sheet:
                        mrow = int(re.sub(r"[A-Z$]", "", coord) or 0)
                        if mrow in polluted_rows:
                            bad.append(f"{cell.coordinate}→{sw_sheet}!{coord}")
            checks.append({"name": "估值对比引用隔离",
                           "status": "pass" if not bad else "fail",
                           "detail": "各情景列未引用当前情景行或开关行" if not bad
                           else f"{len(bad)} 处引用当前情景行或开关行(抽样 {bad[:6]});请改为引用对应的 Bear、Base 或 Bull 参数行"})
            # 9: 每情景 block 有 隐含 forward P/E 行
            ipe_rows = 0
            n_cases = 0
            for row in wcmp.iter_rows():
                rtext = " ".join(str(c.value) for c in row if isinstance(c.value, str))
                if "隐含" in rtext and re.search(r"(forward\s*)?P\s*/?\s*E", rtext, re.I):
                    ipe_rows += 1
                if re.search(r"案\s*—|案 —", rtext):
                    n_cases += 1
            need = max(n_cases, 3)
            st = "pass" if ipe_rows >= need else ("warn" if ipe_rows else "fail")
            checks.append({"name": "每情景隐含forward P/E",
                           "status": st,
                           "detail": f"估值对比检出 {ipe_rows} 行隐含 P/E(需≥{need}: 每案 block 一行)"})
            # 10: 历史列内置回测 ≈0%(需缓存计算值)
            try:
                wbv = openpyxl.load_workbook(path, data_only=True)
                wcv = wbv[cmp_sheet]
                hdr_year = {}
                for row in wcv.iter_rows(max_row=min(wcv.max_row, 40)):
                    for c in row:
                        if isinstance(c.value, str) and re.match(r"^20\d\d[AE]?$", c.value.strip()):
                            hdr_year.setdefault(get_column_letter(c.column), c.value.strip())
                hist_cols_cmp = [cl for cl, y in hdr_year.items() if y.endswith("A") or re.match(r"^20\d\d$", y)]
                vals, none_cnt = [], 0
                for row in wcv.iter_rows():
                    rtext = " ".join(str(c.value) for c in row if isinstance(c.value, str))
                    if "回测" not in rtext:
                        continue
                    for c in row:
                        if get_column_letter(c.column) in hist_cols_cmp:
                            if isinstance(c.value, (int, float)):
                                vals.append(abs(c.value))
                            elif c.value is None:
                                none_cnt += 1
                if vals:
                    worst = max(vals)
                    st = "pass" if worst <= 0.05 else "warn"
                    checks.append({"name": "估值对比历史回测≈0",
                                   "status": st,
                                   "detail": f"回测行历史列 |偏差| 最大 {worst*100:.1f}%(≤5% 视为复现)"})
                else:
                    checks.append({"name": "估值对比历史回测≈0", "status": "warn",
                                   "detail": "未取到缓存计算值(先用 scripts/recalc.py 再跑)或缺『回测』行"})
            except Exception as e:
                checks.append({"name": "估值对比历史回测≈0", "status": "warn",
                               "detail": f"回测检查异常({e}), 人工复核"})

    # ── 11: $B 口径统一(scale 要求, 见 02-data-grounding.md §29)──────
    # 要求: 营收/利润/EBITDA/权益 等财务量纲行必须统一用 USD $B;
    # 不得用本币百万/亿/兆/EUR/SEK/NT$ 作主口径(市值可双币种、可附 FX 行=例外)。
    FIN_ROW = re.compile(r"营收|总收入|总营收|营业收入|revenue|净利|净利润|营业利润|"
                         r"operating income|ebitda|毛利|gross profit|权益|equity|股东权益")
    EXCL_ROW = re.compile(r"率|%|margin|yoy|同比|环比|per share|每股|eps|bps|占比|增速|"
                          r"mix|p/e|p/b|roe|roic|强度")
    LOCAL_UNIT = re.compile(r"亿|百万(?!美元)|兆|万元|十亿(?!美元)|EUR|SEK|NT\$|人民币|"
                            r"RMB|韩元|日元|JPY|千元|€")
    USD_B = re.compile(r"\$\s?b|\$bn|usd\s?bn|usd\s?billion|十亿美元", re.I)
    try:
        wbv = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        wbv = None
    fin_rows = []   # (sheet, row, label, maxabs)
    for ws in wb.worksheets:
        wsv = wbv[ws.title] if wbv and ws.title in wbv.sheetnames else None
        for row in ws.iter_rows(min_row=2):
            lab = row_label(ws, row[0].row)
            if not lab or not FIN_ROW.search(lab) or EXCL_ROW.search(lab):
                continue
            mx = 0
            if wsv:
                for c in wsv[row[0].row]:
                    if isinstance(c.value, (int, float)):
                        mx = max(mx, abs(c.value))
            fin_rows.append((ws.title, row[0].row, lab, mx))
    usd_rows = [r for r in fin_rows if USD_B.search(r[2])]
    local_rows = [r for r in fin_rows if LOCAL_UNIT.search(r[2]) and not USD_B.search(r[2])]
    huge_rows = [r for r in fin_rows if r[3] >= 5000]  # $B 下营收/利润不可能 >5000
    if not fin_rows:
        scale_status, scale_detail = "warn", "未识别到财务量纲行(营收/利润/权益),无法判 scale"
    elif not usd_rows:
        scale_status = "fail"
        scale_detail = (f"财务量纲行 {len(fin_rows)} 个, 但无一行是 $B 口径(scale 违规)。"
                        f"抽样: {[r[2] for r in fin_rows[:4]]}")
    elif huge_rows:
        scale_status = "fail"
        scale_detail = (f"{len(huge_rows)} 个财务行量级≥5000, 远超 $B 合理区间(疑似本币百万/亿)。"
                        f"抽样: {[(r[2], f'{r[3]:.3g}') for r in huge_rows[:4]]}")
    elif local_rows:
        scale_status = "warn"
        scale_detail = (f"有 $B 主口径 {len(usd_rows)} 行, 但另有 {len(local_rows)} 个本币量纲财务行:"
                        f"确认是次级展示/FX 行而非主口径。抽样: {[r[2] for r in local_rows[:4]]}")
    else:
        scale_status = "pass"
        scale_detail = f"财务量纲行统一 $B 口径({len(usd_rows)}/{len(fin_rows)} 显式标 $B, 无超量级行)"
    checks.append({"name": "$B 口径统一(scale)", "status": scale_status, "detail": scale_detail})

    # ── 12: 结构 = 业务驱动量驱动模型, 非旧 financial-analysis DCF 模板 ────
    # 杜绝管线退回 IS/BS/CF/DCF/Comps 那套(它未建立公式联动、不是 $B 口径、无三情景并排)。
    OLD_TMPL = {"income statement", "balance sheet", "cash flow", "cash flow statement",
                "dcf", "dcf inputs", "dcf valuation", "comps", "sensitivity", "revenue model",
                "valuation summary"}
    STD_SHEETS = {"封面", "估值对比", "情景切换", "综合判断仪表盘"}
    lower_sheets = {n.lower().lstrip("0123456789. ") for n in wb.sheetnames}
    old_hit = sum(1 for n in lower_sheets if n in OLD_TMPL)
    std_hit = sum(1 for n in wb.sheetnames if any(s in n for s in STD_SHEETS))
    if old_hit >= 4 and std_hit < 2:
        struct_status = "fail"
        struct_detail = (f"检出旧 financial-analysis DCF 模板结构({old_hit} 张 IS/BS/CF/DCF/Comps), "
                         f"标准业务驱动量 sheet 仅 {std_hit}/4。必须重建为业务驱动量驱动模型。")
    elif std_hit < 3:
        struct_status = "warn"
        struct_detail = f"标准 sheet(封面/估值对比/情景切换/仪表盘)仅 {std_hit}/4, 复核结构完整性"
    else:
        struct_status = "pass"
        struct_detail = f"业务驱动量驱动结构完整(标准 sheet {std_hit}/4, 无旧 DCF 模板痕迹)"
    checks.append({"name": "结构(业务驱动量驱动, 非旧DCF模板)", "status": struct_status, "detail": struct_detail})

    # ── 13: 公式错误值扫描(#DIV/0! / #REF! 等)— 抓"在场但算坏"的公式/估值方法 ──
    # keyword 检估值方法只看"有没有 DCF/PE 字样", 抓不到估值方法算出 #DIV/0!(实测 ASML DCF
    # 因 mtext 合并吞掉 WACC 单元格 → 全部计算关系 #DIV/0! 却 keyword 仍 pass)。这一关补上。
    ERR_TOKENS = ("#DIV/0!", "#REF!", "#VALUE!", "#NAME?", "#NUM!", "#NULL!", "#N/A")
    err_cells = []
    if wbv is not None:
        for ws in wbv.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.strip() in ERR_TOKENS:
                        err_cells.append(f"{ws.title}!{c.coordinate}={c.value.strip()}")
        err_status = "pass" if not err_cells else "fail"
        err_detail = ("无公式错误值" if not err_cells
                      else f"{len(err_cells)} 个单元格是公式错误值(抽样 {err_cells[:6]}):有公式/估值方法算坏了,必须修")
    else:
        err_status, err_detail = "warn", "未取到缓存值(先 recalc)无法扫错误值"
    checks.append({"name": "公式无错误值(#DIV/0!等)", "status": err_status, "detail": err_detail})

    summary = {s: sum(1 for c in checks if c["status"] == s) for s in ("pass", "warn", "fail")}
    return {"file": path, "lenses": lenses, "checks": checks, "summary": summary,
            "verdict": "FAIL" if summary["fail"] else ("REVIEW" if summary["warn"] else "PASS")}


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"status": "error", "message": "用法: python validate_valuation.py <path.xlsx>"}))
        sys.exit(1)
    print(json.dumps(main(sys.argv[1]), ensure_ascii=False, indent=2))
